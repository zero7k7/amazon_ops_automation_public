from __future__ import annotations

import argparse
print("[BOOT] main.py started", flush=True)
import csv
import json
import os
import re
import shutil
import sys
import time
import urllib.parse
from datetime import datetime
from pathlib import Path

print("[IMPORT] start pandas", flush=True)
import pandas as pd
print("[IMPORT] done pandas", flush=True)
print("[IMPORT] start openpyxl", flush=True)
from openpyxl import load_workbook
print("[IMPORT] done openpyxl", flush=True)

ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

print("[IMPORT] start src modules", flush=True)
from src.analyze_rules import build_analysis_payload, build_no_order_diagnostics
from src.autoopt_feedback import build_autoopt_payload, write_autoopt_outputs
from src.amazon_custom_enhancements import build_enhanced_bundle
from src.build_sku_asin_map import build_sku_asin_map_from_cost_config
from src.data_quality import (
    build_data_quality_issue_detail,
    build_missing_sku_detail,
    write_data_quality_issue_detail,
    write_missing_sku_detail,
)
from src.db import AnalyticsDatabase
from src.generate_excel_report import generate_all_marketplace_excel_report, generate_excel_report
from src.generate_html_report import write_dashboard_html, write_html_report, write_marketplace_report_html, write_recommendations_workbench_html, write_summary_html
from src.inventory_replenishment import build_inventory_replenishment
from src.merge_data import build_daily_dataset
from src.metrics import build_windowed_views
from src.normalize_fields import FieldValidationError
from src.parse_ads_report import load_ads_report
from src.parse_erp_sales import load_erp_report
from src.report_presentation import (
    build_all_enhanced_requests_markdown,
    build_all_marketplace_markdown,
    build_report_view,
    build_marketplace_summary_markdown,
    build_recommendations_markdown,
)
from src.sku_alias import build_alias_review
from src.task_history import annotate_task_history
print("[IMPORT] done src modules", flush=True)

ALL_ADS_REPORT_PATH = ROOT / "data" / "raw_ads" / "ads_report_all.csv"
ALL_ADS_REPORT_XLSX_PATH = ROOT / "data" / "raw_ads" / "ads_report_all.xlsx"
ALL_ERP_REPORT_PATH = ROOT / "data" / "raw_erp" / "sales_report_all.xlsx"
LEGACY_ADS_REPORT_PATH = ROOT / "data" / "raw_ads" / "ads_report_uk.csv"
LEGACY_ADS_REPORT_XLSX_PATH = ROOT / "data" / "raw_ads" / "ads_report_uk.xlsx"
LEGACY_ERP_REPORT_PATH = ROOT / "data" / "raw_erp" / "sales_report_uk.xlsx"
OUTPUT_DIR = ROOT / "data" / "output"
SKU_MAP_PATH = ROOT / "config" / "sku_asin_map.xlsx"
SKU_ALIAS_MAP_PATH = ROOT / "config" / "sku_alias_map.xlsx"
PRODUCT_CONFIG_PATH = ROOT / "config" / "product_config.xlsx"
COST_CONFIG_PATH = ROOT / "config" / "product_cost_config.xlsx"
IGNORED_ISSUES_PATH = ROOT / "config" / "ignored_quality_issues.xlsx"
ARCHIVE_ADS_DIR = ROOT / "data" / "archive" / "ads"
ARCHIVE_ERP_DIR = ROOT / "data" / "archive" / "erp"
CUSTOM_DATA_DIR = ROOT / "data" / "raw_amazon_custom"
SUPPORTED_MARKETPLACES = ["UK", "US", "DE"]
DEBUG_TIMING_EXIT_CODE = 2


def _log_step(step: str, elapsed_sec: float) -> None:
    print(f"[timing] {step}: {elapsed_sec:.3f}s")


def _start(step: str) -> float:
    print(f"[START] {step}", flush=True)
    return time.perf_counter()


def _done(step: str, started: float) -> float:
    elapsed = time.perf_counter() - started
    print(f"[DONE] {step}: {elapsed:.3f}秒", flush=True)
    return elapsed


def _timeout_step(step: str) -> None:
    print(f"[TIMEOUT_STEP] 卡在 {step}", flush=True)


def _to_file_uri(path: Path) -> str:
    posix_path = path.resolve().as_posix()
    drive_pattern = re.match(r"^([A-Za-z]):/(.*)$", posix_path)
    if drive_pattern:
        drive = drive_pattern.group(1)
        tail = drive_pattern.group(2)
        return f"file:///{drive}:/{urllib.parse.quote(tail)}"
    return f"file:///{urllib.parse.quote(posix_path)}"


def _print_report_entrypoints(report_path: Path, output_dir: Path = OUTPUT_DIR) -> None:
    dashboard = output_dir / "dashboard.html"
    latest_html = output_dir / "latest_recommendations.html"
    import_manifest = output_dir / "import_manifest.xlsx"
    print("", flush=True)
    print("报告已生成：", flush=True)
    print("", flush=True)
    print("1. 总览 Dashboard：", flush=True)
    print(_to_file_uri(dashboard), flush=True)
    print("", flush=True)
    print("2. 详细 HTML 报告：", flush=True)
    print(_to_file_uri(latest_html), flush=True)
    print("", flush=True)
    print("3. Excel 报告：", flush=True)
    print(str(report_path.resolve()), flush=True)
    print("", flush=True)
    print("4. 导入日志：", flush=True)
    print(str(import_manifest.resolve()), flush=True)


def _write_autoopt_outputs(results: list[dict], output_dir: Path, fallback_report_date: str | None = None) -> dict[str, object]:
    payload = build_autoopt_payload(results, output_dir=output_dir)
    report_date = str(payload.get("report_date") or fallback_report_date or datetime.now().date().isoformat())
    json_path, xlsx_path = write_autoopt_outputs(output_dir, report_date, payload)
    return {
        "autoopt_json_path": json_path,
        "autoopt_xlsx_path": xlsx_path,
        "autoopt_payload": payload,
    }


def _inject_current_autoopt_reviews(results: list[dict], autoopt_payload: dict[str, object]) -> None:
    action_rows = [
        row for row in autoopt_payload.get("action_review_rows", []) if isinstance(row, dict)
    ]
    keyword_rows = [
        row for row in autoopt_payload.get("keyword_action_review_rows", []) if isinstance(row, dict)
    ]
    for result in results:
        marketplace = str(result.get("marketplace") or "").strip().upper()
        if not marketplace:
            continue
        report_view = result.get("report_view")
        if not isinstance(report_view, dict):
            continue
        report_view["action_effect_review_rows"] = [
            row for row in action_rows if str(row.get("marketplace") or "").strip().upper() == marketplace
        ]
        report_view["keyword_action_effect_review_rows"] = [
            row for row in keyword_rows if str(row.get("marketplace") or "").strip().upper() == marketplace
        ]


def _debug_read_enhanced_headers(custom_dir: Path, marketplace: str) -> tuple[float, float]:
    site_dir = custom_dir / marketplace.upper()
    scan_dir = site_dir if site_dir.exists() and any(site_dir.iterdir()) else custom_dir
    traffic_elapsed = 0.0
    query_elapsed = 0.0
    for path in sorted(scan_dir.iterdir()):
        if not path.is_file() or path.suffix.lower() not in {".xlsx", ".xls", ".csv"}:
            continue
        lower = path.name.lower()
        if "traffic_sales" not in lower and "metric-data" not in lower and "search_query_performance" not in lower:
            continue
        if "traffic_sales" in lower or ("metric-data" in lower and "(1)" not in lower):
            t0 = time.perf_counter()
            if path.suffix.lower() == ".csv":
                with path.open("r", encoding="utf-8-sig", newline="") as f:
                    next(csv.reader(f), [])
            else:
                wb = load_workbook(path, read_only=True, data_only=False)
                try:
                    ws = wb[wb.sheetnames[0]]
                    _ = next(ws.iter_rows(values_only=True), [])
                finally:
                    wb.close()
            traffic_elapsed += time.perf_counter() - t0
        if "search_query_performance" in lower or ("metric-data" in lower and "(1)" in lower):
            t0 = time.perf_counter()
            if path.suffix.lower() == ".csv":
                with path.open("r", encoding="utf-8-sig", newline="") as f:
                    next(csv.reader(f), [])
            else:
                wb = load_workbook(path, read_only=True, data_only=False)
                try:
                    ws = wb[wb.sheetnames[0]]
                    _ = next(ws.iter_rows(values_only=True), [])
                finally:
                    wb.close()
            query_elapsed += time.perf_counter() - t0
    return traffic_elapsed, query_elapsed


def ensure_directories() -> None:
    for relative in [
        "config",
        "data/raw_ads",
        "data/raw_erp",
        "data/inbox",
        "data/raw_amazon_custom",
        "data/raw_amazon_custom/UK",
        "data/raw_amazon_custom/US",
        "data/raw_amazon_custom/DE",
        "data/output",
        "data/archive/ads",
        "data/archive/erp",
        "database",
        "logs",
        "src",
        "tests",
    ]:
        (ROOT / relative).mkdir(parents=True, exist_ok=True)


def ensure_ignored_issues_config() -> None:
    if IGNORED_ISSUES_PATH.exists():
        return
    frame = pd.DataFrame(
        [
            {"marketplace": "UK", "asin": "B0D5D1H28J", "reason": "鐢ㄦ埛纭蹇界暐"},
            {"marketplace": "UK", "asin": "B0DBVJH8KW", "reason": "鐢ㄦ埛纭蹇界暐"},
        ]
    )
    frame.to_excel(IGNORED_ISSUES_PATH, index=False)


def write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _review_snapshot_rows(rows: object) -> list[object]:
    if not isinstance(rows, list):
        return []
    normalized: list[object] = []
    for row in rows:
        if not isinstance(row, dict):
            normalized.append(row)
            continue
        item = dict(row)
        if not str(item.get("effect_metrics") or "").strip():
            metric_text = str(item.get("effect_evidence") or item.get("review_status") or "").strip()
            if metric_text:
                item["effect_metrics"] = metric_text
        normalized.append(item)
    return normalized


def _report_view_snapshot(view: dict) -> dict:
    keys = [
        "analysis_status",
        "issue_summary",
        "strong_recommendation_allowed",
        "summary_lines",
        "today_task_queue_rows",
        "tomorrow_review_rows",
        "search_term_processing_queue_rows",
        "html_search_term_processing_queue_rows",
        "scale_rows",
        "scale_keyword_rows",
        "growth_test_rows",
        "listing_price_diagnosis_rows",
        "frontend_check_queue_rows",
        "inventory_replenishment_rows",
        "product_final_decision_rows",
        "product_operation_cards",
        "cost_profit_diagnosis_rows",
        "true_unsold_diagnosis_rows",
        "recent_conversion_cliff_diagnosis_rows",
        "ad_no_conversion_diagnosis_rows",
        "ad_attribution_weak_diagnosis_rows",
        "yesterday_attribution_rows",
        "action_effect_review_rows",
        "keyword_action_effect_review_rows",
        "enhanced_status_rows",
        "enhanced_request_rows",
        "optimization_notes",
        "runtime_policy_notes",
    ]
    snapshot = {key: view.get(key, [] if key.endswith("_rows") or key.endswith("_notes") else "") for key in keys}
    snapshot["today_action_groups"] = view.get("today_action_groups", {})
    snapshot["final_decision_summary"] = view.get("final_decision_summary", {})
    snapshot["decision_gate_counts"] = view.get("decision_gate_counts", {})
    snapshot["frontend_coverage_summary"] = view.get("frontend_coverage_summary", {})
    snapshot["action_effect_review_rows"] = _review_snapshot_rows(snapshot.get("action_effect_review_rows"))
    snapshot["keyword_action_effect_review_rows"] = _review_snapshot_rows(snapshot.get("keyword_action_effect_review_rows"))
    return snapshot


def _aggregate_frontend_coverage(results: list[dict]) -> dict[str, object]:
    count_fields = [
        "frontend_queue_total",
        "frontend_usable_evidence_count",
        "frontend_decision_ready_count",
        "frontend_reference_evidence_count",
        "frontend_live_success_count",
        "frontend_cached_count",
        "frontend_pending_or_stale_count",
        "frontend_search_success_count",
        "frontend_search_partial_count",
        "frontend_product_page_success_count",
        "frontend_competitor_search_success_count",
        "frontend_own_sellersprite_count",
        "frontend_own_sellersprite_today_count",
        "frontend_own_sellersprite_cache_count",
        "frontend_own_sellersprite_pending_count",
        "frontend_own_sellersprite_failed_count",
        "frontend_sellersprite_trend_ready_count",
        "frontend_competitor_discovery_count",
        "frontend_competitor_pool_count",
        "frontend_competitor_pool_today_count",
        "frontend_competitor_pool_cache_count",
        "frontend_competitor_pool_pending_count",
        "frontend_competitor_pool_failed_count",
        "frontend_competitor_sellersprite_count",
        "frontend_competitor_sellersprite_today_count",
        "frontend_competitor_sellersprite_cache_count",
        "frontend_competitor_sellersprite_pending_count",
        "frontend_competitor_sellersprite_asin_count",
        "frontend_amazon_search_validation_count",
        "frontend_scalable_strong_count",
        "frontend_weak_defensive_count",
        "frontend_insufficient_count",
        "frontend_strong_evidence_count",
        "frontend_background_evidence_count",
        "frontend_unusable_evidence_count",
        "market_survey_complete_count",
        "market_survey_usable_count",
        "market_survey_insufficient_count",
        "market_survey_failed_count",
    ]
    totals = {field: 0 for field in count_fields}
    market_survey_score_total = 0.0
    for result in results:
        coverage = (result.get("report_view") or {}).get("frontend_coverage_summary", {})
        if not isinstance(coverage, dict):
            continue
        queue_total = int(float(coverage.get("frontend_queue_total", 0) or 0))
        market_survey_score_total += float(coverage.get("market_survey_average_score", 0) or 0) * queue_total
        for field in count_fields:
            try:
                totals[field] += int(float(coverage.get(field, 0) or 0))
            except (TypeError, ValueError):
                continue
    total = totals["frontend_queue_total"]
    usable = totals["frontend_usable_evidence_count"]
    strong = totals["frontend_decision_ready_count"]
    background = totals["frontend_reference_evidence_count"]
    live = totals["frontend_live_success_count"]
    search_success = totals["frontend_search_success_count"]
    search_partial = totals["frontend_search_partial_count"]
    product_success = totals["frontend_product_page_success_count"] or live
    competitor_search = totals["frontend_competitor_search_success_count"] or search_success
    own_sellersprite = totals["frontend_own_sellersprite_count"]
    own_sellersprite_today = totals["frontend_own_sellersprite_today_count"]
    own_sellersprite_cache = totals["frontend_own_sellersprite_cache_count"]
    sellersprite_trend_ready = totals["frontend_sellersprite_trend_ready_count"]
    competitor_discovery = totals["frontend_competitor_discovery_count"]
    competitor_pool = totals["frontend_competitor_pool_count"]
    competitor_pool_today = totals["frontend_competitor_pool_today_count"]
    competitor_pool_cache = totals["frontend_competitor_pool_cache_count"]
    competitor_sellersprite = totals["frontend_competitor_sellersprite_count"]
    competitor_sellersprite_today = totals["frontend_competitor_sellersprite_today_count"]
    competitor_sellersprite_cache = totals["frontend_competitor_sellersprite_cache_count"]
    competitor_sellersprite_asins = totals["frontend_competitor_sellersprite_asin_count"]
    amazon_validation = totals["frontend_amazon_search_validation_count"] or competitor_search
    scalable = totals["frontend_scalable_strong_count"]
    weak = totals["frontend_weak_defensive_count"]
    insufficient = totals["frontend_insufficient_count"]
    market_complete = totals["market_survey_complete_count"]
    market_usable = totals["market_survey_usable_count"]
    market_insufficient = totals["market_survey_insufficient_count"]
    market_failed = totals["market_survey_failed_count"]
    market_average = round(market_survey_score_total / total, 1) if total else 0
    return {
        **totals,
        "frontend_product_page_success_count": product_success,
        "frontend_competitor_search_success_count": competitor_search,
        "frontend_amazon_search_validation_count": amazon_validation,
        "market_survey_average_score": market_average,
        "frontend_usable_evidence_rate": (usable / total) if total else 0,
        "frontend_decision_ready_rate": (strong / total) if total else 0,
        "frontend_reference_evidence_rate": (background / total) if total else 0,
        "frontend_live_success_rate": (live / total) if total else 0,
        "frontend_search_success_rate": (search_success / total) if total else 0,
        "frontend_search_observed_rate": ((search_success + search_partial) / total) if total else 0,
        "frontend_product_page_success_label": f"{product_success}/{total}" if total else "无前台队列",
        "frontend_competitor_search_success_label": f"{competitor_search}/{total}" if total else "无前台队列",
        "frontend_own_sellersprite_label": f"{own_sellersprite}/{total}" if total else "无前台队列",
        "frontend_own_sellersprite_today_label": (
            f"今日 {own_sellersprite_today}/{total}，缓存 {own_sellersprite_cache}/{total}"
            if total
            else "无前台队列"
        ),
        "frontend_sellersprite_trend_ready_label": f"{sellersprite_trend_ready}/{total}" if total else "无前台队列",
        "frontend_competitor_discovery_label": f"{competitor_discovery}/{total}" if total else "无前台队列",
        "frontend_competitor_pool_label": f"{competitor_pool}/{total}" if total else "无前台队列",
        "frontend_competitor_pool_freshness_label": (
            f"今日 {competitor_pool_today}/{total}，7天缓存 {competitor_pool_cache}/{total}"
            if total
            else "无前台队列"
        ),
        "frontend_competitor_sellersprite_label": (
            f"{competitor_sellersprite}/{total}，{competitor_sellersprite_asins} ASIN"
            if total
            else "无前台队列"
        ),
        "frontend_competitor_sellersprite_freshness_label": (
            f"今日 {competitor_sellersprite_today}/{total}，缓存 {competitor_sellersprite_cache}/{total}"
            if total
            else "无前台队列"
        ),
        "frontend_amazon_search_validation_label": f"{amazon_validation}/{total}" if total else "无前台队列",
        "frontend_scalable_strong_label": f"{scalable}/{total}" if total else "无前台队列",
        "frontend_weak_defensive_label": f"{weak}/{total}" if total else "无前台队列",
        "frontend_insufficient_label": f"{insufficient}/{total}" if total else "无前台队列",
        "frontend_decision_ready_label": f"{strong}/{total} 强证据，{strong / total:.0%}" if total else "无前台队列",
        "frontend_reference_evidence_label": f"{background}/{total} 背景参考，{background / total:.0%}" if total else "无前台队列",
        "frontend_coverage_label": f"{usable}/{total} 可用，{usable / total:.0%}" if total else "无前台队列",
        "market_survey_complete_label": f"{market_complete}/{total}" if total else "无市场调查队列",
        "market_survey_usable_label": f"{market_usable}/{total}" if total else "无市场调查队列",
        "market_survey_insufficient_label": f"{market_insufficient}/{total}" if total else "无市场调查队列",
        "market_survey_failed_label": f"{market_failed}/{total}" if total else "无市场调查队列",
        "market_survey_average_score_label": f"{market_average}/100" if total else "无市场调查队列",
    }


def archive_raw_file(source_path: Path, archive_dir: Path, prefix: str) -> Path:
    archive_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    archive_path = archive_dir / f"{prefix}_{timestamp}{source_path.suffix.lower()}"
    counter = 1
    while archive_path.exists():
        archive_path = archive_dir / f"{prefix}_{timestamp}_{counter}{source_path.suffix.lower()}"
        counter += 1
    shutil.copy2(source_path, archive_path)
    return archive_path


def _select_preferred_source(primary: Path, legacy: Path, label: str) -> Path:
    if primary.exists():
        return primary
    if legacy.exists():
        return legacy
    raise FileNotFoundError(f"缺少{label}文件，优先读取 {primary}，兼容回退 {legacy}")


def _select_preferred_source_multi(primary_options: list[Path], legacy_options: list[Path], label: str) -> Path:
    for path in primary_options:
        if path.exists():
            return path
    for path in legacy_options:
        if path.exists():
            return path
    primary_text = " / ".join(str(path) for path in primary_options)
    legacy_text = " / ".join(str(path) for path in legacy_options)
    raise FileNotFoundError(f"缺少{label}文件，优先读取 {primary_text}，兼容回退 {legacy_text}")


def prepare_ads_import_frame(ads_df: pd.DataFrame, source_path: Path, archive_path: Path) -> pd.DataFrame:
    frame = ads_df.copy()
    for column in [
        "campaign_id",
        "ad_group_id",
        "sku",
        "asin",
        "search_term",
        "targeting",
        "campaign_name",
        "ad_group_name",
        "match_type",
        "marketplace",
    ]:
        if column not in frame.columns:
            frame[column] = ""
        frame[column] = frame[column].fillna("").astype(str).str.strip()
    for column in [
        "click_orders",
        "click_sales",
        "click_roas",
        "click_cpa",
        "promoted_click_orders",
        "promoted_click_sales",
        "promoted_ad_orders",
        "promoted_ad_sales",
        "halo_click_orders",
        "halo_click_sales",
        "halo_click_units",
        "halo_ad_orders",
        "halo_ad_sales",
        "halo_ad_units",
    ]:
        if column not in frame.columns:
            frame[column] = 0
    frame["source_file"] = str(source_path)
    frame["archive_file"] = str(archive_path)
    frame["raw_modified_at"] = datetime.fromtimestamp(source_path.stat().st_mtime).isoformat(timespec="seconds")
    frame["imported_at"] = datetime.now().isoformat(timespec="seconds")
    return frame[
        [
            "marketplace",
            "date",
            "campaign_id",
            "ad_group_id",
            "sku",
            "asin",
            "search_term",
            "targeting",
            "campaign_name",
            "ad_group_name",
            "match_type",
            "impressions",
            "clicks",
            "spend",
            "ad_orders",
            "ad_sales",
            "click_orders",
            "click_sales",
            "click_roas",
            "click_cpa",
            "promoted_click_orders",
            "promoted_click_sales",
            "promoted_ad_orders",
            "promoted_ad_sales",
            "halo_click_orders",
            "halo_click_sales",
            "halo_click_units",
            "halo_ad_orders",
            "halo_ad_sales",
            "halo_ad_units",
            "source_file",
            "archive_file",
            "raw_modified_at",
            "imported_at",
        ]
    ].copy()


def prepare_erp_import_frame(erp_df: pd.DataFrame, source_path: Path, archive_path: Path) -> pd.DataFrame:
    frame = erp_df.copy()
    for column in ["marketplace", "sku", "asin", "product_name"]:
        if column not in frame.columns:
            frame[column] = ""
        frame[column] = frame[column].fillna("").astype(str).str.strip()
    for column in ["fba_stock", "fbm_stock", "available_stock"]:
        if column not in frame.columns:
            frame[column] = 0
    frame["source_file"] = str(source_path)
    frame["archive_file"] = str(archive_path)
    frame["raw_modified_at"] = datetime.fromtimestamp(source_path.stat().st_mtime).isoformat(timespec="seconds")
    frame["imported_at"] = datetime.now().isoformat(timespec="seconds")
    return frame[
        [
            "marketplace",
            "date",
            "sku",
            "asin",
            "product_name",
            "total_orders",
            "total_sales",
            "fba_stock",
            "fbm_stock",
            "available_stock",
            "source_file",
            "archive_file",
            "raw_modified_at",
            "imported_at",
        ]
    ].copy()


def _apply_stale_file_warning(ads_df: pd.DataFrame, erp_df: pd.DataFrame, ads_source: Path, erp_source: Path) -> None:
    today_local = datetime.now().date()
    warning = "当前文件可能不是今日最新下载，请确认是否已覆盖最新报表。"
    if datetime.fromtimestamp(ads_source.stat().st_mtime).date() != today_local:
        ads_df.attrs.setdefault("validation_messages", []).append(warning)
    if datetime.fromtimestamp(erp_source.stat().st_mtime).date() != today_local:
        erp_df.attrs.setdefault("validation_messages", []).append(warning)


def _make_source_files(
    ads_source: Path,
    erp_source: Path,
    ads_archive_path: Path,
    erp_archive_path: Path,
    target_marketplace: str,
    sku_map_path: Path,
) -> dict[str, str]:
    return {
        "ads_report": str(ads_source),
        "erp_report": str(erp_source),
        "ads_raw": str(ads_source),
        "erp_raw": str(erp_source),
        "ads_archive": str(ads_archive_path),
        "erp_archive": str(erp_archive_path),
        "sku_asin_map": str(sku_map_path),
        "product_config": str(PRODUCT_CONFIG_PATH),
        "product_cost_config": str(COST_CONFIG_PATH),
        "target_marketplace": target_marketplace,
    }


def _marketplace_scopes(ads_df: pd.DataFrame, erp_df: pd.DataFrame, marketplace: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    ads_scope = ads_df[ads_df["marketplace"].astype(str).str.upper() == marketplace.upper()].copy()
    erp_scope = erp_df[erp_df["marketplace"].astype(str).str.upper() == marketplace.upper()].copy()
    return ads_scope, erp_scope


def _scope_counts(ads_scope: pd.DataFrame, erp_scope: pd.DataFrame) -> tuple[int, int]:
    combined = pd.concat(
        [
            ads_scope[["sku", "asin"]] if not ads_scope.empty else pd.DataFrame(columns=["sku", "asin"]),
            erp_scope[["sku", "asin"]] if not erp_scope.empty else pd.DataFrame(columns=["sku", "asin"]),
        ],
        ignore_index=True,
    )
    if combined.empty:
        return 0, 0
    combined["sku"] = combined["sku"].fillna("").astype(str).str.strip()
    combined["asin"] = combined["asin"].fillna("").astype(str).str.strip()
    return int(combined.loc[combined["sku"] != "", "sku"].nunique()), int(combined.loc[combined["asin"] != "", "asin"].nunique())


def _collect_marketplace_quality(
    marketplace: str,
    ads_df: pd.DataFrame,
    erp_df: pd.DataFrame,
    issue_counts: dict[str, int] | None = None,
    dataset=None,
) -> dict[str, object]:
    ads_scope, erp_scope = _marketplace_scopes(ads_df, erp_df, marketplace)
    sku_count, asin_count = _scope_counts(ads_scope, erp_scope)
    issue_counts = issue_counts or {}
    common_date_range = "无数据"
    history_days = 0
    report_date = None
    erp_report_coverage_date_range = None
    erp_observed_sales_date_range = None
    coverage_warning = ""
    zero_fill_applied = False
    if dataset is not None:
        common_date_range = f"{dataset.common_date_range[0]} ~ {dataset.common_date_range[1]}"
        history_days = int(dataset.history_days)
        report_date = dataset.report_date.isoformat()
        erp_report_coverage_date_range = f"{dataset.erp_report_coverage_date_range[0]} ~ {dataset.erp_report_coverage_date_range[1]}"
        erp_observed_sales_date_range = (
            f"{dataset.erp_observed_sales_date_range[0]} ~ {dataset.erp_observed_sales_date_range[1]}"
            if dataset.erp_observed_sales_date_range[0] and dataset.erp_observed_sales_date_range[1]
            else "无销量行"
        )
        coverage_warning = dataset.coverage_warning
        zero_fill_applied = dataset.zero_fill_applied
    return {
        "marketplace": marketplace,
        "ads_row_count": int(len(ads_scope)),
        "erp_row_count": int(len(erp_scope)),
        "ads_date_range": f"{ads_scope['date'].min()} ~ {ads_scope['date'].max()}" if not ads_scope.empty else "无数据",
        "erp_date_range": erp_report_coverage_date_range or (f"{erp_scope['date'].min()} ~ {erp_scope['date'].max()}" if not erp_scope.empty else "无数据"),
        "erp_report_coverage_date_range": erp_report_coverage_date_range or (f"{erp_scope['date'].min()} ~ {erp_scope['date'].max()}" if not erp_scope.empty else "无数据"),
        "erp_observed_sales_date_range": erp_observed_sales_date_range or (f"{erp_scope['date'].min()} ~ {erp_scope['date'].max()}" if not erp_scope.empty else "无数据"),
        "zero_fill_applied": zero_fill_applied,
        "coverage_warning": coverage_warning,
        "sku_count": sku_count,
        "asin_count": asin_count,
        "missing_sku_asin_map": int(issue_counts.get("missing_sku_asin_map", 0) or 0),
        "missing_cost_config": int(issue_counts.get("missing_cost_config", 0) or 0),
        "missing_target_acos": int(issue_counts.get("missing_target_acos", 0) or 0),
        "missing_first_leg_cost": int(issue_counts.get("missing_first_leg_cost", 0) or 0),
        "report_date": report_date,
        "history_days": history_days,
        "common_date_range": common_date_range,
    }


def _empty_result(marketplace: str, ads_df: pd.DataFrame, erp_df: pd.DataFrame) -> dict[str, object]:
    ads_scope, erp_scope = _marketplace_scopes(ads_df, erp_df, marketplace)
    markdown = "\n".join(
        [
            f"# 浜氶┈閫婅繍钀ユ棩鎶ワ綔{marketplace}锝滄棤鏁版嵁",
            "",
            f"{marketplace} 当前无可用于完整运营分析的数据。",
            f"广告行数：{len(ads_scope)}",
            f"ERP 行数：{len(erp_scope)}",
        ]
    ) + "\n"
    return {
        "marketplace": marketplace,
        "has_data": False,
        "summary": _collect_marketplace_quality(marketplace, ads_df, erp_df),
        "analysis_payload": {"enhanced_data_requests": []},
        "report_view": None,
        "no_data_markdown": markdown,
    }


def _run_marketplace_report(
    target_marketplace: str,
    output_dir: Path,
    ads_df: pd.DataFrame,
    erp_df: pd.DataFrame,
    ads_source: Path,
    erp_source: Path,
    ads_archive_path: Path,
    erp_archive_path: Path,
    db: AnalyticsDatabase,
    alias_result,
    import_stats: dict[str, int],
    sku_map_df: pd.DataFrame,
    cost_config_df: pd.DataFrame,
    write_outputs: bool = True,
    sku_map_path: Path = SKU_MAP_PATH,
    alias_map_path: Path = SKU_ALIAS_MAP_PATH,
) -> dict[str, object]:
    timing: dict[str, float] = {}
    marketplace_ads, marketplace_erp = _marketplace_scopes(ads_df, erp_df, target_marketplace)
    if marketplace_ads.empty or erp_df.empty:
        result = _empty_result(target_marketplace, ads_df, erp_df)
        if write_outputs:
            recommendations_path = output_dir / "latest_recommendations.md"
            recommendations_html_path = output_dir / "latest_recommendations.html"
            write_text(recommendations_path, result["no_data_markdown"])
            write_html_report(result["no_data_markdown"], recommendations_html_path, title=f"浜氶┈閫婅繍钀ユ棩鎶ワ綔{target_marketplace}锝滄棤鏁版嵁")
            result["recommendations_path"] = recommendations_path
            result["recommendations_html_path"] = recommendations_html_path
        return result

    t_merge = time.perf_counter()
    dataset = build_daily_dataset(
        ads_df=ads_df,
        erp_df=erp_df,
        sku_map_path=sku_map_path,
        product_config_path=PRODUCT_CONFIG_PATH,
        cost_config_path=COST_CONFIG_PATH,
        alias_map_path=alias_map_path,
        ignored_issues_path=IGNORED_ISSUES_PATH,
        target_marketplace=target_marketplace,
    )
    timing["merge_data"] = time.perf_counter() - t_merge
    _log_step("5. 合并数据", timing["merge_data"])

    history = db.load_history(as_of_date=dataset.report_date, marketplace=target_marketplace)
    for history_key in ["product_daily", "campaign_daily", "search_term_daily"]:
        frame = history[history_key]
        if not frame.empty:
            history[history_key] = frame[
                (frame["date"] >= dataset.common_date_range[0]) & (frame["date"] <= dataset.common_date_range[1])
            ].copy()

    views = build_windowed_views(
        report_date=dataset.report_date,
        current_product_daily=dataset.product_daily,
        current_campaign_daily=dataset.campaign_daily,
        current_search_term_daily=dataset.search_term_daily,
        history=history,
    )

    source_files = _make_source_files(
        ads_source=ads_source,
        erp_source=erp_source,
        ads_archive_path=ads_archive_path,
        erp_archive_path=erp_archive_path,
        target_marketplace=target_marketplace,
        sku_map_path=sku_map_path,
    )
    analysis_payload = build_analysis_payload(
        report_date=dataset.report_date,
        source_files=source_files,
        dataset=dataset,
        views=views,
    )
    analysis_payload["import_summary"] = {
        "ads_imported_rows": int(import_stats.get("ads_imported_rows", 0)),
        "erp_imported_rows": int(import_stats.get("erp_imported_rows", 0)),
        "added_rows": int(import_stats.get("added_rows", 0)),
        "duplicate_skipped_rows": int(import_stats.get("duplicate_skipped_rows", 0)),
        "overwrite_updated_rows": int(import_stats.get("overwrite_updated_rows", 0)),
        "ads_added_rows": int(import_stats.get("ads_added_rows", 0)),
        "ads_duplicate_skipped_rows": int(import_stats.get("ads_duplicate_skipped_rows", 0)),
        "ads_overwrite_updated_rows": int(import_stats.get("ads_overwrite_updated_rows", 0)),
        "erp_added_rows": int(import_stats.get("erp_added_rows", 0)),
        "erp_duplicate_skipped_rows": int(import_stats.get("erp_duplicate_skipped_rows", 0)),
        "erp_overwrite_updated_rows": int(import_stats.get("erp_overwrite_updated_rows", 0)),
    }
    data_quality_detail, issue_counts = build_data_quality_issue_detail(
        dataset.mapping_check,
        cost_config=cost_config_df,
        sku_map=sku_map_df,
        alias_map=alias_result.alias_map,
    )
    analysis_payload["data_quality_issue_summary"] = issue_counts
    missing_sku_detail = build_missing_sku_detail(
        dataset.mapping_check,
        sku_map=sku_map_df,
        cost_config=cost_config_df,
    )
    analysis_payload["missing_sku_detail_count"] = len(missing_sku_detail)
    analysis_payload["sku_mapping_check"] = dataset.mapping_check.to_dict(orient="records")
    analysis_payload["alias_resolution_summary"] = {
        "original_missing_count": int(alias_result.original_missing_count),
        "remaining_missing_count": int(alias_result.remaining_missing_count),
        "auto_alias_count": int(len(alias_result.auto_aliases)),
        "review_count": int(len(alias_result.review)),
    }

    t_enhanced = time.perf_counter()
    enhanced_bundle = build_enhanced_bundle(
        marketplace=target_marketplace,
        custom_dir=CUSTOM_DATA_DIR,
        common_end=dataset.report_date,
        analysis_payload=analysis_payload,
    )
    timing["enhanced_bundle_total"] = time.perf_counter() - t_enhanced
    _log_step("3/4. 读取增强数据(总)", timing["enhanced_bundle_total"])
    enhanced_timing = (enhanced_bundle.status.get("timing") or {}) if isinstance(enhanced_bundle.status, dict) else {}
    if enhanced_timing:
        if "traffic_sales_read_sec" in enhanced_timing:
            _log_step("3. 读取 traffic_sales 增强表", float(enhanced_timing["traffic_sales_read_sec"]))
        if "search_query_read_sec" in enhanced_timing:
            _log_step("4. 读取 search_query_performance 增强表", float(enhanced_timing["search_query_read_sec"]))
    analysis_payload["enhanced_data_status"] = enhanced_bundle.status
    analysis_payload["custom_traffic_sales"] = enhanced_bundle.traffic_sales_detail.to_dict(orient="records")
    analysis_payload["custom_search_query_performance"] = enhanced_bundle.search_query_detail.to_dict(orient="records")
    analysis_payload["natural_decline_enhanced_diagnostics"] = enhanced_bundle.natural_decline_diagnostics.to_dict(orient="records")
    analysis_payload["search_query_opportunities"] = enhanced_bundle.search_query_opportunities.to_dict(orient="records")
    analysis_payload["enhanced_data_requests"] = enhanced_bundle.request_rows.to_dict(orient="records")
    _, erp_scope_for_status = _marketplace_scopes(ads_df, erp_df, target_marketplace)
    analysis_payload["erp_row_count"] = int(len(erp_scope_for_status))
    analysis_payload["无单原因诊断"] = build_no_order_diagnostics(analysis_payload)
    analysis_payload["inventory_replenishment"] = build_inventory_replenishment(
        marketplace=target_marketplace,
        sku_map=sku_map_df,
        cost_config=cost_config_df,
        views=views,
        product_daily=dataset.product_daily,
        report_date=dataset.report_date,
        output_dir=output_dir,
    )

    detected_files = enhanced_bundle.status.get("file_summary", []) or []
    print(f"[enhanced] Recognized files: {len(detected_files)}")
    for record in detected_files:
        print(
            "[enhanced] "
            f"{record.get('file_name')} | marketplace={record.get('marketplace')} | "
            f"data_type={record.get('data_type')} | period={record.get('period_hint')} | format={record.get('format_type')}"
        )
    print(f"[enhanced] traffic_sales ASIN rows: {len(enhanced_bundle.traffic_sales_detail)}")
    print(f"[enhanced] search_query_performance ASIN rows: {len(enhanced_bundle.search_query_detail)}")
    print(f"[enhanced] search_query_opportunities generated: {len(enhanced_bundle.search_query_opportunities)}")
    print(f"[enhanced] natural_decline_diagnostics generated: {len(enhanced_bundle.natural_decline_diagnostics)}")
    print(f"[enhanced] enhanced_data_requests deduped: {bool(enhanced_bundle.status.get('request_deduped'))}")

    validation_messages = analysis_payload.setdefault("data_quality", {}).setdefault("validation_messages", [])
    if not enhanced_bundle.status.get("provided"):
        validation_messages.append("未提供亚马逊定制分析增强数据，无法进一步拆分流量、转化或搜索查询原因。")
    elif detected_files:
        file_names = ", ".join(record.get("file_name") or "N/A" for record in detected_files)
        validation_messages.append(f"宸茶瘑鍒埌澧炲己鏁版嵁鏂囦欢 {len(detected_files)} 涓細{file_names}")
    if enhanced_bundle.status.get("legacy_root_warning"):
        validation_messages.append(str(enhanced_bundle.status.get("legacy_root_warning")))

    report_view = build_report_view(analysis_payload)
    result = {
        "marketplace": target_marketplace,
        "has_data": True,
        "analysis_payload": analysis_payload,
        "report_view": report_view,
        "summary": _collect_marketplace_quality(target_marketplace, ads_df, erp_df, issue_counts=issue_counts, dataset=dataset),
        "recommendations_md": build_recommendations_markdown(analysis_payload),
        "enhanced_requests_md": enhanced_bundle.request_markdown,
        "enhanced_requests_df": enhanced_bundle.request_rows,
        "data_quality_detail": data_quality_detail,
        "missing_sku_detail": missing_sku_detail,
        "mapping_check": dataset.mapping_check,
    }

    if not write_outputs:
        return result

    report_path = output_dir / f"amazon_ops_report_{dataset.report_date.isoformat()}.xlsx"
    analysis_path = output_dir / "latest_analysis.json"
    recommendations_path = output_dir / "latest_recommendations.md"
    recommendations_html_path = output_dir / "latest_recommendations.html"
    enhanced_requests_md_path = output_dir / "enhanced_data_requests.md"
    enhanced_requests_xlsx_path = output_dir / "enhanced_data_requests.xlsx"
    mapping_check_path = output_dir / "sku_mapping_check.xlsx"
    unmatched_review_path = output_dir / "unmatched_sku_review.xlsx"
    data_quality_detail_path = output_dir / "data_quality_issue_detail.xlsx"
    missing_sku_detail_path = output_dir / "missing_sku_detail.xlsx"

    write_json(analysis_path, analysis_payload)
    write_text(recommendations_path, result["recommendations_md"])
    write_text(enhanced_requests_md_path, enhanced_bundle.request_markdown)
    enhanced_bundle.request_rows.to_excel(enhanced_requests_xlsx_path, index=False)
    t_html = time.perf_counter()
    write_html_report(
        markdown_text=result["recommendations_md"],
        output_path=recommendations_html_path,
        title=f"亚马逊运营日报｜{target_marketplace}｜{dataset.report_date.isoformat()}",
    )
    timing["generate_html"] = time.perf_counter() - t_html
    _log_step("7. 生成 HTML", timing["generate_html"])
    t_excel = time.perf_counter()
    actual_report_path = generate_excel_report(
        output_path=report_path,
        report_date=dataset.report_date,
        analysis_payload=analysis_payload,
        views=views,
    )
    timing["generate_excel"] = time.perf_counter() - t_excel
    _log_step("6. 生成 Excel", timing["generate_excel"])
    dataset.mapping_check.to_excel(mapping_check_path, index=False)
    alias_result.review.to_excel(unmatched_review_path, index=False)
    write_data_quality_issue_detail(data_quality_detail_path, result["data_quality_detail"])
    write_missing_sku_detail(missing_sku_detail_path, missing_sku_detail)

    result.update(
        {
            "report_path": actual_report_path,
            "analysis_path": analysis_path,
            "recommendations_path": recommendations_path,
            "recommendations_html_path": recommendations_html_path,
            "enhanced_requests_md_path": enhanced_requests_md_path,
            "enhanced_requests_xlsx_path": enhanced_requests_xlsx_path,
        }
    )
    autoopt_paths = _write_autoopt_outputs([result], output_dir, fallback_report_date=dataset.report_date.isoformat())
    result.update(autoopt_paths)
    print(f"[{target_marketplace}] Report generated: {actual_report_path}")
    print(f"[{target_marketplace}] Analysis JSON: {analysis_path}")
    print(f"[{target_marketplace}] Recommendations MD: {recommendations_path}")
    print(f"[{target_marketplace}] Recommendations HTML: {recommendations_html_path}")
    print(f"[{target_marketplace}] Enhanced requests MD: {enhanced_requests_md_path}")
    print(f"[{target_marketplace}] Enhanced requests XLSX: {enhanced_requests_xlsx_path}")
    print(f"[{target_marketplace}] Autoopt JSON: {autoopt_paths['autoopt_json_path']}")
    print(f"[{target_marketplace}] Autoopt XLSX: {autoopt_paths['autoopt_xlsx_path']}")
    return result


def _write_all_mode_outputs(
    results: list[dict],
    output_dir: Path,
    source_files: dict[str, str],
    import_stats: dict[str, int],
) -> dict[str, Path]:
    report_date = next((result["summary"]["report_date"] for result in results if result["summary"].get("report_date")), None)
    report_date = report_date or datetime.now().date().isoformat()
    annotate_task_history(results, output_dir, report_date)
    autoopt_paths = _write_autoopt_outputs(results, output_dir, fallback_report_date=report_date)
    autoopt_payload = autoopt_paths.get("autoopt_payload")
    if isinstance(autoopt_payload, dict):
        _inject_current_autoopt_reviews(results, autoopt_payload)
    recommendations_md = build_all_marketplace_markdown(results, report_date)
    marketplace_summary_md = build_marketplace_summary_markdown(results, report_date)
    enhanced_requests_md = build_all_enhanced_requests_markdown(results, report_date)

    combined_requests = []
    combined_inventory_rows = []
    combined_product_final_decisions = []
    combined_product_operation_cards = []
    for result in results:
        combined_requests.extend(result.get("analysis_payload", {}).get("enhanced_data_requests", []))
        combined_inventory_rows.extend(result.get("analysis_payload", {}).get("inventory_replenishment", {}).get("rows", []))
        combined_product_final_decisions.extend(result.get("report_view", {}).get("product_final_decision_rows", []))
        combined_product_operation_cards.extend(result.get("report_view", {}).get("product_operation_cards", []))

    combined_payload = {
        "report_date": report_date,
        "target_marketplace": "ALL",
        "source_files": source_files,
        "import_summary": {
            "ads_imported_rows": int(import_stats.get("ads_imported_rows", 0)),
            "erp_imported_rows": int(import_stats.get("erp_imported_rows", 0)),
            "added_rows": int(import_stats.get("added_rows", 0)),
            "duplicate_skipped_rows": int(import_stats.get("duplicate_skipped_rows", 0)),
            "overwrite_updated_rows": int(import_stats.get("overwrite_updated_rows", 0)),
        },
        "marketplace_results": [
            {
                "marketplace": result["marketplace"],
                "has_data": result.get("has_data", False),
                "summary": result["summary"],
                "enhanced_data_status": result.get("analysis_payload", {}).get("enhanced_data_status", {}),
                "data_quality_issue_summary": result.get("analysis_payload", {}).get("data_quality_issue_summary", {}),
                "report_view_snapshot": _report_view_snapshot(result.get("report_view", {})),
            }
            for result in results
        ],
        "enhanced_data_requests": combined_requests,
        "inventory_replenishment_rows": combined_inventory_rows,
        "product_final_decision_rows": combined_product_final_decisions,
        "product_operation_cards": combined_product_operation_cards,
        "final_decision_summary": {
            str(result.get("marketplace") or "N/A"): (result.get("report_view", {}) or {}).get("final_decision_summary", {})
            for result in results
        },
        "decision_gate_counts": {
            str(result.get("marketplace") or "N/A"): (result.get("report_view", {}) or {}).get("decision_gate_counts", {})
            for result in results
        },
        "frontend_coverage_summary": _aggregate_frontend_coverage(results),
    }

    recommendations_path = output_dir / "latest_recommendations.md"
    recommendations_html_path = output_dir / "latest_recommendations.html"
    dashboard_path = output_dir / "dashboard.html"
    summary_html_path = output_dir / "summary.html"
    uk_report_path = output_dir / "uk_report.html"
    us_report_path = output_dir / "us_report.html"
    de_report_path = output_dir / "de_report.html"
    report_path = output_dir / f"amazon_ops_report_{report_date}.xlsx"
    analysis_path = output_dir / "latest_analysis.json"
    enhanced_requests_md_path = output_dir / "enhanced_data_requests.md"
    enhanced_requests_xlsx_path = output_dir / "enhanced_data_requests.xlsx"
    marketplace_summary_path = output_dir / "marketplace_summary.md"
    snapshot_results = [
        {
            **result,
            "report_view": combined_result.get("report_view_snapshot", {}),
        }
        for result, combined_result in zip(results, combined_payload["marketplace_results"])
    ]

    write_text(recommendations_path, recommendations_md)
    write_dashboard_html(snapshot_results, dashboard_path, report_date)
    write_summary_html(snapshot_results, summary_html_path, report_date)
    write_marketplace_report_html(next(result for result in results if result["marketplace"] == "UK"), uk_report_path, report_date)
    write_marketplace_report_html(next(result for result in results if result["marketplace"] == "US"), us_report_path, report_date)
    write_marketplace_report_html(next(result for result in results if result["marketplace"] == "DE"), de_report_path, report_date)
    write_recommendations_workbench_html(results, recommendations_html_path, report_date)
    write_text(enhanced_requests_md_path, enhanced_requests_md)
    pd.DataFrame(combined_requests).to_excel(enhanced_requests_xlsx_path, index=False)
    write_text(marketplace_summary_path, marketplace_summary_md)
    write_json(analysis_path, combined_payload)
    actual_report_path = generate_all_marketplace_excel_report(
        output_path=report_path,
        results=results,
        report_date=report_date,
        source_files=source_files,
        import_summary=combined_payload["import_summary"],
    )
    return {
        "recommendations_path": recommendations_path,
        "recommendations_html_path": recommendations_html_path,
        "dashboard_path": dashboard_path,
        "summary_html_path": summary_html_path,
        "uk_report_path": uk_report_path,
        "us_report_path": us_report_path,
        "de_report_path": de_report_path,
        "report_path": actual_report_path,
        "analysis_path": analysis_path,
        "enhanced_requests_md_path": enhanced_requests_md_path,
        "enhanced_requests_xlsx_path": enhanced_requests_xlsx_path,
        "marketplace_summary_path": marketplace_summary_path,
        **autoopt_paths,
    }


def _safe_output_dir(base_output_dir: Path) -> Path:
    run_id = (os.environ.get("AMAZON_OPS_SAFE_RUN_ID") or "").strip()
    timestamp = run_id or datetime.now().strftime("%Y%m%d_%H%M%S")
    if re.search(r"[^A-Za-z0-9_.-]", timestamp):
        raise FieldValidationError("AMAZON_OPS_SAFE_RUN_ID contains unsafe characters")
    safe_dir = base_output_dir / "safe_run" / timestamp
    safe_dir.mkdir(parents=True, exist_ok=True)
    return safe_dir


def _safe_path(path: Path, safe_output_dir: Path) -> Path:
    return safe_output_dir / path.name


def _formal_run_state_snapshot() -> dict[str, object]:
    from scripts.run_daily_update import report_state_snapshot

    return report_state_snapshot()


def _restore_formal_run_state_snapshot(snapshot: dict[str, object]) -> list[str]:
    from scripts.run_daily_update import restore_report_state_snapshot

    return restore_report_state_snapshot(snapshot)


def main() -> int:
    formal_state_snapshot: dict[str, object] | None = None
    if "--safe-run" not in sys.argv[1:]:
        formal_state_snapshot = _formal_run_state_snapshot()
    try:
        code = _main_impl()
    except Exception:
        if formal_state_snapshot is not None:
            failures = _restore_formal_run_state_snapshot(formal_state_snapshot)
            for failure in failures:
                print(f"[fail] formal state restore blocker: {failure}", flush=True)
        raise
    if formal_state_snapshot is not None and code != 0:
        failures = _restore_formal_run_state_snapshot(formal_state_snapshot)
        for failure in failures:
            print(f"[fail] formal state restore blocker: {failure}", flush=True)
    return code


def _main_impl() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--marketplace", dest="marketplace", default="UK")
    parser.add_argument("--debug-timing", action="store_true")
    parser.add_argument("--safe-run", action="store_true")
    args = parser.parse_args()
    target_marketplace = str(args.marketplace or "UK").upper()
    if target_marketplace not in {*SUPPORTED_MARKETPLACES, "ALL"}:
        raise FieldValidationError(f"涓嶆敮鎸佺殑 marketplace: {target_marketplace}")

    ensure_directories()
    ensure_ignored_issues_config()

    active_output_dir = _safe_output_dir(OUTPUT_DIR) if args.safe_run else OUTPUT_DIR
    active_sku_map_path = (active_output_dir / "sku_asin_map.xlsx") if args.safe_run else SKU_MAP_PATH
    active_alias_map_path = (active_output_dir / "sku_alias_map.xlsx") if args.safe_run else SKU_ALIAS_MAP_PATH

    build_sku_asin_map_from_cost_config(
        cost_config_path=COST_CONFIG_PATH,
        output_path=active_sku_map_path,
        unmapped_output_path=active_output_dir / "unmapped_sku_asin.xlsx",
        allow_overwrite=not args.safe_run,
    )

    ads_source = _select_preferred_source_multi(
        [ALL_ADS_REPORT_PATH, ALL_ADS_REPORT_XLSX_PATH],
        [LEGACY_ADS_REPORT_PATH, LEGACY_ADS_REPORT_XLSX_PATH],
        "广告报表",
    )
    erp_source = _select_preferred_source_multi(
        [ALL_ERP_REPORT_PATH],
        [LEGACY_ERP_REPORT_PATH],
        "ERP销量表",
    )
    t = _start("读取广告表")
    ads_df, ads_source = load_ads_report(ads_source)
    elapsed = _done("读取广告表", t)
    if args.debug_timing and elapsed > 60:
        _timeout_step("读取广告表")
        return DEBUG_TIMING_EXIT_CODE

    t = _start("读取 ERP 表")
    erp_df, erp_source = load_erp_report(erp_source)
    elapsed = _done("读取 ERP 表", t)
    if args.debug_timing and elapsed > 60:
        _timeout_step("读取 ERP 表")
        return DEBUG_TIMING_EXIT_CODE
    _apply_stale_file_warning(ads_df, erp_df, ads_source, erp_source)

    if args.debug_timing:
        _start("读取 traffic_sales 增强表")
        traffic_elapsed, query_elapsed = _debug_read_enhanced_headers(CUSTOM_DATA_DIR, target_marketplace)
        print(f"[DONE] 读取 traffic_sales 增强表: {traffic_elapsed:.3f}秒", flush=True)
        if traffic_elapsed > 60:
            _timeout_step("读取 traffic_sales 增强表")
            return DEBUG_TIMING_EXIT_CODE

        print("[START] 读取 search_query_performance 增强表", flush=True)
        print(f"[DONE] 读取 search_query_performance 增强表: {query_elapsed:.3f}秒", flush=True)
        if query_elapsed > 60:
            _timeout_step("读取 search_query_performance 增强表")
            return DEBUG_TIMING_EXIT_CODE

    ads_archive_prefix = (
        "ads_report_all"
        if ads_source.name in {ALL_ADS_REPORT_PATH.name, ALL_ADS_REPORT_XLSX_PATH.name}
        else "ads_report_uk"
    )
    erp_archive_prefix = "sales_report_all" if erp_source.name == ALL_ERP_REPORT_PATH.name else "sales_report_uk"
    ads_archive_path = ads_source if args.safe_run else archive_raw_file(ads_source, ARCHIVE_ADS_DIR, ads_archive_prefix)
    erp_archive_path = erp_source if args.safe_run else archive_raw_file(erp_source, ARCHIVE_ERP_DIR, erp_archive_prefix)

    t_merge_initial = _start("合并数据")
    initial_dataset = build_daily_dataset(
        ads_df=ads_df,
        erp_df=erp_df,
        sku_map_path=active_sku_map_path,
        product_config_path=PRODUCT_CONFIG_PATH,
        cost_config_path=COST_CONFIG_PATH,
        alias_map_path=active_alias_map_path,
        ignored_issues_path=IGNORED_ISSUES_PATH,
        target_marketplace=None,
    )
    merge_elapsed = _done("合并数据", t_merge_initial)
    if args.debug_timing and merge_elapsed > 60:
        _timeout_step("合并数据")
        return DEBUG_TIMING_EXIT_CODE

    if args.debug_timing:
        print("[START] 生成 Excel", flush=True)
        print("[DONE] 生成 Excel: 0.000秒", flush=True)
        print("[START] 生成 HTML", flush=True)
        print("[DONE] 生成 HTML: 0.000秒", flush=True)
        print("[DEBUG_TIMING] 仅定位耗时，已跳过完整报告生成。", flush=True)
        return DEBUG_TIMING_EXIT_CODE
    sku_map_df = pd.read_excel(active_sku_map_path)
    cost_config_df = pd.read_excel(COST_CONFIG_PATH, sheet_name="product_cost_config")
    alias_result = build_alias_review(
        mapping_check=initial_dataset.mapping_check,
        sku_map=sku_map_df,
        cost_config=cost_config_df,
        alias_map_path=active_alias_map_path,
        review_output_path=active_output_dir / "unmatched_sku_review.xlsx",
    )

    db_path = (active_output_dir / "amazon_ops_safe_run.db") if args.safe_run else (ROOT / "database" / "amazon_ops.db")
    db = AnalyticsDatabase(db_path)
    if args.safe_run:
        import_stats = {
            "ads_imported_rows": int(len(ads_df)),
            "erp_imported_rows": int(len(erp_df)),
            "added_rows": 0,
            "duplicate_skipped_rows": 0,
            "overwrite_updated_rows": 0,
            "ads_added_rows": 0,
            "ads_duplicate_skipped_rows": 0,
            "ads_overwrite_updated_rows": 0,
            "erp_added_rows": 0,
            "erp_duplicate_skipped_rows": 0,
            "erp_overwrite_updated_rows": 0,
        }
    else:
        import_stats = db.import_raw_and_daily_frames(
            ads_raw=prepare_ads_import_frame(ads_df, ads_source, ads_archive_path),
            erp_raw=prepare_erp_import_frame(erp_df, erp_source, erp_archive_path),
            product_daily=initial_dataset.product_daily,
            campaign_daily=initial_dataset.campaign_daily,
            search_term_daily=initial_dataset.search_term_daily,
        )

    source_files = _make_source_files(
        ads_source=ads_source,
        erp_source=erp_source,
        ads_archive_path=ads_archive_path,
        erp_archive_path=erp_archive_path,
        target_marketplace=target_marketplace,
        sku_map_path=active_sku_map_path,
    )

    if target_marketplace == "ALL":
        results = [
            _run_marketplace_report(
                target_marketplace=marketplace,
                output_dir=active_output_dir,
                ads_df=ads_df,
                erp_df=erp_df,
                ads_source=ads_source,
                erp_source=erp_source,
                ads_archive_path=ads_archive_path,
                erp_archive_path=erp_archive_path,
                db=db,
                alias_result=alias_result,
                import_stats=import_stats,
                sku_map_df=sku_map_df,
                cost_config_df=cost_config_df,
                write_outputs=False,
                sku_map_path=active_sku_map_path,
                alias_map_path=active_alias_map_path,
            )
            for marketplace in SUPPORTED_MARKETPLACES
        ]
        written_paths = _write_all_mode_outputs(results, active_output_dir, source_files, import_stats)
        print(f"[ALL] Recommendations MD: {written_paths['recommendations_path']}")
        print(f"[ALL] Dashboard HTML: {written_paths['dashboard_path']}")
        print(f"[ALL] Summary HTML: {written_paths['summary_html_path']}")
        print(f"[ALL] UK report HTML: {written_paths['uk_report_path']}")
        print(f"[ALL] US report HTML: {written_paths['us_report_path']}")
        print(f"[ALL] DE report HTML: {written_paths['de_report_path']}")
        print(f"[ALL] Excel report: {written_paths['report_path']}")
        print(f"[ALL] Marketplace summary: {written_paths['marketplace_summary_path']}")
        print(f"[ALL] Enhanced requests MD: {written_paths['enhanced_requests_md_path']}")
        print(f"[ALL] Enhanced requests XLSX: {written_paths['enhanced_requests_xlsx_path']}")
        print(f"[ALL] Analysis JSON: {written_paths['analysis_path']}")
        print(f"[ALL] Autoopt JSON: {written_paths['autoopt_json_path']}")
        print(f"[ALL] Autoopt XLSX: {written_paths['autoopt_xlsx_path']}")
        _print_report_entrypoints(written_paths["report_path"], output_dir=active_output_dir)
        return 0

    result = _run_marketplace_report(
        target_marketplace=target_marketplace,
        output_dir=active_output_dir,
        ads_df=ads_df,
        erp_df=erp_df,
        ads_source=ads_source,
        erp_source=erp_source,
        ads_archive_path=ads_archive_path,
        erp_archive_path=erp_archive_path,
        db=db,
        alias_result=alias_result,
        import_stats=import_stats,
        sku_map_df=sku_map_df,
        cost_config_df=cost_config_df,
        write_outputs=True,
        sku_map_path=active_sku_map_path,
        alias_map_path=active_alias_map_path,
    )
    if not result.get("has_data"):
        print(f"[{target_marketplace}] 无数据，已生成提示报告。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

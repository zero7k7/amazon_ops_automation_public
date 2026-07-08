from __future__ import annotations

import html
import re
from pathlib import Path
from typing import Any, Callable, TypeVar

from openpyxl import load_workbook

from ..analyze_rules import money_symbol_for_marketplace

ROOT_DIR = Path(__file__).resolve().parents[2]
_COST_DETAIL_LOOKUP: dict[tuple[str, str, str], dict[str, object]] | None = None
_SHARED: Any | None = None
_T = TypeVar("_T")


def with_shared(shared: Any, func: Callable[..., _T], *args: Any, **kwargs: Any) -> _T:
    global _SHARED
    previous = _SHARED
    _SHARED = shared
    try:
        return func(*args, **kwargs)
    finally:
        _SHARED = previous


def _shared() -> Any:
    if _SHARED is None:
        raise RuntimeError("components_cards shared context is not bound")
    return _SHARED


def _inline_markup(text: str) -> str:
    return _shared()._inline_markup(text)


def _product_meta_html(*args: Any, **kwargs: Any) -> str:
    return _shared()._product_meta_html(*args, **kwargs)


def _tag_class(action: str) -> str:
    return _shared()._tag_class(action)


def _confirmed_tag(status: object) -> str:
    return _shared()._confirmed_tag(status)


def _lookup_frontend_evidence(*args: Any, **kwargs: Any) -> dict[str, str] | None:
    return _shared()._lookup_frontend_evidence(*args, **kwargs)


def _render_frontend_evidence_block(*args: Any, **kwargs: Any) -> str:
    return _shared()._render_frontend_evidence_block(*args, **kwargs)


def _render_frontend_gap_block(*args: Any, **kwargs: Any) -> str:
    return _shared()._render_frontend_gap_block(*args, **kwargs)


def _value_list(value: object) -> list[str]:
    return _shared()._value_list(value)


def _boolish(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "是", "已验证"}


_MARKET_SURVEY_LEVEL_LABELS = {
    "complete": "完整",
    "usable": "可用",
    "insufficient": "待补",
    "failed": "失败",
}

_SELLERSPRITE_SOURCE_LABELS = {
    "amazon_search_seed": "搜索页种子",
    "amazon_search_visible": "搜索页可见",
    "sellersprite_keyword_overlap": "共同词",
    "sellersprite_reverse_seed": "反查种子",
    "sellersprite_direct": "直接竞品",
    "sellersprite_competitor_direct": "直接竞品",
    "sellersprite_reversing_sources": "反查来源",
}

_CONFIDENCE_LABELS = {
    "high": "置信高",
    "medium": "置信中",
    "low": "置信低",
    "unknown": "",
}


def _market_survey_level_label(value: object) -> str:
    text = str(value or "").strip()
    return _MARKET_SURVEY_LEVEL_LABELS.get(text.lower(), text)


def _compact_unique_labels(values: list[object], labeler: Callable[[object], str]) -> str:
    labels: list[str] = []
    for value in values:
        label = labeler(value)
        if label and label not in labels:
            labels.append(label)
    return " / ".join(labels)


def _seller_source_label(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    separators = ["、", ",", "，"]
    parts = [text]
    for separator in separators:
        if separator in text:
            parts = [part.strip() for part in re.split(r"[、,，]", text) if part.strip()]
            break
    labels = [_SELLERSPRITE_SOURCE_LABELS.get(part, part) for part in parts]
    return "、".join(label for label in labels if label)


def _confidence_label(value: object) -> str:
    text = str(value or "").strip()
    return _CONFIDENCE_LABELS.get(text.lower(), text)


def _first_present(row: dict[str, object], *fields: str) -> object:
    return _shared()._first_present(row, *fields)


def _product_decision_contract_attrs(row: dict[str, object]) -> str:
    def list_attr(values: object) -> str:
        if isinstance(values, (list, tuple, set)):
            return "|".join(str(item).strip() for item in values if str(item).strip())
        return str(values or "").strip()

    def bool_attr(value: object) -> str:
        if isinstance(value, bool):
            return "true" if value else "false"
        text = str(value or "").strip().lower()
        return "true" if text in {"1", "true", "yes", "y", "是", "已验证"} else "false"

    attrs = {
        "data-product-decision-final": row.get("final_decision") or "",
        "data-product-decision-label": row.get("final_decision_label") or "",
        "data-product-allowed-actions": list_attr(row.get("today_allowed_actions")),
        "data-product-blocked-actions": list_attr(row.get("today_blocked_actions")),
        "data-product-frontend-tier": row.get("frontend_evidence_tier") or "",
        "data-product-frontend-display-tier": row.get("frontend_evidence_display_tier")
        or row.get("frontend_evidence_tier")
        or "",
        "data-product-frontend-decision-tier": row.get("frontend_decision_evidence_tier")
        or row.get("frontend_evidence_display_tier")
        or row.get("frontend_evidence_tier")
        or "",
        "data-product-frontend-strong": bool_attr(row.get("frontend_evidence_is_strong")),
        "data-product-market-survey-score": row.get("market_survey_completeness_score") or "",
        "data-product-market-survey-level": row.get("market_survey_completeness_level") or "",
        "data-product-market-survey-tier": row.get("market_survey_decision_evidence_tier") or "",
    }
    return "".join(
        f' {name}="{html.escape(str(value), quote=True)}"'
        for name, value in attrs.items()
        if str(value).strip()
    )


def _num_from_text(value: object) -> float:
    return _shared()._num_from_text(value)


def _render_collapsed_block(*args: Any, **kwargs: Any) -> str:
    return _shared()._render_collapsed_block(*args, **kwargs)

def _render_decision_summary_block(row: dict[str, object]) -> str:
    label, _label_title = _operation_decision_display(row)
    confidence = str(row.get("fusion_confidence") or row.get("confidence") or "N/A")
    fusion_type = str(row.get("fusion_issue_type") or "自动证据不足")
    return (
        '<div class="card-block decision-summary">'
        '<strong>系统结论</strong>'
        '<div class="decision-summary-list">'
        f'<div><span>系统结论</span><strong>{html.escape(label)}</strong></div>'
        f'<div><span>融合诊断</span><strong>{html.escape(fusion_type)}</strong></div>'
        f'<div><span>证据质量</span><strong>{html.escape(confidence)}</strong></div>'
        '</div>'
        '</div>'
    )


def _render_task_brief_blocks(row: dict[str, object], fallback_action: str) -> str:
    today = str(row.get("fusion_today_action") or fallback_action or "N/A")
    review_label, review = _display_review_instruction(
        row.get("fusion_review_window") or row.get("final_decision_next_review") or row.get("tomorrow_check")
    )
    return (
        '<div class="action-brief-grid">'
        f'<div class="card-block"><strong>今天动作</strong><p>{_inline_markup(today)}</p></div>'
        f'<div class="card-block"><strong>{html.escape(review_label)}</strong><p>{_inline_markup(review)}</p></div>'
        '</div>'
    )


def _display_review_instruction(value: object) -> tuple[str, str]:
    text = str(value or "").strip()
    if "补齐 7 天窗口" in text and "刷新前台" in text:
        return ("补证后复查", "满 7 天或前台证据可用后再定动作。")
    return ("复查时间", text or "3-7 天复查广告点击、花费、订单和转化。")


def _render_task_evidence_details(
    row: dict[str, str],
    *,
    issue_tags: str,
    action_text: str,
    frontend_lookup: dict[tuple[str, str, str], dict[str, str]] | None,
    frontend: dict[str, object] | None,
) -> str:
    blocked = str(row.get("fusion_do_not_do") or row.get("today_blocked_actions") or "无")
    full_reason = str(row.get("fusion_reason") or row.get("final_decision_reason") or row.get("primary_reason") or "N/A")
    flags = _value_list(row.get("fusion_evidence_flags"))
    flag_html = (
        '<div class="metric-row">' + "".join(f'<span class="metric-badge">{html.escape(flag)}</span>' for flag in flags[:6]) + '</div>'
        if flags
        else '<p class="subtle">暂无系统标记。</p>'
    )
    detail_blocks = [
        f'<div class="card-block"><strong>先别做</strong><p>{_inline_markup(blocked)}</p></div>',
        f'<div class="card-block"><strong>完整诊断原因</strong><p>{_inline_markup(full_reason)}</p></div>',
        f'<div class="card-block"><strong>系统标记</strong>{flag_html}</div>',
        f'<div class="card-block"><strong>诊断来源</strong><div>{issue_tags}</div><div class="subtle">{_inline_markup(str(row.get("primary_reason") or "N/A"))}</div></div>',
        f'<div class="card-block"><strong>关键证据</strong><div class="metric-row">{_metric_badges_from_evidence(str(row.get("key_evidence") or "N/A"))}</div></div>',
        _render_frontend_evidence_block(row, frontend_lookup),
        _render_frontend_gap_block(frontend),
        _render_task_review_status(row),
        '<div class="card-block"><strong>需要处理的搜索词 / ASIN</strong>'
        + _render_search_term_items(str(row.get("search_term_top5") or "N/A"))
        + "</div>",
    ]
    detail_html = "".join(block for block in detail_blocks if block)
    return (
        '<details class="evidence-details">'
        '<summary>证据明细</summary>'
        f'<div class="evidence-detail-grid">{detail_html}</div>'
        '</details>'
    )


def _metric_badges_from_evidence(text: str, limit: int = 4) -> str:
    if not text or text == "N/A":
        return '<span class="metric-badge">暂无关键指标</span>'
    parts = [part.strip() for part in str(text).split("；") if part.strip()]
    if not parts:
        return '<span class="metric-badge">暂无关键指标</span>'
    return "".join(f'<span class="metric-badge">{_inline_markup(part)}</span>' for part in parts[:limit])


def _load_cost_detail_lookup() -> dict[tuple[str, str, str], dict[str, object]]:
    global _COST_DETAIL_LOOKUP
    if _COST_DETAIL_LOOKUP is not None:
        return _COST_DETAIL_LOOKUP
    lookup: dict[tuple[str, str, str], dict[str, object]] = {}
    path = ROOT_DIR / "config" / "product_cost_config.xlsx"
    if not path.exists():
        _COST_DETAIL_LOOKUP = lookup
        return lookup
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook["product_cost_config"] if "product_cost_config" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        wanted = {
            "selling_price",
            "purchase_cost_local",
            "first_leg_cost_local",
            "fba_fee",
            "referral_fee",
            "vat",
            "digital_tax",
            "break_even_acos",
            "cost_status",
            "profit_before_ads",
            "suggested_target_acos",
            "currency",
            "marketplace",
            "sku",
            "asin",
        }
        indexes = {name: headers.index(name) for name in headers if name in wanted}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            sku = str(row[indexes["sku"]] if "sku" in indexes else "" or "").strip()
            asin = str(row[indexes["asin"]] if "asin" in indexes else "" or "").strip()
            marketplace = str(row[indexes["marketplace"]] if "marketplace" in indexes else "" or "").strip().upper()
            if not sku and not asin:
                continue
            record = {name: row[index] for name, index in indexes.items()}
            lookup[(marketplace, sku, asin)] = record
            if sku:
                lookup[(marketplace, sku, "")] = record
            if asin:
                lookup[(marketplace, "", asin)] = record
    except Exception:
        lookup = {}
    _COST_DETAIL_LOOKUP = lookup
    return lookup


def _cost_detail_for_row(row: dict[str, str]) -> dict[str, object]:
    lookup = _load_cost_detail_lookup()
    sku = str(row.get("sku") or row.get("SKU") or "").strip()
    asin = str(row.get("asin") or row.get("ASIN") or "").strip()
    marketplace = str(row.get("marketplace") or row.get("站点") or "").strip().upper()
    return (
        lookup.get((marketplace, sku, asin))
        or lookup.get((marketplace, sku, ""))
        or lookup.get((marketplace, "", asin))
        or lookup.get(("", sku, asin))
        or lookup.get(("", sku, ""))
        or lookup.get(("", "", asin))
        or {}
    )


def _split_search_term_item(item: str) -> dict[str, str]:
    text = str(item).strip()
    action_line = ""
    target_line = text
    if "\n" in text:
        first_line, rest = text.split("\n", 1)
        action_line = first_line.strip()
        target_line = rest.strip()
    cells = [cell.strip() for cell in target_line.split("｜")]
    return {
        "action_line": action_line or (cells[4] if len(cells) > 4 else "建议观察"),
        "target": cells[0] if len(cells) > 0 else "N/A",
        "clicks": cells[1] if len(cells) > 1 else "点击 N/A",
        "spend": cells[2] if len(cells) > 2 else "花费 N/A",
        "orders": cells[3] if len(cells) > 3 else "订单 N/A",
        "action": cells[4] if len(cells) > 4 else "观察",
        "reason": cells[5] if len(cells) > 5 else "N/A",
    }


def _render_search_term_items(text: str) -> str:
    if not text or text == "N/A":
        return '<p class="subtle">暂无需要立即处理的搜索词；低点击观察词见 Excel。</p>'
    raw_items = [_split_search_term_item(item) for item in str(text).split("；") if item.strip()]
    action_order = {"否定精准": 0, "降竞价10%-20%": 1, "暂停ASIN定向": 2}
    immediate = [item for item in raw_items if item.get("action") in action_order]
    observe_count = len(raw_items) - len(immediate)
    immediate = sorted(immediate, key=lambda item: action_order.get(item.get("action", ""), 9))[:3]
    if not immediate:
        suffix = f"；另有 {observe_count} 个观察词，见 Excel 低优先级明细" if observe_count else ""
        return f'<p class="subtle">暂无需要立即处理的搜索词{suffix}。</p>'
    parts = ['<div class="search-term-list">']
    for index, parsed in enumerate(immediate, start=1):
        parts.append(
            "\n".join(
                [
                    '<div class="search-term-item">',
                    f'<div class="search-term-word">{index}. {_inline_markup(parsed["target"])}</div>',
                    "</div>",
                ]
            )
        )
    if observe_count:
        parts.append(f'<p class="subtle">另有 {observe_count} 个观察词，见 Excel 低优先级明细。</p>')
    parts.append("</div>")
    return "".join(parts)


def _render_task_cards(
    rows: list[dict[str, str]],
    priority: str | None = None,
    limit: int | None = None,
    empty_message: str | None = "当前没有可展示内容。",
    frontend_lookup: dict[tuple[str, str, str], dict[str, str]] | None = None,
) -> str:
    selected = [row for row in rows if priority is None or row.get("priority") == priority]
    if limit is not None:
        selected = selected[:limit]
    if not selected:
        return f'<p class="subtle">{html.escape(empty_message)}</p>' if empty_message else ""
    parts: list[str] = ['<div class="action-grid">']
    for row in selected:
        row_priority = str(row.get("priority") or priority or "P1")
        css = "p0" if row_priority == "P0" else "p1"
        tags = "".join(f'<span class="tag">{html.escape(part)}</span>' for part in str(row.get("issue_type") or "").split("；") if part)
        confirmed = str(row.get("confirmed_status") or "待确认")
        action_text = str(row.get("today_action") or "N/A")
        if action_text.startswith("暂时不加广告预算"):
            action_text = "见统一临时处理原则；本产品按 Listing 待人工确认补充材料后再决策。"
        elif "近7天转化断崖" in action_text:
            action_text = "近7天转化断崖：先控扩量，按统一原则处理核心词和无效流量；同时复查推荐报价、价格、配送与转化承接。"
        frontend = _lookup_frontend_evidence(row, frontend_lookup)
        parts.append(
            "\n".join(
                [
                    f'<article class="work-card {css}">',
                    '<div class="card-head">',
                    "<div>",
                    f'<h3 class="card-title">{html.escape(str(row.get("product_name") or "N/A"))}</h3>',
                    f'<div class="card-meta">{_product_meta_html(row)}</div>',
                    "</div>",
                    '<div class="card-status-line">'
                    f'<span class="{_tag_class("否定精准") if row_priority == "P0" else _tag_class("降竞价10%-20%")}">{html.escape(row_priority)}</span>'
                    f'<span class="{_confirmed_tag(confirmed)}">{html.escape(confirmed)}</span>'
                    '</div>',
                    "</div>",
                    _render_decision_summary_block(row),
                    _render_task_brief_blocks(row, action_text),
                    _render_task_evidence_details(
                        row,
                        issue_tags=tags,
                        action_text=action_text,
                        frontend_lookup=frontend_lookup,
                        frontend=frontend,
                    ),
                    "</article>",
                ]
            )
        )
    parts.append("</div>")
    return "".join(parts)


def _render_task_review_status(row: dict[str, str]) -> str:
    status = str(row.get("review_status") or "").strip()
    days = str(row.get("consecutive_trigger_days") or "").strip()
    why = str(row.get("why_still_active") or "").strip()
    condition = str(row.get("downgrade_condition") or "").strip()
    if not any([status, days, why, condition]):
        return ""
    chips: list[str] = []
    if status:
        chips.append(f'<span class="metric-badge">复查状态：{html.escape(status)}</span>')
    if days:
        chips.append(f'<span class="metric-badge">连续触发：{html.escape(days)} 天</span>')
    details: list[str] = []
    if why:
        details.append(f'<div><strong>为什么还在：</strong>{_inline_markup(why)}</div>')
    if condition:
        details.append(f'<div><strong>降级条件：</strong>{_inline_markup(condition)}</div>')
    return (
        '<div class="card-block task-review-status"><strong>P0/P1 复查状态</strong>'
        f'<div class="metric-row">{"".join(chips)}</div>'
        f'{"".join(details)}'
        "</div>"
    )


def _task_sku_asin_key(row: dict[str, str]) -> tuple[str, str]:
    return (str(row.get("sku") or ""), str(row.get("asin") or ""))


def _exclude_listing_tasks(rows: list[dict[str, str]], p0_rows: list[dict[str, str]] | None = None) -> list[dict[str, str]]:
    p0_keys = {_task_sku_asin_key(row) for row in (p0_rows or []) if row.get("priority") == "P0"}
    filtered: list[dict[str, str]] = []
    for row in rows:
        if "Listing 待人工确认" in str(row.get("issue_type") or ""):
            continue
        if _task_sku_asin_key(row) in p0_keys:
            continue
        filtered.append(row)
    return filtered


def _exclude_executed_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    return [row for row in rows if str(row.get("confirmed_status") or "") != "已执行"]


def _executed_risk_rows(rows: list[dict[str, str]], limit: int = 6) -> list[dict[str, str]]:
    selected = [
        row
        for row in rows
        if str(row.get("confirmed_status") or "") == "已执行"
        and str(row.get("priority") or "") in {"P0", "P1"}
        and row.get("action_group") != "成本 / 利润动作"
    ]
    priority_rank = {"P0": 0, "P1": 1}
    return sorted(
        selected,
        key=lambda row: (
            priority_rank.get(str(row.get("priority") or ""), 9),
            str(row.get("marketplace") or ""),
            str(row.get("product_name") or row.get("asin") or ""),
        ),
    )[:limit]


def _filter_listing_rows_for_p0(listing_rows: list[dict[str, str]], task_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    p0_keys = {
        (str(row.get("marketplace") or ""), str(row.get("sku") or ""), str(row.get("asin") or ""))
        for row in task_rows
        if row.get("priority") == "P0"
    }
    filtered: list[dict[str, str]] = []
    for row in listing_rows:
        product = str(row.get("产品") or "")
        marketplace = product.split("｜", 1)[0] if "｜" in product else str(row.get("marketplace") or row.get("站点") or "")
        key = (marketplace, str(row.get("SKU") or ""), str(row.get("ASIN") or ""))
        if key in p0_keys:
            continue
        filtered.append(row)
    return filtered


def _first_semicolon_item(text: object, fallback: str = "暂无") -> str:
    items = [item.strip() for item in str(text or "").split("；") if item.strip() and item.strip() != "N/A"]
    return items[0] if items else fallback


def _compact_semicolon_items(text: object, fallback: str = "暂无", limit: int = 2) -> str:
    items = [item.strip() for item in str(text or "").split("；") if item.strip() and item.strip() != "N/A"]
    return "；".join(items[:limit]) if items else fallback


def _compact_keyword_items(text: object, fallback: str = "", limit: int = 1) -> str:
    items = [
        item.strip()
        for item in re.split(r"[；、,]", str(text or ""))
        if item.strip() and item.strip() != "N/A"
    ]
    return "、".join(items[:limit]) if items else fallback


def _listing_basis_label(row: dict[str, str]) -> str:
    if str(row.get("历史人工确认") or "").strip():
        return "历史已确认原因"
    text = "；".join(
        str(row.get(key) or "")
        for key in ["异常信号", "初步方向", "当前证据", "主因", "产品专属下一步"]
    )
    if "加购" in text and ("购买 0" in text or "不购买" in text):
        return "加购后未购买"
    if "近7天" in text and any(marker in text for marker in ["效率变差", "转化恶化", "广告无单", "ACOS 升高"]):
        return "近7天效率变差"
    if "ACOS" in text or "偏贵" in text or "广告效率" in text:
        return "广告效率偏贵"
    if ("推荐报价" in text or "Buy Box" in text) and any(
        marker in text for marker in ["低于", "不稳定", "丢", "风险", "<95"]
    ):
        return "推荐报价 / Buy Box"
    if any(marker in text for marker in ["展示从", "搜索入口变少", "排名", "曝光"]):
        return "曝光 / 排名下降"
    if "购买从" in text or "购买下降" in text or "搜索漏斗变弱" in text:
        return "搜索购买下降"
    if "traffic_sales" in text or "页面转化" in text or "页面成交" in text or "转化率" in text:
        return "页面成交弱"
    if "搜索词" in text or "ASIN" in text:
        return "搜索词 / 定向"
    if "SQP" in text or "搜索漏斗" in text or "展示" in text or "加购" in text:
        return "搜索漏斗待确认"
    return "数据异常待确认"


def _listing_problem_line(row: dict[str, str], evidence: object) -> str:
    history = str(row.get("历史人工确认") or "").strip()
    if history:
        root = str(row.get("初步方向") or "").replace("历史已确认：", "").strip()
        if root:
            return f"历史人工确认：{root}"
        return f"历史人工确认：{history}"
    signals = [
        item.strip()
        for item in str(row.get("异常信号") or row.get("主因") or "").split("；")
        if item.strip() and item.strip() != "N/A"
    ]
    metrics = [item.strip() for item in str(evidence or "").split("；") if item.strip() and item.strip() != "N/A"]
    metric_text = "；".join(metrics)
    problems: list[str] = []
    for signal in signals:
        display_signal = signal.replace("增强数据：", "", 1) if signal.startswith("增强数据：") else signal
        for segment in [part.strip() for part in display_signal.split("；") if part.strip()]:
            if "页面转化率 0.0%" in segment and "推荐报价率" in segment:
                segment = segment.split("，推荐报价率", 1)[0].strip()
            has_problem_token = any(token in segment for token in ["下降", "低于", "不稳定", "变少", "变弱", "购买 0", "偏贵", "效率变差", "ACOS"])
            if "页面转化率 0.0%" in segment:
                has_problem_token = True
            if "推荐报价率 100" in segment and not has_problem_token:
                continue
            if "页面转化率" in segment and "0.0%" not in segment and not has_problem_token:
                continue
            if "展示 " in segment and " / 点击 " in segment and not has_problem_token:
                continue
            if not has_problem_token and any(token in segment for token in ["页面转化率", "推荐报价率", "展示", "点击", "加购", "购买"]):
                continue
            problems.append(segment.replace("页面转化率", "页面成交率"))
    if "页面转化率 0.0%" in metric_text and not any(("页面转化率" in item or "页面成交率" in item) for item in problems):
        problems.insert(0, "页面成交率 0.0%，当前增强数据没有看到页面成交")
    if "搜索漏斗展示从" in metric_text and not any("搜索漏斗展示" in item for item in problems):
        start = metric_text.find("搜索漏斗展示从")
        if start >= 0:
            problems.append(metric_text[start:].split("；", 1)[0])
    if not problems:
        for metric in metrics:
            if any(token in metric for token in ["0 单", "下降", "变少", "变弱", "ACOS", "效率变差", "偏贵"]):
                problems.append(metric.replace("页面转化率", "页面成交率"))
                break
    return "；".join(problems[:2]) if problems else "暂无明确异常，先补截图人工确认"


def _listing_conservative_action(row: dict[str, str]) -> str:
    history = str(row.get("历史人工确认") or "").strip()
    if history:
        return _compact_semicolon_items(row.get("产品专属下一步") or history, "复查已执行动作效果，暂不重复改动", limit=2)
    text = "；".join(
        str(row.get(key) or "")
        for key in ["初步方向", "异常信号", "产品专属下一步", "建议动作", "主因"]
    )
    buy_box_risk = ("推荐报价" in text or "Buy Box" in text) and any(
        marker in text for marker in ["低于", "不稳定", "丢", "风险", "<95", "低 "]
    )
    if buy_box_risk:
        return "先确认推荐报价/配送/售价；未确认前不加预算，只保留核心词低价跑。"
    if "加购后" in text or "加购" in text:
        return "先查价格、Coupon、配送和竞品优惠；广告端只压高花费无单词。"
    if "广告效率" in text or "ACOS" in text or "偏贵" in text:
        return "保留能出单词，优先压贵词和无效 ASIN；不扩大预算。"
    if "流量不准" in text or "搜索词" in text or "ASIN" in text:
        return "先处理广告流量：明显无关才否，相关词降竞价，核心词不否。"
    if "搜索漏斗" in text or "排名" in text or "曝光" in text:
        return "先补搜索结果页/竞品截图；当天不急着改 Listing 或加预算。"
    if "近7天" in text:
        return "先看近7天价格、Coupon、配送、竞品变化；只做小幅保守调整。"
    return _compact_semicolon_items(row.get("产品专属下一步") or row.get("建议动作"), "先补关键截图，再决定是否调广告或页面", limit=1)


def _optional_num_from_text(value: object) -> float | None:
    if value in ("", None):
        return None
    text = str(value or "")
    cleaned = "".join(ch for ch in text if ch.isdigit() or ch in ".-")
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _risk_metric(row: dict[str, str], key: str) -> float:
    return _num_from_text(row.get(key))


def _render_watch_pool(rows: list[dict[str, str]], limit: int = 3) -> str:
    if not rows:
        return '<p class="subtle">当前没有需要放到滞销观察池的对象。</p>'
    high_rows: list[dict[str, str]] = []
    hidden_count = 0
    for row in rows:
        stock = _risk_metric(row, "库存")
        no_order_days = _risk_metric(row, "连续无单天数")
        orders14 = _risk_metric(row, "近14天订单")
        clicks14 = _risk_metric(row, "近14天点击")
        spend14 = _risk_metric(row, "近14天广告花费")
        if spend14 < 5 and clicks14 < 20:
            hidden_count += 1
            continue
        if stock < 5 and spend14 < 5:
            hidden_count += 1
            continue
        if stock >= 30 and no_order_days >= 7 and orders14 == 0 and (clicks14 >= 20 or spend14 >= 5):
            high_rows.append(row)
        else:
            hidden_count += 1
    show_rows = high_rows[:limit]
    if not show_rows:
        return f'<p class="subtle">暂无高风险滞销对象；另有 {hidden_count} 个中低风险或低样本对象，见 Excel。</p>'
    parts = ['<div class="action-grid">']
    for row in show_rows:
        stock = _risk_metric(row, "库存")
        orders14 = _risk_metric(row, "近14天订单")
        clicks14 = _risk_metric(row, "近14天点击")
        spend14 = _risk_metric(row, "近14天广告花费")
        no_order_days = _risk_metric(row, "连续无单天数")
        label = "低动销观察" if orders14 >= 1 else "高风险观察"
        if clicks14 >= 20 or spend14 >= 5:
            advice = "有库存、有流量、有花费但没单，需要复查搜索词是否不准、价格 / 优惠 / 竞品，不加大预算。"
        else:
            advice = "点击少 / 花费低，不急着处理；小预算精准词测试即可，不要直接改 Listing。"
        parts.append(
            "\n".join(
                [
                    '<article class="work-card p2">',
                    '<div class="card-head">',
                    "<div>",
                    f'<h3 class="card-title">{html.escape(str(row.get("产品") or "N/A"))}</h3>',
                    f'<div class="card-meta">{_product_meta_html(row, asin_key="ASIN", sku_key="SKU", marketplace_key="站点")}</div>',
                    "</div>",
                    f'<span class="tag tag-gray">{label}</span>',
                    "</div>",
                    f'<div class="metric-row">{_metric_badges_from_evidence(f"库存 {stock:g}；连续无单 {no_order_days:g} 天；近14天点击 {clicks14:g}；近14天广告花费 {spend14:g}；近14天订单 {orders14:g}", limit=6)}</div>',
                    f'<div class="card-block"><strong>判断</strong><p>{_inline_markup("有库存、有流量、有花费但没单，需要复查" if orders14 == 0 else "已有 1 单或低动销，不直接判定 Listing 崩。")}</p></div>',
                    f'<div class="card-block"><strong>差异化建议</strong><p>{_inline_markup(advice)}</p></div>',
                    "</article>",
                ]
            )
        )
    parts.append("</div>")
    if hidden_count or len(high_rows) > limit:
        parts.append(f'<p class="subtle">另有 {hidden_count + max(len(high_rows) - limit, 0)} 个中低风险滞销对象，见 Excel。</p>')
    return "".join(parts)


def _render_profit_cost_cards(rows: list[dict[str, str]], limit: int = 3) -> str:
    if not rows:
        return '<p class="subtle">当前没有需要优先展示的利润 / 成本待核对对象。</p>'

    def cost_value(row: dict[str, str], key: str, label: str, percent: bool = False) -> str:
        detail = _cost_detail_for_row(row)
        aliases = {
            "purchase_cost": "purchase_cost_local",
            "first_leg_cost": "first_leg_cost_local",
            "profit_before_ads_per_unit": "profit_before_ads",
            "target_acos": "suggested_target_acos",
        }
        value = row.get(key)
        if value in (None, "", "N/A"):
            value = detail.get(key) or detail.get(aliases.get(key, key))
        if value in (None, "", "N/A"):
            return f"{label} 成本字段缺失"
        text = str(value)
        if any(symbol in text for symbol in ["£", "$", "€", "%"]):
            return f"{label} {text}"
        number = _num_from_text(value)
        if percent:
            return f"{label} {number:.1%}" if number else f"{label} 0%"
        marketplace = row.get("marketplace") or row.get("站点") or detail.get("marketplace") or ""
        currency = row.get("currency") or detail.get("currency")
        symbol = money_symbol_for_marketplace(marketplace, currency)
        return f"{label} {symbol}{number:.2f}"

    def cost_evidence(row: dict[str, str], fallback: str) -> str:
        fields = [
            cost_value(row, "selling_price", "售价"),
            cost_value(row, "purchase_cost", "采购"),
            cost_value(row, "first_leg_cost", "头程"),
            cost_value(row, "fba_fee", "FBA"),
            cost_value(row, "vat", "VAT"),
            cost_value(row, "referral_fee", "佣金"),
            cost_value(row, "profit_before_ads_per_unit", "广告前利润"),
            cost_value(row, "target_acos", "target_acos", percent=True),
        ]
        detail = _cost_detail_for_row(row)
        cost_status = row.get("cost_status") or detail.get("cost_status")
        if cost_status not in (None, "", "N/A"):
            fields.append(f"成本状态 {cost_status}")
        if row.get("break_even_acos") not in (None, "", "N/A"):
            fields.append(cost_value(row, "break_even_acos", "盈亏平衡ACOS", percent=True))
        meaningful = [field for field in fields if "成本字段缺失" not in field]
        return "；".join(meaningful or fields[:4] or [fallback])

    parts = ['<div class="action-grid">']
    for row in rows[:limit]:
        reason = str(row.get("primary_reason") or row.get("风险等级") or row.get("risk_level") or row.get("风险原因") or row.get("key_evidence") or "")
        evidence = str(row.get("key_evidence") or "")
        action_text = str(row.get("today_action") or "")
        if "target_acos=0" in reason:
            judgement = "target_acos<=0，不能作为可投放目标"
            advice = "先核对售价、采购成本、头程、FBA、VAT 和佣金；利润边界确认前不建议放量。"
        elif "广告前利润<=0" in reason or "广告前利润为负" in reason or "利润为负" in reason:
            judgement = "广告前利润<=0"
            advice = "先核对成本是否准确；若成本准确，需调价 / 降成本 / 减少广告；不要盲目加预算。"
        elif "高库存" in reason or "库存" in reason:
            judgement = "库存较多但动销偏低"
            advice = "先核对库存、售价和实际毛利；广告花费不高时不放在强动作，作为库存/财务观察项。"
        else:
            judgement = "库存较多但利润边界不清"
            advice = action_text or "标为财务待核对；若近14天广告花费很低，不放在今日重点动作。"
        parts.append(
            "\n".join(
                [
                    '<article class="work-card cost">',
                    '<div class="card-head">',
                    "<div>",
                    f'<h3 class="card-title">{html.escape(str(row.get("product_name") or row.get("产品") or "N/A"))}</h3>',
                    f'<div class="card-meta">{_product_meta_html(row)}</div>',
                    "</div>",
                    '<span class="tag tag-gray">财务待核对</span>',
                    "</div>",
                    f'<div class="card-block"><strong>判断类型</strong><p>{_inline_markup(judgement)}</p></div>',
                    f'<div class="card-block"><strong>当前证据</strong><div class="metric-row">{_metric_badges_from_evidence(cost_evidence(row, evidence or reason), limit=9)}</div></div>',
                    f'<div class="card-block"><strong>建议动作</strong><p>{_inline_markup(advice)}</p></div>',
                    _render_task_review_status(row),
                    "</article>",
                ]
            )
        )
    parts.append("</div>")
    if len(rows) > limit:
        parts.append(f'<p class="subtle">另有 {len(rows) - limit} 条利润 / 成本对象见 Excel。</p>')
    return "".join(parts)


def _render_scale_candidate_cards(rows: list[dict[str, str]], limit: int = 4) -> str:
    if not rows:
        return '<p class="subtle">当前没有达到放量条件的产品。ACOS 略低但样本不足的对象继续观察。</p>'
    parts = ['<div class="action-grid">']
    for row in rows[:limit]:
        level = str(row.get("放量等级") or "谨慎放量候选")
        tag = "tag-green" if level == "可小幅放量" else "tag-yellow"
        title = f"{row.get('站点') or ''}｜{row.get('产品') or 'N/A'}"
        metrics = [
            f"点击 {row.get('点击')}",
            f"广告单 {row.get('订单')}",
            f"总单 {row.get('总单')}",
            f"花费 {row.get('花费')}",
            f"ACOS {row.get('ACOS')}",
            f"目标 {row.get('目标 ACOS')}",
        ]
        parts.append(
            "\n".join(
                [
                    '<article class="work-card status-pass">',
                    '<div class="card-head">',
                    "<div>",
                    f'<h3 class="card-title">{html.escape(title)}</h3>',
                    f'<div class="card-meta">{_product_meta_html(row, asin_key="ASIN", sku_key="SKU", marketplace_key="站点")}</div>',
                    "</div>",
                    f'<span class="tag {tag}">{html.escape(level)}</span>',
                    "</div>",
                    '<div class="metric-row">' + "".join(f'<span class="metric-badge">{html.escape(metric)}</span>' for metric in metrics) + "</div>",
                    f'<div class="card-block"><strong>建议</strong><p>{_inline_markup(str(row.get("建议") or "N/A"))}</p></div>',
                    "</article>",
                ]
            )
        )
    parts.append("</div>")
    if len(rows) > limit:
        parts.append(f'<p class="subtle">另有 {len(rows) - limit} 个放量候选，见 Excel。</p>')
    return "".join(parts)


def _render_scale_keyword_cards(rows: list[dict[str, str]], limit: int = 8) -> str:
    if not rows:
        return '<p class="subtle">当前没有找到可直接放量的出单词 / ASIN。产品候选如果还在上方，先去 Excel 搜索词明细核对具体出单词。</p>'
    parts = ['<div class="queue-evidence-list">']
    for row in rows[:limit]:
        term = str(row.get("search_term_or_target") or "N/A")
        action = str(row.get("scale_action") or "保留出单词")
        product = str(row.get("product_name") or "N/A")
        marketplace = str(row.get("marketplace") or "")
        metrics = (
            f"点击 {row.get('clicks')} ｜ 花费 {row.get('spend')} ｜ "
            f"广告单 {row.get('ad_orders')} ｜ 销售 {row.get('ad_sales')} ｜ "
            f"ACOS {row.get('ACOS')} / 目标 {row.get('target_acos')}"
        )
        parts.append(
            "\n".join(
                [
                    '<div class="queue-evidence-item">',
                    f'<div class="queue-evidence-title">{html.escape(marketplace)}｜{_inline_markup(product)}｜{html.escape(term)}</div>',
                    f'<div class="queue-evidence-body"><strong>{html.escape(action)}</strong></div>',
                    f'<div class="queue-evidence-body">{html.escape(metrics)}</div>',
                    f'<div class="queue-evidence-body">Campaign：{html.escape(str(row.get("campaign_name") or "N/A"))}</div>',
                    "</div>",
                ]
            )
        )
    parts.append("</div>")
    if len(rows) > limit:
        parts.append(f'<p class="subtle">另有 {len(rows) - limit} 个词级放量候选，见 Excel。</p>')
    return "".join(parts)


def _render_product_final_decision_cards(rows: list[dict[str, str]], limit: int = 6) -> str:
    if not rows:
        return '<p class="subtle">当前没有生成产品最终决策。</p>'
    action_labels = {
        "bid_up": "小幅加竞价",
        "bid_down": "降竞价",
        "broad_scale": "放量",
        "budget_up": "加预算",
        "create_exact_low_budget": "低预算精准测试",
        "negative_exact": "否定精准",
        "observe": "只观察，不操作",
        "pause": "暂停",
    }
    blocked_action_labels = {
        "bid_up": "拦截加竞价",
        "bid_down": "拦截降竞价",
        "broad_scale": "拦截放量",
        "budget_up": "拦截加预算",
        "create_exact_low_budget": "拦截低预算精准测试",
        "negative_exact": "拦截否定精准",
        "observe": "拦截观察",
        "pause": "拦截暂停",
    }

    def action_text(values: object, fallback: str, labels_by_action: dict[str, str]) -> str:
        if not isinstance(values, list):
            return str(values or fallback)
        labels = [labels_by_action.get(str(item), str(item)) for item in values if str(item)]
        return " / ".join(labels) if labels else fallback

    def list_text(values: object) -> str:
        if isinstance(values, list):
            return "；".join(str(item) for item in values[:4] if str(item).strip())
        return str(values or "").strip()

    parts = ['<div class="action-grid">']
    for row in rows[:limit]:
        marketplace = str(row.get("marketplace") or "").strip()
        sku = str(row.get("sku") or "").strip()
        asin = str(row.get("asin") or "").strip().upper()
        label, label_title = _operation_decision_display(row)
        decision = str(row.get("final_decision") or "")
        if decision in {"DATA_INSUFFICIENT", "DO_NOT_TOUCH"}:
            tag = "tag tag-gray"
        elif decision in {"WAIT_REVIEW", "FRONTEND_FIRST", "CONSERVATIVE_RUN"}:
            tag = "tag tag-yellow"
        else:
            tag = "tag tag-green"
        evidence_used = row.get("evidence_used") or []
        if isinstance(evidence_used, list):
            evidence_text = "；".join(str(item) for item in evidence_used[:3] if item)
        else:
            evidence_text = str(evidence_used or "")
        blocked_text = action_text(row.get("today_blocked_actions") or [], "无", blocked_action_labels)
        allowed_text = action_text(row.get("today_allowed_actions") or [], "只观察，不操作", action_labels)
        next_review = str(row.get("next_review_date") or "")
        frontend_gate_bits = [
            str(row.get("frontend_check_status") or ""),
            str(row.get("frontend_evidence_display_tier") or row.get("frontend_evidence_tier") or ""),
        ]
        blocking_reason_text = list_text(row.get("frontend_blocking_reasons"))
        audit_detail = str(row.get("frontend_evidence_audit_detail") or "").strip()
        if not audit_detail:
            audit_detail = list_text(row.get("frontend_evidence_audit_reasons"))
        if audit_detail:
            frontend_gate_bits.append(audit_detail)
        def market_score_number(value: object) -> int:
            try:
                return int(float(value))
            except (TypeError, ValueError):
                return 0

        amazon_total = (
            market_score_number(row.get("amazon_product_page_completeness"))
            + market_score_number(row.get("amazon_search_page_completeness"))
        )
        competitor_total = (
            market_score_number(row.get("sellersprite_competitor_pool_completeness"))
            + market_score_number(row.get("sellersprite_competitor_reverse_completeness"))
        )
        market_bits = [
            f"完整度 {row.get('market_survey_completeness_score')}/100" if row.get("market_survey_completeness_score") not in (None, "") else "",
            f"等级 {_market_survey_level_label(row.get('market_survey_completeness_level'))}" if row.get("market_survey_completeness_level") else "",
            f"证据层级 {_market_survey_level_label(row.get('market_survey_decision_evidence_tier'))}" if row.get("market_survey_decision_evidence_tier") else "",
            f"Amazon 前台完整度 {amazon_total}/35" if row.get("amazon_product_page_completeness") not in (None, "") or row.get("amazon_search_page_completeness") not in (None, "") else "",
            f"卖家精灵完整度 {market_score_number(row.get('sellersprite_own_completeness'))}/20" if row.get("sellersprite_own_completeness") not in (None, "") else "",
            f"竞品完整度 {competitor_total}/30" if row.get("sellersprite_competitor_pool_completeness") not in (None, "") or row.get("sellersprite_competitor_reverse_completeness") not in (None, "") else "",
            f"趋势完整度 {market_score_number(row.get('sellersprite_trend_completeness'))}/10" if row.get("sellersprite_trend_completeness") not in (None, "") else "",
            f"缺口 {row.get('market_survey_missing_parts')}" if row.get("market_survey_missing_parts") else "",
            f"补抓 {row.get('market_survey_recommended_fetch_steps')}" if row.get("market_survey_recommended_fetch_steps") else "",
        ]
        market_text = "；".join(bit for bit in market_bits if bit)
        search_status = str(row.get("frontend_search_status") or "").strip()
        search_partial = _boolish(row.get("frontend_search_partial_evidence")) or search_status == "已读取部分结果"
        if search_partial:
            frontend_gate_bits.append(f"搜索页：{search_status or '部分读取'}")
        elif search_status:
            frontend_gate_bits.append(f"搜索页：{search_status}")
        frontend_gate_text = "；".join(bit for bit in frontend_gate_bits if bit)
        parts.append(
            "\n".join(
                [
                    (
                        '<article class="work-card"'
                        f' data-product-decision-marketplace="{html.escape(marketplace, quote=True)}"'
                        f' data-product-decision-sku="{html.escape(sku, quote=True)}"'
                        f' data-product-decision-asin="{html.escape(asin, quote=True)}"'
                        f'{_product_decision_contract_attrs(row)}>'
                    ),
                    '<div class="card-head">',
                    '<div>',
                    f'<h3 class="card-title">{html.escape(marketplace)}｜{html.escape(str(row.get("product_name") or row.get("asin") or "N/A"))}</h3>',
                    f'<div class="card-meta">{_product_meta_html(row)}</div>',
                    '</div>',
                    f'<span class="{tag}" title="{html.escape(label_title)}">{html.escape(label)}</span>',
                    '</div>',
                    f'<div class="card-block"><strong>决策原因</strong><p>{_inline_markup(str(row.get("decision_reason") or row.get("ad_action_summary") or "N/A"))}</p></div>',
                    (f'<div class="card-block"><strong>市场调查完整度</strong><p>{_inline_markup(market_text)}</p></div>' if market_text else ''),
                    (f'<div class="card-block"><strong>前台门禁</strong><p>{_inline_markup(frontend_gate_text)}</p></div>' if frontend_gate_text else ''),
                    (f'<div class="card-block"><strong>阻断原因</strong><p>{_inline_markup(blocking_reason_text)}</p></div>' if blocking_reason_text else ''),
                    f'<div class="card-block"><strong>门禁放行动作</strong><p>{_inline_markup(allowed_text or "只观察，不操作")}</p></div>',
                    f'<div class="card-block"><strong>门禁拦截动作</strong><p>{_inline_markup(blocked_text or "无")}</p></div>',
                    (f'<div class="card-block"><strong>下次复查</strong><p>{html.escape(next_review)}</p></div>' if next_review else ''),
                    (f'<div class="card-block"><strong>依据</strong><p>{_inline_markup(evidence_text)}</p></div>' if evidence_text else ''),
                    '</article>',
                ]
            )
        )
    parts.append('</div>')
    if len(rows) > limit:
        parts.append(f'<p class="subtle">另有 {len(rows) - limit} 个产品决策见 Excel。</p>')
    return ''.join(parts)


def _format_operation_metric(value: object, marketplace: object = "", kind: str = "count") -> str:
    if value in (None, ""):
        return "N/A"
    if kind == "money":
        text = str(value)
        if any(symbol in text for symbol in ["£", "$", "€"]):
            return text
        number = _num_from_text(value)
        if number is None:
            return text
        return f"{money_symbol_for_marketplace(str(marketplace or ''))}{number:.2f}"
    if kind == "percent":
        number = _num_from_text(value)
        if number is None:
            return str(value)
        return f"{number:.1%}" if abs(number) <= 1 else f"{number:.1f}%"
    number = _num_from_text(value)
    if number is None:
        return str(value)
    return str(int(number)) if number == int(number) else f"{number:.1f}"

def _action_list_text(values: object, labels: dict[str, str]) -> str:
    if isinstance(values, list):
        items = [labels.get(str(item), str(item)) for item in values if str(item)]
        return " / ".join(items) if items else "无"
    text = str(values or "").strip()
    if not text:
        return "无"
    return " / ".join(labels.get(part.strip(), part.strip()) for part in re.split(r"[；,/]", text) if part.strip())


def _action_item_count(values: object) -> int:
    if isinstance(values, list):
        return len([item for item in values if str(item).strip()])
    text = str(values or "").strip()
    if not text:
        return 0
    return len([part for part in re.split(r"[；,/]", text) if part.strip()])


def _render_operation_ad_actions(rows: list[dict[str, str]]) -> str:
    if not rows:
        return '<p class="subtle">当前没有卡内可执行广告止损项。</p>'
    parts = ['<div class="queue-evidence-list">']
    for row in rows[:4]:
        term = str(row.get("search_term_or_target") or "N/A")
        action = str(row.get("suggested_action") or row.get("scale_action") or row.get("copy_action_line") or "观察")
        orders = row.get("orders") if row.get("orders") not in (None, "") else row.get("ad_orders")
        metrics = f"点击 {row.get('clicks') or 'N/A'} ｜ 花费 {row.get('spend') or 'N/A'} ｜ 订单 {orders or 'N/A'}"
        parts.append(
            "".join(
                [
                    '<div class="queue-evidence-item">',
                    f'<div class="queue-evidence-title">{html.escape(action)}｜{_inline_markup(term)}</div>',
                    f'<div class="queue-evidence-body">{html.escape(metrics)}</div>',
                    "</div>",
                ]
            )
        )
    hidden_count = max(0, len(rows) - 4)
    if hidden_count:
        parts.append(f'<p class="subtle">另有 {hidden_count} 条广告动作见广告工作台。</p>')
    parts.append("</div>")
    return "".join(parts)


def _render_frontend_coverage_strip(summary: dict[str, object]) -> str:
    if not summary:
        return ""
    total = html.escape(str(summary.get("frontend_queue_total", 0)))
    raw_total = str(summary.get("frontend_queue_total", 0))
    amazon_validation_count = str(
        summary.get(
            "frontend_amazon_search_validation_count",
            summary.get("frontend_competitor_search_success_count", 0),
        )
        or 0
    )
    amazon_validation_label = str(
        summary.get("frontend_amazon_search_validation_label") or f"{amazon_validation_count}/{raw_total}"
    )
    scalable_strong_count = str(summary.get("frontend_scalable_strong_count", 0) or 0)
    scalable_strong_label = str(summary.get("frontend_scalable_strong_label") or f"0/{raw_total}")
    market_average_label = str(summary.get("market_survey_average_score_label") or f'{summary.get("market_survey_average_score", 0)}/100')
    market_complete_label = str(summary.get("market_survey_complete_label") or f'{summary.get("market_survey_complete_count", 0)}/{raw_total}')
    market_usable_label = str(summary.get("market_survey_usable_label") or f'{summary.get("market_survey_usable_count", 0)}/{raw_total}')
    market_insufficient_label = str(summary.get("market_survey_insufficient_label") or f'{summary.get("market_survey_insufficient_count", 0)}/{raw_total}')
    market_failed_label = str(summary.get("market_survey_failed_label") or f'{summary.get("market_survey_failed_count", 0)}/{raw_total}')
    market_complete_count = int(float(summary.get("market_survey_complete_count", 0) or 0))
    market_insufficient_count = int(float(summary.get("market_survey_insufficient_count", 0) or 0))
    market_failed_count = int(float(summary.get("market_survey_failed_count", 0) or 0))
    own_sellersprite_today_label = str(
        summary.get("frontend_own_sellersprite_today_label")
        or (
            f'今日 {summary.get("frontend_own_sellersprite_today_count", 0)}/{raw_total}，'
            f'缓存 {summary.get("frontend_own_sellersprite_cache_count", 0)}/{raw_total}'
        )
    )
    competitor_pool_freshness_label = str(
        summary.get("frontend_competitor_pool_freshness_label")
        or (
            f'今日 {summary.get("frontend_competitor_pool_today_count", 0)}/{raw_total}，'
            f'7天缓存 {summary.get("frontend_competitor_pool_cache_count", 0)}/{raw_total}'
        )
    )
    competitor_sellersprite_freshness_label = str(
        summary.get("frontend_competitor_sellersprite_freshness_label")
        or (
            f'今日 {summary.get("frontend_competitor_sellersprite_today_count", 0)}/{raw_total}，'
            f'缓存 {summary.get("frontend_competitor_sellersprite_cache_count", 0)}/{raw_total}'
        )
    )
    items = [
        (
            "市场调查平均完整度",
            html.escape(market_average_label),
            "综合商品页、搜索页、卖家精灵和趋势",
            "pass" if market_complete_count else "info",
        ),
        (
            "强证据 / 可用证据",
            html.escape(f"{market_complete_label} / {market_usable_label}"),
            "强操作 / 背景止损",
            "info",
        ),
        (
            "产品页成功",
            html.escape(str(summary.get("frontend_product_page_success_label") or f'{summary.get("frontend_live_success_count", 0)}/{total}')),
            "自己产品页",
            "pass" if str(summary.get("frontend_product_page_success_count", summary.get("frontend_live_success_count", 0))) != "0" else "muted",
        ),
        (
            "卖家精灵自己 ASIN",
            html.escape(
                str(
                    own_sellersprite_today_label
                    or summary.get("frontend_own_sellersprite_label")
                    or f'{summary.get("frontend_own_sellersprite_count", 0)}/{total}'
                )
            ),
            "今日抓取 / 沿用缓存",
            "info" if str(summary.get("frontend_own_sellersprite_count", 0)) != "0" else "muted",
        ),
        (
            "卖家精灵趋势",
            html.escape(str(summary.get("frontend_sellersprite_trend_ready_label") or f'{summary.get("frontend_sellersprite_trend_ready_count", 0)}/{total}')),
            "3天或7天趋势",
            "info" if str(summary.get("frontend_sellersprite_trend_ready_count", 0)) != "0" else "muted",
        ),
        (
            "卖家精灵竞品发现",
            html.escape(str(summary.get("frontend_competitor_discovery_label") or f'{summary.get("frontend_competitor_discovery_count", 0)}/{total}')),
            "直达页面发现主竞品",
            "info" if str(summary.get("frontend_competitor_discovery_count", 0)) != "0" else "muted",
        ),
        (
            "卖家精灵竞品池",
            html.escape(
                str(
                    competitor_pool_freshness_label
                    or summary.get("frontend_competitor_pool_label")
                    or f'{summary.get("frontend_competitor_pool_count", 0)}/{total}'
                )
            ),
            "今日快照 / 7天缓存",
            "info" if str(summary.get("frontend_competitor_pool_count", 0)) != "0" else "muted",
        ),
        (
            "竞品 ASIN 反查",
            html.escape(
                str(
                    competitor_sellersprite_freshness_label
                    or summary.get("frontend_competitor_sellersprite_label")
                    or f'{summary.get("frontend_competitor_sellersprite_count", 0)}/{total}，'
                    f'{summary.get("frontend_competitor_sellersprite_asin_count", 0)} ASIN'
                )
            ),
            "竞品 ASIN 反查",
            "info" if str(summary.get("frontend_competitor_sellersprite_count", 0)) != "0" else "muted",
        ),
        (
            "Amazon 搜索页辅助验证",
            html.escape(amazon_validation_label if amazon_validation_count != "0" else "未跑"),
            html.escape("只验证重点词可见性" if amazon_validation_count != "0" else f"{amazon_validation_label}，只在重点词补证"),
            "info" if amazon_validation_count != "0" else "muted",
        ),
        (
            "达到放量准入",
            html.escape(scalable_strong_label if scalable_strong_count != "0" else "暂无"),
            html.escape("同时满足前台和反查" if scalable_strong_count != "0" else f"{scalable_strong_label}，不建议放量"),
            "pass" if scalable_strong_count != "0" else "muted",
        ),
        (
            "弱势止损证据",
            html.escape(str(summary.get("frontend_weak_defensive_label") or f'{summary.get("frontend_weak_defensive_count", 0)}/{total}')),
            "优先控费",
            "warn" if str(summary.get("frontend_weak_defensive_count", 0)) != "0" else "muted",
        ),
        (
            "证据不足",
            html.escape(str(summary.get("frontend_insufficient_label") or f'{summary.get("frontend_insufficient_count", 0)}/{total}')),
            "待补竞品或反查",
            "warn" if str(summary.get("frontend_insufficient_count", 0)) != "0" else "muted",
        ),
    ]
    if market_insufficient_count or market_failed_count:
        items.insert(
            2,
            (
                "待补 / 失败",
                html.escape(f"{market_insufficient_label} / {market_failed_label}"),
                "需要补证据",
                "warn",
            ),
        )
    return "\n".join(
        [
            '<div class="frontend-coverage-strip" aria-label="前台证据覆盖">',
            *(
                '<span class="frontend-coverage-chip '
                + css
                + '"><span class="coverage-label">'
                + label
                + '</span><strong>'
                + value
                + '</strong><span class="coverage-note">'
                + note
                + "</span></span>"
                for label, value, note, css in items
            ),
            "</div>",
        ]
    )


def _seller_sprite_product_summary(row: dict[str, object]) -> str:
    conclusion = str(row.get("product_level_conclusion") or "").strip()
    pressure = str(row.get("competitor_keyword_pressure") or "").strip()
    product_page = "缓存" if _boolish(row.get("frontend_cache_used")) else "已读" if str(row.get("frontend_status") or "") == "已自动检查" else "失败" if str(row.get("frontend_status") or "") else "待补"
    own_status = str(row.get("seller_sprite_check_status") or "").strip()
    discovery_status = str(row.get("competitor_discovery_status") or "").strip()
    discovery_page = str(row.get("competitor_discovery_source_page") or "").strip()
    discovery_error = str(row.get("competitor_discovery_error") or "").strip()
    pool_status = str(row.get("competitor_pool_status") or "").strip()
    pool_source = str(row.get("competitor_discovery_source") or "").strip()
    pool_confidence = str(row.get("competitor_pool_confidence") or "").strip()
    pool_count = str(row.get("competitor_pool_count") or "").strip()
    main_competitor_count = str(row.get("main_competitor_count") or "").strip()
    main_competitor_asins = _compact_keyword_items(row.get("main_competitor_asins"), fallback="", limit=3)
    reference_competitor_count = str(row.get("reference_competitor_count") or "").strip()
    comparability_score = str(row.get("competitor_comparability_score") or "").strip()
    spec_status = str(row.get("competitor_spec_match_status") or "").strip()
    price_status = str(row.get("competitor_price_band_status") or "").strip()
    review_status = str(row.get("competitor_review_tier_status") or "").strip()
    stability_days = str(row.get("competitor_stability_days") or "").strip()
    scalable_status = str(row.get("scalable_evidence_status") or "").strip()
    scalable_blockers = _compact_semicolon_items(row.get("scalable_blockers"), fallback="", limit=2)
    source_keywords = _compact_keyword_items(row.get("competitor_source_keywords"), fallback="", limit=2)
    rejected_count = str(row.get("competitor_rejected_count") or "").strip()
    rejection_reasons = _compact_semicolon_items(row.get("competitor_rejection_reasons"), fallback="", limit=2)
    amazon_validation = str(row.get("amazon_search_validation_status") or "").strip()
    competitor_status = str(row.get("competitor_sellersprite_status") or "").strip()
    missing = _compact_keyword_items(row.get("own_missing_competitor_keywords"), fallback="", limit=2)
    no_match = _compact_keyword_items(row.get("own_ad_terms_not_in_sellersprite"), fallback="", limit=2)
    today_status = str(row.get("sellersprite_today_status") or "").strip()
    cache_date = str(row.get("sellersprite_cache_date") or "").strip()
    history_days = str(row.get("sellersprite_history_days") or "").strip()
    trend_status = str(row.get("sellersprite_trend_status") or "").strip()
    persistent_keywords = _compact_keyword_items(row.get("sellersprite_persistent_keywords"), fallback="", limit=2)
    stable_asins = _compact_keyword_items(row.get("competitor_stable_asins"), fallback="", limit=3)
    ppc_up = _compact_keyword_items(row.get("sellersprite_ppc_up_keywords"), fallback="", limit=2)
    pressure_trend = str(row.get("competitor_pressure_trend") or "").strip()
    missing_trend = _compact_keyword_items(row.get("own_missing_competitor_keywords_trend"), fallback="", limit=2)
    market_score = str(row.get("market_survey_completeness_score") or "").strip()
    market_level = str(row.get("market_survey_completeness_level") or "").strip()
    market_tier = str(row.get("market_survey_decision_evidence_tier") or "").strip()
    market_missing = _compact_semicolon_items(row.get("market_survey_missing_parts"), fallback="", limit=3)
    market_steps = _compact_semicolon_items(row.get("market_survey_recommended_fetch_steps"), fallback="", limit=3)
    market_skip = _compact_semicolon_items(row.get("market_survey_skip_reason"), fallback="", limit=2)
    def score_text(value: object) -> str:
        return "" if value is None or value == "" else str(value).strip()

    amazon_product_score = score_text(row.get("amazon_product_page_completeness"))
    amazon_search_score = score_text(row.get("amazon_search_page_completeness"))
    seller_own_score = score_text(row.get("sellersprite_own_completeness"))
    seller_pool_score = score_text(row.get("sellersprite_competitor_pool_completeness"))
    seller_reverse_score = score_text(row.get("sellersprite_competitor_reverse_completeness"))
    seller_trend_score = score_text(row.get("sellersprite_trend_completeness"))
    seller_penalty = score_text(row.get("sellersprite_data_quality_penalty"))
    chips = []
    if market_score or market_level:
        level_text = _compact_unique_labels([market_level, market_tier], _market_survey_level_label)
        chips.append(f'<span class="seller-chip main">市场调查完整度 {html.escape((market_score or "0") + "/100" + ("｜" + level_text if level_text else ""))}</span>')
    def score_number(value: str) -> int:
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return 0

    dimension_parts = []
    if amazon_product_score or amazon_search_score:
        amazon_total = score_number(amazon_product_score) + score_number(amazon_search_score)
        dimension_parts.append(
            f"Amazon 前台完整度 {amazon_total}/35，产品页 {amazon_product_score or 0}/20，搜索页 {amazon_search_score or 0}/15"
        )
    if seller_own_score:
        dimension_parts.append(f"卖家精灵完整度 {seller_own_score}/20")
    if seller_pool_score or seller_reverse_score:
        competitor_total = score_number(seller_pool_score) + score_number(seller_reverse_score)
        dimension_parts.append(
            f"竞品完整度 {competitor_total}/30，竞品池 {seller_pool_score or 0}/15，竞品反查 {seller_reverse_score or 0}/15"
        )
    if seller_trend_score:
        dimension_parts.append(f"趋势完整度 {seller_trend_score}/10")
    if seller_penalty and seller_penalty != "0":
        dimension_parts.append(f"质量扣分 {seller_penalty}/20")
    if dimension_parts:
        chips.append(f'<span class="seller-chip">完整度拆分 {html.escape("；".join(dimension_parts[:5]))}</span>')
    if market_missing:
        chips.append(f'<span class="seller-chip warn">缺口 {html.escape(market_missing)}</span>')
    if market_steps:
        chips.append(f'<span class="seller-chip">补抓 {html.escape(market_steps)}</span>')
    elif market_skip:
        chips.append(f'<span class="seller-chip muted">跳过 {html.escape(market_skip)}</span>')
    chips.append(f'<span class="seller-chip">产品页 {html.escape(product_page)}</span>')
    if not any([
        conclusion,
        pressure,
        discovery_status,
        pool_status,
        own_status,
        competitor_status,
        amazon_validation,
        missing,
        no_match,
        today_status,
        trend_status,
        main_competitor_count,
        reference_competitor_count,
        comparability_score,
        scalable_status,
        scalable_blockers,
    ]):
        fallback_status = "待补" if str(row.get("frontend_status") or "").strip() else "未入本轮队列"
        chips.append(f'<span class="seller-chip muted">卖家精灵自己 {html.escape(fallback_status)}</span>')
        chips.append(f'<span class="seller-chip muted">竞品池 {html.escape(fallback_status)}</span>')
        chips.append(f'<span class="seller-chip muted">竞品反查 {html.escape(fallback_status)}</span>')
        return '<div class="operation-seller-summary">' + "".join(chips) + "</div>"
    if today_status:
        today_text = today_status
        if today_status == "沿用缓存" and cache_date:
            today_text += f" {cache_date}"
        if history_days and history_days not in {"0", "0.0"}:
            today_text += f"｜{history_days}天"
        chips.append(f'<span class="seller-chip">卖家精灵状态 {html.escape(today_text)}</span>')
    if own_status:
        chips.append(f'<span class="seller-chip">卖家精灵自己 {html.escape(own_status)}</span>')
    if trend_status:
        chips.append(f'<span class="seller-chip">趋势 {html.escape(trend_status)}</span>')
    if discovery_status:
        page_label = _seller_source_label(discovery_page)
        page_part = f"｜{page_label}" if page_label else ""
        error_part = f"｜{discovery_error[:36]}" if discovery_error and discovery_status not in {"已抓取", "沿用缓存", "缓存"} else ""
        chips.append(f'<span class="seller-chip">竞品发现 {html.escape(discovery_status + page_part + error_part)}</span>')
    if pool_status:
        source_label = _seller_source_label(pool_source)
        confidence_label = _confidence_label(pool_confidence)
        source_part = f"｜{source_label}" if source_label else ""
        confidence_part = f"｜{confidence_label}" if confidence_label else ""
        try:
            pool_count_num = int(float(pool_count)) if pool_count else 0
        except ValueError:
            pool_count_num = 0
        count_part = (
            f"｜有效{pool_count_num}/3"
            if pool_count_num > 0 and not re.search(r"\d+\s*/\s*3", pool_status)
            else ""
        )
        chips.append(f'<span class="seller-chip">竞品池 {html.escape(pool_status + count_part + source_part + confidence_part)}</span>')
    if main_competitor_count and main_competitor_count not in {"0", "0.0"}:
        main_text = f"主竞品 {main_competitor_count}/3"
        if main_competitor_asins:
            main_text += f" {main_competitor_asins}"
        chips.append(f'<span class="seller-chip main">{html.escape(main_text)}</span>')
    if reference_competitor_count and reference_competitor_count not in {"0", "0.0"}:
        chips.append(f'<span class="seller-chip">参考竞品 {html.escape(reference_competitor_count)}</span>')
    if comparability_score:
        chips.append(f'<span class="seller-chip">可比性 {html.escape(comparability_score)}/100</span>')
    status_bits = [bit for bit in [spec_status, price_status, review_status] if bit]
    if status_bits:
        chips.append(f'<span class="seller-chip">可比检查 {html.escape("｜".join(status_bits[:3]))}</span>')
    if stability_days and stability_days not in {"0", "0.0"}:
        chips.append(f'<span class="seller-chip">趋势 {html.escape(stability_days)}天</span>')
    if scalable_status:
        css = "main" if scalable_status == "可谨慎放量" else "warn" if scalable_status in {"只能控费止损", "证据不足"} else ""
        chips.append(f'<span class="seller-chip {css}">放量证据 {html.escape(scalable_status)}</span>')
    if scalable_blockers:
        chips.append(f'<span class="seller-chip warn">拦截 {html.escape(scalable_blockers)}</span>')
    if source_keywords:
        chips.append(f'<span class="seller-chip">来源词 {html.escape(source_keywords)}</span>')
    if rejected_count and rejected_count != "0":
        reject_text = f"剔除 {rejected_count}" + (f"｜{rejection_reasons}" if rejection_reasons else "")
        chips.append(f'<span class="seller-chip warn">{html.escape(reject_text)}</span>')
    if competitor_status:
        chips.append(f'<span class="seller-chip">竞品反查 {html.escape(competitor_status)}</span>')
    if amazon_validation:
        chips.append(f'<span class="seller-chip">Amazon 搜索 {html.escape(amazon_validation)}</span>')
    if conclusion:
        chips.append(f'<span class="seller-chip main">卖家精灵：{html.escape(conclusion)}</span>')
    if pressure:
        chips.append(f'<span class="seller-chip">竞品压力 {html.escape(pressure)}</span>')
    if pressure_trend and pressure_trend != "无趋势":
        chips.append(f'<span class="seller-chip">压力趋势 {html.escape(pressure_trend)}</span>')
    if persistent_keywords:
        chips.append(f'<span class="seller-chip">连续词 {html.escape(persistent_keywords)}</span>')
    if stable_asins:
        chips.append(f'<span class="seller-chip">稳定竞品 {html.escape(stable_asins)}</span>')
    if missing:
        chips.append(f'<span class="seller-chip">缺口词 {html.escape(missing)}</span>')
    elif missing_trend:
        chips.append(f'<span class="seller-chip">持续缺口 {html.escape(missing_trend)}</span>')
    elif pool_status in {"", "待补", "卖家精灵证据不足", "竞品证据不足"}:
        chips.append('<span class="seller-chip">缺口词 待补</span>')
    if ppc_up:
        chips.append(f'<span class="seller-chip warn">PPC上升 {html.escape(ppc_up)}</span>')
    if no_match:
        chips.append(f'<span class="seller-chip warn">广告无反查 {html.escape(no_match)}</span>')
    return '<div class="operation-seller-summary">' + "".join(chips) + "</div>"


def _seller_sprite_decision_sentence(row: dict[str, object]) -> str:
    conclusion = str(row.get("product_level_conclusion") or "").strip()
    pressure = str(row.get("competitor_keyword_pressure") or "").strip()
    boundary = str(row.get("product_ad_boundary") or "").strip()
    own_status = str(row.get("seller_sprite_check_status") or "").strip()
    pool_status = str(row.get("competitor_pool_status") or "").strip()
    competitor_status = str(row.get("competitor_sellersprite_status") or "").strip()
    missing = _compact_keyword_items(row.get("own_missing_competitor_keywords"), fallback="", limit=2)
    no_match = _compact_keyword_items(row.get("own_ad_terms_not_in_sellersprite"), fallback="", limit=2)
    today_status = str(row.get("sellersprite_today_status") or "").strip()
    cache_date = str(row.get("sellersprite_cache_date") or "").strip()
    trend_status = str(row.get("sellersprite_trend_status") or "").strip()
    persistent = _compact_keyword_items(row.get("sellersprite_persistent_keywords"), fallback="", limit=2)
    stable_asins = _compact_keyword_items(row.get("competitor_stable_asins"), fallback="", limit=3)
    market_score = str(row.get("market_survey_completeness_score") or "").strip()
    market_level = str(row.get("market_survey_completeness_level") or "").strip()
    market_tier = str(row.get("market_survey_decision_evidence_tier") or "").strip()
    market_missing = _compact_semicolon_items(row.get("market_survey_missing_parts"), fallback="", limit=2)
    main_count = str(row.get("main_competitor_count") or "").strip()
    comparability = str(row.get("competitor_comparability_score") or "").strip()
    scalable_status = str(row.get("scalable_evidence_status") or "").strip()
    scalable_blockers = _compact_semicolon_items(row.get("scalable_blockers"), fallback="", limit=2)
    parts: list[str] = []
    if market_score or market_level:
        level_text = _compact_unique_labels([market_level, market_tier], _market_survey_level_label)
        parts.append(f"市场调查完整度 {market_score or '0'}/100 {level_text}")
    if market_missing:
        parts.append(f"缺口 {market_missing}")
    if conclusion:
        parts.append(f"结论 {conclusion}")
    if today_status:
        if today_status == "沿用缓存" and cache_date:
            parts.append(f"沿用缓存日期 {cache_date}")
        else:
            parts.append(f"今日卖家精灵 {today_status}")
    if trend_status:
        parts.append(f"趋势 {trend_status}")
    if persistent:
        parts.append(f"连续词 {persistent}")
    if stable_asins:
        parts.append(f"稳定竞品 {stable_asins}")
    if main_count:
        parts.append(f"主竞品 {main_count} 个")
    if comparability:
        parts.append(f"可比性 {comparability}/100")
    if scalable_status:
        parts.append(f"放量证据 {scalable_status}")
    if scalable_blockers:
        parts.append(f"拦截 {scalable_blockers}")
    if pressure:
        parts.append(f"竞品词压力 {pressure}")
    if missing:
        parts.append(f"缺口词 {missing}")
    if no_match:
        parts.append(f"广告无反查 {no_match}")
    if boundary:
        parts.append(boundary)
    elif own_status or pool_status or competitor_status:
        evidence_bits = [bit for bit in [own_status, pool_status, competitor_status] if bit]
        parts.append("证据 " + "、".join(evidence_bits[:3]))
    return "卖家精灵融合：" + "；".join(parts) if parts else ""


def _operation_has_external_evidence(row: dict[str, object]) -> bool:
    evidence_fields = [
        "frontend_status",
        "frontend_search_status",
        "seller_sprite_check_status",
        "competitor_discovery_status",
        "competitor_pool_status",
        "competitor_sellersprite_status",
        "amazon_search_validation_status",
        "product_level_conclusion",
        "competitor_keyword_pressure",
    ]
    weak_values = {"", "无缓存", "待补", "未入本轮队列", "卖家精灵证据不足", "竞品证据不足"}
    return any(str(row.get(field) or "").strip() not in weak_values for field in evidence_fields)


def _operation_evidence_profile(row: dict[str, object]) -> tuple[str, str, str, int]:
    if _operation_has_external_evidence(row):
        return (
            "强证据决策",
            "已结合前台、卖家精灵或竞品证据，可用于今天动作边界。",
            "strong",
            0,
        )
    final = str(row.get("final_decision") or "")
    if final == "DATA_INSUFFICIENT":
        return (
            "证据不足",
            "只能观察或补采集，不能做市场和竞品判断。",
            "insufficient",
            2,
        )
    inventory = str(row.get("inventory_constraint") or "")
    has_boundary_data = any(
        str(row.get(field) or "").strip()
        for field in [
            "ad_clicks",
            "ad_spend",
            "ad_orders",
            "total_orders",
            "natural_orders",
            "profit_before_ads_per_unit",
            "cost_status",
            "inventory_reason",
        ]
    )
    if inventory in {"OUT_OF_STOCK", "LOW_STOCK", "RESTOCK_RECOVERY", "REPLENISH_SOON"} or has_boundary_data:
        return (
            "运营边界",
            "基于库存、利润、订单和广告窗口，只用于控费、补货或保守跑。",
            "boundary",
            1,
        )
    return (
        "证据不足",
        "只能观察或补采集，不能做市场和竞品判断。",
        "insufficient",
        2,
    )


def _operation_display_sort_key(row: dict[str, object]) -> tuple[int, int, str, str]:
    _, _, _, profile_rank = _operation_evidence_profile(row)
    decision = str(row.get("final_decision") or "")
    decision_rank = {
        "FRONTEND_FIRST": 0,
        "WAIT_REVIEW": 1,
        "CONSERVATIVE_RUN": 2,
        "DATA_INSUFFICIENT": 3,
        "DO_NOT_TOUCH": 4,
    }.get(decision, 5)
    return profile_rank, decision_rank, str(row.get("marketplace") or ""), str(row.get("product_name") or "")


def _operation_action_boundary_html(allowed_text: str, blocked_text: str, review_label: str, review_text: str) -> str:
    allowed_main = _compact_semicolon_items(allowed_text, fallback="只观察", limit=2)
    blocked_main = _compact_semicolon_items(blocked_text, fallback="无强拦截", limit=2)
    review = f"{review_label}：{review_text}" if review_text else review_label
    return "".join(
        [
            '<div class="operation-action-line">',
            f'<span class="boundary-chip allow">{_inline_markup(allowed_main)}</span>',
            f'<span class="boundary-chip block">{_inline_markup(blocked_main)}</span>',
            f'<span class="boundary-chip review">{_inline_markup(review)}</span>',
            "</div>",
        ]
    )


def _operation_review_instruction(row: dict[str, object]) -> tuple[str, str]:
    raw_text = str(row.get("fusion_review_window") or "").strip()
    is_generic_wait = "补齐 7 天窗口" in raw_text and "刷新前台" in raw_text
    if not is_generic_wait:
        return _display_review_instruction(raw_text)

    if _operation_has_external_evidence(row):
        pressure = str(row.get("competitor_keyword_pressure") or "").strip()
        conclusion = str(row.get("product_level_conclusion") or "").strip()
        if pressure == "高" or conclusion == "暂停扩张":
            ad_orders = _num_from_text(row.get("ad_orders"))
            total_orders = _num_from_text(row.get("total_orders"))
            acos = _num_from_text(_first_present(row, "acos", "ACOS"))
            target_acos = _num_from_text(row.get("target_acos"))
            if (ad_orders or 0) <= 0 and (total_orders or 0) <= 0:
                return ("复查", "7 天看是否破零；没出单继续暂停扩张。")
            if acos is not None and target_acos is not None:
                acos_text = _format_operation_metric(acos, kind="percent")
                if acos <= target_acos:
                    return ("复查", f"7 天看广告单 {int(ad_orders or 0)}、ACOS {acos_text} 能否守住；前台不修不加预算。")
                return ("复查", f"7 天看 ACOS {acos_text} 能否回落；高压竞品词继续控费。")
            if ad_orders and ad_orders > 0:
                return ("复查", f"7 天看广告单 {int(ad_orders)} 是否延续；不稳定就控费。")
            return ("复查", "7 天看订单和竞品词变化；不达标继续控费。")
        return ("复查", "7 天看订单、ACOS 和竞品词变化。")

    inventory = str(row.get("inventory_constraint") or "")
    if inventory in {"OUT_OF_STOCK", "LOW_STOCK", "RESTOCK_RECOVERY", "REPLENISH_SOON"}:
        return ("复查", "补货或库存风险解除后，再评估加预算和竞品测试。")

    return ("补证后复查", "满 7 天或前台证据可用后再定动作。")


def _operation_decision_display(row: dict[str, str]) -> tuple[str, str]:
    decision = str(row.get("final_decision") or "")
    label = str(row.get("final_decision_label") or decision or "N/A")
    frontend_status = str(row.get("frontend_status") or "").strip()
    frontend_label = str(row.get("frontend_auto_conclusion_label") or "").strip()
    fusion_gate = str(row.get("fusion_action_gate") or "").strip()
    evidence_text = "；".join(
        str(row.get(field) or "")
        for field in [
            "frontend_evidence_tier",
            "frontend_evidence_display_tier",
            "frontend_evidence_state",
            "frontend_evidence_audit_detail",
            "frontend_search_findings",
            "fusion_reason",
            "decision_reason",
            "evidence_used",
        ]
    )
    checked = (
        frontend_status and frontend_status not in {"待前台检查", "需要截图确认", "待截图确认"}
    ) or any(token in evidence_text for token in ["前台证据", "今日读取", "沿用", "搜索页已读", "竞品", "质量 "])
    if decision == "FRONTEND_FIRST" and checked:
        if fusion_gate == "fix_both":
            return "止损+修前台", "前台已查；广告先止损，同时处理前台弱项"
        if fusion_gate == "fix_frontend_first":
            return "先修前台", "前台已查；先处理价格、Coupon、Buy Box、评分评论等弱项"
        if fusion_gate == "tighten_ads_first":
            return "先收广告", "前台已查；优先处理广告流量"
        if fusion_gate == "collect_evidence_only" or frontend_label == "自动证据不足，不能用于强诊断":
            return "补证据", "前台证据不足；补齐证据后再做强动作"
        if frontend_label == "明确前台劣势":
            return "修前台+保守跑", "前台已查；按弱项处理，广告不扩预算"
        return "保守跑", f"前台已查；{frontend_label or '继续按证据判断'}"
    return label, label


def _render_product_operation_cards(rows: list[dict[str, str]], coverage_summary: dict[str, object] | None = None, limit: int = 8) -> str:
    if not rows:
        return (
            '<section class="section-card" id="product-operation-cards">'
            '<h2>产品级结论</h2>'
            '<div class="card-block decision-summary">'
            '<strong>系统结论</strong>'
            '<div class="decision-summary-list">'
            '<div><span>系统结论</span><strong>当前没有生成产品级运营结论</strong></div>'
            '<div><span>融合诊断</span><strong>无待处理产品卡</strong></div>'
            '</div>'
            '</div>'
            '<p class="subtle">广告后台动作以下方广告工作台为准。</p>'
            '</section>'
        )
    decision_rows = [row for row in rows if str(row.get("final_decision") or "") not in {"EXECUTE_TODAY"}]
    hidden_executed = len(rows) - len(decision_rows)
    display_rows = sorted(decision_rows or rows, key=_operation_display_sort_key)
    action_labels = {
        "bid_up": "小幅加竞价",
        "bid_down": "降竞价",
        "broad_scale": "放量",
        "budget_up": "加预算",
        "create_exact_low_budget": "低预算精准测试",
        "negative_exact": "否定精准",
        "observe": "只观察",
        "pause": "暂停",
    }
    blocked_labels = {
        "bid_up": "加竞价",
        "bid_down": "降竞价",
        "broad_scale": "放量",
        "budget_up": "加预算",
        "create_exact_low_budget": "低预算精准测试",
        "negative_exact": "否定精准",
        "pause": "暂停",
    }
    parts = [
        '<section class="section-card" id="product-operation-cards">',
        '<div class="ad-section-header"><h2>产品级结论</h2><span class="status-badge status-muted">按产品决策</span></div>',
        '<p class="operation-section-note">只展示需要你判断、复查或拦截的产品。已执行产品不在这里占位，广告后台动作以下方广告工作台为准。</p>',
        _render_frontend_coverage_strip(coverage_summary or {}),
        '<div class="operation-wide-list">',
    ]
    for row in display_rows[:limit]:
        marketplace = str(row.get("marketplace") or "")
        sku = str(row.get("sku") or "").strip()
        asin = str(row.get("asin") or "").strip().upper()
        decision = str(row.get("final_decision") or "")
        label, label_title = _operation_decision_display(row)
        tag = "tag tag-green" if decision in {"EXECUTE_TODAY", "SMALL_SCALE_ALLOWED"} else "tag tag-yellow" if decision in {"FRONTEND_FIRST", "WAIT_REVIEW", "CONSERVATIVE_RUN"} else "tag tag-gray"
        metrics = [
            ("14天点击", _format_operation_metric(row.get("ad_clicks"))),
            ("广告花费", _format_operation_metric(row.get("ad_spend"), marketplace, "money")),
            ("广告单", _format_operation_metric(row.get("ad_orders"))),
            ("总单", _format_operation_metric(row.get("total_orders"))),
            ("自然单", _format_operation_metric(row.get("natural_orders"))),
            ("ACOS", _format_operation_metric(_first_present(row, "acos", "ACOS"), kind="percent")),
            ("目标ACOS", _format_operation_metric(row.get("target_acos"), kind="percent")),
        ]
        metric_html = "".join(
            f'<span class="operation-metric-chip"><span>{html.escape(label_text)}</span><strong>{html.escape(value)}</strong></span>'
            for label_text, value in metrics
        )
        audit_detail = str(row.get("frontend_evidence_audit_detail") or "").strip()
        if not audit_detail:
            audit_reasons = row.get("frontend_evidence_audit_reasons")
            if isinstance(audit_reasons, list):
                audit_detail = "；".join(str(reason).strip() for reason in audit_reasons if str(reason).strip())
            else:
                audit_detail = str(audit_reasons or "").strip()
        frontend_bits = [
            str(row.get("frontend_status") or "无前台队列"),
            str(row.get("frontend_auto_conclusion_label") or ""),
            str(row.get("frontend_evidence_display_tier") or row.get("frontend_evidence_tier") or ""),
            f"质量 {row.get('frontend_evidence_quality_score')}" if row.get("frontend_evidence_quality_score") not in (None, "") else "",
            audit_detail,
            str(row.get("frontend_freshness") or ""),
            f"市场调查完整度：{row.get('market_survey_completeness_score')}/100，{_market_survey_level_label(row.get('market_survey_completeness_level'))}" if row.get("market_survey_completeness_score") not in (None, "") else "",
            f"市场调查证据层级：{_market_survey_level_label(row.get('market_survey_decision_evidence_tier'))}" if row.get("market_survey_decision_evidence_tier") else "",
            f"市场调查缺口：{row.get('market_survey_missing_parts')}" if row.get("market_survey_missing_parts") else "",
            f"建议补抓：{row.get('market_survey_recommended_fetch_steps')}" if row.get("market_survey_recommended_fetch_steps") else "",
            f"卖家精灵产品结论：{row.get('product_level_conclusion')}" if row.get("product_level_conclusion") else "",
            f"竞品发现：{row.get('competitor_discovery_status')}" if row.get("competitor_discovery_status") else "",
            f"竞品发现页：{_seller_source_label(row.get('competitor_discovery_source_page'))}" if row.get("competitor_discovery_source_page") else "",
            f"竞品发现错误：{row.get('competitor_discovery_error')}" if row.get("competitor_discovery_error") else "",
            f"竞品池：{row.get('competitor_pool_status')}" if row.get("competitor_pool_status") else "",
            f"竞品来源：{_seller_source_label(row.get('competitor_discovery_source'))}" if row.get("competitor_discovery_source") else "",
            f"竞品池置信度：{_confidence_label(row.get('competitor_pool_confidence'))}" if row.get("competitor_pool_confidence") else "",
            f"竞品来源词：{row.get('competitor_source_keywords')}" if row.get("competitor_source_keywords") else "",
            f"竞品剔除：{row.get('competitor_rejected_count')}，{row.get('competitor_rejection_reasons')}" if row.get("competitor_rejected_count") else "",
            f"Amazon 搜索验证：{row.get('amazon_search_validation_status')}" if row.get("amazon_search_validation_status") else "",
            f"卖家精灵自己 ASIN：{row.get('seller_sprite_check_status')}" if row.get("seller_sprite_check_status") else "",
            f"竞品 ASIN 反查：{row.get('competitor_sellersprite_status')}" if row.get("competitor_sellersprite_status") else "",
            f"竞品词压力：{row.get('competitor_keyword_pressure')}" if row.get("competitor_keyword_pressure") else "",
            f"竞品重叠词：{row.get('competitor_overlap_keywords')}" if row.get("competitor_overlap_keywords") else "",
            f"竞品共同词：{row.get('competitor_shared_keywords')}" if row.get("competitor_shared_keywords") else "",
            f"竞品缺口词：{row.get('own_missing_competitor_keywords')}" if row.get("own_missing_competitor_keywords") else "",
            f"广告无反查证据词：{row.get('own_ad_terms_not_in_sellersprite')}" if row.get("own_ad_terms_not_in_sellersprite") else "",
        ]
        search_status = str(row.get("frontend_search_status") or "").strip()
        search_findings = str(row.get("frontend_search_findings") or "").strip()
        search_partial = _boolish(row.get("frontend_search_partial_evidence")) or search_status == "已读取部分结果"
        if search_partial:
            frontend_bits.append(f"搜索页：{search_status or '部分读取'}")
        elif search_status:
            frontend_bits.append(f"搜索页：{search_status}")
        if search_findings and any(token in search_findings for token in ["未稳定", "部分", "缺失", "竞品", "搜索页"]):
            frontend_bits.append(search_findings)
        frontend_text = "；".join(bit for bit in frontend_bits if bit)
        cost_bits = [
            str(row.get("cost_status") or ""),
            str(row.get("cost_key_evidence") or ""),
            str(row.get("inventory_constraint") or ""),
            str(row.get("inventory_reason") or ""),
        ]
        cost_text = "；".join(bit for bit in cost_bits if bit) or "当前无成本或库存拦截。"
        decision_reason = str(row.get("decision_reason") or row.get("fusion_reason") or "N/A")
        seller_decision = _seller_sprite_decision_sentence(row)
        if seller_decision and "卖家精灵融合" not in decision_reason:
            decision_reason = f"{seller_decision}；{decision_reason}"
        ad_diagnostic = str(row.get("ad_diagnostic_summary") or "")
        allowed_text = _action_list_text(row.get("today_allowed_actions"), action_labels)
        blocked_text = _action_list_text(row.get("today_blocked_actions"), blocked_labels)
        review_label, review_text = _operation_review_instruction(row)
        action_boundary_html = _operation_action_boundary_html(allowed_text, blocked_text, review_label, review_text)
        decision_summary_html = _render_decision_summary_block(row)
        seller_summary_html = _seller_sprite_product_summary(row)
        profile_label, profile_note, profile_class, _ = _operation_evidence_profile(row)
        profile_html = (
            f'<div class="operation-profile {profile_class}">'
            f'<strong>{html.escape(profile_label)}</strong>'
            f'<span>{html.escape(profile_note)}</span>'
            "</div>"
        )
        parts.append(
            "\n".join(
                [
                    (
                        '<article class="operation-wide-card"'
                        f' data-product-decision-marketplace="{html.escape(marketplace, quote=True)}"'
                        f' data-product-decision-sku="{html.escape(sku, quote=True)}"'
                        f' data-product-decision-asin="{html.escape(asin, quote=True)}"'
                        f'{_product_decision_contract_attrs(row)}>'
                    ),
                    '<div class="operation-wide-head">',
                    '<div class="operation-wide-product product-cell">',
                    f'<strong>{html.escape(marketplace)}｜{html.escape(str(row.get("product_name") or row.get("asin") or "N/A"))}</strong>',
                    f'<div class="card-meta">{_product_meta_html(row)}</div>',
                    profile_html,
                    '</div>',
                    f'<div class="operation-wide-verdict"><span class="{tag}" title="{html.escape(label_title)}">{html.escape(label)}</span></div>',
                    f'<div class="operation-wide-metrics metric-cell"><div class="operation-metric-list">{metric_html}</div></div>',
                    f'<div class="operation-wide-actions action-cell">{action_boundary_html}</div>',
                    '</div>',
                    decision_summary_html,
                    '<div class="operation-wide-body">',
                    f'<div class="operation-wide-market">{seller_summary_html}</div>',
                    '<div class="operation-wide-reason reason-cell">',
                    f'<p class="operation-main-reason">{_inline_markup(decision_reason)}</p>',
                    '<details class="operation-more">',
                    '<summary>展开证据</summary>',
                    f'<p><strong>广告诊断</strong>：{_inline_markup(ad_diagnostic or decision_reason)}</p>',
                    f'<p><strong>前台证据</strong>：{_inline_markup(frontend_text)}</p>',
                    f'<p><strong>成本 / 库存</strong>：{_inline_markup(cost_text)}</p>',
                    f'<div><strong>卡内广告止损项</strong>{_render_operation_ad_actions(row.get("ad_action_items") or [])}</div>',
                    '</details>',
                    '</div>',
                    '</div>',
                    '</article>',
                ]
            )
        )
    parts.append("</div>")
    hidden_other = max(0, len(display_rows) - limit)
    notes = []
    if hidden_other:
        notes.append(f"另有 {hidden_other} 个产品判断见 Excel 或单站报告")
    if hidden_executed:
        notes.append(f"已隐藏 {hidden_executed} 个已执行产品")
    if notes:
        parts.append(f'<p class="subtle">{"；".join(notes)}。</p>')
    parts.append("</section>")
    return "".join(parts)


def _render_inventory_replenishment_cards(rows: list[dict[str, str]], limit: int = 6) -> str:
    if not rows:
        return '<p class="subtle">当前没有库存补货记录。</p>'
    status_order = {
        'OUT_OF_STOCK': 0,
        'LOW_STOCK': 1,
        'RESTOCK_RECOVERY': 2,
        'REPLENISH_SOON': 3,
        'UNKNOWN': 4,
        'HEALTHY': 5,
    }
    def inventory_card_priority(row: dict[str, str]) -> tuple[object, ...]:
        cover = _optional_num_from_text(row.get('days_of_cover'))
        qty = _optional_num_from_text(row.get('recommended_reorder_qty'))
        return (
            status_order.get(str(row.get('stock_risk_level') or ''), 9),
            cover if cover is not None else 99999,
            -qty if qty is not None else 0,
            str(row.get('marketplace') or ''),
            str(row.get('product_name') or ''),
        )

    sorted_rows = sorted(
        rows,
        key=inventory_card_priority,
    )
    parts = ['<div class="action-grid">']
    for row in sorted_rows[:limit]:
        level = str(row.get('stock_risk_level') or '')
        label = str(row.get('stock_status_label') or level or 'N/A')
        if level == 'OUT_OF_STOCK':
            tag = 'tag tag-red'
        elif level in {'LOW_STOCK', 'RESTOCK_RECOVERY', 'REPLENISH_SOON'}:
            tag = 'tag tag-yellow'
        elif level == 'HEALTHY':
            tag = 'tag tag-green'
        else:
            tag = 'tag tag-gray'
        available_value = row.get('available_stock') if row.get('available_stock') not in ('', None) else row.get('current_inventory')
        available_metric = f"可用库存 {available_value}" if available_value not in ('', None) else "可用库存缺失"
        cover_value = row.get('days_of_cover')
        cover_metric = f"覆盖 {cover_value}" if cover_value not in ('', None) else "覆盖缺失"
        metrics = [
            available_metric,
            f"FBA {row.get('fba_stock')}" if row.get("fba_stock") not in ("", None) else "",
            cover_metric,
            f"总提前期 {row.get('total_lead_time_days')}",
            f"目标覆盖 {row.get('target_cover_days')}",
        ]
        parts.append(
            "\n".join(
                [
                    '<article class="work-card">',
                    '<div class="card-head">',
                    '<div>',
                    f'<h3 class="card-title">{html.escape(str(row.get("marketplace") or ""))}｜{html.escape(str(row.get("product_name") or row.get("asin") or "N/A"))}</h3>',
                    f'<div class="card-meta">{_product_meta_html(row)}</div>',
                    '</div>',
                    f'<span class="{tag}">{html.escape(label)}</span>',
                    '</div>',
                    '<div class="metric-row">' + ''.join(f'<span class="metric-badge">{html.escape(metric)}</span>' for metric in metrics if metric) + '</div>',
                    f'<div class="card-block"><strong>风险原因</strong><p>{_inline_markup(str(row.get("stock_risk_reason") or "N/A"))}</p></div>',
                    f'<div class="card-block"><strong>补货建议</strong><p>{_inline_markup(str(row.get("replenishment_advice") or "N/A"))}</p></div>',
                    '</article>',
                ]
            )
        )
    parts.append('</div>')
    if len(sorted_rows) > limit:
        parts.append(f'<p class="subtle">另有 {len(sorted_rows) - limit} 条库存记录见 Excel。</p>')
    return ''.join(parts)


def _render_listing_review_cards(
    rows: list[dict[str, str]],
    limit: int = 5,
    frontend_lookup: dict[tuple[str, str, str], dict[str, str]] | None = None,
) -> str:
    if not rows:
        return '<p class="subtle">当前没有触发 Listing 待人工确认的产品。</p>'
    parts = ['<div class="action-grid">']
    for row in rows[:limit]:
        priority = str(row.get("priority") or "P1") or "P1"
        css = "p0" if priority == "P0" else "p1"
        evidence = row.get("当前证据") or row.get("关键证据") or "N/A"
        confirmed = str(row.get("confirmed_status") or "待确认")
        anomaly_type = _listing_basis_label(row)
        problem_line = _listing_problem_line(row, evidence)
        next_step = _listing_conservative_action(row)
        parts.append(
            "\n".join(
                [
                    f'<article class="work-card {css}">',
                    '<div class="card-head">',
                    "<div>",
                    f'<h3 class="card-title">{html.escape(priority)}｜{html.escape(str(row.get("产品") or "N/A"))}</h3>',
                    f'<div class="card-meta">{_product_meta_html(row, asin_key="ASIN", sku_key="SKU", marketplace_key="站点")}</div>',
                    "</div>",
                    f'<div><span class="tag tag-yellow">Listing 待人工确认</span> <span class="{_confirmed_tag(confirmed)}">{html.escape(confirmed)}</span></div>',
                    "</div>",
                    f'<div class="card-block"><strong>最可能异常方向</strong><p>{_inline_markup(anomaly_type)}</p></div>',
                    f'<div class="card-block"><strong>证据</strong><p>{_inline_markup(problem_line)}</p></div>',
                    _render_frontend_evidence_block(row, frontend_lookup),
                    f'<div class="card-block"><strong>保守动作</strong><p>{_inline_markup(next_step)}</p></div>',
                    "</article>",
                ]
            )
        )
    parts.append("</div>")
    return "".join(parts)


def _render_local_data_submit_tool() -> str:
    cost_example = "\n".join(
        [
            '<div class="submit-example">',
            '<table><caption>product_cost_config.xlsx 示例，sheet 名：product_cost_config</caption>',
            '<thead><tr><th>marketplace</th><th>currency</th><th>sku</th><th>asin</th><th>product_name</th><th>selling_price</th><th>purchase_cost_rmb</th><th>first_leg_cost_rmb</th><th>suggested_target_acos</th></tr></thead>',
            '<tbody><tr><td>DE</td><td>EUR</td><td>SKU-DE-001</td><td>B0EXAMPLE1</td><td>Demo cable ties</td><td>24.99</td><td>52.00</td><td>18.00</td><td>0.12</td></tr></tbody>',
            "</table>",
            "</div>",
        ]
    )
    alias_example = "\n".join(
        [
            '<div class="submit-example">',
            '<table><caption>sku_alias_map.xlsx 示例，sheet 名：Sheet1</caption>',
            '<thead><tr><th>marketplace</th><th>source_sku</th><th>canonical_sku</th><th>asin</th><th>reason</th></tr></thead>',
            '<tbody><tr><td>DE</td><td>SKU-DE-001-FBA</td><td>SKU-DE-001</td><td>B0EXAMPLE1</td><td>ERP/FBA SKU alias</td></tr></tbody>',
            "</table>",
            "</div>",
        ]
    )
    config_tools = "\n".join(
        [
            '<div class="config-submit-list" data-config-submit-panel>',
            '<div class="config-submit-item cost">',
            '<div class="config-submit-header"><strong>产品和成本配置</strong><span class="submit-pill risk">高风险</span></div>',
            '<p class="subtle">上传 product_cost_config.xlsx。系统先保存到 data/config_review 并生成成本差异审计；确认后再应用到正式 config。</p>',
            _render_collapsed_block("查看成本表示例", cost_example, "只在整理配置文件时展开。"),
            '<form class="local-submit-form" data-config-submit-form data-config-kind="cost">',
            '<input type="file" name="file" accept=".xlsx">',
            '<button class="button-link primary" type="submit">上传成本表并审计</button>',
            '<button class="button-link danger" type="button" data-apply-config="cost">确认应用成本表</button>',
            "</form>",
            "</div>",
            '<div class="config-submit-item">',
            '<div class="config-submit-header"><strong>SKU 别名映射</strong><span class="submit-pill">XLSX</span></div>',
            '<p class="subtle">上传 sku_alias_map.xlsx。用于把广告或 ERP 中的别名 SKU 归一到正式 SKU；应用前会检查必填列。</p>',
            _render_collapsed_block("查看别名表示例", alias_example, "只在整理 SKU 归一规则时展开。"),
            '<form class="local-submit-form" data-config-submit-form data-config-kind="alias">',
            '<input type="file" name="file" accept=".xlsx">',
            '<button class="button-link primary" type="submit">上传别名表并检查</button>',
            '<button class="button-link secondary" type="button" data-apply-config="alias">应用别名表</button>',
            "</form>",
            "</div>",
            "</div>",
        ]
    )
    return "\n".join(
        [
            '<section class="section-card local-submit-panel is-collapsed" id="local-data-submit">',
            '<div>',
            '<div class="ad-section-header"><div><p class="local-submit-kicker">LOCAL WORKFLOW</p><h2>提交今日数据</h2></div><button class="collapsible-toggle" type="button" data-collapse-toggle>展开</button></div>',
            '<p class="local-submit-lead">上传日报会先进 inbox 校验，再刷新报告；成本和 SKU 配置仍走审核区。</p>',
            '<div class="collapsible-body">',
            '<div class="daily-submit-card">',
            '<div class="config-submit-header"><strong>日报文件</strong><span class="submit-pill">CSV/XLSX</span></div>',
            '<p class="subtle">上传广告、ERP、Seller Central 自定义分析文件；上传通过后自动刷新报告。</p>',
            '<form class="local-submit-form" data-local-submit-form>',
            '<input type="file" name="files" multiple accept=".csv,.xlsx">',
            '<div class="local-file-list" data-local-file-list>可一次多选多个 CSV / XLSX 文件，按住 Command 或 Shift 选择。</div>',
            '<button class="button-link primary" type="submit">上传并刷新日报</button>',
            "</form>",
            '<p class="subtle">限制：仅本机 127.0.0.1 服务；单个文件 50MB；可一次多选；无法识别的文件会阻塞日报刷新。</p>',
            "</div>",
            _render_collapsed_block("配置文件审核", config_tools, "成本表和 SKU 别名整理时再展开；日报上传不受影响。"),
            "</div>",
            '<div class="local-submit-status-panel">',
            '<div class="local-submit-status" data-local-submit-status>日报：暂无运行任务；上传通过后自动刷新报告。</div>',
            '<div class="local-submit-status" data-config-submit-status>配置：成本表和 SKU 别名只在展开后处理，应用前必须审计。</div>',
            "</div>",
            "</div>",
            "</section>",
        ]
    )

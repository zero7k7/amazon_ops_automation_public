from __future__ import annotations

import re
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook

CUSTOM_DATA_DIRNAME = "data/raw_amazon_custom"
TRAFFIC_PREFIX = "traffic_sales"
QUERY_PREFIX = "search_query_performance"


def _site_enhanced_dir(custom_dir: Path, marketplace: str) -> Path:
    return Path(custom_dir) / marketplace.upper()


def _recognized_report_files(directory: Path) -> list[Path]:
    if not directory.exists():
        return []
    return [
        path
        for path in sorted(directory.iterdir())
        if path.is_file() and path.suffix.lower() in {".xlsx", ".xls", ".csv"}
    ]


def _resolve_enhanced_scan_dir(custom_dir: Path, marketplace: str) -> tuple[Path, Path, list[Path], list[Path]]:
    custom_dir.mkdir(parents=True, exist_ok=True)
    site_dir = _site_enhanced_dir(custom_dir, marketplace)
    site_dir.mkdir(parents=True, exist_ok=True)
    site_files = _recognized_report_files(site_dir)
    root_files = _recognized_report_files(custom_dir)
    scan_dir = site_dir if site_files else custom_dir
    return site_dir, scan_dir, site_files, root_files


def _target_folder_text(custom_dir: Path, marketplace: str) -> str:
    return str(_site_enhanced_dir(custom_dir, marketplace)).replace("\\", "/") + "/"


def _concat_frames_preserving_columns(frames: list[pd.DataFrame]) -> pd.DataFrame:
    column_order: list[object] = []
    prepared: list[pd.DataFrame] = []
    for frame in frames:
        if frame.empty:
            continue
        for column in frame.columns:
            if column not in column_order:
                column_order.append(column)
        trimmed = frame.dropna(axis=1, how="all")
        if not trimmed.empty:
            prepared.append(trimmed)
    if not prepared:
        return pd.DataFrame(columns=column_order)
    result = pd.concat(prepared, ignore_index=True)
    for column in column_order:
        if column not in result.columns:
            result[column] = pd.NA
    return result[column_order]


@dataclass
class EnhancedDataBundle:
    status: dict[str, object]
    traffic_sales_detail: pd.DataFrame
    search_query_detail: pd.DataFrame
    natural_decline_diagnostics: pd.DataFrame
    search_query_opportunities: pd.DataFrame
    request_rows: pd.DataFrame
    request_markdown: str


_DATE_RANGE_PATTERN = re.compile(
    r"(\d{4}[./年]\d{1,2}[./月]\d{1,2}[日]?)\s*[-~—–至]\s*(\d{4}[./年]\d{1,2}[./月]\d{1,2}[日]?)"
)
_ASIN_PATTERN = re.compile(r"^B0[A-Z0-9]{8,}$", re.IGNORECASE)
_WEEK_PATTERN = re.compile(r"\bWeek\s*(\d{1,2})\b", re.IGNORECASE)


def compute_recent_prior_ranges(common_end: date) -> dict[str, tuple[date, date]]:
    recent_end = common_end
    recent_start = recent_end - timedelta(days=6)
    prior_end = recent_start - timedelta(days=1)
    prior_start = prior_end - timedelta(days=6)
    return {"recent": (recent_start, recent_end), "prior": (prior_start, prior_end)}


def expected_enhanced_filenames(marketplace: str, common_end: date) -> dict[str, str]:
    marketplace_lower = marketplace.lower()
    periods = compute_recent_prior_ranges(common_end)
    expected: dict[str, str] = {}
    for period, (start_date, end_date) in periods.items():
        expected[f"traffic_sales_{period}"] = (
            f"{TRAFFIC_PREFIX}_{marketplace_lower}_{period}_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
        )
        expected[f"search_query_{period}"] = (
            f"{QUERY_PREFIX}_{marketplace_lower}_{period}_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
        )
    return expected


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _extract_formula_display(value: object) -> object:
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text.startswith("="):
        return text
    hyperlink_match = re.match(r'=HYPERLINK\("([^"]*)",\s*"([^"]*)"\)', text, flags=re.IGNORECASE)
    if hyperlink_match:
        url, display = hyperlink_match.groups()
        display = display.strip()
        url = url.strip()
        if display:
            return display
        if "/dp/" in url:
            return url.split("/dp/")[-1].split("?")[0].strip()
        return url
    return text


def _read_first_sheet(path: Path) -> tuple[str, list[list[object]]]:
    workbook = load_workbook(path, data_only=False)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    return sheet.title, rows


def _classify_report_file(path: Path) -> tuple[str, str, str]:
    name = path.name.lower()
    if TRAFFIC_PREFIX in name:
        data_type = "traffic_sales"
    elif QUERY_PREFIX in name:
        data_type = "search_query_performance"
    else:
        data_type = "unknown"

    marketplace = "unknown"
    period_hint = "unknown"
    match = re.search(r"_(uk|us|de)_(recent|prior|wow|compare|unknown)_", name)
    if match:
        marketplace = match.group(1).upper()
        period_hint = match.group(2).lower()
    else:
        alt_match = re.search(r"_(uk|us|de)_", name)
        if alt_match:
            marketplace = alt_match.group(1).upper()
    return marketplace, data_type, period_hint


def _find_by_keywords(columns: list[str], include: Iterable[str], exclude: Iterable[str] = ()) -> str | None:
    include_tokens = [token.lower() for token in include]
    exclude_tokens = [token.lower() for token in exclude]
    for column in columns:
        name = str(column).strip()
        lower = name.lower()
        if all(token in lower for token in include_tokens) and not any(token in lower for token in exclude_tokens):
            return name
    return None


def _find_all_by_keywords(columns: list[str], include: Iterable[str], exclude: Iterable[str] = ()) -> list[str]:
    include_tokens = [token.lower() for token in include]
    exclude_tokens = [token.lower() for token in exclude]
    matched: list[str] = []
    for column in columns:
        name = str(column).strip()
        lower = name.lower()
        if all(token in lower for token in include_tokens) and not any(token in lower for token in exclude_tokens):
            matched.append(name)
    return matched


def _parse_date_like_text(text: str) -> date | None:
    normalized = text.strip()
    normalized = normalized.replace("年", "-").replace("月", "-").replace("日", "")
    normalized = normalized.replace("?", "-").replace("?", "-").replace("?", "")
    normalized = normalized.replace(".", "-").replace("/", "-")
    for fmt in ("%Y-%m-%d",):
        try:
            return datetime.strptime(normalized, fmt).date()
        except ValueError:
            continue
    parts = [part for part in normalized.split("-") if part]
    if len(parts) == 3:
        try:
            return date(int(parts[0]), int(parts[1]), int(parts[2]))
        except ValueError:
            return None
    return None


def _extract_header_date_range(text: str) -> tuple[date | None, date | None]:
    match = _DATE_RANGE_PATTERN.search(text)
    if not match:
        return None, None
    start = _parse_date_like_text(match.group(1))
    end = _parse_date_like_text(match.group(2))
    return start, end


def _week_date_range(year: int, week: int) -> tuple[date, date]:
    saturday = date.fromisocalendar(year, week, 6)
    return saturday - timedelta(days=6), saturday


def _expected_week_pair(common_end: date | None) -> tuple[int | None, int | None]:
    if not common_end:
        return None, None
    recent_week = max(common_end.isocalendar().week - 1, 1)
    return recent_week, max(recent_week - 1, 1)


def _extract_week_numbers(text: str) -> list[int]:
    return [int(match.group(1)) for match in _WEEK_PATTERN.finditer(text)]


def _detect_data_type_from_header(columns: list[str], preview_rows: list[list[object]]) -> str:
    text = " | ".join([str(column) for column in columns] + [str(cell) for row in preview_rows[:3] for cell in row if cell is not None])
    lower = text.lower()
    if "搜索查询" in text or "search query" in lower:
        return "search_query_performance"
    traffic_tokens = ["转化率", "推荐报价浏览量", "推荐报价率", "已订购商品数量", "已发货商品数量"]
    if any(token in text for token in traffic_tokens):
        return "traffic_sales"
    return "unknown"


def _extract_date_ranges_from_header(columns: list[str], preview_rows: list[list[object]]) -> list[tuple[date, date]]:
    ranges: list[tuple[date, date]] = []
    seen: set[tuple[date, date]] = set()
    for value in [*columns, *[cell for row in preview_rows[:5] for cell in row if cell is not None]]:
        start, end = _extract_header_date_range(str(value))
        if start and end and (start, end) not in seen:
            seen.add((start, end))
            ranges.append((start, end))
    return sorted(ranges, key=lambda item: item[1])


def _extract_week_pair(columns: list[str], preview_rows: list[list[object]]) -> tuple[int | None, int | None]:
    weeks: set[int] = set()
    for value in [*columns, *[cell for row in preview_rows[:5] for cell in row if cell is not None]]:
        weeks.update(_extract_week_numbers(str(value)))
    if not weeks:
        return None, None
    recent = max(weeks)
    prior = max([week for week in weeks if week < recent], default=recent - 1)
    return recent, prior


def _total_signature(preview_rows: list[list[object]]) -> tuple[object, ...]:
    for row in preview_rows[1:8]:
        if row and _clean_text(row[0]).upper() == "TOTAL":
            values = []
            for cell in row[3:]:
                numeric = _coerce_numeric(cell)
                if numeric is not None:
                    values.append(round(numeric, 4))
            return tuple(values[:8])
    return ()


def _detect_header_metadata(
    path: Path,
    columns: list[str],
    preview_rows: list[list[object]],
    marketplace: str,
    filename_marketplace: str,
    filename_data_type: str,
    filename_period_hint: str,
    common_end: date | None,
) -> dict[str, object]:
    header_data_type = _detect_data_type_from_header(columns, preview_rows)
    data_type = header_data_type if header_data_type != "unknown" else filename_data_type
    detected_from = "header" if header_data_type != "unknown" else ("filename" if filename_data_type != "unknown" or filename_period_hint != "unknown" else "fallback")
    file_marketplace = filename_marketplace if filename_marketplace != "unknown" else marketplace.upper()

    date_ranges = _extract_date_ranges_from_header(columns, preview_rows)
    recent_week, prior_week = _extract_week_pair(columns, preview_rows)
    format_type = _detect_format_type(columns, preview_rows[:5] if preview_rows else [])
    period_hint = filename_period_hint
    recent_start = recent_end = prior_start = prior_end = None
    freshness = "unknown"

    if len(date_ranges) >= 2:
        prior_start, prior_end = date_ranges[-2]
        recent_start, recent_end = date_ranges[-1]
        format_type = "wow"
        period_hint = "recent_vs_prior"
        detected_from = "header"
        if common_end and recent_end >= common_end - timedelta(days=2):
            freshness = "fresh"
        else:
            freshness = "dated"
    elif len(date_ranges) == 1:
        recent_start, recent_end = date_ranges[0]
        period_hint = "recent" if period_hint == "unknown" else period_hint
        detected_from = "header"
        freshness = "fresh" if common_end and recent_end >= common_end - timedelta(days=2) else "dated"
    elif recent_week:
        year = common_end.year if common_end else date.today().year
        recent_start, recent_end = _week_date_range(year, recent_week)
        prior_start, prior_end = _week_date_range(year, prior_week or recent_week - 1)
        format_type = "wow"
        period_hint = f"week{recent_week}_vs_week{prior_week or recent_week - 1}"
        detected_from = "header"
        expected_recent_week, _ = _expected_week_pair(common_end)
        if data_type == "traffic_sales":
            freshness = "fresh" if expected_recent_week and recent_week >= expected_recent_week else "stale"
        else:
            freshness = "weekly_lag"
    elif period_hint == "unknown":
        period_hint = "unknown"

    detected_range = ""
    if recent_start and recent_end and prior_start and prior_end:
        detected_range = f"{recent_start.isoformat()} ~ {recent_end.isoformat()} vs {prior_start.isoformat()} ~ {prior_end.isoformat()}"
    elif recent_start and recent_end:
        detected_range = f"{recent_start.isoformat()} ~ {recent_end.isoformat()}"

    return {
        "file_name": path.name,
        "file_path": str(path),
        "marketplace": file_marketplace,
        "data_type": data_type,
        "period_hint": period_hint,
        "format_type": format_type,
        "detected_from": detected_from,
        "detected_date_range": detected_range,
        "freshness": freshness,
        "recent_start": recent_start.isoformat() if recent_start else "",
        "recent_end": recent_end.isoformat() if recent_end else "",
        "prior_start": prior_start.isoformat() if prior_start else "",
        "prior_end": prior_end.isoformat() if prior_end else "",
        "recent_week": recent_week or "",
        "prior_week": prior_week or "",
        "total_signature": _total_signature(preview_rows),
    }


def _detect_format_type(columns: list[str], preview_rows: list[list[object]]) -> str:
    combined_text = " | ".join([str(col) for col in columns] + [str(cell) for row in preview_rows for cell in row if cell is not None])
    if "周环比" in combined_text or "WOW" in combined_text.upper() or _WEEK_PATTERN.search(combined_text):
        return "wow"
    date_ranges: set[tuple[date | None, date | None]] = set()
    for column in columns:
        start, end = _extract_header_date_range(str(column))
        if start and end:
            date_ranges.add((start, end))
    if len(date_ranges) >= 2:
        return "compare"
    for row in preview_rows[:5]:
        for cell in row:
            if cell is None:
                continue
            start, end = _extract_header_date_range(str(cell))
            if start and end:
                date_ranges.add((start, end))
    if len(date_ranges) >= 2:
        return "compare"
    return "single"


def _choose_metric_columns(columns: list[str], keyword: str) -> tuple[str | None, str | None, str | None]:
    current_candidates = _find_all_by_keywords(columns, [keyword], exclude=["周环比", "wow", "WOW"])
    change_candidates = _find_all_by_keywords(columns, [keyword], exclude=[])
    change_candidates = [c for c in change_candidates if ("周环比" in c or "WOW" in c.upper() or "环比" in c)]

    current_col = None
    prior_col = None
    if len(current_candidates) == 1:
        current_col = current_candidates[0]
    elif len(current_candidates) > 1:
        dated: list[tuple[date | None, str]] = []
        for column in current_candidates:
            _, end = _extract_header_date_range(column)
            dated.append((end, column))
        dated.sort(key=lambda item: (item[0] is None, item[0] or date.min))
        current_col = dated[-1][1]
        if len(dated) >= 2:
            prior_col = dated[-2][1]

    change_col = change_candidates[0] if change_candidates else None
    return current_col, prior_col, change_col


def _coerce_numeric(value: object) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, str):
        text = value.strip().replace(",", "").replace("￥", "").replace("£", "").replace("$", "")
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _to_ratio(value: float | None) -> float | None:
    if value is None:
        return None
    return value


def _derive_prior_from_change(current: float | None, change_pct: float | None) -> float | None:
    if current is None or change_pct is None:
        return None
    denominator = 1 + change_pct
    if denominator == 0:
        return None
    return current / denominator


def _skip_total_row(search_query: object, asin: object, first_cell: object) -> bool:
    if _clean_text(search_query).upper() == "TOTAL":
        return True
    if _clean_text(asin).upper() == "TOTAL":
        return True
    if _clean_text(first_cell).upper() == "TOTAL":
        return True
    return False


def _parse_traffic_sales_file(path: Path, marketplace: str, period_hint: str, format_type: str) -> pd.DataFrame:
    _, rows = _read_first_sheet(path)
    if not rows:
        return pd.DataFrame()

    header_row = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    raw = pd.DataFrame(rows[1:], columns=header_row)
    columns = list(raw.columns)

    asin_col = _find_by_keywords(columns, ["ASIN"])
    product_col = _find_by_keywords(columns, ["商品名称"]) or _find_by_keywords(columns, ["产品名称"]) or _find_by_keywords(columns, ["商品"])
    conv_current_col, conv_prior_col, conv_change_col = _choose_metric_columns(columns, "转化率")
    page_views_current_col, page_views_prior_col, page_views_change_col = _choose_metric_columns(columns, "推荐报价浏览量")
    offer_rate_current_col, offer_rate_prior_col, offer_rate_change_col = _choose_metric_columns(columns, "推荐报价率")
    ordered_current_col, ordered_prior_col, ordered_change_col = _choose_metric_columns(columns, "已订购商品数量")
    shipped_current_col, shipped_prior_col, shipped_change_col = _choose_metric_columns(columns, "已发货商品数量")

    rows_out: list[dict[str, object]] = []
    for _, row in raw.iterrows():
        asin_value = _extract_formula_display(row.get(asin_col, "")) if asin_col else ""
        product_value = _extract_formula_display(row.get(product_col, "")) if product_col else ""
        if _skip_total_row(row.iloc[0] if len(row) else "", asin_value, row.iloc[0] if len(row) else ""):
            continue
        asin = _clean_text(asin_value)
        if not asin:
            continue
        product_name = _clean_text(product_value)

        current_conv = _coerce_numeric(row.get(conv_current_col)) if conv_current_col else None
        prior_conv = _coerce_numeric(row.get(conv_prior_col)) if conv_prior_col else None
        conv_change_pct = None
        conv_change_bps = None
        if conv_change_col:
            raw_change = _coerce_numeric(row.get(conv_change_col))
            if raw_change is not None:
                if "bps" in str(conv_change_col).lower():
                    conv_change_bps = raw_change
                    conv_change_pct = raw_change / 10000.0
                else:
                    conv_change_pct = raw_change
        if prior_conv is None and current_conv is not None and conv_change_pct is not None:
            prior_conv = _derive_prior_from_change(current_conv, conv_change_pct)

        current_views = _coerce_numeric(row.get(page_views_current_col)) if page_views_current_col else None
        prior_views = _coerce_numeric(row.get(page_views_prior_col)) if page_views_prior_col else None
        views_change_pct = None
        if page_views_change_col:
            views_change_pct = _coerce_numeric(row.get(page_views_change_col))
        if prior_views is None and current_views is not None and views_change_pct is not None:
            prior_views = _derive_prior_from_change(current_views, views_change_pct)

        current_offer_rate = _coerce_numeric(row.get(offer_rate_current_col)) if offer_rate_current_col else None
        prior_offer_rate = _coerce_numeric(row.get(offer_rate_prior_col)) if offer_rate_prior_col else None
        offer_rate_change_pct = None
        if offer_rate_change_col:
            offer_rate_change_pct = _coerce_numeric(row.get(offer_rate_change_col))
        if prior_offer_rate is None and current_offer_rate is not None and offer_rate_change_pct is not None:
            prior_offer_rate = _derive_prior_from_change(current_offer_rate, offer_rate_change_pct)

        current_ordered = _coerce_numeric(row.get(ordered_current_col)) if ordered_current_col else None
        prior_ordered = _coerce_numeric(row.get(ordered_prior_col)) if ordered_prior_col else None
        ordered_change_pct = None
        if ordered_change_col:
            ordered_change_pct = _coerce_numeric(row.get(ordered_change_col))
        if prior_ordered is None and current_ordered is not None and ordered_change_pct is not None:
            prior_ordered = _derive_prior_from_change(current_ordered, ordered_change_pct)

        current_shipped = _coerce_numeric(row.get(shipped_current_col)) if shipped_current_col else None
        prior_shipped = _coerce_numeric(row.get(shipped_prior_col)) if shipped_prior_col else None
        shipped_change_pct = None
        if shipped_change_col:
            shipped_change_pct = _coerce_numeric(row.get(shipped_change_col))
        if prior_shipped is None and current_shipped is not None and shipped_change_pct is not None:
            prior_shipped = _derive_prior_from_change(current_shipped, shipped_change_pct)

        rows_out.append(
            {
                "marketplace": marketplace,
                "file_type": "traffic_sales",
                "period_hint": period_hint,
                "format_type": format_type,
                "source_file": str(path),
                "asin": asin,
                "product_name": product_name,
                "recent_conversion_rate": _to_ratio(current_conv),
                "prior_conversion_rate": _to_ratio(prior_conv),
                "conversion_rate_change_pct": _to_ratio(conv_change_pct),
                "conversion_rate_change_bps": _to_ratio(conv_change_bps),
                "recent_featured_offer_page_views": _to_ratio(current_views),
                "prior_featured_offer_page_views": _to_ratio(prior_views),
                "featured_offer_page_views_change_pct": _to_ratio(views_change_pct),
                "recent_featured_offer_rate": _to_ratio(current_offer_rate),
                "prior_featured_offer_rate": _to_ratio(prior_offer_rate),
                "featured_offer_rate_change_pct": _to_ratio(offer_rate_change_pct),
                "recent_units_ordered": _to_ratio(current_ordered),
                "prior_units_ordered": _to_ratio(prior_ordered),
                "units_ordered_change_pct": _to_ratio(ordered_change_pct),
                "recent_units_shipped": _to_ratio(current_shipped),
                "prior_units_shipped": _to_ratio(prior_shipped),
                "units_shipped_change_pct": _to_ratio(shipped_change_pct),
            }
        )

    return pd.DataFrame(rows_out)


def _parse_query_file(path: Path, marketplace: str, period_hint: str, format_type: str) -> pd.DataFrame:
    _, rows = _read_first_sheet(path)
    if not rows:
        return pd.DataFrame()

    header_row = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    raw = pd.DataFrame(rows[1:], columns=header_row)
    columns = list(raw.columns)

    query_col = _find_by_keywords(columns, ["搜索查询"]) or _find_by_keywords(columns, ["Search Query"])
    asin_col = _find_by_keywords(columns, ["ASIN"])
    product_col = _find_by_keywords(columns, ["商品名称"]) or _find_by_keywords(columns, ["产品名称"]) or _find_by_keywords(columns, ["商品"])
    impressions_current_col, impressions_prior_col, impressions_change_col = _choose_metric_columns(columns, "展示次数")
    clicks_current_col, clicks_prior_col, clicks_change_col = _choose_metric_columns(columns, "点击数量")
    cart_current_col, cart_prior_col, cart_change_col = _choose_metric_columns(columns, "添加购物车")
    purchase_current_col, purchase_prior_col, purchase_change_col = _choose_metric_columns(columns, "购买数量")

    rows_out: list[dict[str, object]] = []
    for _, row in raw.iterrows():
        search_query = _extract_formula_display(row.get(query_col, "")) if query_col else ""
        asin_value = _extract_formula_display(row.get(asin_col, "")) if asin_col else ""
        product_value = _extract_formula_display(row.get(product_col, "")) if product_col else ""
        if _skip_total_row(search_query, asin_value, row.iloc[0] if len(row) else ""):
            continue
        search_query_text = _clean_text(search_query)
        asin = _clean_text(asin_value)
        if not asin:
            continue
        product_name = _clean_text(product_value)

        current_impressions = _coerce_numeric(row.get(impressions_current_col)) if impressions_current_col else None
        prior_impressions = _coerce_numeric(row.get(impressions_prior_col)) if impressions_prior_col else None
        impressions_change_pct = _coerce_numeric(row.get(impressions_change_col)) if impressions_change_col else None
        if prior_impressions is None and current_impressions is not None and impressions_change_pct is not None:
            prior_impressions = _derive_prior_from_change(current_impressions, impressions_change_pct)

        current_clicks = _coerce_numeric(row.get(clicks_current_col)) if clicks_current_col else None
        prior_clicks = _coerce_numeric(row.get(clicks_prior_col)) if clicks_prior_col else None
        clicks_change_pct = _coerce_numeric(row.get(clicks_change_col)) if clicks_change_col else None
        if prior_clicks is None and current_clicks is not None and clicks_change_pct is not None:
            prior_clicks = _derive_prior_from_change(current_clicks, clicks_change_pct)

        current_carts = _coerce_numeric(row.get(cart_current_col)) if cart_current_col else None
        prior_carts = _coerce_numeric(row.get(cart_prior_col)) if cart_prior_col else None
        carts_change_pct = _coerce_numeric(row.get(cart_change_col)) if cart_change_col else None
        if prior_carts is None and current_carts is not None and carts_change_pct is not None:
            prior_carts = _derive_prior_from_change(current_carts, carts_change_pct)

        current_purchases = _coerce_numeric(row.get(purchase_current_col)) if purchase_current_col else None
        prior_purchases = _coerce_numeric(row.get(purchase_prior_col)) if purchase_prior_col else None
        purchases_change_pct = _coerce_numeric(row.get(purchase_change_col)) if purchase_change_col else None
        if prior_purchases is None and current_purchases is not None and purchases_change_pct is not None:
            prior_purchases = _derive_prior_from_change(current_purchases, purchases_change_pct)

        rows_out.append(
            {
                "marketplace": marketplace,
                "file_type": "search_query_performance",
                "period_hint": period_hint,
                "format_type": format_type,
                "source_file": str(path),
                "search_query": search_query_text,
                "asin": asin,
                "product_name": product_name,
                "query_impressions": _to_ratio(current_impressions),
                "prior_query_impressions": _to_ratio(prior_impressions),
                "query_impressions_change_pct": _to_ratio(impressions_change_pct),
                "query_clicks": _to_ratio(current_clicks),
                "prior_query_clicks": _to_ratio(prior_clicks),
                "query_clicks_change_pct": _to_ratio(clicks_change_pct),
                "query_cart_adds": _to_ratio(current_carts),
                "prior_query_cart_adds": _to_ratio(prior_carts),
                "query_cart_adds_change_pct": _to_ratio(carts_change_pct),
                "query_purchases": _to_ratio(current_purchases),
                "prior_query_purchases": _to_ratio(prior_purchases),
                "query_purchases_change_pct": _to_ratio(purchases_change_pct),
            }
        )

    return pd.DataFrame(rows_out)


def _load_dataset_rows(
    custom_dir: Path, marketplace: str, common_end: date | None = None
) -> tuple[list[dict[str, object]], pd.DataFrame, pd.DataFrame, dict[str, float]]:
    site_dir, scan_dir, _, _ = _resolve_enhanced_scan_dir(custom_dir, marketplace)
    file_records: list[dict[str, object]] = []
    traffic_frames: list[pd.DataFrame] = []
    query_frames: list[pd.DataFrame] = []
    traffic_read_sec = 0.0
    query_read_sec = 0.0
    seen_file_keys: set[tuple[object, ...]] = set()

    for path in sorted(scan_dir.iterdir()):
        if not path.is_file() or path.suffix.lower() not in {".xlsx", ".xls", ".csv"}:
            continue
        file_marketplace, data_type, period_hint = _classify_report_file(path)
        file_sheet, preview_rows = _read_first_sheet(path)
        columns = [str(cell).strip() if cell is not None else "" for cell in preview_rows[0]] if preview_rows else []
        file_info = _detect_header_metadata(
            path=path,
            columns=columns,
            preview_rows=preview_rows[:8] if preview_rows else [],
            marketplace=marketplace,
            filename_marketplace=file_marketplace,
            filename_data_type=data_type,
            filename_period_hint=period_hint,
            common_end=common_end,
        )
        if file_info.get("marketplace") != marketplace.upper():
            continue
        duplicate_key = (
            file_info.get("marketplace"),
            file_info.get("data_type"),
            file_info.get("format_type"),
            file_info.get("recent_week") or file_info.get("recent_start") or file_info.get("period_hint"),
            file_info.get("recent_end"),
            file_info.get("total_signature"),
        )
        if duplicate_key in seen_file_keys:
            continue
        seen_file_keys.add(duplicate_key)
        file_info = {
            **file_info,
            "file_name": path.name,
            "file_path": str(path),
            "sheet_name": file_sheet,
            "source_folder": str(path.parent).replace('\\', '/'),
            "site_folder": str(site_dir).replace('\\', '/'),
        }
        file_records.append(file_info)
        data_type = str(file_info.get("data_type") or "unknown")
        period_hint = str(file_info.get("period_hint") or "unknown")
        format_type = str(file_info.get("format_type") or "unknown")
        file_marketplace = str(file_info.get("marketplace") or marketplace.upper())
        if data_type == "traffic_sales":
            t0 = time.perf_counter()
            traffic_frames.append(_parse_traffic_sales_file(path, file_marketplace, period_hint, format_type))
            traffic_read_sec += time.perf_counter() - t0
        elif data_type == "search_query_performance":
            t0 = time.perf_counter()
            query_frames.append(_parse_query_file(path, file_marketplace, period_hint, format_type))
            query_read_sec += time.perf_counter() - t0

    traffic_df = _concat_frames_preserving_columns(traffic_frames)
    query_df = _concat_frames_preserving_columns(query_frames)
    return file_records, traffic_df, query_df, {
        "traffic_sales_read_sec": float(traffic_read_sec),
        "search_query_read_sec": float(query_read_sec),
    }

def _is_comparable(format_type: str, file_records: list[dict[str, object]], data_type: str) -> bool:
    if format_type in {"wow", "compare"}:
        return True
    hints = {str(record.get("period_hint") or "") for record in file_records if record.get("data_type") == data_type}
    return "recent" in hints and "prior" in hints


def _is_dated_single_file(record: dict[str, object]) -> bool:
    name = str(record.get("file_name") or "")
    return bool(re.search(r"\d{4}-\d{2}-\d{2}", name))


def _parse_iso_date(value: object) -> date | None:
    if value in (None, ""):
        return None
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def _record_recency_key(record: dict[str, object]) -> tuple[int, int, date]:
    period_hint = str(record.get("period_hint") or "")
    format_type = str(record.get("format_type") or "")
    identifiable = (
        period_hint != "unknown"
        or format_type in {"wow", "compare"}
        or _is_dated_single_file(record)
    )
    week = int(_coerce_numeric(record.get("recent_week")) or 0)
    recent_end = _parse_iso_date(record.get("recent_end")) or date.min
    return (1 if identifiable else 0, week, recent_end)


def _mark_diagnosis_files(file_records: list[dict[str, object]]) -> set[str]:
    used_paths: set[str] = set()
    for data_type in ["traffic_sales", "search_query_performance"]:
        records = [record for record in file_records if record.get("data_type") == data_type]
        if not records:
            continue
        best_key = max(_record_recency_key(record) for record in records)
        selected = [record for record in records if _record_recency_key(record) == best_key]
        selected_paths = {str(record.get("file_path") or "") for record in selected}
        for record in records:
            used = str(record.get("file_path") or "") in selected_paths
            freshness = str(record.get("freshness") or "")
            if used and data_type == "traffic_sales" and freshness == "stale":
                record["used_in_diagnosis"] = "仅背景参考"
                continue
            record["used_in_diagnosis"] = "是" if used else "否"
            if used:
                used_paths.add(str(record.get("file_path") or ""))
    return used_paths


def _filter_detail_for_diagnosis(frame: pd.DataFrame, used_paths: set[str]) -> pd.DataFrame:
    if frame.empty or "source_file" not in frame.columns:
        return frame
    if not used_paths:
        return frame.iloc[0:0].copy()
    return frame[frame["source_file"].astype(str).isin(used_paths)].copy()


def _build_file_request_rows(
    marketplace: str,
    file_records: list[dict[str, object]],
    common_end: date,
) -> pd.DataFrame:
    periods = compute_recent_prior_ranges(common_end)
    rows: list[dict[str, object]] = []
    seen: set[tuple[str, str, str, str]] = set()

    files_by_type: dict[str, list[dict[str, object]]] = {
        "traffic_sales": [record for record in file_records if record.get("data_type") == "traffic_sales"],
        "search_query_performance": [record for record in file_records if record.get("data_type") == "search_query_performance"],
    }

    for data_type, records in files_by_type.items():
        for record in records:
            key = (
                marketplace,
                data_type,
                str(record.get("period_hint") or "unknown"),
                str(record.get("file_name") or ""),
            )
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                {
                    "marketplace": marketplace,
                    "trigger_sku": "",
                    "trigger_asin": "",
                    "trigger_product_name": "",
                    "issue_type": "增强数据文件导入",
                    "report_type": "流量和销售数据" if data_type == "traffic_sales" else "搜索查询绩效",
                    "period": record.get("period_hint") or "unknown",
                    "start_date": "",
                    "end_date": "",
                    "expected_filename": record.get("file_name") or "",
                    "target_path": f"{CUSTOM_DATA_DIRNAME}/",
                    "required": "否",
                    "seller_central_page": "定制分析",
                    "instruction": "已导入",
                    "status": "已导入",
                    "file_type": data_type,
                    "format_type": record.get("format_type") or "unknown",
                }
            )

    comparable_traffic = _is_comparable(
        next((str(record.get("format_type") or "") for record in files_by_type["traffic_sales"]), "unknown"),
        file_records,
        "traffic_sales",
    )
    comparable_query = _is_comparable(
        next((str(record.get("format_type") or "") for record in files_by_type["search_query_performance"]), "unknown"),
        file_records,
        "search_query_performance",
    )

    if not comparable_traffic:
        recent_exists = any(record.get("period_hint") == "recent" for record in files_by_type["traffic_sales"])
        prior_exists = any(record.get("period_hint") == "prior" for record in files_by_type["traffic_sales"])
        if recent_exists and not prior_exists:
            start_date, end_date = periods["prior"]
            rows.append(
                {
                    "marketplace": marketplace,
                    "trigger_sku": "",
                    "trigger_asin": "",
                    "trigger_product_name": "",
                    "issue_type": "流量和销售数据缺少对比文件",
                    "report_type": "流量和销售数据",
                    "period": "prior",
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                    "expected_filename": f"{TRAFFIC_PREFIX}_{marketplace.lower()}_prior_{start_date.isoformat()}_{end_date.isoformat()}.xlsx",
                    "target_path": f"{CUSTOM_DATA_DIRNAME}/",
                    "required": "是",
                    "seller_central_page": "定制分析 > 流量和销售数据",
                    "instruction": "请导出对比周期文件并放入 data/raw_amazon_custom/",
                    "status": "待导出",
                    "file_type": "traffic_sales",
                    "format_type": "single",
                }
            )
        elif prior_exists and not recent_exists:
            start_date, end_date = periods["recent"]
            rows.append(
                {
                    "marketplace": marketplace,
                    "trigger_sku": "",
                    "trigger_asin": "",
                    "trigger_product_name": "",
                    "issue_type": "流量和销售数据缺少对比文件",
                    "report_type": "流量和销售数据",
                    "period": "recent",
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                    "expected_filename": f"{TRAFFIC_PREFIX}_{marketplace.lower()}_recent_{start_date.isoformat()}_{end_date.isoformat()}.xlsx",
                    "target_path": f"{CUSTOM_DATA_DIRNAME}/",
                    "required": "是",
                    "seller_central_page": "定制分析 > 流量和销售数据",
                    "instruction": "请导出对比周期文件并放入 data/raw_amazon_custom/",
                    "status": "待导出",
                    "file_type": "traffic_sales",
                    "format_type": "single",
                }
            )

    if not comparable_query:
        recent_exists = any(record.get("period_hint") == "recent" for record in files_by_type["search_query_performance"])
        prior_exists = any(record.get("period_hint") == "prior" for record in files_by_type["search_query_performance"])
        if recent_exists and not prior_exists:
            start_date, end_date = periods["prior"]
            rows.append(
                {
                    "marketplace": marketplace,
                    "trigger_sku": "",
                    "trigger_asin": "",
                    "trigger_product_name": "",
                    "issue_type": "搜索查询绩效缺少对比文件",
                    "report_type": "搜索查询绩效",
                    "period": "prior",
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                    "expected_filename": f"{QUERY_PREFIX}_{marketplace.lower()}_prior_{start_date.isoformat()}_{end_date.isoformat()}.xlsx",
                    "target_path": f"{CUSTOM_DATA_DIRNAME}/",
                    "required": "可选",
                    "seller_central_page": "定制分析 > 搜索查询绩效",
                    "instruction": "如需对比分析，请导出 prior 文件并放入 data/raw_amazon_custom/",
                    "status": "待导出",
                    "file_type": "search_query_performance",
                    "format_type": "single",
                }
            )
        elif prior_exists and not recent_exists:
            start_date, end_date = periods["recent"]
            rows.append(
                {
                    "marketplace": marketplace,
                    "trigger_sku": "",
                    "trigger_asin": "",
                    "trigger_product_name": "",
                    "issue_type": "搜索查询绩效缺少对比文件",
                    "report_type": "搜索查询绩效",
                    "period": "recent",
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                    "expected_filename": f"{QUERY_PREFIX}_{marketplace.lower()}_recent_{start_date.isoformat()}_{end_date.isoformat()}.xlsx",
                    "target_path": f"{CUSTOM_DATA_DIRNAME}/",
                    "required": "可选",
                    "seller_central_page": "定制分析 > 搜索查询绩效",
                    "instruction": "如需对比分析，请导出 recent 文件并放入 data/raw_amazon_custom/",
                    "status": "待导出",
                    "file_type": "search_query_performance",
                    "format_type": "single",
                }
            )

    frame = pd.DataFrame(
        rows,
        columns=[
            "marketplace",
            "trigger_sku",
            "trigger_asin",
            "trigger_product_name",
            "issue_type",
            "report_type",
            "period",
            "start_date",
            "end_date",
            "expected_filename",
            "target_path",
            "required",
            "seller_central_page",
            "instruction",
            "status",
            "file_type",
            "format_type",
        ],
    )
    if frame.empty:
        return frame
    return frame.drop_duplicates(subset=["marketplace", "report_type", "period", "expected_filename"], keep="first").reset_index(drop=True)


def _build_request_markdown(marketplace: str, file_records: list[dict[str, object]], request_rows: pd.DataFrame) -> str:
    lines = ["# 需要补充导出的增强数据", ""]
    imported_rows = request_rows[request_rows["status"] == "已导入"] if not request_rows.empty else pd.DataFrame()
    pending_rows = request_rows[request_rows["status"] == "待导出"] if not request_rows.empty else pd.DataFrame()

    lines.append(f"## {marketplace}｜已导入增强数据")
    if imported_rows.empty:
        lines.append("当前没有识别到已导入的增强数据文件。")
    else:
        lines.append("| marketplace | 数据类型 | 周期类型 | 格式类型 | 文件名 | 状态 |")
        lines.append("|---|---|---|---|---|---|")
        for _, row in imported_rows.iterrows():
            lines.append(
                f"| {row['marketplace']} | {row['report_type']} | {row['period']} | {row['format_type']} | {row['expected_filename']} | {row['status']} |"
            )

    lines.append("")
    lines.append(f"## {marketplace}｜待导出请求")
    if pending_rows.empty:
        lines.append("当前没有真正缺少的必要对比文件。")
    else:
        lines.append("| 序号 | 报表 | 日期范围 | 导出后文件名 | 放置路径 | 必需 |")
        lines.append("|---|---|---|---|---|---|")
        for idx, (_, row) in enumerate(pending_rows.iterrows(), start=1):
            lines.append(
                f"| {idx} | {row['report_type']} | {row['start_date']} ~ {row['end_date']} | {row['expected_filename']} | {row['target_path']} | {row['required']} |"
            )

    lines.append("")
    lines.append("操作说明：")
    lines.append("1. 进入亚马逊后台对应定制分析页面。")
    lines.append("2. 日期选择表格里的日期范围。")
    lines.append("3. 点击导出。")
    lines.append("4. 把下载文件改成表格里的文件名。")
    lines.append(f"5. 放入 {CUSTOM_DATA_DIRNAME}/。")
    lines.append(f"6. 重新运行：`python main.py --marketplace {marketplace}`")
    return "\n".join(lines).rstrip() + "\n"


def _trigger_lookup(analysis_payload: dict) -> dict[tuple[str, str], dict]:
    lookup: dict[tuple[str, str], dict] = {}
    for value in analysis_payload.values():
        if isinstance(value, dict) and {"1d", "3d", "7d", "14d"}.issubset(value.keys()):
            for row in value.get("1d", []):
                sku = _clean_text(row.get("sku"))
                asin = _clean_text(row.get("asin"))
                if sku or asin:
                    lookup[(sku, asin)] = row
            break
    return lookup


def _analysis_records(analysis_payload: dict) -> list[dict]:
    for value in analysis_payload.values():
        if isinstance(value, dict) and {"1d", "3d", "7d", "14d"}.issubset(value.keys()):
            records = value.get("1d", [])
            if isinstance(records, list):
                return records
    return []


def _build_cost_lookup(product_records: list[dict]) -> dict[tuple[str, str], dict[str, object]]:
    lookup: dict[tuple[str, str], dict[str, object]] = {}
    for record in product_records:
        sku = _clean_text(record.get("sku"))
        asin = _clean_text(record.get("asin"))
        profit = record.get("profit_before_ads_per_unit")
        target_acos = record.get("target_acos")
        has_issue = False
        try:
            if profit not in (None, "") and float(profit) < 0:
                has_issue = True
        except (TypeError, ValueError):
            pass
        if has_issue:
            lookup[(sku, asin)] = {
                "profit_before_ads_per_unit": profit,
                "target_acos": target_acos,
            }
    return lookup


def _natural_decline_triggers(analysis_payload: dict) -> list[dict]:
    triggers: list[dict] = []
    product_lookup = _trigger_lookup(analysis_payload)
    anomaly_records: list[dict] = []
    for value in analysis_payload.values():
        if isinstance(value, list) and value:
            first = value[0]
            if isinstance(first, dict) and {"rule", "target", "evidence"}.issubset(first.keys()):
                anomaly_records = value
                break

    for anomaly in anomaly_records:
        evidence = anomaly.get("evidence", {})
        if "recent_period_natural_orders" not in evidence or "prior_period_natural_orders" not in evidence:
            continue
        sku = _clean_text(evidence.get("sku") or anomaly.get("target"))
        asin = _clean_text(evidence.get("asin"))
        product = product_lookup.get((sku, asin), {})
        triggers.append(
            {
                "marketplace": _clean_text(evidence.get("marketplace")),
                "sku": sku,
                "asin": asin,
                "product_name": product.get("product_name") or sku,
                "recent_period_days": int(float(evidence.get("recent_period_days", 7) or 7)),
                "recent_period_natural_orders": float(evidence.get("recent_period_natural_orders", 0) or 0),
                "prior_period_natural_orders": float(evidence.get("prior_period_natural_orders", 0) or 0),
                "natural_order_drop_abs": float(evidence.get("natural_order_drop_abs", 0) or 0),
                "natural_order_drop_pct": float(evidence.get("natural_order_drop_pct", 0) or 0),
            }
        )
    return triggers


def _resolve_current_prior(row: pd.Series, current_key: str, prior_key: str, change_key: str) -> tuple[float | None, float | None, float | None]:
    current = row.get(current_key, None)
    prior = row.get(prior_key, None)
    change = row.get(change_key, None)
    current_value = _coerce_numeric(current)
    prior_value = _coerce_numeric(prior)
    change_value = _coerce_numeric(change)
    if prior_value is None and current_value is not None and change_value is not None:
        prior_value = _derive_prior_from_change(current_value, change_value)
    return current_value, prior_value, change_value


def _pick_traffic_record(traffic_df: pd.DataFrame, marketplace: str, asin: str) -> dict[str, object] | None:
    if traffic_df.empty:
        return None
    matches = traffic_df[
        (traffic_df["marketplace"].astype(str).str.upper() == marketplace.upper())
        & (traffic_df["asin"].astype(str).str.upper() == asin.upper())
    ]
    if matches.empty:
        return None
    return matches.iloc[0].to_dict()


def _build_enhanced_diagnostics(marketplace: str, analysis_payload: dict, traffic_df: pd.DataFrame, cost_lookup: dict[tuple[str, str], dict[str, object]]) -> pd.DataFrame:
    triggers = _natural_decline_triggers(analysis_payload)
    if not triggers:
        return pd.DataFrame(columns=[
            "marketplace",
            "sku",
            "asin",
            "product_name",
            "recent_period_days",
            "recent_featured_offer_page_views",
            "prior_featured_offer_page_views",
            "featured_offer_page_views_change_pct",
            "recent_conversion_rate",
            "prior_conversion_rate",
            "conversion_rate_change_pct",
            "recent_featured_offer_rate",
            "prior_featured_offer_rate",
            "featured_offer_rate_change_pct",
            "diagnosis",
            "recommendation",
            "has_cost_issue",
        ])

    rows: list[dict[str, object]] = []
    for trigger in triggers:
        row = _pick_traffic_record(traffic_df, marketplace, trigger["asin"])
        cost_issue = (trigger["sku"], trigger["asin"]) in cost_lookup
        if row is None:
            rows.append(
                {
                    "marketplace": marketplace,
                    "sku": trigger["sku"],
                    "asin": trigger["asin"],
                    "product_name": trigger["product_name"],
                    "recent_period_days": trigger["recent_period_days"],
                    "recent_featured_offer_page_views": None,
                    "prior_featured_offer_page_views": None,
                    "featured_offer_page_views_change_pct": None,
                    "recent_conversion_rate": None,
                    "prior_conversion_rate": None,
                    "conversion_rate_change_pct": None,
                    "recent_featured_offer_rate": None,
                    "prior_featured_offer_rate": None,
                    "featured_offer_rate_change_pct": None,
                    "diagnosis": "数据不足",
                    "recommendation": "未导入可比较增强数据，当前只能判断自然单下降，无法拆分流量/转化/推荐报价原因。",
                    "has_cost_issue": cost_issue,
                }
            )
            continue

        recent_views = row.get("recent_featured_offer_page_views")
        prior_views = row.get("prior_featured_offer_page_views")
        views_change_pct = row.get("featured_offer_page_views_change_pct")
        recent_cr = row.get("recent_conversion_rate")
        prior_cr = row.get("prior_conversion_rate")
        cr_change_pct = row.get("conversion_rate_change_pct")
        recent_offer_rate = row.get("recent_featured_offer_rate")
        prior_offer_rate = row.get("prior_featured_offer_rate")
        offer_change_pct = row.get("featured_offer_rate_change_pct")

        if recent_views is not None and prior_views is None and views_change_pct is not None:
            prior_views = _derive_prior_from_change(float(recent_views), float(views_change_pct))
        if recent_cr is not None and prior_cr is None and cr_change_pct is not None:
            prior_cr = _derive_prior_from_change(float(recent_cr), float(cr_change_pct))
        if recent_offer_rate is not None and prior_offer_rate is None and offer_change_pct is not None:
            prior_offer_rate = _derive_prior_from_change(float(recent_offer_rate), float(offer_change_pct))

        diagnosis = "数据不足"
        recommendation = "请结合广告与页面表现继续观察。"
        recent_views_f = _coerce_numeric(recent_views)
        prior_views_f = _coerce_numeric(prior_views)
        recent_cr_f = _coerce_numeric(recent_cr)
        prior_cr_f = _coerce_numeric(prior_cr)
        recent_offer_rate_f = _coerce_numeric(recent_offer_rate)
        prior_offer_rate_f = _coerce_numeric(prior_offer_rate)
        views_change = _coerce_numeric(views_change_pct)
        cr_change = _coerce_numeric(cr_change_pct)

        if recent_offer_rate_f is not None and recent_offer_rate_f < 0.9:
            diagnosis = "推荐报价 / Buy Box 风险"
            recommendation = "检查价格、报价竞争、库存和配送方式。"
        elif prior_offer_rate_f is not None and recent_offer_rate_f is not None and (prior_offer_rate_f - recent_offer_rate_f) >= 0.1:
            diagnosis = "推荐报价 / Buy Box 风险"
            recommendation = "检查价格、报价竞争、库存和配送方式。"
        elif views_change is not None and views_change <= -0.3 and (cr_change is None or cr_change > -0.3):
            diagnosis = "流量问题"
            recommendation = "检查自然曝光、广告承接、关键词覆盖、预算和竞价。"
        elif cr_change is not None and cr_change <= -0.3 and (views_change is None or views_change > -0.3):
            diagnosis = "转化问题"
            recommendation = "检查价格、Coupon、主图、评价、竞品和 Listing。"
        elif views_change is not None and cr_change is not None and views_change <= -0.3 and cr_change <= -0.3:
            diagnosis = "流量和转化同时下降"
            recommendation = "先查价格/Coupon/评价，再查广告承接。"
        elif recent_views_f is not None and recent_views_f < 100 and recent_cr_f is not None and recent_cr_f >= 0.1:
            diagnosis = "流量不足但转化尚可"
            recommendation = "可小幅提高核心词精准广告或商品投放竞价 10%-15%，观察 2-3 天。"

        if cost_issue:
            recommendation = "先核对售价、采购成本、头程成本、FBA费用；如果利润为负属实，不建议加广告放量，只保留高相关低价精准词，暂停泛词和无转化商品投放。"

        rows.append(
            {
                "marketplace": marketplace,
                "sku": trigger["sku"],
                "asin": trigger["asin"],
                "product_name": trigger["product_name"],
                "recent_period_days": trigger["recent_period_days"],
                "recent_featured_offer_page_views": recent_views_f,
                "prior_featured_offer_page_views": prior_views_f,
                "featured_offer_page_views_change_pct": views_change,
                "recent_conversion_rate": recent_cr_f,
                "prior_conversion_rate": prior_cr_f,
                "conversion_rate_change_pct": cr_change,
                "recent_featured_offer_rate": recent_offer_rate_f,
                "prior_featured_offer_rate": prior_offer_rate_f,
                "featured_offer_rate_change_pct": _coerce_numeric(offer_change_pct),
                "diagnosis": diagnosis,
                "recommendation": recommendation,
                "has_cost_issue": cost_issue,
            }
        )

    return pd.DataFrame(rows)


def _build_search_query_opportunities(marketplace: str, query_df: pd.DataFrame, analysis_payload: dict) -> pd.DataFrame:
    if query_df.empty:
        return pd.DataFrame(columns=[
            "marketplace",
            "type",
            "search_query",
            "asin",
            "product_name",
            "query_impressions",
            "query_clicks",
            "query_cart_adds",
            "query_purchases",
            "suggestion",
            "priority_score",
        ])

    recent = query_df.copy()
    recent["priority_score"] = 0.0
    recent["type"] = ""
    recent["suggestion"] = ""

    ad_term_lookup: dict[str, dict] = {}
    search_summary = analysis_payload.get("搜索词分析", {})
    if isinstance(search_summary, dict):
        for item in search_summary.get("7d", []):
            term = _clean_text(item.get("search_term")).lower()
            if term:
                ad_term_lookup[term] = item

    for idx, row in recent.iterrows():
        impressions = _coerce_numeric(row.get("query_impressions")) or 0.0
        clicks = _coerce_numeric(row.get("query_clicks")) or 0.0
        carts = _coerce_numeric(row.get("query_cart_adds")) or 0.0
        purchases = _coerce_numeric(row.get("query_purchases")) or 0.0
        search_term = _clean_text(row.get("search_query")).lower()

        if purchases > 0:
            recent.at[idx, "type"] = "有购买词"
            recent.at[idx, "suggestion"] = "可加入精准广告或提高关注。"
            recent.at[idx, "priority_score"] = purchases * 100 + clicks
        elif carts > 0 and purchases == 0:
            recent.at[idx, "type"] = "高加购低购买词"
            recent.at[idx, "suggestion"] = "检查价格、页面转化、Coupon。"
            recent.at[idx, "priority_score"] = carts * 50 + clicks
        elif clicks >= 5 and purchases == 0:
            recent.at[idx, "type"] = "高点击无购买词"
            recent.at[idx, "suggestion"] = "广告谨慎投放或降竞价。"
            recent.at[idx, "priority_score"] = clicks * 10 + impressions / 100
        elif impressions >= 100 and clicks <= 1:
            recent.at[idx, "type"] = "高展示低点击词"
            recent.at[idx, "suggestion"] = "检查主图、标题、价格吸引力。"
            recent.at[idx, "priority_score"] = impressions / 10
        else:
            continue

        if search_term in ad_term_lookup:
            note = "有购买表现" if _coerce_numeric(ad_term_lookup[search_term].get("ad_orders")) else "无购买表现"
            recent.at[idx, "suggestion"] = f"{recent.at[idx, 'suggestion']} 该词在搜索词建议中{note}。"

    filtered = recent[recent["type"] != ""].copy()
    if filtered.empty:
        return filtered
    filtered = filtered.sort_values(["priority_score", "query_impressions", "query_clicks"], ascending=[False, False, False]).head(15)
    filtered["marketplace"] = marketplace
    return filtered[
        [
            "marketplace",
            "type",
            "search_query",
            "asin",
            "product_name",
            "query_impressions",
            "query_clicks",
            "query_cart_adds",
            "query_purchases",
            "suggestion",
            "priority_score",
        ]
    ].reset_index(drop=True)


def _build_requests_dataframe(marketplace: str, file_records: list[dict[str, object]], analysis_payload: dict, common_end: date) -> pd.DataFrame:
    periods = compute_recent_prior_ranges(common_end)
    rows: list[dict[str, object]] = []
    target_folder = _target_folder_text(CUSTOM_DATA_DIRNAME, marketplace)

    traffic_records = [record for record in file_records if record.get("data_type") == "traffic_sales"]
    query_records = [record for record in file_records if record.get("data_type") == "search_query_performance"]
    traffic_comparable = _is_comparable(next((str(record.get("format_type") or "") for record in traffic_records), "unknown"), file_records, "traffic_sales")
    query_comparable = _is_comparable(next((str(record.get("format_type") or "") for record in query_records), "unknown"), file_records, "search_query_performance")
    expected_recent_week, expected_prior_week = _expected_week_pair(common_end)
    expected_period = (
        f"week{expected_recent_week}_vs_week{expected_prior_week}"
        if expected_recent_week and expected_prior_week
        else ""
    )
    expected_recent_range = _week_date_range(common_end.year, expected_recent_week) if expected_recent_week else None
    expected_prior_range = _week_date_range(common_end.year, expected_prior_week) if expected_prior_week else None

    def _has_expected_week(records: list[dict[str, object]]) -> bool:
        if not expected_recent_week:
            return True
        expected_end = expected_recent_range[1] if expected_recent_range else None
        return any(
            int(_coerce_numeric(record.get("recent_week")) or 0) >= expected_recent_week
            or (expected_end is not None and (_parse_iso_date(record.get("recent_end")) or date.min) >= expected_end)
            for record in records
        )

    def _append_expected_week_request(data_type: str, required: str) -> None:
        if not expected_period or not expected_recent_range or not expected_prior_range:
            return
        recent_start, recent_end = expected_recent_range
        prior_start, prior_end = expected_prior_range
        report_type = "流量和销售数据" if data_type == "traffic_sales" else "搜索查询绩效"
        prefix = TRAFFIC_PREFIX if data_type == "traffic_sales" else QUERY_PREFIX
        page = "流量和销售数据" if data_type == "traffic_sales" else "搜索查询绩效"
        rows.append(
            {
                "marketplace": marketplace,
                "trigger_sku": "",
                "trigger_asin": "",
                "trigger_product_name": "",
                "issue_type": f"缺少 {data_type}_{marketplace.lower()}_{expected_period}",
                "report_type": report_type,
                "period": expected_period,
                "start_date": recent_start.isoformat(),
                "end_date": recent_end.isoformat(),
                "expected_filename": f"{prefix}_{marketplace.lower()}_wow_{recent_start.isoformat()}_{recent_end.isoformat()}_vs_{prior_start.isoformat()}_{prior_end.isoformat()}.xlsx",
                "target_path": target_folder,
                "target_folder": target_folder,
                "required": required,
                "seller_central_page": f"定制分析 > {page}",
                "instruction": f"请补充 {expected_period} 文件并放入 {target_folder}",
                "status": "待导出",
                "file_type": data_type,
                "format_type": "wow",
                "detected_from": "",
                "detected_date_range": f"{recent_start.isoformat()} ~ {recent_end.isoformat()} vs {prior_start.isoformat()} ~ {prior_end.isoformat()}",
                "freshness": "missing",
                "used_in_diagnosis": "否",
            }
        )

    def _append_file_record(record: dict[str, object]) -> None:
        rows.append(
            {
                "marketplace": marketplace,
                "trigger_sku": "",
                "trigger_asin": "",
                "trigger_product_name": "",
                "issue_type": "增强数据文件导入",
                "report_type": "流量和销售数据" if record.get("data_type") == "traffic_sales" else "搜索查询绩效",
                "period": record.get("period_hint") or "unknown",
                "start_date": "",
                "end_date": "",
                "expected_filename": record.get("file_name") or "",
                "target_path": target_folder,
                "target_folder": target_folder,
                "required": "否",
                "seller_central_page": "定制分析",
                "instruction": "已导入",
                "status": "已导入",
                "file_type": record.get("data_type") or "unknown",
                "format_type": record.get("format_type") or "unknown",
                "detected_from": record.get("detected_from") or "fallback",
                "detected_date_range": record.get("detected_date_range") or "",
                "freshness": record.get("freshness") or "unknown",
                "used_in_diagnosis": record.get("used_in_diagnosis", "否"),
            }
        )

    for record in traffic_records + query_records:
        _append_file_record(record)

    if traffic_records and not _has_expected_week(traffic_records):
        _append_expected_week_request("traffic_sales", "是")

    if query_records and not _has_expected_week(query_records):
        _append_expected_week_request("search_query_performance", "否")

    if not traffic_comparable and traffic_records:
        if any(record.get("period_hint") == "recent" for record in traffic_records) and not any(record.get("period_hint") == "prior" for record in traffic_records):
            start_date, end_date = periods["prior"]
            rows.append(
                {
                    "marketplace": marketplace,
                    "trigger_sku": "",
                    "trigger_asin": "",
                    "trigger_product_name": "",
                    "issue_type": "流量和销售数据缺少对比文件",
                    "report_type": "流量和销售数据",
                    "period": "prior",
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                    "expected_filename": f"{TRAFFIC_PREFIX}_{marketplace.lower()}_prior_{start_date.isoformat()}_{end_date.isoformat()}.xlsx",
                    "target_path": target_folder,
                    "target_folder": target_folder,
                    "required": "是",
                    "seller_central_page": "定制分析 > 流量和销售数据",
                    "instruction": f"请导出对比周期文件并放入 {target_folder}",
                    "status": "待导出",
                    "file_type": "traffic_sales",
                    "format_type": "single",
                }
            )
        elif any(record.get("period_hint") == "prior" for record in traffic_records) and not any(record.get("period_hint") == "recent" for record in traffic_records):
            start_date, end_date = periods["recent"]
            rows.append(
                {
                    "marketplace": marketplace,
                    "trigger_sku": "",
                    "trigger_asin": "",
                    "trigger_product_name": "",
                    "issue_type": "流量和销售数据缺少对比文件",
                    "report_type": "流量和销售数据",
                    "period": "recent",
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                    "expected_filename": f"{TRAFFIC_PREFIX}_{marketplace.lower()}_recent_{start_date.isoformat()}_{end_date.isoformat()}.xlsx",
                    "target_path": target_folder,
                    "target_folder": target_folder,
                    "required": "是",
                    "seller_central_page": "定制分析 > 流量和销售数据",
                    "instruction": f"请导出对比周期文件并放入 {target_folder}",
                    "status": "待导出",
                    "file_type": "traffic_sales",
                    "format_type": "single",
                }
            )

    if not query_comparable and query_records:
        if any(record.get("period_hint") == "recent" for record in query_records) and not any(record.get("period_hint") == "prior" for record in query_records):
            start_date, end_date = periods["prior"]
            rows.append(
                {
                    "marketplace": marketplace,
                    "trigger_sku": "",
                    "trigger_asin": "",
                    "trigger_product_name": "",
                    "issue_type": "搜索查询绩效缺少对比文件",
                    "report_type": "搜索查询绩效",
                    "period": "prior",
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                    "expected_filename": f"{QUERY_PREFIX}_{marketplace.lower()}_prior_{start_date.isoformat()}_{end_date.isoformat()}.xlsx",
                    "target_path": target_folder,
                    "target_folder": target_folder,
                    "required": "否",
                    "seller_central_page": "定制分析 > 搜索查询绩效",
                    "instruction": f"如需对比分析，请导出 prior 文件并放入 {target_folder}",
                    "status": "待导出",
                    "file_type": "search_query_performance",
                    "format_type": "single",
                }
            )
        elif any(record.get("period_hint") == "prior" for record in query_records) and not any(record.get("period_hint") == "recent" for record in query_records):
            start_date, end_date = periods["recent"]
            rows.append(
                {
                    "marketplace": marketplace,
                    "trigger_sku": "",
                    "trigger_asin": "",
                    "trigger_product_name": "",
                    "issue_type": "搜索查询绩效缺少对比文件",
                    "report_type": "搜索查询绩效",
                    "period": "recent",
                    "start_date": start_date.isoformat(),
                    "end_date": end_date.isoformat(),
                    "expected_filename": f"{QUERY_PREFIX}_{marketplace.lower()}_recent_{start_date.isoformat()}_{end_date.isoformat()}.xlsx",
                    "target_path": target_folder,
                    "target_folder": target_folder,
                    "required": "否",
                    "seller_central_page": "定制分析 > 搜索查询绩效",
                    "instruction": f"如需对比分析，请导出 recent 文件并放入 {target_folder}",
                    "status": "待导出",
                    "file_type": "search_query_performance",
                    "format_type": "single",
                }
            )

    frame = pd.DataFrame(rows)
    if frame.empty:
        return frame
    return frame.drop_duplicates(subset=["marketplace", "report_type", "period", "expected_filename"], keep="first").reset_index(drop=True)

def _request_markdown(marketplace: str, file_records: list[dict[str, object]], request_rows: pd.DataFrame) -> str:
    lines = ["# 需要补充导出的增强数据", ""]
    lines.append(f"## {marketplace}｜文件识别结果")
    if file_records:
        lines.append("| 文件名 | marketplace | 数据类型 | 周期类型 | 格式类型 | 目录 |")
        lines.append("|---|---|---|---|---|---|")
        for record in file_records:
            lines.append(
                f"| {record.get('file_name') or 'N/A'} | {record.get('marketplace') or 'N/A'} | {record.get('data_type') or 'unknown'} | {record.get('period_hint') or 'unknown'} | {record.get('format_type') or 'unknown'} | {record.get('source_folder') or _target_folder_text(CUSTOM_DATA_DIRNAME, marketplace)} |"
            )
    else:
        lines.append("当前目录没有识别到增强数据文件。")

    lines.append("")
    pending_rows = request_rows[request_rows["status"] == "待导出"] if not request_rows.empty else pd.DataFrame()
    imported_rows = request_rows[request_rows["status"] == "已导入"] if not request_rows.empty else pd.DataFrame()

    lines.append(f"## {marketplace}｜待导出请求")
    if pending_rows.empty:
        lines.append("当前没有真正缺少的必要对比文件。")
    else:
        lines.append("| 序号 | 报表 | 日期范围 | 导出后文件名 | 目标文件夹 | 必需 |")
        lines.append("|---|---|---|---|---|---|")
        for idx, (_, row) in enumerate(pending_rows.iterrows(), start=1):
            lines.append(
                f"| {idx} | {row.get('report_type') or 'N/A'} | {row.get('start_date') or ''} ~ {row.get('end_date') or ''} | {row.get('expected_filename') or 'N/A'} | {row.get('target_folder') or row.get('target_path') or 'N/A'} | {row.get('required') or 'N/A'} |"
            )

    lines.append("")
    lines.append(f"## {marketplace}｜已导入文件")
    if imported_rows.empty:
        lines.append("当前没有已导入文件。")
    else:
        lines.append("| 报表 | 周期 | 格式 | 文件名 | 状态 | 目标文件夹 |")
        lines.append("|---|---|---|---|---|---|")
        for _, row in imported_rows.iterrows():
            lines.append(
                f"| {row.get('report_type') or 'N/A'} | {row.get('period') or 'N/A'} | {row.get('format_type') or 'N/A'} | {row.get('expected_filename') or 'N/A'} | {row.get('status') or 'N/A'} | {row.get('target_folder') or row.get('target_path') or 'N/A'} |"
            )

    lines.extend(
        [
            "",
            "请把文件放入对应站点文件夹，而不是直接放在 raw_amazon_custom 根目录。",
            "操作说明：",
            "1. 进入亚马逊后台对应定制分析页面。",
            "2. 选择表格中的日期范围。",
            "3. 点击导出。",
            "4. 把下载文件改成表格里的文件名。",
            f"5. 请把文件放入对应站点文件夹，而不是直接放在 raw_amazon_custom 根目录：{_target_folder_text(CUSTOM_DATA_DIRNAME, marketplace)}",
            f"6. 重新运行：`python main.py --marketplace {marketplace}`",
        ]
    )
    return "\n".join(lines).rstrip() + "\n"

def load_enhanced_data(custom_dir: Path, marketplace: str, common_end: date) -> tuple[dict[str, object], pd.DataFrame, pd.DataFrame]:
    site_dir, scan_dir, site_files, root_files = _resolve_enhanced_scan_dir(custom_dir, marketplace)
    file_records, traffic_df, query_df, timing = _load_dataset_rows(custom_dir, marketplace, common_end)
    diagnosis_paths = _mark_diagnosis_files(file_records)
    traffic_df = _filter_detail_for_diagnosis(traffic_df, diagnosis_paths)
    query_df = _filter_detail_for_diagnosis(query_df, diagnosis_paths)
    traffic_records = [record for record in file_records if record.get("data_type") == "traffic_sales"]
    query_records = [record for record in file_records if record.get("data_type") == "search_query_performance"]
    traffic_format = traffic_records[0]["format_type"] if traffic_records else "unknown"
    query_format = query_records[0]["format_type"] if query_records else "unknown"
    legacy_root_detected = bool(root_files)

    status = {
        "custom_dir": str(custom_dir),
        "site_dir": str(site_dir),
        "scan_dir": str(scan_dir),
        "target_folder": _target_folder_text(custom_dir, marketplace),
        "provided": bool(file_records),
        "recognized_files": len(file_records),
        "detected_files": file_records,
        "traffic_sales_files": traffic_records,
        "search_query_files": query_records,
        "traffic_sales_recent_exists": any(record.get("period_hint") == "recent" for record in traffic_records),
        "traffic_sales_prior_exists": any(record.get("period_hint") == "prior" for record in traffic_records),
        "search_query_recent_exists": any(record.get("period_hint") == "recent" for record in query_records),
        "search_query_prior_exists": any(record.get("period_hint") == "prior" for record in query_records),
        "traffic_sales_format": traffic_format,
        "search_query_format": query_format,
        "traffic_sales_comparable": _is_comparable(str(traffic_format), file_records, "traffic_sales"),
        "search_query_comparable": _is_comparable(str(query_format), file_records, "search_query_performance"),
        "legacy_root_files_detected": legacy_root_detected,
        "legacy_root_files": [str(path) for path in root_files],
        "site_folder_files": [str(path) for path in site_files],
        "timing": timing,
    }

    status["traffic_sales_needs_prior"] = (
        not status["traffic_sales_comparable"]
        and status["traffic_sales_recent_exists"]
        and not status["traffic_sales_prior_exists"]
    )
    status["search_query_needs_prior"] = (
        not status["search_query_comparable"]
        and status["search_query_recent_exists"]
        and not status["search_query_prior_exists"]
    )
    status["file_summary"] = [
        {
            "file_name": record.get("file_name"),
            "marketplace": record.get("marketplace"),
            "data_type": record.get("data_type"),
            "period_hint": record.get("period_hint"),
            "format_type": record.get("format_type"),
            "source_folder": record.get("source_folder"),
            "status": "已导入",
            "used_in_diagnosis": record.get("used_in_diagnosis", "否"),
            "detected_from": record.get("detected_from", "fallback"),
            "detected_date_range": record.get("detected_date_range", ""),
            "freshness": record.get("freshness", "unknown"),
            "recent_week": record.get("recent_week", ""),
            "prior_week": record.get("prior_week", ""),
            "recent_start": record.get("recent_start", ""),
            "recent_end": record.get("recent_end", ""),
            "prior_start": record.get("prior_start", ""),
            "prior_end": record.get("prior_end", ""),
        }
        for record in file_records
    ]
    status["common_end"] = common_end.isoformat()
    if legacy_root_detected:
        status["legacy_root_warning"] = "检测到旧目录增强文件，建议移动到对应站点文件夹。"
    return status, traffic_df, query_df

def build_natural_decline_diagnostics(
    marketplace: str,
    analysis_payload: dict,
    traffic_df: pd.DataFrame,
) -> pd.DataFrame:
    triggers = _natural_decline_triggers(analysis_payload)
    product_records = _analysis_records(analysis_payload)
    cost_lookup = _build_cost_lookup(product_records)

    if not triggers:
        return pd.DataFrame(
            columns=[
                "marketplace",
                "sku",
                "asin",
                "product_name",
                "recent_period_days",
                "recent_featured_offer_page_views",
                "prior_featured_offer_page_views",
                "featured_offer_page_views_change_pct",
                "recent_conversion_rate",
                "prior_conversion_rate",
                "conversion_rate_change_pct",
                "recent_featured_offer_rate",
                "prior_featured_offer_rate",
                "featured_offer_rate_change_pct",
                "diagnosis",
                "recommendation",
                "has_cost_issue",
            ]
        )

    rows: list[dict[str, object]] = []
    for trigger in triggers:
        row = _pick_traffic_record(traffic_df, marketplace, trigger["asin"])
        cost_issue = (trigger["sku"], trigger["asin"]) in cost_lookup
        if row is None:
            rows.append(
                {
                    "marketplace": marketplace,
                    "sku": trigger["sku"],
                    "asin": trigger["asin"],
                    "product_name": trigger["product_name"],
                    "recent_period_days": trigger["recent_period_days"],
                    "recent_featured_offer_page_views": None,
                    "prior_featured_offer_page_views": None,
                    "featured_offer_page_views_change_pct": None,
                    "recent_conversion_rate": None,
                    "prior_conversion_rate": None,
                    "conversion_rate_change_pct": None,
                    "recent_featured_offer_rate": None,
                    "prior_featured_offer_rate": None,
                    "featured_offer_rate_change_pct": None,
                    "diagnosis": "数据不足",
                    "recommendation": "未导入可比较增强数据，当前只能判断自然单下降，无法拆分流量/转化/推荐报价原因。",
                    "has_cost_issue": cost_issue,
                }
            )
            continue

        recent_views = _coerce_numeric(row.get("recent_featured_offer_page_views"))
        prior_views = _coerce_numeric(row.get("prior_featured_offer_page_views"))
        views_change_pct = _coerce_numeric(row.get("featured_offer_page_views_change_pct"))
        recent_cr = _coerce_numeric(row.get("recent_conversion_rate"))
        prior_cr = _coerce_numeric(row.get("prior_conversion_rate"))
        cr_change_pct = _coerce_numeric(row.get("conversion_rate_change_pct"))
        recent_offer_rate = _coerce_numeric(row.get("recent_featured_offer_rate"))
        prior_offer_rate = _coerce_numeric(row.get("prior_featured_offer_rate"))
        offer_change_pct = _coerce_numeric(row.get("featured_offer_rate_change_pct"))

        if recent_views is not None and prior_views is None and views_change_pct is not None:
            prior_views = _derive_prior_from_change(recent_views, views_change_pct)
        if recent_cr is not None and prior_cr is None and cr_change_pct is not None:
            prior_cr = _derive_prior_from_change(recent_cr, cr_change_pct)
        if recent_offer_rate is not None and prior_offer_rate is None and offer_change_pct is not None:
            prior_offer_rate = _derive_prior_from_change(recent_offer_rate, offer_change_pct)

        diagnosis = "数据不足"
        recommendation = "请结合广告与页面表现继续观察。"
        if recent_offer_rate is not None and recent_offer_rate < 0.9:
            diagnosis = "推荐报价 / Buy Box 风险"
            recommendation = "检查价格、报价竞争、库存和配送方式。"
        elif prior_offer_rate is not None and recent_offer_rate is not None and (prior_offer_rate - recent_offer_rate) >= 0.1:
            diagnosis = "推荐报价 / Buy Box 风险"
            recommendation = "检查价格、报价竞争、库存和配送方式。"
        elif views_change_pct is not None and views_change_pct <= -0.3 and (cr_change_pct is None or cr_change_pct > -0.3):
            diagnosis = "流量问题"
            recommendation = "检查自然曝光、广告承接、关键词覆盖、预算和竞价。"
        elif cr_change_pct is not None and cr_change_pct <= -0.3 and (views_change_pct is None or views_change_pct > -0.3):
            diagnosis = "转化问题"
            recommendation = "检查价格、Coupon、主图、评价、竞品和 Listing。"
        elif views_change_pct is not None and cr_change_pct is not None and views_change_pct <= -0.3 and cr_change_pct <= -0.3:
            diagnosis = "流量和转化同时下降"
            recommendation = "先查价格/Coupon/评价，再查广告承接。"
        elif recent_views is not None and recent_views < 100 and recent_cr is not None and recent_cr >= 0.1:
            diagnosis = "流量不足但转化尚可"
            recommendation = "可小幅提高核心词精准广告或商品投放竞价 10%-15%，观察 2-3 天。"

        if cost_issue:
            recommendation = "先核对售价、采购成本、头程成本、FBA费用；如果利润为负属实，不建议加广告放量，只保留高相关低价精准词，暂停泛词和无转化商品投放。"

        rows.append(
            {
                "marketplace": marketplace,
                "sku": trigger["sku"],
                "asin": trigger["asin"],
                "product_name": trigger["product_name"],
                "recent_period_days": trigger["recent_period_days"],
                "recent_featured_offer_page_views": recent_views,
                "prior_featured_offer_page_views": prior_views,
                "featured_offer_page_views_change_pct": views_change_pct,
                "recent_conversion_rate": recent_cr,
                "prior_conversion_rate": prior_cr,
                "conversion_rate_change_pct": cr_change_pct,
                "recent_featured_offer_rate": recent_offer_rate,
                "prior_featured_offer_rate": prior_offer_rate,
                "featured_offer_rate_change_pct": offer_change_pct,
                "diagnosis": diagnosis,
                "recommendation": recommendation,
                "has_cost_issue": cost_issue,
            }
        )

    return pd.DataFrame(rows)


def build_search_query_opportunities(
    marketplace: str,
    query_df: pd.DataFrame,
    analysis_payload: dict,
) -> pd.DataFrame:
    if query_df.empty:
        return pd.DataFrame(
            columns=[
                "marketplace",
                "type",
                "search_query",
                "asin",
                "product_name",
                "query_impressions",
                "query_clicks",
                "query_cart_adds",
                "query_purchases",
                "suggestion",
                "priority_score",
            ]
        )

    recent = query_df.copy()
    recent["type"] = ""
    recent["suggestion"] = ""
    recent["priority_score"] = 0.0

    ad_term_lookup: dict[str, dict] = {}
    search_summary = analysis_payload.get("搜索词分析", {})
    if isinstance(search_summary, dict):
        for item in search_summary.get("7d", []):
            term = _clean_text(item.get("search_term")).lower()
            if term:
                ad_term_lookup[term] = item

    for idx, row in recent.iterrows():
        impressions = _coerce_numeric(row.get("query_impressions")) or 0.0
        clicks = _coerce_numeric(row.get("query_clicks")) or 0.0
        carts = _coerce_numeric(row.get("query_cart_adds")) or 0.0
        purchases = _coerce_numeric(row.get("query_purchases")) or 0.0

        if purchases > 0:
            recent.at[idx, "type"] = "有购买词"
            recent.at[idx, "suggestion"] = "可加入精准广告或提高关注。"
            recent.at[idx, "priority_score"] = purchases * 100 + clicks
        elif carts > 0 and purchases == 0:
            recent.at[idx, "type"] = "高加购低购买词"
            recent.at[idx, "suggestion"] = "检查价格、页面转化、Coupon。"
            recent.at[idx, "priority_score"] = carts * 50 + clicks
        elif clicks >= 5 and purchases == 0:
            recent.at[idx, "type"] = "高点击无购买词"
            recent.at[idx, "suggestion"] = "广告谨慎投放或降竞价。"
            recent.at[idx, "priority_score"] = clicks * 10 + impressions / 100
        elif impressions >= 100 and clicks <= 1:
            recent.at[idx, "type"] = "高展示低点击词"
            recent.at[idx, "suggestion"] = "检查主图、标题、价格吸引力。"
            recent.at[idx, "priority_score"] = impressions / 10
        else:
            continue

        term = _clean_text(row.get("search_query")).lower()
        if term in ad_term_lookup:
            note = "有购买表现" if _coerce_numeric(ad_term_lookup[term].get("ad_orders")) else "无购买表现"
            recent.at[idx, "suggestion"] = f"{recent.at[idx, 'suggestion']} 该词在广告搜索词中{note}。"

    filtered = recent[recent["type"] != ""].copy()
    if filtered.empty:
        return filtered
    filtered = filtered.sort_values(["priority_score", "query_impressions", "query_clicks"], ascending=[False, False, False]).head(15)
    filtered["marketplace"] = marketplace
    return filtered[
        [
            "marketplace",
            "type",
            "search_query",
            "asin",
            "product_name",
            "query_impressions",
            "query_clicks",
            "query_cart_adds",
            "query_purchases",
            "suggestion",
            "priority_score",
        ]
    ].reset_index(drop=True)


def build_enhanced_bundle(
    marketplace: str,
    custom_dir: Path,
    common_end: date,
    analysis_payload: dict,
) -> EnhancedDataBundle:
    site_dir, scan_dir, site_files, root_files = _resolve_enhanced_scan_dir(custom_dir, marketplace)
    file_records, traffic_df, query_df, timing = _load_dataset_rows(custom_dir, marketplace, common_end)
    diagnosis_paths = _mark_diagnosis_files(file_records)
    traffic_df = _filter_detail_for_diagnosis(traffic_df, diagnosis_paths)
    query_df = _filter_detail_for_diagnosis(query_df, diagnosis_paths)
    request_rows = _build_requests_dataframe(marketplace, file_records, analysis_payload, common_end)
    request_markdown = _request_markdown(marketplace, file_records, request_rows)
    comparable_traffic = _is_comparable(
        next((str(record.get("format_type") or "") for record in file_records if record.get("data_type") == "traffic_sales"), "unknown"),
        file_records,
        "traffic_sales",
    )
    enhanced_status = {
        "custom_dir": str(custom_dir),
        "site_dir": str(site_dir),
        "scan_dir": str(scan_dir),
        "target_folder": _target_folder_text(custom_dir, marketplace),
        "provided": bool(file_records),
        "recognized_files": len(file_records),
        "detected_files": file_records,
        "traffic_sales_files": [record for record in file_records if record.get("data_type") == "traffic_sales"],
        "search_query_files": [record for record in file_records if record.get("data_type") == "search_query_performance"],
        "traffic_sales_recent_exists": any(record.get("period_hint") == "recent" for record in file_records if record.get("data_type") == "traffic_sales"),
        "traffic_sales_prior_exists": any(record.get("period_hint") == "prior" for record in file_records if record.get("data_type") == "traffic_sales"),
        "search_query_recent_exists": any(record.get("period_hint") == "recent" for record in file_records if record.get("data_type") == "search_query_performance"),
        "search_query_prior_exists": any(record.get("period_hint") == "prior" for record in file_records if record.get("data_type") == "search_query_performance"),
        "traffic_sales_format": next((record.get("format_type") for record in file_records if record.get("data_type") == "traffic_sales"), "unknown"),
        "search_query_format": next((record.get("format_type") for record in file_records if record.get("data_type") == "search_query_performance"), "unknown"),
        "traffic_sales_comparable": comparable_traffic,
        "search_query_comparable": _is_comparable(
            next((str(record.get("format_type") or "") for record in file_records if record.get("data_type") == "search_query_performance"), "unknown"),
            file_records,
            "search_query_performance",
        ),
        "request_deduped": True,
        "legacy_root_files_detected": bool(root_files),
        "legacy_root_files": [str(path) for path in root_files],
        "site_folder_files": [str(path) for path in site_files],
        "file_summary": [
            {
                "file_name": record.get("file_name"),
                "marketplace": record.get("marketplace"),
                "data_type": record.get("data_type"),
                "period_hint": record.get("period_hint"),
                "format_type": record.get("format_type"),
                "source_folder": record.get("source_folder"),
                "status": "已导入",
                "used_in_diagnosis": record.get("used_in_diagnosis", "否"),
                "detected_from": record.get("detected_from", "fallback"),
                "detected_date_range": record.get("detected_date_range", ""),
                "freshness": record.get("freshness", "unknown"),
                "recent_week": record.get("recent_week", ""),
                "prior_week": record.get("prior_week", ""),
                "recent_start": record.get("recent_start", ""),
                "recent_end": record.get("recent_end", ""),
                "prior_start": record.get("prior_start", ""),
                "prior_end": record.get("prior_end", ""),
            }
            for record in file_records
        ],
        "common_end": common_end.isoformat(),
        "timing": timing,
    }
    if root_files:
        enhanced_status["legacy_root_warning"] = "检测到旧目录增强文件，建议移动到对应站点文件夹。"

    cost_lookup = _build_cost_lookup(_analysis_records(analysis_payload))
    natural_decline_diagnostics = _build_enhanced_diagnostics(marketplace, analysis_payload, traffic_df, cost_lookup)
    search_query_opportunities = _build_search_query_opportunities(marketplace, query_df, analysis_payload)

    return EnhancedDataBundle(
        status=enhanced_status,
        traffic_sales_detail=traffic_df,
        search_query_detail=query_df,
        natural_decline_diagnostics=natural_decline_diagnostics,
        search_query_opportunities=search_query_opportunities,
        request_rows=request_rows,
        request_markdown=request_markdown,
    )

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .report_presentation import build_report_view
from .analyze_rules import money_symbol_for_marketplace
from .sellersprite_fusion import FUSION_COLUMNS

PERCENT_COLUMNS = {
    "CTR",
    "CVR",
    "ACOS",
    "TACOS",
    "target_acos",
    "suggested_target_acos",
    "break_even_acos",
    "目标 ACOS",
    "target_acos",
    "current_7d_acos",
    "current_7d_target_acos",
    "current_7d_tacos",
    "current_14d_acos",
    "current_14d_tacos",
    "pre_7d_tacos",
    "post_3d_acos",
    "post_3d_tacos",
    "post_7d_acos",
    "post_7d_tacos",
    "seller_sprite_traffic_share",
    "seller_sprite_purchase_rate",
    "ad_acos",
}
CURRENCY_COLUMNS = {
    "spend",
    "ad_sales",
    "total_sales",
    "unit_cost",
    "shipping_cost",
    "handling_fee",
    "profit_before_ads_per_unit",
    "广告前利润/件",
    "花费",
    "销售额",
    "current_7d_spend",
    "current_7d_ad_sales",
    "current_7d_promoted_ad_sales",
    "current_7d_halo_ad_sales",
    "current_14d_spend",
    "current_14d_ad_sales",
    "current_14d_promoted_ad_sales",
    "current_14d_halo_ad_sales",
    "ad_spend",
    "ad_sales",
    "seller_sprite_ppc",
}
COUNT_COLUMNS = {
    "impressions",
    "clicks",
    "ad_orders",
    "total_orders",
    "natural_orders",
    "available_stock",
    "recent_7d_total_orders",
    "recent_14d_total_orders",
    "recent_30d_total_orders",
    "recent_14d_ad_orders",
    "recent_14d_clicks",
    "recent_14d_natural_orders",
    "consecutive_no_order_days",
    "query_impressions",
    "query_clicks",
    "query_cart_adds",
    "query_purchases",
    "featured_offer_page_views",
    "units_ordered",
    "units_shipped",
    "recent_period_days",
    "recent_period_natural_orders",
    "prior_period_natural_orders",
    "natural_order_drop_abs",
    "历史天数",
    "广告行数",
    "ERP行数",
    "SKU数",
    "ASIN数",
    "点击",
    "订单",
    "展示",
    "加购",
    "购买",
    "current_7d_clicks",
    "current_7d_ad_orders",
    "current_7d_promoted_ad_orders",
    "current_7d_halo_ad_orders",
    "current_7d_total_orders",
    "current_7d_available_stock",
    "current_14d_clicks",
    "current_14d_ad_orders",
    "current_14d_promoted_ad_orders",
    "current_14d_halo_ad_orders",
    "current_14d_total_orders",
    "current_14d_available_stock",
    "pre_7d_promoted_ad_orders",
    "pre_7d_total_orders",
    "post_3d_days",
    "post_3d_promoted_ad_orders",
    "post_3d_total_orders",
    "post_3d_available_stock",
    "post_7d_days",
    "post_7d_promoted_ad_orders",
    "post_7d_total_orders",
    "post_7d_available_stock",
    "seller_sprite_monthly_searches",
    "seller_sprite_purchases",
    "seller_sprite_natural_rank",
    "seller_sprite_spr",
}
TODAY_TASK_QUEUE_COLUMNS = [
    "marketplace",
    "sku",
    "asin",
    "priority",
    "issue_type",
    "action_group",
    "today_action",
    "search_term_or_target",
    "suggested_action",
    "normalized_action",
    "action_id",
    "product_name",
    "confirmed_status",
    "primary_reason",
    "key_evidence",
    "tomorrow_check",
    "source_section",
    "copy_action_line",
    "copy_block",
    "action_scope",
    "final_decision",
    "final_decision_label",
    "today_allowed_actions",
    "today_blocked_actions",
    "fusion_action_gate",
    "fusion_today_action",
    "fusion_do_not_do",
    "fusion_review_window",
    "review_status",
    "why_still_active",
    "downgrade_condition",
]
FILL_RED = PatternFill(fill_type="solid", fgColor="FDE2E1")
FILL_GREEN = PatternFill(fill_type="solid", fgColor="E2F4E8")
FILL_YELLOW = PatternFill(fill_type="solid", fgColor="FFF4CC")


def _frame_currency_symbol(frame: pd.DataFrame) -> str:
    marketplace = None
    currency = None
    if "currency" in frame.columns:
        values = [value for value in frame["currency"].tolist() if value not in (None, "")]
        if values:
            currency = values[0]
    if "marketplace" in frame.columns:
        values = [value for value in frame["marketplace"].tolist() if value not in (None, "")]
        if values:
            marketplace = values[0]
    return money_symbol_for_marketplace(marketplace=marketplace, currency=currency)


def _apply_number_format(sheet, frame: pd.DataFrame) -> None:
    money_format = f'"{_frame_currency_symbol(frame)}"#,##0.00'
    for col_idx, column_name in enumerate(frame.columns, start=1):
        if column_name in PERCENT_COLUMNS:
            fmt = "0.0%"
        elif column_name in CURRENCY_COLUMNS:
            fmt = money_format
        elif column_name in COUNT_COLUMNS:
            fmt = "0"
        else:
            continue
        for row_idx in range(2, sheet.max_row + 1):
            sheet.cell(row=row_idx, column=col_idx).number_format = fmt


def _apply_row_fill(sheet, frame: pd.DataFrame, title: str) -> None:
    if frame.empty:
        return
    level_column = None
    for candidate in ["优先级", "风险等级", "板块", "分析状态", "状态"]:
        if candidate in frame.columns:
            level_column = candidate
            break
    if level_column is None:
        return
    level_idx = frame.columns.get_loc(level_column) + 1
    for row_idx in range(2, sheet.max_row + 1):
        value = str(sheet.cell(row=row_idx, column=level_idx).value or "")
        if "今日必须处理" in value or "高优先级" in value:
            fill = FILL_RED
        elif "严重风险" in value:
            fill = FILL_RED
        elif "可以放量" in value:
            fill = FILL_GREEN
        elif "明天观察" in value or "中度风险" in value or "仅广告数据" in value or "仅 ERP 数据" in value:
            fill = FILL_YELLOW
        else:
            continue
        for col_idx in range(1, sheet.max_column + 1):
            sheet.cell(row=row_idx, column=col_idx).fill = fill


def _append_dataframe_sheet(workbook: Workbook, title: str, frame: pd.DataFrame) -> None:
    sheet = workbook.create_sheet(title=title[:31])
    if frame.empty:
        sheet.append(["暂无数据"])
        sheet.freeze_panes = "A2"
        return

    def _excel_cell_value(value):
        if isinstance(value, (list, dict)):
            return str(value)
        if pd.isna(value):
            return None
        return value

    sheet.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        sheet.append([_excel_cell_value(value) for value in row])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"

    for index, column_name in enumerate(frame.columns, start=1):
        values = [len(str(column_name))] + [len(str(value)) for value in frame.iloc[:, index - 1].tolist()]
        sheet.column_dimensions[get_column_letter(index)].width = min(max(values) + 2, 48)

    _apply_number_format(sheet, frame)
    _apply_row_fill(sheet, frame, title)


def _append_dataframe_sheet_keep_headers(workbook: Workbook, title: str, frame: pd.DataFrame) -> None:
    if frame.empty and list(frame.columns):
        frame = pd.DataFrame([{column: "" for column in frame.columns}])
    _append_dataframe_sheet(workbook, title, frame)


def _ensure_frame_columns(frame: pd.DataFrame, required_columns: list[str]) -> pd.DataFrame:
    if frame.empty and not list(frame.columns):
        return pd.DataFrame(columns=required_columns)
    frame = frame.copy()
    for column in required_columns:
        if column not in frame.columns:
            frame[column] = ""
    ordered_columns = [column for column in required_columns if column in frame.columns]
    ordered_columns.extend(column for column in frame.columns if column not in ordered_columns)
    return frame.loc[:, ordered_columns]


def _add_sheet_index(workbook: Workbook) -> None:
    rows: list[dict[str, str]] = []
    daily_keywords = ["总览", "今日动作清单", "明日复查清单", "广告处理队列", "Listing待确认", "成本利润诊断", "Metrics_Validation", "增强数据请求"]
    evidence_keywords = ["产品汇总", "搜索词明细", "搜索词建议", "搜索查询", "自定义流量", "滞销风险", "广告消耗无转化", "库存利润压力", "近期转化断崖", "真无单", "广告无转化", "广告归因弱", "无单原因", "自然单下降"]
    for idx, name in enumerate(workbook.sheetnames, start=1):
        if name == "索引":
            continue
        if any(keyword in name for keyword in daily_keywords):
            group = "每日必看"
            note = "日常运营优先查看"
        elif any(keyword in name for keyword in evidence_keywords):
            group = "明细证据"
            note = "用于复盘和排查原因"
        else:
            group = "配置/历史明细"
            note = "辅助检查"
        rows.append({"顺序": str(idx), "Sheet": name, "分组": group, "用途": note})
    frame = pd.DataFrame(rows)
    _append_dataframe_sheet(workbook, "索引", frame)
    sheet = workbook["索引"]
    workbook._sheets.remove(sheet)
    workbook._sheets.insert(0, sheet)


def _json_list_to_frame(items: list[dict]) -> pd.DataFrame:
    if not items:
        return pd.DataFrame()
    rows = []
    for item in items:
        row = {}
        for key, value in item.items():
            row[key] = value if not isinstance(value, dict) else str(value)
        rows.append(row)
    return pd.DataFrame(rows)


def _window_payload_to_frame(window_payload: dict[str, list[dict]]) -> pd.DataFrame:
    frames = []
    for window_name, items in window_payload.items():
        frame = pd.DataFrame(items)
        if frame.empty:
            continue
        frame.insert(0, "window", window_name)
        frames.append(frame)
    return _concat_frames_preserving_columns(frames)


def _concat_frames_preserving_columns(frames: list[pd.DataFrame]) -> pd.DataFrame:
    column_order: list[object] = []
    prepared: list[pd.DataFrame] = []
    for frame in frames:
        if frame.empty:
            continue
        for column in frame.columns:
            if column not in column_order:
                column_order.append(column)
        trimmed = frame.dropna(axis=1, how="all")
        if not trimmed.empty:
            prepared.append(trimmed)
    if not prepared:
        return pd.DataFrame(columns=column_order)
    result = pd.concat(prepared, ignore_index=True)
    for column in column_order:
        if column not in result.columns:
            result[column] = pd.NA
    return result[column_order]


def _select_prefixed_columns(frame: pd.DataFrame, prefixes: tuple[str, ...]) -> pd.DataFrame:
    if frame.empty:
        return frame
    selected: list[str] = []
    for column in frame.columns:
        name = str(column)
        lower = name.lower()
        if lower in {"marketplace", "sku", "asin", "product_name", "search_query", "type", "suggestion", "priority_score", "status", "file_type", "period_hint", "format_type", "source_file"}:
            selected.append(name)
            continue
        if any(lower.startswith(prefix.lower()) for prefix in prefixes):
            selected.append(name)
    selected = list(dict.fromkeys(selected))
    return frame[selected].copy() if selected else frame.copy()




def _fallback_output_path(output_path: Path) -> Path:
    timestamp = datetime.now().strftime("%H%M%S")
    candidate = output_path.with_name(f"{output_path.stem}_{timestamp}{output_path.suffix}")
    counter = 1
    while candidate.exists():
        candidate = output_path.with_name(f"{output_path.stem}_{timestamp}_{counter}{output_path.suffix}")
        counter += 1
    return candidate


def _save_workbook(workbook: Workbook, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(output_path)
        return output_path
    except PermissionError:
        fallback_path = _fallback_output_path(output_path)
        workbook.save(fallback_path)
        return fallback_path


def _build_single_overview_frame(analysis_payload: dict, report_view: dict) -> pd.DataFrame:
    import_summary = analysis_payload.get("import_summary", {})
    enhanced_status = analysis_payload.get("enhanced_data_status", {})
    return pd.DataFrame(
        [
            {"分组": "本次运行信息", "指标": "报告日期", "值": analysis_payload["report_date"]},
            {"分组": "本次运行信息", "指标": "广告原始文件", "值": analysis_payload["source_files"].get("ads_raw", "")},
            {"分组": "本次运行信息", "指标": "ERP 原始文件", "值": analysis_payload["source_files"].get("erp_raw", "")},
            {"分组": "本次运行信息", "指标": "广告归档文件", "值": analysis_payload["source_files"].get("ads_archive", "")},
            {"分组": "本次运行信息", "指标": "ERP 归档文件", "值": analysis_payload["source_files"].get("erp_archive", "")},
            {"分组": "本次运行信息", "指标": "导入广告行数", "值": import_summary.get("ads_imported_rows", 0)},
            {"分组": "本次运行信息", "指标": "导入 ERP 行数", "值": import_summary.get("erp_imported_rows", 0)},
            {"分组": "本次运行信息", "指标": "新增行数", "值": import_summary.get("added_rows", 0)},
            {"分组": "本次运行信息", "指标": "重复跳过行数", "值": import_summary.get("duplicate_skipped_rows", 0)},
            {"分组": "本次运行信息", "指标": "覆盖更新行数", "值": import_summary.get("overwrite_updated_rows", 0)},
            {"分组": "分析区间", "指标": "广告日期范围", "值": f"{analysis_payload['ads_date_range']['start']} ~ {analysis_payload['ads_date_range']['end']}"},
            {"分组": "分析区间", "指标": "ERP 报表覆盖范围", "值": f"{analysis_payload.get('erp_report_coverage_date_range', analysis_payload['erp_date_range'])['start']} ~ {analysis_payload.get('erp_report_coverage_date_range', analysis_payload['erp_date_range'])['end']}"},
            {"分组": "分析区间", "指标": "ERP 实际有销量日期", "值": f"{analysis_payload.get('erp_observed_sales_date_range', {}).get('start') or '无'} ~ {analysis_payload.get('erp_observed_sales_date_range', {}).get('end') or '无'}"},
            {"分组": "分析区间", "指标": "ERP 补零说明", "值": analysis_payload.get("coverage_warning") or "无需补零"},
            {"分组": "分析区间", "指标": "共同日期范围", "值": f"{analysis_payload['common_date_range']['start']} ~ {analysis_payload['common_date_range']['end']}"},
            {"分组": "分析区间", "指标": "历史天数", "值": analysis_payload["history_days"]},
            {"分组": "增强数据", "指标": "是否提供增强数据", "值": "已提供" if enhanced_status.get("provided") else "未提供"},
            {"分组": "增强数据", "指标": "traffic_sales recent", "值": "已提供" if enhanced_status.get("traffic_sales_recent_exists") else "未提供"},
            {"分组": "增强数据", "指标": "traffic_sales prior", "值": "已提供" if enhanced_status.get("traffic_sales_prior_exists") else "未提供"},
            {"分组": "增强数据", "指标": "search_query recent", "值": "已提供" if enhanced_status.get("search_query_recent_exists") else "未提供"},
            {"分组": "增强数据", "指标": "search_query prior", "值": "已提供" if enhanced_status.get("search_query_prior_exists") else "未提供"},
            {"分组": "摘要", "指标": "Markdown 展示数量", "值": report_view["displayed_count"]},
            {"分组": "摘要", "指标": "隐藏低优先级数量", "值": report_view["hidden_count"]},
        ]
    )


def _build_single_summary_frame(report_view: dict) -> pd.DataFrame:
    rows: list[dict[str, str]] = []
    for sentence in report_view["summary_lines"]:
        rows.append({"板块": "今日总判断", "内容": sentence})
    rows.append(
        {
            "板块": "数据质量",
            "内容": "✅ 数据质量通过，可以用于运营判断" if report_view["quality_pass"] else "⚠️ 数据质量存在问题，本报告需谨慎使用",
        }
    )
    rows.append({"板块": "数据质量", "内容": f"低优先级隐藏项 {report_view['hidden_count']} 条，完整明细请看 Excel。"})
    for row in report_view["today_rows"]:
        rows.append({"板块": "今日必须处理", "内容": f"{row['对象']}：{row['建议动作']}"})
    for row in report_view["scale_rows"]:
        rows.append({"板块": "可以放量", "内容": f"{row['产品']}：{row['建议']}"})
    for row in report_view["watch_rows"]:
        rows.append({"板块": "明天观察", "内容": f"{row['产品']}：{row['明天看什么']}"})
    for row in report_view["cost_rows"]:
        rows.append({"板块": "成本/定价异常", "内容": f"{row['SKU']}：{row['问题']}。{row['建议']}"})
    return pd.DataFrame(rows)


def generate_excel_report(output_path: Path, report_date, analysis_payload: dict, views) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    report_view = build_report_view(analysis_payload)

    traffic_frame = pd.DataFrame(analysis_payload.get("custom_traffic_sales", []))
    traffic_compare_frame = _select_prefixed_columns(
        traffic_frame,
        (
            "recent_",
            "prior_",
            "conversion_rate_",
            "featured_offer_",
            "units_",
        ),
    )
    query_frame = pd.DataFrame(analysis_payload.get("custom_search_query_performance", []))
    search_opportunity_frame = pd.DataFrame(analysis_payload.get("search_query_opportunities", []))
    natural_frame = pd.DataFrame(analysis_payload.get("natural_decline_enhanced_diagnostics", []))
    risk_frame = pd.DataFrame(analysis_payload.get("滞销风险", []))
    ad_no_conversion_frame = pd.DataFrame(analysis_payload.get("广告消耗无转化风险", []))
    inventory_profit_frame = pd.DataFrame(analysis_payload.get("库存 / 利润压力风险", []))
    search_term_suggestion_frame = pd.DataFrame(report_view.get("search_term_suggestion_rows", []))
    search_term_processing_frame = pd.DataFrame(report_view.get("search_term_processing_queue_rows", []))
    seller_sprite_ads_fusion_frame = pd.DataFrame(
        report_view.get("seller_sprite_ads_fusion_rows", []),
        columns=FUSION_COLUMNS,
    )
    no_order_diagnosis_frame = pd.DataFrame(analysis_payload.get("无单原因诊断", []))
    inventory_replenishment_frame = pd.DataFrame(analysis_payload.get("inventory_replenishment", {}).get("rows", []))
    product_final_decision_frame = pd.DataFrame(report_view.get("product_final_decision_rows", []))
    product_operation_card_frame = pd.DataFrame(report_view.get("product_operation_cards", []))
    frontend_check_queue_frame = pd.DataFrame(report_view.get("frontend_check_queue_rows", []))
    market_survey_fetch_plan_frame = pd.DataFrame(report_view.get("market_survey_selective_fetch_plan", []))
    today_task_queue_frame = _ensure_frame_columns(
        pd.DataFrame(report_view.get("today_task_queue_rows", [])),
        TODAY_TASK_QUEUE_COLUMNS,
    )
    tomorrow_review_frame = pd.DataFrame(report_view.get("tomorrow_review_rows", []))
    true_unsold_diagnosis_frame = pd.DataFrame(report_view.get("true_unsold_diagnosis_rows", []))
    recent_conversion_cliff_diagnosis_frame = pd.DataFrame(report_view.get("recent_conversion_cliff_diagnosis_rows", []))
    ad_no_conversion_diagnosis_frame = pd.DataFrame(report_view.get("ad_no_conversion_diagnosis_rows", []))
    ad_attribution_weak_diagnosis_frame = pd.DataFrame(report_view.get("ad_attribution_weak_diagnosis_rows", []))
    listing_price_diagnosis_frame = pd.DataFrame(report_view.get("listing_price_diagnosis_rows", []))
    cost_profit_diagnosis_frame = pd.DataFrame(report_view.get("cost_profit_diagnosis_rows", []))
    request_frame = pd.DataFrame(analysis_payload.get("enhanced_data_requests", []))
    scale_candidate_frame = pd.DataFrame(report_view.get("scale_rows", []))
    scale_keyword_frame = pd.DataFrame(report_view.get("scale_keyword_rows", []))
    growth_test_frame = pd.DataFrame(report_view.get("growth_test_rows", []))

    _append_dataframe_sheet(workbook, "今日总览", _build_single_overview_frame(analysis_payload, report_view))
    _append_dataframe_sheet_keep_headers(workbook, "今日动作清单", today_task_queue_frame)
    _append_dataframe_sheet(workbook, "明日复查清单", tomorrow_review_frame)
    _append_dataframe_sheet(workbook, "库存补货提醒", inventory_replenishment_frame)
    _append_dataframe_sheet(workbook, "产品最终决策", product_final_decision_frame)
    _append_dataframe_sheet(workbook, "产品运营卡", product_operation_card_frame)
    _append_dataframe_sheet(workbook, "前台证据队列", frontend_check_queue_frame)
    _append_dataframe_sheet(workbook, "市场调查采集计划", market_survey_fetch_plan_frame)
    _append_dataframe_sheet(workbook, "昨日异常归因摘要", pd.DataFrame(report_view.get("yesterday_attribution_rows", [])))
    _append_dataframe_sheet(workbook, "执行后效果复盘", pd.DataFrame(report_view.get("action_effect_review_rows", [])))
    _append_dataframe_sheet(workbook, "词级执行复盘", pd.DataFrame(report_view.get("keyword_action_effect_review_rows", [])))
    _append_dataframe_sheet(workbook, "放量候选", scale_candidate_frame)
    _append_dataframe_sheet(workbook, "放量词候选", scale_keyword_frame)
    _append_dataframe_sheet(workbook, "小预算试投", growth_test_frame)
    _append_dataframe_sheet(workbook, "运营日报摘要", _build_single_summary_frame(report_view))
    _append_dataframe_sheet(workbook, "产品汇总", _window_payload_to_frame(analysis_payload["产品汇总"]))
    _append_dataframe_sheet(workbook, "广告活动汇总", _window_payload_to_frame(analysis_payload["广告活动汇总"]))
    _append_dataframe_sheet(workbook, "搜索词分析", _window_payload_to_frame(analysis_payload["搜索词分析"]))
    _append_dataframe_sheet(workbook, "异常提醒", _json_list_to_frame(analysis_payload["异常提醒"]))
    _append_dataframe_sheet(workbook, "操作建议", _json_list_to_frame(analysis_payload["操作建议"]))
    _append_dataframe_sheet(workbook, "SKU映射检查", pd.DataFrame(analysis_payload.get("sku_mapping_check", [])))
    _append_dataframe_sheet(workbook, "自定义流量销售", traffic_frame)
    _append_dataframe_sheet(workbook, "自定义流量销售对比", traffic_compare_frame)
    _append_dataframe_sheet(workbook, "搜索查询绩效", query_frame)
    _append_dataframe_sheet(workbook, "搜索查询机会", search_opportunity_frame)
    _append_dataframe_sheet(workbook, "自然单下降增强诊断", natural_frame)
    _append_dataframe_sheet(workbook, "滞销风险", risk_frame)
    _append_dataframe_sheet(workbook, "广告消耗无转化风险", ad_no_conversion_frame)
    _append_dataframe_sheet(workbook, "库存利润压力风险", inventory_profit_frame)
    _append_dataframe_sheet(workbook, "搜索词建议", search_term_suggestion_frame)
    _append_dataframe_sheet(workbook, "广告处理队列", search_term_processing_frame)
    _append_dataframe_sheet_keep_headers(workbook, "卖家精灵广告融合明细", seller_sprite_ads_fusion_frame)
    _append_dataframe_sheet(workbook, "近期转化断崖诊断", recent_conversion_cliff_diagnosis_frame)
    _append_dataframe_sheet(workbook, "真无单滞销诊断", true_unsold_diagnosis_frame)
    _append_dataframe_sheet(workbook, "广告无转化诊断", ad_no_conversion_diagnosis_frame)
    _append_dataframe_sheet(workbook, "广告归因弱诊断", ad_attribution_weak_diagnosis_frame)
    _append_dataframe_sheet(workbook, "Listing待确认", listing_price_diagnosis_frame)
    _append_dataframe_sheet(workbook, "成本利润诊断", cost_profit_diagnosis_frame)
    _append_dataframe_sheet(workbook, "无单原因诊断", no_order_diagnosis_frame)
    _append_dataframe_sheet(workbook, "增强数据请求", request_frame)
    _add_sheet_index(workbook)
    return _save_workbook(workbook, output_path)



def _build_all_overview_frame(results: list[dict], source_files: dict, import_summary: dict, report_date: str) -> pd.DataFrame:
    frontend_coverage_totals = {
        "frontend_queue_total": 0,
        "frontend_product_page_success_count": 0,
        "frontend_competitor_search_success_count": 0,
        "frontend_own_sellersprite_count": 0,
        "frontend_own_sellersprite_today_count": 0,
        "frontend_own_sellersprite_cache_count": 0,
        "frontend_own_sellersprite_pending_count": 0,
        "frontend_own_sellersprite_failed_count": 0,
        "frontend_sellersprite_trend_ready_count": 0,
        "frontend_competitor_discovery_count": 0,
        "frontend_competitor_pool_count": 0,
        "frontend_competitor_pool_today_count": 0,
        "frontend_competitor_pool_cache_count": 0,
        "frontend_competitor_pool_pending_count": 0,
        "frontend_competitor_pool_failed_count": 0,
        "frontend_competitor_sellersprite_count": 0,
        "frontend_competitor_sellersprite_today_count": 0,
        "frontend_competitor_sellersprite_cache_count": 0,
        "frontend_competitor_sellersprite_pending_count": 0,
        "frontend_competitor_sellersprite_asin_count": 0,
        "frontend_amazon_search_validation_count": 0,
        "frontend_scalable_strong_count": 0,
        "frontend_weak_defensive_count": 0,
        "frontend_insufficient_count": 0,
        "market_survey_complete_count": 0,
        "market_survey_usable_count": 0,
        "market_survey_insufficient_count": 0,
        "market_survey_failed_count": 0,
    }
    market_survey_score_total = 0.0
    rows = [
        {"分组": "本次运行信息", "指标": "报告日期", "值": report_date},
        {"分组": "本次运行信息", "指标": "广告原始文件", "值": source_files.get("ads_raw", "")},
        {"分组": "本次运行信息", "指标": "ERP 原始文件", "值": source_files.get("erp_raw", "")},
        {"分组": "本次运行信息", "指标": "广告归档文件", "值": source_files.get("ads_archive", "")},
        {"分组": "本次运行信息", "指标": "ERP 归档文件", "值": source_files.get("erp_archive", "")},
        {"分组": "本次运行信息", "指标": "导入广告行数", "值": import_summary.get("ads_imported_rows", 0)},
        {"分组": "本次运行信息", "指标": "导入 ERP 行数", "值": import_summary.get("erp_imported_rows", 0)},
        {"分组": "本次运行信息", "指标": "新增行数", "值": import_summary.get("added_rows", 0)},
        {"分组": "本次运行信息", "指标": "重复跳过行数", "值": import_summary.get("duplicate_skipped_rows", 0)},
        {"分组": "本次运行信息", "指标": "覆盖更新行数", "值": import_summary.get("overwrite_updated_rows", 0)},
    ]
    for result in results:
        summary = result["summary"]
        coverage = (result.get("report_view") or {}).get("frontend_coverage_summary", {})
        if isinstance(coverage, dict):
            for key in frontend_coverage_totals:
                frontend_coverage_totals[key] += int(float(coverage.get(key, 0) or 0))
            market_survey_score_total += float(coverage.get("market_survey_average_score", 0) or 0) * int(float(coverage.get("frontend_queue_total", 0) or 0))
        rows.extend(
            [
                {"分组": f"{summary['marketplace']} 状态", "指标": "广告行数", "值": summary["ads_row_count"]},
                {"分组": f"{summary['marketplace']} 状态", "指标": "ERP 行数", "值": summary["erp_row_count"]},
                {"分组": f"{summary['marketplace']} 状态", "指标": "SKU 数", "值": summary["sku_count"]},
                {"分组": f"{summary['marketplace']} 状态", "指标": "ASIN 数", "值": summary["asin_count"]},
                {"分组": f"{summary['marketplace']} 状态", "指标": "广告日期范围", "值": summary["ads_date_range"]},
                {"分组": f"{summary['marketplace']} 状态", "指标": "ERP 报表覆盖范围", "值": summary.get("erp_report_coverage_date_range", summary["erp_date_range"])},
                {"分组": f"{summary['marketplace']} 状态", "指标": "ERP 实际有销量日期", "值": summary.get("erp_observed_sales_date_range", summary["erp_date_range"])},
                {"分组": f"{summary['marketplace']} 状态", "指标": "ERP 补零说明", "值": summary.get("coverage_warning", "") or "无需补零"},
            ]
        )
    frontend_total = frontend_coverage_totals["frontend_queue_total"]
    if frontend_total:
        market_survey_average = round(market_survey_score_total / frontend_total, 1)
        rows.extend(
            [
                {"分组": "前台证据覆盖", "指标": "ALL 前台队列", "值": frontend_total},
                {"分组": "前台证据覆盖", "指标": "ALL 市场调查平均完整度", "值": f"{market_survey_average}/100"},
                {"分组": "前台证据覆盖", "指标": "ALL 市场调查完整", "值": f'{frontend_coverage_totals["market_survey_complete_count"]}/{frontend_total}'},
                {"分组": "前台证据覆盖", "指标": "ALL 市场调查可用", "值": f'{frontend_coverage_totals["market_survey_usable_count"]}/{frontend_total}'},
                {"分组": "前台证据覆盖", "指标": "ALL 市场调查待补", "值": f'{frontend_coverage_totals["market_survey_insufficient_count"]}/{frontend_total}'},
                {"分组": "前台证据覆盖", "指标": "ALL 市场调查失败", "值": f'{frontend_coverage_totals["market_survey_failed_count"]}/{frontend_total}'},
                {"分组": "前台证据覆盖", "指标": "ALL 产品页成功", "值": f'{frontend_coverage_totals["frontend_product_page_success_count"]}/{frontend_total}'},
                {
                    "分组": "前台证据覆盖",
                    "指标": "ALL 卖家精灵自己 ASIN",
                    "值": f'今日 {frontend_coverage_totals["frontend_own_sellersprite_today_count"]}/{frontend_total}，缓存 {frontend_coverage_totals["frontend_own_sellersprite_cache_count"]}/{frontend_total}',
                },
                {"分组": "前台证据覆盖", "指标": "ALL 卖家精灵趋势", "值": f'{frontend_coverage_totals["frontend_sellersprite_trend_ready_count"]}/{frontend_total}'},
                {"分组": "前台证据覆盖", "指标": "ALL 卖家精灵竞品发现", "值": f'{frontend_coverage_totals["frontend_competitor_discovery_count"]}/{frontend_total}'},
                {
                    "分组": "前台证据覆盖",
                    "指标": "ALL 卖家精灵竞品池",
                    "值": f'今日 {frontend_coverage_totals["frontend_competitor_pool_today_count"]}/{frontend_total}，7天缓存 {frontend_coverage_totals["frontend_competitor_pool_cache_count"]}/{frontend_total}',
                },
                {
                    "分组": "前台证据覆盖",
                    "指标": "ALL 竞品 ASIN 反查",
                    "值": f'今日 {frontend_coverage_totals["frontend_competitor_sellersprite_today_count"]}/{frontend_total}，缓存 {frontend_coverage_totals["frontend_competitor_sellersprite_cache_count"]}/{frontend_total}',
                },
                {"分组": "前台证据覆盖", "指标": "ALL Amazon 搜索页辅助验证", "值": f'{frontend_coverage_totals["frontend_amazon_search_validation_count"]}/{frontend_total}'},
                {"分组": "前台证据覆盖", "指标": "ALL 达到放量准入", "值": f'{frontend_coverage_totals["frontend_scalable_strong_count"]}/{frontend_total}'},
                {"分组": "前台证据覆盖", "指标": "ALL 弱势止损证据", "值": f'{frontend_coverage_totals["frontend_weak_defensive_count"]}/{frontend_total}'},
                {"分组": "前台证据覆盖", "指标": "ALL 证据不足", "值": f'{frontend_coverage_totals["frontend_insufficient_count"]}/{frontend_total}'},
            ]
        )
    return pd.DataFrame(rows)


def _build_marketplace_status_frame(result: dict) -> pd.DataFrame:
    summary = result["summary"]
    if result.get("has_data"):
        view = result.get("report_view") or {}
        status = view.get("analysis_status") or "正式分析"
        issue = view.get("issue_summary") or "正常"
    elif summary["ads_row_count"] > 0 and summary["erp_row_count"] == 0:
        status = "仅广告数据"
        issue = f"{summary['marketplace']} 有广告数据，但 ERP 销量数据缺失。"
    elif summary["ads_row_count"] == 0 and summary["erp_row_count"] > 0:
        status = "仅 ERP 数据"
        issue = f"{summary['marketplace']} 有 ERP 数据，但广告数据缺失。"
    else:
        status = "无数据"
        issue = f"{summary['marketplace']} 当前无可分析数据。"
    return pd.DataFrame(
        [
            {"站点": summary["marketplace"], "广告行数": summary["ads_row_count"], "ERP行数": summary["erp_row_count"], "SKU数": summary["sku_count"], "ASIN数": summary["asin_count"], "分析状态": status, "问题": issue}
        ]
    )


def _build_marketplace_today_frame(result: dict) -> pd.DataFrame:
    payload = result["analysis_payload"]
    view = result["report_view"]
    return _build_single_overview_frame(payload, view)


def _pct_diff(a: float, b: float) -> float:
    return abs(a - b) / max(abs(a), abs(b), 1.0)


def _build_metrics_validation_frame(results: list[dict]) -> pd.DataFrame:
    rows: list[dict] = []
    for result in results:
        if not result.get("has_data"):
            continue
        payload = result["analysis_payload"]
        metrics = payload.get("product_window_metrics", {})
        rows7 = {(row.get("marketplace"), row.get("sku"), row.get("asin")): row for row in metrics.get("7d", [])}
        rows14 = {(row.get("marketplace"), row.get("sku"), row.get("asin")): row for row in metrics.get("14d", [])}
        rows30 = {(row.get("marketplace"), row.get("sku"), row.get("asin")): row for row in metrics.get("30d", [])}
        report_used: dict[tuple[object, object, object], dict] = {}
        for section in ["滞销风险", "广告消耗无转化风险", "库存 / 利润压力风险"]:
            for row in payload.get(section, []):
                report_used[(row.get("marketplace"), row.get("sku"), row.get("asin"))] = row
        for key, raw14 in rows14.items():
            raw7 = rows7.get(key, {})
            raw30 = rows30.get(key, {})
            used = report_used.get(key, {})
            erp_coverage = payload.get("erp_report_coverage_date_range") or payload.get("erp_date_range") or {}
            erp_observed = payload.get("erp_observed_sales_date_range") or {}
            used_clicks = float(used.get("recent_14d_clicks", raw14.get("ad_clicks") or 0) or 0)
            used_spend = float(used.get("recent_14d_ad_spend", raw14.get("ad_spend") or 0) or 0)
            used_orders = float(used.get("recent_14d_ad_orders", raw14.get("ad_orders") or 0) or 0)
            raw_clicks = float(raw14.get("ad_clicks") or 0)
            raw_spend = float(raw14.get("ad_spend") or 0)
            raw_orders = float(raw14.get("ad_orders") or 0)
            mismatches = []
            if _pct_diff(used_clicks, raw_clicks) > 0.01:
                mismatches.append("14d clicks mismatch")
            if _pct_diff(used_spend, raw_spend) > 0.01:
                mismatches.append("14d spend mismatch")
            if _pct_diff(used_orders, raw_orders) > 0.01:
                mismatches.append("14d orders mismatch")
            rows.append(
                {
                    "marketplace": key[0],
                    "sku": key[1],
                    "asin": key[2],
                    "product_name": raw14.get("product_name"),
                    "raw_7d_ad_clicks": raw7.get("ad_clicks", 0),
                    "raw_7d_ad_spend": raw7.get("ad_spend", 0),
                    "raw_7d_ad_orders": raw7.get("ad_orders", 0),
                    "raw_14d_ad_clicks": raw14.get("ad_clicks", 0),
                    "raw_14d_ad_spend": raw14.get("ad_spend", 0),
                    "raw_14d_ad_orders": raw14.get("ad_orders", 0),
                    "raw_30d_ad_clicks": raw30.get("ad_clicks", 0),
                    "raw_30d_ad_spend": raw30.get("ad_spend", 0),
                    "raw_30d_ad_orders": raw30.get("ad_orders", 0),
                    "raw_7d_total_orders": raw7.get("total_orders", 0),
                    "raw_14d_total_orders": raw14.get("total_orders", 0),
                    "raw_30d_total_orders": raw30.get("total_orders", 0),
                    "used_in_report_14d_clicks": used_clicks,
                    "used_in_report_14d_spend": used_spend,
                    "used_in_report_14d_orders": used_orders,
                    "mismatch_flag": bool(mismatches),
                    "mismatch_reason": "；".join(mismatches),
                    "erp_report_coverage_start": erp_coverage.get("start"),
                    "erp_report_coverage_end": erp_coverage.get("end"),
                    "erp_observed_sales_start": erp_observed.get("start"),
                    "erp_observed_sales_end": erp_observed.get("end"),
                    "erp_zero_filled_days": payload.get("erp_zero_filled_days", 0),
                    "zero_fill_applied": bool(payload.get("zero_fill_applied")),
                    "coverage_warning": payload.get("coverage_warning", ""),
                }
            )
    return pd.DataFrame(rows)


def generate_all_marketplace_excel_report(
    output_path: Path,
    results: list[dict],
    report_date: str,
    source_files: dict,
    import_summary: dict,
) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _append_dataframe_sheet(workbook, "总览", _build_all_overview_frame(results, source_files, import_summary, report_date))
    _append_dataframe_sheet(workbook, "Metrics_Validation", _build_metrics_validation_frame(results))
    _append_dataframe_sheet_keep_headers(
        workbook,
        "今日动作清单",
        _ensure_frame_columns(
            pd.DataFrame(
                [
                    row
                    for result in results
                    for row in result.get("report_view", {}).get("today_task_queue_rows", [])
                ]
            ),
            TODAY_TASK_QUEUE_COLUMNS,
        ),
    )
    _append_dataframe_sheet(
        workbook,
        "明日复查清单",
        pd.DataFrame(
            [
                row
                for result in results
                for row in result.get("report_view", {}).get("tomorrow_review_rows", [])
            ]
        ),
    )
    _append_dataframe_sheet(
        workbook,
        "产品最终决策",
        pd.DataFrame(
            [
                row
                for result in results
                for row in result.get("report_view", {}).get("product_final_decision_rows", [])
            ]
        ),
    )
    _append_dataframe_sheet(
        workbook,
        "产品运营卡",
        pd.DataFrame(
            [
                row
                for result in results
                for row in result.get("report_view", {}).get("product_operation_cards", [])
            ]
        ),
    )
    _append_dataframe_sheet(
        workbook,
        "前台证据队列",
        pd.DataFrame(
            [
                row
                for result in results
                for row in result.get("report_view", {}).get("frontend_check_queue_rows", [])
            ]
        ),
    )
    _append_dataframe_sheet(
        workbook,
        "Listing待确认",
        pd.DataFrame(
            [
                row
                for result in results
                for row in result.get("report_view", {}).get("listing_price_diagnosis_rows", [])
            ]
        ),
    )
    _append_dataframe_sheet(
        workbook,
        "成本利润诊断",
        pd.DataFrame(
            [
                row
                for result in results
                for row in result.get("report_view", {}).get("cost_profit_diagnosis_rows", [])
            ]
        ),
    )
    _append_dataframe_sheet(
        workbook,
        "执行后效果复盘",
        pd.DataFrame(
            [
                row
                for result in results
                for row in result.get("report_view", {}).get("action_effect_review_rows", [])
            ]
        ),
    )
    _append_dataframe_sheet(
        workbook,
        "词级执行复盘",
        pd.DataFrame(
            [
                row
                for result in results
                for row in result.get("report_view", {}).get("keyword_action_effect_review_rows", [])
            ]
        ),
    )
    _append_dataframe_sheet(
        workbook,
        "放量候选",
        pd.DataFrame(
            [
                row
                for result in results
                for row in result.get("report_view", {}).get("scale_rows", [])
            ]
        ),
    )
    _append_dataframe_sheet(
        workbook,
        "放量词候选",
        pd.DataFrame(
            [
                row
                for result in results
                for row in result.get("report_view", {}).get("scale_keyword_rows", [])
            ]
        ),
    )
    _append_dataframe_sheet(
        workbook,
        "广告处理队列",
        pd.DataFrame(
            [
                row
                for result in results
                for row in result.get("report_view", {}).get("search_term_processing_queue_rows", [])
            ]
        ),
    )
    _append_dataframe_sheet_keep_headers(
        workbook,
        "卖家精灵广告融合明细",
        pd.DataFrame(
            [
                row
                for result in results
                for row in result.get("report_view", {}).get("seller_sprite_ads_fusion_rows", [])
            ],
            columns=FUSION_COLUMNS,
        ),
    )
    _append_dataframe_sheet(
        workbook,
        "小预算试投",
        pd.DataFrame(
            [
                row
                for result in results
                for row in result.get("report_view", {}).get("growth_test_rows", [])
            ]
        ),
    )

    for result in results:
        marketplace = result["marketplace"]
        if result.get("has_data"):
            payload = result["analysis_payload"]
            traffic_frame = pd.DataFrame(payload.get("custom_traffic_sales", []))
            traffic_compare_frame = _select_prefixed_columns(
                traffic_frame,
                (
                    "recent_",
                    "prior_",
                    "conversion_rate_",
                    "featured_offer_",
                    "units_",
                ),
            )
            query_frame = pd.DataFrame(payload.get("custom_search_query_performance", []))
            search_opportunity_frame = pd.DataFrame(payload.get("search_query_opportunities", []))
            natural_frame = pd.DataFrame(payload.get("natural_decline_enhanced_diagnostics", []))
            risk_frame = pd.DataFrame(payload.get("滞销风险", []))
            ad_no_conversion_frame = pd.DataFrame(payload.get("广告消耗无转化风险", []))
            inventory_profit_frame = pd.DataFrame(payload.get("库存 / 利润压力风险", []))
            search_term_suggestion_frame = pd.DataFrame(result["report_view"].get("search_term_suggestion_rows", []))
            search_term_processing_frame = pd.DataFrame(result["report_view"].get("search_term_processing_queue_rows", []))
            seller_sprite_ads_fusion_frame = pd.DataFrame(
                result["report_view"].get("seller_sprite_ads_fusion_rows", []),
                columns=FUSION_COLUMNS,
            )
            no_order_diagnosis_frame = pd.DataFrame(payload.get("无单原因诊断", []))
            inventory_replenishment_frame = pd.DataFrame(payload.get("inventory_replenishment", {}).get("rows", []))
            product_final_decision_frame = pd.DataFrame(result["report_view"].get("product_final_decision_rows", []))
            product_operation_card_frame = pd.DataFrame(result["report_view"].get("product_operation_cards", []))
            frontend_check_queue_frame = pd.DataFrame(result["report_view"].get("frontend_check_queue_rows", []))
            market_survey_fetch_plan_frame = pd.DataFrame(result["report_view"].get("market_survey_selective_fetch_plan", []))
            today_task_queue_frame = _ensure_frame_columns(
                pd.DataFrame(result["report_view"].get("today_task_queue_rows", [])),
                TODAY_TASK_QUEUE_COLUMNS,
            )
            tomorrow_review_frame = pd.DataFrame(result["report_view"].get("tomorrow_review_rows", []))
            yesterday_attribution_frame = pd.DataFrame(result["report_view"].get("yesterday_attribution_rows", []))
            action_effect_review_frame = pd.DataFrame(
                [
                    row
                    for row in result["report_view"].get("action_effect_review_rows", [])
                    if str(row.get("marketplace") or "").upper() == str(marketplace).upper()
                ]
            )
            scale_candidate_frame = pd.DataFrame(result["report_view"].get("scale_rows", []))
            true_unsold_diagnosis_frame = pd.DataFrame(result["report_view"].get("true_unsold_diagnosis_rows", []))
            recent_conversion_cliff_diagnosis_frame = pd.DataFrame(result["report_view"].get("recent_conversion_cliff_diagnosis_rows", []))
            ad_no_conversion_diagnosis_frame = pd.DataFrame(result["report_view"].get("ad_no_conversion_diagnosis_rows", []))
            ad_attribution_weak_diagnosis_frame = pd.DataFrame(result["report_view"].get("ad_attribution_weak_diagnosis_rows", []))
            listing_price_diagnosis_frame = pd.DataFrame(result["report_view"].get("listing_price_diagnosis_rows", []))
            cost_profit_diagnosis_frame = pd.DataFrame(result["report_view"].get("cost_profit_diagnosis_rows", []))
            request_frame = pd.DataFrame(payload.get("enhanced_data_requests", []))
            scale_keyword_frame = pd.DataFrame(result["report_view"].get("scale_keyword_rows", []))
            growth_test_frame = pd.DataFrame(result["report_view"].get("growth_test_rows", []))

            _append_dataframe_sheet(workbook, f"{marketplace}_今日总览", _build_marketplace_today_frame(result))
            _append_dataframe_sheet_keep_headers(workbook, f"{marketplace}_今日动作清单", today_task_queue_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_明日复查清单", tomorrow_review_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_库存补货提醒", inventory_replenishment_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_产品最终决策", product_final_decision_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_产品运营卡", product_operation_card_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_前台证据队列", frontend_check_queue_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_市场调查采集计划", market_survey_fetch_plan_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_昨日异常归因", yesterday_attribution_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_执行后复盘", action_effect_review_frame)
            _append_dataframe_sheet(
                workbook,
                f"{marketplace}_词级执行复盘",
                pd.DataFrame(
                    [
                        row
                        for row in result["report_view"].get("keyword_action_effect_review_rows", [])
                        if str(row.get("marketplace") or "").upper() == str(marketplace).upper()
                    ]
                ),
            )
            _append_dataframe_sheet(workbook, f"{marketplace}_放量候选", scale_candidate_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_放量词候选", scale_keyword_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_小预算试投", growth_test_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_操作建议", _json_list_to_frame(payload.get("操作建议", [])))
            _append_dataframe_sheet(workbook, f"{marketplace}_搜索词明细", _window_payload_to_frame(payload.get("搜索词分析", {})))
            _append_dataframe_sheet(workbook, f"{marketplace}_产品汇总", _window_payload_to_frame(payload.get("产品汇总", {})))
            _append_dataframe_sheet(workbook, f"{marketplace}_自定义流量销售", traffic_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_自定义流量销售对比", traffic_compare_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_搜索查询绩效", query_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_搜索查询机会", search_opportunity_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_自然单下降增强诊断", natural_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_滞销风险", risk_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_广告消耗无转化风险", ad_no_conversion_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_库存利润压力风险", inventory_profit_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_搜索词建议", search_term_suggestion_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_广告处理队列", search_term_processing_frame)
            _append_dataframe_sheet_keep_headers(workbook, f"{marketplace}_卖家精灵融合明细", seller_sprite_ads_fusion_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_近期转化断崖诊断", recent_conversion_cliff_diagnosis_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_真无单滞销诊断", true_unsold_diagnosis_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_广告无转化诊断", ad_no_conversion_diagnosis_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_广告归因弱诊断", ad_attribution_weak_diagnosis_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_Listing待确认", listing_price_diagnosis_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_成本利润诊断", cost_profit_diagnosis_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_无单原因诊断", no_order_diagnosis_frame)
            _append_dataframe_sheet(workbook, f"{marketplace}_增强数据请求", request_frame)
        else:
            _append_dataframe_sheet(workbook, f"{marketplace}_数据状态", _build_marketplace_status_frame(result))

    all_requests = []
    for result in results:
        all_requests.extend(result.get("analysis_payload", {}).get("enhanced_data_requests", []))
    _append_dataframe_sheet(workbook, "增强数据请求", pd.DataFrame(all_requests))
    all_inventory = []
    for result in results:
        all_inventory.extend(result.get("analysis_payload", {}).get("inventory_replenishment", {}).get("rows", []))
    _append_dataframe_sheet(workbook, "库存补货提醒", pd.DataFrame(all_inventory))
    _add_sheet_index(workbook)
    return _save_workbook(workbook, output_path)

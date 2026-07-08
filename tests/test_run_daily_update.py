from __future__ import annotations

import html
import json

from openpyxl import Workbook, load_workbook

import scripts.run_daily_update as daily_update


def _excel_value(value: object) -> object:
    if isinstance(value, (list, tuple, set, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _write_workbook(path, sheet_names: list[str]) -> None:
    workbook = Workbook()
    workbook.active.title = sheet_names[0]
    for sheet_name in sheet_names[1:]:
        workbook.create_sheet(sheet_name)
    workbook.save(path)


def _append_rows_sheet(workbook: Workbook, sheet_name: str, rows: list[dict[str, object]] | None = None) -> None:
    rows = rows or []
    sheet = workbook.create_sheet(sheet_name)
    headers = sorted({field for row in rows for field in row.keys()}) or ["marketplace", "sku", "asin"]
    sheet.append(headers)
    for row in rows:
        sheet.append([_excel_value(row.get(header, "")) for header in headers])


def _write_daily_report_workbook(
    path,
    report_date: str = "2026-06-24",
    *,
    today_task_rows: list[dict[str, object]] | None = None,
    tomorrow_review_rows: list[dict[str, object]] | None = None,
    listing_diagnosis_rows: list[dict[str, object]] | None = None,
    cost_diagnosis_rows: list[dict[str, object]] | None = None,
    product_rows: list[dict[str, object]] | None = None,
    product_operation_rows: list[dict[str, object]] | None = None,
    inventory_rows: list[dict[str, object]] | None = None,
    frontend_rows: list[dict[str, object]] | None = None,
    ad_processing_rows: list[dict[str, object]] | None = None,
    scale_rows: list[dict[str, object]] | None = None,
    scale_keyword_rows: list[dict[str, object]] | None = None,
    growth_test_rows: list[dict[str, object]] | None = None,
    action_review_rows: list[dict[str, object]] | None = None,
    keyword_review_rows: list[dict[str, object]] | None = None,
) -> None:
    workbook = Workbook()
    workbook.active.title = "总览"
    workbook["总览"].append(["分组", "指标", "值"])
    workbook["总览"].append(["本次运行信息", "报告日期", report_date])
    for sheet_name in ["Metrics_Validation", "UK_今日总览", "US_今日总览", "DE_今日总览"]:
        workbook.create_sheet(sheet_name)
    _append_rows_sheet(workbook, "今日动作清单", today_task_rows)
    _append_rows_sheet(workbook, "明日复查清单", tomorrow_review_rows)
    _append_rows_sheet(workbook, "Listing待确认", listing_diagnosis_rows)
    _append_rows_sheet(workbook, "成本利润诊断", cost_diagnosis_rows)
    _append_rows_sheet(workbook, "产品最终决策", product_rows)
    if product_operation_rows is None and product_rows:
        product_operation_rows = [_product_operation_card(row) for row in product_rows]
    _append_rows_sheet(workbook, "产品运营卡", product_operation_rows)
    _append_rows_sheet(workbook, "库存补货提醒", inventory_rows)
    _append_rows_sheet(workbook, "前台证据队列", frontend_rows)
    _append_rows_sheet(workbook, "广告处理队列", ad_processing_rows)
    _append_rows_sheet(workbook, "放量候选", scale_rows)
    _append_rows_sheet(workbook, "放量词候选", scale_keyword_rows)
    _append_rows_sheet(workbook, "小预算试投", growth_test_rows)
    _append_rows_sheet(workbook, "执行后效果复盘", action_review_rows)
    _append_rows_sheet(workbook, "词级执行复盘", keyword_review_rows)
    today_task_rows = today_task_rows or []
    tomorrow_review_rows = tomorrow_review_rows or []
    listing_diagnosis_rows = listing_diagnosis_rows or []
    cost_diagnosis_rows = cost_diagnosis_rows or []
    frontend_rows = frontend_rows or []
    product_operation_rows = product_operation_rows or []
    ad_processing_rows = ad_processing_rows or []
    scale_rows = scale_rows or []
    scale_keyword_rows = scale_keyword_rows or []
    growth_test_rows = growth_test_rows or []
    action_review_rows = action_review_rows or []
    keyword_review_rows = keyword_review_rows or []
    for marketplace in ["UK", "US", "DE"]:
        market_task_rows = [
            row for row in today_task_rows if str(row.get("marketplace") or "").upper() == marketplace
        ]
        market_tomorrow_rows = [
            row for row in tomorrow_review_rows if str(row.get("marketplace") or "").upper() == marketplace
        ]
        market_listing_rows = [
            row for row in listing_diagnosis_rows if str(row.get("marketplace") or "").upper() == marketplace
        ]
        market_cost_rows = [
            row for row in cost_diagnosis_rows if str(row.get("marketplace") or "").upper() == marketplace
        ]
        market_frontend_rows = [
            row for row in frontend_rows if str(row.get("marketplace") or "").upper() == marketplace
        ]
        market_product_operation_rows = [
            row for row in product_operation_rows if str(row.get("marketplace") or "").upper() == marketplace
        ]
        market_ad_processing_rows = [
            row for row in ad_processing_rows if str(row.get("marketplace") or "").upper() == marketplace
        ]
        market_scale_rows = [
            row for row in scale_rows if str(row.get("站点") or row.get("marketplace") or "").upper() == marketplace
        ]
        market_scale_keyword_rows = [
            row for row in scale_keyword_rows if str(row.get("marketplace") or "").upper() == marketplace
        ]
        market_growth_test_rows = [
            row for row in growth_test_rows if str(row.get("marketplace") or "").upper() == marketplace
        ]
        market_action_rows = [
            row for row in action_review_rows if str(row.get("marketplace") or "").upper() == marketplace
        ]
        market_keyword_rows = [
            row for row in keyword_review_rows if str(row.get("marketplace") or "").upper() == marketplace
        ]
        _append_rows_sheet(workbook, f"{marketplace}_今日动作清单", market_task_rows)
        _append_rows_sheet(workbook, f"{marketplace}_明日复查清单", market_tomorrow_rows)
        _append_rows_sheet(workbook, f"{marketplace}_Listing待确认", market_listing_rows)
        _append_rows_sheet(workbook, f"{marketplace}_成本利润诊断", market_cost_rows)
        _append_rows_sheet(workbook, f"{marketplace}_产品运营卡", market_product_operation_rows)
        _append_rows_sheet(workbook, f"{marketplace}_前台证据队列", market_frontend_rows)
        _append_rows_sheet(workbook, f"{marketplace}_广告处理队列", market_ad_processing_rows)
        _append_rows_sheet(workbook, f"{marketplace}_放量候选", market_scale_rows)
        _append_rows_sheet(workbook, f"{marketplace}_放量词候选", market_scale_keyword_rows)
        _append_rows_sheet(workbook, f"{marketplace}_小预算试投", market_growth_test_rows)
        _append_rows_sheet(workbook, f"{marketplace}_执行后复盘", market_action_rows)
        _append_rows_sheet(workbook, f"{marketplace}_词级执行复盘", market_keyword_rows)
    workbook.save(path)


def _write_autoopt_workbook(
    path,
    report_date: str = "2026-06-24",
    *,
    sheet_rows: dict[str, list[dict[str, object]]] | None = None,
) -> None:
    workbook = Workbook()
    workbook.active.title = "autoopt_log"
    summary = workbook.create_sheet("summary")
    summary.append(["report_date"])
    summary.append([report_date])
    for sheet_name in ["action_review", "keyword_action_review", "final_decisions"]:
        workbook.create_sheet(sheet_name)
    for sheet_name, rows in (sheet_rows or {}).items():
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        _append_rows_sheet(workbook, sheet_name, rows)
    workbook.save(path)
    date_token = path.stem.replace("autoopt_", "")
    autoopt_log_path = path.parent / f"autoopt_log_{date_token}.json"
    action_rows: list[object] = []
    keyword_rows: list[object] = []
    learned_rows: list[object] = []
    manual_learning_rows: list[object] = []
    product_profiles: list[object] = []
    keyword_memory: list[object] = []
    positive_patterns: list[object] = []
    negative_patterns: list[object] = []
    if autoopt_log_path.exists():
        try:
            payload = json.loads(autoopt_log_path.read_text(encoding="utf-8"))
        except Exception:
            payload = {}
        if isinstance(payload, dict):
            action_rows = payload.get("action_review_rows") if isinstance(payload.get("action_review_rows"), list) else []
            keyword_rows = (
                payload.get("keyword_action_review_rows")
                if isinstance(payload.get("keyword_action_review_rows"), list)
                else []
            )
            learned_rows = payload.get("learned_rules") if isinstance(payload.get("learned_rules"), list) else []
            manual_learning_rows = (
                payload.get("manual_learning_rows") if isinstance(payload.get("manual_learning_rows"), list) else []
            )
            product_profiles = (
                payload.get("product_strategy_profiles")
                if isinstance(payload.get("product_strategy_profiles"), list)
                else []
            )
            keyword_memory = (
                payload.get("keyword_strategy_memory")
                if isinstance(payload.get("keyword_strategy_memory"), list)
                else []
            )
            positive_patterns = (
                payload.get("positive_action_patterns")
                if isinstance(payload.get("positive_action_patterns"), list)
                else []
            )
            negative_patterns = (
                payload.get("negative_action_patterns")
                if isinstance(payload.get("negative_action_patterns"), list)
                else []
            )
    (path.parent / f"action_review_{date_token}.json").write_text(json.dumps(action_rows), encoding="utf-8")
    (path.parent / f"keyword_action_review_{date_token}.json").write_text(json.dumps(keyword_rows), encoding="utf-8")
    (path.parent / f"learned_rules_{date_token}.json").write_text(json.dumps(learned_rows), encoding="utf-8")
    (path.parent / f"manual_learning_log_{date_token}.json").write_text(
        json.dumps(manual_learning_rows),
        encoding="utf-8",
    )
    (path.parent / f"product_strategy_profiles_{date_token}.json").write_text(
        json.dumps(product_profiles),
        encoding="utf-8",
    )
    (path.parent / f"keyword_strategy_memory_{date_token}.json").write_text(
        json.dumps(keyword_memory),
        encoding="utf-8",
    )
    (path.parent / f"self_optimization_log_{date_token}.json").write_text(
        json.dumps(
            {
                "report_date": report_date,
                "learned_rules": learned_rows,
                "manual_learning_rows": manual_learning_rows,
                "action_review_rows": action_rows,
                "keyword_action_review_rows": keyword_rows,
                "positive_action_patterns": positive_patterns,
                "negative_action_patterns": negative_patterns,
                "product_strategy_profiles": product_profiles,
                "keyword_strategy_memory": keyword_memory,
            }
        ),
        encoding="utf-8",
    )


def _write_import_manifest_workbook(path, rows: list[dict[str, object]] | None = None) -> None:
    rows = rows or []
    headers = ["original_filename", "status", "created_at", "target_path", "archive_path"]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "import_manifest"
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    workbook.save(path)


def _enhanced_request_row(**overrides) -> dict[str, object]:
    row: dict[str, object] = {
        "marketplace": "UK",
        "trigger_sku": "SKU-ENH",
        "trigger_asin": "B0ENHREQ01",
        "trigger_product_name": "Enhanced request product",
        "issue_type": "增强数据文件导入",
        "report_type": "流量和销售数据",
        "period": "recent_vs_prior",
        "start_date": "2026-06-17",
        "end_date": "2026-06-24",
        "expected_filename": "traffic_sales_uk_recent.xlsx",
        "target_path": "data/raw_amazon_custom/UK/",
        "target_folder": "data/raw_amazon_custom/UK/",
        "required": "是",
        "seller_central_page": "定制分析",
        "instruction": "待导出",
        "status": "待导出",
        "file_type": "traffic_sales",
        "format_type": "single",
        "detected_from": "",
        "detected_date_range": "",
        "freshness": "missing",
        "used_in_diagnosis": "否",
    }
    row.update(overrides)
    if "frontend_decision_evidence_tier" not in overrides:
        row["frontend_decision_evidence_tier"] = (
            row.get("frontend_evidence_display_tier") or row.get("frontend_evidence_tier") or ""
        )
    return row


def _write_enhanced_requests_workbook(path, rows: list[dict[str, object]] | None = None) -> None:
    rows = rows or []
    headers = [
        "marketplace",
        "trigger_sku",
        "trigger_asin",
        "trigger_product_name",
        "issue_type",
        "report_type",
        "period",
        "start_date",
        "end_date",
        "expected_filename",
        "target_path",
        "target_folder",
        "required",
        "seller_central_page",
        "instruction",
        "status",
        "file_type",
        "format_type",
        "detected_from",
        "detected_date_range",
        "freshness",
        "used_in_diagnosis",
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(headers)
    for row in rows:
        sheet.append([_excel_value(row.get(header, "")) for header in headers])
    workbook.save(path)
    workbook.close()


def _write_enhanced_requests_markdown(path, rows: list[dict[str, object]] | None = None, report_date: str = "2026-06-24") -> None:
    rows = rows or []
    lines = [
        f"# 需要补充导出的增强数据｜{report_date}",
        "",
        "请把文件放入对应站点文件夹，而不是直接放在 raw_amazon_custom 根目录。",
        "",
    ]
    if not rows:
        lines.append("当前没有需要额外导出的增强数据请求。")
        path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
        return
    lines.extend(
        [
            "| 站点 | 状态 | 报表类型 | 周期 | 日期范围 | 导出后文件名 | 目标文件夹 | 必需 |",
            "|---|---|---|---|---|---|---|---|",
        ]
    )
    for row in rows:
        required_value = str(row.get("required") or "").strip().lower()
        required_text = "是" if required_value in {"1", "true", "yes", "是"} else "否"
        lines.append(
            f"| {row.get('marketplace') or 'N/A'} | {row.get('status') or 'N/A'} | "
            f"{row.get('report_type') or 'N/A'} | {row.get('period') or 'N/A'} | "
            f"{row.get('start_date') or 'N/A'} ~ {row.get('end_date') or 'N/A'} | "
            f"{row.get('expected_filename') or 'N/A'} | "
            f"{row.get('target_folder') or row.get('target_path') or 'N/A'} | {required_text} |"
        )
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def _review_row(review_type: str = "keyword_action_review", **overrides) -> dict[str, object]:
    required_fields = (
        daily_update.KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS
        if review_type == "keyword_action_review"
        else daily_update.ACTION_REVIEW_REQUIRED_FIELDS
    )
    row = {field: "" for field in required_fields}
    row.update(
        {
            "action_id": "UK||SKU-1||B0REVIEW1||search_term||desk lamp clip||bid_down",
            "normalized_action": "bid_down",
            "action_scope": "search_term",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0REVIEW1",
            "product_name": "Review product",
            "search_term_or_target": "desk lamp clip",
            "action_detail": "降竞价",
            "executed_at": "2026-06-20T09:00:00",
            "report_date": "2026-06-24",
            "review_date": "2026-06-24",
            "days_since_execution": 4,
            "review_window": "3d_check",
            "current_7d_clicks": 12,
            "current_7d_spend": 6.5,
            "current_7d_ad_orders": 1,
            "current_7d_promoted_ad_orders": 1,
            "current_7d_promoted_ad_sales": 19.99,
            "current_7d_halo_ad_orders": 0,
            "current_7d_halo_ad_sales": 0,
            "current_7d_total_orders": 2,
            "current_7d_acos": 0.08,
            "current_7d_target_acos": 0.1,
            "current_7d_tacos": 0.06,
            "current_7d_available_stock": 20,
            "current_14d_clicks": 20,
            "current_14d_spend": 10,
            "current_14d_ad_orders": 2,
            "current_14d_promoted_ad_orders": 2,
            "current_14d_promoted_ad_sales": 39.98,
            "current_14d_halo_ad_orders": 0,
            "current_14d_halo_ad_sales": 0,
            "current_14d_total_orders": 4,
            "current_14d_acos": 0.08,
            "current_14d_tacos": 0.06,
            "current_14d_available_stock": 20,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            "attribution_effect_status": "promoted_sku_converted",
            "attribution_effect_note": "promoted SKU 有订单",
            "outcome": "待7天确认",
            "effect_evidence": "3天复查口径，7天结论待补；3 天复盘",
            "review_status": "待7天复盘",
            "review_phase": "3d",
            "review_outcome": "not_ready",
            "effectiveness_score": "",
            "cooldown_status": "cooldown",
            "cooldown_until": "2026-06-27",
            "block_reason": "",
            "rule_adjustment": "",
            "learning_scope": "keyword",
        }
    )
    row.update(overrides)
    if review_type == "action_review":
        row.pop("search_term_or_target", None)
        row.setdefault("action_type", "广告动作")
    return row


def _product_decision_row(**overrides) -> dict[str, object]:
    row = {
        "marketplace": "UK",
        "sku": "SKU-DECISION",
        "asin": "B0DECISION",
        "product_name": "Decision product",
        "final_decision": "WAIT_REVIEW",
        "final_decision_label": "等待复查",
        "decision_priority": "P1",
        "decision_reason": "复盘窗口未完成",
        "today_allowed_actions": ["observe"],
        "today_blocked_actions": ["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
        "frontend_evidence_state": "weak",
        "frontend_evidence_display_tier": "仅背景参考",
        "frontend_decision_evidence_tier": "仅背景参考",
        "frontend_check_status": "待前台检查",
        "frontend_cache_used": False,
        "frontend_failure_category": "",
        "frontend_price_currency_warning": "",
        "frontend_location_warning": "",
        "frontend_search_status": "",
        "frontend_search_partial_evidence": False,
        "frontend_auto_conclusion": "FRONTEND_WEAK",
        "inventory_constraint": "",
        "feedback_cooldown_status": "",
        "ad_action_summary": "",
        "next_review_date": "2026-06-27",
        "confidence": "medium",
        "fusion_action_gate": "observe",
        "fusion_today_action": "观察",
        "fusion_do_not_do": "不放量",
        "fusion_review_window": "3d_check",
    }
    row.update(overrides)
    return row


def _product_decision_contract_attrs(row: dict[str, object], **overrides: object) -> str:
    def list_attr(value: object) -> str:
        if isinstance(value, list):
            return "|".join(str(item).strip() for item in value if str(item).strip())
        return str(value or "").strip()

    attrs = {
        "data-product-decision-final": row.get("final_decision") or "",
        "data-product-decision-label": row.get("final_decision_label") or "",
        "data-product-allowed-actions": list_attr(row.get("today_allowed_actions")),
        "data-product-blocked-actions": list_attr(row.get("today_blocked_actions")),
        "data-product-frontend-tier": row.get("frontend_evidence_tier") or "",
        "data-product-frontend-display-tier": row.get("frontend_evidence_display_tier")
        or row.get("frontend_evidence_tier")
        or "",
        "data-product-frontend-decision-tier": row.get("frontend_decision_evidence_tier")
        or row.get("frontend_evidence_display_tier")
        or row.get("frontend_evidence_tier")
        or "",
    }
    attrs.update(overrides)
    return "".join(
        f' {name}="{html.escape(str(value), quote=True)}"'
        for name, value in attrs.items()
        if str(value).strip()
    )


def _boolish_test_flag(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "是", "已验证"}


def _attach_product_decision(payload: dict[str, object], row: dict[str, object]) -> None:
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["product_final_decision_rows"] = [row]
    snapshot["product_operation_cards"] = [_product_operation_card(row)]
    payload["product_final_decision_rows"] = [row]
    payload["product_operation_cards"] = list(snapshot["product_operation_cards"])
    snapshot["final_decision_summary"] = {str(row["final_decision"]): 1}
    snapshot["decision_gate_counts"] = {str(row["final_decision"]): 1}
    payload["final_decision_summary"]["UK"] = {str(row["final_decision"]): 1}
    payload["decision_gate_counts"]["UK"] = {str(row["final_decision"]): 1}


def _product_operation_card(row: dict[str, object], **overrides: object) -> dict[str, object]:
    card = {
        "marketplace": row.get("marketplace") or "UK",
        "sku": row.get("sku") or "",
        "asin": row.get("asin") or "",
        "product_name": row.get("product_name") or "",
        "final_decision": row.get("final_decision") or "",
        "final_decision_label": row.get("final_decision_label") or "",
        "decision_reason": row.get("decision_reason") or "",
        "operation_main_reason": row.get("decision_reason") or "",
        "fusion_issue_type": row.get("fusion_issue_type") or "",
        "fusion_action_gate": row.get("fusion_action_gate") or "",
        "fusion_do_not_do": row.get("fusion_do_not_do") or "",
        "fusion_review_window": row.get("fusion_review_window") or row.get("next_review_date") or "",
        "today_allowed_actions": row.get("today_allowed_actions") or [],
        "today_blocked_actions": row.get("today_blocked_actions") or [],
        "frontend_status": row.get("frontend_check_status") or "",
        "frontend_evidence_tier": row.get("frontend_evidence_tier") or "",
        "frontend_evidence_display_tier": row.get("frontend_evidence_display_tier") or "",
        "frontend_decision_evidence_tier": row.get("frontend_decision_evidence_tier")
        or row.get("frontend_evidence_display_tier")
        or row.get("frontend_evidence_tier")
        or "",
        "frontend_cache_used": _boolish_test_flag(row.get("frontend_cache_used")),
        "frontend_evidence_audit_summary": row.get("frontend_evidence_audit_summary") or "",
        "inventory_constraint": row.get("inventory_constraint") or "",
        "ad_action_count": 0,
        "ad_action_display_limit": 4,
        "ad_action_more_count": 0,
        "ad_action_items": [],
        "ad_diagnostic_summary": "",
    }
    card.update(overrides)
    return card


def test_product_operation_card_helper_treats_string_false_cache_flag_as_false() -> None:
    card = _product_operation_card(_product_decision_row(frontend_cache_used="False"))

    assert card["frontend_cache_used"] is False


def _analysis_payload(report_date: str) -> dict[str, object]:
    def snapshot() -> dict[str, object]:
        return {
            "today_task_queue_rows": [],
            "tomorrow_review_rows": [],
            "search_term_processing_queue_rows": [],
            "html_search_term_processing_queue_rows": [],
            "scale_rows": [],
            "scale_keyword_rows": [],
            "growth_test_rows": [],
            "today_action_groups": {},
            "frontend_check_queue_rows": [],
            "inventory_replenishment_rows": [],
            "product_final_decision_rows": [],
            "product_operation_cards": [],
            "final_decision_summary": {},
            "decision_gate_counts": {},
            "frontend_coverage_summary": daily_update._frontend_coverage_expected([]),
            "action_effect_review_rows": [],
            "keyword_action_effect_review_rows": [],
            "cost_profit_diagnosis_rows": [],
            "listing_price_diagnosis_rows": [],
        }

    payload = {
        "report_date": report_date,
        "final_decision_summary": {marketplace: {} for marketplace in ["UK", "US", "DE"]},
        "decision_gate_counts": {marketplace: {} for marketplace in ["UK", "US", "DE"]},
        "product_final_decision_rows": [],
        "product_operation_cards": [],
        "inventory_replenishment_rows": [],
        "marketplace_results": [
            {
                "marketplace": marketplace,
                "summary": {"report_date": report_date},
                "report_view_snapshot": snapshot(),
            }
            for marketplace in ["UK", "US", "DE"]
        ],
    }
    payload["frontend_coverage_summary"] = daily_update._frontend_coverage_aggregate_from_snapshots(
        payload["marketplace_results"]
    )
    return payload


def _manifest_core_row(
    filename: str = "ads_2026-06-24.csv",
    *,
    detected_type: str = "ads_report_all",
    detected_date_range: str = "2026-06-24~2026-06-24",
    status: str = "imported",
) -> dict[str, object]:
    return {
        "original_filename": filename,
        "status": status,
        "detected_type": detected_type,
        "detected_date_range": detected_date_range,
        "created_at": daily_update.datetime.now().replace(microsecond=0).isoformat(sep=" "),
        "target_path": __file__,
        "archive_path": __file__,
    }


def test_frontend_wrapper_refresh_failures_allows_missing_results_when_queue_is_empty(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")

    failures = daily_update.frontend_wrapper_refresh_failures(
        tmp_path / "missing_frontend_check_results.json",
        analysis,
    )

    assert failures == []


def test_frontend_wrapper_refresh_failures_blocks_stale_results_when_queue_exists(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [
        {"marketplace": "UK", "sku": "SKU-FRONTEND", "asin": "B0FRONTEND1"}
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    results = tmp_path / "frontend_check_results.json"
    old_payload = {
        "generated_at": "2026-06-23T10:00:00",
        "refresh_summary": {"frontend_refresh_total": 1},
        "items": [{"marketplace": "UK", "sku": "SKU-FRONTEND", "asin": "B0FRONTEND1"}],
    }
    results.write_text(json.dumps(old_payload), encoding="utf-8")

    failures = daily_update.frontend_wrapper_refresh_failures(
        results,
        analysis,
        previous_mtime_ns=results.stat().st_mtime_ns,
        previous_generated_at="2026-06-23T10:00:00",
    )

    assert failures == [f"frontend check results were not refreshed by frontend wrapper: {results}"]


def test_frontend_wrapper_refresh_failures_blocks_gate_field_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [
        {
            "marketplace": "UK",
            "sku": "SKU-FRONTEND",
            "asin": "B0FRONTEND1",
            "frontend_check_status": "沿用 2026-06-23 前台数据",
            "frontend_cache_used": True,
            "frontend_evidence_tier": "仅背景参考",
            "frontend_search_status": "读取失败",
            "frontend_evidence_quality_score": "58",
        }
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    results = tmp_path / "frontend_check_results.json"
    results.write_text(
        json.dumps(
            {
                "generated_at": "2026-06-24T10:00:00",
                "items": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-FRONTEND",
                        "asin": "B0FRONTEND1",
                        "frontend_check_status": "沿用 2026-06-23 前台数据",
                        "frontend_cache_used": False,
                        "frontend_evidence_tier": "仅背景参考",
                        "frontend_search_status": "读取失败",
                        "frontend_evidence_quality_score": "58",
                    }
                ],
            }
        ),
        encoding="utf-8",
    )

    failures = daily_update.frontend_wrapper_refresh_failures(results, analysis)

    assert failures == [
        "frontend check results gate field mismatch after wrapper: "
        "UK/SKU-FRONTEND/B0FRONTEND1 field frontend_cache_used expected true, got false"
    ]


def test_frontend_wrapper_refresh_failures_allows_missing_result_gate_fields(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [
        {
            "marketplace": "UK",
            "sku": "SKU-FRONTEND",
            "asin": "B0FRONTEND1",
            "frontend_check_status": "已自动检查",
            "frontend_price_currency_warning": "币种异常",
            "frontend_location_exact": False,
            "frontend_failure_category": "currency_warning",
            "competitor_comparability": "low",
        }
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    results = tmp_path / "frontend_check_results.json"
    results.write_text(
        json.dumps(
            {
                "generated_at": "2026-06-24T10:00:00",
                "items": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-FRONTEND",
                        "asin": "B0FRONTEND1",
                        "frontend_check_status": "已自动检查",
                    }
                ],
            }
        ),
        encoding="utf-8",
    )

    failures = daily_update.frontend_wrapper_refresh_failures(results, analysis)

    assert failures == []


def test_frontend_wrapper_refresh_failures_blocks_competitor_count_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [
        {
            "marketplace": "UK",
            "sku": "SKU-FRONTEND",
            "asin": "B0FRONTEND1",
            "frontend_check_status": "已自动检查",
            "frontend_competitor_count": 1,
            "comparable_competitor_count": 1,
        }
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    results = tmp_path / "frontend_check_results.json"
    results.write_text(
        json.dumps(
            {
                "generated_at": "2026-06-24T10:00:00",
                "items": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-FRONTEND",
                        "asin": "B0FRONTEND1",
                        "frontend_check_status": "已自动检查",
                        "frontend_competitor_count": 3,
                        "comparable_competitor_count": 1,
                    }
                ],
            }
        ),
        encoding="utf-8",
    )

    failures = daily_update.frontend_wrapper_refresh_failures(results, analysis)

    assert failures == [
        "frontend check results gate field mismatch after wrapper: "
        "UK/SKU-FRONTEND/B0FRONTEND1 field frontend_competitor_count expected 1, got 3"
    ]


def test_frontend_wrapper_refresh_failures_blocks_strong_flag_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [
        {
            "marketplace": "UK",
            "sku": "SKU-FRONTEND",
            "asin": "B0FRONTEND1",
            "frontend_check_status": "已自动检查",
            "frontend_evidence_tier": "强诊断可用",
            "frontend_evidence_display_tier": "强诊断可用",
            "frontend_decision_evidence_tier": "强诊断可用",
            "frontend_evidence_is_strong": True,
        }
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    results = tmp_path / "frontend_check_results.json"
    results.write_text(
        json.dumps(
            {
                "generated_at": "2026-06-24T10:00:00",
                "items": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-FRONTEND",
                        "asin": "B0FRONTEND1",
                        "frontend_check_status": "已自动检查",
                        "frontend_evidence_tier": "强诊断可用",
                        "frontend_evidence_display_tier": "强诊断可用",
                        "frontend_decision_evidence_tier": "强诊断可用",
                        "frontend_evidence_is_strong": False,
                    }
                ],
            }
        ),
        encoding="utf-8",
    )

    failures = daily_update.frontend_wrapper_refresh_failures(results, analysis)

    assert failures == [
        "frontend check results gate field mismatch after wrapper: "
        "UK/SKU-FRONTEND/B0FRONTEND1 field frontend_evidence_is_strong expected true, got false"
    ]


def test_frontend_wrapper_refresh_failures_blocks_audit_reasons_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [
        {
            "marketplace": "UK",
            "sku": "SKU-FRONTEND",
            "asin": "B0FRONTEND1",
            "frontend_check_status": "已自动检查",
            "frontend_evidence_audit_summary": "前台证据可辅助判断，不能单独放量",
            "frontend_evidence_audit_reasons": ["地区未确认", "竞品可比性未达强诊断"],
            "frontend_evidence_audit_detail": "地区未确认；竞品可比性未达强诊断",
        }
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    results = tmp_path / "frontend_check_results.json"
    results.write_text(
        json.dumps(
            {
                "generated_at": "2026-06-24T10:00:00",
                "items": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-FRONTEND",
                        "asin": "B0FRONTEND1",
                        "frontend_check_status": "已自动检查",
                        "frontend_evidence_audit_summary": "前台证据可辅助判断，不能单独放量",
                        "frontend_evidence_audit_reasons": ["产品页和搜索页均可用"],
                        "frontend_evidence_audit_detail": "地区未确认；竞品可比性未达强诊断",
                    }
                ],
            }
        ),
        encoding="utf-8",
    )

    failures = daily_update.frontend_wrapper_refresh_failures(results, analysis)

    assert failures == [
        "frontend check results gate field mismatch after wrapper: "
        "UK/SKU-FRONTEND/B0FRONTEND1 field frontend_evidence_audit_reasons "
        "expected 地区未确认；竞品可比性未达强诊断, got 产品页和搜索页均可用"
    ]


def test_frontend_queue_content_fields_include_raw_competitor_count() -> None:
    assert "frontend_competitor_count" in daily_update.FRONTEND_QUEUE_CONTENT_FIELDS
    assert "frontend_competitor_count" in daily_update.FRONTEND_GATE_CONSISTENCY_FIELDS
    assert "frontend_evidence_is_strong" in daily_update.FRONTEND_GATE_CONSISTENCY_FIELDS
    assert "frontend_evidence_audit_reasons" in daily_update.FRONTEND_QUEUE_CONTENT_FIELDS
    assert "frontend_evidence_audit_summary" in daily_update.FRONTEND_GATE_CONSISTENCY_FIELDS
    assert "frontend_evidence_audit_reasons" in daily_update.FRONTEND_GATE_CONSISTENCY_FIELDS
    assert "frontend_evidence_audit_detail" in daily_update.FRONTEND_GATE_CONSISTENCY_FIELDS


def _text_report_fixture(name: str, report_date: str) -> str:
    markers = daily_update.TEXT_REPORT_REQUIRED_MARKERS.get(name, [])
    anchors = daily_update.TEXT_REPORT_REQUIRED_ANCHORS.get(name, [])
    anchor_html = [f'<section id="{anchor}"></section>' for anchor in anchors]
    return f"{report_date}\n" + "\n".join([*markers, *anchor_html])


def test_import_manifest_failures_allows_clean_manifest(tmp_path) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(
        json.dumps(
            [
                {"original_filename": "ads.csv", "status": "imported"},
                {"original_filename": "erp.xlsx", "status": "split_imported"},
                {"original_filename": "old_ads.csv", "status": "skipped_old_duplicate"},
            ]
        ),
        encoding="utf-8",
    )

    assert daily_update.import_manifest_failures(manifest) == []


def test_import_manifest_failures_blocks_unknown_and_error(tmp_path) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(
        json.dumps(
            [
                {"original_filename": "unknown.xlsx", "status": "unknown", "reason": "cannot classify"},
                {"original_filename": "locked.xlsx", "status": "error", "reason": "permission"},
            ]
        ),
        encoding="utf-8",
    )

    failures = daily_update.import_manifest_failures(manifest)

    assert "unknown.xlsx: unknown | cannot classify" in failures
    assert "locked.xlsx: error | permission" in failures


def test_import_manifest_failures_blocks_missing_and_unexpected_status(tmp_path) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(
        json.dumps(
            [
                {"original_filename": "missing_status.csv"},
                {"original_filename": "partial.xlsx", "status": "partial"},
            ]
        ),
        encoding="utf-8",
    )

    failures = daily_update.import_manifest_failures(manifest)

    assert "missing_status.csv: unexpected import status <missing>" in failures
    assert "partial.xlsx: unexpected import status partial" in failures


def test_import_manifest_failures_blocks_missing_manifest(tmp_path) -> None:
    failures = daily_update.import_manifest_failures(tmp_path / "missing.json")

    assert len(failures) == 1
    assert "import manifest missing" in failures[0]


def test_import_manifest_failures_blocks_stale_manifest(tmp_path) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    previous_mtime = manifest.stat().st_mtime_ns

    failures = daily_update.import_manifest_failures(manifest, previous_mtime_ns=previous_mtime)

    assert len(failures) == 1
    assert "import manifest was not refreshed by current import step" in failures[0]


def test_import_manifest_audit_failures_blocks_missing_xlsx(tmp_path) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")

    failures = daily_update.import_manifest_audit_failures(manifest, tmp_path / "import_manifest.xlsx")

    assert len(failures) == 1
    assert "import manifest xlsx missing" in failures[0]


def test_import_manifest_audit_failures_blocks_stale_xlsx(tmp_path) -> None:
    rows = [{"original_filename": "ads.csv", "status": "imported"}]
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps(rows), encoding="utf-8")
    xlsx = tmp_path / "import_manifest.xlsx"
    _write_import_manifest_workbook(xlsx, rows)
    previous_mtime = xlsx.stat().st_mtime_ns

    failures = daily_update.import_manifest_audit_failures(
        manifest,
        xlsx,
        previous_xlsx_mtime_ns=previous_mtime,
    )

    assert len(failures) == 1
    assert "import manifest xlsx was not refreshed by current import step" in failures[0]


def test_import_manifest_audit_failures_blocks_xlsx_json_row_count_mismatch(tmp_path) -> None:
    rows = [{"original_filename": "ads.csv", "status": "imported"}]
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps(rows), encoding="utf-8")
    xlsx = tmp_path / "import_manifest.xlsx"
    _write_import_manifest_workbook(xlsx, [])

    failures = daily_update.import_manifest_audit_failures(manifest, xlsx)

    assert len(failures) == 1
    assert "import manifest xlsx row count mismatch: expected 1, got 0" in failures


def test_import_manifest_audit_failures_blocks_xlsx_json_content_mismatch(tmp_path) -> None:
    rows = [
        {
            "original_filename": "new_ads.csv",
            "status": "imported",
            "created_at": "2026-06-24T09:30:00",
            "target_path": "data/raw/ads/new_ads.csv",
            "archive_path": "data/archive/ads/new_ads.csv",
        }
    ]
    stale_rows = [
        {
            "original_filename": "old_ads.csv",
            "status": "imported",
            "created_at": "2026-06-23T09:30:00",
            "target_path": "data/raw/ads/old_ads.csv",
            "archive_path": "data/archive/ads/old_ads.csv",
        }
    ]
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps(rows), encoding="utf-8")
    xlsx = tmp_path / "import_manifest.xlsx"
    _write_import_manifest_workbook(xlsx, stale_rows)

    failures = daily_update.import_manifest_audit_failures(manifest, xlsx)

    assert len(failures) == 1
    assert "import manifest xlsx identity mismatch vs import manifest json for import manifest rows" in failures[0]
    assert "new_ads.csv" in failures[0]
    assert "old_ads.csv" in failures[0]


def test_import_manifest_audit_failures_blocks_missing_success_row_files(tmp_path) -> None:
    rows = [
        {
            "original_filename": "new_ads.csv",
            "status": "imported",
            "created_at": "2026-06-24 09:30:00",
            "target_path": str(tmp_path / "raw" / "new_ads.csv"),
            "archive_path": str(tmp_path / "archive" / "new_ads.csv"),
        }
    ]
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps(rows), encoding="utf-8")
    xlsx = tmp_path / "import_manifest.xlsx"
    _write_import_manifest_workbook(xlsx, rows)

    failures = daily_update.import_manifest_audit_failures(
        manifest,
        xlsx,
        import_started_at=daily_update.datetime.fromisoformat("2026-06-24 09:00:00"),
    )

    assert failures == [
        f"import manifest successful row 1 new_ads.csv target_path does not exist: {tmp_path / 'raw' / 'new_ads.csv'}",
        f"import manifest successful row 1 new_ads.csv archive_path does not exist: {tmp_path / 'archive' / 'new_ads.csv'}",
    ]


def test_output_refresh_failures_blocks_missing_and_stale_outputs(tmp_path) -> None:
    output = tmp_path / "latest_analysis.json"
    output.write_text("{}", encoding="utf-8")
    previous_mtime = output.stat().st_mtime_ns
    missing = tmp_path / "missing.html"

    failures = daily_update.output_refresh_failures(
        [output, missing],
        previous_mtimes_ns={output: previous_mtime, missing: None},
    )

    assert f"required output was not refreshed by report step: {output}" in failures
    assert f"required output missing after report refresh: {missing}" in failures


def test_restore_output_content_snapshot_removes_new_date_scoped_outputs(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    html_report = tmp_path / "latest_recommendations.html"
    analysis.write_text("old analysis", encoding="utf-8")
    html_report.write_text("old html", encoding="utf-8")
    snapshot = daily_update.output_content_snapshot([analysis, html_report])

    new_daily_report = tmp_path / "amazon_ops_report_2026-06-24.xlsx"
    new_autoopt_log = tmp_path / "autoopt_log_20260624.json"
    new_daily_report.write_bytes(b"partial workbook")
    new_autoopt_log.write_text("partial log", encoding="utf-8")
    analysis.write_text("partial analysis", encoding="utf-8")
    html_report.write_text("partial html", encoding="utf-8")

    failures = daily_update.restore_output_content_snapshot(snapshot)

    assert failures == []
    assert analysis.read_text(encoding="utf-8") == "old analysis"
    assert html_report.read_text(encoding="utf-8") == "old html"
    assert not new_daily_report.exists()
    assert not new_autoopt_log.exists()


def test_restore_report_state_snapshot_restores_database_and_archive_files(monkeypatch, tmp_path) -> None:
    output = tmp_path / "latest_analysis.json"
    output.write_text("old analysis", encoding="utf-8")
    db_path = tmp_path / "amazon_ops.db"
    db_path.write_bytes(b"old db")
    ads_archive = tmp_path / "archive" / "ads"
    erp_archive = tmp_path / "archive" / "erp"
    existing_archive = ads_archive / "existing.csv"
    existing_archive.parent.mkdir(parents=True, exist_ok=True)
    existing_archive.write_text("old archive", encoding="utf-8")
    erp_archive.mkdir(parents=True, exist_ok=True)

    monkeypatch.setattr(daily_update, "REQUIRED_REFRESHED_OUTPUTS", [output])
    monkeypatch.setattr(daily_update, "DB_STATE_FILES", [db_path])
    monkeypatch.setattr(daily_update, "ARCHIVE_STATE_DIRS", [ads_archive, erp_archive])

    snapshot = daily_update.report_state_snapshot([output])
    output.write_text("partial analysis", encoding="utf-8")
    db_path.write_bytes(b"partial db")
    existing_archive.write_text("partial archive overwrite", encoding="utf-8")
    new_ads_archive = ads_archive / "new_ads.csv"
    new_erp_archive = erp_archive / "new_erp.csv"
    new_ads_archive.write_text("new ads archive", encoding="utf-8")
    new_erp_archive.write_text("new erp archive", encoding="utf-8")

    failures = daily_update.restore_report_state_snapshot(snapshot)

    assert failures == []
    assert output.read_text(encoding="utf-8") == "old analysis"
    assert db_path.read_bytes() == b"old db"
    assert existing_archive.read_text(encoding="utf-8") == "old archive"
    assert not new_ads_archive.exists()
    assert not new_erp_archive.exists()


def test_required_refreshed_outputs_include_enhanced_requests_and_assets() -> None:
    required = {path.relative_to(daily_update.OUTPUT_DIR).as_posix() for path in daily_update.REQUIRED_REFRESHED_OUTPUTS}

    assert "enhanced_data_requests.md" in required
    assert "enhanced_data_requests.xlsx" in required
    assert "assets/report.css" in required
    assert "assets/report.js" in required


def test_report_refresh_failures_blocks_missing_enhanced_requests_and_assets(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    assets_dir = tmp_path / "assets"
    paths = [
        analysis,
        tmp_path / "enhanced_data_requests.md",
        tmp_path / "enhanced_data_requests.xlsx",
        assets_dir / "report.css",
        assets_dir / "report.js",
    ]

    failures = daily_update.report_refresh_failures(paths, previous_mtimes_ns={})

    assert f"required output missing after report refresh: {tmp_path / 'enhanced_data_requests.md'}" in failures
    assert f"required output missing after report refresh: {tmp_path / 'enhanced_data_requests.xlsx'}" in failures
    assert f"required output missing after report refresh: {assets_dir / 'report.css'}" in failures
    assert f"required output missing after report refresh: {assets_dir / 'report.js'}" in failures


def test_report_refresh_failures_blocks_malformed_enhanced_requests_markdown(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    enhanced_md = tmp_path / "enhanced_data_requests.md"
    enhanced_md.write_text("2026-06-24 stale enhanced request text without the required table", encoding="utf-8")

    failures = daily_update.report_refresh_failures([analysis, enhanced_md], previous_mtimes_ns={})

    assert f"text report missing required marker # 需要补充导出的增强数据: {enhanced_md}" in failures
    assert "enhanced data requests Markdown missing empty-state message" in failures


def test_report_refresh_failures_blocks_stale_report_assets(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    assets_dir = tmp_path / "assets"
    assets_dir.mkdir()
    css = assets_dir / "report.css"
    js = assets_dir / "report.js"
    css.write_text("body{}", encoding="utf-8")
    js.write_text("console.log('old')", encoding="utf-8")

    failures = daily_update.report_refresh_failures(
        [analysis, css, js],
        previous_mtimes_ns={css: css.stat().st_mtime_ns, js: js.stat().st_mtime_ns},
    )

    assert f"required output was not refreshed by report step: {css}" in failures
    assert f"required output was not refreshed by report step: {js}" in failures


def test_report_refresh_failures_blocks_malformed_report_assets(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    assets_dir = tmp_path / "assets"
    assets_dir.mkdir()
    css = assets_dir / "report.css"
    js = assets_dir / "report.js"
    css.write_text("body{}", encoding="utf-8")
    js.write_text("console.log('ok')", encoding="utf-8")

    failures = daily_update.report_refresh_failures([analysis, css, js], previous_mtimes_ns={})

    assert "report asset report.css missing required token .ad-task-card" in failures
    assert "report asset report.css missing required token .ad-copy-box" in failures
    assert "report asset report.js missing required token data-ad-complete-checkbox" in failures
    assert "report asset report.js missing required token data-ad-filter-summary" in failures


def test_report_refresh_failures_blocks_enhanced_request_excel_row_count_mismatch(tmp_path) -> None:
    row = _enhanced_request_row()
    payload = _analysis_payload("2026-06-24")
    payload["enhanced_data_requests"] = [row]
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    xlsx = tmp_path / "enhanced_data_requests.xlsx"
    _write_enhanced_requests_workbook(xlsx, [])

    failures = daily_update.report_refresh_failures([analysis, xlsx], previous_mtimes_ns={})

    assert (
        "enhanced data requests Excel row count mismatch vs latest analysis for enhanced data requests: "
        "expected 1, got 0"
    ) in failures


def test_report_refresh_failures_blocks_enhanced_request_excel_identity_mismatch(tmp_path) -> None:
    expected_row = _enhanced_request_row()
    stale_row = dict(expected_row)
    stale_row["expected_filename"] = "old_traffic_sales.xlsx"
    payload = _analysis_payload("2026-06-24")
    payload["enhanced_data_requests"] = [expected_row]
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    xlsx = tmp_path / "enhanced_data_requests.xlsx"
    _write_enhanced_requests_workbook(xlsx, [stale_row])

    failures = daily_update.report_refresh_failures([analysis, xlsx], previous_mtimes_ns={})

    assert any(
        "enhanced data requests Excel identity mismatch vs latest analysis for enhanced data requests" in failure
        and "old_traffic_sales.xlsx" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_enhanced_request_excel_field_mismatch(tmp_path) -> None:
    expected_row = _enhanced_request_row()
    stale_row = dict(expected_row)
    stale_row["status"] = "已导入"
    stale_row["freshness"] = "stale"
    payload = _analysis_payload("2026-06-24")
    payload["enhanced_data_requests"] = [expected_row]
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    xlsx = tmp_path / "enhanced_data_requests.xlsx"
    _write_enhanced_requests_workbook(xlsx, [stale_row])

    failures = daily_update.report_refresh_failures([analysis, xlsx], previous_mtimes_ns={})

    assert any(
        "enhanced data requests Excel field mismatch vs latest analysis for enhanced data requests" in failure
        and "field status: expected '待导出', got '已导入'" in failure
        for failure in failures
    )
    assert any(
        "enhanced data requests Excel field mismatch vs latest analysis for enhanced data requests" in failure
        and "field freshness: expected 'missing', got 'stale'" in failure
        for failure in failures
    )


def test_report_refresh_failures_allows_empty_enhanced_requests_markdown(tmp_path) -> None:
    payload = _analysis_payload("2026-06-24")
    payload["enhanced_data_requests"] = []
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    markdown = tmp_path / "enhanced_data_requests.md"
    _write_enhanced_requests_markdown(markdown, [])

    failures = daily_update.report_refresh_failures([analysis, markdown], previous_mtimes_ns={})

    assert not any("enhanced data requests Markdown" in failure for failure in failures)


def test_report_refresh_failures_blocks_enhanced_request_markdown_row_count_mismatch(tmp_path) -> None:
    row = _enhanced_request_row()
    payload = _analysis_payload("2026-06-24")
    payload["enhanced_data_requests"] = [row]
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    markdown = tmp_path / "enhanced_data_requests.md"
    _write_enhanced_requests_markdown(markdown, [])

    failures = daily_update.report_refresh_failures([analysis, markdown], previous_mtimes_ns={})

    assert (
        "enhanced data requests Markdown row count mismatch vs latest analysis display "
        "for enhanced data requests Markdown display: expected 1, got 0"
    ) in failures


def test_report_refresh_failures_blocks_enhanced_request_markdown_identity_mismatch(tmp_path) -> None:
    expected_row = _enhanced_request_row()
    stale_row = dict(expected_row)
    stale_row["expected_filename"] = "old_traffic_sales.xlsx"
    payload = _analysis_payload("2026-06-24")
    payload["enhanced_data_requests"] = [expected_row]
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    markdown = tmp_path / "enhanced_data_requests.md"
    _write_enhanced_requests_markdown(markdown, [stale_row])

    failures = daily_update.report_refresh_failures([analysis, markdown], previous_mtimes_ns={})

    assert any(
        "enhanced data requests Markdown identity mismatch vs latest analysis display "
        "for enhanced data requests Markdown display" in failure
        and "old_traffic_sales.xlsx" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_enhanced_request_markdown_field_mismatch(tmp_path) -> None:
    expected_row = _enhanced_request_row()
    stale_row = dict(expected_row)
    stale_row["status"] = "已导入"
    stale_row["required"] = "否"
    payload = _analysis_payload("2026-06-24")
    payload["enhanced_data_requests"] = [expected_row]
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    markdown = tmp_path / "enhanced_data_requests.md"
    _write_enhanced_requests_markdown(markdown, [stale_row])

    failures = daily_update.report_refresh_failures([analysis, markdown], previous_mtimes_ns={})

    assert any(
        "enhanced data requests Markdown field mismatch vs latest analysis display "
        "for enhanced data requests Markdown display" in failure
        and "field 状态: expected '待导出', got '已导入'" in failure
        for failure in failures
    )
    assert any(
        "enhanced data requests Markdown field mismatch vs latest analysis display "
        "for enhanced data requests Markdown display" in failure
        and "field 必需: expected '是', got '否'" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_invalid_latest_analysis(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text("not json", encoding="utf-8")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert len(failures) >= 1
    assert "latest analysis cannot be read" in failures[0]


def test_report_refresh_failures_blocks_future_report_date(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps({"report_date": "2999-01-01"}), encoding="utf-8")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis report_date 2999-01-01 is after current date" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_missing_date_scoped_outputs(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert f"required output missing after report refresh: {tmp_path / 'amazon_ops_report_2026-06-24.xlsx'}" in failures
    assert f"required output missing after report refresh: {tmp_path / 'autoopt_log_20260624.json'}" in failures
    assert f"required output missing after report refresh: {tmp_path / 'autoopt_20260624.xlsx'}" in failures
    assert f"required output missing after report refresh: {tmp_path / 'action_review_20260624.json'}" in failures
    assert f"required output missing after report refresh: {tmp_path / 'keyword_action_review_20260624.json'}" in failures
    assert f"required output missing after report refresh: {tmp_path / 'learned_rules_20260624.json'}" in failures
    assert f"required output missing after report refresh: {tmp_path / 'manual_learning_log_20260624.json'}" in failures
    assert f"required output missing after report refresh: {tmp_path / 'product_strategy_profiles_20260624.json'}" in failures
    assert f"required output missing after report refresh: {tmp_path / 'keyword_strategy_memory_20260624.json'}" in failures
    assert f"required output missing after report refresh: {tmp_path / 'self_optimization_log_20260624.json'}" in failures


def test_report_refresh_failures_blocks_stale_date_scoped_outputs(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    daily_report = tmp_path / "amazon_ops_report_2026-06-24.xlsx"
    _write_daily_report_workbook(daily_report)
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "learned_rules": [],
                "manual_learning_rows": [],
                "keyword_action_review_rows": [],
                "positive_action_patterns": [],
                "negative_action_patterns": [],
                "product_strategy_profiles": [],
                "keyword_strategy_memory": [],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures(
        [analysis],
        previous_mtimes_ns={daily_report: daily_report.stat().st_mtime_ns},
    )

    assert f"required output was not refreshed by report step: {daily_report}" in failures


def test_report_refresh_failures_blocks_stale_learning_sidecar_json(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "learned_rules": [{"marketplace": "UK", "sku": "SKU-NEW"}],
                "manual_learning_rows": [],
                "product_strategy_profiles": [],
                "keyword_strategy_memory": [],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
    (tmp_path / "learned_rules_20260624.json").write_text(
        json.dumps([{"marketplace": "UK", "sku": "SKU-OLD"}]),
        encoding="utf-8",
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "learned_rules_20260624.json content mismatch: expected autoopt log learned_rules" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_stale_self_optimization_learning_payload(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "learned_rules": [{"marketplace": "UK", "sku": "SKU-NEW"}],
                "manual_learning_rows": [],
                "keyword_action_review_rows": [],
                "positive_action_patterns": [],
                "negative_action_patterns": [],
                "product_strategy_profiles": [],
                "keyword_strategy_memory": [],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
    (tmp_path / "self_optimization_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "learned_rules": [{"marketplace": "UK", "sku": "SKU-OLD"}],
                "manual_learning_rows": [],
                "keyword_action_review_rows": [],
                "positive_action_patterns": [],
                "negative_action_patterns": [],
                "product_strategy_profiles": [],
                "keyword_strategy_memory": [],
            }
        ),
        encoding="utf-8",
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert "self_optimization_log_20260624.json learned_rules mismatch vs autoopt log" in failures


def test_report_refresh_failures_blocks_stale_self_optimization_action_review_rows(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "action_review_rows": [{"marketplace": "UK", "sku": "SKU-NEW", "action_id": "a-new"}],
                "learned_rules": [],
                "manual_learning_rows": [],
                "keyword_action_review_rows": [],
                "positive_action_patterns": [],
                "negative_action_patterns": [],
                "product_strategy_profiles": [],
                "keyword_strategy_memory": [],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
    (tmp_path / "self_optimization_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "action_review_rows": [{"marketplace": "UK", "sku": "SKU-OLD", "action_id": "a-old"}],
                "learned_rules": [],
                "manual_learning_rows": [],
                "keyword_action_review_rows": [],
                "positive_action_patterns": [],
                "negative_action_patterns": [],
                "product_strategy_profiles": [],
                "keyword_strategy_memory": [],
            }
        ),
        encoding="utf-8",
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert "self_optimization_log_20260624.json action_review_rows mismatch vs autoopt log" in failures


def test_report_refresh_failures_blocks_execution_rate_as_effectiveness_learning(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "rule_adjustments": [
                    {
                        "rule_scope": "广告动作",
                        "observed_status": "已执行 3/3",
                        "suggested_adjustment": "保留并前置该类建议；可继续维持当前阈值。",
                    }
                ],
                "action_adjustments": [
                    {
                        "rule_scope": "广告动作 / 提高竞价",
                        "observed_status": "已执行 2/2",
                        "suggested_adjustment": "这类动作执行率较高，可继续保留为默认推荐。",
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "autoopt log rule_adjustments row 1 广告动作 treats execution rate as effectiveness via 前置"
        in failures
    )
    assert any("action_adjustments row 1 广告动作 / 提高竞价 treats execution rate as effectiveness" in failure for failure in failures)


def test_report_refresh_failures_allows_execution_rate_learning_when_promoted_sku_gated(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "rule_adjustments": [
                    {
                        "rule_scope": "广告动作",
                        "observed_status": "已执行 3/3",
                        "suggested_adjustment": "执行接受度较高，先保留当前展示；等3天/7天 promoted SKU 复盘证明有效后再提高优先级或维持阈值。",
                    }
                ],
                "action_adjustments": [
                    {
                        "rule_scope": "广告动作 / 提高竞价",
                        "observed_status": "已执行 2/2",
                        "suggested_adjustment": "这类动作执行率较高，只能说明容易落地；等3天/7天 promoted SKU 复盘有效后再作为默认推荐。",
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert not [failure for failure in failures if "treats execution rate as effectiveness" in failure]


def test_report_refresh_failures_blocks_autoopt_log_report_date_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-23"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert "autoopt log report_date mismatch: expected 2026-06-24, got 2026-06-23" in failures


def test_report_refresh_failures_blocks_corrupt_date_scoped_excel(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    (tmp_path / "amazon_ops_report_2026-06-24.xlsx").write_text("not an xlsx", encoding="utf-8")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any("daily Excel report cannot be opened" in failure for failure in failures)


def test_report_refresh_failures_blocks_incomplete_latest_analysis_marketplaces(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "marketplace_results": [
                    {"marketplace": "UK", "summary": {"report_date": "2026-06-24"}},
                    {"marketplace": "US", "summary": {"report_date": "2026-06-24"}},
                ],
            }
        ),
        encoding="utf-8",
    )
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert "latest analysis marketplace_results must contain UK, US, DE; got ['UK', 'US']" in failures


def test_report_refresh_failures_blocks_duplicate_latest_analysis_marketplaces(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    payload["marketplace_results"].append(payload["marketplace_results"][0])
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis marketplace_results must contain exactly one UK, one US, and one DE; "
        "got ['UK', 'US', 'DE', 'UK']"
    ) in failures


def test_report_refresh_failures_blocks_empty_marketplace_result(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    payload["marketplace_results"][1]["has_data"] = False
    payload["marketplace_results"][1]["summary"].update(
        {
            "ads_row_count": 0,
            "erp_row_count": 12,
        }
    )
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert "latest analysis US marketplace result has_data is false" in failures
    assert "latest analysis US summary ads_row_count must be positive, got 0" in failures


def test_report_refresh_failures_blocks_missing_report_view_snapshot_keys(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot.pop("keyword_action_effect_review_rows")
    snapshot.pop("inventory_replenishment_rows")
    snapshot.pop("product_operation_cards")
    snapshot.pop("search_term_processing_queue_rows")
    snapshot.pop("scale_keyword_rows")
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK report_view_snapshot missing inventory_replenishment_rows, "
        "keyword_action_effect_review_rows, product_operation_cards, scale_keyword_rows, "
        "search_term_processing_queue_rows"
        in failures
    )


def test_report_refresh_failures_blocks_report_view_snapshot_type_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["today_task_queue_rows"] = {}
    snapshot["today_action_groups"] = []
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert "latest analysis UK report_view_snapshot today_task_queue_rows must be a list" in failures
    assert "latest analysis UK report_view_snapshot today_action_groups must be an object" in failures


def test_report_refresh_failures_blocks_product_operation_card_decision_field_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        asin="B0OPCARDMISMATCH",
        product_name="Operation card mismatch product",
        today_blocked_actions=["bid_up", "budget_up", "broad_scale"],
    )
    _attach_product_decision(payload, product_row)
    payload["marketplace_results"][0]["report_view_snapshot"]["product_operation_cards"] = [
        _product_operation_card(product_row, today_blocked_actions=["bid_up"])
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[product_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "product_final_decisions": [product_row],
                "final_decision_summary": {"WAIT_REVIEW": 1},
                "decision_gate_counts": {"WAIT_REVIEW": 1},
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"final_decisions": [product_row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis UK product_operation_cards row 1 Operation card mismatch product "
        "field today_blocked_actions mismatch vs product_final_decision_rows" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_product_operation_card_ad_item_without_action_id(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        asin="B0OPCARDTRACE",
        product_name="Operation card trace product",
    )
    ad_item = {
        "marketplace": "UK",
        "sku": product_row["sku"],
        "asin": product_row["asin"],
        "product_name": product_row["product_name"],
        "search_term_or_target": "operation card trace term",
        "suggested_action": "降竞价",
        "normalized_action": "bid_down",
        "action_scope": "search_term",
    }
    _attach_product_decision(payload, product_row)
    payload["marketplace_results"][0]["report_view_snapshot"]["product_operation_cards"] = [
        _product_operation_card(product_row, ad_action_count=1, ad_action_items=[ad_item])
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[product_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis UK product_operation_cards row 1 Operation card trace product "
        "ad_action_items row 1 missing action_id" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_product_operation_card_ad_item_from_wrong_product(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        asin="B0OPCARDTRACE",
        product_name="Operation card trace product",
    )
    ad_item = {
        "marketplace": "UK",
        "sku": product_row["sku"],
        "asin": "B0OPCARDOTHER",
        "product_name": product_row["product_name"],
        "search_term_or_target": "operation card trace term",
        "suggested_action": "降竞价",
        "normalized_action": "bid_down",
        "action_scope": "search_term",
    }
    ad_item["action_id"] = daily_update.make_action_id(ad_item, "bid_down", "search_term")
    _attach_product_decision(payload, product_row)
    payload["marketplace_results"][0]["report_view_snapshot"]["product_operation_cards"] = [
        _product_operation_card(product_row, ad_action_count=1, ad_action_items=[ad_item])
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[product_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis UK product_operation_cards row 1 Operation card trace product "
        "ad_action_items row 1 asin mismatch: expected B0OPCARDTRACE, got B0OPCARDOTHER" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_product_operation_card_missing_pending_ad_item(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        asin="B0OPCARDTRACE",
        product_name="Operation card trace product",
    )
    ad_row = {
        field: ""
        for field in [
            *daily_update.AD_PROCESSING_QUEUE_IDENTITY_FIELDS,
            *daily_update.AD_PROCESSING_QUEUE_CONTENT_FIELDS,
        ]
    }
    ad_row.update(
        {
            "marketplace": "UK",
            "sku": product_row["sku"],
            "asin": product_row["asin"],
            "product_name": product_row["product_name"],
            "search_term_or_target": "operation card pending term",
            "campaign_name": "Campaign A",
            "ad_group_name": "Ad group A",
            "match_type_or_targeting": "exact",
            "suggested_action": "降竞价",
            "normalized_action": "bid_down",
            "action_scope": "search_term",
            "confirmed_status": "待确认",
        }
    )
    ad_row["action_id"] = daily_update.make_action_id(ad_row, "bid_down", "search_term")
    _attach_product_decision(payload, product_row)
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["html_search_term_processing_queue_rows"] = [ad_row]
    snapshot["product_operation_cards"] = [
        _product_operation_card(product_row, ad_action_count=0, ad_action_items=[])
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[product_row],
        ad_processing_rows=[ad_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis UK product_operation_cards row 1 Operation card trace product "
        f"missing ad_action_items action_id from ad workbench rows: {ad_row['action_id']}" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_product_operation_card_missing_fifth_pending_ad_item(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        asin="B0OPCARDTRACE",
        product_name="Operation card trace product",
    )
    ad_rows = []
    for index in range(5):
        ad_row = {
            field: ""
            for field in [
                *daily_update.AD_PROCESSING_QUEUE_IDENTITY_FIELDS,
                *daily_update.AD_PROCESSING_QUEUE_CONTENT_FIELDS,
            ]
        }
        ad_row.update(
            {
                "marketplace": "UK",
                "sku": product_row["sku"],
                "asin": product_row["asin"],
                "product_name": product_row["product_name"],
                "search_term_or_target": f"operation card pending term {index + 1}",
                "campaign_name": "Campaign A",
                "ad_group_name": "Ad group A",
                "match_type_or_targeting": "exact",
                "suggested_action": "降竞价",
                "normalized_action": "bid_down",
                "action_scope": "search_term",
                "confirmed_status": "待确认",
            }
        )
        ad_row["action_id"] = daily_update.make_action_id(ad_row, "bid_down", "search_term")
        ad_rows.append(ad_row)
    _attach_product_decision(payload, product_row)
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["html_search_term_processing_queue_rows"] = ad_rows
    snapshot["product_operation_cards"] = [
        _product_operation_card(
            product_row,
            ad_action_count=4,
            ad_action_display_limit=4,
            ad_action_more_count=0,
            ad_action_items=ad_rows[:4],
        )
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[product_row],
        ad_processing_rows=ad_rows,
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis UK product_operation_cards row 1 Operation card trace product "
        f"missing ad_action_items action_id from ad workbench rows: {ad_rows[4]['action_id']}" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_daily_excel_missing_product_operation_sheet(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        asin="B0OPCARDEXCEL",
        product_name="Operation card Excel product",
    )
    _attach_product_decision(payload, product_row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    workbook_path = tmp_path / "amazon_ops_report_2026-06-24.xlsx"
    _write_daily_report_workbook(
        workbook_path,
        product_rows=[product_row],
    )
    workbook = load_workbook(workbook_path)
    del workbook["产品运营卡"]
    workbook.save(workbook_path)
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert "daily Excel report missing 产品运营卡 sheet for product operation cards consistency" in failures


def test_report_refresh_failures_blocks_missing_required_empty_daily_excel_sheet(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    workbook_path = tmp_path / "amazon_ops_report_2026-06-24.xlsx"
    _write_daily_report_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    del workbook["今日动作清单"]
    workbook.save(workbook_path)
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert "daily Excel report missing required sheet even when empty: 今日动作清单" in failures


def test_report_refresh_failures_blocks_missing_required_market_frontend_empty_sheet(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    workbook_path = tmp_path / "amazon_ops_report_2026-06-24.xlsx"
    _write_daily_report_workbook(workbook_path)
    workbook = load_workbook(workbook_path)
    del workbook["UK_前台证据队列"]
    workbook.save(workbook_path)
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert "daily Excel report missing required sheet even when empty: UK_前台证据队列" in failures


def test_report_refresh_failures_blocks_decision_summary_count_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["product_final_decision_rows"] = [{"final_decision": "EXECUTE_TODAY"}]
    payload["product_final_decision_rows"] = snapshot["product_final_decision_rows"]
    snapshot["final_decision_summary"] = {"WAIT_REVIEW": 1}
    snapshot["decision_gate_counts"] = {"EXECUTE_TODAY": 1}
    payload["final_decision_summary"]["UK"] = {"WAIT_REVIEW": 1}
    payload["decision_gate_counts"]["UK"] = {"EXECUTE_TODAY": 1}
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK report_view_snapshot final_decision_summary does not match "
        "product_final_decision_rows final_decision counts"
    ) in failures


def test_report_refresh_failures_blocks_top_level_decision_summary_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    payload["final_decision_summary"]["UK"] = {"BROKEN": 1}
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert "latest analysis top-level final_decision_summary mismatch for UK report_view_snapshot" in failures


def test_report_refresh_failures_blocks_top_level_product_decision_union_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    snapshot_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0SNAPSHOT",
        "final_decision": "WAIT_REVIEW",
        "today_allowed_actions": ["observe"],
    }
    top_level_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0TOPLEVEL",
        "final_decision": "WAIT_REVIEW",
        "today_allowed_actions": ["observe"],
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["product_final_decision_rows"] = [snapshot_row]
    snapshot["final_decision_summary"] = {"WAIT_REVIEW": 1}
    snapshot["decision_gate_counts"] = {"WAIT_REVIEW": 1}
    payload["product_final_decision_rows"] = [top_level_row]
    payload["final_decision_summary"]["UK"] = {"WAIT_REVIEW": 1}
    payload["decision_gate_counts"]["UK"] = {"WAIT_REVIEW": 1}
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis top-level product_final_decision_rows identity mismatch for product final decisions"
        in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_top_level_product_decision_content_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    snapshot_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0TOPCONTENT",
        "final_decision": "EXECUTE_TODAY",
        "final_decision_label": "今天执行",
        "today_allowed_actions": ["bid_down"],
    }
    top_level_row = dict(snapshot_row)
    top_level_row["final_decision"] = "WAIT_REVIEW"
    top_level_row["final_decision_label"] = "等复盘"
    top_level_row["today_allowed_actions"] = ["observe"]
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["product_final_decision_rows"] = [snapshot_row]
    snapshot["final_decision_summary"] = {"EXECUTE_TODAY": 1}
    snapshot["decision_gate_counts"] = {"EXECUTE_TODAY": 1}
    payload["product_final_decision_rows"] = [top_level_row]
    payload["final_decision_summary"]["UK"] = {"EXECUTE_TODAY": 1}
    payload["decision_gate_counts"]["UK"] = {"EXECUTE_TODAY": 1}
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis top-level product_final_decision_rows field mismatch for product final decisions "
        "('UK', 'SKU-1', 'B0TOPCONTENT') field final_decision: "
        "expected 'EXECUTE_TODAY' from marketplace snapshots, got 'WAIT_REVIEW'"
        in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_product_decision_blank_marketplace(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = _product_decision_row(
        marketplace="",
        asin="B0BLANKDECISION",
        product_name="Blank marketplace decision",
    )
    _attach_product_decision(payload, row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK product_final_decision_rows row 1 Blank marketplace decision missing marketplace value"
        in failures
    )


def test_report_refresh_failures_blocks_top_level_inventory_union_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    snapshot_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0STOCKSNAP",
        "stock_risk_level": "LOW_STOCK",
    }
    top_level_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0STOCKTOP",
        "stock_risk_level": "LOW_STOCK",
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["inventory_replenishment_rows"] = [snapshot_row]
    payload["inventory_replenishment_rows"] = [top_level_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        inventory_rows=payload["inventory_replenishment_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis top-level inventory_replenishment_rows identity mismatch for inventory replenishment rows"
        in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_top_level_inventory_content_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    snapshot_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0STOCKCONTENT",
        "product_name": "Top stock product",
        "stock_risk_level": "URGENT_REORDER",
        "stock_status_label": "紧急补货",
        "available_stock": 2,
        "days_of_cover": 3,
    }
    top_level_row = dict(snapshot_row)
    top_level_row["stock_risk_level"] = "HEALTHY"
    top_level_row["stock_status_label"] = "健康"
    top_level_row["available_stock"] = 200
    top_level_row["days_of_cover"] = 365
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["inventory_replenishment_rows"] = [snapshot_row]
    payload["inventory_replenishment_rows"] = [top_level_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        inventory_rows=payload["inventory_replenishment_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis top-level inventory_replenishment_rows field mismatch for inventory replenishment rows "
        "('UK', 'SKU-1', 'B0STOCKCONTENT') field available_stock: "
        "expected 2.0 from marketplace snapshots, got 200.0"
        in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_inventory_blank_marketplace(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = {
        "marketplace": "",
        "sku": "SKU-BLANK-INV",
        "asin": "B0BLANKINV",
        "product_name": "Blank marketplace inventory",
        "stock_risk_level": "LOW_STOCK",
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["inventory_replenishment_rows"] = [row]
    payload["inventory_replenishment_rows"] = [row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        inventory_rows=payload["inventory_replenishment_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK inventory_replenishment_rows row 1 Blank marketplace inventory missing marketplace value"
        in failures
    )


def test_report_refresh_failures_blocks_daily_excel_product_decision_identity_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0ANALYSIS",
        "final_decision": "WAIT_REVIEW",
        "today_allowed_actions": ["observe"],
    }
    excel_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0STALEEXCEL",
        "final_decision": "WAIT_REVIEW",
        "today_allowed_actions": ["observe"],
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["product_final_decision_rows"] = [row]
    payload["product_final_decision_rows"] = [row]
    snapshot["final_decision_summary"] = {"WAIT_REVIEW": 1}
    snapshot["decision_gate_counts"] = {"WAIT_REVIEW": 1}
    payload["final_decision_summary"]["UK"] = {"WAIT_REVIEW": 1}
    payload["decision_gate_counts"]["UK"] = {"WAIT_REVIEW": 1}
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 产品最终决策 identity mismatch for product final decisions" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_daily_excel_stale_rows_when_analysis_is_empty(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    for index, result in enumerate(payload["marketplace_results"], start=1):
        result["summary"].update(
            {
                "marketplace": result["marketplace"],
                "ads_row_count": 9 + index,
                "erp_row_count": 19 + index,
                "sku_count": 2 + index,
                "asin_count": 3 + index,
            }
        )
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-STALE",
                "asin": "B0STALEEMPTY",
                "final_decision": "EXECUTE_TODAY",
                "today_allowed_actions": ["bid_up"],
            }
        ],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "daily Excel 产品最终决策 row count mismatch for product final decisions: expected 0, got 1"
        in failures
    )


def test_report_refresh_failures_blocks_daily_excel_today_task_identity_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = {
        "marketplace": "UK",
        "sku": "SKU-ACTION",
        "asin": "B0ACTIONJSON",
        "priority": "P0",
        "issue_type": "广告消耗无转化",
        "action_group": "广告动作",
        "today_action": "降竞价10%-20%",
        "search_term_or_target": "dimmer desk lamp",
        "suggested_action": "降竞价10%-20%",
        "normalized_action": "bid_down",
        "action_id": "UK||SKU-ACTION||B0ACTIONJSON||dimmer desk lamp||bid_down",
    }
    excel_row = dict(row)
    excel_row["asin"] = "B0ACTIONOLD"
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["today_task_queue_rows"] = [row]
    snapshot["today_action_groups"] = {"广告动作": [row]}
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        today_task_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 今日动作清单 identity mismatch for today task queue rows" in failure
        and "B0ACTIONJSON" in failure
        and "B0ACTIONOLD" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_daily_excel_today_task_content_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = {
        "marketplace": "UK",
        "sku": "SKU-ACTION",
        "asin": "B0ACTIONJSON",
        "product_name": "Action product",
        "confirmed_status": "待确认",
        "priority": "P0",
        "issue_type": "广告消耗无转化",
        "primary_reason": "近14天广告无单且点击>=20",
        "key_evidence": "点击 31；广告订单 0",
        "action_group": "广告动作",
        "today_action": "降竞价10%-20%",
        "search_term_or_target": "dimmer desk lamp",
        "suggested_action": "降竞价10%-20%",
        "copy_action_line": "dimmer desk lamp 降竞价10%-20%",
        "normalized_action": "bid_down",
        "action_scope": "search_term",
        "action_id": "UK||SKU-ACTION||B0ACTIONJSON||dimmer desk lamp||bid_down",
        "final_decision": "CONSERVATIVE_RUN",
        "today_allowed_actions": "bid_down / negative_exact / observe",
        "today_blocked_actions": "bid_up / budget_up / broad_scale",
    }
    excel_row = dict(row)
    excel_row["primary_reason"] = "旧原因"
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["today_task_queue_rows"] = [row]
    snapshot["today_action_groups"] = {"广告动作": [row]}
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        today_task_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 今日动作清单 field mismatch for today task queue rows "
        "('UK', 'SKU-ACTION', 'B0ACTIONJSON', 'P0', '广告消耗无转化', '广告动作', "
        "'降竞价10%-20%', 'dimmer desk lamp', '降竞价10%-20%', 'bid_down', "
        "'UK||SKU-ACTION||B0ACTIONJSON||dimmer desk lamp||bid_down') field primary_reason: "
        "expected '近14天广告无单且点击>=20', got '旧原因'"
        in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_daily_excel_today_task_copy_block_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = {
        "marketplace": "UK",
        "sku": "SKU-COPY",
        "asin": "B0COPYJSON",
        "product_name": "Copy block product",
        "confirmed_status": "待确认",
        "priority": "P1",
        "issue_type": "广告动作",
        "primary_reason": "后台复制文本必须可追溯",
        "key_evidence": "点击 31；广告订单 0",
        "action_group": "广告动作",
        "today_action": "降竞价10%-20%",
        "search_term_or_target": "copy block term",
        "suggested_action": "降竞价10%-20%",
        "copy_action_line": "copy block term 降竞价10%-20%",
        "copy_block": "copy block term 降竞价10%-20%\n后台复制文本",
        "normalized_action": "bid_down",
        "action_scope": "search_term",
        "action_id": "UK||SKU-COPY||B0COPYJSON||copy block term||bid_down",
    }
    excel_row = dict(row)
    excel_row["copy_block"] = "旧后台复制文本"
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["today_task_queue_rows"] = [row]
    snapshot["today_action_groups"] = {"广告动作": [row]}
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        today_task_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 今日动作清单 field mismatch for today task queue rows" in failure
        and "field copy_block" in failure
        and "旧后台复制文本" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_daily_excel_tomorrow_review_identity_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = {
        "marketplace": "UK",
        "sku": "SKU-REVIEW",
        "asin": "B0REVIEWJSON",
        "product_name": "Review product",
        "review_reason": "3-5点击无单，未达到强动作阈值",
        "current_evidence": "点击 4；花费 £0.88；订单 0",
        "tomorrow_check": "复查是否新增点击、花费或订单；未改善再降竞价",
        "trigger_action": "达到阈值后升级今日动作",
    }
    excel_row = dict(row)
    excel_row["asin"] = "B0REVIEWOLD"
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["tomorrow_review_rows"] = [row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        tomorrow_review_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 明日复查清单 identity mismatch for tomorrow review rows" in failure
        and "B0REVIEWJSON" in failure
        and "B0REVIEWOLD" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_daily_excel_tomorrow_review_content_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = {
        "marketplace": "UK",
        "sku": "SKU-REVIEW",
        "asin": "B0REVIEWJSON",
        "product_name": "Review product",
        "review_reason": "3-5点击无单，未达到强动作阈值",
        "current_evidence": "点击 4；花费 £0.88；订单 0",
        "tomorrow_check": "复查是否新增点击、花费或订单；未改善再降竞价",
        "trigger_action": "达到阈值后升级今日动作",
    }
    excel_row = dict(row)
    excel_row["current_evidence"] = "点击 0；花费 £0.00；订单 0"
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["tomorrow_review_rows"] = [row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        tomorrow_review_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 明日复查清单 field mismatch for tomorrow review rows" in failure
        and "field current_evidence: expected '点击 4；花费 £0.88；订单 0', got '点击 0；花费 £0.00；订单 0'"
        in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_daily_excel_listing_diagnosis_identity_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = {
        "marketplace": "UK",
        "产品": "Listing product",
        "SKU": "SKU-LISTING",
        "ASIN": "B0LISTJSON",
        "诊断类型": "Listing 待人工确认",
        "主因": "加购后不购买",
        "关键证据": "加购 7；购买 0",
        "建议动作": "先确认价格、Coupon、配送和竞品。",
    }
    excel_row = dict(row)
    excel_row["ASIN"] = "B0LISTOLD"
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["listing_price_diagnosis_rows"] = [row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        listing_diagnosis_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel Listing待确认 identity mismatch for listing price diagnosis rows" in failure
        and "B0LISTJSON" in failure
        and "B0LISTOLD" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_daily_excel_cost_diagnosis_content_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = {
        "marketplace": "UK",
        "产品": "Cost product",
        "SKU": "SKU-COST",
        "ASIN": "B0COSTJSON",
        "诊断类型": "成本 / 利润压力诊断",
        "主因": "利润不允许加广告",
        "关键证据": "近14天广告点击 5；广告订单 0；广告花费 £2.94",
        "建议动作": "核对采购成本、头程、FBA、售价；利润未修正前不放量。",
    }
    excel_row = dict(row)
    excel_row["关键证据"] = "近14天广告点击 0；广告订单 0；广告花费 £0.00"
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["cost_profit_diagnosis_rows"] = [row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        cost_diagnosis_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 成本利润诊断 field mismatch for cost profit diagnosis rows" in failure
        and "field 关键证据: expected '近14天广告点击 5；广告订单 0；广告花费 £2.94', got '近14天广告点击 0；广告订单 0；广告花费 £0.00'"
        in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_daily_excel_product_decision_content_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0ANALYSIS",
        "final_decision": "EXECUTE_TODAY",
        "final_decision_label": "今天执行",
        "today_allowed_actions": ["bid_down"],
        "today_blocked_actions": ["budget_up"],
        "frontend_required": True,
        "frontend_posture": "frontend_blocked",
        "frontend_blocking_reasons": ["地区待确认，当前前台证据不能用于放量"],
        "frontend_auto_conclusion_label": "明确前台劣势",
        "frontend_location_uncertain": True,
        "frontend_location_block_reason": "地区待确认，当前前台证据不能用于放量",
    }
    excel_row = dict(row)
    excel_row["final_decision"] = "WAIT_REVIEW"
    excel_row["today_allowed_actions"] = ["observe"]
    excel_row["frontend_blocking_reasons"] = []
    excel_row["frontend_location_uncertain"] = False
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["product_final_decision_rows"] = [row]
    payload["product_final_decision_rows"] = [row]
    snapshot["final_decision_summary"] = {"EXECUTE_TODAY": 1}
    snapshot["decision_gate_counts"] = {"EXECUTE_TODAY": 1}
    payload["final_decision_summary"]["UK"] = {"EXECUTE_TODAY": 1}
    payload["decision_gate_counts"]["UK"] = {"EXECUTE_TODAY": 1}
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 产品最终决策 field mismatch for product final decisions "
        "('UK', 'SKU-1', 'B0ANALYSIS') field final_decision: expected 'EXECUTE_TODAY', got 'WAIT_REVIEW'"
        in failure
        for failure in failures
    )
    assert any(
        "daily Excel 产品最终决策 field mismatch for product final decisions "
        "('UK', 'SKU-1', 'B0ANALYSIS') field today_allowed_actions"
        in failure
        for failure in failures
    )
    assert any(
        "daily Excel 产品最终决策 field mismatch for product final decisions "
        "('UK', 'SKU-1', 'B0ANALYSIS') field frontend_blocking_reasons"
        in failure
        for failure in failures
    )
    assert any(
        "daily Excel 产品最终决策 field mismatch for product final decisions "
        "('UK', 'SKU-1', 'B0ANALYSIS') field frontend_location_uncertain"
        in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_daily_excel_frontend_evidence_field_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = _product_decision_row(
        asin="B0FRONTEXCEL",
        frontend_evidence_state="ok_high",
        frontend_evidence_tier="仅背景参考",
        frontend_evidence_display_tier="强诊断可用",
        frontend_evidence_is_strong=True,
        frontend_check_status="已自动检查",
        frontend_search_status="已自动检查",
        frontend_auto_conclusion="FRONTEND_OK",
        frontend_cache_used=False,
        frontend_evidence_quality_score=88,
        competitor_comparability="medium",
        comparable_competitor_count=3,
        frontend_location_scope="exact",
        frontend_location_verified=True,
        frontend_location_exact=True,
    )
    excel_row = dict(row)
    excel_row["competitor_comparability"] = "low"
    excel_row["comparable_competitor_count"] = 1
    excel_row["frontend_location_scope"] = "marketplace"
    excel_row["frontend_evidence_is_strong"] = False
    _attach_product_decision(payload, row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 产品最终决策 field mismatch for product final decisions "
        "('UK', 'SKU-DECISION', 'B0FRONTEXCEL') field competitor_comparability: expected 'medium', got 'low'"
        in failure
        for failure in failures
    )
    assert any(
        "daily Excel 产品最终决策 field mismatch for product final decisions "
        "('UK', 'SKU-DECISION', 'B0FRONTEXCEL') field frontend_location_scope: expected 'exact', got 'marketplace'"
        in failure
        for failure in failures
    )
    assert any(
        "daily Excel 产品最终决策 field mismatch for product final decisions "
        "('UK', 'SKU-DECISION', 'B0FRONTEXCEL') field frontend_evidence_is_strong: expected True, got False"
        in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_daily_excel_product_decision_fusion_field_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = _product_decision_row(
        asin="B0FUSIONXLS",
        keyword_memory_summary="历史同词复盘：未验证",
        evidence_used=["近14天点击 42", "前台证据仅背景参考"],
        coupon={"coupon_type": "percent", "coupon_value": 10, "coupon_confidence": "high"},
        fusion_issue_type="前台和广告共同问题",
        fusion_confidence="high",
        fusion_reason="广告证据：ad_no_order_clicks；前台证据：明确前台劣势",
        fusion_evidence_flags=["ad_no_order_clicks", "frontend_competitiveness"],
        fusion_missing_evidence=["current_frontend_screenshot"],
        last_updated="2026-06-24",
    )
    excel_row = dict(row)
    excel_row["fusion_reason"] = "广告证据：ad_no_order_clicks"
    excel_row["evidence_used"] = ["近14天点击 42"]
    _attach_product_decision(payload, row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 产品最终决策 field mismatch for product final decisions "
        "('UK', 'SKU-DECISION', 'B0FUSIONXLS') field evidence_used"
        in failure
        for failure in failures
    )
    assert any(
        "daily Excel 产品最终决策 field mismatch for product final decisions "
        "('UK', 'SKU-DECISION', 'B0FUSIONXLS') field fusion_reason"
        in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_daily_excel_inventory_identity_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0STOCKANALYSIS",
        "stock_risk_level": "LOW_STOCK",
    }
    excel_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0STOCKSTALE",
        "stock_risk_level": "LOW_STOCK",
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["inventory_replenishment_rows"] = [row]
    payload["inventory_replenishment_rows"] = [row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        inventory_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 库存补货提醒 identity mismatch for inventory replenishment rows" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_daily_excel_inventory_content_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0STOCKANALYSIS",
        "product_name": "Inventory product",
        "stock_risk_level": "URGENT_REORDER",
        "stock_status_label": "紧急补货",
        "available_stock": 3,
        "days_of_cover": 2,
    }
    excel_row = dict(row)
    excel_row["stock_risk_level"] = "HEALTHY"
    excel_row["days_of_cover"] = 180
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["inventory_replenishment_rows"] = [row]
    payload["inventory_replenishment_rows"] = [row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        inventory_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 库存补货提醒 field mismatch for inventory replenishment rows "
        "('UK', 'SKU-1', 'B0STOCKANALYSIS') field stock_risk_level: expected 'URGENT_REORDER', got 'HEALTHY'"
        in failure
        for failure in failures
    )
    assert any(
        "daily Excel 库存补货提醒 field mismatch for inventory replenishment rows "
        "('UK', 'SKU-1', 'B0STOCKANALYSIS') field days_of_cover: expected 2.0, got 180.0"
        in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_daily_excel_keyword_review_identity_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    expected_row = _review_row(action_id="k1", search_term_or_target="desk lamp clip")
    stale_row = dict(expected_row)
    stale_row["search_term_or_target"] = "old stale term"
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [expected_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[stale_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 词级执行复盘 identity mismatch for keyword action reviews" in failure
        and "desk lamp clip" in failure
        and "old stale term" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_daily_excel_keyword_review_promoted_sku_field_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    expected_row = _review_row(
        action_id="k1",
        search_term_or_target="desk lamp clip",
        current_7d_promoted_ad_orders=2,
        current_7d_total_orders=3,
        current_7d_available_stock=18,
    )
    stale_row = dict(expected_row)
    stale_row["current_7d_promoted_ad_orders"] = 0
    stale_row["current_7d_total_orders"] = 0
    stale_row["current_7d_available_stock"] = 0
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [expected_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[stale_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 词级执行复盘 field mismatch for keyword action reviews" in failure
        and "field current_7d_promoted_ad_orders: expected 2.0, got 0.0" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_daily_excel_action_review_target_acos_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    expected_row = _review_row(
        "action_review",
        action_id="a-target-acos",
        normalized_action="bid_down",
        action_scope="product",
        current_7d_target_acos=0.2,
    )
    stale_row = dict(expected_row)
    stale_row["current_7d_target_acos"] = 0.1
    payload["marketplace_results"][0]["report_view_snapshot"]["action_effect_review_rows"] = [expected_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        action_review_rows=[stale_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 执行后效果复盘 field mismatch for action reviews" in failure
        and "field current_7d_target_acos: expected 0.2, got 0.1" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_daily_excel_keyword_review_missing_empty_field_header(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    expected_row = _review_row(
        action_id="k-header",
        search_term_or_target="missing header term",
        block_reason="",
    )
    excel_row = dict(expected_row)
    excel_row.pop("block_reason", None)
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [expected_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 词级执行复盘 missing fields for keyword action reviews: block_reason" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_growth_decision_with_cached_frontend_evidence(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0CACHE",
        "product_name": "Cached frontend product",
        "final_decision": "SMALL_SCALE_ALLOWED",
        "today_allowed_actions": ["observe", "bid_up"],
        "today_blocked_actions": ["budget_up", "broad_scale"],
        "frontend_evidence_state": "ok_high",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_display_tier": "强诊断可用",
        "frontend_check_status": "沿用 2026-06-20 前台数据",
        "frontend_search_status": "已自动检查",
        "frontend_search_partial_evidence": False,
        "frontend_cache_used": True,
        "frontend_failure_category": "none",
        "frontend_price_currency_warning": "",
        "frontend_location_warning": "",
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_location_scope": "exact",
        "frontend_location_exact": True,
        "frontend_location_verified": True,
        "competitor_comparability": "medium",
        "comparable_competitor_count": 3,
        "frontend_evidence_quality_score": 88,
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["product_final_decision_rows"] = [row]
    payload["product_final_decision_rows"] = [row]
    snapshot["final_decision_summary"] = {"SMALL_SCALE_ALLOWED": 1}
    snapshot["decision_gate_counts"] = {"SMALL_SCALE_ALLOWED": 1}
    payload["final_decision_summary"]["UK"] = {"SMALL_SCALE_ALLOWED": 1}
    payload["decision_gate_counts"]["UK"] = {"SMALL_SCALE_ALLOWED": 1}
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK product_final_decision_rows row 1 B0CACHE allows growth without current frontend check"
        in failures
    )
    assert (
        "latest analysis UK product_final_decision_rows row 1 B0CACHE allows growth with cached frontend evidence"
        in failures
    )


def test_report_refresh_failures_blocks_product_decision_allowed_blocked_action_overlap(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = _product_decision_row(
        marketplace="UK",
        sku="SKU-1",
        asin="B0OVERLAP",
        product_name="Contradictory gate product",
        final_decision="SMALL_SCALE_ALLOWED",
        today_allowed_actions=["observe", "bid_up"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale"],
        frontend_evidence_state="ok_high",
        frontend_evidence_tier="仅背景参考",
        frontend_evidence_display_tier="强诊断可用",
        frontend_evidence_is_strong=True,
        frontend_check_status="已自动检查",
        frontend_search_status="已自动检查",
        frontend_search_partial_evidence=False,
        frontend_cache_used=False,
        frontend_failure_category="none",
        frontend_price_currency_warning="",
        frontend_location_warning="",
        frontend_auto_conclusion="FRONTEND_OK",
        frontend_location_scope="exact",
        frontend_location_exact=True,
        frontend_location_verified=True,
        competitor_comparability="medium",
        comparable_competitor_count=3,
        frontend_evidence_quality_score=88,
    )
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["product_final_decision_rows"] = [row]
    payload["product_final_decision_rows"] = [row]
    snapshot["final_decision_summary"] = {"SMALL_SCALE_ALLOWED": 1}
    snapshot["decision_gate_counts"] = {"SMALL_SCALE_ALLOWED": 1}
    payload["final_decision_summary"]["UK"] = {"SMALL_SCALE_ALLOWED": 1}
    payload["decision_gate_counts"]["UK"] = {"SMALL_SCALE_ALLOWED": 1}
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK product_final_decision_rows row 1 B0OVERLAP "
        "has actions both allowed and blocked: bid_up"
    ) in failures


def test_report_refresh_failures_blocks_listing_growth_claim_with_weak_frontend(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    listing_row = {
        "marketplace": "UK",
        "产品": "Weak frontend listing product",
        "SKU": "SKU-LISTING-WEAK",
        "ASIN": "B0LISTWEAK",
        "诊断类型": "Listing 待人工确认",
        "主因": "点击后不转化",
        "关键证据": "近14天产品级广告点击 18；广告订单 0；总单 0",
        "建议动作": "立即改 Listing 并加预算",
        "confirmed_status": "待确认",
        "suggested_action": "先查前台",
        "normalized_action": "observe",
        "action_scope": "product",
        "action_id": "UK||SKU-LISTING-WEAK||B0LISTWEAK||product||||observe",
    }
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-LISTING-WEAK",
        "asin": "B0LISTWEAK",
        "frontend_check_status": "沿用 2026-06-20 前台数据",
        "frontend_data_freshness": "沿用 2026-06-20 前台数据",
        "frontend_findings": "沿用缓存，自动证据不足，不能用于强诊断",
        "frontend_evidence_display_tier": "仅背景参考",
        "frontend_evidence_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_cache_used": True,
        "frontend_search_status": "已自动检查",
        "frontend_search_partial_evidence": False,
        "frontend_failure_category": "none",
        "frontend_price_currency_warning": "",
        "frontend_location_warning": "",
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_location_scope": "exact",
        "frontend_location_exact": True,
        "frontend_location_verified": True,
        "competitor_comparability": "medium",
        "comparable_competitor_count": 3,
        "frontend_evidence_quality_score": 88,
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["listing_price_diagnosis_rows"] = [listing_row]
    snapshot["frontend_check_queue_rows"] = [frontend_row]
    snapshot["frontend_coverage_summary"] = daily_update._frontend_coverage_expected([frontend_row])
    payload["frontend_coverage_summary"] = daily_update._frontend_coverage_aggregate_from_snapshots(
        payload["marketplace_results"]
    )
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        listing_diagnosis_rows=[listing_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis UK listing_price_diagnosis_rows row 1 B0LISTWEAK "
        "contains frontend-backed strong listing or growth claim"
        in failure
        and "under weak frontend evidence" in failure
        and "frontend check is stale or cached" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_listing_price_action_with_weak_frontend(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    listing_row = {
        "marketplace": "UK",
        "产品": "Weak frontend price product",
        "SKU": "SKU-LISTING-PRICE",
        "ASIN": "B0LISTPRICE",
        "诊断类型": "Listing 待人工确认",
        "主因": "价格竞争力不足",
        "关键证据": "沿用缓存显示疑似高于竞品",
        "建议动作": "前台显示价格无优势，立即降价到竞品价",
        "confirmed_status": "待确认",
        "suggested_action": "先查前台",
        "normalized_action": "observe",
        "action_scope": "product",
        "action_id": "UK||SKU-LISTING-PRICE||B0LISTPRICE||product||||observe",
    }
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-LISTING-PRICE",
        "asin": "B0LISTPRICE",
        "frontend_check_status": "沿用 2026-06-20 前台数据",
        "frontend_data_freshness": "沿用 2026-06-20 前台数据",
        "frontend_findings": "沿用缓存，自动证据不足，不能用于强诊断",
        "frontend_evidence_display_tier": "仅背景参考",
        "frontend_evidence_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_cache_used": True,
        "frontend_search_status": "已自动检查",
        "frontend_search_partial_evidence": False,
        "frontend_failure_category": "none",
        "frontend_price_currency_warning": "",
        "frontend_location_warning": "",
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_location_scope": "exact",
        "frontend_location_exact": True,
        "frontend_location_verified": True,
        "competitor_comparability": "medium",
        "comparable_competitor_count": 3,
        "frontend_evidence_quality_score": 88,
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["listing_price_diagnosis_rows"] = [listing_row]
    snapshot["frontend_check_queue_rows"] = [frontend_row]
    snapshot["frontend_coverage_summary"] = daily_update._frontend_coverage_expected([frontend_row])
    payload["frontend_coverage_summary"] = daily_update._frontend_coverage_aggregate_from_snapshots(
        payload["marketplace_results"]
    )
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        listing_diagnosis_rows=[listing_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis UK listing_price_diagnosis_rows row 1 B0LISTPRICE "
        "contains frontend-backed strong listing or growth claim 立即降价"
        in failure
        and "under weak frontend evidence" in failure
        for failure in failures
    )


def test_report_refresh_failures_allows_listing_price_confirmation_copy_with_weak_frontend(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    listing_row = {
        "marketplace": "UK",
        "产品": "Weak frontend safe price product",
        "SKU": "SKU-LISTING-SAFE-PRICE",
        "ASIN": "B0LISTSAFE",
        "诊断类型": "Listing 待人工确认",
        "主因": "价格竞争力待确认",
        "关键证据": "沿用缓存显示疑似高于竞品",
        "建议动作": "先对比前三竞品价格/Coupon/配送，再决定是否调价",
        "confirmed_status": "待确认",
        "suggested_action": "先查前台",
        "normalized_action": "observe",
        "action_scope": "product",
        "action_id": "UK||SKU-LISTING-SAFE-PRICE||B0LISTSAFE||product||||observe",
    }
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-LISTING-SAFE-PRICE",
        "asin": "B0LISTSAFE",
        "frontend_check_status": "沿用 2026-06-20 前台数据",
        "frontend_data_freshness": "沿用 2026-06-20 前台数据",
        "frontend_findings": "沿用缓存，自动证据不足，不能用于强诊断",
        "frontend_evidence_display_tier": "仅背景参考",
        "frontend_evidence_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_cache_used": True,
        "frontend_search_status": "已自动检查",
        "frontend_failure_category": "none",
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_location_scope": "exact",
        "frontend_location_exact": True,
        "frontend_location_verified": True,
        "competitor_comparability": "medium",
        "comparable_competitor_count": 3,
        "frontend_evidence_quality_score": 88,
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["listing_price_diagnosis_rows"] = [listing_row]
    snapshot["frontend_check_queue_rows"] = [frontend_row]
    snapshot["frontend_coverage_summary"] = daily_update._frontend_coverage_expected([frontend_row])
    payload["frontend_coverage_summary"] = daily_update._frontend_coverage_aggregate_from_snapshots(
        payload["marketplace_results"]
    )
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        listing_diagnosis_rows=[listing_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert not any("B0LISTSAFE" in failure and "frontend-backed strong listing" in failure for failure in failures)


def test_report_refresh_failures_blocks_strong_product_frontend_without_strong_tier(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = _product_decision_row(
        asin="B0STRONGTIERMISMATCH",
        final_decision="OBSERVE",
        today_allowed_actions=["observe"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale"],
        frontend_evidence_state="ok_high",
        frontend_evidence_tier="仅背景参考",
        frontend_evidence_display_tier="强诊断可用",
        frontend_check_status="已自动检查",
        frontend_search_status="已自动检查",
        frontend_cache_used=False,
        frontend_failure_category="none",
        frontend_price_currency_warning="",
        frontend_location_warning="",
        frontend_auto_conclusion="FRONTEND_OK",
        frontend_location_scope="exact",
        frontend_location_exact=True,
        frontend_location_verified=True,
        competitor_comparability="medium",
        comparable_competitor_count=3,
        frontend_evidence_quality_score=88,
    )
    _attach_product_decision(payload, row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK product_final_decision_rows row 1 B0STRONGTIERMISMATCH "
        "marks strong frontend evidence without explicit strong flag"
    ) in failures


def test_report_refresh_failures_blocks_growth_decision_with_currency_warning(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = _product_decision_row(
        asin="B0CURRENCY",
        final_decision="SMALL_SCALE_ALLOWED",
        today_allowed_actions=["observe", "bid_up"],
        today_blocked_actions=["budget_up", "broad_scale"],
        frontend_evidence_state="ok_high",
        frontend_evidence_tier="仅背景参考",
        frontend_evidence_display_tier="强诊断可用",
        frontend_check_status="已自动检查",
        frontend_search_status="已自动检查",
        frontend_cache_used=False,
        frontend_failure_category="none",
        frontend_price_currency_warning="价格币种异常：TWD594.77，已忽略",
        frontend_auto_conclusion="FRONTEND_OK",
        frontend_location_scope="exact",
        frontend_location_exact=True,
        frontend_location_verified=True,
        competitor_comparability="medium",
        comparable_competitor_count=3,
        frontend_evidence_quality_score=88,
    )
    _attach_product_decision(payload, row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK product_final_decision_rows row 1 B0CURRENCY allows growth with currency warning"
        in failures
    )


def test_report_refresh_failures_blocks_growth_decision_with_location_warning(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = _product_decision_row(
        asin="B0LOCATION",
        final_decision="SMALL_SCALE_ALLOWED",
        today_allowed_actions=["observe", "bid_up"],
        today_blocked_actions=["budget_up", "broad_scale"],
        frontend_evidence_state="ok_high",
        frontend_evidence_tier="仅背景参考",
        frontend_evidence_display_tier="强诊断可用",
        frontend_check_status="已自动检查",
        frontend_search_status="已自动检查",
        frontend_cache_used=False,
        frontend_failure_category="none",
        frontend_location_warning="UK 地区异常：United States",
        frontend_auto_conclusion="FRONTEND_OK",
        frontend_location_scope="exact",
        frontend_location_exact=True,
        frontend_location_verified=True,
        competitor_comparability="medium",
        comparable_competitor_count=3,
        frontend_evidence_quality_score=88,
    )
    _attach_product_decision(payload, row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK product_final_decision_rows row 1 B0LOCATION allows growth with location warning"
        in failures
    )


def test_report_refresh_failures_blocks_growth_decision_with_weak_competitor_evidence(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = _product_decision_row(
        asin="B0WEAKCOMP",
        final_decision="SMALL_SCALE_ALLOWED",
        today_allowed_actions=["observe", "bid_up"],
        today_blocked_actions=["budget_up", "broad_scale"],
        frontend_evidence_state="ok_high",
        frontend_evidence_tier="强诊断可用",
        frontend_evidence_display_tier="强诊断可用",
        frontend_check_status="已自动检查",
        frontend_search_status="已自动检查",
        frontend_cache_used=False,
        frontend_failure_category="none",
        frontend_auto_conclusion="FRONTEND_OK",
        frontend_location_scope="exact",
        frontend_location_exact=True,
        frontend_location_verified=True,
        competitor_comparability="low",
        comparable_competitor_count=1,
        frontend_evidence_quality_score=88,
    )
    _attach_product_decision(payload, row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK product_final_decision_rows row 1 B0WEAKCOMP allows growth with weak competitor evidence"
        in failures
    )


def test_report_refresh_failures_allows_observe_only_with_weak_frontend_evidence(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0OBSERVE",
        "product_name": "Observe frontend product",
        "final_decision": "OBSERVE",
        "today_allowed_actions": ["observe"],
        "today_blocked_actions": ["bid_up", "budget_up", "broad_scale"],
        "frontend_evidence_state": "weak",
        "frontend_evidence_tier": "仅背景参考",
        "frontend_evidence_display_tier": "仅背景参考",
        "frontend_check_status": "待前台检查",
        "frontend_search_status": "",
        "frontend_search_partial_evidence": False,
        "frontend_cache_used": False,
        "frontend_failure_category": "",
        "frontend_price_currency_warning": "",
        "frontend_location_warning": "",
        "frontend_auto_conclusion": "FRONTEND_WEAK",
        "frontend_location_scope": "",
        "frontend_location_exact": False,
        "frontend_location_verified": False,
        "competitor_comparability": "",
        "comparable_competitor_count": "",
        "frontend_evidence_quality_score": "",
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["product_final_decision_rows"] = [row]
    payload["product_final_decision_rows"] = [row]
    snapshot["final_decision_summary"] = {"OBSERVE": 1}
    snapshot["decision_gate_counts"] = {"OBSERVE": 1}
    payload["final_decision_summary"]["UK"] = {"OBSERVE": 1}
    payload["decision_gate_counts"]["UK"] = {"OBSERVE": 1}
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert not any("B0OBSERVE allows growth" in failure for failure in failures)


def test_report_refresh_failures_blocks_strong_frontend_boolean_without_strong_tier(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = _product_decision_row(
        asin="B0STRONGBOOL",
        today_allowed_actions=["observe"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale"],
        frontend_evidence_state="weak",
        frontend_evidence_tier="仅背景参考",
        frontend_evidence_display_tier="仅背景参考",
        frontend_evidence_is_strong=True,
        frontend_check_status="待前台检查",
        frontend_search_status="",
        frontend_auto_conclusion="FRONTEND_WEAK",
        competitor_comparability="unknown",
        comparable_competitor_count=0,
        frontend_evidence_quality_score=30,
    )
    _attach_product_decision(payload, row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis UK product_final_decision_rows row 1 B0STRONGBOOL "
        "marks strong frontend evidence without strong frontend display tier" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_strong_tier_with_downgraded_display_tier(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = _product_decision_row(
        asin="B0STRONGDISPLAYDOWN",
        today_allowed_actions=["observe"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale"],
        frontend_evidence_state="weak",
        frontend_evidence_tier="强诊断可用",
        frontend_evidence_display_tier="仅背景参考",
        frontend_decision_evidence_tier="强诊断可用",
        frontend_check_status="已自动检查",
        frontend_search_status="已自动检查",
        frontend_auto_conclusion="FRONTEND_WEAK",
        frontend_failure_category="none",
        frontend_location_scope="exact",
        frontend_location_exact=True,
        frontend_location_verified=True,
        competitor_comparability="medium",
        comparable_competitor_count=3,
        frontend_evidence_quality_score=88,
    )
    _attach_product_decision(payload, row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis UK product_final_decision_rows row 1 B0STRONGDISPLAYDOWN "
        "marks strong frontend evidence without strong frontend display tier" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_weak_frontend_without_explicit_growth_blocks(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = _product_decision_row(
        asin="B0WEAKBLOCKS",
        today_allowed_actions=["observe"],
        today_blocked_actions=[],
        frontend_evidence_state="weak",
        frontend_evidence_display_tier="仅背景参考",
        frontend_auto_conclusion="FRONTEND_WEAK",
    )
    _attach_product_decision(payload, row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis UK product_final_decision_rows row 1 B0WEAKBLOCKS "
        "does not explicitly block growth actions under weak frontend evidence" in failure
        and "bid_up" in failure
        and "budget_up" in failure
        and "broad_scale" in failure
        and "create_exact_low_budget" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_task_queue_growth_action_when_product_gate_blocks_it(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        sku="SKU-GATE",
        asin="B0TASKGATE",
        today_allowed_actions=["observe"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale"],
        frontend_evidence_state="weak",
        frontend_evidence_display_tier="仅背景参考",
        frontend_auto_conclusion="FRONTEND_WEAK",
    )
    _attach_product_decision(payload, product_row)
    payload["marketplace_results"][0]["report_view_snapshot"]["today_task_queue_rows"] = [
        {
            "marketplace": "UK",
            "sku": "SKU-GATE",
            "asin": "B0TASKGATE",
            "product_name": "Blocked growth task product",
            "suggested_action": "加价 5%-10%",
            "copy_action_line": "建议加价 5%-10%",
        }
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK today_task_queue_rows row 1 B0TASKGATE "
        "contains growth action bid_up blocked by product final decision"
    ) in failures
    assert (
        "latest analysis UK today_task_queue_rows row 1 B0TASKGATE "
        "contains growth action bid_up not allowed by product final decision"
    ) in failures


def test_report_refresh_failures_blocks_aux_ad_workbench_growth_action_when_product_gate_blocks_it(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        sku="SKU-AUX-GATE",
        asin="B0AUXGATE",
        today_allowed_actions=["observe"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale"],
        frontend_evidence_state="weak",
        frontend_evidence_display_tier="仅背景参考",
        frontend_auto_conclusion="FRONTEND_WEAK",
    )
    _attach_product_decision(payload, product_row)
    payload["marketplace_results"][0]["report_view_snapshot"]["scale_keyword_rows"] = [
        {
            "marketplace": "UK",
            "sku": "SKU-AUX-GATE",
            "asin": "B0AUXGATE",
            "product_name": "Aux blocked growth product",
            "suggested_action": "加价 5%-10%",
            "copy_action_line": "建议加价 5%-10%",
        }
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK ad_workbench_rows row 1 B0AUXGATE "
        "contains growth action bid_up blocked by product final decision"
    ) in failures
    assert (
        "latest analysis UK ad_workbench_rows row 1 B0AUXGATE "
        "contains growth action bid_up not allowed by product final decision"
    ) in failures


def test_report_refresh_failures_blocks_growth_test_when_product_gate_blocks_it(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        sku="SKU-GROWTH-GATE",
        asin="B0GROWTHTEST",
        today_allowed_actions=["observe"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
        frontend_evidence_state="weak",
        frontend_evidence_display_tier="仅背景参考",
        frontend_auto_conclusion="FRONTEND_WEAK",
    )
    _attach_product_decision(payload, product_row)
    payload["marketplace_results"][0]["report_view_snapshot"]["growth_test_rows"] = [
        {
            "marketplace": "UK",
            "sku": "SKU-GROWTH-GATE",
            "asin": "B0GROWTHTEST",
            "product_name": "Growth test blocked product",
            "search_term_or_target": "blocked exact test",
            "suggested_action": "小预算试投",
            "manual_action_taken": "小预算试投",
            "normalized_action": "growth_test",
            "experiment_type": "growth_test",
        }
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK ad_workbench_rows row 1 B0GROWTHTEST "
        "contains growth action create_exact_low_budget blocked by product final decision"
    ) in failures
    assert (
        "latest analysis UK ad_workbench_rows row 1 B0GROWTHTEST "
        "contains growth action create_exact_low_budget not allowed by product final decision"
    ) in failures


def test_report_refresh_failures_treats_percent_add_text_as_bid_growth_action(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        sku="SKU-PERCENT-GATE",
        asin="B0PERCENTGATE",
        today_allowed_actions=["observe"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale"],
        frontend_evidence_state="weak",
        frontend_evidence_display_tier="仅背景参考",
        frontend_auto_conclusion="FRONTEND_WEAK",
    )
    _attach_product_decision(payload, product_row)
    payload["marketplace_results"][0]["report_view_snapshot"]["scale_keyword_rows"] = [
        {
            "marketplace": "UK",
            "sku": "SKU-PERCENT-GATE",
            "asin": "B0PERCENTGATE",
            "product_name": "Percent add blocked growth product",
            "search_term_or_target": "percent add target",
            "suggested_action": "加 5%-10%",
            "copy_action_line": "percent add target 加 5%-10%",
        }
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK ad_workbench_rows row 1 B0PERCENTGATE "
        "contains growth action bid_up blocked by product final decision"
    ) in failures
    assert (
        "latest analysis UK ad_workbench_rows row 1 B0PERCENTGATE "
        "contains growth action bid_up not allowed by product final decision"
    ) in failures


def test_report_refresh_failures_does_not_treat_conservative_no_budget_text_as_growth(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        sku="SKU-NO-BUDGET",
        asin="B0NOBUDGET",
        today_allowed_actions=["observe"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale"],
        frontend_evidence_state="weak",
        frontend_evidence_display_tier="仅背景参考",
        frontend_auto_conclusion="FRONTEND_WEAK",
    )
    _attach_product_decision(payload, product_row)
    payload["marketplace_results"][0]["report_view_snapshot"]["today_task_queue_rows"] = [
        {
            "marketplace": "UK",
            "sku": "SKU-NO-BUDGET",
            "asin": "B0NOBUDGET",
            "product_name": "Conservative task product",
            "today_action": "保守跑：不加预算，不推大词放量，只观察。",
            "suggested_action": "观察",
        }
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert not any("B0NOBUDGET contains growth action" in failure for failure in failures)


def test_report_refresh_failures_does_not_block_already_executed_growth_trace(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        sku="SKU-EXECUTED",
        asin="B0EXECUTED",
        today_allowed_actions=["observe"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale"],
        frontend_evidence_state="weak",
        frontend_evidence_display_tier="仅背景参考",
        frontend_auto_conclusion="FRONTEND_WEAK",
    )
    _attach_product_decision(payload, product_row)
    payload["marketplace_results"][0]["report_view_snapshot"]["today_task_queue_rows"] = [
        {
            "marketplace": "UK",
            "sku": "SKU-EXECUTED",
            "asin": "B0EXECUTED",
            "product_name": "Executed growth trace product",
            "suggested_action": "加价 5%-10%",
            "normalized_action": "bid_up",
            "confirmed_status": "已执行",
        }
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert not any("B0EXECUTED contains growth action" in failure for failure in failures)


def test_report_refresh_failures_blocks_today_action_groups_missing_task_row(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    payload["marketplace_results"][0]["report_view_snapshot"]["today_task_queue_rows"] = [
        {
            "marketplace": "UK",
            "sku": "SKU-GROUP",
            "asin": "B0GROUPMISS",
            "product_name": "Missing group product",
            "today_action": "核心词降竞价 10%-20%",
            "action_group": "广告动作",
            "confirmed_status": "待确认",
        }
    ]
    payload["marketplace_results"][0]["report_view_snapshot"]["today_action_groups"] = {"广告动作": []}
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis UK today_action_groups missing task rows" in failure
        and "B0GROUPMISS" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_today_action_groups_stale_task_row(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    payload["marketplace_results"][0]["report_view_snapshot"]["today_task_queue_rows"] = [
        {
            "marketplace": "UK",
            "sku": "SKU-GROUP",
            "asin": "B0GROUPLIVE",
            "product_name": "Live group product",
            "today_action": "核心词降竞价 10%-20%",
            "action_group": "广告动作",
            "confirmed_status": "待确认",
        }
    ]
    payload["marketplace_results"][0]["report_view_snapshot"]["today_action_groups"] = {
        "广告动作": [
            {
                "marketplace": "UK",
                "sku": "SKU-GROUP",
                "asin": "B0GROUPLIVE",
                "today_action": "核心词降竞价 10%-20%",
                "action_group": "广告动作",
            },
            {
                "marketplace": "UK",
                "sku": "SKU-OLD",
                "asin": "B0GROUPOLD",
                "today_action": "旧动作",
                "action_group": "广告动作",
            },
        ]
    }
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis UK today_action_groups contains stale task rows" in failure
        and "B0GROUPOLD" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_frontend_queue_cached_strong_evidence(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0QUEUECACHE",
        "frontend_check_status": "沿用 2026-06-20 前台数据",
        "frontend_data_freshness": "沿用 2026-06-20 前台数据",
        "frontend_findings": "沿用 2026-06-20 前台数据",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_is_strong": True,
        "frontend_cache_used": True,
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_failure_category": "none",
        "frontend_location_scope": "exact",
        "frontend_location_verified": True,
        "frontend_location_exact": True,
        "frontend_search_status": "已自动检查",
        "frontend_competitor_count": 3,
        "competitor_comparability": "medium",
        "frontend_evidence_quality_score": 82,
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["frontend_check_queue_rows"] = [frontend_row]
    snapshot["frontend_coverage_summary"] = daily_update._frontend_coverage_expected([frontend_row])
    payload["frontend_coverage_summary"] = daily_update._frontend_coverage_aggregate_from_snapshots(
        payload["marketplace_results"]
    )
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK frontend_check_queue_rows row 1 B0QUEUECACHE "
        "cached or pending evidence marked strong"
    ) in failures


def test_report_refresh_failures_blocks_frontend_queue_strong_without_explicit_flag(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0QUEUEFLAG",
        "frontend_check_status": "已自动检查",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_cache_used": False,
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_failure_category": "none",
        "frontend_location_scope": "exact",
        "frontend_location_verified": True,
        "frontend_location_exact": True,
        "frontend_search_status": "已自动检查",
        "frontend_competitor_count": 3,
        "competitor_comparability": "medium",
        "frontend_evidence_quality_score": 82,
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [frontend_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK frontend_check_queue_rows row 1 B0QUEUEFLAG "
        "strong evidence missing explicit strong flag"
    ) in failures


def test_report_refresh_failures_blocks_frontend_queue_missing_from_daily_excel(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-FRONTEXCEL",
        "asin": "B0FRONTEXCEL",
        "product_name": "Frontend Excel product",
        "frontend_check_status": "待前台检查",
        "frontend_data_freshness": "无可用前台数据",
        "frontend_evidence_tier": "仅背景参考",
        "frontend_evidence_display_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_cache_used": False,
        "frontend_auto_conclusion": "FRONTEND_WEAK",
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [frontend_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 前台证据队列 missing fields for frontend check queue rows" in failure
        and "frontend_check_status" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_frontend_queue_daily_excel_field_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-FRONTEXCEL",
        "asin": "B0FRONTFIELD",
        "product_name": "Frontend Excel field product",
        "frontend_check_status": "待前台检查",
        "frontend_data_freshness": "无可用前台数据",
        "frontend_evidence_tier": "仅背景参考",
        "frontend_evidence_display_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_cache_used": False,
        "frontend_auto_conclusion": "FRONTEND_WEAK",
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["frontend_check_queue_rows"] = [frontend_row]
    snapshot["frontend_coverage_summary"] = daily_update._frontend_coverage_expected([frontend_row])
    payload["frontend_coverage_summary"] = daily_update._frontend_coverage_aggregate_from_snapshots(
        payload["marketplace_results"]
    )
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    excel_row = dict(frontend_row)
    excel_row["frontend_evidence_tier"] = "强诊断可用"
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        frontend_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 前台证据队列 field mismatch for frontend check queue rows "
        "('UK', 'SKU-FRONTEXCEL', 'B0FRONTFIELD') field frontend_evidence_tier" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_frontend_queue_excel_next_step_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-FRONTEXPLAIN",
        "asin": "B0FRONTEXPLAIN",
        "product_name": "Frontend explanation product",
        "trigger_reason": "近14天广告点击高但本 SKU 无单",
        "key_metrics": "点击 34；广告订单 0；总单 1",
        "frontend_check_status": "待前台检查",
        "frontend_data_freshness": "无可用前台数据",
        "frontend_findings": "自动证据不足，不能用于强诊断；尚未读取到可用前台字段。",
        "suspected_issue": "前台转化阻力待确认",
        "conservative_action": "广告先降竞价，不加预算。",
        "recommended_next_step": "先刷新前台证据，再决定是否修 Listing。",
        "frontend_evidence_tier": "仅背景参考",
        "frontend_evidence_display_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_cache_used": False,
        "frontend_auto_conclusion": "FRONTEND_WEAK",
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["frontend_check_queue_rows"] = [frontend_row]
    snapshot["frontend_coverage_summary"] = daily_update._frontend_coverage_expected([frontend_row])
    payload["frontend_coverage_summary"] = daily_update._frontend_coverage_aggregate_from_snapshots(
        payload["marketplace_results"]
    )
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    excel_row = dict(frontend_row)
    excel_row["recommended_next_step"] = "Excel 错误改成直接加预算。"
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        frontend_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 前台证据队列 field mismatch for frontend check queue rows "
        "('UK', 'SKU-FRONTEXPLAIN', 'B0FRONTEXPLAIN') field recommended_next_step" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_frontend_queue_excel_audit_reasons_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-FRONTAUDIT",
        "asin": "B0FRONTAUDIT",
        "product_name": "Frontend audit product",
        "frontend_check_status": "已自动检查",
        "frontend_data_freshness": "当前读取",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_display_tier": "仅背景参考",
        "frontend_decision_evidence_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_cache_used": False,
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_evidence_audit_summary": "前台证据可辅助判断，不能单独放量",
        "frontend_evidence_audit_reasons": ["地区未确认", "竞品可比性未达强诊断"],
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["frontend_check_queue_rows"] = [frontend_row]
    snapshot["frontend_coverage_summary"] = daily_update._frontend_coverage_expected([frontend_row])
    payload["frontend_coverage_summary"] = daily_update._frontend_coverage_aggregate_from_snapshots(
        payload["marketplace_results"]
    )
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    excel_row = dict(frontend_row)
    excel_row["frontend_evidence_audit_reasons"] = ["产品页和搜索页均可用"]
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        frontend_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 前台证据队列 field mismatch for frontend check queue rows "
        "('UK', 'SKU-FRONTAUDIT', 'B0FRONTAUDIT') field frontend_evidence_audit_reasons" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_ad_processing_queue_missing_from_daily_excel(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    ad_row = {
        "marketplace": "UK",
        "sku": "SKU-ADEXCEL",
        "asin": "B0ADEXCEL",
        "product_name": "Ad Excel product",
        "search_term_or_target": "dimmer desk lamp",
        "campaign_name": "Manual campaign",
        "ad_group_name": "Exact group",
        "match_type_or_targeting": "EXACT",
        "suggested_action": "降竞价10%-20%",
        "copy_action_line": "建议降竞价 10%-20%",
        "reason": "5 次点击 0 单",
        "clicks": "5",
        "orders": "0",
        "normalized_action": "bid_down",
        "action_scope": "search_term",
        "action_id": "UK||SKU-ADEXCEL||B0ADEXCEL||search_term||dimmer desk lamp||bid_down",
        "confirmed_status": "待确认",
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["search_term_processing_queue_rows"] = [ad_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 广告处理队列 missing fields for ad processing queue rows" in failure
        and "search_term_or_target" in failure
        and "suggested_action" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_scale_candidate_daily_excel_field_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    scale_row = {
        "站点": "UK",
        "产品": "Scale product",
        "SKU": "SKU-SCALE",
        "ASIN": "B0SCALE",
        "点击": "19",
        "花费": "£6.28",
        "订单": "3",
        "总单": "3",
        "ACOS": "6.7%",
        "目标 ACOS": "9.1%",
        "放量等级": "谨慎放量候选",
        "建议": "只恢复核心词展示",
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["scale_rows"] = [scale_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    excel_row = dict(scale_row)
    excel_row["放量等级"] = "可小幅放量"
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        scale_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 放量候选 field mismatch for scale candidate rows "
        "('UK', 'SKU-SCALE', 'B0SCALE') field 放量等级" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_scale_keyword_daily_excel_field_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    scale_keyword_row = {
        "marketplace": "UK",
        "sku": "SKU-SCALEKW",
        "asin": "B0SCALEKW",
        "product_name": "Scale keyword product",
        "search_term_or_target": "serving board",
        "campaign_name": "Auto campaign",
        "ad_group_name": "Auto group",
        "match_type_or_targeting": "TARGETING_EXPRESSION_PREDEFINED",
        "clicks": "4",
        "ad_orders": "1",
        "target_acos": "10.0%",
        "scale_action": "试探提高竞价 3%-5%",
        "normalized_action": "bid_up",
        "action_scope": "search_term",
        "action_id": "UK||SKU-SCALEKW||B0SCALEKW||search_term||serving board||bid_up",
        "reason": "低 ACOS 小样本",
        "product_scale_level": "谨慎放量候选",
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["scale_keyword_rows"] = [scale_keyword_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    excel_row = dict(scale_keyword_row)
    excel_row["reason"] = "Excel 中被改写"
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        scale_keyword_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 放量词候选 field mismatch for scale keyword rows "
        "('UK', 'SKU-SCALEKW', 'B0SCALEKW', 'serving board', 'Auto campaign', 'Auto group', "
        "'TARGETING_EXPRESSION_PREDEFINED', '试探提高竞价 3%-5%') field reason" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_ad_processing_daily_excel_action_id_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    ad_row = {
        "marketplace": "UK",
        "sku": "SKU-ADID",
        "asin": "B0ADIDENT",
        "product_name": "Ad identity product",
        "search_term_or_target": "dimmer desk lamp",
        "campaign_name": "Manual campaign",
        "ad_group_name": "Exact group",
        "match_type_or_targeting": "EXACT",
        "suggested_action": "降竞价10%-20%",
        "copy_action_line": "建议降竞价 10%-20%",
        "reason": "5 次点击 0 单",
        "clicks": "5",
        "orders": "0",
        "normalized_action": "bid_down",
        "action_scope": "search_term",
        "action_id": "UK||SKU-ADID||B0ADIDENT||search_term||dimmer desk lamp||bid_down",
        "confirmed_status": "待确认",
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["search_term_processing_queue_rows"] = [ad_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    excel_row = dict(ad_row)
    excel_row["action_id"] = "UK||SKU-ADID||B0ADIDENT||search_term||dimmer desk lamp||observe"
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        ad_processing_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 广告处理队列 field mismatch for ad processing queue rows "
        "('UK', 'SKU-ADID', 'B0ADIDENT', 'dimmer desk lamp', 'Manual campaign', 'Exact group', "
        "'EXACT', '降竞价10%-20%') field action_id" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_scale_keyword_daily_excel_action_scope_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    scale_keyword_row = {
        "marketplace": "UK",
        "sku": "SKU-SCALEID",
        "asin": "B0SCALEID",
        "product_name": "Scale keyword identity product",
        "search_term_or_target": "serving board",
        "campaign_name": "Auto campaign",
        "ad_group_name": "Auto group",
        "match_type_or_targeting": "EXACT",
        "clicks": "4",
        "ad_orders": "1",
        "target_acos": "10.0%",
        "scale_action": "试探提高竞价 3%-5%",
        "normalized_action": "bid_up",
        "action_scope": "search_term",
        "action_id": "UK||SKU-SCALEID||B0SCALEID||search_term||serving board||bid_up",
        "reason": "低 ACOS 小样本",
        "product_scale_level": "谨慎放量候选",
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["scale_keyword_rows"] = [scale_keyword_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    excel_row = dict(scale_keyword_row)
    excel_row["action_scope"] = "product"
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        scale_keyword_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 放量词候选 field mismatch for scale keyword rows "
        "('UK', 'SKU-SCALEID', 'B0SCALEID', 'serving board', 'Auto campaign', 'Auto group', "
        "'EXACT', '试探提高竞价 3%-5%') field action_scope" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_ad_memory_blocked_row_reusing_original_action_identity(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    blocked_row = {
        "marketplace": "UK",
        "sku": "SKU-MEMBLOCK",
        "asin": "B0MEMBLOCK",
        "product_name": "Memory blocked product",
        "search_term_or_target": "blocked bid up term",
        "campaign_name": "Manual campaign",
        "ad_group_name": "Exact group",
        "match_type_or_targeting": "EXACT",
        "suggested_action": "观察",
        "copy_action_line": "建议观察",
        "copy_block": "建议观察\nblocked bid up term",
        "ad_memory_blocked": True,
        "blocked_action_id": "UK||SKU-MEMBLOCK||B0MEMBLOCK||search_term||blocked bid up term||bid_up",
        "blocked_original_action": "加价 5%-10%",
        "normalized_action": "bid_up",
        "action_scope": "search_term",
        "action_id": "UK||SKU-MEMBLOCK||B0MEMBLOCK||search_term||blocked bid up term||bid_up",
        "confirmed_status": "待确认",
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["search_term_processing_queue_rows"] = [blocked_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        ad_processing_rows=[blocked_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK search_term_processing_queue_rows row 1 B0MEMBLOCK "
        "ad memory blocked row must normalize current action to observe"
    ) in failures
    assert (
        "latest analysis UK search_term_processing_queue_rows row 1 B0MEMBLOCK "
        "ad memory blocked row reuses blocked action_id as current action_id"
    ) in failures


def test_report_refresh_failures_blocks_growth_test_daily_excel_field_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    growth_row = {
        "marketplace": "UK",
        "sku": "SKU-GROWTHEX",
        "asin": "B0GROWTHEX",
        "product_name": "Growth Excel product",
        "search_term_or_target": "adjustable desk lamp",
        "suggested_action": "小预算试投",
        "manual_action_taken": "小预算试投",
        "normalized_action": "growth_test",
        "action_scope": "search_term",
        "action_id": "UK||SKU-GROWTHEX||B0GROWTHEX||search_term||adjustable desk lamp||growth_test",
        "experiment_type": "growth_test",
        "term_source": "search_term_report",
        "evidence_level": "核心强相关",
        "test_days": "7",
        "report_date": "2026-06-24",
        "next_review": "2026-07-01",
        "cooldown_days": 7,
        "suggested_daily_budget": "£3.00",
        "suggested_bid_min": "£0.18",
        "suggested_bid_max": "£0.28",
        "stop_loss_rule": "7天无本 SKU 订单则停止",
        "success_rule": "7天内至少出现本 SKU 订单",
        "confirmed_status": "待确认",
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["growth_test_rows"] = [growth_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    excel_row = dict(growth_row)
    excel_row["success_rule"] = "Excel 中丢失本 SKU 复盘口径"
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        growth_test_rows=[excel_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "daily Excel 小预算试投 field mismatch for growth test rows "
        "('UK', 'SKU-GROWTHEX', 'B0GROWTHEX', 'adjustable desk lamp', "
        "'UK||SKU-GROWTHEX||B0GROWTHEX||search_term||adjustable desk lamp||growth_test') field success_rule"
        in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_action_id_scope_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    growth_row = {
        "marketplace": "UK",
        "sku": "SKU-GROWTHID",
        "asin": "B0GROWTHID",
        "product_name": "Growth identity product",
        "search_term_or_target": "adjustable desk lamp",
        "suggested_action": "小预算试投",
        "manual_action_taken": "小预算试投",
        "normalized_action": "growth_test",
        "action_scope": "keyword",
        "action_id": "UK||SKU-GROWTHID||B0GROWTHID||search_term||adjustable desk lamp||growth_test",
        "experiment_type": "growth_test",
        "term_source": "search_term_report",
        "evidence_level": "核心强相关",
        "test_days": "7",
        "report_date": "2026-06-24",
        "next_review": "2026-07-01",
        "cooldown_days": 7,
        "suggested_daily_budget": "£3.00",
        "suggested_bid_min": "£0.18",
        "suggested_bid_max": "£0.28",
        "stop_loss_rule": "7天无本 SKU 订单则停止",
        "success_rule": "7天内至少出现本 SKU 订单",
        "confirmed_status": "待确认",
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["growth_test_rows"] = [growth_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        growth_test_rows=[growth_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK growth_test_rows row 1 B0GROWTHID has invalid action_scope keyword; "
        "expected one of asin_target, campaign, product, search_term"
    ) in failures


def test_report_refresh_failures_blocks_frontend_coverage_summary_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0QUEUECOVER",
        "frontend_check_status": "沿用 2026-06-20 前台数据",
        "frontend_data_freshness": "沿用 2026-06-20 前台数据",
        "frontend_findings": "沿用 2026-06-20 前台数据",
        "frontend_evidence_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_cache_used": True,
        "frontend_auto_conclusion": "FRONTEND_WEAK",
        "frontend_search_status": "已读取部分结果",
        "frontend_search_partial_evidence": True,
        "frontend_evidence_quality_score": 58,
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["frontend_check_queue_rows"] = [frontend_row]
    snapshot["frontend_coverage_summary"] = daily_update._frontend_coverage_expected([frontend_row])
    snapshot["frontend_coverage_summary"]["frontend_decision_ready_count"] = 1
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis UK frontend_coverage_summary field frontend_decision_ready_count: "
        "expected 0, got 1" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_top_level_frontend_coverage_summary_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0TOPCOVER",
        "frontend_check_status": "沿用 2026-06-20 前台数据",
        "frontend_data_freshness": "沿用 2026-06-20 前台数据",
        "frontend_findings": "沿用 2026-06-20 前台数据",
        "frontend_evidence_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_cache_used": True,
        "frontend_auto_conclusion": "FRONTEND_WEAK",
        "frontend_search_status": "已读取部分结果",
        "frontend_search_partial_evidence": True,
        "frontend_evidence_quality_score": 58,
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["frontend_check_queue_rows"] = [frontend_row]
    snapshot["frontend_coverage_summary"] = daily_update._frontend_coverage_expected([frontend_row])
    payload["frontend_coverage_summary"] = daily_update._frontend_coverage_aggregate_from_snapshots(
        payload["marketplace_results"]
    )
    payload["frontend_coverage_summary"]["frontend_reference_evidence_count"] = 0
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis top-level frontend_coverage_summary field frontend_reference_evidence_count: "
        "expected 1, got 0"
    ) in failures


def test_dashboard_frontend_coverage_validation_uses_top_level_summary(tmp_path) -> None:
    payload = _analysis_payload("2026-06-24")
    payload["frontend_coverage_summary"] = {
        "frontend_queue_total": 2,
        "frontend_decision_ready_count": 1,
        "frontend_reference_evidence_count": 1,
    }
    (tmp_path / "latest_analysis.json").write_text(json.dumps(payload), encoding="utf-8")
    (tmp_path / "dashboard.html").write_text(
        "2026-06-24 latest_recommendations.html summary.html UK US DE",
        encoding="utf-8",
    )

    failures = daily_update._dashboard_marketplace_summary_failures("2026-06-24", tmp_path)

    assert "dashboard.html missing frontend coverage token 产品页成功" in failures
    assert "dashboard.html missing frontend coverage token 0/2" in failures


def test_frontend_coverage_expected_downgrades_unsafe_strong_display_tier() -> None:
    summary = daily_update._frontend_coverage_expected(
        [
            {
                "frontend_check_status": "已自动检查",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_evidence_display_tier": "仅背景参考",
                "frontend_evidence_is_strong": False,
                "frontend_search_status": "已自动检查",
            }
        ]
    )

    assert summary["frontend_usable_evidence_count"] == 1
    assert summary["frontend_decision_ready_count"] == 0
    assert summary["frontend_reference_evidence_count"] == 1
    assert summary["frontend_strong_evidence_count"] == 0
    assert summary["frontend_background_evidence_count"] == 1
    assert summary["frontend_decision_ready_label"] == "0/1 强证据，0%"
    assert summary["frontend_reference_evidence_label"] == "1/1 背景参考，100%"


def test_frontend_coverage_expected_counts_cached_rows_as_stale_or_pending() -> None:
    summary = daily_update._frontend_coverage_expected(
        [
            {
                "frontend_check_status": "沿用 2026-06-20 前台数据",
                "frontend_cache_used": True,
                "frontend_evidence_quality_score": 58,
                "frontend_evidence_tier": "仅背景参考",
                "frontend_evidence_display_tier": "仅背景参考",
            }
        ]
    )

    assert summary["frontend_cached_count"] == 1
    assert summary["frontend_pending_or_stale_count"] == 1
    assert summary["frontend_decision_ready_count"] == 0
    assert summary["frontend_reference_evidence_count"] == 1


def test_report_refresh_failures_blocks_latest_recommendations_missing_frontend_coverage_label(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0HTMLCOVER",
        "frontend_check_status": "沿用 2026-06-20 前台数据",
        "frontend_data_freshness": "沿用 2026-06-20 前台数据",
        "frontend_findings": "沿用 2026-06-20 前台数据",
        "frontend_evidence_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_cache_used": True,
        "frontend_auto_conclusion": "FRONTEND_WEAK",
        "frontend_search_status": "已读取部分结果",
        "frontend_search_partial_evidence": True,
        "frontend_evidence_quality_score": 58,
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["frontend_check_queue_rows"] = [frontend_row]
    snapshot["frontend_coverage_summary"] = daily_update._frontend_coverage_expected([frontend_row])
    payload["frontend_coverage_summary"] = daily_update._frontend_coverage_aggregate_from_snapshots(
        payload["marketplace_results"]
    )
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24"),
        encoding="utf-8",
    )
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert "latest_recommendations.html missing frontend coverage token 产品页成功" in failures
    assert "latest_recommendations.html missing frontend coverage token 0/1" in failures


def test_report_refresh_failures_blocks_frontend_queue_failed_status_without_fallback(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0QUEUEFAILED",
        "frontend_check_status": "读取失败",
        "frontend_data_freshness": "",
        "frontend_findings": "自动读取失败",
        "frontend_evidence_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_cache_used": False,
        "frontend_auto_conclusion": "FRONTEND_WEAK",
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [frontend_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK frontend_check_queue_rows row 1 B0QUEUEFAILED "
        "missing cached-date or pending-check status"
    ) in failures


def test_report_refresh_failures_blocks_frontend_queue_cache_status_without_date(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0QUEUECACHEDATE",
        "frontend_check_status": "沿用缓存前台数据",
        "frontend_data_freshness": "沿用缓存前台数据",
        "frontend_findings": "沿用缓存前台数据",
        "frontend_evidence_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_cache_used": True,
        "frontend_auto_conclusion": "FRONTEND_WEAK",
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [frontend_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK frontend_check_queue_rows row 1 B0QUEUECACHEDATE "
        "cache status missing date"
    ) in failures


def test_report_refresh_failures_blocks_frontend_queue_currency_warning_strong_evidence(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0QUEUECURRENCY",
        "frontend_check_status": "已自动检查",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_is_strong": True,
        "frontend_cache_used": False,
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_price_currency_warning": "价格币种异常：TWD594.77，已忽略",
        "frontend_failure_category": "none",
        "frontend_location_scope": "exact",
        "frontend_location_verified": True,
        "frontend_location_exact": True,
        "frontend_search_status": "已自动检查",
        "frontend_competitor_count": 3,
        "competitor_comparability": "medium",
        "frontend_evidence_quality_score": 82,
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [frontend_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK frontend_check_queue_rows row 1 B0QUEUECURRENCY "
        "currency warning marked strong"
    ) in failures


def test_report_refresh_failures_blocks_frontend_queue_location_warning_strong_evidence(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0QUEUELOCWARN",
        "frontend_check_status": "已自动检查",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_is_strong": True,
        "frontend_cache_used": False,
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_location_warning": "UK 地区异常：United States",
        "frontend_failure_category": "none",
        "frontend_location_scope": "exact",
        "frontend_location_verified": True,
        "frontend_location_exact": True,
        "frontend_search_status": "已自动检查",
        "frontend_competitor_count": 3,
        "competitor_comparability": "medium",
        "frontend_evidence_quality_score": 82,
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [frontend_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK frontend_check_queue_rows row 1 B0QUEUELOCWARN "
        "location warning marked strong"
    ) in failures


def test_report_refresh_failures_blocks_frontend_queue_failure_category_strong_evidence(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0QUEUEFAIL",
        "frontend_check_status": "已自动检查",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_is_strong": True,
        "frontend_cache_used": False,
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_failure_category": "captcha_blocked",
        "frontend_location_scope": "exact",
        "frontend_location_verified": True,
        "frontend_location_exact": True,
        "frontend_search_status": "已自动检查",
        "frontend_competitor_count": 3,
        "competitor_comparability": "medium",
        "frontend_evidence_quality_score": 82,
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [frontend_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK frontend_check_queue_rows row 1 B0QUEUEFAIL "
        "failure category marked strong"
    ) in failures


def test_report_refresh_failures_blocks_frontend_queue_low_quality_strong_evidence(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0QUEUEQUALITY",
        "frontend_check_status": "已自动检查",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_is_strong": True,
        "frontend_cache_used": False,
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_failure_category": "none",
        "frontend_location_scope": "exact",
        "frontend_location_verified": True,
        "frontend_location_exact": True,
        "frontend_search_status": "已自动检查",
        "frontend_competitor_count": 3,
        "competitor_comparability": "medium",
        "frontend_evidence_quality_score": 72,
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [frontend_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK frontend_check_queue_rows row 1 B0QUEUEQUALITY "
        "strong evidence quality score below threshold"
    ) in failures


def test_report_refresh_failures_blocks_frontend_queue_missing_comparability_strong_evidence(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0QUEUECOMP",
        "frontend_check_status": "已自动检查",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_is_strong": True,
        "frontend_cache_used": False,
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_failure_category": "none",
        "frontend_location_scope": "exact",
        "frontend_location_verified": True,
        "frontend_location_exact": True,
        "frontend_search_status": "已自动检查",
        "frontend_competitor_count": 3,
        "frontend_evidence_quality_score": 82,
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [frontend_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK frontend_check_queue_rows row 1 B0QUEUECOMP "
        "strong evidence missing competitor comparability"
    ) in failures


def test_report_refresh_failures_blocks_frontend_queue_frontend_ok_without_strong_tier(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0QUEUETIER",
        "frontend_check_status": "已自动检查",
        "frontend_evidence_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_cache_used": False,
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_failure_category": "none",
        "frontend_location_scope": "exact",
        "frontend_location_verified": True,
        "frontend_location_exact": True,
        "frontend_search_status": "已自动检查",
        "frontend_competitor_count": 3,
        "competitor_comparability": "medium",
        "frontend_evidence_quality_score": 82,
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [frontend_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK frontend_check_queue_rows row 1 B0QUEUETIER "
        "strong evidence missing strong decision tier"
    ) in failures


def test_report_refresh_failures_blocks_frontend_queue_partial_search_strong_evidence(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0QUEUEPARTIAL",
        "frontend_check_status": "已自动检查",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_is_strong": True,
        "frontend_cache_used": False,
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_failure_category": "none",
        "frontend_location_scope": "exact",
        "frontend_location_verified": True,
        "frontend_location_exact": True,
        "frontend_search_status": "已读取部分结果",
        "frontend_search_partial_evidence": True,
        "frontend_competitor_count": 3,
        "competitor_comparability": "medium",
        "frontend_evidence_quality_score": 82,
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [frontend_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK frontend_check_queue_rows row 1 B0QUEUEPARTIAL "
        "partial search evidence marked strong"
    ) in failures


def test_report_refresh_failures_blocks_frontend_queue_uncertain_location_strong_evidence(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0QUEUELOCATION",
        "frontend_check_status": "已自动检查",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_is_strong": True,
        "frontend_cache_used": False,
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_failure_category": "none",
        "frontend_location_scope": "marketplace",
        "frontend_location_verified": True,
        "frontend_location_exact": False,
        "frontend_location_uncertain": True,
        "frontend_search_status": "已自动检查",
        "frontend_competitor_count": 3,
        "competitor_comparability": "medium",
        "frontend_evidence_quality_score": 82,
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [frontend_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK frontend_check_queue_rows row 1 B0QUEUELOCATION "
        "uncertain location marked strong"
    ) in failures


def test_report_refresh_failures_blocks_frontend_queue_missing_search_success_strong_evidence(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0QUEUEMISSINGSEARCH",
        "frontend_check_status": "已自动检查",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_is_strong": True,
        "frontend_cache_used": False,
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_failure_category": "none",
        "frontend_location_scope": "exact",
        "frontend_location_verified": True,
        "frontend_location_exact": True,
        "frontend_search_status": "",
        "frontend_competitor_count": 3,
        "competitor_comparability": "medium",
        "frontend_evidence_quality_score": 82,
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [frontend_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK frontend_check_queue_rows row 1 B0QUEUEMISSINGSEARCH "
        "strong evidence missing successful search page"
    ) in failures


def test_report_refresh_failures_blocks_frontend_queue_missing_competitor_count_strong_evidence(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0QUEUEMISSINGCOMP",
        "frontend_check_status": "已自动检查",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_is_strong": True,
        "frontend_cache_used": False,
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_failure_category": "none",
        "frontend_location_scope": "exact",
        "frontend_location_verified": True,
        "frontend_location_exact": True,
        "frontend_search_status": "已自动检查",
        "competitor_comparability": "medium",
        "frontend_evidence_quality_score": 82,
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [frontend_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK frontend_check_queue_rows row 1 B0QUEUEMISSINGCOMP "
        "strong evidence missing competitor count"
    ) in failures


def test_report_refresh_failures_blocks_frontend_queue_weak_competitor_strong_evidence(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0QUEUEWEAKCOMP",
        "frontend_check_status": "已自动检查",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_is_strong": True,
        "frontend_cache_used": False,
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_failure_category": "none",
        "frontend_location_scope": "exact",
        "frontend_location_verified": True,
        "frontend_location_exact": True,
        "frontend_search_status": "已自动检查",
        "frontend_competitor_count": 1,
        "competitor_comparability": "low",
        "frontend_evidence_quality_score": 82,
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [frontend_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK frontend_check_queue_rows row 1 B0QUEUEWEAKCOMP "
        "weak competitor evidence marked strong"
    ) in failures


def test_report_refresh_failures_blocks_frontend_queue_low_comparable_competitor_strong_evidence(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0QUEUELOWCOMPARABLE",
        "frontend_check_status": "已自动检查",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_is_strong": True,
        "frontend_cache_used": False,
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_failure_category": "none",
        "frontend_location_scope": "exact",
        "frontend_location_verified": True,
        "frontend_location_exact": True,
        "frontend_search_status": "已自动检查",
        "frontend_competitor_count": 3,
        "competitor_comparability": "medium",
        "comparable_competitor_count": 0,
        "frontend_evidence_quality_score": 82,
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["frontend_check_queue_rows"] = [frontend_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK frontend_check_queue_rows row 1 B0QUEUELOWCOMPARABLE "
        "weak competitor evidence marked strong"
    ) in failures


def test_report_refresh_failures_blocks_strong_frontend_conclusion_with_cached_evidence(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0STRONGCACHE",
        "product_name": "Cached strong frontend product",
        "final_decision": "WAIT_REVIEW",
        "today_allowed_actions": ["observe"],
        "today_blocked_actions": ["bid_up", "budget_up", "broad_scale"],
        "frontend_evidence_state": "ok_high",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_display_tier": "强诊断可用",
        "frontend_check_status": "沿用 2026-06-20 前台数据",
        "frontend_search_status": "已自动检查",
        "frontend_search_partial_evidence": False,
        "frontend_cache_used": True,
        "frontend_failure_category": "none",
        "frontend_price_currency_warning": "",
        "frontend_location_warning": "",
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_location_scope": "exact",
        "frontend_location_exact": True,
        "frontend_location_verified": True,
        "competitor_comparability": "medium",
        "comparable_competitor_count": 3,
        "frontend_evidence_quality_score": 88,
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["product_final_decision_rows"] = [row]
    payload["product_final_decision_rows"] = [row]
    snapshot["final_decision_summary"] = {"WAIT_REVIEW": 1}
    snapshot["decision_gate_counts"] = {"WAIT_REVIEW": 1}
    payload["final_decision_summary"]["UK"] = {"WAIT_REVIEW": 1}
    payload["decision_gate_counts"]["UK"] = {"WAIT_REVIEW": 1}
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK product_final_decision_rows row 1 B0STRONGCACHE marks strong frontend evidence "
        "with cached frontend evidence"
    ) in failures


def test_report_refresh_failures_blocks_growth_without_explicit_frontend_strong_flag(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0GROWNOFLAG",
        "product_name": "Growth without explicit frontend flag product",
        "final_decision": "EXECUTE_TODAY",
        "today_allowed_actions": ["observe", "bid_up"],
        "today_blocked_actions": ["budget_up", "broad_scale"],
        "frontend_evidence_state": "ok_high",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_display_tier": "强诊断可用",
        "frontend_evidence_is_strong": False,
        "frontend_check_status": "已自动检查",
        "frontend_search_status": "已自动检查",
        "frontend_search_partial_evidence": False,
        "frontend_cache_used": False,
        "frontend_failure_category": "none",
        "frontend_price_currency_warning": "",
        "frontend_location_warning": "",
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_location_scope": "exact",
        "frontend_location_exact": True,
        "frontend_location_verified": True,
        "competitor_comparability": "medium",
        "comparable_competitor_count": 3,
        "frontend_evidence_quality_score": 88,
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["product_final_decision_rows"] = [row]
    payload["product_final_decision_rows"] = [row]
    snapshot["final_decision_summary"] = {"EXECUTE_TODAY": 1}
    snapshot["decision_gate_counts"] = {"EXECUTE_TODAY": 1}
    payload["final_decision_summary"]["UK"] = {"EXECUTE_TODAY": 1}
    payload["decision_gate_counts"]["UK"] = {"EXECUTE_TODAY": 1}
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK product_final_decision_rows row 1 B0GROWNOFLAG "
        "allows growth without explicit strong frontend flag"
    ) in failures


def test_report_refresh_failures_blocks_strong_product_decision_without_competitor_comparability(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0STRONGNOCOMP",
        "product_name": "Missing comparability product",
        "final_decision": "WAIT_REVIEW",
        "today_allowed_actions": ["observe"],
        "today_blocked_actions": ["bid_up", "budget_up", "broad_scale"],
        "frontend_evidence_state": "ok_high",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_display_tier": "强诊断可用",
        "frontend_check_status": "已自动检查",
        "frontend_search_status": "已自动检查",
        "frontend_search_partial_evidence": False,
        "frontend_cache_used": False,
        "frontend_failure_category": "none",
        "frontend_price_currency_warning": "",
        "frontend_location_warning": "",
        "frontend_auto_conclusion": "FRONTEND_OK",
        "frontend_location_scope": "exact",
        "frontend_location_exact": True,
        "frontend_location_verified": True,
        "comparable_competitor_count": 3,
        "frontend_evidence_quality_score": 88,
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["product_final_decision_rows"] = [row]
    payload["product_final_decision_rows"] = [row]
    snapshot["final_decision_summary"] = {"WAIT_REVIEW": 1}
    snapshot["decision_gate_counts"] = {"WAIT_REVIEW": 1}
    payload["final_decision_summary"]["UK"] = {"WAIT_REVIEW": 1}
    payload["decision_gate_counts"]["UK"] = {"WAIT_REVIEW": 1}
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=payload["product_final_decision_rows"],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis UK product_final_decision_rows row 1 B0STRONGNOCOMP "
        "marks strong frontend evidence without competitor comparability" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_text_report_date_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    html = tmp_path / "latest_recommendations.html"
    html.write_text("<h1>亚马逊运营工作台｜2026-06-23</h1>", encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, html], previous_mtimes_ns={})

    assert f"text report missing report_date 2026-06-24: {html}" in failures


def test_report_refresh_failures_blocks_daily_excel_report_date_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx", report_date="2026-06-23")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any("daily Excel report_date mismatch: expected 2026-06-24" in failure for failure in failures)


def test_report_refresh_failures_blocks_autoopt_excel_report_date_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx", report_date="2026-06-23")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any("autoopt Excel report_date mismatch: expected 2026-06-24, got 2026-06-23" in failure for failure in failures)


def test_report_refresh_failures_blocks_autoopt_excel_json_row_count_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "action_review_rows": [{"action_id": "a1"}],
                "keyword_action_review_rows": [{"action_id": "k1"}, {"action_id": "k2"}],
                "product_final_decisions": [{"marketplace": "UK"}],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert "autoopt Excel action_review row count mismatch: expected 1 from action_review_rows, got 0" in failures
    assert (
        "autoopt Excel keyword_action_review row count mismatch: "
        "expected 2 from keyword_action_review_rows, got 0"
    ) in failures
    assert "autoopt Excel final_decisions row count mismatch: expected 1 from product_final_decisions, got 0" in failures


def test_report_refresh_failures_blocks_autoopt_product_decision_identity_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    expected_row = _product_decision_row()
    stale_row = dict(expected_row)
    stale_row["asin"] = "B0STALEDECISION"
    _attach_product_decision(payload, expected_row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[expected_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "product_final_decisions": [stale_row]}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"final_decisions": [stale_row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "autoopt log product_final_decisions identity mismatch vs "
        "latest analysis product_final_decision_rows for product final decisions" in failure
        and "B0DECISION" in failure
        and "B0STALEDECISION" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_autoopt_product_decision_field_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    expected_row = _product_decision_row(final_decision="WAIT_REVIEW", final_decision_label="等待复查")
    stale_row = dict(expected_row)
    stale_row["final_decision"] = "EXECUTE_TODAY"
    stale_row["final_decision_label"] = "今日执行"
    _attach_product_decision(payload, expected_row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[expected_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "product_final_decisions": [stale_row]}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"final_decisions": [stale_row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "autoopt log product_final_decisions field mismatch vs "
        "latest analysis product_final_decision_rows for product final decisions" in failure
        and "field final_decision: expected 'WAIT_REVIEW', got 'EXECUTE_TODAY'" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_autoopt_keyword_review_snapshot_row_count_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    expected_row = _review_row(
        action_id="k-analysis-review-count",
        search_term_or_target="analysis review count term",
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [expected_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[expected_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "action_review_rows": [],
                "keyword_action_review_rows": [],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "autoopt log keyword_action_review_rows row count mismatch vs "
        "latest analysis marketplace snapshots keyword_action_effect_review_rows for keyword action reviews: "
        "expected 1, got 0"
    ) in failures


def test_report_refresh_failures_blocks_autoopt_keyword_review_snapshot_field_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    expected_row = _review_row(
        action_id="k-analysis-review-field",
        search_term_or_target="analysis review field term",
        current_7d_tacos=0.06,
    )
    stale_row = dict(expected_row)
    stale_row["current_7d_tacos"] = 0.25
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [expected_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[expected_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "action_review_rows": [],
                "keyword_action_review_rows": [stale_row],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"keyword_action_review": [stale_row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "autoopt log keyword_action_review_rows field mismatch vs "
        "latest analysis marketplace snapshots keyword_action_effect_review_rows for keyword action reviews"
        in failure
        and "field current_7d_tacos: expected 0.06, got 0.25" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_autoopt_action_review_snapshot_row_count_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    expected_row = _review_row(
        "action_review",
        action_id="a-analysis-review-count",
        normalized_action="bid_down",
        action_scope="product",
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["action_effect_review_rows"] = [expected_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        action_review_rows=[expected_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "action_review_rows": [],
                "keyword_action_review_rows": [],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "autoopt log action_review_rows row count mismatch vs "
        "latest analysis marketplace snapshots action_effect_review_rows for action reviews: "
        "expected 1, got 0"
    ) in failures


def test_report_refresh_failures_blocks_autoopt_action_review_snapshot_field_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    expected_row = _review_row(
        "action_review",
        action_id="a-analysis-review-field",
        normalized_action="bid_down",
        action_scope="product",
        current_7d_tacos=0.06,
    )
    stale_row = dict(expected_row)
    stale_row["current_7d_tacos"] = 0.25
    payload["marketplace_results"][0]["report_view_snapshot"]["action_effect_review_rows"] = [expected_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        action_review_rows=[expected_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "action_review_rows": [stale_row],
                "keyword_action_review_rows": [],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"action_review": [stale_row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "autoopt log action_review_rows field mismatch vs "
        "latest analysis marketplace snapshots action_effect_review_rows for action reviews"
        in failure
        and "field current_7d_tacos: expected 0.06, got 0.25" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_autoopt_final_decision_summary_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    expected_row = _product_decision_row(final_decision="WAIT_REVIEW")
    _attach_product_decision(payload, expected_row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[expected_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "product_final_decisions": [expected_row],
                "final_decision_summary": {"EXECUTE_TODAY": 1},
                "decision_gate_counts": {"WAIT_REVIEW": 1},
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"final_decisions": [expected_row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "autoopt log final_decision_summary mismatch vs collapsed latest analysis final_decision_summary" in failure
        and "WAIT_REVIEW" in failure
        and "EXECUTE_TODAY" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_autoopt_decision_gate_counts_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    expected_row = _product_decision_row(final_decision="WAIT_REVIEW")
    _attach_product_decision(payload, expected_row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[expected_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "product_final_decisions": [expected_row],
                "final_decision_summary": {"WAIT_REVIEW": 1},
                "decision_gate_counts": {"EXECUTE_TODAY": 1},
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"final_decisions": [expected_row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "autoopt log decision_gate_counts mismatch vs collapsed latest analysis decision_gate_counts" in failure
        and "WAIT_REVIEW" in failure
        and "EXECUTE_TODAY" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_autoopt_log_effective_review_before_7_days(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    row = _review_row(
        action_id="k-effective-early",
        search_term_or_target="early effective term",
        review_outcome="effective",
        effectiveness_score=2,
        outcome="明确改善",
        days_since_execution=4,
        promoted_conversion_improved=True,
        halo_only_conversion=False,
        target_sku_not_converted=False,
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "keyword_action_review_rows": [row]}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"keyword_action_review": [row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert "autoopt log keyword action review row 1 early effective term marks effective before 7-day review window" in failures


def test_report_refresh_failures_blocks_autoopt_log_negative_display_before_3_days(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    row = _review_row(
        action_id="k-negative-too-early",
        search_term_or_target="too early negative term",
        review_outcome="not_ready",
        effectiveness_score="",
        outcome="暂未改善",
        days_since_execution=2,
        promoted_conversion_improved=False,
        halo_only_conversion=False,
        target_sku_not_converted=True,
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "keyword_action_review_rows": [row]}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"keyword_action_review": [row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "autoopt log keyword action review row 1 too early negative term "
        "shows negative display outcome before 3-day review window"
    ) in failures


def test_report_refresh_failures_blocks_autoopt_log_insufficient_sample_after_7_days_with_traffic(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    row = _review_row(
        action_id="k-stale-sample",
        search_term_or_target="stale sample term",
        review_outcome="insufficient_sample",
        effectiveness_score=0,
        outcome="样本不足",
        days_since_execution=8,
        current_7d_clicks=9,
        current_7d_spend=5.5,
        current_7d_promoted_ad_orders=0,
        promoted_conversion_improved=False,
        halo_only_conversion=False,
        target_sku_not_converted=True,
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "keyword_action_review_rows": [row]}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"keyword_action_review": [row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "autoopt log keyword action review row 1 stale sample term "
        "shows insufficient sample after 7-day sufficient traffic"
    ) in failures


def test_report_refresh_failures_blocks_autoopt_log_blank_review_marketplace(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    row = _review_row(
        action_id="k-autoopt-blank-marketplace",
        marketplace="",
        search_term_or_target="autoopt blank marketplace term",
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "keyword_action_review_rows": [row]}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"keyword_action_review": [row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "autoopt log keyword action review row 1 autoopt blank marketplace term missing marketplace value"
        in failures
    )


def test_report_refresh_failures_blocks_autoopt_log_blank_review_action_id(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    row = _review_row(
        action_id="",
        search_term_or_target="autoopt blank action id term",
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "keyword_action_review_rows": [row]}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"keyword_action_review": [row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "autoopt log keyword action review row 1 autoopt blank action id term missing action_id value"
        in failures
    )
    assert (
        "keyword_action_review_20260624.json row 1 autoopt blank action id term missing action_id value"
        in failures
    )


def test_report_refresh_failures_blocks_autoopt_log_blank_action_review_action_id(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    row = _review_row(
        "action_review",
        action_id="",
        product_name="Autoopt blank product action id",
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "action_review_rows": [row]}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"action_review": [row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "autoopt log action review row 1 Autoopt blank product action id missing action_id value"
        in failures
    )
    assert (
        "action_review_20260624.json row 1 Autoopt blank product action id missing action_id value"
        in failures
    )


def test_report_refresh_failures_blocks_autoopt_log_unsupported_review_marketplace(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    row = _review_row(
        action_id="k-autoopt-bad-marketplace",
        marketplace="FR",
        search_term_or_target="autoopt unsupported marketplace term",
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "keyword_action_review_rows": [row]}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"keyword_action_review": [row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "autoopt log keyword action review row 1 autoopt unsupported marketplace term contains unsupported marketplace FR"
        in failures
    )


def test_report_refresh_failures_blocks_autoopt_log_halo_only_effective_review(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    row = _review_row(
        action_id="k-halo-only",
        search_term_or_target="halo only term",
        review_outcome="effective",
        effectiveness_score=2,
        outcome="明确改善",
        days_since_execution=8,
        promoted_conversion_improved=False,
        halo_only_conversion=True,
        target_sku_not_converted=True,
        current_7d_promoted_ad_orders=0,
        current_7d_halo_ad_orders=2,
        current_14d_promoted_ad_orders=0,
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "keyword_action_review_rows": [row]}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"keyword_action_review": [row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert "autoopt log keyword action review row 1 halo only term marks halo-only conversion as effective" in failures
    assert (
        "autoopt log keyword action review row 1 halo only term effective outcome missing promoted SKU conversion"
        in failures
    )
    assert (
        "autoopt log keyword action review row 1 halo only term effective outcome missing positive promoted SKU orders"
        in failures
    )


def test_report_refresh_failures_blocks_latest_analysis_review_window_overstating_seven_day_readiness(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = _review_row(
        action_id="k-overstated-7d-window",
        search_term_or_target="formal overstated review window term",
        review_outcome="not_ready",
        outcome="待7天确认",
        days_since_execution=4,
        review_window="7d_check",
        review_status="待7天复盘",
        review_phase="not_ready",
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK keyword action review row 1 formal overstated review window term "
        "claims 7-day review window before 7 days"
    ) in failures


def test_report_refresh_failures_allows_not_ready_review_missing_tracked_metrics(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = _review_row(
        action_id="k-missing-review-metrics",
        search_term_or_target="missing review metrics term",
        review_outcome="not_ready",
        outcome="待7天确认",
        days_since_execution=4,
        review_window="3d_check",
        review_status="待7天复盘",
        review_phase="3d",
        current_7d_promoted_ad_orders="",
        current_7d_acos="",
        current_7d_tacos="",
        current_7d_total_orders="",
        current_7d_available_stock="",
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert not any("missing review metrics term reached review window but missing numeric" in failure for failure in failures)


def test_report_refresh_failures_blocks_day_three_review_without_early_window_evidence(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = _review_row(
        action_id="k-missing-day-three-qualifier",
        search_term_or_target="missing day three qualifier term",
        review_outcome="not_ready",
        outcome="待7天确认",
        days_since_execution=4,
        review_window="3d_check",
        review_status="待7天复盘",
        review_phase="day_3_check",
        effect_evidence="近7天点击 8，广告单 1，本 SKU 单 1",
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK keyword action review row 1 missing day three qualifier term "
        "3-day review evidence missing early-window qualifier"
    ) in failures


def test_report_refresh_failures_blocks_ineffective_review_missing_tracked_metrics(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = _review_row(
        action_id="k-ineffective-missing-review-metrics",
        search_term_or_target="ineffective missing review metrics term",
        review_outcome="ineffective",
        outcome="暂未改善",
        effectiveness_score=-2,
        days_since_execution=8,
        review_window="7天后复盘",
        review_status="可做7天复盘",
        review_phase="day_7_review",
        current_7d_promoted_ad_orders="",
        current_7d_acos="",
        current_7d_tacos="",
        current_7d_total_orders="",
        current_7d_available_stock="",
        current_14d_promoted_ad_orders="",
        current_14d_acos="",
        current_14d_tacos="",
        current_14d_total_orders="",
        current_14d_available_stock="",
        promoted_conversion_improved=False,
        halo_only_conversion=False,
        target_sku_not_converted=True,
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK keyword action review row 1 ineffective missing review metrics term "
        "reached review window but missing numeric current_7d_promoted_ad_orders"
    ) in failures
    assert (
        "latest analysis UK keyword action review row 1 ineffective missing review metrics term "
        "reached review window but missing numeric current_14d_promoted_ad_orders"
    ) in failures


def test_report_refresh_failures_blocks_effective_review_with_impossible_promoted_order_totals(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    row = _review_row(
        action_id="k-impossible-promoted-orders",
        search_term_or_target="impossible promoted orders",
        review_outcome="effective",
        effectiveness_score=2,
        outcome="明确改善",
        days_since_execution=8,
        promoted_conversion_improved=True,
        halo_only_conversion=False,
        target_sku_not_converted=False,
        current_7d_ad_orders=1,
        current_7d_promoted_ad_orders=2,
        current_7d_total_orders=1,
        current_14d_ad_orders=2,
        current_14d_promoted_ad_orders=3,
        current_14d_total_orders=2,
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "keyword_action_review_rows": [row]}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"keyword_action_review": [row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "autoopt log keyword action review row 1 impossible promoted orders "
        "effective outcome has promoted SKU orders above ad orders"
    ) in failures
    assert (
        "autoopt log keyword action review row 1 impossible promoted orders "
        "effective outcome has promoted SKU orders above total orders"
    ) in failures
    assert (
        "autoopt log keyword action review row 1 impossible promoted orders "
        "effective outcome has 14-day promoted SKU orders above ad orders"
    ) in failures
    assert (
        "autoopt log keyword action review row 1 impossible promoted orders "
        "effective outcome has 14-day promoted SKU orders above total orders"
    ) in failures


def test_report_refresh_failures_allows_not_ready_review_with_impossible_promoted_order_totals(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    row = _review_row(
        action_id="k-not-ready-impossible-promoted-orders",
        search_term_or_target="not ready impossible promoted orders",
        review_outcome="not_ready",
        effectiveness_score="",
        outcome="待7天确认",
        judgement="待7天确认",
        days_since_execution=8,
        promoted_conversion_improved=False,
        halo_only_conversion=False,
        target_sku_not_converted=False,
        current_7d_ad_orders=1,
        current_7d_promoted_ad_orders=2,
        current_7d_total_orders=1,
        current_14d_ad_orders=2,
        current_14d_promoted_ad_orders=3,
        current_14d_total_orders=2,
        review_window="7天后复盘",
        review_phase="7d",
        effect_evidence="7天复盘口径，等待人工确认。",
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "keyword_action_review_rows": [row]}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"keyword_action_review": [row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert not any("not ready impossible promoted orders review metrics have" in failure for failure in failures)


def test_report_refresh_failures_blocks_effective_review_missing_ad_order_denominator(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = _review_row(
        action_id="k-missing-ad-order-denominator",
        search_term_or_target="missing ad order denominator term",
        review_outcome="effective",
        effectiveness_score=2,
        outcome="明确改善",
        days_since_execution=8,
        promoted_conversion_improved=True,
        halo_only_conversion=False,
        target_sku_not_converted=False,
        current_7d_ad_orders="",
        current_14d_ad_orders="",
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK keyword action review row 1 missing ad order denominator term "
        "effective outcome missing numeric current_7d_ad_orders"
    ) in failures
    assert (
        "latest analysis UK keyword action review row 1 missing ad order denominator term "
        "effective outcome missing numeric current_14d_ad_orders"
    ) in failures


def test_report_refresh_failures_blocks_final_review_outcome_missing_click_or_spend(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    row = _review_row(
        action_id="k-final-missing-click-spend",
        search_term_or_target="final review missing traffic term",
        review_outcome="effective",
        outcome="有改善迹象",
        judgement="有改善迹象",
        effectiveness_score=2,
        days_since_execution=8,
        review_window="7天后复盘",
        review_status="可做7天复盘",
        review_phase="day_7_review",
        current_7d_clicks="nan",
        current_7d_spend="NaN",
        current_7d_promoted_ad_orders=2,
        current_7d_promoted_ad_sales=39.98,
        current_7d_ad_orders=2,
        current_7d_total_orders=3,
        current_7d_acos=0.08,
        current_7d_tacos=0.06,
        current_7d_available_stock=20,
        current_14d_promoted_ad_orders=3,
        current_14d_promoted_ad_sales=59.97,
        current_14d_ad_orders=3,
        current_14d_total_orders=4,
        current_14d_acos=0.08,
        current_14d_tacos=0.06,
        current_14d_available_stock=20,
        promoted_conversion_improved=True,
        halo_only_conversion=False,
        target_sku_not_converted=False,
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK keyword action review row 1 final review missing traffic term "
        "final review outcome missing numeric current_7d_clicks or current_7d_spend"
    ) in failures


def test_report_refresh_failures_blocks_latest_analysis_keyword_review_missing_contract_fields(tmp_path) -> None:
    payload = _analysis_payload("2026-06-24")
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0REVIEW1",
            "search_term_or_target": "desk lamp clip",
        }
    ]
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis UK keyword action review row 1 desk lamp clip missing fields:" in failure
        and "current_7d_promoted_ad_orders" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_latest_analysis_keyword_review_blank_marketplace(tmp_path) -> None:
    row = _review_row(
        action_id="k-blank-marketplace",
        marketplace="",
        search_term_or_target="blank marketplace term",
    )
    payload = _analysis_payload("2026-06-24")
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [row]
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK keyword action review row 1 blank marketplace term missing marketplace value"
        in failures
    )


def test_report_refresh_failures_blocks_latest_analysis_keyword_review_blank_action_id(tmp_path) -> None:
    row = _review_row(
        action_id="",
        search_term_or_target="blank action id term",
    )
    payload = _analysis_payload("2026-06-24")
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [row]
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK keyword action review row 1 blank action id term missing action_id value"
        in failures
    )


def test_report_refresh_failures_blocks_latest_analysis_action_review_blank_action_id(tmp_path) -> None:
    row = _review_row(
        "action_review",
        action_id="",
        product_name="Blank product action id",
    )
    payload = _analysis_payload("2026-06-24")
    payload["marketplace_results"][0]["report_view_snapshot"]["action_effect_review_rows"] = [row]
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK action review row 1 Blank product action id missing action_id value"
        in failures
    )


def test_report_refresh_failures_blocks_latest_analysis_keyword_review_halo_only_effective(tmp_path) -> None:
    row = _review_row(
        action_id="k1",
        search_term_or_target="desk lamp clip",
        review_outcome="effective",
        effectiveness_score=2,
        outcome="明确改善",
        days_since_execution=8,
        promoted_conversion_improved=False,
        halo_only_conversion=True,
        target_sku_not_converted=True,
        current_7d_promoted_ad_orders=0,
        current_7d_halo_ad_orders=2,
        current_14d_promoted_ad_orders=0,
    )
    payload = _analysis_payload("2026-06-24")
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [row]
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert "latest analysis UK keyword action review row 1 desk lamp clip marks halo-only conversion as effective" in failures
    assert "latest analysis UK keyword action review row 1 desk lamp clip marks target SKU not converted as effective" in failures


def test_report_refresh_failures_blocks_positive_rule_adjustment_without_promoted_sku(tmp_path) -> None:
    row = _review_row(
        action_id="policy-positive-without-promoted",
        search_term_or_target="policy positive term",
        review_outcome="not_ready",
        outcome="待7天确认",
        judgement="待7天确认",
        effectiveness_score="",
        days_since_execution=8,
        rule_adjustment="可保留当前竞价",
        promoted_conversion_improved=False,
        halo_only_conversion=False,
        target_sku_not_converted=True,
        current_7d_promoted_ad_orders=0,
        current_14d_promoted_ad_orders=0,
    )
    payload = _analysis_payload("2026-06-24")
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [row]
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "latest analysis UK keyword action review row 1 policy positive term "
        "effective outcome missing promoted SKU conversion"
        in failure
        for failure in failures
    )


def test_report_refresh_failures_allows_not_ready_positive_policy_before_7_days(tmp_path) -> None:
    row = _review_row(
        action_id="early-policy-positive",
        search_term_or_target="early policy term",
        review_outcome="not_ready",
        outcome="待7天确认",
        judgement="待7天确认",
        effectiveness_score="",
        days_since_execution=4,
        rule_adjustment="保留当前竞价，继续观察。",
        promoted_conversion_improved=True,
        halo_only_conversion=False,
        target_sku_not_converted=False,
        current_7d_promoted_ad_orders=1,
        current_14d_promoted_ad_orders=1,
        effect_evidence="3天复查口径，7天结论待补；已有订单，继续观察。",
    )
    payload = _analysis_payload("2026-06-24")
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [row]
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert not any("early policy term marks effective before 7-day review window" in failure for failure in failures)
    assert not any("early policy term effective outcome" in failure for failure in failures)


def test_report_refresh_failures_blocks_latest_analysis_negative_display_before_3_days(tmp_path) -> None:
    row = _review_row(
        action_id="k-negative-snapshot-too-early",
        search_term_or_target="too early snapshot negative term",
        review_outcome="not_ready",
        effectiveness_score="",
        outcome="暂未改善",
        days_since_execution=2,
        promoted_conversion_improved=False,
        halo_only_conversion=False,
        target_sku_not_converted=True,
    )
    payload = _analysis_payload("2026-06-24")
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [row]
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "latest analysis UK keyword action review row 1 too early snapshot negative term "
        "shows negative display outcome before 3-day review window"
    ) in failures


def test_report_refresh_failures_blocks_autoopt_excel_keyword_review_identity_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    expected_row = _review_row(action_id="k1", search_term_or_target="desk lamp clip")
    stale_row = dict(expected_row)
    stale_row["search_term_or_target"] = "old stale term"
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "keyword_action_review_rows": [expected_row],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"keyword_action_review": [stale_row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "autoopt Excel keyword_action_review identity mismatch for keyword action reviews" in failure
        and "desk lamp clip" in failure
        and "old stale term" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_autoopt_excel_keyword_review_promoted_sku_field_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    expected_row = _review_row(
        action_id="k1",
        search_term_or_target="desk lamp clip",
        current_7d_promoted_ad_orders=2,
        current_7d_acos=0.07,
        current_7d_tacos=0.05,
        current_7d_total_orders=3,
        current_7d_available_stock=18,
    )
    stale_row = dict(expected_row)
    stale_row["current_7d_promoted_ad_orders"] = 0
    stale_row["current_7d_acos"] = 0.3
    stale_row["current_7d_tacos"] = 0.25
    stale_row["current_7d_total_orders"] = 0
    stale_row["current_7d_available_stock"] = 0
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "keyword_action_review_rows": [expected_row],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"keyword_action_review": [stale_row]},
    )

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "autoopt Excel keyword_action_review field mismatch for keyword action reviews" in failure
        and "field current_7d_promoted_ad_orders: expected 2.0, got 0.0" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_autoopt_log_review_action_id_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    row = _review_row(
        action_id="UK||SKU-1||B0REVIEW1||search_term||wrong term||bid_down",
        search_term_or_target="desk lamp clip",
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "keyword_action_review_rows": [row],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "autoopt log keyword action review row 1 desk lamp clip action_id mismatch: "
        "expected UK||SKU-1||B0REVIEW1||search_term||desk lamp clip||bid_down, "
        "got UK||SKU-1||B0REVIEW1||search_term||wrong term||bid_down"
    ) in failures


def test_report_refresh_failures_blocks_standalone_review_json_row_count_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "action_review_rows": [{"action_id": "a1"}],
                "keyword_action_review_rows": [{"action_id": "k1"}, {"action_id": "k2"}],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
    (tmp_path / "keyword_action_review_20260624.json").write_text(json.dumps([]), encoding="utf-8")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert "keyword_action_review_20260624.json row count mismatch: expected 2 from keyword_action_review_rows, got 0" in failures


def test_report_refresh_failures_blocks_standalone_review_json_identity_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    expected_row = _review_row(action_id="k1", search_term_or_target="desk lamp clip")
    stale_row = dict(expected_row)
    stale_row["search_term_or_target"] = "old stale term"
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "keyword_action_review_rows": [expected_row],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
    (tmp_path / "keyword_action_review_20260624.json").write_text(json.dumps([stale_row]), encoding="utf-8")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "keyword_action_review_20260624.json identity mismatch" in failure
        and "desk lamp clip" in failure
        and "old stale term" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_standalone_review_json_action_scope_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    expected_row = _review_row(search_term_or_target="desk lamp clip")
    stale_row = dict(expected_row)
    stale_row["action_scope"] = "keyword"
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "keyword_action_review_rows": [expected_row],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
    (tmp_path / "keyword_action_review_20260624.json").write_text(json.dumps([stale_row]), encoding="utf-8")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "keyword_action_review_20260624.json row 1 desk lamp clip has invalid action_scope keyword; "
        "expected one of asin_target, campaign, product, search_term"
    ) in failures


def test_report_refresh_failures_blocks_standalone_review_json_promoted_sku_field_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    expected_row = _review_row(
        action_id="k1",
        search_term_or_target="desk lamp clip",
        current_7d_promoted_ad_orders=2,
        current_7d_total_orders=3,
        current_7d_available_stock=18,
    )
    stale_row = dict(expected_row)
    stale_row["current_7d_promoted_ad_orders"] = 0
    stale_row["current_7d_total_orders"] = 0
    stale_row["current_7d_available_stock"] = 0
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "keyword_action_review_rows": [expected_row],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
    (tmp_path / "keyword_action_review_20260624.json").write_text(json.dumps([stale_row]), encoding="utf-8")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "keyword_action_review_20260624.json field mismatch" in failure
        and "field current_7d_promoted_ad_orders: expected 2.0, got 0.0" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_standalone_review_json_missing_promoted_sku_contract(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "keyword_action_review_rows": [
                    {
                        "action_id": "k1",
                        "marketplace": "UK",
                        "search_term_or_target": "desk lamp clip",
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert any(
        "keyword_action_review_20260624.json row 1 desk lamp clip missing fields:" in failure
        and "current_7d_promoted_ad_orders" in failure
        and "current_7d_acos" in failure
        and "current_7d_tacos" in failure
        and "current_7d_available_stock" in failure
        and "current_14d_promoted_ad_orders" in failure
        and "current_14d_available_stock" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_executed_feedback_without_keyword_review_trace(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    feedback_row = {
        "marketplace": "UK",
        "sku": "SKU-FEEDBACK",
        "asin": "B0FEEDBACK",
        "product_name": "Feedback trace product",
        "search_term_or_target": "missing review keyword",
        "manual_action_taken": "加价 5%-10%",
        "confirmed_status": "已执行",
        "confirmed_at": "2026-06-24T09:00:00",
        "report_date": "2026-06-24",
    }
    (tmp_path / "autoopt_feedback_input.json").write_text(json.dumps([feedback_row]), encoding="utf-8")
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "keyword_action_review_rows": []}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    expected_action_id = "UK||SKU-FEEDBACK||B0FEEDBACK||search_term||missing review keyword||bid_up"
    assert (
        "executed feedback row 1 missing review keyword missing keyword action review trace "
        f"{expected_action_id}"
    ) in failures


def test_report_refresh_failures_normalizes_feedback_missing_action_id_before_trace_check(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    feedback_row = {
        "marketplace": "UK",
        "sku": "SKU-FEEDBACK-NO-ID",
        "asin": "B0FEEDNOID",
        "product_name": "Feedback missing id product",
        "search_term_or_target": "missing id keyword",
        "manual_action_taken": "加价 5%-10%",
        "confirmed_status": "已执行",
        "confirmed_at": "2026-06-24T09:00:00",
        "report_date": "2026-06-24",
    }
    (tmp_path / "autoopt_feedback_input.json").write_text(json.dumps([feedback_row]), encoding="utf-8")
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "keyword_action_review_rows": []}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "executed feedback row 1 missing id keyword missing keyword action review trace "
        "UK||SKU-FEEDBACK-NO-ID||B0FEEDNOID||search_term||missing id keyword||bid_up"
    ) in failures


def test_report_refresh_failures_normalizes_feedback_action_id_mismatch_before_trace_check(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    feedback_row = {
        "marketplace": "UK",
        "sku": "SKU-FEEDBACK-BAD-ID",
        "asin": "B0FEEDBADID",
        "product_name": "Feedback bad id product",
        "search_term_or_target": "actual feedback keyword",
        "manual_action_taken": "加价 5%-10%",
        "normalized_action": "bid_up",
        "action_scope": "search_term",
        "action_id": "UK||SKU-FEEDBACK-BAD-ID||B0FEEDBADID||search_term||wrong feedback keyword||bid_up",
        "confirmed_status": "已执行",
        "confirmed_at": "2026-06-24T09:00:00",
        "report_date": "2026-06-24",
    }
    (tmp_path / "autoopt_feedback_input.json").write_text(json.dumps([feedback_row]), encoding="utf-8")
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "keyword_action_review_rows": []}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "executed feedback row 1 actual feedback keyword missing keyword action review trace "
        "UK||SKU-FEEDBACK-BAD-ID||B0FEEDBADID||search_term||actual feedback keyword||bid_up"
    ) in failures


def test_report_refresh_failures_does_not_require_review_trace_for_executed_observation_feedback(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    feedback_row = {
        "marketplace": "UK",
        "sku": "SKU-OBSERVE",
        "asin": "B0OBSERVEFB",
        "product_name": "Observation feedback product",
        "search_term_or_target": "observe only keyword",
        "manual_action_taken": "观察",
        "confirmed_status": "已执行",
        "confirmed_at": "2026-06-24T09:00:00",
        "report_date": "2026-06-24",
    }
    (tmp_path / "autoopt_feedback_input.json").write_text(json.dumps([feedback_row]), encoding="utf-8")
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "keyword_action_review_rows": []}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert not any("B0OBSERVEFB" in failure and "missing keyword action review trace" in failure for failure in failures)


def test_report_refresh_failures_blocks_autoopt_executed_rows_without_review_trace(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    autoopt_row = {
        "action_id": "autoopt-product-bid-up",
        "normalized_action": "bid_up",
        "action_scope": "product",
        "marketplace": "UK",
        "sku": "SKU-AUTOOPT",
        "asin": "B0AUTOOPT",
        "product_name": "Autoopt executed product",
        "today_action": "加价 5%-10%",
        "confirmed_status": "已执行",
        "confirmed_at": "2026-06-24T09:00:00",
        "report_date": "2026-06-24",
    }
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "rows": [autoopt_row],
                "action_review_rows": [],
                "keyword_action_review_rows": [],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "autoopt log executed row 1 Autoopt executed product "
        "missing action review trace autoopt-product-bid-up"
    ) in failures


def test_report_refresh_failures_blocks_autoopt_executed_rows_missing_action_id(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    autoopt_row = {
        "normalized_action": "bid_up",
        "action_scope": "product",
        "marketplace": "UK",
        "sku": "SKU-NO-ID",
        "asin": "B0NOACTIONID",
        "product_name": "Missing action id product",
        "today_action": "加价 5%-10%",
        "confirmed_status": "已执行",
        "confirmed_at": "2026-06-24T09:00:00",
        "report_date": "2026-06-24",
    }
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "rows": [autoopt_row],
                "action_review_rows": [],
                "keyword_action_review_rows": [],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "autoopt log executed row 1 Missing action id product missing action_id for executed action trace"
        in failures
    )


def test_report_refresh_failures_does_not_require_action_id_for_autoopt_executed_observation(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    autoopt_row = {
        "normalized_action": "observe",
        "action_scope": "product",
        "marketplace": "UK",
        "sku": "SKU-OBS-NO-ID",
        "asin": "B0OBSNOID",
        "product_name": "Observation without action id",
        "today_action": "观察",
        "confirmed_status": "已执行",
        "confirmed_at": "2026-06-24T09:00:00",
        "report_date": "2026-06-24",
    }
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "rows": [autoopt_row],
                "action_review_rows": [],
                "keyword_action_review_rows": [],
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert not any("B0OBSNOID" in failure and "missing action_id" in failure for failure in failures)


def test_report_refresh_failures_blocks_standalone_keyword_review_json_halo_only_effective(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    row = _review_row(
        action_id="k-standalone-halo",
        search_term_or_target="standalone halo term",
        review_outcome="effective",
        outcome="明确改善",
        effectiveness_score=2,
        days_since_execution=8,
        promoted_conversion_improved=False,
        halo_only_conversion=True,
        target_sku_not_converted=True,
        current_7d_promoted_ad_orders=0,
        current_7d_halo_ad_orders=2,
        current_14d_promoted_ad_orders=0,
        current_14d_halo_ad_orders=3,
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "keyword_action_review_rows": [row]}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "keyword_action_review_20260624.json keyword action review row 1 standalone halo term "
        "marks halo-only conversion as effective"
    ) in failures
    assert (
        "keyword_action_review_20260624.json keyword action review row 1 standalone halo term "
        "effective outcome missing promoted SKU conversion"
    ) in failures


def test_report_refresh_failures_blocks_standalone_action_review_json_halo_only_effective(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    row = _review_row(
        review_type="action_review",
        action_id="a-standalone-halo",
        product_name="Standalone halo product",
        review_outcome="effective",
        outcome="明确改善",
        effectiveness_score=2,
        days_since_execution=8,
        promoted_conversion_improved=False,
        halo_only_conversion=True,
        target_sku_not_converted=True,
        current_7d_promoted_ad_orders=0,
        current_7d_halo_ad_orders=2,
        current_14d_promoted_ad_orders=0,
        current_14d_halo_ad_orders=3,
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "action_review_rows": [row]}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "action_review_20260624.json action review row 1 Standalone halo product "
        "marks halo-only conversion as effective"
    ) in failures
    assert (
        "action_review_20260624.json action review row 1 Standalone halo product "
        "effective outcome missing promoted SKU conversion"
    ) in failures


def test_report_refresh_failures_blocks_text_report_missing_required_marker(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    html = tmp_path / "latest_recommendations.html"
    html.write_text("2026-06-24\n产品级结论\n市场调查\n执行后效果复盘", encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, html], previous_mtimes_ns={})

    assert f"text report missing required marker today-ad-actions-all: {html}" in failures


def test_report_refresh_failures_blocks_text_report_missing_required_anchor(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    html = tmp_path / "latest_recommendations.html"
    html.write_text(
        "\n".join(
            [
                "2026-06-24",
                '<section id="product-operation-cards">产品级结论</section>',
                '<section id="today-ad-actions-all">today-ad-actions-all</section>',
                "市场调查",
                '<section id="action-effect-review">执行后效果复盘</section>',
            ]
        ),
        encoding="utf-8",
    )
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, html], previous_mtimes_ns={})

    assert f"text report missing required anchor #frontend-evidence-status: {html}" in failures


def test_report_refresh_failures_blocks_marketplace_report_missing_product_operation_anchor(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    html = tmp_path / "uk_report.html"
    html.write_text(
        "\n".join(
            [
                "2026-06-24",
                "亚马逊运营日报｜UK",
                "站点状态",
                "市场调查",
                "执行后效果复盘",
                '<section id="frontend-evidence-status"></section>',
                '<section id="action-effect-review"></section>',
            ]
        ),
        encoding="utf-8",
    )
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, html], previous_mtimes_ns={})

    assert f"text report missing required anchor #product-operation-cards: {html}" in failures


def test_report_refresh_failures_blocks_dashboard_marketplace_summary_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    for index, result in enumerate(payload["marketplace_results"], start=1):
        result["summary"].update(
            {
                "marketplace": result["marketplace"],
                "ads_row_count": 9 + index,
                "erp_row_count": 19 + index,
                "sku_count": 2 + index,
                "asin_count": 3 + index,
            }
        )
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    dashboard = tmp_path / "dashboard.html"
    dashboard.write_text(
        "\n".join(
            [
                "2026-06-24",
                "latest_recommendations.html",
                "summary.html",
                '<a href="uk_report.html">UK</a>',
                '<a href="us_report.html">US</a>',
                '<a href="de_report.html">DE</a>',
                "<tr><td>UK</td><td>999</td><td>20</td><td>3</td><td>4</td></tr>",
                "<tr><td>US</td><td>11</td><td>21</td><td>4</td><td>5</td></tr>",
                "<tr><td>DE</td><td>12</td><td>22</td><td>5</td><td>6</td></tr>",
            ]
        ),
        encoding="utf-8",
    )
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, dashboard], previous_mtimes_ns={})

    assert (
        "dashboard.html missing marketplace summary row for UK: ads=10 erp=20 sku=3 asin=4"
        in failures
    )


def test_report_refresh_failures_blocks_dashboard_missing_frontend_coverage_label(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-DASH-FRONTEND",
        "asin": "B0DASHFE",
        "product_name": "Dashboard frontend coverage product",
        "frontend_check_status": "沿用 2026-06-20 前台数据",
        "frontend_data_freshness": "沿用 2026-06-20 前台数据",
        "frontend_findings": "沿用 2026-06-20 前台数据",
        "frontend_evidence_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_cache_used": True,
        "frontend_auto_conclusion": "FRONTEND_WEAK",
        "frontend_search_status": "已读取部分结果",
        "frontend_search_partial_evidence": True,
        "frontend_evidence_quality_score": 58,
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["frontend_check_queue_rows"] = [frontend_row]
    snapshot["frontend_coverage_summary"] = daily_update._frontend_coverage_expected([frontend_row])
    payload["frontend_coverage_summary"] = daily_update._frontend_coverage_aggregate_from_snapshots(
        payload["marketplace_results"]
    )
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    dashboard = tmp_path / "dashboard.html"
    dashboard.write_text(
        "2026-06-24 latest_recommendations.html summary.html "
        "uk_report.html us_report.html de_report.html UK US DE",
        encoding="utf-8",
    )
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, dashboard], previous_mtimes_ns={})

    assert "dashboard.html missing frontend coverage token 前台证据覆盖" in failures
    assert "dashboard.html missing frontend coverage token 弱势止损证据" in failures


def test_report_refresh_failures_blocks_marketplace_summary_markdown_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    for index, result in enumerate(payload["marketplace_results"], start=1):
        result["summary"].update(
            {
                "marketplace": result["marketplace"],
                "ads_row_count": 9 + index,
                "erp_row_count": 19 + index,
                "sku_count": 2 + index,
                "asin_count": 3 + index,
            }
        )
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    marketplace_summary = tmp_path / "marketplace_summary.md"
    marketplace_summary.write_text(
        "\n".join(
            [
                "# 站点汇总｜2026-06-24",
                "",
                "| 站点 | 广告行数 | ERP行数 | SKU数 | ASIN数 | 状态 | 说明 |",
                "| --- | ---: | ---: | ---: | ---: | --- | --- |",
                "| UK | 999 | 20 | 3 | 4 | 正式分析 | 无 |",
                "| US | 11 | 21 | 4 | 5 | 正式分析 | 无 |",
                "| DE | 12 | 22 | 5 | 6 | 正式分析 | 无 |",
            ]
        ),
        encoding="utf-8",
    )
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, marketplace_summary], previous_mtimes_ns={})

    assert (
        "marketplace_summary.md missing marketplace summary row for UK: ads=10 erp=20 sku=3 asin=4"
        in failures
    )


def test_report_refresh_failures_blocks_marketplace_summary_missing_frontend_coverage_label(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-MD-FRONTEND",
        "asin": "B0MDFE",
        "product_name": "Markdown frontend coverage product",
        "frontend_check_status": "沿用 2026-06-20 前台数据",
        "frontend_data_freshness": "沿用 2026-06-20 前台数据",
        "frontend_findings": "沿用 2026-06-20 前台数据",
        "frontend_evidence_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_cache_used": True,
        "frontend_auto_conclusion": "FRONTEND_WEAK",
        "frontend_search_status": "已读取部分结果",
        "frontend_search_partial_evidence": True,
        "frontend_evidence_quality_score": 58,
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["frontend_check_queue_rows"] = [frontend_row]
    snapshot["frontend_coverage_summary"] = daily_update._frontend_coverage_expected([frontend_row])
    payload["frontend_coverage_summary"] = daily_update._frontend_coverage_aggregate_from_snapshots(
        payload["marketplace_results"]
    )
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    marketplace_summary = tmp_path / "marketplace_summary.md"
    marketplace_summary.write_text("# 站点汇总｜2026-06-24\n\n无前台覆盖表\n", encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, marketplace_summary], previous_mtimes_ns={})

    assert "marketplace_summary.md missing frontend coverage token 前台证据覆盖" in failures
    assert "marketplace_summary.md missing frontend coverage token 弱势止损证据" in failures


def test_report_refresh_failures_blocks_daily_excel_missing_frontend_coverage_label(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-XLSX-FRONTEND",
        "asin": "B0XLSXFE",
        "product_name": "Excel frontend coverage product",
        "frontend_check_status": "沿用 2026-06-20 前台数据",
        "frontend_data_freshness": "沿用 2026-06-20 前台数据",
        "frontend_findings": "沿用 2026-06-20 前台数据",
        "frontend_evidence_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_cache_used": True,
        "frontend_auto_conclusion": "FRONTEND_WEAK",
        "frontend_search_status": "已读取部分结果",
        "frontend_search_partial_evidence": True,
        "frontend_evidence_quality_score": 58,
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["frontend_check_queue_rows"] = [frontend_row]
    snapshot["frontend_coverage_summary"] = daily_update._frontend_coverage_expected([frontend_row])
    payload["frontend_coverage_summary"] = daily_update._frontend_coverage_aggregate_from_snapshots(
        payload["marketplace_results"]
    )
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis], previous_mtimes_ns={})

    assert (
        "daily Excel 总览 missing frontend coverage metric ALL 产品页成功: "
        "expected 0/1, got missing"
    ) in failures


def test_report_refresh_failures_blocks_summary_missing_keyword_review_identity(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(search_term_or_target="summary formal missing keyword")
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    summary = tmp_path / "summary.html"
    summary.write_text(
        "2026-06-24 三分钟摘要 今日开工结论 今天先做 昨天动作今天要盯",
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, summary], previous_mtimes_ns={})

    assert "summary.html missing review watch row 1 target summary formal missing keyword" in failures


def test_report_refresh_failures_blocks_summary_missing_frontend_coverage_label(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-SUMMARY-FRONTEND",
        "asin": "B0SUMMARYFE",
        "product_name": "Summary frontend coverage product",
        "frontend_check_status": "沿用 2026-06-20 前台数据",
        "frontend_data_freshness": "沿用 2026-06-20 前台数据",
        "frontend_findings": "沿用 2026-06-20 前台数据",
        "frontend_evidence_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_cache_used": True,
        "frontend_auto_conclusion": "FRONTEND_WEAK",
        "frontend_search_status": "已读取部分结果",
        "frontend_search_partial_evidence": True,
        "frontend_evidence_quality_score": 58,
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["frontend_check_queue_rows"] = [frontend_row]
    snapshot["frontend_coverage_summary"] = daily_update._frontend_coverage_expected([frontend_row])
    payload["frontend_coverage_summary"] = daily_update._frontend_coverage_aggregate_from_snapshots(
        payload["marketplace_results"]
    )
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    summary = tmp_path / "summary.html"
    summary.write_text(
        "2026-06-24 三分钟摘要 今日开工结论 今天先做 昨天动作今天要盯",
        encoding="utf-8",
    )
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, summary], previous_mtimes_ns={})

    assert "summary.html missing frontend coverage token 产品页成功" in failures
    assert "summary.html missing frontend coverage token 0/1" in failures


def test_report_refresh_failures_blocks_summary_missing_action_review_identity(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(
        "action_review",
        product_name="Summary formal missing action product",
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    summary = tmp_path / "summary.html"
    summary.write_text(
        "2026-06-24 三分钟摘要 今日开工结论 今天先做 昨天动作今天要盯",
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        action_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, summary], previous_mtimes_ns={})

    assert "summary.html missing review watch row 1 target Summary formal missing action product" in failures


def test_report_refresh_failures_blocks_summary_missing_prioritized_not_validated_review(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    ordinary_row = _review_row(
        search_term_or_target="ordinary sample summary term",
        judgement="样本不足",
        review_outcome="insufficient_sample",
        days_since_execution=14,
    )
    not_validated_row = _review_row(
        search_term_or_target="halo only summary priority term",
        judgement="本 SKU 未验证",
        review_outcome="needs_manual_review",
        next_step="有光环订单也不能证明本 SKU 有效，今天不追加预算或竞价。",
        days_since_execution=12,
        current_7d_promoted_ad_orders=0,
        current_7d_halo_ad_orders=1,
        current_7d_total_orders=1,
        promoted_conversion_improved=False,
        halo_only_conversion=True,
        target_sku_not_converted=True,
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [
        ordinary_row,
        not_validated_row,
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    summary = tmp_path / "summary.html"
    summary.write_text(
        "2026-06-24 三分钟摘要 今日开工结论 今天先做 昨天动作今天要盯 ordinary sample summary term",
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[ordinary_row, not_validated_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, summary], previous_mtimes_ns={})

    assert "summary.html missing review watch row 1 target halo only summary priority term" in failures


def test_report_refresh_failures_blocks_summary_missing_not_validated_review_text(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(
        search_term_or_target="halo only summary text term",
        judgement="初步有效",
        review_outcome="needs_manual_review",
        next_step="继续加价观察",
        days_since_execution=12,
        current_7d_promoted_ad_orders=0,
        current_7d_halo_ad_orders=1,
        current_7d_total_orders=1,
        promoted_conversion_improved=False,
        halo_only_conversion=True,
        target_sku_not_converted=True,
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    summary = tmp_path / "summary.html"
    summary.write_text(
        "2026-06-24 三分钟摘要 今日开工结论 今天先做 昨天动作今天要盯 "
        '<div class="summary-item"><strong>halo only summary text term</strong>'
        "<div>总单 1</div></div>",
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, summary], previous_mtimes_ns={})

    assert "summary.html missing review watch row 1 judgement 本 SKU 未验证" in failures
    assert (
        "summary.html missing review watch row 1 next step 仅光环成交，不算本 SKU 有效；今天不追加预算或竞价。"
        in failures
    )


def test_report_refresh_failures_blocks_summary_positive_review_without_promoted_conversion(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(
        search_term_or_target="no promoted summary text term",
        judgement="初步有效",
        review_outcome="effective",
        next_step="继续加价观察",
        days_since_execution=12,
        current_7d_promoted_ad_orders=0,
        current_7d_halo_ad_orders=0,
        current_7d_total_orders=1,
        promoted_conversion_improved=False,
        halo_only_conversion=False,
        target_sku_not_converted=False,
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    summary = tmp_path / "summary.html"
    summary.write_text(
        "2026-06-24 三分钟摘要 今日开工结论 今天先做 昨天动作今天要盯 "
        '<div class="summary-item"><strong>no promoted summary text term</strong>'
        "<div>本 SKU 单 0｜总单 1</div></div>",
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, summary], previous_mtimes_ns={})

    assert "summary.html missing review watch row 1 judgement 本 SKU 未验证" in failures
    assert (
        "summary.html missing review watch row 1 next step 缺少本 SKU 转化证据；今天不追加预算或竞价。"
        in failures
    )


def test_summary_review_sort_prioritizes_positive_review_without_promoted_conversion() -> None:
    no_promoted_row = _review_row(
        search_term_or_target="summary sort no promoted term",
        judgement="初步有效",
        next_step="继续加价观察",
        days_since_execution=12,
        promoted_conversion_improved=False,
        halo_only_conversion=False,
        target_sku_not_converted=False,
    )
    promoted_row = _review_row(
        search_term_or_target="summary sort promoted term",
        judgement="初步有效",
        next_step="继续加价观察",
        days_since_execution=12,
        promoted_conversion_improved=True,
        halo_only_conversion=False,
        target_sku_not_converted=False,
    )
    wait_row = _review_row(
        search_term_or_target="summary sort wait term",
        judgement="待7天确认",
        next_step="继续观察",
        days_since_execution=12,
        promoted_conversion_improved=False,
        halo_only_conversion=False,
        target_sku_not_converted=False,
    )

    rows = sorted([promoted_row, wait_row, no_promoted_row], key=daily_update._summary_review_watch_sort_key)

    assert rows[0]["search_term_or_target"] == "summary sort no promoted term"
    assert daily_update._summary_review_display_judgement(rows[0]) == "本 SKU 未验证"


def test_summary_review_metric_tokens_skip_14d_before_7_day_window() -> None:
    row = _review_row(
        days_since_execution=4,
        current_7d_promoted_ad_orders=1,
        current_7d_acos=0.08,
        current_7d_tacos=0.06,
        current_7d_total_orders=2,
        current_7d_available_stock=20,
        current_14d_promoted_ad_orders=3,
        current_14d_acos=0.09,
        current_14d_tacos=0.07,
    )

    tokens = daily_update._summary_review_metric_tokens(row)

    assert "本 SKU 单 1" in tokens
    assert "14天本 SKU 单 3" not in tokens
    assert "14天 TACOS 7%" not in tokens


def test_report_refresh_failures_blocks_summary_early_review_metric_token(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(
        search_term_or_target="summary too early metric keyword",
        review_window="未满3天",
        judgement="样本不足",
        days_since_execution=2,
        current_7d_promoted_ad_orders=1,
        current_7d_acos=0.08,
        current_7d_tacos=0.06,
        current_7d_total_orders=2,
        current_7d_available_stock=20,
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    summary = tmp_path / "summary.html"
    summary.write_text(
        "2026-06-24 三分钟摘要 今日开工结论 今天先做 昨天动作今天要盯 "
        "summary too early metric keyword 未满3天 本 SKU 单 1 ACOS 8%",
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, summary], previous_mtimes_ns={})

    assert "summary.html early review row 1 should not show metric token 本 SKU 单 1" in failures


def test_report_refresh_failures_blocks_summary_missing_review_metric_token(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(
        search_term_or_target="summary metric keyword",
        current_7d_promoted_ad_orders=1,
        current_7d_acos=0.08,
        current_7d_tacos=0.06,
        current_7d_total_orders=2,
        current_7d_available_stock=20,
        current_14d_promoted_ad_orders=2,
        days_since_execution=7,
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    summary = tmp_path / "summary.html"
    summary.write_text(
        "2026-06-24 三分钟摘要 今日开工结论 今天先做 昨天动作今天要盯 summary metric keyword 本 SKU 单 1",
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, summary], previous_mtimes_ns={})

    assert "summary.html missing review watch row 1 metric token ACOS 8%" in failures


def test_report_refresh_failures_blocks_summary_missing_review_14d_promoted_token(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(
        search_term_or_target="summary 14d metric keyword",
        current_7d_promoted_ad_orders=1,
        current_7d_acos=0.08,
        current_7d_tacos=0.06,
        current_7d_total_orders=2,
        current_7d_available_stock=20,
        current_14d_promoted_ad_orders=3,
        days_since_execution=7,
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    summary = tmp_path / "summary.html"
    summary.write_text(
        "2026-06-24 三分钟摘要 今日开工结论 今天先做 昨天动作今天要盯 "
        "summary 14d metric keyword 本 SKU 单 1 ACOS 8% TACOS 6% 总单 2 库存 20",
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, summary], previous_mtimes_ns={})

    assert "summary.html missing review watch row 1 metric token 14天本 SKU 单 3" in failures


def test_report_refresh_failures_blocks_summary_missing_review_14d_cost_token(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(
        search_term_or_target="summary 14d cost keyword",
        current_7d_promoted_ad_orders=1,
        current_7d_acos=0.08,
        current_7d_tacos=0.06,
        current_7d_total_orders=2,
        current_7d_available_stock=20,
        current_14d_promoted_ad_orders=3,
        current_14d_acos=0.09,
        current_14d_tacos=0.07,
        current_14d_total_orders=5,
        current_14d_available_stock=18,
        days_since_execution=7,
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    summary = tmp_path / "summary.html"
    summary.write_text(
        "2026-06-24 三分钟摘要 今日开工结论 今天先做 昨天动作今天要盯 "
        "summary 14d cost keyword 本 SKU 单 1 ACOS 8% TACOS 6% 总单 2 库存 20 "
        "14天本 SKU 单 3 14天 ACOS 9%",
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, summary], previous_mtimes_ns={})

    assert "summary.html missing review watch row 1 metric token 14天 TACOS 7%" in failures


def test_report_refresh_failures_blocks_summary_missing_replenishment_identity(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    inventory_row = {
        "marketplace": "UK",
        "sku": "SKU-SUMMARY-INV",
        "asin": "B0SUMMARYINV",
        "product_name": "Formal summary stockout product",
        "stock_risk_level": "OUT_OF_STOCK",
        "stock_status_label": "断货",
        "available_stock": 0,
        "days_of_cover": 0,
        "recommended_reorder_qty": 120,
        "replenishment_advice": "先核对在途库存，今天处理补货。",
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["inventory_replenishment_rows"] = [inventory_row]
    payload["inventory_replenishment_rows"] = [inventory_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    summary = tmp_path / "summary.html"
    summary.write_text(
        "2026-06-24 三分钟摘要 今日开工结论 今天先做 昨天动作今天要盯",
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        inventory_rows=[inventory_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, summary], previous_mtimes_ns={})

    assert "summary.html missing replenishment row 1 product name Formal summary stockout product" in failures
    assert "summary.html missing replenishment row 1 status 断货" in failures


def test_report_refresh_failures_blocks_latest_recommendations_missing_product_identity(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        asin="B0FORMALDECISION",
        product_name="Formal missing decision product",
    )
    _attach_product_decision(payload, product_row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24"),
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[product_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-24",
                "product_final_decisions": [product_row],
                "final_decision_summary": {"WAIT_REVIEW": 1},
                "decision_gate_counts": {"WAIT_REVIEW": 1},
            }
        ),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"final_decisions": [product_row]},
    )

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert "latest_recommendations.html missing product decision row 1 ASIN B0FORMALDECISION" in failures
    assert (
        "latest_recommendations.html missing product decision row 1 product name Formal missing decision product"
        in failures
    )


def test_report_refresh_failures_blocks_latest_recommendations_missing_product_blocking_reason(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        asin="B0FORMALGATE1",
        product_name="Formal gate product",
        frontend_blocking_reasons=["地区待确认，当前前台证据不能用于放量"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale"],
    )
    _attach_product_decision(payload, product_row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24")
        + "\nB0FORMALGATE1 Formal gate product 拦截加竞价 拦截加预算 拦截放量",
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[product_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert (
        "latest_recommendations.html missing product gate row 1 blocking reason "
        "地区待确认，当前前台证据不能用于放量"
    ) in failures


def test_report_refresh_failures_blocks_latest_recommendations_missing_product_blocked_action_label(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        asin="B0FORMALGATE2",
        product_name="Formal blocked action product",
        frontend_blocking_reasons=["竞品可比性不足"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale"],
    )
    _attach_product_decision(payload, product_row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24")
        + "\nB0FORMALGATE2 Formal blocked action product 竞品可比性不足",
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[product_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert "latest_recommendations.html missing product decision row 1 blocked action label bid_up" in failures


def test_report_refresh_failures_blocks_latest_recommendations_missing_low_budget_blocked_action_label(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        asin="B0FORMALLOWBUDGET",
        product_name="Formal low budget blocked product",
        frontend_blocking_reasons=["前台证据不足，需刷新后再放量"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
    )
    _attach_product_decision(payload, product_row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24")
        + "\nB0FORMALLOWBUDGET Formal low budget blocked product 前台证据不足，需刷新后再放量 拦截加竞价 拦截加预算 拦截放量",
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[product_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert (
        "latest_recommendations.html missing product decision row 1 blocked action label create_exact_low_budget"
    ) in failures


def test_report_refresh_failures_blocks_latest_recommendations_product_decision_data_attr_mismatch(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        asin="B0FORMALATTR",
        product_name="Formal attr product",
        frontend_evidence_tier="仅背景参考",
    )
    _attach_product_decision(payload, product_row)
    payload["marketplace_results"][0]["report_view_snapshot"]["product_operation_cards"] = [
        _product_operation_card(product_row)
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    visible_tokens = (
        "B0FORMALATTR Formal attr product 待前台检查 仅背景参考 "
        "拦截加竞价 拦截加预算 拦截放量 拦截低预算精准测试"
    )
    mismatched_attrs = _product_decision_contract_attrs(
        product_row,
        **{"data-product-decision-final": "EXECUTE_TODAY"},
    )
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24")
        + (
            f'\n<tr data-product-decision-marketplace="UK" data-product-decision-sku="SKU-DECISION" '
            f'data-product-decision-asin="B0FORMALATTR"{mismatched_attrs}><td>{visible_tokens}</td></tr>'
        )
        + (
            f'\n<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-DECISION" '
            f'data-product-decision-asin="B0FORMALATTR"{mismatched_attrs}><p>{visible_tokens}</p></article>'
        ),
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[product_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert any(
        "latest_recommendations.html product decision row 1 B0FORMALATTR "
        "data attr data-product-decision-final mismatch for field final_decision" in failure
        for failure in failures
    )


def test_report_refresh_failures_blocks_latest_recommendations_product_tokens_bound_to_wrong_structured_row(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        asin="B0FORMALBOUND",
        product_name="Formal bound product",
        frontend_blocking_reasons=["地区待确认，当前前台证据不能用于放量"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale"],
    )
    _attach_product_decision(payload, product_row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    attrs = _product_decision_contract_attrs(product_row)
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24")
        + (
            f'\n<tr data-product-decision-marketplace="UK" data-product-decision-sku="SKU-DECISION" '
            f'data-product-decision-asin="B0FORMALBOUND"{attrs}><td>B0FORMALBOUND Formal bound product</td></tr>'
        )
        + '\n<tr data-product-decision-marketplace="US" data-product-decision-sku="SKU-US" data-product-decision-asin="B0FORMALBOUND"><td>地区待确认，当前前台证据不能用于放量 禁止加竞价 禁止加预算 禁止放量</td></tr>'
        + (
            f'\n<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-DECISION" '
            f'data-product-decision-asin="B0FORMALBOUND"{attrs}><p>B0FORMALBOUND Formal bound product</p></article>'
        )
        + '\n<article data-product-decision-marketplace="US" data-product-decision-sku="SKU-US" data-product-decision-asin="B0FORMALBOUND"><p>地区待确认，当前前台证据不能用于放量 禁止加竞价 禁止加预算 禁止放量</p></article>',
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[product_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert (
        "latest_recommendations.html product gate row 1 blocking reason not bound to ASIN B0FORMALBOUND"
        in failures
    )
    assert (
        "latest_recommendations.html product gate row 1 blocked action label bid_up not bound to ASIN B0FORMALBOUND"
        in failures
    )


def test_report_refresh_failures_blocks_latest_recommendations_product_gate_tokens_bound_to_wrong_article(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        asin="B0FORMALGATE",
        product_name="Formal gate product",
        frontend_blocking_reasons=["竞品可比性不足，不能用于强动作"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale"],
    )
    _attach_product_decision(payload, product_row)
    payload["marketplace_results"][0]["report_view_snapshot"]["product_operation_cards"] = [
        _product_operation_card(product_row)
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24")
        + '\n<tr data-product-decision-marketplace="UK" data-product-decision-sku="SKU-DECISION" data-product-decision-asin="B0FORMALGATE"><td>B0FORMALGATE Formal gate product 竞品可比性不足，不能用于强动作 禁止加竞价 禁止加预算 禁止放量</td></tr>'
        + '\n<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-DECISION" data-product-decision-asin="B0FORMALGATE"><p>B0FORMALGATE Formal gate product</p></article>'
        + '\n<article data-product-decision-marketplace="US" data-product-decision-sku="SKU-US" data-product-decision-asin="B0FORMALGATE"><p>竞品可比性不足，不能用于强动作 禁止加竞价 禁止加预算 禁止放量</p></article>',
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[product_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert (
        "latest_recommendations.html product gate row 1 blocking reason not bound to ASIN B0FORMALGATE"
        in failures
    )
    assert (
        "latest_recommendations.html product gate row 1 blocked action label bid_up not bound to ASIN B0FORMALGATE"
        in failures
    )


def test_report_refresh_failures_blocks_latest_recommendations_frontend_tokens_bound_to_wrong_article(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        asin="B0FRONTGATE",
        product_name="Formal frontend evidence product",
        frontend_blocking_reasons=["前台缓存或读取失败待确认，仅背景参考，不能单独支持放量"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
        frontend_check_status="沿用 2026-06-20 前台数据",
        frontend_status="沿用 2026-06-20 前台数据",
        frontend_cache_used=True,
        frontend_evidence_tier="仅背景参考",
        frontend_evidence_display_tier="仅背景参考",
        frontend_evidence_audit_detail="沿用缓存；搜索页仅部分读取",
        frontend_search_status="已读取部分结果",
        frontend_search_partial_evidence=True,
    )
    _attach_product_decision(payload, product_row)
    payload["marketplace_results"][0]["report_view_snapshot"]["product_operation_cards"] = [
        _product_operation_card(product_row)
    ]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24")
        + '\n<tr data-product-decision-marketplace="UK" data-product-decision-sku="SKU-DECISION" data-product-decision-asin="B0FRONTGATE"><td>B0FRONTGATE Formal frontend evidence product 前台缓存或读取失败待确认，仅背景参考，不能单独支持放量 禁止加竞价 禁止加预算 禁止放量 禁止低预算精准测试 沿用 2026-06-20 前台数据 仅背景参考 搜索页：已读取部分结果</td></tr>'
        + '\n<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-DECISION" data-product-decision-asin="B0FRONTGATE"><p>B0FRONTGATE Formal frontend evidence product 前台缓存或读取失败待确认，仅背景参考，不能单独支持放量 禁止加竞价 禁止加预算 禁止放量 禁止低预算精准测试</p></article>'
        + '\n<article data-product-decision-marketplace="US" data-product-decision-sku="SKU-US" data-product-decision-asin="B0FRONTGATE"><p>沿用 2026-06-20 前台数据 仅背景参考 沿用缓存；搜索页仅部分读取 搜索页：已读取部分结果</p></article>',
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[product_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert (
        "latest_recommendations.html product gate row 1 frontend status "
        "沿用 2026-06-20 前台数据 not bound to ASIN B0FRONTGATE"
    ) in failures
    assert (
        "latest_recommendations.html product gate row 1 frontend search partial marker "
        "已读取部分结果 not bound to ASIN B0FRONTGATE"
    ) in failures
    assert (
        "latest_recommendations.html product gate row 1 frontend audit detail "
        "沿用缓存；搜索页仅部分读取 not bound to ASIN B0FRONTGATE"
    ) in failures


def test_report_refresh_failures_blocks_latest_recommendations_missing_keyword_review_identity(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(search_term_or_target="formal all missing keyword")
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24"),
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert "latest_recommendations.html missing keyword review row 1 target formal all missing keyword" in failures
    assert "latest_recommendations.html missing keyword review row 1 ASIN B0REVIEW1" in failures


def test_report_refresh_failures_blocks_latest_recommendations_missing_action_review_identity(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(
        "action_review",
        asin="B0ACTIONALL",
        product_name="Formal all missing action product",
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24"),
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        action_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert "latest_recommendations.html missing action review row 1 ASIN B0ACTIONALL" in failures
    assert (
        "latest_recommendations.html missing action review row 1 product name Formal all missing action product"
        in failures
    )


def test_report_refresh_failures_blocks_latest_recommendations_missing_review_metric_token(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(
        "keyword_action_review",
        search_term_or_target="formal metric missing keyword",
        asin="B0METRICHTML",
        current_7d_promoted_ad_orders=1,
        current_7d_acos=0.08,
        current_7d_tacos=0.06,
        current_7d_total_orders=2,
        current_7d_available_stock=20,
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24")
        + "\nformal metric missing keyword B0METRICHTML",
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert "latest_recommendations.html missing keyword review row 1 metric token 本 SKU 单 1" in failures


def test_report_refresh_failures_blocks_latest_recommendations_missing_early_review_policy_token(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(
        "keyword_action_review",
        search_term_or_target="formal early policy keyword",
        asin="B0EARLYHTML",
        current_7d_promoted_ad_orders=0,
        current_7d_acos=0.08,
        current_7d_target_acos=0.2,
        current_7d_tacos=0.06,
        current_7d_total_orders=0,
        current_7d_available_stock=20,
        days_since_execution=4,
        review_window="3d_check",
        review_phase="3d",
        effect_evidence="3天复查口径，7天结论待补；继续观察。",
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24")
        + "\nformal early policy keyword B0EARLYHTML"
        + "\n本 SKU 单 0 ACOS 0.08 目标 ACOS 0.2 TACOS 0.06 总单 0 库存 20",
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert (
        "latest_recommendations.html missing keyword review row 1 metric token 3 天窗口只做初步判断"
        in failures
    )


def test_report_refresh_failures_blocks_latest_recommendations_early_review_metric_token(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(
        "keyword_action_review",
        action_id="UK||SKU-1||B0TOOEARLYHTML||search_term||formal too early metric keyword||bid_down",
        search_term_or_target="formal too early metric keyword",
        asin="B0TOOEARLYHTML",
        review_window="未满3天",
        judgement="样本不足",
        current_7d_promoted_ad_orders=1,
        current_7d_acos=0.08,
        current_7d_target_acos=0.2,
        current_7d_tacos=0.06,
        current_7d_total_orders=2,
        current_7d_available_stock=20,
        days_since_execution=2,
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24")
        + "\nformal too early metric keyword B0TOOEARLYHTML "
        "UK||SKU-1||B0TOOEARLYHTML||search_term||formal too early metric keyword||bid_down "
        "7天 本 SKU 单 1 7天 ACOS 0.08",
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(
        json.dumps({"report_date": "2026-06-24", "keyword_action_review_rows": [review_row]}),
        encoding="utf-8",
    )
    _write_autoopt_workbook(
        tmp_path / "autoopt_20260624.xlsx",
        sheet_rows={"keyword_action_review": [review_row]},
    )

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert (
        "latest_recommendations.html keyword review row 1 should not show metric token 7天 本 SKU 单 1"
        in failures
    )


def test_review_html_metric_tokens_skip_anchor_7d_before_7_day_window() -> None:
    row = {
        "review_data_source": "execution_anchored_daily",
        "days_since_execution": 4,
        "pre_7d_start": "2026-06-11",
        "pre_7d_end": "2026-06-17",
        "post_3d_start": "2026-06-18",
        "post_3d_end": "2026-06-20",
        "post_7d_start": "2026-06-18",
        "post_7d_end": "2026-06-24",
        "pre_7d_promoted_ad_orders": 0,
        "pre_7d_total_orders": 1,
        "pre_7d_tacos": 0.12,
        "post_3d_days": 3,
        "post_3d_promoted_ad_orders": 1,
        "post_3d_total_orders": 2,
        "post_3d_acos": 0.10,
        "post_3d_tacos": 0.07,
        "post_3d_available_stock": 18,
        "post_7d_days": 7,
        "post_7d_promoted_ad_orders": 2,
        "post_7d_total_orders": 4,
        "post_7d_acos": 0.12,
        "post_7d_tacos": 0.08,
        "post_7d_available_stock": 18,
    }

    required_tokens = daily_update._review_html_metric_tokens(row)
    forbidden_tokens = daily_update._review_html_forbidden_metric_tokens(row)

    assert "执行后3天本 SKU 单 1" in required_tokens
    assert "执行后3天 ACOS 0.1" in required_tokens
    assert "执行后本 SKU 单 2" not in required_tokens
    assert "执行后7天 2026-06-18 至 2026-06-24" not in required_tokens
    assert "执行后本 SKU 单 2" in forbidden_tokens
    assert "执行后7天 2026-06-18 至 2026-06-24" in forbidden_tokens
    assert "执行后3天本 SKU 单 1" not in forbidden_tokens


def test_report_refresh_failures_blocks_latest_recommendations_missing_review_14d_metric_token(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(
        "keyword_action_review",
        search_term_or_target="formal 14d metric missing keyword",
        asin="B014DMETRIC",
        current_7d_promoted_ad_orders=1,
        current_7d_acos=0.08,
        current_7d_tacos=0.06,
        current_7d_total_orders=2,
        current_7d_available_stock=20,
        current_14d_promoted_ad_orders=3,
        current_14d_acos=0.09,
        current_14d_tacos=0.07,
        current_14d_total_orders=5,
        current_14d_available_stock=18,
        days_since_execution=8,
        review_window="7天后复盘",
        review_phase="7d",
        effect_evidence="7天复盘口径，等待人工确认。",
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24")
        + "\nformal 14d metric missing keyword B014DMETRIC"
        + "\n本 SKU 单 1 ACOS 0.08 TACOS 0.06 总单 2 库存 20"
        + "\n14天 本 SKU 单 3 14天 ACOS 0.09 14天 总单 5 14天 库存 18",
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert "latest_recommendations.html missing keyword review row 1 metric token 14天 TACOS 0.07" in failures


def test_report_refresh_failures_blocks_latest_recommendations_review_metric_bound_to_wrong_row(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(
        "keyword_action_review",
        search_term_or_target="formal bound keyword",
        asin="B0METRICBOUND",
        current_7d_promoted_ad_orders=1,
        current_7d_acos=0.08,
        current_7d_tacos=0.06,
        current_7d_total_orders=2,
        current_7d_available_stock=20,
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24")
        + "\nformal bound keyword B0METRICBOUND"
        + "\nwrong metric keyword 本 SKU 单 1 ACOS 8% TACOS 6% 总单 2 库存 20",
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert (
        "latest_recommendations.html keyword review row 1 metric token 本 SKU 单 1 not bound to formal bound keyword"
        in failures
    )


def test_report_refresh_failures_blocks_latest_recommendations_missing_pending_ad_identity(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    task_row = {
        "marketplace": "UK",
        "sku": "SKU-PENDING-FORMAL",
        "asin": "B0PENDINGFORMAL",
        "product_name": "Formal pending ad product",
        "confirmed_status": "待确认",
        "priority": "P1",
        "issue_type": "广告动作",
        "action_group": "广告动作",
        "today_action": "降竞价10%-20%",
        "suggested_action": "降竞价10%-20%",
        "copy_action_line": "formal missing pending target 降竞价10%-20%",
        "search_term_or_target": "formal missing pending target",
        "normalized_action": "bid_down",
        "action_id": "UK||SKU-PENDING-FORMAL||B0PENDINGFORMAL||formal missing pending target||bid_down",
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["today_task_queue_rows"] = [task_row]
    snapshot["today_action_groups"] = {"广告动作": [task_row]}
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24"),
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        today_task_rows=[task_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert "latest_recommendations.html missing pending ad row 1 target formal missing pending target" in failures


def test_report_refresh_failures_blocks_latest_recommendations_non_executable_ad_completion_payload(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    expected_action_id = "UK||SKU-HTMLPAY||B0HTMLPAY||search_term||formal html payload term||bid_down"
    observe_action_id = "UK||SKU-HTMLPAY||B0HTMLPAY||search_term||formal html payload term||observe"
    ad_row = {
        "marketplace": "UK",
        "sku": "SKU-HTMLPAY",
        "asin": "B0HTMLPAY",
        "product_name": "Formal HTML payload product",
        "confirmed_status": "待确认",
        "suggested_action": "降竞价10%-20%",
        "copy_action_line": "建议降竞价 10%-20%",
        "search_term_or_target": "formal html payload term",
        "campaign_name": "Payload campaign",
        "ad_group_name": "Payload group",
        "match_type_or_targeting": "EXACT",
        "clicks": "7",
        "orders": "0",
        "normalized_action": "bid_down",
        "action_scope": "search_term",
        "action_id": expected_action_id,
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["html_search_term_processing_queue_rows"] = [ad_row]
    snapshot["search_term_processing_queue_rows"] = [ad_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    bad_payload = {
        **ad_row,
        "suggested_action": "观察",
        "manual_action_taken": "观察",
        "normalized_action": "observe",
        "action_id": observe_action_id,
    }
    payload_attr = html.escape(
        json.dumps(bad_payload, ensure_ascii=False, separators=(",", ":")),
        quote=True,
    )
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        "\n".join(
            [
                "2026-06-24 产品级结论 市场调查 执行后效果复盘",
                '<section id="product-operation-cards">产品级结论</section>',
                '<section id="today-ad-actions-all">复制到广告后台 formal html payload term B0HTMLPAY</section>',
                '<section id="frontend-evidence-status">市场调查</section>',
                '<section id="action-effect-review">执行后效果复盘</section>',
                f'<div data-ad-complete-payload="{payload_attr}" data-action-id="{observe_action_id}"></div>',
            ]
        ),
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        ad_processing_rows=[ad_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert (
        "latest_recommendations.html ad completion payload 1 formal html payload term "
        "is not executable but appears in ad completion payload"
    ) in failures
    assert (
        f"latest_recommendations.html missing ad completion payload action_id {expected_action_id}"
        in failures
    )


def test_report_refresh_failures_blocks_marketplace_report_non_executable_ad_completion_payload(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    expected_action_id = "UK||SKU-MARKETPAY||B0MARKETPAY||search_term||market html payload term||bid_down"
    observe_action_id = "UK||SKU-MARKETPAY||B0MARKETPAY||search_term||market html payload term||observe"
    ad_row = {
        "marketplace": "UK",
        "sku": "SKU-MARKETPAY",
        "asin": "B0MARKETPAY",
        "product_name": "Market HTML payload product",
        "confirmed_status": "待确认",
        "suggested_action": "降竞价10%-20%",
        "copy_action_line": "建议降竞价 10%-20%",
        "search_term_or_target": "market html payload term",
        "campaign_name": "Market payload campaign",
        "ad_group_name": "Market payload group",
        "match_type_or_targeting": "EXACT",
        "clicks": "7",
        "orders": "0",
        "normalized_action": "bid_down",
        "action_scope": "search_term",
        "action_id": expected_action_id,
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["html_search_term_processing_queue_rows"] = [ad_row]
    snapshot["search_term_processing_queue_rows"] = [ad_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    bad_payload = {
        **ad_row,
        "suggested_action": "观察",
        "manual_action_taken": "观察",
        "normalized_action": "observe",
        "action_id": observe_action_id,
    }
    payload_attr = html.escape(
        json.dumps(bad_payload, ensure_ascii=False, separators=(",", ":")),
        quote=True,
    )
    uk_report = tmp_path / "uk_report.html"
    uk_report.write_text(
        "\n".join(
            [
                "2026-06-24 亚马逊运营日报｜UK 站点状态 市场调查 执行后效果复盘",
                '<section id="frontend-evidence-status">市场调查</section>',
                '<section id="action-effect-review">执行后效果复盘</section>',
                "复制到广告后台 market html payload term B0MARKETPAY",
                f'<div data-ad-complete-payload="{payload_attr}" data-action-id="{observe_action_id}"></div>',
            ]
        ),
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        ad_processing_rows=[ad_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, uk_report], previous_mtimes_ns={})

    assert (
        "uk_report.html ad completion payload 1 market html payload term "
        "is not executable but appears in ad completion payload"
    ) in failures
    assert f"uk_report.html missing ad completion payload action_id {expected_action_id}" in failures


def test_report_refresh_failures_blocks_copy_block_only_pending_ad_missing_copy_area(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    task_row = {
        "marketplace": "UK",
        "sku": "SKU-COPY-ONLY",
        "asin": "B0COPYONLY",
        "product_name": "Formal copy block only product",
        "confirmed_status": "待确认",
        "priority": "P1",
        "issue_type": "广告动作",
        "action_group": "广告动作",
        "today_action": "观察",
        "suggested_action": "观察",
        "copy_action_line": "",
        "copy_block": "formal copy block only target 降竞价10%-20%",
        "search_term_or_target": "formal copy block only target",
        "normalized_action": "observe",
        "action_id": "UK||SKU-COPY-ONLY||B0COPYONLY||formal copy block only target||observe",
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["today_task_queue_rows"] = [task_row]
    snapshot["today_action_groups"] = {"广告动作": [task_row]}
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24"),
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        today_task_rows=[task_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert "latest_recommendations.html missing copy area for pending ad rows" in failures


def test_report_refresh_failures_blocks_pending_growth_test_missing_copy_area(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    growth_row = {
        "marketplace": "UK",
        "sku": "SKU-GROWTH-COPY",
        "asin": "B0GROWTHCOPY",
        "product_name": "Formal growth copy product",
        "search_term_or_target": "formal growth copy target",
        "confirmed_status": "待确认",
        "suggested_action": "小预算试投",
        "experiment_type": "growth_test",
        "normalized_action": "growth_test",
        "action_id": "UK||SKU-GROWTH-COPY||B0GROWTHCOPY||search_term||formal growth copy target||growth_test",
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["growth_test_rows"] = [growth_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24"),
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        growth_test_rows=[growth_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert "latest_recommendations.html missing copy area for pending ad rows" in failures


def test_report_refresh_failures_blocks_latest_recommendations_missing_tomorrow_review_identity(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = {
        "marketplace": "UK",
        "sku": "SKU-TOMORROW-FORMAL",
        "asin": "B0TOMORROWFORMAL",
        "product_name": "Formal tomorrow missing product",
        "review_reason": "执行后复查",
        "current_evidence": "点击 4；花费 £0.88；订单 0",
        "tomorrow_check": "复查该词点击、花费、订单和 ACOS",
        "search_term_or_target": "formal tomorrow missing target",
        "normalized_action": "bid_down",
        "action_id": "UK||SKU-TOMORROW-FORMAL||B0TOMORROWFORMAL||formal tomorrow missing target||bid_down",
    }
    payload["marketplace_results"][0]["report_view_snapshot"]["tomorrow_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    latest_recommendations = tmp_path / "latest_recommendations.html"
    latest_recommendations.write_text(
        _text_report_fixture("latest_recommendations.html", "2026-06-24"),
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        tomorrow_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, latest_recommendations], previous_mtimes_ns={})

    assert "latest_recommendations.html missing tomorrow review row 1 target formal tomorrow missing target" in failures


def test_report_refresh_failures_blocks_marketplace_report_missing_keyword_review_identity(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(search_term_or_target="formal market missing keyword")
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    uk_report = tmp_path / "uk_report.html"
    uk_report.write_text(
        "\n".join(
            [
                "2026-06-24",
                "亚马逊运营日报｜UK",
                "站点状态",
                "市场调查",
                "执行后效果复盘",
                '<section id="frontend-evidence-status"></section>',
                '<section id="action-effect-review"></section>',
            ]
        ),
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, uk_report], previous_mtimes_ns={})

    assert "uk_report.html missing keyword review row 1 target formal market missing keyword" in failures
    assert "uk_report.html missing keyword review row 1 ASIN B0REVIEW1" in failures


def test_report_refresh_failures_blocks_marketplace_report_missing_action_review_identity(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(
        "action_review",
        asin="B0ACTIONMARKET",
        product_name="Formal market missing action product",
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    uk_report = tmp_path / "uk_report.html"
    uk_report.write_text(
        "\n".join(
            [
                "2026-06-24",
                "亚马逊运营日报｜UK",
                "站点状态",
                "市场调查",
                "执行后效果复盘",
                '<section id="frontend-evidence-status"></section>',
                '<section id="action-effect-review"></section>',
            ]
        ),
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        action_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, uk_report], previous_mtimes_ns={})

    assert "uk_report.html missing action review row 1 ASIN B0ACTIONMARKET" in failures
    assert "uk_report.html missing action review row 1 product name Formal market missing action product" in failures


def test_report_refresh_failures_blocks_marketplace_report_missing_review_metric_token(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(
        "keyword_action_review",
        search_term_or_target="formal market metric keyword",
        asin="B0MARKETMETRIC",
        current_7d_promoted_ad_orders=1,
        current_7d_acos=0.08,
        current_7d_tacos=0.06,
        current_7d_total_orders=2,
        current_7d_available_stock=20,
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    uk_report = tmp_path / "uk_report.html"
    uk_report.write_text(
        "\n".join(
            [
                "2026-06-24",
                "亚马逊运营日报｜UK",
                "站点状态",
                "市场调查",
                "执行后效果复盘",
                '<section id="frontend-evidence-status"></section>',
                '<section id="action-effect-review"></section>',
                "formal market metric keyword B0MARKETMETRIC",
            ]
        ),
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, uk_report], previous_mtimes_ns={})

    assert "uk_report.html missing keyword review row 1 metric token 本 SKU 单 1" in failures


def test_report_refresh_failures_blocks_marketplace_report_missing_review_14d_metric_token(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    review_row = _review_row(
        "keyword_action_review",
        search_term_or_target="formal market 14d metric keyword",
        asin="B0MARKET14D",
        current_7d_promoted_ad_orders=1,
        current_7d_acos=0.08,
        current_7d_tacos=0.06,
        current_7d_total_orders=2,
        current_7d_available_stock=20,
        current_14d_promoted_ad_orders=3,
        current_14d_acos=0.09,
        current_14d_tacos=0.07,
        current_14d_total_orders=5,
        current_14d_available_stock=18,
        days_since_execution=8,
        review_window="7天后复盘",
        review_phase="7d",
        effect_evidence="7天复盘口径，等待人工确认。",
    )
    payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [review_row]
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    uk_report = tmp_path / "uk_report.html"
    uk_report.write_text(
        "\n".join(
            [
                "2026-06-24",
                "亚马逊运营日报｜UK",
                "站点状态",
                "市场调查",
                "执行后效果复盘",
                '<section id="frontend-evidence-status"></section>',
                '<section id="action-effect-review"></section>',
                "formal market 14d metric keyword B0MARKET14D",
                "本 SKU 单 1 ACOS 0.08 TACOS 0.06 总单 2 库存 20",
                "14天 本 SKU 单 3 14天 ACOS 0.09 14天 总单 5 14天 库存 18",
            ]
        ),
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        keyword_review_rows=[review_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, uk_report], previous_mtimes_ns={})

    assert "uk_report.html missing keyword review row 1 metric token 14天 TACOS 0.07" in failures


def test_report_refresh_failures_blocks_marketplace_report_missing_frontend_coverage_label(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    frontend_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0MARKETCOVER",
        "frontend_check_status": "沿用 2026-06-20 前台数据",
        "frontend_data_freshness": "沿用 2026-06-20 前台数据",
        "frontend_findings": "沿用 2026-06-20 前台数据",
        "frontend_evidence_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_cache_used": True,
        "frontend_auto_conclusion": "FRONTEND_WEAK",
        "frontend_search_status": "已读取部分结果",
        "frontend_search_partial_evidence": True,
        "frontend_evidence_quality_score": 58,
    }
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["frontend_check_queue_rows"] = [frontend_row]
    snapshot["frontend_coverage_summary"] = daily_update._frontend_coverage_expected([frontend_row])
    payload["frontend_coverage_summary"] = daily_update._frontend_coverage_aggregate_from_snapshots(
        payload["marketplace_results"]
    )
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    uk_report = tmp_path / "uk_report.html"
    uk_report.write_text(
        "\n".join(
            [
                "2026-06-24",
                "亚马逊运营日报｜UK",
                "站点状态",
                "市场调查",
                "执行后效果复盘",
                '<section id="frontend-evidence-status"></section>',
                '<section id="action-effect-review"></section>',
            ]
        ),
        encoding="utf-8",
    )
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, uk_report], previous_mtimes_ns={})

    assert "uk_report.html missing frontend coverage token 产品页成功" in failures
    assert "uk_report.html missing frontend coverage token 0/1" in failures


def test_report_refresh_failures_blocks_marketplace_report_missing_product_blocking_reason(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        asin="B0MARKETGATE1",
        product_name="Formal market gate product",
        frontend_blocking_reasons=["地区待确认，当前前台证据不能用于放量"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale"],
    )
    _attach_product_decision(payload, product_row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    uk_report = tmp_path / "uk_report.html"
    uk_report.write_text(
        "\n".join(
            [
                "2026-06-24",
                "亚马逊运营日报｜UK",
                "站点状态",
                "市场调查",
                "产品广告门禁",
                "B0MARKETGATE1 Formal market gate product 拦截加竞价 拦截加预算 拦截放量",
                '<section id="frontend-evidence-status"></section>',
            ]
        ),
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[product_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, uk_report], previous_mtimes_ns={})

    assert (
        "uk_report.html missing product decision row 1 blocking reason "
        "地区待确认，当前前台证据不能用于放量"
    ) in failures


def test_report_refresh_failures_blocks_marketplace_report_missing_product_blocked_action_label(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        asin="B0MARKETGATE2",
        product_name="Formal market blocked action product",
        frontend_blocking_reasons=["竞品可比性不足"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale"],
    )
    _attach_product_decision(payload, product_row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    uk_report = tmp_path / "uk_report.html"
    uk_report.write_text(
        "\n".join(
            [
                "2026-06-24",
                "亚马逊运营日报｜UK",
                "站点状态",
                "市场调查",
                "产品广告门禁",
                "B0MARKETGATE2 Formal market blocked action product 竞品可比性不足 拦截加预算 拦截放量",
                '<section id="frontend-evidence-status"></section>',
            ]
        ),
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[product_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, uk_report], previous_mtimes_ns={})

    assert "uk_report.html missing product decision row 1 blocked action label bid_up" in failures


def test_report_refresh_failures_blocks_marketplace_report_missing_low_budget_blocked_action_label(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        asin="B0MARKETLOWBUDGET",
        product_name="Formal market low budget blocked product",
        frontend_blocking_reasons=["前台证据不足，需刷新后再放量"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
    )
    _attach_product_decision(payload, product_row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    uk_report = tmp_path / "uk_report.html"
    uk_report.write_text(
        "\n".join(
            [
                "2026-06-24",
                "亚马逊运营日报｜UK",
                "站点状态",
                "市场调查",
                "产品广告门禁",
                (
                    "B0MARKETLOWBUDGET Formal market low budget blocked product "
                    "前台证据不足，需刷新后再放量 拦截加竞价 拦截加预算 拦截放量"
                ),
                '<section id="frontend-evidence-status"></section>',
            ]
        ),
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[product_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, uk_report], previous_mtimes_ns={})

    assert "uk_report.html missing product decision row 1 blocked action label create_exact_low_budget" in failures


def test_report_refresh_failures_blocks_marketplace_report_frontend_tokens_bound_to_wrong_article(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    payload = _analysis_payload("2026-06-24")
    product_row = _product_decision_row(
        asin="B0MARKETFRONT",
        product_name="Formal market frontend evidence product",
        frontend_blocking_reasons=["前台缓存或读取失败待确认，仅背景参考，不能单独支持放量"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
        frontend_check_status="沿用 2026-06-20 前台数据",
        frontend_cache_used=True,
        frontend_evidence_tier="仅背景参考",
        frontend_evidence_display_tier="仅背景参考",
        frontend_evidence_audit_detail="沿用缓存；搜索页仅部分读取",
        frontend_search_status="已读取部分结果",
        frontend_search_partial_evidence=True,
    )
    _attach_product_decision(payload, product_row)
    analysis.write_text(json.dumps(payload), encoding="utf-8")
    uk_report = tmp_path / "uk_report.html"
    uk_report.write_text(
        "\n".join(
            [
                "2026-06-24",
                "亚马逊运营日报｜UK",
                "站点状态",
                "市场调查",
                "产品广告门禁",
                '<section id="frontend-evidence-status"></section>',
                (
                    '<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-DECISION" '
                    'data-product-decision-asin="B0MARKETFRONT">'
                    "B0MARKETFRONT Formal market frontend evidence product 前台缓存或读取失败待确认，仅背景参考，不能单独支持放量 "
                    "拦截加竞价 拦截加预算 拦截放量 拦截低预算精准测试</article>"
                ),
                (
                    '<article data-product-decision-marketplace="US" data-product-decision-sku="SKU-US" '
                    'data-product-decision-asin="B0MARKETFRONT">'
                    "沿用 2026-06-20 前台数据 仅背景参考 沿用缓存；搜索页仅部分读取 搜索页：已读取部分结果</article>"
                ),
            ]
        ),
        encoding="utf-8",
    )
    _write_daily_report_workbook(
        tmp_path / "amazon_ops_report_2026-06-24.xlsx",
        product_rows=[product_row],
    )
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    failures = daily_update.report_refresh_failures([analysis, uk_report], previous_mtimes_ns={})

    assert (
        "uk_report.html product decision row 1 frontend status "
        "沿用 2026-06-20 前台数据 not bound to ASIN B0MARKETFRONT"
    ) in failures
    assert (
        "uk_report.html product decision row 1 frontend search partial marker "
        "已读取部分结果 not bound to ASIN B0MARKETFRONT"
    ) in failures
    assert (
        "uk_report.html product decision row 1 frontend audit detail "
        "沿用缓存；搜索页仅部分读取 not bound to ASIN B0MARKETFRONT"
    ) in failures


def test_report_refresh_failures_blocks_core_import_newer_than_report_date(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(
        json.dumps(
            [
                {
                    "original_filename": "ads_2026-06-25.csv",
                    "detected_type": "ads_report_all",
                    "detected_date_range": "2026-06-25~2026-06-25",
                    "status": "imported",
                }
            ]
        ),
        encoding="utf-8",
    )

    failures = daily_update.report_refresh_failures(
        [analysis],
        previous_mtimes_ns={},
        import_manifest_path=manifest,
    )

    assert "latest analysis report_date 2026-06-24 is older than imported core data ending 2026-06-25" in failures


def test_report_refresh_failures_blocks_successful_import_missing_detected_type(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads_2026-06-24.csv", "status": "imported"}]), encoding="utf-8")

    failures = daily_update.report_refresh_failures(
        [analysis],
        previous_mtimes_ns={},
        import_manifest_path=manifest,
    )

    assert "import manifest successful row 1 ads_2026-06-24.csv missing detected_type" in failures


def test_report_refresh_failures_blocks_core_filename_misclassified_as_enhanced_import(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(
        json.dumps(
            [
                {
                    "original_filename": "ads_report_all_2026-06-25.csv",
                    "target_path": "data/raw_ads/ads_report_all.csv",
                    "status": "imported",
                    "detected_type": "traffic_sales",
                    "detected_date_range": "2026-06-25~2026-06-25",
                }
            ]
        ),
        encoding="utf-8",
    )

    failures = daily_update.report_refresh_failures(
        [analysis],
        previous_mtimes_ns={},
        import_manifest_path=manifest,
    )

    assert (
        "import manifest successful row 1 ads_report_all_2026-06-25.csv "
        "looks like ads_report_all but detected_type is traffic_sales"
    ) in failures


def test_report_refresh_failures_blocks_successful_core_import_missing_parseable_date(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(
        json.dumps(
            [
                {
                    "original_filename": "ads_latest.csv",
                    "status": "imported",
                    "detected_type": "ads_report_all",
                }
            ]
        ),
        encoding="utf-8",
    )

    failures = daily_update.report_refresh_failures(
        [analysis],
        previous_mtimes_ns={},
        import_manifest_path=manifest,
    )

    assert "import manifest successful core row 1 ads_latest.csv missing parseable date" in failures


def test_report_refresh_failures_allows_enhanced_import_newer_than_report_date(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(
        json.dumps(
            [
                {
                    "original_filename": "traffic_sales_uk_wow_2026-06-25_2026-07-01.xlsx",
                    "detected_type": "traffic_sales",
                    "detected_date_range": "2026-06-25~2026-07-01",
                    "status": "imported",
                }
            ]
        ),
        encoding="utf-8",
    )

    failures = daily_update.report_refresh_failures(
        [analysis],
        previous_mtimes_ns={},
        import_manifest_path=manifest,
    )

    assert failures == []


def test_report_refresh_failures_blocks_missing_successful_core_import_when_required(tmp_path) -> None:
    analysis = tmp_path / "latest_analysis.json"
    analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(
        json.dumps(
            [
                {
                    "original_filename": "old_ads_2026-06-24.csv",
                    "detected_type": "ads_report_all",
                    "detected_date_range": "2026-06-24~2026-06-24",
                    "status": "skipped_old_duplicate",
                }
            ]
        ),
        encoding="utf-8",
    )

    failures = daily_update.report_refresh_failures(
        [analysis],
        previous_mtimes_ns={},
        import_manifest_path=manifest,
        require_successful_core_import=True,
    )

    assert "import manifest has no successful core ads or ERP import rows for this daily update" in failures


def test_daily_update_runs_import_then_regenerates_without_frontend_fetch(monkeypatch, tmp_path) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    receipt = tmp_path / "daily_update_validation_receipt.json"
    outputs = [tmp_path / "latest_analysis.json", tmp_path / "latest_recommendations.html", tmp_path / "dashboard.html"]
    for output in outputs:
        output.write_text(json.dumps(_analysis_payload("2026-06-23")) if output.name == "latest_analysis.json" else "old", encoding="utf-8")
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)
    monkeypatch.setattr(daily_update, "DAILY_UPDATE_VALIDATION_RECEIPT", receipt)
    monkeypatch.setattr(daily_update, "REQUIRED_REFRESHED_OUTPUTS", outputs)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[1] == "scripts/import_inbox_files.py":
            rows = [_manifest_core_row("new_ads_2026-06-24.csv")]
            manifest.write_text(json.dumps(rows), encoding="utf-8")
            _write_import_manifest_workbook(manifest_xlsx, rows)
        if args[1] == "main.py":
            for output in outputs:
                output.write_text(
                    json.dumps(_analysis_payload("2026-06-24"))
                    if output.name == "latest_analysis.json"
                    else _text_report_fixture(output.name, "2026-06-24"),
                    encoding="utf-8",
                )
            _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
            (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
            _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()

    assert code == 0
    assert calls == [
        ["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        ["/python", "scripts/import_inbox_files.py"],
        ["/python", "main.py", "--marketplace", "ALL"],
    ]
    joined = " ".join(" ".join(call) for call in calls)
    assert "scripts/run_all_with_frontend_checks.py" not in joined
    assert "--no-live-browser-frontend" not in joined
    assert "--frontend-method urllib" not in joined
    assert "--live-browser-frontend" not in joined
    assert "playwright" not in joined
    receipt_payload = json.loads(receipt.read_text(encoding="utf-8"))
    assert receipt_payload["schema_version"] == daily_update.DAILY_UPDATE_RECEIPT_SCHEMA_VERSION
    assert receipt_payload["result"] == "passed"
    assert receipt_payload["report_date"] == "2026-06-24"
    assert receipt_payload["import_manifest"]["success_row_count"] == 1
    assert receipt_payload["import_manifest"]["json"]["exists"] is True
    assert receipt_payload["import_manifest"]["xlsx"]["exists"] is True
    analysis_receipt = next(
        value for key, value in receipt_payload["outputs"].items() if key.endswith("latest_analysis.json")
    )
    assert analysis_receipt["sha256"]


def test_daily_update_clears_old_success_receipt_when_import_step_fails(monkeypatch, tmp_path) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    receipt = tmp_path / "daily_update_validation_receipt.json"
    receipt.write_text(json.dumps({"result": "passed", "report_date": "2026-06-23"}), encoding="utf-8")
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "DAILY_UPDATE_VALIDATION_RECEIPT", receipt)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        return 19 if args[1] == "scripts/import_inbox_files.py" else 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)

    code = daily_update.main()

    assert code == 19
    assert calls == [
        ["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        ["/python", "scripts/import_inbox_files.py"],
    ]
    assert not receipt.exists()


def test_daily_update_stops_when_report_outputs_are_not_refreshed(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    outputs = [tmp_path / "latest_analysis.json", tmp_path / "latest_recommendations.html", tmp_path / "dashboard.html"]
    for output in outputs:
        output.write_text(json.dumps(_analysis_payload("2026-06-24")) if output.name == "latest_analysis.json" else "old", encoding="utf-8")
    _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
    (tmp_path / "autoopt_log_20260624.json").write_text("{}", encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)
    monkeypatch.setattr(daily_update, "REQUIRED_REFRESHED_OUTPUTS", outputs)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[1] == "scripts/import_inbox_files.py":
            rows = [_manifest_core_row("new_ads_2026-06-24.csv")]
            manifest.write_text(json.dumps(rows), encoding="utf-8")
            _write_import_manifest_workbook(manifest_xlsx, rows)
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 1
    assert calls == [
        ["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        ["/python", "scripts/import_inbox_files.py"],
        ["/python", "main.py", "--marketplace", "ALL"],
    ]
    assert "report refresh blocker" in output
    assert "required output was not refreshed by report step" in output
    assert "[restore] report outputs restored to pre-report snapshot after failure" in output
    assert "[done] daily update completed" not in output


def test_daily_update_restores_report_outputs_when_report_step_fails(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    analysis = tmp_path / "latest_analysis.json"
    html = tmp_path / "latest_recommendations.html"
    dashboard = tmp_path / "dashboard.html"
    outputs = [analysis, html, dashboard]
    analysis.write_text(json.dumps(_analysis_payload("2026-06-23")), encoding="utf-8")
    html.write_text("old recommendations", encoding="utf-8")
    dashboard.write_text("old dashboard", encoding="utf-8")
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)
    monkeypatch.setattr(daily_update, "REQUIRED_REFRESHED_OUTPUTS", outputs)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[1] == "scripts/import_inbox_files.py":
            rows = [_manifest_core_row("new_ads_2026-06-24.csv")]
            manifest.write_text(json.dumps(rows), encoding="utf-8")
            _write_import_manifest_workbook(manifest_xlsx, rows)
        if args[1] == "main.py":
            analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
            html.write_text("partial recommendations", encoding="utf-8")
            dashboard.write_text("partial dashboard", encoding="utf-8")
            return 17
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 17
    assert calls[-1][1] == "main.py"
    assert json.loads(analysis.read_text(encoding="utf-8"))["report_date"] == "2026-06-23"
    assert html.read_text(encoding="utf-8") == "old recommendations"
    assert dashboard.read_text(encoding="utf-8") == "old dashboard"
    assert "[restore] report outputs restored to pre-report snapshot after failure" in output
    assert "[done] daily update completed" not in output


def test_daily_update_allows_stale_frontend_results_when_report_outputs_refresh(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    analysis = tmp_path / "latest_analysis.json"
    html = tmp_path / "latest_recommendations.html"
    dashboard = tmp_path / "dashboard.html"
    frontend_results = tmp_path / "frontend_check_results.json"
    old_results_payload = {
        "generated_at": "2026-06-23T10:00:00",
        "refresh_summary": {"frontend_refresh_total": 1},
        "items": [{"marketplace": "UK", "sku": "SKU-FRONTEND", "asin": "B0FRONTEND1"}],
    }
    frontend_results.write_text(json.dumps(old_results_payload), encoding="utf-8")
    outputs = [analysis, html, dashboard]
    analysis.write_text(json.dumps(_analysis_payload("2026-06-23")), encoding="utf-8")
    html.write_text("old recommendations", encoding="utf-8")
    dashboard.write_text("old dashboard", encoding="utf-8")
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)
    monkeypatch.setattr(daily_update, "FRONTEND_RESULTS_JSON", frontend_results)
    monkeypatch.setattr(daily_update, "REQUIRED_REFRESHED_OUTPUTS", outputs)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[1] == "scripts/import_inbox_files.py":
            rows = [_manifest_core_row("new_ads_2026-06-24.csv")]
            manifest.write_text(json.dumps(rows), encoding="utf-8")
            _write_import_manifest_workbook(manifest_xlsx, rows)
        if args[1] == "main.py":
            payload = _analysis_payload("2026-06-24")
            analysis.write_text(json.dumps(payload), encoding="utf-8")
            html.write_text(_text_report_fixture("latest_recommendations.html", "2026-06-24"), encoding="utf-8")
            dashboard.write_text(_text_report_fixture("dashboard.html", "2026-06-24"), encoding="utf-8")
            _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
            (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
            _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 0
    assert calls[-1][1] == "main.py"
    assert json.loads(analysis.read_text(encoding="utf-8"))["report_date"] == "2026-06-24"
    assert "frontend wrapper blocker" not in output
    assert "frontend check results were not refreshed by frontend wrapper" not in output
    assert "[restore] report outputs restored to pre-report snapshot after failure" not in output
    assert "[done] daily update completed with inbox import and cached market evidence" in output


def test_daily_update_restores_report_outputs_when_final_validation_fails(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    analysis = tmp_path / "latest_analysis.json"
    html = tmp_path / "latest_recommendations.html"
    dashboard = tmp_path / "dashboard.html"
    outputs = [analysis, html, dashboard]
    analysis.write_text(json.dumps(_analysis_payload("2026-06-23")), encoding="utf-8")
    html.write_text("old recommendations", encoding="utf-8")
    dashboard.write_text("old dashboard", encoding="utf-8")

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)
    monkeypatch.setattr(daily_update, "REQUIRED_REFRESHED_OUTPUTS", outputs)
    monkeypatch.setattr(daily_update, "report_refresh_failures", lambda **kwargs: ["synthetic validation failure"])

    def fake_run_step(args: list[str]) -> int:
        if args[1] == "scripts/import_inbox_files.py":
            rows = [_manifest_core_row("new_ads_2026-06-24.csv")]
            manifest.write_text(json.dumps(rows), encoding="utf-8")
            _write_import_manifest_workbook(manifest_xlsx, rows)
        if args[1] == "main.py":
            analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
            html.write_text("validated later but should restore", encoding="utf-8")
            dashboard.write_text("validated later but should restore", encoding="utf-8")
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 1
    assert json.loads(analysis.read_text(encoding="utf-8"))["report_date"] == "2026-06-23"
    assert html.read_text(encoding="utf-8") == "old recommendations"
    assert dashboard.read_text(encoding="utf-8") == "old dashboard"
    assert "synthetic validation failure" in output
    assert "[restore] report outputs restored to pre-report snapshot after failure" in output
    assert "[done] daily update completed" not in output


def test_daily_update_removes_new_date_scoped_outputs_when_final_validation_fails(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    analysis = tmp_path / "latest_analysis.json"
    html = tmp_path / "latest_recommendations.html"
    dashboard = tmp_path / "dashboard.html"
    outputs = [analysis, html, dashboard]
    analysis.write_text(json.dumps(_analysis_payload("2026-06-23")), encoding="utf-8")
    html.write_text("old recommendations", encoding="utf-8")
    dashboard.write_text("old dashboard", encoding="utf-8")
    new_daily_excel = tmp_path / "amazon_ops_report_2026-06-24.xlsx"
    new_autoopt_json = tmp_path / "autoopt_log_20260624.json"
    new_keyword_review = tmp_path / "keyword_action_review_20260624.json"

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)
    monkeypatch.setattr(daily_update, "REQUIRED_REFRESHED_OUTPUTS", outputs)
    monkeypatch.setattr(daily_update, "report_refresh_failures", lambda **kwargs: ["synthetic validation failure"])

    def fake_run_step(args: list[str]) -> int:
        if args[1] == "scripts/import_inbox_files.py":
            rows = [_manifest_core_row("new_ads_2026-06-24.csv")]
            manifest.write_text(json.dumps(rows), encoding="utf-8")
            _write_import_manifest_workbook(manifest_xlsx, rows)
        if args[1] == "main.py":
            analysis.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
            html.write_text("new recommendations", encoding="utf-8")
            dashboard.write_text("new dashboard", encoding="utf-8")
            new_daily_excel.write_text("partial excel", encoding="utf-8")
            new_autoopt_json.write_text("partial autoopt", encoding="utf-8")
            new_keyword_review.write_text("partial keyword review", encoding="utf-8")
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 1
    assert json.loads(analysis.read_text(encoding="utf-8"))["report_date"] == "2026-06-23"
    assert html.read_text(encoding="utf-8") == "old recommendations"
    assert dashboard.read_text(encoding="utf-8") == "old dashboard"
    assert not new_daily_excel.exists()
    assert not new_autoopt_json.exists()
    assert not new_keyword_review.exists()
    assert "[restore] report outputs restored to pre-report snapshot after failure" in output
    assert "[done] daily update completed" not in output


def test_daily_update_stops_when_date_scoped_outputs_are_not_refreshed(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    outputs = [tmp_path / "latest_analysis.json", tmp_path / "latest_recommendations.html", tmp_path / "dashboard.html"]
    for output in outputs:
        output.write_text(
            json.dumps(_analysis_payload("2026-06-23"))
            if output.name == "latest_analysis.json"
            else _text_report_fixture(output.name, "2026-06-23"),
            encoding="utf-8",
        )
    stale_daily_report = tmp_path / "amazon_ops_report_2026-06-24.xlsx"
    _write_daily_report_workbook(stale_daily_report)
    (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
    _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)
    monkeypatch.setattr(daily_update, "REQUIRED_REFRESHED_OUTPUTS", outputs)

    def fake_run_step(args: list[str]) -> int:
        if args[1] == "scripts/import_inbox_files.py":
            rows = [_manifest_core_row("new_ads_2026-06-24.csv")]
            manifest.write_text(json.dumps(rows), encoding="utf-8")
            _write_import_manifest_workbook(manifest_xlsx, rows)
        if args[1] == "main.py":
            for output in outputs:
                output.write_text(
                    json.dumps(_analysis_payload("2026-06-24"))
                    if output.name == "latest_analysis.json"
                    else _text_report_fixture(output.name, "2026-06-24"),
                    encoding="utf-8",
                )
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 1
    assert f"required output was not refreshed by report step: {stale_daily_report}" in output
    assert "[done] daily update completed" not in output


def test_daily_update_stops_when_refreshed_latest_recommendations_is_malformed(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    outputs = [tmp_path / "latest_analysis.json", tmp_path / "latest_recommendations.html", tmp_path / "dashboard.html"]
    for output in outputs:
        output.write_text(json.dumps(_analysis_payload("2026-06-23")) if output.name == "latest_analysis.json" else "old", encoding="utf-8")
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)
    monkeypatch.setattr(daily_update, "REQUIRED_REFRESHED_OUTPUTS", outputs)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[1] == "scripts/import_inbox_files.py":
            rows = [_manifest_core_row("new_ads_2026-06-24.csv")]
            manifest.write_text(json.dumps(rows), encoding="utf-8")
            _write_import_manifest_workbook(manifest_xlsx, rows)
        if args[1] == "main.py":
            for output in outputs:
                if output.name == "latest_analysis.json":
                    output.write_text(json.dumps(_analysis_payload("2026-06-24")), encoding="utf-8")
                elif output.name == "latest_recommendations.html":
                    output.write_text("2026-06-24 产品级结论 市场调查 执行后效果复盘", encoding="utf-8")
                else:
                    output.write_text(_text_report_fixture(output.name, "2026-06-24"), encoding="utf-8")
            _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
            (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
            _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 1
    assert calls == [
        ["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        ["/python", "scripts/import_inbox_files.py"],
        ["/python", "main.py", "--marketplace", "ALL"],
    ]
    assert "report refresh blocker" in output
    assert "text report missing required marker today-ad-actions-all" in output
    assert "[done] daily update completed" not in output


def test_daily_update_stops_when_autoopt_log_content_is_stale(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    outputs = [tmp_path / "latest_analysis.json", tmp_path / "latest_recommendations.html", tmp_path / "dashboard.html"]
    for output in outputs:
        output.write_text(json.dumps(_analysis_payload("2026-06-23")) if output.name == "latest_analysis.json" else "old", encoding="utf-8")
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)
    monkeypatch.setattr(daily_update, "REQUIRED_REFRESHED_OUTPUTS", outputs)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[1] == "scripts/import_inbox_files.py":
            rows = [_manifest_core_row("new_ads_2026-06-24.csv")]
            manifest.write_text(json.dumps(rows), encoding="utf-8")
            _write_import_manifest_workbook(manifest_xlsx, rows)
        if args[1] == "main.py":
            for output in outputs:
                output.write_text(
                    json.dumps(_analysis_payload("2026-06-24"))
                    if output.name == "latest_analysis.json"
                    else _text_report_fixture(output.name, "2026-06-24"),
                    encoding="utf-8",
                )
            _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
            (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-23"}), encoding="utf-8")
            _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 1
    assert calls == [
        ["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        ["/python", "scripts/import_inbox_files.py"],
        ["/python", "main.py", "--marketplace", "ALL"],
    ]
    assert "report refresh blocker" in output
    assert "autoopt log report_date mismatch: expected 2026-06-24, got 2026-06-23" in output
    assert "[done] daily update completed" not in output


def test_daily_update_stops_when_daily_excel_content_date_is_stale(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    outputs = [tmp_path / "latest_analysis.json", tmp_path / "latest_recommendations.html", tmp_path / "dashboard.html"]
    for output in outputs:
        output.write_text(json.dumps(_analysis_payload("2026-06-23")) if output.name == "latest_analysis.json" else "old", encoding="utf-8")
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)
    monkeypatch.setattr(daily_update, "REQUIRED_REFRESHED_OUTPUTS", outputs)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[1] == "scripts/import_inbox_files.py":
            rows = [_manifest_core_row("new_ads_2026-06-24.csv")]
            manifest.write_text(json.dumps(rows), encoding="utf-8")
            _write_import_manifest_workbook(manifest_xlsx, rows)
        if args[1] == "main.py":
            for output in outputs:
                output.write_text(
                    json.dumps(_analysis_payload("2026-06-24"))
                    if output.name == "latest_analysis.json"
                    else _text_report_fixture(output.name, "2026-06-24"),
                    encoding="utf-8",
                )
            _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx", report_date="2026-06-23")
            (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
            _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 1
    assert calls == [
        ["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        ["/python", "scripts/import_inbox_files.py"],
        ["/python", "main.py", "--marketplace", "ALL"],
    ]
    assert "report refresh blocker" in output
    assert "daily Excel report_date mismatch: expected 2026-06-24" in output
    assert "[done] daily update completed" not in output


def test_daily_update_stops_when_autoopt_excel_content_date_is_stale(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    outputs = [tmp_path / "latest_analysis.json", tmp_path / "latest_recommendations.html", tmp_path / "dashboard.html"]
    for output in outputs:
        output.write_text(json.dumps(_analysis_payload("2026-06-23")) if output.name == "latest_analysis.json" else "old", encoding="utf-8")
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)
    monkeypatch.setattr(daily_update, "REQUIRED_REFRESHED_OUTPUTS", outputs)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[1] == "scripts/import_inbox_files.py":
            rows = [_manifest_core_row("new_ads_2026-06-24.csv")]
            manifest.write_text(json.dumps(rows), encoding="utf-8")
            _write_import_manifest_workbook(manifest_xlsx, rows)
        if args[1] == "main.py":
            for output in outputs:
                output.write_text(
                    json.dumps(_analysis_payload("2026-06-24"))
                    if output.name == "latest_analysis.json"
                    else _text_report_fixture(output.name, "2026-06-24"),
                    encoding="utf-8",
                )
            _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
            (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
            _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx", report_date="2026-06-23")
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 1
    assert calls == [
        ["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        ["/python", "scripts/import_inbox_files.py"],
        ["/python", "main.py", "--marketplace", "ALL"],
    ]
    assert "report refresh blocker" in output
    assert "autoopt Excel report_date mismatch: expected 2026-06-24, got 2026-06-23" in output
    assert "[done] daily update completed" not in output


def test_daily_update_stops_when_autoopt_excel_row_counts_are_stale(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    outputs = [tmp_path / "latest_analysis.json", tmp_path / "latest_recommendations.html", tmp_path / "dashboard.html"]
    for output in outputs:
        output.write_text(json.dumps(_analysis_payload("2026-06-23")) if output.name == "latest_analysis.json" else "old", encoding="utf-8")
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)
    monkeypatch.setattr(daily_update, "REQUIRED_REFRESHED_OUTPUTS", outputs)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[1] == "scripts/import_inbox_files.py":
            rows = [_manifest_core_row("new_ads_2026-06-24.csv")]
            manifest.write_text(json.dumps(rows), encoding="utf-8")
            _write_import_manifest_workbook(manifest_xlsx, rows)
        if args[1] == "main.py":
            for output in outputs:
                output.write_text(
                    json.dumps(_analysis_payload("2026-06-24"))
                    if output.name == "latest_analysis.json"
                    else _text_report_fixture(output.name, "2026-06-24"),
                    encoding="utf-8",
                )
            _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
            (tmp_path / "autoopt_log_20260624.json").write_text(
                json.dumps({"report_date": "2026-06-24", "keyword_action_review_rows": [{"action_id": "k1"}]}),
                encoding="utf-8",
            )
            _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 1
    assert calls == [
        ["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        ["/python", "scripts/import_inbox_files.py"],
        ["/python", "main.py", "--marketplace", "ALL"],
    ]
    assert "report refresh blocker" in output
    assert "autoopt Excel keyword_action_review row count mismatch: expected 1 from keyword_action_review_rows, got 0" in output
    assert "[done] daily update completed" not in output


def test_daily_update_stops_when_standalone_review_json_is_stale(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    outputs = [tmp_path / "latest_analysis.json", tmp_path / "latest_recommendations.html", tmp_path / "dashboard.html"]
    for output in outputs:
        output.write_text(json.dumps(_analysis_payload("2026-06-23")) if output.name == "latest_analysis.json" else "old", encoding="utf-8")
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)
    monkeypatch.setattr(daily_update, "REQUIRED_REFRESHED_OUTPUTS", outputs)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[1] == "scripts/import_inbox_files.py":
            rows = [_manifest_core_row("new_ads_2026-06-24.csv")]
            manifest.write_text(json.dumps(rows), encoding="utf-8")
            _write_import_manifest_workbook(manifest_xlsx, rows)
        if args[1] == "main.py":
            for output in outputs:
                output.write_text(
                    json.dumps(_analysis_payload("2026-06-24"))
                    if output.name == "latest_analysis.json"
                    else _text_report_fixture(output.name, "2026-06-24"),
                    encoding="utf-8",
                )
            _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
            (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
            _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
            (tmp_path / "keyword_action_review_20260624.json").write_text(
                json.dumps([{"action_id": "old-k1"}]),
                encoding="utf-8",
            )
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 1
    assert calls == [
        ["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        ["/python", "scripts/import_inbox_files.py"],
        ["/python", "main.py", "--marketplace", "ALL"],
    ]
    assert "report refresh blocker" in output
    assert (
        "keyword_action_review_20260624.json row count mismatch: "
        "expected 0 from keyword_action_review_rows, got 1"
    ) in output
    assert "[done] daily update completed" not in output


def test_daily_update_stops_when_review_window_overstates_seven_day_readiness(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    outputs = [tmp_path / "latest_analysis.json", tmp_path / "latest_recommendations.html", tmp_path / "dashboard.html"]
    for output in outputs:
        output.write_text(
            json.dumps(_analysis_payload("2026-06-23")) if output.name == "latest_analysis.json" else "old",
            encoding="utf-8",
        )
    calls: list[list[str]] = []
    review_row = _review_row(
        action_id="k-overstated-7d-main",
        search_term_or_target="main overstated review window term",
        review_outcome="not_ready",
        outcome="待7天确认",
        days_since_execution=4,
        review_window="7d_check",
        review_status="待7天复盘",
        review_phase="not_ready",
    )

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)
    monkeypatch.setattr(daily_update, "REQUIRED_REFRESHED_OUTPUTS", outputs)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[1] == "scripts/import_inbox_files.py":
            rows = [
                _manifest_core_row("ads_2026-06-24.csv")
            ]
            manifest.write_text(json.dumps(rows), encoding="utf-8")
            _write_import_manifest_workbook(manifest_xlsx, rows)
        if args[1] == "main.py":
            payload = _analysis_payload("2026-06-24")
            payload["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [
                review_row
            ]
            for output in outputs:
                output.write_text(
                    json.dumps(payload) if output.name == "latest_analysis.json" else _text_report_fixture(output.name, "2026-06-24"),
                    encoding="utf-8",
                )
            _write_daily_report_workbook(
                tmp_path / "amazon_ops_report_2026-06-24.xlsx",
                keyword_review_rows=[review_row],
            )
            (tmp_path / "autoopt_log_20260624.json").write_text(
                json.dumps({"report_date": "2026-06-24"}),
                encoding="utf-8",
            )
            _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 1
    assert calls == [
        ["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        ["/python", "scripts/import_inbox_files.py"],
        ["/python", "main.py", "--marketplace", "ALL"],
    ]
    assert "report refresh blocker" in output
    assert "main overstated review window term claims 7-day review window before 7 days" in output
    assert "[done] daily update completed" not in output


def test_daily_update_stops_when_imported_core_data_is_newer_than_report(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    outputs = [tmp_path / "latest_analysis.json", tmp_path / "latest_recommendations.html", tmp_path / "dashboard.html"]
    for output in outputs:
        output.write_text(json.dumps(_analysis_payload("2026-06-23")) if output.name == "latest_analysis.json" else "old", encoding="utf-8")
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)
    monkeypatch.setattr(daily_update, "REQUIRED_REFRESHED_OUTPUTS", outputs)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[1] == "scripts/import_inbox_files.py":
            rows = [
                _manifest_core_row("ads_2026-06-25.csv", detected_date_range="2026-06-25~2026-06-25")
            ]
            manifest.write_text(json.dumps(rows), encoding="utf-8")
            _write_import_manifest_workbook(manifest_xlsx, rows)
        if args[1] == "main.py":
            for output in outputs:
                output.write_text(
                    json.dumps(_analysis_payload("2026-06-24"))
                    if output.name == "latest_analysis.json"
                    else _text_report_fixture(output.name, "2026-06-24"),
                    encoding="utf-8",
                )
            _write_daily_report_workbook(tmp_path / "amazon_ops_report_2026-06-24.xlsx")
            (tmp_path / "autoopt_log_20260624.json").write_text(json.dumps({"report_date": "2026-06-24"}), encoding="utf-8")
            _write_autoopt_workbook(tmp_path / "autoopt_20260624.xlsx")
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 1
    assert calls == [
        ["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        ["/python", "scripts/import_inbox_files.py"],
        ["/python", "main.py", "--marketplace", "ALL"],
    ]
    assert "report refresh blocker" in output
    assert "latest analysis report_date 2026-06-24 is older than imported core data ending 2026-06-25" in output
    assert "[done] daily update completed" not in output


def test_daily_update_stops_when_preflight_fails(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        return 23

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 23
    assert calls == [["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"]]
    assert "[done] daily update completed" not in output


def test_daily_update_stops_when_step_cannot_start(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    def fake_run(args: list[str], cwd=None):
        calls.append(args)
        raise OSError("python missing")

    monkeypatch.setattr(daily_update.subprocess, "run", fake_run)

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 127
    assert calls == [["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"]]
    assert "[fail] cannot start step: python missing" in output
    assert "scripts/import_inbox_files.py" not in output
    assert "[done] daily update completed" not in output


def test_daily_update_stops_when_import_step_fails(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        return 19 if args[1] == "scripts/import_inbox_files.py" else 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 19
    assert calls == [
        ["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        ["/python", "scripts/import_inbox_files.py"],
    ]
    assert "[done] daily update completed" not in output


def test_daily_update_stops_when_import_manifest_has_blockers(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    blocker_rows = [{"original_filename": "unknown.xlsx", "status": "unknown", "reason": "cannot classify"}]
    manifest.write_text(json.dumps([{"original_filename": "old_ads.csv", "status": "imported"}]), encoding="utf-8")
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[1] == "scripts/import_inbox_files.py":
            manifest.write_text(json.dumps(blocker_rows), encoding="utf-8")
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 1
    assert calls == [
        ["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        ["/python", "scripts/import_inbox_files.py"],
    ]
    assert "import manifest blocker" in output
    assert "scripts/run_all_with_frontend_checks.py" not in " ".join(" ".join(call) for call in calls)
    assert "[done] daily update completed" not in output


def test_daily_update_stops_when_import_step_does_not_refresh_manifest(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "old_ads.csv", "status": "imported"}]), encoding="utf-8")
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "run_step", lambda args: calls.append(args) or 0)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 1
    assert calls == [
        ["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        ["/python", "scripts/import_inbox_files.py"],
    ]
    assert "import manifest was not refreshed by current import step" in output
    assert "scripts/run_all_with_frontend_checks.py" not in " ".join(" ".join(call) for call in calls)
    assert "[done] daily update completed" not in output


def test_daily_update_stops_when_import_manifest_xlsx_content_is_stale(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "old_ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    _write_import_manifest_workbook(
        manifest_xlsx,
        [
            {
                "original_filename": "old_ads.csv",
                "status": "imported",
                "created_at": "2026-06-23T09:30:00",
                "target_path": "data/raw/ads/old_ads.csv",
                "archive_path": "data/archive/ads/old_ads.csv",
            }
        ],
    )
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[1] == "scripts/import_inbox_files.py":
            manifest.write_text(
                json.dumps(
                    [
                        {
                            "original_filename": "new_ads.csv",
                            "status": "imported",
                            "created_at": "2026-06-24T09:30:00",
                            "target_path": "data/raw/ads/new_ads.csv",
                            "archive_path": "data/archive/ads/new_ads.csv",
                        }
                    ]
                ),
                encoding="utf-8",
            )
            _write_import_manifest_workbook(
                manifest_xlsx,
                [
                    {
                        "original_filename": "old_ads.csv",
                        "status": "imported",
                        "created_at": "2026-06-23T09:30:00",
                        "target_path": "data/raw/ads/old_ads.csv",
                        "archive_path": "data/archive/ads/old_ads.csv",
                    }
                ],
            )
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 1
    assert calls == [
        ["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        ["/python", "scripts/import_inbox_files.py"],
    ]
    assert "import manifest blocker" in output
    assert "import manifest xlsx identity mismatch vs import manifest json for import manifest rows" in output
    assert "scripts/run_all_with_frontend_checks.py" not in " ".join(" ".join(call) for call in calls)
    assert "[done] daily update completed" not in output


def test_daily_update_stops_when_import_manifest_xlsx_is_missing_after_import(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "old_ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[1] == "scripts/import_inbox_files.py":
            manifest.write_text(
                json.dumps(
                    [
                        {
                            "original_filename": "new_ads.csv",
                            "status": "imported",
                            "created_at": "2026-06-24T09:30:00",
                            "target_path": "data/raw/ads/new_ads.csv",
                            "archive_path": "data/archive/ads/new_ads.csv",
                        }
                    ]
                ),
                encoding="utf-8",
            )
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 1
    assert calls == [
        ["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        ["/python", "scripts/import_inbox_files.py"],
    ]
    assert "import manifest blocker" in output
    assert "import manifest xlsx missing" in output
    assert "scripts/run_all_with_frontend_checks.py" not in " ".join(" ".join(call) for call in calls)
    assert "[done] daily update completed" not in output


def test_daily_update_stops_when_import_manifest_success_rows_are_not_from_current_run(
    monkeypatch,
    tmp_path,
    capsys,
) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "old_ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    target = tmp_path / "raw" / "new_ads.csv"
    archive = tmp_path / "archive" / "new_ads.csv"
    target.parent.mkdir(parents=True)
    archive.parent.mkdir(parents=True)
    target.write_text("target", encoding="utf-8")
    archive.write_text("archive", encoding="utf-8")
    stale_rows = [
        {
            "original_filename": "new_ads.csv",
            "status": "imported",
            "created_at": "2026-06-23 09:30:00",
            "target_path": str(target),
            "archive_path": str(archive),
        }
    ]
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[1] == "scripts/import_inbox_files.py":
            manifest.write_text(json.dumps(stale_rows), encoding="utf-8")
            _write_import_manifest_workbook(manifest_xlsx, stale_rows)
        return 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 1
    assert calls == [
        ["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        ["/python", "scripts/import_inbox_files.py"],
    ]
    assert "import manifest blocker" in output
    assert "created_at 2026-06-23 09:30:00 is older than current import step started at" in output
    assert "scripts/run_all_with_frontend_checks.py" not in " ".join(" ".join(call) for call in calls)
    assert "[done] daily update completed" not in output


def test_daily_update_stops_when_report_step_fails(monkeypatch, tmp_path, capsys) -> None:
    manifest = tmp_path / "import_manifest.json"
    manifest.write_text(json.dumps([{"original_filename": "ads.csv", "status": "imported"}]), encoding="utf-8")
    manifest_xlsx = tmp_path / "import_manifest.xlsx"
    calls: list[list[str]] = []

    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_JSON", manifest)
    monkeypatch.setattr(daily_update, "IMPORT_MANIFEST_XLSX", manifest_xlsx)
    monkeypatch.setattr(daily_update.sys, "executable", "/python")

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[1] == "scripts/import_inbox_files.py":
            rows = [_manifest_core_row("new_ads_2026-06-24.csv")]
            manifest.write_text(json.dumps(rows), encoding="utf-8")
            _write_import_manifest_workbook(manifest_xlsx, rows)
        return 17 if args[1] == "main.py" else 0

    monkeypatch.setattr(daily_update, "run_step", fake_run_step)

    code = daily_update.main()
    output = capsys.readouterr().out

    assert code == 17
    assert calls == [
        ["/python", "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        ["/python", "scripts/import_inbox_files.py"],
        ["/python", "main.py", "--marketplace", "ALL"],
    ]
    assert "[done] daily update completed" not in output
    assert "[open]" not in output

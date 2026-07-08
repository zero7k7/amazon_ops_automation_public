from __future__ import annotations

import html
import json
from pathlib import Path

import scripts.run_daily_update as daily_update
import scripts.validate_showcase_mvp as showcase_validate
from src.autoopt_feedback import add_action_identity


def test_validate_canonical_report_value_treats_nan_as_blank_like_formal_daily() -> None:
    for value in [float("nan"), "NaN"]:
        assert showcase_validate._canonical_report_value(value) == ""
        assert showcase_validate._canonical_report_value(value) == daily_update._canonical_report_value(value)


def test_safe_run_required_outputs_reuse_formal_daily_required_outputs() -> None:
    formal_names = [
        path.relative_to(daily_update.OUTPUT_DIR).as_posix()
        for path in daily_update.REQUIRED_REFRESHED_OUTPUTS
    ]

    assert showcase_validate.DAILY_REFRESH_REQUIRED_FILES == formal_names
    assert showcase_validate.REQUIRED_FILES == [
        name for name in formal_names if not name.startswith("assets/")
    ]
    assert showcase_validate.REQUIRED_ASSET_FILES == [
        name for name in formal_names if name.startswith("assets/")
    ]


def test_validate_safe_run_blocks_malformed_report_assets(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
    )
    (safe_dir / "assets" / "report.css").write_text("body{}", encoding="utf-8")
    (safe_dir / "assets" / "report.js").write_text("console.log('ok')", encoding="utf-8")

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "safe-run report asset report.css missing required token .ad-task-card" in output


def test_safe_run_snapshot_keys_reuse_formal_daily_snapshot_keys() -> None:
    assert showcase_validate.REQUIRED_SNAPSHOT_KEYS == (
        daily_update.REQUIRED_REPORT_VIEW_SNAPSHOT_LIST_KEYS
        | daily_update.REQUIRED_REPORT_VIEW_SNAPSHOT_DICT_KEYS
    )


def test_validate_frontend_rows_accepts_pending_and_cached_status() -> None:
    rows = [
        {
            "asin": "B0PENDING",
            "frontend_check_status": "待前台检查",
            "frontend_data_freshness": "无可用前台数据",
            "frontend_findings": "暂无最近成功前台数据",
        },
        {
            "asin": "B0CACHED",
            "frontend_check_status": "沿用 2026-06-05 前台数据",
            "frontend_data_freshness": "沿用 2026-06-05 前台数据",
            "frontend_findings": "沿用 2026-06-05 前台数据",
        },
    ]

    assert showcase_validate._validate_frontend_rows("US", rows) == 0


def test_validate_frontend_rows_blocks_failed_status_without_fallback(capsys) -> None:
    rows = [
        {
            "asin": "B0BAD",
            "frontend_check_status": "读取失败",
            "frontend_data_freshness": "",
            "frontend_findings": "自动读取失败",
        }
    ]

    code = showcase_validate._validate_frontend_rows("US", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "missing cached-date or pending-check status" in output


def test_validate_frontend_rows_blocks_cache_status_without_date(capsys) -> None:
    rows = [
        {
            "asin": "B0BAD",
            "frontend_check_status": "沿用缓存前台数据",
            "frontend_data_freshness": "沿用缓存前台数据",
            "frontend_findings": "沿用缓存前台数据",
        }
    ]

    code = showcase_validate._validate_frontend_rows("US", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "cache status missing date" in output


def test_validate_frontend_rows_blocks_cache_status_with_number_but_no_date(capsys) -> None:
    rows = [
        {
            "asin": "B0BADCACHEDATE",
            "frontend_check_status": "沿用最近 1 天前台数据",
            "frontend_data_freshness": "沿用最近 1 天前台数据",
            "frontend_findings": "沿用最近 1 天前台数据",
        }
    ]

    code = showcase_validate._validate_frontend_rows("US", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "cache status missing date" in output


def test_validate_frontend_rows_blocks_cached_strong_evidence(capsys) -> None:
    rows = [
        {
            "asin": "B0BADSTRONG",
            "frontend_check_status": "沿用 2026-06-05 前台数据",
            "frontend_data_freshness": "沿用 2026-06-05 前台数据",
            "frontend_findings": "沿用 2026-06-05 前台数据",
            "frontend_evidence_tier": "强诊断可用",
            "frontend_evidence_is_strong": True,
            "frontend_auto_conclusion": "FRONTEND_OK",
        }
    ]

    code = showcase_validate._validate_frontend_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "cached or pending evidence marked strong" in output


def test_validate_frontend_rows_allows_live_strong_evidence() -> None:
    rows = [
        {
            "asin": "B0LIVEGOOD",
            "frontend_check_status": "已自动检查",
            "frontend_evidence_tier": "强诊断可用",
            "frontend_evidence_is_strong": True,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "competitor_comparability": "high",
            "frontend_evidence_quality_score": 82,
        }
    ]

    assert showcase_validate._validate_frontend_rows("UK", rows) == 0


def test_validate_frontend_rows_blocks_live_strong_without_explicit_flag(capsys) -> None:
    rows = [
        {
            "asin": "B0MISSINGFLAG",
            "frontend_check_status": "已自动检查",
            "frontend_evidence_tier": "强诊断可用",
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "competitor_comparability": "high",
            "frontend_evidence_quality_score": 82,
        }
    ]

    code = showcase_validate._validate_frontend_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "strong evidence missing explicit strong flag" in output


def test_validate_frontend_rows_blocks_live_strong_currency_warning(capsys) -> None:
    rows = [
        {
            "asin": "B0BADCURRENCY",
            "frontend_check_status": "已自动检查",
            "frontend_evidence_tier": "强诊断可用",
            "frontend_evidence_is_strong": True,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_price_currency_warning": "价格币种异常：TWD594.77，已忽略",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_exact": True,
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
        }
    ]

    code = showcase_validate._validate_frontend_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "currency warning marked strong" in output


def test_validate_frontend_rows_blocks_live_strong_location_warning(capsys) -> None:
    rows = [
        {
            "asin": "B0BADLOCATIONWARN",
            "frontend_check_status": "已自动检查",
            "frontend_evidence_tier": "强诊断可用",
            "frontend_evidence_is_strong": True,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_location_warning": "UK 地区异常：United States",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_exact": True,
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
        }
    ]

    code = showcase_validate._validate_frontend_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "location warning marked strong" in output


def test_validate_frontend_rows_blocks_live_strong_uncertain_location(capsys) -> None:
    rows = [
        {
            "asin": "B0BADLOCATION",
            "frontend_check_status": "已自动检查",
            "frontend_evidence_tier": "强诊断可用",
            "frontend_evidence_is_strong": True,
            "frontend_auto_conclusion": "FRONTEND_WEAK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "marketplace",
            "frontend_location_exact": False,
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
        }
    ]

    code = showcase_validate._validate_frontend_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "uncertain location marked strong" in output


def test_validate_frontend_rows_blocks_live_strong_missing_exact_location(capsys) -> None:
    rows = [
        {
            "asin": "B0MISSINGLOC",
            "frontend_check_status": "已自动检查",
            "frontend_evidence_tier": "强诊断可用",
            "frontend_evidence_is_strong": True,
            "frontend_auto_conclusion": "FRONTEND_WEAK",
            "frontend_failure_category": "none",
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
        }
    ]

    code = showcase_validate._validate_frontend_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "strong evidence missing exact location" in output


def test_validate_frontend_rows_blocks_live_strong_missing_verified_location(capsys) -> None:
    rows = [
        {
            "asin": "B0MISSINGLOCVERIFY",
            "frontend_check_status": "已自动检查",
            "frontend_evidence_tier": "强诊断可用",
            "frontend_evidence_is_strong": True,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_exact": True,
            "frontend_location_verified": False,
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
        }
    ]

    code = showcase_validate._validate_frontend_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "strong evidence missing verified location" in output


def test_validate_frontend_rows_blocks_live_strong_weak_competitor_sample(capsys) -> None:
    rows = [
        {
            "asin": "B0BADCOMP",
            "frontend_check_status": "已自动检查",
            "frontend_evidence_tier": "强诊断可用",
            "frontend_evidence_is_strong": True,
            "frontend_auto_conclusion": "FRONTEND_WEAK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 1,
            "competitor_comparability": "low",
        }
    ]

    code = showcase_validate._validate_frontend_rows("US", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "weak competitor evidence marked strong" in output


def test_validate_frontend_rows_blocks_live_strong_low_comparable_competitor_count(capsys) -> None:
    rows = [
        {
            "asin": "B0LOWCOMPARABLE",
            "frontend_check_status": "已自动检查",
            "frontend_evidence_tier": "强诊断可用",
            "frontend_evidence_is_strong": True,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "competitor_comparability": "high",
            "comparable_competitor_count": 0,
            "frontend_evidence_quality_score": 88,
        }
    ]

    code = showcase_validate._validate_frontend_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "weak competitor evidence marked strong" in output


def test_validate_frontend_rows_blocks_live_strong_missing_competitor_count(capsys) -> None:
    rows = [
        {
            "asin": "B0MISSINGCOMP",
            "frontend_check_status": "已自动检查",
            "frontend_evidence_tier": "强诊断可用",
            "frontend_evidence_is_strong": True,
            "frontend_auto_conclusion": "FRONTEND_WEAK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_search_status": "已自动检查",
        }
    ]

    code = showcase_validate._validate_frontend_rows("US", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "strong evidence missing competitor count" in output


def test_validate_frontend_rows_blocks_frontend_ok_without_strong_tier(capsys) -> None:
    rows = [
        {
            "asin": "B0MISSINGTIER",
            "frontend_check_status": "已自动检查",
            "frontend_evidence_tier": "仅背景参考",
            "frontend_evidence_is_strong": False,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "competitor_comparability": "high",
            "frontend_evidence_quality_score": 88,
        }
    ]

    code = showcase_validate._validate_frontend_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "strong evidence missing strong decision tier" in output


def test_validate_frontend_rows_blocks_live_strong_missing_search_success(capsys) -> None:
    rows = [
        {
            "asin": "B0MISSINGSEARCH",
            "frontend_check_status": "已自动检查",
            "frontend_evidence_tier": "强诊断可用",
            "frontend_evidence_is_strong": True,
            "frontend_auto_conclusion": "FRONTEND_WEAK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_competitor_count": 3,
        }
    ]

    code = showcase_validate._validate_frontend_rows("DE", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "strong evidence missing successful search page" in output


def test_validate_frontend_rows_blocks_live_strong_missing_competitor_comparability(capsys) -> None:
    rows = [
        {
            "asin": "B0MISSINGCOMPARABILITY",
            "frontend_check_status": "已自动检查",
            "frontend_evidence_tier": "强诊断可用",
            "frontend_evidence_is_strong": True,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "frontend_evidence_quality_score": 88,
        }
    ]

    code = showcase_validate._validate_frontend_rows("US", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "strong evidence missing competitor comparability" in output


def test_validate_frontend_rows_blocks_live_strong_low_quality_score(capsys) -> None:
    rows = [
        {
            "asin": "B0LOWQUALITY",
            "frontend_check_status": "已自动检查",
            "frontend_evidence_tier": "强诊断可用",
            "frontend_evidence_is_strong": True,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "competitor_comparability": "high",
            "frontend_evidence_quality_score": 54,
        }
    ]

    code = showcase_validate._validate_frontend_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "strong evidence quality score below threshold" in output


def test_validate_frontend_rows_blocks_live_strong_partial_search(capsys) -> None:
    rows = [
        {
            "asin": "B0BADSEARCH",
            "frontend_check_status": "已自动检查",
            "frontend_evidence_tier": "强诊断可用",
            "frontend_evidence_is_strong": True,
            "frontend_auto_conclusion": "FRONTEND_WEAK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_search_status": "已读取部分结果",
            "frontend_competitor_count": 3,
        }
    ]

    code = showcase_validate._validate_frontend_rows("DE", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "partial search evidence marked strong" in output


def test_validate_ad_observation_rows_blocks_executed_watch_row(capsys) -> None:
    rows = [
        {
            "search_term_or_target": "low sample keyword",
            "suggested_action": "观察",
            "confirmed_status": "已执行",
        }
    ]

    code = showcase_validate._validate_ad_observation_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "ad observation row" in output


def test_validate_ad_observation_rows_allows_executable_done_row() -> None:
    rows = [
        {
            "search_term_or_target": "bad keyword",
            "suggested_action": "否定精准",
            "confirmed_status": "已执行",
        }
    ]

    assert showcase_validate._validate_ad_observation_rows("UK", rows) == 0


def _product_decision_row(**overrides) -> dict[str, object]:
    from src.product_decision_layer import PRODUCT_FINAL_DECISION_REQUIRED_FIELDS

    row = {field: "" for field in PRODUCT_FINAL_DECISION_REQUIRED_FIELDS}
    row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0DEFAULT",
            "product_name": "Default product",
            "final_decision": "SMALL_SCALE_ALLOWED",
            "final_decision_label": "可小放量",
            "decision_priority": 7,
            "decision_reason": "test",
            "today_allowed_actions": ["observe"],
            "today_blocked_actions": ["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
            "frontend_required": False,
            "frontend_posture": "ok",
            "frontend_evidence_state": "mixed",
            "frontend_blocking_reasons": [],
            "frontend_evidence_tier": "仅背景参考",
            "frontend_evidence_display_tier": "仅背景参考",
            "frontend_decision_evidence_tier": "仅背景参考",
            "frontend_evidence_quality_score": "",
            "frontend_check_status": "已自动检查",
            "frontend_cache_used": False,
            "frontend_failure_category": "none",
            "frontend_price_currency_warning": "",
            "frontend_location_warning": "",
            "frontend_search_status": "已自动检查",
            "frontend_search_partial_evidence": False,
            "frontend_auto_conclusion": "",
            "frontend_auto_conclusion_label": "",
            "inventory_constraint": "NONE",
            "evidence_used": [],
            "confidence": "medium",
            "competitor_comparability": "medium",
            "comparable_competitor_count": 3,
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_location_uncertain": False,
            "last_updated": "2026-06-24",
        }
    )
    row.update(overrides)
    if "frontend_evidence_display_tier" not in overrides:
        row["frontend_evidence_display_tier"] = row.get("frontend_evidence_tier") or ""
    if "frontend_decision_evidence_tier" not in overrides:
        row["frontend_decision_evidence_tier"] = (
            row.get("frontend_evidence_display_tier") or row.get("frontend_evidence_tier") or ""
        )
    if "frontend_evidence_is_strong" not in overrides:
        row["frontend_evidence_is_strong"] = (
            row.get("frontend_evidence_tier") == "强诊断可用"
            and row.get("frontend_evidence_display_tier") == "强诊断可用"
        )
    return row


def _product_decision_contract_attrs(row: dict, **overrides: object) -> str:
    def list_attr(value: object) -> str:
        if isinstance(value, list):
            return "|".join(str(item).strip() for item in value if str(item).strip())
        return str(value or "").strip()

    def bool_attr(value: object) -> str:
        if isinstance(value, bool):
            return "true" if value else "false"
        text = str(value or "").strip().lower()
        return "true" if text in {"1", "true", "yes", "y", "是", "已验证"} else "false"

    attrs = {
        "data-product-decision-final": row.get("final_decision") or "",
        "data-product-decision-label": row.get("final_decision_label") or "",
        "data-product-allowed-actions": list_attr(row.get("today_allowed_actions")),
        "data-product-blocked-actions": list_attr(row.get("today_blocked_actions")),
        "data-product-frontend-tier": row.get("frontend_evidence_tier") or "",
        "data-product-frontend-display-tier": row.get("frontend_evidence_display_tier")
        or row.get("frontend_evidence_tier")
        or "",
        "data-product-frontend-decision-tier": row.get("frontend_decision_evidence_tier")
        or row.get("frontend_evidence_display_tier")
        or row.get("frontend_evidence_tier")
        or "",
        "data-product-frontend-strong": bool_attr(row.get("frontend_evidence_is_strong")),
    }
    attrs.update(overrides)
    return "".join(
        f' {name}="{html.escape(str(value), quote=True)}"'
        for name, value in attrs.items()
        if str(value).strip()
    )


def _boolish_test_flag(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "是", "已验证"}


def _product_operation_card(row: dict, **overrides: object) -> dict:
    card = {
        "marketplace": row.get("marketplace") or "UK",
        "sku": row.get("sku") or "",
        "asin": row.get("asin") or "",
        "product_name": row.get("product_name") or "",
        "final_decision": row.get("final_decision") or "",
        "final_decision_label": row.get("final_decision_label") or "",
        "decision_reason": row.get("decision_reason") or "",
        "operation_main_reason": row.get("decision_reason") or "",
        "fusion_issue_type": row.get("fusion_issue_type") or "",
        "fusion_action_gate": row.get("fusion_action_gate") or "",
        "fusion_do_not_do": row.get("fusion_do_not_do") or "",
        "fusion_review_window": row.get("fusion_review_window") or row.get("next_review_date") or "",
        "today_allowed_actions": row.get("today_allowed_actions") or [],
        "today_blocked_actions": row.get("today_blocked_actions") or [],
        "frontend_status": row.get("frontend_check_status") or "",
        "frontend_evidence_tier": row.get("frontend_evidence_tier") or "",
        "frontend_evidence_display_tier": row.get("frontend_evidence_display_tier") or "",
        "frontend_decision_evidence_tier": row.get("frontend_decision_evidence_tier")
        or row.get("frontend_evidence_display_tier")
        or row.get("frontend_evidence_tier")
        or "",
        "frontend_evidence_is_strong": _boolish_test_flag(row.get("frontend_evidence_is_strong")),
        "frontend_cache_used": _boolish_test_flag(row.get("frontend_cache_used")),
        "frontend_evidence_audit_summary": row.get("frontend_evidence_audit_summary") or "",
        "inventory_constraint": row.get("inventory_constraint") or "",
        "frontend_evidence_audit_reasons": row.get("frontend_evidence_audit_reasons") or [],
        "ad_action_count": 0,
        "ad_action_display_limit": 4,
        "ad_action_more_count": 0,
        "ad_action_items": [],
        "ad_diagnostic_summary": "",
    }
    card.update(overrides)
    return card


def test_product_operation_card_helper_treats_string_false_booleans_as_false() -> None:
    card = _product_operation_card(
        _product_decision_row(frontend_evidence_is_strong="False", frontend_cache_used="False")
    )

    assert card["frontend_evidence_is_strong"] is False
    assert card["frontend_cache_used"] is False


def test_validate_product_final_decision_blocks_missing_contract_fields(capsys) -> None:
    rows = [{"asin": "B0MISSING", "today_allowed_actions": ["observe"]}]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "product final decision row" in output
    assert "missing fields" in output


def test_validate_workbook_product_decision_content_includes_frontend_gate_fields(capsys) -> None:
    from openpyxl import Workbook

    expected_row = _product_decision_row(
        frontend_required=True,
        frontend_posture="frontend_blocked",
        frontend_blocking_reasons=["地区待确认，当前前台证据不能用于放量"],
        frontend_location_uncertain=True,
        frontend_location_block_reason="地区待确认，当前前台证据不能用于放量",
    )
    actual_row = dict(expected_row)
    actual_row["frontend_blocking_reasons"] = []
    actual_row["frontend_location_uncertain"] = False
    headers = [
        *showcase_validate.PRODUCT_DECISION_IDENTITY_FIELDS,
        *showcase_validate.PRODUCT_DECISION_EXCEL_CONTENT_FIELDS,
    ]
    headers = list(dict.fromkeys(headers))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "产品最终决策"
    sheet.append(headers)
    sheet.append(
        [
            json.dumps(actual_row.get(header), ensure_ascii=False)
            if isinstance(actual_row.get(header), list)
            else actual_row.get(header, "")
            for header in headers
        ]
    )

    code = showcase_validate._validate_workbook_content(
        workbook,
        "产品最终决策",
        [expected_row],
        showcase_validate.PRODUCT_DECISION_IDENTITY_FIELDS,
        showcase_validate.PRODUCT_DECISION_EXCEL_CONTENT_FIELDS,
        "product final decisions",
    )
    output = capsys.readouterr().out

    assert code == 1
    assert "field frontend_blocking_reasons" in output


def test_validate_workbook_product_decision_content_includes_frontend_strong_flag(capsys) -> None:
    from openpyxl import Workbook

    expected_row = _product_decision_row(frontend_evidence_is_strong=True)
    actual_row = dict(expected_row)
    actual_row["frontend_evidence_is_strong"] = False
    headers = [
        *showcase_validate.PRODUCT_DECISION_IDENTITY_FIELDS,
        *showcase_validate.PRODUCT_DECISION_EXCEL_CONTENT_FIELDS,
    ]
    headers = list(dict.fromkeys(headers))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "产品最终决策"
    sheet.append(headers)
    sheet.append(
        [
            json.dumps(actual_row.get(header), ensure_ascii=False)
            if isinstance(actual_row.get(header), list)
            else actual_row.get(header, "")
            for header in headers
        ]
    )

    code = showcase_validate._validate_workbook_content(
        workbook,
        "产品最终决策",
        [expected_row],
        showcase_validate.PRODUCT_DECISION_IDENTITY_FIELDS,
        showcase_validate.PRODUCT_DECISION_EXCEL_CONTENT_FIELDS,
        "product final decisions",
    )
    output = capsys.readouterr().out

    assert code == 1
    assert "field frontend_evidence_is_strong" in output


def test_validate_workbook_content_blocks_duplicate_expected_identity(capsys) -> None:
    from openpyxl import Workbook

    expected_row = _product_decision_row(frontend_evidence_is_strong=True)
    duplicate_row = dict(expected_row)
    duplicate_row["final_decision"] = "CONSERVATIVE_RUN"
    headers = [
        *showcase_validate.PRODUCT_DECISION_IDENTITY_FIELDS,
        *showcase_validate.PRODUCT_DECISION_EXCEL_CONTENT_FIELDS,
    ]
    headers = list(dict.fromkeys(headers))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "产品最终决策"
    sheet.append(headers)
    sheet.append(
        [
            json.dumps(expected_row.get(header), ensure_ascii=False)
            if isinstance(expected_row.get(header), list)
            else expected_row.get(header, "")
            for header in headers
        ]
    )

    code = showcase_validate._validate_workbook_content(
        workbook,
        "产品最终决策",
        [expected_row, duplicate_row],
        showcase_validate.PRODUCT_DECISION_IDENTITY_FIELDS,
        showcase_validate.PRODUCT_DECISION_EXCEL_CONTENT_FIELDS,
        "product final decisions",
    )
    output = capsys.readouterr().out

    assert code == 1
    assert "latest_analysis.json duplicate identity for product final decisions" in output


def test_validate_workbook_content_blocks_duplicate_excel_identity(capsys) -> None:
    from openpyxl import Workbook

    expected_row = _product_decision_row(frontend_evidence_is_strong=True)
    conflicting_excel_row = dict(expected_row)
    conflicting_excel_row["final_decision"] = "CONSERVATIVE_RUN"
    headers = [
        *showcase_validate.PRODUCT_DECISION_IDENTITY_FIELDS,
        *showcase_validate.PRODUCT_DECISION_EXCEL_CONTENT_FIELDS,
    ]
    headers = list(dict.fromkeys(headers))
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "产品最终决策"
    sheet.append(headers)
    for row in [expected_row, conflicting_excel_row]:
        sheet.append(
            [
                json.dumps(row.get(header), ensure_ascii=False)
                if isinstance(row.get(header), list)
                else row.get(header, "")
                for header in headers
            ]
        )

    code = showcase_validate._validate_workbook_content(
        workbook,
        "产品最终决策",
        [expected_row],
        showcase_validate.PRODUCT_DECISION_IDENTITY_FIELDS,
        showcase_validate.PRODUCT_DECISION_EXCEL_CONTENT_FIELDS,
        "product final decisions",
    )
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 产品最终决策 duplicate identity for product final decisions" in output


def test_showcase_product_decision_excel_fields_reuse_daily_refresh_spec() -> None:
    import scripts.run_daily_update as daily_update

    assert showcase_validate.PRODUCT_DECISION_IDENTITY_FIELDS == list(
        daily_update.DAILY_EXCEL_CONSISTENCY_SPECS["product_final_decision_rows"]["identity_fields"]
    )
    assert showcase_validate.PRODUCT_DECISION_EXCEL_CONTENT_FIELDS == list(
        daily_update.DAILY_EXCEL_CONSISTENCY_SPECS["product_final_decision_rows"]["content_fields"]
    )
    assert showcase_validate.INVENTORY_REPLENISHMENT_IDENTITY_FIELDS == list(
        daily_update.DAILY_EXCEL_CONSISTENCY_SPECS["inventory_replenishment_rows"]["identity_fields"]
    )
    assert showcase_validate.INVENTORY_REPLENISHMENT_EXCEL_CONTENT_FIELDS == list(
        daily_update.DAILY_EXCEL_CONSISTENCY_SPECS["inventory_replenishment_rows"]["content_fields"]
    )


def test_frontend_decision_evidence_fields_stay_in_report_contracts() -> None:
    import scripts.run_daily_update as daily_update
    from src.product_decision_layer import FRONTEND_EVIDENCE_DECISION_FIELDS, PRODUCT_FINAL_DECISION_REQUIRED_FIELDS

    for field in FRONTEND_EVIDENCE_DECISION_FIELDS:
        assert field in PRODUCT_FINAL_DECISION_REQUIRED_FIELDS
        assert field in daily_update.DAILY_EXCEL_CONSISTENCY_SPECS["product_final_decision_rows"]["content_fields"]
        assert field in daily_update.FRONTEND_QUEUE_CONTENT_FIELDS
        assert field in daily_update.PRODUCT_OPERATION_CARD_CONTENT_FIELDS
        assert field in daily_update.PRODUCT_OPERATION_CARD_REQUIRED_FIELDS
    assert daily_update.PRODUCT_DECISION_CONTRACT_ATTRS["frontend_decision_evidence_tier"] == (
        "data-product-frontend-decision-tier"
    )
    assert daily_update.PRODUCT_DECISION_CONTRACT_ATTRS["frontend_evidence_is_strong"] == (
        "data-product-frontend-strong"
    )
    assert "frontend_evidence_audit_reasons" in daily_update.PRODUCT_OPERATION_CARD_CONTENT_FIELDS
    assert "frontend_evidence_audit_reasons" in daily_update.PRODUCT_OPERATION_CARD_REQUIRED_FIELDS
    assert "frontend_evidence_audit_reasons" in daily_update.FRONTEND_QUEUE_CONTENT_FIELDS
    assert "ad_action_display_limit" in daily_update.PRODUCT_OPERATION_CARD_CONTENT_FIELDS
    assert "ad_action_display_limit" in daily_update.PRODUCT_OPERATION_CARD_REQUIRED_FIELDS
    assert "ad_action_more_count" in daily_update.PRODUCT_OPERATION_CARD_CONTENT_FIELDS
    assert "ad_action_more_count" in daily_update.PRODUCT_OPERATION_CARD_REQUIRED_FIELDS


def test_product_identity_fields_stay_in_report_contracts() -> None:
    import scripts.run_daily_update as daily_update
    from src.product_decision_layer import PRODUCT_FINAL_DECISION_REQUIRED_FIELDS, PRODUCT_IDENTITY_FIELDS

    expected = list(PRODUCT_IDENTITY_FIELDS)
    assert expected == ["marketplace", "sku", "asin"]
    assert PRODUCT_FINAL_DECISION_REQUIRED_FIELDS[: len(expected)] == expected
    assert daily_update.PRODUCT_OPERATION_CARD_IDENTITY_FIELDS == expected
    assert daily_update.FRONTEND_QUEUE_IDENTITY_FIELDS == expected
    assert daily_update.DAILY_EXCEL_CONSISTENCY_SPECS["product_final_decision_rows"]["identity_fields"] == expected
    assert daily_update.DAILY_EXCEL_CONSISTENCY_SPECS["inventory_replenishment_rows"]["identity_fields"] == expected
    assert daily_update.TOP_LEVEL_SNAPSHOT_UNION_KEYS["product_final_decision_rows"][0] == expected
    assert daily_update.TOP_LEVEL_SNAPSHOT_UNION_KEYS["product_operation_cards"][0] == expected
    assert daily_update.TOP_LEVEL_SNAPSHOT_UNION_KEYS["inventory_replenishment_rows"][0] == expected
    assert daily_update.AUTOOPT_EXCEL_CONSISTENCY_SPECS["final_decisions"]["identity_fields"] == expected
    assert daily_update.AUTOOPT_LATEST_ANALYSIS_CONSISTENCY_SPECS["product_final_decisions"]["identity_fields"] == expected


def test_showcase_product_operation_validation_checks_all_pending_ad_items() -> None:
    product_row = _product_decision_row(
        asin="B0SHOWCASEOPS",
        product_name="Showcase operation product",
    )
    ad_rows = []
    for index in range(5):
        ad_row = {
            "marketplace": "UK",
            "sku": product_row["sku"],
            "asin": product_row["asin"],
            "product_name": product_row["product_name"],
            "search_term_or_target": f"showcase operation term {index + 1}",
            "suggested_action": "降竞价",
            "normalized_action": "bid_down",
            "action_scope": "search_term",
            "confirmed_status": "待确认",
        }
        ad_row["action_id"] = daily_update.make_action_id(ad_row, "bid_down", "search_term")
        ad_rows.append(ad_row)
    card = _product_operation_card(
        product_row,
        ad_action_count=4,
        ad_action_display_limit=4,
        ad_action_more_count=0,
        ad_action_items=ad_rows[:4],
    )

    failures = showcase_validate.daily_product_operation_card_failures(
        "UK",
        [card],
        [product_row],
        ad_rows,
    )

    assert any(
        f"missing ad_action_items action_id from ad workbench rows: {ad_rows[4]['action_id']}" in failure
        for failure in failures
    )


def test_report_consistency_specs_do_not_have_duplicate_fields() -> None:
    import scripts.run_daily_update as daily_update

    spec_groups = {
        "DAILY_EXCEL_CONSISTENCY_SPECS": daily_update.DAILY_EXCEL_CONSISTENCY_SPECS,
        "DAILY_EXCEL_REVIEW_CONSISTENCY_SPECS": daily_update.DAILY_EXCEL_REVIEW_CONSISTENCY_SPECS,
        "AUTOOPT_EXCEL_CONSISTENCY_SPECS": daily_update.AUTOOPT_EXCEL_CONSISTENCY_SPECS,
        "AUTOOPT_LATEST_ANALYSIS_CONSISTENCY_SPECS": daily_update.AUTOOPT_LATEST_ANALYSIS_CONSISTENCY_SPECS,
        "AUTOOPT_LATEST_ANALYSIS_SNAPSHOT_CONSISTENCY_SPECS": daily_update.AUTOOPT_LATEST_ANALYSIS_SNAPSHOT_CONSISTENCY_SPECS,
    }
    for group_name, specs in spec_groups.items():
        for spec_name, spec in specs.items():
            for field_list_name in ["identity_fields", "content_fields"]:
                fields = list(spec.get(field_list_name) or [])
                duplicates = sorted({field for field in fields if fields.count(field) > 1})
                assert duplicates == [], f"{group_name}.{spec_name}.{field_list_name} duplicates {duplicates}"


def test_showcase_action_queue_excel_fields_reuse_daily_refresh_spec() -> None:
    import scripts.run_daily_update as daily_update

    assert showcase_validate.TODAY_TASK_QUEUE_IDENTITY_FIELDS == list(
        daily_update.TODAY_TASK_QUEUE_IDENTITY_FIELDS
    )
    assert showcase_validate.TODAY_TASK_QUEUE_CONTENT_FIELDS == list(
        daily_update.TODAY_TASK_QUEUE_CONTENT_FIELDS
    )
    assert showcase_validate.TOMORROW_REVIEW_IDENTITY_FIELDS == list(
        daily_update.TOMORROW_REVIEW_IDENTITY_FIELDS
    )
    assert showcase_validate.TOMORROW_REVIEW_CONTENT_FIELDS == list(
        daily_update.TOMORROW_REVIEW_CONTENT_FIELDS
    )
    assert showcase_validate.DIAGNOSIS_ROW_IDENTITY_FIELDS == list(
        daily_update.DIAGNOSIS_ROW_IDENTITY_FIELDS
    )
    assert showcase_validate.DIAGNOSIS_ROW_CONTENT_FIELDS == list(
        daily_update.DIAGNOSIS_ROW_CONTENT_FIELDS
    )


def test_showcase_action_review_identity_fields_reuse_daily_refresh_spec() -> None:
    import scripts.run_daily_update as daily_update

    assert showcase_validate.ACTION_REVIEW_IDENTITY_FIELDS == list(
        daily_update.ACTION_REVIEW_IDENTITY_FIELDS
    )
    assert showcase_validate.KEYWORD_REVIEW_IDENTITY_FIELDS == list(
        daily_update.KEYWORD_REVIEW_IDENTITY_FIELDS
    )


def test_validate_product_final_decision_blocks_growth_without_ok_high_frontend(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0SCALEBAD1",
            today_allowed_actions=["observe", "bid_up"],
            today_blocked_actions=["budget_up", "broad_scale"],
            frontend_evidence_state="background",
            frontend_evidence_tier="仅背景参考",
            frontend_location_scope="exact",
            frontend_location_exact=True,
            competitor_comparability="high",
            comparable_competitor_count=3,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "allows growth without ok_high frontend evidence" in output


def test_validate_product_final_decision_blocks_growth_without_frontend_ok_code(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0NOFRONTENDOK",
            today_allowed_actions=["observe", "bid_up"],
            today_blocked_actions=["budget_up", "broad_scale"],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_check_status="已自动检查",
            frontend_search_status="已自动检查",
            frontend_auto_conclusion="",
            frontend_auto_conclusion_label="未见明显前台劣势",
            frontend_location_scope="exact",
            frontend_location_exact=True,
            frontend_location_verified=True,
            competitor_comparability="high",
            comparable_competitor_count=3,
            frontend_evidence_quality_score=88,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "allows growth without FRONTEND_OK conclusion" in output


def test_validate_product_final_decision_blocks_growth_without_explicit_strong_flag(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0NOSTRONGFLAG",
            today_allowed_actions=["observe", "bid_up"],
            today_blocked_actions=["budget_up", "broad_scale"],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_evidence_display_tier="强诊断可用",
            frontend_evidence_is_strong=False,
            frontend_check_status="已自动检查",
            frontend_search_status="已自动检查",
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_exact=True,
            frontend_location_verified=True,
            competitor_comparability="high",
            comparable_competitor_count=3,
            frontend_evidence_quality_score=88,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "allows growth without explicit strong frontend flag" in output


def test_validate_product_final_decision_blocks_strong_mark_without_explicit_strong_flag(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0STRONGNOFLAG",
            today_allowed_actions=["observe"],
            today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_evidence_display_tier="强诊断可用",
            frontend_evidence_is_strong=False,
            frontend_check_status="已自动检查",
            frontend_search_status="已自动检查",
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_exact=True,
            frontend_location_verified=True,
            competitor_comparability="high",
            comparable_competitor_count=3,
            frontend_evidence_quality_score=88,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "marks strong frontend evidence without explicit strong frontend flag" in output


def test_validate_product_final_decision_blocks_allowed_and_blocked_overlap(capsys) -> None:
    rows = [
            _product_decision_row(
                marketplace="US",
                asin="B0OVERLAP",
                today_allowed_actions=["observe", "bid_up"],
                today_blocked_actions=["bid_up", "budget_up"],
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("US", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "has actions both allowed and blocked: bid_up" in output


def test_validate_product_final_decision_blocks_strong_listing_with_cached_frontend(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0STRONGCACHE",
            today_allowed_actions=["observe"],
            today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_evidence_display_tier="强诊断可用",
            frontend_failure_category="none",
            frontend_check_status="沿用 2026-06-10 前台数据",
            frontend_cache_used=True,
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_exact=True,
            competitor_comparability="high",
            comparable_competitor_count=3,
            frontend_evidence_quality_score=88,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "marks strong frontend evidence with cached frontend evidence" in output


def test_validate_product_final_decision_blocks_strong_listing_without_strong_tier(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0STRONGTIERMISMATCH",
            today_allowed_actions=["observe"],
            today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="仅背景参考",
            frontend_evidence_display_tier="强诊断可用",
            frontend_failure_category="none",
            frontend_check_status="已自动检查",
            frontend_search_status="已自动检查",
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_verified=True,
            frontend_location_exact=True,
            competitor_comparability="high",
            comparable_competitor_count=3,
            frontend_evidence_quality_score=88,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "marks strong frontend evidence without explicit strong frontend flag" in output


def test_validate_product_final_decision_blocks_strong_boolean_without_strong_tier(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0STRONGBOOL",
            today_allowed_actions=["observe"],
            today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
            frontend_evidence_state="weak",
            frontend_evidence_tier="仅背景参考",
            frontend_evidence_display_tier="仅背景参考",
            frontend_evidence_is_strong=True,
            frontend_check_status="待前台检查",
            frontend_search_status="",
            frontend_auto_conclusion="FRONTEND_WEAK",
            frontend_cache_used=False,
            competitor_comparability="unknown",
            comparable_competitor_count=0,
            frontend_evidence_quality_score=30,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "marks strong frontend evidence without strong frontend display tier" in output


def test_validate_product_final_decision_blocks_strong_listing_with_currency_warning(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0STRONGCURR",
            today_allowed_actions=["observe"],
            today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_evidence_display_tier="强诊断可用",
            frontend_failure_category="none",
            frontend_check_status="已自动检查",
            frontend_price_currency_warning="价格币种异常：TWD594.77，已忽略",
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_exact=True,
            competitor_comparability="high",
            comparable_competitor_count=3,
            frontend_evidence_quality_score=88,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "marks strong frontend evidence with currency warning" in output


def test_validate_product_final_decision_blocks_strong_listing_with_partial_search(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0STRONGPARTIALSEARCH",
            today_allowed_actions=["observe"],
            today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_evidence_display_tier="强诊断可用",
            frontend_failure_category="none",
            frontend_check_status="已自动检查",
            frontend_search_status="已读取部分结果",
            frontend_search_partial_evidence=True,
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_verified=True,
            frontend_location_exact=True,
            competitor_comparability="medium",
            comparable_competitor_count=3,
            frontend_evidence_quality_score=88,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "marks strong frontend evidence with partial search evidence" in output


def test_validate_product_final_decision_blocks_strong_listing_with_uncertain_location(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0STRONGLOCATION",
            today_allowed_actions=["observe"],
            today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_evidence_display_tier="强诊断可用",
            frontend_failure_category="none",
            frontend_check_status="已自动检查",
            frontend_search_status="已自动检查",
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="marketplace",
            frontend_location_verified=True,
            frontend_location_exact=False,
            competitor_comparability="medium",
            comparable_competitor_count=3,
            frontend_evidence_quality_score=88,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "marks strong frontend evidence without exact frontend location" in output


def test_validate_product_final_decision_blocks_strong_listing_with_weak_competitors(capsys) -> None:
    rows = [
            _product_decision_row(
                marketplace="US",
                asin="B0STRONGWEAKCOMP",
                today_allowed_actions=["observe"],
                today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_evidence_display_tier="强诊断可用",
            frontend_failure_category="none",
            frontend_check_status="已自动检查",
            frontend_search_status="已自动检查",
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_verified=True,
            frontend_location_exact=True,
            competitor_comparability="low",
            comparable_competitor_count=1,
            frontend_evidence_quality_score=88,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("US", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "marks strong frontend evidence with weak competitor evidence" in output


def test_validate_product_final_decision_blocks_strong_listing_without_competitor_comparability(capsys) -> None:
    rows = [
            _product_decision_row(
                marketplace="US",
                asin="B0STRONGNOCOMP",
                today_allowed_actions=["observe"],
                today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_evidence_display_tier="强诊断可用",
            frontend_failure_category="none",
            frontend_check_status="已自动检查",
            frontend_search_status="已自动检查",
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_verified=True,
            frontend_location_exact=True,
            competitor_comparability="",
            comparable_competitor_count=3,
            frontend_evidence_quality_score=88,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("US", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "marks strong frontend evidence without competitor comparability" in output


def test_validate_product_final_decision_blocks_growth_when_display_tier_is_downgraded(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0DISPLAYTIER",
            today_allowed_actions=["observe", "bid_up"],
            today_blocked_actions=["budget_up", "broad_scale"],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_evidence_display_tier="仅背景参考",
            frontend_failure_category="none",
            frontend_check_status="已自动检查",
            frontend_search_status="已自动检查",
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_exact=True,
            frontend_location_verified=True,
            competitor_comparability="medium",
            comparable_competitor_count=3,
            frontend_evidence_quality_score=88,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "allows growth without strong frontend display tier" in output


def test_validate_product_final_decision_blocks_strong_tier_with_downgraded_display_tier(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0STRONGDISPLAYDOWN",
            today_allowed_actions=["observe"],
            today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
            frontend_evidence_state="weak",
            frontend_evidence_tier="强诊断可用",
            frontend_evidence_display_tier="仅背景参考",
            frontend_decision_evidence_tier="强诊断可用",
            frontend_failure_category="none",
            frontend_check_status="已自动检查",
            frontend_search_status="已自动检查",
            frontend_auto_conclusion="FRONTEND_WEAK",
            frontend_location_scope="exact",
            frontend_location_exact=True,
            frontend_location_verified=True,
            competitor_comparability="medium",
            comparable_competitor_count=3,
            frontend_evidence_quality_score=88,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "marks strong frontend evidence without strong frontend display tier" in output


def test_validate_product_final_decision_blocks_growth_with_weak_competitors(capsys) -> None:
    rows = [
            _product_decision_row(
                marketplace="DE",
                asin="B0SCALEBAD2",
                today_allowed_actions=["bid_up"],
                today_blocked_actions=[],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_failure_category="none",
            frontend_check_status="已自动检查",
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_exact=True,
            competitor_comparability="low",
            comparable_competitor_count=1,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("DE", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "allows growth with weak competitor evidence" in output


def test_validate_product_final_decision_blocks_weak_frontend_without_explicit_growth_blocks(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0WEAKBLOCKS",
            today_allowed_actions=["observe"],
            today_blocked_actions=["bid_up"],
            frontend_evidence_state="weak",
            frontend_evidence_tier="仅背景参考",
            frontend_evidence_display_tier="仅背景参考",
            frontend_check_status="待前台检查",
            frontend_auto_conclusion="FRONTEND_WEAK",
            competitor_comparability="unknown",
            comparable_competitor_count=0,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "does not explicitly block growth actions under weak frontend evidence" in output
    assert "budget_up" in output
    assert "broad_scale" in output


def test_validate_product_final_decision_allows_growth_with_strong_frontend() -> None:
    rows = [
        _product_decision_row(
            asin="B0SCALEGOOD",
            today_allowed_actions=["observe", "bid_up"],
            today_blocked_actions=["budget_up", "broad_scale"],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_failure_category="none",
            frontend_check_status="已自动检查",
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_evidence_quality_score=88,
            frontend_location_scope="exact",
            frontend_location_exact=True,
            competitor_comparability="high",
            comparable_competitor_count=3,
        )
    ]

    assert showcase_validate._validate_product_final_decision_rows("UK", rows) == 0


def test_validate_product_final_decision_blocks_growth_without_verified_location(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0LOCVERIFY",
            today_allowed_actions=["observe", "bid_up"],
            today_blocked_actions=["budget_up", "broad_scale"],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_failure_category="none",
            frontend_check_status="已自动检查",
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_exact=True,
            frontend_location_verified=False,
            competitor_comparability="medium",
            comparable_competitor_count=3,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "allows growth without verified frontend location" in output


def test_validate_product_final_decision_blocks_growth_with_low_frontend_quality_score(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0LOWQUALITYGROWTH",
            today_allowed_actions=["observe", "bid_up"],
            today_blocked_actions=["budget_up", "broad_scale"],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_evidence_quality_score=60,
            frontend_failure_category="none",
            frontend_check_status="已自动检查",
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_verified=True,
            frontend_location_exact=True,
            competitor_comparability="high",
            comparable_competitor_count=3,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "allows growth without high frontend quality score" in output


def test_validate_product_final_decision_blocks_growth_without_competitor_comparability(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0NOCOMPGROWTH",
            today_allowed_actions=["observe", "bid_up"],
            today_blocked_actions=["budget_up", "broad_scale"],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_evidence_display_tier="强诊断可用",
            frontend_evidence_quality_score=88,
            frontend_failure_category="none",
            frontend_check_status="已自动检查",
            frontend_search_status="已自动检查",
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_verified=True,
            frontend_location_exact=True,
            competitor_comparability="",
            comparable_competitor_count=3,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "allows growth without competitor comparability" in output


def test_validate_product_final_decision_blocks_scope_exact_when_explicit_exact_false(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0LOCEXPLICIT",
            today_allowed_actions=["observe", "bid_up"],
            today_blocked_actions=["budget_up", "broad_scale"],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_failure_category="none",
            frontend_check_status="已自动检查",
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_verified=True,
            frontend_location_exact=False,
            competitor_comparability="medium",
            comparable_competitor_count=3,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "allows growth without exact frontend location" in output


def test_validate_product_final_decision_blocks_growth_with_cached_frontend() -> None:
    rows = [
        _product_decision_row(
            asin="B0CACHEBAD",
            today_allowed_actions=["bid_up"],
            today_blocked_actions=[],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_failure_category="none",
            frontend_check_status="沿用 2026-06-10 前台数据",
            frontend_cache_used=True,
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_exact=True,
            competitor_comparability="medium",
            comparable_competitor_count=3,
        )
    ]

    assert showcase_validate._validate_product_final_decision_rows("UK", rows) == 1


def test_validate_product_final_decision_splits_slash_joined_growth_actions(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0SLASHGROWTH",
            today_allowed_actions="bid_up/budget_up",
            today_blocked_actions=["broad_scale"],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_evidence_display_tier="强诊断可用",
            frontend_failure_category="none",
            frontend_check_status="沿用 2026-06-10 前台数据",
            frontend_cache_used=True,
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_verified=True,
            frontend_location_exact=True,
            competitor_comparability="medium",
            comparable_competitor_count=3,
            frontend_evidence_quality_score=88,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "allows growth without current frontend check" in output


def test_validate_product_final_decision_blocks_growth_with_partial_search(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0PARTIALSEARCH",
            today_allowed_actions=["bid_up"],
            today_blocked_actions=[],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_failure_category="none",
            frontend_check_status="已自动检查",
            frontend_search_status="已读取部分结果",
            frontend_search_partial_evidence=True,
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_exact=True,
            competitor_comparability="medium",
            comparable_competitor_count=3,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "allows growth with partial search evidence" in output


def test_validate_product_final_decision_blocks_growth_without_search_success(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0NOSEARCHOK",
            today_allowed_actions=["bid_up"],
            today_blocked_actions=[],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_failure_category="none",
            frontend_check_status="已自动检查",
            frontend_search_status="待前台检查",
            frontend_search_partial_evidence=False,
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_exact=True,
            competitor_comparability="medium",
            comparable_competitor_count=3,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "allows growth without successful search page" in output


def test_validate_product_final_decision_blocks_growth_with_currency_warning(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0CURRENCYBAD",
            today_allowed_actions=["bid_up"],
            today_blocked_actions=[],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_failure_category="none",
            frontend_check_status="已自动检查",
            frontend_price_currency_warning="价格币种异常：TWD594.77，已忽略",
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_exact=True,
            competitor_comparability="medium",
            comparable_competitor_count=3,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "allows growth with currency warning" in output


def test_validate_product_final_decision_blocks_growth_with_location_warning(capsys) -> None:
    rows = [
        _product_decision_row(
            asin="B0LOCATIONWARNBAD",
            today_allowed_actions=["bid_up"],
            today_blocked_actions=[],
            frontend_evidence_state="ok_high",
            frontend_evidence_tier="强诊断可用",
            frontend_failure_category="none",
            frontend_check_status="已自动检查",
            frontend_location_warning="UK 地区异常：United States",
            frontend_auto_conclusion="FRONTEND_OK",
            frontend_location_scope="exact",
            frontend_location_exact=True,
            competitor_comparability="medium",
            comparable_competitor_count=3,
        )
    ]

    code = showcase_validate._validate_product_final_decision_rows("UK", rows)
    output = capsys.readouterr().out

    assert code == 1
    assert "allows growth with location warning" in output


def _effective_review_metrics() -> dict[str, object]:
    return {
        "current_7d_clicks": 8,
        "current_7d_spend": 3.2,
        "current_7d_ad_orders": 2,
        "current_7d_promoted_ad_orders": 2,
        "current_7d_acos": 0.12,
        "current_7d_target_acos": 0.20,
        "current_7d_tacos": 0.08,
        "current_7d_total_orders": 4,
        "current_7d_available_stock": 18,
        "current_14d_ad_orders": 3,
        "current_14d_promoted_ad_orders": 3,
        "current_14d_acos": 0.14,
        "current_14d_total_orders": 7,
        "current_14d_tacos": 0.09,
        "current_14d_available_stock": 18,
        "review_data_source": "execution_anchored_daily",
        "pre_7d_start": "2026-05-13",
        "pre_7d_end": "2026-05-19",
        "post_3d_start": "2026-05-20",
        "post_3d_end": "2026-05-22",
        "post_7d_start": "2026-05-20",
        "post_7d_end": "2026-05-26",
        "pre_7d_promoted_ad_orders": 1,
        "pre_7d_total_orders": 2,
        "pre_7d_tacos": 0.10,
        "post_3d_days": 3,
        "post_3d_promoted_ad_orders": 1,
        "post_3d_total_orders": 2,
        "post_3d_acos": 0.13,
        "post_3d_tacos": 0.09,
        "post_3d_available_stock": 18,
        "post_7d_days": 7,
        "post_7d_promoted_ad_orders": 2,
        "post_7d_total_orders": 4,
        "post_7d_acos": 0.12,
        "post_7d_tacos": 0.08,
        "post_7d_available_stock": 18,
    }


def _write_minimal_safe_run(
    safe_dir: Path,
    latest_html: str,
    ad_rows: list[dict] | None = None,
    product_final_decision_rows: list[dict] | None = None,
    action_review_rows: list[dict] | None = None,
    keyword_review_rows: list[dict] | None = None,
    inventory_replenishment_rows: list[dict] | None = None,
    tomorrow_review_rows: list[dict] | None = None,
    listing_diagnosis_rows: list[dict] | None = None,
    cost_diagnosis_rows: list[dict] | None = None,
    excel_today_task_headers: list[str] | None = None,
    excel_today_task_rows: list[dict] | None = None,
    excel_market_today_task_headers: list[str] | None = None,
    excel_market_today_task_rows: list[dict] | None = None,
    excel_tomorrow_review_headers: list[str] | None = None,
    excel_tomorrow_review_rows: list[dict] | None = None,
    excel_market_tomorrow_review_headers: list[str] | None = None,
    excel_market_tomorrow_review_rows: list[dict] | None = None,
    excel_listing_diagnosis_headers: list[str] | None = None,
    excel_listing_diagnosis_rows: list[dict] | None = None,
    excel_market_listing_diagnosis_headers: list[str] | None = None,
    excel_market_listing_diagnosis_rows: list[dict] | None = None,
    excel_cost_diagnosis_headers: list[str] | None = None,
    excel_cost_diagnosis_rows: list[dict] | None = None,
    excel_market_cost_diagnosis_headers: list[str] | None = None,
    excel_market_cost_diagnosis_rows: list[dict] | None = None,
    excel_product_headers: list[str] | None = None,
    excel_product_rows: list[dict] | None = None,
    excel_market_product_headers: list[str] | None = None,
    excel_market_product_rows: list[dict] | None = None,
    excel_inventory_headers: list[str] | None = None,
    excel_inventory_rows: list[dict] | None = None,
    excel_market_inventory_headers: list[str] | None = None,
    excel_market_inventory_rows: list[dict] | None = None,
    excel_action_headers: list[str] | None = None,
    excel_keyword_headers: list[str] | None = None,
    excel_action_rows: list[dict] | None = None,
    excel_keyword_rows: list[dict] | None = None,
    excel_market_action_headers: list[str] | None = None,
    excel_market_keyword_headers: list[str] | None = None,
    excel_market_action_rows: list[dict] | None = None,
    excel_market_keyword_rows: list[dict] | None = None,
    write_market_review_sheets: bool = True,
    omit_snapshot_keys: set[str] | None = None,
    include_html_snapshot_tokens: bool = True,
    include_frontend_coverage_tokens: bool = True,
    include_dashboard_frontend_coverage_tokens: bool = True,
    include_marketplace_summary_frontend_coverage_tokens: bool = True,
    include_excel_frontend_coverage_tokens: bool = True,
    include_market_html_snapshot_tokens: bool = True,
    include_summary_snapshot_tokens: bool = True,
    include_summary_frontend_coverage_tokens: bool = True,
    sync_top_level_product_final_decision_rows: bool = True,
    sync_top_level_inventory_replenishment_rows: bool = True,
    top_level_product_final_decision_rows: list[dict] | None = None,
    top_level_inventory_replenishment_rows: list[dict] | None = None,
    sync_top_level_final_decision_summary: bool = True,
    sync_top_level_decision_gate_counts: bool = True,
    corrupt_snapshot_final_decision_summary: bool = False,
    corrupt_snapshot_decision_gate_counts: bool = False,
    fill_blank_review_action_ids: bool = True,
) -> None:
    safe_dir.mkdir()
    (safe_dir / "assets").mkdir()
    (safe_dir / "assets" / "report.css").write_text(".ad-task-card{} .ad-copy-box{}", encoding="utf-8")
    (safe_dir / "assets" / "report.js").write_text(
        "document.querySelector('[data-ad-complete-checkbox]');"
        "document.querySelector('[data-ad-filter-summary]');",
        encoding="utf-8",
    )
    report_date = "2026-06-08"

    def review_rows_with_default_action_ids(rows: list[dict] | None, prefix: str) -> list[dict] | None:
        if rows is None:
            return None
        normalized_rows: list[dict] = []
        is_keyword_review = "keyword" in prefix
        for idx, row in enumerate(rows, start=1):
            normalized = dict(row)
            normalized.setdefault("marketplace", "UK")
            normalized.setdefault("sku", f"SKU-REVIEW-{idx}")
            normalized.setdefault("asin", f"B0REVIEW{idx:03d}")
            if is_keyword_review:
                normalized.setdefault("search_term_or_target", f"review target {idx}")
            normalized.setdefault("normalized_action", "bid_down")
            normalized.setdefault("action_scope", "search_term" if is_keyword_review else "product")
            if (
                fill_blank_review_action_ids
                and "action_id" in normalized
                and not str(normalized.get("action_id") or "").strip()
            ):
                normalized = add_action_identity(normalized, normalized.get("normalized_action"), normalized["action_scope"])
            normalized_rows.append(normalized)
        return normalized_rows

    action_review_rows = review_rows_with_default_action_ids(action_review_rows, "action-review")
    keyword_review_rows = review_rows_with_default_action_ids(keyword_review_rows, "keyword-review")
    excel_action_rows = review_rows_with_default_action_ids(excel_action_rows, "excel-action-review")
    excel_keyword_rows = review_rows_with_default_action_ids(excel_keyword_rows, "excel-keyword-review")
    excel_market_action_rows = review_rows_with_default_action_ids(
        excel_market_action_rows,
        "market-action-review",
    )
    excel_market_keyword_rows = review_rows_with_default_action_ids(
        excel_market_keyword_rows,
        "market-keyword-review",
    )

    def task_rows_with_contract_defaults(rows: list[dict] | None) -> list[dict] | None:
        if rows is None:
            return None
        normalized_rows: list[dict] = []
        for idx, row in enumerate(rows, start=1):
            normalized = dict(row)
            normalized.setdefault("marketplace", "UK")
            normalized.setdefault("sku", f"SKU-TASK-{idx}")
            normalized.setdefault("asin", f"B0TASK{idx:04d}")
            normalized.setdefault("product_name", f"Task product {idx}")
            normalized.setdefault("priority", "P1")
            normalized.setdefault("issue_type", "广告动作")
            normalized.setdefault("action_group", "广告动作")
            copy_action = str(normalized.get("copy_action_line") or "")
            default_action = "否定精准" if "否" in copy_action else "观察"
            normalized.setdefault("today_action", default_action)
            normalized.setdefault("search_term_or_target", f"task target {idx}")
            normalized.setdefault("suggested_action", normalized.get("today_action") or "观察")
            normalized.setdefault("normalized_action", "negative_exact" if default_action == "否定精准" else "observe")
            if not str(normalized.get("action_id") or "").strip():
                normalized = add_action_identity(normalized, normalized.get("normalized_action"))
            else:
                normalized.setdefault("action_scope", "search_term")
            normalized_rows.append(normalized)
        return normalized_rows

    ad_rows = task_rows_with_contract_defaults(ad_rows)
    excel_today_task_rows = task_rows_with_contract_defaults(excel_today_task_rows)
    excel_market_today_task_rows = task_rows_with_contract_defaults(excel_market_today_task_rows)

    def rows_for_market(rows: list[dict] | None, market: str) -> list[dict]:
        return [
            row
            for row in rows or []
            if str(row.get("marketplace") or market).upper() == market
        ]

    omit_snapshot_keys = omit_snapshot_keys or set()

    def snapshot_for_market(market: str) -> dict:
        product_rows = rows_for_market(product_final_decision_rows, market)
        market_ad_rows = rows_for_market(ad_rows, market)
        decision_summary = {}
        for row in product_rows:
            final_decision = str(row.get("final_decision") or "").strip()
            decision_summary[final_decision] = decision_summary.get(final_decision, 0) + 1
        gate_counts = dict(decision_summary)
        if corrupt_snapshot_final_decision_summary and market == "UK":
            decision_summary = {"BROKEN": 999}
        if corrupt_snapshot_decision_gate_counts and market == "UK":
            gate_counts = {"BROKEN": 999}
        snapshot = {
            "today_task_queue_rows": market_ad_rows,
            "search_term_processing_queue_rows": [],
            "html_search_term_processing_queue_rows": [],
            "scale_rows": [],
            "scale_keyword_rows": [],
            "growth_test_rows": [],
            "frontend_check_queue_rows": [
                {
                    "marketplace": market,
                    "sku": f"SKU-{market}-FRONTEND",
                    "asin": f"B0{market}TEST",
                    "frontend_check_status": "待前台检查",
                    "frontend_data_freshness": "无可用前台数据",
                    "frontend_findings": "自动证据不足，不能用于强诊断",
                }
            ],
            "listing_price_diagnosis_rows": rows_for_market(listing_diagnosis_rows, market),
            "cost_profit_diagnosis_rows": rows_for_market(cost_diagnosis_rows, market),
            "inventory_replenishment_rows": rows_for_market(inventory_replenishment_rows, market),
            "tomorrow_review_rows": rows_for_market(tomorrow_review_rows, market),
            "today_action_groups": {"广告动作": market_ad_rows},
            "product_final_decision_rows": product_rows,
            "product_operation_cards": [_product_operation_card(row) for row in product_rows],
            "final_decision_summary": decision_summary,
            "decision_gate_counts": gate_counts,
            "frontend_coverage_summary": {},
            "action_effect_review_rows": rows_for_market(action_review_rows, market),
            "keyword_action_effect_review_rows": rows_for_market(keyword_review_rows, market),
        }
        snapshot["frontend_coverage_summary"] = daily_update._frontend_coverage_expected(
            snapshot["frontend_check_queue_rows"]
        )
        for key in omit_snapshot_keys:
            snapshot.pop(key, None)
        return snapshot

    snapshots_by_market = {market: snapshot_for_market(market) for market in ["UK", "US", "DE"]}
    product_operation_cards = [
        row
        for snapshot in snapshots_by_market.values()
        for row in snapshot.get("product_operation_cards", [])
        if isinstance(row, dict)
    ]
    enhanced_request_rows = [
        {
            "marketplace": "UK",
            "trigger_sku": "",
            "trigger_asin": "",
            "trigger_product_name": "",
            "issue_type": "增强数据文件导入",
            "report_type": "流量和销售数据",
            "period": "recent_vs_prior",
            "start_date": "",
            "end_date": "",
            "expected_filename": "traffic_sales_uk.xlsx",
            "target_path": "data/raw_amazon_custom/UK/",
            "target_folder": "data/raw_amazon_custom/UK/",
            "required": "否",
            "seller_central_page": "定制分析",
            "instruction": "已导入",
            "status": "已导入",
            "file_type": "traffic_sales",
            "format_type": "single",
            "detected_from": "",
            "detected_date_range": "",
            "freshness": "stale",
            "used_in_diagnosis": "否",
        }
    ]
    analysis = {
        "report_date": report_date,
        "import_summary": {"ads_imported_rows": 10, "erp_imported_rows": 10},
        "enhanced_data_requests": enhanced_request_rows,
        "product_final_decision_rows": top_level_product_final_decision_rows
        if top_level_product_final_decision_rows is not None
        else ((product_final_decision_rows or []) if sync_top_level_product_final_decision_rows else []),
        "inventory_replenishment_rows": top_level_inventory_replenishment_rows
        if top_level_inventory_replenishment_rows is not None
        else ((inventory_replenishment_rows or []) if sync_top_level_inventory_replenishment_rows else []),
        "product_operation_cards": product_operation_cards,
        "final_decision_summary": {
            market: snapshots_by_market[market].get("final_decision_summary", {})
            for market in ["UK", "US", "DE"]
        }
        if sync_top_level_final_decision_summary
        else {},
        "decision_gate_counts": {
            market: snapshots_by_market[market].get("decision_gate_counts", {})
            for market in ["UK", "US", "DE"]
        }
        if sync_top_level_decision_gate_counts
        else {},
        "marketplace_results": [
            {
                "marketplace": market,
                "summary": {
                    "marketplace": market,
                    "report_date": report_date,
                    "ads_row_count": 10 + index,
                    "erp_row_count": 20 + index,
                    "sku_count": 3 + index,
                    "asin_count": 4 + index,
                },
                "report_view_snapshot": snapshots_by_market[market],
            }
            for index, market in enumerate(["UK", "US", "DE"], start=1)
        ],
    }
    analysis["frontend_coverage_summary"] = showcase_validate.daily_frontend_coverage_aggregate_from_snapshots(
        analysis["marketplace_results"]
    )
    (safe_dir / "latest_analysis.json").write_text(json.dumps(analysis, ensure_ascii=False), encoding="utf-8")

    def snapshot_html_tokens() -> str:
        if not include_html_snapshot_tokens:
            return ""
        blocked_action_labels = {
            "bid_up": "加竞价",
            "budget_up": "加预算",
            "broad_scale": "放量",
            "create_exact_low_budget": "低预算精准测试",
        }
        tokens: list[str] = []
        product_articles: list[str] = []
        for row in (product_final_decision_rows or [])[:6]:
            marketplace = str(row.get("marketplace") or "").strip().upper()
            sku = str(row.get("sku") or "").strip()
            asin = str(row.get("asin") or "").strip().upper()
            article_tokens = [asin, str(row.get("product_name") or "")]
            reasons = row.get("frontend_blocking_reasons")
            if isinstance(reasons, list):
                article_tokens.extend(str(reason) for reason in reasons if str(reason).strip())
            elif str(reasons or "").strip():
                article_tokens.append(str(reasons))
            blocked_actions = row.get("today_blocked_actions")
            if isinstance(blocked_actions, list):
                article_tokens.extend(
                    blocked_action_labels[action]
                    for action in blocked_actions
                    if action in blocked_action_labels
                )
            product_articles.append(
                (
                    '<article class="work-card"'
                    f' data-product-decision-marketplace="{html.escape(marketplace, quote=True)}"'
                    f' data-product-decision-sku="{html.escape(sku, quote=True)}"'
                    f' data-product-decision-asin="{html.escape(asin, quote=True)}"'
                    f'{_product_decision_contract_attrs(row)}>'
                    + " ".join(html.escape(token) for token in article_tokens if token)
                    + "</article>"
                )
            )
        tokens.extend(product_articles)
        for row in action_review_rows or []:
            tokens.extend([str(row.get("asin") or ""), str(row.get("product_name") or "")])
            tokens.extend(showcase_validate._review_html_metric_tokens(row))
        for row in keyword_review_rows or []:
            tokens.extend([str(row.get("asin") or ""), str(row.get("search_term_or_target") or "")])
            tokens.extend(showcase_validate._review_html_metric_tokens(row))
        for row in tomorrow_review_rows or []:
            tokens.extend(
                [
                    str(row.get("search_term_or_target") or ""),
                    str(row.get("asin") or ""),
                    str(row.get("product_name") or ""),
                ]
            )
        for row in ad_rows or []:
            tokens.extend(
                [
                    str(row.get("search_term_or_target") or ""),
                    str(row.get("asin") or ""),
                    str(row.get("product_name") or ""),
                ]
            )
        return " " + " ".join(token for token in tokens if token)

    def frontend_coverage_html_tokens() -> str:
        if not include_frontend_coverage_tokens:
            return ""
        return " " + " ".join(daily_update._frontend_coverage_display_tokens(analysis))

    def ad_completion_html_tokens(market: str | None = None) -> str:
        parts: list[str] = []
        for row in rows_for_market(ad_rows, market) if market else (ad_rows or []):
            if str(row.get("confirmed_status") or "") in {"已执行", "已核查", "已忽略", "仅背景参考"}:
                continue
            if str(row.get("normalized_action") or "") in {"", "observe"}:
                continue
            identified = add_action_identity(
                row,
                row.get("suggested_action")
                or row.get("today_action")
                or row.get("copy_action_line")
                or row.get("normalized_action"),
                row.get("action_scope") or None,
            )
            payload = {
                **identified,
                "manual_action_taken": identified.get("suggested_action") or identified.get("today_action") or "",
                "confirmed_note": "网页勾选已完成；未满3天不判断，3天初查，7天正式复盘。",
            }
            action_id = str(payload.get("action_id") or "")
            payload_attr = html.escape(
                json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
                quote=True,
            )
            parts.append(
                f'复制到广告后台 <div data-ad-complete-payload="{payload_attr}" data-action-id="{html.escape(action_id, quote=True)}"></div>'
            )
        return " " + " ".join(parts) if parts else ""

    latest_html = (
        latest_html
        + ' 产品级结论 执行后效果复盘 id="product-operation-cards" '
        + 'id="frontend-evidence-status" id="action-effect-review" '
        + snapshot_html_tokens()
        + frontend_coverage_html_tokens()
        + (
            ad_completion_html_tokens()
            if include_html_snapshot_tokens
            and "复制到广告后台" in latest_html
            and "data-ad-complete-payload" not in latest_html
            else ""
        )
    )

    def dashboard_html() -> str:
        rows = []
        for result in analysis["marketplace_results"]:
            summary = result["summary"]
            rows.append(
                f"<tr><td>{summary['marketplace']}</td><td>{summary['ads_row_count']}</td>"
                f"<td>{summary['erp_row_count']}</td><td>{summary['sku_count']}</td>"
                f"<td>{summary['asin_count']}</td><td>正式分析</td><td>无</td></tr>"
                f"<a href=\"{str(summary['marketplace']).lower()}_report.html\">打开报告</a>"
            )
        tokens: list[str] = []
        if include_dashboard_frontend_coverage_tokens:
            tokens.extend(daily_update._frontend_coverage_display_tokens(analysis, include_title=True))
        return (
            "2026-06-08 运营状态入口 打开三分钟摘要 打开 ALL 运营控制台 "
            "latest_recommendations.html summary.html "
            + " ".join(tokens)
            + "".join(rows)
        )

    def market_html_tokens(market: str) -> str:
        if not include_market_html_snapshot_tokens:
            return ""
        blocked_action_labels = {
            "bid_up": "加竞价",
            "budget_up": "加预算",
            "broad_scale": "放量",
        }
        tokens: list[str] = []
        for row in rows_for_market(product_final_decision_rows, market):
            tokens.extend([str(row.get("asin") or ""), str(row.get("product_name") or "")])
            reasons = row.get("frontend_blocking_reasons")
            if isinstance(reasons, list):
                tokens.extend(str(reason) for reason in reasons if str(reason).strip())
            elif str(reasons or "").strip():
                tokens.append(str(reasons))
            blocked_actions = row.get("today_blocked_actions")
            if isinstance(blocked_actions, list):
                tokens.extend(
                    blocked_action_labels[action]
                    for action in blocked_actions
                    if action in blocked_action_labels
                )
        for row in rows_for_market(action_review_rows, market):
            tokens.extend([str(row.get("asin") or ""), str(row.get("product_name") or "")])
            tokens.extend(showcase_validate._review_html_metric_tokens(row))
        for row in rows_for_market(keyword_review_rows, market):
            tokens.extend([str(row.get("asin") or ""), str(row.get("search_term_or_target") or "")])
            tokens.extend(showcase_validate._review_html_metric_tokens(row))
        coverage = snapshots_by_market[market].get("frontend_coverage_summary") or {}
        if isinstance(coverage, dict):
            tokens.extend(daily_update._frontend_coverage_display_tokens({"frontend_coverage_summary": coverage}))
        return " " + " ".join(token for token in tokens if token)

    def summary_html_tokens() -> str:
        tokens: list[str] = []
        if include_summary_snapshot_tokens:
            for row in action_review_rows or []:
                tokens.append(str(row.get("product_name") or ""))
                if showcase_validate.daily_summary_review_needs_display_guard(row):
                    tokens.append(showcase_validate.daily_summary_review_display_judgement(row))
                    tokens.append(showcase_validate.daily_summary_review_display_next_step(row))
                tokens.extend(showcase_validate._summary_review_metric_tokens(row))
            for row in keyword_review_rows or []:
                tokens.append(str(row.get("search_term_or_target") or row.get("product_name") or ""))
                if showcase_validate.daily_summary_review_needs_display_guard(row):
                    tokens.append(showcase_validate.daily_summary_review_display_judgement(row))
                    tokens.append(showcase_validate.daily_summary_review_display_next_step(row))
                tokens.extend(showcase_validate._summary_review_metric_tokens(row))
            for row in inventory_replenishment_rows or []:
                tokens.extend([str(row.get("product_name") or ""), str(row.get("stock_status_label") or "")])
        if include_summary_frontend_coverage_tokens:
            tokens.extend(daily_update._frontend_coverage_display_tokens(analysis))
        return " " + " ".join(token for token in tokens if token)

    for name, text in {
        "latest_recommendations.html": latest_html,
        "dashboard.html": dashboard_html(),
        "summary.html": "2026-06-08 三分钟摘要 今日开工结论 今天先做 昨天动作今天要盯" + summary_html_tokens(),
        "uk_report.html": (
            '2026-06-08 亚马逊运营日报｜UK 广告状态 数据质量与增强数据 站点状态 市场调查 '
            '执行后效果复盘 id="product-operation-cards" id="frontend-evidence-status" id="action-effect-review"'
            + market_html_tokens("UK")
            + ad_completion_html_tokens("UK")
        ),
        "us_report.html": (
            '2026-06-08 亚马逊运营日报｜US 广告状态 数据质量与增强数据 站点状态 市场调查 '
            '执行后效果复盘 id="product-operation-cards" id="frontend-evidence-status" id="action-effect-review"'
            + market_html_tokens("US")
            + ad_completion_html_tokens("US")
        ),
        "de_report.html": (
            '2026-06-08 亚马逊运营日报｜DE 广告状态 数据质量与增强数据 站点状态 市场调查 '
            '执行后效果复盘 id="product-operation-cards" id="frontend-evidence-status" id="action-effect-review"'
            + market_html_tokens("DE")
            + ad_completion_html_tokens("DE")
        ),
    }.items():
        (safe_dir / name).write_text(text, encoding="utf-8")
    (safe_dir / "latest_recommendations.md").write_text(
        "# 亚马逊运营日报汇总建议\n\n"
        "报告日期 2026-06-08\n\n"
        "## 1. 各站点状态摘要\n\n"
        "## 2. 今日动作清单\n\n"
        "## 3. 明日复查清单\n",
        encoding="utf-8",
    )
    marketplace_summary_lines = [
        "# 站点汇总｜2026-06-08",
        "",
        "| 站点 | 广告行数 | ERP行数 | SKU数 | ASIN数 | 状态 | 说明 |",
        "| --- | ---: | ---: | ---: | ---: | --- | --- |",
        *[
            f"| {result['summary']['marketplace']} | {result['summary']['ads_row_count']} | "
            f"{result['summary']['erp_row_count']} | {result['summary']['sku_count']} | "
            f"{result['summary']['asin_count']} | 正式分析 | 无 |"
            for result in analysis["marketplace_results"]
        ],
        "",
    ]
    if include_marketplace_summary_frontend_coverage_tokens:
        coverage_rows: list[str] = []
        for market, snapshot in snapshots_by_market.items():
            coverage = snapshot.get("frontend_coverage_summary") or {}
            if not isinstance(coverage, dict):
                continue
            total = int(coverage.get("frontend_queue_total") or 0)
            if not total:
                continue
            product_success = int(coverage.get("frontend_product_page_success_count") or coverage.get("frontend_live_success_count") or 0)
            own_seller = int(coverage.get("frontend_own_sellersprite_count") or 0)
            own_seller_today = int(coverage.get("frontend_own_sellersprite_today_count") or 0)
            own_seller_cache = int(coverage.get("frontend_own_sellersprite_cache_count") or 0)
            sellersprite_trend = int(coverage.get("frontend_sellersprite_trend_ready_count") or 0)
            competitor_discovery = int(coverage.get("frontend_competitor_discovery_count") or 0)
            competitor_pool = int(coverage.get("frontend_competitor_pool_count") or 0)
            competitor_pool_today = int(coverage.get("frontend_competitor_pool_today_count") or 0)
            competitor_pool_cache = int(coverage.get("frontend_competitor_pool_cache_count") or 0)
            competitor_seller = int(coverage.get("frontend_competitor_sellersprite_count") or 0)
            competitor_seller_asins = int(coverage.get("frontend_competitor_sellersprite_asin_count") or 0)
            competitor_seller_today = int(coverage.get("frontend_competitor_sellersprite_today_count") or 0)
            competitor_seller_cache = int(coverage.get("frontend_competitor_sellersprite_cache_count") or 0)
            amazon_validation = int(coverage.get("frontend_amazon_search_validation_count") or coverage.get("frontend_competitor_search_success_count") or coverage.get("frontend_search_success_count") or 0)
            scalable = int(coverage.get("frontend_scalable_strong_count") or 0)
            weak = int(coverage.get("frontend_weak_defensive_count") or 0)
            insufficient = int(coverage.get("frontend_insufficient_count") or 0)
            coverage_rows.append(
                f"| {market} | {total} | {product_success}/{total} | 今日 {own_seller_today}/{total}，缓存 {own_seller_cache}/{total} | "
                f"{sellersprite_trend}/{total} | {competitor_discovery}/{total} | 今日 {competitor_pool_today}/{total}，7天缓存 {competitor_pool_cache}/{total} | "
                f"今日 {competitor_seller_today}/{total}，缓存 {competitor_seller_cache}/{total} | "
                f"{amazon_validation}/{total} | "
                f"{scalable}/{total} | {weak}/{total} | {insufficient}/{total} |"
            )
        expected = daily_update._frontend_coverage_excel_expected(analysis)
        frontend_total = expected.get("ALL 前台队列")
        if frontend_total:
            marketplace_summary_lines.extend(
                [
                    "## 前台证据覆盖",
                    "",
                    "| 站点 | 前台队列 | 产品页成功 | 卖家精灵自己 ASIN | 卖家精灵趋势 | 卖家精灵竞品发现 | 卖家精灵竞品池 | 竞品 ASIN 反查 | Amazon 搜索页辅助验证 | 达到放量准入 | 弱势止损证据 | 证据不足 |",
                    "| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
                    *coverage_rows,
                    f"| ALL | {frontend_total} | {expected['ALL 产品页成功']} | {expected['ALL 卖家精灵自己 ASIN']} | "
                    f"{expected['ALL 卖家精灵趋势']} | "
                    f"{expected['ALL 卖家精灵竞品发现']} | "
                    f"{expected['ALL 卖家精灵竞品池']} | {expected['ALL 竞品 ASIN 反查']} | "
                    f"{expected['ALL Amazon 搜索页辅助验证']} | "
                    f"{expected['ALL 达到放量准入']} | {expected['ALL 弱势止损证据']} | {expected['ALL 证据不足']} |",
                    "",
                    " ".join(daily_update._frontend_coverage_display_tokens(analysis, include_title=True)),
                    "",
                ]
            )
    (safe_dir / "marketplace_summary.md").write_text("\n".join(marketplace_summary_lines), encoding="utf-8")
    from openpyxl import Workbook

    enhanced_workbook = Workbook()
    enhanced_sheet = enhanced_workbook.active
    enhanced_sheet.title = "Sheet1"
    enhanced_headers = list(enhanced_request_rows[0].keys())
    enhanced_sheet.append(enhanced_headers)
    for row in enhanced_request_rows:
        enhanced_sheet.append([row.get(header, "") for header in enhanced_headers])
    enhanced_workbook.save(safe_dir / "enhanced_data_requests.xlsx")
    enhanced_workbook.close()
    (safe_dir / "enhanced_data_requests.md").write_text(
        "# 需要补充导出的增强数据｜2026-06-08\n\n"
        "| 站点 | 状态 | 报表类型 | 周期 | 日期范围 | 导出后文件名 | 目标文件夹 | 必需 |\n"
        "|---|---|---|---|---|---|---|---|\n"
        "| UK | 已导入 | 流量和销售数据 | recent_vs_prior | N/A ~ N/A | traffic_sales_uk.xlsx | data/raw_amazon_custom/UK/ | 否 |\n",
        encoding="utf-8",
    )

    workbook = Workbook()
    workbook.active.title = "Metrics_Validation"
    overview = workbook.create_sheet("总览")
    overview.append(["分组", "指标", "值"])
    overview.append(["本次运行信息", "报告日期", report_date])
    if include_excel_frontend_coverage_tokens:
        for metric, value in daily_update._frontend_coverage_excel_expected(analysis).items():
            overview.append(["前台证据覆盖", metric, value])
    for market in ["UK", "US", "DE"]:
        sheet = workbook.create_sheet(f"{market}_今日总览")
        sheet.append(["marketplace", "report_date"])
        sheet.append([market, report_date])

    def append_sheet_rows(title: str, rows: list[dict], headers: list[str] | None) -> None:
        rows = rows or []
        sheet = workbook[title] if title in workbook.sheetnames else workbook.create_sheet(title)
        selected_headers = headers or (list(rows[0].keys()) if rows else [])
        sheet.append(selected_headers)
        for row in rows:
            sheet.append(
                [
                    json.dumps(row.get(header), ensure_ascii=False)
                    if isinstance(row.get(header), (list, dict))
                    else row.get(header)
                    for header in selected_headers
                ]
            )

    product_headers = list(showcase_validate.PRODUCT_FINAL_DECISION_REQUIRED_FIELDS)
    inventory_headers = [
        "marketplace",
        "sku",
        "asin",
        *showcase_validate.INVENTORY_REPLENISHMENT_EXCEL_CONTENT_FIELDS,
    ]
    append_sheet_rows(
        "产品最终决策",
        product_final_decision_rows if excel_product_rows is None else excel_product_rows,
        excel_product_headers or product_headers,
    )
    product_operation_headers = [
        *daily_update.PRODUCT_OPERATION_CARD_IDENTITY_FIELDS,
        *daily_update.PRODUCT_OPERATION_CARD_CONTENT_FIELDS,
    ]
    append_sheet_rows(
        "产品运营卡",
        product_operation_cards,
        product_operation_headers,
    )
    append_sheet_rows(
        "库存补货提醒",
        inventory_replenishment_rows if excel_inventory_rows is None else excel_inventory_rows,
        excel_inventory_headers or inventory_headers,
    )
    today_headers = [
        *daily_update.TODAY_TASK_QUEUE_IDENTITY_FIELDS,
        *daily_update.TODAY_TASK_QUEUE_CONTENT_FIELDS,
    ]
    append_sheet_rows(
        "今日动作清单",
        ad_rows if excel_today_task_rows is None else excel_today_task_rows,
        excel_today_task_headers or (list(ad_rows[0].keys()) if ad_rows else today_headers),
    )
    for market in ["UK", "US", "DE"]:
        market_ad_rows = rows_for_market(ad_rows, market)
        market_excel_rows = (
            market_ad_rows
            if excel_market_today_task_rows is None
            else rows_for_market(excel_market_today_task_rows, market)
        )
        if market_ad_rows:
            append_sheet_rows(
                f"{market}_今日动作清单",
                market_excel_rows,
                excel_market_today_task_headers or list(market_ad_rows[0].keys()),
            )
    tomorrow_headers = [
        *daily_update.TOMORROW_REVIEW_IDENTITY_FIELDS,
        *daily_update.TOMORROW_REVIEW_CONTENT_FIELDS,
    ]
    append_sheet_rows(
        "明日复查清单",
        tomorrow_review_rows if excel_tomorrow_review_rows is None else excel_tomorrow_review_rows,
        excel_tomorrow_review_headers or (list(tomorrow_review_rows[0].keys()) if tomorrow_review_rows else tomorrow_headers),
    )
    for market in ["UK", "US", "DE"]:
        market_review_rows = rows_for_market(tomorrow_review_rows, market)
        market_excel_rows = (
            market_review_rows
            if excel_market_tomorrow_review_rows is None
            else rows_for_market(excel_market_tomorrow_review_rows, market)
        )
        if market_review_rows:
            append_sheet_rows(
                f"{market}_明日复查清单",
                market_excel_rows,
                excel_market_tomorrow_review_headers or list(market_review_rows[0].keys()),
            )
    if listing_diagnosis_rows:
        append_sheet_rows(
            "Listing待确认",
            listing_diagnosis_rows if excel_listing_diagnosis_rows is None else excel_listing_diagnosis_rows,
            excel_listing_diagnosis_headers or list(listing_diagnosis_rows[0].keys()),
        )
        for market in ["UK", "US", "DE"]:
            market_listing_rows = rows_for_market(listing_diagnosis_rows, market)
            if market_listing_rows:
                market_excel_rows = (
                    market_listing_rows
                    if excel_market_listing_diagnosis_rows is None
                    else rows_for_market(excel_market_listing_diagnosis_rows, market)
                )
                append_sheet_rows(
                    f"{market}_Listing待确认",
                    market_excel_rows,
                    excel_market_listing_diagnosis_headers or list(market_listing_rows[0].keys()),
                )
    if cost_diagnosis_rows:
        append_sheet_rows(
            "成本利润诊断",
            cost_diagnosis_rows if excel_cost_diagnosis_rows is None else excel_cost_diagnosis_rows,
            excel_cost_diagnosis_headers or list(cost_diagnosis_rows[0].keys()),
        )
        for market in ["UK", "US", "DE"]:
            market_cost_rows = rows_for_market(cost_diagnosis_rows, market)
            if market_cost_rows:
                market_excel_rows = (
                    market_cost_rows
                    if excel_market_cost_diagnosis_rows is None
                    else rows_for_market(excel_market_cost_diagnosis_rows, market)
                )
                append_sheet_rows(
                    f"{market}_成本利润诊断",
                    market_excel_rows,
                    excel_market_cost_diagnosis_headers or list(market_cost_rows[0].keys()),
                )
    if product_final_decision_rows:
        for market in ["UK", "US", "DE"]:
            market_product_rows = rows_for_market(product_final_decision_rows, market)
            if market_product_rows:
                market_excel_rows = (
                    market_product_rows
                    if excel_market_product_rows is None
                    else rows_for_market(excel_market_product_rows, market)
                )
                append_sheet_rows(
                    f"{market}_产品最终决策",
                    market_excel_rows,
                    excel_market_product_headers or list(market_product_rows[0].keys()),
                )
    if product_operation_cards:
        for market in ["UK", "US", "DE"]:
            market_operation_rows = rows_for_market(product_operation_cards, market)
            if market_operation_rows:
                append_sheet_rows(
                    f"{market}_产品运营卡",
                    market_operation_rows,
                    product_operation_headers,
                )
    if inventory_replenishment_rows:
        for market in ["UK", "US", "DE"]:
            market_inventory_rows = rows_for_market(inventory_replenishment_rows, market)
            if market_inventory_rows:
                market_excel_rows = (
                    market_inventory_rows
                    if excel_market_inventory_rows is None
                    else rows_for_market(excel_market_inventory_rows, market)
                )
                append_sheet_rows(
                    f"{market}_库存补货提醒",
                    market_excel_rows,
                    excel_market_inventory_headers or list(market_inventory_rows[0].keys()),
                )
    frontend_rows = [
        row
        for snapshot in snapshots_by_market.values()
        for row in snapshot.get("frontend_check_queue_rows", [])
        if isinstance(row, dict)
    ]
    frontend_headers = [
        *daily_update.FRONTEND_QUEUE_IDENTITY_FIELDS,
        *daily_update.FRONTEND_QUEUE_CONTENT_FIELDS,
    ]
    append_sheet_rows("前台证据队列", frontend_rows, frontend_headers)
    for market in ["UK", "US", "DE"]:
        market_frontend_rows = rows_for_market(frontend_rows, market)
        append_sheet_rows(f"{market}_前台证据队列", market_frontend_rows, frontend_headers)
    if action_review_rows:
        append_sheet_rows(
            "执行后效果复盘",
            action_review_rows if excel_action_rows is None else excel_action_rows,
            excel_action_headers or list(action_review_rows[0].keys()),
        )
    if keyword_review_rows:
        append_sheet_rows(
            "词级执行复盘",
            keyword_review_rows if excel_keyword_rows is None else excel_keyword_rows,
            excel_keyword_headers or list(keyword_review_rows[0].keys()),
        )
    if write_market_review_sheets:
        for market in ["UK", "US", "DE"]:
            market_action_rows = rows_for_market(action_review_rows, market)
            if market_action_rows:
                market_excel_rows = (
                    market_action_rows
                    if excel_market_action_rows is None
                    else rows_for_market(excel_market_action_rows, market)
                )
                append_sheet_rows(
                    f"{market}_执行后复盘",
                    market_excel_rows,
                    excel_market_action_headers or list(market_action_rows[0].keys()),
                )
            market_keyword_rows = rows_for_market(keyword_review_rows, market)
            if market_keyword_rows:
                market_excel_rows = (
                    market_keyword_rows
                    if excel_market_keyword_rows is None
                    else rows_for_market(excel_market_keyword_rows, market)
                )
                append_sheet_rows(
                    f"{market}_词级执行复盘",
                    market_excel_rows,
                    excel_market_keyword_headers or list(market_keyword_rows[0].keys()),
                )
    workbook.save(safe_dir / "amazon_ops_report_20260608.xlsx")
    workbook.save(safe_dir / f"amazon_ops_report_{report_date}.xlsx")

    date_token = report_date.replace("-", "")
    action_rows = action_review_rows or []
    keyword_rows = keyword_review_rows or []
    final_decision_rows = analysis.get("product_final_decision_rows") or []

    def collapsed_counts(value: object) -> dict[str, int]:
        counts: dict[str, int] = {}
        if not isinstance(value, dict):
            return counts
        for market_counts in value.values():
            if not isinstance(market_counts, dict):
                continue
            for key, count in market_counts.items():
                counts[str(key)] = counts.get(str(key), 0) + int(count or 0)
        return counts

    autoopt_payload = {
        "report_date": report_date,
        "rows": [],
        "action_review_rows": action_rows,
        "keyword_action_review_rows": keyword_rows,
        "product_final_decisions": final_decision_rows,
        "final_decision_summary": collapsed_counts(analysis.get("final_decision_summary")),
        "decision_gate_counts": collapsed_counts(analysis.get("decision_gate_counts")),
        "learned_rules": [],
        "manual_learning_rows": [],
        "positive_action_patterns": [],
        "negative_action_patterns": [],
        "product_strategy_profiles": [],
        "keyword_strategy_memory": [],
    }
    (safe_dir / f"autoopt_log_{date_token}.json").write_text(
        json.dumps(autoopt_payload, ensure_ascii=False),
        encoding="utf-8",
    )
    for filename, payload in {
        f"action_review_{date_token}.json": action_rows,
        f"keyword_action_review_{date_token}.json": keyword_rows,
        f"learned_rules_{date_token}.json": [],
        f"manual_learning_log_{date_token}.json": [],
        f"product_strategy_profiles_{date_token}.json": [],
        f"keyword_strategy_memory_{date_token}.json": [],
    }.items():
        (safe_dir / filename).write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    (safe_dir / f"self_optimization_log_{date_token}.json").write_text(
        json.dumps(
            {
                "report_date": report_date,
                "learned_rules": [],
                "manual_learning_rows": [],
                "action_review_rows": action_rows,
                "keyword_action_review_rows": keyword_rows,
                "positive_action_patterns": [],
                "negative_action_patterns": [],
                "product_strategy_profiles": [],
                "keyword_strategy_memory": [],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    autoopt_workbook = Workbook()
    autoopt_workbook.active.title = "autoopt_log"
    autoopt_workbook["autoopt_log"].append(["placeholder"])
    summary = autoopt_workbook.create_sheet("summary")
    summary.append(["report_date"])
    summary.append([report_date])

    def append_autoopt_sheet(title: str, rows: list[dict], headers: list[str]) -> None:
        sheet = autoopt_workbook.create_sheet(title)
        sheet.append(headers)
        for row in rows:
            sheet.append(
                [
                    json.dumps(row.get(header), ensure_ascii=False)
                    if isinstance(row.get(header), (list, dict))
                    else row.get(header)
                    for header in headers
                ]
            )

    append_autoopt_sheet("action_review", action_rows, showcase_validate.ACTION_REVIEW_REQUIRED_FIELDS)
    append_autoopt_sheet("keyword_action_review", keyword_rows, showcase_validate.KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS)
    append_autoopt_sheet("final_decisions", final_decision_rows, product_headers)
    autoopt_workbook.save(safe_dir / f"autoopt_{date_token}.xlsx")


def test_validate_safe_run_blocks_generic_confirmation_copy(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 需要确认的问题',
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "forbidden marker" in output


def test_validate_safe_run_blocks_missing_enhanced_request_outputs(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
    )
    (safe_dir / "enhanced_data_requests.md").unlink()

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "safe-run missing files: enhanced_data_requests.md" in output


def test_validate_safe_run_blocks_missing_core_snapshot_sections(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        omit_snapshot_keys={"product_final_decision_rows", "keyword_action_effect_review_rows", "product_operation_cards"},
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "report_view_snapshot missing" in output
    assert "product_final_decision_rows" in output
    assert "product_operation_cards" in output
    assert "keyword_action_effect_review_rows" in output


def test_validate_safe_run_blocks_frontend_coverage_summary_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
    )
    analysis_path = safe_dir / "latest_analysis.json"
    payload = json.loads(analysis_path.read_text(encoding="utf-8"))
    snapshot = payload["marketplace_results"][0]["report_view_snapshot"]
    snapshot["frontend_coverage_summary"]["frontend_decision_ready_count"] = 1
    analysis_path.write_text(json.dumps(payload), encoding="utf-8")

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "frontend_coverage_summary field frontend_decision_ready_count" in output


def test_validate_safe_run_blocks_top_level_frontend_coverage_summary_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
    )
    analysis_path = safe_dir / "latest_analysis.json"
    payload = json.loads(analysis_path.read_text(encoding="utf-8"))
    payload["frontend_coverage_summary"]["frontend_queue_total"] = 999
    analysis_path.write_text(json.dumps(payload), encoding="utf-8")

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "latest_analysis.json top-level frontend_coverage_summary field frontend_queue_total" in output


def test_validate_safe_run_blocks_latest_recommendations_missing_frontend_coverage_label(
    capsys,
    tmp_path,
) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        include_frontend_coverage_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "latest_recommendations.html missing frontend coverage token 产品页成功" in output


def test_validate_safe_run_blocks_top_level_product_decision_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(marketplace="UK", asin="B0TOPLEVEL1")
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
        sync_top_level_product_final_decision_rows=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "top-level product_final_decision_rows count mismatch" in output
    assert "expected 1 from marketplace snapshots, got 0" in output


def test_validate_safe_run_blocks_product_decision_blank_marketplace(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="",
        asin="B0BLANKDECISION",
        product_name="Blank marketplace decision",
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "UK product final decision row 1 B0BLANKDECISION missing marketplace value" in output


def test_validate_safe_run_blocks_top_level_inventory_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    inventory_row = {
        "marketplace": "UK",
        "sku": "SKU-INV",
        "asin": "B0TOPINV1",
        "product_name": "Top level stock check",
        "stock_risk_level": "LOW_STOCK",
        "stock_status_label": "低库存",
        "available_stock": 2,
        "days_of_cover": 3,
        "recommended_reorder_qty": 80,
        "replenishment_advice": "进入补货窗口。",
    }
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        inventory_replenishment_rows=[inventory_row],
        sync_top_level_inventory_replenishment_rows=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "top-level inventory_replenishment_rows count mismatch" in output
    assert "expected 1 from marketplace snapshots, got 0" in output


def test_validate_safe_run_blocks_inventory_blank_marketplace(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    inventory_row = {
        "marketplace": "",
        "sku": "SKU-BLANK-INV",
        "asin": "B0BLANKINV",
        "product_name": "Blank marketplace inventory",
        "stock_risk_level": "LOW_STOCK",
        "stock_status_label": "低库存",
        "available_stock": 2,
        "days_of_cover": 3,
        "recommended_reorder_qty": 80,
        "replenishment_advice": "进入补货窗口。",
    }
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        inventory_replenishment_rows=[inventory_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "UK inventory replenishment row 1 Blank marketplace inventory missing marketplace value" in output


def test_validate_safe_run_blocks_top_level_product_decision_content_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        asin="B0TOPCONTENT",
        final_decision="EXECUTE_TODAY",
        final_decision_label="今天执行",
        today_allowed_actions=["bid_down"],
    )
    top_level_row = dict(decision_row)
    top_level_row["final_decision"] = "WAIT_REVIEW"
    top_level_row["final_decision_label"] = "等复盘"
    top_level_row["today_allowed_actions"] = ["observe"]
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
        top_level_product_final_decision_rows=[top_level_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "latest_analysis.json top-level product_final_decision_rows field mismatch for product final decisions" in output
    assert "field final_decision: expected 'EXECUTE_TODAY' from marketplace snapshots, got 'WAIT_REVIEW'" in output


def test_validate_safe_run_blocks_top_level_inventory_content_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    inventory_row = {
        "marketplace": "UK",
        "sku": "SKU-INV",
        "asin": "B0TOPINVCONTENT",
        "product_name": "Top stock check",
        "stock_risk_level": "URGENT_REORDER",
        "stock_status_label": "紧急补货",
        "available_stock": 2,
        "days_of_cover": 3,
        "recommended_reorder_qty": 80,
        "replenishment_advice": "进入补货窗口。",
    }
    top_level_row = dict(inventory_row)
    top_level_row["stock_risk_level"] = "HEALTHY"
    top_level_row["stock_status_label"] = "健康"
    top_level_row["available_stock"] = 200
    top_level_row["days_of_cover"] = 365
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        inventory_replenishment_rows=[inventory_row],
        top_level_inventory_replenishment_rows=[top_level_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "latest_analysis.json top-level inventory_replenishment_rows field mismatch for inventory replenishment rows" in output
    assert "field available_stock: expected 2.0 from marketplace snapshots, got 200.0" in output


def test_validate_safe_run_blocks_inventory_replenishment_excel_missing_fields(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    inventory_row = {
        "marketplace": "UK",
        "sku": "SKU-INV",
        "asin": "B0EXCELINV1",
        "product_name": "Excel stock check",
        "stock_risk_level": "LOW_STOCK",
        "stock_status_label": "低库存",
        "available_stock": 2,
        "days_of_cover": 3,
        "recommended_reorder_qty": 80,
        "replenishment_advice": "进入补货窗口。",
    }
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        inventory_replenishment_rows=[inventory_row],
        excel_inventory_headers=["marketplace", "sku", "asin", "product_name"],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 库存补货提醒 missing fields" in output
    assert "available_stock" in output


def test_validate_safe_run_blocks_inventory_replenishment_excel_identity_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    inventory_row = {
        "marketplace": "UK",
        "sku": "SKU-INV",
        "asin": "B0EXCELINV1",
        "product_name": "Excel stock check",
        "stock_risk_level": "LOW_STOCK",
        "stock_status_label": "低库存",
        "available_stock": 2,
        "days_of_cover": 3,
        "recommended_reorder_qty": 80,
        "replenishment_advice": "进入补货窗口。",
    }
    excel_row = dict(inventory_row)
    excel_row["asin"] = "B0EXCELINV2"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        inventory_replenishment_rows=[inventory_row],
        excel_inventory_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 库存补货提醒 row identity mismatch" in output
    assert "B0EXCELINV1" in output
    assert "B0EXCELINV2" in output


def test_validate_safe_run_blocks_inventory_replenishment_excel_content_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    inventory_row = {
        "marketplace": "UK",
        "sku": "SKU-INV",
        "asin": "B0EXCELINV1",
        "product_name": "Excel stock check",
        "stock_risk_level": "LOW_STOCK",
        "stock_status_label": "低库存",
        "available_stock": 2,
        "days_of_cover": 3,
        "recommended_reorder_qty": 80,
        "replenishment_advice": "进入补货窗口。",
    }
    excel_row = dict(inventory_row)
    excel_row["stock_risk_level"] = "HEALTHY"
    excel_row["days_of_cover"] = 180
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        inventory_replenishment_rows=[inventory_row],
        excel_inventory_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 库存补货提醒 field mismatch for inventory replenishment rows" in output
    assert "field days_of_cover: expected 3.0, got 180.0" in output


def test_validate_safe_run_blocks_market_inventory_excel_wrong_marketplace(capsys, tmp_path) -> None:
    from openpyxl import load_workbook

    safe_dir = tmp_path / "safe"
    inventory_row = {
        "marketplace": "UK",
        "sku": "SKU-INV-MARKET",
        "asin": "B0MARKETINV",
        "product_name": "Market inventory product",
        "stock_risk_level": "LOW_STOCK",
        "stock_status_label": "低库存",
        "available_stock": 2,
        "days_of_cover": 3,
        "recommended_reorder_qty": 80,
        "replenishment_advice": "进入补货窗口。",
    }
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        inventory_replenishment_rows=[inventory_row],
    )
    for workbook_path in safe_dir.glob("amazon_ops_report_*.xlsx"):
        workbook = load_workbook(workbook_path)
        sheet = workbook["UK_库存补货提醒"]
        headers = [cell.value for cell in sheet[1]]
        sheet.cell(row=2, column=headers.index("marketplace") + 1).value = "US"
        workbook.save(workbook_path)
        workbook.close()

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet UK_库存补货提醒 row 1 Market inventory product contains US marketplace data" in output


def test_validate_safe_run_blocks_top_level_final_decision_summary_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(marketplace="UK", asin="B0SUMMARYTOP")
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
        sync_top_level_final_decision_summary=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "top-level final_decision_summary marketplace mismatch" in output


def test_validate_safe_run_blocks_top_level_decision_gate_counts_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(marketplace="UK", asin="B0GATETOP")
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
        sync_top_level_decision_gate_counts=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "top-level decision_gate_counts marketplace mismatch" in output


def test_validate_safe_run_blocks_final_decision_summary_count_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        asin="B0SUMMARYCOUNT",
        final_decision="EXECUTE_TODAY",
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
        corrupt_snapshot_final_decision_summary=True,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "UK final_decision_summary does not match product_final_decision_rows final_decision counts" in output
    assert "EXECUTE_TODAY" in output


def test_validate_safe_run_blocks_decision_gate_counts_count_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        asin="B0GATECOUNT",
        final_decision="WAIT_REVIEW",
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
        corrupt_snapshot_decision_gate_counts=True,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "UK decision_gate_counts does not match product_final_decision_rows final_decision counts" in output
    assert "WAIT_REVIEW" in output


def test_validate_safe_run_blocks_dashboard_marketplace_summary_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
    )
    dashboard = safe_dir / "dashboard.html"
    dashboard.write_text(
        dashboard.read_text(encoding="utf-8").replace("<td>UK</td><td>11</td>", "<td>UK</td><td>999</td>"),
        encoding="utf-8",
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "dashboard.html missing marketplace summary row for UK" in output


def test_validate_safe_run_blocks_dashboard_missing_frontend_coverage_label(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        include_dashboard_frontend_coverage_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "dashboard.html missing frontend coverage token 前台证据覆盖" in output


def test_validate_safe_run_blocks_marketplace_summary_missing_frontend_coverage_label(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        include_marketplace_summary_frontend_coverage_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "marketplace_summary.md missing frontend coverage token 前台证据覆盖" in output


def test_validate_safe_run_blocks_excel_overview_missing_frontend_coverage_label(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        include_excel_frontend_coverage_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 总览 missing frontend coverage metric ALL 前台队列" in output


def test_validate_safe_run_blocks_product_decision_excel_missing_contract_fields(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(marketplace="UK", asin="B0DECISION1")
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
        excel_product_headers=["marketplace", "sku", "asin", "product_name"],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 产品最终决策 missing fields" in output


def test_validate_safe_run_blocks_latest_recommendations_missing_product_decision_identity(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(marketplace="UK", asin="B0HTMLMISS1", product_name="HTML Missing Product")
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
        include_html_snapshot_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "latest_recommendations.html missing product decision row 1 ASIN B0HTMLMISS1" in output


def test_validate_safe_run_blocks_latest_recommendations_missing_product_decision_blocking_reason(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        asin="B0HTMLGATE1",
        product_name="HTML Gate Product",
        frontend_blocking_reasons=["地区待确认，当前前台证据不能用于放量"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
    )
    _write_minimal_safe_run(
        safe_dir,
        (
            '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 '
            '<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-1" '
            'data-product-decision-asin="B0HTMLGATE1">'
            "B0HTMLGATE1 HTML Gate Product 拦截加竞价 拦截加预算 拦截放量</article>"
        ),
        product_final_decision_rows=[decision_row],
        include_html_snapshot_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert (
        "latest_recommendations.html missing product decision row 1 blocking reason "
        "地区待确认，当前前台证据不能用于放量"
    ) in output


def test_validate_safe_run_blocks_latest_recommendations_missing_product_decision_blocked_action_label(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        asin="B0HTMLGATE2",
        product_name="HTML Blocked Action Product",
        frontend_blocking_reasons=["竞品可比性不足"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
    )
    _write_minimal_safe_run(
        safe_dir,
        (
            '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 '
            '<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-1" '
            'data-product-decision-asin="B0HTMLGATE2">'
            "B0HTMLGATE2 HTML Blocked Action Product 竞品可比性不足</article>"
        ),
        product_final_decision_rows=[decision_row],
        include_html_snapshot_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "latest_recommendations.html missing product decision row 1 blocked action label bid_up" in output


def test_validate_safe_run_blocks_latest_recommendations_product_gate_tokens_bound_to_wrong_article(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        asin="B0HTMLGATEBIND",
        product_name="HTML Gate Bound Product",
        frontend_blocking_reasons=["竞品不可比，不能用于放量"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
    )
    _write_minimal_safe_run(
        safe_dir,
            (
                '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 '
                '<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-1" '
                'data-product-decision-asin="B0HTMLGATEBIND">B0HTMLGATEBIND HTML Gate Bound Product</article> '
                '<article data-product-decision-marketplace="US" data-product-decision-sku="SKU-US" '
            'data-product-decision-asin="B0HTMLGATEBIND">竞品不可比，不能用于放量 拦截加竞价 拦截加预算 拦截放量</article>'
        ),
        product_final_decision_rows=[decision_row],
        include_html_snapshot_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert (
        "latest_recommendations.html product gate row 1 blocking reason not bound to ASIN B0HTMLGATEBIND"
        in output
    )


def test_validate_safe_run_blocks_latest_recommendations_frontend_tokens_bound_to_wrong_article(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        asin="B0HTMLFRONTBIND",
        product_name="HTML Frontend Evidence Product",
        frontend_blocking_reasons=["前台缓存或读取失败待确认，仅背景参考，不能单独支持放量"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
        frontend_check_status="沿用 2026-06-20 前台数据",
        frontend_cache_used=True,
        frontend_evidence_tier="仅背景参考",
        frontend_evidence_display_tier="仅背景参考",
        frontend_search_status="已读取部分结果",
        frontend_search_partial_evidence=True,
    )
    _write_minimal_safe_run(
        safe_dir,
        (
            '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 '
            '<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-1" '
            'data-product-decision-asin="B0HTMLFRONTBIND">'
            "B0HTMLFRONTBIND HTML Frontend Evidence Product 前台缓存或读取失败待确认，仅背景参考，不能单独支持放量 "
            "拦截加竞价 拦截加预算 拦截放量 拦截低预算精准测试</article> "
            '<article data-product-decision-marketplace="US" data-product-decision-sku="SKU-US" '
            'data-product-decision-asin="B0HTMLFRONTBIND">沿用 2026-06-20 前台数据 仅背景参考 搜索页：已读取部分结果</article>'
        ),
        product_final_decision_rows=[decision_row],
        include_html_snapshot_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert (
        "latest_recommendations.html product gate row 1 frontend status "
        "沿用 2026-06-20 前台数据 not bound to ASIN B0HTMLFRONTBIND"
    ) in output


def test_validate_safe_run_blocks_latest_recommendations_missing_low_budget_blocked_action_label(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        asin="B0HTMLLOWBUDGET",
        product_name="HTML Low Budget Blocked Product",
        frontend_blocking_reasons=["前台证据不足，需刷新后再放量"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
    )
    _write_minimal_safe_run(
        safe_dir,
        (
            '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 '
            '<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-1" '
            'data-product-decision-asin="B0HTMLLOWBUDGET">'
            "B0HTMLLOWBUDGET HTML Low Budget Blocked Product 前台证据不足，需刷新后再放量 "
            "拦截加竞价 拦截加预算 拦截放量</article>"
        ),
        product_final_decision_rows=[decision_row],
        include_html_snapshot_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert (
        "latest_recommendations.html missing product decision row 1 "
        "blocked action label create_exact_low_budget"
    ) in output


def test_validate_safe_run_blocks_latest_recommendations_product_decision_data_attr_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        asin="B0HTMLATTR",
        product_name="HTML Attr Product",
        frontend_blocking_reasons=["前台证据不足，需刷新后再放量"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
    )
    mismatched_attrs = _product_decision_contract_attrs(
        decision_row,
        **{"data-product-decision-final": "EXECUTE_TODAY"},
    )
    _write_minimal_safe_run(
        safe_dir,
        (
            '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 '
            '<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-1" '
            f'data-product-decision-asin="B0HTMLATTR"{mismatched_attrs}>'
            "B0HTMLATTR HTML Attr Product 前台证据不足，需刷新后再放量 "
            "拦截加竞价 拦截加预算 拦截放量 拦截低预算精准测试 已自动检查 仅背景参考</article>"
        ),
        product_final_decision_rows=[decision_row],
        include_html_snapshot_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert (
        "latest_recommendations.html product gate row 1 B0HTMLATTR "
        "data attr data-product-decision-final mismatch for field final_decision"
    ) in output


def test_validate_safe_run_blocks_latest_recommendations_missing_review_metric_token(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    review_row = {
        "marketplace": "UK",
        "sku": "SKU-HTML-METRIC",
        "asin": "B0HTMLMETRIC",
        "product_name": "HTML Metric Product",
        "search_term_or_target": "html metric keyword",
        "action_id": "UK||SKU-HTML-METRIC||B0HTMLMETRIC||html metric keyword||bid_up",
        "normalized_action": "bid_up",
        "action_scope": "keyword",
        "action_type": "加竞价",
        "action_detail": "加竞价 5%-10%",
        "executed_at": "2026-05-20T09:00:00",
        "report_date": "2026-05-20",
        "review_date": "2026-06-08",
        "days_since_execution": 19,
        "review_window": "7天后复盘",
        "review_outcome": "effective",
        "judgement": "有改善迹象",
        "effectiveness_score": 2,
        "outcome": "有改善迹象",
        "effect_evidence": "7 天本 SKU 有单",
        "review_status": "可做7天复盘",
        "review_phase": "7d_review",
        "cooldown_status": "观察中",
        "cooldown_until": "",
        "block_reason": "",
        "rule_adjustment": "保留当前竞价",
        "learning_scope": "keyword",
        "current_7d_clicks": 8,
        "current_7d_spend": 3.2,
        "current_7d_ad_orders": 2,
        "current_7d_ad_sales": 30,
        "current_7d_promoted_ad_sales": 30,
        "current_7d_halo_ad_orders": 0,
        "current_7d_halo_ad_sales": 0,
        "current_14d_clicks": 14,
        "current_14d_spend": 5.1,
        "current_14d_ad_orders": 3,
        "current_14d_ad_sales": 45,
        "current_14d_promoted_ad_sales": 45,
        "current_14d_halo_ad_orders": 0,
        "current_14d_halo_ad_sales": 0,
        "current_14d_tacos": 0.09,
        "promoted_conversion_improved": True,
        "halo_only_conversion": False,
        "target_sku_not_converted": False,
        "attribution_effect_status": "promoted_sku_converted",
        "attribution_effect_note": "本 SKU 成交优先",
        **_effective_review_metrics(),
    }
    _write_minimal_safe_run(
        safe_dir,
        (
            '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 '
            "html metric keyword B0HTMLMETRIC"
        ),
        keyword_review_rows=[review_row],
        include_html_snapshot_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "latest_recommendations.html missing keyword review row 1 metric token 本 SKU 单 2" in output


def test_validate_safe_run_blocks_effective_review_without_execution_anchored_source(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "kw-anchored-source-1",
            "normalized_action": "bid_up",
            "action_scope": "search_term",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0ANCHORSRC",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "anchored source keyword",
            "executed_at": "2026-06-01",
            "report_date": "2026-06-01",
            "review_date": "2026-06-08",
            "review_window": "7天后复盘",
            "judgement": "有改善迹象",
            "outcome": "有改善迹象",
            "days_since_execution": 7,
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "effect_evidence": "7 天本 SKU 有单。",
            "review_phase": "7d_review",
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
            "review_data_source": "rolling_window",
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        (
            '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 '
            "anchored source keyword B0ANCHORSRC 本 SKU 单 2 ACOS 0.12 目标 ACOS 0.2 TACOS 0.08 总单 4 库存 18 "
            "14天 本 SKU 单 3 14天 ACOS 0.14 14天 TACOS 0.09 14天 总单 7 14天 库存 18 "
            "执行后本 SKU 单 2 执行后 TACOS 0.08"
        ),
        keyword_review_rows=[review_row],
        include_html_snapshot_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "effective outcome missing execution anchored daily source" in output


def test_validate_safe_run_blocks_latest_recommendations_missing_anchor_metric_token(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "kw-anchor-html-1",
            "normalized_action": "bid_up",
            "action_scope": "search_term",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0ANCHORHTML",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "anchor html keyword",
            "executed_at": "2026-06-01",
            "report_date": "2026-06-01",
            "review_date": "2026-06-08",
            "review_window": "7天后复盘",
            "judgement": "有改善迹象",
            "outcome": "有改善迹象",
            "days_since_execution": 7,
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "effect_evidence": "7 天本 SKU 有单。",
            "review_phase": "7d_review",
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        (
            '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 '
            "anchor html keyword B0ANCHORHTML 本 SKU 单 2 ACOS 0.12 目标 ACOS 0.2 TACOS 0.08 总单 4 库存 18 "
            "14天 本 SKU 单 3 14天 ACOS 0.14 14天 TACOS 0.09 14天 总单 7 14天 库存 18 "
            "执行前7天 2026-05-13 至 2026-05-19 执行前本 SKU 单 1 执行前总单 2 执行前 TACOS 0.1 "
            "执行后3天 2026-05-20 至 2026-05-22 执行后3天覆盖天数 3 执行后3天本 SKU 单 1 "
            "执行后3天总单 2 执行后3天 ACOS 0.13 执行后3天 TACOS 0.09 执行后3天库存 18 "
            "执行后7天 2026-05-20 至 2026-05-26 执行后覆盖天数 7 执行后总单 4 "
            "执行后 ACOS 0.12 执行后 TACOS 0.08 执行后库存 18"
        ),
        keyword_review_rows=[review_row],
        include_html_snapshot_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "latest_recommendations.html missing keyword review row 1 metric token 执行后本 SKU 单 2" in output


def test_validate_safe_run_blocks_latest_recommendations_missing_review_target_acos_token(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "kw-html-target-acos-1",
            "normalized_action": "bid_down",
            "action_scope": "keyword",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0HTMLTARGET",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "html target acos keyword",
            "executed_at": "2026-06-01",
            "report_date": "2026-06-01",
            "review_window": "7天后复盘",
            "judgement": "待人工复查",
            "days_since_execution": 8,
            "review_outcome": "not_ready",
            "effect_evidence": "7天复盘口径，等待人工确认。",
            "review_phase": "7d",
            **_effective_review_metrics(),
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        (
            '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 '
            "html target acos keyword B0HTMLTARGET 本 SKU 单 2 ACOS 0.12 TACOS 0.08 总单 4 库存 18 "
            "14天 本 SKU 单 3 14天 ACOS 0.09 14天 TACOS 0.09 14天 总单 5 14天 库存 18"
        ),
        keyword_review_rows=[review_row],
        include_html_snapshot_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "latest_recommendations.html missing keyword review row 1 metric token 目标 ACOS 0.2" in output


def test_validate_safe_run_blocks_latest_recommendations_missing_early_review_policy_token(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "kw-html-early-policy-1",
            "normalized_action": "bid_down",
            "action_scope": "keyword",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0HTMLEARLY",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "html early policy keyword",
            "executed_at": "2026-06-01",
            "report_date": "2026-06-01",
            "review_window": "3d_check",
            "judgement": "待7天确认",
            "days_since_execution": 4,
            "review_outcome": "not_ready",
            "effect_evidence": "3天复查口径，7天结论待补；继续观察。",
            "review_phase": "3d",
            **_effective_review_metrics(),
            "current_7d_promoted_ad_orders": 0,
            "current_7d_total_orders": 0,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        (
            '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 '
            "html early policy keyword B0HTMLEARLY 本 SKU 单 0 ACOS 0.12 目标 ACOS 0.2 TACOS 0.08 总单 0 库存 18"
        ),
        keyword_review_rows=[review_row],
        include_html_snapshot_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "latest_recommendations.html missing keyword review row 1 metric token 3 天窗口只做初步判断" in output


def test_validate_safe_run_blocks_three_day_review_evidence_without_early_qualifier(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "kw-json-early-policy-1",
            "normalized_action": "bid_up",
            "action_scope": "keyword",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0JSONEARLY",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "json early policy keyword",
            "executed_at": "2026-06-01",
            "report_date": "2026-06-01",
            "review_window": "3d_check",
            "judgement": "待7天确认",
            "days_since_execution": 4,
            "review_outcome": "not_ready",
            "effect_evidence": "继续观察，暂不调整。",
            "review_phase": "3d",
            **_effective_review_metrics(),
            "current_7d_promoted_ad_orders": 0,
            "current_7d_total_orders": 0,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        (
            '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 '
            "json early policy keyword B0JSONEARLY 本 SKU 单 0 ACOS 0.12 目标 ACOS 0.2 TACOS 0.08 总单 0 库存 18 "
            "3 天窗口只做初步判断"
        ),
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "UK review row 1 json early policy keyword 3-day review evidence missing early-window qualifier" in output


def test_validate_safe_run_blocks_latest_recommendations_missing_review_14d_metric_token(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "kw-html-14d-metric-1",
            "normalized_action": "bid_down",
            "action_scope": "keyword",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0HTML14D",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "html 14d metric keyword",
            "executed_at": "2026-06-01",
            "report_date": "2026-06-01",
            "review_window": "7天后复盘",
            "judgement": "待人工复查",
            "days_since_execution": 8,
            "review_outcome": "not_ready",
            "effect_evidence": "7天复盘口径，等待人工确认。",
            "review_phase": "7d",
            **_effective_review_metrics(),
            "current_14d_promoted_ad_orders": 3,
            "current_14d_acos": 0.09,
            "current_14d_tacos": 0.07,
            "current_14d_total_orders": 5,
            "current_14d_available_stock": 18,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        (
            '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 '
            "html 14d metric keyword B0HTML14D 本 SKU 单 2 ACOS 0.12 目标 ACOS 0.2 TACOS 0.08 总单 4 库存 18 "
            "14天 本 SKU 单 3 14天 ACOS 0.09 14天 总单 5 14天 库存 18"
        ),
        keyword_review_rows=[review_row],
        include_html_snapshot_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "latest_recommendations.html missing keyword review row 1 metric token 14天 TACOS 0.07" in output


def test_validate_safe_run_blocks_latest_recommendations_missing_tomorrow_review_identity(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    review_row = {
        "marketplace": "UK",
        "sku": "SKU-TOMORROW",
        "asin": "B0TOMORROW",
        "product_name": "Tomorrow missing product",
        "review_reason": "执行后复查",
        "current_evidence": "点击 4；花费 £0.88；订单 0",
        "tomorrow_check": "复查该词点击、花费、订单和 ACOS",
        "trigger_action": "未改善则降竞价",
        "search_term_or_target": "tomorrow missing target",
        "normalized_action": "bid_down",
        "action_id": "UK||SKU-TOMORROW||B0TOMORROW||tomorrow missing target||bid_down",
    }
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        tomorrow_review_rows=[review_row],
        include_html_snapshot_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "latest_recommendations.html missing tomorrow review row 1 target tomorrow missing target" in output


def test_validate_safe_run_blocks_today_task_excel_identity_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    task_row = {
        "marketplace": "UK",
        "sku": "SKU-ACTION",
        "asin": "B0ACTIONJSON",
        "product_name": "Action product",
        "confirmed_status": "已执行",
        "priority": "P0",
        "issue_type": "广告消耗无转化",
        "action_group": "广告动作",
        "today_action": "降竞价10%-20%",
        "search_term_or_target": "dimmer desk lamp",
        "suggested_action": "降竞价10%-20%",
        "normalized_action": "bid_down",
        "action_id": "UK||SKU-ACTION||B0ACTIONJSON||dimmer desk lamp||bid_down",
    }
    excel_row = dict(task_row)
    excel_row["asin"] = "B0ACTIONOLD"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        ad_rows=[task_row],
        excel_today_task_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 今日动作清单 row identity mismatch for today task queue rows" in output
    assert "B0ACTIONJSON" in output
    assert "B0ACTIONOLD" in output


def test_validate_safe_run_blocks_today_task_excel_content_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    task_row = {
        "marketplace": "UK",
        "sku": "SKU-ACTION",
        "asin": "B0ACTIONJSON",
        "product_name": "Action product",
        "confirmed_status": "已执行",
        "priority": "P0",
        "issue_type": "广告消耗无转化",
        "primary_reason": "近14天广告无单且点击>=20",
        "key_evidence": "点击 31；广告订单 0",
        "action_group": "广告动作",
        "today_action": "降竞价10%-20%",
        "search_term_or_target": "dimmer desk lamp",
        "suggested_action": "降竞价10%-20%",
        "copy_action_line": "dimmer desk lamp 降竞价10%-20%",
        "normalized_action": "bid_down",
        "action_scope": "search_term",
        "action_id": "UK||SKU-ACTION||B0ACTIONJSON||dimmer desk lamp||bid_down",
    }
    excel_row = dict(task_row)
    excel_row["primary_reason"] = "旧原因"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        ad_rows=[task_row],
        excel_today_task_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 今日动作清单 field mismatch for today task queue rows" in output
    assert "field primary_reason" in output
    assert "expected '近14天广告无单且点击>=20', got '旧原因'" in output


def test_validate_safe_run_blocks_today_task_excel_copy_block_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    task_row = {
        "marketplace": "UK",
        "sku": "SKU-COPY",
        "asin": "B0COPYJSON",
        "product_name": "Copy block product",
        "confirmed_status": "已执行",
        "priority": "P1",
        "issue_type": "广告动作",
        "primary_reason": "复制文本必须和 JSON 快照一致",
        "key_evidence": "点击 31；广告订单 0",
        "action_group": "广告动作",
        "today_action": "降竞价10%-20%",
        "search_term_or_target": "copy block term",
        "suggested_action": "降竞价10%-20%",
        "copy_action_line": "copy block term 降竞价10%-20%",
        "copy_block": "copy block term 降竞价10%-20%\n后台复制文本",
        "normalized_action": "bid_down",
        "action_scope": "search_term",
        "action_id": "UK||SKU-COPY||B0COPYJSON||copy block term||bid_down",
    }
    excel_row = dict(task_row)
    excel_row["copy_block"] = "旧后台复制文本"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        ad_rows=[task_row],
        excel_today_task_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 今日动作清单 field mismatch for today task queue rows" in output
    assert "field copy_block" in output
    assert "旧后台复制文本" in output


def test_validate_safe_run_blocks_tomorrow_review_excel_identity_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    review_row = {
        "marketplace": "UK",
        "sku": "SKU-REVIEW",
        "asin": "B0REVIEWJSON",
        "product_name": "Review product",
        "review_reason": "3-5点击无单，未达到强动作阈值",
        "current_evidence": "点击 4；花费 £0.88；订单 0",
        "tomorrow_check": "复查是否新增点击、花费或订单；未改善再降竞价",
        "trigger_action": "达到阈值后升级今日动作",
    }
    excel_row = dict(review_row)
    excel_row["asin"] = "B0REVIEWOLD"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        tomorrow_review_rows=[review_row],
        excel_tomorrow_review_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 明日复查清单 row identity mismatch for tomorrow review rows" in output
    assert "B0REVIEWJSON" in output
    assert "B0REVIEWOLD" in output


def test_validate_safe_run_blocks_tomorrow_review_excel_content_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    review_row = {
        "marketplace": "UK",
        "sku": "SKU-REVIEW",
        "asin": "B0REVIEWJSON",
        "product_name": "Review product",
        "review_reason": "3-5点击无单，未达到强动作阈值",
        "current_evidence": "点击 4；花费 £0.88；订单 0",
        "tomorrow_check": "复查是否新增点击、花费或订单；未改善再降竞价",
        "trigger_action": "达到阈值后升级今日动作",
    }
    excel_row = dict(review_row)
    excel_row["current_evidence"] = "点击 0；花费 £0.00；订单 0"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        tomorrow_review_rows=[review_row],
        excel_tomorrow_review_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 明日复查清单 field mismatch for tomorrow review rows" in output
    assert "field current_evidence" in output
    assert "expected '点击 4；花费 £0.88；订单 0', got '点击 0；花费 £0.00；订单 0'" in output


def test_validate_safe_run_blocks_listing_diagnosis_excel_identity_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    diagnosis_row = {
        "marketplace": "UK",
        "产品": "Listing product",
        "SKU": "SKU-LISTING",
        "ASIN": "B0LISTJSON",
        "诊断类型": "Listing 待人工确认",
        "主因": "加购后不购买",
        "关键证据": "近14天广告点击 12；广告订单 0；总单 0",
        "建议动作": "暂时不加广告预算；人工检查价格、评分和主图",
        "confirmed_status": "待确认",
        "suggested_action": "先查前台",
        "normalized_action": "observe",
        "action_scope": "product",
        "action_id": "UK||SKU-LISTING||B0LISTJSON||product||||observe",
    }
    excel_row = dict(diagnosis_row)
    excel_row["ASIN"] = "B0LISTOLD"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        listing_diagnosis_rows=[diagnosis_row],
        excel_listing_diagnosis_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet Listing待确认 row identity mismatch for listing price diagnosis rows" in output
    assert "B0LISTJSON" in output
    assert "B0LISTOLD" in output


def test_validate_safe_run_blocks_listing_growth_claim_with_weak_frontend(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    diagnosis_row = {
        "marketplace": "UK",
        "产品": "Listing weak frontend product",
        "SKU": "SKU-LISTING",
        "ASIN": "B0UKTEST",
        "诊断类型": "Listing 待人工确认",
        "主因": "点击后不转化",
        "关键证据": "近14天广告点击 12；广告订单 0；总单 0",
        "建议动作": "立即改 Listing 并加预算",
        "confirmed_status": "待确认",
        "suggested_action": "先查前台",
        "normalized_action": "observe",
        "action_scope": "product",
        "action_id": "UK||SKU-LISTING||B0UKTEST||product||||observe",
    }
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        listing_diagnosis_rows=[diagnosis_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert (
        "latest analysis UK listing_price_diagnosis_rows row 1 B0UKTEST "
        "contains frontend-backed strong listing or growth claim"
    ) in output
    assert "under weak frontend evidence" in output


def test_validate_safe_run_blocks_cost_diagnosis_excel_content_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    diagnosis_row = {
        "marketplace": "UK",
        "产品": "Cost product",
        "SKU": "SKU-COST",
        "ASIN": "B0COSTJSON",
        "诊断类型": "成本 / 利润压力诊断",
        "主因": "利润不允许加广告",
        "关键证据": "毛利不足；广告 ACOS 高于目标；库存 18",
        "建议动作": "不加预算；先复核成本、售价和优惠券",
        "confirmed_status": "待确认",
        "suggested_action": "不加预算",
        "normalized_action": "observe",
        "action_scope": "product",
        "action_id": "UK||SKU-COST||B0COSTJSON||product||||observe",
    }
    excel_row = dict(diagnosis_row)
    excel_row["关键证据"] = "旧证据"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        cost_diagnosis_rows=[diagnosis_row],
        excel_cost_diagnosis_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 成本利润诊断 field mismatch for cost profit diagnosis rows" in output
    assert "field 关键证据" in output
    assert "expected '毛利不足；广告 ACOS 高于目标；库存 18', got '旧证据'" in output


def test_validate_safe_run_blocks_product_decision_excel_growth_without_strong_frontend(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(marketplace="UK", asin="B0DECISION1")
    excel_row = dict(decision_row)
    excel_row.update(
        {
            "today_allowed_actions": ["observe", "bid_up"],
            "today_blocked_actions": ["budget_up", "broad_scale"],
            "frontend_evidence_state": "background",
            "frontend_evidence_tier": "仅背景参考",
            "frontend_check_status": "沿用 2026-06-05 前台数据",
            "frontend_cache_used": True,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_location_scope": "exact",
            "frontend_location_exact": True,
            "competitor_comparability": "medium",
            "comparable_competitor_count": 3,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
        excel_product_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel product final decision row 1 B0DECISION1 allows growth without ok_high frontend evidence" in output


def test_validate_safe_run_blocks_product_decision_excel_identity_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(marketplace="UK", asin="B0DECISION1")
    excel_row = dict(decision_row)
    excel_row["asin"] = "B0DECISION2"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
        excel_product_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 产品最终决策 row identity mismatch" in output
    assert "B0DECISION1" in output
    assert "B0DECISION2" in output


def test_validate_safe_run_blocks_task_queue_growth_action_when_product_gate_blocks_it(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        sku="SKU-GATE",
        asin="B0TASKGATE",
        today_allowed_actions=["observe"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
        frontend_evidence_state="weak",
        frontend_evidence_tier="仅背景参考",
        frontend_evidence_display_tier="仅背景参考",
        frontend_auto_conclusion="FRONTEND_WEAK",
    )
    task_row = {
        "marketplace": "UK",
        "sku": "SKU-GATE",
        "asin": "B0TASKGATE",
        "product_name": "Blocked growth task product",
        "suggested_action": "加价 5%-10%",
        "copy_action_line": "建议加价 5%-10%",
    }
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        ad_rows=[task_row],
        product_final_decision_rows=[decision_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "UK today_task_queue_rows row 1 B0TASKGATE contains growth action bid_up blocked by product final decision" in output


def test_validate_safe_run_blocks_aux_ad_workbench_growth_action_when_product_gate_blocks_it(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        sku="SKU-AUX-GATE",
        asin="B0AUXGATE",
        today_allowed_actions=["observe"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
        frontend_evidence_state="weak",
        frontend_evidence_tier="仅背景参考",
        frontend_evidence_display_tier="仅背景参考",
        frontend_auto_conclusion="FRONTEND_WEAK",
    )
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-AUX-GATE",
        "asin": "B0AUXGATE",
        "product_name": "Aux blocked growth product",
        "suggested_action": "加价 5%-10%",
        "copy_action_line": "建议加价 5%-10%",
    }
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
    )
    analysis_path = safe_dir / "latest_analysis.json"
    analysis = json.loads(analysis_path.read_text(encoding="utf-8"))
    analysis["marketplace_results"][0]["report_view_snapshot"]["scale_keyword_rows"] = [scale_row]
    analysis_path.write_text(json.dumps(analysis, ensure_ascii=False), encoding="utf-8")

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "UK ad_workbench_rows row 1 B0AUXGATE contains growth action bid_up blocked by product final decision" in output


def test_validate_safe_run_blocks_growth_test_when_product_gate_blocks_it(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        sku="SKU-GROWTH-GATE",
        asin="B0GROWTHTEST",
        today_allowed_actions=["observe"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
        frontend_evidence_state="weak",
        frontend_evidence_tier="仅背景参考",
        frontend_evidence_display_tier="仅背景参考",
        frontend_auto_conclusion="FRONTEND_WEAK",
    )
    growth_row = {
        "marketplace": "UK",
        "sku": "SKU-GROWTH-GATE",
        "asin": "B0GROWTHTEST",
        "product_name": "Growth test blocked product",
        "search_term_or_target": "blocked exact test",
        "suggested_action": "小预算试投",
        "manual_action_taken": "小预算试投",
        "normalized_action": "growth_test",
        "experiment_type": "growth_test",
    }
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
    )
    analysis_path = safe_dir / "latest_analysis.json"
    analysis = json.loads(analysis_path.read_text(encoding="utf-8"))
    analysis["marketplace_results"][0]["report_view_snapshot"]["growth_test_rows"] = [growth_row]
    analysis_path.write_text(json.dumps(analysis, ensure_ascii=False), encoding="utf-8")

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert (
        "UK ad_workbench_rows row 1 B0GROWTHTEST "
        "contains growth action create_exact_low_budget blocked by product final decision"
    ) in output


def test_validate_safe_run_treats_new_exact_text_as_growth_test(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        sku="SKU-NEW-EXACT-GATE",
        asin="B0NEWEXACT",
        today_allowed_actions=["observe"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
        frontend_evidence_state="weak",
        frontend_evidence_tier="仅背景参考",
        frontend_evidence_display_tier="仅背景参考",
        frontend_auto_conclusion="FRONTEND_WEAK",
    )
    growth_row = {
        "marketplace": "UK",
        "sku": "SKU-NEW-EXACT-GATE",
        "asin": "B0NEWEXACT",
        "product_name": "New exact blocked product",
        "search_term_or_target": "new exact target",
        "suggested_action": "新建精准小预算",
        "copy_action_line": "新建精准小预算测试",
        "normalized_action": "observe",
    }
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
    )
    analysis_path = safe_dir / "latest_analysis.json"
    analysis = json.loads(analysis_path.read_text(encoding="utf-8"))
    analysis["marketplace_results"][0]["report_view_snapshot"]["growth_test_rows"] = [growth_row]
    analysis_path.write_text(json.dumps(analysis, ensure_ascii=False), encoding="utf-8")

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert (
        "UK ad_workbench_rows row 1 B0NEWEXACT "
        "contains growth action create_exact_low_budget blocked by product final decision"
    ) in output


def test_validate_safe_run_treats_percent_add_text_as_bid_growth_action(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        sku="SKU-PERCENT-GATE",
        asin="B0PERCENTGATE",
        today_allowed_actions=["observe"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
        frontend_evidence_state="weak",
        frontend_evidence_tier="仅背景参考",
        frontend_evidence_display_tier="仅背景参考",
        frontend_auto_conclusion="FRONTEND_WEAK",
    )
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-PERCENT-GATE",
        "asin": "B0PERCENTGATE",
        "product_name": "Percent add blocked growth product",
        "search_term_or_target": "percent add target",
        "suggested_action": "加 5%-10%",
        "copy_action_line": "percent add target 加 5%-10%",
    }
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
    )
    analysis_path = safe_dir / "latest_analysis.json"
    analysis = json.loads(analysis_path.read_text(encoding="utf-8"))
    analysis["marketplace_results"][0]["report_view_snapshot"]["scale_keyword_rows"] = [scale_row]
    analysis_path.write_text(json.dumps(analysis, ensure_ascii=False), encoding="utf-8")

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "UK ad_workbench_rows row 1 B0PERCENTGATE contains growth action bid_up blocked by product final decision" in output


def test_validate_safe_run_blocks_product_decision_excel_content_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        asin="B0DECISION1",
        final_decision="EXECUTE_TODAY",
        final_decision_label="今天执行",
        today_allowed_actions=["bid_down"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
    )
    excel_row = dict(decision_row)
    excel_row["final_decision"] = "WAIT_REVIEW"
    excel_row["today_allowed_actions"] = ["observe"]
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
        excel_product_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 产品最终决策 field mismatch for product final decisions" in output
    assert "field final_decision: expected 'EXECUTE_TODAY', got 'WAIT_REVIEW'" in output


def test_validate_safe_run_blocks_product_decision_excel_frontend_evidence_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        asin="B0FRONTEXCEL",
        frontend_evidence_state="ok_high",
        frontend_evidence_tier="强诊断可用",
        frontend_evidence_display_tier="强诊断可用",
        frontend_check_status="已自动检查",
        frontend_search_status="已自动检查",
        frontend_auto_conclusion="FRONTEND_OK",
        frontend_cache_used=False,
        frontend_evidence_quality_score=88,
        frontend_evidence_audit_summary="前台证据通过质量门",
        frontend_evidence_audit_detail="搜索页已读；竞品 3 个",
        competitor_comparability="high",
        comparable_competitor_count=3,
        frontend_location_scope="exact",
        frontend_location_verified=True,
        frontend_location_exact=True,
    )
    excel_row = dict(decision_row)
    excel_row["frontend_evidence_audit_summary"] = "前台证据可辅助判断，不能单独放量"
    excel_row["comparable_competitor_count"] = 4
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
        excel_product_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 产品最终决策 field mismatch for product final decisions" in output
    assert "field frontend_evidence_audit_summary" in output


def test_validate_safe_run_blocks_product_decision_excel_fusion_reason_mismatch(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        asin="B0FUSIONXLS",
        keyword_memory_summary="历史同词复盘：未验证",
        evidence_used=["近14天点击 42", "前台证据仅背景参考"],
        coupon={"coupon_type": "percent", "coupon_value": 10, "coupon_confidence": "high"},
        fusion_issue_type="前台和广告共同问题",
        fusion_confidence="high",
        fusion_reason="广告证据：ad_no_order_clicks；前台证据：明确前台劣势",
        fusion_evidence_flags=["ad_no_order_clicks", "frontend_competitiveness"],
        fusion_missing_evidence=["current_frontend_screenshot"],
        last_updated="2026-06-24",
    )
    excel_row = dict(decision_row)
    excel_row["fusion_reason"] = "广告证据：ad_no_order_clicks"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
        excel_product_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 产品最终决策 field mismatch for product final decisions" in output
    assert "field fusion_reason" in output


def test_validate_safe_run_blocks_today_task_excel_missing_empty_content_header(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    task_row = {
        "marketplace": "UK",
        "sku": "SKU-ACTION",
        "asin": "B0ACTIONHEADER",
        "product_name": "Action product",
        "confirmed_status": "已执行",
        "priority": "P0",
        "issue_type": "广告消耗无转化",
        "action_group": "广告动作",
        "today_action": "降竞价10%-20%",
        "search_term_or_target": "missing copy block term",
        "suggested_action": "降竞价10%-20%",
        "copy_block": "",
        "normalized_action": "bid_down",
        "action_scope": "search_term",
        "action_id": "UK||SKU-ACTION||B0ACTIONHEADER||missing copy block term||bid_down",
    }
    excel_row = dict(task_row)
    excel_row.pop("copy_block", None)
    excel_headers = list(excel_row.keys())
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        ad_rows=[task_row],
        excel_today_task_headers=excel_headers,
        excel_today_task_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 今日动作清单 missing fields for today task queue rows: copy_block" in output


def test_validate_safe_run_blocks_market_product_decision_excel_wrong_marketplace(capsys, tmp_path) -> None:
    from openpyxl import load_workbook

    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        asin="B0MARKETDECISION",
        product_name="Market decision product",
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
    )
    for workbook_path in safe_dir.glob("amazon_ops_report_*.xlsx"):
        workbook = load_workbook(workbook_path)
        sheet = workbook["UK_产品最终决策"]
        headers = [cell.value for cell in sheet[1]]
        sheet.cell(row=2, column=headers.index("marketplace") + 1).value = "US"
        workbook.save(workbook_path)
        workbook.close()

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet UK_产品最终决策 row 1 Market decision product contains US marketplace data" in output


def test_validate_safe_run_blocks_market_product_decision_excel_growth_without_strong_frontend(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(marketplace="UK", asin="B0DECISION1")
    excel_market_row = dict(decision_row)
    excel_market_row.update(
        {
            "today_allowed_actions": "['observe', 'bid_up']",
            "today_blocked_actions": "['budget_up', 'broad_scale']",
            "frontend_evidence_state": "background",
            "frontend_evidence_tier": "仅背景参考",
            "frontend_check_status": "沿用 2026-06-05 前台数据",
            "frontend_cache_used": True,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_location_scope": "exact",
            "frontend_location_exact": True,
            "competitor_comparability": "medium",
            "comparable_competitor_count": 3,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
        excel_market_product_rows=[excel_market_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel product final decision row 1 B0DECISION1 allows growth without ok_high frontend evidence" in output


def test_validate_safe_run_accepts_decision_markers(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 前台缓存工具',
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 0
    assert "safe-run outputs validated" in output


def test_validate_safe_run_blocks_stale_formal_learning_sidecar(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 前台缓存工具',
    )
    (safe_dir / "learned_rules_20260608.json").write_text(
        json.dumps([{"stale": True}], ensure_ascii=False),
        encoding="utf-8",
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "safe-run formal daily validation failed" in output
    assert "learned_rules_20260608.json content mismatch" in output


def test_validate_safe_run_requires_copy_area_for_pending_ads(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 1 市场调查 提交今日数据 系统结论 融合诊断',
        ad_rows=[{"confirmed_status": "待确认", "copy_action_line": "建议否词"}],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "missing copy area for pending ad rows" in output


def test_validate_safe_run_requires_copy_area_for_copy_block_only_pending_ads(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 1 市场调查 提交今日数据 系统结论 融合诊断',
        ad_rows=[
            {
                "confirmed_status": "待确认",
                "suggested_action": "观察",
                "copy_action_line": "",
                "copy_block": "copy block only term 降竞价10%-20%",
                "search_term_or_target": "copy block only term",
            }
        ],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "missing copy area for pending ad rows" in output


def test_validate_safe_run_requires_copy_area_for_pending_growth_test_rows(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        sku="SKU-GROWTH-COPY",
        asin="B0GROWTHCOPY",
        today_allowed_actions=["observe", "create_exact_low_budget"],
        today_blocked_actions=[],
        frontend_evidence_state="ok_high",
        frontend_evidence_tier="强诊断可用",
        frontend_evidence_display_tier="强诊断可用",
        frontend_decision_evidence_tier="强诊断可用",
        frontend_evidence_quality_score=88,
        frontend_auto_conclusion="FRONTEND_OK",
        frontend_auto_conclusion_label="未见明显前台劣势",
        competitor_comparability="high",
        comparable_competitor_count=3,
        frontend_location_scope="exact",
        frontend_location_verified=True,
        frontend_location_exact=True,
        frontend_cache_used=False,
        frontend_check_status="已自动检查",
        frontend_search_status="已自动检查",
    )
    growth_row = {
        "marketplace": "UK",
        "sku": "SKU-GROWTH-COPY",
        "asin": "B0GROWTHCOPY",
        "product_name": "Growth copy product",
        "search_term_or_target": "growth copy target",
        "confirmed_status": "待确认",
        "suggested_action": "小预算试投",
        "experiment_type": "growth_test",
        "normalized_action": "growth_test",
    }
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 1 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[decision_row],
    )
    analysis_path = safe_dir / "latest_analysis.json"
    analysis = json.loads(analysis_path.read_text(encoding="utf-8"))
    analysis["marketplace_results"][0]["report_view_snapshot"]["growth_test_rows"] = [growth_row]
    analysis_path.write_text(json.dumps(analysis, ensure_ascii=False), encoding="utf-8")

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "missing copy area for pending ad rows" in output


def test_validate_safe_run_allows_copy_area_for_pending_ads(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    ad_row = {
        "marketplace": "UK",
        "sku": "SKU-COPY-OK",
        "asin": "B0COPYOK",
        "product_name": "Copy OK product",
        "confirmed_status": "待确认",
        "copy_action_line": "建议否词",
        "suggested_action": "否定精准",
        "today_action": "否定精准",
        "search_term_or_target": "copy ok target",
    }
    ad_row = add_action_identity(ad_row, "否定精准", "search_term")
    completion_payload = {
        **ad_row,
        "manual_action_taken": "否定精准",
        "confirmed_note": "网页勾选已完成；未满3天不判断，3天初查，7天正式复盘。",
    }
    payload_attr = html.escape(
        json.dumps(completion_payload, ensure_ascii=False, separators=(",", ":")),
        quote=True,
    )
    _write_minimal_safe_run(
        safe_dir,
            (
                '2026-06-08 今天广告动作 id="today-ad-actions-all" 复制到广告后台 copy ok target B0COPYOK '
                f'市场调查 提交今日数据 系统结论 融合诊断 '
                f'<div data-ad-complete-payload="{payload_attr}" data-action-id="{ad_row["action_id"]}"></div>'
            ),
        ad_rows=[ad_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 0
    assert "safe-run outputs validated" in output


def test_validate_safe_run_blocks_latest_recommendations_missing_pending_ad_identity(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    task_row = {
        "marketplace": "UK",
        "sku": "SKU-PENDING-HTML",
        "asin": "B0PENDINGHTML",
        "product_name": "Pending ad HTML product",
        "confirmed_status": "待确认",
        "today_action": "降竞价10%-20%",
        "suggested_action": "降竞价10%-20%",
        "copy_action_line": "missing pending ad target 降竞价10%-20%",
        "search_term_or_target": "missing pending ad target",
        "normalized_action": "bid_down",
    }
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 复制到广告后台 市场调查 提交今日数据 系统结论 融合诊断',
        ad_rows=[task_row],
        include_html_snapshot_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "latest_recommendations.html missing pending ad row 1 target missing pending ad target" in output


def test_validate_safe_run_treats_background_ad_note_as_zero_pending(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        ad_rows=[
            {
                "confirmed_status": "仅背景参考",
                "today_action": "保守跑：不加预算，只保留高相关精准词和必要降竞价/否词。",
            }
        ],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 0
    assert "safe-run outputs validated" in output


def test_validate_safe_run_blocks_growth_decision_without_strong_frontend(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[
            _product_decision_row(
                marketplace="UK",
                asin="B0SCALEBAD1",
                today_allowed_actions=["observe", "bid_up"],
                today_blocked_actions=["budget_up", "broad_scale"],
                frontend_evidence_state="background",
                frontend_evidence_tier="仅背景参考",
                frontend_location_scope="exact",
                frontend_location_exact=True,
                competitor_comparability="medium",
                comparable_competitor_count=3,
            )
        ],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "allows growth without ok_high frontend evidence" in output


def test_validate_safe_run_blocks_weak_frontend_without_explicit_growth_blocks(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        product_final_decision_rows=[
            _product_decision_row(
                marketplace="UK",
                asin="B0WEAKBLOCKS",
                today_allowed_actions=["observe"],
                today_blocked_actions=["bid_up"],
                frontend_evidence_state="weak",
                frontend_evidence_tier="仅背景参考",
                frontend_evidence_display_tier="仅背景参考",
                frontend_check_status="待前台检查",
                frontend_auto_conclusion="FRONTEND_WEAK",
                competitor_comparability="unknown",
                comparable_competitor_count=0,
            )
        ],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "does not explicitly block growth actions under weak frontend evidence" in output
    assert "budget_up" in output
    assert "broad_scale" in output


def test_validate_safe_run_accepts_empty_copy_area_with_growth_pending(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        (
            '2026-06-08 今天广告动作 id="today-ad-actions-all" 待处理</span><strong>0</strong> '
            "复制到广告后台 无待确认动作 当前没有需要复制执行的广告动作 小预算投词 待确认 1 "
            "市场调查 提交今日数据 系统结论 融合诊断"
        ),
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 0
    assert "safe-run outputs validated" in output


def test_validate_safe_run_blocks_keyword_review_rows_missing_contract_fields(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "dimmer desk lamp",
            }
        ],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "keyword action review row" in output
    assert "missing fields" in output


def test_validate_safe_run_blocks_action_review_rows_missing_contract_fields(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        action_review_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "product_name": "Demo desk lamp",
            }
        ],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "action review row" in output
    assert "missing fields" in output


def test_validate_safe_run_accepts_action_review_contract_fields(capsys, tmp_path) -> None:
    from src.autoopt_feedback import ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "product_name": "Demo desk lamp",
        }
    )
    review_row.update(add_action_identity(review_row, "bid_down", "product"))
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        action_review_rows=[review_row],
        fill_blank_review_action_ids=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 0
    assert "safe-run outputs validated" in output


def test_validate_safe_run_blocks_action_review_blank_action_id(capsys, tmp_path) -> None:
    from src.autoopt_feedback import ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0BLANKACTION",
            "product_name": "Blank action product",
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        action_review_rows=[review_row],
        fill_blank_review_action_ids=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "UK action review row 1 Blank action product missing action_id value" in output


def test_validate_safe_run_blocks_action_review_excel_missing_contract_fields(capsys, tmp_path) -> None:
    from src.autoopt_feedback import ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "accept-keyword-review-contract",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "product_name": "Demo desk lamp",
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        action_review_rows=[review_row],
        excel_action_headers=["marketplace", "sku", "asin", "product_name"],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 执行后效果复盘 missing fields" in output


def test_validate_safe_run_blocks_market_action_review_excel_missing_contract_fields(capsys, tmp_path) -> None:
    from src.autoopt_feedback import ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "accept-effective-promoted-review",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "product_name": "Demo desk lamp",
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        action_review_rows=[review_row],
        excel_market_action_headers=["marketplace", "sku", "asin", "product_name"],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet UK_执行后复盘 missing fields" in output


def test_validate_safe_run_blocks_action_review_excel_missing_data_rows(capsys, tmp_path) -> None:
    from src.autoopt_feedback import ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "accept-effective-integer-percent-review",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "product_name": "Demo desk lamp",
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        action_review_rows=[review_row],
        excel_action_rows=[],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 执行后效果复盘 has 0 data rows, expected exactly 1" in output


def test_validate_safe_run_blocks_action_review_excel_extra_data_rows(capsys, tmp_path) -> None:
    from src.autoopt_feedback import ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "product_name": "Demo desk lamp",
        }
    )
    extra_row = dict(review_row)
    extra_row["asin"] = "B0STALE999"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        action_review_rows=[review_row],
        excel_action_rows=[review_row, extra_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 执行后效果复盘 has 2 data rows, expected exactly 1" in output


def test_validate_safe_run_accepts_keyword_review_contract_fields(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "dimmer desk lamp",
        }
    )
    review_row.update(add_action_identity(review_row, "bid_down", "search_term"))
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
        fill_blank_review_action_ids=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 0
    assert "safe-run outputs validated" in output


def test_validate_safe_run_blocks_keyword_review_blank_action_id(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0BLANKACTION",
            "search_term_or_target": "blank action id term",
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
        fill_blank_review_action_ids=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "UK keyword action review row 1 blank action id term missing action_id value" in output


def test_validate_safe_run_blocks_keyword_review_excel_missing_contract_fields(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "dimmer desk lamp",
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
        excel_keyword_headers=["marketplace", "sku", "asin", "search_term_or_target"],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 词级执行复盘 missing fields" in output


def test_validate_safe_run_blocks_market_keyword_review_excel_missing_contract_fields(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "US",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "dimmer desk lamp",
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
        excel_market_keyword_headers=["marketplace", "sku", "asin", "search_term_or_target"],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet US_词级执行复盘 missing fields" in output


def test_validate_safe_run_blocks_keyword_review_excel_halo_only_effective(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    json_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    json_row.update(
        {
            "action_id": "accept-effective-promoted-review",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "promoted term",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
        }
    )
    excel_row = dict(json_row)
    excel_row.update(
        {
            "search_term_or_target": "halo term",
            "promoted_conversion_improved": False,
            "halo_only_conversion": True,
            "target_sku_not_converted": True,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[json_row],
        excel_keyword_rows=[excel_row],
        excel_keyword_headers=[*list(json_row.keys()), "judgement"],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel review row 1 halo term marks halo-only conversion as effective" in output


def test_validate_safe_run_blocks_halo_only_keyword_review_marked_effective(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "halo term",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": False,
            "halo_only_conversion": True,
            "target_sku_not_converted": True,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "marks halo-only conversion as effective" in output


def test_validate_safe_run_blocks_halo_only_keyword_review_with_positive_display_outcome(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "halo display term",
            "outcome": "初步有效",
            "days_since_execution": 8,
            "promoted_conversion_improved": False,
            "halo_only_conversion": True,
            "target_sku_not_converted": True,
            **_effective_review_metrics(),
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "halo display term marks halo-only conversion as effective" in output


def test_validate_safe_run_blocks_halo_only_keyword_review_with_positive_judgement(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "halo judgement term",
            "judgement": "初步有效",
            "days_since_execution": 8,
            "promoted_conversion_improved": False,
            "halo_only_conversion": True,
            "target_sku_not_converted": True,
            **_effective_review_metrics(),
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "halo judgement term marks halo-only conversion as effective" in output


def test_validate_safe_run_blocks_positive_rule_adjustment_without_promoted_sku(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "positive policy term",
            "review_outcome": "not_ready",
            "outcome": "待7天确认",
            "judgement": "待7天确认",
            "effectiveness_score": "",
            "days_since_execution": 8,
            "rule_adjustment": "可保留当前竞价",
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": True,
            **_effective_review_metrics(),
            "current_7d_promoted_ad_orders": 0,
            "current_14d_promoted_ad_orders": 0,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "positive policy term marks target SKU not converted as effective" in output


def test_validate_safe_run_blocks_product_review_effective_without_promoted_conversion(capsys, tmp_path) -> None:
    from src.autoopt_feedback import ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "product_name": "Demo desk lamp",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        action_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "effective outcome missing promoted SKU conversion" in output


def test_validate_safe_run_blocks_review_window_overstating_seven_day_readiness(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "overstated-7d-window",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "overstated review window term",
            "review_outcome": "not_ready",
            "outcome": "待7天确认",
            "days_since_execution": 4,
            "review_window": "7d_check",
            "review_status": "待7天复盘",
            "review_phase": "not_ready",
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "claims 7-day review window before 7 days" in output


def test_validate_safe_run_accepts_effective_review_with_promoted_conversion(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "promoted term",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
        }
    )
    review_row.update(add_action_identity(review_row, "bid_down", "search_term"))
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 0
    assert "safe-run outputs validated" in output


def test_validate_safe_run_accepts_effective_review_with_integer_percent_under_default_target(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "integer percent term",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
            "current_7d_acos": 8,
            "current_7d_tacos": 8,
            "current_7d_target_acos": "",
        }
    )
    review_row.update(add_action_identity(review_row, "bid_down", "search_term"))
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 0
    assert "safe-run outputs validated" in output


def test_validate_safe_run_blocks_effective_review_with_integer_percent_above_default_target(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "integer percent high term",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
            "current_7d_acos": 8,
            "current_7d_tacos": 12,
            "current_7d_target_acos": "",
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "effective outcome has TACOS above target" in output


def test_validate_safe_run_blocks_effective_review_without_positive_promoted_orders(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "promoted term",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
            "current_7d_promoted_ad_orders": 0,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "effective outcome missing positive promoted SKU orders" in output


def test_validate_safe_run_blocks_effective_review_without_numeric_tacos(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "promoted term",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
            "current_7d_tacos": "",
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "effective outcome missing numeric current_7d_tacos" in output


def test_validate_safe_run_blocks_effective_review_without_14d_support_metrics(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "promoted term",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
            "current_14d_total_orders": "",
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "effective outcome missing numeric current_14d_total_orders" in output


def test_validate_safe_run_blocks_effective_review_with_14d_promoted_orders_below_7d(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "conflicting promoted term",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
            "current_14d_promoted_ad_orders": 1,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "effective outcome has 14-day promoted SKU orders below 7-day orders" in output


def test_validate_safe_run_blocks_effective_review_with_promoted_orders_above_ad_orders(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "overstated promoted term",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
            "current_7d_ad_orders": 1,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "effective outcome has promoted SKU orders above ad orders" in output


def test_validate_safe_run_blocks_not_ready_review_with_promoted_orders_above_ad_orders(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "not ready overstated promoted term",
            "review_outcome": "not_ready",
            "outcome": "待7天确认",
            "judgement": "待7天确认",
            "effectiveness_score": "",
            "days_since_execution": 8,
            "review_window": "7天后复盘",
            "effect_evidence": "7天复盘口径，等待人工确认。",
            "review_phase": "7d",
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
            "current_7d_ad_orders": 1,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "review metrics have promoted SKU orders above ad orders" in output


def test_validate_safe_run_blocks_effective_review_missing_ad_order_denominator(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "missing ad order denominator term",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
            "current_7d_ad_orders": "",
            "current_14d_ad_orders": "",
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "effective outcome missing numeric current_7d_ad_orders" in output
    assert "effective outcome missing numeric current_14d_ad_orders" in output


def test_validate_safe_run_blocks_effective_review_with_promoted_orders_above_total_orders(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "overstated total term",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
            "current_7d_ad_orders": 3,
            "current_7d_total_orders": 1,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "effective outcome has promoted SKU orders above total orders" in output


def test_validate_safe_run_blocks_effective_review_with_14d_promoted_orders_above_ad_orders(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "overstated 14d ad term",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
            "current_14d_ad_orders": 2,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "effective outcome has 14-day promoted SKU orders above ad orders" in output


def test_validate_safe_run_blocks_effective_review_with_14d_promoted_orders_above_total_orders(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "overstated 14d total term",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
            "current_14d_ad_orders": 4,
            "current_14d_total_orders": 2,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "effective outcome has 14-day promoted SKU orders above total orders" in output


def test_validate_safe_run_blocks_effective_review_without_positive_total_orders(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "zero total term",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
            "current_7d_total_orders": 0,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "effective outcome missing positive total orders" in output


def test_validate_safe_run_blocks_excel_effective_review_with_tacos_above_target(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    json_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    json_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "promoted term",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
        }
    )
    excel_row = dict(json_row)
    excel_row.update(
        {
            "search_term_or_target": "high tacos term",
            "current_7d_target_acos": 0.10,
            "current_7d_acos": 0.08,
            "current_7d_tacos": 0.18,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[json_row],
        excel_keyword_rows=[excel_row],
        excel_keyword_headers=[*list(json_row.keys()), "judgement"],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel review row 1 high tacos term effective outcome has TACOS above target" in output


def test_validate_safe_run_blocks_effective_review_before_seven_days(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "early term",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 3,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "marks effective before 7-day review window" in output


def test_validate_safe_run_treats_percent_effectiveness_score_as_positive(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "early percent score term",
            "effectiveness_score": "1%",
            "days_since_execution": 3,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "marks effective before 7-day review window" in output


def test_validate_safe_run_blocks_keyword_review_in_wrong_marketplace_snapshot(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
    )
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "US",
            "sku": "SKU-1",
            "asin": "B0WRONGMARKET",
            "search_term_or_target": "wrong market term",
            "days_since_execution": 4,
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
        }
    )
    analysis_path = safe_dir / "latest_analysis.json"
    analysis = json.loads(analysis_path.read_text(encoding="utf-8"))
    analysis["marketplace_results"][0]["report_view_snapshot"]["keyword_action_effect_review_rows"] = [review_row]
    analysis_path.write_text(json.dumps(analysis, ensure_ascii=False), encoding="utf-8")

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "UK keyword action review row 1 wrong market term contains US marketplace data" in output


def test_validate_safe_run_blocks_keyword_review_blank_marketplace(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "",
            "sku": "SKU-1",
            "asin": "B0BLANKMARKET",
            "search_term_or_target": "blank marketplace term",
            "days_since_execution": 4,
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            **_effective_review_metrics(),
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "UK keyword action review row 1 blank marketplace term missing marketplace value" in output


def test_validate_safe_run_blocks_review_window_missing_tracked_metrics(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0MISSINGMETRICS",
            "search_term_or_target": "missing metrics term",
            "days_since_execution": 4,
            "review_window": "3d_check",
            "review_status": "待7天复盘",
            "review_phase": "3d",
            "review_outcome": "not_ready",
            "outcome": "待7天确认",
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "UK review row 1 missing metrics term reached review window but missing numeric current_7d_promoted_ad_orders" in output


def test_validate_safe_run_allows_nonfinal_review_with_explicit_missing_ratio_evidence(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "UK||SKU-1||B0MISSINGRATIO||search_term||missing ratio term||bid_down",
            "normalized_action": "bid_down",
            "action_scope": "search_term",
            "action_detail": "降竞价 10%",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0MISSINGRATIO",
            "search_term_or_target": "missing ratio term",
            "days_since_execution": 8,
            "review_window": "7天后复盘",
            "review_status": "可做7天复盘",
            "review_phase": "day_7_review",
            "review_outcome": "needs_manual_review",
            "outcome": "待人工复查",
            "effectiveness_score": -1,
            "effect_evidence": "7天：点击 3，花费 0.66，订单 1，本 SKU 单 1，光环单 0，销售 0.00，ACOS N/A；总单 0",
            "effect_metrics": "7天：点击 3，订单 1，本 SKU 单 1，总单 0，TACOS 不可计算",
            "block_reason": "本 SKU 有单但缺少TACOS、库存，需人工复查",
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            "current_7d_clicks": 3,
            "current_7d_spend": 0.66,
            "current_7d_ad_orders": 1,
            "current_7d_promoted_ad_orders": 1,
            "current_7d_acos": 0.08,
            "current_7d_tacos": "",
            "current_7d_total_orders": 0,
            "current_7d_available_stock": "",
            "current_14d_clicks": 4,
            "current_14d_spend": 0.88,
            "current_14d_ad_orders": 1,
            "current_14d_promoted_ad_orders": 1,
            "current_14d_acos": 0.09,
            "current_14d_tacos": "",
            "current_14d_total_orders": 0,
            "current_14d_available_stock": "",
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)

    assert code == 0


def test_validate_safe_run_allows_insufficient_sample_review_missing_stock(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "UK||SKU-1||B0ZEROSAMPLE||search_term||zero sample term||bid_down",
            "normalized_action": "bid_down",
            "action_scope": "search_term",
            "action_detail": "降竞价15%-25%",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0ZEROSAMPLE",
            "search_term_or_target": "zero sample term",
            "days_since_execution": 40,
            "review_window": "7天后复盘",
            "review_status": "可做7天复盘",
            "review_phase": "day_7_review",
            "review_outcome": "insufficient_sample",
            "outcome": "待观察",
            "effectiveness_score": 0,
            "effect_evidence": "7天：点击 0，花费 0.00，订单 0，本 SKU 单 0，光环单 0，销售 0.00，ACOS N/A；14天：点击 0，花费 0.00，订单 0，本 SKU 单 0，光环单 0，销售 0.00，ACOS N/A",
            "effect_metrics": "7天：点击 0，花费 0.00，订单 0，本 SKU 单 0，光环单 0",
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": True,
            "current_7d_clicks": 0,
            "current_7d_spend": 0,
            "current_7d_ad_orders": 0,
            "current_7d_promoted_ad_orders": 0,
            "current_7d_acos": 0,
            "current_7d_tacos": 0,
            "current_7d_total_orders": 0,
            "current_7d_available_stock": "",
            "current_14d_clicks": 0,
            "current_14d_spend": 0,
            "current_14d_ad_orders": 0,
            "current_14d_promoted_ad_orders": 0,
            "current_14d_acos": 0,
            "current_14d_tacos": "",
            "current_14d_total_orders": 0,
            "current_14d_available_stock": "",
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)

    assert code == 0


def test_validate_safe_run_blocks_ineffective_review_window_missing_tracked_metrics(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0INEFFMETRICS",
            "search_term_or_target": "ineffective missing metrics term",
            "days_since_execution": 8,
            "review_window": "7天后复盘",
            "review_status": "可做7天复盘",
            "review_phase": "day_7_review",
            "review_outcome": "ineffective",
            "outcome": "暂未改善",
            "effectiveness_score": -2,
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": True,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert (
        "UK review row 1 ineffective missing metrics term "
        "reached review window but missing numeric current_7d_promoted_ad_orders"
    ) in output


def test_validate_safe_run_blocks_ineffective_review_before_seven_days(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "early bad term",
            "review_outcome": "ineffective",
            "effectiveness_score": -2,
            "days_since_execution": 4,
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": True,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "marks ineffective before 7-day review window" in output


def test_validate_safe_run_blocks_excel_ineffective_review_before_seven_days(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    json_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    json_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "early excel bad term",
            "review_outcome": "not_ready",
            "effectiveness_score": "",
            "effect_evidence": "3天复查口径，7天结论待补；继续观察。",
            "days_since_execution": 4,
            **_effective_review_metrics(),
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": True,
        }
    )
    excel_row = dict(json_row)
    excel_row["review_outcome"] = "ineffective"
    excel_row["effectiveness_score"] = -2
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[json_row],
        excel_keyword_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel review row 1 early excel bad term marks ineffective before 7-day review window" in output


def test_validate_safe_run_blocks_positive_display_outcome_before_seven_days(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "early display term",
            "outcome": "有改善迹象",
            "review_outcome": "not_ready",
            "effectiveness_score": "",
            "effect_evidence": "3天复查口径，7天结论待补；继续观察。",
            "days_since_execution": 4,
            **_effective_review_metrics(),
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "shows positive display outcome before 7-day review window" in output


def test_validate_safe_run_blocks_keep_current_outcome_before_seven_days(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "early keep term",
            "outcome": "可保留",
            "review_outcome": "not_ready",
            "effectiveness_score": "",
            "effect_evidence": "3天复查口径，7天结论待补；继续观察。",
            "days_since_execution": 4,
            **_effective_review_metrics(),
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "early keep term shows positive display outcome before 7-day review window" in output


def test_validate_safe_run_blocks_excel_positive_display_outcome_before_seven_days(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    json_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    json_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "early display term",
            "outcome": "待7天确认",
            "review_outcome": "not_ready",
            "effectiveness_score": "",
            "effect_evidence": "3天复查口径，7天结论待补；继续观察。",
            "days_since_execution": 4,
            **_effective_review_metrics(),
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
        }
    )
    excel_row = dict(json_row)
    excel_row["outcome"] = "初步有效"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[json_row],
        excel_keyword_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel review row 1 early display term shows positive display outcome before 7-day review window" in output


def test_validate_safe_run_blocks_excel_keep_current_judgement_before_seven_days(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    json_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    json_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "early excel keep term",
            "outcome": "待7天确认",
            "review_outcome": "not_ready",
            "effectiveness_score": "",
            "effect_evidence": "3天复查口径，7天结论待补；继续观察。",
            "days_since_execution": 4,
            **_effective_review_metrics(),
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
        }
    )
    excel_row = dict(json_row)
    excel_row["outcome"] = ""
    excel_row["judgement"] = "可保留"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[json_row],
        excel_keyword_rows=[excel_row],
        excel_keyword_headers=[*list(json_row.keys()), "judgement"],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel review row 1 early excel keep term shows positive display outcome before 7-day review window" in output


def test_validate_safe_run_blocks_excel_positive_judgement_before_seven_days(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    json_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    json_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "early judgement term",
            "outcome": "待7天确认",
            "review_outcome": "not_ready",
            "effectiveness_score": "",
            "effect_evidence": "3天复查口径，7天结论待补；继续观察。",
            "days_since_execution": 4,
            **_effective_review_metrics(),
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
        }
    )
    excel_row = dict(json_row)
    excel_row["outcome"] = ""
    excel_row["judgement"] = "有改善迹象"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[json_row],
        excel_keyword_rows=[excel_row],
        excel_keyword_headers=[*list(json_row.keys()), "judgement"],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel review row 1 early judgement term shows positive display outcome before 7-day review window" in output


def test_validate_safe_run_blocks_negative_display_outcome_before_three_days(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "too early bad term",
            "outcome": "暂未改善",
            "review_outcome": "not_ready",
            "effectiveness_score": "",
            "days_since_execution": 2,
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": True,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "shows negative display outcome before 3-day review window" in output


def test_validate_safe_run_blocks_stop_append_outcome_before_three_days(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "too early stop append term",
            "outcome": "停止追加",
            "review_outcome": "not_ready",
            "effectiveness_score": "",
            "days_since_execution": 2,
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": True,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "too early stop append term shows negative display outcome before 3-day review window" in output


def test_validate_safe_run_blocks_excel_negative_display_outcome_before_three_days(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    json_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    json_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "too early excel bad term",
            "outcome": "样本不足",
            "review_outcome": "not_ready",
            "effectiveness_score": "",
            "days_since_execution": 2,
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": True,
        }
    )
    excel_row = dict(json_row)
    excel_row["outcome"] = "暂未改善"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[json_row],
        excel_keyword_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel review row 1 too early excel bad term shows negative display outcome before 3-day review window" in output


def test_validate_safe_run_blocks_excel_deprioritize_judgement_before_three_days(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    json_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    json_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "too early excel deprioritize term",
            "outcome": "待7天确认",
            "review_outcome": "not_ready",
            "effectiveness_score": "",
            "days_since_execution": 2,
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": True,
        }
    )
    excel_row = dict(json_row)
    excel_row["outcome"] = ""
    excel_row["judgement"] = "降优先级"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[json_row],
        excel_keyword_rows=[excel_row],
        excel_keyword_headers=[*list(json_row.keys()), "judgement"],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert (
        "Excel review row 1 too early excel deprioritize term "
        "shows negative display outcome before 3-day review window"
    ) in output


def test_validate_safe_run_blocks_insufficient_sample_after_seven_days_with_sufficient_traffic(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "stale sample term",
            "outcome": "样本不足",
            "review_outcome": "insufficient_sample",
            "effectiveness_score": 0,
            "days_since_execution": 8,
            "current_7d_clicks": 9,
            "current_7d_spend": 5.5,
            "current_7d_promoted_ad_orders": 0,
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": True,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "stale sample term shows insufficient sample after 7-day sufficient traffic" in output


def test_validate_safe_run_blocks_keyword_review_excel_identity_mismatch(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    json_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    json_row.update(
        {
            "action_id": "kw-1",
            "normalized_action": "bid_down",
            "action_scope": "keyword",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "dimmer desk lamp",
            "executed_at": "2026-06-01",
            "report_date": "2026-06-01",
            "review_outcome": "not_ready",
        }
    )
    excel_row = dict(json_row)
    excel_row["search_term_or_target"] = "unrelated storage box"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[json_row],
        excel_keyword_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 词级执行复盘 row identity mismatch" in output
    assert "dimmer desk lamp" in output
    assert "unrelated storage box" in output


def test_validate_safe_run_blocks_keyword_review_excel_content_mismatch(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    json_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    json_row.update(
        {
            "action_id": "kw-content-1",
            "normalized_action": "bid_down",
            "action_scope": "keyword",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "dimmer desk lamp",
            "executed_at": "2026-06-01",
            "report_date": "2026-06-01",
            "review_outcome": "not_ready",
            "current_7d_promoted_ad_orders": 2,
            "current_7d_acos": 0.08,
            "current_7d_tacos": 0.05,
            "current_7d_total_orders": 3,
            "current_7d_available_stock": 18,
        }
    )
    excel_row = dict(json_row)
    excel_row["current_7d_promoted_ad_orders"] = 0
    excel_row["current_7d_acos"] = 0.3
    excel_row["current_7d_tacos"] = 0.25
    excel_row["current_7d_total_orders"] = 0
    excel_row["current_7d_available_stock"] = 0
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[json_row],
        excel_keyword_rows=[excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet 词级执行复盘 field mismatch for keyword action reviews" in output
    assert "field current_7d_promoted_ad_orders: expected 2.0, got 0.0" in output


def test_validate_safe_run_blocks_market_keyword_review_excel_content_mismatch(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    json_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    json_row.update(
        {
            "action_id": "kw-market-content-1",
            "normalized_action": "bid_down",
            "action_scope": "keyword",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "dimmer desk lamp market",
            "executed_at": "2026-06-01",
            "report_date": "2026-06-01",
            "review_outcome": "not_ready",
            "current_7d_promoted_ad_orders": 2,
            "current_7d_acos": 0.08,
            "current_7d_tacos": 0.05,
            "current_7d_total_orders": 3,
            "current_7d_available_stock": 18,
        }
    )
    market_excel_row = dict(json_row)
    market_excel_row["current_7d_promoted_ad_orders"] = 0
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[json_row],
        excel_market_keyword_rows=[market_excel_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet UK_词级执行复盘 field mismatch for UK keyword action reviews" in output
    assert "field current_7d_promoted_ad_orders: expected 2.0, got 0.0" in output


def test_validate_safe_run_blocks_market_keyword_review_excel_wrong_marketplace(capsys, tmp_path) -> None:
    from openpyxl import load_workbook
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    json_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    json_row.update(
        {
            "action_id": "kw-market-wrong-market-1",
            "normalized_action": "bid_down",
            "action_scope": "keyword",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "wrong market excel term",
            "executed_at": "2026-06-01",
            "report_date": "2026-06-01",
            "review_outcome": "not_ready",
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[json_row],
    )
    for workbook_path in safe_dir.glob("amazon_ops_report_*.xlsx"):
        workbook = load_workbook(workbook_path)
        sheet = workbook["UK_词级执行复盘"]
        headers = [cell.value for cell in sheet[1]]
        sheet.cell(row=2, column=headers.index("marketplace") + 1).value = "US"
        workbook.save(workbook_path)
        workbook.close()

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "Excel sheet UK_词级执行复盘 row 1 wrong market excel term contains US marketplace data" in output


def test_validate_safe_run_blocks_marketplace_html_missing_keyword_review_identity(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "kw-market-1",
            "normalized_action": "bid_down",
            "action_scope": "keyword",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0MARKETHTML",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "market missing keyword",
            "executed_at": "2026-06-01",
            "report_date": "2026-06-01",
            "review_outcome": "not_ready",
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
        include_market_html_snapshot_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "uk_report.html missing keyword review row 1 target market missing keyword" in output


def test_validate_safe_run_blocks_marketplace_html_missing_review_metric_token(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "kw-market-metric-1",
            "normalized_action": "bid_up",
            "action_scope": "keyword",
            "action_type": "加竞价",
            "action_detail": "加竞价 5%-10%",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0MARKETMETRIC",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "market metric keyword",
            "executed_at": "2026-05-20T09:00:00",
            "report_date": "2026-05-20",
            "review_date": "2026-06-08",
            "days_since_execution": 19,
            "review_window": "7天后复盘",
            "review_outcome": "effective",
            "judgement": "有改善迹象",
            "outcome": "有改善迹象",
            "effectiveness_score": 2,
            "effect_evidence": "7 天本 SKU 有单",
            "review_status": "可做7天复盘",
            "review_phase": "7d_review",
            "cooldown_status": "观察中",
            "cooldown_until": "",
            "block_reason": "",
            "rule_adjustment": "保留当前竞价",
            "learning_scope": "keyword",
            "current_7d_clicks": 8,
            "current_7d_spend": 3.2,
            "current_7d_ad_orders": 2,
            "current_7d_ad_sales": 30,
            "current_7d_promoted_ad_sales": 30,
            "current_7d_halo_ad_orders": 0,
            "current_7d_halo_ad_sales": 0,
            "current_14d_clicks": 14,
            "current_14d_spend": 5.1,
            "current_14d_ad_orders": 3,
            "current_14d_ad_sales": 45,
            "current_14d_promoted_ad_sales": 45,
            "current_14d_halo_ad_orders": 0,
            "current_14d_halo_ad_sales": 0,
            "current_14d_tacos": 0.09,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            "attribution_effect_status": "promoted_sku_converted",
            "attribution_effect_note": "本 SKU 成交优先",
            **_effective_review_metrics(),
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
        include_market_html_snapshot_tokens=False,
    )
    (safe_dir / "uk_report.html").write_text(
        (
            '2026-06-08 亚马逊运营日报｜UK 广告状态 数据质量与增强数据 站点状态 市场调查 '
            '执行后效果复盘 id="product-operation-cards" id="frontend-evidence-status" id="action-effect-review" '
            "market metric keyword B0MARKETMETRIC"
        ),
        encoding="utf-8",
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "uk_report.html missing keyword review row 1 metric token 本 SKU 单 2" in output


def test_validate_safe_run_blocks_marketplace_html_missing_review_14d_metric_token(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "kw-market-14d-metric-1",
            "normalized_action": "bid_down",
            "action_scope": "keyword",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0MARKET14D",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "market 14d metric keyword",
            "executed_at": "2026-06-01",
            "report_date": "2026-06-01",
            "review_window": "7天后复盘",
            "judgement": "待人工复查",
            "days_since_execution": 8,
            "review_outcome": "not_ready",
            "effect_evidence": "7天复盘口径，等待人工确认。",
            "review_phase": "7d",
            **_effective_review_metrics(),
            "current_14d_promoted_ad_orders": 3,
            "current_14d_acos": 0.09,
            "current_14d_tacos": 0.07,
            "current_14d_total_orders": 5,
            "current_14d_available_stock": 18,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
        include_market_html_snapshot_tokens=False,
    )
    (safe_dir / "uk_report.html").write_text(
        (
            '2026-06-08 亚马逊运营日报｜UK 广告状态 数据质量与增强数据 站点状态 市场调查 '
            '执行后效果复盘 id="product-operation-cards" id="frontend-evidence-status" id="action-effect-review" '
            "market 14d metric keyword B0MARKET14D 本 SKU 单 2 ACOS 0.12 目标 ACOS 0.2 TACOS 0.08 总单 4 库存 18 "
            "14天 本 SKU 单 3 14天 ACOS 0.09 14天 总单 5 14天 库存 18"
        ),
        encoding="utf-8",
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "uk_report.html missing keyword review row 1 metric token 14天 TACOS 0.07" in output


def test_validate_safe_run_blocks_marketplace_html_missing_product_decision_blocking_reason(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        asin="B0MARKETGATE1",
        product_name="Market Gate Product",
        frontend_blocking_reasons=["地区待确认，当前前台证据不能用于放量"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
    )
    _write_minimal_safe_run(
        safe_dir,
        (
            '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 '
            "B0MARKETGATE1 Market Gate Product 地区待确认，当前前台证据不能用于放量 拦截加竞价 拦截加预算 拦截放量"
        ),
        product_final_decision_rows=[decision_row],
        include_market_html_snapshot_tokens=False,
    )
    (safe_dir / "latest_recommendations.html").write_text(
        (
            safe_dir / "latest_recommendations.html"
        ).read_text(encoding="utf-8")
        + (
            '<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-1" '
            'data-product-decision-asin="B0MARKETGATE1">'
            "B0MARKETGATE1 Market Gate Product 地区待确认，当前前台证据不能用于放量 "
            "拦截加竞价 拦截加预算 拦截放量 拦截低预算精准测试 已自动检查 仅背景参考</article>"
        ),
        encoding="utf-8",
    )
    (safe_dir / "uk_report.html").write_text(
        (
            safe_dir / "uk_report.html"
        ).read_text(encoding="utf-8")
        + (
            '<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-1" '
            'data-product-decision-asin="B0MARKETGATE1">'
            "B0MARKETGATE1 Market Gate Product 拦截加竞价 拦截加预算 拦截放量 "
            "拦截低预算精准测试 已自动检查 仅背景参考</article>"
        ),
        encoding="utf-8",
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert (
        "uk_report.html missing product decision row 1 blocking reason "
        "地区待确认，当前前台证据不能用于放量"
    ) in output


def test_validate_safe_run_blocks_marketplace_html_missing_product_decision_blocked_action_label(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        asin="B0MARKETGATE2",
        product_name="Market Blocked Action Product",
        frontend_blocking_reasons=["竞品可比性不足"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
    )
    _write_minimal_safe_run(
        safe_dir,
        (
            '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 '
            "B0MARKETGATE2 Market Blocked Action Product 竞品可比性不足 拦截加预算 拦截放量"
        ),
        product_final_decision_rows=[decision_row],
        include_market_html_snapshot_tokens=False,
    )
    (safe_dir / "latest_recommendations.html").write_text(
        (
            safe_dir / "latest_recommendations.html"
        ).read_text(encoding="utf-8")
        + (
            '<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-1" '
            'data-product-decision-asin="B0MARKETGATE2">'
            "B0MARKETGATE2 Market Blocked Action Product 竞品可比性不足 "
            "拦截加竞价 拦截加预算 拦截放量 拦截低预算精准测试 已自动检查 仅背景参考</article>"
        ),
        encoding="utf-8",
    )
    (safe_dir / "uk_report.html").write_text(
        (
            safe_dir / "uk_report.html"
        ).read_text(encoding="utf-8")
        + (
            '<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-1" '
            'data-product-decision-asin="B0MARKETGATE2">'
            "B0MARKETGATE2 Market Blocked Action Product 竞品可比性不足 "
            "拦截加预算 拦截放量 拦截低预算精准测试 已自动检查 仅背景参考</article>"
        ),
        encoding="utf-8",
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "uk_report.html missing product decision row 1 blocked action label bid_up" in output


def test_validate_safe_run_blocks_marketplace_html_missing_low_budget_blocked_action_label(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        asin="B0MARKETLOWBUDGET",
        product_name="Market Low Budget Blocked Product",
        frontend_blocking_reasons=["前台证据不足，需刷新后再放量"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
    )
    _write_minimal_safe_run(
        safe_dir,
        (
            '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 '
            "B0MARKETLOWBUDGET Market Low Budget Blocked Product 前台证据不足，需刷新后再放量 "
            "拦截加竞价 拦截加预算 拦截放量"
        ),
        product_final_decision_rows=[decision_row],
        include_market_html_snapshot_tokens=False,
    )
    (safe_dir / "latest_recommendations.html").write_text(
        (
            safe_dir / "latest_recommendations.html"
        ).read_text(encoding="utf-8")
        + (
            '<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-1" '
            'data-product-decision-asin="B0MARKETLOWBUDGET">'
            "B0MARKETLOWBUDGET Market Low Budget Blocked Product 前台证据不足，需刷新后再放量 "
            "拦截加竞价 拦截加预算 拦截放量 禁止低预算精准测试 已自动检查 仅背景参考</article>"
        ),
        encoding="utf-8",
    )
    (safe_dir / "uk_report.html").write_text(
        (
            safe_dir / "uk_report.html"
        ).read_text(encoding="utf-8")
        + (
            '<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-1" '
            'data-product-decision-asin="B0MARKETLOWBUDGET">'
            "B0MARKETLOWBUDGET Market Low Budget Blocked Product 前台证据不足，需刷新后再放量 "
            "拦截加竞价 拦截加预算 拦截放量 已自动检查 仅背景参考</article>"
        ),
        encoding="utf-8",
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "uk_report.html missing product decision row 1 blocked action label create_exact_low_budget" in output


def test_validate_safe_run_blocks_marketplace_html_frontend_tokens_bound_to_wrong_article(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    decision_row = _product_decision_row(
        marketplace="UK",
        asin="B0MARKETFRONTBIND",
        product_name="Market Frontend Evidence Product",
        frontend_blocking_reasons=["前台缓存或读取失败待确认，仅背景参考，不能单独支持放量"],
        today_blocked_actions=["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
        frontend_check_status="沿用 2026-06-20 前台数据",
        frontend_cache_used=True,
        frontend_evidence_tier="仅背景参考",
        frontend_evidence_display_tier="仅背景参考",
        frontend_search_status="已读取部分结果",
        frontend_search_partial_evidence=True,
    )
    _write_minimal_safe_run(
        safe_dir,
            (
                '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断 '
                '<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-1" '
                'data-product-decision-asin="B0MARKETFRONTBIND">'
                "B0MARKETFRONTBIND Market Frontend Evidence Product 前台缓存或读取失败待确认，仅背景参考，不能单独支持放量 "
                "拦截加竞价 拦截加预算 拦截放量 拦截低预算精准测试 "
                "沿用 2026-06-20 前台数据 仅背景参考 搜索页：已读取部分结果</article>"
            ),
        product_final_decision_rows=[decision_row],
        include_market_html_snapshot_tokens=False,
    )
    (safe_dir / "uk_report.html").write_text(
        (
            safe_dir / "uk_report.html"
        ).read_text(encoding="utf-8")
        + (
            '<article data-product-decision-marketplace="UK" data-product-decision-sku="SKU-1" '
            'data-product-decision-asin="B0MARKETFRONTBIND">'
            "B0MARKETFRONTBIND Market Frontend Evidence Product 前台缓存或读取失败待确认，仅背景参考，不能单独支持放量 "
            "拦截加竞价 拦截加预算 拦截放量 拦截低预算精准测试</article>"
        )
        + (
            '<article data-product-decision-marketplace="US" data-product-decision-sku="SKU-US" '
            'data-product-decision-asin="B0MARKETFRONTBIND">沿用 2026-06-20 前台数据 仅背景参考 搜索页：已读取部分结果</article>'
        ),
        encoding="utf-8",
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert (
        "uk_report.html product decision row 1 frontend status "
        "沿用 2026-06-20 前台数据 not bound to ASIN B0MARKETFRONTBIND"
    ) in output


def test_validate_safe_run_blocks_summary_missing_keyword_review_identity(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "kw-summary-1",
            "normalized_action": "bid_down",
            "action_scope": "keyword",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0SUMMARYKW",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "summary missing keyword",
            "executed_at": "2026-06-01",
            "report_date": "2026-06-01",
            "review_window": "3天后复盘",
            "judgement": "暂未改善",
            "days_since_execution": 4,
            "effect_metrics": "promoted SKU 0 单",
            "next_step": "不要继续加价",
            "review_outcome": "not_ready",
            "effect_evidence": "3天复查口径，7天结论待补；继续观察。",
            **_effective_review_metrics(),
            "promoted_conversion_improved": False,
            "current_7d_promoted_ad_orders": 0,
            "current_14d_promoted_ad_orders": 0,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
        include_summary_snapshot_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "summary.html missing review watch row 1 target summary missing keyword" in output


def test_validate_safe_run_blocks_summary_missing_frontend_coverage_label(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        include_summary_frontend_coverage_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "summary.html missing frontend coverage token 产品页成功" in output


def test_validate_safe_run_blocks_summary_missing_review_metric_token(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "kw-summary-metric-1",
            "normalized_action": "bid_down",
            "action_scope": "keyword",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0SUMMARYMETRIC",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "summary metric keyword",
            "executed_at": "2026-06-01",
            "report_date": "2026-06-01",
            "review_window": "7天后复盘",
            "judgement": "待人工复查",
            "days_since_execution": 7,
            "review_outcome": "not_ready",
            **_effective_review_metrics(),
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
        include_summary_snapshot_tokens=False,
    )
    (safe_dir / "summary.html").write_text(
        "2026-06-08 三分钟摘要 今日开工结论 今天先做 昨天动作今天要盯 "
        "产品页成功 0/3 卖家精灵自己 ASIN 今日 0/3，缓存 0/3 卖家精灵趋势 0/3 卖家精灵竞品发现 0/3 卖家精灵竞品池 今日 0/3，7天缓存 0/3 竞品 ASIN 反查 今日 0/3，缓存 0/3 Amazon 搜索页辅助验证 0/3 达到放量准入 0/3 弱势止损证据 0/3 证据不足 3/3 市场调查平均完整度 0.0/100 强证据 / 可用证据 0/3 / 0/3 "
        "summary metric keyword 待人工复查 继续观察，等待足够样本。本 SKU 单 2",
        encoding="utf-8",
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "summary.html missing review watch row 1 metric token ACOS 12%" in output


def test_validate_safe_run_blocks_summary_positive_review_without_promoted_conversion_text(
    capsys,
    tmp_path,
) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "kw-summary-no-promoted-1",
            "normalized_action": "bid_up",
            "action_scope": "keyword",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0SUMMARYNOPROMO",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "summary no promoted keyword",
            "executed_at": "2026-06-01",
            "report_date": "2026-06-01",
            "review_window": "7天后复盘",
            "judgement": "初步有效",
            "days_since_execution": 10,
            "review_outcome": "not_ready",
            "next_step": "继续加价观察",
            **_effective_review_metrics(),
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            "current_7d_promoted_ad_orders": 0,
            "current_7d_total_orders": 1,
        }
    )
    data = {
        "marketplace_results": [
            {
                "marketplace": "UK",
                "report_view_snapshot": {
                    "keyword_action_effect_review_rows": [review_row],
                    "action_effect_review_rows": [],
                    "inventory_replenishment_rows": [],
                },
            }
        ]
    }
    text = (
        "2026-06-08 三分钟摘要 今日开工结论 今天先做 昨天动作今天要盯 "
        "summary no promoted keyword 本 SKU 单 0 ACOS 12% TACOS 8% 总单 1 库存 22 "
        "目标 ACOS 20% 14天本 SKU 单 3 14天 ACOS 14% 14天 TACOS 9% 14天总单 7 14天库存 18"
    )

    code = showcase_validate._validate_summary_snapshot_rows(text, data)
    output = capsys.readouterr().out

    assert code == 1
    assert "summary.html missing review watch row 1 judgement 本 SKU 未验证" in output


def test_validate_safe_run_blocks_summary_missing_review_target_acos_token(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "kw-summary-target-acos-1",
            "normalized_action": "bid_down",
            "action_scope": "keyword",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0SUMMARYTARGET",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "summary target acos keyword",
            "executed_at": "2026-06-01",
            "report_date": "2026-06-01",
            "review_window": "7天后复盘",
            "judgement": "待人工复查",
            "days_since_execution": 7,
            "review_outcome": "not_ready",
            **_effective_review_metrics(),
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
        include_summary_snapshot_tokens=False,
    )
    (safe_dir / "summary.html").write_text(
        "2026-06-08 三分钟摘要 今日开工结论 今天先做 昨天动作今天要盯 "
        "产品页成功 0/3 卖家精灵自己 ASIN 今日 0/3，缓存 0/3 卖家精灵趋势 0/3 卖家精灵竞品发现 0/3 卖家精灵竞品池 今日 0/3，7天缓存 0/3 竞品 ASIN 反查 今日 0/3，缓存 0/3 Amazon 搜索页辅助验证 0/3 达到放量准入 0/3 弱势止损证据 0/3 证据不足 3/3 市场调查平均完整度 0.0/100 强证据 / 可用证据 0/3 / 0/3 "
        "summary target acos keyword 待人工复查 继续观察，等待足够样本。"
        "本 SKU 单 2 ACOS 12% TACOS 8% 总单 4 库存 18 "
        "14天本 SKU 单 3 14天 ACOS 9% 14天 TACOS 9% 14天总单 5 14天库存 18",
        encoding="utf-8",
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "summary.html missing review watch row 1 metric token 目标 ACOS 20%" in output


def test_summary_review_metric_tokens_skip_14d_before_7_day_window() -> None:
    row = {
        "days_since_execution": 4,
        "current_7d_promoted_ad_orders": 2,
        "current_7d_acos": 0.12,
        "current_7d_target_acos": 0.2,
        "current_7d_tacos": 0.08,
        "current_7d_total_orders": 4,
        "current_7d_available_stock": 18,
        "current_14d_promoted_ad_orders": 3,
        "current_14d_acos": 0.09,
        "current_14d_tacos": 0.07,
        "current_14d_total_orders": 5,
        "current_14d_available_stock": 20,
    }

    tokens = showcase_validate._summary_review_metric_tokens(row)

    assert "本 SKU 单 2" in tokens
    assert "ACOS 12%" in tokens
    assert not any(token.startswith("14天") for token in tokens)


def test_validate_summary_snapshot_blocks_early_review_metric_token(capsys) -> None:
    row = {
        "marketplace": "UK",
        "search_term_or_target": "summary too early showcase keyword",
        "review_window": "未满3天",
        "judgement": "样本不足",
        "days_since_execution": 2,
        "current_7d_promoted_ad_orders": 1,
        "current_7d_acos": 0.12,
        "current_7d_tacos": 0.08,
        "current_7d_total_orders": 2,
        "current_7d_available_stock": 18,
    }
    data = {
        "marketplace_results": [
            {
                "marketplace": "UK",
                "report_view_snapshot": {
                    "keyword_action_effect_review_rows": [row],
                    "action_effect_review_rows": [],
                    "inventory_replenishment_rows": [],
                },
            }
        ]
    }
    text = (
        "2026-06-08 三分钟摘要 今日开工结论 今天先做 昨天动作今天要盯 "
        "summary too early showcase keyword 未满3天 本 SKU 单 1 ACOS 12%"
    )

    code = showcase_validate._validate_summary_snapshot_rows(text, data)
    output = capsys.readouterr().out

    assert code == 1
    assert "summary.html early review row 1 should not show metric token 本 SKU 单 1" in output


def test_validate_latest_recommendations_blocks_early_review_metric_token(capsys) -> None:
    row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0TOOEARLYHTML",
        "product_name": "Too early product",
        "search_term_or_target": "too early detail keyword",
        "review_window": "未满3天",
        "judgement": "样本不足",
        "days_since_execution": 2,
        "current_7d_promoted_ad_orders": 1,
        "current_7d_acos": 0.12,
        "current_7d_tacos": 0.08,
        "current_7d_total_orders": 2,
        "current_7d_available_stock": 18,
    }
    data = {
        "marketplace_results": [
            {
                "marketplace": "UK",
                "report_view_snapshot": {
                    "keyword_action_effect_review_rows": [row],
                    "action_effect_review_rows": [],
                    "product_final_decision_rows": [],
                    "tomorrow_review_rows": [],
                    "today_task_queue_rows": [],
                    "html_search_term_processing_queue_rows": [],
                },
            }
        ]
    }
    text = (
        '2026-06-08 今日广告动作 id="today-ad-actions-all" 待确认 0 '
        "市场调查 提交今日数据 系统结论 融合诊断 "
        "too early detail keyword B0TOOEARLYHTML 7天 本 SKU 单 1 7天 ACOS 0.12"
    )

    code = showcase_validate._validate_latest_recommendations_snapshot_rows(text, data)
    output = capsys.readouterr().out

    assert code == 1
    assert (
        "latest_recommendations.html keyword review row 1 should not show metric token 7天 本 SKU 单 1"
        in output
    )


def test_validate_latest_recommendations_blocks_anchor_7d_metric_before_7_day_window(capsys) -> None:
    row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0ANCHOR4D",
        "product_name": "Anchor Early Product",
        "search_term_or_target": "anchor early detail keyword",
        "review_window": "3d_check",
        "judgement": "样本不足",
        "days_since_execution": 4,
        "review_data_source": "execution_anchored_daily",
        "pre_7d_start": "2026-06-11",
        "pre_7d_end": "2026-06-17",
        "post_3d_start": "2026-06-18",
        "post_3d_end": "2026-06-20",
        "post_7d_start": "2026-06-18",
        "post_7d_end": "2026-06-24",
        "pre_7d_promoted_ad_orders": 0,
        "pre_7d_total_orders": 1,
        "pre_7d_tacos": 0.12,
        "post_3d_days": 3,
        "post_3d_promoted_ad_orders": 1,
        "post_3d_total_orders": 2,
        "post_3d_acos": 0.10,
        "post_3d_tacos": 0.07,
        "post_3d_available_stock": 18,
        "post_7d_days": 7,
        "post_7d_promoted_ad_orders": 2,
        "post_7d_total_orders": 4,
        "post_7d_acos": 0.12,
        "post_7d_tacos": 0.08,
        "post_7d_available_stock": 18,
    }
    data = {
        "marketplace_results": [
            {
                "marketplace": "UK",
                "report_view_snapshot": {
                    "keyword_action_effect_review_rows": [row],
                    "action_effect_review_rows": [],
                    "product_final_decision_rows": [],
                    "tomorrow_review_rows": [],
                    "today_task_queue_rows": [],
                    "html_search_term_processing_queue_rows": [],
                },
            }
        ]
    }
    text = (
        '2026-06-08 今日广告动作 id="today-ad-actions-all" 待确认 0 '
        "市场调查 提交今日数据 系统结论 融合诊断 "
        "anchor early detail keyword B0ANCHOR4D "
        "执行前7天 2026-06-11 至 2026-06-17 执行后3天 2026-06-18 至 2026-06-20 "
        "执行前本 SKU 单 0 执行前总单 1 执行前 TACOS 0.12 "
        "执行后3天覆盖天数 3 执行后3天本 SKU 单 1 执行后3天总单 2 "
        "执行后3天 ACOS 0.1 执行后3天 TACOS 0.07 执行后3天库存 18 "
        "3 天窗口只做初步判断 执行后本 SKU 单 2"
    )

    code = showcase_validate._validate_latest_recommendations_snapshot_rows(text, data)
    output = capsys.readouterr().out

    assert code == 1
    assert (
        "latest_recommendations.html keyword review row 1 should not show metric token 执行后本 SKU 单 2"
        in output
    )


def test_validate_latest_recommendations_uses_exact_keyword_review_title_context(capsys) -> None:
    older_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0ANCHOROLD",
        "product_name": "Anchor Older Product",
        "search_term_or_target": "stationery boxen for bags",
        "review_window": "7天后复盘",
        "judgement": "数据不足",
        "days_since_execution": 52,
        "review_data_source": "execution_anchored_daily",
        "post_7d_start": "2026-05-15",
        "post_7d_end": "2026-05-21",
        "post_7d_acos": "N/A",
    }
    early_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0ANCHOREARLY",
        "product_name": "Anchor Early Product",
        "search_term_or_target": "stationery box",
        "review_window": "3d_check",
        "judgement": "样本不足",
        "days_since_execution": 4,
        "review_data_source": "execution_anchored_daily",
        "post_7d_start": "2026-07-02",
        "post_7d_end": "2026-07-08",
        "post_7d_acos": "N/A",
    }
    data = {
        "marketplace_results": [
            {
                "marketplace": "UK",
                "report_view_snapshot": {
                    "keyword_action_effect_review_rows": [older_row, early_row],
                    "action_effect_review_rows": [],
                    "product_final_decision_rows": [],
                    "tomorrow_review_rows": [],
                    "today_task_queue_rows": [],
                    "html_search_term_processing_queue_rows": [],
                },
            }
        ]
    }
    text = (
        '2026-06-08 今日广告动作 id="today-ad-actions-all" 待确认 0 '
        "市场调查 提交今日数据 系统结论 融合诊断 "
        '<div class="review-item"><strong>UK｜stationery boxen for bags</strong> '
        "B0ANCHOROLD 执行后7天 2026-05-15 至 2026-05-21 执行后 ACOS N/A</div>"
        '<div class="review-item"><strong>UK｜stationery box</strong> '
        "B0ANCHOREARLY 3 天窗口只做初步判断 "
        "执行后3天 2026-07-02 至 2026-07-04 执行后3天 ACOS N/A</div>"
        '<div class="review-item"><strong>UK｜other term</strong></div>'
    )

    code = showcase_validate._validate_latest_recommendations_snapshot_rows(text, data)

    assert code == 0


def test_validate_safe_run_blocks_summary_missing_review_14d_promoted_token(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "kw-summary-14d-1",
            "normalized_action": "bid_down",
            "action_scope": "keyword",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0SUMMARY14D",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "summary 14d keyword",
            "executed_at": "2026-06-01",
            "report_date": "2026-06-01",
            "review_window": "7天后复盘",
            "judgement": "待人工复查",
            "days_since_execution": 7,
            "review_outcome": "not_ready",
            **_effective_review_metrics(),
            "current_14d_promoted_ad_orders": 3,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
        include_summary_snapshot_tokens=False,
    )
    (safe_dir / "summary.html").write_text(
        "2026-06-08 三分钟摘要 今日开工结论 今天先做 昨天动作今天要盯 "
        "产品页成功 0/3 卖家精灵自己 ASIN 今日 0/3，缓存 0/3 卖家精灵趋势 0/3 卖家精灵竞品发现 0/3 卖家精灵竞品池 今日 0/3，7天缓存 0/3 竞品 ASIN 反查 今日 0/3，缓存 0/3 Amazon 搜索页辅助验证 0/3 达到放量准入 0/3 弱势止损证据 0/3 证据不足 3/3 市场调查平均完整度 0.0/100 强证据 / 可用证据 0/3 / 0/3 "
        "summary 14d keyword 待人工复查 继续观察，等待足够样本。"
        "本 SKU 单 2 ACOS 12% 目标 ACOS 20% TACOS 8% 总单 4 库存 18",
        encoding="utf-8",
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "summary.html missing review watch row 1 metric token 14天本 SKU 单 3" in output


def test_validate_safe_run_blocks_summary_missing_review_14d_cost_token(capsys, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "kw-summary-14d-cost-1",
            "normalized_action": "bid_down",
            "action_scope": "keyword",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0SUMMARY14COST",
            "product_name": "Demo desk lamp",
            "search_term_or_target": "summary 14d cost keyword",
            "executed_at": "2026-06-01",
            "report_date": "2026-06-01",
            "review_window": "7天后复盘",
            "judgement": "待人工复查",
            "days_since_execution": 7,
            "review_outcome": "not_ready",
            **_effective_review_metrics(),
            "current_14d_promoted_ad_orders": 3,
            "current_14d_acos": 0.09,
            "current_14d_tacos": 0.07,
            "current_14d_total_orders": 5,
            "current_14d_available_stock": 18,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        keyword_review_rows=[review_row],
        include_summary_snapshot_tokens=False,
    )
    (safe_dir / "summary.html").write_text(
        "2026-06-08 三分钟摘要 今日开工结论 今天先做 昨天动作今天要盯 "
        "产品页成功 0/3 卖家精灵自己 ASIN 今日 0/3，缓存 0/3 卖家精灵趋势 0/3 卖家精灵竞品发现 0/3 卖家精灵竞品池 今日 0/3，7天缓存 0/3 竞品 ASIN 反查 今日 0/3，缓存 0/3 Amazon 搜索页辅助验证 0/3 达到放量准入 0/3 弱势止损证据 0/3 证据不足 3/3 市场调查平均完整度 0.0/100 强证据 / 可用证据 0/3 / 0/3 "
        "summary 14d cost keyword 待人工复查 继续观察，等待足够样本。"
        "本 SKU 单 2 ACOS 12% 目标 ACOS 20% TACOS 8% 总单 4 库存 18 "
        "14天本 SKU 单 3 14天 ACOS 9%",
        encoding="utf-8",
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "summary.html missing review watch row 1 metric token 14天 TACOS 7%" in output


def test_validate_safe_run_blocks_summary_missing_replenishment_identity(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    inventory_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0SUMMARYINV",
        "product_name": "Summary stockout board",
        "stock_risk_level": "OUT_OF_STOCK",
        "stock_status_label": "断货",
        "available_stock": 0,
        "days_of_cover": 0,
        "recommended_reorder_qty": 120,
        "replenishment_advice": "先核对在途库存，今天处理补货。",
    }
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        inventory_replenishment_rows=[inventory_row],
        include_summary_snapshot_tokens=False,
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "summary.html missing replenishment row 1 product name Summary stockout board" in output


def test_validate_safe_run_blocks_effective_review_without_days_since(capsys, tmp_path) -> None:
    from src.autoopt_feedback import ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "product_name": "Demo desk lamp",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        action_review_rows=[review_row],
    )

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "effective outcome missing days_since_execution" in output


def test_validate_safe_run_blocks_duplicate_marketplace_results(capsys, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
    )
    analysis_path = safe_dir / "latest_analysis.json"
    analysis = json.loads(analysis_path.read_text(encoding="utf-8"))
    analysis["marketplace_results"][2]["marketplace"] = "UK"
    analysis_path.write_text(json.dumps(analysis, ensure_ascii=False), encoding="utf-8")

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "marketplace_results must contain exactly one UK, one US, and one DE; got ['UK', 'US', 'UK']" in output


def test_validate_safe_run_blocks_stale_self_optimization_action_review_rows(capsys, tmp_path) -> None:
    from src.autoopt_feedback import ACTION_REVIEW_REQUIRED_FIELDS

    safe_dir = tmp_path / "safe"
    review_row = {field: "" for field in ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "action_id": "current-action-review",
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "product_name": "Demo desk lamp",
        }
    )
    _write_minimal_safe_run(
        safe_dir,
        '2026-06-08 今天广告动作 id="today-ad-actions-all" 待确认 0 市场调查 提交今日数据 系统结论 融合诊断',
        action_review_rows=[review_row],
        fill_blank_review_action_ids=False,
    )
    self_opt_path = safe_dir / "self_optimization_log_20260608.json"
    stale_payload = json.loads(self_opt_path.read_text(encoding="utf-8"))
    stale_payload["action_review_rows"] = [
        {
            "action_id": "stale-action-review",
            "marketplace": "UK",
            "sku": "SKU-OLD",
            "asin": "B0OLD12345",
        }
    ]
    self_opt_path.write_text(json.dumps(stale_payload, ensure_ascii=False), encoding="utf-8")

    code = showcase_validate.validate_safe_run_outputs(safe_dir)
    output = capsys.readouterr().out

    assert code == 1
    assert "self_optimization_log_20260608.json action_review_rows mismatch vs autoopt log" in output


def test_workspace_fingerprint_tracks_untracked_file_content(monkeypatch, tmp_path) -> None:
    monkeypatch.setattr(showcase_validate, "ROOT", tmp_path)
    monkeypatch.setattr(showcase_validate, "git_text", lambda args: "HEAD123\n" if args == ["rev-parse", "HEAD"] else "")
    monkeypatch.setattr(showcase_validate, "git_lines", lambda args: ["new.txt"] if args == ["ls-files", "--others", "--exclude-standard"] else [])
    monkeypatch.setattr(showcase_validate, "git_bytes", lambda args: b"")
    (tmp_path / "new.txt").write_text("before", encoding="utf-8")

    before = showcase_validate.workspace_fingerprint()["workspace_hash"]
    (tmp_path / "new.txt").write_text("after", encoding="utf-8")
    after = showcase_validate.workspace_fingerprint()["workspace_hash"]

    assert before != after


def test_write_validation_receipt_records_current_workspace(monkeypatch, tmp_path) -> None:
    safe_dir = tmp_path / "safe"
    safe_dir.mkdir()
    (safe_dir / "latest_analysis.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-08",
                "marketplace_results": [
                    {"marketplace": "DE"},
                    {"marketplace": "UK"},
                    {"marketplace": "US"},
                ],
            }
        ),
        encoding="utf-8",
    )
    receipt = tmp_path / "receipt.json"
    monkeypatch.setattr(showcase_validate, "VALIDATION_RECEIPT", receipt)
    monkeypatch.setattr(
        showcase_validate,
        "workspace_fingerprint",
        lambda: {
            "git_head": "HEAD123",
            "workspace_hash": "hash123",
            "changed_paths": ["main.py"],
            "status_porcelain": [" M main.py"],
        },
    )

    showcase_validate.write_validation_receipt(
        safe_dir,
        pytest_exit_code=0,
        safe_run_exit_code=0,
        output_validation_exit_code=0,
    )
    payload = json.loads(receipt.read_text(encoding="utf-8"))

    assert payload["result"] == "passed"
    assert payload["git_head"] == "HEAD123"
    assert payload["workspace_hash"] == "hash123"
    assert payload["report_date"] == "2026-06-08"
    assert payload["marketplaces"] == ["DE", "UK", "US"]
    assert payload["safe_run_dir"] == str(safe_dir)

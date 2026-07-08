from __future__ import annotations

import json
from datetime import date, timedelta
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pandas as pd
import pytest


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "data" / "output"
LATEST_ANALYSIS_PATH = OUTPUT_DIR / "latest_analysis.json"


def latest_report_xlsx_path() -> Path:
    reports = sorted(OUTPUT_DIR.glob("amazon_ops_report_*.xlsx"))
    if not reports:
        pytest.skip("official data/output report is not present in this clean demo workspace")
    return reports[-1]


def load_latest_analysis() -> dict:
    if not LATEST_ANALYSIS_PATH.exists():
        pytest.skip("official data/output latest_analysis.json is not present in this clean demo workspace")
    return json.loads(LATEST_ANALYSIS_PATH.read_text(encoding="utf-8"))


def marketplace_result(payload: dict, marketplace: str) -> dict:
    for result in payload["marketplace_results"]:
        if result["marketplace"] == marketplace:
            return result
    raise AssertionError(f"missing marketplace {marketplace}")


def number(value: object) -> float:
    if value in (None, ""):
        return 0.0
    return float(str(value).replace("£", "").replace("$", "").replace("€", "").replace(",", ""))


def anchored_keyword_review_payload(
    terms: list[str],
    confirmed_at: str,
    *,
    marketplace: str = "UK",
    sku: str = "SKU-1",
    asin: str = "B0TEST1234",
    clicks: float = 4,
    spend: float = 4,
    promoted_sales: float = 50,
    total_orders: float = 3,
    total_sales: float = 50,
    stock: float = 24,
    target_acos: object = "10%",
) -> dict[str, list[dict[str, object]]]:
    start_day = date.fromisoformat(confirmed_at)
    keyword_rows: list[dict[str, object]] = []
    product_rows: list[dict[str, object]] = []
    for offset in range(7):
        row_date = (start_day + timedelta(days=offset)).isoformat()
        for term in terms:
            active_day = offset == 0
            keyword_rows.append(
                {
                    "date": row_date,
                    "marketplace": marketplace,
                    "sku": sku,
                    "asin": asin,
                    "search_term": term,
                    "clicks": clicks if active_day else 0,
                    "spend": spend if active_day else 0,
                    "ad_orders": 1 if active_day else 0,
                    "ad_sales": promoted_sales if active_day else 0,
                    "promoted_ad_orders": 1 if active_day else 0,
                    "promoted_ad_sales": promoted_sales if active_day else 0,
                    "halo_ad_orders": 0,
                    "halo_ad_sales": 0,
                    "target_acos": target_acos,
                }
            )
        product_rows.append(
            {
                "date": row_date,
                "marketplace": marketplace,
                "sku": sku,
                "asin": asin,
                "clicks": clicks if offset == 0 else 0,
                "spend": spend if offset == 0 else 0,
                "ad_orders": 1 if offset == 0 else 0,
                "ad_sales": promoted_sales if offset == 0 else 0,
                "promoted_ad_orders": 1 if offset == 0 else 0,
                "promoted_ad_sales": promoted_sales if offset == 0 else 0,
                "halo_ad_orders": 0,
                "halo_ad_sales": 0,
                "total_orders": total_orders if offset == 0 else 0,
                "total_sales": total_sales if offset == 0 else 0,
                "available_stock": stock,
            }
        )
    return {"review_search_term_daily": keyword_rows, "review_product_daily": product_rows}


def anchored_review_contract_fields(promoted_orders: object = 1) -> dict[str, object]:
    return {
        "review_data_source": "execution_anchored_daily",
        "pre_7d_start": "2026-06-01",
        "pre_7d_end": "2026-06-07",
        "post_3d_start": "2026-06-08",
        "post_3d_end": "2026-06-10",
        "post_7d_start": "2026-06-08",
        "post_7d_end": "2026-06-14",
        "pre_7d_promoted_ad_orders": 0,
        "pre_7d_total_orders": 1,
        "pre_7d_tacos": "5%",
        "post_3d_days": 3,
        "post_3d_promoted_ad_orders": promoted_orders,
        "post_3d_total_orders": 2,
        "post_3d_acos": "7%",
        "post_3d_tacos": "5%",
        "post_3d_available_stock": 18,
        "post_7d_days": 7,
        "post_7d_promoted_ad_orders": promoted_orders,
        "post_7d_total_orders": 3,
        "post_7d_acos": "8%",
        "post_7d_tacos": "6%",
        "post_7d_available_stock": 18,
    }


def test_scale_keyword_action_keeps_tiny_sample_out_of_bid_up_queue() -> None:
    from src.report_presentation import _scale_keyword_action

    action = _scale_keyword_action("B0DEMOSCAL", ad_orders=1, acos=0.011, target_acos=0.273, clicks=1)

    assert action == "小样本保留观察"


def test_scale_keyword_action_uses_probe_bid_up_for_four_click_one_order() -> None:
    from src.report_presentation import _scale_keyword_action

    action = _scale_keyword_action("wooden dimmer desk lamp with dimmer switch", ad_orders=1, acos=0.062, target_acos=0.198, clicks=4)

    assert action == "试探提高竞价 3%-5%"


def test_scale_keyword_action_caps_automatic_bid_up_at_five_to_ten_percent() -> None:
    from src.report_presentation import _scale_keyword_action

    action = _scale_keyword_action("metal desk lamp", ad_orders=5, acos=0.05, target_acos=0.30, clicks=20)

    assert action == "提高竞价 5%-10%"
    assert "10%-15%" not in action


def test_scale_keyword_html_conversion_preserves_action_identity() -> None:
    from src.html_pages.components_ad_workbench import _ad_completion_payload, _scale_keywords_as_ad_queue_rows

    class Shared:
        @staticmethod
        def load_feedback_input():
            return []

    rows = _scale_keywords_as_ad_queue_rows(
        Shared(),
        [
            {
                "marketplace": "UK",
                "sku": "SKU-SCALE-HTML",
                "asin": "B0SCALEHTML",
                "product_name": "Scale HTML product",
                "search_term_or_target": "serving board",
                "campaign_name": "Scale campaign",
                "ad_group_name": "Scale group",
                "match_type_or_targeting": "EXACT",
                "scale_action": "试探提高竞价 3%-5%",
                "clicks": "4",
                "ad_orders": "1",
                "ACOS": "4.0%",
                "target_acos": "10.0%",
            }
        ],
    )

    assert rows[0]["suggested_action"] == "加价3%-5%"
    assert rows[0]["normalized_action"] == "bid_up"
    assert rows[0]["action_scope"] == "search_term"
    assert rows[0]["action_id"] == "UK||SKU-SCALE-HTML||B0SCALEHTML||search_term||serving board||bid_up"
    assert rows[0]["campaign_name"] == "Scale campaign"
    assert rows[0]["ad_group_name"] == "Scale group"
    assert rows[0]["match_type_or_targeting"] == "EXACT"
    payload = _ad_completion_payload(rows[0], "加价 3%-5%")
    assert payload["action_id"] == rows[0]["action_id"]
    assert payload["normalized_action"] == "bid_up"


def test_scale_keyword_html_conversion_keeps_blocked_action_as_audit_only() -> None:
    from src.html_pages.components_ad_workbench import _scale_keywords_as_ad_queue_rows

    class Shared:
        @staticmethod
        def load_feedback_input():
            return []

    blocked_id = "UK||SKU-SCALE-BLOCK||B0SCALEBLOCK||search_term||blocked scale term||bid_up"
    rows = _scale_keywords_as_ad_queue_rows(
        Shared(),
        [
            {
                "marketplace": "UK",
                "sku": "SKU-SCALE-BLOCK",
                "asin": "B0SCALEBLOCK",
                "product_name": "Scale blocked product",
                "search_term_or_target": "blocked scale term",
                "scale_action": "观察",
                "suggested_action": "观察",
                "copy_action_line": "建议观察",
                "ad_memory_blocked": True,
                "blocked_action_id": blocked_id,
                "blocked_original_action": "试探提高竞价 3%-5%",
                "keyword_memory_summary": "只有光环成交，需人工复查",
                "normalized_action": "bid_up",
                "action_scope": "search_term",
                "action_id": blocked_id,
                "confirmed_status": "仅背景参考",
            }
        ],
    )

    assert rows[0]["suggested_action"] == "观察"
    assert rows[0]["normalized_action"] == "observe"
    assert rows[0]["action_id"] != blocked_id
    assert rows[0]["blocked_action_id"] == blocked_id
    assert rows[0]["ad_memory_blocked"] is True
    assert rows[0]["confirmed_status"] == "仅背景参考"


def test_write_autoopt_outputs_includes_report_date_in_summary_sheet(tmp_path) -> None:
    from src.autoopt_feedback import write_autoopt_outputs

    payload = {
        "report_date": "2026-06-24",
        "rows": [],
        "summary": {"total_rows": 0},
        "action_review_rows": [{"marketplace": "UK", "sku": "SKU-ACTION", "action_id": "a1"}],
        "keyword_action_review_rows": [],
        "product_strategy_profiles": [{"marketplace": "UK", "sku": "SKU-PROFILE"}],
        "keyword_strategy_memory": [{"marketplace": "UK", "search_term_or_target": "dimmer desk lamp"}],
    }

    _, xlsx_path = write_autoopt_outputs(tmp_path, "2026-06-24", payload)
    self_optimization = json.loads((tmp_path / "self_optimization_log_20260624.json").read_text(encoding="utf-8"))

    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet = workbook["summary"]
        headers = [str(value or "").strip() for value in next(sheet.iter_rows(values_only=True))]
        values = [str(value or "").strip() for value in next(sheet.iter_rows(min_row=2, values_only=True))]
    finally:
        workbook.close()
    assert "report_date" in headers
    assert values[headers.index("report_date")] == "2026-06-24"
    assert self_optimization["action_review_rows"] == payload["action_review_rows"]
    assert self_optimization["product_strategy_profiles"] == payload["product_strategy_profiles"]
    assert self_optimization["keyword_strategy_memory"] == payload["keyword_strategy_memory"]


def test_rule_adjustments_do_not_treat_execution_rate_as_effectiveness() -> None:
    from src.autoopt_feedback import derive_rule_adjustments

    rows = [
        {
            "diagnosis_type": "广告动作",
            "today_action": "提高竞价 5%-10%",
            "confirmed_status": "已执行",
        },
        {
            "diagnosis_type": "广告动作",
            "today_action": "提高竞价 5%-10%",
            "confirmed_status": "已执行",
        },
        {
            "diagnosis_type": "广告动作",
            "today_action": "降低竞价 10%",
            "confirmed_status": "已执行",
        },
    ]

    payload = derive_rule_adjustments(rows)
    suggestions = [
        str(item.get("suggested_adjustment") or "")
        for item in [*payload["adjustments"], *payload["action_adjustments"]]
    ]

    assert suggestions
    assert all("promoted SKU" in suggestion for suggestion in suggestions)
    assert all("前置" not in suggestion for suggestion in suggestions)
    assert all("可继续保留为默认推荐" not in suggestion for suggestion in suggestions)


def validation_row(marketplace: str, asin: str) -> dict:
    workbook = openpyxl.load_workbook(latest_report_xlsx_path(), read_only=True, data_only=True)
    sheet = workbook["Metrics_Validation"]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    for row in rows:
        record = dict(zip(headers, row))
        if record.get("marketplace") == marketplace and record.get("asin") == asin:
            return record
    raise AssertionError(f"missing Metrics_Validation row for {marketplace} {asin}")


def validation_row_matching(marketplace: str, predicate) -> dict:
    workbook = openpyxl.load_workbook(latest_report_xlsx_path(), read_only=True, data_only=True)
    sheet = workbook["Metrics_Validation"]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    for row in rows:
        record = dict(zip(headers, row))
        if record.get("marketplace") == marketplace and predicate(record):
            return record
    raise AssertionError(f"missing Metrics_Validation row for marketplace {marketplace}")


def test_all_marketplace_excel_includes_global_action_review_sheet(tmp_path) -> None:
    from src.autoopt_feedback import ACTION_REVIEW_REQUIRED_FIELDS, KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS
    from src.generate_excel_report import generate_all_marketplace_excel_report
    from src.product_decision_layer import PRODUCT_FINAL_DECISION_REQUIRED_FIELDS

    review_row = {field: "" for field in ACTION_REVIEW_REQUIRED_FIELDS}
    review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B000TEST01",
            "manual_action_taken": "降竞价 10%",
            "review_window": "3天后复盘",
            "current_7d_promoted_ad_orders": 2,
            "current_7d_promoted_ad_sales": 30.0,
            "current_7d_tacos": 0.08,
            "current_7d_available_stock": 18,
            "pre_7d_promoted_ad_orders": 1,
            "pre_7d_tacos": 0.10,
            "post_3d_days": 3,
            "post_3d_promoted_ad_orders": 1,
            "post_3d_acos": 0.09,
            "post_3d_tacos": 0.07,
            "post_3d_available_stock": 18,
            "post_7d_days": 7,
            "post_7d_promoted_ad_orders": 2,
            "post_7d_acos": 0.10,
            "post_7d_tacos": 0.08,
            "post_7d_available_stock": 18,
        }
    )
    keyword_review_row = {field: "" for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS}
    keyword_review_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B000TEST01",
            "search_term_or_target": "dimmer desk lamp",
            "manual_action_taken": "降竞价 10%",
            "review_window": "3天后复盘",
            "current_7d_promoted_ad_orders": 1,
            "current_7d_promoted_ad_sales": 15.0,
            "current_7d_tacos": 0.06,
            "current_7d_available_stock": 18,
            "pre_7d_promoted_ad_orders": 0,
            "pre_7d_tacos": 0.12,
            "post_3d_days": 3,
            "post_3d_promoted_ad_orders": 1,
            "post_3d_acos": 0.08,
            "post_3d_tacos": 0.06,
            "post_3d_available_stock": 18,
            "post_7d_days": 7,
            "post_7d_promoted_ad_orders": 1,
            "post_7d_acos": 0.08,
            "post_7d_tacos": 0.06,
            "post_7d_available_stock": 18,
        }
    )
    decision_row = {field: "" for field in PRODUCT_FINAL_DECISION_REQUIRED_FIELDS}
    decision_row.update(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B000TEST01",
            "product_name": "Test product",
            "final_decision": "DO_NOT_TOUCH",
            "today_allowed_actions": ["observe"],
            "today_blocked_actions": [],
        }
    )
    task_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B000TEST01",
        "priority": "P0",
        "issue_type": "广告消耗无转化",
        "action_group": "广告动作",
        "today_action": "降竞价10%-20%",
        "search_term_or_target": "dimmer desk lamp",
        "suggested_action": "降竞价10%-20%",
        "normalized_action": "bid_down",
        "action_id": "UK||SKU-1||B000TEST01||dimmer desk lamp||bid_down",
    }
    tomorrow_review_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B000TEST01",
        "product_name": "Test product",
        "review_reason": "3-5点击无单，未达到强动作阈值",
        "current_evidence": "点击 4；花费 £0.88；订单 0",
        "tomorrow_check": "复查是否新增点击、花费或订单；未改善再降竞价",
        "trigger_action": "达到阈值后升级今日动作",
    }
    listing_diagnosis_row = {
        "marketplace": "UK",
        "产品": "Test product",
        "SKU": "SKU-1",
        "ASIN": "B000TEST01",
        "诊断类型": "Listing 待人工确认",
        "主因": "加购后不购买",
        "关键证据": "近14天广告点击 12；广告订单 0；总单 0",
        "建议动作": "暂时不加广告预算；人工检查价格、评分和主图",
    }
    cost_diagnosis_row = {
        "marketplace": "UK",
        "产品": "Test product",
        "SKU": "SKU-1",
        "ASIN": "B000TEST01",
        "诊断类型": "成本 / 利润压力诊断",
        "主因": "利润不允许加广告",
        "关键证据": "毛利不足；广告 ACOS 高于目标；库存 18",
        "建议动作": "不加预算；先复核成本、售价和优惠券",
    }
    result = {
        "marketplace": "UK",
        "has_data": False,
        "summary": {
            "marketplace": "UK",
            "ads_row_count": 0,
            "erp_row_count": 0,
            "sku_count": 0,
            "asin_count": 0,
            "ads_date_range": "无",
            "erp_date_range": "无",
        },
        "report_view": {
            "today_task_queue_rows": [task_row],
            "tomorrow_review_rows": [tomorrow_review_row],
            "product_final_decision_rows": [decision_row],
            "listing_price_diagnosis_rows": [listing_diagnosis_row],
            "cost_profit_diagnosis_rows": [cost_diagnosis_row],
            "action_effect_review_rows": [review_row],
            "keyword_action_effect_review_rows": [keyword_review_row],
            "frontend_coverage_summary": {
                "frontend_queue_total": 2,
                "frontend_decision_ready_count": 1,
                "frontend_reference_evidence_count": 1,
            },
        },
        "analysis_payload": {},
    }

    output_path = generate_all_marketplace_excel_report(
        tmp_path / "all.xlsx",
        [result],
        "2026-06-24",
        source_files={},
        import_summary={},
    )

    workbook = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    assert "今日动作清单" in workbook.sheetnames
    task_sheet = workbook["今日动作清单"]
    task_rows = task_sheet.iter_rows(values_only=True)
    task_headers = set(next(task_rows))
    assert {"marketplace", "sku", "asin", "today_action", "action_id"}.issubset(task_headers)
    assert "明日复查清单" in workbook.sheetnames
    tomorrow_sheet = workbook["明日复查清单"]
    tomorrow_rows = tomorrow_sheet.iter_rows(values_only=True)
    tomorrow_headers = set(next(tomorrow_rows))
    assert {"marketplace", "sku", "asin", "tomorrow_check", "trigger_action"}.issubset(tomorrow_headers)
    assert "Listing待确认" in workbook.sheetnames
    listing_sheet = workbook["Listing待确认"]
    listing_rows = listing_sheet.iter_rows(values_only=True)
    listing_headers = set(next(listing_rows))
    assert {"marketplace", "SKU", "ASIN", "诊断类型", "主因", "关键证据", "建议动作"}.issubset(listing_headers)
    assert "成本利润诊断" in workbook.sheetnames
    cost_sheet = workbook["成本利润诊断"]
    cost_rows = cost_sheet.iter_rows(values_only=True)
    cost_headers = set(next(cost_rows))
    assert {"marketplace", "SKU", "ASIN", "诊断类型", "主因", "关键证据", "建议动作"}.issubset(cost_headers)
    assert "产品最终决策" in workbook.sheetnames
    decision_sheet = workbook["产品最终决策"]
    decision_rows = decision_sheet.iter_rows(values_only=True)
    decision_headers = set(next(decision_rows))
    assert set(PRODUCT_FINAL_DECISION_REQUIRED_FIELDS).issubset(decision_headers)
    assert "执行后效果复盘" in workbook.sheetnames
    sheet = workbook["执行后效果复盘"]
    rows = sheet.iter_rows(values_only=True)
    headers = set(next(rows))
    assert set(ACTION_REVIEW_REQUIRED_FIELDS).issubset(headers)
    assert "词级执行复盘" in workbook.sheetnames
    keyword_sheet = workbook["词级执行复盘"]
    keyword_rows = keyword_sheet.iter_rows(values_only=True)
    keyword_headers = set(next(keyword_rows))
    assert set(KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS).issubset(keyword_headers)
    formatted_workbook = openpyxl.load_workbook(output_path, data_only=True)
    formatted_review_sheet = formatted_workbook["执行后效果复盘"]
    formatted_review_headers = {
        str(cell.value or ""): index
        for index, cell in enumerate(formatted_review_sheet[1], start=1)
    }
    assert formatted_review_sheet.cell(
        row=2,
        column=formatted_review_headers["current_7d_tacos"],
    ).number_format == "0.0%"
    assert formatted_review_sheet.cell(
        row=2,
        column=formatted_review_headers["current_7d_available_stock"],
    ).number_format == "0"
    assert formatted_review_sheet.cell(
        row=2,
        column=formatted_review_headers["current_7d_promoted_ad_orders"],
    ).number_format == "0"
    assert "#,##0.00" in formatted_review_sheet.cell(
        row=2,
        column=formatted_review_headers["current_7d_promoted_ad_sales"],
    ).number_format
    assert formatted_review_sheet.cell(
        row=2,
        column=formatted_review_headers["pre_7d_tacos"],
    ).number_format == "0.0%"
    assert formatted_review_sheet.cell(
        row=2,
        column=formatted_review_headers["post_7d_tacos"],
    ).number_format == "0.0%"
    assert formatted_review_sheet.cell(
        row=2,
        column=formatted_review_headers["post_3d_acos"],
    ).number_format == "0.0%"
    assert formatted_review_sheet.cell(
        row=2,
        column=formatted_review_headers["post_7d_acos"],
    ).number_format == "0.0%"
    assert formatted_review_sheet.cell(
        row=2,
        column=formatted_review_headers["post_7d_days"],
    ).number_format == "0"
    assert formatted_review_sheet.cell(
        row=2,
        column=formatted_review_headers["post_3d_available_stock"],
    ).number_format == "0"
    assert formatted_review_sheet.cell(
        row=2,
        column=formatted_review_headers["post_7d_promoted_ad_orders"],
    ).number_format == "0"
    formatted_keyword_sheet = formatted_workbook["词级执行复盘"]
    formatted_keyword_headers = {
        str(cell.value or ""): index
        for index, cell in enumerate(formatted_keyword_sheet[1], start=1)
    }
    assert formatted_keyword_sheet.cell(
        row=2,
        column=formatted_keyword_headers["current_7d_tacos"],
    ).number_format == "0.0%"
    assert formatted_keyword_sheet.cell(
        row=2,
        column=formatted_keyword_headers["current_7d_promoted_ad_orders"],
    ).number_format == "0"
    assert formatted_keyword_sheet.cell(
        row=2,
        column=formatted_keyword_headers["pre_7d_tacos"],
    ).number_format == "0.0%"
    assert formatted_keyword_sheet.cell(
        row=2,
        column=formatted_keyword_headers["post_7d_promoted_ad_orders"],
    ).number_format == "0"
    assert formatted_keyword_sheet.cell(
        row=2,
        column=formatted_keyword_headers["post_3d_tacos"],
    ).number_format == "0.0%"
    assert formatted_keyword_sheet.cell(
        row=2,
        column=formatted_keyword_headers["post_7d_available_stock"],
    ).number_format == "0"
    overview_sheet = workbook["总览"]
    overview_rows = list(overview_sheet.iter_rows(values_only=True))
    overview_records = {
        str(row[1] or ""): str(row[2] or "")
        for row in overview_rows
        if row and len(row) >= 3 and str(row[0] or "") == "前台证据覆盖"
    }
    assert overview_records["ALL 产品页成功"] == "0/2"
    assert overview_records["ALL 卖家精灵自己 ASIN"] == "今日 0/2，缓存 0/2"
    assert overview_records["ALL 卖家精灵竞品池"] == "今日 0/2，7天缓存 0/2"
    assert overview_records["ALL 竞品 ASIN 反查"] == "今日 0/2，缓存 0/2"
    assert overview_records["ALL Amazon 搜索页辅助验证"] == "0/2"


def test_keyword_review_contract_downgrades_effective_without_support_metrics() -> None:
    from src.autoopt_feedback import ensure_keyword_action_review_contract

    row = ensure_keyword_action_review_contract(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "legacy term",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            "current_7d_promoted_ad_orders": 1,
            "current_7d_acos": 0.08,
            "current_7d_tacos": "",
            "current_7d_total_orders": "",
            "current_7d_available_stock": "",
            **anchored_review_contract_fields(),
            "block_reason": "历史判断有效，继续提高优先级",
        }
    )

    assert row["review_outcome"] == "needs_manual_review"
    assert row["effectiveness_score"] == -1
    assert "缺少有效复盘指标" in row["block_reason"]
    assert "继续提高优先级" not in row["block_reason"]


def test_review_contract_treats_nan_strings_as_missing_support_metrics() -> None:
    from src.autoopt_feedback import ensure_action_review_contract, ensure_keyword_action_review_contract

    base_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0TEST1234",
        "review_outcome": "effective",
        "effectiveness_score": 2,
        "days_since_execution": 8,
        "promoted_conversion_improved": True,
        "halo_only_conversion": False,
        "target_sku_not_converted": False,
        "current_7d_promoted_ad_orders": 1,
        "current_7d_acos": "nan",
        "current_7d_tacos": "NaN",
        "current_7d_total_orders": "nan",
        "current_7d_available_stock": "NaN",
        "current_14d_promoted_ad_orders": "nan",
        "current_14d_acos": "NaN",
        "current_14d_total_orders": "nan",
        "current_14d_tacos": "NaN",
        "current_14d_available_stock": "nan",
        **anchored_review_contract_fields(),
        "block_reason": "历史判断有效，继续提高优先级",
    }

    product_row = ensure_action_review_contract(base_row)
    keyword_row = ensure_keyword_action_review_contract(
        {
            **base_row,
            "search_term_or_target": "nan support term",
            "action_detail": "加价 5%-10%",
        }
    )

    for row in [product_row, keyword_row]:
        assert row["review_outcome"] == "needs_manual_review"
        assert row["effectiveness_score"] == -1
        assert "缺少有效复盘指标" in row["block_reason"]
        assert "current_7d_acos" in row["block_reason"]
        assert "current_14d_available_stock" in row["block_reason"]
        assert "继续提高优先级" not in row["block_reason"]


def test_review_contract_requires_current_7d_click_and_spend_for_effective() -> None:
    from src.autoopt_feedback import ensure_action_review_contract, ensure_keyword_action_review_contract

    base_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0TEST1234",
        "review_outcome": "effective",
        "effectiveness_score": 2,
        "days_since_execution": 8,
        "promoted_conversion_improved": True,
        "halo_only_conversion": False,
        "target_sku_not_converted": False,
        "current_7d_clicks": "",
        "current_7d_spend": "NaN",
        "current_7d_promoted_ad_orders": 1,
        "current_7d_acos": "8%",
        "current_7d_tacos": "6%",
        "current_7d_total_orders": 3,
        "current_7d_available_stock": 18,
        "current_14d_promoted_ad_orders": 2,
        "current_14d_acos": "9%",
        "current_14d_total_orders": 5,
        "current_14d_tacos": "7%",
        "current_14d_available_stock": 18,
        **anchored_review_contract_fields(),
        "block_reason": "历史判断有效，继续提高优先级",
    }

    product_row = ensure_action_review_contract(base_row)
    keyword_row = ensure_keyword_action_review_contract(
        {
            **base_row,
            "search_term_or_target": "missing traffic term",
            "action_detail": "加价 5%-10%",
        }
    )

    for row in [product_row, keyword_row]:
        assert row["review_outcome"] == "needs_manual_review"
        assert row["effectiveness_score"] == -1
        assert "current_7d_clicks" in row["block_reason"]
        assert "current_7d_spend" in row["block_reason"]
        assert "继续提高优先级" not in row["block_reason"]


def test_review_contract_downgrades_effective_when_acos_exceeds_target() -> None:
    from src.autoopt_feedback import ensure_action_review_contract, ensure_keyword_action_review_contract

    base_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0TEST1234",
        "review_outcome": "effective",
        "effectiveness_score": 2,
        "days_since_execution": 8,
        "promoted_conversion_improved": True,
        "halo_only_conversion": False,
        "target_sku_not_converted": False,
        "current_7d_clicks": 8,
        "current_7d_spend": 3.2,
        "current_7d_promoted_ad_orders": 1,
        "current_7d_acos": "25%",
        "target_acos": "10%",
        "current_7d_tacos": "8%",
        "current_7d_total_orders": 3,
        "current_7d_available_stock": 18,
        "current_14d_promoted_ad_orders": 2,
        "current_14d_acos": "12%",
        "current_14d_total_orders": 5,
        "current_14d_tacos": "9%",
        "current_14d_available_stock": 18,
        **anchored_review_contract_fields(),
        "block_reason": "历史判断有效，继续提高优先级",
    }

    product_row = ensure_action_review_contract(base_row)
    keyword_row = ensure_keyword_action_review_contract(
        {
            **base_row,
            "search_term_or_target": "dimmer desk lamp",
            "action_detail": "加价 5%-10%",
        }
    )

    for row in [product_row, keyword_row]:
        assert row["review_outcome"] == "needs_manual_review"
        assert row["effectiveness_score"] == -1
        assert row["block_reason"] == "本 SKU 有单但 ACOS 高于目标，需人工复查"


def test_review_contract_downgrades_effective_when_14d_promoted_orders_conflict_with_7d() -> None:
    from src.autoopt_feedback import ensure_keyword_action_review_contract

    row = ensure_keyword_action_review_contract(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "conflicting promoted orders",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            "current_7d_clicks": 8,
            "current_7d_spend": 3.2,
            "current_7d_promoted_ad_orders": 2,
            "current_7d_acos": "8%",
            "target_acos": "10%",
            "current_7d_tacos": "6%",
            "current_7d_total_orders": 4,
            "current_7d_available_stock": 18,
            "current_14d_promoted_ad_orders": 1,
            "current_14d_acos": "9%",
            "current_14d_total_orders": 7,
            "current_14d_tacos": "7%",
            "current_14d_available_stock": 18,
            **anchored_review_contract_fields(2),
        }
    )

    assert row["review_outcome"] == "needs_manual_review"
    assert row["effectiveness_score"] == -1
    assert row["block_reason"] == "本 SKU 14天订单数小于7天订单数，复盘口径需复查"


def test_review_contract_downgrades_effective_when_promoted_orders_exceed_totals() -> None:
    from src.autoopt_feedback import ensure_keyword_action_review_contract

    row = ensure_keyword_action_review_contract(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "impossible promoted orders",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            "current_7d_clicks": 8,
            "current_7d_spend": 3.2,
            "current_7d_ad_orders": 1,
            "current_7d_promoted_ad_orders": 2,
            "current_7d_acos": "8%",
            "target_acos": "10%",
            "current_7d_tacos": "6%",
            "current_7d_total_orders": 3,
            "current_7d_available_stock": 18,
            "current_14d_ad_orders": 3,
            "current_14d_promoted_ad_orders": 3,
            "current_14d_acos": "9%",
            "current_14d_total_orders": 5,
            "current_14d_tacos": "7%",
            "current_14d_available_stock": 18,
            **anchored_review_contract_fields(2),
        }
    )

    assert row["review_outcome"] == "needs_manual_review"
    assert row["effectiveness_score"] == -1
    assert row["block_reason"] == "本 SKU 7天订单数高于广告总订单数，复盘口径需复查"


def test_review_contract_downgrades_positive_display_halo_only_without_standard_outcome() -> None:
    from src.autoopt_feedback import ensure_action_review_contract

    row = ensure_action_review_contract(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "outcome": "初步有效",
            "days_since_execution": 8,
            "promoted_conversion_improved": False,
            "halo_only_conversion": True,
            "target_sku_not_converted": True,
            "current_7d_promoted_ad_orders": 0,
            "current_7d_acos": "8%",
            "current_7d_tacos": "6%",
            "current_7d_total_orders": 3,
            "current_7d_available_stock": 18,
        }
    )

    assert row["review_outcome"] == "needs_manual_review"
    assert row["effectiveness_score"] == -1
    assert row["block_reason"] == "只有光环成交，需人工复查"
    assert row["outcome"] == "待人工复查"
    assert row["judgement"] == "待人工复查"


def test_review_contract_recomputes_halo_only_from_current_fields() -> None:
    from src.autoopt_feedback import ensure_keyword_action_review_contract

    row = ensure_keyword_action_review_contract(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "halo term",
            "outcome": "有改善迹象",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            "current_7d_clicks": 8,
            "current_7d_spend": 3.2,
            "current_7d_ad_orders": 1,
            "current_7d_promoted_ad_orders": 0,
            "current_7d_halo_ad_orders": 1,
            "current_7d_acos": "8%",
            "current_7d_target_acos": "20%",
            "current_7d_tacos": "6%",
            "current_7d_total_orders": 3,
            "current_7d_available_stock": 18,
            "current_14d_promoted_ad_orders": 2,
            "current_14d_acos": "9%",
            "current_14d_total_orders": 5,
            "current_14d_tacos": "7%",
            "current_14d_available_stock": 18,
        }
    )

    assert row["promoted_conversion_improved"] is False
    assert row["halo_only_conversion"] is True
    assert row["target_sku_not_converted"] is True
    assert row["attribution_effect_status"] == "halo_only_conversion"
    assert row["review_outcome"] == "needs_manual_review"
    assert row["effectiveness_score"] == -1
    assert row["outcome"] == "待人工复查"


def test_keyword_review_contract_downgrades_positive_display_before_7_days() -> None:
    from src.autoopt_feedback import ensure_keyword_action_review_contract

    row = ensure_keyword_action_review_contract(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "dimmer desk lamp",
            "outcome": "有改善迹象",
            "days_since_execution": 4,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            "current_7d_promoted_ad_orders": 1,
            "current_7d_acos": "8%",
            "current_7d_tacos": "6%",
            "current_7d_total_orders": 3,
            "current_7d_available_stock": 18,
        }
    )

    assert row["review_outcome"] == "not_ready"
    assert row["effectiveness_score"] is None
    assert row["block_reason"] == "未满7天或缺执行日期，暂不判断有效"
    assert row["outcome"] == "待7天确认"
    assert row["judgement"] == "待7天确认"


def test_keyword_review_contract_treats_keep_current_as_positive_before_7_days() -> None:
    from src.autoopt_feedback import ensure_keyword_action_review_contract

    row = ensure_keyword_action_review_contract(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "dimmer desk lamp keep",
            "outcome": "可保留",
            "judgement": "可保留",
            "days_since_execution": 4,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            "current_7d_promoted_ad_orders": 1,
            "current_7d_acos": "8%",
            "current_7d_tacos": "6%",
            "current_7d_total_orders": 3,
            "current_7d_available_stock": 18,
        }
    )

    assert row["review_outcome"] == "not_ready"
    assert row["effectiveness_score"] is None
    assert row["block_reason"] == "未满7天或缺执行日期，暂不判断有效"
    assert row["outcome"] == "待7天确认"
    assert row["judgement"] == "待7天确认"


def test_keyword_review_contract_requires_14d_support_for_effective_claim() -> None:
    from src.autoopt_feedback import ensure_keyword_action_review_contract

    row = ensure_keyword_action_review_contract(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "dimmer desk lamp",
            "review_outcome": "effective",
            "effectiveness_score": 2,
            "days_since_execution": 8,
            "promoted_conversion_improved": True,
            "halo_only_conversion": False,
            "target_sku_not_converted": False,
            "current_7d_promoted_ad_orders": 1,
            "current_7d_acos": "8%",
            "current_7d_tacos": "6%",
            "current_7d_total_orders": 3,
            "current_7d_available_stock": 18,
            "current_14d_promoted_ad_orders": 2,
            "current_14d_acos": "9%",
            "current_14d_total_orders": "",
            "current_14d_tacos": "7%",
            "current_14d_available_stock": 18,
            **anchored_review_contract_fields(),
        }
    )

    assert row["review_outcome"] == "needs_manual_review"
    assert row["effectiveness_score"] == -1
    assert "current_14d_total_orders" in row["block_reason"]
    assert row["outcome"] == "待人工复查"
    assert row["judgement"] == "待人工复查"


def test_keyword_review_contract_downgrades_ineffective_before_7_days() -> None:
    from src.autoopt_feedback import ensure_keyword_action_review_contract

    row = ensure_keyword_action_review_contract(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "dimmer desk lamp",
            "outcome": "暂未改善",
            "review_outcome": "ineffective",
            "effectiveness_score": -2,
            "days_since_execution": 4,
            "current_7d_clicks": 9,
            "current_7d_spend": 6,
            "current_7d_promoted_ad_orders": 0,
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": True,
        }
    )

    assert row["review_outcome"] == "not_ready"
    assert row["effectiveness_score"] is None
    assert row["block_reason"] == "未满7天或缺执行日期，暂不判断无效"
    assert row["outcome"] == "待7天确认"
    assert row["judgement"] == "待7天确认"


def test_keyword_review_contract_downgrades_negative_display_before_3_days() -> None:
    from src.autoopt_feedback import ensure_keyword_action_review_contract

    row = ensure_keyword_action_review_contract(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "dimmer desk lamp",
            "outcome": "暂未改善",
            "review_outcome": "not_ready",
            "effectiveness_score": "",
            "days_since_execution": 2,
            "current_7d_clicks": 9,
            "current_7d_spend": 6,
            "current_7d_promoted_ad_orders": 0,
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": True,
        }
    )

    assert row["review_outcome"] == "not_ready"
    assert row["effectiveness_score"] is None
    assert row["block_reason"] == "未满3天或缺执行日期，暂不判断负面效果"
    assert row["outcome"] == "待7天确认"
    assert row["judgement"] == "待7天确认"


def test_keyword_review_contract_treats_stop_append_as_negative_before_3_days() -> None:
    from src.autoopt_feedback import ensure_keyword_action_review_contract

    row = ensure_keyword_action_review_contract(
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "stop append term",
            "outcome": "停止追加",
            "judgement": "停止追加",
            "review_outcome": "not_ready",
            "effectiveness_score": "",
            "days_since_execution": 2,
            "current_7d_clicks": 9,
            "current_7d_spend": 6,
            "current_7d_promoted_ad_orders": 0,
            "promoted_conversion_improved": False,
            "halo_only_conversion": False,
            "target_sku_not_converted": True,
        }
    )

    assert row["review_outcome"] == "not_ready"
    assert row["effectiveness_score"] is None
    assert row["block_reason"] == "未满3天或缺执行日期，暂不判断负面效果"
    assert row["outcome"] == "待7天确认"
    assert row["judgement"] == "待7天确认"


def test_keyword_strategy_memory_does_not_keep_high_acos_effective_legacy_row(tmp_path) -> None:
    from src.autoopt_feedback import build_keyword_strategy_memory

    rows = build_keyword_strategy_memory(
        output_dir=tmp_path,
        current_keyword_reviews=[
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "product_name": "Demo desk lamp",
                "search_term_or_target": "dimmer desk lamp",
                "action_detail": "加价 5%-10%",
                "review_outcome": "effective",
                "effectiveness_score": 2,
                "days_since_execution": 8,
                "promoted_conversion_improved": True,
                "halo_only_conversion": False,
                "target_sku_not_converted": False,
                "current_7d_promoted_ad_orders": 1,
                "current_7d_acos": "25%",
                "target_acos": "10%",
                "current_7d_tacos": "8%",
                "current_7d_total_orders": 3,
                "current_7d_available_stock": 18,
            }
        ],
    )

    assert len(rows) == 1
    assert rows[0]["review_outcome"] == "needs_manual_review"
    assert rows[0]["effectiveness_score"] == -1
    assert rows[0]["recommended_future_policy"] == "recheck_frontend_before_action"
    assert rows[0]["should_keep"] is False
    assert rows[0]["should_block_repeating"] is True
    assert rows[0]["should_recheck_frontend"] is True


def test_de_common_date_range_uses_erp_coverage_not_last_sales_date() -> None:
    payload = load_latest_analysis()
    report_date = payload["report_date"]
    summary = marketplace_result(payload, "DE")["summary"]

    assert summary["erp_report_coverage_date_range"].endswith(report_date)
    assert summary["common_date_range"].endswith(report_date)
    assert summary["history_days"] >= 30
    assert summary["zero_fill_applied"] is True


def test_us_product_level_14d_metrics_match_validation_sheet() -> None:
    row = validation_row_matching(
        "US",
        lambda record: float(record["raw_14d_ad_clicks"]) > 50
        and float(record["raw_14d_ad_spend"]) > 10
        and float(record["raw_14d_ad_orders"]) > 0,
    )

    assert row["raw_14d_ad_clicks"] == row["used_in_report_14d_clicks"]
    assert row["raw_14d_ad_spend"] == row["used_in_report_14d_spend"]
    assert row["raw_14d_ad_orders"] == row["used_in_report_14d_orders"]
    assert row["mismatch_flag"] is False
    assert float(row["raw_14d_ad_clicks"]) > 50
    assert float(row["raw_14d_ad_spend"]) > 10
    assert float(row["raw_14d_ad_orders"]) > 0


def test_low_click_zero_order_search_terms_are_hidden_from_html_queue() -> None:
    payload = load_latest_analysis()

    for result in payload["marketplace_results"]:
        rows = result["report_view_snapshot"].get("html_search_term_processing_queue_rows", [])
        for row in rows:
            clicks = number(row.get("clicks"))
            orders = number(row.get("orders"))
            assert not (clicks <= 2 and orders == 0), row


def test_executed_actions_within_cooldown_do_not_occupy_p0_or_p1_main_slots() -> None:
    payload = load_latest_analysis()
    for result in payload["marketplace_results"]:
        view = result["report_view_snapshot"]
        today_keys = {
            (row.get("sku"), row.get("asin"))
            for row in view.get("today_task_queue_rows", [])
            if row.get("priority") in {"P0", "P1"}
        }
        cooldown_keys = {
            (row.get("sku"), row.get("asin"))
            for row in view.get("tomorrow_review_rows", [])
            if row.get("review_reason") == "已执行动作冷却期"
        }
        assert not (today_keys & cooldown_keys), result.get("marketplace")


def test_report_view_snapshot_preserves_keyword_action_review_rows() -> None:
    from main import _report_view_snapshot

    row = {
        "marketplace": "UK",
        "search_term_or_target": "dimmer desk lamp",
        "judgement": "样本不足",
        "review_window": "未满3天",
    }
    snapshot = _report_view_snapshot({"keyword_action_effect_review_rows": [row]})

    assert snapshot["keyword_action_effect_review_rows"] == [row]


def test_report_view_snapshot_fills_review_effect_metrics_from_evidence() -> None:
    from main import _report_view_snapshot

    row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0SNAPSHOT",
        "effect_evidence": "7天：点击 3，订单 0，本 SKU 单 0",
    }
    snapshot = _report_view_snapshot({"action_effect_review_rows": [row]})

    assert snapshot["action_effect_review_rows"][0]["effect_metrics"] == row["effect_evidence"]
    assert "effect_metrics" not in row


def test_current_autoopt_reviews_replace_historical_report_view_rows_by_marketplace() -> None:
    from main import _inject_current_autoopt_reviews

    results = [
        {
            "marketplace": "UK",
            "report_view": {
                "action_effect_review_rows": [
                    {
                        "marketplace": "UK",
                        "product_name": "stale product",
                        "current_7d_tacos": "",
                        "current_7d_available_stock": "",
                    }
                ],
                "keyword_action_effect_review_rows": [
                    {
                        "marketplace": "UK",
                        "search_term_or_target": "stale term",
                        "current_7d_tacos": "",
                    }
                ],
            },
        },
        {"marketplace": "US", "report_view": {}},
    ]
    payload = {
        "action_review_rows": [
            {
                "marketplace": "UK",
                "product_name": "current product",
                "current_7d_tacos": 0.08,
                "current_7d_available_stock": 15,
            },
            {
                "marketplace": "US",
                "product_name": "us current product",
                "current_7d_tacos": 0.09,
                "current_7d_available_stock": 20,
            },
        ],
        "keyword_action_review_rows": [
            {
                "marketplace": "UK",
                "search_term_or_target": "current term",
                "current_7d_tacos": 0.07,
                "current_7d_available_stock": 15,
            }
        ],
    }

    _inject_current_autoopt_reviews(results, payload)

    uk_view = results[0]["report_view"]
    us_view = results[1]["report_view"]
    assert uk_view["action_effect_review_rows"] == [payload["action_review_rows"][0]]
    assert uk_view["keyword_action_effect_review_rows"] == [payload["keyword_action_review_rows"][0]]
    assert us_view["action_effect_review_rows"] == [payload["action_review_rows"][1]]
    assert us_view["keyword_action_effect_review_rows"] == []


def test_manual_learning_log_contains_heat_pad_4_root_cause() -> None:
    learning_logs = sorted(OUTPUT_DIR.glob("manual_learning_log_*.json"))
    if not learning_logs:
        pytest.skip("official manual learning log is not present in this clean demo workspace")
    assert learning_logs
    rows = json.loads(learning_logs[-1].read_text(encoding="utf-8"))
    matching = [
        row
        for row in rows
        if row.get("manual_root_cause") == "新品评分低；价格无优势"
        and row.get("manual_action_taken") == "价格调整到24.99"
    ]

    assert matching
    assert matching[0]["manual_root_cause"] == "新品评分低；价格无优势"
    assert matching[0]["manual_action_taken"] == "价格调整到24.99"


def test_unknown_or_stale_enhanced_data_does_not_override_fresh_files() -> None:
    payload = load_latest_analysis()

    for result in payload["marketplace_results"]:
        files = result["enhanced_data_status"].get("detected_files", [])
        by_type: dict[str, list[dict]] = {}
        for record in files:
            by_type.setdefault(record.get("data_type"), []).append(record)
        for records in by_type.values():
            has_fresh_used = any(
                record.get("used_in_diagnosis") == "是"
                and record.get("period_hint") != "unknown"
                and record.get("freshness") in {"fresh", "weekly_lag", "dated"}
                for record in records
            )
            if not has_fresh_used:
                continue
            for record in records:
                if record.get("period_hint") == "unknown" or record.get("freshness") == "stale":
                    assert record.get("used_in_diagnosis") != "是", record


def test_markdown_export_functions_have_single_effective_definition() -> None:
    import src.report_presentation as report_presentation

    source = Path(report_presentation.__file__).read_text(encoding="utf-8")
    assert source.count("def build_recommendations_markdown(") == 1
    assert source.count("def build_marketplace_summary_markdown(") == 1
    assert source.count("def build_all_marketplace_markdown(") == 1
    assert source.count("def build_all_enhanced_requests_markdown(") == 1

    analyze_rules_source = (ROOT / "src" / "analyze_rules.py").read_text(encoding="utf-8")
    markdown_reports_source = (ROOT / "src" / "markdown_reports.py").read_text(encoding="utf-8")
    assert "def build_report_view(" not in analyze_rules_source
    assert "def build_recommendations_markdown(" not in analyze_rules_source
    assert "def build_recommendations_markdown(" not in markdown_reports_source


def test_marketplace_summary_markdown_exposes_frontend_coverage() -> None:
    from src.report_presentation import build_marketplace_summary_markdown

    results = [
        {
            "marketplace": "UK",
            "has_data": True,
            "summary": {
                "marketplace": "UK",
                "ads_row_count": 10,
                "erp_row_count": 20,
                "sku_count": 3,
                "asin_count": 4,
                "common_date_range": "2026-06-20 至 2026-06-24",
            },
            "report_view": {
                "frontend_coverage_summary": {
                    "frontend_queue_total": 2,
                    "frontend_product_page_success_count": 2,
                        "frontend_competitor_search_success_count": 1,
                        "frontend_own_sellersprite_count": 1,
                        "frontend_competitor_discovery_count": 1,
                        "frontend_competitor_pool_count": 1,
                        "frontend_competitor_sellersprite_count": 1,
                        "frontend_competitor_sellersprite_asin_count": 3,
                        "frontend_amazon_search_validation_count": 1,
                        "frontend_scalable_strong_count": 0,
                    "frontend_weak_defensive_count": 1,
                    "frontend_insufficient_count": 1,
                }
            },
        },
        {
            "marketplace": "US",
            "has_data": True,
            "summary": {
                "marketplace": "US",
                "ads_row_count": 5,
                "erp_row_count": 6,
                "sku_count": 2,
                "asin_count": 2,
                "common_date_range": "2026-06-20 至 2026-06-24",
            },
            "report_view": {
                "frontend_coverage_summary": {
                    "frontend_queue_total": 1,
                    "frontend_product_page_success_count": 1,
                        "frontend_competitor_search_success_count": 0,
                        "frontend_own_sellersprite_count": 1,
                        "frontend_competitor_discovery_count": 0,
                        "frontend_competitor_pool_count": 0,
                        "frontend_competitor_sellersprite_count": 0,
                        "frontend_competitor_sellersprite_asin_count": 0,
                        "frontend_amazon_search_validation_count": 0,
                        "frontend_scalable_strong_count": 0,
                    "frontend_weak_defensive_count": 0,
                    "frontend_insufficient_count": 1,
                }
            },
        },
    ]

    markdown = build_marketplace_summary_markdown(results, "2026-06-24")

    assert "## 前台证据覆盖" in markdown
    assert "产品页成功" in markdown
    assert "竞品 ASIN 反查" in markdown
    assert "| UK | 2 | 2/2 | 今日 0/2，缓存 0/2 | 0/2 | 1/2 | 今日 0/2，7天缓存 0/2 | 今日 0/2，缓存 0/2 | 1/2 | 0/2 | 1/2 | 1/2 |" in markdown
    assert "| ALL | 3 | 3/3 | 今日 0/3，缓存 0/3 | 0/3 | 1/3 | 今日 0/3，7天缓存 0/3 | 今日 0/3，缓存 0/3 | 1/3 | 0/3 | 1/3 | 2/3 |" in markdown


def test_today_task_queue_contains_review_status_fields() -> None:
    payload = load_latest_analysis()
    rows = [
        row
        for result in payload["marketplace_results"]
        for row in result["report_view_snapshot"].get("today_task_queue_rows", [])
    ]
    assert rows
    for row in rows:
        assert row.get("review_status"), row
        assert row.get("consecutive_trigger_days"), row
        assert row.get("why_still_active"), row
        assert row.get("downgrade_condition"), row


def test_frontend_check_queue_excludes_pure_cost_and_does_not_fabricate_findings() -> None:
    from src.report_presentation import _build_frontend_check_queue

    rows = _build_frontend_check_queue(
        [
            {
                "marketplace": "UK",
                "product_name": "演示笔记本3片",
                "sku": "SKU-DEMO-TRAY-01",
                "asin": "B0DEMOFR01",
                "priority": "P1",
                "issue_type": "广告归因弱诊断",
                "primary_reason": "近14天广告有点击但广告订单弱",
                "key_evidence": "近14天点击 31；广告订单 0；总单 2",
                "confirmed_status": "待确认",
                "search_term_or_target": "serving board",
                "suggested_action": "降竞价10%-20%",
                "confirmed_note": "用户反馈：已执行 serving board 降竞价。",
            },
            {
                "marketplace": "UK",
                "product_name": "90L演示耗材100只",
                "sku": "ABT-90-100-C",
                "asin": "B0DEMOTRSH",
                "priority": "P1",
                "issue_type": "库存 / 利润压力",
                "action_group": "成本 / 利润动作",
                "primary_reason": "广告前利润<=0",
                "key_evidence": "广告前利润<=0",
                "confirmed_status": "待确认",
            },
        ],
        [],
        "UK",
        output_dir=OUTPUT_DIR / "__missing_frontend_fixture__",
    )

    assert [row["asin"] for row in rows] == ["B0DEMOFR01"]
    assert rows[0]["product_url"] == "https://www.amazon.co.uk/dp/B0DEMOFR01?th=1"
    assert rows[0]["frontend_core_keyword"]
    assert rows[0]["frontend_search_url"].startswith("https://www.amazon.co.uk/s?k=")
    assert rows[0]["frontend_check_status"] == "待前台检查"
    assert "自动证据不足，不能用于强诊断" in rows[0]["frontend_findings"]
    assert rows[0]["search_term_or_target"] == "serving board"
    assert rows[0]["suggested_action"] == "降竞价10%-20%"
    assert rows[0]["confirmed_note"] == "用户反馈：已执行 serving board 降竞价。"


def test_frontend_check_queue_only_keeps_p0_p1_listing_candidates() -> None:
    from src.report_presentation import _build_frontend_check_queue

    rows = _build_frontend_check_queue(
        [],
        [
            {
                "marketplace": "UK",
                "product_name": "演示笔记本3片",
                "sku": "SKU-DEMO-TRAY-01",
                "asin": "B0DEMOFR01",
                "primary_reason": "Listing 待人工确认",
                "key_evidence": "近14天点击 31；广告订单 0",
            }
        ],
        "UK",
        output_dir=OUTPUT_DIR / "__missing_frontend_fixture__",
    )

    assert rows == []


def test_enhanced_status_rows_preserve_background_reference_bucket() -> None:
    from src.report_presentation import _enhanced_status_rows

    rows = _enhanced_status_rows(
        {
            "enhanced_data_status": {
                "traffic_sales_comparable": True,
                "search_query_comparable": True,
                "traffic_sales_files": [
                    {
                        "filename": "traffic_fresh.xlsx",
                        "period_hint": "week23",
                        "freshness": "fresh",
                        "used_in_diagnosis": "是",
                    },
                    {
                        "filename": "traffic_stale.xlsx",
                        "period_hint": "week21",
                        "freshness": "stale",
                        "used_in_diagnosis": "仅背景参考",
                    },
                    {
                        "filename": "traffic_unused.xlsx",
                        "period_hint": "week20",
                        "freshness": "dated",
                        "used_in_diagnosis": "否",
                    },
                ],
                "search_query_files": [],
            }
        }
    )

    by_filename = {row["文件名"]: row for row in rows}
    assert by_filename["traffic_fresh.xlsx"]["诊断使用类型"] == "强诊断使用"
    assert by_filename["traffic_fresh.xlsx"]["是否参与诊断"] == "是"
    assert by_filename["traffic_stale.xlsx"]["诊断使用类型"] == "仅背景参考"
    assert by_filename["traffic_stale.xlsx"]["是否参与诊断"] == "仅背景参考"
    assert by_filename["traffic_unused.xlsx"]["诊断使用类型"] == "未参与诊断"
    assert by_filename["traffic_unused.xlsx"]["是否参与诊断"] == "否"


def test_ad_copy_boxes_split_pending_actions_by_marketplace() -> None:
    from src.generate_html_report import _render_ad_copy_boxes

    html = _render_ad_copy_boxes(
        [
            {
                "marketplace": "US",
                "search_term_or_target": "B0US111111",
                "suggested_action": "加价 5%-10%",
                "confirmed_status": "待确认",
            },
            {
                "marketplace": "UK",
                "search_term_or_target": "B0UK222222",
                "suggested_action": "加价 5%-10%",
                "confirmed_status": "待确认",
            },
        ],
        "pending-ad",
    )

    assert html.count('class="ad-copy-box"') == 2
    assert 'data-marketplace="US"' in html
    assert 'data-marketplace="UK"' in html
    assert '<span class="status-badge status-muted ad-copy-market">US</span><strong>加价 5%-10%</strong>' in html
    assert '<span class="status-badge status-muted ad-copy-market">UK</span><strong>加价 5%-10%</strong>' in html
    assert 'class="ad-copy-visual"' in html
    assert 'class="ad-copy-text copy-source"' in html
    us_block = html.split('data-copy-group="pending-ad-copy-1"', 1)[1].split('data-copy-group="pending-ad-copy-2"', 1)[0]
    uk_block = html.split('data-copy-group="pending-ad-copy-2"', 1)[1]
    assert "B0US111111" in us_block
    assert "B0UK222222" not in us_block
    assert "站点=" not in us_block
    assert "\t" not in us_block
    assert "B0UK222222" in uk_block
    assert "B0US111111" not in uk_block
    assert "站点=" not in uk_block
    assert "\t" not in uk_block


def test_growth_test_section_has_copy_boxes_for_pending_terms() -> None:
    from src.html_pages import components_ad_workbench as ad_components
    from src import generate_html_report as shared

    html = ad_components._render_growth_test_section(
        shared,
        [
            {
                "marketplace": "US",
                "sku": "SKU-1",
                "asin": "B0TEST1111",
                "product_name": "演示台灯",
                "search_term_or_target": "led desk lamp with tray",
                "suggested_action": "小预算试投",
                "confirmed_status": "待确认",
                "suggested_daily_budget": "$2.00/天",
                "suggested_bid_min": "$0.20",
                "suggested_bid_max": "$0.30",
                "test_days": "7",
                "evidence_level": "核心强相关",
                "traffic_origin": "手动广告",
                "operation_label": "广泛出词，拉精准",
                "match_type": "BROAD",
            }
        ],
    )

    assert 'class="ad-copy-box growth-copy-box"' in html
    assert 'data-copy-target="growth-test-copy-1"' in html
    copy_text = html.split('id="growth-test-copy-1">', 1)[1].split("</pre>", 1)[0]
    assert copy_text == "led desk lamp with tray"
    assert "站点=" not in copy_text
    assert "动作=" not in copy_text
    assert "来源=" not in copy_text
    assert "预算=" not in copy_text
    assert "\t" not in copy_text
    assert "来源 手动广告" in html
    assert "操作 广泛出词，拉精准" in html
    assert "匹配 BROAD" in html
    assert "竞价 $0.20-$0.30" in html
    assert "依据 主词相关" in html
    assert "总预算 $2.00/天" in html
    assert "证据 核心强相关" not in html
    visual_html = html.split('class="ad-copy-visual"', 1)[1].split("</pre>", 1)[0]
    assert "预算 $2.00/天" not in visual_html
    assert "<span>小预算试投</span>" not in visual_html
    assert '<span class="subtle" data-ad-complete-message></span>' in html
    assert '<div class="ad-task-meta" data-ad-complete-message></div>' not in html


def test_ad_complete_control_is_attached_inside_task_card() -> None:
    from src.generate_html_report import REPORT_UI_CSS

    assert ".ad-task-card .ad-complete-row" in REPORT_UI_CSS
    assert "position: absolute;" in REPORT_UI_CSS
    assert "bottom: 12px;" in REPORT_UI_CSS
    assert ".ad-task-card [data-ad-complete-message]" in REPORT_UI_CSS


def test_report_js_moves_legacy_complete_message_into_complete_row() -> None:
    from src.generate_html_report import REPORT_JS

    assert "function attachCompletionRows()" in REPORT_JS
    assert "row.appendChild(span);" in REPORT_JS
    assert "message.remove();" in REPORT_JS


def test_report_js_copy_button_has_textarea_fallback_and_clear_prompt() -> None:
    from src.generate_html_report import REPORT_JS

    assert "legacyTextareaCopy" in REPORT_JS
    assert "document.createElement('textarea')" in REPORT_JS
    assert "按Cmd+C复制" in REPORT_JS
    assert "event.stopImmediatePropagation()" in REPORT_JS
    assert "}, true);" in REPORT_JS


def test_report_js_builds_growth_copy_boxes_for_existing_reports() -> None:
    from src.generate_html_report import REPORT_JS

    assert "function ensureGrowthCopyBoxes()" in REPORT_JS
    assert "growth-test-dynamic-copy-" in REPORT_JS
    assert "growthCopyLine" in REPORT_JS
    assert "function operationLabelDisplay(item)" in REPORT_JS
    assert "自动出词，拉精准" in REPORT_JS
    assert "已在精准，管理原广告" in REPORT_JS
    growth_copy_line = REPORT_JS.split("function growthCopyLine(item)", 1)[1].split("function createEl", 1)[0]
    assert "search_term_or_target" in growth_copy_line
    assert "预算=" not in growth_copy_line
    assert "竞价=" not in growth_copy_line
    dynamic_growth = REPORT_JS.split("function ensureGrowthCopyBoxes()", 1)[1]
    assert "总预算 " in dynamic_growth
    assert "'预算 ' + safeText(item.suggested_daily_budget" not in dynamic_growth


def test_report_js_filters_ad_workbench_and_growth_boxes() -> None:
    from src.generate_html_report import REPORT_JS, REPORT_UI_CSS

    assert "function applyWorkbenchFilter(workbench)" in REPORT_JS
    assert "data-ad-filter-summary" in REPORT_JS
    assert ".ad-task-card, .ad-copy-box" in REPORT_JS
    assert "itemAction(item) !== filters.action" in REPORT_JS
    assert "itemMarketplace(item) !== filters.marketplace" in REPORT_JS
    assert "筛选命中 " in REPORT_JS
    assert "ad-filter-hidden" in REPORT_UI_CSS
    assert ".ad-filter-summary.is-active" in REPORT_UI_CSS


def test_ad_workbench_shows_boundary_reminder_without_price_action_filter() -> None:
    from src.generate_html_report import _render_ad_workbench

    html = _render_ad_workbench(
        [
            {
                "marketplace": "US",
                "search_term_or_target": "B0US111111",
                "suggested_action": "降竞价 10%-20%",
                "confirmed_status": "待确认",
                "clicks": "5",
                "orders": "0",
                "spend": "$1.20",
            }
        ],
        all_marketplaces=True,
        anchor_id="today-ad-actions-all",
        keyword_review_count=12,
    )

    assert 'class="ad-head-left"' in html
    assert 'class="ad-head-right"' in html
    assert 'class="ad-summary-grid"' in html
    assert "只复制待确认动作" in html
    assert "观察项不操作" in html
    assert '<option value="price-down">降价</option>' not in html
    assert '<section class="ad-section primary-action">' in html
    assert 'data-ad-complete-checkbox' in html
    assert 'data-ad-complete-payload=' in html
    head_html = html.split('<div class="ad-workbench-head">', 1)[1].split('<div class="ad-summary-grid">', 1)[0]
    assert "小预算投词 0" in head_html
    assert "历史复盘 12" in head_html
    assert 'href="#action-effect-review"' in head_html
    assert 'data-ad-feedback-refresh' in head_html
    assert 'data-run-report-action="report-refresh"' in head_html
    assert 'data-run-report-status="report-refresh"' in head_html
    assert "筛选栏只筛当前工作台和小预算投词" in html
    assert "完整词级历史复盘见" in html
    assert "低点击隐藏项 0 条在 Excel" in html
    copy_section = html.split('<section class="ad-section primary-action">', 1)[1].split("</section>", 1)[0]
    assert 'data-ad-feedback-refresh' not in copy_section


def test_ad_workbench_hides_internal_action_gate_keys() -> None:
    from src.generate_html_report import _render_ad_workbench

    html = _render_ad_workbench(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-GATE",
                "asin": "B0GATE1234",
                "product_name": "Gate product",
                "search_term_or_target": "gate keyword",
                "suggested_action": "降竞价 10%-20%",
                "confirmed_status": "待确认",
                "clicks": "16",
                "orders": "0",
                "spend": "£8.20",
                "product_level_conclusion": "暂停扩张",
                "product_ad_boundary": "前台和竞品证据不足，今天只允许广告止损动作。",
                "today_allowed_actions": "bid_down / negative_exact / observe",
                "today_blocked_actions": "bid_up / budget_up / broad_scale / create_exact_low_budget",
            }
        ],
        all_marketplaces=False,
        anchor_id="today-ad-actions",
    )

    assert "产品门禁" in html
    assert "仅止损" in html
    assert "前台和竞品证据不足，今天只允许广告止损动作。" in html
    assert "允许 bid_down" not in html
    assert "禁止 bid_up" not in html
    assert "create_exact_low_budget" not in html


def test_ad_workbench_does_not_render_completion_control_for_observation_rows() -> None:
    from src.generate_html_report import _render_ad_workbench

    html = _render_ad_workbench(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "product_name": "Demo notebook",
                "search_term_or_target": "demo notebook",
                "suggested_action": "观察",
                "confirmed_status": "待确认",
                "clicks": "3",
                "orders": "0",
                "spend": "£0.70",
            }
        ],
        all_marketplaces=False,
        anchor_id="today-ad-actions",
    )

    assert "demo notebook" in html
    assert 'data-ad-complete-checkbox' not in html
    assert 'data-ad-complete-payload=' not in html


def test_report_js_can_submit_ad_completion_and_show_local_service_failure() -> None:
    from src.generate_html_report import REPORT_JS

    assert "fetch('http://127.0.0.1:8765/feedback/ad-action-complete'" in REPORT_JS
    assert "fetch('http://127.0.0.1:8765/feedback/ad-action-cancel'" in REPORT_JS
    assert "X-Report-Action-Token" in REPORT_JS
    assert "report-action-token" in REPORT_JS
    assert "JSON.stringify(payloads.length === 1 ? payloads[0] : { actions: payloads })" in REPORT_JS
    assert "data-ad-complete-checkbox" in REPORT_JS
    assert "取消勾选可撤销" in REPORT_JS
    assert "data-ad-complete-cancel" not in REPORT_JS
    assert "本地确认服务未启动" in REPORT_JS
    assert "本地确认服务版本过旧" in REPORT_JS
    assert "data-ad-feedback-refresh" in REPORT_JS
    assert "removeCopyRow" not in REPORT_JS


def test_ad_memory_gate_blocks_cooldown_action_from_copy_area() -> None:
    from src.autoopt_feedback import add_action_identity
    from src.generate_html_report import _render_ad_workbench
    from src.report_presentation import _apply_ad_memory_hard_gate

    row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0TEST1234",
        "search_term_or_target": "dimmer desk lamp",
        "suggested_action": "降竞价10%-20%",
        "copy_action_line": "建议降竞价 10%-20%",
        "confirmed_status": "待确认",
        "clicks": "8",
        "orders": "0",
        "spend": "£4.20",
        "html_visible": "是",
    }
    action_id = add_action_identity(row, row["suggested_action"])["action_id"]

    rows = _apply_ad_memory_hard_gate(
        [row],
        {"blocked_action_ids": [action_id], "keyword_strategy_memory": []},
    )

    assert rows[0]["suggested_action"] == "观察"
    assert rows[0]["ad_memory_blocked"] is True
    assert rows[0]["blocked_action_id"] == action_id
    assert rows[0]["normalized_action"] == "observe"
    assert rows[0]["action_id"] != action_id
    assert rows[0]["confirmed_status"] == "仅背景参考"
    assert rows[0]["copy_block"] == "建议观察\ndimmer desk lamp"

    html = _render_ad_workbench(rows, all_marketplaces=False)
    assert "当前没有需要复制执行的广告动作。" in html
    assert 'data-ad-copy-row' not in html


def test_ad_memory_gate_blocks_cooldown_action_by_term_key_when_action_id_changes() -> None:
    from src.report_presentation import _apply_ad_memory_hard_gate

    row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0TEST1234",
        "search_term_or_target": "dimmer desk lamp",
        "suggested_action": "降竞价10%-20%",
        "copy_action_line": "dimmer desk lamp 降竞价10%-20%",
        "confirmed_status": "待确认",
    }
    cooldown = {
        "action_id": "legacy-action-id",
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0TEST1234",
        "search_term_or_target": "dimmer desk lamp",
        "normalized_action": "bid_down",
        "reason": "已执行，未满3天，不重复操作",
    }

    rows = _apply_ad_memory_hard_gate(
        [row],
        {
            "blocked_action_ids": ["legacy-action-id"],
            "action_cooldowns": {"legacy-action-id": cooldown},
            "keyword_strategy_memory": [],
        },
    )

    assert rows[0]["suggested_action"] == "观察"
    assert rows[0]["ad_memory_blocked"] is True
    assert rows[0]["blocked_action_id"] == "legacy-action-id"
    assert rows[0]["normalized_action"] == "observe"
    assert rows[0]["action_id"] != "legacy-action-id"
    assert rows[0]["confirmed_status"] == "仅背景参考"
    assert rows[0]["keyword_memory_summary"] == "已执行，未满3天，不重复操作"


def test_ad_memory_gate_does_not_block_different_action_with_same_term_key() -> None:
    from src.report_presentation import _apply_ad_memory_hard_gate

    row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0TEST1234",
        "search_term_or_target": "dimmer desk lamp",
        "suggested_action": "加价 5%-10%",
        "copy_action_line": "dimmer desk lamp 加价 5%-10%",
        "confirmed_status": "待确认",
    }
    cooldown = {
        "action_id": "legacy-bid-down-action",
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0TEST1234",
        "search_term_or_target": "dimmer desk lamp",
        "normalized_action": "bid_down",
        "reason": "已执行降竞价冷却中",
    }

    rows = _apply_ad_memory_hard_gate(
        [row],
        {
            "blocked_action_ids": ["legacy-bid-down-action"],
            "action_cooldowns": {"legacy-bid-down-action": cooldown},
            "keyword_strategy_memory": [],
        },
    )

    assert rows[0]["suggested_action"] == "加价 5%-10%"
    assert "ad_memory_blocked" not in rows[0]


def test_ad_memory_gate_blocks_negative_effectiveness_even_without_cooldown_id() -> None:
    from src.autoopt_feedback import add_action_identity
    from src.report_presentation import _apply_ad_memory_hard_gate

    row = {
        "marketplace": "US",
        "sku": "SKU-1",
        "asin": "B0TEST1234",
        "search_term_or_target": "metal board",
        "suggested_action": "加价 5%-10%",
        "copy_action_line": "建议加价 5%-10%",
        "confirmed_status": "待确认",
    }
    action_id = add_action_identity(row, row["suggested_action"])["action_id"]

    rows = _apply_ad_memory_hard_gate(
        [row],
        {
            "blocked_action_ids": [],
            "keyword_strategy_memory": [
                {
                    "action_id": action_id,
                    "effectiveness_score": -1,
                    "should_block_repeating": False,
                    "recommended_future_policy": "recheck_frontend_before_action",
                }
            ],
        },
    )

    assert rows[0]["suggested_action"] == "观察"
    assert rows[0]["scale_action"] == "观察"
    assert rows[0]["copy_action_line"] == "建议观察"
    assert rows[0]["normalized_action"] == "observe"
    assert rows[0]["action_id"] != action_id
    assert rows[0]["confirmed_status"] == "仅背景参考"
    assert rows[0]["keyword_memory_summary"] == "recheck_frontend_before_action"
    assert "自我优化拦截：recheck_frontend_before_action" in rows[0]["classification_reason"]


def test_ad_memory_gate_blocks_halo_only_bid_up_from_copy_area() -> None:
    from src.autoopt_feedback import add_action_identity
    from src.generate_html_report import _render_ad_workbench
    from src.report_presentation import _apply_ad_memory_hard_gate

    row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0TEST1234",
        "search_term_or_target": "dimmer desk lamp",
        "suggested_action": "加价 5%-10%",
        "copy_action_line": "建议加价 5%-10%",
        "confirmed_status": "待确认",
        "html_visible": "是",
    }
    action_id = add_action_identity(row, row["suggested_action"])["action_id"]

    rows = _apply_ad_memory_hard_gate(
        [row],
        {
            "blocked_action_ids": [],
            "keyword_strategy_memory": [
                {
                    "action_id": action_id,
                    "review_outcome": "needs_manual_review",
                    "effectiveness_score": -1,
                    "halo_only_conversion": True,
                    "target_sku_not_converted": True,
                    "attribution_effect_status": "halo_only_conversion",
                    "should_block_repeating": True,
                    "block_reason": "只有光环成交，本 SKU 未验证成交",
                }
            ],
        },
    )

    assert rows[0]["suggested_action"] == "观察"
    assert rows[0]["copy_action_line"] == "建议观察"
    assert rows[0]["ad_memory_blocked"] is True
    assert rows[0]["normalized_action"] == "observe"
    assert rows[0]["action_id"] != action_id
    assert rows[0]["confirmed_status"] == "仅背景参考"
    assert rows[0]["keyword_memory_summary"] == "只有光环成交，需人工复查"
    assert "加价" not in rows[0]["copy_block"]

    html = _render_ad_workbench(rows, all_marketplaces=False)
    assert "当前没有需要复制执行的广告动作。" in html
    assert "建议加价" not in html
    assert 'data-ad-copy-row' not in html


def test_ad_memory_gate_does_not_treat_not_ready_as_historical_failure() -> None:
    from src.autoopt_feedback import add_action_identity
    from src.report_presentation import _apply_ad_memory_hard_gate

    row = {
        "marketplace": "US",
        "sku": "SKU-1",
        "asin": "B0TEST1234",
        "search_term_or_target": "metal board",
        "suggested_action": "加价 5%-10%",
        "copy_action_line": "建议加价 5%-10%",
        "confirmed_status": "待确认",
    }
    action_id = add_action_identity(row, row["suggested_action"])["action_id"]

    rows = _apply_ad_memory_hard_gate(
        [row],
        {
            "blocked_action_ids": [],
            "keyword_strategy_memory": [
                {
                    "action_id": action_id,
                    "review_outcome": "not_ready",
                    "effectiveness_score": None,
                    "should_block_repeating": False,
                    "recommended_future_policy": "review_after_7_days",
                }
            ],
        },
    )

    assert rows[0]["suggested_action"] == "加价 5%-10%"
    assert rows[0].get("ad_memory_blocked") is not True


def test_product_action_review_contract_keeps_structured_metrics() -> None:
    from src.autoopt_feedback import ACTION_REVIEW_REQUIRED_FIELDS, build_action_review_rows

    old_date = (date.today() - timedelta(days=8)).isoformat()
    rows = build_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "ad_clicks": 8,
                                "ad_spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 30,
                                "halo_ad_orders": 0,
                                "total_orders": 3,
                                "ACOS": 0.12,
                                "TACOS": 0.08,
                                "available_stock": 18,
                            }
                        ],
                        "14d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "ad_clicks": 14,
                                "ad_spend": 7,
                                "ad_orders": 2,
                                "promoted_ad_orders": 2,
                                "promoted_ad_sales": 60,
                                "halo_ad_orders": 0,
                                "total_orders": 5,
                                "ACOS": 0.13,
                                "TACOS": 0.09,
                                "available_stock": 18,
                            }
                        ],
                    }
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "product_name": "Demo desk lamp",
                "diagnosis_type": "广告复盘",
                "today_action": "降竞价",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            }
        ],
    )

    assert set(ACTION_REVIEW_REQUIRED_FIELDS).issubset(rows[0])
    assert rows[0]["review_window"] == "7天后复盘"
    assert rows[0]["current_7d_promoted_ad_orders"] == 1.0
    assert rows[0]["current_7d_total_orders"] == 3
    assert rows[0]["current_7d_tacos"] == 0.08
    assert rows[0]["current_7d_available_stock"] == 18
    assert rows[0]["current_14d_promoted_ad_orders"] == 2.0
    assert rows[0]["current_14d_tacos"] == 0.09
    assert rows[0]["current_14d_available_stock"] == 18


def test_product_action_review_does_not_treat_total_orders_as_ad_effective_without_promoted_sku() -> None:
    from src.autoopt_feedback import build_action_review_rows

    old_date = (date.today() - timedelta(days=8)).isoformat()
    rows = build_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "ad_clicks": 4,
                                "ad_spend": 4,
                                "ad_orders": 0,
                                "promoted_ad_orders": 0,
                                "promoted_ad_sales": 0,
                                "halo_ad_orders": 0,
                                "total_orders": 3,
                                "ACOS": "",
                                "TACOS": 0.08,
                                "available_stock": 18,
                            }
                        ],
                        "14d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "ad_clicks": 8,
                                "ad_spend": 6,
                                "ad_orders": 0,
                                "promoted_ad_orders": 0,
                                "promoted_ad_sales": 0,
                                "halo_ad_orders": 0,
                                "total_orders": 5,
                                "TACOS": 0.09,
                                "available_stock": 18,
                            }
                        ],
                    }
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "product_name": "Demo desk lamp",
                "diagnosis_type": "广告复盘",
                "today_action": "降竞价",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            }
        ],
    )

    assert rows[0]["outcome"] == "待观察"
    assert rows[0]["review_outcome"] == "insufficient_sample"
    assert rows[0]["effectiveness_score"] == 0
    assert "本 SKU 广告成交未验证" in rows[0]["rule_adjustment"]
    assert rows[0]["promoted_conversion_improved"] is False


def test_product_action_review_day_three_keeps_positive_signals_as_pending_display() -> None:
    from src.autoopt_feedback import build_action_review_rows

    review_date = (date.today() - timedelta(days=4)).isoformat()
    rows = build_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "ad_clicks": 8,
                                "ad_spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 30,
                                "halo_ad_orders": 0,
                                "halo_ad_sales": 0,
                                "total_orders": 3,
                                "ACOS": "8%",
                                "TACOS": 0.08,
                                "available_stock": 18,
                            }
                        ],
                    }
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "product_name": "Demo desk lamp",
                "diagnosis_type": "广告复盘",
                "today_action": "降竞价",
                "confirmed_status": "已执行",
                "confirmed_at": review_date,
            }
        ],
    )

    assert rows[0]["review_phase"] == "day_3_check"
    assert rows[0]["review_outcome"] == "not_ready"
    assert rows[0]["effectiveness_score"] is None
    assert rows[0]["outcome"] == "待7天确认"
    assert rows[0]["block_reason"] == "等待7天复盘"
    assert rows[0]["promoted_conversion_improved"] is True
    assert rows[0]["current_7d_promoted_ad_orders"] == 1.0
    assert "3天复查口径，7天结论待补" in rows[0]["effect_evidence"]


def test_action_review_uses_report_date_instead_of_system_date_for_review_window() -> None:
    from src.autoopt_feedback import build_action_review_rows

    rows = build_action_review_rows(
        [
            {
                "marketplace": "UK",
                "summary": {"report_date": "2026-06-24"},
                "analysis_payload": {
                    "report_date": "2026-06-24",
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "ad_clicks": 8,
                                "ad_spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 30,
                                "halo_ad_orders": 0,
                                "halo_ad_sales": 0,
                                "total_orders": 3,
                                "ACOS": "8%",
                                "TACOS": 0.08,
                                "available_stock": 18,
                            }
                        ],
                    },
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "product_name": "Demo desk lamp",
                "diagnosis_type": "广告复盘",
                "today_action": "降竞价",
                "confirmed_status": "已执行",
                "confirmed_at": "2026-06-22",
            }
        ],
    )

    assert rows[0]["review_date"] == "2026-06-24"
    assert rows[0]["days_since_execution"] == 2
    assert rows[0]["review_window"] == "未满3天"
    assert rows[0]["review_phase"] == "under_3_days"
    assert rows[0]["outcome"] == "样本不足"
    assert rows[0]["review_outcome"] == "not_ready"


def test_product_action_review_uses_execution_anchored_post_window_over_rolling_metrics() -> None:
    from src.autoopt_feedback import build_action_review_rows

    rows = build_action_review_rows(
        [
            {
                "marketplace": "UK",
                "summary": {"report_date": "2026-06-24"},
                "analysis_payload": {
                    "report_date": "2026-06-24",
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "ad_clicks": 8,
                                "ad_spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 30,
                                "halo_ad_orders": 0,
                                "total_orders": 3,
                                "ACOS": "8%",
                                "TACOS": 0.08,
                                "available_stock": 18,
                            }
                        ],
                    },
                    "review_product_daily": [
                        {
                            "date": "2026-06-17",
                            "marketplace": "UK",
                            "sku": "SKU-1",
                            "asin": "B0TEST1234",
                            "clicks": 4,
                            "spend": 2,
                            "ad_orders": 1,
                            "ad_sales": 30,
                            "promoted_ad_orders": 1,
                            "promoted_ad_sales": 30,
                            "halo_ad_orders": 0,
                            "total_orders": 2,
                            "total_sales": 60,
                            "available_stock": 18,
                        },
                        {
                            "date": "2026-06-20",
                            "marketplace": "UK",
                            "sku": "SKU-1",
                            "asin": "B0TEST1234",
                            "clicks": 10,
                            "spend": 6,
                            "ad_orders": 0,
                            "ad_sales": 0,
                            "promoted_ad_orders": 0,
                            "promoted_ad_sales": 0,
                            "halo_ad_orders": 0,
                            "total_orders": 0,
                            "total_sales": 0,
                            "available_stock": 18,
                        },
                    ],
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "product_name": "Demo desk lamp",
                "diagnosis_type": "广告复盘",
                "today_action": "加价 5%-10%",
                "confirmed_status": "已执行",
                "confirmed_at": "2026-06-18",
            }
        ],
    )

    assert rows[0]["review_data_source"] == "execution_anchored_daily"
    assert rows[0]["post_3d_start"] == "2026-06-18"
    assert rows[0]["post_3d_end"] == "2026-06-20"
    assert rows[0]["post_3d_days"] != ""
    assert rows[0]["post_3d_available_stock"] != ""
    assert rows[0]["post_7d_start"] == "2026-06-18"
    assert rows[0]["post_7d_end"] == "2026-06-24"
    assert rows[0]["pre_7d_promoted_ad_orders"] == 1.0
    assert rows[0]["current_7d_promoted_ad_orders"] == 0.0
    assert rows[0]["post_7d_promoted_ad_orders"] == 0.0
    assert rows[0]["outcome"] not in {"初步有效", "有改善迹象", "明确改善", "可保留"}
    assert rows[0]["promoted_conversion_improved"] is False


def test_product_action_review_high_acos_promoted_order_requires_manual_review_display() -> None:
    from src.autoopt_feedback import build_action_review_rows

    old_date = (date.today() - timedelta(days=8)).isoformat()
    rows = build_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "ad_clicks": 8,
                                "ad_spend": 10,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 20,
                                "halo_ad_orders": 0,
                                "total_orders": 3,
                                "ACOS": 0.50,
                                "TACOS": 0.20,
                                "available_stock": 18,
                            }
                        ],
                    }
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "product_name": "Demo desk lamp",
                "diagnosis_type": "广告复盘",
                "today_action": "降竞价",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            }
        ],
    )

    assert rows[0]["review_phase"] == "day_7_review"
    assert rows[0]["review_outcome"] == "needs_manual_review"
    assert rows[0]["effectiveness_score"] == -1
    assert rows[0]["outcome"] == "待人工复查"
    assert rows[0]["block_reason"] == "本 SKU 有单但 ACOS 高于目标，需人工复查"
    assert rows[0]["rule_adjustment"] == "本 SKU 有单但 ACOS 高于目标，需人工复查；不能把该动作判为可复用。"
    assert "类似动作保留" not in rows[0]["rule_adjustment"]


def test_product_action_review_promoted_order_missing_support_metrics_requires_manual_review() -> None:
    from src.autoopt_feedback import build_action_review_rows

    old_date = (date.today() - timedelta(days=8)).isoformat()
    rows = build_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "ad_clicks": 8,
                                "ad_spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 50,
                                "halo_ad_orders": 0,
                                "ACOS": "8%",
                                "target_acos": "10%",
                            }
                        ],
                    }
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "product_name": "Demo desk lamp",
                "diagnosis_type": "广告复盘",
                "today_action": "加价 5%-10%",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            }
        ],
    )

    assert rows[0]["review_phase"] == "day_7_review"
    assert rows[0]["review_outcome"] == "needs_manual_review"
    assert rows[0]["effectiveness_score"] == -1
    assert rows[0]["outcome"] == "待人工复查"
    assert rows[0]["promoted_conversion_improved"] is True
    assert rows[0]["current_7d_promoted_ad_orders"] == 1.0
    assert rows[0]["current_7d_tacos"] == ""
    assert rows[0]["current_7d_total_orders"] == ""
    assert rows[0]["current_7d_available_stock"] == ""
    assert "缺少TACOS、总单、库存" in rows[0]["block_reason"]


def test_keyword_review_separates_cooldown_from_effectiveness_outcome() -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS, build_keyword_action_review_rows

    today = date.today()
    under_three_rows = build_keyword_action_review_rows(
        [],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "dimmer desk lamp",
                "manual_action_taken": "小预算试投",
                "confirmed_status": "已执行",
                "confirmed_at": today.isoformat(),
            }
        ],
    )

    assert under_three_rows[0]["cooldown_status"] == "cooldown_active"
    assert under_three_rows[0]["review_phase"] == "under_3_days"
    assert under_three_rows[0]["review_outcome"] == "not_ready"
    assert under_three_rows[0]["effectiveness_score"] is None
    assert under_three_rows[0]["block_reason"] == "已执行，冷却中"
    assert under_three_rows[0]["cooldown_until"] == (today + timedelta(days=3)).isoformat()
    assert set(KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS).issubset(under_three_rows[0])

    old_date = (today - timedelta(days=8)).isoformat()
    ineffective_rows = build_keyword_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                        "搜索词分析": {
                            "7d": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "dimmer desk lamp",
                                "clicks": 9,
                                "spend": 6,
                                "ad_orders": 0,
                            }
                        ]
                    }
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "dimmer desk lamp",
                "manual_action_taken": "加价 5%-10%",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            }
        ],
    )

    assert ineffective_rows[0]["cooldown_status"] == "review_ready"
    assert ineffective_rows[0]["review_phase"] == "day_7_review"
    assert ineffective_rows[0]["review_outcome"] == "ineffective"
    assert ineffective_rows[0]["effectiveness_score"] == -2
    assert ineffective_rows[0]["block_reason"] == "历史效果差，不重复推送"

    halo_rows = build_keyword_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "搜索词分析": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "dimmer desk lamp",
                                "clicks": 4,
                                "spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 0,
                                "halo_ad_orders": 1,
                            }
                        ]
                    }
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "dimmer desk lamp",
                "manual_action_taken": "加价 5%-10%",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            }
        ],
    )

    assert halo_rows[0]["review_outcome"] == "needs_manual_review"
    assert halo_rows[0]["effectiveness_score"] == -1
    assert halo_rows[0]["block_reason"] == "只有光环成交，需人工复查"

    effective_rows = build_keyword_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "total_orders": 3,
                                "TACOS": 0.08,
                                "available_stock": 24,
                            }
                        ],
                        "14d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "total_orders": 5,
                                "TACOS": 0.09,
                                "available_stock": 24,
                            }
                        ],
                    },
                    "搜索词分析": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "low acos term",
                                "clicks": 4,
                                "spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 50,
                                "ACOS": "8%",
                                "target_acos": "10%",
                            }
                        ],
                        "14d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "low acos term",
                                "clicks": 8,
                                "spend": 8,
                                "ad_orders": 2,
                                "promoted_ad_orders": 2,
                                "promoted_ad_sales": 90,
                                "ACOS": "9%",
                            }
                        ],
                    },
                    **anchored_keyword_review_payload(["low acos term"], old_date),
                }
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "low acos term",
                "manual_action_taken": "加价 5%-10%",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            }
        ],
    )

    assert effective_rows[0]["review_outcome"] == "effective"
    assert effective_rows[0]["effectiveness_score"] == 2
    assert effective_rows[0]["current_7d_total_orders"] == 3
    assert effective_rows[0]["current_7d_tacos"] == 0.08
    assert effective_rows[0]["current_7d_available_stock"] == 24
    assert effective_rows[0]["current_14d_total_orders"] == 3.0
    assert effective_rows[0]["current_14d_tacos"] == 0.08
    assert effective_rows[0]["current_14d_available_stock"] == 24

    growth_test_effective_rows = build_keyword_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "total_orders": 3,
                                "TACOS": 0.08,
                                "available_stock": 24,
                            }
                        ],
                        "14d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "total_orders": 5,
                                "TACOS": 0.09,
                                "available_stock": 24,
                            }
                        ],
                    },
                    "搜索词分析": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "small test term",
                                "clicks": 4,
                                "spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 50,
                                "ACOS": "8%",
                                "target_acos": "10%",
                            },
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "pull exact term",
                                "clicks": 4,
                                "spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 50,
                                "ACOS": "8%",
                                "target_acos": "10%",
                            },
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "new exact term",
                                "clicks": 4,
                                "spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 50,
                                "ACOS": "8%",
                                "target_acos": "10%",
                            },
                        ],
                        "14d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "small test term",
                                "clicks": 8,
                                "spend": 8,
                                "ad_orders": 2,
                                "promoted_ad_orders": 2,
                                "promoted_ad_sales": 90,
                                "ACOS": "9%",
                            },
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "pull exact term",
                                "clicks": 8,
                                "spend": 8,
                                "ad_orders": 2,
                                "promoted_ad_orders": 2,
                                "promoted_ad_sales": 90,
                                "ACOS": "9%",
                            },
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "new exact term",
                                "clicks": 8,
                                "spend": 8,
                                "ad_orders": 2,
                                "promoted_ad_orders": 2,
                                "promoted_ad_sales": 90,
                                "ACOS": "9%",
                            },
                        ],
                    },
                    **anchored_keyword_review_payload(
                        ["small test term", "pull exact term", "new exact term"],
                        old_date,
                    ),
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "small test term",
                "manual_action_taken": "小预算试投",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            },
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "pull exact term",
                "manual_action_taken": "拉精准",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            },
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "new exact term",
                "manual_action_taken": "新建精准",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            },
        ],
    )

    growth_rows_by_term = {row["search_term_or_target"]: row for row in growth_test_effective_rows}
    assert growth_rows_by_term["small test term"]["review_outcome"] == "effective"
    assert growth_rows_by_term["small test term"]["effectiveness_score"] == 2
    assert growth_rows_by_term["small test term"]["normalized_action"] == "growth_test"
    assert "小预算试投后本 SKU 出单" in growth_rows_by_term["small test term"]["rule_adjustment"]
    assert "不追加预算" in growth_rows_by_term["small test term"]["rule_adjustment"]
    assert growth_rows_by_term["small test term"]["current_7d_promoted_ad_orders"] == 1
    assert growth_rows_by_term["small test term"]["current_7d_tacos"] == 0.08

    assert growth_rows_by_term["pull exact term"]["review_outcome"] == "effective"
    assert growth_rows_by_term["pull exact term"]["effectiveness_score"] == 2
    assert growth_rows_by_term["pull exact term"]["normalized_action"] == "create_exact"
    assert "拉精准" in growth_rows_by_term["pull exact term"]["rule_adjustment"]
    assert "不追加预算" in growth_rows_by_term["pull exact term"]["rule_adjustment"]
    assert growth_rows_by_term["pull exact term"]["current_7d_promoted_ad_orders"] == 1
    assert growth_rows_by_term["pull exact term"]["current_7d_tacos"] == 0.08

    assert growth_rows_by_term["new exact term"]["review_outcome"] == "effective"
    assert growth_rows_by_term["new exact term"]["effectiveness_score"] == 2
    assert growth_rows_by_term["new exact term"]["normalized_action"] == "create_exact"
    assert "拉精准" in growth_rows_by_term["new exact term"]["rule_adjustment"]
    assert "不追加预算" in growth_rows_by_term["new exact term"]["rule_adjustment"]
    assert growth_rows_by_term["new exact term"]["current_7d_promoted_ad_orders"] == 1
    assert growth_rows_by_term["new exact term"]["current_7d_tacos"] == 0.08

    rolling_only_positive_rows = build_keyword_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-ROLLING",
                                "asin": "B0ROLLING1",
                                "total_orders": 3,
                                "TACOS": 0.08,
                                "available_stock": 24,
                            }
                        ],
                        "14d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-ROLLING",
                                "asin": "B0ROLLING1",
                                "total_orders": 5,
                                "TACOS": 0.09,
                                "available_stock": 24,
                            }
                        ],
                    },
                    "搜索词分析": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-ROLLING",
                                "asin": "B0ROLLING1",
                                "search_term": "rolling good term",
                                "clicks": 4,
                                "spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 50,
                                "ACOS": "8%",
                                "target_acos": "10%",
                            }
                        ],
                        "14d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-ROLLING",
                                "asin": "B0ROLLING1",
                                "search_term": "rolling good term",
                                "clicks": 8,
                                "spend": 8,
                                "ad_orders": 2,
                                "promoted_ad_orders": 2,
                                "promoted_ad_sales": 90,
                                "ACOS": "9%",
                            }
                        ],
                    },
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-ROLLING",
                "asin": "B0ROLLING1",
                "search_term_or_target": "rolling good term",
                "manual_action_taken": "加价 5%-10%",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            }
        ],
    )

    assert rolling_only_positive_rows[0]["review_data_source"] == "rolling_window"
    assert rolling_only_positive_rows[0]["review_outcome"] == "needs_manual_review"
    assert rolling_only_positive_rows[0]["effectiveness_score"] == -1
    assert rolling_only_positive_rows[0]["outcome"] == "待人工复查"
    assert rolling_only_positive_rows[0]["block_reason"] == "缺少执行锚定7天复盘证据，需人工复查"
    assert rolling_only_positive_rows[0]["outcome"] not in {"初步有效", "有改善迹象", "明确改善", "可保留"}

    zero_stock_rows = build_keyword_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "total_orders": 3,
                                "TACOS": 0.16,
                                "available_stock": 0,
                            }
                        ],
                    },
                    "搜索词分析": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "zero stock low acos term",
                                "clicks": 4,
                                "spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 50,
                                "ACOS": "8%",
                                "target_acos": "10%",
                            }
                        ]
                    },
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "zero stock low acos term",
                "manual_action_taken": "加价 5%-10%",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            }
        ],
    )

    assert zero_stock_rows[0]["review_outcome"] == "needs_manual_review"
    assert zero_stock_rows[0]["effectiveness_score"] == -1
    assert zero_stock_rows[0]["current_7d_available_stock"] == 0
    assert zero_stock_rows[0]["block_reason"] == "本 SKU 有单但当前库存不足，需人工复查"
    assert zero_stock_rows[0]["outcome"] == "待人工复查"

    zero_total_orders_rows = build_keyword_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "total_orders": 0,
                                "TACOS": 0.08,
                                "available_stock": 24,
                            }
                        ],
                        "14d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "total_orders": 5,
                                "TACOS": 0.09,
                                "available_stock": 24,
                            }
                        ],
                    },
                    "搜索词分析": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "zero total low acos term",
                                "clicks": 4,
                                "spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 50,
                                "ACOS": "8%",
                                "target_acos": "10%",
                            }
                        ]
                    },
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "zero total low acos term",
                "manual_action_taken": "加价 5%-10%",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            }
        ],
    )

    assert zero_total_orders_rows[0]["review_outcome"] == "needs_manual_review"
    assert zero_total_orders_rows[0]["effectiveness_score"] == -1
    assert zero_total_orders_rows[0]["block_reason"] == "本 SKU 有广告单但总单未验证，需人工复查"
    assert zero_total_orders_rows[0]["outcome"] == "待人工复查"

    high_tacos_rows = build_keyword_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "total_orders": 3,
                                "TACOS": 0.18,
                                "available_stock": 24,
                            }
                        ],
                    },
                    "搜索词分析": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "high tacos low acos term",
                                "clicks": 4,
                                "spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 50,
                                "ACOS": "8%",
                                "target_acos": "10%",
                            }
                        ]
                    },
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "high tacos low acos term",
                "manual_action_taken": "加价 5%-10%",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            }
        ],
    )

    assert high_tacos_rows[0]["review_outcome"] == "needs_manual_review"
    assert high_tacos_rows[0]["effectiveness_score"] == -1
    assert high_tacos_rows[0]["current_7d_tacos"] == 0.18
    assert high_tacos_rows[0]["current_7d_target_acos"] == "10%"
    assert high_tacos_rows[0]["block_reason"] == "本 SKU 有单但 TACOS 高于目标，需人工复查"
    assert high_tacos_rows[0]["outcome"] == "待人工复查"
    assert high_tacos_rows[0]["rule_adjustment"] == "本 SKU 有单但 TACOS 高于目标，需人工复查；暂不继续加价或放量。"
    assert "可保留当前竞价" not in high_tacos_rows[0]["rule_adjustment"]

    effective_integer_percent_rows = build_keyword_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "total_orders": 3,
                                "TACOS": 0.08,
                                "available_stock": 24,
                            }
                        ],
                        "14d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "total_orders": 5,
                                "TACOS": 0.09,
                                "available_stock": 24,
                            }
                        ],
                    },
                    "搜索词分析": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "integer acos term",
                                "clicks": 4,
                                "spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 50,
                                "ACOS": 8,
                                "target_acos": 10,
                            }
                        ],
                        "14d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "integer acos term",
                                "clicks": 8,
                                "spend": 8,
                                "ad_orders": 2,
                                "promoted_ad_orders": 2,
                                "promoted_ad_sales": 90,
                                "ACOS": 9,
                            }
                        ],
                    },
                    **anchored_keyword_review_payload(["integer acos term"], old_date, target_acos=10),
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "integer acos term",
                "manual_action_taken": "加价 5%-10%",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            }
        ],
    )

    assert effective_integer_percent_rows[0]["review_outcome"] == "effective"
    assert effective_integer_percent_rows[0]["effectiveness_score"] == 2

    day_three_rows = build_keyword_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "total_orders": 3,
                                "TACOS": 0.16,
                                "available_stock": 24,
                            }
                        ],
                    },
                    "搜索词分析": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "day three low acos term",
                                "clicks": 4,
                                "spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 50,
                                "halo_ad_orders": 0,
                                "halo_ad_sales": 0,
                                "ACOS": "8%",
                                "target_acos": "10%",
                            }
                        ]
                    },
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "day three low acos term",
                "manual_action_taken": "加价 5%-10%",
                "confirmed_status": "已执行",
                "confirmed_at": (today - timedelta(days=4)).isoformat(),
            }
        ],
    )

    assert day_three_rows[0]["review_phase"] == "day_3_check"
    assert day_three_rows[0]["review_outcome"] == "not_ready"
    assert day_three_rows[0]["effectiveness_score"] is None
    assert day_three_rows[0]["outcome"] == "待7天确认"
    assert day_three_rows[0]["block_reason"] == "等待7天复盘"
    assert day_three_rows[0]["promoted_conversion_improved"] is True
    assert day_three_rows[0]["current_7d_promoted_ad_orders"] == 1.0
    assert day_three_rows[0]["current_7d_acos"] == "8%"
    assert day_three_rows[0]["current_7d_tacos"] == 0.16
    assert day_three_rows[0]["current_7d_total_orders"] == 3
    assert day_three_rows[0]["current_7d_available_stock"] == 24
    assert "3天复查口径，7天结论待补" in day_three_rows[0]["effect_evidence"]
    assert "未满7天" in day_three_rows[0]["rule_adjustment"]
    assert "保留当前竞价" not in day_three_rows[0]["rule_adjustment"]

    day_three_bid_down_rows = build_keyword_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "total_orders": 3,
                                "TACOS": 0.16,
                                "available_stock": 24,
                            }
                        ],
                    },
                    "搜索词分析": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "day three bid down term",
                                "clicks": 4,
                                "spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 50,
                                "halo_ad_orders": 0,
                                "halo_ad_sales": 0,
                                "ACOS": "8%",
                                "target_acos": "10%",
                            }
                        ]
                    },
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "day three bid down term",
                "manual_action_taken": "降竞价 10%-20%",
                "confirmed_status": "已执行",
                "confirmed_at": (today - timedelta(days=4)).isoformat(),
            }
        ],
    )

    assert day_three_bid_down_rows[0]["review_outcome"] == "not_ready"
    assert day_three_bid_down_rows[0]["outcome"] == "待7天确认"
    assert "未满7天" in day_three_bid_down_rows[0]["rule_adjustment"]
    assert "保留当前竞价" not in day_three_bid_down_rows[0]["rule_adjustment"]

    halo_attribution_rows = build_keyword_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "total_orders": 3,
                                "TACOS": 0.16,
                                "available_stock": 24,
                            }
                        ],
                        "14d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "total_orders": 4,
                                "TACOS": 0.15,
                                "available_stock": 24,
                            }
                        ],
                    },
                    "搜索词分析": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "halo attribution term",
                                "clicks": 8,
                                "spend": 5,
                                "ad_orders": 1,
                                "promoted_ad_orders": 0,
                                "promoted_ad_sales": 0,
                                "halo_ad_orders": 1,
                                "halo_ad_sales": 40,
                                "ACOS": "12%",
                                "target_acos": "20%",
                            }
                        ],
                        "14d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "halo attribution term",
                                "clicks": 14,
                                "spend": 8,
                                "ad_orders": 2,
                                "promoted_ad_orders": 0,
                                "promoted_ad_sales": 0,
                                "halo_ad_orders": 2,
                                "halo_ad_sales": 80,
                                "ACOS": "10%",
                            }
                        ],
                    },
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "halo attribution term",
                "manual_action_taken": "加价 5%-10%",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            }
        ],
    )

    assert halo_attribution_rows[0]["current_7d_ad_orders"] == 1
    assert halo_attribution_rows[0]["current_7d_promoted_ad_orders"] == 0.0
    assert halo_attribution_rows[0]["current_7d_halo_ad_orders"] == 1.0
    assert halo_attribution_rows[0]["halo_only_conversion"] is True
    assert halo_attribution_rows[0]["target_sku_not_converted"] is True
    assert halo_attribution_rows[0]["review_outcome"] == "needs_manual_review"
    assert halo_attribution_rows[0]["effectiveness_score"] == -1
    assert "只有光环成交" in halo_attribution_rows[0]["block_reason"]

    missing_product_metrics_rows = build_keyword_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "搜索词分析": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "low acos term",
                                "clicks": 4,
                                "spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 50,
                                "ACOS": "8%",
                                "target_acos": "10%",
                            }
                        ]
                    },
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "low acos term",
                "manual_action_taken": "加价 5%-10%",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            }
        ],
    )

    assert missing_product_metrics_rows[0]["review_outcome"] == "needs_manual_review"
    assert missing_product_metrics_rows[0]["effectiveness_score"] == -1
    assert "缺少TACOS、总单、库存" in missing_product_metrics_rows[0]["block_reason"]

    missing_14d_support_rows = build_keyword_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "total_orders": 3,
                                "TACOS": 0.08,
                                "available_stock": 24,
                            }
                        ],
                        "14d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "total_orders": 5,
                                "TACOS": 0.09,
                            }
                        ],
                    },
                    "搜索词分析": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "low acos term",
                                "clicks": 4,
                                "spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 50,
                                "ACOS": "8%",
                                "target_acos": "10%",
                            }
                        ],
                        "14d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "low acos term",
                                "clicks": 8,
                                "spend": 8,
                                "ad_orders": 2,
                                "promoted_ad_orders": 2,
                                "promoted_ad_sales": 90,
                                "ACOS": "9%",
                                "target_acos": "10%",
                            }
                        ],
                    },
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "low acos term",
                "manual_action_taken": "加价 5%-10%",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            }
        ],
    )

    assert missing_14d_support_rows[0]["review_outcome"] == "needs_manual_review"
    assert missing_14d_support_rows[0]["effectiveness_score"] == -1
    assert missing_14d_support_rows[0]["outcome"] == "待人工复查"
    assert "14天本 SKU 有单但缺少库存" in missing_14d_support_rows[0]["block_reason"]
    assert "暂不继续加价或放量" in missing_14d_support_rows[0]["rule_adjustment"]

    high_acos_rows = build_keyword_action_review_rows(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "搜索词分析": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "high acos term",
                                "clicks": 4,
                                "spend": 4,
                                "ad_orders": 1,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 20,
                                "ACOS": "25%",
                                "target_acos": "10%",
                            }
                        ]
                    }
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "high acos term",
                "manual_action_taken": "加价 5%-10%",
                "confirmed_status": "已执行",
                "confirmed_at": old_date,
            }
        ],
    )

    assert high_acos_rows[0]["review_outcome"] == "needs_manual_review"
    assert high_acos_rows[0]["effectiveness_score"] == -1
    assert high_acos_rows[0]["outcome"] == "待人工复查"
    assert high_acos_rows[0]["block_reason"] == "本 SKU 有单但 ACOS 高于目标，需人工复查"
    assert high_acos_rows[0]["rule_adjustment"] == "本 SKU 有单但 ACOS 高于目标，需人工复查；暂不继续加价或放量。"
    assert "可保留当前竞价" not in high_acos_rows[0]["rule_adjustment"]


def test_keyword_strategy_memory_uses_standard_outcome_for_blocking(tmp_path) -> None:
    from src.autoopt_feedback import build_keyword_strategy_memory

    memory = build_keyword_strategy_memory(
        output_dir=tmp_path,
        current_keyword_reviews=[
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "not ready term",
                "action_detail": "小预算试投",
                "review_outcome": "not_ready",
                "effectiveness_score": None,
                "review_status": "未满3天，暂不判断效果",
                "review_phase": "under_3_days",
            },
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "bad term",
                "action_detail": "加价 5%-10%",
                "review_outcome": "ineffective",
                "effectiveness_score": -2,
                "days_since_execution": 8,
                "review_status": "可做7天复盘",
                "block_reason": "历史效果差，不重复推送",
            },
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "halo term",
                "action_detail": "加价 5%-10%",
                "review_outcome": "needs_manual_review",
                "effectiveness_score": -1,
                "halo_only_conversion": True,
                "attribution_effect_status": "halo_only_conversion",
            },
        ],
    )

    by_term = {row["search_term_or_target"]: row for row in memory}
    assert by_term["not ready term"]["effectiveness_score"] is None
    assert by_term["not ready term"]["should_block_repeating"] is False
    assert by_term["bad term"]["should_block_repeating"] is True
    assert by_term["bad term"]["review_outcome"] == "ineffective"
    assert by_term["halo term"]["should_block_repeating"] is True
    assert by_term["halo term"]["review_outcome"] == "needs_manual_review"


def test_keyword_strategy_memory_downgrades_legacy_effective_without_support_metrics(tmp_path) -> None:
    from src.autoopt_feedback import build_keyword_strategy_memory

    memory = build_keyword_strategy_memory(
        output_dir=tmp_path,
        current_keyword_reviews=[
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "legacy effective term",
                "action_detail": "加价 5%-10%",
                "review_outcome": "effective",
                "effectiveness_score": 2,
                "days_since_execution": 8,
                "promoted_conversion_improved": True,
                "halo_only_conversion": False,
                "target_sku_not_converted": False,
                "current_7d_promoted_ad_orders": 1,
                "current_7d_acos": "8%",
                "current_7d_tacos": "",
                "current_7d_total_orders": "",
                "current_7d_available_stock": "",
                **anchored_review_contract_fields(),
                "block_reason": "历史判断有效，继续提高优先级",
            }
        ],
    )

    row = memory[0]

    assert row["review_outcome"] == "needs_manual_review"
    assert row["effectiveness_score"] == -1
    assert row["should_keep"] is False
    assert row["should_block_repeating"] is True
    assert row["should_recheck_frontend"] is True
    assert "缺少有效复盘指标" in row["block_reason"]
    assert "继续提高优先级" not in row["block_reason"]


def test_action_learning_policy_uses_standard_review_outcome_before_display_outcome(tmp_path) -> None:
    from src.autoopt_feedback import build_action_learning_policy

    policy = build_action_learning_policy(
        output_dir=tmp_path,
        current_action_reviews=[
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0PRODUCT1",
                "product_name": "Product 1",
                "action_detail": "降竞价",
                "outcome": "有改善迹象",
                "review_outcome": "not_ready",
                "effectiveness_score": None,
                "review_phase": "day_3_check",
                "review_window": "3天后复盘",
            }
        ],
        current_keyword_reviews=[
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0PRODUCT1",
                "product_name": "Product 1",
                "search_term_or_target": "dimmer desk lamp",
                "action_detail": "加价 5%-10%",
                "outcome": "有改善迹象",
                "review_outcome": "not_ready",
                "effectiveness_score": None,
                "review_phase": "day_3_check",
                "review_window": "3天后复盘",
            }
        ],
    )

    assert policy["action_learning_summary"]["positive_count"] == 0
    assert policy["action_learning_summary"]["pending_count"] == 2
    assert policy["product_scores"]["UK||SKU-1||B0PRODUCT1"] == 0
    assert policy["term_scores"]["UK||SKU-1||B0PRODUCT1||dimmer desk lamp"] == 0
    assert policy["positive_action_patterns"] == []
    assert len(policy["pending_review_objects"]) == 2


def test_action_learning_policy_keeps_seven_day_effective_as_positive(tmp_path) -> None:
    from src.autoopt_feedback import build_action_learning_policy

    policy = build_action_learning_policy(
        output_dir=tmp_path,
        current_keyword_reviews=[
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0PRODUCT1",
                "product_name": "Product 1",
                "search_term_or_target": "low acos term",
                "action_detail": "加价 5%-10%",
                "outcome": "待7天确认",
                "review_outcome": "effective",
                "effectiveness_score": 2,
                "days_since_execution": 8,
                "review_phase": "day_7_review",
                "review_window": "7天后复盘",
                "promoted_conversion_improved": True,
                "halo_only_conversion": False,
                "target_sku_not_converted": False,
                "current_7d_clicks": 8,
                "current_7d_spend": 3.2,
                "current_7d_promoted_ad_orders": 1,
                "current_7d_acos": "8%",
                "current_7d_target_acos": "20%",
                "current_7d_tacos": 0.16,
                "current_7d_total_orders": 3,
                "current_7d_available_stock": 24,
                "current_14d_promoted_ad_orders": 2,
                "current_14d_acos": "9%",
                "current_14d_total_orders": 5,
                "current_14d_tacos": 0.12,
                "current_14d_available_stock": 24,
                **anchored_review_contract_fields(),
            }
        ],
    )

    assert policy["action_learning_summary"]["positive_count"] == 1
    assert policy["action_learning_summary"]["pending_count"] == 0
    assert policy["term_scores"]["UK||SKU-1||B0PRODUCT1||low acos term"] == 2
    assert policy["positive_action_patterns"][0]["search_term_or_target"] == "low acos term"


def test_action_learning_policy_downgrades_effective_when_tacos_exceeds_target(tmp_path) -> None:
    from src.autoopt_feedback import build_action_learning_policy

    policy = build_action_learning_policy(
        output_dir=tmp_path,
        current_keyword_reviews=[
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0PRODUCT1",
                "product_name": "Product 1",
                "search_term_or_target": "high tacos term",
                "action_detail": "加价 5%-10%",
                "outcome": "待7天确认",
                "review_outcome": "effective",
                "effectiveness_score": 2,
                "days_since_execution": 8,
                "review_phase": "day_7_review",
                "review_window": "7天后复盘",
                "promoted_conversion_improved": True,
                "halo_only_conversion": False,
                "target_sku_not_converted": False,
                "current_7d_clicks": 8,
                "current_7d_spend": 3.2,
                "current_7d_promoted_ad_orders": 1,
                "current_7d_acos": "8%",
                "current_7d_target_acos": "10%",
                "current_7d_tacos": "18%",
                "current_7d_total_orders": 3,
                "current_7d_available_stock": 24,
                "current_14d_promoted_ad_orders": 2,
                "current_14d_acos": "9%",
                "current_14d_total_orders": 5,
                "current_14d_tacos": "12%",
                "current_14d_available_stock": 24,
                **anchored_review_contract_fields(),
            }
        ],
    )

    assert policy["action_learning_summary"]["positive_count"] == 0
    assert policy["action_learning_summary"]["negative_count"] == 1
    assert policy["term_scores"]["UK||SKU-1||B0PRODUCT1||high tacos term"] == -1
    assert policy["positive_action_patterns"] == []
    assert policy["negative_action_patterns"][0]["block_reason"] == "本 SKU 有单但 TACOS 高于目标，需人工复查"


def test_learned_rules_require_promoted_sku_effective_contract() -> None:
    from src.autoopt_feedback import build_learned_rules

    rows = build_learned_rules(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-VALID",
                "asin": "B0VALID",
                "product_name": "Valid product",
                "action_type": "广告动作",
                "action_detail": "加价 5%-10%",
                "outcome": "有改善迹象",
                "review_outcome": "effective",
                "effectiveness_score": 2,
                "days_since_execution": 8,
                "promoted_conversion_improved": True,
                "halo_only_conversion": False,
                "target_sku_not_converted": False,
                "current_7d_clicks": 8,
                "current_7d_spend": 3.2,
                "current_7d_promoted_ad_orders": 1,
                "current_7d_ad_orders": 1,
                "current_7d_acos": "8%",
                "current_7d_target_acos": "15%",
                "current_7d_tacos": "6%",
                "current_7d_total_orders": 3,
                "current_7d_available_stock": 18,
                "current_14d_promoted_ad_orders": 2,
                "current_14d_ad_orders": 2,
                "current_14d_acos": "9%",
                "current_14d_total_orders": 5,
                "current_14d_tacos": "7%",
                "current_14d_available_stock": 18,
                **anchored_review_contract_fields(),
            },
            {
                "marketplace": "UK",
                "sku": "SKU-HALO",
                "asin": "B0HALO",
                "product_name": "Halo product",
                "action_type": "广告动作",
                "action_detail": "加价 5%-10%",
                "outcome": "初步有效",
                "days_since_execution": 8,
                "promoted_conversion_improved": False,
                "halo_only_conversion": True,
                "target_sku_not_converted": True,
                "current_7d_ad_orders": 1,
                "current_7d_promoted_ad_orders": 0,
                "current_7d_halo_ad_orders": 1,
                "current_7d_acos": "8%",
                "current_7d_tacos": "6%",
                "current_7d_total_orders": 3,
                "current_7d_available_stock": 18,
            },
            {
                "marketplace": "UK",
                "sku": "SKU-HIGH-ACOS",
                "asin": "B0HIGHACOS",
                "product_name": "High ACOS product",
                "action_type": "广告动作",
                "action_detail": "加价 5%-10%",
                "outcome": "有改善迹象",
                "review_outcome": "effective",
                "effectiveness_score": 2,
                "days_since_execution": 8,
                "promoted_conversion_improved": True,
                "halo_only_conversion": False,
                "target_sku_not_converted": False,
                "current_7d_clicks": 8,
                "current_7d_spend": 3.2,
                "current_7d_promoted_ad_orders": 1,
                "current_7d_ad_orders": 1,
                "current_7d_acos": "25%",
                "current_7d_target_acos": "10%",
                "current_7d_tacos": "8%",
                "current_7d_total_orders": 3,
                "current_7d_available_stock": 18,
                "current_14d_promoted_ad_orders": 2,
                "current_14d_ad_orders": 2,
                "current_14d_acos": "12%",
                "current_14d_total_orders": 5,
                "current_14d_tacos": "9%",
                "current_14d_available_stock": 18,
                **anchored_review_contract_fields(),
            },
        ]
    )

    assert [row["sku"] for row in rows] == ["SKU-VALID"]
    assert rows[0]["promoted_conversion_improved"] is True
    assert rows[0]["halo_only_conversion"] is False


def test_action_learning_policy_downgrades_effective_when_stock_is_zero(tmp_path) -> None:
    from src.autoopt_feedback import build_action_learning_policy

    policy = build_action_learning_policy(
        output_dir=tmp_path,
        current_keyword_reviews=[
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0PRODUCT1",
                "product_name": "Product 1",
                "search_term_or_target": "zero stock term",
                "action_detail": "加价 5%-10%",
                "outcome": "待7天确认",
                "review_outcome": "effective",
                "effectiveness_score": 2,
                "days_since_execution": 8,
                "review_phase": "day_7_review",
                "review_window": "7天后复盘",
                "promoted_conversion_improved": True,
                "halo_only_conversion": False,
                "target_sku_not_converted": False,
                "current_7d_clicks": 8,
                "current_7d_spend": 3.2,
                "current_7d_promoted_ad_orders": 1,
                "current_7d_acos": "8%",
                "current_7d_tacos": 0.16,
                "current_7d_total_orders": 3,
                "current_7d_available_stock": 0,
                "current_14d_promoted_ad_orders": 2,
                "current_14d_acos": "9%",
                "current_14d_total_orders": 5,
                "current_14d_tacos": 0.12,
                "current_14d_available_stock": 24,
                **anchored_review_contract_fields(),
                "block_reason": "历史判断有效，继续提高优先级",
            }
        ],
    )

    assert policy["action_learning_summary"]["positive_count"] == 0
    assert policy["action_learning_summary"]["negative_count"] == 1
    assert policy["term_scores"]["UK||SKU-1||B0PRODUCT1||zero stock term"] == -1
    assert policy["positive_action_patterns"] == []
    assert policy["negative_action_patterns"][0]["score"] == -1
    assert policy["negative_action_patterns"][0]["block_reason"] == "本 SKU 有单但当前库存不足，需人工复查"


def test_action_learning_policy_downgrades_legacy_effective_without_support_metrics(tmp_path) -> None:
    from src.autoopt_feedback import build_action_learning_policy

    policy = build_action_learning_policy(
        output_dir=tmp_path,
        current_keyword_reviews=[
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0PRODUCT1",
                "product_name": "Product 1",
                "search_term_or_target": "legacy effective term",
                "action_detail": "加价 5%-10%",
                "outcome": "待7天确认",
                "review_outcome": "effective",
                "effectiveness_score": 2,
                "days_since_execution": 8,
                "review_phase": "day_7_review",
                "review_window": "7天后复盘",
                "promoted_conversion_improved": True,
                "halo_only_conversion": False,
                "target_sku_not_converted": False,
                "current_7d_promoted_ad_orders": 1,
                "current_7d_acos": "8%",
                "current_7d_tacos": "",
                "current_7d_total_orders": "",
                "current_7d_available_stock": "",
                "block_reason": "历史判断有效，继续提高优先级",
            }
        ],
    )

    assert policy["action_learning_summary"]["positive_count"] == 0
    assert policy["action_learning_summary"]["negative_count"] == 1
    assert policy["term_scores"]["UK||SKU-1||B0PRODUCT1||legacy effective term"] == -1
    assert policy["positive_action_patterns"] == []
    assert policy["negative_action_patterns"][0]["score"] == -1


def test_recommendation_guard_blocks_manual_review_keyword_memory() -> None:
    from src.autoopt_feedback import build_recommendation_guard

    guard = build_recommendation_guard(
        product_profiles=[],
        keyword_memory=[
            {
                "action_id": "UK||SKU-1||B0PRODUCT1||legacy effective term||bid_up",
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0PRODUCT1",
                "search_term_or_target": "legacy effective term",
                "normalized_action": "bid_up",
                "review_outcome": "needs_manual_review",
                "effectiveness_score": -1,
                "recommended_future_policy": "recheck_frontend_before_action",
                "should_block_repeating": True,
            }
        ],
    )

    assert guard["recommendation_guard_summary"]["blocked_count"] == 1
    assert guard["recommendation_guard_summary"]["downgraded_count"] == 1
    assert guard["blocked_recommendations"][0]["search_term_or_target"] == "legacy effective term"


def test_recommendation_guard_normalizes_legacy_effective_keyword_memory() -> None:
    from src.autoopt_feedback import build_recommendation_guard

    guard = build_recommendation_guard(
        product_profiles=[],
        keyword_memory=[
            {
                "action_id": "UK||SKU-1||B0PRODUCT1||legacy effective term||bid_up",
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0PRODUCT1",
                "search_term_or_target": "legacy effective term",
                "normalized_action": "bid_up",
                "review_outcome": "effective",
                "effectiveness_score": 2,
                "days_since_execution": 8,
                "promoted_conversion_improved": True,
                "halo_only_conversion": False,
                "target_sku_not_converted": False,
                "current_7d_promoted_ad_orders": 1,
                "current_7d_acos": "8%",
                "current_7d_tacos": "",
                "current_7d_total_orders": "",
                "current_7d_available_stock": "",
                    **anchored_review_contract_fields(),
                    "recommended_future_policy": "keep_current_bid",
                    "should_keep": True,
                "should_block_repeating": False,
            }
        ],
    )

    memory = guard["keyword_memory_updates"][0]

    assert guard["recommendation_guard_summary"]["blocked_count"] == 1
    assert guard["recommendation_guard_summary"]["allowed_with_memory_count"] == 0
    assert guard["recommendation_guard_summary"]["downgraded_count"] == 1
    assert memory["review_outcome"] == "needs_manual_review"
    assert memory["should_keep"] is False
    assert memory["should_block_repeating"] is True


def test_runtime_policy_normalizes_loaded_legacy_keyword_memory(tmp_path) -> None:
    from src.autoopt_feedback import build_runtime_policy

    action_id = "UK||SKU-1||B0PRODUCT1||legacy effective term||bid_up"
    (tmp_path / "keyword_strategy_memory_20260624.json").write_text(
        json.dumps(
            [
                {
                    "action_id": action_id,
                    "marketplace": "UK",
                    "sku": "SKU-1",
                    "asin": "B0PRODUCT1",
                    "search_term_or_target": "legacy effective term",
                    "normalized_action": "bid_up",
                    "review_outcome": "effective",
                    "effectiveness_score": 2,
                    "days_since_execution": 8,
                    "promoted_conversion_improved": True,
                    "halo_only_conversion": False,
                    "target_sku_not_converted": False,
                    "current_7d_promoted_ad_orders": 1,
                    "current_7d_acos": "8%",
                    "current_7d_tacos": "",
                    "current_7d_total_orders": "",
                    "current_7d_available_stock": "",
                    **anchored_review_contract_fields(),
                    "recommended_future_policy": "keep_current_bid",
                    "should_keep": True,
                    "should_block_repeating": False,
                    "block_reason": "历史判断有效，继续提高优先级",
                }
            ],
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    policy = build_runtime_policy(tmp_path)
    memory = policy["keyword_strategy_memory"][0]

    assert memory["action_id"] == action_id
    assert memory["review_outcome"] == "needs_manual_review"
    assert memory["effectiveness_score"] == -1
    assert memory["should_keep"] is False
    assert memory["should_block_repeating"] is True
    assert memory["should_recheck_frontend"] is True
    assert "缺少有效复盘指标" in memory["block_reason"]
    assert policy["keyword_memory_by_action_id"][action_id]["should_block_repeating"] is True


def test_product_strategy_profile_normalizes_legacy_effective_keyword_memory() -> None:
    from src.autoopt_feedback import build_product_strategy_profiles

    profiles = build_product_strategy_profiles(
        [
            {
                "marketplace": "UK",
                "analysis_payload": {
                    "product_window_metrics": {
                        "14d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0PRODUCT1",
                                "product_name": "Product 1",
                                "clicks": 4,
                                "ad_orders": 1,
                                "total_orders": 3,
                            }
                        ]
                    }
                },
            }
        ],
        keyword_memory=[
            {
                "action_id": "UK||SKU-1||B0PRODUCT1||legacy effective term||bid_up",
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0PRODUCT1",
                "search_term_or_target": "legacy effective term",
                "normalized_action": "bid_up",
                "review_outcome": "effective",
                "effectiveness_score": 2,
                "days_since_execution": 8,
                "promoted_conversion_improved": True,
                "halo_only_conversion": False,
                "target_sku_not_converted": False,
                "current_7d_promoted_ad_orders": 1,
                "current_7d_acos": "8%",
                "current_7d_tacos": "",
                "current_7d_total_orders": "",
                "current_7d_available_stock": "",
                "recommended_future_policy": "keep_current_bid",
                "should_keep": True,
                "should_block_repeating": False,
            }
        ],
    )

    profile = profiles[0]

    assert profile["product_stage"] == "CONSERVATIVE_RUN"
    assert profile["ad_strategy_mode"] == "suppress_broad"
    assert "bid_up" in profile["blocked_actions"]
    assert profile["reusable_success_patterns"] == []
    assert any("legacy effective term" in item for item in profile["failed_action_patterns"])


def test_action_learning_policy_parses_string_boolean_attribution_flags(tmp_path) -> None:
    from src.autoopt_feedback import build_action_learning_policy

    policy = build_action_learning_policy(
        output_dir=tmp_path,
        current_keyword_reviews=[
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0PRODUCT1",
                "product_name": "Product 1",
                "search_term_or_target": "string flags term",
                "action_detail": "加价 5%-10%",
                "outcome": "待7天确认",
                "review_outcome": "effective",
                "effectiveness_score": 2,
                "days_since_execution": 8,
                "review_phase": "day_7_review",
                "review_window": "7天后复盘",
                "promoted_conversion_improved": "False",
                "halo_only_conversion": "False",
                "target_sku_not_converted": "False",
                "current_7d_promoted_ad_orders": 1,
                "current_7d_acos": "8%",
                "current_7d_tacos": 0.16,
                "current_7d_total_orders": 3,
                "current_7d_available_stock": 24,
            }
        ],
    )

    item = policy["negative_action_patterns"][0]

    assert item["promoted_conversion_improved"] is False
    assert item["halo_only_conversion"] is False
    assert item["target_sku_not_converted"] is False


def test_attribution_learning_summary_parses_string_boolean_flags() -> None:
    from src.autoopt_feedback import _attribution_learning_summary

    summary = _attribution_learning_summary(
        action_review_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0PRODUCT1",
                "promoted_conversion_improved": "False",
                "halo_only_conversion": "False",
                "target_sku_not_converted": "False",
            },
            {
                "marketplace": "UK",
                "sku": "SKU-2",
                "asin": "B0PRODUCT2",
                "promoted_conversion_improved": "True",
                "halo_only_conversion": "False",
                "target_sku_not_converted": "False",
            },
        ],
        keyword_action_review_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-3",
                "asin": "B0PRODUCT3",
                "search_term_or_target": "halo term",
                "promoted_conversion_improved": "False",
                "halo_only_conversion": "True",
                "target_sku_not_converted": "True",
            }
        ],
        keyword_strategy_memory=[
            {
                "marketplace": "UK",
                "sku": "SKU-4",
                "asin": "B0PRODUCT4",
                "search_term_or_target": "memory term",
                "halo_only_conversion": "False",
            }
        ],
    )

    assert summary["promoted_conversion_improved_count"] == 1
    assert summary["halo_only_conversion_count"] == 1
    assert summary["target_sku_not_converted_count"] == 1
    assert summary["keyword_memory_halo_only_count"] == 0


def test_growth_test_rows_require_intent_and_respect_cooldown() -> None:
    from src.autoopt_feedback import add_action_identity
    from src.report_presentation import _build_growth_test_rows

    source_row = {
        "marketplace": "UK",
        "sku": "SKU-DEMO-LAMP-01",
        "asin": "B0DEMOLMP1",
        "product_name": "Demo desk lamp",
        "search_term_or_target": "adjustable desk lamp",
        "suggested_action": "观察",
        "manual_level": "核心词",
        "classification_reason": "命中 dimmer desk lamp 核心词",
        "match_type": "close-match",
        "matched_target": "close-match",
        "targeting": "close-match",
        "campaign_name": "自动投放测试",
        "ad_group_name": "自动广告组",
        "clicks": "4",
        "orders": "0",
        "spend": "1.20",
    }
    product_card = {
        "marketplace": "UK",
        "sku": "SKU-DEMO-LAMP-01",
        "asin": "B0DEMOLMP1",
        "product_name": "Demo desk lamp",
        "today_allowed_actions": ["observe", "create_exact_low_budget"],
        "today_blocked_actions": ["bid_up", "budget_up", "broad_scale"],
        "target_acos": "19.8%",
        "selling_price": "19.09",
        "recent_7d_clicks": "41",
        "recent_7d_orders": "1",
        "recent_7d_ad_spend": "13.70",
    }

    rows = _build_growth_test_rows([source_row], [], [product_card], [], {}, report_date="2026-06-17")

    assert len(rows) == 1
    assert rows[0]["experiment_type"] == "growth_test"
    assert rows[0]["suggested_action"] == "小预算试投"
    assert rows[0]["normalized_action"] == "growth_test"
    assert rows[0]["action_scope"] == "search_term"
    assert (
        rows[0]["action_id"]
        == "UK||SKU-DEMO-LAMP-01||B0DEMOLMP1||search_term||adjustable desk lamp||growth_test"
    )
    assert rows[0]["term_source"] == "search_term_report"
    assert rows[0]["evidence_level"] in {"核心强相关", "强意图长尾"}
    assert rows[0]["traffic_origin"] == "自动广告"
    assert rows[0]["operation_label"] == "自动出词，拉精准"
    assert rows[0]["campaign_name"] == "自动投放测试"
    assert rows[0]["ad_group_name"] == "自动广告组"
    assert rows[0]["match_type"] == "close-match"
    assert rows[0]["matched_target"] == "close-match"
    assert rows[0]["targeting"] == "close-match"
    assert rows[0]["suggested_daily_budget"].startswith("£")
    assert rows[0]["suggested_bid_min"].startswith("£")
    assert rows[0]["suggested_bid_max"].startswith("£")
    assert rows[0]["test_days"] == "7"
    assert "本 SKU 订单" in rows[0]["success_rule"]
    assert "CPC" in rows[0]["stop_loss_rule"]

    product_gate_blocked_rows = _build_growth_test_rows(
        [source_row],
        [],
        [
            {
                **product_card,
                "today_allowed_actions": ["observe"],
                "today_blocked_actions": ["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
            }
        ],
        [],
        {},
        report_date="2026-06-17",
    )
    assert product_gate_blocked_rows == []

    missing_product_gate_rows = _build_growth_test_rows(
        [source_row],
        [],
        [
            {
                key: value
                for key, value in product_card.items()
                if key not in {"today_allowed_actions", "today_blocked_actions"}
            }
        ],
        [],
        {},
        report_date="2026-06-17",
    )
    assert missing_product_gate_rows == []

    missing_product_card_rows = _build_growth_test_rows(
        [
            {
                **source_row,
                "asin": "B0NOFINALGATE",
            }
        ],
        [],
        [product_card],
        [],
        {},
        report_date="2026-06-17",
    )
    assert missing_product_card_rows == []

    blocked_id = add_action_identity(rows[0], "小预算试投")["action_id"]
    blocked_rows = _build_growth_test_rows(
        [source_row],
        [],
        [product_card],
        [],
        {"blocked_action_ids": [blocked_id], "keyword_strategy_memory": []},
        report_date="2026-06-17",
    )
    assert blocked_rows == []

    low_related_rows = _build_growth_test_rows(
        [
            {
                **source_row,
                "search_term_or_target": "office desk lamp",
                "manual_level": "泛词",
                "classification_reason": "低相关泛词",
            }
        ],
        [],
        [product_card],
        [],
        {},
        report_date="2026-06-17",
    )
    assert low_related_rows == []

    broad_core_rows = _build_growth_test_rows(
        [
            {
                **source_row,
                "search_term_or_target": "desk lamp",
                "classification_reason": "命中 demo_desk_lamp 的核心词词：desk lamp",
                "clicks": "1",
                "orders": "0",
            }
        ],
        [],
        [product_card],
        [],
        {},
        report_date="2026-06-17",
    )
    assert broad_core_rows == []

    historical_only_rows = _build_growth_test_rows(
        [],
        [],
        [product_card],
        [
            {
                **source_row,
                "search_term_or_target": "desk dimmer desk lamp",
                "normalized_action": "create_exact",
                "action_detail": "拉精准小预算",
                "review_outcome": "insufficient_sample",
                "effectiveness_score": 0,
            }
        ],
        {},
        report_date="2026-06-17",
    )
    assert historical_only_rows == []

    history_blocked_rows = _build_growth_test_rows(
        [
            {
                **source_row,
                "search_term_or_target": "led desk lamp with brightness setting",
                "classification_reason": "命中 dimmer_board 的核心词词：led desk lamp",
            }
        ],
        [],
        [product_card],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-DEMO-LAMP-01",
                "asin": "B0DEMOLMP1",
                "search_term_or_target": "led desk lamp with brightness setting",
                "normalized_action": "bid_up",
                "action_detail": "加价5%-10%",
                "review_outcome": "insufficient_sample",
                "effectiveness_score": 0,
            }
        ],
        {},
        report_date="2026-06-17",
    )
    assert history_blocked_rows == []

    active_product_blocked_rows = _build_growth_test_rows(
        [
            {
                **source_row,
                "search_term_or_target": "wooden dimmer desk lamp with dimmer switch",
                "classification_reason": "命中 dimmer_board 的核心词词：dimmer desk lamp",
            }
        ],
        [],
        [product_card],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-DEMO-LAMP-01",
                "asin": "B0DEMOLMP1",
                "search_term_or_target": "adjustable desk lamp",
                "normalized_action": "growth_test",
                "action_detail": "小预算试投",
                "cooldown_status": "cooldown_active",
                "review_phase": "under_3_days",
                "review_outcome": "not_ready",
                "days_since_execution": 0,
            }
        ],
        {},
        report_date="2026-06-17",
    )
    assert active_product_blocked_rows == []

    many_rows = _build_growth_test_rows(
        [
            {**source_row, "search_term_or_target": "adjustable desk lamp"},
            {**source_row, "search_term_or_target": "wooden dimmer desk lamp"},
            {**source_row, "search_term_or_target": "metal dimmer desk lamp"},
        ],
        [],
        [product_card],
        [],
        {},
        report_date="2026-06-17",
    )
    assert [row["search_term_or_target"] for row in many_rows] == ["adjustable desk lamp", "wooden dimmer desk lamp"]

    exact_rows = _build_growth_test_rows(
        [
            {
                **source_row,
                "search_term_or_target": "adjustable desk lamp exact",
                "match_type": "EXACT",
                "matched_target": "adjustable desk lamp exact",
                "targeting": "adjustable desk lamp exact",
                "campaign_name": "手动精准演示台灯",
            }
        ],
        [],
        [product_card],
        [],
        {},
        report_date="2026-06-17",
    )
    assert exact_rows == []

    phrase_rows = _build_growth_test_rows(
        [
            {
                **source_row,
                "search_term_or_target": "wooden dimmer desk lamp with dimmer switch",
                "match_type": "PHRASE",
                "matched_target": "wooden dimmer desk lamp",
                "targeting": "wooden dimmer desk lamp",
                "campaign_name": "手动词组演示台灯",
            }
        ],
        [],
        [product_card],
        [],
        {},
        report_date="2026-06-17",
    )
    assert phrase_rows[0]["operation_label"] == "词组出词，拉精准"

    broad_rows = _build_growth_test_rows(
        [
            {
                **source_row,
                "search_term_or_target": "metal dimmer desk lamp with dimmer switch",
                "match_type": "BROAD",
                "matched_target": "metal dimmer desk lamp",
                "targeting": "metal dimmer desk lamp",
                "campaign_name": "手动广泛演示台灯",
            }
        ],
        [],
        [product_card],
        [],
        {},
        report_date="2026-06-17",
    )
    assert broad_rows[0]["operation_label"] == "广泛出词，拉精准"


def test_build_report_view_blocks_growth_outputs_without_strong_frontend_gate() -> None:
    from src.report_presentation import build_report_view

    payload = {
        "target_marketplace": "UK",
        "report_date": "2026-06-24",
        "ads_date_range": {"start": "2026-06-11", "end": "2026-06-24"},
        "erp_report_coverage_date_range": {"start": "2026-06-11", "end": "2026-06-24"},
        "erp_observed_sales_date_range": {"start": "2026-06-11", "end": "2026-06-24"},
        "common_date_range": {"start": "2026-06-11", "end": "2026-06-24"},
        "history_days": 14,
        "ads_row_count": 20,
        "erp_row_count": 20,
        "data_quality": {"validation_messages": []},
        "data_quality_issue_summary": {},
        "enhanced_data_status": {"provided": False},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-GATED-SCALE",
                    "asin": "B0GATEDSCL",
                    "product_name": "Gated scale product",
                    "clicks": 42,
                    "spend": 12.0,
                    "ad_orders": 4,
                    "ad_sales": 160.0,
                    "total_orders": 5,
                    "ACOS": 0.075,
                    "target_acos": 0.20,
                    "profit_before_ads_per_unit": 10.0,
                }
            ],
            "7d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-GATED-SCALE",
                    "asin": "B0GATEDSCL",
                    "product_name": "Gated scale product",
                    "ad_orders": 2,
                    "total_orders": 3,
                }
            ],
        },
        "操作建议": [
            {
                "level": "可以放量",
                "category": "产品",
                "target": "SKU-GATED-SCALE",
                "action": "表现稳定，可小幅放量",
                "evidence": {
                    "marketplace": "UK",
                    "sku": "SKU-GATED-SCALE",
                    "asin": "B0GATEDSCL",
                    "product_name": "Gated scale product",
                    "clicks": 42,
                    "spend": 12.0,
                    "ad_orders": 4,
                    "ad_sales": 160.0,
                    "total_orders": 5,
                    "ACOS": 0.075,
                    "target_acos": 0.20,
                    "currency": "GBP",
                },
                "markdown_section": "可以放量",
                "markdown_visible": True,
                "priority_rank": 1,
            }
        ],
        "搜索词分析": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-GATED-SCALE",
                    "asin": "B0GATEDSCL",
                    "product_name": "Gated scale product",
                    "search_term": "led desk lamp with brightness setting",
                    "campaign_name": "Auto campaign",
                    "ad_group_name": "Auto group",
                    "match_type": "close-match",
                    "clicks": 8,
                    "spend": 3.0,
                    "ad_orders": 1,
                    "ad_sales": 80.0,
                    "ACOS": 0.0375,
                    "CVR": 0.125,
                    "currency": "GBP",
                }
            ]
        },
    }

    view = build_report_view(payload)

    assert view["scale_rows"] == []
    decision = view["product_final_decision_rows"][0]
    assert "bid_up" in decision["today_blocked_actions"]
    assert "create_exact_low_budget" in decision["today_blocked_actions"]
    assert "bid_up" not in decision["today_allowed_actions"]
    assert "create_exact_low_budget" not in decision["today_allowed_actions"]
    assert len(view["scale_keyword_rows"]) == 1
    scale_keyword = view["scale_keyword_rows"][0]
    assert scale_keyword["scale_action"] == "观察"
    assert scale_keyword["suggested_action"] == "观察"
    assert scale_keyword["ad_gate_blocked"] is True
    assert "提高竞价" in scale_keyword["blocked_original_action"]
    assert view["growth_test_rows"] == []


def test_growth_traffic_origin_mapping() -> None:
    from src.report_presentation import _growth_traffic_origin

    assert _growth_traffic_origin({"match_type": "close-match"}) == "自动广告"
    assert _growth_traffic_origin({"matched_target": "loose-match"}) == "自动广告"
    assert _growth_traffic_origin({"targeting": "close-match"}) == "自动广告"
    assert _growth_traffic_origin({"targeting": "substitutes"}) == "自动广告"
    assert _growth_traffic_origin({"match_type": "exact", "campaign_name": "自动兜底"}) == "手动广告"
    assert _growth_traffic_origin({"match_type": "phrase"}) == "手动广告"
    assert _growth_traffic_origin({"match_type": "broad"}) == "手动广告"
    assert _growth_traffic_origin({"search_term_or_target": "B0DEMOLMP1"}) == "ASIN定向"
    assert _growth_traffic_origin({"targeting": "ASIN定向"}) == "ASIN定向"
    assert _growth_traffic_origin({"campaign_name": "自动-演示台灯"}) == "自动广告"
    assert _growth_traffic_origin({"confirmed_note": "用户反馈：美国演示台灯自动出单词拉入精准小预算。"}) == "自动广告"
    assert _growth_traffic_origin({"confirmed_note": "用户反馈：美国演示台灯泛核心词降竞价。"}) == "手动广告"
    assert _growth_traffic_origin({"normalized_action": "bid_down", "action_detail": "核心词降竞价15%-20%"}) == "手动广告"
    assert _growth_traffic_origin({}) == "未识别"


def test_growth_test_section_is_separate_from_copy_area() -> None:
    from src.generate_html_report import _render_ad_workbench

    normal_row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0TEST1234",
        "product_name": "Demo desk lamp",
        "search_term_or_target": "irrelevant dimmer desk lamp",
        "suggested_action": "降竞价10%-20%",
        "copy_action_line": "建议降竞价 10%-20%",
        "confirmed_status": "待确认",
        "clicks": "8",
        "orders": "0",
        "spend": "£4.20",
    }
    growth_row = {
        "marketplace": "UK",
        "sku": "SKU-DEMO-LAMP-01",
        "asin": "B0DEMOLMP1",
        "product_name": "Demo desk lamp",
        "search_term_or_target": "adjustable desk lamp",
        "suggested_action": "小预算试投",
        "manual_action_taken": "小预算试投",
        "experiment_type": "growth_test",
        "term_source": "search_term_report",
        "evidence_level": "强意图长尾",
        "suggested_daily_budget": "£1.50/天",
        "suggested_bid_min": "£0.30",
        "suggested_bid_max": "£0.38",
        "test_days": "7",
        "stop_loss_rule": "7天点击达到 12 次仍无本 SKU 订单则停。",
        "success_rule": "7天内至少出现本 SKU 订单。",
        "confirmed_status": "待确认",
    }
    second_growth_row = {**growth_row, "search_term_or_target": "led desk lamp with brightness setting"}

    html = _render_ad_workbench(
        [normal_row],
        all_marketplaces=True,
        anchor_id="today-ad-actions-all",
        growth_test_rows=[growth_row, second_growth_row],
        keyword_review_count=45,
    )

    assert "小预算试投" in html
    head_html = html.split('<div class="ad-workbench-head">', 1)[1].split('<div class="ad-summary-grid">', 1)[0]
    assert "待处理 3" in head_html
    assert "广告动作 1" in head_html
    assert "小预算投词 2" in head_html
    assert "历史复盘 45" in head_html
    assert 'data-run-report-action="report-refresh"' in head_html
    assert "<span>小预算投词</span><strong>2</strong>" in html
    assert "<span>历史复盘</span><strong>45</strong>" in html
    assert "完整词级历史复盘见" in html
    assert "UK Demo desk lamp" in html
    assert "adjustable desk lamp" in html
    assert "led desk lamp with brightness setting" in html
    assert "normalized_action&quot;:&quot;growth_test" in html
    growth_section = html.split("growth-test-actions", 1)[1].split("</section>", 1)[0]
    assert growth_section.count('data-ad-complete-checkbox') == 1
    assert growth_section.count("normalized_action&quot;:&quot;growth_test") == 2
    copy_section = html.split("复制到广告后台", 1)[1].split("growth-test-actions", 1)[0]
    assert "adjustable desk lamp" not in copy_section
    assert "irrelevant dimmer desk lamp" in copy_section


def test_autoopt_observe_product_policy_does_not_become_executed_action() -> None:
    from src.autoopt_feedback import _build_action_rows_from_view

    rows = _build_action_rows_from_view(
        {"report_date": "2026-06-10", "source_files": {}},
        {
            "today_task_queue_rows": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-1",
                    "asin": "B0OBSERVE01",
                    "product_name": "Observe product",
                    "priority": "P0",
                    "issue_type": "广告处理",
                    "today_action": "保守跑：不加预算，只保留必要降竞价/否词；观察",
                    "confirmed_status": "已执行",
                }
            ],
            "search_term_processing_queue_rows": [],
        },
        "UK",
    )

    assert rows[0]["normalized_action"] == "observe"
    assert rows[0]["confirmed_status"] == "仅背景参考"


def test_product_level_feedback_does_not_execute_observation_like_today_action(monkeypatch) -> None:
    import src.report_presentation as report_presentation

    monkeypatch.setattr(
        report_presentation,
        "load_feedback_input",
        lambda output_dir: [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0OBSERVE01",
                "confirmed_status": "已执行",
                "confirmed_note": "今天广告后台操作都做了",
            }
        ],
    )

    rows = report_presentation._apply_manual_feedback_to_rows(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0OBSERVE01",
                "today_action": "保守跑：不加预算，不推大词放量，只保留高相关精准词和必要降竞价/否词。",
                "confirmed_status": "待确认",
            }
        ]
    )

    assert rows[0]["confirmed_status"] == "仅背景参考"
    assert rows[0]["confirmed_note"] == "今天广告后台操作都做了"


def test_final_observation_like_today_action_status_is_normalized() -> None:
    from src.report_presentation import _normalize_observation_feedback_statuses

    rows = _normalize_observation_feedback_statuses(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0OBSERVE01",
                "today_action": "保守跑：不加预算，不推大词放量，只保留高相关精准词和必要降竞价/否词。",
                "confirmed_status": "已执行",
            }
        ]
    )

    assert rows[0]["confirmed_status"] == "仅背景参考"


def test_summary_html_uses_three_minute_summary_label(tmp_path) -> None:
    from src.generate_html_report import write_summary_html

    output_path = tmp_path / "summary.html"

    write_summary_html([], output_path, "2026-06-08")

    html = output_path.read_text(encoding="utf-8")
    assert "<title>三分钟摘要｜2026-06-08</title>" in html
    assert "<h1>三分钟摘要｜2026-06-08</h1>" in html


def test_summary_html_surfaces_replenishment_items(tmp_path) -> None:
    from src.generate_html_report import write_summary_html

    output_path = tmp_path / "summary.html"
    results = [
        {
            "has_data": True,
            "marketplace": "UK",
            "report_view": {
                "today_task_queue_rows": [],
                "tomorrow_review_rows": [],
                "html_search_term_processing_queue_rows": [],
                "listing_price_diagnosis_rows": [],
                "yesterday_attribution_rows": [],
                "action_effect_review_rows": [],
                "keyword_action_effect_review_rows": [],
                "inventory_replenishment_rows": [
                    {
                        "marketplace": "UK",
                        "product_name": "演示台灯",
                        "stock_status_label": "低库存",
                        "current_inventory": 4,
                        "days_of_cover": 28,
                        "recommended_reorder_qty": 15,
                        "replenishment_advice": "建议补货约 15 件。",
                    }
                ],
                "analysis_status": "正式分析",
                "issue_summary": "ok",
            },
        }
    ]

    write_summary_html(results, output_path, "2026-06-08")

    html = output_path.read_text(encoding="utf-8")
    assert "补货先看" in html
    assert "UK｜演示台灯" in html
    assert "建议补货约 15 件。" in html
    assert "打开详细报告" not in html
    assert "打开 US" not in html


def test_summary_html_keeps_review_metric_brief_long_enough_for_14d_support(tmp_path) -> None:
    from src.generate_html_report import write_summary_html

    output_path = tmp_path / "summary.html"
    results = [
        {
            "has_data": True,
            "marketplace": "UK",
            "report_view": {
                "today_task_queue_rows": [],
                "tomorrow_review_rows": [],
                "html_search_term_processing_queue_rows": [],
                "listing_price_diagnosis_rows": [],
                "yesterday_attribution_rows": [],
                "inventory_replenishment_rows": [],
                "action_effect_review_rows": [
                    {
                        "marketplace": "UK",
                        "product_name": "Review target product",
                        "review_window": "7天后复盘",
                        "judgement": "待复查",
                        "current_7d_target_acos": 0.003232416623042491,
                        "current_7d_tacos": 0.222,
                        "current_7d_total_orders": 1,
                        "current_7d_available_stock": 37,
                        "current_14d_tacos": 0.331,
                        "current_14d_total_orders": 1,
                        "current_14d_available_stock": 37,
                    }
                ],
                "keyword_action_effect_review_rows": [],
                "analysis_status": "正式分析",
                "issue_summary": "ok",
            },
        }
    ]

    write_summary_html(results, output_path, "2026-06-08")

    html = output_path.read_text(encoding="utf-8")
    assert "目标 ACOS 0.3%" in html
    assert "14天总单 1" in html
    review_section = html.split("<h2>昨天动作今天要盯</h2>", 1)[1].split("<h2>补货先看</h2>", 1)[0]
    assert "..." not in review_section


def test_summary_html_hides_14d_review_metrics_before_7_day_window(tmp_path) -> None:
    from src.generate_html_report import write_summary_html

    output_path = tmp_path / "summary.html"
    results = [
        {
            "has_data": True,
            "marketplace": "UK",
            "report_view": {
                "today_task_queue_rows": [],
                "tomorrow_review_rows": [],
                "html_search_term_processing_queue_rows": [],
                "listing_price_diagnosis_rows": [],
                "yesterday_attribution_rows": [],
                "inventory_replenishment_rows": [],
                "action_effect_review_rows": [],
                "keyword_action_effect_review_rows": [
                    {
                        "marketplace": "UK",
                        "search_term_or_target": "early review keyword",
                        "review_window": "3天后复盘",
                        "days_since_execution": 4,
                        "judgement": "待7天确认",
                        "current_7d_promoted_ad_orders": 1,
                        "current_7d_acos": 0.12,
                        "current_7d_tacos": 0.08,
                        "current_7d_total_orders": 2,
                        "current_7d_available_stock": 16,
                        "current_14d_promoted_ad_orders": 3,
                        "current_14d_acos": 0.09,
                        "current_14d_tacos": 0.07,
                        "current_14d_total_orders": 5,
                        "current_14d_available_stock": 18,
                    }
                ],
                "analysis_status": "正式分析",
                "issue_summary": "ok",
            },
        }
    ]

    write_summary_html(results, output_path, "2026-06-08")

    html = output_path.read_text(encoding="utf-8")
    review_section = html.split("<h2>昨天动作今天要盯</h2>", 1)[1].split("<h2>补货先看</h2>", 1)[0]
    assert "early review keyword" in review_section
    assert "本 SKU 单 1" in review_section
    assert "14天本 SKU 单" not in review_section
    assert "14天 ACOS" not in review_section


def test_summary_html_hides_review_metrics_before_3_day_window(tmp_path) -> None:
    from src.generate_html_report import write_summary_html

    output_path = tmp_path / "summary.html"
    results = [
        {
            "has_data": True,
            "marketplace": "UK",
            "report_view": {
                "today_task_queue_rows": [],
                "tomorrow_review_rows": [],
                "html_search_term_processing_queue_rows": [],
                "listing_price_diagnosis_rows": [],
                "yesterday_attribution_rows": [],
                "inventory_replenishment_rows": [],
                "action_effect_review_rows": [],
                "keyword_action_effect_review_rows": [
                    {
                        "marketplace": "UK",
                        "search_term_or_target": "too early review keyword",
                        "review_window": "未满3天",
                        "days_since_execution": 2,
                        "judgement": "样本不足",
                        "current_7d_promoted_ad_orders": 1,
                        "current_7d_acos": 0.12,
                        "current_7d_tacos": 0.08,
                        "current_7d_total_orders": 2,
                        "current_7d_available_stock": 16,
                    }
                ],
                "analysis_status": "正式分析",
                "issue_summary": "ok",
            },
        }
    ]

    write_summary_html(results, output_path, "2026-06-08")

    html = output_path.read_text(encoding="utf-8")
    review_section = html.split("<h2>昨天动作今天要盯</h2>", 1)[1].split("<h2>补货先看</h2>", 1)[0]
    assert "too early review keyword" in review_section
    assert "未满3天" in review_section
    assert "本 SKU 单 1" not in review_section
    assert "ACOS 12%" not in review_section
    assert "TACOS 8%" not in review_section


def test_report_counters_count_bid_up_and_growth_test_as_pending_ads() -> None:
    from src.generate_html_report import _collect_report_counters

    counters = _collect_report_counters(
        [],
        [
            {
                "confirmed_status": "待确认",
                "copy_action_line": "建议加价 5%-10%",
                "suggested_action": "加价5%-10%",
                "search_term_or_target": "bid up term",
            },
            {
                "confirmed_status": "待确认",
                "suggested_action": "小预算试投",
                "experiment_type": "growth_test",
                "search_term_or_target": "growth term",
            },
            {
                "confirmed_status": "待确认",
                "suggested_action": "观察",
                "search_term_or_target": "watch term",
            },
        ],
        [],
        [],
    )

    assert counters["pending_ad"] == 2


def test_summary_html_downgrades_halo_only_positive_review_judgement(tmp_path) -> None:
    from src.generate_html_report import write_summary_html

    output_path = tmp_path / "summary.html"
    results = [
        {
            "has_data": True,
            "marketplace": "UK",
            "report_view": {
                "today_task_queue_rows": [],
                "tomorrow_review_rows": [],
                "html_search_term_processing_queue_rows": [],
                "listing_price_diagnosis_rows": [],
                "yesterday_attribution_rows": [],
                "inventory_replenishment_rows": [],
                "action_effect_review_rows": [],
                "keyword_action_effect_review_rows": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-1",
                        "asin": "B0HALOONLY",
                        "product_name": "Halo only product",
                        "search_term_or_target": "halo target",
                        "review_window": "7天后复盘",
                        "judgement": "初步有效",
                        "next_step": "继续加价观察",
                        "days_since_execution": "10",
                        "current_7d_clicks": "8",
                        "current_7d_spend": "3.20",
                        "current_7d_ad_orders": "1",
                        "current_7d_promoted_ad_orders": "0",
                        "current_7d_halo_ad_orders": "1",
                        "current_7d_total_orders": "3",
                        "current_7d_tacos": "6%",
                        "current_7d_available_stock": "18",
                        "halo_only_conversion": "True",
                        "target_sku_not_converted": "True",
                    }
                ],
                "analysis_status": "正式分析",
                "issue_summary": "ok",
            },
        }
    ]

    write_summary_html(results, output_path, "2026-06-08")

    html = output_path.read_text(encoding="utf-8")
    review_section = html.split("<h2>昨天动作今天要盯</h2>", 1)[1].split("<h2>补货先看</h2>", 1)[0]
    dont_do_section = html.split("<h2>今天先别做</h2>", 1)[1].split("<h2>昨天动作今天要盯</h2>", 1)[0]
    assert "本 SKU 未验证" in review_section
    assert "本 SKU 未验证" in dont_do_section
    assert "仅光环成交，不算本 SKU 有效；今天不追加预算或竞价。" in review_section
    assert "初步有效" not in review_section
    assert "继续加价观察" not in review_section


def test_summary_html_downgrades_positive_review_without_promoted_conversion(tmp_path) -> None:
    from src.generate_html_report import write_summary_html

    output_path = tmp_path / "summary.html"
    results = [
        {
            "has_data": True,
            "marketplace": "UK",
            "report_view": {
                "today_task_queue_rows": [],
                "tomorrow_review_rows": [],
                "html_search_term_processing_queue_rows": [],
                "listing_price_diagnosis_rows": [],
                "yesterday_attribution_rows": [],
                "inventory_replenishment_rows": [],
                "action_effect_review_rows": [],
                "keyword_action_effect_review_rows": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-1",
                        "asin": "B0NOPROMO",
                        "product_name": "No promoted product",
                        "search_term_or_target": "no promoted target",
                        "review_window": "7天后复盘",
                        "judgement": "初步有效",
                        "next_step": "继续加价观察",
                        "days_since_execution": "10",
                        "current_7d_clicks": "9",
                        "current_7d_spend": "4.10",
                        "current_7d_ad_orders": "1",
                        "current_7d_promoted_ad_orders": "0",
                        "current_7d_halo_ad_orders": "0",
                        "current_7d_total_orders": "1",
                        "promoted_conversion_improved": "False",
                        "halo_only_conversion": "False",
                        "target_sku_not_converted": "False",
                    }
                ],
                "analysis_status": "正式分析",
                "issue_summary": "ok",
            },
        }
    ]

    write_summary_html(results, output_path, "2026-06-08")

    html = output_path.read_text(encoding="utf-8")
    review_section = html.split("<h2>昨天动作今天要盯</h2>", 1)[1].split("<h2>补货先看</h2>", 1)[0]
    assert "本 SKU 未验证" in review_section
    assert "缺少本 SKU 转化证据；今天不追加预算或竞价。" in review_section
    assert "初步有效" not in review_section
    assert "继续加价观察" not in review_section


def test_summary_html_keeps_background_rows_out_of_today_do(tmp_path) -> None:
    from src.generate_html_report import write_summary_html

    output_path = tmp_path / "summary.html"
    results = [
        {
            "has_data": True,
            "marketplace": "US",
            "report_view": {
                "today_task_queue_rows": [
                    {
                        "marketplace": "US",
                        "product_name": "背景观察品",
                        "priority": "P0",
                        "today_action": "保守跑：不加预算，只观察。",
                        "primary_reason": "样本不足",
                        "confirmed_status": "仅背景参考",
                    },
                    {
                        "marketplace": "US",
                        "product_name": "成本阻断品",
                        "priority": "P0",
                        "action_group": "成本 / 利润动作",
                        "today_action": "保守跑：不加预算。",
                        "primary_reason": "利润不允许加广告",
                        "confirmed_status": "仅背景参考",
                        "fusion_action_gate": "cost_blocked",
                    },
                ],
                "tomorrow_review_rows": [],
                "html_search_term_processing_queue_rows": [],
                "listing_price_diagnosis_rows": [],
                "yesterday_attribution_rows": [],
                "action_effect_review_rows": [],
                "keyword_action_effect_review_rows": [],
                "inventory_replenishment_rows": [],
                "analysis_status": "正式分析",
                "issue_summary": "ok",
            },
        }
    ]

    write_summary_html(results, output_path, "2026-06-08")

    html = output_path.read_text(encoding="utf-8")
    today_do = html.split("<h2>今天先做</h2>", 1)[1].split("<h2>今天先别做</h2>", 1)[0]
    dont_do = html.split("<h2>今天先别做</h2>", 1)[1].split("<h2>昨天动作今天要盯</h2>", 1)[0]
    assert "背景观察品" not in today_do
    assert "成本阻断品" not in today_do
    assert "背景观察品" in dont_do
    assert "成本阻断品" in dont_do
    assert "仅背景参考" in dont_do


def test_summary_html_conclusion_uses_risk_sources_instead_of_generic_copy(tmp_path) -> None:
    from src.generate_html_report import write_summary_html

    output_path = tmp_path / "summary.html"
    results = [
        {
            "has_data": True,
            "marketplace": "UK",
            "report_view": {
                "today_task_queue_rows": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-3",
                        "asin": "B0SUM12345",
                        "product_name": "结论样本品",
                        "priority": "P1",
                        "today_action": "先补前台和库存证据。",
                        "primary_reason": "广告有花费但转化弱",
                        "confirmed_status": "待确认",
                    }
                ],
                "product_operation_cards": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-3",
                        "asin": "B0SUM12345",
                        "product_name": "结论样本品",
                        "final_decision": "CONSERVATIVE_RUN",
                        "decision_reason": "广告有花费；总单 2，自然单 2；利润<=0；库存：低库存；前台：沿用缓存；今天只做止损。",
                        "profit_before_ads_per_unit": -1,
                        "inventory_constraint": "LOW_STOCK",
                        "frontend_status": "沿用 2026-06-16 前台数据",
                        "frontend_auto_conclusion_label": "自动证据不足，不能用于强诊断",
                        "total_orders": 2,
                        "natural_orders": 2,
                        "ad_orders": 0,
                    }
                ],
                "tomorrow_review_rows": [],
                "html_search_term_processing_queue_rows": [],
                "listing_price_diagnosis_rows": [],
                "yesterday_attribution_rows": [],
                "action_effect_review_rows": [],
                "keyword_action_effect_review_rows": [],
                "inventory_replenishment_rows": [],
                "analysis_status": "正式分析",
                "issue_summary": "ok",
            },
        }
    ]

    write_summary_html(results, output_path, "2026-06-08")

    html = output_path.read_text(encoding="utf-8")
    assert "今天没有新的 P0，先查 UK｜结论样本品。" in html
    assert "利润不支持放量" in html
    assert "低库存" in html
    assert "前台证据不足" in html


def test_dashboard_html_exposes_showcase_entry_points(tmp_path) -> None:
    from src.generate_html_report import write_dashboard_html

    output_path = tmp_path / "dashboard.html"
    results = [
        {
            "marketplace": "US",
            "has_data": True,
            "summary": {
                "marketplace": "US",
                "ads_row_count": 1,
                "erp_row_count": 1,
                "sku_count": 1,
                "asin_count": 1,
            },
            "analysis_payload": {"产品汇总": {"1d": [{"currency": "USD"}]}},
            "report_view": {
                "today_task_queue_rows": [],
                "html_search_term_processing_queue_rows": [],
                "tomorrow_review_rows": [],
                "listing_price_diagnosis_rows": [],
                "frontend_check_queue_rows": [],
                "frontend_coverage_summary": {
                    "frontend_queue_total": 2,
                    "frontend_decision_ready_count": 1,
                    "frontend_reference_evidence_count": 1,
                    "frontend_live_success_count": 1,
                    "frontend_search_success_count": 1,
                    "frontend_search_partial_count": 0,
                },
                "analysis_status": "正式分析",
                "issue_summary": "样例数据",
            },
        }
    ]

    write_dashboard_html(results, output_path, "2026-06-08")

    html = output_path.read_text(encoding="utf-8")
    assert "运营状态入口" in html
    assert "打开三分钟摘要" in html
    assert "打开 ALL 运营控制台" in html
    assert "前台证据覆盖" in html
    assert "产品页成功" in html
    assert "Amazon 搜索页辅助验证" in html
    assert "竞品 ASIN 反查" in html
    hero = html.split("运营状态入口", 1)[0]
    assert "请把文件放入对应站点文件夹" not in hero
    assert "打开摘要" not in hero
    assert "打开 ALL 汇总建议" not in hero
    assert "打开广告工作台" not in html
    assert "<strong>广告行数</strong>" not in html
    assert "<strong>ERP行数</strong>" not in html


def test_marketplace_report_hides_empty_review_sections(tmp_path) -> None:
    from src.generate_html_report import write_marketplace_report_html

    output_path = tmp_path / "uk_report.html"
    result = {
        "marketplace": "UK",
        "has_data": True,
        "analysis_payload": {"产品汇总": {"1d": [{"currency": "GBP"}]}},
        "summary": {"ads_row_count": 1, "erp_row_count": 1, "sku_count": 1, "asin_count": 1},
        "report_view": {
            "quality_pass": True,
            "analysis_status": "正式分析",
            "issue_summary": "ok",
            "today_task_queue_rows": [],
            "tomorrow_review_rows": [],
            "html_search_term_processing_queue_rows": [],
            "scale_keyword_rows": [],
            "hidden_low_click_search_terms": 0,
            "frontend_check_queue_rows": [],
            "listing_price_diagnosis_rows": [],
            "yesterday_attribution_rows": [],
            "keyword_action_effect_review_rows": [],
            "action_effect_review_rows": [],
            "inventory_replenishment_rows": [],
            "cost_profit_diagnosis_rows": [],
            "data_quality_rows": [],
            "enhanced_status_rows": [],
            "enhanced_data_request_rows": [],
            "low_priority_watch_rows": [],
            "stale_data_notes": [],
            "runtime_policy_notes": [],
        },
    }

    write_marketplace_report_html(result, output_path, "2026-06-08")

    html = output_path.read_text(encoding="utf-8")
    assert "广告状态" in html
    assert "待确认 0" in html
    assert "section-card is-collapsed" in html[html.index("广告状态") - 120 : html.index("广告状态")]
    assert "已执行但仍需复查" not in html
    assert "明日复查清单" not in html
    assert "复制到广告后台" not in html


def test_marketplace_report_separates_executable_p0_from_observation(tmp_path) -> None:
    from src.generate_html_report import write_marketplace_report_html

    output_path = tmp_path / "uk_report.html"
    result = {
        "marketplace": "UK",
        "has_data": True,
        "analysis_payload": {"产品汇总": {"1d": [{"currency": "GBP"}]}},
        "summary": {"ads_row_count": 1, "erp_row_count": 1, "sku_count": 1, "asin_count": 1},
        "report_view": {
            "quality_pass": True,
            "analysis_status": "正式分析",
            "issue_summary": "ok",
            "today_task_queue_rows": [
                {
                    "marketplace": "UK",
                    "product_name": "观察产品",
                    "sku": "SKU-OBS",
                    "asin": "B0OBSERVE1",
                    "priority": "P0",
                    "action_group": "Listing / 价格动作",
                    "today_action": "只观察或补足自动证据，不输出强运营动作。",
                    "primary_reason": "自动证据不足",
                    "confirmed_status": "仅背景参考",
                    "fusion_action_gate": "collect_evidence_only",
                    "search_term_top5": "N/A",
                }
            ],
            "tomorrow_review_rows": [],
            "html_search_term_processing_queue_rows": [],
            "scale_keyword_rows": [],
            "hidden_low_click_search_terms": 0,
            "frontend_check_queue_rows": [
                {
                    "marketplace": "UK",
                    "product_name": "观察产品",
                    "sku": "SKU-OBS",
                    "asin": "B0OBSERVE1",
                    "frontend_check_status": "待前台检查",
                }
            ],
            "listing_price_diagnosis_rows": [],
            "yesterday_attribution_rows": [],
            "keyword_action_effect_review_rows": [],
            "action_effect_review_rows": [],
            "inventory_replenishment_rows": [],
            "cost_profit_diagnosis_rows": [],
            "data_quality_rows": [],
            "enhanced_status_rows": [],
            "enhanced_data_request_rows": [],
            "low_priority_watch_rows": [],
            "stale_data_notes": [],
            "runtime_policy_notes": [],
        },
    }

    write_marketplace_report_html(result, output_path, "2026-06-08")

    html = output_path.read_text(encoding="utf-8")
    assert "P0 今日必须执行" not in html
    observe = html.split("<h2>P0 待核查 / 观察</h2>", 1)[1].split("前台证据", 1)[0]
    assert "观察产品" in observe
    assert "section-card is-collapsed" in html[html.index("P0 待核查 / 观察") - 140 : html.index("P0 待核查 / 观察")]
    assert "需要处理的搜索词 / ASIN\"" not in html
    frontend_section = html.split('id="frontend-evidence-status"', 1)[1].split("</section>", 1)[0]
    assert '<div class="priority-grid">' not in frontend_section
    assert frontend_section.index('<div class="collapsible-body">') < frontend_section.index("frontend-retry")


def test_marketplace_report_collapses_review_and_watch_sections(tmp_path) -> None:
    from src.generate_html_report import write_marketplace_report_html

    output_path = tmp_path / "uk_report.html"
    result = {
        "marketplace": "UK",
        "has_data": True,
        "analysis_payload": {"产品汇总": {"1d": [{"currency": "GBP"}]}},
        "summary": {"ads_row_count": 1, "erp_row_count": 1, "sku_count": 1, "asin_count": 1},
        "report_view": {
            "quality_pass": True,
            "analysis_status": "正式分析",
            "issue_summary": "ok",
            "today_task_queue_rows": [],
            "tomorrow_review_rows": [
                {
                    "marketplace": "UK",
                    "product_name": "待复查产品",
                    "today_action": "明日复查广告花费。",
                }
            ],
            "html_search_term_processing_queue_rows": [],
            "scale_keyword_rows": [],
            "hidden_low_click_search_terms": 0,
            "frontend_check_queue_rows": [],
            "listing_price_diagnosis_rows": [],
            "yesterday_attribution_rows": [],
            "keyword_action_effect_review_rows": [],
            "action_effect_review_rows": [],
            "inventory_replenishment_rows": [],
            "cost_profit_diagnosis_rows": [],
            "data_quality_rows": [],
            "enhanced_status_rows": [],
            "enhanced_data_request_rows": [],
            "low_priority_watch_rows": [],
            "stale_data_notes": [],
            "runtime_policy_notes": [],
            "risk_rows": [
                {
                    "marketplace": "UK",
                    "product_name": "滞销产品",
                    "sku": "SKU-WATCH",
                    "asin": "B0WATCH123",
                    "risk_level": "P2",
                    "risk_type": "滞销观察",
                    "reason": "低销量观察",
                    "suggested_action": "观察",
                }
            ],
        },
    }

    write_marketplace_report_html(result, output_path, "2026-06-08")

    html = output_path.read_text(encoding="utf-8")
    assert "明日复查清单" in html
    assert "滞销观察池" in html
    assert html.count('section-card is-collapsed') >= 2
    assert "复查对象默认收起" in html
    assert "观察池默认收起" in html


def test_marketplace_ad_workbench_collapses_when_no_pending_actions() -> None:
    from src.generate_html_report import _render_ad_workbench_status_only

    html = _render_ad_workbench_status_only(
        [
            {
                "marketplace": "UK",
                "search_term_or_target": "observe only",
                "suggested_action": "观察",
                "confirmed_status": "仅背景参考",
            }
        ],
        hidden_low_click_count=2,
        marketplace_hint="UK",
    )

    assert "待确认 0" in html
    assert "UK 今天没有新的广告后台执行动作" in html
    assert "观察项不复制到后台" in html
    assert "待执行 0" in html
    assert "观察项 1" in html
    assert "低点击隐藏项 2" in html
    assert "打开 ALL 广告工作台" not in html
    assert "可以展开核对广告线索" not in html
    assert "复制到广告后台" not in html
    assert "ad-toolbar" not in html
    html_with_anchor = _render_ad_workbench_status_only([], marketplace_hint="ALL", section_id="today-ad-actions-all")
    assert 'id="today-ad-actions-all"' in html_with_anchor
    assert "查看广告区" in html_with_anchor
    assert "已执行留档 0" not in html_with_anchor
    assert "观察项 0" not in html_with_anchor


def test_all_ad_status_can_expand_observation_rows_without_copy_area() -> None:
    from src.generate_html_report import _render_ad_workbench_status_only

    html = _render_ad_workbench_status_only(
        [
            {
                "marketplace": "UK",
                "product_name": "六格演示线夹",
                "search_term_or_target": "demo notebook",
                "suggested_action": "观察",
                "confirmed_status": "待确认",
                "clicks": "4",
                "spend": "£0.94",
                "orders": "0",
            },
            {
                "marketplace": "US",
                "product_name": "演示笔记本",
                "search_term_or_target": "metal desk lamps for kitchen",
                "suggested_action": "降竞价10%-20%",
                "confirmed_status": "已执行",
                "clicks": "6",
                "spend": "$4.02",
                "orders": "0",
            },
        ],
        marketplace_hint="ALL",
        section_id="today-ad-actions-all",
        show_details=True,
    )

    assert "待确认 0" in html
    assert "查看观察项和已执行留档" in html
    assert "观察项，不操作" in html
    assert "demo notebook" in html
    assert "已执行留档" in html
    assert "metal desk lamps for kitchen" in html
    assert "复制到广告后台" not in html


def test_all_ad_status_without_rows_avoids_zero_noise() -> None:
    from src.generate_html_report import _render_ad_workbench_status_only

    html = _render_ad_workbench_status_only(
        [],
        marketplace_hint="ALL",
        section_id="today-ad-actions-all",
        show_details=True,
    )

    assert "待执行 0" in html
    assert "已执行留档 0" not in html
    assert "观察项 0" not in html
    assert "查看观察项和已执行留档" not in html
    assert "可以展开核对广告线索" not in html


def test_all_zero_pending_ad_details_can_be_collapsed() -> None:
    from src.generate_html_report import _render_ad_workbench_status_only

    html = _render_ad_workbench_status_only(
        [
            {
                "marketplace": "US",
                "product_name": "演示台灯",
                "search_term_or_target": "led desk lamp",
                "suggested_action": "观察",
                "confirmed_status": "待确认",
            }
        ],
        marketplace_hint="ALL",
        section_id="today-ad-actions-all",
        show_details=True,
        collapsed=True,
    )

    assert 'id="today-ad-actions-all"' in html
    assert "section-card is-collapsed" in html
    assert "collapsible-body" in html
    assert "查看观察项和已执行留档" in html


def test_task_cards_explain_waiting_for_evidence_as_condition() -> None:
    from src.generate_html_report import _render_task_cards

    html = _render_task_cards(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "product_name": "Test product",
                "priority": "P0",
                "today_action": "只观察或补足自动证据，不输出强运营动作。",
                "fusion_review_window": "补齐 7 天窗口或刷新前台后再判断。",
            }
        ],
        "P0",
    )

    assert "补证后复查" in html
    assert "满 7 天或前台证据可用后再定动作" in html
    assert "先不做强操作" not in html
    assert "补齐 7 天窗口或刷新前台后再判断" not in html


def test_action_effect_review_renders_structured_metrics_and_attribution() -> None:
    from src.generate_html_report import _render_action_effect_review_rows

    html = _render_action_effect_review_rows(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0PRODUCT",
                "product_name": "演示笔记本",
                "executed_action": "降竞价",
                "executed_at": "2026-05-20",
                "review_window": "7天后复盘",
                "judgement": "初步有效",
                "next_step": "保留当前竞价。",
                "current_7d_clicks": "8",
                "current_7d_spend": "3.20",
                "current_7d_ad_orders": "2",
                "current_7d_promoted_ad_orders": "2",
                "current_7d_promoted_ad_sales": "30.00",
                "current_7d_halo_ad_orders": "0",
                "current_7d_total_orders": "4",
                "current_7d_acos": "0.12",
                "current_7d_target_acos": "0.2",
                "current_7d_tacos": "0.09",
                "current_7d_available_stock": "18",
                "current_14d_clicks": "14",
                "current_14d_spend": "5.10",
                "current_14d_ad_orders": "3",
                "current_14d_promoted_ad_orders": "3",
                "current_14d_halo_ad_orders": "1",
                "current_14d_total_orders": "7",
                "current_14d_acos": "0.14",
                "current_14d_tacos": "0.11",
                "current_14d_available_stock": "18",
                "promoted_conversion_improved": "True",
                "halo_only_conversion": "False",
                "target_sku_not_converted": "False",
                "attribution_effect_note": "本 SKU 成交优先",
            }
        ]
    )

    assert "归因判断：本 SKU 成交有效：本 SKU 成交优先" in html
    assert "7天 点击 8" in html
    assert "本 SKU 单 2" in html
    assert "光环单 0" in html
    assert "总单 4" in html
    assert "7天 目标 ACOS 0.2" in html
    assert "TACOS 0.09" in html
    assert "库存 18" in html
    assert "14天 点击 14" in html
    assert "复盘结论</strong><p>可保留：保留当前动作，不重复加价或追加预算</p>" in html
    assert "触发标准</strong><p>触发标准：按复盘窗口、本 SKU 单、光环单、点击和花费共同判断。</p>" in html


def test_action_effect_review_marks_three_day_window_as_early_judgement() -> None:
    from src.generate_html_report import _render_action_effect_review_rows

    html = _render_action_effect_review_rows(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0EARLY",
                "product_name": "早期复盘产品",
                "executed_action": "降竞价",
                "executed_at": "2026-06-01",
                "days_since_execution": "4",
                "review_window": "3d_check",
                "judgement": "待7天确认",
                "next_step": "继续观察到 7 天窗口。",
                "current_7d_clicks": "4",
                "current_7d_spend": "2.10",
                "current_7d_ad_orders": "0",
                "current_7d_promoted_ad_orders": "0",
                "current_7d_total_orders": "0",
                "current_7d_acos": "0",
                "current_7d_target_acos": "0.2",
                "current_7d_tacos": "0",
                "current_7d_available_stock": "18",
            }
        ]
    )

    assert "复盘口径：3d_check；3 天窗口只做初步判断，低点击或低花费仍按样本不足处理" in html


def test_keyword_action_effect_review_uses_early_trigger_standard_before_seven_days() -> None:
    from src.generate_html_report import _render_keyword_action_effect_review_rows

    html = _render_keyword_action_effect_review_rows(
        [
            {
                "marketplace": "US",
                "sku": "SKU-1",
                "asin": "B0EARLYBIDUP",
                "product_name": "早期加竞价产品",
                "search_term_or_target": "desk lamp clip",
                "normalized_action": "bid_up",
                "action_scope": "search_term",
                "executed_at": "2026-06-01",
                "days_since_execution": "4",
                "review_window": "3d_check",
                "judgement": "待7天确认",
                "next_step": "继续观察到 7 天窗口。",
                "confirmed_note": "用户反馈：已执行 desk lamp clip 加竞价",
                "report_date": "2026-06-01",
                "review_date": "2026-06-05",
                "current_7d_clicks": "8",
                "current_7d_spend": "6.20",
                "current_7d_ad_orders": "1",
                "current_7d_ad_sales": "19.99",
                "current_7d_promoted_ad_orders": "1",
                "current_7d_promoted_ad_sales": "19.99",
                "current_7d_halo_ad_orders": "0",
                "current_7d_acos": "0.12",
                "current_7d_target_acos": "0.2",
                "current_7d_total_orders": "2",
                "current_7d_tacos": "0.08",
                "current_7d_available_stock": "18",
            }
        ],
        limit=1,
    )

    assert "触发标准：3 天初查，只看是否继续消耗、本 SKU 是否已有初步订单，不给有效结论，不追加预算或竞价。" in html
    assert "可保留当前竞价" not in html
    assert "7天 本 SKU 单 1" in html
    assert "14天 本 SKU 单" not in html


def test_keyword_action_effect_review_hides_metrics_before_three_days() -> None:
    from src.generate_html_report import _render_keyword_action_effect_review_rows

    html = _render_keyword_action_effect_review_rows(
        [
            {
                "marketplace": "US",
                "sku": "SKU-1",
                "asin": "B0TOOEARLY",
                "product_name": "未满三天产品",
                "search_term_or_target": "too early keyword",
                "normalized_action": "bid_up",
                "action_scope": "search_term",
                "executed_at": "2026-06-01",
                "days_since_execution": "2",
                "review_window": "未满3天",
                "judgement": "样本不足",
                "next_step": "等满 3 天再看。",
                "current_7d_clicks": "8",
                "current_7d_spend": "6.20",
                "current_7d_ad_orders": "1",
                "current_7d_promoted_ad_orders": "1",
                "current_7d_promoted_ad_sales": "19.99",
                "current_7d_halo_ad_orders": "0",
                "current_7d_acos": "0.12",
                "current_7d_target_acos": "0.2",
                "current_7d_total_orders": "2",
                "current_7d_tacos": "0.08",
                "current_7d_available_stock": "18",
            }
        ],
        limit=1,
    )

    assert "未满3天" in html
    assert "触发标准：执行未满 3 天，禁止判定有效或失败。" in html
    assert "7天 本 SKU 单 1" not in html
    assert "7天 ACOS 0.12" not in html
    assert "7天 TACOS 0.08" not in html


def test_action_effect_review_warns_halo_only_is_not_promoted_sku_effective() -> None:
    from src.generate_html_report import _render_action_effect_review_rows

    html = _render_action_effect_review_rows(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-2",
                "asin": "B0HALOONLY",
                "product_name": "演示线夹",
                "executed_action": "加价 5%-10%",
                "executed_at": "2026-05-21",
                "review_window": "7天后复盘",
                "judgement": "样本不足",
                "next_step": "不追加预算。",
                "current_7d_clicks": "3",
                "current_7d_spend": "1.20",
                "current_7d_ad_orders": "1",
                "current_7d_promoted_ad_orders": "0",
                "current_7d_promoted_ad_sales": "0",
                "current_7d_halo_ad_orders": "1",
                "current_7d_total_orders": "1",
                "current_7d_acos": "0.20",
                "current_7d_tacos": "0.08",
                "current_7d_available_stock": "6",
                "current_14d_clicks": "5",
                "current_14d_spend": "2.00",
                "current_14d_ad_orders": "1",
                "current_14d_promoted_ad_orders": "0",
                "current_14d_halo_ad_orders": "1",
                "current_14d_total_orders": "2",
                "current_14d_acos": "0.20",
                "current_14d_tacos": "0.10",
                "current_14d_available_stock": "6",
                "promoted_conversion_improved": "False",
                "halo_only_conversion": "True",
                "target_sku_not_converted": "True",
            }
        ]
    )

    assert "复盘结论</strong><p>停止追加：有光环订单也不能证明本 SKU 有效，今天不追加预算或竞价</p>" in html
    assert "触发标准</strong><p>触发标准：光环单 &gt; 0 且本 SKU 单 = 0，不能作为本 SKU 有效证据。</p>" in html
    assert "归因判断：仅光环成交，不算本 SKU 有效" in html
    assert "本 SKU 单 0" in html
    assert "光环单 1" in html


def test_keyword_action_review_downgrades_halo_only_chinese_truthy_flags() -> None:
    from src.generate_html_report import _render_keyword_action_effect_review_rows

    html = _render_keyword_action_effect_review_rows(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-HALO-CN",
                "asin": "B0HALOCN01",
                "product_name": "中文标记光环单产品",
                "search_term_or_target": "halo chinese flag term",
                "executed_action": "加价 5%-10%",
                "executed_at": "2026-05-21",
                "days_since_execution": "9",
                "review_window": "7天后复盘",
                "judgement": "初步有效",
                "next_step": "保留当前竞价。",
                "current_7d_clicks": "8",
                "current_7d_spend": "6.20",
                "current_7d_ad_orders": "1",
                "current_7d_promoted_ad_orders": "0",
                "current_7d_halo_ad_orders": "1",
                "current_7d_total_orders": "1",
                "current_7d_acos": "0.20",
                "current_7d_tacos": "0.08",
                "current_7d_available_stock": "6",
                "promoted_conversion_improved": "否",
                "halo_only_conversion": "是",
                "target_sku_not_converted": "已验证",
            }
        ],
        limit=1,
    )

    assert "本 SKU 未验证" in html
    assert "停止追加：缺少本 SKU 转化证据，今天不追加预算或竞价" in html
    assert "归因判断：仅光环成交，不算本 SKU 有效" in html
    assert "可保留：保留当前动作" not in html


def test_keyword_action_review_halo_only_overrides_conflicting_promoted_flag() -> None:
    from src.generate_html_report import _render_keyword_action_effect_review_rows

    html = _render_keyword_action_effect_review_rows(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-HALO-CONFLICT",
                "asin": "B0HALOCONF",
                "product_name": "冲突归因产品",
                "search_term_or_target": "halo conflict term",
                "executed_action": "加价 5%-10%",
                "executed_at": "2026-05-21",
                "days_since_execution": "9",
                "review_window": "7天后复盘",
                "judgement": "初步有效",
                "next_step": "保留当前竞价。",
                "current_7d_clicks": "8",
                "current_7d_spend": "6.20",
                "current_7d_ad_orders": "1",
                "current_7d_promoted_ad_orders": "0",
                "current_7d_halo_ad_orders": "1",
                "current_7d_total_orders": "1",
                "current_7d_acos": "0.20",
                "current_7d_tacos": "0.08",
                "current_7d_available_stock": "6",
                "promoted_conversion_improved": "True",
                "halo_only_conversion": "True",
                "target_sku_not_converted": "True",
                "attribution_effect_note": "冲突字段，光环单优先",
            }
        ],
        limit=1,
    )

    assert "本 SKU 未验证" in html
    assert "停止追加：缺少本 SKU 转化证据，今天不追加预算或竞价" in html
    assert "归因判断：仅光环成交，不算本 SKU 有效：冲突字段，光环单优先" in html
    assert "归因判断：本 SKU 成交有效" not in html
    assert "可保留：保留当前动作" not in html


def test_action_effect_review_shows_missing_manual_review_metrics() -> None:
    from src.generate_html_report import _render_action_effect_review_rows

    html = _render_action_effect_review_rows(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-MISS",
                "asin": "B0MISSMETRIC",
                "product_name": "缺指标产品",
                "executed_action": "加价 5%-10%",
                "executed_at": "2026-05-20",
                "review_window": "7天后复盘",
                "review_outcome": "needs_manual_review",
                "judgement": "待人工复查",
                "days_since_execution": "12",
                "current_7d_clicks": "8",
                "current_7d_spend": "3.20",
                "current_7d_ad_orders": "1",
                "current_7d_promoted_ad_orders": "1",
                "current_7d_acos": "0.08",
                "current_7d_tacos": "",
                "current_7d_total_orders": "3",
                "current_7d_available_stock": "",
                "current_14d_promoted_ad_orders": "2",
                "current_14d_acos": "0.09",
                "current_14d_tacos": "",
                "current_14d_total_orders": "5",
                "current_14d_available_stock": "",
                "promoted_conversion_improved": "True",
                "halo_only_conversion": "False",
                "block_reason": "缺少有效复盘指标：current_7d_tacos, current_7d_available_stock",
            }
        ]
    )

    assert "缺失复盘指标：7天 TACOS、7天库存、14天 TACOS、14天库存" in html
    assert "缺少有效复盘指标" in html


def test_action_effect_review_treats_nan_metrics_as_missing() -> None:
    from src.generate_html_report import _render_action_effect_review_rows

    html = _render_action_effect_review_rows(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-NAN",
                "asin": "B0NANMETRIC",
                "product_name": "NaN 指标产品",
                "executed_action": "加价 5%-10%",
                "executed_at": "2026-05-20",
                "review_window": "7天后复盘",
                "review_outcome": "needs_manual_review",
                "judgement": "待人工复查",
                "days_since_execution": "12",
                "current_7d_clicks": "nan",
                "current_7d_spend": "NaN",
                "current_7d_ad_orders": "nan",
                "current_7d_promoted_ad_orders": "1",
                "current_7d_acos": "nan",
                "current_7d_tacos": "NaN",
                "current_7d_total_orders": "3",
                "current_7d_available_stock": "nan",
                "current_14d_promoted_ad_orders": "2",
                "current_14d_acos": "0.09",
                "current_14d_tacos": "nan",
                "current_14d_total_orders": "5",
                "current_14d_available_stock": "18",
            }
        ]
    )

    assert "缺失复盘指标：7天 ACOS、7天 TACOS、7天库存、14天 TACOS" in html
    assert "7天 点击 N/A" in html
    assert "花费 N/A" in html
    assert "ACOS N/A" in html
    assert "TACOS N/A" in html
    assert "库存 N/A" in html
    assert "点击 nan" not in html.lower()
    assert "花费 nan" not in html.lower()
    assert "库存 nan" not in html.lower()


def test_action_effect_review_downgrades_positive_judgement_without_promoted_sku_evidence() -> None:
    from src.generate_html_report import _render_action_effect_review_rows

    html = _render_action_effect_review_rows(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-2",
                "asin": "B0HALOPOS",
                "product_name": "演示线夹",
                "executed_action": "加价 5%-10%",
                "executed_at": "2026-05-21",
                "review_window": "7天后复盘",
                "judgement": "有改善迹象",
                "current_7d_clicks": "5",
                "current_7d_spend": "2.40",
                "current_7d_ad_orders": "1",
                "current_7d_promoted_ad_orders": "0",
                "current_7d_promoted_ad_sales": "0",
                "current_7d_halo_ad_orders": "1",
                "current_7d_total_orders": "1",
                "current_7d_acos": "0.20",
                "current_7d_tacos": "0.08",
                "current_7d_available_stock": "6",
                "promoted_conversion_improved": "False",
                "halo_only_conversion": "True",
                "target_sku_not_converted": "True",
            }
        ]
    )

    assert '<span class="tag tag-yellow">本 SKU 未验证</span>' in html
    assert "复盘结论</strong><p>停止追加：有光环订单也不能证明本 SKU 有效，今天不追加预算或竞价</p>" in html
    assert '<span class="tag tag-green">有改善迹象</span>' not in html


def test_action_effect_review_prioritizes_halo_only_not_validated_when_limited() -> None:
    from src.generate_html_report import _render_action_effect_review_rows

    html = _render_action_effect_review_rows(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-SAMPLE",
                "asin": "B0SAMPLE",
                "product_name": "普通样本不足产品",
                "executed_action": "观察",
                "review_window": "7天后复盘",
                "judgement": "样本不足",
                "days_since_execution": "12",
            },
            {
                "marketplace": "UK",
                "sku": "SKU-HALO",
                "asin": "B0HALOPRI",
                "product_name": "光环未验证产品",
                "executed_action": "加价 5%-10%",
                "review_window": "7天后复盘",
                "judgement": "有改善迹象",
                "days_since_execution": "12",
                "current_7d_ad_orders": "1",
                "current_7d_promoted_ad_orders": "0",
                "current_7d_halo_ad_orders": "1",
                "current_7d_total_orders": "1",
                "promoted_conversion_improved": "False",
                "halo_only_conversion": "True",
                "target_sku_not_converted": "True",
            },
        ],
        limit=1,
    )

    assert "光环未验证产品" in html
    assert "普通样本不足产品" not in html
    assert '<span class="tag tag-yellow">本 SKU 未验证</span>' in html


def test_keyword_action_review_renders_execution_anchored_window_metrics() -> None:
    from src.generate_html_report import _render_keyword_action_effect_review_rows

    html = _render_keyword_action_effect_review_rows(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-ANCHOR",
                "asin": "B0ANCHOR01",
                "product_name": "Anchored Product",
                "search_term_or_target": "anchored keyword",
                "executed_action": "加竞价 5%-10%",
                "executed_at": "2026-06-18",
                "review_window": "7天后复盘",
                "days_since_execution": "7",
                "judgement": "有改善迹象",
                "normalized_action": "bid_up",
                "review_data_source": "execution_anchored_daily",
                "pre_7d_start": "2026-06-11",
                "pre_7d_end": "2026-06-17",
                "post_3d_start": "2026-06-18",
                "post_3d_end": "2026-06-20",
                "post_7d_start": "2026-06-18",
                "post_7d_end": "2026-06-24",
                "pre_7d_promoted_ad_orders": "0",
                "pre_7d_total_orders": "1",
                "pre_7d_tacos": "0.12",
                "post_3d_days": "3",
                "post_3d_promoted_ad_orders": "1",
                "post_3d_total_orders": "2",
                "post_3d_acos": "0.10",
                "post_3d_tacos": "0.07",
                "post_3d_available_stock": "18",
                "post_7d_days": "7",
                "post_7d_promoted_ad_orders": "2",
                "post_7d_total_orders": "4",
                "post_7d_acos": "0.12",
                "post_7d_tacos": "0.08",
                "post_7d_available_stock": "18",
                "current_7d_clicks": "8",
                "current_7d_spend": "3.2",
                "current_7d_ad_orders": "2",
                "current_7d_promoted_ad_orders": "2",
                "current_7d_promoted_ad_sales": "30",
                "current_7d_halo_ad_orders": "0",
                "current_7d_total_orders": "4",
                "current_7d_acos": "0.12",
                "current_7d_tacos": "0.08",
                "current_7d_available_stock": "18",
                "promoted_conversion_improved": "True",
            }
        ]
    )

    assert "执行前7天 2026-06-11 至 2026-06-17" in html
    assert "执行后3天 2026-06-18 至 2026-06-20" in html
    assert "执行后3天 ACOS 0.10" in html
    assert "执行后3天库存 18" in html
    assert "执行后7天 2026-06-18 至 2026-06-24" in html
    assert "执行后本 SKU 单 2" in html
    assert "执行后 ACOS 0.12" in html
    assert "执行后 TACOS 0.08" in html
    assert "执行后库存 18" in html


def test_keyword_action_review_hides_anchor_7d_metrics_before_7_day_window() -> None:
    from src.generate_html_report import _render_keyword_action_effect_review_rows

    html = _render_keyword_action_effect_review_rows(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-ANCHOR",
                "asin": "B0ANCHOR04",
                "product_name": "Anchored Early Product",
                "search_term_or_target": "anchored early keyword",
                "executed_action": "加竞价 5%-10%",
                "executed_at": "2026-06-18",
                "review_window": "3d_check",
                "days_since_execution": "4",
                "judgement": "样本不足",
                "normalized_action": "bid_up",
                "review_data_source": "execution_anchored_daily",
                "pre_7d_start": "2026-06-11",
                "pre_7d_end": "2026-06-17",
                "post_3d_start": "2026-06-18",
                "post_3d_end": "2026-06-20",
                "post_7d_start": "2026-06-18",
                "post_7d_end": "2026-06-24",
                "pre_7d_promoted_ad_orders": "0",
                "pre_7d_total_orders": "1",
                "pre_7d_tacos": "0.12",
                "post_3d_days": "3",
                "post_3d_promoted_ad_orders": "1",
                "post_3d_total_orders": "2",
                "post_3d_acos": "0.10",
                "post_3d_tacos": "0.07",
                "post_3d_available_stock": "18",
                "post_7d_days": "7",
                "post_7d_promoted_ad_orders": "2",
                "post_7d_total_orders": "4",
                "post_7d_acos": "0.12",
                "post_7d_tacos": "0.08",
                "post_7d_available_stock": "18",
            }
        ]
    )

    assert "执行前7天 2026-06-11 至 2026-06-17" in html
    assert "执行后3天 2026-06-18 至 2026-06-20" in html
    assert "执行后3天本 SKU 单 1" in html
    assert "执行后3天 ACOS 0.10" in html
    assert "执行后7天 2026-06-18 至 2026-06-24" not in html
    assert "执行后本 SKU 单 2" not in html
    assert "执行后 ACOS 0.12" not in html


def test_keyword_action_review_shows_missing_manual_review_metrics() -> None:
    from src.generate_html_report import _render_keyword_action_effect_review_rows

    html = _render_keyword_action_effect_review_rows(
        [
            {
                "marketplace": "US",
                "sku": "SKU-MISS",
                "asin": "B0KEYMISSMETRIC",
                "product_name": "缺指标产品",
                "search_term_or_target": "missing metrics term",
                "executed_action": "加价 5%-10%",
                "executed_at": "2026-05-20",
                "review_window": "7天后复盘",
                "review_outcome": "needs_manual_review",
                "judgement": "待人工复查",
                "normalized_action": "bid_up",
                "days_since_execution": "12",
                "current_7d_clicks": "8",
                "current_7d_spend": "3.20",
                "current_7d_ad_orders": "1",
                "current_7d_promoted_ad_orders": "1",
                "current_7d_acos": "0.08",
                "current_7d_tacos": "",
                "current_7d_total_orders": "",
                "current_7d_available_stock": "",
                "current_14d_promoted_ad_orders": "2",
                "current_14d_acos": "0.09",
                "current_14d_tacos": "",
                "current_14d_total_orders": "5",
                "current_14d_available_stock": "",
                "promoted_conversion_improved": "True",
                "halo_only_conversion": "False",
                "block_reason": "本 SKU 有单但缺少TACOS、总单、库存，需人工复查",
            }
        ],
        limit=1,
    )

    assert "缺失复盘指标：7天 TACOS、7天总单、7天库存、14天 TACOS、14天库存" in html
    assert "本 SKU 有单但缺少TACOS、总单、库存" in html


def test_keyword_action_review_treats_nan_metrics_as_missing() -> None:
    from src.generate_html_report import _render_keyword_action_effect_review_rows

    html = _render_keyword_action_effect_review_rows(
        [
            {
                "marketplace": "US",
                "sku": "SKU-NAN",
                "asin": "B0KEYNAN",
                "product_name": "NaN 关键词产品",
                "search_term_or_target": "nan metrics term",
                "executed_action": "加价 5%-10%",
                "executed_at": "2026-05-20",
                "review_window": "7天后复盘",
                "review_outcome": "needs_manual_review",
                "judgement": "待人工复查",
                "normalized_action": "bid_up",
                "days_since_execution": "12",
                "current_7d_clicks": "nan",
                "current_7d_spend": "NaN",
                "current_7d_ad_orders": "nan",
                "current_7d_promoted_ad_orders": "1",
                "current_7d_acos": "nan",
                "current_7d_tacos": "NaN",
                "current_7d_total_orders": "3",
                "current_7d_available_stock": "nan",
                "current_14d_promoted_ad_orders": "2",
                "current_14d_acos": "0.09",
                "current_14d_tacos": "nan",
                "current_14d_total_orders": "5",
                "current_14d_available_stock": "18",
            }
        ],
        limit=1,
    )

    assert "缺失复盘指标：7天 ACOS、7天 TACOS、7天库存、14天 TACOS" in html
    assert "7天 点击 N/A" in html
    assert "花费 N/A" in html
    assert "ACOS N/A" in html
    assert "TACOS N/A" in html
    assert "库存 N/A" in html
    assert "点击 nan" not in html.lower()
    assert "花费 nan" not in html.lower()
    assert "库存 nan" not in html.lower()


def test_keyword_action_review_nan_trigger_metrics_do_not_imply_remaining_spend() -> None:
    from src.generate_html_report import _render_keyword_action_effect_review_rows

    html = _render_keyword_action_effect_review_rows(
        [
            {
                "marketplace": "US",
                "sku": "SKU-NAN-TRIGGER",
                "asin": "B0NANTRIGGER",
                "product_name": "NaN 触发标准产品",
                "search_term_or_target": "nan trigger term",
                "executed_action": "否定精准",
                "executed_at": "2026-05-20",
                "review_window": "7天后复盘",
                "review_outcome": "needs_manual_review",
                "judgement": "待人工复查",
                "normalized_action": "negative_exact",
                "days_since_execution": "12",
                "current_7d_clicks": "nan",
                "current_7d_spend": "NaN",
                "current_7d_promoted_ad_orders": "0",
                "current_7d_acos": "nan",
                "current_7d_tacos": "NaN",
                "current_7d_total_orders": "0",
                "current_7d_available_stock": "18",
                "current_14d_promoted_ad_orders": "0",
                "current_14d_acos": "nan",
                "current_14d_tacos": "nan",
                "current_14d_total_orders": "0",
                "current_14d_available_stock": "18",
            }
        ],
        limit=1,
    )

    assert "触发标准：7 天点击或花费缺失，不能判断动作效果。" in html
    assert "消耗已停止" not in html
    assert "仍有点击或花费" not in html


def test_keyword_action_review_renders_decision_summary_and_attribution() -> None:
    from src.generate_html_report import _render_keyword_action_effect_review_rows

    html = _render_keyword_action_effect_review_rows(
        [
            {
                "marketplace": "US",
                "sku": "SKU-1",
                "asin": "B0PROMOTED",
                "product_name": "演示笔记本",
                "search_term_or_target": "metal board",
                "executed_action": "加价 5%-10%",
                "executed_at": "2026-05-20",
                "confirmed_note": "用户反馈：已执行 metal board 加价",
                "report_date": "2026-05-19",
                "review_date": "2026-06-02",
                "review_status": "可做7天复查",
                "normalized_action": "bid_up",
                "action_scope": "search_term",
                "action_id": "US||SKU-1||B0PROMOTED||search_term||metal board||bid_up",
                "days_since_execution": "13",
                "review_window": "7天后复盘",
                "judgement": "有改善迹象",
                "next_step": "保留当前竞价。",
                "current_7d_clicks": "8",
                "current_7d_spend": "3.20",
                "current_7d_ad_orders": "2",
                "current_7d_ad_sales": "30.00",
                "current_7d_promoted_ad_orders": "2",
                "current_7d_promoted_ad_sales": "30.00",
                "current_7d_halo_ad_orders": "0",
                "current_7d_acos": "0.12",
                "current_7d_total_orders": "4",
                "current_7d_tacos": "0.09",
                "current_7d_available_stock": "18",
                "current_14d_clicks": "14",
                "current_14d_spend": "5.10",
                "current_14d_ad_orders": "3",
                "current_14d_promoted_ad_orders": "3",
                "current_14d_halo_ad_orders": "0",
                "current_14d_acos": "0.14",
                "current_14d_total_orders": "7",
                "current_14d_tacos": "0.11",
                "current_14d_available_stock": "18",
                "promoted_conversion_improved": "True",
                "halo_only_conversion": "False",
                "target_sku_not_converted": "False",
            },
            {
                "marketplace": "UK",
                "sku": "SKU-2",
                "asin": "B0HALOONLY",
                "product_name": "演示线夹",
                "search_term_or_target": "B0TARGET",
                "executed_action": "加价 5%-10%",
                "executed_at": "2026-05-21",
                "confirmed_note": "用户反馈：已执行 B0TARGET 加价",
                "report_date": "2026-05-20",
                "review_date": "2026-06-02",
                "review_status": "可做7天复查",
                "normalized_action": "bid_up",
                "action_scope": "asin_target",
                "action_id": "UK||SKU-2||B0HALOONLY||asin_target||B0TARGET||bid_up",
                "days_since_execution": "12",
                "review_window": "7天后复盘",
                "judgement": "样本不足",
                "next_step": "暂不追加预算。",
                "current_7d_clicks": "3",
                "current_7d_spend": "1.20",
                "current_7d_ad_orders": "1",
                "current_7d_ad_sales": "10.00",
                "current_7d_promoted_ad_orders": "0",
                "current_7d_promoted_ad_sales": "0",
                "current_7d_halo_ad_orders": "1",
                "current_7d_acos": "0.20",
                "current_7d_total_orders": "1",
                "current_7d_tacos": "0.08",
                "current_7d_available_stock": "6",
                "current_14d_clicks": "5",
                "current_14d_spend": "2.00",
                "current_14d_ad_orders": "1",
                "current_14d_promoted_ad_orders": "0",
                "current_14d_halo_ad_orders": "1",
                "current_14d_acos": "0.20",
                "current_14d_total_orders": "2",
                "current_14d_tacos": "0.10",
                "current_14d_available_stock": "6",
                "promoted_conversion_improved": "False",
                "halo_only_conversion": "True",
                "target_sku_not_converted": "True",
            },
        ],
        limit=2,
    )

    assert '<span>有效或改善</span><strong>1</strong>' in html
    assert '<span>样本不足</span><strong>1</strong>' in html
    assert "判断硬标准" in html
    assert "复盘天数" in html
    assert "&lt;3 天：不判有效；3-6 天：初判；&gt;=7 天：正式判断。" in html
    assert "加竞价有效" in html
    assert "7 天内本 SKU 单 &gt; 0 且销售额 &gt; 0：有改善迹象。风险：销量低仍可能被判有效。" in html
    assert "加竞价停止" in html
    assert "点击 &gt;= 8 或花费 &gt;= 5 且无本 SKU 单：暂未改善，停止加价。风险：低样本可能误判。" in html
    assert "低流量 SKU 易误判收敛" in html
    assert "仅光环单：标记目标 SKU 未验证，不允许加价。" in html
    assert "ACOS 约束" in html
    assert "本 SKU 有单但 ACOS 高于目标时，只能算有成交，不能继续加价或放量。" in html
    assert "前后对比" in html
    assert "正式复盘应比较执行前 7 天和执行后 7 天；只有执行后数据时，结论需降级。" in html
    assert "库存约束" in html
    assert "库存不足时禁止放量" in html
    assert "前台约束" in html
    assert "复盘结论分布" in html
    assert '<span>可保留</span>' in html
    assert '<span>停止追加</span>' in html
    assert "复盘结论</strong><p>可保留：只保留当前竞价，今天不重复加价，避免用小样本追高</p>" in html
    assert "复盘结论</strong><p>停止追加：有光环订单也不能证明本 SKU 有效，今天不追加预算或竞价</p>" in html
    assert "触发标准</strong><p>触发标准：加竞价后 7 天本 SKU 单 &gt; 0 且销售额 &gt; 0，可保留当前竞价。</p>" in html
    assert "触发标准</strong><p>触发标准：光环单 &gt; 0 且本 SKU 单 = 0，不能作为本 SKU 有效证据。</p>" in html
    assert "归因判断：本 SKU 成交有效" in html
    assert "仅光环成交，不算本 SKU 有效" in html
    assert "执行记录：用户反馈：已执行 metal board 加价" in html
    assert "报告日：2026-05-19；复盘日：2026-06-02" in html
    assert "7 天窗口可做效果判断" in html
    assert "动作口径：搜索词｜加竞价" in html
    assert "追踪 ID：US||SKU-1||B0PROMOTED||search_term||metal board||bid_up" in html
    assert "7天 点击 8" in html
    assert "总单 4" in html
    assert "TACOS 0.09" in html
    assert "库存 18" in html
    assert "14天 点击 14" in html
    assert "本 SKU 单 0" in html


def test_keyword_action_review_renders_keep_current_as_positive_judgement() -> None:
    from src.generate_html_report import _render_keyword_action_effect_review_rows

    html = _render_keyword_action_effect_review_rows(
        [
            {
                "marketplace": "US",
                "sku": "SKU-KEEP",
                "asin": "B0KEEP0001",
                "product_name": "保留复盘产品",
                "search_term_or_target": "keep current term",
                "executed_action": "加价 5%-10%",
                "executed_at": "2026-05-20",
                "normalized_action": "bid_up",
                "action_scope": "search_term",
                "action_id": "US||SKU-KEEP||B0KEEP0001||search_term||keep current term||bid_up",
                "days_since_execution": "13",
                "review_window": "7天后复盘",
                "judgement": "可保留",
                "current_7d_clicks": "8",
                "current_7d_spend": "3.20",
                "current_7d_ad_orders": "2",
                "current_7d_promoted_ad_orders": "2",
                "current_7d_promoted_ad_sales": "30.00",
                "current_7d_halo_ad_orders": "0",
                "current_7d_acos": "0.12",
                "current_7d_total_orders": "4",
                "current_7d_tacos": "0.09",
                "current_7d_available_stock": "18",
                "current_14d_promoted_ad_orders": "3",
                "current_14d_tacos": "0.11",
                "current_14d_available_stock": "18",
                "promoted_conversion_improved": "True",
                "halo_only_conversion": "False",
                "target_sku_not_converted": "False",
            }
        ],
        limit=1,
    )

    assert '<span class="tag tag-green">可保留</span>' in html
    assert "复盘结论</strong><p>可保留：只保留当前竞价，今天不重复加价，避免用小样本追高</p>" in html
    assert "待判断" not in html


def test_keyword_action_review_renders_clear_improvement_as_positive_judgement() -> None:
    from src.generate_html_report import _render_keyword_action_effect_review_rows

    html = _render_keyword_action_effect_review_rows(
        [
            {
                "marketplace": "US",
                "sku": "SKU-CLEAR",
                "asin": "B0CLEAR0001",
                "product_name": "明确改善复盘产品",
                "search_term_or_target": "clear improvement term",
                "executed_action": "加价 5%-10%",
                "executed_at": "2026-05-20",
                "normalized_action": "bid_up",
                "action_scope": "search_term",
                "action_id": "US||SKU-CLEAR||B0CLEAR0001||search_term||clear improvement term||bid_up",
                "days_since_execution": "13",
                "review_window": "7天后复盘",
                "judgement": "明确改善",
                "current_7d_clicks": "8",
                "current_7d_spend": "3.20",
                "current_7d_ad_orders": "2",
                "current_7d_promoted_ad_orders": "2",
                "current_7d_promoted_ad_sales": "30.00",
                "current_7d_halo_ad_orders": "0",
                "current_7d_acos": "0.12",
                "current_7d_total_orders": "4",
                "current_7d_tacos": "0.09",
                "current_7d_available_stock": "18",
                "current_14d_promoted_ad_orders": "3",
                "current_14d_tacos": "0.11",
                "current_14d_available_stock": "18",
                "promoted_conversion_improved": "True",
                "halo_only_conversion": "False",
                "target_sku_not_converted": "False",
            }
        ],
        limit=1,
    )

    assert '<span class="tag tag-green">明确改善</span>' in html
    assert '<span class="tag tag-red">明确改善</span>' not in html


def test_keyword_action_review_downgrades_halo_only_positive_judgement_in_summary() -> None:
    from src.generate_html_report import _render_keyword_action_effect_review_rows

    html = _render_keyword_action_effect_review_rows(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-2",
                "asin": "B0HALOONLY",
                "product_name": "演示线夹",
                "search_term_or_target": "B0TARGET",
                "executed_action": "加价 5%-10%",
                "executed_at": "2026-05-21",
                "review_window": "7天后复盘",
                "judgement": "初步有效",
                "normalized_action": "bid_up",
                "days_since_execution": "12",
                "current_7d_clicks": "3",
                "current_7d_spend": "1.20",
                "current_7d_ad_orders": "1",
                "current_7d_promoted_ad_orders": "0",
                "current_7d_promoted_ad_sales": "0",
                "current_7d_halo_ad_orders": "1",
                "current_7d_total_orders": "1",
                "current_7d_acos": "0.20",
                "current_7d_tacos": "0.08",
                "current_7d_available_stock": "6",
                "promoted_conversion_improved": "False",
                "halo_only_conversion": "True",
                "target_sku_not_converted": "True",
            }
        ],
        limit=1,
    )

    assert '<span>有效或改善</span><strong>0</strong>' in html
    assert '<span>本 SKU 未验证</span><strong>1</strong>' in html
    assert '<span>停止追加</span>' in html
    assert '<span class="tag tag-yellow">本 SKU 未验证</span>' in html
    assert '<span class="tag tag-green">初步有效</span>' not in html


def test_keyword_action_review_prioritizes_not_validated_before_sample_when_limited() -> None:
    from src.generate_html_report import _render_keyword_action_effect_review_rows

    html = _render_keyword_action_effect_review_rows(
        [
            {
                "marketplace": "US",
                "sku": "SKU-SAMPLE",
                "asin": "B0SAMPLEKW",
                "product_name": "样本不足产品",
                "search_term_or_target": "ordinary sample term",
                "executed_action": "观察",
                "review_window": "7天后复盘",
                "judgement": "样本不足",
                "days_since_execution": "14",
            },
            {
                "marketplace": "US",
                "sku": "SKU-HALO",
                "asin": "B0HALOKW",
                "product_name": "光环未验证产品",
                "search_term_or_target": "halo only target",
                "executed_action": "加价 5%-10%",
                "review_window": "7天后复盘",
                "judgement": "初步有效",
                "days_since_execution": "12",
                "normalized_action": "bid_up",
                "current_7d_ad_orders": "1",
                "current_7d_promoted_ad_orders": "0",
                "current_7d_halo_ad_orders": "1",
                "current_7d_total_orders": "1",
                "promoted_conversion_improved": "False",
                "halo_only_conversion": "True",
                "target_sku_not_converted": "True",
            },
        ],
        limit=1,
    )

    assert "halo only target" in html
    assert "ordinary sample term" not in html
    assert '<span class="tag tag-yellow">本 SKU 未验证</span>' in html


def test_latest_action_review_rows_filter_by_marketplace(monkeypatch, tmp_path) -> None:
    from src.autoopt_feedback import ACTION_REVIEW_REQUIRED_FIELDS
    import src.report_presentation as presentation

    output = tmp_path / "output"
    output.mkdir()
    (output / "action_review_20260616.json").write_text(
        json.dumps(
            [
                {
                    "marketplace": "UK",
                    "sku": "SKU-UK",
                    "asin": "B0UK",
                    "product_name": "UK product",
                    "action_type": "观察",
                    "days_since_execution": 8,
                    "outcome": "样本不足",
                },
                {
                    "marketplace": "US",
                    "sku": "SKU-US",
                    "asin": "B0US",
                    "product_name": "US product",
                    "action_type": "观察",
                    "days_since_execution": 8,
                    "review_window": "7天后复盘",
                    "outcome": "初步有效",
                    "current_7d_promoted_ad_orders": 2,
                    "current_7d_total_orders": 5,
                    "current_7d_tacos": 0.08,
                    "current_7d_target_acos": 0.2,
                    "current_7d_available_stock": 18,
                    "current_14d_promoted_ad_orders": 3,
                    "current_14d_total_orders": 8,
                    "current_14d_tacos": 0.09,
                    "current_14d_available_stock": 18,
                },
            ],
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(presentation, "OUTPUT_DIR", output)

    rows = presentation._latest_action_review_rows("US")

    assert [row["marketplace"] for row in rows] == ["US"]
    assert rows[0]["sku"] == "SKU-US"
    assert set(ACTION_REVIEW_REQUIRED_FIELDS).issubset(rows[0])
    assert rows[0]["review_window"] == "7天后复盘"
    assert rows[0]["current_7d_promoted_ad_orders"] == "2"
    assert rows[0]["current_7d_total_orders"] == "5"
    assert rows[0]["current_7d_tacos"] == "0.08"
    assert rows[0]["current_7d_target_acos"] == "0.2"
    assert rows[0]["current_7d_available_stock"] == "18"
    assert rows[0]["current_14d_promoted_ad_orders"] == "3"
    assert rows[0]["current_14d_total_orders"] == "8"
    assert rows[0]["current_14d_tacos"] == "0.09"
    assert rows[0]["current_14d_available_stock"] == "18"


def test_latest_keyword_action_review_rows_filter_by_marketplace(monkeypatch, tmp_path) -> None:
    from src.autoopt_feedback import KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS
    import src.report_presentation as presentation

    output = tmp_path / "output"
    output.mkdir()
    (output / "keyword_action_review_20260616.json").write_text(
        json.dumps(
            [
                {
                    "marketplace": "UK",
                    "sku": "SKU-UK",
                    "asin": "B0UK",
                    "search_term_or_target": "spiral notebook",
                    "outcome": "样本不足",
                },
                {
                    "marketplace": "DE",
                    "sku": "SKU-DE",
                    "asin": "B0DE",
                    "search_term_or_target": "stationery box",
                    "outcome": "有改善迹象",
                    "current_7d_total_orders": 4,
                    "current_7d_tacos": 0.12,
                    "current_7d_target_acos": "20%",
                    "current_7d_available_stock": 22,
                    "current_14d_total_orders": 7,
                    "current_14d_tacos": 0.14,
                    "current_14d_available_stock": 22,
                },
            ],
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(presentation, "OUTPUT_DIR", output)

    rows = presentation._latest_keyword_action_review_rows("DE")

    assert [row["marketplace"] for row in rows] == ["DE"]
    assert rows[0]["search_term_or_target"] == "stationery box"
    assert rows[0]["current_7d_total_orders"] == "4"
    assert rows[0]["current_7d_tacos"] == "0.12"
    assert rows[0]["current_7d_target_acos"] == "20%"
    assert rows[0]["current_7d_available_stock"] == "22"
    assert rows[0]["current_14d_total_orders"] == "7"
    assert rows[0]["current_14d_tacos"] == "0.14"
    assert rows[0]["current_14d_available_stock"] == "22"
    assert set(KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS).issubset(rows[0])


def test_keyword_action_review_uses_report_date_for_review_window() -> None:
    from src.autoopt_feedback import build_keyword_action_review_rows

    rows = build_keyword_action_review_rows(
        [
            {
                "marketplace": "UK",
                "summary": {"report_date": "2026-06-24"},
                "analysis_payload": {
                    "report_date": "2026-06-24",
                    "搜索词分析": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "dimmer desk lamp",
                                "clicks": 8,
                                "spend": 4,
                                "ad_orders": 1,
                                "ad_sales": 30,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 30,
                                "halo_ad_orders": 0,
                                "halo_ad_sales": 0,
                                "ACOS": "8%",
                            }
                        ],
                    },
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "total_orders": 3,
                                "TACOS": 0.08,
                                "available_stock": 18,
                            }
                        ],
                    },
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "product_name": "Demo desk lamp",
                "search_term_or_target": "dimmer desk lamp",
                "today_action": "加价 5%-10%",
                "confirmed_status": "已执行",
                "confirmed_at": "2026-06-22",
            }
        ],
    )

    assert rows[0]["review_date"] == "2026-06-24"
    assert rows[0]["days_since_execution"] == 2
    assert rows[0]["review_window"] == "未满3天"
    assert rows[0]["review_phase"] == "under_3_days"
    assert rows[0]["outcome"] == "样本不足"
    assert rows[0]["review_outcome"] == "not_ready"


def test_keyword_action_review_uses_execution_anchored_post_window_over_rolling_metrics() -> None:
    from src.autoopt_feedback import build_keyword_action_review_rows

    rows = build_keyword_action_review_rows(
        [
            {
                "marketplace": "UK",
                "summary": {"report_date": "2026-06-24"},
                "analysis_payload": {
                    "report_date": "2026-06-24",
                    "搜索词分析": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "search_term": "dimmer desk lamp",
                                "clicks": 8,
                                "spend": 4,
                                "ad_orders": 1,
                                "ad_sales": 30,
                                "promoted_ad_orders": 1,
                                "promoted_ad_sales": 30,
                                "halo_ad_orders": 0,
                                "ACOS": "8%",
                            }
                        ],
                    },
                    "product_window_metrics": {
                        "7d": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "total_orders": 3,
                                "TACOS": 0.08,
                                "available_stock": 18,
                            }
                        ],
                    },
                    "review_search_term_daily": [
                        {
                            "date": "2026-06-17",
                            "marketplace": "UK",
                            "sku": "SKU-1",
                            "asin": "B0TEST1234",
                            "search_term": "dimmer desk lamp",
                            "clicks": 4,
                            "spend": 2,
                            "ad_orders": 1,
                            "ad_sales": 30,
                            "promoted_ad_orders": 1,
                            "promoted_ad_sales": 30,
                            "halo_ad_orders": 0,
                        },
                        {
                            "date": "2026-06-20",
                            "marketplace": "UK",
                            "sku": "SKU-1",
                            "asin": "B0TEST1234",
                            "search_term": "dimmer desk lamp",
                            "clicks": 10,
                            "spend": 6,
                            "ad_orders": 0,
                            "ad_sales": 0,
                            "promoted_ad_orders": 0,
                            "promoted_ad_sales": 0,
                            "halo_ad_orders": 0,
                        },
                    ],
                    "review_product_daily": [
                        {
                            "date": "2026-06-17",
                            "marketplace": "UK",
                            "sku": "SKU-1",
                            "asin": "B0TEST1234",
                            "clicks": 4,
                            "spend": 2,
                            "ad_orders": 1,
                            "ad_sales": 30,
                            "promoted_ad_orders": 1,
                            "promoted_ad_sales": 30,
                            "halo_ad_orders": 0,
                            "total_orders": 2,
                            "total_sales": 60,
                            "available_stock": 18,
                        },
                        {
                            "date": "2026-06-20",
                            "marketplace": "UK",
                            "sku": "SKU-1",
                            "asin": "B0TEST1234",
                            "clicks": 10,
                            "spend": 6,
                            "ad_orders": 0,
                            "ad_sales": 0,
                            "promoted_ad_orders": 0,
                            "promoted_ad_sales": 0,
                            "halo_ad_orders": 0,
                            "total_orders": 0,
                            "total_sales": 0,
                            "available_stock": 18,
                        },
                    ],
                },
            }
        ],
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "product_name": "Demo desk lamp",
                "search_term_or_target": "dimmer desk lamp",
                "today_action": "加价 5%-10%",
                "confirmed_status": "已执行",
                "confirmed_at": "2026-06-18",
            }
        ],
    )

    assert rows[0]["review_data_source"] == "execution_anchored_daily"
    assert rows[0]["post_3d_start"] == "2026-06-18"
    assert rows[0]["post_3d_end"] == "2026-06-20"
    assert rows[0]["post_3d_days"] != ""
    assert rows[0]["post_3d_available_stock"] != ""
    assert rows[0]["post_7d_start"] == "2026-06-18"
    assert rows[0]["pre_7d_promoted_ad_orders"] == 1.0
    assert rows[0]["current_7d_promoted_ad_orders"] == 0.0
    assert rows[0]["post_7d_promoted_ad_orders"] == 0.0
    assert rows[0]["outcome"] not in {"初步有效", "有改善迹象", "明确改善", "可保留"}
    assert rows[0]["promoted_conversion_improved"] is False


def test_parse_amazon_search_results_extracts_competitors() -> None:
    from scripts.run_frontend_checks import parse_amazon_search_results

    html = """
    <div data-asin="B0OWN12345" class="s-result-item"><h2><span>Own product</span></h2></div>
    <div data-asin="B0COMP1111" class="s-result-item"><h2><span>Competitor One Metal Board</span></h2><span class="a-offscreen">£12.99</span><span class="a-icon-alt">4.5 out of 5 stars</span><span aria-label="120 ratings"></span></div>
    <div data-asin="B0COMP2222" class="s-result-item"><h2><span>Competitor Two</span></h2><span class="a-offscreen">£14.99</span></div>
    <div data-asin="B0COMP3333" class="s-result-item"><h2><span>Competitor Three</span></h2><span class="a-offscreen">£9.99</span></div>
    <div data-asin="B0COMP4444" class="s-result-item"><h2><span>Competitor Four</span></h2><span class="a-offscreen">£8.99</span></div>
    """

    parsed = parse_amazon_search_results(html, own_asin="B0OWN12345")

    assert parsed["own_search_position"] == "1"
    assert [item["asin"] for item in parsed["competitors"]] == ["B0COMP1111", "B0COMP2222", "B0COMP3333"]
    assert parsed["search_result_count"] == 4


def test_search_partial_evidence_scores_without_strong_conclusion() -> None:
    from scripts.run_frontend_checks import _frontend_quality_payload

    payload = _frontend_quality_payload(
        {"price": "£12.99", "rating": "4.5 out of 5 stars", "reviews": "120 ratings", "coupon": "10%", "buy_box": "识别到购买按钮"},
        {"marketplace": "UK", "frontend_search_url": "https://www.amazon.co.uk/s?k=board"},
        {
            "frontend_search_status": "已读取部分结果",
            "frontend_search_url": "https://www.amazon.co.uk/s?k=board",
            "frontend_search_result_count": 12,
            "frontend_search_partial_evidence": True,
            "frontend_competitors": [],
            "own_search_position": "",
        },
        "",
        "",
        status="已自动检查",
        method="urllib",
    )

    assert payload["frontend_search_partial_evidence"] is True
    assert payload["frontend_search_result_count"] == 12
    assert payload["frontend_search_quality_score"] > 0
    assert payload["frontend_auto_conclusion"] != "FRONTEND_OK"


def test_frontend_search_gap_requires_two_competitor_samples() -> None:
    from scripts.run_frontend_checks import _frontend_quality_payload

    payload = _frontend_quality_payload(
        {"price": "$8.99", "rating": "3.9 out of 5 stars", "reviews": "(52)", "coupon": "待确认", "buy_box": "识别到购买按钮"},
        {"marketplace": "US", "frontend_search_url": "https://www.amazon.com/s?k=desk+board"},
        {
            "frontend_search_status": "已自动检查",
            "frontend_search_url": "https://www.amazon.com/s?k=desk+board",
            "frontend_competitor_count": 1,
            "frontend_competitors": [
                {"asin": "B0COMP1111", "price": "$11.99", "rating": "4.6 out of 5 stars", "reviews": "1,200 ratings"}
            ],
        },
        "",
        "",
        status="已自动检查",
        method="chrome-cdp",
    )

    assert "competitor_sample_small" in payload["frontend_search_quality_flags"]
    assert payload["frontend_price_delta_pct"] == ""
    assert payload["frontend_rating_delta"] == ""
    assert payload["frontend_review_delta_pct"] == ""


def test_cache_frontend_quality_never_marks_frontend_ok() -> None:
    from scripts.run_frontend_checks import _quality_payload_from_record

    payload = _quality_payload_from_record(
        {
            "marketplace": "UK",
            "frontend_check_status": "沿用 2026-06-01 前台数据",
            "frontend_price": "£12.99",
            "frontend_rating": "4.8 out of 5 stars",
            "frontend_reviews": "1,200 ratings",
            "frontend_coupon": "10%",
            "frontend_buy_box": "识别到购买按钮",
            "frontend_search_status": "已自动检查",
            "frontend_search_result_count": 12,
            "frontend_competitor_count": 3,
            "comparable_competitor_count": 3,
            "competitor_comparability": "high",
            "frontend_competitors": [
                {"asin": "B0COMP1111", "price": "£13.99", "rating": "4.6 out of 5 stars", "reviews": "1,100 ratings"},
                {"asin": "B0COMP2222", "price": "£14.99", "rating": "4.7 out of 5 stars", "reviews": "1,300 ratings"},
                {"asin": "B0COMP3333", "price": "£15.99", "rating": "4.8 out of 5 stars", "reviews": "1,500 ratings"},
            ],
        },
        "network failed",
        basis="cache",
    )

    assert payload["frontend_auto_conclusion"] == "INSUFFICIENT_EVIDENCE"
    assert payload["frontend_auto_conclusion_label"] == "缓存证据，仅背景参考，不能支持放量"
    assert "bid_up" in payload["frontend_auto_conclusion_blocked_ad_actions"]


def test_frontend_product_price_rejects_wrong_marketplace_currency() -> None:
    from scripts.run_frontend_checks import _findings, _front_structured_fields, parse_amazon_frontend

    html = """
    <span id="productTitle">UK product</span>
    <span class="a-offscreen">TWD$ 499.00</span>
    <span class="a-icon-alt">4.5 out of 5 stars</span>
    """

    parsed = parse_amazon_frontend(html, marketplace="UK")

    assert parsed["price"] == ""
    assert "价格币种异常" in parsed["price_currency_warning"]
    assert _front_structured_fields(parsed)["frontend_price"] == ""
    assert "TWD$ 499.00" in _findings(parsed, "")


def test_frontend_search_competitor_prices_reject_wrong_marketplace_currency() -> None:
    from scripts.run_frontend_checks import parse_amazon_search_results

    html = """
    <div data-asin="B0COMP1111" class="s-result-item"><h2><span>Wrong currency</span></h2><span class="a-offscreen">$12.99</span></div>
    <div data-asin="B0COMP2222" class="s-result-item"><h2><span>Correct currency</span></h2><span class="a-offscreen">£14.99</span></div>
    """

    parsed = parse_amazon_search_results(html, marketplace="UK")

    assert parsed["competitors"][0]["price"] == ""
    assert "价格币种异常" in parsed["competitors"][0]["price_currency_warning"]
    assert parsed["competitors"][1]["price"] == "£14.99"


def test_frontend_search_competitor_prices_reject_cad_on_us_marketplace() -> None:
    from scripts.run_frontend_checks import parse_amazon_search_results

    html = """
    <div data-asin="B0COMP1111" class="s-result-item"><h2><span>Canadian currency</span></h2><span class="a-offscreen">CAD 20.95</span></div>
    <div data-asin="B0COMP2222" class="s-result-item"><h2><span>Correct currency</span></h2><span class="a-offscreen">$14.99</span></div>
    """

    parsed = parse_amazon_search_results(html, marketplace="US")

    assert parsed["competitors"][0]["price"] == ""
    assert "价格币种异常" in parsed["competitors"][0]["price_currency_warning"]
    assert parsed["competitors"][1]["price"] == "$14.99"


def test_frontend_product_price_rejects_jpy_cache_contamination() -> None:
    from scripts.run_frontend_checks import _has_currency_contamination, parse_amazon_frontend

    parsed = parse_amazon_frontend(
        '<span id="productTitle">UK product</span><span class="a-offscreen">JPY4,658</span>',
        marketplace="UK",
    )

    assert parsed["price"] == ""
    assert "价格币种异常" in parsed["price_currency_warning"]
    assert _has_currency_contamination(
        {
            "marketplace": "UK",
            "frontend_check_status": "已自动检查",
            "frontend_findings": "售价：JPY4,658；评分：4.0 out of 5 stars",
            "frontend_price": "JPY4,658",
        }
    )


def test_frontend_urllib_wrong_visible_location_is_not_success() -> None:
    from scripts.run_frontend_checks import _frontend_quality_payload, _status_from_parsed, parse_amazon_frontend

    html = """
    <span id="productTitle">UK product</span>
    <span id="glow-ingress-line2">United States</span>
    <span class="a-offscreen">£14.99</span>
    <span class="a-icon-alt">4.5 out of 5 stars</span>
    """

    parsed = parse_amazon_frontend(html, marketplace="UK")
    payload = _frontend_quality_payload(
        parsed,
        {"marketplace": "UK", "product_url": "https://www.amazon.co.uk/dp/B0TEST1234"},
        {"frontend_search_status": "待前台检查", "frontend_competitors": []},
        "",
        status="待前台检查",
        method="urllib",
        location_note=parsed["visible_location"],
    )

    assert _status_from_parsed(parsed) == "待前台检查"
    assert "地区异常" in parsed["location_warning"]
    assert payload["frontend_failure_category"] == "location_unverified"
    assert payload["frontend_evidence_quality_score"] <= 35
    assert payload["frontend_location_verified"] is False


def test_frontend_missing_location_keeps_product_read_but_not_strong_evidence() -> None:
    from scripts.run_frontend_checks import _frontend_quality_payload, _status_from_parsed, parse_amazon_frontend

    html = """
    <span id="productTitle">UK product</span>
    <span class="a-offscreen">£17.89</span>
    <span class="a-icon-alt">4.2 out of 5 stars</span>
    <span id="acrCustomerReviewText">(38)</span>
    <input id="add-to-cart-button" />
    """
    parsed = parse_amazon_frontend(html, marketplace="UK")
    assert _status_from_parsed(parsed) == "已自动检查"

    payload = _frontend_quality_payload(
        parsed,
        {"marketplace": "UK", "frontend_search_url": ""},
        {"frontend_search_status": "待前台检查", "frontend_competitors": []},
        "UK 地区未确认",
        status="已自动检查",
        location_note="",
        method="chrome",
    )

    assert payload["frontend_location_scope"] == "missing"
    assert payload["frontend_location_verified"] is False
    assert payload["frontend_evidence_quality_score"] <= 65
    assert payload["frontend_auto_conclusion"] != "FRONTEND_OK"


def test_frontend_same_marketplace_non_exact_location_is_background_only() -> None:
    from scripts.run_frontend_checks import _frontend_quality_payload, parse_amazon_frontend
    from src.report_presentation import _frontend_evidence_audit

    html = """
    <span id="productTitle">UK product</span>
    <span id="glow-ingress-line2">Aberdeen AB10 1</span>
    <span class="a-offscreen">£18.99</span>
    <span class="a-icon-alt">4.5 out of 5 stars</span>
    <span id="acrCustomerReviewText">120 ratings</span>
    """
    search_payload = {
        "frontend_search_status": "已自动检查",
        "frontend_competitor_count": 3,
        "frontend_competitors": [
            {"asin": "B0COMP1111", "price": "£14.99", "rating": "4.6 out of 5 stars", "reviews": "900 ratings"},
            {"asin": "B0COMP2222", "price": "£15.99", "rating": "4.5 out of 5 stars", "reviews": "700 ratings"},
            {"asin": "B0COMP3333", "price": "£16.99", "rating": "4.4 out of 5 stars", "reviews": "500 ratings"},
        ],
    }

    parsed = parse_amazon_frontend(html, marketplace="UK")
    payload = _frontend_quality_payload(
        parsed,
        {"marketplace": "UK", "product_url": "https://www.amazon.co.uk/dp/B0TEST1234"},
        search_payload,
        "",
        status="已自动检查",
        method="chrome",
        location_note="Aberdeen AB10 1",
    )
    audit = _frontend_evidence_audit(
        {
            "frontend_check_status": "已自动检查",
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            **payload,
        }
    )

    assert payload["frontend_location_scope"] == "marketplace"
    assert payload["frontend_location_verified"] is True
    assert payload["frontend_location_exact"] is False
    assert payload["frontend_evidence_quality_score"] <= 65
    assert audit["frontend_evidence_tier"] == "仅背景参考"
    assert any("地区非配置邮编" in reason for reason in audit["frontend_evidence_audit_reasons"])


def test_frontend_weak_conclusion_never_becomes_strong_evidence() -> None:
    from src.report_presentation import _frontend_evidence_audit

    audit = _frontend_evidence_audit(
        {
            "frontend_check_status": "已自动检查",
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "frontend_evidence_quality_score": 90,
            "frontend_auto_conclusion": "FRONTEND_WEAK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_search_partial_evidence": False,
        }
    )

    assert audit["frontend_evidence_tier"] == "强诊断可用"
    assert audit["frontend_evidence_display_tier"] == "仅背景参考"
    assert audit["frontend_decision_evidence_tier"] == "仅背景参考"
    assert audit["frontend_evidence_is_strong"] is False


def test_frontend_audit_summary_downgrades_when_decision_tier_is_background() -> None:
    from src.report_presentation import _frontend_evidence_audit

    audit = _frontend_evidence_audit(
        {
            "frontend_check_status": "已自动检查",
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "competitor_comparability": "high",
            "comparable_competitor_count": 3,
            "frontend_evidence_quality_score": 72,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_search_partial_evidence": False,
        }
    )

    assert audit["frontend_evidence_tier"] == "强诊断可用"
    assert audit["frontend_evidence_display_tier"] == "仅背景参考"
    assert audit["frontend_evidence_audit_summary"] == "前台证据可辅助判断，不能单独放量"
    assert audit["frontend_evidence_is_strong"] is False


def test_frontend_cache_flag_never_becomes_strong_evidence() -> None:
    from src.report_presentation import _frontend_evidence_audit

    audit = _frontend_evidence_audit(
        {
            "frontend_check_status": "已自动检查",
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "competitor_comparability": "medium",
            "comparable_competitor_count": 3,
            "frontend_evidence_quality_score": 90,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_cache_used": True,
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_search_partial_evidence": False,
        }
    )

    assert audit["frontend_evidence_tier"] == "强诊断可用"
    assert audit["frontend_evidence_display_tier"] == "仅背景参考"
    assert audit["frontend_decision_evidence_tier"] == "仅背景参考"
    assert audit["frontend_evidence_is_strong"] is False
    assert "沿用缓存" in audit["frontend_evidence_audit_reasons"]


def test_frontend_currency_warning_never_sets_strong_flag() -> None:
    from src.report_presentation import _frontend_evidence_audit

    audit = _frontend_evidence_audit(
        {
            "frontend_check_status": "已自动检查",
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "competitor_comparability": "high",
            "comparable_competitor_count": 3,
            "frontend_evidence_quality_score": 90,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_failure_category": "none",
            "frontend_price_currency_warning": "价格币种异常：TWD594.77，已忽略",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_search_partial_evidence": False,
        }
    )

    assert audit["frontend_evidence_tier"] == "不可用"
    assert audit["frontend_evidence_display_tier"] == "不可用"
    assert audit["frontend_decision_evidence_tier"] == "不可用"
    assert audit["frontend_evidence_is_strong"] is False
    assert "币种异常" in audit["frontend_evidence_audit_reasons"]


def test_frontend_failure_category_never_sets_strong_flag() -> None:
    from src.report_presentation import _frontend_evidence_audit

    audit = _frontend_evidence_audit(
        {
            "frontend_check_status": "已自动检查",
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "competitor_comparability": "high",
            "comparable_competitor_count": 3,
            "frontend_evidence_quality_score": 90,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_failure_category": "captcha_or_blocked",
            "frontend_failure_reason": "Amazon 页面触发验证码",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_search_partial_evidence": False,
        }
    )

    assert audit["frontend_evidence_tier"] == "不可用"
    assert audit["frontend_evidence_display_tier"] == "不可用"
    assert audit["frontend_decision_evidence_tier"] == "不可用"
    assert audit["frontend_evidence_is_strong"] is False
    assert "Amazon 页面触发验证码" in audit["frontend_evidence_audit_reasons"]


def test_frontend_low_competitor_comparability_is_not_strong_evidence() -> None:
    from src.report_presentation import _frontend_evidence_audit

    audit = _frontend_evidence_audit(
        {
            "frontend_check_status": "已自动检查",
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "competitor_comparability": "low",
            "comparable_competitor_count": 1,
            "frontend_evidence_quality_score": 90,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_search_partial_evidence": False,
        }
    )

    assert audit["frontend_evidence_tier"] == "强诊断可用"
    assert audit["frontend_evidence_display_tier"] == "仅背景参考"
    assert audit["frontend_decision_evidence_tier"] == "仅背景参考"
    assert audit["frontend_evidence_is_strong"] is False
    assert "竞品可比性不足" in audit["frontend_evidence_audit_reasons"]


def test_frontend_medium_competitor_comparability_is_background_evidence() -> None:
    from src.report_presentation import _frontend_evidence_audit

    audit = _frontend_evidence_audit(
        {
            "frontend_check_status": "已自动检查",
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "competitor_comparability": "medium",
            "comparable_competitor_count": 3,
            "frontend_evidence_quality_score": 90,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_search_partial_evidence": False,
        }
    )

    assert audit["frontend_evidence_tier"] == "强诊断可用"
    assert audit["frontend_evidence_display_tier"] == "仅背景参考"
    assert audit["frontend_decision_evidence_tier"] == "仅背景参考"
    assert audit["frontend_evidence_is_strong"] is False
    assert "竞品可比性未达强诊断" in audit["frontend_evidence_audit_reasons"]


def test_frontend_missing_competitor_comparability_is_not_strong_evidence() -> None:
    from src.report_presentation import _frontend_evidence_audit

    audit = _frontend_evidence_audit(
        {
            "frontend_check_status": "已自动检查",
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "frontend_evidence_quality_score": 90,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_search_partial_evidence": False,
        }
    )

    assert audit["frontend_evidence_tier"] == "强诊断可用"
    assert audit["frontend_evidence_display_tier"] == "仅背景参考"
    assert audit["frontend_decision_evidence_tier"] == "仅背景参考"
    assert audit["frontend_evidence_is_strong"] is False
    assert "竞品可比性未验证" in audit["frontend_evidence_audit_reasons"]


def test_frontend_exact_scope_without_verified_location_is_not_strong_evidence() -> None:
    from src.report_presentation import _frontend_evidence_audit

    audit = _frontend_evidence_audit(
        {
            "frontend_check_status": "已自动检查",
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "frontend_evidence_quality_score": 90,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_exact": True,
            "frontend_search_partial_evidence": False,
        }
    )

    assert audit["frontend_evidence_tier"] == "强诊断可用"
    assert audit["frontend_evidence_display_tier"] == "仅背景参考"
    assert audit["frontend_decision_evidence_tier"] == "仅背景参考"
    assert audit["frontend_evidence_is_strong"] is False
    assert "地区未确认" in audit["frontend_evidence_audit_reasons"]


def test_frontend_partial_search_string_false_does_not_downgrade_strong_evidence() -> None:
    from src.report_presentation import _frontend_evidence_audit

    audit = _frontend_evidence_audit(
        {
            "frontend_check_status": "已自动检查",
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "competitor_comparability": "high",
            "comparable_competitor_count": 3,
            "frontend_evidence_quality_score": 88,
            "frontend_auto_conclusion": "FRONTEND_OK",
            "frontend_failure_category": "none",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "frontend_search_partial_evidence": "False",
        }
    )

    assert audit["frontend_evidence_tier"] == "强诊断可用"
    assert audit["frontend_evidence_display_tier"] == "强诊断可用"
    assert audit["frontend_decision_evidence_tier"] == "强诊断可用"
    assert audit["frontend_evidence_is_strong"] is True
    assert "搜索页仅部分读取" not in audit["frontend_evidence_audit_reasons"]


def test_frontend_location_failure_text_does_not_count_as_exact_location() -> None:
    from scripts.run_frontend_checks import _location_exact, _location_scope, _location_verified

    note = "UK SW1A 1AA 未确认：timeout"

    assert _location_scope(note, "UK") == "missing"
    assert _location_verified(note, "UK") is False
    assert _location_exact(note, "UK") is False


def test_apply_delivery_location_reads_back_visible_location_when_postcode_not_confirmed(monkeypatch) -> None:
    import scripts.run_frontend_checks as frontend

    class FakeLocator:
        @property
        def first(self):
            return self

        def click(self, timeout=None):
            return None

        def fill(self, value, timeout=None):
            return None

        def inner_text(self, timeout=None):
            return "Deliver to Aberdeen AB10 1"

    class FakePage:
        def locator(self, selector):
            return FakeLocator()

        def wait_for_timeout(self, timeout):
            return None

        def reload(self, wait_until=None, timeout=None):
            return None

        def wait_for_load_state(self, state, timeout=None):
            return None

    note = frontend._apply_delivery_location(FakePage(), "UK", "SW1A 1AA")

    assert note == "Aberdeen AB10 1"
    assert frontend._location_scope(note, "UK") == "marketplace"


def test_frontend_location_basis_clears_stale_visible_location_when_no_basis() -> None:
    import scripts.run_frontend_checks as frontend

    parsed = {
        "visible_location": "United States",
        "location_warning": "UK 地区异常：United States",
        "location_hard_failure": "UK 地区异常：United States",
        "location_scope": "wrong",
    }

    note = frontend._apply_location_basis(parsed, "", "UK", allow_html_visible=False)

    assert note == ""
    assert parsed.get("visible_location") is None
    assert parsed["location_scope"] == "missing"
    assert parsed["location_warning"] == "UK 地区未确认"
    assert parsed["location_hard_failure"] == "UK 地区未确认"


def test_frontend_report_audit_blocks_unverified_location() -> None:
    from src.report_presentation import _frontend_evidence_audit

    audit = _frontend_evidence_audit(
        {
            "frontend_check_status": "已自动检查",
            "frontend_evidence_quality_score": 88,
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "frontend_location_verified": False,
        }
    )

    assert audit["frontend_evidence_tier"] == "不可用"
    assert "地区未确认" in audit["frontend_evidence_audit_reasons"]


def test_safe_target_acos_keeps_profit_buffer_and_caps_high_margin_products() -> None:
    from scripts.sync_frontend_prices import safe_target_acos_from_break_even

    assert round(safe_target_acos_from_break_even(0.545716), 6) == 0.272858
    assert safe_target_acos_from_break_even(0.80) == 0.30
    assert safe_target_acos_from_break_even(0.20) == 0.10
    assert safe_target_acos_from_break_even(0) == 0.0
    assert safe_target_acos_from_break_even(-0.10) == 0.0


def test_sync_frontend_prices_rejects_blocked_currency_markers() -> None:
    from scripts.sync_frontend_prices import _parse_money

    assert _parse_money("JPY4,658") is None
    assert _parse_money("CAD 20.95") is None
    assert _parse_money("A$12.99") is None
    assert _parse_money("£14.99") == ("£", 14.99)


def test_frontend_check_findings_hide_technical_fetch_errors() -> None:
    from scripts.run_frontend_checks import _findings

    message = _findings({}, "<urlopen error [Errno 8] nodename nor servname provided, or not known>")

    assert "urlopen" not in message
    assert "nodename" not in message
    assert "自动证据不足，不能用于强诊断" in message


def test_frontend_suspected_issue_uses_frontend_evidence() -> None:
    from scripts.run_frontend_checks import _suspected_issue

    issue = _suspected_issue(
        {
            "rating": "4.0 out of 5 stars",
            "reviews": "(42)",
            "coupon": "11%",
            "buy_box": "识别到购买按钮",
            "price": "£15.60",
        },
        "疑似广告流量不准或价格竞争力不足",
        {
            "trigger_reason": "近14天广告无单且点击>=20",
            "key_metrics": "近14天总单 2；广告订单 0；点击 31",
        },
    )

    assert "前台基础项未见明显硬伤" in issue
    assert "广告搜索词或ASIN定向" in issue


def test_frontend_cache_fallback_reuses_last_success(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                        "frontend_check_queue_rows": [
                            {
                                "marketplace": "UK",
                                "sku": "SKU-1",
                                "asin": "B0TEST1234",
                                "product_name": "Test product",
                                "product_url": "https://www.amazon.co.uk/dp/B0TEST1234?th=1",
                                "frontend_core_keyword": "dimmer desk lamp",
                                "frontend_search_url": "https://www.amazon.co.uk/s?k=dimmer+board",
                            }
                        ]
                    }
                }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(
        json.dumps(
            {
                "items": [],
                "cache": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-1",
                        "asin": "B0TEST1234",
                        "frontend_check_status": "已自动检查",
                        "frontend_findings": "售价：£14.99；评分：4.5 out of 5 stars",
                        "frontend_location_note": "UK SW1A 1AA 已设置",
                        "frontend_location_verified": True,
                        "suspected_issue": "历史前台判断",
                        "checked_at": "2026-05-14T09:00:00",
                    }
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)
    def fake_fetch(url, *args, **kwargs):
        if "/s?" in url:
            return (
                '<div data-asin="B0COMP1111" class="s-result-item"><h2><span>Competitor</span></h2><span class="a-offscreen">£11.99</span></div>',
                "",
                "playwright",
                "",
            )
        return "", "network failed", "playwright", ""

    monkeypatch.setattr(frontend, "_fetch_html", fake_fetch)

    rows = frontend.check_frontend_queue(method="urllib", timeout=1, sleep_seconds=0, retries=2)

    assert rows[0]["frontend_check_status"] == "沿用 2026-05-14 前台数据"
    assert "售价：£14.99" in rows[0]["frontend_findings"]
    assert rows[0]["frontend_cache_used"] is True
    assert rows[0]["frontend_search_status"] == "已自动检查"
    assert rows[0]["frontend_competitor_count"] == 1


def test_frontend_cache_fallback_rejects_wrong_currency_cache(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "US",
                                    "sku": "SKU-1",
                                    "asin": "B0TEST1234",
                                    "product_name": "Test product",
                                    "product_url": "https://www.amazon.com/dp/B0TEST1234",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(
        json.dumps(
            {
                "items": [],
                "cache": [
                    {
                        "marketplace": "US",
                        "sku": "SKU-1",
                        "asin": "B0TEST1234",
                        "frontend_check_status": "已自动检查",
                        "frontend_findings": "价格币种异常：TWD594.77，已忽略；评分：4.2 out of 5 stars",
                        "frontend_price_currency_warning": "价格币种异常：TWD594.77，已忽略",
                        "suspected_issue": "历史前台判断",
                        "checked_at": "2026-06-02T09:00:00",
                    }
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)
    monkeypatch.setattr(frontend, "_fetch_html", lambda *args, **kwargs: ("", "dns failed", "urllib", ""))

    rows = frontend.check_frontend_queue(timeout=1, sleep_seconds=0, retries=1)

    assert rows[0]["frontend_check_status"] == "待前台检查"
    assert "TWD594.77" not in rows[0]["frontend_findings"]
    assert "自动证据不足，不能用于强诊断" in rows[0]["frontend_findings"]


def test_chrome_cdp_frontend_check_reads_search_competitor_samples(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "priority": "P0",
                                    "marketplace": "US",
                                    "sku": "SKU-P0",
                                    "asin": "B0P0TEST01",
                                    "product_name": "P0 product",
                                    "product_url": "https://www.amazon.com/dp/B0P0TEST01",
                                    "frontend_core_keyword": "desk lamp",
                                    "frontend_search_url": "https://www.amazon.com/s?k=desk+board",
                                    "trigger_reason": "近7天广告无单",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(json.dumps({"items": [], "cache": []}), encoding="utf-8")
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)
    monkeypatch.setattr(frontend, "_today_iso", lambda: "2026-06-17")
    monkeypatch.setattr(
        frontend,
        "run_chrome_cdp_probe",
        lambda **kwargs: {
            "attempts": [
                {
                    "attempt": 1,
                    "marketplace": "US",
                    "asin": "B0P0TEST01",
                    "success": True,
                    "title": "P0 product",
                    "price": "$8.99",
                    "rating": "3.9 out of 5 stars",
                    "reviews": "(52)",
                    "location": "US 10001 已设置",
                    "buy_box": "识别到购买按钮",
                }
            ]
        },
    )
    search_calls: list[str] = []

    def fake_search(url, timeout, marketplace, own_asin, endpoint, limit=3):
        search_calls.append(url)
        return (
            {
                "competitors": [
                    {"asin": "B0COMP1111", "price": "$10.99", "rating": "4.5 out of 5 stars", "reviews": "1,200 ratings"},
                    {"asin": "B0COMP2222", "price": "$11.99", "rating": "4.4 out of 5 stars", "reviews": "900 ratings"},
                ],
                "own_search_position": "4",
                "search_result_count": 8,
            },
            "",
            "chrome-cdp-dom",
        )

    monkeypatch.setattr(frontend, "_parse_search_results_chrome_cdp", fake_search)

    rows = frontend.check_frontend_queue(
        timeout=1,
        sleep_seconds=0,
        method="chrome-cdp",
        cdp_attempts=1,
        search_policy="always",
        priority="P0",
        require_competitor_samples=True,
    )

    assert search_calls == ["https://www.amazon.com/s?k=desk+board"]
    assert rows[0]["frontend_check_status"] == "已自动检查"
    assert rows[0]["frontend_search_status"] == "已自动检查"
    assert rows[0]["frontend_competitor_count"] == 2
    assert rows[0]["frontend_rating_delta"] < 0


def test_chrome_cdp_only_stale_refreshes_when_competitor_samples_are_missing(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "priority": "P0",
                                    "marketplace": "US",
                                    "sku": "SKU-P0",
                                    "asin": "B0P0TEST01",
                                    "product_name": "P0 product",
                                    "product_url": "https://www.amazon.com/dp/B0P0TEST01",
                                    "frontend_core_keyword": "desk lamp",
                                    "frontend_search_url": "https://www.amazon.com/s?k=desk+board",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(
        json.dumps(
            {
                "items": [
                    {
                        "marketplace": "US",
                        "sku": "SKU-P0",
                        "asin": "B0P0TEST01",
                        "frontend_check_status": "已自动检查",
                        "frontend_check_method": "chrome-cdp",
                        "frontend_data_date": "2026-06-17",
                        "frontend_findings": "售价：$8.99；评分：3.9 out of 5 stars",
                        "frontend_price": "$8.99",
                        "frontend_rating": "3.9 out of 5 stars",
                        "frontend_reviews": "(52)",
                        "frontend_location_note": "US 10001 已设置",
                        "frontend_competitor_count": 0,
                        "frontend_competitors": [],
                    }
                ],
                "cache": [],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)
    monkeypatch.setattr(frontend, "_today_iso", lambda: "2026-06-17")
    probe_calls: list[str] = []

    def fake_probe(**kwargs):
        probe_calls.append(kwargs["asin"])
        return {
            "attempts": [
                {
                    "attempt": 1,
                    "marketplace": "US",
                    "asin": "B0P0TEST01",
                    "success": True,
                    "title": "P0 product",
                    "price": "$8.99",
                    "rating": "3.9 out of 5 stars",
                    "reviews": "(52)",
                    "location": "US 10001 已设置",
                    "buy_box": "识别到购买按钮",
                }
            ]
        }

    monkeypatch.setattr(frontend, "run_chrome_cdp_probe", fake_probe)
    monkeypatch.setattr(
        frontend,
        "_parse_search_results_chrome_cdp",
        lambda *args, **kwargs: (
            {
                "competitors": [
                    {"asin": "B0COMP1111", "price": "$10.99", "rating": "4.5 out of 5 stars", "reviews": "1,200 ratings"},
                    {"asin": "B0COMP2222", "price": "$11.99", "rating": "4.4 out of 5 stars", "reviews": "900 ratings"},
                ],
                "search_result_count": 8,
            },
            "",
            "chrome-cdp-dom",
        ),
    )

    rows = frontend.check_frontend_queue(
        timeout=1,
        sleep_seconds=0,
        method="chrome-cdp",
        cdp_attempts=1,
        search_policy="always",
        priority="P0",
        only_stale=True,
        require_competitor_samples=True,
    )

    assert probe_calls == ["B0P0TEST01"]
    assert rows[0]["frontend_refresh_action"] == "live_checked"
    assert rows[0]["frontend_competitor_count"] == 2


def test_frontend_current_wrong_currency_is_not_success(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "DE",
                                    "sku": "SKU-2",
                                    "asin": "B0TEST5678",
                                    "product_name": "Test product",
                                    "product_url": "https://www.amazon.de/dp/B0TEST5678",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(json.dumps({"items": [], "cache": []}), encoding="utf-8")
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)
    monkeypatch.setattr(
        frontend,
        "_fetch_html",
        lambda *args, **kwargs: (
            """
            <span id="productTitle">DE product</span>
            <span class="a-offscreen">TWD432.81</span>
            <span class="a-icon-alt">4.2 out of 5 stars</span>
            """,
            "",
            "urllib",
            "",
        ),
    )

    rows = frontend.check_frontend_queue(timeout=1, sleep_seconds=0, retries=1)

    assert rows[0]["frontend_check_status"] == "待前台检查"
    assert "TWD432.81" in rows[0]["frontend_findings"]
    assert "价格币种异常" in rows[0]["frontend_last_error"]
    assert rows[0]["frontend_price"] == ""


def test_frontend_retry_keeps_strongest_failure_signal(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-UK",
                                    "asin": "B0TESTUK12",
                                    "product_name": "UK product",
                                    "product_url": "https://www.amazon.co.uk/dp/B0TESTUK12?th=1",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(json.dumps({"items": [], "cache": []}), encoding="utf-8")
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)

    responses = iter(
        [
            (
                """
                <span id="productTitle">UK product</span>
                <span id="glow-ingress-line2">Germany</span>
                <span class="a-offscreen">EUR20.70</span>
                <span class="a-icon-alt">4.2 out of 5 stars</span>
                """,
                "",
                "urllib",
                "",
            ),
            ("", "", "urllib", ""),
        ]
    )
    monkeypatch.setattr(frontend, "_fetch_html", lambda *args, **kwargs: next(responses))

    rows = frontend.check_frontend_queue(timeout=1, sleep_seconds=0, retries=2, search_policy="never")

    assert rows[0]["frontend_check_status"] == "待前台检查"
    assert "价格币种异常" in rows[0]["frontend_last_error"]
    assert "EUR20.70" in rows[0]["frontend_last_error"]
    assert "UK 地区异常：Germany" in rows[0]["frontend_last_error"]
    assert rows[0]["frontend_failure_category"] == "currency_mismatch"
    assert rows[0]["frontend_evidence_quality_score"] == 0


def test_frontend_merged_items_drop_previous_currency_contamination(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    results = tmp_path / "frontend_check_results.json"
    results.write_text(
        json.dumps(
            {
                "items": [
                    {
                        "marketplace": "DE",
                        "sku": "SKU-OLD",
                        "asin": "B0OLD12345",
                        "frontend_check_status": "已自动检查",
                        "frontend_findings": "价格币种异常：TWD432.81，已忽略；评分：4.2 out of 5 stars",
                        "frontend_price_currency_warning": "价格币种异常：TWD432.81，已忽略",
                    },
                    {
                        "marketplace": "UK",
                        "sku": "SKU-RAW",
                        "asin": "B0RAW12345",
                        "frontend_check_status": "已自动检查",
                        "frontend_findings": "售价：TWD761.55；评分：4.2 out of 5 stars",
                        "frontend_price": "TWD761.55",
                    },
                    {
                        "marketplace": "DE",
                        "sku": "SKU-GOOD",
                        "asin": "B0GOOD1234",
                        "frontend_check_status": "已自动检查",
                        "frontend_findings": "售价：€12.99；评分：4.5 out of 5 stars",
                        "frontend_location_note": "DE 10115 已设置",
                        "frontend_location_verified": True,
                    },
                    {
                        "marketplace": "UK",
                        "sku": "SKU-NO-LOCATION",
                        "asin": "B0NOLOC123",
                        "frontend_check_status": "已自动检查",
                        "frontend_findings": "售价：£12.99；评分：4.5 out of 5 stars",
                    },
                ],
                "cache": [],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)

    merged = frontend._merged_items([])

    assert [row["asin"] for row in merged] == ["B0GOOD1234"]


def test_frontend_merged_items_keep_previous_success_when_current_probe_fails(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    results = tmp_path / "frontend_check_results.json"
    results.write_text(
        json.dumps(
            {
                "items": [
                    {
                        "marketplace": "US",
                        "sku": "SKU-1",
                        "asin": "B0TEST1234",
                        "frontend_check_status": "已自动检查",
                        "frontend_findings": "售价：$16.79；评分：3.9 out of 5 stars；配送：US 10001 已设置",
                        "frontend_price": "$16.79",
                        "frontend_location_note": "US 10001 已设置",
                        "frontend_check_method": "chrome-cdp",
                        "frontend_stability_total_attempts": 20,
                        "frontend_stability_success_count": 20,
                        "frontend_stability_failure_count": 0,
                        "frontend_stability_success_rate": 1.0,
                        "frontend_stability_passed": True,
                    }
                ],
                "cache": [],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)

    merged = frontend._merged_items(
        [
            {
                "marketplace": "US",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "frontend_check_status": "待前台检查",
                "frontend_findings": "自动证据不足，不能用于强诊断；Chrome CDP 20次验收未通过：0/20，失败 20",
                "frontend_check_method": "chrome-cdp",
                "frontend_stability_total_attempts": 20,
                "frontend_stability_success_count": 0,
                "frontend_stability_failure_count": 20,
                "frontend_stability_success_rate": 0,
                "frontend_stability_passed": False,
            }
        ]
    )

    assert len(merged) == 1
    assert merged[0]["frontend_check_status"] == "已自动检查"
    assert merged[0]["frontend_stability_success_count"] == 20
    assert merged[0]["frontend_stability_passed"] is True


def test_frontend_delivery_currency_contamination_is_not_success() -> None:
    from scripts.run_frontend_checks import _status_from_parsed

    parsed = {
        "title": "UK product",
        "rating": "4.2 out of 5 stars",
        "delivery": "TWD 495.16 delivery Thursday, 18 June . Details",
        "captcha_or_block": "否",
    }

    assert _status_from_parsed(parsed) == "待前台检查"


def test_frontend_no_cache_asks_for_first_manual_check(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "DE",
                                    "sku": "SKU-2",
                                    "asin": "B0TEST5678",
                                    "product_name": "Test product",
                                    "product_url": "https://www.amazon.de/dp/B0TEST5678",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)
    monkeypatch.setattr(frontend, "_fetch_html", lambda *args, **kwargs: ("", "blocked", "playwright", ""))

    rows = frontend.check_frontend_queue(timeout=1, sleep_seconds=0, retries=1)

    assert rows[0]["frontend_check_status"] == "待前台检查"
    assert "自动证据不足，不能用于强诊断" in rows[0]["frontend_findings"]
    assert rows[0]["frontend_auto_conclusion_label"] == "自动证据不足，不能用于强诊断"


def test_frontend_missing_core_fields_are_insufficient_not_weak() -> None:
    from src.report_presentation import _derive_frontend_display_quality

    payload = _derive_frontend_display_quality(
        {
            "marketplace": "US",
            "frontend_check_status": "待前台检查",
            "frontend_findings": "自动证据不足，不能用于强诊断；尚未读取到可用前台字段。",
        }
    )

    assert payload["frontend_auto_conclusion_label"] == "自动证据不足，不能用于强诊断"
    assert payload["frontend_auto_conclusion"] == "INSUFFICIENT_EVIDENCE"
    assert payload["frontend_failure_category"] == "missing_core_fields"
    assert "price_missing" in payload["frontend_product_quality_flags"]
    assert "rating_missing" in payload["frontend_product_quality_flags"]


def test_frontend_auto_conclusion_scores_current_product_and_search_page() -> None:
    from scripts.run_frontend_checks import _frontend_quality_payload

    parsed = {
        "price": "£18.99",
        "rating": "3.8 out of 5 stars",
        "reviews": "12 ratings",
        "coupon": "未稳定识别 Coupon",
        "buy_box": "未稳定识别 Buy Box/购买按钮",
        "delivery": "FREE delivery",
        "captcha_or_block": "否",
    }
    search_payload = {
        "frontend_search_status": "已自动检查",
        "frontend_competitors": [
            {"asin": "B0COMP1111", "price": "£14.99", "rating": "4.5 out of 5 stars", "reviews": "1,200 ratings"},
            {"asin": "B0COMP2222", "price": "£15.99", "rating": "4.4 out of 5 stars", "reviews": "900 ratings"},
            {"asin": "B0COMP3333", "price": "£16.99", "rating": "4.3 out of 5 stars", "reviews": "600 ratings"},
        ],
        "own_search_position": "18",
    }

    payload = _frontend_quality_payload(
        parsed,
        {"marketplace": "UK", "product_url": "https://www.amazon.co.uk/dp/B0TEST1234"},
        search_payload,
        "",
        status="已自动检查",
        location_note="UK SW1A 1AA 已设置",
    )

    assert payload["frontend_auto_conclusion_label"] == "明确前台劣势"
    assert payload["frontend_evidence_quality_score"] > 0
    assert "price_gap" in payload["frontend_search_quality_flags"]
    assert payload["frontend_price_delta_pct"] > 0


def test_frontend_browser_session_reuses_marketplace_context_and_closes(monkeypatch) -> None:
    import scripts.run_frontend_checks as frontend

    events: list[tuple[str, str]] = []

    class FakePage:
        def goto(self, url, wait_until=None, timeout=None):
            events.append(("goto", url))

        def close(self):
            events.append(("page_close", ""))

    class FakeContext:
        def __init__(self, name: str):
            self.name = name

        def new_page(self):
            events.append(("new_page", self.name))
            return FakePage()

        def close(self):
            events.append(("context_close", self.name))

    class FakeBrowser:
        def __init__(self):
            self.count = 0

        def new_context(self, **kwargs):
            self.count += 1
            name = f"context-{self.count}"
            events.append(("new_context", name))
            return FakeContext(name)

        def close(self):
            events.append(("browser_close", ""))

    class FakeManager:
        def stop(self):
            events.append(("manager_stop", ""))

    fake_browser = FakeBrowser()
    def fake_ensure_browser(self):
        self._browser = fake_browser
        return fake_browser

    monkeypatch.setattr(frontend.FrontendBrowserSession, "_ensure_browser", fake_ensure_browser)
    monkeypatch.setattr(frontend, "_apply_delivery_location", lambda page, marketplace, postcode: f"{marketplace}:{postcode}")

    session = frontend.FrontendBrowserSession()
    session._playwright_manager = FakeManager()
    context_1 = session.get_context("UK")
    context_2 = session.get_context("UK")
    context_3 = session.get_context("DE")
    session.close()

    assert context_1 is context_2
    assert context_1 is not context_3
    assert [event for event in events if event[0] == "new_context"] == [("new_context", "context-1"), ("new_context", "context-2")]
    assert ("browser_close", "") in events
    assert ("manager_stop", "") in events


def test_parse_search_results_preserves_partial_dom_evidence_when_fallback_empty(monkeypatch) -> None:
    import scripts.run_frontend_checks as frontend

    monkeypatch.setattr(
        frontend,
        "_parse_amazon_search_results_playwright",
        lambda *args, **kwargs: (
            {"search_result_count": 10, "own_search_position": "5", "competitors": []},
            "dom partial",
            "playwright-dom",
        ),
    )
    monkeypatch.setattr(frontend, "_fetch_html", lambda *args, **kwargs: ("", "urllib failed", "urllib", ""))

    parsed, error, parser = frontend._parse_search_results(
        "https://www.amazon.co.uk/s?k=test",
        10,
        "auto",
        "UK",
        "B0TEST1234",
        1,
    )

    assert parsed["search_result_count"] == 10
    assert parsed["own_search_position"] == "5"
    assert parsed["competitors"] == []
    assert error == "urllib failed"
    assert parser == "urllib"


def test_record_from_cache_keeps_stronger_search_status(monkeypatch) -> None:
    import scripts.run_frontend_checks as frontend

    cached = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0TEST1234",
        "frontend_check_status": "已自动检查",
        "frontend_findings": "售价：£14.99；评分：4.5 out of 5 stars",
        "frontend_search_status": "已自动检查",
        "frontend_search_findings": "历史强证据",
        "frontend_competitors": [{"asin": "B0COMP1111"}],
        "frontend_location_note": "UK SW1A 1AA 已设置",
        "frontend_location_verified": True,
        "checked_at": "2026-06-01T10:00:00",
    }

    row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0TEST1234",
        "product_name": "Test product",
    }

    merged = frontend._record_from_cache(
        row,
        cached,
        "network failed",
        extra={
            "frontend_search_status": "待前台检查",
            "frontend_search_findings": "本次没读到",
            "frontend_competitors": [],
        },
    )

    assert merged["frontend_search_status"] == "已自动检查"
    assert merged["frontend_search_findings"] == "历史强证据"
    assert merged["frontend_competitors"] == [{"asin": "B0COMP1111"}]


def test_record_from_cache_downgrades_search_evidence_when_context_changes(monkeypatch) -> None:
    import scripts.run_frontend_checks as frontend

    cached = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0TEST1234",
        "frontend_check_status": "已自动检查",
        "frontend_findings": "售价：£14.99；评分：4.5 out of 5 stars",
        "frontend_core_keyword": "old keyword",
        "frontend_search_keyword": "old keyword",
        "frontend_search_url": "https://www.amazon.co.uk/s?k=old",
        "frontend_search_status": "已自动检查",
        "frontend_search_findings": "历史强证据",
        "frontend_competitor_count": 1,
        "frontend_competitors": [{"asin": "B0OLD11111"}],
        "frontend_price_delta_pct": 0.18,
        "frontend_rating_delta": -0.4,
        "frontend_review_delta_pct": -0.7,
        "frontend_competitor_price_median": 17.99,
        "frontend_competitor_rating_avg": 4.7,
        "frontend_competitor_review_median": 1200,
        "own_search_position": "2",
        "frontend_location_note": "UK SW1A 1AA 已设置",
        "frontend_location_verified": True,
        "checked_at": "2026-06-01T10:00:00",
    }

    row = {
        "marketplace": "UK",
        "sku": "SKU-1",
        "asin": "B0TEST1234",
        "product_name": "Test product",
        "frontend_core_keyword": "new keyword",
        "frontend_search_url": "https://www.amazon.co.uk/s?k=new",
    }

    merged = frontend._record_from_cache(
        row,
        cached,
        "network failed",
        extra={
            "frontend_search_status": "待前台检查",
            "frontend_search_keyword": "new keyword",
            "frontend_search_url": "https://www.amazon.co.uk/s?k=new",
            "frontend_search_findings": "本次没读到",
            "frontend_competitors": [],
        },
    )

    assert merged["frontend_search_status"] == "待前台检查"
    assert merged["frontend_search_keyword"] == "new keyword"
    assert merged["frontend_search_url"] == "https://www.amazon.co.uk/s?k=new"
    assert "历史搜索页证据仅作背景参考" in merged["frontend_search_findings"]
    assert "本次没读到" in merged["frontend_search_findings"]
    assert merged["frontend_competitors"] == []
    assert merged["frontend_competitor_count"] == 0
    assert merged["frontend_price_delta_pct"] == ""
    assert merged["frontend_rating_delta"] == ""
    assert merged["frontend_review_delta_pct"] == ""
    assert merged["frontend_competitor_price_median"] == ""
    assert merged["frontend_competitor_rating_avg"] == ""
    assert merged["frontend_competitor_review_median"] == ""


def test_check_frontend_queue_reuses_browser_session_and_closes_on_error(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-1",
                                    "asin": "B0TEST1234",
                                    "product_name": "Test product",
                                    "product_url": "https://www.amazon.co.uk/dp/B0TEST1234?th=1",
                                    "frontend_core_keyword": "dimmer desk lamp",
                                    "frontend_search_url": "https://www.amazon.co.uk/s?k=dimmer+board",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(json.dumps({"items": [], "cache": []}), encoding="utf-8")
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)

    events: list[tuple[str, object]] = []

    class FakeSession:
        def __init__(self, *, use_chrome=False, persistent=False, profile_root=None, headless=False, apply_location=True):
            events.append(("init", use_chrome))

        def close(self):
            events.append(("close", None))

    monkeypatch.setattr(frontend, "FrontendBrowserSession", FakeSession)

    def fake_fetch_html(url, timeout, method="auto", marketplace="", attempt=1, browser_session=None):
        events.append(("fetch", browser_session.__class__.__name__ if browser_session else None))
        return (
            '<span id="productTitle">UK product</span><span class="a-offscreen">£19.99</span><span class="a-icon-alt">4.3 out of 5 stars</span>',
            "",
            "playwright",
            "",
        )

    monkeypatch.setattr(frontend, "_fetch_html", fake_fetch_html)

    def fake_parse_search(url, timeout, method, marketplace, own_asin, attempt, browser_session=None):
        events.append(("search", browser_session.__class__.__name__ if browser_session else None))
        raise RuntimeError("boom")

    monkeypatch.setattr(frontend, "_parse_search_results", fake_parse_search)

    try:
        frontend.check_frontend_queue(timeout=1, sleep_seconds=0, retries=1, reuse_browser=True, method="playwright")
    except RuntimeError:
        pass

    assert ("init", False) in events
    assert ("fetch", "FakeSession") in events
    assert ("search", "FakeSession") in events
    assert ("close", None) in events


def test_check_frontend_queue_uses_persistent_chrome_session(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-1",
                                    "asin": "B0TEST1234",
                                    "product_name": "Test product",
                                    "product_url": "https://www.amazon.co.uk/dp/B0TEST1234?th=1",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(json.dumps({"items": [], "cache": []}), encoding="utf-8")
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)

    events: list[tuple[str, object]] = []

    class FakeSession:
        def __init__(self, *, use_chrome=False, persistent=False, profile_root=None, headless=False, apply_location=True):
            events.append(("init", use_chrome, persistent, bool(profile_root), headless, apply_location))

        def close(self):
            events.append(("close", None))

    monkeypatch.setattr(frontend, "FrontendBrowserSession", FakeSession)
    monkeypatch.setattr(
        frontend,
        "_fetch_html",
        lambda *args, **kwargs: (
            '<span id="productTitle">UK product</span><span class="a-offscreen">£19.99</span><span class="a-icon-alt">4.3 out of 5 stars</span>',
            "",
            "chrome-persistent",
            "UK SW1A 1AA 已设置",
        ),
    )

    rows = frontend.check_frontend_queue(
        timeout=1,
        sleep_seconds=0,
        retries=1,
        reuse_browser=True,
        method="chrome-persistent",
        search_policy="never",
    )

    assert ("init", True, True, False, False, False) in events
    assert ("close", None) in events
    assert rows[0]["frontend_check_method"] == "chrome-persistent"


def test_import_manual_frontend_evidence_writes_success_cache(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-1",
                                    "asin": "B0TEST1234",
                                    "product_name": "Test product",
                                    "product_url": "https://www.amazon.co.uk/dp/B0TEST1234?th=1",
                                    "frontend_core_keyword": "spiral notebook",
                                    "frontend_search_url": "https://www.amazon.co.uk/s?k=tea+box",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(json.dumps({"items": [], "cache": []}), encoding="utf-8")
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)

    row = frontend.import_manual_frontend_evidence(
        marketplace="UK",
        sku="SKU-1",
        asin="B0TEST1234",
        title="Tea organizer",
        price="£17.89",
        rating="4.2 out of 5 stars",
        reviews="(38)",
        location_note="Aberdeen AB10 1",
        method="chrome-extension",
        stability_total_attempts=20,
        stability_success_count=19,
        stability_failure_count=1,
        stability_success_rate=0.95,
        stability_passed=True,
    )

    payload = json.loads(results.read_text(encoding="utf-8"))
    assert row["frontend_check_status"] == "已自动检查"
    assert row["frontend_price"] == "£17.89"
    assert row["frontend_check_method"] == "chrome-extension"
    assert row["frontend_stability_total_attempts"] == 20
    assert row["frontend_stability_success_count"] == 19
    assert row["frontend_stability_failure_count"] == 1
    assert row["frontend_stability_success_rate"] == 0.95
    assert row["frontend_stability_passed"] is True
    assert row["frontend_location_scope"] == "marketplace"
    assert payload["items"][0]["asin"] == "B0TEST1234"
    assert payload["cache"][0]["frontend_price"] == "£17.89"
    assert payload["cache"][0]["frontend_stability_passed"] is True


def test_check_frontend_queue_urllib_mode_does_not_create_browser_session(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-1",
                                    "asin": "B0TEST1234",
                                    "product_name": "Test product",
                                    "product_url": "https://www.amazon.co.uk/dp/B0TEST1234?th=1",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(json.dumps({"items": [], "cache": []}), encoding="utf-8")
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)

    def fail_init(*args, **kwargs):
        raise AssertionError("browser session should not be created in urllib mode")

    monkeypatch.setattr(frontend, "FrontendBrowserSession", fail_init)
    monkeypatch.setattr(
        frontend,
        "_fetch_html",
        lambda *args, **kwargs: (
            '<span id="productTitle">UK product</span><span class="a-offscreen">£19.99</span><span class="a-icon-alt">4.3 out of 5 stars</span>',
            "",
            "urllib",
            "",
        ),
    )

    rows = frontend.check_frontend_queue(timeout=1, sleep_seconds=0, retries=1, reuse_browser=True, method="urllib")

    assert rows[0]["frontend_check_method"] == "urllib"


def test_check_frontend_queue_chrome_cdp_single_attempt_is_operational_refresh(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-1",
                                    "asin": "B0TEST1234",
                                    "product_name": "Test product",
                                    "product_url": "https://www.amazon.co.uk/dp/B0TEST1234?th=1",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(json.dumps({"items": [], "cache": []}), encoding="utf-8")
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)
    captured_kwargs: dict[str, object] = {}

    def fake_probe(**kwargs):
        captured_kwargs.update(kwargs)
        return {
            "generated_at": "2026-06-16T23:30:00",
            "attempts": [
                {
                    "attempt": 1,
                    "marketplace": "UK",
                    "asin": "B0TEST1234",
                    "success": True,
                    "title": "UK product",
                    "price": "£19.99",
                    "rating": "4.3 out of 5 stars",
                    "reviews": "(38)",
                    "location": "London SW1A 1AA",
                    "buy_box": "识别到购买按钮",
                }
            ],
        }

    monkeypatch.setattr(frontend, "run_chrome_cdp_probe", fake_probe)

    rows = frontend.check_frontend_queue(
        timeout=1,
        sleep_seconds=0,
        retries=1,
        method="chrome-cdp",
        search_policy="never",
        cdp_endpoint="http://127.0.0.1:9222",
        cdp_attempts=1,
    )

    assert captured_kwargs["attempts"] == 1
    assert rows[0]["frontend_check_status"] == "已自动检查"
    assert rows[0]["frontend_check_method"] == "chrome-cdp"
    assert rows[0]["frontend_price"] == "£19.99"
    assert "frontend_stability_total_attempts" not in rows[0]
    assert "frontend_stability_passed" not in rows[0]


def test_chrome_cdp_probe_treats_string_false_success_as_failed(monkeypatch) -> None:
    import scripts.run_frontend_checks as frontend

    row = {
        "marketplace": "UK",
        "sku": "SKU-CDP-FALSE",
        "asin": "B0CDPFALSE",
        "product_url": "https://www.amazon.co.uk/dp/B0CDPFALSE?th=1",
    }

    monkeypatch.setattr(
        frontend,
        "run_chrome_cdp_probe",
        lambda **kwargs: {
            "attempts": [
                {
                    "attempt": 1,
                    "success": "False",
                    "title": "False success product",
                    "price": "£19.99",
                    "rating": "4.7 out of 5 stars",
                    "reviews": "999 ratings",
                    "location": "London SW1A 1AA",
                    "captcha_or_block": "False",
                }
            ]
        },
    )

    result = frontend._result_row_from_cdp_probe(
        row,
        endpoint="http://127.0.0.1:9222",
        timeout=1,
        attempts=1,
        sleep_seconds=0,
        search_policy="never",
    )

    assert result["frontend_check_status"] == "待前台检查"
    assert result["frontend_check_method"] == "chrome-cdp"
    assert result.get("frontend_price", "") == ""
    assert "Chrome CDP 实时读取失败" in result["frontend_last_error"]


def test_chrome_cdp_stability_treats_string_false_passed_as_failed(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    row = {
        "marketplace": "UK",
        "sku": "SKU-CDP-STABILITY",
        "asin": "B0CDPSTABLE",
        "product_url": "https://www.amazon.co.uk/dp/B0CDPSTABLE?th=1",
    }

    monkeypatch.setattr(frontend, "OUTPUT_DIR", tmp_path)
    monkeypatch.setattr(
        frontend,
        "run_chrome_cdp_probe",
        lambda **kwargs: {
            "attempts": [
                {
                    "attempt": 1,
                    "success": "True",
                    "title": "Stability product",
                    "price": "£19.99",
                    "rating": "4.7 out of 5 stars",
                    "reviews": "999 ratings",
                    "location": "London SW1A 1AA",
                }
            ]
        },
    )
    monkeypatch.setattr(
        frontend,
        "build_stability_report",
        lambda *args, **kwargs: {
            "passed": "False",
            "total_attempts": 20,
            "success_count": 20,
            "failure_count": 0,
            "success_rate": 1.0,
            "attempts": [{"attempt": 1, "success": "True"}],
        },
    )

    result = frontend._result_row_from_cdp_probe(
        row,
        endpoint="http://127.0.0.1:9222",
        timeout=1,
        attempts=20,
        sleep_seconds=0,
        search_policy="never",
    )

    assert result["frontend_check_status"] == "待前台检查"
    assert result["frontend_stability_passed"] is False
    assert result.get("frontend_price", "") == ""
    assert "Chrome CDP 20次验收未通过" in result["frontend_last_error"]


def test_check_frontend_queue_only_stale_skips_today_chrome_cdp_rows(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-FRESH",
                                    "asin": "B0FRESH001",
                                    "product_name": "Fresh product",
                                    "product_url": "https://www.amazon.co.uk/dp/B0FRESH001?th=1",
                                },
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-STALE",
                                    "asin": "B0STALE001",
                                    "product_name": "Stale product",
                                    "product_url": "https://www.amazon.co.uk/dp/B0STALE001?th=1",
                                },
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(
        json.dumps(
            {
                "items": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-FRESH",
                        "asin": "B0FRESH001",
                        "frontend_check_status": "已自动检查",
                        "frontend_check_method": "chrome-cdp",
                        "frontend_data_date": "2026-06-17",
                        "checked_at": "2026-06-17T09:00:00",
                        "frontend_findings": "售价：£19.99；评分：4.3 out of 5 stars；配送：London SW1A 1AA",
                        "frontend_price": "£19.99",
                        "frontend_rating": "4.3 out of 5 stars",
                        "frontend_location_note": "London SW1A 1AA",
                        "frontend_delivery": "London SW1A 1AA",
                    }
                ],
                "cache": [],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)
    monkeypatch.setattr(frontend, "_today_iso", lambda: "2026-06-17")
    probed_asins: list[str] = []

    def fake_probe(**kwargs):
        probed_asins.append(str(kwargs["asin"]))
        return {
            "attempts": [
                {
                    "attempt": 1,
                    "marketplace": "UK",
                    "asin": kwargs["asin"],
                    "success": True,
                    "title": "UK product",
                    "price": "£18.99",
                    "rating": "4.2 out of 5 stars",
                    "reviews": "(31)",
                    "location": "London SW1A 1AA",
                    "buy_box": "识别到购买按钮",
                }
            ]
        }

    monkeypatch.setattr(frontend, "run_chrome_cdp_probe", fake_probe)

    rows = frontend.check_frontend_queue(
        timeout=1,
        sleep_seconds=0,
        retries=1,
        method="chrome-cdp",
        search_policy="never",
        cdp_endpoint="http://127.0.0.1:9222",
        cdp_attempts=1,
        only_stale=True,
    )

    assert probed_asins == ["B0STALE001"]
    actions = {row["asin"]: row.get("frontend_refresh_action") for row in rows}
    assert actions["B0FRESH001"] == "skipped_fresh"
    assert actions["B0STALE001"] == "live_checked"
    summary = frontend._refresh_summary(rows)
    assert summary["frontend_refresh_live_checked"] == 1
    assert summary["frontend_refresh_skipped"] == 1
    assert summary["frontend_refresh_cache_used"] == 0
    assert summary["frontend_refresh_failed"] == 0


def test_frontend_refresh_summary_treats_string_false_cache_flag_as_false() -> None:
    import scripts.run_frontend_checks as frontend

    summary = frontend._refresh_summary(
        [
            {
                "marketplace": "UK",
                "asin": "B0LIVEFALSE",
                "frontend_refresh_action": "live_checked",
                "frontend_check_status": "已自动检查",
                "frontend_cache_used": "False",
            }
        ]
    )

    assert summary["frontend_refresh_live_checked"] == 1
    assert summary["frontend_refresh_cache_used"] == 0
    assert summary["frontend_refresh_failed"] == 0


def test_frontend_today_cdp_record_treats_string_false_cache_flag_as_false() -> None:
    import scripts.run_frontend_checks as frontend

    row = {
        "marketplace": "UK",
        "asin": "B0LIVEFALSE",
        "frontend_check_status": "已自动检查",
        "frontend_check_method": "chrome-cdp",
        "frontend_cache_used": "False",
        "frontend_findings": "售价：£18.99；配送：London SW1A 1AA",
        "frontend_location_note": "London SW1A 1AA",
        "checked_at": "2026-06-24T08:00:00",
    }

    assert frontend._is_today_chrome_cdp_record(row, today="2026-06-24") is True


def test_report_action_server_today_frontend_evidence_treats_string_false_cache_flag_as_false() -> None:
    import scripts.report_action_server as server

    row = {
        "marketplace": "UK",
        "asin": "B0LIVEFALSE",
        "frontend_check_status": "已自动检查",
        "frontend_check_method": "urllib",
        "frontend_cache_used": "False",
        "frontend_competitor_count": 2,
        "checked_at": "2026-06-24T08:00:00",
    }

    assert server._is_today_frontend_evidence(row, today="2026-06-24", require_competitor_samples=True) is True


def test_frontend_score_search_page_treats_string_false_partial_flag_as_false() -> None:
    import scripts.run_frontend_checks as frontend

    scored = frontend._score_search_page(
        {"price": "£18.99", "rating": "4.5 out of 5 stars", "reviews": "120 ratings"},
        {
            "frontend_search_status": "已自动检查",
            "frontend_competitors": [],
            "frontend_search_partial_evidence": "False",
        },
    )

    assert scored["frontend_search_partial_evidence"] is False
    assert scored["frontend_search_quality_components"]["competitor_sample_score"] == 0
    assert scored["frontend_search_quality_components"]["own_position_score"] == 0


def test_frontend_quality_from_record_treats_string_false_partial_flag_as_false() -> None:
    import scripts.run_frontend_checks as frontend

    payload = frontend._quality_payload_from_record(
        {
            "marketplace": "UK",
            "frontend_check_status": "已自动检查",
            "frontend_check_method": "chrome-cdp",
            "frontend_findings": "售价：£18.99；评分：4.5 out of 5 stars；评论数：120 ratings；配送：London SW1A 1AA",
            "frontend_search_status": "已自动检查",
            "frontend_search_partial_evidence": "False",
            "frontend_competitors": [],
        },
        basis="live",
    )

    assert payload["frontend_search_partial_evidence"] is False


def test_check_frontend_queue_only_stale_still_refreshes_explicit_single_product(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-FRESH",
                                    "asin": "B0FRESH001",
                                    "product_name": "Fresh product",
                                    "product_url": "https://www.amazon.co.uk/dp/B0FRESH001?th=1",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(
        json.dumps(
            {
                "items": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-FRESH",
                        "asin": "B0FRESH001",
                        "frontend_check_status": "已自动检查",
                        "frontend_check_method": "chrome-cdp",
                        "frontend_data_date": "2026-06-17",
                        "checked_at": "2026-06-17T09:00:00",
                        "frontend_findings": "售价：£19.99；评分：4.3 out of 5 stars；配送：London SW1A 1AA",
                        "frontend_price": "£19.99",
                        "frontend_rating": "4.3 out of 5 stars",
                        "frontend_location_note": "London SW1A 1AA",
                        "frontend_delivery": "London SW1A 1AA",
                    }
                ],
                "cache": [],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)
    monkeypatch.setattr(frontend, "_today_iso", lambda: "2026-06-17")
    probed_asins: list[str] = []

    def fake_probe(**kwargs):
        probed_asins.append(str(kwargs["asin"]))
        return {
            "attempts": [
                {
                    "attempt": 1,
                    "marketplace": "UK",
                    "asin": kwargs["asin"],
                    "success": True,
                    "title": "UK product refreshed",
                    "price": "£18.99",
                    "rating": "4.2 out of 5 stars",
                    "reviews": "(31)",
                    "location": "London SW1A 1AA",
                    "buy_box": "识别到购买按钮",
                }
            ]
        }

    monkeypatch.setattr(frontend, "run_chrome_cdp_probe", fake_probe)

    rows = frontend.check_frontend_queue(
        timeout=1,
        sleep_seconds=0,
        retries=1,
        method="chrome-cdp",
        search_policy="never",
        marketplace="UK",
        asin="B0FRESH001",
        cdp_endpoint="http://127.0.0.1:9222",
        cdp_attempts=1,
        only_stale=True,
    )

    assert probed_asins == ["B0FRESH001"]
    assert rows[0]["frontend_refresh_action"] == "live_checked"
    assert rows[0]["frontend_price"] == "£18.99"


def test_check_frontend_queue_only_stale_daily_failure_uses_cache(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-STALE",
                                    "asin": "B0STALE001",
                                    "product_name": "Stale product",
                                    "product_url": "https://www.amazon.co.uk/dp/B0STALE001?th=1",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(
        json.dumps(
            {
                "items": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-STALE",
                        "asin": "B0STALE001",
                        "frontend_check_status": "已自动检查",
                        "frontend_check_method": "chrome-cdp",
                        "frontend_data_date": "2026-06-16",
                        "checked_at": "2026-06-16T09:00:00",
                        "frontend_findings": "售价：£19.99；评分：4.3 out of 5 stars；配送：London SW1A 1AA",
                        "frontend_price": "£19.99",
                        "frontend_rating": "4.3 out of 5 stars",
                        "frontend_location_note": "London SW1A 1AA",
                        "frontend_delivery": "London SW1A 1AA",
                    }
                ],
                "cache": [],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)
    monkeypatch.setattr(frontend, "_today_iso", lambda: "2026-06-17")

    def fail_probe(**kwargs):
        raise RuntimeError("Chrome CDP endpoint is not available: http://127.0.0.1:9222")

    monkeypatch.setattr(frontend, "run_chrome_cdp_probe", fail_probe)

    rows = frontend.check_frontend_queue(
        timeout=1,
        sleep_seconds=0,
        retries=1,
        method="chrome-cdp",
        search_policy="never",
        cdp_endpoint="http://127.0.0.1:9222",
        cdp_attempts=1,
        only_stale=True,
    )

    assert rows[0]["frontend_check_status"] == "沿用 2026-06-16 前台数据"
    assert rows[0]["frontend_refresh_action"] == "cache_fallback"
    assert rows[0]["frontend_cache_used"] is True
    assert "Chrome CDP endpoint is not available" in rows[0]["frontend_last_error"]
    summary = frontend._refresh_summary(rows)
    assert summary["frontend_refresh_cache_used"] == 1
    assert summary["frontend_refresh_failed"] == 0


def test_check_frontend_queue_only_stale_does_not_skip_20_run_gate(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-FRESH",
                                    "asin": "B0FRESH001",
                                    "product_name": "Fresh product",
                                    "product_url": "https://www.amazon.co.uk/dp/B0FRESH001?th=1",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(
        json.dumps(
            {
                "items": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-FRESH",
                        "asin": "B0FRESH001",
                        "frontend_check_status": "已自动检查",
                        "frontend_check_method": "chrome-cdp",
                        "frontend_data_date": "2026-06-17",
                        "checked_at": "2026-06-17T09:00:00",
                        "frontend_findings": "售价：£19.99；评分：4.3 out of 5 stars；配送：London SW1A 1AA",
                        "frontend_price": "£19.99",
                        "frontend_rating": "4.3 out of 5 stars",
                        "frontend_location_note": "London SW1A 1AA",
                        "frontend_delivery": "London SW1A 1AA",
                    }
                ],
                "cache": [],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)
    monkeypatch.setattr(frontend, "_today_iso", lambda: "2026-06-17")
    attempts = [
        {
            "attempt": index,
            "marketplace": "UK",
            "asin": "B0FRESH001",
            "success": True,
            "title": "UK product",
            "price": "£19.99",
            "rating": "4.3 out of 5 stars",
            "reviews": "(38)",
            "location": "London SW1A 1AA",
            "buy_box": "识别到购买按钮",
        }
        for index in range(1, 21)
    ]
    captured_attempts: list[int] = []

    def fake_probe(**kwargs):
        captured_attempts.append(int(kwargs["attempts"]))
        return {"attempts": attempts}

    monkeypatch.setattr(frontend, "run_chrome_cdp_probe", fake_probe)

    rows = frontend.check_frontend_queue(
        timeout=1,
        sleep_seconds=0,
        retries=1,
        method="chrome-cdp",
        search_policy="never",
        cdp_endpoint="http://127.0.0.1:9222",
        cdp_attempts=20,
        only_stale=True,
    )

    assert captured_attempts == [20]
    assert rows[0]["frontend_stability_total_attempts"] == 20
    assert rows[0]["frontend_refresh_action"] == "live_checked"


def test_check_frontend_queue_chrome_cdp_requires_20_run_gate(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-1",
                                    "asin": "B0TEST1234",
                                    "product_name": "Test product",
                                    "product_url": "https://www.amazon.co.uk/dp/B0TEST1234?th=1",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(json.dumps({"items": [], "cache": []}), encoding="utf-8")
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)
    attempts = [
        {
            "attempt": index,
            "marketplace": "UK",
            "asin": "B0TEST1234",
            "success": index != 16,
            "title": "UK product" if index != 16 else "",
            "price": "£19.99" if index != 16 else "",
            "rating": "4.3 out of 5 stars" if index != 16 else "",
            "reviews": "(38)" if index != 16 else "",
            "location": "London SW1A 1AA" if index != 16 else "",
            "buy_box": "识别到购买按钮" if index != 16 else "",
        }
        for index in range(1, 21)
    ]
    monkeypatch.setattr(
        frontend,
        "run_chrome_cdp_probe",
        lambda **kwargs: {"generated_at": "2026-06-16T23:30:00", "attempts": attempts},
    )

    rows = frontend.check_frontend_queue(
        timeout=1,
        sleep_seconds=0,
        retries=1,
        method="chrome-cdp",
        search_policy="never",
        cdp_endpoint="http://127.0.0.1:9222",
        cdp_attempts=20,
    )

    assert rows[0]["frontend_check_status"] == "已自动检查"
    assert rows[0]["frontend_check_method"] == "chrome-cdp"
    assert rows[0]["frontend_stability_total_attempts"] == 20
    assert rows[0]["frontend_stability_success_count"] == 19
    assert rows[0]["frontend_stability_failure_count"] == 1
    assert rows[0]["frontend_stability_passed"] is True
    assert rows[0]["frontend_price"] == "£19.99"


def test_check_frontend_queue_chrome_cdp_uses_last_valid_attempt(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-1",
                                    "asin": "B0TEST1234",
                                    "product_name": "Test product",
                                    "product_url": "https://www.amazon.co.uk/dp/B0TEST1234?th=1",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(json.dumps({"items": [], "cache": []}), encoding="utf-8")
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)
    attempts = [
        {
            "attempt": index,
            "marketplace": "UK",
            "asin": "B0TEST1234",
            "success": index not in {16, 20},
            "title": "UK product" if index not in {16, 20} else "",
            "price": f"£19.{index:02d}" if index not in {16, 20} else "",
            "rating": "4.3 out of 5 stars" if index not in {16, 20} else "",
            "reviews": "(38)" if index not in {16, 20} else "",
            "location": "London SW1A 1AA" if index not in {16, 20} else "",
            "buy_box": "识别到购买按钮" if index not in {16, 20} else "",
        }
        for index in range(1, 21)
    ]
    monkeypatch.setattr(
        frontend,
        "run_chrome_cdp_probe",
        lambda **kwargs: {"generated_at": "2026-06-16T23:30:00", "attempts": attempts},
    )

    rows = frontend.check_frontend_queue(
        timeout=1,
        sleep_seconds=0,
        retries=1,
        method="chrome-cdp",
        search_policy="never",
        cdp_endpoint="http://127.0.0.1:9222",
        cdp_attempts=20,
    )

    assert rows[0]["frontend_check_status"] == "已自动检查"
    assert rows[0]["frontend_stability_success_count"] == 18
    assert rows[0]["frontend_stability_failure_count"] == 2
    assert rows[0]["frontend_price"] == "£19.19"
    assert rows[0]["frontend_location_note"] == "London SW1A 1AA"


def test_check_frontend_queue_chrome_cdp_writes_attempt_diagnostics(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    output_dir = tmp_path / "output"
    output_dir.mkdir()
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "US",
                                    "sku": "SKU-1",
                                    "asin": "B0TEST1234",
                                    "product_name": "Test product",
                                    "product_url": "https://www.amazon.com/dp/B0TEST1234",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)
    monkeypatch.setattr(frontend, "OUTPUT_DIR", output_dir)
    attempts = [
        {
            "attempt": index,
            "marketplace": "US",
            "asin": "B0TEST1234",
            "success": True,
            "title": "US product",
            "price": "$19.99",
            "rating": "4.3 out of 5 stars",
            "reviews": "(38)",
            "location": "US 10001 已设置",
            "buy_box": "识别到购买按钮",
        }
        for index in range(1, 21)
    ]
    monkeypatch.setattr(
        frontend,
        "run_chrome_cdp_probe",
        lambda **kwargs: {"generated_at": "2026-06-16T23:30:00", "attempts": attempts},
    )

    rows = frontend.check_frontend_queue(
        timeout=1,
        sleep_seconds=0,
        retries=1,
        method="chrome-cdp",
        search_policy="never",
        cdp_endpoint="http://127.0.0.1:9222",
        cdp_attempts=20,
    )

    assert rows[0]["frontend_check_status"] == "已自动检查"
    attempts_path = output_dir / "frontend_stability_attempts_us_b0test1234.json"
    report_path = output_dir / "frontend_stability_report_us_b0test1234.json"
    assert attempts_path.exists()
    assert report_path.exists()
    assert json.loads(report_path.read_text(encoding="utf-8"))["passed"] is True


def test_check_frontend_queue_chrome_cdp_accepts_80pct_gate(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-1",
                                    "asin": "B0TEST1234",
                                    "product_name": "Test product",
                                    "product_url": "https://www.amazon.co.uk/dp/B0TEST1234?th=1",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(json.dumps({"items": [], "cache": []}), encoding="utf-8")
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)
    attempts = [
        {
            "attempt": index,
            "marketplace": "UK",
            "asin": "B0TEST1234",
            "success": index not in {4, 9, 16, 20},
            "title": "UK product" if index not in {4, 9, 16, 20} else "",
            "price": "£19.99" if index not in {4, 9, 16, 20} else "",
            "rating": "4.3 out of 5 stars" if index not in {4, 9, 16, 20} else "",
            "reviews": "(38)" if index not in {4, 9, 16, 20} else "",
            "location": "London SW1A 1AA" if index not in {4, 9, 16, 20} else "",
            "buy_box": "识别到购买按钮" if index not in {4, 9, 16, 20} else "",
        }
        for index in range(1, 21)
    ]
    monkeypatch.setattr(
        frontend,
        "run_chrome_cdp_probe",
        lambda **kwargs: {"generated_at": "2026-06-16T23:30:00", "attempts": attempts},
    )

    rows = frontend.check_frontend_queue(
        timeout=1,
        sleep_seconds=0,
        retries=1,
        method="chrome-cdp",
        search_policy="never",
        cdp_endpoint="http://127.0.0.1:9222",
        cdp_attempts=20,
        strict_live_pass=True,
    )

    assert rows[0]["frontend_check_status"] == "已自动检查"
    assert rows[0]["frontend_stability_success_count"] == 16
    assert rows[0]["frontend_stability_failure_count"] == 4
    assert rows[0]["frontend_stability_success_rate"] == 0.8
    assert rows[0]["frontend_stability_passed"] is True


def test_check_frontend_queue_strict_live_rejects_cached_cdp_failure(monkeypatch, tmp_path) -> None:
    import pytest

    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-1",
                                    "asin": "B0TEST1234",
                                    "product_name": "Test product",
                                    "product_url": "https://www.amazon.co.uk/dp/B0TEST1234?th=1",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(
        json.dumps(
            {
                "items": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-1",
                        "asin": "B0TEST1234",
                        "frontend_check_status": "已自动检查",
                        "frontend_findings": "售价：£17.89；评分：4.2 out of 5 stars；配送：Aberdeen AB10 1",
                        "frontend_price": "£17.89",
                        "frontend_rating": "4.2 out of 5 stars",
                        "frontend_location_note": "Aberdeen AB10 1",
                        "frontend_delivery": "Aberdeen AB10 1",
                        "frontend_check_method": "chrome-extension-20x-95pct",
                        "checked_at": "2026-06-16T22:48:14",
                    }
                ],
                "cache": [],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)

    def fail_probe(**kwargs):
        raise RuntimeError("Chrome CDP endpoint is not available: http://127.0.0.1:9222")

    monkeypatch.setattr(frontend, "run_chrome_cdp_probe", fail_probe)

    with pytest.raises(RuntimeError, match="严格实时前台验收未通过"):
        frontend.check_frontend_queue(
            timeout=1,
            sleep_seconds=0,
            retries=1,
            method="chrome-cdp",
            search_policy="never",
            cdp_endpoint="http://127.0.0.1:9222",
            cdp_attempts=20,
            strict_live_pass=True,
        )


def test_run_frontend_checks_writes_rows_before_strict_live_failure(monkeypatch, tmp_path, capsys) -> None:
    import sys

    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "US",
                                    "sku": "SKU-1",
                                    "asin": "B0TEST1234",
                                    "product_name": "Test product",
                                    "product_url": "https://www.amazon.com/dp/B0TEST1234",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)
    attempts = [
        {
            "attempt": index,
            "marketplace": "US",
            "asin": "B0TEST1234",
            "success": False,
            "title": "",
            "price": "",
            "location": "",
            "error": "timeout",
        }
        for index in range(1, 21)
    ]
    monkeypatch.setattr(
        frontend,
        "run_chrome_cdp_probe",
        lambda **kwargs: {"generated_at": "2026-06-16T23:30:00", "attempts": attempts},
    )
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "run_frontend_checks.py",
            "--method",
            "chrome-cdp",
            "--cdp-endpoint",
            "http://127.0.0.1:9222",
            "--cdp-attempts",
            "20",
            "--timeout",
            "1",
            "--sleep",
            "0",
            "--search-policy",
            "never",
            "--strict-live-pass",
        ],
    )

    assert frontend.main() == 1
    captured = capsys.readouterr().out
    assert "严格实时前台验收未通过" in captured
    payload = json.loads(results.read_text(encoding="utf-8"))
    row = payload["items"][0]
    assert row["marketplace"] == "US"
    assert row["asin"] == "B0TEST1234"
    assert row["frontend_check_method"] == "chrome-cdp"
    assert row["frontend_stability_total_attempts"] == 20
    assert row["frontend_stability_success_count"] == 0
    assert row["frontend_stability_passed"] is False


def test_check_frontend_queue_chrome_cdp_failure_uses_cache(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-1",
                                    "asin": "B0TEST1234",
                                    "product_name": "Test product",
                                    "product_url": "https://www.amazon.co.uk/dp/B0TEST1234?th=1",
                                }
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(
        json.dumps(
            {
                "items": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-1",
                        "asin": "B0TEST1234",
                        "frontend_check_status": "已自动检查",
                        "frontend_findings": "售价：£17.89；评分：4.2 out of 5 stars；配送：Aberdeen AB10 1",
                        "frontend_price": "£17.89",
                        "frontend_rating": "4.2 out of 5 stars",
                        "frontend_location_note": "Aberdeen AB10 1",
                        "frontend_delivery": "Aberdeen AB10 1",
                        "frontend_check_method": "chrome-extension-20x-95pct",
                        "checked_at": "2026-06-16T22:48:14",
                    }
                ],
                "cache": [],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)

    def fail_probe(**kwargs):
        raise RuntimeError("Chrome CDP endpoint is not available: http://127.0.0.1:9222")

    monkeypatch.setattr(frontend, "run_chrome_cdp_probe", fail_probe)

    rows = frontend.check_frontend_queue(
        timeout=1,
        sleep_seconds=0,
        retries=1,
        method="chrome-cdp",
        search_policy="never",
        cdp_endpoint="http://127.0.0.1:9222",
        cdp_attempts=20,
    )

    assert rows[0]["frontend_check_status"] == "沿用 2026-06-16 前台数据"
    assert rows[0]["source"] == "auto_frontend_check_cache"
    assert rows[0]["frontend_check_method"] == "chrome-extension-20x-95pct"
    assert "Chrome CDP endpoint is not available" in rows[0]["frontend_last_error"]


def test_check_frontend_queue_priority_filter_only_runs_p0(monkeypatch, tmp_path) -> None:
    import scripts.run_frontend_checks as frontend

    latest = tmp_path / "latest_analysis.json"
    results = tmp_path / "frontend_check_results.json"
    latest.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "priority": "P0",
                                    "marketplace": "UK",
                                    "sku": "SKU-P0",
                                    "asin": "B0P0TEST01",
                                    "product_name": "P0 product",
                                    "product_url": "https://www.amazon.co.uk/dp/B0P0TEST01?th=1",
                                },
                                {
                                    "priority": "P1",
                                    "marketplace": "US",
                                    "sku": "SKU-P1",
                                    "asin": "B0P1TEST01",
                                    "product_name": "P1 product",
                                    "product_url": "https://www.amazon.com/dp/B0P1TEST01",
                                },
                            ]
                        }
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    results.write_text(json.dumps({"items": [], "cache": []}), encoding="utf-8")
    monkeypatch.setattr(frontend, "LATEST_ANALYSIS", latest)
    monkeypatch.setattr(frontend, "RESULTS_PATH", results)

    rows = frontend.check_frontend_queue(
        timeout=1,
        sleep_seconds=0,
        retries=1,
        dry_run=True,
        method="urllib",
        search_policy="never",
        priority="P0",
    )

    assert [row["asin"] for row in rows] == ["B0P0TEST01"]
    assert rows[0]["marketplace"] == "UK"
    assert rows[0]["frontend_check_status"] == "待前台检查"
    assert rows[0]["frontend_check_method"] == "dry-run"


def _stub_report_output_restore(monkeypatch, workflow):
    snapshot = {
        "outputs": {workflow.ROOT / "data" / "output" / "latest_analysis.json": b"old"},
        "state_files": {},
        "archive_files": set(),
    }
    restored: list[dict] = []

    monkeypatch.setattr(workflow, "report_state_snapshot", lambda: snapshot)
    monkeypatch.setattr(workflow, "restore_report_state_snapshot", lambda value: restored.append(value) or [])
    return snapshot, restored


def test_main_debug_timing_timeout_returns_non_success(monkeypatch, tmp_path, capsys) -> None:
    import main as app_main

    source_path = tmp_path / "ads.csv"
    monkeypatch.setattr(app_main.sys, "argv", ["main.py", "--marketplace", "UK", "--debug-timing", "--safe-run"])
    monkeypatch.setattr(app_main, "ensure_directories", lambda: None)
    monkeypatch.setattr(app_main, "ensure_ignored_issues_config", lambda: None)
    monkeypatch.setattr(app_main, "_safe_output_dir", lambda output_dir: tmp_path)
    monkeypatch.setattr(app_main, "build_sku_asin_map_from_cost_config", lambda **kwargs: None)
    monkeypatch.setattr(app_main, "_select_preferred_source_multi", lambda *args, **kwargs: source_path)
    monkeypatch.setattr(
        app_main,
        "load_ads_report",
        lambda path: (pd.DataFrame({"marketplace": ["UK"]}), path),
    )
    monkeypatch.setattr(app_main, "_done", lambda step, started: 61.0 if step == "读取广告表" else 0.0)

    code = app_main.main()
    output = capsys.readouterr().out

    assert code == app_main.DEBUG_TIMING_EXIT_CODE
    assert "[TIMEOUT_STEP] 卡在 读取广告表" in output


def test_main_debug_timing_skip_report_generation_returns_non_success(monkeypatch, tmp_path, capsys) -> None:
    import main as app_main

    source_path = tmp_path / "source.csv"
    monkeypatch.setattr(app_main.sys, "argv", ["main.py", "--marketplace", "ALL", "--debug-timing", "--safe-run"])
    monkeypatch.setattr(app_main, "ensure_directories", lambda: None)
    monkeypatch.setattr(app_main, "ensure_ignored_issues_config", lambda: None)
    monkeypatch.setattr(app_main, "_safe_output_dir", lambda output_dir: tmp_path)
    monkeypatch.setattr(app_main, "build_sku_asin_map_from_cost_config", lambda **kwargs: None)
    monkeypatch.setattr(app_main, "_select_preferred_source_multi", lambda *args, **kwargs: source_path)
    monkeypatch.setattr(
        app_main,
        "load_ads_report",
        lambda path: (pd.DataFrame({"marketplace": ["UK"], "date": ["2026-06-24"]}), path),
    )
    monkeypatch.setattr(
        app_main,
        "load_erp_report",
        lambda path: (pd.DataFrame({"marketplace": ["UK"], "date": ["2026-06-24"]}), path),
    )
    monkeypatch.setattr(app_main, "_apply_stale_file_warning", lambda *args, **kwargs: None)
    monkeypatch.setattr(app_main, "_debug_read_enhanced_headers", lambda *args, **kwargs: (0.0, 0.0))
    monkeypatch.setattr(app_main, "_done", lambda step, started: 0.0)
    monkeypatch.setattr(app_main, "build_daily_dataset", lambda **kwargs: SimpleNamespace())

    code = app_main.main()
    output = capsys.readouterr().out

    assert code == app_main.DEBUG_TIMING_EXIT_CODE
    assert "[DEBUG_TIMING] 仅定位耗时，已跳过完整报告生成。" in output


def test_main_restores_formal_state_when_non_safe_run_returns_non_success(monkeypatch) -> None:
    import main as app_main

    snapshot = {"outputs": {}, "state_files": {}, "archive_files": set()}
    restored: list[dict[str, object]] = []

    monkeypatch.setattr(app_main.sys, "argv", ["main.py", "--marketplace", "ALL"])
    monkeypatch.setattr(app_main, "_formal_run_state_snapshot", lambda: snapshot)
    monkeypatch.setattr(app_main, "_restore_formal_run_state_snapshot", lambda value: restored.append(value) or [])
    monkeypatch.setattr(app_main, "_main_impl", lambda: app_main.DEBUG_TIMING_EXIT_CODE)

    code = app_main.main()

    assert code == app_main.DEBUG_TIMING_EXIT_CODE
    assert restored == [snapshot]


def test_main_restores_formal_state_when_non_safe_run_raises(monkeypatch) -> None:
    import main as app_main

    snapshot = {"outputs": {}, "state_files": {}, "archive_files": set()}
    restored: list[dict[str, object]] = []

    def fail_impl() -> int:
        raise RuntimeError("report write failed")

    monkeypatch.setattr(app_main.sys, "argv", ["main.py", "--marketplace", "ALL"])
    monkeypatch.setattr(app_main, "_formal_run_state_snapshot", lambda: snapshot)
    monkeypatch.setattr(app_main, "_restore_formal_run_state_snapshot", lambda value: restored.append(value) or [])
    monkeypatch.setattr(app_main, "_main_impl", fail_impl)

    with pytest.raises(RuntimeError, match="report write failed"):
        app_main.main()

    assert restored == [snapshot]


def test_main_safe_run_failure_does_not_restore_formal_state(monkeypatch) -> None:
    import main as app_main

    snapshots: list[str] = []
    restored: list[dict[str, object]] = []

    monkeypatch.setattr(app_main.sys, "argv", ["main.py", "--marketplace", "ALL", "--safe-run"])
    monkeypatch.setattr(app_main, "_formal_run_state_snapshot", lambda: snapshots.append("snapshot") or {})
    monkeypatch.setattr(app_main, "_restore_formal_run_state_snapshot", lambda value: restored.append(value) or [])
    monkeypatch.setattr(app_main, "_main_impl", lambda: app_main.DEBUG_TIMING_EXIT_CODE)

    code = app_main.main()

    assert code == app_main.DEBUG_TIMING_EXIT_CODE
    assert snapshots == []
    assert restored == []


def test_run_all_with_frontend_checks_live_browser_enables_reuse_session(monkeypatch) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    calls: list[list[str]] = []
    _, restored = _stub_report_output_restore(monkeypatch, workflow)

    monkeypatch.setattr(workflow, "run_step", lambda args: calls.append(args) or 0)
    monkeypatch.setattr(workflow, "run_frontend_price_sync", lambda python: 0)
    monkeypatch.setattr(workflow, "_output_refresh_snapshot", lambda: {})
    monkeypatch.setattr(workflow, "frontend_results_refresh_failures", lambda **kwargs: [])
    monkeypatch.setattr(workflow, "report_refresh_failures", lambda previous_mtimes_ns=None: [])
    monkeypatch.setattr(workflow, "frontend_results_absorption_failures", lambda: [])
    monkeypatch.setattr(workflow.sys, "argv", ["run_all_with_frontend_checks.py", "--live-browser-frontend"])

    code = workflow.main()

    assert code == 0
    frontend_step = calls[1]
    assert frontend_step[:3] == [workflow.sys.executable, "scripts/run_frontend_checks.py", "--method"]
    assert "--reuse-browser-session" in frontend_step
    assert "auto" in frontend_step
    assert "chrome-persistent" not in frontend_step
    assert restored == []


def test_run_frontend_price_sync_allows_playwright_fallback_success(monkeypatch) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    calls: list[list[str]] = []

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        return 17 if "chrome" in args else 0

    monkeypatch.setattr(workflow, "run_step", fake_run_step)

    code = workflow.run_frontend_price_sync("/python")

    assert code == 0
    assert len(calls) == 2
    assert calls[0][:2] == ["/python", "scripts/sync_frontend_prices.py"]
    assert "chrome" in calls[0]
    assert calls[1][:2] == ["/python", "scripts/sync_frontend_prices.py"]
    assert "playwright" in calls[1]


def test_run_all_with_frontend_checks_stops_when_live_price_sync_fails(monkeypatch, capsys) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    calls: list[list[str]] = []
    snapshot, restored = _stub_report_output_restore(monkeypatch, workflow)

    monkeypatch.setattr(workflow, "run_step", lambda args: calls.append(args) or 0)
    monkeypatch.setattr(workflow, "run_frontend_price_sync", lambda python: 31)
    monkeypatch.setattr(workflow.sys, "argv", ["run_all_with_frontend_checks.py", "--live-browser-frontend"])

    code = workflow.main()
    output = capsys.readouterr().out

    assert code == 31
    assert calls == [[workflow.sys.executable, "main.py", "--marketplace", "ALL"]]
    assert restored == [snapshot]
    assert "[done]" not in output


def test_run_all_with_frontend_checks_no_browser_keeps_urllib_without_reuse(monkeypatch) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    calls: list[list[str]] = []
    _, restored = _stub_report_output_restore(monkeypatch, workflow)

    monkeypatch.setattr(workflow, "run_step", lambda args: calls.append(args) or 0)
    monkeypatch.setattr(workflow, "_output_refresh_snapshot", lambda: {})
    monkeypatch.setattr(workflow, "frontend_results_refresh_failures", lambda **kwargs: [])
    monkeypatch.setattr(workflow, "report_refresh_failures", lambda previous_mtimes_ns=None: [])
    monkeypatch.setattr(workflow, "frontend_results_absorption_failures", lambda: [])
    monkeypatch.setattr(workflow.sys, "argv", ["run_all_with_frontend_checks.py", "--no-live-browser-frontend"])

    code = workflow.main()

    assert code == 0
    frontend_step = calls[1]
    assert "--reuse-browser-session" not in frontend_step
    assert "urllib" in frontend_step
    assert restored == []


def test_run_all_with_frontend_checks_stops_when_frontend_step_fails(monkeypatch) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    calls: list[list[str]] = []
    snapshot, restored = _stub_report_output_restore(monkeypatch, workflow)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[:2] == [workflow.sys.executable, "scripts/run_frontend_checks.py"]:
            return 7
        if len(calls) > 2:
            raise AssertionError("workflow continued after frontend failure")
        return 0

    monkeypatch.setattr(workflow, "run_step", fake_run_step)
    monkeypatch.setattr(workflow.sys, "argv", ["run_all_with_frontend_checks.py", "--no-live-browser-frontend"])

    code = workflow.main()

    assert code == 7
    assert restored == [snapshot]
    assert len(calls) == 2
    assert calls[0] == [workflow.sys.executable, "main.py", "--marketplace", "ALL"]
    assert calls[1][:2] == [workflow.sys.executable, "scripts/run_frontend_checks.py"]


def test_run_all_with_frontend_checks_stops_when_final_report_outputs_are_not_refreshed(monkeypatch, capsys) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    calls: list[list[str]] = []
    snapshot, restored = _stub_report_output_restore(monkeypatch, workflow)

    monkeypatch.setattr(workflow, "run_step", lambda args: calls.append(args) or 0)
    monkeypatch.setattr(workflow, "frontend_results_refresh_failures", lambda **kwargs: [])
    monkeypatch.setattr(workflow, "_output_refresh_snapshot", lambda: {"snapshot": 1})
    monkeypatch.setattr(
        workflow,
        "report_refresh_failures",
        lambda previous_mtimes_ns=None: ["required output was not refreshed by report step: latest_analysis.json"],
    )
    monkeypatch.setattr(workflow.sys, "argv", ["run_all_with_frontend_checks.py", "--no-live-browser-frontend"])

    code = workflow.main()
    output = capsys.readouterr().out

    assert code == 1
    assert restored == [snapshot]
    assert calls == [
        [workflow.sys.executable, "main.py", "--marketplace", "ALL"],
        [
            workflow.sys.executable,
            "scripts/run_frontend_checks.py",
            "--method",
            "urllib",
            "--timeout",
            "30",
            "--retries",
            "3",
            "--search-policy",
            "always",
        ],
        [workflow.sys.executable, "main.py", "--marketplace", "ALL"],
    ]
    assert "report refresh blocker" in output
    assert "[restore] report outputs restored to pre-run snapshot after failure" in output
    assert "[done]" not in output


def test_run_all_with_frontend_checks_stops_when_frontend_results_are_not_refreshed(
    monkeypatch, tmp_path, capsys
) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    results_path = tmp_path / "frontend_check_results.json"
    results_path.write_text(
        json.dumps({"generated_at": "2026-06-24T10:00:00", "refresh_summary": {}, "items": []}),
        encoding="utf-8",
    )
    calls: list[list[str]] = []
    snapshot, restored = _stub_report_output_restore(monkeypatch, workflow)

    monkeypatch.setattr(workflow, "FRONTEND_RESULTS_JSON", results_path)
    monkeypatch.setattr(workflow, "run_step", lambda args: calls.append(args) or 0)
    monkeypatch.setattr(workflow.sys, "argv", ["run_all_with_frontend_checks.py", "--no-live-browser-frontend"])

    code = workflow.main()
    output = capsys.readouterr().out

    assert code == 1
    assert restored == [snapshot]
    assert calls == [
        [workflow.sys.executable, "main.py", "--marketplace", "ALL"],
        [
            workflow.sys.executable,
            "scripts/run_frontend_checks.py",
            "--method",
            "urllib",
            "--timeout",
            "30",
            "--retries",
            "3",
            "--search-policy",
            "always",
        ],
    ]
    assert "frontend refresh blocker" in output
    assert "frontend check results were not refreshed by frontend step" in output
    assert "[restore] report outputs restored to pre-run snapshot after failure" in output
    assert "[done]" not in output


def test_run_all_with_frontend_checks_stops_when_frontend_results_rewrite_old_generated_at(
    monkeypatch, tmp_path, capsys
) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    results_path = tmp_path / "frontend_check_results.json"
    old_payload = {"generated_at": "2026-06-24T10:00:00", "refresh_summary": {}, "items": []}
    results_path.write_text(json.dumps(old_payload), encoding="utf-8")
    calls: list[list[str]] = []
    snapshot, restored = _stub_report_output_restore(monkeypatch, workflow)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        if args[:2] == [workflow.sys.executable, "scripts/run_frontend_checks.py"]:
            results_path.write_text(json.dumps(old_payload), encoding="utf-8")
        return 0

    monkeypatch.setattr(workflow, "FRONTEND_RESULTS_JSON", results_path)
    monkeypatch.setattr(workflow, "run_step", fake_run_step)
    monkeypatch.setattr(workflow.sys, "argv", ["run_all_with_frontend_checks.py", "--no-live-browser-frontend"])

    code = workflow.main()
    output = capsys.readouterr().out

    assert code == 1
    assert restored == [snapshot]
    assert calls == [
        [workflow.sys.executable, "main.py", "--marketplace", "ALL"],
        [
            workflow.sys.executable,
            "scripts/run_frontend_checks.py",
            "--method",
            "urllib",
            "--timeout",
            "30",
            "--retries",
            "3",
            "--search-policy",
            "always",
        ],
    ]
    assert "frontend refresh blocker" in output
    assert "frontend check results generated_at was not refreshed by frontend step" in output
    assert "[restore] report outputs restored to pre-run snapshot after failure" in output
    assert "[done]" not in output


def test_frontend_results_refresh_failures_blocks_empty_results_when_queue_was_present(tmp_path) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    results_path = tmp_path / "frontend_check_results.json"
    analysis_path = tmp_path / "latest_analysis.json"
    results_path.write_text(
        json.dumps({"generated_at": "2026-06-24T10:30:00", "refresh_summary": {}, "items": []}),
        encoding="utf-8",
    )
    analysis_path.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "marketplace": "UK",
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-FRONTEND-QUEUE",
                                    "asin": "B0QUEUE001",
                                }
                            ]
                        },
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    failures = workflow.frontend_results_refresh_failures(
        results_path,
        expected_analysis_path=analysis_path,
    )

    assert failures == [
        "frontend check results did not cover queued frontend identities: "
        "UK/SKU-FRONTEND-QUEUE/B0QUEUE001"
    ]


def test_frontend_results_refresh_failures_blocks_all_failed_frontend_rows(tmp_path) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    results_path = tmp_path / "frontend_check_results.json"
    analysis_path = tmp_path / "latest_analysis.json"
    results_path.write_text(
        json.dumps(
            {
                "generated_at": "2026-06-24T10:30:00",
                "refresh_summary": {
                    "frontend_refresh_total": 1,
                    "frontend_refresh_live_checked": 0,
                    "frontend_refresh_skipped": 0,
                    "frontend_refresh_cache_used": 0,
                    "frontend_refresh_failed": 1,
                },
                "items": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-FRONTEND-FAILED",
                        "asin": "B0FAILED01",
                        "frontend_refresh_action": "live_failed",
                    }
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    analysis_path.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "marketplace": "UK",
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-FRONTEND-FAILED",
                                    "asin": "B0FAILED01",
                                }
                            ]
                        },
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    failures = workflow.frontend_results_refresh_failures(
        results_path,
        expected_analysis_path=analysis_path,
    )

    assert failures == [
        f"frontend check results contain only failed frontend refresh rows: {results_path}"
    ]


def test_frontend_results_refresh_failures_blocks_stale_covered_queue_row_without_refresh_action(tmp_path) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    results_path = tmp_path / "frontend_check_results.json"
    analysis_path = tmp_path / "latest_analysis.json"
    results_path.write_text(
        json.dumps(
            {
                "generated_at": "2026-06-24T10:30:00",
                "refresh_summary": {
                    "frontend_refresh_total": 1,
                    "frontend_refresh_live_checked": 1,
                    "frontend_refresh_skipped": 0,
                    "frontend_refresh_cache_used": 0,
                    "frontend_refresh_failed": 0,
                },
                "items": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-FRONTEND-STALE-COVERED",
                        "asin": "B0STALECOV",
                        "frontend_data_date": "2026-06-20",
                    }
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    analysis_path.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "marketplace": "UK",
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-FRONTEND-STALE-COVERED",
                                    "asin": "B0STALECOV",
                                }
                            ]
                        },
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    failures = workflow.frontend_results_refresh_failures(
        results_path,
        expected_analysis_path=analysis_path,
    )

    assert failures == [
        "frontend check results covered queued identities with stale rows lacking refresh action: "
        "UK/SKU-FRONTEND-STALE-COVERED/B0STALECOV"
    ]


def test_run_all_with_frontend_checks_validates_against_pre_final_refresh_snapshot(monkeypatch) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    calls: list[list[str]] = []
    snapshot = {"latest_analysis": 123}
    received_snapshots: list[dict[str, int]] = []
    _, restored = _stub_report_output_restore(monkeypatch, workflow)

    def fake_run_step(args: list[str]) -> int:
        calls.append(args)
        return 0

    def fake_snapshot() -> dict[str, int]:
        assert calls == [
            [workflow.sys.executable, "main.py", "--marketplace", "ALL"],
            [
                workflow.sys.executable,
                "scripts/run_frontend_checks.py",
                "--method",
                "urllib",
                "--timeout",
                "30",
                "--retries",
                "3",
                "--search-policy",
                "always",
            ],
        ]
        return snapshot

    def fake_report_refresh_failures(previous_mtimes_ns=None):
        received_snapshots.append(previous_mtimes_ns)
        return []

    monkeypatch.setattr(workflow, "run_step", fake_run_step)
    monkeypatch.setattr(workflow, "frontend_results_refresh_failures", lambda **kwargs: [])
    monkeypatch.setattr(workflow, "_output_refresh_snapshot", fake_snapshot)
    monkeypatch.setattr(workflow, "report_refresh_failures", fake_report_refresh_failures)
    monkeypatch.setattr(workflow, "frontend_results_absorption_failures", lambda: [])
    monkeypatch.setattr(workflow.sys, "argv", ["run_all_with_frontend_checks.py", "--no-live-browser-frontend"])

    code = workflow.main()

    assert code == 0
    assert calls == [
        [workflow.sys.executable, "main.py", "--marketplace", "ALL"],
        [
            workflow.sys.executable,
            "scripts/run_frontend_checks.py",
            "--method",
            "urllib",
            "--timeout",
            "30",
            "--retries",
            "3",
            "--search-policy",
            "always",
        ],
        [workflow.sys.executable, "main.py", "--marketplace", "ALL"],
    ]
    assert received_snapshots == [snapshot]
    assert restored == []


def test_frontend_results_absorption_failures_allows_result_row_dropped_from_latest_analysis_queue(tmp_path) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    results_path = tmp_path / "frontend_check_results.json"
    analysis_path = tmp_path / "latest_analysis.json"
    results_path.write_text(
        json.dumps(
            {
                "generated_at": "2026-06-24T10:30:00",
                "refresh_summary": {},
                "items": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-FRONTEND-ABSORB",
                        "asin": "B0ABSORB001",
                        "frontend_check_status": "已自动检查",
                        "frontend_refresh_action": "live_checked",
                        "frontend_data_date": "2026-06-24",
                    }
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    analysis_path.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "marketplace": "UK",
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-OTHER",
                                    "asin": "B0OTHER001",
                                }
                            ]
                        },
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    failures = workflow.frontend_results_absorption_failures(results_path, analysis_path)

    assert failures == []


def test_frontend_results_absorption_failures_blocks_stale_latest_analysis_frontend_fields(tmp_path) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    results_path = tmp_path / "frontend_check_results.json"
    analysis_path = tmp_path / "latest_analysis.json"
    results_path.write_text(
        json.dumps(
            {
                "generated_at": "2026-06-24T10:30:00",
                "refresh_summary": {},
                "items": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-FRONTEND-STALE",
                        "asin": "B0STALE001",
                        "frontend_check_status": "已自动检查",
                        "frontend_refresh_action": "live_checked",
                        "frontend_data_date": "2026-06-24",
                        "frontend_cache_used": False,
                        "frontend_evidence_tier": "强诊断可用",
                        "frontend_search_status": "已自动检查",
                    }
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    analysis_path.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "marketplace": "UK",
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-FRONTEND-STALE",
                                    "asin": "B0STALE001",
                                    "frontend_check_status": "沿用 2026-06-20 前台数据",
                                    "frontend_refresh_action": "cache_used",
                                    "frontend_data_date": "2026-06-20",
                                    "frontend_cache_used": True,
                                    "frontend_evidence_tier": "仅背景参考",
                                    "frontend_search_status": "待前台检查",
                                }
                            ]
                        },
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    failures = workflow.frontend_results_absorption_failures(results_path, analysis_path)

    assert failures == [
        "frontend check results field was not absorbed into latest_analysis frontend queue "
        "for UK/SKU-FRONTEND-STALE/B0STALE001 field frontend_check_status: "
        "expected '已自动检查', got '沿用 2026-06-20 前台数据'"
    ]


def test_frontend_results_absorption_failures_blocks_stale_decision_tier(tmp_path) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    results_path = tmp_path / "frontend_check_results.json"
    analysis_path = tmp_path / "latest_analysis.json"
    base_row = {
        "marketplace": "UK",
        "sku": "SKU-FRONTEND-TIER",
        "asin": "B0TIER001",
        "frontend_check_status": "已自动检查",
        "frontend_refresh_action": "live_checked",
        "frontend_data_date": "2026-06-24",
        "frontend_cache_used": False,
        "frontend_evidence_tier": "强诊断可用",
        "frontend_search_status": "已自动检查",
    }
    results_row = dict(base_row, frontend_decision_evidence_tier="强诊断可用")
    analysis_row = dict(base_row, frontend_decision_evidence_tier="仅背景参考")
    results_path.write_text(
        json.dumps(
            {
                "generated_at": "2026-06-24T10:30:00",
                "refresh_summary": {},
                "items": [results_row],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    analysis_path.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "marketplace": "UK",
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [analysis_row]
                        },
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    failures = workflow.frontend_results_absorption_failures(results_path, analysis_path)

    assert failures == [
        "frontend check results field was not absorbed into latest_analysis frontend queue "
        "for UK/SKU-FRONTEND-TIER/B0TIER001 field frontend_decision_evidence_tier: "
        "expected '强诊断可用', got '仅背景参考'"
    ]


def test_frontend_results_absorption_failures_blocks_stale_gate_quality_fields(tmp_path) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    results_path = tmp_path / "frontend_check_results.json"
    analysis_path = tmp_path / "latest_analysis.json"
    base_row = {
        "marketplace": "UK",
        "sku": "SKU-FRONTEND-GATE",
        "asin": "B0GATE001",
        "frontend_check_status": "已自动检查",
        "frontend_refresh_action": "live_checked",
        "frontend_data_date": "2026-06-24",
        "frontend_cache_used": False,
        "frontend_evidence_tier": "强诊断可用",
        "frontend_decision_evidence_tier": "强诊断可用",
        "frontend_search_status": "已自动检查",
        "frontend_evidence_quality_score": "88",
        "competitor_comparability": "high",
        "comparable_competitor_count": "3",
        "frontend_location_scope": "exact",
        "frontend_location_verified": True,
    }
    results_path.write_text(
        json.dumps(
            {
                "generated_at": "2026-06-24T10:30:00",
                "refresh_summary": {},
                "items": [base_row],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    analysis_row = dict(base_row, frontend_evidence_quality_score="58")
    analysis_path.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "marketplace": "UK",
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [analysis_row]
                        },
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    failures = workflow.frontend_results_absorption_failures(results_path, analysis_path)

    assert failures == [
        "frontend check results field was not absorbed into latest_analysis frontend queue "
        "for UK/SKU-FRONTEND-GATE/B0GATE001 field frontend_evidence_quality_score: "
        "expected '88', got '58'"
    ]


def test_frontend_results_absorption_failures_blocks_competitor_count_mismatch(tmp_path) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    results_path = tmp_path / "frontend_check_results.json"
    analysis_path = tmp_path / "latest_analysis.json"
    base_row = {
        "marketplace": "UK",
        "sku": "SKU-FRONTEND-COMP",
        "asin": "B0COMP001",
        "frontend_check_status": "已自动检查",
        "frontend_refresh_action": "live_checked",
        "frontend_data_date": "2026-06-24",
        "frontend_competitor_count": "3",
        "comparable_competitor_count": "1",
    }
    results_path.write_text(
        json.dumps(
            {
                "generated_at": "2026-06-24T10:30:00",
                "refresh_summary": {},
                "items": [base_row],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    analysis_row = dict(base_row, frontend_competitor_count="1")
    analysis_path.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "marketplace": "UK",
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [analysis_row]
                        },
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    failures = workflow.frontend_results_absorption_failures(results_path, analysis_path)

    assert failures == [
        "frontend check results field was not absorbed into latest_analysis frontend queue "
        "for UK/SKU-FRONTEND-COMP/B0COMP001 field frontend_competitor_count: "
        "expected '3', got '1'"
    ]


def test_frontend_results_absorption_failures_blocks_audit_reasons_mismatch(tmp_path) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    results_path = tmp_path / "frontend_check_results.json"
    analysis_path = tmp_path / "latest_analysis.json"
    base_row = {
        "marketplace": "UK",
        "sku": "SKU-FRONTEND-AUDIT",
        "asin": "B0AUDIT001",
        "frontend_check_status": "已自动检查",
        "frontend_refresh_action": "live_checked",
        "frontend_data_date": "2026-06-24",
        "frontend_evidence_audit_summary": "前台证据可辅助判断，不能单独放量",
        "frontend_evidence_audit_detail": "地区未确认；竞品可比性未达强诊断",
    }
    results_row = dict(
        base_row,
        frontend_evidence_audit_reasons=["地区未确认", "竞品可比性未达强诊断"],
    )
    analysis_row = dict(base_row, frontend_evidence_audit_reasons=["产品页和搜索页均可用"])
    results_path.write_text(
        json.dumps(
            {
                "generated_at": "2026-06-24T10:30:00",
                "refresh_summary": {},
                "items": [results_row],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    analysis_path.write_text(
        json.dumps(
            {
                "marketplace_results": [
                    {
                        "marketplace": "UK",
                        "report_view_snapshot": {
                            "frontend_check_queue_rows": [analysis_row]
                        },
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    failures = workflow.frontend_results_absorption_failures(results_path, analysis_path)

    assert failures == [
        "frontend check results field was not absorbed into latest_analysis frontend queue "
        "for UK/SKU-FRONTEND-AUDIT/B0AUDIT001 field frontend_evidence_audit_reasons: "
        "expected '地区未确认；竞品可比性未达强诊断', got '产品页和搜索页均可用'"
    ]


def test_run_all_with_frontend_checks_stops_when_frontend_results_are_not_absorbed(
    monkeypatch, capsys
) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    calls: list[list[str]] = []
    snapshot, restored = _stub_report_output_restore(monkeypatch, workflow)

    monkeypatch.setattr(workflow, "run_step", lambda args: calls.append(args) or 0)
    monkeypatch.setattr(workflow, "frontend_results_refresh_failures", lambda **kwargs: [])
    monkeypatch.setattr(workflow, "_output_refresh_snapshot", lambda: {"latest_analysis": 1})
    monkeypatch.setattr(workflow, "report_refresh_failures", lambda previous_mtimes_ns=None: [])
    monkeypatch.setattr(
        workflow,
        "frontend_results_absorption_failures",
        lambda: ["frontend check results were not absorbed into latest_analysis frontend queue: UK/SKU/B0ASIN"],
    )
    monkeypatch.setattr(workflow.sys, "argv", ["run_all_with_frontend_checks.py", "--no-live-browser-frontend"])

    code = workflow.main()
    output = capsys.readouterr().out

    assert code == 1
    assert restored == [snapshot]
    assert calls == [
        [workflow.sys.executable, "main.py", "--marketplace", "ALL"],
        [
            workflow.sys.executable,
            "scripts/run_frontend_checks.py",
            "--method",
            "urllib",
            "--timeout",
            "30",
            "--retries",
            "3",
            "--search-policy",
            "always",
        ],
        [workflow.sys.executable, "main.py", "--marketplace", "ALL"],
    ]
    assert "frontend absorption blocker" in output
    assert "[restore] report outputs restored to pre-run snapshot after failure" in output
    assert "[done]" not in output


def test_run_all_with_frontend_checks_returns_127_when_step_cannot_start(monkeypatch, capsys) -> None:
    import scripts.run_all_with_frontend_checks as workflow

    calls: list[list[str]] = []
    snapshot, restored = _stub_report_output_restore(monkeypatch, workflow)

    def fake_run(args: list[str], cwd=None):
        calls.append(args)
        raise OSError("python missing")

    monkeypatch.setattr(workflow.subprocess, "run", fake_run)
    monkeypatch.setattr(workflow.sys, "argv", ["run_all_with_frontend_checks.py", "--no-live-browser-frontend"])

    code = workflow.main()
    output = capsys.readouterr().out

    assert code == 127
    assert restored == [snapshot]
    assert calls == [[workflow.sys.executable, "main.py", "--marketplace", "ALL"]]
    assert "[fail] cannot start step: python missing" in output
    assert "scripts/run_frontend_checks.py" not in output
    assert "[done]" not in output


def test_frontend_queue_transmits_auto_analysis_fields(monkeypatch, tmp_path) -> None:
    from src.report_presentation import _build_frontend_check_queue

    output_dir = tmp_path
    (output_dir / "frontend_check_results.json").write_text(
        json.dumps(
            {
                "items": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-1",
                        "asin": "B0TEST1234",
                        "frontend_check_status": "已自动检查",
                        "frontend_findings": "售价：£18.99；评分：3.8 out of 5 stars",
                        "frontend_auto_conclusion_label": "明确前台劣势",
                        "frontend_evidence_quality_score": 58,
                        "frontend_check_method": "chrome-extension-20x-95pct",
                        "frontend_stability_total_attempts": 20,
                        "frontend_stability_success_count": 19,
                        "frontend_stability_failure_count": 1,
                        "frontend_stability_success_rate": 0.95,
                        "frontend_stability_passed": True,
                            "frontend_price_delta_pct": 0.15,
                            "frontend_search_status": "已读取部分结果",
                            "frontend_search_partial_evidence": True,
                            "frontend_competitor_count": 1,
                            "frontend_location_note": "UK SW1A 1AA 已设置",
                            "frontend_location_verified": True,
                        }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    rows = _build_frontend_check_queue(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "product_name": "Test product",
                "priority": "P0",
                "issue_type": "广告消耗无转化诊断",
                "primary_reason": "广告无单",
                "key_evidence": "点击 31；广告订单 0",
            }
        ],
        [],
        "UK",
        output_dir=output_dir,
    )

    assert rows[0]["frontend_auto_conclusion_label"] == "明确前台劣势"
    assert rows[0]["frontend_evidence_quality_score"] == 58
    assert rows[0]["frontend_check_method"] == "chrome-extension-20x-95pct"
    assert rows[0]["frontend_stability_total_attempts"] == 20
    assert rows[0]["frontend_stability_success_count"] == 19
    assert rows[0]["frontend_stability_failure_count"] == 1
    assert rows[0]["frontend_stability_success_rate"] == 0.95
    assert rows[0]["frontend_stability_passed"] is True
    assert rows[0]["frontend_price_delta_pct"] == 0.15
    assert rows[0]["frontend_evidence_tier"] == "仅背景参考"
    assert "搜索页仅部分读取" in rows[0]["frontend_evidence_audit_reasons"]
    assert "可比竞品少于2个" in rows[0]["frontend_evidence_audit_reasons"]


def test_frontend_queue_derives_location_gate_from_legacy_location_note(tmp_path) -> None:
    from src.report_presentation import _build_frontend_check_queue

    output_dir = tmp_path
    (output_dir / "frontend_check_results.json").write_text(
        json.dumps(
            {
                "items": [
                        {
                            "marketplace": "US",
                            "sku": "SKU-US",
                            "asin": "B0US123456",
                        "frontend_check_status": "已自动检查",
                        "frontend_findings": "售价：$18.99；评分：4.5 out of 5 stars；评论数：120 ratings；配送：US 10001 已设置",
                        "frontend_location_note": "US 10001 已设置",
                        "frontend_evidence_quality_score": 82,
                        "frontend_auto_conclusion": "FRONTEND_OK",
                            "frontend_search_status": "已自动检查",
                            "frontend_competitor_count": 3,
                            "competitor_comparability": "high",
                            "comparable_competitor_count": 3,
                        },
                    {
                        "marketplace": "UK",
                        "sku": "SKU-UK",
                        "asin": "B0UK123456",
                        "frontend_check_status": "已自动检查",
                        "frontend_findings": "售价：£18.99；评分：4.5 out of 5 stars；评论数：120 ratings；配送：Aberdeen AB10 1",
                        "frontend_location_note": "Aberdeen AB10 1",
                        "frontend_evidence_quality_score": 82,
                        "frontend_search_status": "已自动检查",
                        "frontend_competitor_count": 3,
                    },
                    {
                        "marketplace": "UK",
                        "sku": "SKU-UK-MIXED",
                        "asin": "B0UKMIXED1",
                        "frontend_check_status": "已自动检查",
                        "frontend_findings": "售价：£18.99；评分：4.5 out of 5 stars；配送：UK SW1A 1AA 已设置 United States",
                        "frontend_location_note": "UK SW1A 1AA 已设置 United States",
                        "frontend_location_scope": "exact",
                        "frontend_location_exact": True,
                        "frontend_location_verified": True,
                        "frontend_evidence_quality_score": 82,
                        "frontend_auto_conclusion": "FRONTEND_OK",
                            "frontend_search_status": "已自动检查",
                            "frontend_competitor_count": 3,
                            "competitor_comparability": "high",
                            "comparable_competitor_count": 3,
                        },
                    {
                        "marketplace": "UK",
                        "sku": "SKU-UK-UNVERIFIED",
                        "asin": "B0UKUNVER1",
                        "frontend_check_status": "已自动检查",
                        "frontend_findings": "售价：£18.99；评分：4.5 out of 5 stars；配送：UK SW1A 1AA 已设置",
                        "frontend_location_note": "UK SW1A 1AA 已设置",
                        "frontend_location_scope": "exact",
                        "frontend_location_exact": True,
                        "frontend_location_verified": False,
                        "frontend_evidence_quality_score": 82,
                        "frontend_auto_conclusion": "FRONTEND_OK",
                        "frontend_search_status": "已自动检查",
                        "frontend_competitor_count": 3,
                        "competitor_comparability": "high",
                        "comparable_competitor_count": 3,
                    },
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    rows = _build_frontend_check_queue(
        [
            {
                "marketplace": "US",
                "sku": "SKU-US",
                "asin": "B0US123456",
                "product_name": "US product",
                "priority": "P0",
                "issue_type": "广告消耗无转化诊断",
                "primary_reason": "广告无单",
            },
            {
                "marketplace": "UK",
                "sku": "SKU-UK",
                "asin": "B0UK123456",
                "product_name": "UK product",
                "priority": "P0",
                "issue_type": "广告消耗无转化诊断",
                "primary_reason": "广告无单",
            },
            {
                "marketplace": "UK",
                "sku": "SKU-UK-MIXED",
                "asin": "B0UKMIXED1",
                "product_name": "UK mixed location product",
                "priority": "P0",
                "issue_type": "广告消耗无转化诊断",
                "primary_reason": "广告无单",
            },
            {
                "marketplace": "UK",
                "sku": "SKU-UK-UNVERIFIED",
                "asin": "B0UKUNVER1",
                "product_name": "UK unverified exact location product",
                "priority": "P0",
                "issue_type": "广告消耗无转化诊断",
                "primary_reason": "广告无单",
            },
        ],
        [],
        "ALL",
        output_dir=output_dir,
        limit=5,
    )
    by_asin = {row["asin"]: row for row in rows}

    assert by_asin["B0US123456"]["frontend_location_scope"] == "exact"
    assert by_asin["B0US123456"]["frontend_location_exact"] is True
    assert by_asin["B0US123456"]["frontend_evidence_tier"] == "强诊断可用"
    assert by_asin["B0UK123456"]["frontend_location_scope"] == "marketplace"
    assert by_asin["B0UK123456"]["frontend_location_exact"] is False
    assert by_asin["B0UK123456"]["frontend_evidence_tier"] == "仅背景参考"
    assert by_asin["B0UKMIXED1"]["frontend_location_scope"] == "wrong"
    assert by_asin["B0UKMIXED1"]["frontend_location_exact"] is False
    assert "地区异常" in by_asin["B0UKMIXED1"]["frontend_location_warning"]
    assert by_asin["B0UKMIXED1"]["frontend_evidence_tier"] == "不可用"
    assert by_asin["B0UKUNVER1"]["frontend_location_scope"] == "exact"
    assert by_asin["B0UKUNVER1"]["frontend_location_exact"] is True
    assert by_asin["B0UKUNVER1"]["frontend_location_verified"] is False
    assert by_asin["B0UKUNVER1"]["frontend_evidence_display_tier"] == "仅背景参考"
    assert by_asin["B0UKUNVER1"]["frontend_evidence_is_strong"] is False
    assert "地区未确认" in by_asin["B0UKUNVER1"]["frontend_evidence_audit_reasons"]


def test_frontend_coverage_summary_counts_quality_gate_tiers() -> None:
    from src.report_presentation import _build_frontend_coverage_summary

    summary = _build_frontend_coverage_summary(
        [
            {
                "frontend_check_status": "已自动检查",
                "frontend_evidence_quality_score": 82,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_evidence_is_strong": True,
                "frontend_search_status": "已自动检查",
            },
            {
                "frontend_check_status": "沿用 2026-06-10 前台数据",
                "frontend_evidence_quality_score": 58,
                "frontend_evidence_tier": "仅背景参考",
                "frontend_search_status": "已读取部分结果",
                "frontend_search_partial_evidence": True,
            },
            {
                "frontend_check_status": "待前台检查",
                "frontend_evidence_quality_score": 0,
                "frontend_evidence_tier": "不可用",
                "frontend_price_currency_warning": "价格币种异常",
            },
        ]
    )

    assert summary["frontend_usable_evidence_count"] == 2
    assert summary["frontend_decision_ready_count"] == 1
    assert summary["frontend_reference_evidence_count"] == 1
    assert summary["frontend_strong_evidence_count"] == 1
    assert summary["frontend_background_evidence_count"] == 1
    assert summary["frontend_unusable_evidence_count"] == 1
    assert summary["frontend_live_success_count"] == 1
    assert summary["frontend_cached_count"] == 1
    assert summary["frontend_pending_or_stale_count"] == 2
    assert summary["frontend_search_success_count"] == 1
    assert summary["frontend_search_partial_count"] == 1
    assert summary["frontend_decision_ready_label"] == "1/3 强证据，33%"
    assert summary["frontend_reference_evidence_label"] == "1/3 背景参考，33%"


def test_frontend_coverage_summary_downgrades_unsafe_strong_display_tier() -> None:
    from src.report_presentation import _build_frontend_coverage_summary

    summary = _build_frontend_coverage_summary(
        [
            {
                "frontend_check_status": "已自动检查",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_evidence_display_tier": "仅背景参考",
                "frontend_evidence_is_strong": False,
                "frontend_search_status": "已自动检查",
            }
        ]
    )

    assert summary["frontend_usable_evidence_count"] == 1
    assert summary["frontend_decision_ready_count"] == 0
    assert summary["frontend_reference_evidence_count"] == 1
    assert summary["frontend_strong_evidence_count"] == 0
    assert summary["frontend_background_evidence_count"] == 1
    assert summary["frontend_decision_ready_label"] == "0/1 强证据，0%"
    assert summary["frontend_reference_evidence_label"] == "1/1 背景参考，100%"


def test_frontend_coverage_summary_does_not_count_unusable_tier_as_usable() -> None:
    from src.report_presentation import _build_frontend_coverage_summary

    summary = _build_frontend_coverage_summary(
        [
            {
                "frontend_check_status": "已自动检查",
                "frontend_evidence_quality_score": 58,
                "frontend_evidence_tier": "不可用",
                "frontend_evidence_display_tier": "不可用",
                "frontend_search_status": "已自动检查",
            }
        ]
    )

    assert summary["frontend_usable_evidence_count"] == 0
    assert summary["frontend_unusable_evidence_count"] == 1
    assert summary["frontend_coverage_label"] == "0/1 可用，0%"


def test_frontend_coverage_summary_treats_string_false_partial_flag_as_false() -> None:
    from src.report_presentation import _build_frontend_coverage_summary

    summary = _build_frontend_coverage_summary(
        [
            {
                "frontend_check_status": "已自动检查",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_evidence_display_tier": "强诊断可用",
                "frontend_evidence_is_strong": True,
                "frontend_search_status": "已自动检查",
                "frontend_search_partial_evidence": "False",
            }
        ]
    )

    assert summary["frontend_search_success_count"] == 1
    assert summary["frontend_search_partial_count"] == 0
    assert summary["frontend_decision_ready_count"] == 1
    assert summary["frontend_reference_evidence_count"] == 0


def test_frontend_coverage_summary_requires_explicit_strong_flag() -> None:
    from src.report_presentation import _build_frontend_coverage_summary

    summary = _build_frontend_coverage_summary(
        [
            {
                "frontend_check_status": "已自动检查",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_evidence_display_tier": "强诊断可用",
                "frontend_search_status": "已自动检查",
            }
        ]
    )

    assert summary["frontend_decision_ready_count"] == 0
    assert summary["frontend_reference_evidence_count"] == 1
    assert summary["frontend_strong_evidence_count"] == 0
    assert summary["frontend_background_evidence_count"] == 1
    assert summary["frontend_decision_ready_label"] == "0/1 强证据，0%"
    assert summary["frontend_reference_evidence_label"] == "1/1 背景参考，100%"


def test_product_final_decisions_classify_ad_traffic_problem() -> None:
    from src.product_decision_layer import build_product_final_decisions

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-1",
                    "asin": "B0TEST1234",
                    "product_name": "Test product",
                    "ad_clicks": 24,
                    "ad_spend": 18,
                    "ad_orders": 0,
                    "total_orders": 2,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    rows = build_product_final_decisions(
        payload,
        search_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "suggested_action": "否定精准",
                "search_term_or_target": "wrong size keyword",
                "relevance_level": "明显不相关",
                "clicks": "4",
                "spend": "£2.00",
            }
        ],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 82,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_failure_category": "none",
                "frontend_product_quality_score": 82,
                "frontend_search_quality_score": 80,
                "frontend_buy_box": "识别到购买按钮",
                "frontend_coupon": "10%",
            }
        ],
    )

    assert rows[0]["fusion_issue_type"] == "广告流量问题"
    assert rows[0]["fusion_action_gate"] == "tighten_ads_first"
    assert rows[0]["final_decision"] == "EXECUTE_TODAY"
    assert "negative_exact" in rows[0]["today_allowed_actions"]
    assert "bid_up" in rows[0]["today_blocked_actions"]
    assert "budget_up" in rows[0]["today_blocked_actions"]
    assert "broad_scale" in rows[0]["today_blocked_actions"]
    assert rows[0]["frontend_posture"] != "insufficient_evidence"
    assert "no_fresh_frontend" not in rows[0]["fusion_evidence_flags"]


def test_product_final_decisions_classify_joint_frontend_and_ad_problem() -> None:
    from src.product_decision_layer import build_product_final_decisions

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-2",
                    "asin": "B0TEST5678",
                    "product_name": "Weak product",
                    "ad_clicks": 31,
                    "ad_spend": 22,
                    "ad_orders": 0,
                    "total_orders": 1,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    rows = build_product_final_decisions(
        payload,
        search_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-2",
                "asin": "B0TEST5678",
                "suggested_action": "降竞价10%-20%",
                "search_term_or_target": "broad keyword",
                "relevance_level": "泛词/竞品/待确认",
                "clicks": "12",
                "spend": "£8.00",
            }
        ],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-2",
                "asin": "B0TEST5678",
                "frontend_check_status": "已自动检查",
                "frontend_auto_conclusion": "FRONTEND_WEAK",
                "frontend_auto_conclusion_label": "明确前台劣势",
                "frontend_evidence_quality_score": 55,
                "frontend_product_quality_score": 48,
                "frontend_search_quality_score": 62,
                "frontend_buy_box": "未稳定识别 Buy Box/购买按钮",
                "frontend_coupon": "未稳定识别 Coupon",
                "frontend_price_delta_pct": 0.18,
            }
        ],
    )

    assert rows[0]["fusion_issue_type"] == "前台和广告共同问题"
    assert rows[0]["fusion_action_gate"] == "fix_both"
    assert "不扩预算" in rows[0]["fusion_do_not_do"]


def test_frontend_ok_coupon_pending_does_not_become_frontend_problem() -> None:
    from src.product_decision_layer import build_product_final_decisions

    payload = {
        "target_marketplace": "US",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "US",
                    "sku": "SKU-OK",
                    "asin": "B0OK123456",
                    "product_name": "Frontend OK product",
                    "ad_clicks": 28,
                    "ad_spend": 15,
                    "ad_orders": 0,
                    "total_orders": 2,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }

    rows = build_product_final_decisions(
        payload,
        search_rows=[
            {
                "marketplace": "US",
                "sku": "SKU-OK",
                "asin": "B0OK123456",
                "suggested_action": "否定精准",
                "search_term_or_target": "wrong keyword",
                "relevance_level": "明显不相关",
                "clicks": "4",
                "spend": "$2.00",
            }
        ],
        frontend_rows=[
            {
                "marketplace": "US",
                "sku": "SKU-OK",
                "asin": "B0OK123456",
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 90,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_buy_box": "识别到购买按钮",
                "frontend_coupon": "待确认",
                "frontend_competitors": [
                    {"asin": "B0COMP1111", "title": "Frontend OK product alternative"},
                    {"asin": "B0COMP2222", "title": "Frontend OK product premium"},
                ],
            }
        ],
    )

    assert rows[0]["fusion_issue_type"] == "广告流量问题"
    assert rows[0]["fusion_action_gate"] == "tighten_ads_first"
    assert "frontend_competitiveness" not in rows[0]["fusion_evidence_flags"]
    assert rows[0]["final_decision"] == "EXECUTE_TODAY"


def test_frontend_failed_read_is_insufficient_evidence_not_competitiveness() -> None:
    from src.product_decision_layer import build_product_final_decisions

    payload = {
        "target_marketplace": "DE",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "DE",
                    "sku": "SKU-BLOCKED",
                    "asin": "B0BAD12345",
                    "product_name": "Blocked frontend product",
                    "ad_clicks": 24,
                    "ad_spend": 16,
                    "ad_orders": 0,
                    "total_orders": 0,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }

    rows = build_product_final_decisions(
        payload,
        frontend_rows=[
            {
                "marketplace": "DE",
                "sku": "SKU-BLOCKED",
                "asin": "B0BAD12345",
                "frontend_auto_conclusion": "INSUFFICIENT_EVIDENCE",
                "frontend_auto_conclusion_label": "自动证据不足，不能用于强诊断",
                "frontend_evidence_quality_score": 0,
                "frontend_price_currency_warning": "价格币种异常：TWD432.81，已忽略",
                "frontend_coupon": "待确认",
                "frontend_buy_box": "未稳定识别 Buy Box/购买按钮",
            }
        ],
    )

    assert rows[0]["fusion_issue_type"] == "广告流量问题"
    assert "no_fresh_frontend" in rows[0]["fusion_evidence_flags"]
    assert "frontend_competitiveness" not in rows[0]["fusion_evidence_flags"]
    assert rows[0]["frontend_posture"] == "insufficient_evidence"


def test_product_decision_explicitly_blocks_growth_when_frontend_missing() -> None:
    from src.product_decision_layer import build_product_final_decisions

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-NO-FRONTEND",
                    "asin": "B0NOFRONT",
                    "product_name": "No frontend product",
                    "ad_clicks": 24,
                    "ad_spend": 18,
                    "ad_orders": 0,
                    "total_orders": 2,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }

    rows = build_product_final_decisions(
        payload,
        search_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-NO-FRONTEND",
                "asin": "B0NOFRONT",
                "suggested_action": "否定精准",
                "search_term_or_target": "wrong size keyword",
                "relevance_level": "明显不相关",
                "clicks": "4",
                "spend": "£2.00",
            }
        ],
        frontend_rows=[],
    )

    assert rows[0]["final_decision"] == "EXECUTE_TODAY"
    assert "negative_exact" in rows[0]["today_allowed_actions"]
    assert "bid_up" in rows[0]["today_blocked_actions"]
    assert "budget_up" in rows[0]["today_blocked_actions"]
    assert "broad_scale" in rows[0]["today_blocked_actions"]
    assert "bid_up" not in rows[0]["today_allowed_actions"]


def test_non_executable_product_gate_keeps_ad_loss_desk_actions() -> None:
    from src.product_decision_layer import filter_ad_queue_by_decision

    decisions = [
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "final_decision": "CONSERVATIVE_RUN",
            "final_decision_label": "保守跑",
            "decision_reason": "前台或库存限制，不加预算。",
            "today_allowed_actions": ["observe", "bid_down", "negative_exact", "pause"],
            "today_blocked_actions": ["bid_up", "budget_up", "broad_scale"],
        }
    ]
    rows = [
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "wrong keyword",
            "suggested_action": "否定精准",
            "copy_action_line": "建议否词",
        },
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "broad keyword",
            "suggested_action": "降竞价10%-20%",
            "copy_action_line": "建议降竞价 10%-20%",
        },
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "B0TARGET123",
            "suggested_action": "暂停ASIN定向",
            "copy_action_line": "建议暂停 ASIN 定向",
        },
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "search_term_or_target": "core keyword",
            "suggested_action": "加价 5%-10%",
            "copy_action_line": "建议加价 5%-10%",
        },
    ]

    filtered = filter_ad_queue_by_decision(rows, decisions)

    assert [row["suggested_action"] for row in filtered] == [
        "否定精准",
        "降竞价10%-20%",
        "暂停ASIN定向",
        "观察",
    ]


def test_executable_product_gate_still_blocks_bid_up_when_action_blocked() -> None:
    from src.product_decision_layer import filter_ad_queue_by_decision

    decisions = [
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "final_decision": "EXECUTE_TODAY",
            "today_allowed_actions": ["observe", "bid_down", "negative_exact", "pause"],
            "today_blocked_actions": ["bid_up", "budget_up", "broad_scale"],
        }
    ]

    filtered = filter_ad_queue_by_decision(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "search_term_or_target": "verified keyword",
                "scale_action": "提高竞价 5%-10%",
                "copy_action_line": "建议加价 5%-10%",
                "action_id": "UK||SKU-1||B0TEST1234||search_term||verified keyword||bid_up",
            }
        ],
        decisions,
    )

    assert filtered[0]["suggested_action"] == "观察"
    assert filtered[0]["scale_action"] == "观察"
    assert filtered[0]["normalized_action"] == "observe"
    assert filtered[0]["copy_block"] == "建议观察\nverified keyword"
    assert filtered[0]["ad_gate_blocked"] is True
    assert filtered[0]["blocked_original_action"] == "提高竞价 5%-10%"
    assert filtered[0]["blocked_original_action_id"] == "UK||SKU-1||B0TEST1234||search_term||verified keyword||bid_up"
    assert filtered[0]["action_id"] == "UK||SKU-1||B0TEST1234||search_term||verified keyword||observe"


def test_product_gate_observation_action_id_does_not_use_na_for_empty_target() -> None:
    from src.product_decision_layer import filter_ad_queue_by_decision

    decisions = [
        {
            "marketplace": "US",
            "sku": "SKU-EMPTY-TARGET",
            "asin": "B0EMPTYTGT",
            "final_decision": "CONSERVATIVE_RUN",
            "today_allowed_actions": ["observe", "bid_down", "negative_exact", "pause"],
            "today_blocked_actions": ["bid_up", "budget_up", "broad_scale"],
        }
    ]

    filtered = filter_ad_queue_by_decision(
        [
            {
                "marketplace": "US",
                "sku": "SKU-EMPTY-TARGET",
                "asin": "B0EMPTYTGT",
                "search_term_or_target": "",
                "action_scope": "search_term",
                "scale_action": "提高竞价 5%-10%",
                "copy_action_line": "建议加价 5%-10%",
                "action_id": "US||SKU-EMPTY-TARGET||B0EMPTYTGT||search_term||N/A||bid_up",
            }
        ],
        decisions,
    )

    assert filtered[0]["suggested_action"] == "观察"
    assert filtered[0]["action_id"] == "US||SKU-EMPTY-TARGET||B0EMPTYTGT||search_term||||observe"
    assert "N/A" not in filtered[0]["action_id"]


def test_product_gate_observation_action_id_lowercases_asin_target() -> None:
    from src.product_decision_layer import filter_ad_queue_by_decision

    decisions = [
        {
            "marketplace": "US",
            "sku": "SKU-ASIN-TARGET",
            "asin": "B0PARENT001",
            "final_decision": "CONSERVATIVE_RUN",
            "today_allowed_actions": ["observe", "bid_down", "negative_exact", "pause"],
            "today_blocked_actions": ["bid_up", "budget_up", "broad_scale"],
        }
    ]

    filtered = filter_ad_queue_by_decision(
        [
            {
                "marketplace": "US",
                "sku": "SKU-ASIN-TARGET",
                "asin": "B0PARENT001",
                "search_term_or_target": "B0CHILD001",
                "action_scope": "asin_target",
                "scale_action": "提高竞价 5%-10%",
                "copy_action_line": "建议加价 5%-10%",
                "action_id": "US||SKU-ASIN-TARGET||B0PARENT001||asin_target||B0CHILD001||bid_up",
            }
        ],
        decisions,
    )

    assert filtered[0]["suggested_action"] == "观察"
    assert filtered[0]["action_id"] == "US||SKU-ASIN-TARGET||B0PARENT001||asin_target||b0child001||observe"


def _assert_growth_gate_blocks_small_budget_test(decision: dict[str, object]) -> None:
    from src.product_decision_layer import filter_ad_queue_by_decision

    assert "create_exact_low_budget" in decision["today_blocked_actions"]
    assert "create_exact_low_budget" not in decision["today_allowed_actions"]

    filtered = filter_ad_queue_by_decision(
        [
            {
                "marketplace": decision["marketplace"],
                "sku": decision["sku"],
                "asin": decision["asin"],
                "search_term_or_target": "verified keyword",
                "suggested_action": "小预算试投",
                "copy_action_line": "创建精准小预算",
            }
        ],
        [decision],
    )

    assert filtered[0]["suggested_action"] == "观察"
    assert filtered[0]["normalized_action"] == "observe"
    assert filtered[0]["copy_block"] == "建议观察\nverified keyword"
    assert filtered[0]["ad_gate_blocked"] is True
    assert filtered[0]["blocked_original_action"] == "小预算试投"


def test_product_decision_clears_blocked_growth_fields_from_today_task_rows() -> None:
    from src.product_decision_layer import apply_decisions_to_rows

    rows = [
        {
            "marketplace": "UK",
            "sku": "SKU-TASK-GATE",
            "asin": "B0TASKGATE",
            "product_name": "Blocked task product",
            "issue_type": "Listing 待人工确认",
            "today_action": "暂时不加广告预算；等人工确认后再决定。",
            "suggested_action": "小预算试投",
            "normalized_action": "growth_test",
            "manual_action_taken": "小预算试投",
            "action_id": "UK||SKU-TASK-GATE||B0TASKGATE||search_term||blocked exact test||growth_test",
            "search_term_or_target": "blocked exact test",
            "confirmed_status": "待复查",
        }
    ]
    decisions = [
        {
            "marketplace": "UK",
            "sku": "SKU-TASK-GATE",
            "asin": "B0TASKGATE",
            "final_decision": "WAIT_REVIEW",
            "final_decision_label": "等复盘",
            "today_allowed_actions": ["observe"],
            "today_blocked_actions": ["bid_up", "budget_up", "broad_scale", "create_exact_low_budget"],
        }
    ]

    updated = apply_decisions_to_rows(rows, decisions)

    assert updated[0]["suggested_action"] == "观察"
    assert updated[0]["normalized_action"] == "observe"
    assert updated[0]["manual_action_taken"] == "观察"
    assert updated[0]["copy_block"] == "建议观察\nblocked exact test"
    assert updated[0]["ad_gate_blocked"] is True
    assert updated[0]["blocked_original_action"] == "小预算试投"
    assert updated[0]["blocked_original_action_id"] == "UK||SKU-TASK-GATE||B0TASKGATE||search_term||blocked exact test||growth_test"
    assert updated[0]["action_id"] == "UK||SKU-TASK-GATE||B0TASKGATE||search_term||blocked exact test||observe"


def test_product_decision_allows_bid_up_only_with_frontend_ok_and_verified_term() -> None:
    from src.product_decision_layer import build_product_final_decisions, filter_ad_queue_by_decision

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-SCALE",
                    "asin": "B0SCALE123",
                    "product_name": "Metal dimmer desk lamp",
                    "ad_clicks": 42,
                    "ad_spend": 12,
                    "ad_orders": 4,
                    "ad_sales": 80,
                    "total_orders": 6,
                    "ACOS": 0.12,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-SCALE",
        "asin": "B0SCALE123",
        "search_term_or_target": "verified keyword",
        "scale_action": "提高竞价 5%-10%",
        "ad_orders": "3",
        "promoted_ad_orders": "3",
        "ad_sales": "£60.00",
        "ACOS": "12%",
        "target_acos": "25%",
    }

    decisions = build_product_final_decisions(
        payload,
        scale_rows=[scale_row],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-SCALE",
                "asin": "B0SCALE123",
                "frontend_check_status": "已自动检查",
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "market_survey_completeness_score": 92,
                "market_survey_completeness_level": "complete",
                "market_survey_decision_evidence_tier": "strong",
                "frontend_buy_box": "识别到购买按钮",
                "frontend_coupon": "10%",
                "frontend_search_status": "已自动检查",
                "frontend_location_scope": "exact",
                "frontend_location_verified": True,
                "frontend_location_exact": True,
                "competitor_comparability": "high",
                "comparable_competitor_count": 3,
                "frontend_competitors": [
                    {"asin": "B0COMP0001", "title": "Metal dimmer desk lamp with brightness setting"},
                    {"asin": "B0COMP0002", "title": "Large metal dimmer desk lamp slicer"},
                    {"asin": "B0COMP0003", "title": "Metal desk lamp for dimmer"},
                ],
            }
        ],
    )
    filtered = filter_ad_queue_by_decision([scale_row], decisions)

    assert "bid_up" in decisions[0]["today_allowed_actions"]
    assert decisions[0]["frontend_check_status"] == "已自动检查"
    assert decisions[0]["frontend_auto_conclusion"] == "FRONTEND_OK"
    assert decisions[0]["frontend_cache_used"] is False
    assert filtered[0]["scale_action"] == "提高竞价 5%-10%"


def test_product_scale_rows_are_filtered_by_final_growth_gate() -> None:
    from src import report_presentation as presentation

    scale_rows = [
        {
            "站点": "UK",
            "产品": "Metal dimmer desk lamp",
            "SKU": "SKU-SCALE",
            "ASIN": "B0SCALE123",
            "放量等级": "可小幅放量",
            "建议": "表现稳定，可逐步加预算或提竞价。",
        }
    ]
    decision_rows = [
        {
            "marketplace": "UK",
            "sku": "SKU-SCALE",
            "asin": "B0SCALE123",
            "today_allowed_actions": ["observe"],
            "today_blocked_actions": ["bid_up", "budget_up", "broad_scale"],
        }
    ]

    assert presentation._filter_product_scale_rows_by_decision(scale_rows, decision_rows) == []


def test_product_scale_rows_remain_when_growth_gate_allows_actions() -> None:
    from src import report_presentation as presentation

    scale_row = {
        "站点": "UK",
        "产品": "Metal dimmer desk lamp",
        "SKU": "SKU-SCALE",
        "ASIN": "B0SCALE123",
        "放量等级": "可小幅放量",
        "建议": "表现稳定，可逐步加预算或提竞价。",
    }
    decision_rows = [
        {
            "marketplace": "UK",
            "sku": "SKU-SCALE",
            "asin": "B0SCALE123",
            "today_allowed_actions": ["observe", "bid_up", "budget_up"],
            "today_blocked_actions": [],
        }
    ]

    assert presentation._filter_product_scale_rows_by_decision([scale_row], decision_rows) == [scale_row]


def test_product_decision_blocks_bid_up_when_orders_are_halo_only() -> None:
    from src.product_decision_layer import build_product_final_decisions, filter_ad_queue_by_decision

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-SCALE",
                    "asin": "B0SCALE123",
                    "product_name": "Metal dimmer desk lamp",
                    "ad_clicks": 42,
                    "ad_spend": 12,
                    "ad_orders": 4,
                    "ad_sales": 80,
                    "total_orders": 6,
                    "ACOS": 0.12,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-SCALE",
        "asin": "B0SCALE123",
        "search_term_or_target": "halo only keyword",
        "scale_action": "提高竞价 5%-10%",
        "ad_orders": "3",
        "promoted_ad_orders": "0",
        "halo_ad_orders": "3",
        "halo_only_conversion": True,
        "ad_sales": "£60.00",
        "ACOS": "12%",
        "target_acos": "25%",
    }

    decisions = build_product_final_decisions(
        payload,
        scale_rows=[scale_row],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-SCALE",
                "asin": "B0SCALE123",
                "frontend_check_status": "已自动检查",
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "market_survey_completeness_score": 92,
                "market_survey_completeness_level": "complete",
                "market_survey_decision_evidence_tier": "strong",
                "frontend_buy_box": "识别到购买按钮",
                "frontend_coupon": "10%",
                "frontend_search_status": "已自动检查",
                "frontend_location_scope": "exact",
                "frontend_location_verified": True,
                "frontend_location_exact": True,
                "competitor_comparability": "high",
                "comparable_competitor_count": 3,
                "frontend_competitors": [
                    {"asin": "B0COMP0001", "title": "Metal dimmer desk lamp with brightness setting"},
                    {"asin": "B0COMP0002", "title": "Large metal dimmer desk lamp slicer"},
                    {"asin": "B0COMP0003", "title": "Metal desk lamp for dimmer"},
                ],
            }
        ],
    )
    filtered = filter_ad_queue_by_decision([scale_row], decisions)

    assert "bid_up" not in decisions[0]["today_allowed_actions"]
    assert "bid_up" in decisions[0]["today_blocked_actions"]
    assert filtered[0]["scale_action"] == "观察"
    assert filtered[0]["ad_gate_blocked"] is True


def test_product_decision_blocks_bid_up_when_frontend_ok_code_is_missing() -> None:
    from src.product_decision_layer import build_product_final_decisions, filter_ad_queue_by_decision

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-SCALE",
                    "asin": "B0SCALE123",
                    "product_name": "Metal dimmer desk lamp",
                    "ad_clicks": 42,
                    "ad_spend": 12,
                    "ad_orders": 4,
                    "ad_sales": 80,
                    "total_orders": 6,
                    "ACOS": 0.12,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-SCALE",
        "asin": "B0SCALE123",
        "search_term_or_target": "verified keyword",
        "scale_action": "提高竞价 5%-10%",
        "ad_orders": "3",
        "promoted_ad_orders": "3",
        "ad_sales": "£60.00",
        "ACOS": "12%",
        "target_acos": "25%",
    }

    decisions = build_product_final_decisions(
        payload,
        scale_rows=[scale_row],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-SCALE",
                "asin": "B0SCALE123",
                "frontend_check_status": "已自动检查",
                "frontend_auto_conclusion": "",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "market_survey_completeness_score": 92,
                "market_survey_completeness_level": "complete",
                "market_survey_decision_evidence_tier": "strong",
                "frontend_buy_box": "识别到购买按钮",
                "frontend_coupon": "10%",
                "frontend_search_status": "已自动检查",
                "frontend_location_scope": "exact",
                "frontend_location_verified": True,
                "frontend_location_exact": True,
                "competitor_comparability": "high",
                "comparable_competitor_count": 3,
                "frontend_competitors": [
                    {"asin": "B0COMP0001", "title": "Metal dimmer desk lamp with brightness setting"},
                    {"asin": "B0COMP0002", "title": "Large metal dimmer desk lamp slicer"},
                    {"asin": "B0COMP0003", "title": "Metal desk lamp for dimmer"},
                ],
            }
        ],
    )
    filtered = filter_ad_queue_by_decision([scale_row], decisions)

    assert decisions[0]["frontend_evidence_state"] == "mixed"
    assert "bid_up" not in decisions[0]["today_allowed_actions"]
    assert "bid_up" in decisions[0]["today_blocked_actions"]
    assert filtered[0]["scale_action"] == "观察"


def test_product_decision_blocks_bid_up_when_frontend_search_is_partial() -> None:
    from src.product_decision_layer import build_product_final_decisions, filter_ad_queue_by_decision

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-SCALE",
                    "asin": "B0SCALE123",
                    "product_name": "Metal dimmer desk lamp",
                    "ad_clicks": 42,
                    "ad_spend": 12,
                    "ad_orders": 4,
                    "ad_sales": 80,
                    "total_orders": 6,
                    "ACOS": 0.12,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-SCALE",
        "asin": "B0SCALE123",
        "search_term_or_target": "verified keyword",
        "scale_action": "提高竞价 5%-10%",
        "ad_orders": "3",
        "promoted_ad_orders": "3",
        "ad_sales": "£60.00",
        "ACOS": "12%",
        "target_acos": "25%",
    }

    decisions = build_product_final_decisions(
        payload,
        scale_rows=[scale_row],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-SCALE",
                "asin": "B0SCALE123",
                "frontend_check_status": "已自动检查",
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "market_survey_completeness_score": 92,
                "market_survey_completeness_level": "complete",
                "market_survey_decision_evidence_tier": "strong",
                "frontend_buy_box": "识别到购买按钮",
                "frontend_coupon": "10%",
                "frontend_search_status": "已读取部分结果",
                "frontend_search_partial_evidence": True,
                "frontend_location_scope": "exact",
                "frontend_location_verified": True,
                "frontend_location_exact": True,
                "competitor_comparability": "high",
                "comparable_competitor_count": 3,
                "frontend_competitors": [
                    {"asin": "B0COMP0001", "title": "Metal dimmer desk lamp with brightness setting"},
                    {"asin": "B0COMP0002", "title": "Large metal dimmer desk lamp slicer"},
                    {"asin": "B0COMP0003", "title": "Metal desk lamp for dimmer"},
                ],
            }
        ],
    )
    filtered = filter_ad_queue_by_decision([scale_row], decisions)

    assert decisions[0]["frontend_evidence_state"] == "mixed"
    assert decisions[0]["frontend_search_status"] == "已读取部分结果"
    assert decisions[0]["frontend_search_partial_evidence"] is True
    assert "搜索页仅部分读取，竞品证据仅背景参考" in decisions[0]["frontend_blocking_reasons"]
    assert "bid_up" in decisions[0]["today_blocked_actions"]
    assert "budget_up" in decisions[0]["today_blocked_actions"]
    assert "bid_up" not in decisions[0]["today_allowed_actions"]
    _assert_growth_gate_blocks_small_budget_test(decisions[0])
    assert filtered[0]["scale_action"] == "观察"


def test_product_decision_allows_bid_up_with_percent_acos_and_decimal_target() -> None:
    from src.product_decision_layer import build_product_final_decisions, filter_ad_queue_by_decision

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-SCALE",
                    "asin": "B0SCALE123",
                    "product_name": "Metal dimmer desk lamp",
                    "ad_clicks": 42,
                    "ad_spend": 12,
                    "ad_orders": 4,
                    "ad_sales": 80,
                    "total_orders": 6,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-SCALE",
        "asin": "B0SCALE123",
        "search_term_or_target": "verified keyword",
        "scale_action": "提高竞价 5%-10%",
        "ad_orders": "3",
        "promoted_ad_orders": "3",
        "ad_sales": "£60.00",
        "ACOS": "12%",
    }

    decisions = build_product_final_decisions(
        payload,
        scale_rows=[scale_row],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-SCALE",
                "asin": "B0SCALE123",
                "frontend_check_status": "已自动检查",
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "market_survey_completeness_score": 92,
                "market_survey_completeness_level": "complete",
                "market_survey_decision_evidence_tier": "strong",
                "frontend_buy_box": "识别到购买按钮",
                "frontend_search_status": "已自动检查",
                "frontend_location_scope": "exact",
                "frontend_location_verified": True,
                "frontend_location_exact": True,
                "competitor_comparability": "high",
                "comparable_competitor_count": 3,
                "frontend_competitors": [
                    {"asin": "B0COMP0001", "title": "Metal dimmer desk lamp with brightness setting"},
                    {"asin": "B0COMP0002", "title": "Large metal dimmer desk lamp slicer"},
                    {"asin": "B0COMP0003", "title": "Metal desk lamp for dimmer"},
                ],
            }
        ],
    )
    filtered = filter_ad_queue_by_decision([scale_row], decisions)

    assert "bid_up" in decisions[0]["today_allowed_actions"]
    assert filtered[0]["scale_action"] == "提高竞价 5%-10%"


def test_product_decision_allows_bid_up_with_integer_percent_acos_and_decimal_target() -> None:
    from src.product_decision_layer import build_product_final_decisions, filter_ad_queue_by_decision

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-SCALE",
                    "asin": "B0SCALE123",
                    "product_name": "Metal dimmer desk lamp",
                    "ad_clicks": 42,
                    "ad_spend": 12,
                    "ad_orders": 4,
                    "ad_sales": 80,
                    "total_orders": 6,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-SCALE",
        "asin": "B0SCALE123",
        "search_term_or_target": "verified keyword",
        "scale_action": "提高竞价 5%-10%",
        "ad_orders": "3",
        "promoted_ad_orders": "3",
        "ad_sales": "£60.00",
        "ACOS": 12,
    }

    decisions = build_product_final_decisions(
        payload,
        scale_rows=[scale_row],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-SCALE",
                "asin": "B0SCALE123",
                "frontend_check_status": "已自动检查",
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "market_survey_completeness_score": 92,
                "market_survey_completeness_level": "complete",
                "market_survey_decision_evidence_tier": "strong",
                "frontend_buy_box": "识别到购买按钮",
                "frontend_search_status": "已自动检查",
                "frontend_location_scope": "exact",
                "frontend_location_verified": True,
                "frontend_location_exact": True,
                "competitor_comparability": "high",
                "comparable_competitor_count": 3,
                "frontend_competitors": [
                    {"asin": "B0COMP0001", "title": "Metal dimmer desk lamp with brightness setting"},
                    {"asin": "B0COMP0002", "title": "Large metal dimmer desk lamp slicer"},
                    {"asin": "B0COMP0003", "title": "Metal desk lamp for dimmer"},
                ],
            }
        ],
    )
    filtered = filter_ad_queue_by_decision([scale_row], decisions)

    assert "bid_up" in decisions[0]["today_allowed_actions"]
    assert filtered[0]["scale_action"] == "提高竞价 5%-10%"


def test_product_decision_blocks_bid_up_when_frontend_is_background_only() -> None:
    from src.product_decision_layer import build_product_final_decisions, filter_ad_queue_by_decision

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-SCALE",
                    "asin": "B0SCALE123",
                    "product_name": "Scale product",
                    "ad_clicks": 42,
                    "ad_spend": 12,
                    "ad_orders": 4,
                    "ad_sales": 80,
                    "total_orders": 6,
                    "ACOS": 0.12,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-SCALE",
        "asin": "B0SCALE123",
        "search_term_or_target": "verified keyword",
        "scale_action": "提高竞价 5%-10%",
        "ad_orders": "3",
        "ad_sales": "£60.00",
        "ACOS": "12%",
        "target_acos": "25%",
    }

    decisions = build_product_final_decisions(
        payload,
        scale_rows=[scale_row],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-SCALE",
                "asin": "B0SCALE123",
                "frontend_check_status": "已自动检查",
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 58,
                "frontend_evidence_tier": "仅背景参考",
                "frontend_evidence_audit_reasons": ["沿用缓存", "搜索页仅部分读取"],
            }
        ],
    )
    filtered = filter_ad_queue_by_decision([scale_row], decisions)

    assert "bid_up" not in decisions[0]["today_allowed_actions"]
    assert "bid_up" in decisions[0]["today_blocked_actions"]
    assert decisions[0]["frontend_evidence_state"] == "background"
    assert "仅背景参考" in decisions[0]["fusion_reason"]
    assert filtered[0]["scale_action"] == "观察"


def test_product_decision_blocks_bid_up_when_frontend_ok_is_from_cache() -> None:
    from src.product_decision_layer import build_product_final_decisions, filter_ad_queue_by_decision

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-SCALE",
                    "asin": "B0SCALE123",
                    "product_name": "Metal dimmer desk lamp",
                    "ad_clicks": 42,
                    "ad_spend": 12,
                    "ad_orders": 4,
                    "ad_sales": 80,
                    "total_orders": 6,
                    "ACOS": 0.12,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-SCALE",
        "asin": "B0SCALE123",
        "search_term_or_target": "verified keyword",
        "scale_action": "提高竞价 5%-10%",
        "ad_orders": "3",
        "ad_sales": "£60.00",
        "ACOS": "12%",
        "target_acos": "25%",
    }

    decisions = build_product_final_decisions(
        payload,
        scale_rows=[scale_row],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-SCALE",
                "asin": "B0SCALE123",
                "frontend_check_status": "沿用 2026-06-10 前台数据",
                "frontend_cache_used": True,
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_buy_box": "识别到购买按钮",
                "frontend_coupon": "10%",
                "frontend_location_scope": "exact",
                "frontend_location_verified": True,
                "frontend_location_exact": True,
                "frontend_competitors": [
                    {"asin": "B0COMP0001", "title": "Metal dimmer desk lamp with brightness setting"},
                    {"asin": "B0COMP0002", "title": "Large metal dimmer desk lamp slicer"},
                ],
            }
        ],
    )
    filtered = filter_ad_queue_by_decision([scale_row], decisions)

    assert decisions[0]["final_decision"] == "FRONTEND_FIRST"
    assert decisions[0]["frontend_evidence_state"] == "background"
    assert decisions[0]["frontend_evidence_tier"] == "强诊断可用"
    assert decisions[0]["frontend_evidence_display_tier"] == "仅背景参考"
    assert decisions[0]["frontend_decision_evidence_tier"] == "仅背景参考"
    assert "前台缓存或读取失败待确认，仅背景参考" in "；".join(decisions[0]["frontend_blocking_reasons"])
    assert "bid_up" in decisions[0]["today_blocked_actions"]
    assert "budget_up" in decisions[0]["today_blocked_actions"]
    assert "broad_scale" in decisions[0]["today_blocked_actions"]
    assert "create_exact_low_budget" in decisions[0]["today_blocked_actions"]
    assert "bid_up" not in decisions[0]["today_allowed_actions"]
    assert "create_exact_low_budget" not in decisions[0]["today_allowed_actions"]
    _assert_growth_gate_blocks_small_budget_test(decisions[0])
    assert filtered[0]["scale_action"] == "观察"


def test_product_final_decision_source_exports_frontend_strong_flag() -> None:
    from src.product_decision_layer import PRODUCT_FINAL_DECISION_REQUIRED_FIELDS, build_product_final_decisions

    assert "frontend_evidence_is_strong" in PRODUCT_FINAL_DECISION_REQUIRED_FIELDS
    assert "frontend_decision_evidence_tier" in PRODUCT_FINAL_DECISION_REQUIRED_FIELDS

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-STRONG",
                    "asin": "B0STRONG01",
                    "product_name": "Metal dimmer desk lamp",
                    "ad_clicks": 18,
                    "ad_spend": 8,
                    "ad_orders": 2,
                    "ad_sales": 60,
                    "total_orders": 4,
                    "ACOS": 0.12,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }

    rows = build_product_final_decisions(
        payload,
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-STRONG",
                "asin": "B0STRONG01",
                "frontend_check_status": "已自动检查",
                "frontend_search_status": "已自动检查",
                "frontend_cache_used": False,
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_failure_category": "none",
                "frontend_location_scope": "exact",
                "frontend_location_verified": True,
                "frontend_location_exact": True,
                "competitor_comparability": "high",
                "comparable_competitor_count": 3,
                "frontend_competitors": [
                    {"asin": "B0COMP0001", "title": "Metal dimmer desk lamp with brightness setting"},
                    {"asin": "B0COMP0002", "title": "Large metal dimmer desk lamp slicer"},
                    {"asin": "B0COMP0003", "title": "Metal desk lamp for dimmer"},
                ],
            }
        ],
    )

    assert rows[0]["frontend_evidence_is_strong"] is True


def test_product_decision_blocks_bid_up_when_frontend_currency_is_wrong() -> None:
    from src.product_decision_layer import build_product_final_decisions, filter_ad_queue_by_decision

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-CURRENCY",
                    "asin": "B0CURR1234",
                    "product_name": "Metal dimmer desk lamp",
                    "ad_clicks": 42,
                    "ad_spend": 12,
                    "ad_orders": 4,
                    "ad_sales": 80,
                    "total_orders": 6,
                    "ACOS": 0.12,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-CURRENCY",
        "asin": "B0CURR1234",
        "search_term_or_target": "verified keyword",
        "scale_action": "提高竞价 5%-10%",
        "ad_orders": "3",
        "ad_sales": "£60.00",
        "ACOS": "12%",
        "target_acos": "25%",
    }

    decisions = build_product_final_decisions(
        payload,
        scale_rows=[scale_row],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-CURRENCY",
                "asin": "B0CURR1234",
                "frontend_check_status": "已自动检查",
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_price_currency_warning": "价格币种异常：TWD594.77，已忽略",
                "frontend_failure_category": "currency_mismatch",
                "frontend_buy_box": "识别到购买按钮",
                "frontend_coupon": "10%",
                "frontend_location_scope": "exact",
                "frontend_location_verified": True,
                "frontend_location_exact": True,
                "frontend_competitors": [
                    {"asin": "B0COMP0001", "title": "Metal dimmer desk lamp with brightness setting"},
                    {"asin": "B0COMP0002", "title": "Large metal dimmer desk lamp slicer"},
                    {"asin": "B0COMP0003", "title": "Metal desk lamp for dimmer"},
                ],
            }
        ],
    )
    filtered = filter_ad_queue_by_decision([scale_row], decisions)

    assert decisions[0]["final_decision"] in {"FRONTEND_FIRST", "CONSERVATIVE_RUN"}
    assert decisions[0]["frontend_evidence_state"] == "insufficient"
    assert decisions[0]["frontend_evidence_tier"] == "强诊断可用"
    assert decisions[0]["frontend_evidence_display_tier"] == "不可用"
    assert decisions[0]["frontend_posture"] == "insufficient_evidence"
    assert "前台证据不足，需刷新后再放量" in decisions[0]["frontend_blocking_reasons"]
    assert "bid_up" in decisions[0]["today_blocked_actions"]
    assert "budget_up" in decisions[0]["today_blocked_actions"]
    assert "broad_scale" in decisions[0]["today_blocked_actions"]
    assert "bid_up" not in decisions[0]["today_allowed_actions"]
    _assert_growth_gate_blocks_small_budget_test(decisions[0])
    assert filtered[0]["scale_action"] == "观察"


def test_product_decision_blocks_bid_up_when_competitors_are_not_comparable() -> None:
    from src.product_decision_layer import build_product_final_decisions, filter_ad_queue_by_decision

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-SCALE",
                    "asin": "B0SCALE123",
                    "product_name": "Metal dimmer desk lamp",
                    "ad_clicks": 42,
                    "ad_spend": 12,
                    "ad_orders": 4,
                    "ad_sales": 80,
                    "total_orders": 6,
                    "ACOS": 0.12,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-SCALE",
        "asin": "B0SCALE123",
        "search_term_or_target": "verified keyword",
        "scale_action": "提高竞价 5%-10%",
        "ad_orders": "3",
        "ad_sales": "£60.00",
        "ACOS": "12%",
        "target_acos": "25%",
    }

    decisions = build_product_final_decisions(
        payload,
        scale_rows=[scale_row],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-SCALE",
                "asin": "B0SCALE123",
                "frontend_check_status": "已自动检查",
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_buy_box": "识别到购买按钮",
                "frontend_coupon": "10%",
                "frontend_competitors": [
                    {"asin": "B0COMP0001", "title": "Metal wall shelf"},
                    {"asin": "B0COMP0002", "title": "Plastic storage bin"},
                    {"asin": "B0COMP0003", "title": "Kitchen towel rack"},
                ],
            }
        ],
    )
    filtered = filter_ad_queue_by_decision([scale_row], decisions)

    assert decisions[0]["final_decision"] == "FRONTEND_FIRST"
    assert decisions[0]["competitor_comparability"] == "low"
    assert decisions[0]["frontend_evidence_tier"] == "强诊断可用"
    assert decisions[0]["frontend_evidence_display_tier"] == "仅背景参考"
    assert "竞品可比性不足" in decisions[0]["frontend_blocking_reasons"]
    assert "bid_up" in decisions[0]["today_blocked_actions"]
    assert "budget_up" in decisions[0]["today_blocked_actions"]
    assert "broad_scale" in decisions[0]["today_blocked_actions"]
    assert "bid_up" not in decisions[0]["today_allowed_actions"]
    _assert_growth_gate_blocks_small_budget_test(decisions[0])
    assert filtered[0]["scale_action"] == "观察"


def test_product_decision_respects_explicit_low_competitor_comparability() -> None:
    from src.product_decision_layer import build_product_final_decisions, filter_ad_queue_by_decision

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-EXPLICIT-COMP",
                    "asin": "B0EXPLCOMP",
                    "product_name": "Metal dimmer desk lamp",
                    "ad_clicks": 42,
                    "ad_spend": 12,
                    "ad_orders": 4,
                    "ad_sales": 80,
                    "total_orders": 6,
                    "ACOS": 0.12,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-EXPLICIT-COMP",
        "asin": "B0EXPLCOMP",
        "search_term_or_target": "verified keyword",
        "scale_action": "提高竞价 5%-10%",
        "ad_orders": "3",
        "ad_sales": "£60.00",
        "ACOS": "12%",
        "target_acos": "25%",
    }

    decisions = build_product_final_decisions(
        payload,
        scale_rows=[scale_row],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-EXPLICIT-COMP",
                "asin": "B0EXPLCOMP",
                "frontend_check_status": "已自动检查",
                "frontend_search_status": "已自动检查",
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_buy_box": "识别到购买按钮",
                "frontend_coupon": "10%",
                "frontend_location_scope": "exact",
                "frontend_location_verified": True,
                "frontend_location_exact": True,
                "competitor_comparability": "low",
                "comparable_competitor_count": 1,
                "competitor_mismatch_reason": "前台搜索页可比竞品少于2个",
                "frontend_competitors": [
                    {"asin": "B0COMP0001", "title": "Metal dimmer desk lamp with brightness setting"},
                    {"asin": "B0COMP0002", "title": "Large metal dimmer desk lamp slicer"},
                    {"asin": "B0COMP0003", "title": "Metal desk lamp for dimmer"},
                ],
            }
        ],
    )
    filtered = filter_ad_queue_by_decision([scale_row], decisions)

    assert decisions[0]["final_decision"] == "FRONTEND_FIRST"
    assert decisions[0]["competitor_comparability"] == "low"
    assert decisions[0]["comparable_competitor_count"] == 1
    assert decisions[0]["competitor_mismatch_reason"] == "前台搜索页可比竞品少于2个"
    assert decisions[0]["frontend_evidence_display_tier"] == "仅背景参考"
    assert "竞品可比性不足" in decisions[0]["frontend_blocking_reasons"]
    assert "bid_up" in decisions[0]["today_blocked_actions"]
    assert "budget_up" in decisions[0]["today_blocked_actions"]
    assert "broad_scale" in decisions[0]["today_blocked_actions"]
    assert "bid_up" not in decisions[0]["today_allowed_actions"]
    _assert_growth_gate_blocks_small_budget_test(decisions[0])
    assert filtered[0]["scale_action"] == "观察"


def test_product_decision_treats_medium_competitor_comparability_as_background_only() -> None:
    from src.product_decision_layer import build_product_final_decisions, filter_ad_queue_by_decision

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-MEDIUM-COMP",
                    "asin": "B0MEDCOMP",
                    "product_name": "Metal dimmer desk lamp",
                    "ad_clicks": 42,
                    "ad_spend": 12,
                    "ad_orders": 4,
                    "ad_sales": 80,
                    "total_orders": 6,
                    "ACOS": 0.12,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-MEDIUM-COMP",
        "asin": "B0MEDCOMP",
        "search_term_or_target": "verified keyword",
        "scale_action": "提高竞价 5%-10%",
        "ad_orders": "3",
        "ad_sales": "£60.00",
        "ACOS": "12%",
        "target_acos": "25%",
    }

    decisions = build_product_final_decisions(
        payload,
        scale_rows=[scale_row],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-MEDIUM-COMP",
                "asin": "B0MEDCOMP",
                "frontend_check_status": "已自动检查",
                "frontend_search_status": "已自动检查",
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_buy_box": "识别到购买按钮",
                "frontend_coupon": "10%",
                "frontend_location_scope": "exact",
                "frontend_location_verified": True,
                "frontend_location_exact": True,
                "competitor_comparability": "medium",
                "comparable_competitor_count": 3,
            }
        ],
    )
    filtered = filter_ad_queue_by_decision([scale_row], decisions)

    assert decisions[0]["final_decision"] == "CONSERVATIVE_RUN"
    assert decisions[0]["competitor_comparability"] == "medium"
    assert decisions[0]["frontend_evidence_display_tier"] == "仅背景参考"
    assert decisions[0]["frontend_decision_evidence_tier"] == "仅背景参考"
    assert "竞品可比性未达强诊断，前台证据仅背景参考" in decisions[0]["frontend_blocking_reasons"]
    assert "bid_up" in decisions[0]["today_blocked_actions"]
    assert "bid_up" not in decisions[0]["today_allowed_actions"]
    _assert_growth_gate_blocks_small_budget_test(decisions[0])
    assert filtered[0]["scale_action"] == "观察"


def test_product_decision_blocks_bid_up_when_competitor_samples_are_missing() -> None:
    from src.product_decision_layer import build_product_final_decisions, filter_ad_queue_by_decision

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-SCALE",
                    "asin": "B0SCALE123",
                    "product_name": "Metal dimmer desk lamp",
                    "ad_clicks": 42,
                    "ad_spend": 12,
                    "ad_orders": 4,
                    "ad_sales": 80,
                    "total_orders": 6,
                    "ACOS": 0.12,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-SCALE",
        "asin": "B0SCALE123",
        "search_term_or_target": "verified keyword",
        "scale_action": "提高竞价 5%-10%",
        "ad_orders": "3",
        "ad_sales": "£60.00",
        "ACOS": "12%",
        "target_acos": "25%",
    }

    decisions = build_product_final_decisions(
        payload,
        scale_rows=[scale_row],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-SCALE",
                "asin": "B0SCALE123",
                "frontend_check_status": "已自动检查",
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_buy_box": "识别到购买按钮",
                "frontend_coupon": "10%",
                "frontend_location_scope": "exact",
                "frontend_location_verified": True,
                "frontend_location_exact": True,
            }
        ],
    )
    filtered = filter_ad_queue_by_decision([scale_row], decisions)

    assert decisions[0]["final_decision"] == "FRONTEND_FIRST"
    assert decisions[0]["competitor_comparability"] == "unknown"
    assert "竞品可比性不足" in decisions[0]["frontend_blocking_reasons"]
    assert "bid_up" in decisions[0]["today_blocked_actions"]
    assert "budget_up" in decisions[0]["today_blocked_actions"]
    assert "broad_scale" in decisions[0]["today_blocked_actions"]
    assert "bid_up" not in decisions[0]["today_allowed_actions"]
    _assert_growth_gate_blocks_small_budget_test(decisions[0])
    assert filtered[0]["scale_action"] == "观察"


def test_product_decision_blocks_bid_up_when_product_name_has_no_comparable_tokens() -> None:
    from src.product_decision_layer import build_product_final_decisions, filter_ad_queue_by_decision

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-SCALE",
                    "asin": "B0SCALE123",
                    "product_name": "90L",
                    "ad_clicks": 42,
                    "ad_spend": 12,
                    "ad_orders": 4,
                    "ad_sales": 80,
                    "total_orders": 6,
                    "ACOS": 0.12,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-SCALE",
        "asin": "B0SCALE123",
        "search_term_or_target": "verified keyword",
        "scale_action": "提高竞价 5%-10%",
        "ad_orders": "3",
        "ad_sales": "£60.00",
        "ACOS": "12%",
        "target_acos": "25%",
    }

    decisions = build_product_final_decisions(
        payload,
        scale_rows=[scale_row],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-SCALE",
                "asin": "B0SCALE123",
                "frontend_check_status": "已自动检查",
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_buy_box": "识别到购买按钮",
                "frontend_coupon": "10%",
                "frontend_location_scope": "exact",
                "frontend_location_verified": True,
                "frontend_location_exact": True,
                "frontend_search_status": "已自动检查",
                "frontend_competitors": [
                    {"asin": "B0COMP0001", "title": "Heavy duty bin liners"},
                    {"asin": "B0COMP0002", "title": "Large cable sacks"},
                    {"asin": "B0COMP0003", "title": "Kitchen waste bags"},
                ],
            }
        ],
    )
    filtered = filter_ad_queue_by_decision([scale_row], decisions)

    assert decisions[0]["final_decision"] == "FRONTEND_FIRST"
    assert decisions[0]["competitor_comparability"] == "unknown"
    assert decisions[0]["comparable_competitor_count"] == 0
    assert "产品名缺少可比关键词" in decisions[0]["competitor_mismatch_reason"]
    assert "竞品可比性不足" in decisions[0]["frontend_blocking_reasons"]
    assert "bid_up" in decisions[0]["today_blocked_actions"]
    assert "bid_up" not in decisions[0]["today_allowed_actions"]
    _assert_growth_gate_blocks_small_budget_test(decisions[0])
    assert filtered[0]["scale_action"] == "观察"


def test_product_decision_blocks_bid_up_when_frontend_location_is_not_exact() -> None:
    from src.product_decision_layer import build_product_final_decisions, filter_ad_queue_by_decision

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-SCALE",
                    "asin": "B0SCALE123",
                    "product_name": "Metal dimmer desk lamp",
                    "ad_clicks": 42,
                    "ad_spend": 12,
                    "ad_orders": 4,
                    "ad_sales": 80,
                    "total_orders": 6,
                    "ACOS": 0.12,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-SCALE",
        "asin": "B0SCALE123",
        "search_term_or_target": "verified keyword",
        "scale_action": "提高竞价 5%-10%",
        "ad_orders": "3",
        "ad_sales": "£60.00",
        "ACOS": "12%",
        "target_acos": "25%",
    }

    decisions = build_product_final_decisions(
        payload,
        scale_rows=[scale_row],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-SCALE",
                "asin": "B0SCALE123",
                "frontend_check_status": "已自动检查",
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_buy_box": "识别到购买按钮",
                "frontend_coupon": "10%",
                "frontend_location_scope": "marketplace",
                "frontend_location_note": "Aberdeen AB10 1",
                "frontend_location_verified": True,
                "frontend_location_exact": False,
                "frontend_competitors": [
                    {"asin": "B0COMP0001", "title": "Metal dimmer desk lamp with brightness setting"},
                    {"asin": "B0COMP0002", "title": "Metal led desk lamp"},
                    {"asin": "B0COMP0003", "title": "Large metal serving board"},
                ],
            }
        ],
    )
    filtered = filter_ad_queue_by_decision([scale_row], decisions)

    assert decisions[0]["frontend_location_uncertain"] is True
    assert decisions[0]["frontend_evidence_state"] == "mixed"
    assert decisions[0]["frontend_evidence_tier"] == "强诊断可用"
    assert decisions[0]["frontend_evidence_display_tier"] == "仅背景参考"
    assert "地区非配置邮编" in "；".join(decisions[0]["frontend_blocking_reasons"])
    assert "bid_up" in decisions[0]["today_blocked_actions"]
    assert "budget_up" in decisions[0]["today_blocked_actions"]
    assert "broad_scale" in decisions[0]["today_blocked_actions"]
    assert "bid_up" not in decisions[0]["today_allowed_actions"]
    _assert_growth_gate_blocks_small_budget_test(decisions[0])
    assert filtered[0]["scale_action"] == "观察"


def test_product_decision_blocks_bid_up_when_exact_location_is_not_verified() -> None:
    from src.product_decision_layer import build_product_final_decisions, filter_ad_queue_by_decision

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-SCALE",
                    "asin": "B0SCALE123",
                    "product_name": "Metal dimmer desk lamp",
                    "ad_clicks": 42,
                    "ad_spend": 12,
                    "ad_orders": 4,
                    "ad_sales": 80,
                    "total_orders": 6,
                    "ACOS": 0.12,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-SCALE",
        "asin": "B0SCALE123",
        "search_term_or_target": "verified keyword",
        "scale_action": "提高竞价 5%-10%",
        "ad_orders": "3",
        "ad_sales": "£60.00",
        "ACOS": "12%",
        "target_acos": "25%",
    }

    decisions = build_product_final_decisions(
        payload,
        scale_rows=[scale_row],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-SCALE",
                "asin": "B0SCALE123",
                "frontend_check_status": "已自动检查",
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_buy_box": "识别到购买按钮",
                "frontend_coupon": "10%",
                "frontend_search_status": "已自动检查",
                "frontend_location_scope": "exact",
                "frontend_location_verified": False,
                "frontend_location_exact": True,
                "frontend_competitors": [
                    {"asin": "B0COMP0001", "title": "Metal dimmer desk lamp with brightness setting"},
                    {"asin": "B0COMP0002", "title": "Metal led desk lamp"},
                    {"asin": "B0COMP0003", "title": "Large metal serving board"},
                ],
            }
        ],
    )
    filtered = filter_ad_queue_by_decision([scale_row], decisions)

    assert decisions[0]["frontend_location_verified"] is False
    assert decisions[0]["frontend_location_uncertain"] is True
    assert "地区待确认" in "；".join(decisions[0]["frontend_blocking_reasons"])
    assert "bid_up" in decisions[0]["today_blocked_actions"]
    assert "budget_up" in decisions[0]["today_blocked_actions"]
    assert "broad_scale" in decisions[0]["today_blocked_actions"]
    assert "bid_up" not in decisions[0]["today_allowed_actions"]
    _assert_growth_gate_blocks_small_budget_test(decisions[0])
    assert filtered[0]["scale_action"] == "观察"


def test_product_decision_blocks_bid_up_when_exact_location_verification_is_missing() -> None:
    from src.product_decision_layer import build_product_final_decisions, filter_ad_queue_by_decision

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-SCALE",
                    "asin": "B0SCALE123",
                    "product_name": "Metal dimmer desk lamp",
                    "ad_clicks": 42,
                    "ad_spend": 12,
                    "ad_orders": 4,
                    "ad_sales": 80,
                    "total_orders": 6,
                    "ACOS": 0.12,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-SCALE",
        "asin": "B0SCALE123",
        "search_term_or_target": "verified keyword",
        "scale_action": "提高竞价 5%-10%",
        "ad_orders": "3",
        "ad_sales": "£60.00",
        "ACOS": "12%",
        "target_acos": "25%",
    }

    decisions = build_product_final_decisions(
        payload,
        scale_rows=[scale_row],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-SCALE",
                "asin": "B0SCALE123",
                "frontend_check_status": "已自动检查",
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_buy_box": "识别到购买按钮",
                "frontend_coupon": "10%",
                "frontend_search_status": "已自动检查",
                "frontend_location_scope": "exact",
                "frontend_location_exact": True,
                "frontend_competitors": [
                    {"asin": "B0COMP0001", "title": "Metal dimmer desk lamp with brightness setting"},
                    {"asin": "B0COMP0002", "title": "Metal led desk lamp"},
                    {"asin": "B0COMP0003", "title": "Large metal serving board"},
                ],
            }
        ],
    )
    filtered = filter_ad_queue_by_decision([scale_row], decisions)

    assert decisions[0]["frontend_location_verified"] is False
    assert decisions[0]["frontend_location_uncertain"] is True
    assert decisions[0]["frontend_evidence_display_tier"] == "仅背景参考"
    assert "地区待确认" in "；".join(decisions[0]["frontend_blocking_reasons"])
    assert "bid_up" in decisions[0]["today_blocked_actions"]
    assert "bid_up" not in decisions[0]["today_allowed_actions"]
    _assert_growth_gate_blocks_small_budget_test(decisions[0])
    assert filtered[0]["scale_action"] == "观察"


def test_product_decision_blocks_bid_up_when_exact_location_has_warning() -> None:
    from src.product_decision_layer import build_product_final_decisions, filter_ad_queue_by_decision

    payload = {
        "target_marketplace": "UK",
        "history_days": 14,
        "summary": {"ads_row_count": 20, "erp_row_count": 20},
        "data_quality": {"validation_messages": []},
        "product_window_metrics": {
            "14d": [
                {
                    "marketplace": "UK",
                    "sku": "SKU-SCALE",
                    "asin": "B0SCALE123",
                    "product_name": "Metal dimmer desk lamp",
                    "ad_clicks": 42,
                    "ad_spend": 12,
                    "ad_orders": 4,
                    "ad_sales": 80,
                    "total_orders": 6,
                    "ACOS": 0.12,
                    "target_acos": 0.25,
                    "profit_before_ads_per_unit": 8,
                }
            ],
            "7d": [],
        },
    }
    scale_row = {
        "marketplace": "UK",
        "sku": "SKU-SCALE",
        "asin": "B0SCALE123",
        "search_term_or_target": "verified keyword",
        "scale_action": "提高竞价 5%-10%",
        "ad_orders": "3",
        "ad_sales": "£60.00",
        "ACOS": "12%",
        "target_acos": "25%",
    }

    decisions = build_product_final_decisions(
        payload,
        scale_rows=[scale_row],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-SCALE",
                "asin": "B0SCALE123",
                "frontend_check_status": "已自动检查",
                "frontend_auto_conclusion": "FRONTEND_OK",
                "frontend_auto_conclusion_label": "未见明显前台劣势",
                "frontend_evidence_quality_score": 88,
                "frontend_evidence_tier": "强诊断可用",
                "frontend_buy_box": "识别到购买按钮",
                "frontend_coupon": "10%",
                "frontend_location_scope": "exact",
                "frontend_location_warning": "UK 地区异常：United States",
                "frontend_location_verified": True,
                "frontend_location_exact": True,
                "frontend_competitors": [
                    {"asin": "B0COMP0001", "title": "Metal dimmer desk lamp with brightness setting"},
                    {"asin": "B0COMP0002", "title": "Metal led desk lamp"},
                    {"asin": "B0COMP0003", "title": "Large metal serving board"},
                ],
            }
        ],
    )
    filtered = filter_ad_queue_by_decision([scale_row], decisions)

    assert decisions[0]["frontend_location_uncertain"] is True
    assert decisions[0]["frontend_evidence_state"] == "mixed"
    assert decisions[0]["frontend_evidence_tier"] == "强诊断可用"
    assert decisions[0]["frontend_evidence_display_tier"] == "仅背景参考"
    assert "UK 地区异常：United States" in "；".join(decisions[0]["frontend_blocking_reasons"])
    assert "bid_up" in decisions[0]["today_blocked_actions"]
    assert "bid_up" not in decisions[0]["today_allowed_actions"]
    _assert_growth_gate_blocks_small_budget_test(decisions[0])
    assert filtered[0]["scale_action"] == "观察"


def test_product_operation_cards_read_uppercase_acos_and_build_ad_summary() -> None:
    from src.report_presentation import _build_product_operation_cards

    cards = _build_product_operation_cards(
        {
            "target_marketplace": "UK",
            "product_window_metrics": {
                "14d": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-1",
                        "asin": "B0TEST1234",
                        "product_name": "Test product",
                        "ad_clicks": 24,
                        "ad_spend": 18,
                        "ad_orders": 0,
                        "total_orders": 2,
                        "natural_orders": 2,
                        "ACOS": 0.45,
                        "TACOS": 0.18,
                        "ad_CVR": 0.0,
                    }
                ],
                "7d": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-1",
                        "asin": "B0TEST1234",
                        "ad_orders": 0,
                        "total_orders": 1,
                        "natural_orders": 1,
                    }
                ],
            },
        },
        decision_rows=[],
        task_rows=[],
        search_rows=[],
        frontend_rows=[],
        inventory_rows=[],
    )

    assert cards[0]["acos"] == 0.45
    assert cards[0]["tacos"] == 0.18
    assert cards[0]["ad_cvr"] == 0.0
    assert cards[0]["natural_orders"] == 2
    assert cards[0]["recent_7d_total_orders"] == 1
    assert cards[0]["recent_7d_natural_orders"] == 1
    assert "广告需继续观察" in cards[0]["ad_diagnostic_summary"]


def test_product_operation_cards_keep_all_pending_ad_action_items() -> None:
    from src.report_presentation import _build_product_operation_cards

    search_rows = []
    for index in range(5):
        search_rows.append(
            {
                "marketplace": "UK",
                "sku": "SKU-OPS-AD",
                "asin": "B0OPSADFULL",
                "product_name": "Operation full ad product",
                "search_term_or_target": f"operation full ad term {index + 1}",
                "suggested_action": "降竞价",
                "action_id": f"action-{index + 1}",
            }
        )

    cards = _build_product_operation_cards(
        {
            "target_marketplace": "UK",
            "product_window_metrics": {
                "14d": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-OPS-AD",
                        "asin": "B0OPSADFULL",
                        "product_name": "Operation full ad product",
                        "ad_clicks": 25,
                        "ad_orders": 0,
                        "total_orders": 1,
                    }
                ],
                "7d": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-OPS-AD",
                        "asin": "B0OPSADFULL",
                        "ad_orders": 0,
                    }
                ],
            },
        },
        decision_rows=[],
        task_rows=[],
        search_rows=search_rows,
        frontend_rows=[],
        inventory_rows=[],
    )

    card = cards[0]
    assert [row["action_id"] for row in card["ad_action_items"]] == [
        "action-1",
        "action-2",
        "action-3",
        "action-4",
        "action-5",
    ]
    assert card["ad_action_count"] == 5
    assert card["ad_action_display_limit"] == 4
    assert card["ad_action_more_count"] == 1
    assert "页面预览前 4 条，另有 1 条见广告工作台" in card["ad_diagnostic_summary"]


def test_product_operation_cards_preserve_frontend_display_tier_for_cached_evidence() -> None:
    from src.report_presentation import _build_product_operation_cards

    cards = _build_product_operation_cards(
        {
            "target_marketplace": "UK",
            "product_window_metrics": {
                "14d": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-CACHED-FRONTEND",
                        "asin": "B0CACHEOPS",
                        "product_name": "Cached frontend product",
                    }
                ],
            },
        },
        decision_rows=[],
        task_rows=[],
        search_rows=[],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-CACHED-FRONTEND",
                "asin": "B0CACHEOPS",
                "frontend_check_status": "沿用 2026-06-20 前台数据",
                "frontend_evidence_tier": "强诊断可用",
                "frontend_evidence_display_tier": "仅背景参考",
                "frontend_evidence_is_strong": "False",
                "frontend_cache_used": True,
                "frontend_search_partial_evidence": "False",
                "frontend_evidence_audit_reasons": ["沿用缓存"],
            }
        ],
        inventory_rows=[],
    )

    assert cards[0]["frontend_evidence_tier"] == "强诊断可用"
    assert cards[0]["frontend_evidence_display_tier"] == "仅背景参考"
    assert cards[0]["frontend_evidence_is_strong"] is False
    assert cards[0]["frontend_search_partial_evidence"] is False
    assert cards[0]["frontend_cache_used"] is True
    assert cards[0]["frontend_evidence_audit_reasons"] == ["沿用缓存"]


def test_product_operation_cards_treat_string_false_cache_flag_as_false() -> None:
    from src.report_presentation import _build_product_operation_cards

    cards = _build_product_operation_cards(
        {
            "target_marketplace": "UK",
            "product_window_metrics": {
                "14d": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-LIVE-FRONTEND",
                        "asin": "B0LIVEOPS1",
                        "product_name": "Live frontend product",
                    }
                ],
            },
        },
        decision_rows=[],
        task_rows=[],
        search_rows=[],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-LIVE-FRONTEND",
                "asin": "B0LIVEOPS1",
                "frontend_check_status": "已自动检查",
                "frontend_evidence_tier": "强诊断可用",
                "frontend_evidence_display_tier": "强诊断可用",
                "frontend_evidence_is_strong": "True",
                "frontend_cache_used": "False",
                "frontend_search_partial_evidence": "False",
            }
        ],
        inventory_rows=[],
    )

    assert cards[0]["frontend_cache_used"] is False
    assert cards[0]["frontend_evidence_is_strong"] is True
    assert cards[0]["frontend_search_partial_evidence"] is False


def test_product_operation_cards_build_richer_main_reason() -> None:
    from src.report_presentation import _build_product_operation_cards

    cards = _build_product_operation_cards(
        {
            "target_marketplace": "UK",
            "product_window_metrics": {
                "14d": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-2",
                        "asin": "B0RICH12345",
                        "product_name": "Rich product",
                        "ad_clicks": 24,
                        "ad_spend": 18,
                        "ad_orders": 0,
                        "total_orders": 2,
                        "natural_orders": 2,
                        "profit_before_ads_per_unit": -1,
                    }
                ],
                "7d": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-2",
                        "asin": "B0RICH12345",
                        "total_orders": 1,
                        "natural_orders": 1,
                    }
                ],
            },
        },
        decision_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-2",
                "asin": "B0RICH12345",
                "final_decision": "CONSERVATIVE_RUN",
                "decision_reason": "前台证据不足，今天只允许广告止损动作。",
                "fusion_action_gate": "fix_both",
                "feedback_cooldown_status": "已执行动作冷却中",
            }
        ],
        task_rows=[],
        search_rows=[],
        frontend_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-2",
                "asin": "B0RICH12345",
                "frontend_check_status": "沿用 2026-06-16 前台数据",
                "frontend_auto_conclusion_label": "自动证据不足，不能用于强诊断",
                "frontend_search_findings": "搜索页前三竞品未稳定解析。",
            }
        ],
        inventory_rows=[
            {
                "marketplace": "UK",
                "sku": "SKU-2",
                "asin": "B0RICH12345",
                "stock_risk_level": "LOW_STOCK",
                "stock_risk_reason": "覆盖天数偏低。",
            }
        ],
    )

    reason = cards[0]["decision_reason"]
    assert "前台证据不足，今天只允许广告止损动作。" in reason
    assert "总单 2" in reason
    assert "自然单 2" in reason
    assert "利润" in reason
    assert "库存" in reason
    assert "前台" in reason
    assert "历史" in reason
    assert cards[0]["fusion_action_gate"] == "fix_both"


def test_task_and_frontend_cards_render_automatic_decisions_without_generic_confirmation() -> None:
    from src.generate_html_report import _build_frontend_lookup, _render_frontend_status_summary, _render_task_cards

    frontend_rows = [
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "product_name": "Test product",
            "frontend_check_status": "已自动检查",
            "frontend_findings": "售价：£18.99；评分：3.8 out of 5 stars",
            "frontend_auto_conclusion_label": "明确前台劣势",
            "frontend_evidence_quality_score": "58",
            "frontend_price_delta_pct": "0.15",
            "frontend_rating_delta": "-0.5",
            "frontend_review_delta_pct": "-0.8",
            "frontend_coupon": "未稳定识别 Coupon",
            "frontend_buy_box": "未稳定识别 Buy Box/购买按钮",
            "trigger_reason": "广告无单",
            "key_metrics": "点击 31；订单 0",
            "suspected_issue": "明确前台劣势",
            "conservative_action": "不加预算，只降无效流量。",
            "recommended_next_step": "刷新前台缓存后复查。",
            "priority": "P0",
        }
    ]
    task_rows = [
        {
            "marketplace": "UK",
            "sku": "SKU-1",
            "asin": "B0TEST1234",
            "product_name": "Test product",
            "priority": "P0",
            "issue_type": "广告消耗无转化诊断",
            "primary_reason": "广告无单",
            "key_evidence": "点击 31；广告订单 0",
            "today_action": "处理广告流量",
            "fusion_issue_type": "前台和广告共同问题",
            "fusion_confidence": "高",
            "fusion_reason": "广告证据：ad_no_order_clicks；前台证据：明确前台劣势",
            "fusion_today_action": "同时收紧广告流量并修正前台弱项。",
            "fusion_do_not_do": "不扩预算；不推大词。",
            "fusion_review_window": "3 天复查广告浪费，7 天复查订单。",
        }
    ]

    html = _render_task_cards(task_rows, "P0", frontend_lookup=_build_frontend_lookup(frontend_rows))
    frontend_html = _render_frontend_status_summary(frontend_rows)

    combined = html + frontend_html
    assert "系统结论" in combined
    assert "融合诊断" in combined
    assert "价格差距" in combined
    assert "先别做" in combined
    assert 'class="action-brief-grid"' in html
    assert '<summary>证据明细</summary>' in html
    assert html.index("今天动作") < html.index("证据明细")
    details_html = html[html.index('<details class="evidence-details">') :]
    assert "完整诊断原因" in details_html
    assert "前台证据" in details_html
    assert "需要处理的搜索词 / ASIN" in details_html
    assert "需要确认的问题" not in combined
    assert "疑似前台竞争力需要确认" not in combined
    assert "<h2>前台数据更新</h2>" not in combined
    assert 'id="frontend-evidence-status"' in frontend_html
    assert 'section-card is-collapsed" id="frontend-evidence-status"' in frontend_html
    assert 'data-collapse-toggle>展开</button>' in frontend_html
    assert '<div class="collapsible-body">' in frontend_html
    assert "市场调查" in frontend_html
    assert "外部证据" in frontend_html
    assert "刷新调查队列" in frontend_html
    assert "运行状态" in frontend_html
    assert "市场信号" not in frontend_html
    assert "刷新信号队列" not in frontend_html
    assert "信号采集" not in frontend_html
    assert "市场验证中心" not in frontend_html
    assert "刷新市场验证队列" not in frontend_html
    assert "前台证据状态" not in frontend_html
    assert "刷新当前前台队列" not in frontend_html
    assert "缓存可参考，强操作只看完整证据" in frontend_html
    assert "更新前台缓存" not in frontend_html
    assert "待运行，点击后一次完成商品页和卖家精灵调查" in frontend_html


def test_frontend_html_downgrades_unsafe_strong_tier_to_background_reference() -> None:
    from src.generate_html_report import _build_frontend_lookup, _render_frontend_check_cards, _render_task_cards

    task_rows = [
        {
            "marketplace": "UK",
            "sku": "SKU-CACHED",
            "asin": "B0CACHED01",
            "product_name": "Cached frontend product",
            "priority": "P0",
            "issue_type": "广告消耗无转化诊断",
            "primary_reason": "近7天广告无单",
            "key_evidence": "点击 31；广告订单 0",
            "today_action": "先查前台",
        }
    ]
    frontend_rows = [
        {
            "marketplace": "UK",
            "sku": "SKU-CACHED",
            "asin": "B0CACHED01",
            "product_name": "Cached frontend product",
            "priority": "P0",
            "frontend_check_status": "沿用上次前台数据",
            "frontend_cache_used": True,
            "frontend_findings": "沿用上次前台数据（2026-06-20T08:00:00）；售价：£18.99",
            "frontend_auto_conclusion_label": "自动证据不足，不能用于强诊断",
            "frontend_evidence_quality_score": "82",
            "frontend_evidence_tier": "强诊断可用",
            "frontend_evidence_display_tier": "强诊断可用",
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "competitor_comparability": "high",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "trigger_reason": "广告无单",
            "key_metrics": "点击 31；订单 0",
            "conservative_action": "不加预算，只降无效流量。",
            "recommended_next_step": "刷新前台后再判断。",
        }
    ]

    task_html = _render_task_cards(task_rows, "P0", frontend_lookup=_build_frontend_lookup(frontend_rows))
    frontend_html = _render_frontend_check_cards(frontend_rows)
    combined = task_html + frontend_html

    assert "口径：强诊断可用" not in combined
    assert "证据口径：强诊断可用" not in combined
    assert "口径：仅背景参考" in combined
    assert "证据口径：仅背景参考" in combined


def test_frontend_html_downgrades_medium_competitor_comparability_to_background_reference() -> None:
    from src.generate_html_report import _build_frontend_lookup, _render_frontend_check_cards, _render_task_cards

    task_rows = [
        {
            "marketplace": "US",
            "sku": "SKU-MEDIUM-COMP",
            "asin": "B0MEDCOMP1",
            "product_name": "Medium competitor product",
            "priority": "P1",
            "issue_type": "前台竞争力复核",
            "primary_reason": "前台显示可比竞品证据不足",
            "key_evidence": "点击 22；广告订单 0",
            "today_action": "先查前台",
        }
    ]
    frontend_rows = [
        {
            "marketplace": "US",
            "sku": "SKU-MEDIUM-COMP",
            "asin": "B0MEDCOMP1",
            "product_name": "Medium competitor product",
            "priority": "P1",
            "frontend_check_status": "已自动检查",
            "frontend_findings": "售价：$18.99；评分：4.3 out of 5 stars",
            "frontend_auto_conclusion_label": "前台证据通过质量门",
            "frontend_evidence_quality_score": "82",
            "frontend_evidence_tier": "强诊断可用",
            "frontend_evidence_display_tier": "强诊断可用",
            "frontend_evidence_is_strong": True,
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 3,
            "comparable_competitor_count": 3,
            "competitor_comparability": "medium",
            "frontend_location_scope": "exact",
            "frontend_location_verified": True,
            "frontend_location_exact": True,
            "trigger_reason": "前台竞争力复核",
            "key_metrics": "点击 22；订单 0",
            "conservative_action": "竞品可比性未达强诊断前不加预算。",
            "recommended_next_step": "人工核对竞品是否同类同价位。",
        }
    ]

    task_html = _render_task_cards(task_rows, "P1", frontend_lookup=_build_frontend_lookup(frontend_rows))
    frontend_html = _render_frontend_check_cards(frontend_rows)
    combined = task_html + frontend_html

    assert "口径：强诊断可用" not in combined
    assert "证据口径：强诊断可用" not in combined
    assert "口径：仅背景参考" in combined
    assert "证据口径：仅背景参考" in combined


def test_frontend_gap_labels_explain_missing_competitor_samples() -> None:
    from src.generate_html_report import _build_frontend_lookup, _render_task_cards

    task_rows = [
        {
            "marketplace": "US",
            "sku": "SKU-P0",
            "asin": "B0P0TEST01",
            "product_name": "P0 product",
            "priority": "P0",
            "issue_type": "广告消耗无转化诊断",
            "primary_reason": "近7天广告无单",
            "key_evidence": "点击 12；订单 0",
            "today_action": "先查前台",
        }
    ]
    frontend_rows = [
        {
            "marketplace": "US",
            "sku": "SKU-P0",
            "asin": "B0P0TEST01",
            "frontend_check_status": "已自动检查",
            "frontend_findings": "售价：$8.99；评分：3.9 out of 5 stars；评论数：(52)",
            "frontend_price": "$8.99",
            "frontend_rating": "3.9 out of 5 stars",
            "frontend_reviews": "(52)",
            "frontend_search_status": "按广告信号跳过",
            "frontend_search_findings": "当前不是广告点击/转化异常触发，跳过搜索页前三竞品以降低浏览器和报告开销。",
            "frontend_competitor_count": 0,
            "frontend_competitors": [],
            "frontend_coupon": "待确认",
            "frontend_buy_box": "识别到购买按钮",
        }
    ]

    html = _render_task_cards(task_rows, "P0", frontend_lookup=_build_frontend_lookup(frontend_rows))

    assert "价格差距：搜索页未读取" in html
    assert "评分差距：搜索页未读取" in html
    assert "评论差距：搜索页未读取" in html
    assert "差距计算需要至少 2 个可比竞品样本。" in html


def test_frontend_gap_labels_competitor_sample_shortage_after_search_read() -> None:
    from src.generate_html_report import _render_frontend_gap_block

    html = _render_frontend_gap_block(
        {
            "frontend_price": "$8.99",
            "frontend_rating": "3.9 out of 5 stars",
            "frontend_reviews": "(52)",
            "frontend_search_status": "已自动检查",
            "frontend_competitor_count": 1,
            "frontend_competitors": [{"asin": "B0COMP1111"}],
            "frontend_coupon": "待确认",
            "frontend_buy_box": "识别到购买按钮",
        }
    )

    assert "价格差距：竞品样本不足" in html
    assert "评分差距：竞品样本不足" in html
    assert "评论差距：竞品样本不足" in html


def test_product_operation_cards_render_combined_ads_frontend_and_cost() -> None:
    from src.generate_html_report import _render_product_operation_cards

    html = _render_product_operation_cards(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "product_name": "Test product",
                "final_decision": "CONSERVATIVE_RUN",
                "final_decision_label": "保守跑",
                "decision_reason": "前台证据不足，今天只允许广告止损动作。",
                "ad_clicks": 24,
                "ad_spend": 18,
                "ad_orders": 0,
                "total_orders": 2,
                "natural_orders": 2,
                "ACOS": 0.45,
                "target_acos": 0.2,
                "tacos": 0.18,
                "ad_cvr": 0.04,
                "profit_before_ads_per_unit": 8,
                "ad_diagnostic_summary": "广告端优先，近14天点击 24、广告单 0、近7天广告单 0；卡内 1 条词级/ASIN 止损项按广告证据执行。",
                "frontend_status": "沿用 2026-06-10 前台数据",
                "frontend_auto_conclusion_label": "自动证据不足，不能用于强诊断",
                "frontend_evidence_quality_score": 45,
                "frontend_search_status": "已读取部分结果",
                "frontend_search_partial_evidence": True,
                "frontend_search_findings": "核心词搜索页未稳定解析前三竞品。",
                "seller_sprite_check_status": "已抓取",
                "competitor_frontend_status": "部分",
                "competitor_sellersprite_status": "竞品卖家精灵证据不足",
                "competitor_keyword_pressure": "高",
                "product_level_conclusion": "暂停扩张",
                "product_ad_boundary": "前台弱且竞品词压力高，停止新增词，优先处理高花费 0 单词。",
                "own_missing_competitor_keywords": "wrong keyword",
                "cost_status": "广告前利润<=0",
                "cost_key_evidence": "近14天总单 2；广告订单 0；点击 24",
                "inventory_constraint": "LOW_STOCK",
                "today_allowed_actions": ["observe", "bid_down", "negative_exact", "pause"],
                "today_blocked_actions": ["bid_up", "budget_up", "broad_scale"],
                "fusion_review_window": "3 天复查广告浪费，7 天复查订单。",
                "ad_action_items": [
                    {
                        "search_term_or_target": "wrong keyword",
                        "suggested_action": "否定精准",
                        "clicks": "4",
                        "spend": "£2.00",
                        "orders": "0",
                    }
                ],
            }
        ],
        {
            "frontend_queue_total": 1,
            "frontend_usable_evidence_count": 1,
            "frontend_decision_ready_count": 0,
            "frontend_reference_evidence_count": 1,
            "frontend_live_success_count": 0,
            "frontend_cached_count": 1,
            "frontend_search_success_count": 0,
            "frontend_product_page_success_count": 0,
            "frontend_competitor_search_success_count": 0,
            "frontend_own_sellersprite_count": 1,
            "frontend_competitor_sellersprite_count": 0,
            "frontend_competitor_sellersprite_asin_count": 0,
            "frontend_scalable_strong_count": 0,
            "frontend_weak_defensive_count": 0,
            "frontend_insufficient_count": 1,
            "frontend_usable_evidence_rate": 1,
            "frontend_decision_ready_rate": 0,
            "frontend_reference_evidence_rate": 1,
            "frontend_live_success_rate": 0,
            "frontend_search_success_rate": 0,
            "frontend_coverage_label": "1/1 可用，100%",
            "frontend_product_page_success_label": "0/1",
            "frontend_competitor_search_success_label": "0/1",
            "frontend_own_sellersprite_label": "1/1",
            "frontend_competitor_sellersprite_label": "0/1，0 ASIN",
            "frontend_scalable_strong_label": "0/1",
            "frontend_weak_defensive_label": "0/1",
            "frontend_insufficient_label": "1/1",
        },
    )

    assert "产品级结论" in html
    assert "按产品决策" in html
    assert "卖家精灵融合：结论 暂停扩张" in html
    assert "竞品词压力 高" in html
    assert "前台弱且竞品词压力高" in html
    assert "前台证据不足，今天只允许广告止损动作。" in html
    assert "卡内广告止损项" in html
    assert "广告端优先，近14天点击 24" in html
    assert "ACOS" in html
    assert "45.0%" in html
    assert "自然单" in html
    assert "禁止加预算" not in html
    assert "加预算" in html
    assert "否定精准" in html
    assert "成本 / 库存" in html
    assert "搜索页：已读取部分结果" in html
    assert "核心词搜索页未稳定解析前三竞品。" in html
    assert "产品页成功" in html
    assert "Amazon 搜索页辅助验证" in html
    assert "卖家精灵自己 ASIN" in html
    assert "竞品 ASIN 反查" in html
    assert "证据不足" in html
    assert "强前台证据" not in html
    assert "可用前台证据" not in html


def test_product_operation_cards_show_missing_sellersprite_queue_status() -> None:
    from src.generate_html_report import _render_product_operation_cards

    html = _render_product_operation_cards(
        [
            {
                "marketplace": "US",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "product_name": "No seller evidence product",
                "final_decision": "CONSERVATIVE_RUN",
                "final_decision_label": "保守跑",
                "decision_reason": "广告样本不足，先保守观察。",
                "today_allowed_actions": ["observe"],
                "today_blocked_actions": ["bid_up", "budget_up"],
            }
        ]
    )

    assert "产品页 待补" in html
    assert "卖家精灵自己 未入本轮队列" in html
    assert "竞品池 未入本轮队列" in html
    assert "竞品反查 未入本轮队列" in html
    assert "operation-seller-summary" in html


def test_product_operation_cards_prioritize_strong_evidence_over_boundary_rows() -> None:
    from src.generate_html_report import _render_product_operation_cards

    html = _render_product_operation_cards(
        [
            {
                "marketplace": "US",
                "sku": "SKU-LOW",
                "asin": "B0LOWSTOCK1",
                "product_name": "Low stock product",
                "final_decision": "CONSERVATIVE_RUN",
                "final_decision_label": "保守跑",
                "decision_reason": "低库存，广告不要大幅放量。",
                "inventory_constraint": "LOW_STOCK",
                "inventory_reason": "库存覆盖低于提前期。",
                "fusion_review_window": "补齐 7 天窗口或刷新前台后再判断。",
                "ad_clicks": 7,
                "ad_spend": 2.03,
                "ad_orders": 2,
                "total_orders": 3,
                "today_allowed_actions": ["observe", "bid_down"],
                "today_blocked_actions": ["bid_up", "budget_up"],
            },
            {
                "marketplace": "US",
                "sku": "SKU-EVIDENCE",
                "asin": "B0EVIDENCE1",
                "product_name": "Evidence product",
                "final_decision": "CONSERVATIVE_RUN",
                "final_decision_label": "保守跑",
                "decision_reason": "前台和卖家精灵显示竞品压力高。",
                "frontend_status": "已自动检查",
                "seller_sprite_check_status": "已抓取",
                "competitor_pool_status": "有效 3/3",
                "competitor_sellersprite_status": "已抓 3 个",
                "amazon_search_validation_status": "已验证",
                "competitor_keyword_pressure": "高",
                "fusion_review_window": "补齐 7 天窗口或刷新前台后再判断。",
                "today_allowed_actions": ["observe", "bid_down"],
                "today_blocked_actions": ["bid_up", "budget_up"],
            },
        ]
    )

    assert "强证据决策" in html
    assert "运营边界" in html
    assert html.index("US｜Evidence product") < html.index("US｜Low stock product")
    assert "只用于控费、补货或保守跑" in html
    assert "7 天看是否破零" in html
    assert "补货或库存风险解除后" in html
    assert "先不做强操作" not in html


def test_product_operation_cards_show_next_step_after_frontend_checked() -> None:
    from src.generate_html_report import _render_product_final_decision_cards, _render_product_operation_cards

    row = {
        "marketplace": "US",
        "sku": "SKU-1",
        "asin": "B0TEST1234",
        "product_name": "Test product",
        "final_decision": "FRONTEND_FIRST",
        "final_decision_label": "先查前台",
        "decision_reason": "先确认Coupon 到手价待确认，不直接加预算。",
        "frontend_status": "已自动检查",
        "frontend_check_status": "已自动检查",
        "frontend_auto_conclusion_label": "明确前台劣势",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_display_tier": "仅背景参考",
        "frontend_decision_evidence_tier": "仅背景参考",
        "frontend_evidence_is_strong": False,
        "frontend_evidence_quality_score": 68,
        "frontend_evidence_audit_detail": "搜索页已读；竞品 3 个；价差 -26%；评分差 -1.5",
        "frontend_search_status": "已读取部分结果",
        "frontend_search_partial_evidence": True,
        "frontend_blocking_reasons": ["搜索页仅部分读取，竞品证据仅背景参考"],
        "fusion_action_gate": "fix_both",
        "today_allowed_actions": ["observe", "bid_down"],
        "today_blocked_actions": ["bid_up", "budget_up"],
    }
    html = _render_product_operation_cards([row])

    decision_cell = html
    assert 'class="operation-wide-card"' in decision_cell
    assert 'data-product-decision-marketplace="US"' in decision_cell
    assert 'data-product-decision-sku="SKU-1"' in decision_cell
    assert 'data-product-decision-asin="B0TEST1234"' in decision_cell
    assert 'data-product-frontend-decision-tier="仅背景参考"' in decision_cell
    assert 'data-product-frontend-strong="false"' in decision_cell
    assert "止损+修前台" in decision_cell
    assert ">先查前台<" not in decision_cell
    assert "前台已查" in decision_cell
    assert "搜索页已读；竞品 3 个；价差 -26%；评分差 -1.5" in decision_cell
    assert 'title="前台已查；广告先止损，同时处理前台弱项"' in decision_cell

    final_html = _render_product_final_decision_cards([row])
    assert 'data-product-decision-marketplace="US"' in final_html
    assert 'data-product-decision-sku="SKU-1"' in final_html
    assert 'data-product-decision-asin="B0TEST1234"' in final_html
    assert 'data-product-frontend-decision-tier="仅背景参考"' in final_html
    assert 'data-product-frontend-strong="false"' in final_html
    assert "止损+修前台" in final_html
    assert ">先查前台<" not in final_html
    assert "前台门禁" in final_html
    assert "仅背景参考" in final_html
    assert ">强诊断可用<" not in final_html
    assert "搜索页已读；竞品 3 个；价差 -26%；评分差 -1.5" in final_html
    assert "搜索页：已读取部分结果" in final_html
    assert "阻断原因" in final_html
    assert "搜索页仅部分读取，竞品证据仅背景参考" in final_html
    assert "拦截加竞价" in final_html
    assert "拦截加预算" in final_html


def test_product_operation_cards_treat_string_false_partial_search_as_false() -> None:
    from src.generate_html_report import _render_product_final_decision_cards, _render_product_operation_cards

    row = {
        "marketplace": "US",
        "sku": "SKU-1",
        "asin": "B0TEST1234",
        "product_name": "Test product",
        "final_decision": "CONSERVATIVE_RUN",
        "final_decision_label": "保守跑",
        "decision_reason": "前台证据已检查，广告继续保守观察。",
        "frontend_status": "已自动检查",
        "frontend_check_status": "已自动检查",
        "frontend_auto_conclusion_label": "未见明显前台劣势",
        "frontend_evidence_tier": "强诊断可用",
        "frontend_evidence_display_tier": "强诊断可用",
        "frontend_decision_evidence_tier": "强诊断可用",
        "frontend_evidence_is_strong": True,
        "frontend_evidence_quality_score": 88,
        "frontend_search_status": "已自动检查",
        "frontend_search_partial_evidence": "False",
        "today_allowed_actions": ["observe"],
        "today_blocked_actions": [],
    }

    operation_html = _render_product_operation_cards([row])
    final_html = _render_product_final_decision_cards([row])
    combined = operation_html + final_html

    assert "搜索页：已自动检查" in combined
    assert "搜索页：部分读取" not in combined
    assert "搜索页仅部分读取" not in combined


def test_report_js_syncs_submission_status_and_redirects_stale_safe_run_page() -> None:
    from src.generate_html_report import REPORT_JS

    assert "syncLocalSubmissionStatus({ redirectOnDone: true" in REPORT_JS
    assert "fetch('http://127.0.0.1:8765/submission/status?t=' + Date.now(), { cache: 'no-store' })" in REPORT_JS
    assert "function startDailyUpdate(options)" in REPORT_JS
    assert "fetch('http://127.0.0.1:8765/run/daily-update', { method: 'POST' })" in REPORT_JS
    assert "'/run/report-refresh': true" in REPORT_JS
    assert "上传通过，正在自动启动 daily update" in REPORT_JS
    assert "reloadLatestReport(payload)" in REPORT_JS
    assert "'v=updated-'" in REPORT_JS
    assert "fetch('http://127.0.0.1:8765/health?t=' + Date.now(), { cache: 'no-store' })" in REPORT_JS
    assert "window.location.href = latestReportUrl(payload)" in REPORT_JS
    assert "payload.report_links.latest_recommendations" in REPORT_JS
    assert "function localServiceErrorMessage(error)" in REPORT_JS
    assert "error.payload.message" in REPORT_JS
    assert "function installLocalSubmitStatusStabilizer()" in REPORT_JS
    assert "data-local-status-primary" in REPORT_JS
    assert "data-local-status-frontend" not in REPORT_JS
    assert "已运行 ' + payload.elapsed_seconds + ' 秒" in REPORT_JS
    assert "elapsed + '秒'" in REPORT_JS
    assert ".replace(/；?\\s*已运行\\s*\\d+\\s*秒/g, '')" not in REPORT_JS
    assert "function mainWorkflowButtonLabel(payload)" in REPORT_JS
    assert "function updateMainWorkflowButtons(payload)" in REPORT_JS
    assert "updateMainWorkflowButtons(payload);" in REPORT_JS
    assert "本机服务未开启：请先双击 start_report_action_server.command，再点此按钮。" not in REPORT_JS


def test_report_js_disables_local_action_buttons_on_file_reports() -> None:
    from src.generate_html_report import REPORT_JS

    assert "function disableFileReportActions()" in REPORT_JS
    assert "window.location.protocol !== 'file:'" in REPORT_JS
    assert "document.querySelectorAll('[data-run-report-action]')" in REPORT_JS
    assert "button.disabled = true" in REPORT_JS
    assert "静态 file:// 报告只读" in REPORT_JS
    assert "http://127.0.0.1:8765/report/latest_recommendations.html" in REPORT_JS


def test_local_submit_panel_uses_compact_collapsed_status_style() -> None:
    from src.generate_html_report import REPORT_UI_CSS

    assert ".local-submit-panel.is-collapsed [data-config-submit-status]" in REPORT_UI_CSS
    assert ".local-submit-panel.is-collapsed .local-submit-links" in REPORT_UI_CSS
    assert ".local-submit-panel.is-collapsed [data-workflow-progress-status].status-ok" in REPORT_UI_CSS
    assert "display: none;" in REPORT_UI_CSS
    assert "max-height: 40px;" in REPORT_UI_CSS
    assert "white-space: nowrap;" in REPORT_UI_CSS
    assert "text-overflow: ellipsis;" in REPORT_UI_CSS
    assert "grid-template-columns: minmax(0, 1fr) minmax(260px, 0.32fr);" not in REPORT_UI_CSS
    assert ".local-submit-panel:not(.is-collapsed)" in REPORT_UI_CSS
    assert "grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));" in REPORT_UI_CSS
    assert ".section-card > *" in REPORT_UI_CSS
    assert "min-width: 0;" in REPORT_UI_CSS


def test_report_js_treats_urllib_frontend_block_as_soft_failure() -> None:
    from src.generate_html_report import REPORT_JS

    assert "function isSoftFrontendRetryFailure(payload)" in REPORT_JS
    assert "payload.soft_failure" in REPORT_JS
    assert "payload.failure_mode === 'urllib_frontend_blocked'" in REPORT_JS
    assert "payload.failure_mode === 'chrome_cdp_frontend_check_partial'" in REPORT_JS
    assert "payload.failure_mode === 'chrome_cdp_frontend_check_passed_with_pending'" in REPORT_JS
    assert "&& !isSoftFrontendRetryFailure(payload)" in REPORT_JS
    assert "payload.returncode && payload.returncode !== 0 && !isSoftFrontendRetryFailure(payload)" in REPORT_JS


def test_report_js_routes_frontend_retry_status_to_frontend_button() -> None:
    from src.generate_html_report import REPORT_JS

    assert "function isFrontendRetryStatus(payload)" in REPORT_JS
    assert "payload.status_scope === 'frontend_retry'" in REPORT_JS
    assert "调查队列 20次验收" in REPORT_JS
    assert "function moveFrontendRetryStatus(payload)" in REPORT_JS
    assert 'document.querySelector(\'[data-run-report-status="frontend-retry"]\')' in REPORT_JS
    assert "if (!statusBox && !retryStatus) return;" in REPORT_JS
    assert "调查结果见市场调查" in REPORT_JS
    assert "statusBox.innerHTML = '日报文件：当前没有新的上传或 daily update 状态；调查结果见市场调查。'" in REPORT_JS
    assert "信号采集结果见市场信号" not in REPORT_JS
    assert "市场验证结果见市场验证中心" not in REPORT_JS
    assert "前台重试结果见前台证据状态" not in REPORT_JS


def test_report_js_displays_p0_frontend_async_status_separately() -> None:
    from src.generate_html_report import REPORT_JS

    assert "function frontendAsyncStatusHtml(payload)" in REPORT_JS
    assert "payload.frontend_async_status" in REPORT_JS
    assert "P0 前台后台检查运行中：" in REPORT_JS
    assert "P0 前台后台检查：" in REPORT_JS
    assert "formatDailyPayload(payload) + reportLinksHtml(payload) + frontendAsyncStatusHtml(payload)" in REPORT_JS
    assert "message + reportLinksHtml(payload) + frontendAsyncStatusHtml(payload)" in REPORT_JS


def test_report_js_supports_single_product_frontend_check_with_reload() -> None:
    from src.generate_html_report import REPORT_JS

    assert "function reportActionUrl(runButton, action)" in REPORT_JS
    assert "runButton.dataset.runReportQuery || ''" in REPORT_JS
    assert "data-run-report-reload-on-done" not in REPORT_JS
    assert "var reloadOnDone = runButton.dataset.runReportReloadOnDone === 'true';" in REPORT_JS
    assert "fetch(reportActionUrl(runButton, action), { method: 'POST' })" in REPORT_JS
    assert "if (!isError && reloadOnDone) {" in REPORT_JS
    assert "reloadLatestReport(payload);" in REPORT_JS


def test_frontend_status_summary_exposes_single_product_status_slot() -> None:
    from src.generate_html_report import _render_frontend_status_summary

    html = _render_frontend_status_summary(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-1",
                "asin": "B0TEST1234",
                "product_name": "Test product",
                "frontend_check_status": "待前台检查",
            }
        ]
    )

    assert 'data-run-report-status="frontend-check-one"' in html
    assert "单产品按钮只检查当前 ASIN" in html


def test_frontend_status_summary_exposes_public_demo_reverse_lookup_button_only_for_demo_rows() -> None:
    from src.generate_html_report import _render_frontend_status_summary

    demo_html = _render_frontend_status_summary(
        [{"marketplace": "US", "sku": "SKU-DEMO-US-001", "asin": "B0DEMOUS01"}]
    )
    private_html = _render_frontend_status_summary(
        [{"marketplace": "US", "sku": "SKU-REAL-001", "asin": "B0TEST1234"}]
    )

    assert "公开 ASIN 测试" in demo_html
    assert "B084Z8CXXN" in demo_html
    assert 'data-run-report-action="battle-diagnosis-one"' in demo_html
    assert "公开 ASIN 测试" not in private_html
    assert "B084Z8CXXN" not in private_html


def test_frontend_check_queue_synthesizes_ad_hoc_asin_test_row() -> None:
    from scripts.frontend_check_queue import fallback_rows_from_payload

    rows = fallback_rows_from_payload(
        {"marketplace_results": [{"report_view_snapshot": {"frontend_check_queue_rows": []}}]},
        {"marketplace": "US", "sku": "PUBLIC-LIVE-ASIN-SMOKE", "asin": "B084Z8CXXN"},
    )

    assert len(rows) == 1
    row = rows[0]
    assert row["marketplace"] == "US"
    assert row["sku"] == "PUBLIC-LIVE-ASIN-SMOKE"
    assert row["asin"] == "B084Z8CXXN"
    assert row["product_url"] == "https://www.amazon.com/dp/B084Z8CXXN"
    assert row["frontend_search_url"] == "https://www.amazon.com/s?k=desk+lamp"
    assert row["source_role"] == "ad_hoc_public_test"


def test_frontend_check_queue_does_not_synthesize_without_marketplace_or_asin() -> None:
    from scripts.frontend_check_queue import fallback_rows_from_payload

    assert fallback_rows_from_payload({}, {"marketplace": "US", "sku": "", "asin": ""}) == []
    assert fallback_rows_from_payload({}, {"marketplace": "", "sku": "", "asin": "B084Z8CXXN"}) == []


def test_frontend_status_summary_counts_read_failures_as_pending_not_checked() -> None:
    from src.generate_html_report import _frontend_queue_counts, _render_frontend_status_summary

    rows = [
        {
            "marketplace": "UK",
            "sku": "SKU-FAIL",
            "asin": "B0FAIL0001",
            "product_name": "Frontend failed product",
            "frontend_check_status": "读取失败",
            "frontend_findings": "自动读取失败",
        },
        {
            "marketplace": "UK",
            "sku": "SKU-OK",
            "asin": "B0OK000001",
            "product_name": "Frontend ok product",
            "frontend_check_status": "已自动检查",
        },
    ]

    counts = _frontend_queue_counts(rows)
    html = _render_frontend_status_summary(rows)

    assert counts["待前台检查"] == 1
    assert counts["读取失败"] == 1
    assert counts["已自动检查"] == 1
    assert "运行状态" in html
    assert "前台：已读 1，待检 1，缓存 0，失败 1。" not in html


def test_frontend_check_cards_render_single_product_frontend_button() -> None:
    from src.generate_html_report import _render_frontend_check_cards

    html = _render_frontend_check_cards(
        [
            {
                "priority": "P1",
                "marketplace": "US",
                "sku": "SKU-DEMO-BAG-01",
                "asin": "B0DEMOBAG1",
                "product_name": "演示笔记本 12片",
                "frontend_check_status": "待前台检查",
                "trigger_reason": "近7天广告无单",
                "key_metrics": "近7天点击 11；订单 0；总单 4",
            }
        ]
    )

    assert "检查这个产品前台" in html
    assert 'data-run-report-action="frontend-check-one"' in html
    assert 'data-run-report-reload-on-done="true"' in html
    assert "marketplace=US" in html
    assert "sku=SKU-DEMO-BAG-01" in html
    assert "asin=B0DEMOBAG1" in html


def test_frontend_check_cards_show_executed_action_detail() -> None:
    from src.generate_html_report import _render_frontend_check_cards

    html = _render_frontend_check_cards(
        [
            {
                "priority": "P0",
                "marketplace": "US",
                "sku": "SKU-DEMO-BAG-BULK",
                "asin": "B0DEMOBAG2",
                "product_name": "Demo cable ties 100 count",
                "frontend_check_status": "已自动检查",
                "trigger_reason": "近14天总单 0",
                "key_metrics": "近14天总单 0；广告订单 0；点击 0",
                "confirmed_status": "已执行",
                "search_term_or_target": "20 gallon cable holder",
                "suggested_action": "降竞价10%-20%",
                "action_scope": "search_term",
                "confirmed_note": "用户反馈：2026-05-18 今天广告已调整完成",
            }
        ]
    )

    assert "广告：20 gallon cable holder 降竞价10%-20%" in html
    assert ">已执行<" not in html
    assert 'title="广告：20 gallon cable holder 降竞价10%-20%"' in html


def test_frontend_check_cards_fuse_sellersprite_into_auto_conclusion() -> None:
    from src.generate_html_report import _render_frontend_check_cards

    html = _render_frontend_check_cards(
        [
            {
                "priority": "P0",
                "marketplace": "US",
                "sku": "SKU-DEMO-BAG-100",
                "asin": "B0DEMOBAG3",
                "product_name": "Demo cable ties 100 count",
                "frontend_check_status": "已自动检查",
                "frontend_auto_conclusion_label": "明确前台劣势",
                "frontend_findings": "售价 $8.59；评分 3.0",
                "seller_sprite_check_status": "已抓取",
                "seller_sprite_keyword_count": 4,
                "product_level_conclusion": "暂停扩张",
                "product_ad_boundary": "前台弱且竞品词压力高，停止新增词。",
                "competitor_keyword_pressure": "高",
                "own_missing_competitor_keywords": "cable ties",
                "trigger_reason": "近14天总单 0",
                "key_metrics": "近14天总单 0；广告订单 0；点击 0",
            }
        ]
    )

    assert "自动前台结论" in html
    assert "卖家精灵融合" in html
    assert "结论 暂停扩张" in html
    assert "竞品词压力 高" in html
    assert "前台弱且竞品词压力高" in html


def test_recommendations_workbench_exposes_inventory_replenishment_anchor(tmp_path) -> None:
    from src.generate_html_report import write_recommendations_workbench_html

    output_path = tmp_path / "latest_recommendations.html"
    results = [
        {
            "has_data": True,
            "marketplace": "UK",
            "report_view": {
                "today_task_queue_rows": [],
                "tomorrow_review_rows": [],
                "html_search_term_processing_queue_rows": [],
                "listing_price_diagnosis_rows": [],
                "yesterday_attribution_rows": [],
                "action_effect_review_rows": [],
                "keyword_action_effect_review_rows": [],
                "frontend_check_queue_rows": [],
                "inventory_replenishment_rows": [
                    {
                        "marketplace": "UK",
                        "product_name": "演示台灯",
                        "stock_risk_level": "LOW_STOCK",
                        "stock_status_label": "低库存",
                        "current_inventory": 4,
                        "days_of_cover": 28,
                        "total_lead_time_days": 100,
                        "target_cover_days": 130,
                        "replenishment_advice": "建议补货约 15 件。",
                    }
                ],
                "product_final_decision_rows": [],
                "product_operation_cards": [],
                "frontend_coverage_summary": {},
                "analysis_status": "正式分析",
                "issue_summary": "ok",
            },
        }
    ]

    write_recommendations_workbench_html(results, output_path, "2026-06-08")

    html = output_path.read_text(encoding="utf-8")
    assert 'href="#inventory-replenishment">补货提醒</a>' in html
    assert 'id="inventory-replenishment"' in html
    assert "建议补货约 15 件。" in html
    assert "LOCAL WORKFLOW" in html
    assert "今日数据入口" not in html
    assert "上传并刷新日报" in html
    assert "手动运行 daily update" not in html
    assert "data-run-daily-update" not in html
    assert "上传通过后自动刷新报告" in html
    assert "上传并检查" not in html


def test_recommendations_workbench_surfaces_zero_pending_ad_status(tmp_path) -> None:
    from src.generate_html_report import write_recommendations_workbench_html

    output_path = tmp_path / "latest_recommendations.html"
    results = [
        {
            "has_data": True,
            "marketplace": "US",
            "report_view": {
                "today_task_queue_rows": [],
                "tomorrow_review_rows": [],
                "html_search_term_processing_queue_rows": [
                    {
                        "marketplace": "US",
                        "product_name": "演示台灯",
                        "search_term_or_target": "led desk lamp",
                        "suggested_action": "观察",
                        "confirmed_status": "待确认",
                    },
                    {
                        "marketplace": "US",
                        "product_name": "演示笔记本",
                        "search_term_or_target": "metal desk lamps for kitchen",
                        "suggested_action": "降竞价10%-20%",
                        "confirmed_status": "已执行",
                    },
                ],
                "scale_keyword_rows": [],
                "hidden_low_click_search_terms": 5,
                "listing_price_diagnosis_rows": [],
                "yesterday_attribution_rows": [],
                "action_effect_review_rows": [],
                "keyword_action_effect_review_rows": [],
                "frontend_check_queue_rows": [],
                "inventory_replenishment_rows": [],
                "product_final_decision_rows": [],
                "product_operation_cards": [],
                "frontend_coverage_summary": {},
                "analysis_status": "正式分析",
                "issue_summary": "ok",
            },
        }
    ]

    write_recommendations_workbench_html(results, output_path, "2026-06-08")

    html = output_path.read_text(encoding="utf-8")
    assert "广告动作状态" in html
    assert "待执行 0" in html
    assert "观察项 1" in html
    assert "已执行留档 1" in html
    assert "低点击隐藏项 5" in html
    summary_index = html.index("<h2>今日运营摘要</h2>")
    local_submit_index = html.index('id="local-data-submit"')
    ad_status_index = html.index("广告动作状态")
    assert summary_index < local_submit_index < ad_status_index
    assert html.index("广告动作状态") < html.index("产品级结论")
    ad_detail_index = html.index('id="today-ad-actions-all"')
    assert "section-card is-collapsed" in html[ad_detail_index - 120 : ad_detail_index + 120]
    assert "复制到广告后台" not in html


def test_recommendations_workbench_counts_listing_without_rendering_secondary_block(tmp_path) -> None:
    from src.generate_html_report import write_recommendations_workbench_html

    output_path = tmp_path / "latest_recommendations.html"
    results = [
        {
            "has_data": True,
            "marketplace": "US",
            "report_view": {
                "today_task_queue_rows": [],
                "tomorrow_review_rows": [],
                "html_search_term_processing_queue_rows": [],
                "scale_keyword_rows": [],
                "hidden_low_click_search_terms": 0,
                "listing_price_diagnosis_rows": [
                    {
                        "站点": "US",
                        "产品": "Demo cable ties 100 count",
                        "ASIN": "B0LISTING1",
                        "SKU": "SKU-LISTING",
                        "priority": "P1",
                        "confirmed_status": "待确认",
                        "最可能异常方向": "搜索漏斗待确认",
                    }
                ],
                "yesterday_attribution_rows": [],
                "action_effect_review_rows": [],
                "keyword_action_effect_review_rows": [],
                "frontend_check_queue_rows": [],
                "inventory_replenishment_rows": [],
                "product_final_decision_rows": [],
                "product_operation_cards": [],
                "frontend_coverage_summary": {},
                "analysis_status": "正式分析",
                "issue_summary": "ok",
            },
        }
    ]

    write_recommendations_workbench_html(results, output_path, "2026-06-08")

    html = output_path.read_text(encoding="utf-8")
    assert "P1 合计<strong>1</strong>" in html
    assert "Listing 待确认<strong>1</strong>" in html
    assert 'href="#listing-review">P1 Listing 待确认</a>' not in html
    assert 'id="listing-review"' not in html
    assert 'id="p1-check"' not in html
    assert "通用确认材料" not in html
    assert "建议发给 ChatGPT" not in html
    assert "P1｜US｜Demo cable ties 100 count" not in html


def test_recommendations_workbench_does_not_hide_cross_market_listing_with_same_sku_asin(tmp_path) -> None:
    from src.generate_html_report import write_recommendations_workbench_html

    output_path = tmp_path / "latest_recommendations.html"
    results = [
        {
            "has_data": True,
            "marketplace": "UK",
            "report_view": {
                "today_task_queue_rows": [
                    {
                        "marketplace": "UK",
                        "product_name": "Demo cable ties 20 count",
                        "sku": "SKU-SHARED",
                        "asin": "B0SHARED12",
                        "priority": "P0",
                        "issue_type": "库存 / 利润压力",
                        "confirmed_status": "仅背景参考",
                    }
                ],
                "tomorrow_review_rows": [],
                "html_search_term_processing_queue_rows": [],
                "scale_keyword_rows": [],
                "hidden_low_click_search_terms": 0,
                "listing_price_diagnosis_rows": [],
                "yesterday_attribution_rows": [],
                "action_effect_review_rows": [],
                "keyword_action_effect_review_rows": [],
                "frontend_check_queue_rows": [],
                "inventory_replenishment_rows": [],
                "product_final_decision_rows": [],
                "product_operation_cards": [],
                "frontend_coverage_summary": {},
                "analysis_status": "正式分析",
                "issue_summary": "ok",
            },
        },
        {
            "has_data": True,
            "marketplace": "US",
            "report_view": {
                "today_task_queue_rows": [],
                "tomorrow_review_rows": [],
                "html_search_term_processing_queue_rows": [],
                "scale_keyword_rows": [],
                "hidden_low_click_search_terms": 0,
                "listing_price_diagnosis_rows": [
                    {
                        "站点": "US",
                        "产品": "90L cable ties 20 count",
                        "ASIN": "B0SHARED12",
                        "SKU": "SKU-SHARED",
                        "priority": "P1",
                        "confirmed_status": "待确认",
                        "最可能异常方向": "价格/配送待确认",
                    }
                ],
                "yesterday_attribution_rows": [],
                "action_effect_review_rows": [],
                "keyword_action_effect_review_rows": [],
                "frontend_check_queue_rows": [],
                "inventory_replenishment_rows": [],
                "product_final_decision_rows": [],
                "product_operation_cards": [],
                "frontend_coverage_summary": {},
                "analysis_status": "正式分析",
                "issue_summary": "ok",
            },
        },
    ]

    write_recommendations_workbench_html(results, output_path, "2026-06-08")

    html = output_path.read_text(encoding="utf-8")
    assert "Listing 待确认<strong>1</strong>" in html
    assert 'id="listing-review"' not in html
    assert "P1｜US｜90L cable ties 20 count" not in html


def test_recommendations_workbench_collapses_p0_detail_after_product_summary(tmp_path) -> None:
    from src.generate_html_report import write_recommendations_workbench_html

    output_path = tmp_path / "latest_recommendations.html"
    results = [
        {
            "has_data": True,
            "marketplace": "US",
            "report_view": {
                "today_task_queue_rows": [
                    {
                        "marketplace": "US",
                        "product_name": "演示笔记本3",
                        "sku": "SKU-P0",
                        "asin": "B0P0TEST12",
                        "priority": "P0",
                        "today_action": "只做广告止损，不加预算。",
                        "primary_reason": "14天广告花费高，转化不足。",
                    }
                ],
                "tomorrow_review_rows": [],
                "html_search_term_processing_queue_rows": [],
                "scale_keyword_rows": [],
                "hidden_low_click_search_terms": 0,
                "listing_price_diagnosis_rows": [],
                "yesterday_attribution_rows": [],
                "action_effect_review_rows": [],
                "keyword_action_effect_review_rows": [],
                "frontend_check_queue_rows": [],
                "inventory_replenishment_rows": [],
                "product_final_decision_rows": [],
                "product_operation_cards": [
                    {
                        "marketplace": "US",
                        "product_name": "演示笔记本3",
                        "sku": "SKU-P0",
                        "asin": "B0P0TEST12",
                        "final_decision": "保守跑",
                        "main_reason": "前台自动证据不足，今天只允许广告止损动作。",
                    }
                ],
                "frontend_coverage_summary": {},
                "analysis_status": "正式分析",
                "issue_summary": "ok",
            },
        }
    ]

    write_recommendations_workbench_html(results, output_path, "2026-06-08")

    html = output_path.read_text(encoding="utf-8")
    assert "产品级结论" in html
    assert 'id="p0-actions"' in html
    assert "P0 明细" in html
    assert "首屏以产品级结论为准" in html
    assert html.index("产品级结论") < html.index("<h2>P0 明细</h2>")
    p0_section_start = html.index('id="p0-actions"')
    assert "section-card is-collapsed" in html[p0_section_start - 80 : p0_section_start + 80]


def test_inventory_replenishment_uses_available_stock_before_fba_stock(tmp_path) -> None:
    from src.inventory_replenishment import build_inventory_replenishment

    sku_map = pd.DataFrame(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-AVAILABLE",
                "asin": "B0AVAILABLE",
                "product_name": "可用库存测试品",
            }
        ]
    )
    cost_config = pd.DataFrame()
    product_daily = pd.DataFrame(
        [
            {
                "date": "2026-06-16",
                "marketplace": "UK",
                "sku": "SKU-AVAILABLE",
                "asin": "B0AVAILABLE",
                "fba_stock": 23,
                "available_stock": 35,
                "fbm_stock": 0,
            }
        ]
    )
    views = SimpleNamespace(
        product_windows={
            7: pd.DataFrame(
                [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-AVAILABLE",
                        "asin": "B0AVAILABLE",
                        "total_orders": 7,
                        "clicks": 7,
                        "ad_orders": 1,
                    }
                ]
            ),
            14: pd.DataFrame(
                [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-AVAILABLE",
                        "asin": "B0AVAILABLE",
                        "total_orders": 14,
                        "clicks": 14,
                        "ad_orders": 2,
                    }
                ]
            ),
            30: pd.DataFrame(
                [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-AVAILABLE",
                        "asin": "B0AVAILABLE",
                        "total_orders": 30,
                        "clicks": 30,
                        "ad_orders": 4,
                    }
                ]
            ),
        }
    )

    result = build_inventory_replenishment(
        marketplace="UK",
        sku_map=sku_map,
        cost_config=cost_config,
        views=views,
        product_daily=product_daily,
        report_date=date(2026, 6, 16),
        inventory_path=tmp_path / "missing_inventory.xlsx",
        output_dir=tmp_path,
    )

    row = result["rows"][0]
    assert row["current_inventory"] == 35
    assert row["available_stock"] == 35
    assert row["fba_stock"] == 23
    assert row["days_of_cover"] == 35
    assert row["recommended_reorder_qty"] == 95
    assert row["recommended_reorder_qty"] != 107
    assert "可用库存覆盖约 35" in row["stock_risk_reason"]


def test_inventory_replenishment_restock_recovery_does_not_emit_purchase_instruction(tmp_path) -> None:
    from src.inventory_replenishment import build_inventory_replenishment

    sku_map = pd.DataFrame(
        [
            {
                "marketplace": "UK",
                "sku": "SKU-RECOVERY",
                "asin": "B0RECOVERY",
                "product_name": "恢复期测试品",
            }
        ]
    )
    product_daily = pd.DataFrame(
        [
            {
                "date": "2026-06-16",
                "marketplace": "UK",
                "sku": "SKU-RECOVERY",
                "asin": "B0RECOVERY",
                "fba_stock": 37,
                "available_stock": 133,
                "fbm_stock": 0,
            }
        ]
    )
    views = SimpleNamespace(
        product_windows={
            days: pd.DataFrame(
                [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-RECOVERY",
                        "asin": "B0RECOVERY",
                        "total_orders": days,
                        "clicks": days,
                        "ad_orders": 1,
                    }
                ]
            )
            for days in [7, 14, 30]
        }
    )
    (tmp_path / "autoopt_feedback_input.json").write_text(
        json.dumps(
            {
                "rows": [
                    {
                        "marketplace": "UK",
                        "sku": "SKU-RECOVERY",
                        "asin": "B0RECOVERY",
                        "product_name": "恢复期测试品",
                        "diagnosis_type": "断货刚到货",
                        "confirmed_note": "此前断货，当前刚到货，按补货恢复期观察。",
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    result = build_inventory_replenishment(
        marketplace="UK",
        sku_map=sku_map,
        cost_config=pd.DataFrame(),
        views=views,
        product_daily=product_daily,
        report_date=date(2026, 6, 16),
        inventory_path=tmp_path / "missing_inventory.xlsx",
        output_dir=tmp_path,
    )

    row = result["rows"][0]
    assert row["stock_risk_level"] == "RESTOCK_RECOVERY"
    assert row["recommended_reorder_qty"] is None
    assert row["reference_reorder_qty"] == 0
    assert "不作为今日采购指令" in row["replenishment_advice"]
    assert "建议补货约" not in row["replenishment_advice"]


def test_inventory_replenishment_card_labels_available_stock_and_fba() -> None:
    from src.generate_html_report import _render_inventory_replenishment_cards

    html = _render_inventory_replenishment_cards(
        [
            {
                "marketplace": "UK",
                "product_name": "可用库存测试品",
                "stock_risk_level": "LOW_STOCK",
                "stock_status_label": "低库存",
                "current_inventory": 35,
                "available_stock": 35,
                "fba_stock": 23,
                "days_of_cover": 35,
                "total_lead_time_days": 100,
                "target_cover_days": 130,
                "stock_risk_reason": "可用库存覆盖约 35 天，低于总提前期 100 天。",
                "replenishment_advice": "建议补货约 95 件。",
            }
        ]
    )

    assert "可用库存 35" in html
    assert "FBA 23" in html
    assert "建议补货约 95 件。" in html


def test_inventory_replenishment_cards_sort_low_stock_by_cover_days() -> None:
    from src.generate_html_report import _render_inventory_replenishment_cards

    html = _render_inventory_replenishment_cards(
        [
            {
                "marketplace": "UK",
                "product_name": "覆盖84天产品",
                "stock_risk_level": "LOW_STOCK",
                "stock_status_label": "低库存",
                "current_inventory": 12,
                "available_stock": 12,
                "fba_stock": 12,
                "days_of_cover": 84,
                "recommended_reorder_qty": 7,
                "total_lead_time_days": 100,
                "target_cover_days": 130,
                "stock_risk_reason": "可用库存覆盖约 84 天，低于总提前期 100 天。",
                "replenishment_advice": "建议补货约 7 件。",
            },
            {
                "marketplace": "UK",
                "product_name": "覆盖18天产品",
                "stock_risk_level": "LOW_STOCK",
                "stock_status_label": "低库存",
                "current_inventory": 52,
                "available_stock": 52,
                "fba_stock": 51,
                "days_of_cover": 18.67,
                "recommended_reorder_qty": 311,
                "total_lead_time_days": 100,
                "target_cover_days": 130,
                "stock_risk_reason": "可用库存覆盖约 18.7 天，低于总提前期 100 天。",
                "replenishment_advice": "建议补货约 311 件。",
            },
        ]
    )

    assert html.index("覆盖18天产品") < html.index("覆盖84天产品")


def test_inventory_replenishment_card_marks_missing_available_stock() -> None:
    from src.generate_html_report import _render_inventory_replenishment_cards

    html = _render_inventory_replenishment_cards(
        [
            {
                "marketplace": "DE",
                "product_name": "缺库存测试品",
                "stock_risk_level": "UNKNOWN",
                "stock_status_label": "销量基准不足",
                "current_inventory": None,
                "available_stock": None,
                "fba_stock": None,
                "days_of_cover": None,
                "total_lead_time_days": 100,
                "target_cover_days": 130,
                "stock_risk_reason": "ERP销量表未读取到可用库存，不能判断断货或覆盖天数。",
                "replenishment_advice": "缺少可用库存，需先核对销量表库存字段。",
            }
        ]
    )

    assert "可用库存缺失" in html
    assert "覆盖缺失" in html
    assert "可用库存 None" not in html
    assert "覆盖 None" not in html

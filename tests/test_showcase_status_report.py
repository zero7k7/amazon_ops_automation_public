from __future__ import annotations

import json
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook

import scripts.build_showcase_status_report as status_report
import scripts.run_daily_update as daily_update


MARKETPLACE_SUMMARIES = [
    {
        "marketplace": "UK",
        "report_date": "2026-06-08",
        "ads_row_count": 11,
        "erp_row_count": 21,
        "sku_count": 3,
        "asin_count": 4,
    },
    {
        "marketplace": "US",
        "report_date": "2026-06-08",
        "ads_row_count": 12,
        "erp_row_count": 22,
        "sku_count": 5,
        "asin_count": 6,
    },
    {
        "marketplace": "DE",
        "report_date": "2026-06-08",
        "ads_row_count": 13,
        "erp_row_count": 23,
        "sku_count": 7,
        "asin_count": 8,
    },
]


def _enhanced_request_rows() -> list[dict[str, object]]:
    return [
        {
            "marketplace": "UK",
            "trigger_sku": "",
            "trigger_asin": "",
            "trigger_product_name": "",
            "issue_type": "增强数据文件导入",
            "report_type": "流量和销售数据",
            "period": "recent_vs_prior",
            "start_date": "",
            "end_date": "",
            "expected_filename": "traffic_sales_uk.xlsx",
            "target_path": "data/raw_amazon_custom/UK/",
            "target_folder": "data/raw_amazon_custom/UK/",
            "required": "否",
            "seller_central_page": "定制分析",
            "instruction": "已导入",
            "status": "已导入",
            "file_type": "traffic_sales",
            "format_type": "single",
            "detected_from": "",
            "detected_date_range": "",
            "freshness": "stale",
            "used_in_diagnosis": "否",
        }
    ]


def _patch_paths(monkeypatch, tmp_path: Path) -> Path:
    root = tmp_path / "repo"
    output = root / "data" / "output"
    output.mkdir(parents=True)
    monkeypatch.setattr(status_report, "ROOT", root)
    monkeypatch.setattr(status_report, "OUTPUT", output)
    monkeypatch.setattr(status_report, "LATEST_ANALYSIS", output / "latest_analysis.json")
    monkeypatch.setattr(status_report, "COST_AUDIT", output / "cost_config_diff_summary.json")
    monkeypatch.setattr(status_report, "UNKNOWN_AUDIT", output / "unknown_inbox_audit.json")
    monkeypatch.setattr(status_report, "UNKNOWN_INBOX_DIR", root / "data" / "inbox" / "_unknown")
    monkeypatch.setattr(status_report, "OWNER_DRAFT", root / "docs" / "parallel_lock_owners.draft.json")
    monkeypatch.setattr(status_report, "OWNER_MANIFEST", root / "docs" / "parallel_lock_owners.json")
    monkeypatch.setattr(status_report, "BUSINESS_CONFIRMATION", root / "docs" / "business_review_confirmation.json")
    monkeypatch.setattr(status_report, "JSON_OUTPUT", output / "showcase_status_report.json")
    monkeypatch.setattr(status_report, "MARKDOWN_OUTPUT", output / "showcase_status_report.md")
    return output


def _write_required_outputs(output: Path) -> None:
    report_date = "2026-06-08"
    dashboard_rows = "".join(
        f"<tr><td>{summary['marketplace']}</td><td>{summary['ads_row_count']}</td>"
        f"<td>{summary['erp_row_count']}</td><td>{summary['sku_count']}</td>"
        f"<td>{summary['asin_count']}</td><td>正式分析</td><td>无</td></tr>"
        f"<a href=\"{str(summary['marketplace']).lower()}_report.html\">打开报告</a>"
        for summary in MARKETPLACE_SUMMARIES
    )
    html = {
        "dashboard.html": [
            "运营状态入口",
            "打开三分钟摘要",
            "打开 ALL 运营控制台",
            "latest_recommendations.html",
            "summary.html",
            dashboard_rows,
        ],
        "latest_recommendations.html": [
            "今天广告动作",
            'id="today-ad-actions-all"',
            'id="product-operation-cards"',
            'id="frontend-evidence-status"',
            'id="action-effect-review"',
            "待确认 0",
            "市场调查",
            "产品级结论",
            "执行后效果复盘",
            "提交今日数据",
            "系统结论",
            "融合诊断",
        ],
        "summary.html": ["三分钟摘要", "今日开工结论", "今天先做", "昨天动作今天要盯"],
        "uk_report.html": [
            "亚马逊运营日报｜UK",
            "广告状态",
            "数据质量与增强数据",
            "站点状态",
            "市场调查",
            "执行后效果复盘",
            'id="product-operation-cards"',
            'id="frontend-evidence-status"',
            'id="action-effect-review"',
        ],
        "us_report.html": [
            "亚马逊运营日报｜US",
            "广告状态",
            "数据质量与增强数据",
            "站点状态",
            "市场调查",
            "执行后效果复盘",
            'id="product-operation-cards"',
            'id="frontend-evidence-status"',
            'id="action-effect-review"',
        ],
        "de_report.html": [
            "亚马逊运营日报｜DE",
            "广告状态",
            "数据质量与增强数据",
            "站点状态",
            "市场调查",
            "执行后效果复盘",
            'id="product-operation-cards"',
            'id="frontend-evidence-status"',
            'id="action-effect-review"',
        ],
    }
    for name, markers in html.items():
        (output / name).write_text(" ".join([report_date, *markers]), encoding="utf-8")
    (output / "latest_recommendations.md").write_text(
        "# 亚马逊运营日报汇总建议\n\n"
        "报告日期 2026-06-08\n\n"
        "## 1. 各站点状态摘要\n\n"
        "## 2. 今日动作清单\n\n"
        "## 3. 明日复查清单\n",
        encoding="utf-8",
    )
    (output / "marketplace_summary.md").write_text(
        "\n".join(
            [
                "# 站点汇总｜2026-06-08",
                "",
                "| 站点 | 广告行数 | ERP行数 | SKU数 | ASIN数 | 状态 | 说明 |",
                "| --- | ---: | ---: | ---: | ---: | --- | --- |",
                *[
                    f"| {summary['marketplace']} | {summary['ads_row_count']} | {summary['erp_row_count']} | "
                    f"{summary['sku_count']} | {summary['asin_count']} | 正式分析 | 无 |"
                    for summary in MARKETPLACE_SUMMARIES
                ],
                "",
            ]
        ),
        encoding="utf-8",
    )
    assets = output / "assets"
    assets.mkdir()
    (assets / "report.css").write_text(".ad-task-card{} .ad-copy-box{}", encoding="utf-8")
    (assets / "report.js").write_text(
        "document.querySelector('[data-ad-complete-checkbox]');"
        "document.querySelector('[data-ad-filter-summary]');",
        encoding="utf-8",
    )
    (output / "enhanced_data_requests.md").write_text(
        "# 需要补充导出的增强数据｜2026-06-08\n\n"
        "| 站点 | 状态 | 报表类型 | 周期 | 日期范围 | 导出后文件名 | 目标文件夹 | 必需 |\n"
        "|---|---|---|---|---|---|---|---|\n"
        "| UK | 已导入 | 流量和销售数据 | recent_vs_prior | N/A ~ N/A | traffic_sales_uk.xlsx | data/raw_amazon_custom/UK/ | 否 |\n",
        encoding="utf-8",
    )
    enhanced_workbook = Workbook()
    enhanced_sheet = enhanced_workbook.active
    enhanced_sheet.title = "Sheet1"
    enhanced_rows = _enhanced_request_rows()
    enhanced_headers = list(enhanced_rows[0].keys())
    enhanced_sheet.append(enhanced_headers)
    for row in enhanced_rows:
        enhanced_sheet.append([row.get(header, "") for header in enhanced_headers])
    enhanced_workbook.save(output / "enhanced_data_requests.xlsx")
    enhanced_workbook.close()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Metrics_Validation"
    worksheet.append(["check", "status"])
    worksheet.append(["placeholder", "ok"])
    overview = workbook.create_sheet("总览")
    overview.append(["报告日期", report_date])
    for marketplace in ["UK", "US", "DE"]:
        sheet = workbook.create_sheet(f"{marketplace}_今日总览")
        sheet.append(["marketplace", "report_date"])
        sheet.append([marketplace, report_date])
    for sheet_name in [
        "产品最终决策",
        "产品运营卡",
        "前台证据队列",
        "今日动作清单",
        "明日复查清单",
        "库存补货提醒",
        "UK_前台证据队列",
        "US_前台证据队列",
        "DE_前台证据队列",
    ]:
        workbook.create_sheet(sheet_name).append(["marketplace", "sku", "asin"])
    workbook.save(output / "amazon_ops_report_2026-06-08.xlsx")


def _write_latest_analysis(output: Path) -> None:
    report_view_snapshot = {
        "today_task_queue_rows": [],
        "today_action_groups": {},
        "tomorrow_review_rows": [],
        "search_term_processing_queue_rows": [],
        "html_search_term_processing_queue_rows": [],
        "scale_rows": [],
        "scale_keyword_rows": [],
        "growth_test_rows": [],
        "frontend_check_queue_rows": [],
        "inventory_replenishment_rows": [],
        "product_final_decision_rows": [],
        "product_operation_cards": [],
        "final_decision_summary": {},
        "decision_gate_counts": {},
        "frontend_coverage_summary": daily_update._frontend_coverage_expected([]),
        "action_effect_review_rows": [],
        "keyword_action_effect_review_rows": [],
        "cost_profit_diagnosis_rows": [],
        "listing_price_diagnosis_rows": [],
    }
    report_date = "2026-06-08"
    payload = {
        "report_date": report_date,
        "import_summary": {
            "ads_imported_rows": 10,
            "erp_imported_rows": 5,
        },
        "enhanced_data_requests": _enhanced_request_rows(),
        "inventory_replenishment_rows": [],
        "product_final_decision_rows": [],
        "product_operation_cards": [],
        "final_decision_summary": {
            summary["marketplace"]: report_view_snapshot["final_decision_summary"]
            for summary in MARKETPLACE_SUMMARIES
        },
        "decision_gate_counts": {
            summary["marketplace"]: report_view_snapshot["decision_gate_counts"]
            for summary in MARKETPLACE_SUMMARIES
        },
        "marketplace_results": [
            {
                "marketplace": summary["marketplace"],
                "summary": summary,
                "report_view_snapshot": report_view_snapshot,
            }
            for summary in MARKETPLACE_SUMMARIES
        ],
    }
    payload["frontend_coverage_summary"] = daily_update._frontend_coverage_aggregate_from_snapshots(
        payload["marketplace_results"]
    )
    (output / "latest_analysis.json").write_text(
        json.dumps(payload),
        encoding="utf-8",
    )
    date_token = report_date.replace("-", "")
    autoopt_payload = {
        "report_date": report_date,
        "rows": [],
        "action_review_rows": [],
        "keyword_action_review_rows": [],
        "product_final_decisions": [],
        "final_decision_summary": {},
        "decision_gate_counts": {},
        "learned_rules": [],
        "manual_learning_rows": [],
        "positive_action_patterns": [],
        "negative_action_patterns": [],
        "product_strategy_profiles": [],
        "keyword_strategy_memory": [],
    }
    (output / f"autoopt_log_{date_token}.json").write_text(json.dumps(autoopt_payload), encoding="utf-8")
    for filename in [
        f"action_review_{date_token}.json",
        f"keyword_action_review_{date_token}.json",
        f"learned_rules_{date_token}.json",
        f"manual_learning_log_{date_token}.json",
        f"product_strategy_profiles_{date_token}.json",
        f"keyword_strategy_memory_{date_token}.json",
    ]:
        (output / filename).write_text("[]", encoding="utf-8")
    (output / f"self_optimization_log_{date_token}.json").write_text(
        json.dumps(
            {
                "report_date": report_date,
                "learned_rules": [],
                "manual_learning_rows": [],
                "keyword_action_review_rows": [],
                "positive_action_patterns": [],
                "negative_action_patterns": [],
                "product_strategy_profiles": [],
                "keyword_strategy_memory": [],
            }
        ),
        encoding="utf-8",
    )
    workbook = Workbook()
    workbook.active.title = "autoopt_log"
    workbook["autoopt_log"].append(["placeholder"])
    summary = workbook.create_sheet("summary")
    summary.append(["report_date"])
    summary.append([report_date])
    workbook.create_sheet("action_review").append(["action_id"])
    workbook.create_sheet("keyword_action_review").append(["action_id"])
    workbook.create_sheet("final_decisions").append(["marketplace", "sku", "asin"])
    workbook.save(output / f"autoopt_{date_token}.xlsx")


def test_run_strict_readiness_uses_business_confirmation_flags(monkeypatch, tmp_path) -> None:
    root = tmp_path / "repo"
    script = root / "scripts" / "check_showcase_commit_readiness.py"
    confirmation = root / "docs" / "business_review_confirmation.json"
    script.parent.mkdir(parents=True)
    confirmation.parent.mkdir(parents=True)
    script.write_text("placeholder", encoding="utf-8")
    confirmation.write_text(
        json.dumps(
            {
                "confirmation_status": "confirmed",
                "allowed_strict_flags": {
                    "manual_config_confirmed": True,
                    "business_review_confirmed": True,
                    "parallel_lock_owner_confirmed": True,
                    "daily_update_verified": True,
                },
            }
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(status_report, "ROOT", root)
    monkeypatch.setattr(status_report, "STRICT_READINESS_SCRIPT", script)
    monkeypatch.setattr(status_report, "BUSINESS_CONFIRMATION", confirmation)
    captured: dict[str, list[str]] = {}

    def fake_run(command, **kwargs):
        captured["command"] = [str(item) for item in command]
        return SimpleNamespace(returncode=0, stdout="[check] strict confirmations accepted\n")

    monkeypatch.setattr(status_report.subprocess, "run", fake_run)

    result = status_report._run_strict_readiness()

    assert result["passed"] is True
    assert "--strict" in captured["command"]
    assert "--manual-config-confirmed" in captured["command"]
    assert "--business-review-confirmed" in captured["command"]
    assert "--parallel-lock-owner-confirmed" in captured["command"]
    assert "--daily-update-verified" in captured["command"]


def test_showcase_status_report_flags_showcase_ready_but_not_business_ready(monkeypatch, tmp_path) -> None:
    output = _patch_paths(monkeypatch, tmp_path)
    monkeypatch.setattr(
        status_report,
        "_run_strict_readiness",
        lambda: {"passed": False, "exit_code": 1, "summary": ["strict failure"]},
    )
    _write_required_outputs(output)
    _write_latest_analysis(output)
    (output / "cost_config_diff_summary.json").write_text(
        json.dumps(
            {
                "total_changed_cells": 940,
                "sheets": {
                    "product_cost_config": {
                        "keyed_diff": {
                            "changed_record_count": 21,
                            "sample_changed_records": [
                                {
                                    "marketplace": "UK",
                                    "sku": "SKU-1",
                                    "asin": "ASIN-1",
                                    "product_name": "产品 1",
                                    "changed_fields": [
                                        {"field": "selling_price", "before": 20, "after": 22},
                                        {"field": "suggested_target_acos", "before": 0.12, "after": 0.18},
                                        {"field": "current_inventory", "before": 10, "after": 5},
                                    ],
                                }
                            ],
                        },
                    },
                },
            }
        ),
        encoding="utf-8",
    )
    (output / "unknown_inbox_audit.json").write_text(
        json.dumps(
            [
                {
                    "blocks_daily_update": True,
                    "original_path": str(status_report.UNKNOWN_INBOX_DIR / "metric.xlsx"),
                    "likely_report_type": "seller_central_promotion_metrics",
                    "likely_marketplace": "DE_or_EUR",
                    "business_interpretation": "Seller Central 促销表现 metric-data",
                    "recommendation": "当前不参与 daily import",
                }
            ]
        ),
        encoding="utf-8",
    )
    unknown_dir = status_report.UNKNOWN_INBOX_DIR
    unknown_dir.mkdir(parents=True)
    (unknown_dir / "metric.xlsx").write_text("placeholder", encoding="utf-8")
    status_report.OWNER_DRAFT.parent.mkdir(parents=True)
    status_report.OWNER_DRAFT.write_text(
        json.dumps(
            {
                "confirmation_status": "draft_requires_manual_owner_confirmation",
                "owners": [
                    {
                        "owner": "REVIEW_REQUIRED_report_display",
                        "work_package": "report_display",
                        "files": ["src/generate_html_report.py", "scripts/validate_showcase_mvp.py"],
                        "confirmation_status": "pending",
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    status_report.OWNER_MANIFEST.write_text(
        json.dumps(
            {
                "confirmation_status": "lock_ownership_confirmed_by_main_thread_only",
                "owners": [
                    {
                        "owner": "codex_main_thread_integration",
                        "work_package": "report_display",
                        "files": ["src/generate_html_report.py", "scripts/validate_showcase_mvp.py"],
                        "confirmation_status": "lock_owner_confirmed",
                    }
                ],
            }
        ),
        encoding="utf-8",
    )

    code = status_report.main(["--fail-if-not-business-ready"])
    payload = json.loads(status_report.JSON_OUTPUT.read_text(encoding="utf-8"))
    markdown = status_report.MARKDOWN_OUTPUT.read_text(encoding="utf-8")

    assert code == 1
    assert payload["showcase_ready"] is True
    assert payload["business_submit_ready"] is False
    assert payload["daily_update_blocked"] is True
    assert payload["cost_config_changed_cells"] == 940
    assert payload["cost_config_changed_product_records"] == 21
    assert payload["strict_readiness_passed"] is False
    assert payload["strict_readiness_exit_code"] == 1
    assert payload["cost_config_sample_records"][0]["sku"] == "SKU-1"
    assert "售价变化会影响利润和 ACOS 目标" in payload["cost_config_sample_records"][0]["risk_hints"]
    assert payload["owner_draft_confirmation_status"] == "draft_requires_manual_owner_confirmation"
    assert payload["owner_draft_groups"][0]["work_package"] == "report_display"
    assert payload["owner_draft_groups"][0]["file_count"] == 2
    assert payload["owner_manifest_confirmation_status"] == "lock_ownership_confirmed_by_main_thread_only"
    assert payload["owner_manifest_groups"][0]["owner"] == "codex_main_thread_integration"
    assert payload["owner_manifest_groups"][0]["file_count"] == 2
    assert payload["business_confirmation_valid"] is False
    assert payload["unknown_inbox_details"][0]["likely_report_type"] == "seller_central_promotion_metrics"
    assert "展示状态：可以展示" in markdown
    assert "业务提交状态：不能业务提交" in markdown
    assert "展示产物内容校验：通过" in markdown
    assert "严格提交门禁：未通过" in markdown
    assert "Seller Central 促销表现 metric-data" in markdown
    assert "## 成本配置样本" in markdown
    assert "SKU-1" in markdown
    assert "selling_price" in markdown
    assert "## 并行 owner 草案" in markdown
    assert "REVIEW_REQUIRED_report_display" in markdown
    assert "## 并行 owner 正式清单" in markdown
    assert "codex_main_thread_integration" in markdown
    assert "业务确认记录：未确认" in markdown


def test_showcase_status_report_fails_when_outputs_missing(monkeypatch, tmp_path) -> None:
    output = _patch_paths(monkeypatch, tmp_path)
    monkeypatch.setattr(
        status_report,
        "_run_strict_readiness",
        lambda: {"passed": True, "exit_code": 0, "summary": ["strict ok"]},
    )
    (output / "latest_analysis.json").write_text(
        json.dumps({"report_date": "", "marketplace_results": []}),
        encoding="utf-8",
    )

    code = status_report.main(["--fail-if-not-showcase-ready"])
    payload = json.loads(status_report.JSON_OUTPUT.read_text(encoding="utf-8"))

    assert code == 1
    assert payload["showcase_ready"] is False
    assert "summary.html" in payload["missing_outputs"]


def test_strict_readiness_failure_blocks_business_submit(monkeypatch, tmp_path) -> None:
    output = _patch_paths(monkeypatch, tmp_path)
    monkeypatch.setattr(
        status_report,
        "_run_strict_readiness",
        lambda: {
            "passed": False,
            "exit_code": 1,
            "summary": ["[fail] strict mode requires confirmations"],
        },
    )
    _write_required_outputs(output)
    _write_latest_analysis(output)
    (output / "cost_config_diff_summary.json").write_text(
        json.dumps({"total_changed_cells": 0, "sheets": {}}),
        encoding="utf-8",
    )
    (output / "unknown_inbox_audit.json").write_text("[]", encoding="utf-8")

    code = status_report.main(["--fail-if-not-business-ready"])
    payload = json.loads(status_report.JSON_OUTPUT.read_text(encoding="utf-8"))
    markdown = status_report.MARKDOWN_OUTPUT.read_text(encoding="utf-8")

    assert code == 1
    assert payload["showcase_ready"] is True
    assert payload["business_submit_ready"] is False
    assert payload["daily_update_blocked"] is False
    assert payload["cost_config_changed_cells"] == 0
    assert payload["strict_readiness_passed"] is False
    assert "处理 scripts/check_showcase_commit_readiness.py --strict 的失败项" in markdown


def test_status_report_does_not_treat_placeholder_outputs_as_showcase_ready(monkeypatch, tmp_path) -> None:
    output = _patch_paths(monkeypatch, tmp_path)
    monkeypatch.setattr(
        status_report,
        "_run_strict_readiness",
        lambda: {"passed": True, "exit_code": 0, "summary": ["strict ok"]},
    )
    for name in status_report.REQUIRED_FORMAL_OUTPUTS:
        path = output / name
        if path.suffix != ".json":
            path.write_text("placeholder", encoding="utf-8")
    (output / "latest_analysis.json").write_text(
        json.dumps(
            {
                "report_date": "2026-06-08",
                "marketplace_results": [
                    {"marketplace": "UK"},
                    {"marketplace": "US"},
                    {"marketplace": "DE"},
                ],
            }
        ),
        encoding="utf-8",
    )
    (output / "amazon_ops_report_2026-06-08.xlsx").write_text("placeholder", encoding="utf-8")

    code = status_report.main(["--fail-if-not-showcase-ready"])
    payload = json.loads(status_report.JSON_OUTPUT.read_text(encoding="utf-8"))

    assert code == 1
    assert payload["showcase_ready"] is False
    assert payload["output_validation_passed"] is False


def test_status_report_scans_unknown_inbox_when_audit_json_is_missing(monkeypatch, tmp_path) -> None:
    output = _patch_paths(monkeypatch, tmp_path)
    monkeypatch.setattr(
        status_report,
        "_run_strict_readiness",
        lambda: {"passed": True, "exit_code": 0, "summary": ["strict ok"]},
    )
    _write_required_outputs(output)
    _write_latest_analysis(output)
    (output / "cost_config_diff_summary.json").write_text(json.dumps({"total_changed_cells": 0}), encoding="utf-8")
    status_report.UNKNOWN_INBOX_DIR.mkdir(parents=True)
    (status_report.UNKNOWN_INBOX_DIR / "metric.xlsx").write_text("placeholder", encoding="utf-8")

    code = status_report.main(["--fail-if-not-business-ready"])
    payload = json.loads(status_report.JSON_OUTPUT.read_text(encoding="utf-8"))

    assert code == 1
    assert payload["showcase_ready"] is True
    assert payload["business_submit_ready"] is False
    assert payload["daily_update_blocked"] is True
    assert payload["unknown_inbox_blockers"] == 1


def test_status_report_accepts_confirmed_cost_changes_when_strict_passes(monkeypatch, tmp_path) -> None:
    output = _patch_paths(monkeypatch, tmp_path)
    monkeypatch.setattr(
        status_report,
        "_run_strict_readiness",
        lambda: {"passed": True, "exit_code": 0, "summary": ["strict ok"]},
    )
    _write_required_outputs(output)
    _write_latest_analysis(output)
    (output / "cost_config_diff_summary.json").write_text(
        json.dumps(
            {
                "total_changed_cells": 940,
                "sheets": {
                    "product_cost_config": {
                        "keyed_diff": {
                            "changed_record_count": 21,
                            "sample_changed_records": [],
                        }
                    }
                },
            }
        ),
        encoding="utf-8",
    )
    (output / "unknown_inbox_audit.json").write_text("[]", encoding="utf-8")
    status_report.BUSINESS_CONFIRMATION.parent.mkdir(parents=True)
    status_report.BUSINESS_CONFIRMATION.write_text(
        json.dumps(
            {
                "confirmation_status": "confirmed",
                "allowed_strict_flags": {
                    "manual_config_confirmed": True,
                    "business_review_confirmed": True,
                    "parallel_lock_owner_confirmed": True,
                    "daily_update_verified": True,
                },
            }
        ),
        encoding="utf-8",
    )

    code = status_report.main(["--fail-if-not-business-ready"])
    payload = json.loads(status_report.JSON_OUTPUT.read_text(encoding="utf-8"))
    markdown = status_report.MARKDOWN_OUTPUT.read_text(encoding="utf-8")

    assert code == 0
    assert payload["showcase_ready"] is True
    assert payload["business_submit_ready"] is True
    assert payload["cost_config_changed_cells"] == 940
    assert payload["business_confirmation_valid"] is True
    assert not any("业务 owner 按 docs/business_confirmation_template.md" in item for item in payload["required_next_actions"])
    assert "业务提交状态：可以业务提交" in markdown
    assert "业务确认记录：有效" in markdown

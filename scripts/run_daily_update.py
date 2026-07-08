from __future__ import annotations

import ast
import html
import math
import json
import re
import subprocess
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from hashlib import sha256
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from src.autoopt_feedback import (
    ACTION_REVIEW_REQUIRED_FIELDS,
    KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS,
    NEGATIVE_DISPLAY_OUTCOMES,
    POSITIVE_DISPLAY_OUTCOMES,
    is_executable_action,
    load_feedback_input,
    make_action_id,
)
from src.html_pages.components_review import (
    _action_effect_review_sort_key as _html_action_effect_review_sort_key,
    _keyword_review_sort_key as _html_keyword_review_sort_key,
    _review_display_judgement as _html_review_display_judgement,
)
from src.product_decision_layer import FRONTEND_EVIDENCE_DECISION_FIELDS
from src.product_decision_layer import FRONTEND_GATED_GROWTH_ACTIONS as GROWTH_ACTIONS
from src.product_decision_layer import PRODUCT_IDENTITY_FIELDS
from src.market_survey_completeness import MARKET_SURVEY_COMPLETENESS_FIELDS
from src.sellersprite_history import HISTORY_TREND_FIELDS


PRODUCT_GATED_AD_ACTIONS = GROWTH_ACTIONS | {"create_exact_low_budget"}
IMPORT_MANIFEST_JSON = ROOT / "data" / "output" / "import_manifest.json"
IMPORT_MANIFEST_XLSX = ROOT / "data" / "output" / "import_manifest.xlsx"
OUTPUT_DIR = ROOT / "data" / "output"
LATEST_ANALYSIS_JSON = OUTPUT_DIR / "latest_analysis.json"
FRONTEND_RESULTS_JSON = OUTPUT_DIR / "frontend_check_results.json"
DAILY_UPDATE_VALIDATION_RECEIPT = OUTPUT_DIR / "daily_update_validation_receipt.json"
DAILY_UPDATE_RECEIPT_SCHEMA_VERSION = 1
DB_STATE_FILES = [
    ROOT / "database" / "amazon_ops.db",
    ROOT / "database" / "amazon_ops.db-wal",
    ROOT / "database" / "amazon_ops.db-shm",
]
ARCHIVE_STATE_DIRS = [
    ROOT / "data" / "archive" / "ads",
    ROOT / "data" / "archive" / "erp",
]
REQUIRED_REFRESHED_OUTPUTS = [
    LATEST_ANALYSIS_JSON,
    OUTPUT_DIR / "latest_recommendations.html",
    OUTPUT_DIR / "latest_recommendations.md",
    OUTPUT_DIR / "dashboard.html",
    OUTPUT_DIR / "summary.html",
    OUTPUT_DIR / "uk_report.html",
    OUTPUT_DIR / "us_report.html",
    OUTPUT_DIR / "de_report.html",
    OUTPUT_DIR / "marketplace_summary.md",
    OUTPUT_DIR / "enhanced_data_requests.md",
    OUTPUT_DIR / "enhanced_data_requests.xlsx",
    OUTPUT_DIR / "assets" / "report.css",
    OUTPUT_DIR / "assets" / "report.js",
]
DATE_SCOPED_OUTPUT_PATTERNS = [
    "amazon_ops_report_*.xlsx",
    "autoopt_log_*.json",
    "autoopt_*.xlsx",
    "action_review_*.json",
    "keyword_action_review_*.json",
    "learned_rules_*.json",
    "manual_learning_log_*.json",
    "product_strategy_profiles_*.json",
    "keyword_strategy_memory_*.json",
    "self_optimization_log_*.json",
]
ALLOWED_IMPORT_STATUSES = {"imported", "split_imported", "skipped_old_duplicate"}
CORE_IMPORT_TYPES = {"ads_report_all", "erp_sales_all"}
IMPORT_SUCCESS_STATUSES = {"imported", "split_imported"}
VALID_MARKETPLACES = {"UK", "US", "DE"}
MARKETPLACE_DISPLAY_ORDER = {"US": 0, "UK": 1, "DE": 2}
SUSPICIOUS_CORE_IMPORT_PATTERNS = [
    (re.compile(r"(?:^|[/\\])(?:ads_report_all|amazon_ads_all)(?:[._-]|$)", re.IGNORECASE), "ads_report_all"),
    (re.compile(r"(?:^|[/\\])(?:erp_sales_all|sales_report_all)(?:[._-]|$)", re.IGNORECASE), "erp_sales_all"),
]
ACTION_GROUPS = ["广告动作", "Listing / 价格动作", "成本 / 利润动作"]
DATE_RE = re.compile(r"20\d{2}-\d{2}-\d{2}")
TEXT_REPORT_REQUIRED_MARKERS = {
    "latest_recommendations.html": ["today-ad-actions-all", "产品级结论", "市场调查", "执行后效果复盘"],
    "latest_recommendations.md": ["# 亚马逊运营日报汇总建议", "## 1. 各站点状态摘要", "## 2. 今日动作清单", "## 3. 明日复查清单"],
    "dashboard.html": ["latest_recommendations.html", "summary.html", "UK", "US", "DE"],
    "summary.html": ["三分钟摘要", "今日开工结论", "今天先做", "昨天动作今天要盯"],
    "uk_report.html": ["亚马逊运营日报｜UK", "站点状态", "市场调查", "执行后效果复盘"],
    "us_report.html": ["亚马逊运营日报｜US", "站点状态", "市场调查", "执行后效果复盘"],
    "de_report.html": ["亚马逊运营日报｜DE", "站点状态", "市场调查", "执行后效果复盘"],
    "marketplace_summary.md": ["| UK |", "| US |", "| DE |"],
    "enhanced_data_requests.md": ["# 需要补充导出的增强数据"],
}
TEXT_REPORT_REQUIRED_ANCHORS = {
    "latest_recommendations.html": [
        "product-operation-cards",
        "today-ad-actions-all",
        "frontend-evidence-status",
        "action-effect-review",
    ],
    "uk_report.html": ["product-operation-cards", "frontend-evidence-status", "action-effect-review"],
    "us_report.html": ["product-operation-cards", "frontend-evidence-status", "action-effect-review"],
    "de_report.html": ["product-operation-cards", "frontend-evidence-status", "action-effect-review"],
}
REPORT_ASSET_REQUIRED_TOKENS = {
    "report.css": [".ad-task-card", ".ad-copy-box"],
    "report.js": ["data-ad-complete-checkbox", "data-ad-filter-summary"],
}
REQUIRED_REPORT_VIEW_SNAPSHOT_LIST_KEYS = {
    "today_task_queue_rows",
    "tomorrow_review_rows",
    "search_term_processing_queue_rows",
    "html_search_term_processing_queue_rows",
    "scale_rows",
    "scale_keyword_rows",
    "growth_test_rows",
    "frontend_check_queue_rows",
    "inventory_replenishment_rows",
    "product_final_decision_rows",
    "product_operation_cards",
    "action_effect_review_rows",
    "keyword_action_effect_review_rows",
    "cost_profit_diagnosis_rows",
    "listing_price_diagnosis_rows",
}
REQUIRED_REPORT_VIEW_SNAPSHOT_DICT_KEYS = {
    "today_action_groups",
    "final_decision_summary",
    "decision_gate_counts",
    "frontend_coverage_summary",
}
ACTION_ID_CONSISTENCY_SNAPSHOT_KEYS = REQUIRED_REPORT_VIEW_SNAPSHOT_LIST_KEYS - {
    "frontend_check_queue_rows",
    "inventory_replenishment_rows",
    "product_final_decision_rows",
    "product_operation_cards",
    "scale_rows",
}
ACTION_ID_REQUIRED_SNAPSHOT_KEYS = {
    "search_term_processing_queue_rows",
    "html_search_term_processing_queue_rows",
    "scale_keyword_rows",
    "growth_test_rows",
}
VALID_ACTION_SCOPES = {"product", "campaign", "search_term", "asin_target"}
REPORT_VIEW_MARKETPLACE_SUMMARY_KEYS = {"final_decision_summary", "decision_gate_counts"}
FRONTEND_COVERAGE_COUNT_FIELDS = [
    "frontend_queue_total",
    "frontend_usable_evidence_count",
    "frontend_decision_ready_count",
    "frontend_reference_evidence_count",
    "frontend_live_success_count",
    "frontend_cached_count",
    "frontend_pending_or_stale_count",
    "frontend_search_success_count",
    "frontend_search_partial_count",
    "frontend_product_page_success_count",
    "frontend_competitor_search_success_count",
    "frontend_own_sellersprite_count",
    "frontend_own_sellersprite_today_count",
    "frontend_own_sellersprite_cache_count",
    "frontend_own_sellersprite_pending_count",
    "frontend_own_sellersprite_failed_count",
    "frontend_sellersprite_trend_ready_count",
    "frontend_competitor_discovery_count",
    "frontend_competitor_pool_count",
    "frontend_competitor_pool_today_count",
    "frontend_competitor_pool_cache_count",
    "frontend_competitor_pool_pending_count",
    "frontend_competitor_pool_failed_count",
    "frontend_competitor_sellersprite_count",
    "frontend_competitor_sellersprite_today_count",
    "frontend_competitor_sellersprite_cache_count",
    "frontend_competitor_sellersprite_pending_count",
    "frontend_competitor_sellersprite_asin_count",
    "frontend_amazon_search_validation_count",
    "frontend_scalable_strong_count",
    "frontend_weak_defensive_count",
    "frontend_insufficient_count",
    "frontend_strong_evidence_count",
    "frontend_background_evidence_count",
    "frontend_unusable_evidence_count",
    "market_survey_complete_count",
    "market_survey_usable_count",
    "market_survey_insufficient_count",
    "market_survey_failed_count",
]
FRONTEND_COVERAGE_RATE_FIELDS = [
    "frontend_usable_evidence_rate",
    "frontend_decision_ready_rate",
    "frontend_reference_evidence_rate",
    "frontend_live_success_rate",
    "frontend_search_success_rate",
    "frontend_search_observed_rate",
    "market_survey_average_score",
]
PRODUCT_OPERATION_CARD_IDENTITY_FIELDS = list(PRODUCT_IDENTITY_FIELDS)
PRODUCT_OPERATION_CARD_CONTENT_FIELDS = [
    "product_name",
    "final_decision",
    "final_decision_label",
    "decision_reason",
    "operation_main_reason",
    "fusion_issue_type",
    "fusion_confidence",
    "fusion_action_gate",
    "fusion_reason",
    "fusion_today_action",
    "fusion_do_not_do",
    "fusion_review_window",
    "today_allowed_actions",
    "today_blocked_actions",
    "frontend_status",
    "frontend_freshness",
    "frontend_auto_conclusion_label",
    "frontend_evidence_quality_score",
    "frontend_findings",
    "frontend_search_status",
    "frontend_search_findings",
    "frontend_search_result_count",
    "frontend_search_partial_evidence",
    "amazon_search_validation_status",
    "amazon_search_visible_competitors",
    "competitor_discovery_status",
    "competitor_discovery_error",
    "competitor_discovery_source_page",
    "competitor_discovery_source",
    "competitor_pool_status",
    "competitor_pool_asins",
    "competitor_pool_count",
    "competitor_pool_confidence",
    "competitor_overlap_keywords",
    "competitor_frontend_status",
    "competitor_frontend_asins",
    "competitor_frontend_count",
    "comparable_competitor_count",
    "competitor_sellersprite_status",
    "competitor_sellersprite_asin_count",
    "competitor_sellersprite_keyword_count",
    "competitor_keyword_pressure",
    "competitor_shared_keywords",
    "own_missing_competitor_keywords",
    "own_ad_terms_not_in_sellersprite",
    "seller_sprite_check_status",
    "seller_sprite_keyword_count",
    *HISTORY_TREND_FIELDS,
    *MARKET_SURVEY_COMPLETENESS_FIELDS,
    *FRONTEND_EVIDENCE_DECISION_FIELDS,
    "frontend_cache_used",
    "frontend_evidence_audit_summary",
    "frontend_evidence_audit_detail",
    "frontend_evidence_audit_reasons",
    "inventory_constraint",
    "inventory_reason",
    "ad_action_count",
    "ad_action_display_limit",
    "ad_action_more_count",
    "ad_action_items",
    "ad_diagnostic_summary",
    "cost_status",
    "cost_key_evidence",
    "ad_clicks",
    "ad_spend",
    "ad_orders",
    "total_orders",
    "natural_orders",
    "acos",
    "target_acos",
    "tacos",
    "ad_cvr",
    "ad_order_share",
    "recent_7d_clicks",
    "recent_7d_orders",
    "recent_7d_total_orders",
    "recent_7d_natural_orders",
]
TOP_LEVEL_SNAPSHOT_UNION_KEYS = {
    "product_final_decision_rows": (list(PRODUCT_IDENTITY_FIELDS), "product final decisions"),
    "product_operation_cards": (PRODUCT_OPERATION_CARD_IDENTITY_FIELDS, "product operation cards"),
    "inventory_replenishment_rows": (list(PRODUCT_IDENTITY_FIELDS), "inventory replenishment rows"),
}
DAILY_EXCEL_CONSISTENCY_SPECS = {
    "product_final_decision_rows": {
        "sheet_name": "产品最终决策",
        "identity_fields": list(PRODUCT_IDENTITY_FIELDS),
        "label": "product final decisions",
        "content_fields": [
            "product_name",
            "final_decision",
            "final_decision_label",
            "decision_priority",
            "decision_reason",
            "today_allowed_actions",
            "today_blocked_actions",
            "frontend_required",
            "frontend_posture",
            "frontend_evidence_state",
            "frontend_blocking_reasons",
            *FRONTEND_EVIDENCE_DECISION_FIELDS,
            "frontend_check_status",
            "frontend_cache_used",
            "frontend_failure_category",
            "frontend_price_currency_warning",
            "frontend_location_warning",
            "frontend_search_status",
            "frontend_search_partial_evidence",
            "frontend_auto_conclusion",
            "frontend_auto_conclusion_label",
            "frontend_evidence_quality_score",
            "frontend_evidence_audit_summary",
            "frontend_evidence_audit_detail",
            "competitor_comparability",
            "comparable_competitor_count",
            "competitor_mismatch_reason",
            "competitor_frontend_status",
            "competitor_frontend_asins",
            "competitor_frontend_count",
            "competitor_sellersprite_status",
            "competitor_sellersprite_asin_count",
            "competitor_sellersprite_keyword_count",
            "competitor_keyword_pressure",
            "competitor_shared_keywords",
            "own_missing_competitor_keywords",
            "own_ad_terms_not_in_sellersprite",
            *MARKET_SURVEY_COMPLETENESS_FIELDS,
            *HISTORY_TREND_FIELDS,
            "frontend_location_scope",
            "frontend_location_verified",
            "frontend_location_exact",
            "frontend_location_uncertain",
            "frontend_location_block_reason",
            "inventory_constraint",
            "feedback_cooldown_status",
            "keyword_memory_summary",
            "ad_action_summary",
            "next_review_date",
            "evidence_used",
            "confidence",
            "coupon",
            "fusion_action_gate",
            "fusion_issue_type",
            "fusion_confidence",
            "fusion_reason",
            "fusion_today_action",
            "fusion_do_not_do",
            "fusion_review_window",
            "fusion_evidence_flags",
            "fusion_missing_evidence",
            "last_updated",
        ],
    },
    "inventory_replenishment_rows": {
        "sheet_name": "库存补货提醒",
        "identity_fields": list(PRODUCT_IDENTITY_FIELDS),
        "label": "inventory replenishment rows",
        "content_fields": [
            "product_name",
            "current_inventory",
            "available_stock",
            "avg_daily_units_used",
            "avg_daily_units_source",
            "days_of_cover",
            "recommended_reorder_qty",
            "reference_reorder_qty",
            "reorder_deadline",
            "stock_risk_level",
            "stock_status_label",
            "stock_risk_reason",
            "replenishment_advice",
            "inventory_match_status",
            "inventory_source",
            "inventory_date",
            "recent_14d_clicks",
            "recent_14d_orders",
            "recent_14d_ad_orders",
        ],
    },
}
TOP_LEVEL_SNAPSHOT_CONTENT_FIELDS = {
    "product_final_decision_rows": DAILY_EXCEL_CONSISTENCY_SPECS["product_final_decision_rows"]["content_fields"],
    "product_operation_cards": PRODUCT_OPERATION_CARD_CONTENT_FIELDS,
    "inventory_replenishment_rows": DAILY_EXCEL_CONSISTENCY_SPECS["inventory_replenishment_rows"]["content_fields"],
}
PRODUCT_OPERATION_CARD_REQUIRED_FIELDS = [
    "marketplace",
    "sku",
    "asin",
    "product_name",
    "final_decision",
    "final_decision_label",
    "decision_reason",
    "operation_main_reason",
    "fusion_issue_type",
    "fusion_action_gate",
    "fusion_do_not_do",
    "fusion_review_window",
    "today_allowed_actions",
    "today_blocked_actions",
    "frontend_status",
    *FRONTEND_EVIDENCE_DECISION_FIELDS,
    "frontend_cache_used",
    "frontend_evidence_audit_summary",
    "frontend_evidence_audit_reasons",
    "inventory_constraint",
    "ad_action_count",
    "ad_action_display_limit",
    "ad_action_more_count",
    "ad_action_items",
    "ad_diagnostic_summary",
]
ENHANCED_REQUEST_IDENTITY_FIELDS = [
    "marketplace",
    "report_type",
    "period",
    "expected_filename",
    "target_path",
    "trigger_sku",
    "trigger_asin",
]
ENHANCED_REQUEST_CONTENT_FIELDS = [
    "trigger_product_name",
    "issue_type",
    "start_date",
    "end_date",
    "target_folder",
    "required",
    "seller_central_page",
    "instruction",
    "status",
    "file_type",
    "format_type",
    "detected_from",
    "detected_date_range",
    "freshness",
    "used_in_diagnosis",
]
ENHANCED_REQUEST_MARKDOWN_HEADERS = ["站点", "状态", "报表类型", "周期", "日期范围", "导出后文件名", "目标文件夹", "必需"]
ENHANCED_REQUEST_MARKDOWN_IDENTITY_FIELDS = ["站点", "报表类型", "周期", "导出后文件名"]
ENHANCED_REQUEST_MARKDOWN_CONTENT_FIELDS = ["状态", "日期范围", "目标文件夹", "必需"]
AUTOOPT_EXCEL_ROW_COUNT_SHEETS = {
    "autoopt_log": "rows",
    "action_review": "action_review_rows",
    "keyword_action_review": "keyword_action_review_rows",
    "final_decisions": "product_final_decisions",
}
ACTION_REVIEW_IDENTITY_FIELDS = [
    "action_id",
    "marketplace",
    "sku",
    "asin",
    "action_scope",
    "normalized_action",
    "executed_at",
    "report_date",
]
KEYWORD_REVIEW_IDENTITY_FIELDS = [
    *ACTION_REVIEW_IDENTITY_FIELDS,
    "search_term_or_target",
]
TODAY_TASK_QUEUE_IDENTITY_FIELDS = [
    "marketplace",
    "sku",
    "asin",
    "priority",
    "issue_type",
    "action_group",
    "today_action",
    "search_term_or_target",
    "suggested_action",
    "normalized_action",
    "action_id",
]
TODAY_TASK_QUEUE_CONTENT_FIELDS = [
    "product_name",
    "confirmed_status",
    "primary_reason",
    "key_evidence",
    "today_action",
    "action_group",
    "tomorrow_check",
    "source_section",
    "search_term_or_target",
    "suggested_action",
    "copy_action_line",
    "copy_block",
    "normalized_action",
    "action_scope",
    "action_id",
    "final_decision",
    "final_decision_label",
    "today_allowed_actions",
    "today_blocked_actions",
    "fusion_action_gate",
    "fusion_today_action",
    "fusion_do_not_do",
    "fusion_review_window",
    "review_status",
    "why_still_active",
    "downgrade_condition",
]
TOMORROW_REVIEW_IDENTITY_FIELDS = [
    "marketplace",
    "sku",
    "asin",
    "review_reason",
    "tomorrow_check",
    "trigger_action",
    "search_term_or_target",
    "normalized_action",
    "action_id",
]
TOMORROW_REVIEW_CONTENT_FIELDS = [
    "product_name",
    "review_reason",
    "current_evidence",
    "tomorrow_check",
    "trigger_action",
    "confirmed_status",
    "priority",
    "issue_type",
    "today_action",
    "search_term_or_target",
    "suggested_action",
    "normalized_action",
    "action_id",
    "review_status",
    "why_still_active",
    "downgrade_condition",
    "next_review_date",
    "cooldown_until",
]
DIAGNOSIS_ROW_IDENTITY_FIELDS = [
    "marketplace",
    "SKU",
    "ASIN",
    "诊断类型",
    "主因",
    "search_term_or_target",
    "suggested_action",
    "normalized_action",
    "action_id",
]
DIAGNOSIS_ROW_CONTENT_FIELDS = [
    "产品",
    "SKU",
    "ASIN",
    "诊断类型",
    "主因",
    "关键证据",
    "建议动作",
    "confirmed_status",
    "priority",
    "source_section",
    "marketplace",
    "异常信号",
    "初步方向",
    "需要人工确认",
    "发给 ChatGPT 的材料",
    "产品专属下一步",
    "confirmed_note",
    "search_term_or_target",
    "suggested_action",
    "normalized_action",
    "action_scope",
    "action_id",
    "confirmed_at",
    "report_date",
]
FRONTEND_QUEUE_IDENTITY_FIELDS = list(PRODUCT_IDENTITY_FIELDS)
FRONTEND_QUEUE_CONTENT_FIELDS = [
    "product_name",
    "product_url",
    "frontend_core_keyword",
    "frontend_search_url",
    "trigger_reason",
    "key_metrics",
    "frontend_check_status",
    "frontend_check_focus",
    "frontend_data_freshness",
    "frontend_cache_used",
    "frontend_data_date",
    "frontend_refresh_action",
    "frontend_check_method",
    "frontend_findings",
    "suspected_issue",
    "questions_to_check",
    "conservative_action",
    "recommended_next_step",
    "frontend_failure_category",
    "frontend_failure_reason",
    *FRONTEND_EVIDENCE_DECISION_FIELDS,
    "frontend_evidence_quality_score",
    "frontend_evidence_quality_grade",
    "frontend_price_currency_warning",
    "frontend_location_scope",
    "frontend_location_verified",
    "frontend_location_exact",
    "frontend_location_uncertain",
    "frontend_location_warning",
    "frontend_location_block_reason",
    "frontend_search_status",
    "frontend_search_keyword",
    "frontend_search_findings",
    "frontend_search_partial_evidence",
    "frontend_auto_conclusion",
    "frontend_auto_conclusion_label",
    "frontend_auto_conclusion_reasons",
    "competitor_comparability",
    "frontend_competitor_count",
    "comparable_competitor_count",
    "competitor_mismatch_reason",
    "frontend_evidence_audit_summary",
    "frontend_evidence_audit_reasons",
    "frontend_evidence_audit_detail",
    *MARKET_SURVEY_COMPLETENESS_FIELDS,
]
AD_PROCESSING_QUEUE_IDENTITY_FIELDS = [
    "marketplace",
    "sku",
    "asin",
    "search_term_or_target",
    "campaign_name",
    "ad_group_name",
    "match_type_or_targeting",
    "suggested_action",
]
AD_PROCESSING_QUEUE_CONTENT_FIELDS = [
    "product_name",
    "campaign",
    "ad_group",
    "match_type",
    "matched_target",
    "targeting",
    "match_type_or_targeting",
    "clicks",
    "spend",
    "orders",
    "sales",
    "relevance_level",
    "manual_level",
    "keyword_level",
    "matched_keyword",
    "classification_reason",
    "copy_action_line",
    "copy_block",
    "reason",
    "html_visible",
    "confirmed_status",
    "manual_action_taken",
    "normalized_action",
    "action_scope",
    "action_id",
    "ad_memory_blocked",
    "blocked_action_id",
    "blocked_original_action",
    "keyword_memory_summary",
    "scale_action",
    "final_decision",
    "final_decision_label",
    "final_decision_reason",
    "today_allowed_actions",
    "today_blocked_actions",
]
SCALE_CANDIDATE_IDENTITY_FIELDS = ["站点", "SKU", "ASIN"]
SCALE_CANDIDATE_CONTENT_FIELDS = [
    "产品",
    "点击",
    "花费",
    "订单",
    "总单",
    "销售额",
    "ACOS",
    "目标 ACOS",
    "放量等级",
    "建议",
]
SCALE_KEYWORD_IDENTITY_FIELDS = [
    "marketplace",
    "sku",
    "asin",
    "search_term_or_target",
    "campaign_name",
    "ad_group_name",
    "match_type_or_targeting",
    "scale_action",
]
SCALE_KEYWORD_CONTENT_FIELDS = [
    "product_name",
    "campaign",
    "ad_group",
    "match_type",
    "matched_target",
    "targeting",
    "match_type_or_targeting",
    "clicks",
    "spend",
    "ad_orders",
    "ad_sales",
    "ACOS",
    "CVR",
    "target_acos",
    "suggested_action",
    "copy_action_line",
    "copy_block",
    "reason",
    "product_scale_level",
    "confirmed_status",
    "manual_action_taken",
    "normalized_action",
    "action_scope",
    "action_id",
    "ad_memory_blocked",
    "blocked_action_id",
    "blocked_original_action",
    "keyword_memory_summary",
    "final_decision",
    "final_decision_label",
    "final_decision_reason",
    "today_allowed_actions",
    "today_blocked_actions",
]
GROWTH_TEST_IDENTITY_FIELDS = ["marketplace", "sku", "asin", "search_term_or_target", "action_id"]
GROWTH_TEST_CONTENT_FIELDS = [
    "product_name",
    "suggested_action",
    "manual_action_taken",
    "normalized_action",
    "action_scope",
    "experiment_type",
    "term_source",
    "evidence_level",
    "traffic_origin",
    "operation_label",
    "campaign_name",
    "ad_group_name",
    "match_type",
    "matched_target",
    "targeting",
    "test_days",
    "report_date",
    "next_review",
    "cooldown_days",
    "clicks",
    "spend",
    "orders",
    "suggested_daily_budget",
    "suggested_bid_min",
    "suggested_bid_max",
    "stop_loss_rule",
    "success_rule",
    "reason",
    "confirmed_status",
    "html_visible",
]
DAILY_EXCEL_REVIEW_CONSISTENCY_SPECS = {
    "frontend_check_queue_rows": {
        "sheet_name": "前台证据队列",
        "market_sheet_suffix": "前台证据队列",
        "identity_fields": FRONTEND_QUEUE_IDENTITY_FIELDS,
        "label": "frontend check queue rows",
        "content_fields": FRONTEND_QUEUE_CONTENT_FIELDS,
    },
    "product_operation_cards": {
        "sheet_name": "产品运营卡",
        "market_sheet_suffix": "产品运营卡",
        "identity_fields": PRODUCT_OPERATION_CARD_IDENTITY_FIELDS,
        "label": "product operation cards",
        "content_fields": PRODUCT_OPERATION_CARD_CONTENT_FIELDS,
    },
    "search_term_processing_queue_rows": {
        "sheet_name": "广告处理队列",
        "market_sheet_suffix": "广告处理队列",
        "identity_fields": AD_PROCESSING_QUEUE_IDENTITY_FIELDS,
        "label": "ad processing queue rows",
        "content_fields": AD_PROCESSING_QUEUE_CONTENT_FIELDS,
    },
    "scale_rows": {
        "sheet_name": "放量候选",
        "market_sheet_suffix": "放量候选",
        "identity_fields": SCALE_CANDIDATE_IDENTITY_FIELDS,
        "label": "scale candidate rows",
        "content_fields": SCALE_CANDIDATE_CONTENT_FIELDS,
    },
    "scale_keyword_rows": {
        "sheet_name": "放量词候选",
        "market_sheet_suffix": "放量词候选",
        "identity_fields": SCALE_KEYWORD_IDENTITY_FIELDS,
        "label": "scale keyword rows",
        "content_fields": SCALE_KEYWORD_CONTENT_FIELDS,
    },
    "growth_test_rows": {
        "sheet_name": "小预算试投",
        "market_sheet_suffix": "小预算试投",
        "identity_fields": GROWTH_TEST_IDENTITY_FIELDS,
        "label": "growth test rows",
        "content_fields": GROWTH_TEST_CONTENT_FIELDS,
    },
    "today_task_queue_rows": {
        "sheet_name": "今日动作清单",
        "market_sheet_suffix": "今日动作清单",
        "identity_fields": TODAY_TASK_QUEUE_IDENTITY_FIELDS,
        "label": "today task queue rows",
        "content_fields": TODAY_TASK_QUEUE_CONTENT_FIELDS,
    },
    "tomorrow_review_rows": {
        "sheet_name": "明日复查清单",
        "market_sheet_suffix": "明日复查清单",
        "identity_fields": TOMORROW_REVIEW_IDENTITY_FIELDS,
        "label": "tomorrow review rows",
        "content_fields": TOMORROW_REVIEW_CONTENT_FIELDS,
    },
    "listing_price_diagnosis_rows": {
        "sheet_name": "Listing待确认",
        "market_sheet_suffix": "Listing待确认",
        "identity_fields": DIAGNOSIS_ROW_IDENTITY_FIELDS,
        "label": "listing price diagnosis rows",
        "content_fields": DIAGNOSIS_ROW_CONTENT_FIELDS,
    },
    "cost_profit_diagnosis_rows": {
        "sheet_name": "成本利润诊断",
        "market_sheet_suffix": "成本利润诊断",
        "identity_fields": DIAGNOSIS_ROW_IDENTITY_FIELDS,
        "label": "cost profit diagnosis rows",
        "content_fields": DIAGNOSIS_ROW_CONTENT_FIELDS,
    },
    "action_effect_review_rows": {
        "sheet_name": "执行后效果复盘",
        "market_sheet_suffix": "执行后复盘",
        "identity_fields": ACTION_REVIEW_IDENTITY_FIELDS,
        "label": "action reviews",
        "content_fields": ACTION_REVIEW_REQUIRED_FIELDS,
    },
    "keyword_action_effect_review_rows": {
        "sheet_name": "词级执行复盘",
        "market_sheet_suffix": "词级执行复盘",
        "identity_fields": KEYWORD_REVIEW_IDENTITY_FIELDS,
        "label": "keyword action reviews",
        "content_fields": KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS,
    },
}
DAILY_EXCEL_REQUIRED_EMPTY_SHEETS = [
    "产品最终决策",
    "产品运营卡",
    "前台证据队列",
    "今日动作清单",
    "明日复查清单",
]
DAILY_EXCEL_REQUIRED_MARKET_EMPTY_SHEETS = [
    "UK_前台证据队列",
    "US_前台证据队列",
    "DE_前台证据队列",
]
AUTOOPT_EXCEL_CONSISTENCY_SPECS = {
    "action_review": {
        "payload_key": "action_review_rows",
        "identity_fields": ACTION_REVIEW_IDENTITY_FIELDS,
        "label": "action reviews",
        "content_fields": ACTION_REVIEW_REQUIRED_FIELDS,
    },
    "keyword_action_review": {
        "payload_key": "keyword_action_review_rows",
        "identity_fields": KEYWORD_REVIEW_IDENTITY_FIELDS,
        "label": "keyword action reviews",
        "content_fields": KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS,
    },
    "final_decisions": {
        "payload_key": "product_final_decisions",
        "identity_fields": list(PRODUCT_IDENTITY_FIELDS),
        "label": "product final decisions",
        "content_fields": DAILY_EXCEL_CONSISTENCY_SPECS["product_final_decision_rows"]["content_fields"],
    },
}
STANDALONE_REVIEW_JSON_REQUIRED_FIELDS = {
    "action_review": ACTION_REVIEW_REQUIRED_FIELDS,
    "keyword_action_review": KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS,
}
STANDALONE_REVIEW_JSON_IDENTITY_FIELDS = {
    "action_review": ACTION_REVIEW_IDENTITY_FIELDS,
    "keyword_action_review": KEYWORD_REVIEW_IDENTITY_FIELDS,
}
AUTOOPT_LATEST_ANALYSIS_CONSISTENCY_SPECS = {
    "product_final_decisions": {
        "expected_payload_key": "product_final_decision_rows",
        "identity_fields": list(PRODUCT_IDENTITY_FIELDS),
        "label": "product final decisions",
        "content_fields": DAILY_EXCEL_CONSISTENCY_SPECS["product_final_decision_rows"]["content_fields"],
    },
}
AUTOOPT_LATEST_ANALYSIS_SNAPSHOT_CONSISTENCY_SPECS = {
    "action_review_rows": {
        "snapshot_key": "action_effect_review_rows",
        "identity_fields": ACTION_REVIEW_IDENTITY_FIELDS,
        "label": "action reviews",
        "content_fields": ACTION_REVIEW_REQUIRED_FIELDS,
    },
    "keyword_action_review_rows": {
        "snapshot_key": "keyword_action_effect_review_rows",
        "identity_fields": KEYWORD_REVIEW_IDENTITY_FIELDS,
        "label": "keyword action reviews",
        "content_fields": KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS,
    },
}
AUTOOPT_LATEST_ANALYSIS_SUMMARY_KEYS = ["final_decision_summary", "decision_gate_counts"]


def _marketplace_sort_key(value: object) -> tuple[int, str]:
    marketplace = str(value or "N/A").upper()
    return (MARKETPLACE_DISPLAY_ORDER.get(marketplace, 99), marketplace)


def _truthy(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "是", "已验证"}


def _number(value: object) -> float | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if text.endswith("%"):
        try:
            number = float(text[:-1].strip()) / 100
        except ValueError:
            return None
        return None if math.isnan(number) else number
    try:
        number = float(text)
    except ValueError:
        return None
    return None if math.isnan(number) else number


def _ratio_number(value: object) -> float | None:
    number = _number(value)
    if number is None:
        return None
    return number / 100 if number > 1 else number


def _positive_score(value: object) -> bool:
    number = _number(value)
    return bool(number is not None and number > 0)


def _action_set(value: object) -> set[str]:
    if isinstance(value, (list, tuple, set)):
        return {str(item).strip() for item in value if str(item).strip()}
    text = str(value or "").strip()
    if not text:
        return set()
    if text.startswith("[") and text.endswith("]"):
        for parser in (json.loads, ast.literal_eval):
            try:
                parsed = parser(text)
            except (ValueError, SyntaxError, TypeError, json.JSONDecodeError):
                continue
            if isinstance(parsed, (list, tuple, set)):
                return {str(item).strip() for item in parsed if str(item).strip()}
    for separator in ["；", ";", ",", "，", "|", "/"]:
        text = text.replace(separator, "\n")
    return {item.strip() for item in text.splitlines() if item.strip()}


def _growth_action_bucket(row: dict[str, object]) -> str:
    normalized = str(row.get("normalized_action") or "").strip()
    if normalized in GROWTH_ACTIONS:
        return normalized
    if normalized in {"growth_test", "create_exact", "create_exact_low_budget"}:
        return "create_exact_low_budget"
    text = " ".join(
        str(row.get(field) or "")
        for field in [
            "suggested_action",
            "scale_action",
            "copy_action_line",
            "copy_block",
            "today_action",
            "action_detail",
            "manual_action_taken",
        ]
    ).strip().lower()
    if not text:
        return ""
    negative_markers = [
        "不加价",
        "不提高竞价",
        "停止加价",
        "禁止加价",
        "不能加价",
        "不追加预算",
        "不加预算",
        "不提高预算",
        "停止追加",
        "禁止放量",
        "不放量",
        "不能放量",
        "不推大词放量",
    ]
    if any(marker in text for marker in negative_markers):
        return ""
    if any(
        marker in text
        for marker in [
            "小预算",
            "精准测试",
            "拉精准",
            "创建精准",
            "create exact",
            "exact test",
            "low budget",
        ]
    ):
        return "create_exact_low_budget"
    if (
        "budget up" in text
        or "increase budget" in text
        or "raise budget" in text
        or "加预算" in text
        or "追加预算" in text
        or "提高预算" in text
    ):
        return "budget_up"
    if "broad scale" in text or "放量" in text:
        return "broad_scale"
    bid_raise_markers = [
        "bid up",
        "increase bid",
        "raise bid",
        "加价",
        "提高竞价",
        "上调竞价",
        "调高竞价",
        "提高出价",
        "上调出价",
        "调高出价",
    ]
    price_or_coupon_context = any(marker in text for marker in ["优惠券", "coupon", "价格", "售价", "price"])
    if any(marker in text for marker in bid_raise_markers) or (
        not price_or_coupon_context and re.search(r"加\s*\d+(?:\.\d+)?\s*%", text)
    ):
        return "bid_up"
    return ""


def _identity_counter(rows: list[object], identity_fields: list[str]) -> Counter[tuple[str, ...]]:
    identities: list[tuple[str, ...]] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        identities.append(tuple(str(row.get(field) or "").strip() for field in identity_fields))
    return Counter(identities)


def _canonical_report_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if math.isnan(float(value)):
            return ""
        return round(float(value), 6)
    if isinstance(value, dict):
        return {str(key): _canonical_report_value(item) for key, item in sorted(value.items())}
    if isinstance(value, (list, tuple, set)):
        return [_canonical_report_value(item) for item in value]
    text = str(value).strip()
    if not text:
        return ""
    if text.lower() in {"true", "false"}:
        return text.lower() == "true"
    try:
        numeric = float(text)
        if math.isnan(numeric):
            return ""
        return round(numeric, 6)
    except ValueError:
        pass
    if text.startswith(("[", "{", "(")) and text.endswith(("]", "}", ")")):
        for parser in (json.loads, ast.literal_eval):
            try:
                parsed = parser(text)
            except (ValueError, SyntaxError, TypeError, json.JSONDecodeError):
                continue
            return _canonical_report_value(parsed)
    return text


def _rows_by_unique_identity(
    rows: list[object],
    identity_fields: list[str],
    *,
    label: str,
    source: str,
) -> tuple[dict[tuple[str, ...], dict], list[str]]:
    mapped: dict[tuple[str, ...], dict] = {}
    counts = _identity_counter(rows, identity_fields)
    failures = [
        f"{source} duplicate {label} identity: {identity}"
        for identity, count in sorted(counts.items())
        if count > 1
    ]
    for row in rows:
        if not isinstance(row, dict):
            continue
        identity = tuple(str(row.get(field) or "").strip() for field in identity_fields)
        if counts.get(identity) == 1:
            mapped[identity] = row
    return mapped, failures


def run_step(args: list[str]) -> int:
    print("[run]", " ".join(args), flush=True)
    try:
        completed = subprocess.run(args, cwd=ROOT)
    except OSError as exc:
        print(f"[fail] cannot start step: {exc}", flush=True)
        return 127
    print("[exit]", completed.returncode, flush=True)
    return completed.returncode


def _manifest_mtime_ns(path: Path) -> int | None:
    if not path.exists():
        return None
    return path.stat().st_mtime_ns


def _mtime_ns(path: Path) -> int | None:
    if not path.exists():
        return None
    return path.stat().st_mtime_ns


def _frontend_identity(row: dict, fallback_marketplace: object = "") -> tuple[str, str, str]:
    return (
        str(row.get("marketplace") or fallback_marketplace or "").strip().upper(),
        str(row.get("sku") or row.get("SKU") or "").strip(),
        str(row.get("asin") or row.get("ASIN") or "").strip().upper(),
    )


FRONTEND_GATE_CONSISTENCY_FIELDS = [
    "frontend_check_status",
    "frontend_cache_used",
    "frontend_evidence_tier",
    "frontend_evidence_display_tier",
    "frontend_decision_evidence_tier",
    "frontend_evidence_is_strong",
    "frontend_search_status",
    "frontend_search_partial_evidence",
    "frontend_price_currency_warning",
    "frontend_location_warning",
    "frontend_failure_category",
    "frontend_auto_conclusion",
    "frontend_evidence_quality_score",
    "frontend_evidence_audit_summary",
    "frontend_evidence_audit_reasons",
    "frontend_evidence_audit_detail",
    "competitor_comparability",
    "frontend_competitor_count",
    "comparable_competitor_count",
    "frontend_location_scope",
    "frontend_location_verified",
    "frontend_location_exact",
    "frontend_location_uncertain",
]
FRONTEND_GATE_NUMERIC_FIELDS = {
    "frontend_evidence_quality_score",
    "frontend_competitor_count",
    "comparable_competitor_count",
}


def _latest_analysis_frontend_rows(path: Path) -> tuple[dict[tuple[str, str, str], dict[str, object]], list[str]]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        return {}, [f"latest analysis cannot be read for frontend wrapper validation: {exc}"]
    if not isinstance(payload, dict):
        return {}, [f"latest analysis root must be an object for frontend wrapper validation: {path}"]
    rows: dict[tuple[str, str, str], dict[str, object]] = {}
    for result in payload.get("marketplace_results") or []:
        if not isinstance(result, dict):
            continue
        marketplace = str(result.get("marketplace") or "").strip().upper()
        snapshot = result.get("report_view_snapshot") or {}
        if not isinstance(snapshot, dict):
            continue
        for row in snapshot.get("frontend_check_queue_rows") or []:
            if not isinstance(row, dict):
                continue
            identity = _frontend_identity(row, marketplace)
            if all(identity):
                rows[identity] = row
    return rows, []


def _latest_analysis_frontend_identities(path: Path) -> tuple[set[tuple[str, str, str]], list[str]]:
    rows, failures = _latest_analysis_frontend_rows(path)
    return set(rows), failures


def _frontend_gate_value(value: object) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (list, tuple)):
        return "；".join(_frontend_gate_value(item) for item in value if _frontend_gate_value(item))
    if isinstance(value, set):
        return "；".join(sorted(_frontend_gate_value(item) for item in value if _frontend_gate_value(item)))
    text = str(value or "").strip()
    lower = text.lower()
    if lower in {"true", "false", "1", "0", "yes", "no", "y", "n", "是", "否", "已验证"}:
        return "true" if _truthy(value) else "false"
    number = _number(value)
    if number is not None:
        return f"{number:g}"
    return text


def _frontend_gate_field_value(field: str, value: object) -> str:
    if field in FRONTEND_GATE_NUMERIC_FIELDS:
        number = _number(value)
        if number is not None:
            return f"{number:g}"
    return _frontend_gate_value(value)


def _frontend_results_generated_at(path: Path) -> str:
    if not path.exists():
        return ""
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return ""
    if not isinstance(payload, dict):
        return ""
    return str(payload.get("generated_at") or "").strip()


def frontend_wrapper_refresh_failures(
    results_path: Path | None = None,
    analysis_path: Path | None = None,
    *,
    previous_mtime_ns: int | None = None,
    previous_generated_at: str = "",
) -> list[str]:
    results_path = results_path or FRONTEND_RESULTS_JSON
    analysis_path = analysis_path or LATEST_ANALYSIS_JSON
    expected_rows, analysis_failures = _latest_analysis_frontend_rows(analysis_path)
    if analysis_failures:
        return analysis_failures
    expected = set(expected_rows)
    if not expected:
        return []
    current_mtime = _mtime_ns(results_path)
    if current_mtime is None:
        return [f"frontend check results missing after frontend wrapper: {results_path}"]
    if previous_mtime_ns is not None and current_mtime == previous_mtime_ns:
        return [f"frontend check results were not refreshed by frontend wrapper: {results_path}"]
    try:
        payload = json.loads(results_path.read_text(encoding="utf-8"))
    except Exception as exc:
        return [f"frontend check results cannot be read after frontend wrapper: {exc}"]
    if not isinstance(payload, dict):
        return [f"frontend check results root must be an object after frontend wrapper: {results_path}"]
    generated_at = str(payload.get("generated_at") or "").strip()
    if not generated_at:
        return [f"frontend check results missing generated_at after frontend wrapper: {results_path}"]
    if previous_generated_at and generated_at == previous_generated_at:
        return [f"frontend check results generated_at was not refreshed by frontend wrapper: {results_path}"]
    items = payload.get("items")
    if not isinstance(items, list):
        return [f"frontend check results items must be a list after frontend wrapper: {results_path}"]
    actual_rows = {
        identity: row
        for row in items
        if isinstance(row, dict)
        for identity in [_frontend_identity(row)]
        if all(identity)
    }
    actual = set(actual_rows)
    missing = sorted(expected.difference(actual))
    if missing:
        labels = ["/".join(identity) for identity in missing[:5]]
        return ["frontend check results did not cover latest analysis frontend queue after wrapper: " + "; ".join(labels)]
    for identity in sorted(expected):
        expected_row = expected_rows.get(identity) or {}
        actual_row = actual_rows.get(identity) or {}
        for field in FRONTEND_GATE_CONSISTENCY_FIELDS:
            if field not in expected_row:
                continue
            if field not in actual_row:
                continue
            expected_value = _frontend_gate_field_value(field, expected_row.get(field))
            actual_value = _frontend_gate_field_value(field, actual_row.get(field))
            if expected_value != actual_value:
                label = "/".join(identity)
                return [
                    "frontend check results gate field mismatch after wrapper: "
                    f"{label} field {field} expected {expected_value}, got {actual_value}"
                ]
    return []


def output_refresh_failures(
    paths: list[Path] | None = None,
    *,
    previous_mtimes_ns: dict[Path, int | None] | None = None,
) -> list[str]:
    paths = paths or REQUIRED_REFRESHED_OUTPUTS
    previous_mtimes_ns = previous_mtimes_ns or {}
    failures: list[str] = []
    for path in paths:
        current_mtime = _mtime_ns(path)
        if current_mtime is None:
            failures.append(f"required output missing after report refresh: {path}")
            continue
        if path in previous_mtimes_ns and previous_mtimes_ns[path] == current_mtime:
            failures.append(f"required output was not refreshed by report step: {path}")
    return failures


def _asset_content_failures(paths: list[Path]) -> list[str]:
    failures: list[str] = []
    for path in paths:
        required_tokens = REPORT_ASSET_REQUIRED_TOKENS.get(path.name)
        if not required_tokens or not path.exists():
            continue
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
        except Exception as exc:
            failures.append(f"report asset {path.name} cannot be read: {exc}")
            continue
        if not text.strip():
            failures.append(f"report asset {path.name} is empty")
            continue
        for token in required_tokens:
            if token not in text:
                failures.append(f"report asset {path.name} missing required token {token}")
    return failures


def _analysis_path_from_required_outputs(paths: list[Path]) -> Path:
    for path in paths:
        if path.name == "latest_analysis.json":
            return path
    return LATEST_ANALYSIS_JSON


def _report_date_from_analysis(path: Path) -> tuple[str | None, list[str]]:
    if not path.exists():
        return None, [f"latest analysis missing after report refresh: {path}"]
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        return None, [f"latest analysis cannot be read: {exc}"]
    if not isinstance(payload, dict):
        return None, ["latest analysis root must be an object"]
    report_date = str(payload.get("report_date") or "").strip()
    if not report_date:
        return None, ["latest analysis missing report_date"]
    try:
        report_day = date.fromisoformat(report_date)
    except ValueError:
        return None, [f"latest analysis has invalid report_date: {report_date}"]
    today = date.today()
    if report_day > today:
        return None, [
            f"latest analysis report_date {report_date} is after current date {today.isoformat()}"
        ]
    return report_date, []


def _product_growth_frontend_failures(marketplace: str, rows: list[object]) -> list[str]:
    failures: list[str] = []
    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            continue
        allowed = _action_set(row.get("today_allowed_actions"))
        blocked_actions = _action_set(row.get("today_blocked_actions"))
        overlap = sorted(allowed & blocked_actions)
        growth_allowed = sorted(allowed & GROWTH_ACTIONS)
        asin = str(row.get("asin") or "unknown_asin")
        prefix = f"latest analysis {marketplace} product_final_decision_rows row {idx} {asin}"
        if overlap:
            failures.append(f"{prefix} has actions both allowed and blocked: {', '.join(overlap)}")
        state = str(row.get("frontend_evidence_state") or "").strip()
        tier = str(row.get("frontend_evidence_tier") or "").strip()
        display_tier = str(row.get("frontend_evidence_display_tier") or "").strip()
        decision_tier = str(row.get("frontend_decision_evidence_tier") or display_tier or tier).strip()
        check_status = str(row.get("frontend_check_status") or "").strip()
        search_status = str(row.get("frontend_search_status") or "").strip()
        partial_search = _truthy(row.get("frontend_search_partial_evidence")) or search_status == "已读取部分结果"
        currency_warning = str(row.get("frontend_price_currency_warning") or "").strip()
        location_warning = str(row.get("frontend_location_warning") or "").strip()
        failure = str(row.get("frontend_failure_category") or "").strip()
        auto_conclusion = str(row.get("frontend_auto_conclusion") or "").strip()
        location_scope = str(row.get("frontend_location_scope") or "").strip().lower()
        location_exact = _truthy(row.get("frontend_location_exact")) if "frontend_location_exact" in row else location_scope == "exact"
        location_verified = _truthy(row.get("frontend_location_verified"))
        competitor_count = _number(row.get("comparable_competitor_count"))
        comparability = str(row.get("competitor_comparability") or "").strip().lower()
        quality_score = _number(row.get("frontend_evidence_quality_score"))

        strong_marked = (
            state == "ok_high"
            or decision_tier == "强诊断可用"
            or _truthy(row.get("frontend_evidence_is_strong"))
        )
        if not growth_allowed and strong_marked:
            if display_tier != "强诊断可用":
                failures.append(f"{prefix} marks strong frontend evidence without strong frontend display tier")
            if decision_tier != "强诊断可用":
                failures.append(f"{prefix} marks strong frontend evidence without strong frontend decision tier")
            if not _truthy(row.get("frontend_evidence_is_strong")):
                failures.append(f"{prefix} marks strong frontend evidence without explicit strong flag")
            if check_status != "已自动检查" or _truthy(row.get("frontend_cache_used")):
                failures.append(f"{prefix} marks strong frontend evidence with cached frontend evidence")
            if partial_search:
                failures.append(f"{prefix} marks strong frontend evidence with partial search evidence")
            if search_status != "已自动检查":
                failures.append(f"{prefix} marks strong frontend evidence without successful search page")
            if currency_warning:
                failures.append(f"{prefix} marks strong frontend evidence with currency warning")
            if location_warning:
                failures.append(f"{prefix} marks strong frontend evidence with location warning")
            if failure and failure != "none":
                failures.append(f"{prefix} marks strong frontend evidence with frontend failure")
            if auto_conclusion != "FRONTEND_OK":
                failures.append(f"{prefix} marks strong frontend evidence without FRONTEND_OK conclusion")
            if not location_exact:
                failures.append(f"{prefix} marks strong frontend evidence without exact frontend location")
            if not location_verified:
                failures.append(f"{prefix} marks strong frontend evidence without verified frontend location")
            if not comparability:
                failures.append(f"{prefix} marks strong frontend evidence without competitor comparability")
            if competitor_count is None or competitor_count < 2 or comparability != "high":
                failures.append(f"{prefix} marks strong frontend evidence with weak competitor evidence")
            if quality_score is None or quality_score < 75:
                failures.append(f"{prefix} marks strong frontend evidence without high frontend quality score")
            continue
        if (
            not growth_allowed
            and (state != "ok_high" or decision_tier != "强诊断可用" or auto_conclusion != "FRONTEND_OK")
            and not GROWTH_ACTIONS.issubset(blocked_actions)
        ):
            missing = sorted(GROWTH_ACTIONS.difference(blocked_actions))
            failures.append(f"{prefix} does not explicitly block growth actions under weak frontend evidence: {missing}")
        if not growth_allowed:
            continue

        if state != "ok_high":
            failures.append(f"{prefix} allows growth without ok_high frontend evidence")
        if display_tier != "强诊断可用":
            failures.append(f"{prefix} allows growth without strong frontend display tier")
        if decision_tier != "强诊断可用":
            failures.append(f"{prefix} allows growth without strong frontend decision tier")
        if not _truthy(row.get("frontend_evidence_is_strong")):
            failures.append(f"{prefix} allows growth without explicit strong frontend flag")
        if check_status != "已自动检查":
            failures.append(f"{prefix} allows growth without current frontend check")
        if partial_search:
            failures.append(f"{prefix} allows growth with partial search evidence")
        if search_status != "已自动检查":
            failures.append(f"{prefix} allows growth without successful search page")
        if _truthy(row.get("frontend_cache_used")):
            failures.append(f"{prefix} allows growth with cached frontend evidence")
        if currency_warning:
            failures.append(f"{prefix} allows growth with currency warning")
        if location_warning:
            failures.append(f"{prefix} allows growth with location warning")
        if failure and failure != "none":
            failures.append(f"{prefix} allows growth with frontend failure")
        if auto_conclusion != "FRONTEND_OK":
            failures.append(f"{prefix} allows growth without FRONTEND_OK conclusion")
        if not location_exact:
            failures.append(f"{prefix} allows growth without exact frontend location")
        if not location_verified:
            failures.append(f"{prefix} allows growth without verified frontend location")
        if not comparability:
            failures.append(f"{prefix} allows growth without competitor comparability")
        if competitor_count is None or competitor_count < 2 or comparability != "high":
            failures.append(f"{prefix} allows growth with weak competitor evidence")
        if quality_score is None or quality_score < 75:
            failures.append(f"{prefix} allows growth without high frontend quality score")
    return failures


def _row_market_sku_asin(row: dict[str, object], marketplace: str) -> tuple[str, str, str]:
    return (
        str(row.get("marketplace") or row.get("站点") or marketplace).strip().upper(),
        str(row.get("sku") or row.get("SKU") or "").strip(),
        str(row.get("asin") or row.get("ASIN") or "").strip().upper(),
    )


def _frontend_weak_evidence_reason(row: dict[str, object] | None) -> str:
    if row is None:
        return "missing matching frontend evidence"
    reasons: list[str] = []
    display_tier = str(row.get("frontend_evidence_display_tier") or row.get("frontend_evidence_tier") or "").strip()
    status = str(row.get("frontend_check_status") or "").strip()
    search_status = str(row.get("frontend_search_status") or "").strip()
    failure = str(row.get("frontend_failure_category") or "").strip()
    comparability = str(row.get("competitor_comparability") or "").strip().lower()
    competitor_count = _number(row.get("frontend_competitor_count"))
    comparable_competitor_count = _number(row.get("comparable_competitor_count"))
    if competitor_count is None:
        competitor_count = comparable_competitor_count
    quality_score = _number(row.get("frontend_evidence_quality_score"))
    if display_tier != "强诊断可用" or not _truthy(row.get("frontend_evidence_is_strong")):
        reasons.append("frontend evidence is not strong")
    if status != "已自动检查" or _truthy(row.get("frontend_cache_used")):
        reasons.append("frontend check is stale or cached")
    if search_status != "已自动检查":
        reasons.append("search page evidence is not current")
    if _truthy(row.get("frontend_search_partial_evidence")) or search_status == "已读取部分结果":
        reasons.append("search page evidence is partial")
    if str(row.get("frontend_price_currency_warning") or "").strip():
        reasons.append("currency warning")
    if str(row.get("frontend_location_warning") or "").strip() or _truthy(row.get("frontend_location_uncertain")):
        reasons.append("location warning")
    if failure and failure != "none":
        reasons.append("frontend failure")
    if str(row.get("frontend_auto_conclusion") or "").strip() != "FRONTEND_OK":
        reasons.append("frontend conclusion is not FRONTEND_OK")
    if not _truthy(row.get("frontend_location_verified")):
        reasons.append("location is not verified")
    location_scope = str(row.get("frontend_location_scope") or "").strip().lower()
    if not (_truthy(row.get("frontend_location_exact")) or location_scope == "exact"):
        reasons.append("location is not exact")
    if not comparability:
        reasons.append("competitor comparability is missing")
    elif comparability in {"low", "unknown"}:
        reasons.append("competitor comparability is weak")
    if competitor_count is None or competitor_count < 2:
        reasons.append("comparable competitor count is below 2")
    elif comparable_competitor_count is not None and comparable_competitor_count < 2:
        reasons.append("comparable competitor count is below 2")
    if quality_score is None or quality_score < 75:
        reasons.append("frontend quality score is below 75")
    return "; ".join(reasons)


def _listing_risky_frontend_claim(row: dict[str, object]) -> str:
    fields = [
        "主因",
        "关键证据",
        "建议动作",
        "异常信号",
        "初步方向",
        "需要人工确认",
        "产品专属下一步",
    ]
    text = " ".join(str(row.get(field) or "") for field in fields)
    compact = re.sub(r"\s+", "", text)
    risky_phrases = [
        "强诊断可用",
        "确认Listing问题",
        "确定Listing问题",
        "Listing崩",
        "立即改Listing",
        "必须改Listing",
        "重做Listing",
    ]
    for phrase in risky_phrases:
        if phrase in compact:
            return phrase
    safe_budget_phrases = [
        "不加预算",
        "不急着加预算",
        "暂时不加广告预算",
        "不扩大预算",
        "不放量",
    ]
    budget_phrases = ["加预算", "提高预算", "扩大预算", "放量", "大幅加价", "大幅提价"]
    for phrase in budget_phrases:
        if phrase in compact and not any(safe in compact for safe in safe_budget_phrases):
            return phrase
    safe_price_phrases = [
        "先确认",
        "待确认",
        "人工确认",
        "先核查",
        "先对比",
        "再决定",
        "确认差异后",
        "确认异常后",
        "暂不调价",
        "不急着调价",
    ]
    price_action_phrases = [
        "立即降价",
        "直接降价",
        "马上降价",
        "立刻降价",
        "降到竞品价",
        "跟随竞品降价",
        "调低售价",
        "降低售价",
        "改价到",
        "调价到",
        "立即调价",
        "直接调价",
        "马上调价",
        "立刻调价",
        "加Coupon",
        "提高Coupon",
        "加优惠券",
        "提高优惠券",
    ]
    for phrase in price_action_phrases:
        if phrase in compact and not any(safe in compact for safe in safe_price_phrases):
            return phrase
    return ""


def _listing_frontend_evidence_failures(
    marketplace: str,
    listing_rows: list[object],
    frontend_rows: list[object],
) -> list[str]:
    frontend_by_key: dict[tuple[str, str, str], dict[str, object]] = {}
    frontend_by_asin: dict[tuple[str, str], dict[str, object]] = {}
    for row in frontend_rows:
        if not isinstance(row, dict):
            continue
        key = _row_market_sku_asin(row, marketplace)
        if key[0] and key[2]:
            frontend_by_key[key] = row
            frontend_by_asin[(key[0], key[2])] = row

    failures: list[str] = []
    for idx, row in enumerate(listing_rows, start=1):
        if not isinstance(row, dict):
            continue
        confirmed_status = str(row.get("confirmed_status") or "").strip()
        if confirmed_status in {"已执行", "已核查", "已忽略", "仅背景参考"}:
            continue
        risky_claim = _listing_risky_frontend_claim(row)
        if not risky_claim:
            continue
        key = _row_market_sku_asin(row, marketplace)
        frontend_row = frontend_by_key.get(key) or frontend_by_asin.get((key[0], key[2]))
        weak_reason = _frontend_weak_evidence_reason(frontend_row)
        if weak_reason:
            asin = key[2] or "unknown_asin"
            failures.append(
                f"latest analysis {marketplace} listing_price_diagnosis_rows row {idx} {asin} "
                f"contains frontend-backed strong listing or growth claim {risky_claim} "
                f"under weak frontend evidence: {weak_reason}"
            )
    return failures


def _task_queue_growth_gate_failures(
    marketplace: str,
    task_rows: list[object],
    product_decision_rows: list[object],
    *,
    row_source: str = "today_task_queue_rows",
) -> list[str]:
    decisions: dict[tuple[str, str, str], dict[str, object]] = {}
    for row in product_decision_rows:
        if not isinstance(row, dict):
            continue
        key = (
            str(row.get("marketplace") or marketplace).strip().upper(),
            str(row.get("sku") or "").strip(),
            str(row.get("asin") or "").strip().upper(),
        )
        if key[0] and key[2]:
            decisions[key] = row

    failures: list[str] = []
    for idx, row in enumerate(task_rows, start=1):
        if not isinstance(row, dict):
            continue
        confirmed_status = str(row.get("confirmed_status") or "").strip()
        if confirmed_status in {"已执行", "已核查", "已忽略", "仅背景参考"}:
            continue
        bucket = _growth_action_bucket(row)
        if bucket not in PRODUCT_GATED_AD_ACTIONS:
            continue
        key = (
            str(row.get("marketplace") or marketplace).strip().upper(),
            str(row.get("sku") or "").strip(),
            str(row.get("asin") or "").strip().upper(),
        )
        asin = key[2] or "unknown_asin"
        prefix = f"latest analysis {marketplace} {row_source} row {idx} {asin}"
        decision = decisions.get(key) or decisions.get((key[0], "", key[2]))
        if not decision:
            failures.append(f"{prefix} contains growth action {bucket} without matching product final decision gate")
            continue
        allowed = _action_set(decision.get("today_allowed_actions"))
        blocked = _action_set(decision.get("today_blocked_actions"))
        if bucket in blocked:
            failures.append(f"{prefix} contains growth action {bucket} blocked by product final decision")
        if bucket not in allowed:
            failures.append(f"{prefix} contains growth action {bucket} not allowed by product final decision")
    return failures


def _aux_ad_workbench_rows_from_snapshot(snapshot: dict[str, object]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for key in ["html_search_term_processing_queue_rows", "scale_keyword_rows", "growth_test_rows"]:
        value = snapshot.get(key) or []
        if isinstance(value, list):
            rows.extend(row for row in value if isinstance(row, dict))
    return rows


def _action_identity_source_row(row: dict[str, object], marketplace: str) -> dict[str, object]:
    source = dict(row)
    source["marketplace"] = source.get("marketplace") or source.get("站点") or marketplace
    source["sku"] = source.get("sku") or source.get("SKU") or ""
    source["asin"] = source.get("asin") or source.get("ASIN") or ""
    source["search_term_or_target"] = (
        source.get("search_term_or_target")
        or source.get("search_term")
        or source.get("targeting")
        or source.get("target")
        or ""
    )
    return source


def _action_identity_failure_for_row(
    row: dict[str, object],
    prefix: str,
    marketplace: str,
    *,
    require_action_id: bool = False,
) -> list[str]:
    action_id = str(row.get("action_id") or "").strip()
    if not action_id:
        if require_action_id:
            return [f"{prefix} missing action_id"]
        return []
    normalized_action = str(row.get("normalized_action") or "").strip()
    action_scope = str(row.get("action_scope") or "").strip()
    if not normalized_action:
        return [f"{prefix} has action_id but missing normalized_action"]
    if not action_scope:
        return [f"{prefix} has action_id but missing action_scope"]
    if action_scope not in VALID_ACTION_SCOPES:
        return [
            f"{prefix} has invalid action_scope {action_scope}; expected one of {', '.join(sorted(VALID_ACTION_SCOPES))}"
        ]
    expected_action_id = make_action_id(
        _action_identity_source_row(row, marketplace),
        normalized_action,
        action_scope,
    )
    if action_id != expected_action_id:
        return [f"{prefix} action_id mismatch: expected {expected_action_id}, got {action_id}"]
    return []


def _action_id_consistency_failures(
    marketplace: str,
    rows: list[object],
    *,
    row_source: str,
    require_action_id: bool = False,
) -> list[str]:
    failures: list[str] = []
    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            continue
        asin = str(row.get("asin") or row.get("ASIN") or "unknown_asin").strip().upper()
        prefix = f"latest analysis {marketplace} {row_source} row {idx} {asin}"
        failures.extend(
            _action_identity_failure_for_row(
                row,
                prefix,
                marketplace,
                require_action_id=require_action_id,
            )
        )
        if _truthy(row.get("ad_memory_blocked")):
            normalized = str(row.get("normalized_action") or "").strip()
            current_action_id = str(row.get("action_id") or "").strip()
            blocked_action_id = str(row.get("blocked_action_id") or "").strip()
            if not blocked_action_id:
                failures.append(f"{prefix} ad memory blocked row missing blocked_action_id")
            if normalized != "observe":
                failures.append(f"{prefix} ad memory blocked row must normalize current action to observe")
            if blocked_action_id and current_action_id and blocked_action_id == current_action_id:
                failures.append(f"{prefix} ad memory blocked row reuses blocked action_id as current action_id")
    return failures


def _today_action_group_for(row: dict[str, object]) -> str:
    group = str(row.get("action_group") or "").strip()
    if group:
        return group
    action = str(row.get("today_action") or "")
    issue_type = str(row.get("issue_type") or "")
    text = f"{action}；{issue_type}"
    if "成本 / 利润" in issue_type or any(
        token in text for token in ["广告前利润<=0", "广告前利润为负", "利润为负", "利润不允许"]
    ):
        return "成本 / 利润动作"
    if "近7天转化断崖" in text:
        return "广告动作"
    if any(token in text for token in ["Listing", "价格", "主图", "Coupon", "评价", "A+", "疑点"]):
        return "Listing / 价格动作"
    return "广告动作"


def _today_action_group_identity_counter(
    groups: dict[str, object],
    *,
    default_marketplace: str = "",
) -> Counter[tuple[str, str, str, str]]:
    identities: list[tuple[str, str, str, str]] = []
    for group, rows in groups.items():
        if not isinstance(rows, list):
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            identities.append(
                (
                    str(group),
                    str(row.get("marketplace") or default_marketplace).strip().upper(),
                    str(row.get("sku") or "").strip(),
                    str(row.get("asin") or "").strip().upper(),
                )
            )
    return Counter(identities)


def _today_action_group_consistency_failures(
    marketplace: str,
    task_rows: list[object],
    groups: dict[str, object],
) -> list[str]:
    failures: list[str] = []
    rebuilt: dict[str, list[dict[str, object]]] = {group: [] for group in ACTION_GROUPS}
    seen: set[tuple[str, str, str, str]] = set()
    for row in task_rows:
        if not isinstance(row, dict):
            continue
        group = _today_action_group_for(row)
        key = (
            group,
            str(row.get("marketplace") or marketplace).strip().upper(),
            str(row.get("sku") or "").strip(),
            str(row.get("asin") or "").strip().upper(),
        )
        if key in seen or len(rebuilt.get(group, [])) >= 3:
            continue
        seen.add(key)
        rebuilt.setdefault(group, []).append(row)
    for group, rows in groups.items():
        if not isinstance(rows, list):
            failures.append(f"latest analysis {marketplace} today_action_groups {group} must be a list")
    expected_counter = _today_action_group_identity_counter(rebuilt, default_marketplace=marketplace)
    actual_counter = _today_action_group_identity_counter(groups, default_marketplace=marketplace)
    if actual_counter != expected_counter:
        missing = sorted((expected_counter - actual_counter).elements())
        extra = sorted((actual_counter - expected_counter).elements())
        if missing:
            failures.append(f"latest analysis {marketplace} today_action_groups missing task rows: {missing[:5]}")
        if extra:
            failures.append(f"latest analysis {marketplace} today_action_groups contains stale task rows: {extra[:5]}")
    return failures


def _is_frontend_fallback_status(status: str, freshness: str, findings: str) -> bool:
    text = " ".join([status, freshness, findings])
    return "待前台检查" in text or "沿用" in text


def _frontend_queue_failures(marketplace: str, rows: list[object]) -> list[str]:
    failures: list[str] = []
    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            failures.append(f"latest analysis {marketplace} frontend_check_queue_rows row {idx} must be an object")
            continue
        status = str(row.get("frontend_check_status") or "").strip()
        freshness = str(row.get("frontend_data_freshness") or "").strip()
        findings = str(row.get("frontend_findings") or "").strip()
        asin = str(row.get("asin") or "unknown_asin")
        prefix = f"latest analysis {marketplace} frontend_check_queue_rows row {idx} {asin}"
        tier = str(row.get("frontend_evidence_tier") or "").strip()
        display_tier = str(row.get("frontend_evidence_display_tier") or "").strip()
        decision_tier = str(row.get("frontend_decision_evidence_tier") or display_tier or tier).strip()
        auto_conclusion = str(row.get("frontend_auto_conclusion") or "").strip()
        strong_marked = (
            decision_tier == "强诊断可用"
            or _truthy(row.get("frontend_evidence_is_strong"))
            or auto_conclusion == "FRONTEND_OK"
        )
        cached_or_pending = (
            status != "已自动检查"
            or status.startswith("沿用")
            or _truthy(row.get("frontend_cache_used"))
        )
        if cached_or_pending and strong_marked:
            failures.append(f"{prefix} cached or pending evidence marked strong")
        if strong_marked:
            warning = str(row.get("frontend_price_currency_warning") or "").strip()
            location_warning = str(row.get("frontend_location_warning") or "").strip()
            failure = str(row.get("frontend_failure_category") or "").strip()
            location_scope = str(row.get("frontend_location_scope") or "").strip().lower()
            location_exact_explicit = "frontend_location_exact" in row
            location_verified = _truthy(row.get("frontend_location_verified"))
            location_uncertain = _truthy(row.get("frontend_location_uncertain"))
            search_status = str(row.get("frontend_search_status") or "").strip()
            partial_search = (
                _truthy(row.get("frontend_search_partial_evidence"))
                or search_status == "已读取部分结果"
            )
            competitor_count = _number(
                row.get("frontend_competitor_count")
                if row.get("frontend_competitor_count") not in (None, "")
                else row.get("comparable_competitor_count")
            )
            comparable_competitor_count = _number(row.get("comparable_competitor_count"))
            comparability = str(row.get("competitor_comparability") or "").strip().lower()
            quality_score = _number(row.get("frontend_evidence_quality_score"))
            if decision_tier != "强诊断可用":
                failures.append(f"{prefix} strong evidence missing strong decision tier")
            if not _truthy(row.get("frontend_evidence_is_strong")):
                failures.append(f"{prefix} strong evidence missing explicit strong flag")
            if warning:
                failures.append(f"{prefix} currency warning marked strong")
            if location_warning:
                failures.append(f"{prefix} location warning marked strong")
            if failure and failure != "none":
                failures.append(f"{prefix} failure category marked strong")
            if location_uncertain or location_scope in {"wrong", "missing", "unknown", "marketplace"} or (
                location_exact_explicit and not _truthy(row.get("frontend_location_exact"))
            ):
                failures.append(f"{prefix} uncertain location marked strong")
            if location_scope != "exact" and not _truthy(row.get("frontend_location_exact")):
                failures.append(f"{prefix} strong evidence missing exact location")
            if not location_verified:
                failures.append(f"{prefix} strong evidence missing verified location")
            if partial_search:
                failures.append(f"{prefix} partial search evidence marked strong")
            if search_status != "已自动检查":
                failures.append(f"{prefix} strong evidence missing successful search page")
            if competitor_count is None:
                failures.append(f"{prefix} strong evidence missing competitor count")
            if not comparability:
                failures.append(f"{prefix} strong evidence missing competitor comparability")
            if (
                comparability != "high"
                or (competitor_count is not None and competitor_count < 2)
                or (comparable_competitor_count is not None and comparable_competitor_count < 2)
            ):
                failures.append(f"{prefix} weak competitor evidence marked strong")
            if quality_score is None or quality_score < 75:
                failures.append(f"{prefix} strong evidence quality score below threshold")
        if status == "已自动检查":
            continue
        if not _is_frontend_fallback_status(status, freshness, findings):
            failures.append(f"{prefix} missing cached-date or pending-check status")
        if status.startswith("沿用") and not DATE_RE.search(" ".join([status, freshness, findings])):
            failures.append(f"{prefix} cache status missing date")
    return failures


def _frontend_coverage_expected(rows: list[object]) -> dict[str, object]:
    total = len(rows)
    usable = 0
    live = 0
    cached = 0
    search_success = 0
    search_partial = 0
    own_sellersprite = 0
    own_sellersprite_today = 0
    own_sellersprite_cache = 0
    own_sellersprite_pending = 0
    own_sellersprite_failed = 0
    sellersprite_trend_ready = 0
    competitor_discovery = 0
    competitor_pool = 0
    competitor_pool_today = 0
    competitor_pool_cache = 0
    competitor_pool_pending = 0
    competitor_pool_failed = 0
    competitor_sellersprite = 0
    competitor_sellersprite_today = 0
    competitor_sellersprite_cache = 0
    competitor_sellersprite_pending = 0
    competitor_sellersprite_asins = 0
    amazon_search_validation = 0
    scalable_strong = 0
    weak_defensive = 0
    insufficient = 0
    stale_or_pending = 0
    strong = 0
    background = 0
    unusable = 0
    market_complete = 0
    market_usable = 0
    market_insufficient = 0
    market_failed = 0
    market_score_total = 0.0
    for row in rows:
        if not isinstance(row, dict):
            continue
        status = str(row.get("frontend_check_status") or "").strip()
        score = _number(row.get("frontend_evidence_quality_score"))
        warning = str(row.get("frontend_price_currency_warning") or "").strip()
        tier = str(row.get("frontend_evidence_tier") or "").strip()
        display_tier = str(row.get("frontend_evidence_display_tier") or "").strip()
        decision_tier = str(row.get("frontend_decision_evidence_tier") or display_tier or tier).strip()
        strong_flag = _truthy(row.get("frontend_evidence_is_strong"))
        effective_tier = decision_tier
        strong_safe = effective_tier == "强诊断可用" and strong_flag
        is_live = status == "已自动检查"
        is_cached = status.startswith("沿用") or _truthy(row.get("frontend_cache_used"))
        if is_live:
            live += 1
        if is_cached:
            cached += 1
        search_status = str(row.get("frontend_search_status") or "").strip()
        if search_status == "已自动检查":
            search_success += 1
        if search_status == "已读取部分结果" or _truthy(row.get("frontend_search_partial_evidence")):
            search_partial += 1
        if is_cached or status in {"待前台检查", "读取失败", ""}:
            stale_or_pending += 1
        seller_status = str(row.get("seller_sprite_check_status") or "").strip()
        if seller_status and seller_status != "无缓存":
            own_sellersprite += 1
        today_status = str(row.get("sellersprite_today_status") or "").strip()
        if today_status == "今日已抓":
            own_sellersprite_today += 1
        elif today_status == "沿用缓存":
            own_sellersprite_cache += 1
        elif today_status == "失败":
            own_sellersprite_failed += 1
        else:
            own_sellersprite_pending += 1
        if str(row.get("sellersprite_trend_status") or "") in {"3天趋势可用", "7天趋势可用"}:
            sellersprite_trend_ready += 1
        discovery_status = str(row.get("competitor_discovery_status") or "").strip()
        if discovery_status in {"已抓取", "沿用缓存", "缓存"}:
            competitor_discovery += 1
        pool_count = int(_number(row.get("competitor_pool_count")) or 0)
        pool_status = str(row.get("competitor_pool_status") or "").strip()
        if pool_count > 0:
            competitor_pool += 1
            if today_status == "今日已抓":
                competitor_pool_today += 1
            else:
                competitor_pool_cache += 1
        elif pool_status in {"卖家精灵竞品发现失败", "竞品证据不足", "卖家精灵证据不足"}:
            competitor_pool_failed += 1
        else:
            competitor_pool_pending += 1
        comp_seller_count = int(_number(row.get("competitor_sellersprite_asin_count")) or 0)
        if comp_seller_count > 0:
            competitor_sellersprite += 1
            competitor_sellersprite_asins += comp_seller_count
            if today_status == "今日已抓":
                competitor_sellersprite_today += 1
            else:
                competitor_sellersprite_cache += 1
        else:
            competitor_sellersprite_pending += 1
        amazon_validation_status = str(row.get("amazon_search_validation_status") or "").strip()
        if amazon_validation_status in {"已验证", "已读，无池内竞品", "部分"}:
            amazon_search_validation += 1
        product_conclusion = str(row.get("product_level_conclusion") or "").strip()
        competitor_pool_missing = pool_count <= 0 or pool_status in {"", "待补", "卖家精灵证据不足", "卖家精灵竞品发现失败", "竞品证据不足"}
        competitor_seller_missing = str(row.get("competitor_sellersprite_status") or "").strip() in {
            "",
            "待补",
            "竞品反查待补",
            "竞品卖家精灵证据不足",
        }
        amazon_verified = amazon_validation_status == "已验证"
        if strong_safe and product_conclusion == "可放量" and not competitor_pool_missing and not competitor_seller_missing and amazon_verified:
            scalable_strong += 1
        if product_conclusion in {"产品问题优先", "暂停扩张", "只防守"} or str(row.get("frontend_auto_conclusion") or "") == "FRONTEND_WEAK":
            weak_defensive += 1
        if effective_tier == "不可用" or competitor_pool_missing or competitor_seller_missing or not seller_status or seller_status == "无缓存":
            insufficient += 1
        if strong_safe:
            strong += 1
        elif effective_tier == "仅背景参考" or (tier == "强诊断可用" and not strong_safe):
            background += 1
        elif effective_tier == "不可用":
            unusable += 1
        if strong_safe or effective_tier == "仅背景参考" or ((is_live or is_cached) and not warning and score is not None and score >= 45):
            usable += 1
        market_level = str(row.get("market_survey_completeness_level") or "").strip()
        market_score = _number(row.get("market_survey_completeness_score")) or 0
        market_score_total += market_score
        if market_level == "complete":
            market_complete += 1
        elif market_level == "usable":
            market_usable += 1
        elif market_level == "insufficient":
            market_insufficient += 1
        elif market_level == "failed":
            market_failed += 1
    usable_rate = (usable / total) if total else 0
    strong_rate = (strong / total) if total else 0
    background_rate = (background / total) if total else 0
    return {
        "frontend_queue_total": total,
        "frontend_usable_evidence_count": usable,
        "frontend_decision_ready_count": strong,
        "frontend_reference_evidence_count": background,
        "frontend_live_success_count": live,
        "frontend_cached_count": cached,
        "frontend_pending_or_stale_count": stale_or_pending,
        "frontend_search_success_count": search_success,
        "frontend_search_partial_count": search_partial,
        "frontend_product_page_success_count": live,
        "frontend_competitor_search_success_count": search_success,
        "frontend_own_sellersprite_count": own_sellersprite,
        "frontend_own_sellersprite_today_count": own_sellersprite_today,
        "frontend_own_sellersprite_cache_count": own_sellersprite_cache,
        "frontend_own_sellersprite_pending_count": own_sellersprite_pending,
        "frontend_own_sellersprite_failed_count": own_sellersprite_failed,
        "frontend_sellersprite_trend_ready_count": sellersprite_trend_ready,
        "frontend_competitor_discovery_count": competitor_discovery,
        "frontend_competitor_pool_count": competitor_pool,
        "frontend_competitor_pool_today_count": competitor_pool_today,
        "frontend_competitor_pool_cache_count": competitor_pool_cache,
        "frontend_competitor_pool_pending_count": competitor_pool_pending,
        "frontend_competitor_pool_failed_count": competitor_pool_failed,
        "frontend_competitor_sellersprite_count": competitor_sellersprite,
        "frontend_competitor_sellersprite_today_count": competitor_sellersprite_today,
        "frontend_competitor_sellersprite_cache_count": competitor_sellersprite_cache,
        "frontend_competitor_sellersprite_pending_count": competitor_sellersprite_pending,
        "frontend_competitor_sellersprite_asin_count": competitor_sellersprite_asins,
        "frontend_amazon_search_validation_count": amazon_search_validation,
        "frontend_scalable_strong_count": scalable_strong,
        "frontend_weak_defensive_count": weak_defensive,
        "frontend_insufficient_count": insufficient,
        "frontend_strong_evidence_count": strong,
        "frontend_background_evidence_count": background,
        "frontend_unusable_evidence_count": unusable,
        "market_survey_complete_count": market_complete,
        "market_survey_usable_count": market_usable,
        "market_survey_insufficient_count": market_insufficient,
        "market_survey_failed_count": market_failed,
        "market_survey_average_score": round(market_score_total / total, 1) if total else 0,
        "frontend_usable_evidence_rate": usable_rate,
        "frontend_decision_ready_rate": strong_rate,
        "frontend_reference_evidence_rate": background_rate,
        "frontend_live_success_rate": (live / total) if total else 0,
        "frontend_search_success_rate": (search_success / total) if total else 0,
        "frontend_search_observed_rate": ((search_success + search_partial) / total) if total else 0,
        "frontend_product_page_success_label": f"{live}/{total}" if total else "无前台队列",
        "frontend_competitor_search_success_label": f"{search_success}/{total}" if total else "无前台队列",
        "frontend_own_sellersprite_label": f"{own_sellersprite}/{total}" if total else "无前台队列",
        "frontend_own_sellersprite_today_label": f"今日 {own_sellersprite_today}/{total}，缓存 {own_sellersprite_cache}/{total}" if total else "无前台队列",
        "frontend_sellersprite_trend_ready_label": f"{sellersprite_trend_ready}/{total}" if total else "无前台队列",
        "frontend_competitor_discovery_label": f"{competitor_discovery}/{total}" if total else "无前台队列",
        "frontend_competitor_pool_label": f"{competitor_pool}/{total}" if total else "无前台队列",
        "frontend_competitor_pool_freshness_label": f"今日 {competitor_pool_today}/{total}，7天缓存 {competitor_pool_cache}/{total}" if total else "无前台队列",
        "frontend_competitor_sellersprite_label": f"{competitor_sellersprite}/{total}，{competitor_sellersprite_asins} ASIN" if total else "无前台队列",
        "frontend_competitor_sellersprite_freshness_label": f"今日 {competitor_sellersprite_today}/{total}，缓存 {competitor_sellersprite_cache}/{total}" if total else "无前台队列",
        "frontend_amazon_search_validation_label": f"{amazon_search_validation}/{total}" if total else "无前台队列",
        "frontend_scalable_strong_label": f"{scalable_strong}/{total}" if total else "无前台队列",
        "frontend_weak_defensive_label": f"{weak_defensive}/{total}" if total else "无前台队列",
        "frontend_insufficient_label": f"{insufficient}/{total}" if total else "无前台队列",
        "frontend_decision_ready_label": f"{strong}/{total} 强证据，{strong_rate:.0%}" if total else "无前台队列",
        "frontend_reference_evidence_label": f"{background}/{total} 背景参考，{background_rate:.0%}" if total else "无前台队列",
        "frontend_coverage_label": f"{usable}/{total} 可用，{usable_rate:.0%}" if total else "无前台队列",
        "market_survey_complete_label": f"{market_complete}/{total}" if total else "无市场调查队列",
        "market_survey_usable_label": f"{market_usable}/{total}" if total else "无市场调查队列",
        "market_survey_insufficient_label": f"{market_insufficient}/{total}" if total else "无市场调查队列",
        "market_survey_failed_label": f"{market_failed}/{total}" if total else "无市场调查队列",
        "market_survey_average_score_label": f"{round(market_score_total / total, 1)}/100" if total else "无市场调查队列",
    }


def _frontend_coverage_aggregate_from_snapshots(results: list[object]) -> dict[str, object]:
    totals = {field: 0 for field in FRONTEND_COVERAGE_COUNT_FIELDS}
    market_score_total = 0.0
    for result in results:
        if not isinstance(result, dict):
            continue
        snapshot = result.get("report_view_snapshot") or {}
        if not isinstance(snapshot, dict):
            continue
        coverage = snapshot.get("frontend_coverage_summary") or {}
        if not isinstance(coverage, dict):
            continue
        for field in FRONTEND_COVERAGE_COUNT_FIELDS:
            totals[field] += int(_number(coverage.get(field)) or 0)
        queue_total = int(_number(coverage.get("frontend_queue_total")) or 0)
        market_score_total += (_number(coverage.get("market_survey_average_score")) or 0) * queue_total
    total = totals["frontend_queue_total"]
    usable = totals["frontend_usable_evidence_count"]
    strong = totals["frontend_decision_ready_count"]
    background = totals["frontend_reference_evidence_count"]
    live = totals["frontend_live_success_count"]
    search_success = totals["frontend_search_success_count"]
    search_partial = totals["frontend_search_partial_count"]
    product_success = totals["frontend_product_page_success_count"]
    own_sellersprite = totals["frontend_own_sellersprite_count"]
    own_sellersprite_today = totals["frontend_own_sellersprite_today_count"]
    own_sellersprite_cache = totals["frontend_own_sellersprite_cache_count"]
    sellersprite_trend_ready = totals["frontend_sellersprite_trend_ready_count"]
    competitor_discovery = totals["frontend_competitor_discovery_count"]
    competitor_pool = totals["frontend_competitor_pool_count"]
    competitor_pool_today = totals["frontend_competitor_pool_today_count"]
    competitor_pool_cache = totals["frontend_competitor_pool_cache_count"]
    competitor_sellersprite = totals["frontend_competitor_sellersprite_count"]
    competitor_sellersprite_today = totals["frontend_competitor_sellersprite_today_count"]
    competitor_sellersprite_cache = totals["frontend_competitor_sellersprite_cache_count"]
    competitor_sellersprite_asins = totals["frontend_competitor_sellersprite_asin_count"]
    amazon_search_validation = totals["frontend_amazon_search_validation_count"]
    scalable_strong = totals["frontend_scalable_strong_count"]
    weak_defensive = totals["frontend_weak_defensive_count"]
    insufficient = totals["frontend_insufficient_count"]
    market_complete = totals["market_survey_complete_count"]
    market_usable = totals["market_survey_usable_count"]
    market_insufficient = totals["market_survey_insufficient_count"]
    market_failed = totals["market_survey_failed_count"]
    market_average = round(market_score_total / total, 1) if total else 0
    return {
        **totals,
        "frontend_usable_evidence_rate": (usable / total) if total else 0,
        "frontend_decision_ready_rate": (strong / total) if total else 0,
        "frontend_reference_evidence_rate": (background / total) if total else 0,
        "frontend_live_success_rate": (live / total) if total else 0,
        "frontend_search_success_rate": (search_success / total) if total else 0,
        "frontend_search_observed_rate": ((search_success + search_partial) / total) if total else 0,
        "market_survey_average_score": market_average,
        "frontend_product_page_success_label": f"{product_success}/{total}" if total else "无前台队列",
        "frontend_competitor_search_success_label": f"{search_success}/{total}" if total else "无前台队列",
        "frontend_own_sellersprite_label": f"{own_sellersprite}/{total}" if total else "无前台队列",
        "frontend_own_sellersprite_today_label": f"今日 {own_sellersprite_today}/{total}，缓存 {own_sellersprite_cache}/{total}" if total else "无前台队列",
        "frontend_sellersprite_trend_ready_label": f"{sellersprite_trend_ready}/{total}" if total else "无前台队列",
        "frontend_competitor_discovery_label": f"{competitor_discovery}/{total}" if total else "无前台队列",
        "frontend_competitor_pool_label": f"{competitor_pool}/{total}" if total else "无前台队列",
        "frontend_competitor_pool_freshness_label": f"今日 {competitor_pool_today}/{total}，7天缓存 {competitor_pool_cache}/{total}" if total else "无前台队列",
        "frontend_competitor_sellersprite_label": f"{competitor_sellersprite}/{total}，{competitor_sellersprite_asins} ASIN" if total else "无前台队列",
        "frontend_competitor_sellersprite_freshness_label": f"今日 {competitor_sellersprite_today}/{total}，缓存 {competitor_sellersprite_cache}/{total}" if total else "无前台队列",
        "frontend_amazon_search_validation_label": f"{amazon_search_validation}/{total}" if total else "无前台队列",
        "frontend_scalable_strong_label": f"{scalable_strong}/{total}" if total else "无前台队列",
        "frontend_weak_defensive_label": f"{weak_defensive}/{total}" if total else "无前台队列",
        "frontend_insufficient_label": f"{insufficient}/{total}" if total else "无前台队列",
        "frontend_decision_ready_label": f"{strong}/{total} 强证据，{strong / total:.0%}" if total else "无前台队列",
        "frontend_reference_evidence_label": f"{background}/{total} 背景参考，{background / total:.0%}" if total else "无前台队列",
        "market_survey_complete_label": f"{market_complete}/{total}" if total else "无市场调查队列",
        "market_survey_usable_label": f"{market_usable}/{total}" if total else "无市场调查队列",
        "market_survey_insufficient_label": f"{market_insufficient}/{total}" if total else "无市场调查队列",
        "market_survey_failed_label": f"{market_failed}/{total}" if total else "无市场调查队列",
        "market_survey_average_score_label": f"{market_average}/100" if total else "无市场调查队列",
        "frontend_coverage_label": f"{usable}/{total} 可用，{usable / total:.0%}" if total else "无前台队列",
    }


def _top_level_frontend_coverage_for_display(payload: dict) -> dict[str, object]:
    coverage = payload.get("frontend_coverage_summary")
    if isinstance(coverage, dict):
        return coverage
    return _frontend_coverage_aggregate_from_snapshots(payload.get("marketplace_results") or [])


def _frontend_coverage_display_counts(payload: dict) -> tuple[int, int, int]:
    coverage = _top_level_frontend_coverage_for_display(payload)
    total = int(_number(coverage.get("frontend_queue_total")) or 0)
    strong = int(_number(coverage.get("frontend_decision_ready_count")) or 0)
    background = int(_number(coverage.get("frontend_reference_evidence_count")) or 0)
    return total, strong, background


def _frontend_coverage_display_tokens(payload: dict, *, include_title: bool = False) -> list[str]:
    coverage = _top_level_frontend_coverage_for_display(payload)
    total = int(_number(coverage.get("frontend_queue_total")) or 0)
    if not total:
        return []
    product_success = int(_number(coverage.get("frontend_product_page_success_count") or coverage.get("frontend_live_success_count")) or 0)
    own_seller = int(_number(coverage.get("frontend_own_sellersprite_count")) or 0)
    own_seller_label = str(coverage.get("frontend_own_sellersprite_today_label") or f"{own_seller}/{total}")
    sellersprite_trend = int(_number(coverage.get("frontend_sellersprite_trend_ready_count")) or 0)
    sellersprite_trend_label = str(coverage.get("frontend_sellersprite_trend_ready_label") or f"{sellersprite_trend}/{total}")
    competitor_discovery = int(_number(coverage.get("frontend_competitor_discovery_count")) or 0)
    competitor_pool = int(_number(coverage.get("frontend_competitor_pool_count")) or 0)
    competitor_pool_label = str(coverage.get("frontend_competitor_pool_freshness_label") or f"{competitor_pool}/{total}")
    competitor_seller = int(_number(coverage.get("frontend_competitor_sellersprite_count")) or 0)
    competitor_seller_asins = int(_number(coverage.get("frontend_competitor_sellersprite_asin_count")) or 0)
    competitor_seller_label = str(
        coverage.get("frontend_competitor_sellersprite_freshness_label")
        or f"{competitor_seller}/{total}，{competitor_seller_asins} ASIN"
    )
    amazon_validation = int(_number(coverage.get("frontend_amazon_search_validation_count") or coverage.get("frontend_competitor_search_success_count") or coverage.get("frontend_search_success_count")) or 0)
    scalable = int(_number(coverage.get("frontend_scalable_strong_count")) or 0)
    weak = int(_number(coverage.get("frontend_weak_defensive_count")) or 0)
    insufficient = int(_number(coverage.get("frontend_insufficient_count")) or 0)
    market_complete = int(_number(coverage.get("market_survey_complete_count")) or 0)
    market_usable = int(_number(coverage.get("market_survey_usable_count")) or 0)
    market_insufficient = int(_number(coverage.get("market_survey_insufficient_count")) or 0)
    market_failed = int(_number(coverage.get("market_survey_failed_count")) or 0)
    market_average = str(coverage.get("market_survey_average_score_label") or f'{coverage.get("market_survey_average_score", 0)}/100')
    tokens = [
        "产品页成功",
        f"{product_success}/{total}",
        "卖家精灵自己 ASIN",
        own_seller_label,
        "卖家精灵趋势",
        sellersprite_trend_label,
        "卖家精灵竞品发现",
        f"{competitor_discovery}/{total}",
        "卖家精灵竞品池",
        competitor_pool_label,
        "竞品 ASIN 反查",
        competitor_seller_label,
        "Amazon 搜索页辅助验证",
        f"{amazon_validation}/{total}",
        "达到放量准入",
        f"{scalable}/{total}",
        "弱势止损证据",
        f"{weak}/{total}",
        "证据不足",
        f"{insufficient}/{total}",
        "市场调查平均完整度",
        market_average,
        "强证据 / 可用证据",
        f"{market_complete}/{total} / {market_usable}/{total}",
    ]
    if market_insufficient or market_failed:
        tokens.extend(
            [
                "待补 / 失败",
                f"{market_insufficient}/{total} / {market_failed}/{total}",
            ]
        )
    return ["前台证据覆盖", *tokens] if include_title else tokens


def _frontend_coverage_excel_expected(payload: dict) -> dict[str, str]:
    coverage = _top_level_frontend_coverage_for_display(payload)
    total = int(_number(coverage.get("frontend_queue_total")) or 0)
    if not total:
        return {}
    product_success = int(_number(coverage.get("frontend_product_page_success_count") or coverage.get("frontend_live_success_count")) or 0)
    own_seller = int(_number(coverage.get("frontend_own_sellersprite_count")) or 0)
    own_seller_today = int(_number(coverage.get("frontend_own_sellersprite_today_count")) or 0)
    own_seller_cache = int(_number(coverage.get("frontend_own_sellersprite_cache_count")) or 0)
    sellersprite_trend = int(_number(coverage.get("frontend_sellersprite_trend_ready_count")) or 0)
    competitor_discovery = int(_number(coverage.get("frontend_competitor_discovery_count")) or 0)
    competitor_pool = int(_number(coverage.get("frontend_competitor_pool_count")) or 0)
    competitor_pool_today = int(_number(coverage.get("frontend_competitor_pool_today_count")) or 0)
    competitor_pool_cache = int(_number(coverage.get("frontend_competitor_pool_cache_count")) or 0)
    competitor_seller = int(_number(coverage.get("frontend_competitor_sellersprite_count")) or 0)
    competitor_seller_asins = int(_number(coverage.get("frontend_competitor_sellersprite_asin_count")) or 0)
    competitor_seller_today = int(_number(coverage.get("frontend_competitor_sellersprite_today_count")) or 0)
    competitor_seller_cache = int(_number(coverage.get("frontend_competitor_sellersprite_cache_count")) or 0)
    amazon_validation = int(_number(coverage.get("frontend_amazon_search_validation_count") or coverage.get("frontend_competitor_search_success_count") or coverage.get("frontend_search_success_count")) or 0)
    scalable = int(_number(coverage.get("frontend_scalable_strong_count")) or 0)
    weak = int(_number(coverage.get("frontend_weak_defensive_count")) or 0)
    insufficient = int(_number(coverage.get("frontend_insufficient_count")) or 0)
    market_complete = int(_number(coverage.get("market_survey_complete_count")) or 0)
    market_usable = int(_number(coverage.get("market_survey_usable_count")) or 0)
    market_insufficient = int(_number(coverage.get("market_survey_insufficient_count")) or 0)
    market_failed = int(_number(coverage.get("market_survey_failed_count")) or 0)
    market_average = str(coverage.get("market_survey_average_score_label") or f'{coverage.get("market_survey_average_score", 0)}/100')
    return {
        "ALL 前台队列": str(total),
        "ALL 市场调查平均完整度": market_average,
        "ALL 市场调查完整": f"{market_complete}/{total}",
        "ALL 市场调查可用": f"{market_usable}/{total}",
        "ALL 市场调查待补": f"{market_insufficient}/{total}",
        "ALL 市场调查失败": f"{market_failed}/{total}",
        "ALL 产品页成功": f"{product_success}/{total}",
        "ALL 卖家精灵自己 ASIN": f"今日 {own_seller_today}/{total}，缓存 {own_seller_cache}/{total}",
        "ALL 卖家精灵趋势": f"{sellersprite_trend}/{total}",
        "ALL 卖家精灵竞品发现": f"{competitor_discovery}/{total}",
        "ALL 卖家精灵竞品池": f"今日 {competitor_pool_today}/{total}，7天缓存 {competitor_pool_cache}/{total}",
        "ALL 竞品 ASIN 反查": f"今日 {competitor_seller_today}/{total}，缓存 {competitor_seller_cache}/{total}",
        "ALL Amazon 搜索页辅助验证": f"{amazon_validation}/{total}",
        "ALL 达到放量准入": f"{scalable}/{total}",
        "ALL 弱势止损证据": f"{weak}/{total}",
        "ALL 证据不足": f"{insufficient}/{total}",
    }


def _frontend_coverage_summary_failures(
    marketplace: str,
    rows: list[object],
    summary: object,
) -> list[str]:
    if not isinstance(summary, dict):
        return [f"latest analysis {marketplace} frontend_coverage_summary must be an object"]
    failures: list[str] = []
    expected = _frontend_coverage_expected(rows)
    numeric_fields = {
        "frontend_usable_evidence_rate",
        "frontend_decision_ready_rate",
        "frontend_reference_evidence_rate",
        "frontend_live_success_rate",
        "frontend_search_success_rate",
        "frontend_search_observed_rate",
        "market_survey_average_score",
    }
    for field, expected_value in expected.items():
        actual_value = summary.get(field)
        if field in numeric_fields:
            actual_number = _number(actual_value)
            expected_number = float(expected_value)
            if actual_number is None or abs(actual_number - expected_number) > 0.0001:
                failures.append(
                    f"latest analysis {marketplace} frontend_coverage_summary field {field}: "
                    f"expected {expected_number}, got {actual_value!r}"
                )
        elif actual_value != expected_value:
            failures.append(
                f"latest analysis {marketplace} frontend_coverage_summary field {field}: "
                f"expected {expected_value!r}, got {actual_value!r}"
            )
    return failures


def _latest_analysis_review_snapshot_failures(
    marketplace: str,
    rows: list[object],
    *,
    required_fields: list[str],
    label: str,
) -> list[str]:
    failures: list[str] = []
    object_rows: list[object] = []
    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            failures.append(f"latest analysis {marketplace} {label} row {idx} must be an object")
            object_rows.append(row)
            continue
        object_rows.append(row)
        row_marketplace = str(row.get("marketplace") or "").strip().upper()
        target = str(
            row.get("search_term_or_target")
            or row.get("product_name")
            or row.get("asin")
            or "unknown_review_row"
        )
        if not row_marketplace:
            failures.append(
                f"latest analysis {marketplace} {label} row {idx} {target} missing marketplace value"
            )
        elif row_marketplace != marketplace:
            failures.append(
                f"latest analysis {marketplace} {label} row {idx} {target} "
                f"contains {row_marketplace} marketplace data"
            )
        missing_fields = [field for field in required_fields if field not in row]
        if missing_fields:
            failures.append(
                f"latest analysis {marketplace} {label} row {idx} {target} "
                f"missing fields: {', '.join(missing_fields)}"
            )
        if "action_id" in row and not str(row.get("action_id") or "").strip():
            failures.append(f"latest analysis {marketplace} {label} row {idx} {target} missing action_id value")
    failures.extend(_review_effect_evidence_failures(object_rows, label, source=f"latest analysis {marketplace}"))
    return failures


def _marketplace_scoped_row_failures(
    marketplace: str,
    rows: list[object],
    *,
    label: str,
    source: str,
) -> list[str]:
    failures: list[str] = []
    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            continue
        target = str(
            row.get("search_term_or_target")
            or row.get("product_name")
            or row.get("asin")
            or "unknown_row"
        )
        row_marketplace = str(row.get("marketplace") or "").strip().upper()
        if not row_marketplace:
            failures.append(f"{source} {marketplace} {label} row {idx} {target} missing marketplace value")
        elif row_marketplace != marketplace:
            failures.append(
                f"{source} {marketplace} {label} row {idx} {target} contains {row_marketplace} marketplace data"
            )
    return failures


def _product_operation_card_failures(
    marketplace: str,
    rows: list[object],
    product_decision_rows: list[object],
    ad_workbench_rows: list[dict[str, object]] | None = None,
) -> list[str]:
    decisions: dict[tuple[str, str, str], dict[str, object]] = {}
    for row in product_decision_rows:
        if not isinstance(row, dict):
            continue
        key = _row_market_sku_asin(row, marketplace)
        if key[0] and key[2]:
            decisions[key] = row

    pending_ad_ids_by_key: dict[tuple[str, str, str], list[str]] = defaultdict(list)
    for ad_row in ad_workbench_rows or []:
        if not isinstance(ad_row, dict) or not _is_pending_ad_workbench_row(ad_row):
            continue
        ad_key = _row_market_sku_asin(ad_row, marketplace)
        action_id = str(ad_row.get("action_id") or "").strip()
        if ad_key[0] != marketplace or not ad_key[2] or not action_id:
            continue
        if action_id not in pending_ad_ids_by_key[ad_key]:
            pending_ad_ids_by_key[ad_key].append(action_id)

    failures: list[str] = []
    seen: Counter[tuple[str, str, str]] = Counter()
    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            failures.append(f"latest analysis {marketplace} product_operation_cards row {idx} must be an object")
            continue
        key = _row_market_sku_asin(row, marketplace)
        asin = key[2] or "unknown_asin"
        target = str(row.get("product_name") or asin)
        seen[key] += 1
        if not key[0]:
            failures.append(f"latest analysis {marketplace} product_operation_cards row {idx} {target} missing marketplace value")
        elif key[0] != marketplace:
            failures.append(
                f"latest analysis {marketplace} product_operation_cards row {idx} {target} "
                f"contains {key[0]} marketplace data"
            )
        missing_fields = [field for field in PRODUCT_OPERATION_CARD_REQUIRED_FIELDS if field not in row]
        if missing_fields:
            failures.append(
                f"latest analysis {marketplace} product_operation_cards row {idx} {target} "
                f"missing fields: {', '.join(missing_fields)}"
            )
        ad_items = row.get("ad_action_items")
        if not isinstance(ad_items, list):
            failures.append(f"latest analysis {marketplace} product_operation_cards row {idx} {target} ad_action_items must be a list")
        else:
            if int(_number(row.get("ad_action_count")) or 0) != len(ad_items):
                failures.append(
                    f"latest analysis {marketplace} product_operation_cards row {idx} {target} "
                    f"ad_action_count mismatch: expected {len(ad_items)}, got {row.get('ad_action_count')!r}"
                )
            display_limit = int(_number(row.get("ad_action_display_limit")) or 0)
            if display_limit <= 0:
                failures.append(
                    f"latest analysis {marketplace} product_operation_cards row {idx} {target} "
                    f"ad_action_display_limit must be positive, got {row.get('ad_action_display_limit')!r}"
                )
            expected_more_count = max(0, len(ad_items) - display_limit) if display_limit > 0 else 0
            actual_more_count = int(_number(row.get("ad_action_more_count")) or 0)
            if actual_more_count != expected_more_count:
                failures.append(
                    f"latest analysis {marketplace} product_operation_cards row {idx} {target} "
                    f"ad_action_more_count mismatch: expected {expected_more_count}, got {row.get('ad_action_more_count')!r}"
                )
            for item_idx, ad_item in enumerate(ad_items, start=1):
                item_prefix = (
                    f"latest analysis {marketplace} product_operation_cards row {idx} {target} "
                    f"ad_action_items row {item_idx}"
                )
                if not isinstance(ad_item, dict):
                    failures.append(f"{item_prefix} must be an object")
                    continue
                item_key = _row_market_sku_asin(ad_item, marketplace)
                if item_key[0] != key[0]:
                    failures.append(
                        f"{item_prefix} marketplace mismatch: expected {key[0] or '<missing>'}, got {item_key[0] or '<missing>'}"
                    )
                if key[1] and item_key[1] and item_key[1] != key[1]:
                    failures.append(f"{item_prefix} sku mismatch: expected {key[1]}, got {item_key[1]}")
                if item_key[2] != key[2]:
                    failures.append(
                        f"{item_prefix} asin mismatch: expected {key[2] or '<missing>'}, got {item_key[2] or '<missing>'}"
                    )
                failures.extend(
                    _action_identity_failure_for_row(
                        ad_item,
                        item_prefix,
                        marketplace,
                        require_action_id=True,
                    )
                )
            shown_action_ids = {
                str(ad_item.get("action_id") or "").strip()
                for ad_item in ad_items
                if isinstance(ad_item, dict) and str(ad_item.get("action_id") or "").strip()
            }
            for expected_action_id in pending_ad_ids_by_key.get(key, []):
                if expected_action_id not in shown_action_ids:
                    failures.append(
                        f"latest analysis {marketplace} product_operation_cards row {idx} {target} "
                        f"missing ad_action_items action_id from ad workbench rows: {expected_action_id}"
                    )
        decision = decisions.get(key)
        if decision:
            for field in [
                "final_decision",
                "final_decision_label",
                "fusion_action_gate",
                "fusion_do_not_do",
                "today_allowed_actions",
                "today_blocked_actions",
            ]:
                expected = _canonical_report_value(decision.get(field))
                actual = _canonical_report_value(row.get(field))
                if expected != actual:
                    failures.append(
                        f"latest analysis {marketplace} product_operation_cards row {idx} {target} "
                        f"field {field} mismatch vs product_final_decision_rows: expected {expected!r}, got {actual!r}"
                    )
        elif str(row.get("final_decision") or "").strip():
            failures.append(
                f"latest analysis {marketplace} product_operation_cards row {idx} {target} "
                "has final_decision without matching product_final_decision_rows identity"
            )

    for identity, count in seen.items():
        if identity[0] and identity[2] and count > 1:
            failures.append(
                f"latest analysis {marketplace} product_operation_cards duplicate identity: {identity}"
            )
    return failures


def _top_level_snapshot_union_failures(payload: dict, snapshot_rows_by_key: dict[str, list[object]]) -> list[str]:
    failures: list[str] = []
    for key, (identity_fields, label) in sorted(TOP_LEVEL_SNAPSHOT_UNION_KEYS.items()):
        top_level_rows = payload.get(key)
        if not isinstance(top_level_rows, list):
            failures.append(f"latest analysis top-level {key} must be a list")
            continue
        union_rows = snapshot_rows_by_key.get(key, [])
        if len(top_level_rows) != len(union_rows):
            failures.append(
                f"latest analysis top-level {key} count mismatch for {label}: "
                f"expected {len(union_rows)}, got {len(top_level_rows)}"
            )
            continue
        top_level_identity = _identity_counter(top_level_rows, identity_fields)
        union_identity = _identity_counter(union_rows, identity_fields)
        if top_level_identity != union_identity:
            failures.append(
                f"latest analysis top-level {key} identity mismatch for {label}: "
                f"expected {sorted(union_identity.elements())}, got {sorted(top_level_identity.elements())}"
            )
            continue
        top_level_object_rows = [row for row in top_level_rows if isinstance(row, dict)]
        union_object_rows = [row for row in union_rows if isinstance(row, dict)]
        if len(top_level_object_rows) != len(top_level_rows):
            failures.append(f"latest analysis top-level {key} contains non-object rows")
            continue
        if len(union_object_rows) != len(union_rows):
            failures.append(f"latest analysis marketplace snapshot {key} contains non-object rows")
            continue
        content_fields = list(TOP_LEVEL_SNAPSHOT_CONTENT_FIELDS.get(key, []))
        if not content_fields:
            continue
        expected_by_identity, expected_duplicate_failures = _rows_by_unique_identity(
            union_object_rows,
            identity_fields,
            label=label,
            source="latest analysis marketplace snapshots",
        )
        actual_by_identity, actual_duplicate_failures = _rows_by_unique_identity(
            top_level_object_rows,
            identity_fields,
            label=label,
            source=f"latest analysis top-level {key}",
        )
        failures.extend(expected_duplicate_failures)
        failures.extend(actual_duplicate_failures)
        if expected_duplicate_failures or actual_duplicate_failures:
            continue
        for identity, expected_row in sorted(expected_by_identity.items()):
            actual_row = actual_by_identity.get(identity)
            if actual_row is None:
                continue
            for field in content_fields:
                if field not in expected_row:
                    continue
                expected_value = _canonical_report_value(expected_row.get(field))
                actual_value = _canonical_report_value(actual_row.get(field))
                if actual_value != expected_value:
                    failures.append(
                        f"latest analysis top-level {key} field mismatch for {label} "
                        f"{identity} field {field}: expected {expected_value!r} "
                        f"from marketplace snapshots, got {actual_value!r}"
                    )
    return failures


def _latest_analysis_content_failures(path: Path, report_date: str) -> list[str]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        return [f"latest analysis cannot be read: {exc}"]
    if not isinstance(payload, dict):
        return ["latest analysis root must be an object"]
    results = payload.get("marketplace_results")
    if not isinstance(results, list):
        return ["latest analysis marketplace_results must be a list"]
    marketplaces = {
        str(result.get("marketplace") or "").upper()
        for result in results
        if isinstance(result, dict) and result.get("marketplace")
    }
    expected_marketplaces = {"UK", "US", "DE"}
    if marketplaces != expected_marketplaces:
        return [f"latest analysis marketplace_results must contain UK, US, DE; got {sorted(marketplaces)}"]
    failures: list[str] = []
    marketplace_sequence = [
        str(result.get("marketplace") or "").upper()
        for result in results
        if isinstance(result, dict)
    ]
    marketplace_counts = Counter(marketplace_sequence)
    duplicate_marketplaces = sorted(
        marketplace
        for marketplace, count in marketplace_counts.items()
        if marketplace and count > 1
    )
    if duplicate_marketplaces or len(marketplace_sequence) != len(expected_marketplaces):
        failures.append(
            "latest analysis marketplace_results must contain exactly one UK, one US, and one DE; "
            f"got {marketplace_sequence}"
        )
    top_level_marketplace_summaries: dict[str, dict] = {}
    snapshot_rows_by_key: dict[str, list[object]] = {key: [] for key in TOP_LEVEL_SNAPSHOT_UNION_KEYS}
    for key in sorted(REPORT_VIEW_MARKETPLACE_SUMMARY_KEYS):
        value = payload.get(key)
        if not isinstance(value, dict):
            failures.append(f"latest analysis top-level {key} must be an object")
            continue
        summary_marketplaces = {str(item).upper() for item in value.keys()}
        if summary_marketplaces != expected_marketplaces:
            failures.append(
                f"latest analysis top-level {key} must contain UK, US, DE; got {sorted(summary_marketplaces)}"
            )
            continue
        top_level_marketplace_summaries[key] = value
    top_level_frontend_coverage = payload.get("frontend_coverage_summary")
    expected_top_level_frontend_coverage = _frontend_coverage_aggregate_from_snapshots(results)
    if not isinstance(top_level_frontend_coverage, dict):
        failures.append("latest analysis top-level frontend_coverage_summary must be an object")
    else:
        for field, expected_value in expected_top_level_frontend_coverage.items():
            actual_value = top_level_frontend_coverage.get(field)
            if field in FRONTEND_COVERAGE_COUNT_FIELDS:
                if int(_number(actual_value) or 0) != int(expected_value):
                    failures.append(
                        f"latest analysis top-level frontend_coverage_summary field {field}: "
                        f"expected {expected_value}, got {actual_value}"
                    )
            elif field in FRONTEND_COVERAGE_RATE_FIELDS:
                actual_number = _number(actual_value)
                if actual_number is None or abs(float(actual_number) - float(expected_value)) > 1e-9:
                    failures.append(
                        f"latest analysis top-level frontend_coverage_summary field {field}: "
                        f"expected {expected_value}, got {actual_value}"
                    )
            elif actual_value != expected_value:
                failures.append(
                    f"latest analysis top-level frontend_coverage_summary field {field}: "
                    f"expected {expected_value}, got {actual_value}"
                )
    for result in results:
        if not isinstance(result, dict):
            failures.append("latest analysis marketplace_results contains non-object row")
            continue
        marketplace = str(result.get("marketplace") or "unknown").upper()
        summary = result.get("summary") or {}
        if not isinstance(summary, dict):
            failures.append(f"latest analysis {marketplace} summary must be an object")
            continue
        if result.get("has_data") is False:
            failures.append(f"latest analysis {marketplace} marketplace result has_data is false")
        for field in ["ads_row_count", "erp_row_count"]:
            if field not in summary:
                continue
            value = _number(summary.get(field))
            if value is None or value <= 0:
                failures.append(f"latest analysis {marketplace} summary {field} must be positive, got {summary.get(field)!r}")
        summary_report_date = str(summary.get("report_date") or "").strip()
        if summary_report_date != report_date:
            failures.append(
                f"latest analysis {marketplace} summary report_date mismatch: expected {report_date}, got {summary_report_date or '<missing>'}"
            )
        snapshot = result.get("report_view_snapshot")
        if not isinstance(snapshot, dict):
            failures.append(f"latest analysis {marketplace} report_view_snapshot must be an object")
            continue
        missing_list_keys = sorted(REQUIRED_REPORT_VIEW_SNAPSHOT_LIST_KEYS.difference(snapshot.keys()))
        missing_dict_keys = sorted(REQUIRED_REPORT_VIEW_SNAPSHOT_DICT_KEYS.difference(snapshot.keys()))
        missing_keys = [*missing_dict_keys, *missing_list_keys]
        if missing_keys:
            failures.append(f"latest analysis {marketplace} report_view_snapshot missing {', '.join(missing_keys)}")
        for key in sorted(REQUIRED_REPORT_VIEW_SNAPSHOT_LIST_KEYS.intersection(snapshot.keys())):
            if not isinstance(snapshot.get(key), list):
                failures.append(f"latest analysis {marketplace} report_view_snapshot {key} must be a list")
            elif key in ACTION_ID_CONSISTENCY_SNAPSHOT_KEYS:
                failures.extend(
                    _action_id_consistency_failures(
                        marketplace,
                        snapshot.get(key) or [],
                        row_source=key,
                        require_action_id=key in ACTION_ID_REQUIRED_SNAPSHOT_KEYS,
                    )
                )
        for key in sorted(REQUIRED_REPORT_VIEW_SNAPSHOT_DICT_KEYS.intersection(snapshot.keys())):
            if not isinstance(snapshot.get(key), dict):
                failures.append(f"latest analysis {marketplace} report_view_snapshot {key} must be an object")
        product_decision_rows = snapshot.get("product_final_decision_rows")
        for key in TOP_LEVEL_SNAPSHOT_UNION_KEYS:
            snapshot_rows = snapshot.get(key)
            if isinstance(snapshot_rows, list):
                snapshot_rows_by_key[key].extend(snapshot_rows)
                failures.extend(
                    _marketplace_scoped_row_failures(
                        marketplace,
                        snapshot_rows,
                        label=key,
                        source="latest analysis",
                    )
                )
        frontend_rows = snapshot.get("frontend_check_queue_rows")
        if isinstance(frontend_rows, list):
            failures.extend(_frontend_queue_failures(marketplace, frontend_rows))
            failures.extend(
                _frontend_coverage_summary_failures(
                    marketplace,
                    frontend_rows,
                    snapshot.get("frontend_coverage_summary"),
                )
            )
        listing_rows = snapshot.get("listing_price_diagnosis_rows")
        if isinstance(listing_rows, list) and isinstance(frontend_rows, list):
            failures.extend(_listing_frontend_evidence_failures(marketplace, listing_rows, frontend_rows))
        task_rows = snapshot.get("today_task_queue_rows")
        if isinstance(task_rows, list) and isinstance(product_decision_rows, list):
            failures.extend(_task_queue_growth_gate_failures(marketplace, task_rows, product_decision_rows))
            failures.extend(
                _task_queue_growth_gate_failures(
                    marketplace,
                    _aux_ad_workbench_rows_from_snapshot(snapshot),
                    product_decision_rows,
                    row_source="ad_workbench_rows",
                )
            )
        action_groups = snapshot.get("today_action_groups")
        if isinstance(task_rows, list) and isinstance(action_groups, dict):
            failures.extend(_today_action_group_consistency_failures(marketplace, task_rows, action_groups))
        action_review_rows = snapshot.get("action_effect_review_rows")
        if isinstance(action_review_rows, list):
            failures.extend(
                _latest_analysis_review_snapshot_failures(
                    marketplace,
                    action_review_rows,
                    required_fields=ACTION_REVIEW_REQUIRED_FIELDS,
                    label="action review",
                )
            )
        keyword_review_rows = snapshot.get("keyword_action_effect_review_rows")
        if isinstance(keyword_review_rows, list):
            failures.extend(
                _latest_analysis_review_snapshot_failures(
                    marketplace,
                    keyword_review_rows,
                    required_fields=KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS,
                    label="keyword action review",
                )
            )
        if isinstance(product_decision_rows, list):
            failures.extend(_product_growth_frontend_failures(marketplace, product_decision_rows))
            expected_decision_counts = dict(
                Counter(
                    str(row.get("final_decision") or "").strip()
                    for row in product_decision_rows
                    if isinstance(row, dict)
                )
            )
            for key in sorted(REPORT_VIEW_MARKETPLACE_SUMMARY_KEYS):
                snapshot_summary = snapshot.get(key)
                if isinstance(snapshot_summary, dict) and snapshot_summary != expected_decision_counts:
                    failures.append(
                        f"latest analysis {marketplace} report_view_snapshot {key} does not match "
                        f"product_final_decision_rows final_decision counts"
                    )
                top_level_summary = top_level_marketplace_summaries.get(key)
                if top_level_summary is not None and snapshot_summary != top_level_summary.get(marketplace):
                    failures.append(
                        f"latest analysis top-level {key} mismatch for {marketplace} report_view_snapshot"
                    )
        product_operation_rows = snapshot.get("product_operation_cards")
        if isinstance(product_operation_rows, list) and isinstance(product_decision_rows, list):
            failures.extend(
                _product_operation_card_failures(
                    marketplace,
                    product_operation_rows,
                    product_decision_rows,
                    _product_operation_ad_rows_from_snapshot(snapshot),
                )
            )
    failures.extend(_top_level_snapshot_union_failures(payload, snapshot_rows_by_key))
    return failures


def _text_report_date_failures(paths: list[Path], report_date: str) -> list[str]:
    failures: list[str] = []
    for path in paths:
        if path.suffix.lower() not in {".html", ".md"}:
            continue
        if not path.exists():
            continue
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
        except Exception as exc:
            failures.append(f"text report cannot be read: {path} | {exc}")
            continue
        if report_date not in text:
            failures.append(f"text report missing report_date {report_date}: {path}")
        for marker in TEXT_REPORT_REQUIRED_MARKERS.get(path.name, []):
            if marker not in text:
                failures.append(f"text report missing required marker {marker}: {path}")
        required_anchors = TEXT_REPORT_REQUIRED_ANCHORS.get(path.name, [])
        if required_anchors:
            anchors = set(re.findall(r"""id=["']([^"']+)["']""", text))
            for anchor in required_anchors:
                if anchor not in anchors:
                    failures.append(f"text report missing required anchor #{anchor}: {path}")
    return failures


def _dashboard_marketplace_summary_failures(report_date: str, output_dir: Path) -> list[str]:
    analysis_path = output_dir / "latest_analysis.json"
    dashboard_path = output_dir / "dashboard.html"
    if not analysis_path.exists() or not dashboard_path.exists():
        return []
    try:
        payload = json.loads(analysis_path.read_text(encoding="utf-8"))
        text = dashboard_path.read_text(encoding="utf-8", errors="replace")
    except Exception as exc:
        return [f"dashboard marketplace summary cannot be validated: {exc}"]
    if not isinstance(payload, dict):
        return []
    if report_date not in text:
        return []
    failures: list[str] = []
    for result in payload.get("marketplace_results") or []:
        if not isinstance(result, dict):
            continue
        summary = result.get("summary") or {}
        if not isinstance(summary, dict):
            continue
        marketplace = str(summary.get("marketplace") or result.get("marketplace") or "").upper()
        if not marketplace:
            failures.append("dashboard.html marketplace summary row missing marketplace")
            continue
        summary_count_fields = ["ads_row_count", "erp_row_count", "sku_count", "asin_count"]
        present_count_fields = [field for field in summary_count_fields if field in summary]
        if not present_count_fields:
            continue
        if len(present_count_fields) != len(summary_count_fields):
            missing_fields = sorted(set(summary_count_fields).difference(present_count_fields))
            failures.append(
                f"dashboard.html marketplace summary for {marketplace} missing source fields: "
                f"{', '.join(missing_fields)}"
            )
            continue
        report_link = f"{marketplace.lower()}_report.html"
        if report_link not in text:
            failures.append(f"dashboard.html missing report link {report_link}")
        fields = [
            marketplace,
            summary.get("ads_row_count"),
            summary.get("erp_row_count"),
            summary.get("sku_count"),
            summary.get("asin_count"),
        ]
        cells = "".join(f"<td>{html.escape(str(value))}</td>" for value in fields)
        pattern = rf"<tr[^>]*>{re.escape(cells)}"
        if not re.search(pattern, text):
            failures.append(
                f"dashboard.html missing marketplace summary row for {marketplace}: "
                f"ads={summary.get('ads_row_count')} erp={summary.get('erp_row_count')} "
                f"sku={summary.get('sku_count')} asin={summary.get('asin_count')}"
            )
    for token in _frontend_coverage_display_tokens(payload, include_title=True):
        if not _html_contains_token(text, token):
            failures.append(f"dashboard.html missing frontend coverage token {token}")
    return failures


def _marketplace_summary_markdown_failures(report_date: str, output_dir: Path) -> list[str]:
    analysis_path = output_dir / "latest_analysis.json"
    summary_path = output_dir / "marketplace_summary.md"
    if not analysis_path.exists() or not summary_path.exists():
        return []
    try:
        payload = json.loads(analysis_path.read_text(encoding="utf-8"))
        text = summary_path.read_text(encoding="utf-8", errors="replace")
    except Exception as exc:
        return [f"marketplace_summary.md cannot be validated: {exc}"]
    if not isinstance(payload, dict):
        return []
    if report_date not in text:
        return []
    failures: list[str] = []
    for result in payload.get("marketplace_results") or []:
        if not isinstance(result, dict):
            continue
        summary = result.get("summary") or {}
        if not isinstance(summary, dict):
            continue
        marketplace = str(summary.get("marketplace") or result.get("marketplace") or "").upper()
        if not marketplace:
            failures.append("marketplace_summary.md summary row missing marketplace")
            continue
        summary_count_fields = ["ads_row_count", "erp_row_count", "sku_count", "asin_count"]
        present_count_fields = [field for field in summary_count_fields if field in summary]
        if not present_count_fields:
            continue
        if len(present_count_fields) != len(summary_count_fields):
            missing_fields = sorted(set(summary_count_fields).difference(present_count_fields))
            failures.append(
                f"marketplace_summary.md summary for {marketplace} missing source fields: "
                f"{', '.join(missing_fields)}"
            )
            continue
        row_pattern = (
            r"\|\s*"
            + r"\s*\|\s*".join(
                re.escape(str(value))
                for value in [
                    marketplace,
                    summary.get("ads_row_count"),
                    summary.get("erp_row_count"),
                    summary.get("sku_count"),
                    summary.get("asin_count"),
                ]
            )
            + r"\s*\|"
        )
        if not re.search(row_pattern, text):
            failures.append(
                f"marketplace_summary.md missing marketplace summary row for {marketplace}: "
                f"ads={summary.get('ads_row_count')} erp={summary.get('erp_row_count')} "
                f"sku={summary.get('sku_count')} asin={summary.get('asin_count')}"
            )
    for token in _frontend_coverage_display_tokens(payload, include_title=True):
        if token not in text:
            failures.append(f"marketplace_summary.md missing frontend coverage token {token}")
    return failures


def _html_contains_token(text: str, token: object) -> bool:
    value = str(token or "").strip()
    return not value or value in text or html.escape(value) in text


def _html_token_contexts(
    text: str,
    anchor: object,
) -> list[str]:
    value = str(anchor or "").strip()
    if not value:
        return []
    contexts: list[str] = []
    for candidate in dict.fromkeys([value, html.escape(value)]):
        start = 0
        while True:
            pos = text.find(candidate, start)
            if pos < 0:
                break
            contexts.append(_html_context_for_position(text, pos))
            start = pos + max(1, len(candidate))
    return contexts


def _html_context_for_position(text: str, pos: int) -> str:
    structured_starts = [
        ("<article", "</article>"),
        ('<div class="summary-item"', '<div class="summary-item"'),
        ('<div class="review-item"', '<div class="review-item"'),
        ('<div class="product-card"', '<div class="product-card"'),
        ('<div class="task-card"', '<div class="task-card"'),
        ("<tr", "</tr>"),
    ]
    candidates: list[tuple[int, int]] = []
    for start_token, end_token in structured_starts:
        start = text.rfind(start_token, 0, pos + 1)
        if start < 0:
            continue
        if end_token == start_token:
            end = text.find(end_token, pos + 1)
            if end < 0:
                end = text.find("</section", pos + 1)
            if end < 0:
                end = text.find("</div></div>", pos + 1)
        else:
            end = text.find(end_token, pos + 1)
            if end >= 0:
                end += len(end_token)
        if end >= 0 and end > pos:
            candidates.append((start, end))
    if candidates:
        start, end = max(candidates, key=lambda item: item[0])
        return text[start:end]
    line_start = text.rfind("\n", 0, pos)
    line_end = text.find("\n", pos)
    return text[(line_start + 1 if line_start >= 0 else 0) : (line_end if line_end >= 0 else len(text))]


def _html_contains_token_near(text: str, anchor: object, token: object) -> bool:
    value = str(token or "").strip()
    if not value:
        return True
    return any(_html_contains_token(context, value) for context in _html_token_contexts(text, anchor))


def _product_decision_contexts(
    text: str,
    asin: object,
    marketplace: object = "",
    sku: object = "",
    container: str = "",
) -> list[str]:
    asin_text = str(asin or "").strip().upper()
    if not asin_text:
        return []
    marketplace_text = str(marketplace or "").strip().upper()
    sku_text = str(sku or "").strip()
    attrs = []
    if marketplace_text:
        attrs.append(f'data-product-decision-marketplace="{html.escape(marketplace_text, quote=True)}"')
    if sku_text:
        attrs.append(f'data-product-decision-sku="{html.escape(sku_text, quote=True)}"')
    attrs.append(f'data-product-decision-asin="{html.escape(asin_text, quote=True)}"')
    contexts: list[str] = []
    start = 0
    while True:
        pos = text.find(attrs[-1], start)
        if pos < 0:
            break
        context = _html_context_for_position(text, pos)
        stripped_context = context.lstrip()
        if container and not stripped_context.startswith(f"<{container}"):
            start = pos + len(attrs[-1])
            continue
        if all(attr in context for attr in attrs):
            contexts.append(context)
        start = pos + len(attrs[-1])
    return contexts


def _html_contains_token_in_contexts(contexts: list[str], token: object) -> bool:
    return any(_html_contains_token(context, token) for context in contexts)


def _html_contains_any_token_in_contexts(contexts: list[str], tokens: list[str]) -> bool:
    return any(_html_contains_token_in_contexts(contexts, token) for token in tokens)


PRODUCT_DECISION_CONTRACT_ATTRS = {
    "final_decision": "data-product-decision-final",
    "final_decision_label": "data-product-decision-label",
    "frontend_evidence_tier": "data-product-frontend-tier",
    "frontend_evidence_display_tier": "data-product-frontend-display-tier",
    "frontend_decision_evidence_tier": "data-product-frontend-decision-tier",
    "frontend_evidence_is_strong": "data-product-frontend-strong",
}


def _html_attr_value(context: str, attr_name: str) -> str:
    match = re.search(rf"""\s{re.escape(attr_name)}=(["'])(.*?)\1""", context, re.DOTALL)
    if not match:
        return ""
    return html.unescape(match.group(2)).strip()


def _html_ad_completion_payload_entries(text: str) -> list[tuple[str, object]]:
    entries: list[tuple[str, object]] = []
    for match in re.finditer(
        r"""(<[^>]*\sdata-ad-complete-payload=(["'])(.*?)\2[^>]*>)""",
        text,
        re.DOTALL,
    ):
        tag = match.group(1)
        raw_payload = html.unescape(match.group(3)).strip()
        try:
            payload = json.loads(raw_payload)
        except json.JSONDecodeError:
            payload = {"_payload_parse_error": raw_payload}
        entries.append((tag, payload))
    return entries


def _html_ad_completion_payload_failures(
    text: str,
    pending_ad_rows: list[dict[str, object]],
    *,
    page_name: str = "latest_recommendations.html",
) -> list[str]:
    failures: list[str] = []
    payload_action_ids: set[str] = set()
    payload_idx = 0
    for tag, payload in _html_ad_completion_payload_entries(text):
        payloads = payload if isinstance(payload, list) else [payload]
        if not isinstance(payloads, list):
            failures.append(f"{page_name} ad completion payload must be an object or list")
            continue
        tag_action_id = _html_attr_value(tag, "data-action-id")
        first_payload_action_id = ""
        for item in payloads:
            payload_idx += 1
            if not isinstance(item, dict):
                failures.append(f"{page_name} ad completion payload {payload_idx} must be an object")
                continue
            target = str(
                item.get("search_term_or_target")
                or item.get("product_name")
                or item.get("asin")
                or "unknown_ad_payload"
            )
            prefix = f"{page_name} ad completion payload {payload_idx} {target}"
            if item.get("_payload_parse_error"):
                failures.append(f"{prefix} cannot parse JSON payload")
                continue
            missing_fields = [
                field
                for field in ["marketplace", "sku", "asin", "search_term_or_target", "action_id", "normalized_action", "action_scope"]
                if not str(item.get(field) or "").strip()
            ]
            if missing_fields:
                failures.append(f"{prefix} missing fields: {', '.join(missing_fields)}")
                continue
            marketplace = str(item.get("marketplace") or "").strip().upper()
            failures.extend(
                _action_identity_failure_for_row(
                    item,
                    prefix,
                    marketplace,
                    require_action_id=True,
                )
            )
            if not is_executable_action(item):
                failures.append(f"{prefix} is not executable but appears in ad completion payload")
            action_id = str(item.get("action_id") or "").strip()
            if action_id:
                payload_action_ids.add(action_id)
                if not first_payload_action_id:
                    first_payload_action_id = action_id
        if tag_action_id and first_payload_action_id and tag_action_id != first_payload_action_id:
            failures.append(
                f"{page_name} ad completion payload data-action-id mismatch: "
                f"expected {first_payload_action_id}, got {tag_action_id}"
            )
    expected_action_ids = {
        str(row.get("action_id") or "").strip()
        for row in pending_ad_rows
        if str(row.get("action_id") or "").strip() and is_executable_action(row)
    }
    missing_action_ids = sorted(expected_action_ids.difference(payload_action_ids))
    for action_id in missing_action_ids[:10]:
        failures.append(f"{page_name} missing ad completion payload action_id {action_id}")
    if len(missing_action_ids) > 10:
        failures.append(
            f"{page_name} missing {len(missing_action_ids) - 10} additional ad completion payload action_ids"
        )
    return failures


def _product_decision_contract_attr_failures(
    page_name: str,
    row: dict[str, object],
    contexts: list[str],
    *,
    row_label: str,
) -> list[str]:
    if not contexts:
        return []
    failures: list[str] = []
    asin = str(row.get("asin") or "").strip().upper() or "unknown_asin"
    for field, attr in PRODUCT_DECISION_CONTRACT_ATTRS.items():
        expected = str(row.get(field) or "").strip()
        if not expected:
            continue
        if not any(_html_attr_value(context, attr) == expected for context in contexts):
            failures.append(
                f"{page_name} {row_label} {asin} data attr {attr} mismatch for field {field}: expected {expected!r}"
            )
    for field, attr in [
        ("today_allowed_actions", "data-product-allowed-actions"),
        ("today_blocked_actions", "data-product-blocked-actions"),
    ]:
        expected_actions = _action_set(row.get(field))
        if not expected_actions:
            continue
        attr_matches = False
        for context in contexts:
            actual_actions = _action_set(_html_attr_value(context, attr).replace("|", "/"))
            if actual_actions == expected_actions:
                attr_matches = True
                break
        if not attr_matches:
            failures.append(
                f"{page_name} {row_label} {asin} data attr {attr} mismatch for field {field}: "
                f"expected {sorted(expected_actions)!r}"
            )
    return failures


PRODUCT_DECISION_BLOCKED_ACTION_LABELS = {
    "bid_up": ["加竞价", "小幅加竞价"],
    "budget_up": ["加预算"],
    "broad_scale": ["放量"],
    "create_exact_low_budget": ["低预算精准测试"],
}


def _html_contains_any_token(text: str, tokens: list[str]) -> bool:
    return any(_html_contains_token(text, token) for token in tokens)


def _html_contains_any_token_near(text: str, anchor: object, tokens: list[str]) -> bool:
    return any(_html_contains_token_near(text, anchor, token) for token in tokens)


def _product_decision_blocking_reason_tokens(row: dict[str, object]) -> list[str]:
    reasons = row.get("frontend_blocking_reasons")
    if isinstance(reasons, list):
        return [str(reason).strip() for reason in reasons if str(reason).strip()]
    reason_text = str(reasons or "").strip()
    if not reason_text:
        return []
    return [token.strip() for token in re.split(r"[；;]\s*", reason_text) if token.strip()]


def _product_frontend_evidence_tokens(row: dict[str, object]) -> list[tuple[str, str]]:
    tokens: list[tuple[str, str]] = []
    status = str(row.get("frontend_check_status") or row.get("frontend_status") or "").strip()
    display_tier = str(row.get("frontend_evidence_display_tier") or row.get("frontend_evidence_tier") or "").strip()
    search_status = str(row.get("frontend_search_status") or "").strip()
    audit_detail = str(row.get("frontend_evidence_audit_detail") or "").strip()
    if not audit_detail:
        audit_reasons = row.get("frontend_evidence_audit_reasons")
        if isinstance(audit_reasons, list):
            audit_detail = "；".join(str(reason).strip() for reason in audit_reasons if str(reason).strip())
        else:
            audit_detail = str(audit_reasons or "").strip()
    if status:
        tokens.append(("frontend status", status))
    if display_tier:
        tokens.append(("frontend display tier", display_tier))
    if audit_detail:
        tokens.append(("frontend audit detail", audit_detail))
    if _truthy(row.get("frontend_cache_used")):
        cache_token = status if "沿用" in status else "沿用"
        tokens.append(("frontend cache marker", cache_token))
    if _truthy(row.get("frontend_search_partial_evidence")) or search_status == "已读取部分结果":
        tokens.append(("frontend search partial marker", search_status or "部分读取"))
    return [(label, token) for label, token in tokens if token]


def _ad_action_label_for_validation(row: dict[str, object]) -> str:
    fallback = ""
    for field in ["suggested_action", "scale_action", "copy_action_line", "copy_block", "today_action"]:
        action = str(row.get(field) or "").strip()
        if not action:
            continue
        if "否" in action and "不直接否" not in action:
            return "否定精准"
        if "暂停" in action and "ASIN" in action.upper():
            return "暂停 ASIN 定向"
        if "降竞价" in action or "降价竞价" in action:
            return "降竞价"
        if "降价" in action:
            return "降价"
        if (
            "growth_test" in action
            or "小预算" in action
            or "新建精准" in action
            or "拉精准" in action
            or "开精准" in action
        ):
            return "小预算试投"
        blocks_bid_up = any(marker in action for marker in ["不加价", "不提高竞价", "不能加价", "禁止加价"])
        if ("加价" in action or "提高竞价" in action) and not blocks_bid_up:
            return "加价"
        if "保留" in action:
            fallback = fallback or "保留观察"
        elif "观察" in action:
            fallback = fallback or "观察"
    return fallback or "观察"


def _is_pending_ad_workbench_row(row: dict[str, object]) -> bool:
    if str(row.get("confirmed_status") or "") in {"已执行", "已核查", "已忽略", "仅背景参考"}:
        return False
    return _ad_action_label_for_validation(row) not in {"观察", "保留观察"}


def _ad_workbench_rows_from_snapshot(snapshot: dict[str, object]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for key in ["html_search_term_processing_queue_rows", "scale_keyword_rows", "growth_test_rows"]:
        value = snapshot.get(key) or []
        if isinstance(value, list):
            rows.extend(row for row in value if isinstance(row, dict))
    groups = snapshot.get("today_action_groups") or {}
    if isinstance(groups, dict):
        group_rows = groups.get("广告动作") or []
        if isinstance(group_rows, list):
            rows.extend(row for row in group_rows if isinstance(row, dict))
    return rows


def _product_operation_ad_rows_from_snapshot(snapshot: dict[str, object]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for key in ["html_search_term_processing_queue_rows", "scale_keyword_rows"]:
        value = snapshot.get(key) or []
        if isinstance(value, list):
            rows.extend(row for row in value if isinstance(row, dict))
    return rows


def _pending_ad_identity_token(row: dict[str, object]) -> tuple[str, str] | None:
    target = str(row.get("search_term_or_target") or "").strip()
    if target:
        return ("target", target)
    asin = str(row.get("asin") or "").strip().upper()
    if asin:
        return ("ASIN", asin)
    product_name = str(row.get("product_name") or "").strip()
    if product_name:
        return ("product name", product_name)
    return None


def _tomorrow_review_identity_token(row: dict[str, object]) -> tuple[str, str] | None:
    target = str(row.get("search_term_or_target") or "").strip()
    if target:
        return ("target", target)
    asin = str(row.get("asin") or "").strip().upper()
    if asin:
        return ("ASIN", asin)
    product_name = str(row.get("product_name") or "").strip()
    if product_name:
        return ("product name", product_name)
    return None


def _review_display_judgement_for_sort(row: dict[str, object]) -> str:
    return _html_review_display_judgement(row)


def _action_review_sort_key(row: dict[str, object]) -> tuple[int, str, str]:
    return _html_action_effect_review_sort_key(row)


def _keyword_review_sort_key(row: dict[str, object]) -> tuple[int, int, str, str]:
    return _html_keyword_review_sort_key(row)


def _review_html_metric_tokens(row: dict[str, object]) -> list[str]:
    days_since = _number(row.get("days_since_execution"))
    if days_since is not None and days_since < 3:
        return []
    tokens: list[str] = []
    if days_since is not None and 3 <= days_since < 7:
        tokens.append("3 天窗口只做初步判断")
    field_labels = [
        ("current_7d_promoted_ad_orders", "本 SKU 单"),
        ("current_7d_acos", "ACOS"),
        ("current_7d_target_acos", "目标 ACOS"),
        ("current_7d_tacos", "TACOS"),
        ("current_7d_total_orders", "总单"),
        ("current_7d_available_stock", "库存"),
    ]
    if days_since is not None and days_since >= 7:
        field_labels.extend(
            [
                ("current_14d_promoted_ad_orders", "14天 本 SKU 单"),
                ("current_14d_acos", "14天 ACOS"),
                ("current_14d_tacos", "14天 TACOS"),
                ("current_14d_total_orders", "14天 总单"),
                ("current_14d_available_stock", "14天 库存"),
            ]
        )
    for field, label in field_labels:
        value = str(row.get(field) or "").strip()
        number = _number(value)
        if value and (number is not None or value.lower() not in {"nan", "none", "null"}):
            tokens.append(f"{label} {value}")
    if str(row.get("review_data_source") or "").strip() == "execution_anchored_daily":
        days_since = _number(row.get("days_since_execution"))
        anchor_fields = [
            ("pre_7d_promoted_ad_orders", "执行前本 SKU 单"),
            ("pre_7d_total_orders", "执行前总单"),
            ("pre_7d_tacos", "执行前 TACOS"),
            ("post_3d_days", "执行后3天覆盖天数"),
            ("post_3d_promoted_ad_orders", "执行后3天本 SKU 单"),
            ("post_3d_total_orders", "执行后3天总单"),
            ("post_3d_acos", "执行后3天 ACOS"),
            ("post_3d_tacos", "执行后3天 TACOS"),
            ("post_3d_available_stock", "执行后3天库存"),
        ]
        if days_since is None or days_since >= 7:
            anchor_fields.extend(
                [
                    ("post_7d_days", "执行后覆盖天数"),
                    ("post_7d_promoted_ad_orders", "执行后本 SKU 单"),
                    ("post_7d_total_orders", "执行后总单"),
                    ("post_7d_acos", "执行后 ACOS"),
                    ("post_7d_tacos", "执行后 TACOS"),
                    ("post_7d_available_stock", "执行后库存"),
                ]
            )
        pre_start = str(row.get("pre_7d_start") or "").strip()
        pre_end = str(row.get("pre_7d_end") or "").strip()
        post_3d_start = str(row.get("post_3d_start") or "").strip()
        post_3d_end = str(row.get("post_3d_end") or "").strip()
        post_start = str(row.get("post_7d_start") or "").strip()
        post_end = str(row.get("post_7d_end") or "").strip()
        if pre_start and pre_end:
            tokens.append(f"执行前7天 {pre_start} 至 {pre_end}")
        if post_3d_start and post_3d_end:
            tokens.append(f"执行后3天 {post_3d_start} 至 {post_3d_end}")
        if (days_since is None or days_since >= 7) and post_start and post_end:
            tokens.append(f"执行后7天 {post_start} 至 {post_end}")
        for field, label in anchor_fields:
            value = str(row.get(field) or "").strip()
            number = _number(value)
            if value and (number is not None or value.lower() not in {"nan", "none", "null"}):
                tokens.append(f"{label} {value}")
    return tokens


def _review_html_forbidden_metric_tokens(row: dict[str, object]) -> list[str]:
    days_since = _number(row.get("days_since_execution"))
    if days_since is None:
        return []
    field_groups: list[tuple[str, str, list[tuple[str, str]]]] = []
    if days_since < 3:
        field_groups.append(
            (
                "current_7d",
                "7天",
                [
                    ("current_7d_promoted_ad_orders", "本 SKU 单"),
                    ("current_7d_acos", "ACOS"),
                    ("current_7d_target_acos", "目标 ACOS"),
                    ("current_7d_tacos", "TACOS"),
                    ("current_7d_total_orders", "总单"),
                    ("current_7d_available_stock", "库存"),
                ],
            )
        )
    if days_since < 7:
        field_groups.append(
            (
                "current_14d",
                "14天",
                [
                    ("current_14d_promoted_ad_orders", "本 SKU 单"),
                    ("current_14d_acos", "ACOS"),
                    ("current_14d_tacos", "TACOS"),
                    ("current_14d_total_orders", "总单"),
                    ("current_14d_available_stock", "库存"),
                ],
            )
        )
    tokens: list[str] = []
    for _prefix, label, fields in field_groups:
        for field, metric_label in fields:
            value = str(row.get(field) or "").strip()
            number = _number(value)
            if value and (number is not None or value.lower() not in {"nan", "none", "null"}):
                tokens.append(f"{label} {metric_label} {value}")
    if str(row.get("review_data_source") or "").strip() == "execution_anchored_daily":
        anchor_fields: list[tuple[str, str]] = []
        if days_since < 3:
            pre_start = str(row.get("pre_7d_start") or "").strip()
            pre_end = str(row.get("pre_7d_end") or "").strip()
            post_3d_start = str(row.get("post_3d_start") or "").strip()
            post_3d_end = str(row.get("post_3d_end") or "").strip()
            if pre_start and pre_end:
                tokens.append(f"执行前7天 {pre_start} 至 {pre_end}")
            if post_3d_start and post_3d_end:
                tokens.append(f"执行后3天 {post_3d_start} 至 {post_3d_end}")
            anchor_fields.extend(
                [
                    ("pre_7d_promoted_ad_orders", "执行前本 SKU 单"),
                    ("pre_7d_total_orders", "执行前总单"),
                    ("pre_7d_tacos", "执行前 TACOS"),
                    ("post_3d_days", "执行后3天覆盖天数"),
                    ("post_3d_promoted_ad_orders", "执行后3天本 SKU 单"),
                    ("post_3d_total_orders", "执行后3天总单"),
                    ("post_3d_acos", "执行后3天 ACOS"),
                    ("post_3d_tacos", "执行后3天 TACOS"),
                    ("post_3d_available_stock", "执行后3天库存"),
                ]
            )
        if days_since < 7:
            post_start = str(row.get("post_7d_start") or "").strip()
            post_end = str(row.get("post_7d_end") or "").strip()
            if post_start and post_end:
                tokens.append(f"执行后7天 {post_start} 至 {post_end}")
            anchor_fields.extend(
                [
                    ("post_7d_days", "执行后覆盖天数"),
                    ("post_7d_promoted_ad_orders", "执行后本 SKU 单"),
                    ("post_7d_total_orders", "执行后总单"),
                    ("post_7d_acos", "执行后 ACOS"),
                    ("post_7d_tacos", "执行后 TACOS"),
                    ("post_7d_available_stock", "执行后库存"),
                ]
            )
        for field, label in anchor_fields:
            value = str(row.get(field) or "").strip()
            number = _number(value)
            if value and (number is not None or value.lower() not in {"nan", "none", "null"}):
                tokens.append(f"{label} {value}")
    return tokens


def _format_summary_metric_number(value: object) -> str:
    text = str(value or "").strip()
    if text.lower() in {"", "nan", "none", "null"}:
        return ""
    number = _number(text)
    if number is None:
        return text
    if number.is_integer():
        return str(int(number))
    return f"{number:.2f}".rstrip("0").rstrip(".")


def _format_summary_metric_percent(value: object) -> str:
    text = str(value or "").strip()
    if text.lower() in {"", "nan", "none", "null"}:
        return ""
    number = _number(text)
    if number is None:
        return text
    percent = number * 100 if abs(number) <= 1 else number
    if percent.is_integer():
        return f"{int(percent)}%"
    return f"{percent:.1f}".rstrip("0").rstrip(".") + "%"


def _summary_review_metric_tokens(row: dict[str, object]) -> list[str]:
    days_since = _number(row.get("days_since_execution"))
    if days_since is not None and days_since < 3:
        return []
    tokens: list[str] = []
    field_labels = [
        ("current_7d_promoted_ad_orders", "本 SKU 单", _format_summary_metric_number),
        ("current_7d_acos", "ACOS", _format_summary_metric_percent),
        ("current_7d_target_acos", "目标 ACOS", _format_summary_metric_percent),
        ("current_7d_tacos", "TACOS", _format_summary_metric_percent),
        ("current_7d_total_orders", "总单", _format_summary_metric_number),
        ("current_7d_available_stock", "库存", _format_summary_metric_number),
    ]
    if days_since is None or days_since >= 7:
        field_labels.extend(
            [
                ("current_14d_promoted_ad_orders", "14天本 SKU 单", _format_summary_metric_number),
                ("current_14d_acos", "14天 ACOS", _format_summary_metric_percent),
                ("current_14d_tacos", "14天 TACOS", _format_summary_metric_percent),
                ("current_14d_total_orders", "14天总单", _format_summary_metric_number),
                ("current_14d_available_stock", "14天库存", _format_summary_metric_number),
            ]
        )
    for field, label, formatter in field_labels:
        value = formatter(row.get(field))
        if value:
            tokens.append(f"{label} {value}")
    return tokens


def _summary_review_forbidden_early_metric_tokens(row: dict[str, object]) -> list[str]:
    days_since = _number(row.get("days_since_execution"))
    if days_since is None or days_since >= 3:
        return []
    tokens: list[str] = []
    field_labels = [
        ("current_7d_promoted_ad_orders", "本 SKU 单", _format_summary_metric_number),
        ("current_7d_acos", "ACOS", _format_summary_metric_percent),
        ("current_7d_target_acos", "目标 ACOS", _format_summary_metric_percent),
        ("current_7d_tacos", "TACOS", _format_summary_metric_percent),
        ("current_7d_total_orders", "总单", _format_summary_metric_number),
        ("current_7d_available_stock", "库存", _format_summary_metric_number),
        ("current_14d_promoted_ad_orders", "14天本 SKU 单", _format_summary_metric_number),
        ("current_14d_acos", "14天 ACOS", _format_summary_metric_percent),
        ("current_14d_tacos", "14天 TACOS", _format_summary_metric_percent),
        ("current_14d_total_orders", "14天总单", _format_summary_metric_number),
        ("current_14d_available_stock", "14天库存", _format_summary_metric_number),
    ]
    for field, label, formatter in field_labels:
        value = formatter(row.get(field))
        if value:
            tokens.append(f"{label} {value}")
    return tokens


def _summary_review_display_judgement(row: dict[str, object]) -> str:
    judgement = str(row.get("judgement") or row.get("outcome") or "待复查")
    if _truthy(row.get("halo_only_conversion")) or _truthy(row.get("target_sku_not_converted")):
        return "本 SKU 未验证"
    if str(row.get("review_outcome") or "").strip() == "needs_manual_review":
        return "待人工复查"
    positive_judgements = {"明确改善", "初步有效", "有改善迹象", "有效", "可保留"}
    if judgement in positive_judgements and not _truthy(row.get("promoted_conversion_improved")):
        return "本 SKU 未验证"
    return judgement


def _summary_review_display_next_step(row: dict[str, object]) -> str:
    if _truthy(row.get("halo_only_conversion")):
        return "仅光环成交，不算本 SKU 有效；今天不追加预算或竞价。"
    if _truthy(row.get("target_sku_not_converted")):
        return "本 SKU 未验证成交；今天不追加预算或竞价。"
    if _summary_review_display_judgement(row) == "本 SKU 未验证":
        return "缺少本 SKU 转化证据；今天不追加预算或竞价。"
    return str(row.get("next_step") or "继续观察，等待足够样本。")


def _summary_review_needs_display_guard(row: dict[str, object]) -> bool:
    return (
        _truthy(row.get("halo_only_conversion"))
        or _truthy(row.get("target_sku_not_converted"))
        or _summary_review_display_judgement(row) in {"本 SKU 未验证", "待人工复查"}
    )


def _summary_review_watch_sort_key(row: dict[str, object]) -> tuple[int, int, int, str]:
    judgement = _summary_review_display_judgement(row)
    next_step = _summary_review_display_next_step(row)
    window = str(row.get("review_window") or "")
    product = str(row.get("product_name") or row.get("search_term_or_target") or "")
    score = 0
    if _truthy(row.get("halo_only_conversion")) or _truthy(row.get("target_sku_not_converted")) or judgement in {
        "本 SKU 未验证",
        "待人工复查",
    }:
        score += 8
    if judgement in {"暂未改善", "初步有效", "有改善迹象"}:
        score += 5
    if judgement in {"待7天确认", "待人工判定有效/无效"}:
        score += 4
    if (
        "不要继续加价" in next_step
        or "不追加预算" in next_step
        or "不追加竞价" in next_step
        or "回到原竞价" in next_step
        or "保留当前竞价" in next_step
    ):
        score += 4
    if "优先要求补竞品/页面证据" in next_step or "确认是否否词匹配类型" in next_step:
        score += 3
    if window == "3天后复盘":
        score += 2
    days_number = _number(row.get("days_since_execution"))
    days = int(days_number) if days_number is not None else 0
    has_target_rank = 0 if str(row.get("search_term_or_target") or "") else 1
    return (-score, -days, has_target_rank, product)


def _summary_inventory_sort_key(row: dict[str, object]) -> tuple[int, float, float, str]:
    status = str(row.get("stock_status_label") or "")
    level = str(row.get("stock_risk_level") or "")
    coverage_source = row.get("days_of_cover")
    if coverage_source in ("", None):
        coverage_source = row.get("coverage_days")
    qty_source = row.get("recommended_reorder_qty")
    if qty_source in ("", None):
        qty_source = row.get("recommended_replenishment_qty")
    coverage = _number(coverage_source)
    qty = _number(qty_source) or 0
    if level == "OUT_OF_STOCK":
        priority = 0
    elif level == "LOW_STOCK" or "低库存" in status:
        priority = 1
    elif level == "REPLENISH_SOON" or "进入补货窗口" in status:
        priority = 2
    else:
        priority = 9
    return (priority, coverage if coverage is not None else 99999, -qty, str(row.get("product_name") or ""))


def _is_summary_inventory_row(row: dict[str, object]) -> bool:
    level = str(row.get("stock_risk_level") or "")
    status = str(row.get("stock_status_label") or "")
    if level in {"OUT_OF_STOCK", "LOW_STOCK", "REPLENISH_SOON"}:
        return True
    return not level and any(token in status for token in ["断货", "低库存", "进入补货窗口"])


def _summary_snapshot_failures(report_date: str, output_dir: Path) -> list[str]:
    analysis_path = output_dir / "latest_analysis.json"
    summary_path = output_dir / "summary.html"
    if not analysis_path.exists() or not summary_path.exists():
        return []
    try:
        payload = json.loads(analysis_path.read_text(encoding="utf-8"))
        text = summary_path.read_text(encoding="utf-8", errors="replace")
    except Exception as exc:
        return [f"summary.html snapshot rows cannot be validated: {exc}"]
    if not isinstance(payload, dict) or report_date not in text:
        return []
    action_review_rows: list[dict[str, object]] = []
    keyword_review_rows: list[dict[str, object]] = []
    inventory_rows: list[dict[str, object]] = []
    marketplace_results = [
        result for result in payload.get("marketplace_results") or [] if isinstance(result, dict)
    ]
    marketplace_results = sorted(marketplace_results, key=lambda result: _marketplace_sort_key(result.get("marketplace")))
    for result in marketplace_results:
        snapshot = result.get("report_view_snapshot") or {}
        if not isinstance(snapshot, dict):
            continue
        action_review_rows.extend(
            row for row in snapshot.get("action_effect_review_rows") or [] if isinstance(row, dict)
        )
        keyword_review_rows.extend(
            row for row in snapshot.get("keyword_action_effect_review_rows") or [] if isinstance(row, dict)
        )
        inventory_rows.extend(
            row for row in snapshot.get("inventory_replenishment_rows") or [] if isinstance(row, dict)
        )

    failures: list[str] = []
    for token in _frontend_coverage_display_tokens(payload):
        if not _html_contains_token(text, token):
            failures.append(f"summary.html missing frontend coverage token {token}")

    review_candidates = sorted([*keyword_review_rows, *action_review_rows], key=_summary_review_watch_sort_key)
    for idx, row in enumerate(review_candidates[:3], start=1):
        target = str(row.get("search_term_or_target") or row.get("product_name") or "").strip()
        if target and not _html_contains_token(text, target):
            failures.append(f"summary.html missing review watch row {idx} target {target}")
        if _summary_review_needs_display_guard(row):
            display_judgement = _summary_review_display_judgement(row)
            display_next_step = _summary_review_display_next_step(row)
            if target:
                if not _html_contains_token_near(text, target, display_judgement):
                    failures.append(
                        f"summary.html missing review watch row {idx} judgement {display_judgement}"
                    )
                if display_next_step and not _html_contains_token_near(text, target, display_next_step):
                    failures.append(
                        f"summary.html missing review watch row {idx} next step {display_next_step}"
                    )
            else:
                if display_judgement and not _html_contains_token(text, display_judgement):
                    failures.append(
                        f"summary.html missing review watch row {idx} judgement {display_judgement}"
                    )
                if display_next_step and not _html_contains_token(text, display_next_step):
                    failures.append(
                        f"summary.html missing review watch row {idx} next step {display_next_step}"
                    )
        for token in _summary_review_metric_tokens(row):
            if not _html_contains_token(text, token):
                failures.append(f"summary.html missing review watch row {idx} metric token {token}")
        for token in _summary_review_forbidden_early_metric_tokens(row):
            if target:
                if _html_contains_token_near(text, target, token):
                    failures.append(
                        f"summary.html early review row {idx} should not show metric token {token}"
                    )
            elif _html_contains_token(text, token):
                failures.append(f"summary.html early review row {idx} should not show metric token {token}")

    replenishment_rows = [
        row
        for row in sorted(inventory_rows, key=_summary_inventory_sort_key)
        if "暂不需要" not in str(row.get("replenishment_advice") or "")
        and _is_summary_inventory_row(row)
    ]
    for idx, row in enumerate(replenishment_rows[:3], start=1):
        product_name = str(row.get("product_name") or "").strip()
        status = str(row.get("stock_status_label") or "").strip()
        if product_name and not _html_contains_token(text, product_name):
            failures.append(f"summary.html missing replenishment row {idx} product name {product_name}")
        if status and not _html_contains_token(text, status):
            failures.append(f"summary.html missing replenishment row {idx} status {status}")
    return failures


def _latest_recommendations_snapshot_failures(report_date: str, output_dir: Path) -> list[str]:
    analysis_path = output_dir / "latest_analysis.json"
    page_path = output_dir / "latest_recommendations.html"
    if not analysis_path.exists() or not page_path.exists():
        return []
    try:
        payload = json.loads(analysis_path.read_text(encoding="utf-8"))
        text = page_path.read_text(encoding="utf-8", errors="replace")
    except Exception as exc:
        return [f"latest_recommendations.html snapshot rows cannot be validated: {exc}"]
    if not isinstance(payload, dict) or report_date not in text:
        return []

    product_rows: list[dict[str, object]] = []
    product_operation_rows: list[dict[str, object]] = []
    action_review_rows: list[dict[str, object]] = []
    keyword_review_rows: list[dict[str, object]] = []
    pending_ad_rows: list[dict[str, object]] = []
    tomorrow_review_rows: list[dict[str, object]] = []
    marketplace_results = [
        result for result in payload.get("marketplace_results") or [] if isinstance(result, dict)
    ]
    marketplace_results = sorted(marketplace_results, key=lambda result: _marketplace_sort_key(result.get("marketplace")))
    for result in marketplace_results:
        snapshot = result.get("report_view_snapshot") or {}
        if not isinstance(snapshot, dict):
            continue
        product_rows.extend(
            row for row in snapshot.get("product_final_decision_rows") or [] if isinstance(row, dict)
        )
        product_operation_rows.extend(
            row for row in snapshot.get("product_operation_cards") or [] if isinstance(row, dict)
        )
        action_review_rows.extend(
            row for row in snapshot.get("action_effect_review_rows") or [] if isinstance(row, dict)
        )
        keyword_review_rows.extend(
            row for row in snapshot.get("keyword_action_effect_review_rows") or [] if isinstance(row, dict)
        )
        tomorrow_review_rows.extend(
            row for row in snapshot.get("tomorrow_review_rows") or [] if isinstance(row, dict)
        )
        pending_ad_rows.extend(
            row
            for row in _ad_workbench_rows_from_snapshot(snapshot)
            if _is_pending_ad_workbench_row(row)
        )
    failures: list[str] = []
    payload_with_coverage = {
        "frontend_coverage_summary": _frontend_coverage_aggregate_from_snapshots(
            marketplace_results
        )
    }
    for token in _frontend_coverage_display_tokens(payload_with_coverage):
        if not _html_contains_token(text, token):
            failures.append(f"latest_recommendations.html missing frontend coverage token {token}")
    if pending_ad_rows and "复制到广告后台" not in text:
        failures.append("latest_recommendations.html missing copy area for pending ad rows")
    failures.extend(_html_ad_completion_payload_failures(text, pending_ad_rows))
    product_display_rows = product_operation_rows or product_rows
    product_display_rows = [
        row for row in product_display_rows if str(row.get("final_decision") or "") not in {"EXECUTE_TODAY"}
    ] or product_display_rows
    for idx, row in enumerate(product_display_rows[:4], start=1):
        marketplace = str(row.get("marketplace") or "").strip().upper()
        sku = str(row.get("sku") or "").strip()
        asin = str(row.get("asin") or "").strip().upper()
        product_name = str(row.get("product_name") or "").strip()
        contexts = _product_decision_contexts(text, asin, marketplace, sku, container="tr")
        if not contexts:
            contexts = _product_decision_contexts(text, asin, marketplace, sku, container="article")
        if asin and not _html_contains_token(text, asin):
            failures.append(f"latest_recommendations.html missing product decision row {idx} ASIN {asin}")
        elif asin and not contexts:
            failures.append(
                f"latest_recommendations.html missing structured product decision row {idx} {marketplace} {sku} {asin}"
            )
        failures.extend(
            _product_decision_contract_attr_failures(
                "latest_recommendations.html",
                row,
                contexts,
                row_label=f"product decision row {idx}",
            )
        )
        if product_name and not _html_contains_token(text, product_name):
            failures.append(f"latest_recommendations.html missing product decision row {idx} product name {product_name}")
        elif product_name and contexts and not _html_contains_token_in_contexts(contexts, product_name):
            failures.append(
                f"latest_recommendations.html product decision row {idx} product name not bound to ASIN {asin}"
            )
        for reason in _product_decision_blocking_reason_tokens(row)[:3]:
            if not _html_contains_token(text, reason):
                failures.append(
                    f"latest_recommendations.html missing product decision row {idx} blocking reason {reason}"
                )
            elif contexts and not _html_contains_token_in_contexts(contexts, reason):
                failures.append(
                    f"latest_recommendations.html product decision row {idx} blocking reason not bound to ASIN {asin}"
                )
        blocked_actions = _action_set(row.get("today_blocked_actions"))
        for action, labels in PRODUCT_DECISION_BLOCKED_ACTION_LABELS.items():
            if action in blocked_actions and not _html_contains_any_token(text, labels):
                failures.append(
                    f"latest_recommendations.html missing product decision row {idx} blocked action label {action}"
                )
            elif action in blocked_actions and contexts and not _html_contains_any_token_in_contexts(contexts, labels):
                failures.append(
                    f"latest_recommendations.html product decision row {idx} blocked action label {action} not bound to ASIN {asin}"
                )
        for label, token in _product_frontend_evidence_tokens(row):
            if not _html_contains_token(text, token):
                failures.append(
                    f"latest_recommendations.html missing product decision row {idx} {label} {token}"
                )
            elif contexts and not _html_contains_token_in_contexts(contexts, token):
                failures.append(
                    f"latest_recommendations.html product decision row {idx} {label} {token} not bound to ASIN {asin}"
                )
    for idx, row in enumerate(product_rows[:6], start=1):
        marketplace = str(row.get("marketplace") or "").strip().upper()
        sku = str(row.get("sku") or "").strip()
        asin = str(row.get("asin") or "").strip().upper()
        product_name = str(row.get("product_name") or "").strip()
        contexts = _product_decision_contexts(text, asin, marketplace, sku, container="article")
        if asin and not _html_contains_token(text, asin):
            failures.append(f"latest_recommendations.html missing product gate row {idx} ASIN {asin}")
        elif asin and not contexts:
            failures.append(
                f"latest_recommendations.html missing structured product gate row {idx} {marketplace} {sku} {asin}"
            )
        failures.extend(
            _product_decision_contract_attr_failures(
                "latest_recommendations.html",
                row,
                contexts,
                row_label=f"product gate row {idx}",
            )
        )
        if product_name and not _html_contains_token(text, product_name):
            failures.append(f"latest_recommendations.html missing product gate row {idx} product name {product_name}")
        elif product_name and contexts and not _html_contains_token_in_contexts(contexts, product_name):
            failures.append(f"latest_recommendations.html product gate row {idx} product name not bound to ASIN {asin}")
        for reason in _product_decision_blocking_reason_tokens(row)[:3]:
            if not _html_contains_token(text, reason):
                failures.append(f"latest_recommendations.html missing product gate row {idx} blocking reason {reason}")
            elif contexts and not _html_contains_token_in_contexts(contexts, reason):
                failures.append(f"latest_recommendations.html product gate row {idx} blocking reason not bound to ASIN {asin}")
        blocked_actions = _action_set(row.get("today_blocked_actions"))
        for action, labels in PRODUCT_DECISION_BLOCKED_ACTION_LABELS.items():
            if action in blocked_actions and not _html_contains_any_token(text, labels):
                failures.append(f"latest_recommendations.html missing product gate row {idx} blocked action label {action}")
            elif action in blocked_actions and contexts and not _html_contains_any_token_in_contexts(contexts, labels):
                failures.append(
                    f"latest_recommendations.html product gate row {idx} blocked action label {action} not bound to ASIN {asin}"
                )
        for label, token in _product_frontend_evidence_tokens(row):
            if not _html_contains_token(text, token):
                failures.append(f"latest_recommendations.html missing product gate row {idx} {label} {token}")
            elif contexts and not _html_contains_token_in_contexts(contexts, token):
                failures.append(
                    f"latest_recommendations.html product gate row {idx} {label} {token} not bound to ASIN {asin}"
                )

    for idx, row in enumerate(sorted(action_review_rows, key=_action_review_sort_key)[:5], start=1):
        asin = str(row.get("asin") or "").strip().upper()
        product_name = str(row.get("product_name") or "").strip()
        anchor = asin or product_name
        if asin and not _html_contains_token(text, asin):
            failures.append(f"latest_recommendations.html missing action review row {idx} ASIN {asin}")
        if product_name and not _html_contains_token(text, product_name):
            failures.append(f"latest_recommendations.html missing action review row {idx} product name {product_name}")
        elif asin and product_name and not _html_contains_token_near(text, asin, product_name):
            failures.append(
                f"latest_recommendations.html action review row {idx} product name not bound to ASIN {asin}"
            )
        for token in _review_html_metric_tokens(row):
            if not _html_contains_token(text, token):
                failures.append(f"latest_recommendations.html missing action review row {idx} metric token {token}")
            elif anchor and not _html_contains_token_near(text, anchor, token):
                failures.append(
                    f"latest_recommendations.html action review row {idx} metric token {token} not bound to {anchor}"
                )
        for token in _review_html_forbidden_metric_tokens(row):
            if anchor and _html_contains_token_near(text, anchor, token):
                failures.append(f"latest_recommendations.html action review row {idx} should not show metric token {token}")
            elif not anchor and _html_contains_token(text, token):
                failures.append(f"latest_recommendations.html action review row {idx} should not show metric token {token}")

    sorted_keyword_review_rows = sorted(keyword_review_rows, key=_keyword_review_sort_key)[:50]
    keyword_review_anchor_counts = Counter(
        str(row.get("search_term_or_target") or row.get("asin") or "").strip()
        for row in sorted_keyword_review_rows
        if str(row.get("search_term_or_target") or row.get("asin") or "").strip()
    )
    for idx, row in enumerate(sorted_keyword_review_rows, start=1):
        target = str(row.get("search_term_or_target") or "").strip()
        asin = str(row.get("asin") or "").strip().upper()
        action_id = str(row.get("action_id") or "").strip()
        anchor = target or asin
        anchor_is_unique = bool(anchor and keyword_review_anchor_counts.get(anchor, 0) == 1)
        row_contexts = _html_token_contexts(text, action_id) if action_id else []
        if target and not _html_contains_token(text, target):
            failures.append(f"latest_recommendations.html missing keyword review row {idx} target {target}")
        if asin and not _html_contains_token(text, asin):
            failures.append(f"latest_recommendations.html missing keyword review row {idx} ASIN {asin}")
        elif target and asin and not _html_contains_token_near(text, target, asin):
            failures.append(
                f"latest_recommendations.html keyword review row {idx} ASIN {asin} not bound to target {target}"
            )
        for token in _review_html_metric_tokens(row):
            if not _html_contains_token(text, token):
                failures.append(f"latest_recommendations.html missing keyword review row {idx} metric token {token}")
            elif row_contexts and not _html_contains_token_in_contexts(row_contexts, token):
                failures.append(
                    f"latest_recommendations.html keyword review row {idx} metric token {token} not bound to action_id {action_id}"
                )
            elif anchor and not _html_contains_token_near(text, anchor, token):
                failures.append(
                    f"latest_recommendations.html keyword review row {idx} metric token {token} not bound to {anchor}"
                )
        for token in _review_html_forbidden_metric_tokens(row):
            if row_contexts and _html_contains_token_in_contexts(row_contexts, token):
                failures.append(f"latest_recommendations.html keyword review row {idx} should not show metric token {token}")
            elif not action_id and anchor_is_unique and _html_contains_token_near(text, anchor, token):
                failures.append(f"latest_recommendations.html keyword review row {idx} should not show metric token {token}")
            elif not action_id and not anchor_is_unique and not anchor and _html_contains_token(text, token):
                failures.append(f"latest_recommendations.html keyword review row {idx} should not show metric token {token}")

    for idx, row in enumerate(tomorrow_review_rows[:50], start=1):
        identity = _tomorrow_review_identity_token(row)
        if identity is None:
            continue
        label, token = identity
        if not _html_contains_token(text, token):
            failures.append(f"latest_recommendations.html missing tomorrow review row {idx} {label} {token}")

    for idx, row in enumerate(pending_ad_rows[:50], start=1):
        identity = _pending_ad_identity_token(row)
        if identity is None:
            continue
        label, token = identity
        if not _html_contains_token(text, token):
            failures.append(f"latest_recommendations.html missing pending ad row {idx} {label} {token}")
        elif row.get("asin") and not _html_contains_token_near(text, token, row.get("asin")):
            failures.append(
                f"latest_recommendations.html pending ad row {idx} ASIN {str(row.get('asin')).strip().upper()} not bound to {label} {token}"
            )
    return failures


def _marketplace_report_snapshot_failures(report_date: str, output_dir: Path) -> list[str]:
    analysis_path = output_dir / "latest_analysis.json"
    if not analysis_path.exists():
        return []
    try:
        payload = json.loads(analysis_path.read_text(encoding="utf-8"))
    except Exception as exc:
        return [f"marketplace report snapshot rows cannot be validated: {exc}"]
    if not isinstance(payload, dict):
        return []
    failures: list[str] = []
    for marketplace in ["UK", "US", "DE"]:
        page_path = output_dir / f"{marketplace.lower()}_report.html"
        if not page_path.exists():
            continue
        try:
            text = page_path.read_text(encoding="utf-8", errors="replace")
        except Exception as exc:
            failures.append(f"{page_path.name} cannot be read for snapshot validation: {exc}")
            continue
        if report_date not in text:
            continue
        result = next(
            (
                item
                for item in payload.get("marketplace_results") or []
                if isinstance(item, dict) and str(item.get("marketplace") or "").upper() == marketplace
            ),
            None,
        )
        if not isinstance(result, dict):
            continue
        snapshot = result.get("report_view_snapshot") or {}
        if not isinstance(snapshot, dict):
            continue
        coverage = snapshot.get("frontend_coverage_summary") or {}
        if isinstance(coverage, dict):
            for token in _frontend_coverage_display_tokens({"frontend_coverage_summary": coverage}):
                if not _html_contains_token(text, token):
                    failures.append(f"{page_path.name} missing frontend coverage token {token}")
        action_rows = [
            row
            for row in snapshot.get("action_effect_review_rows") or []
            if isinstance(row, dict) and str(row.get("marketplace") or "").upper() == marketplace
        ]
        product_rows = [
            row
            for row in snapshot.get("product_final_decision_rows") or []
            if isinstance(row, dict) and str(row.get("marketplace") or "").upper() == marketplace
        ]
        keyword_rows = [
            row
            for row in snapshot.get("keyword_action_effect_review_rows") or []
            if isinstance(row, dict) and str(row.get("marketplace") or "").upper() == marketplace
        ]
        pending_ad_rows = [
            row
            for row in _ad_workbench_rows_from_snapshot(snapshot)
            if isinstance(row, dict)
            and str(row.get("marketplace") or marketplace).upper() == marketplace
            and _is_pending_ad_workbench_row(row)
        ]
        if pending_ad_rows and "复制到广告后台" not in text:
            failures.append(f"{page_path.name} missing copy area for pending ad rows")
        failures.extend(
            _html_ad_completion_payload_failures(
                text,
                pending_ad_rows,
                page_name=page_path.name,
            )
        )
        for idx, row in enumerate(product_rows[:6], start=1):
            sku = str(row.get("sku") or "").strip()
            asin = str(row.get("asin") or "").strip().upper()
            product_name = str(row.get("product_name") or "").strip()
            contexts = _product_decision_contexts(text, asin, marketplace, sku, container="article")
            if asin and not _html_contains_token(text, asin):
                failures.append(f"{page_path.name} missing product decision row {idx} ASIN {asin}")
            elif asin and not contexts:
                failures.append(f"{page_path.name} missing structured product decision row {idx} {marketplace} {sku} {asin}")
            failures.extend(
                _product_decision_contract_attr_failures(
                    page_path.name,
                    row,
                    contexts,
                    row_label=f"product decision row {idx}",
                )
            )
            if product_name and not _html_contains_token(text, product_name):
                failures.append(f"{page_path.name} missing product decision row {idx} product name {product_name}")
            elif product_name and contexts and not _html_contains_token_in_contexts(contexts, product_name):
                failures.append(f"{page_path.name} product decision row {idx} product name not bound to ASIN {asin}")
            for reason in _product_decision_blocking_reason_tokens(row)[:3]:
                if not _html_contains_token(text, reason):
                    failures.append(f"{page_path.name} missing product decision row {idx} blocking reason {reason}")
                elif contexts and not _html_contains_token_in_contexts(contexts, reason):
                    failures.append(f"{page_path.name} product decision row {idx} blocking reason not bound to ASIN {asin}")
            blocked_actions = _action_set(row.get("today_blocked_actions"))
            for action, labels in PRODUCT_DECISION_BLOCKED_ACTION_LABELS.items():
                if action in blocked_actions and not _html_contains_any_token(text, labels):
                    failures.append(f"{page_path.name} missing product decision row {idx} blocked action label {action}")
                elif action in blocked_actions and contexts and not _html_contains_any_token_in_contexts(contexts, labels):
                    failures.append(f"{page_path.name} product decision row {idx} blocked action label {action} not bound to ASIN {asin}")
            for label, token in _product_frontend_evidence_tokens(row):
                if not _html_contains_token(text, token):
                    failures.append(f"{page_path.name} missing product decision row {idx} {label} {token}")
                elif contexts and not _html_contains_token_in_contexts(contexts, token):
                    failures.append(f"{page_path.name} product decision row {idx} {label} {token} not bound to ASIN {asin}")
        for idx, row in enumerate(sorted(action_rows, key=_action_review_sort_key)[:3], start=1):
            asin = str(row.get("asin") or "").strip().upper()
            product_name = str(row.get("product_name") or "").strip()
            anchor = asin or product_name
            if asin and not _html_contains_token(text, asin):
                failures.append(f"{page_path.name} missing action review row {idx} ASIN {asin}")
            if product_name and not _html_contains_token(text, product_name):
                failures.append(f"{page_path.name} missing action review row {idx} product name {product_name}")
            elif asin and product_name and not _html_contains_token_near(text, asin, product_name):
                failures.append(f"{page_path.name} action review row {idx} product name not bound to ASIN {asin}")
            for token in _review_html_metric_tokens(row):
                if not _html_contains_token(text, token):
                    failures.append(f"{page_path.name} missing action review row {idx} metric token {token}")
                elif anchor and not _html_contains_token_near(text, anchor, token):
                    failures.append(f"{page_path.name} action review row {idx} metric token {token} not bound to {anchor}")
            for token in _review_html_forbidden_metric_tokens(row):
                if anchor and _html_contains_token_near(text, anchor, token):
                    failures.append(f"{page_path.name} action review row {idx} should not show metric token {token}")
                elif not anchor and _html_contains_token(text, token):
                    failures.append(f"{page_path.name} action review row {idx} should not show metric token {token}")
        sorted_market_keyword_rows = sorted(keyword_rows, key=_keyword_review_sort_key)[:10]
        market_keyword_anchor_counts = Counter(
            str(row.get("search_term_or_target") or row.get("asin") or "").strip()
            for row in sorted_market_keyword_rows
            if str(row.get("search_term_or_target") or row.get("asin") or "").strip()
        )
        for idx, row in enumerate(sorted_market_keyword_rows, start=1):
            target = str(row.get("search_term_or_target") or "").strip()
            asin = str(row.get("asin") or "").strip().upper()
            action_id = str(row.get("action_id") or "").strip()
            anchor = target or asin
            anchor_is_unique = bool(anchor and market_keyword_anchor_counts.get(anchor, 0) == 1)
            row_contexts = _html_token_contexts(text, action_id) if action_id else []
            if target and not _html_contains_token(text, target):
                failures.append(f"{page_path.name} missing keyword review row {idx} target {target}")
            if asin and not _html_contains_token(text, asin):
                failures.append(f"{page_path.name} missing keyword review row {idx} ASIN {asin}")
            elif target and asin and not _html_contains_token_near(text, target, asin):
                failures.append(f"{page_path.name} keyword review row {idx} ASIN {asin} not bound to target {target}")
            for token in _review_html_metric_tokens(row):
                if not _html_contains_token(text, token):
                    failures.append(f"{page_path.name} missing keyword review row {idx} metric token {token}")
                elif row_contexts and not _html_contains_token_in_contexts(row_contexts, token):
                    failures.append(
                        f"{page_path.name} keyword review row {idx} metric token {token} not bound to action_id {action_id}"
                    )
                elif anchor and not _html_contains_token_near(text, anchor, token):
                    failures.append(f"{page_path.name} keyword review row {idx} metric token {token} not bound to {anchor}")
            for token in _review_html_forbidden_metric_tokens(row):
                if row_contexts and _html_contains_token_in_contexts(row_contexts, token):
                    failures.append(f"{page_path.name} keyword review row {idx} should not show metric token {token}")
                elif not action_id and anchor_is_unique and _html_contains_token_near(text, anchor, token):
                    failures.append(f"{page_path.name} keyword review row {idx} should not show metric token {token}")
                elif not action_id and not anchor_is_unique and not anchor and _html_contains_token(text, token):
                    failures.append(f"{page_path.name} keyword review row {idx} should not show metric token {token}")
    return failures


def _date_scoped_required_outputs(report_date: str, output_dir: Path) -> list[Path]:
    date_token = report_date.replace("-", "")
    return [
        output_dir / f"amazon_ops_report_{report_date}.xlsx",
        output_dir / f"autoopt_log_{date_token}.json",
        output_dir / f"autoopt_{date_token}.xlsx",
        output_dir / f"action_review_{date_token}.json",
        output_dir / f"keyword_action_review_{date_token}.json",
        output_dir / f"learned_rules_{date_token}.json",
        output_dir / f"manual_learning_log_{date_token}.json",
        output_dir / f"product_strategy_profiles_{date_token}.json",
        output_dir / f"keyword_strategy_memory_{date_token}.json",
        output_dir / f"self_optimization_log_{date_token}.json",
    ]


def _autoopt_log_content_failures(report_date: str, output_dir: Path) -> list[str]:
    date_token = report_date.replace("-", "")
    path = output_dir / f"autoopt_log_{date_token}.json"
    if not path.exists():
        return []
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        return [f"autoopt log cannot be read: {exc}"]
    if not isinstance(payload, dict):
        return [f"autoopt log root must be an object: {path}"]
    embedded_report_date = str(payload.get("report_date") or "").strip()
    if not embedded_report_date:
        return [f"autoopt log missing report_date: {path}"]
    if embedded_report_date != report_date:
        return [f"autoopt log report_date mismatch: expected {report_date}, got {embedded_report_date}"]
    failures: list[str] = []
    for payload_key, label in [
        ("action_review_rows", "action review"),
        ("keyword_action_review_rows", "keyword action review"),
    ]:
        rows = payload.get(payload_key, [])
        if rows is None:
            rows = []
        if not isinstance(rows, list):
            failures.append(f"autoopt log {payload_key} must be a list for review evidence validation")
            continue
        for idx, row in enumerate(rows, start=1):
            if not isinstance(row, dict):
                continue
            target = str(
                row.get("search_term_or_target")
                or row.get("product_name")
                or row.get("asin")
                or "unknown_review_row"
            )
            row_marketplace = str(row.get("marketplace") or "").strip().upper()
            if not row_marketplace:
                failures.append(f"autoopt log {label} row {idx} {target} missing marketplace value")
            elif row_marketplace not in VALID_MARKETPLACES:
                failures.append(
                    f"autoopt log {label} row {idx} {target} contains unsupported marketplace {row_marketplace}"
                )
            if "action_id" in row and not str(row.get("action_id") or "").strip():
                failures.append(f"autoopt log {label} row {idx} {target} missing action_id value")
            failures.extend(
                _action_identity_failure_for_row(
                    row,
                    f"autoopt log {label} row {idx} {target}",
                    row_marketplace,
                )
            )
        failures.extend(_review_effect_evidence_failures(rows, label, source="autoopt log"))
    failures.extend(_learning_adjustment_claim_failures(payload))
    return failures


def _feedback_review_trace_failures(report_date: str, output_dir: Path) -> list[str]:
    date_token = report_date.replace("-", "")
    autoopt_path = output_dir / f"autoopt_log_{date_token}.json"
    if not autoopt_path.exists():
        return []
    try:
        feedback_rows = load_feedback_input(output_dir)
        autoopt_payload = json.loads(autoopt_path.read_text(encoding="utf-8"))
    except Exception as exc:
        return [f"feedback review trace cannot be validated: {exc}"]
    if not isinstance(autoopt_payload, dict):
        return []
    product_review_ids = {
        str(row.get("action_id") or "").strip()
        for row in autoopt_payload.get("action_review_rows") or []
        if isinstance(row, dict)
    }
    keyword_review_ids = {
        str(row.get("action_id") or "").strip()
        for row in autoopt_payload.get("keyword_action_review_rows") or []
        if isinstance(row, dict)
    }
    failures: list[str] = []

    def check_rows(rows: list[object], source: str) -> None:
        for idx, row in enumerate(rows, start=1):
            if not isinstance(row, dict):
                continue
            if str(row.get("confirmed_status") or "").strip() != "已执行":
                continue
            if not is_executable_action(row):
                continue
            action_id = str(row.get("action_id") or "").strip()
            target = str(
                row.get("search_term_or_target")
                or row.get("product_name")
                or row.get("asin")
                or action_id
                or f"{source} row {idx}"
            )
            if not action_id:
                failures.append(f"{source} row {idx} {target} missing action_id for executed action trace")
                continue
            scope = str(row.get("action_scope") or "").strip()
            expects_keyword_review = bool(str(row.get("search_term_or_target") or "").strip()) or scope in {
                "search_term",
                "asin_target",
            }
            if expects_keyword_review:
                if action_id and action_id not in keyword_review_ids:
                    failures.append(
                        f"{source} row {idx} {target} missing keyword action review trace {action_id}"
                    )
            elif action_id and action_id not in product_review_ids:
                failures.append(f"{source} row {idx} {target} missing action review trace {action_id}")

    for idx, row in enumerate(feedback_rows, start=1):
        if not isinstance(row, dict):
            continue
        if str(row.get("confirmed_status") or "").strip() != "已执行":
            continue
        if not is_executable_action(row):
            continue
        action_id = str(row.get("action_id") or "").strip()
        target = str(
            row.get("search_term_or_target")
            or row.get("product_name")
            or row.get("asin")
            or action_id
            or f"feedback row {idx}"
        )
        scope = str(row.get("action_scope") or "").strip()
        if not action_id:
            failures.append(f"executed feedback row {idx} {target} missing action_id for executed action trace")
            continue
        failures.extend(
            _action_identity_failure_for_row(
                row,
                f"executed feedback row {idx} {target}",
                str(row.get("marketplace") or "").strip().upper(),
            )
        )
        expects_keyword_review = bool(str(row.get("search_term_or_target") or "").strip()) or scope in {
            "search_term",
            "asin_target",
        }
        if expects_keyword_review:
            if action_id and action_id not in keyword_review_ids:
                failures.append(
                    f"executed feedback row {idx} {target} missing keyword action review trace {action_id}"
                )
        elif action_id and action_id not in product_review_ids:
            failures.append(f"executed feedback row {idx} {target} missing action review trace {action_id}")
    rows = autoopt_payload.get("rows") or []
    if isinstance(rows, list):
        check_rows(rows, "autoopt log executed")
    return failures


def _autoopt_latest_analysis_consistency_failures(report_date: str, output_dir: Path) -> list[str]:
    date_token = report_date.replace("-", "")
    analysis_path = output_dir / "latest_analysis.json"
    autoopt_path = output_dir / f"autoopt_log_{date_token}.json"
    if not analysis_path.exists() or not autoopt_path.exists():
        return []
    try:
        analysis_payload = json.loads(analysis_path.read_text(encoding="utf-8"))
        autoopt_payload = json.loads(autoopt_path.read_text(encoding="utf-8"))
    except Exception:
        return []
    if not isinstance(analysis_payload, dict) or not isinstance(autoopt_payload, dict):
        return []
    failures: list[str] = []
    for actual_payload_key, spec in sorted(AUTOOPT_LATEST_ANALYSIS_CONSISTENCY_SPECS.items()):
        expected_payload_key = str(spec["expected_payload_key"])
        expected_rows = analysis_payload.get(expected_payload_key, [])
        actual_rows = autoopt_payload.get(actual_payload_key, [])
        if expected_rows is None:
            expected_rows = []
        if actual_rows is None:
            actual_rows = []
        if not isinstance(expected_rows, list):
            failures.append(f"latest analysis {expected_payload_key} must be a list for autoopt consistency")
            continue
        if not isinstance(actual_rows, list):
            failures.append(f"autoopt log {actual_payload_key} must be a list for latest analysis consistency")
            continue
        failures.extend(
            _json_rows_match_failures(
                expected_rows=expected_rows,
                actual_rows=actual_rows,
                identity_fields=list(spec["identity_fields"]),
                content_fields=list(spec["content_fields"]),
                label=str(spec["label"]),
                expected_source=f"latest analysis {expected_payload_key}",
                actual_source=f"autoopt log {actual_payload_key}",
            )
        )
    marketplace_results = analysis_payload.get("marketplace_results", [])
    if isinstance(marketplace_results, list):
        for actual_payload_key, spec in sorted(AUTOOPT_LATEST_ANALYSIS_SNAPSHOT_CONSISTENCY_SPECS.items()):
            snapshot_key = str(spec["snapshot_key"])
            expected_rows: list[object] = []
            for result in marketplace_results:
                if not isinstance(result, dict):
                    continue
                snapshot = result.get("report_view_snapshot")
                if not isinstance(snapshot, dict):
                    continue
                snapshot_rows = snapshot.get(snapshot_key, [])
                if snapshot_rows is None:
                    snapshot_rows = []
                if not isinstance(snapshot_rows, list):
                    failures.append(f"latest analysis marketplace snapshot {snapshot_key} must be a list for autoopt consistency")
                    continue
                expected_rows.extend(snapshot_rows)
            actual_rows = autoopt_payload.get(actual_payload_key, [])
            if actual_rows is None:
                actual_rows = []
            if not isinstance(actual_rows, list):
                failures.append(f"autoopt log {actual_payload_key} must be a list for latest analysis snapshot consistency")
                continue
            failures.extend(
                _json_rows_match_failures(
                    expected_rows=expected_rows,
                    actual_rows=actual_rows,
                    identity_fields=list(spec["identity_fields"]),
                    content_fields=list(spec["content_fields"]),
                    label=str(spec["label"]),
                    expected_source=f"latest analysis marketplace snapshots {snapshot_key}",
                    actual_source=f"autoopt log {actual_payload_key}",
                )
            )
    for summary_key in AUTOOPT_LATEST_ANALYSIS_SUMMARY_KEYS:
        if summary_key not in autoopt_payload:
            continue
        expected_summary = _collapsed_marketplace_summary(analysis_payload.get(summary_key))
        actual_summary = _flat_summary(autoopt_payload.get(summary_key))
        if expected_summary is None:
            failures.append(f"latest analysis {summary_key} must be marketplace summary counts for autoopt consistency")
            continue
        if actual_summary is None:
            failures.append(f"autoopt log {summary_key} must be flat summary counts for latest analysis consistency")
            continue
        if actual_summary != expected_summary:
            failures.append(
                f"autoopt log {summary_key} mismatch vs collapsed latest analysis {summary_key}: "
                f"expected {dict(sorted(expected_summary.items()))}, got {dict(sorted(actual_summary.items()))}"
            )
    return failures


def _review_target_ratio(row: dict) -> float:
    target = _ratio_number(row.get("current_7d_target_acos") or row.get("target_acos") or row.get("suggested_target_acos"))
    if target is None or target <= 0:
        return 0.10
    return target


def _review_timing_failure(prefix: str, row: dict[str, object], days_since: float | None) -> str | None:
    if days_since is None:
        return None
    window = str(row.get("review_window") or "").strip()
    phase = str(row.get("review_phase") or "").strip()
    status = str(row.get("review_status") or "").strip()
    combined = " ".join(token for token in [window, phase, status] if token)
    if days_since < 3 and any(token in combined for token in ["3天后复盘", "3d_check", "day_3_check", "可做3天复查"]):
        return f"{prefix} claims 3-day review window before 3 days"
    if days_since < 7 and any(token in combined for token in ["7天后复盘", "day_7_review", "7d_review", "7d_check", "可做7天复查", "进入7天复盘"]):
        return f"{prefix} claims 7-day review window before 7 days"
    if days_since >= 7 and any(token in combined for token in ["未满3天", "under_3_days", "3天后复盘", "3d_check", "day_3_check"]):
        return f"{prefix} keeps early review window after 7 days"
    return None


def _review_metric_tracking_failures(
    prefix: str,
    row: dict[str, object],
    days_since: float | None,
    *,
    require_business_support: bool,
) -> list[str]:
    if days_since is None or days_since < 3:
        return []
    required_fields = [
        "current_7d_promoted_ad_orders",
        "current_7d_total_orders",
    ]
    if require_business_support:
        required_fields.extend(
            [
                "current_7d_acos",
                "current_7d_tacos",
                "current_7d_available_stock",
            ]
        )
    if days_since >= 7:
        required_fields.extend(["current_14d_promoted_ad_orders", "current_14d_total_orders"])
        if require_business_support:
            required_fields.extend(
                [
                    "current_14d_acos",
                    "current_14d_tacos",
                    "current_14d_available_stock",
                ]
            )
    return [
        f"{prefix} reached review window but missing numeric {field}"
        for field in required_fields
        if _number(row.get(field)) is None
    ]


def _review_metric_consistency_failures(prefix: str, row: dict[str, object], days_since: float | None) -> list[str]:
    if days_since is None or days_since < 3:
        return []
    failures: list[str] = []
    promoted_orders = _number(row.get("current_7d_promoted_ad_orders"))
    ad_orders = _number(row.get("current_7d_ad_orders"))
    if promoted_orders is not None and ad_orders is not None and promoted_orders > ad_orders:
        failures.append(f"{prefix} review metrics have promoted SKU orders above ad orders")
    total_orders = _number(row.get("current_7d_total_orders"))
    if promoted_orders is not None and total_orders is not None and promoted_orders > total_orders:
        failures.append(f"{prefix} review metrics have promoted SKU orders above total orders")
    if days_since < 7:
        return failures
    promoted_14d = _number(row.get("current_14d_promoted_ad_orders"))
    if promoted_orders is not None and promoted_14d is not None and promoted_14d < promoted_orders:
        failures.append(f"{prefix} review metrics have 14-day promoted SKU orders below 7-day orders")
    ad_orders_14d = _number(row.get("current_14d_ad_orders"))
    if promoted_14d is not None and ad_orders_14d is not None and promoted_14d > ad_orders_14d:
        failures.append(f"{prefix} review metrics have 14-day promoted SKU orders above ad orders")
    total_orders_14d = _number(row.get("current_14d_total_orders"))
    if promoted_14d is not None and total_orders_14d is not None and promoted_14d > total_orders_14d:
        failures.append(f"{prefix} review metrics have 14-day promoted SKU orders above total orders")
    return failures


def _review_anchor_failures(prefix: str, row: dict[str, object], *, effective_claim: bool) -> list[str]:
    if not effective_claim:
        return []
    failures: list[str] = []
    if str(row.get("review_data_source") or "").strip() != "execution_anchored_daily":
        failures.append(f"{prefix} effective outcome missing execution anchored daily source")
    required_fields = [
        "pre_7d_start",
        "pre_7d_end",
        "post_3d_start",
        "post_3d_end",
        "post_7d_start",
        "post_7d_end",
        "pre_7d_promoted_ad_orders",
        "pre_7d_total_orders",
        "pre_7d_tacos",
        "post_3d_days",
        "post_3d_promoted_ad_orders",
        "post_3d_total_orders",
        "post_3d_acos",
        "post_3d_tacos",
        "post_3d_available_stock",
        "post_7d_days",
        "post_7d_promoted_ad_orders",
        "post_7d_total_orders",
        "post_7d_acos",
        "post_7d_tacos",
        "post_7d_available_stock",
    ]
    for field in required_fields:
        if str(row.get(field) or "").strip() == "":
            failures.append(f"{prefix} effective outcome missing anchored field {field}")
    post_promoted = _number(row.get("post_7d_promoted_ad_orders"))
    current_promoted = _number(row.get("current_7d_promoted_ad_orders"))
    if post_promoted is None or post_promoted <= 0:
        failures.append(f"{prefix} effective outcome missing positive post 7-day promoted SKU orders")
    post_total = _number(row.get("post_7d_total_orders"))
    if post_total is None or post_total <= 0:
        failures.append(f"{prefix} effective outcome missing positive post 7-day total orders")
    if post_total is not None and post_promoted is not None and post_total < post_promoted:
        failures.append(f"{prefix} effective outcome has post 7-day promoted SKU orders above total orders")
    if _ratio_number(row.get("post_7d_acos")) is None:
        failures.append(f"{prefix} effective outcome missing post 7-day ACOS")
    if _ratio_number(row.get("post_7d_tacos")) is None:
        failures.append(f"{prefix} effective outcome missing post 7-day TACOS")
    post_stock = _number(row.get("post_7d_available_stock"))
    if post_stock is None or post_stock <= 0:
        failures.append(f"{prefix} effective outcome missing positive post 7-day available stock")
    if current_promoted is not None and post_promoted is not None and post_promoted != current_promoted:
        failures.append(f"{prefix} effective outcome current 7-day promoted SKU orders differ from post 7-day orders")
    return failures


def _positive_review_policy_claim(row: dict[str, object]) -> str:
    text = " ".join(
        str(row.get(field) or "")
        for field in ["rule_adjustment", "recommended_future_policy", "next_step", "learning_note"]
    )
    compact = re.sub(r"\s+", "", text)
    if not compact:
        return ""
    negated_phrases = [
        "不能把该动作判为可复用",
        "不作为该词/ASIN动作有效",
        "不继续加价或放量",
        "暂不继续加价或放量",
        "不要继续加价",
        "不能继续加价",
        "不因halo-only继续加价或放量",
        "不因光环成交继续放量",
        "今天不重复加价",
    ]
    for phrase in negated_phrases:
        compact = compact.replace(phrase, "")
    for token in [
        "可保留当前竞价",
        "保留当前竞价",
        "类似动作保留",
        "可复用",
        "继续加价",
        "继续放量",
        "keep_current_bid",
    ]:
        if token in compact:
            return token
    return ""


def _learning_adjustment_claim_failures(payload: dict[str, object]) -> list[str]:
    failures: list[str] = []
    strong_execution_tokens = ["前置", "默认推荐", "维持当前阈值"]
    for payload_key in ["rule_adjustments", "action_adjustments"]:
        rows = payload.get(payload_key, [])
        if rows is None:
            rows = []
        if not isinstance(rows, list):
            failures.append(f"autoopt log {payload_key} must be a list for learning adjustment validation")
            continue
        for idx, row in enumerate(rows, start=1):
            if not isinstance(row, dict):
                failures.append(f"autoopt log {payload_key} row {idx} must be an object")
                continue
            observed_status = str(row.get("observed_status") or "")
            suggestion = str(row.get("suggested_adjustment") or "")
            if "已执行" not in observed_status:
                continue
            compact = re.sub(r"\s+", "", suggestion)
            if not compact:
                continue
            strong_token = next((token for token in strong_execution_tokens if token in compact), "")
            if not strong_token:
                continue
            if "promotedSKU" in compact or "promoted_sku" in compact.lower():
                continue
            scope = str(row.get("rule_scope") or "unknown_scope")
            failures.append(
                f"autoopt log {payload_key} row {idx} {scope} treats execution rate as effectiveness via {strong_token}"
            )
    return failures


def _review_effect_evidence_failures(rows: list[object], label: str, *, source: str) -> list[str]:
    failures: list[str] = []
    required_numeric_fields = [
        "current_7d_ad_orders",
        "current_7d_acos",
        "current_7d_tacos",
        "current_7d_total_orders",
        "current_7d_available_stock",
        "current_14d_ad_orders",
        "current_14d_promoted_ad_orders",
        "current_14d_acos",
        "current_14d_total_orders",
        "current_14d_tacos",
        "current_14d_available_stock",
    ]
    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            failures.append(f"{source} {label} row {idx} must be an object for review evidence validation")
            continue
        target = str(
            row.get("search_term_or_target")
            or row.get("product_name")
            or row.get("asin")
            or "unknown_review_row"
        )
        prefix = f"{source} {label} row {idx} {target}"
        review_outcome = str(row.get("review_outcome") or "").strip()
        display_outcome = str(row.get("outcome") or row.get("judgement") or "").strip()
        days_since = _number(row.get("days_since_execution"))
        timing_failure = _review_timing_failure(prefix, row, days_since)
        if timing_failure:
            failures.append(timing_failure)
        if days_since is not None and 3 <= days_since < 7:
            evidence_text = str(row.get("effect_evidence") or row.get("current_evidence") or "")
            if "3天复查口径" not in evidence_text or "7天结论待补" not in evidence_text:
                failures.append(f"{prefix} 3-day review evidence missing early-window qualifier")
        effectiveness_score = _number(row.get("effectiveness_score"))
        downgraded_review_outcomes = {"needs_manual_review", "not_ready", "insufficient_sample"}
        review_is_downgraded = review_outcome in downgraded_review_outcomes
        negative_final_claim = review_outcome == "ineffective" or (
            effectiveness_score is not None
            and effectiveness_score < 0
            and review_outcome not in {"needs_manual_review", "not_ready"}
        )
        negative_display_claim = display_outcome in NEGATIVE_DISPLAY_OUTCOMES and not review_is_downgraded
        if negative_final_claim and (days_since is None or days_since < 7):
            failures.append(f"{prefix} marks ineffective before 7-day review window")
        if display_outcome in POSITIVE_DISPLAY_OUTCOMES and (days_since is None or days_since < 7):
            failures.append(f"{prefix} shows positive display outcome before 7-day review window")
        if display_outcome in NEGATIVE_DISPLAY_OUTCOMES and (days_since is None or days_since < 3):
            failures.append(f"{prefix} shows negative display outcome before 3-day review window")
        clicks7 = _number(row.get("current_7d_clicks"))
        spend7 = _number(row.get("current_7d_spend"))
        if (
            display_outcome in {"样本不足", "数据不足"}
            and days_since is not None
            and days_since >= 7
            and ((clicks7 is not None and clicks7 >= 8) or (spend7 is not None and spend7 >= 5))
        ):
            failures.append(f"{prefix} shows insufficient sample after 7-day sufficient traffic")
        effective_claim = (
            review_outcome == "effective"
            or (display_outcome in POSITIVE_DISPLAY_OUTCOMES and not review_is_downgraded)
            or (_positive_score(row.get("effectiveness_score")) and not review_is_downgraded)
            or (
                bool(_positive_review_policy_claim(row))
                and review_outcome not in {"needs_manual_review", "insufficient_sample"}
                and not (review_outcome == "not_ready" and (days_since is None or days_since < 7))
            )
        )
        final_claim = effective_claim or negative_final_claim or negative_display_claim
        if final_claim:
            failures.extend(
                _review_metric_tracking_failures(
                    prefix,
                    row,
                    days_since,
                    require_business_support=effective_claim,
                )
            )
        if effective_claim or negative_final_claim:
            failures.extend(_review_metric_consistency_failures(prefix, row, days_since))
        if (
            days_since is not None
            and days_since >= 7
            and (
                effective_claim
                or negative_final_claim
                or (display_outcome in POSITIVE_DISPLAY_OUTCOMES and not review_is_downgraded)
                or negative_display_claim
            )
            and (clicks7 is None or spend7 is None)
        ):
            failures.append(f"{prefix} final review outcome missing numeric current_7d_clicks or current_7d_spend")
        if not effective_claim:
            continue
        if days_since is None:
            failures.append(f"{prefix} effective outcome missing days_since_execution")
            continue
        if days_since < 7:
            failures.append(f"{prefix} marks effective before 7-day review window")
        failures.extend(_review_anchor_failures(prefix, row, effective_claim=effective_claim))
        if _truthy(row.get("halo_only_conversion")):
            failures.append(f"{prefix} marks halo-only conversion as effective")
        if _truthy(row.get("target_sku_not_converted")):
            failures.append(f"{prefix} marks target SKU not converted as effective")
        if not _truthy(row.get("promoted_conversion_improved")):
            failures.append(f"{prefix} effective outcome missing promoted SKU conversion")
        promoted_orders = _number(row.get("current_7d_promoted_ad_orders"))
        if promoted_orders is None or promoted_orders <= 0:
            failures.append(f"{prefix} effective outcome missing positive promoted SKU orders")
        for field in required_numeric_fields:
            if _number(row.get(field)) is None:
                failures.append(f"{prefix} effective outcome missing numeric {field}")
        ad_orders = _number(row.get("current_7d_ad_orders"))
        if promoted_orders is not None and ad_orders is not None and promoted_orders > ad_orders:
            failures.append(f"{prefix} effective outcome has promoted SKU orders above ad orders")
        promoted_14d = _number(row.get("current_14d_promoted_ad_orders"))
        if promoted_14d is not None and promoted_14d <= 0:
            failures.append(f"{prefix} effective outcome missing positive 14-day promoted SKU orders")
        if promoted_orders is not None and promoted_14d is not None and promoted_14d < promoted_orders:
            failures.append(f"{prefix} effective outcome has 14-day promoted SKU orders below 7-day orders")
        ad_orders_14d = _number(row.get("current_14d_ad_orders"))
        if promoted_14d is not None and ad_orders_14d is not None and promoted_14d > ad_orders_14d:
            failures.append(f"{prefix} effective outcome has 14-day promoted SKU orders above ad orders")
        total_orders = _number(row.get("current_7d_total_orders"))
        if total_orders is not None and total_orders <= 0:
            failures.append(f"{prefix} effective outcome missing positive total orders")
        if promoted_orders is not None and total_orders is not None and promoted_orders > total_orders:
            failures.append(f"{prefix} effective outcome has promoted SKU orders above total orders")
        total_orders_14d = _number(row.get("current_14d_total_orders"))
        if promoted_14d is not None and total_orders_14d is not None and promoted_14d > total_orders_14d:
            failures.append(f"{prefix} effective outcome has 14-day promoted SKU orders above total orders")
        stock = _number(row.get("current_7d_available_stock"))
        if stock is not None and stock <= 0:
            failures.append(f"{prefix} effective outcome missing positive available stock")
        target_ratio = _review_target_ratio(row)
        acos = _ratio_number(row.get("current_7d_acos"))
        if acos is not None and acos > target_ratio:
            failures.append(f"{prefix} effective outcome has ACOS above target")
        tacos = _ratio_number(row.get("current_7d_tacos"))
        if tacos is not None and tacos > target_ratio:
            failures.append(f"{prefix} effective outcome has TACOS above target")
    return failures


def _standalone_review_json_failures(report_date: str, output_dir: Path) -> list[str]:
    date_token = report_date.replace("-", "")
    autoopt_path = output_dir / f"autoopt_log_{date_token}.json"
    if not autoopt_path.exists():
        return []
    try:
        autoopt_payload = json.loads(autoopt_path.read_text(encoding="utf-8"))
    except Exception:
        return []
    if not isinstance(autoopt_payload, dict):
        return []
    review_files = {
        output_dir / f"action_review_{date_token}.json": ("action_review", "action_review_rows"),
        output_dir / f"keyword_action_review_{date_token}.json": (
            "keyword_action_review",
            "keyword_action_review_rows",
        ),
    }
    failures: list[str] = []
    for path, (review_type, payload_key) in sorted(review_files.items(), key=lambda item: item[0].name):
        if not path.exists():
            continue
        try:
            rows = json.loads(path.read_text(encoding="utf-8"))
        except Exception as exc:
            failures.append(f"{path.name} cannot be read: {exc}")
            continue
        if not isinstance(rows, list):
            failures.append(f"{path.name} root must be a list")
            continue
        expected_rows = autoopt_payload.get(payload_key, [])
        if expected_rows is None:
            expected_rows = []
        if not isinstance(expected_rows, list):
            failures.append(f"autoopt log {payload_key} must be a list for standalone review consistency")
            continue
        if len(rows) != len(expected_rows):
            failures.append(
                f"{path.name} row count mismatch: expected {len(expected_rows)} from {payload_key}, got {len(rows)}"
            )
        required_fields = STANDALONE_REVIEW_JSON_REQUIRED_FIELDS.get(review_type, [])
        identity_fields = STANDALONE_REVIEW_JSON_IDENTITY_FIELDS.get(review_type, [])
        expected_object_rows = [row for row in expected_rows if isinstance(row, dict)]
        actual_object_rows = [row for row in rows if isinstance(row, dict)]
        if len(expected_object_rows) != len(expected_rows):
            failures.append(f"autoopt log {payload_key} contains non-object rows")
        if identity_fields and len(expected_object_rows) == len(expected_rows):
            expected_identity = _identity_counter(expected_object_rows, identity_fields)
            actual_identity = _identity_counter(actual_object_rows, identity_fields)
            if actual_identity != expected_identity:
                failures.append(
                    f"{path.name} identity mismatch: "
                    f"expected {sorted(expected_identity.elements())}, got {sorted(actual_identity.elements())}"
                )
            else:
                expected_by_identity, expected_duplicate_failures = _rows_by_unique_identity(
                    expected_object_rows,
                    identity_fields,
                    label=review_type,
                    source="autoopt log",
                )
                actual_by_identity, actual_duplicate_failures = _rows_by_unique_identity(
                    actual_object_rows,
                    identity_fields,
                    label=review_type,
                    source=path.name,
                )
                failures.extend(expected_duplicate_failures)
                failures.extend(actual_duplicate_failures)
                if not expected_duplicate_failures and not actual_duplicate_failures:
                    for identity, expected_row in sorted(expected_by_identity.items()):
                        actual_row = actual_by_identity.get(identity)
                        if actual_row is None:
                            continue
                        for field in required_fields:
                            if field not in expected_row:
                                continue
                            expected_value = _canonical_report_value(expected_row.get(field))
                            actual_value = _canonical_report_value(actual_row.get(field))
                            if actual_value != expected_value:
                                failures.append(
                                    f"{path.name} field mismatch for {identity} field {field}: "
                                    f"expected {expected_value!r}, got {actual_value!r}"
                                )
        for idx, row in enumerate(rows, start=1):
            if not isinstance(row, dict):
                failures.append(f"{path.name} row {idx} must be an object")
                continue
            missing_fields = [field for field in required_fields if field not in row]
            if missing_fields:
                target = str(
                    row.get("search_term_or_target")
                    or row.get("product_name")
                    or row.get("asin")
                    or "unknown_target"
                )
                failures.append(
                    f"{path.name} row {idx} {target} missing fields: {', '.join(missing_fields)}"
                )
            if "action_id" in row and not str(row.get("action_id") or "").strip():
                target = str(
                    row.get("search_term_or_target")
                    or row.get("product_name")
                    or row.get("asin")
                    or "unknown_target"
                )
                failures.append(f"{path.name} row {idx} {target} missing action_id value")
            target = str(
                row.get("search_term_or_target")
                or row.get("product_name")
                or row.get("asin")
                or "unknown_target"
            )
            failures.extend(
                _action_identity_failure_for_row(
                    row,
                    f"{path.name} row {idx} {target}",
                    str(row.get("marketplace") or "").strip().upper(),
                )
            )
        failures.extend(_review_effect_evidence_failures(rows, review_type.replace("_", " "), source=path.name))
    return failures


def _standalone_learning_json_failures(report_date: str, output_dir: Path) -> list[str]:
    date_token = report_date.replace("-", "")
    autoopt_path = output_dir / f"autoopt_log_{date_token}.json"
    if not autoopt_path.exists():
        return []
    try:
        autoopt_payload = json.loads(autoopt_path.read_text(encoding="utf-8"))
    except Exception:
        return []
    if not isinstance(autoopt_payload, dict):
        return []
    sidecar_specs = {
        output_dir / f"learned_rules_{date_token}.json": "learned_rules",
        output_dir / f"manual_learning_log_{date_token}.json": "manual_learning_rows",
        output_dir / f"product_strategy_profiles_{date_token}.json": "product_strategy_profiles",
        output_dir / f"keyword_strategy_memory_{date_token}.json": "keyword_strategy_memory",
    }
    failures: list[str] = []
    for path, payload_key in sorted(sidecar_specs.items(), key=lambda item: item[0].name):
        if not path.exists():
            continue
        try:
            sidecar_payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception as exc:
            failures.append(f"{path.name} cannot be read: {exc}")
            continue
        expected_payload = autoopt_payload.get(payload_key, [])
        if expected_payload is None:
            expected_payload = []
        if not isinstance(expected_payload, list):
            failures.append(f"autoopt log {payload_key} must be a list for standalone learning consistency")
            continue
        if not isinstance(sidecar_payload, list):
            failures.append(f"{path.name} root must be a list")
            continue
        if _canonical_report_value(sidecar_payload) != _canonical_report_value(expected_payload):
            failures.append(
                f"{path.name} content mismatch: expected autoopt log {payload_key}, got stale or different payload"
            )

    self_opt_path = output_dir / f"self_optimization_log_{date_token}.json"
    if not self_opt_path.exists():
        return failures
    try:
        self_opt_payload = json.loads(self_opt_path.read_text(encoding="utf-8"))
    except Exception as exc:
        failures.append(f"{self_opt_path.name} cannot be read: {exc}")
        return failures
    if not isinstance(self_opt_payload, dict):
        failures.append(f"{self_opt_path.name} root must be an object")
        return failures
    embedded_report_date = str(self_opt_payload.get("report_date") or "").strip()
    if embedded_report_date != report_date:
        failures.append(
            f"{self_opt_path.name} report_date mismatch: expected {report_date}, got {embedded_report_date or '<missing>'}"
        )
    for payload_key in [
        "learned_rules",
        "manual_learning_rows",
        "action_review_rows",
        "keyword_action_review_rows",
        "positive_action_patterns",
        "negative_action_patterns",
        "product_strategy_profiles",
        "keyword_strategy_memory",
    ]:
        expected_payload = autoopt_payload.get(payload_key, [])
        actual_payload = self_opt_payload.get(payload_key, [])
        if expected_payload is None:
            expected_payload = []
        if actual_payload is None:
            actual_payload = []
        if not isinstance(expected_payload, list):
            failures.append(f"autoopt log {payload_key} must be a list for self optimization consistency")
            continue
        if not isinstance(actual_payload, list):
            failures.append(f"{self_opt_path.name} {payload_key} must be a list")
            continue
        if _canonical_report_value(actual_payload) != _canonical_report_value(expected_payload):
            failures.append(f"{self_opt_path.name} {payload_key} mismatch vs autoopt log")
    return failures


def _workbook_content_failures(path: Path, required_sheets: list[str], label: str) -> list[str]:
    if not path.exists():
        return []
    if path.stat().st_size <= 0:
        return [f"{label} is empty: {path}"]
    try:
        workbook = load_workbook(path, read_only=True)
    except Exception as exc:
        return [f"{label} cannot be opened: {exc}"]
    try:
        missing_sheets = [sheet for sheet in required_sheets if sheet not in workbook.sheetnames]
    finally:
        workbook.close()
    if missing_sheets:
        return [f"{label} missing sheets: {', '.join(missing_sheets)}"]
    return []


def _daily_excel_report_date_failures(path: Path, report_date: str) -> list[str]:
    if not path.exists():
        return []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return [f"daily Excel report cannot be opened for report_date validation: {exc}"]
    try:
        if "总览" not in workbook.sheetnames:
            return [f"daily Excel report missing 总览 sheet for report_date validation: {path}"]
        sheet = workbook["总览"]
        for row in sheet.iter_rows(values_only=True):
            values = [str(value or "").strip() for value in row]
            if "报告日期" not in values:
                continue
            if report_date in values:
                return []
            return [f"daily Excel report_date mismatch: expected {report_date}, got row {values}"]
    finally:
        workbook.close()
    return [f"daily Excel report missing 报告日期 row: {path}"]


def _daily_excel_frontend_coverage_failures(path: Path, output_dir: Path) -> list[str]:
    if not path.exists():
        return []
    analysis_path = output_dir / "latest_analysis.json"
    if not analysis_path.exists():
        return []
    try:
        payload = json.loads(analysis_path.read_text(encoding="utf-8"))
    except Exception as exc:
        return [f"daily Excel frontend coverage cannot read latest analysis: {exc}"]
    if not isinstance(payload, dict):
        return []
    expected = _frontend_coverage_excel_expected(payload)
    if not expected:
        return []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return [f"daily Excel report cannot be opened for frontend coverage validation: {exc}"]
    try:
        if "总览" not in workbook.sheetnames:
            return [f"daily Excel report missing 总览 sheet for frontend coverage validation: {path}"]
        actual: dict[str, str] = {}
        for row in workbook["总览"].iter_rows(values_only=True):
            values = [str(value or "").strip() for value in row]
            if len(values) < 3 or values[0] != "前台证据覆盖":
                continue
            actual[values[1]] = values[2]
    finally:
        workbook.close()
    failures: list[str] = []
    for metric, expected_value in expected.items():
        actual_value = actual.get(metric)
        if actual_value != expected_value:
            failures.append(
                f"daily Excel 总览 missing frontend coverage metric {metric}: "
                f"expected {expected_value}, got {actual_value or 'missing'}"
            )
    return failures


def _workbook_sheet_rows(workbook, sheet_name: str) -> list[dict[str, object]]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    parsed: list[dict[str, object]] = []
    for values in rows[1:]:
        row = {
            header: values[idx]
            for idx, header in enumerate(headers)
            if header and idx < len(values)
        }
        if any(str(value or "").strip() for value in row.values()):
            parsed.append(row)
    return parsed


def _workbook_sheet_headers(workbook, sheet_name: str) -> list[str]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(max_row=1, values_only=True))
    if not rows:
        return []
    return [str(value or "").strip() for value in rows[0] if str(value or "").strip()]


def _expected_excel_fields(
    expected_rows: list[dict[str, object]],
    identity_fields: list[str],
    content_fields: list[str],
) -> list[str]:
    fields: list[str] = []
    seen: set[str] = set()
    for field in [*identity_fields, *content_fields]:
        if field in seen:
            continue
        if any(field in row for row in expected_rows):
            fields.append(field)
            seen.add(field)
    return fields


def _excel_rows_match_failures(
    *,
    workbook,
    sheet_name: str,
    expected_rows: list[object],
    identity_fields: list[str],
    content_fields: list[str],
    label: str,
) -> list[str]:
    failures: list[str] = []
    expected_rows = expected_rows or []
    if sheet_name not in workbook.sheetnames:
        if expected_rows:
            failures.append(f"daily Excel report missing {sheet_name} sheet for {label} consistency")
        return failures
    expected_object_rows = [row for row in expected_rows if isinstance(row, dict)]
    if len(expected_object_rows) != len(expected_rows):
        failures.append(f"latest analysis contains non-object rows for {label}")
        return failures
    if expected_object_rows:
        headers = set(_workbook_sheet_headers(workbook, sheet_name))
        missing_headers = [
            field
            for field in _expected_excel_fields(expected_object_rows, identity_fields, content_fields)
            if field not in headers
        ]
        if missing_headers:
            failures.append(f"daily Excel {sheet_name} missing fields for {label}: {', '.join(missing_headers)}")
            return failures
    actual_rows = _workbook_sheet_rows(workbook, sheet_name)
    if len(actual_rows) != len(expected_rows):
        failures.append(
            f"daily Excel {sheet_name} row count mismatch for {label}: "
            f"expected {len(expected_rows)}, got {len(actual_rows)}"
        )
        return failures
    actual_object_rows = [row for row in actual_rows if isinstance(row, dict)]
    expected_identity = _identity_counter(expected_object_rows, identity_fields)
    actual_identity = _identity_counter(actual_object_rows, identity_fields)
    if actual_identity != expected_identity:
        failures.append(
            f"daily Excel {sheet_name} identity mismatch for {label}: "
            f"expected {sorted(expected_identity.elements())}, got {sorted(actual_identity.elements())}"
        )
        return failures
    expected_by_identity, expected_duplicate_failures = _rows_by_unique_identity(
        expected_object_rows,
        identity_fields,
        label=label,
        source="latest analysis",
    )
    actual_by_identity, actual_duplicate_failures = _rows_by_unique_identity(
        actual_object_rows,
        identity_fields,
        label=label,
        source=f"daily Excel {sheet_name}",
    )
    failures.extend(expected_duplicate_failures)
    failures.extend(actual_duplicate_failures)
    if expected_duplicate_failures or actual_duplicate_failures:
        return failures
    for identity, expected_row in sorted(expected_by_identity.items()):
        actual_row = actual_by_identity.get(identity)
        if actual_row is None:
            continue
        for field in content_fields:
            if field not in expected_row:
                continue
            expected_value = _canonical_report_value(expected_row.get(field))
            actual_value = _canonical_report_value(actual_row.get(field))
            if actual_value != expected_value:
                failures.append(
                    f"daily Excel {sheet_name} field mismatch for {label} "
                    f"{identity} field {field}: expected {expected_value!r}, got {actual_value!r}"
                )
    return failures


def _json_rows_match_failures(
    *,
    expected_rows: list[object],
    actual_rows: list[object],
    identity_fields: list[str],
    content_fields: list[str],
    label: str,
    expected_source: str,
    actual_source: str,
) -> list[str]:
    failures: list[str] = []
    if len(actual_rows) != len(expected_rows):
        failures.append(
            f"{actual_source} row count mismatch vs {expected_source} for {label}: "
            f"expected {len(expected_rows)}, got {len(actual_rows)}"
        )
        return failures
    expected_object_rows = [row for row in expected_rows if isinstance(row, dict)]
    actual_object_rows = [row for row in actual_rows if isinstance(row, dict)]
    if len(expected_object_rows) != len(expected_rows):
        failures.append(f"{expected_source} contains non-object rows for {label}")
        return failures
    if len(actual_object_rows) != len(actual_rows):
        failures.append(f"{actual_source} contains non-object rows for {label}")
        return failures
    expected_identity = _identity_counter(expected_object_rows, identity_fields)
    actual_identity = _identity_counter(actual_object_rows, identity_fields)
    if actual_identity != expected_identity:
        failures.append(
            f"{actual_source} identity mismatch vs {expected_source} for {label}: "
            f"expected {sorted(expected_identity.elements())}, got {sorted(actual_identity.elements())}"
        )
        return failures
    expected_by_identity, expected_duplicate_failures = _rows_by_unique_identity(
        expected_object_rows,
        identity_fields,
        label=label,
        source=expected_source,
    )
    actual_by_identity, actual_duplicate_failures = _rows_by_unique_identity(
        actual_object_rows,
        identity_fields,
        label=label,
        source=actual_source,
    )
    failures.extend(expected_duplicate_failures)
    failures.extend(actual_duplicate_failures)
    if expected_duplicate_failures or actual_duplicate_failures:
        return failures
    for identity, expected_row in sorted(expected_by_identity.items()):
        actual_row = actual_by_identity.get(identity)
        if actual_row is None:
            continue
        for field in content_fields:
            if field not in expected_row:
                continue
            expected_value = _canonical_report_value(expected_row.get(field))
            actual_value = _canonical_report_value(actual_row.get(field))
            if actual_value != expected_value:
                failures.append(
                    f"{actual_source} field mismatch vs {expected_source} for {label} "
                    f"{identity} field {field}: expected {expected_value!r}, got {actual_value!r}"
                )
    return failures


def _collapsed_marketplace_summary(value: object) -> dict[str, float] | None:
    if not isinstance(value, dict):
        return None
    collapsed: dict[str, float] = {}
    for marketplace_counts in value.values():
        if not isinstance(marketplace_counts, dict):
            return None
        for key, count in marketplace_counts.items():
            number = _number(count)
            if number is None:
                return None
            collapsed[str(key)] = collapsed.get(str(key), 0.0) + number
    return collapsed


def _flat_summary(value: object) -> dict[str, float] | None:
    if not isinstance(value, dict):
        return None
    flat: dict[str, float] = {}
    for key, count in value.items():
        number = _number(count)
        if number is None:
            return None
        flat[str(key)] = number
    return flat


def _daily_excel_latest_analysis_consistency_failures(report_date: str, output_dir: Path) -> list[str]:
    analysis_path = output_dir / "latest_analysis.json"
    workbook_path = output_dir / f"amazon_ops_report_{report_date}.xlsx"
    if not analysis_path.exists() or not workbook_path.exists():
        return []
    try:
        payload = json.loads(analysis_path.read_text(encoding="utf-8"))
    except Exception:
        return []
    if not isinstance(payload, dict):
        return []
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception:
        return []
    failures: list[str] = []
    try:
        required_empty_sheets = DAILY_EXCEL_REQUIRED_EMPTY_SHEETS + DAILY_EXCEL_REQUIRED_MARKET_EMPTY_SHEETS
        for sheet_name in required_empty_sheets:
            if sheet_name not in workbook.sheetnames:
                failures.append(f"daily Excel report missing required sheet even when empty: {sheet_name}")
        for payload_key, spec in sorted(DAILY_EXCEL_CONSISTENCY_SPECS.items()):
            sheet_name = str(spec["sheet_name"])
            identity_fields = list(spec["identity_fields"])
            label = str(spec["label"])
            content_fields = list(spec["content_fields"])
            expected_rows = payload.get(payload_key, [])
            if not isinstance(expected_rows, list):
                failures.append(f"latest analysis {payload_key} must be a list for daily Excel consistency")
                continue
            if sheet_name not in workbook.sheetnames:
                failures.append(f"daily Excel report missing {sheet_name} sheet for {label} consistency")
                continue
            expected_object_rows = [row for row in expected_rows if isinstance(row, dict)]
            if len(expected_object_rows) != len(expected_rows):
                failures.append(f"latest analysis contains non-object rows for {label}")
                continue
            if expected_object_rows:
                headers = set(_workbook_sheet_headers(workbook, sheet_name))
                missing_headers = [
                    field
                    for field in _expected_excel_fields(expected_object_rows, identity_fields, content_fields)
                    if field not in headers
                ]
                if missing_headers:
                    failures.append(f"daily Excel {sheet_name} missing fields for {label}: {', '.join(missing_headers)}")
                    continue
            actual_rows = _workbook_sheet_rows(workbook, sheet_name)
            if len(actual_rows) != len(expected_rows):
                failures.append(
                    f"daily Excel {sheet_name} row count mismatch for {label}: "
                    f"expected {len(expected_rows)}, got {len(actual_rows)}"
                )
                continue
            expected_identity = _identity_counter(expected_rows, identity_fields)
            actual_identity = _identity_counter(actual_rows, identity_fields)
            if actual_identity != expected_identity:
                failures.append(
                    f"daily Excel {sheet_name} identity mismatch for {label}: "
                    f"expected {sorted(expected_identity.elements())}, got {sorted(actual_identity.elements())}"
                )
                continue
            expected_by_identity, expected_duplicate_failures = _rows_by_unique_identity(
                expected_rows,
                identity_fields,
                label=label,
                source="latest analysis",
            )
            actual_by_identity, actual_duplicate_failures = _rows_by_unique_identity(
                actual_rows,
                identity_fields,
                label=label,
                source=f"daily Excel {sheet_name}",
            )
            failures.extend(expected_duplicate_failures)
            failures.extend(actual_duplicate_failures)
            if expected_duplicate_failures or actual_duplicate_failures:
                continue
            for identity, expected_row in sorted(expected_by_identity.items()):
                actual_row = actual_by_identity.get(identity)
                if actual_row is None:
                    continue
                for field in content_fields:
                    if field not in expected_row:
                        continue
                    expected_value = _canonical_report_value(expected_row.get(field))
                    actual_value = _canonical_report_value(actual_row.get(field))
                    if actual_value != expected_value:
                        failures.append(
                            f"daily Excel {sheet_name} field mismatch for {label} "
                            f"{identity} field {field}: expected {expected_value!r}, got {actual_value!r}"
                        )
        marketplace_results = payload.get("marketplace_results", [])
        if isinstance(marketplace_results, list):
            for snapshot_key, spec in sorted(DAILY_EXCEL_REVIEW_CONSISTENCY_SPECS.items()):
                sheet_name = str(spec["sheet_name"])
                identity_fields = list(spec["identity_fields"])
                label = str(spec["label"])
                content_fields = list(spec["content_fields"])
                market_sheet_suffix = str(spec["market_sheet_suffix"])
                all_expected_rows: list[object] = []
                market_expected_rows: dict[str, list[object]] = {}
                for result in marketplace_results:
                    if not isinstance(result, dict):
                        continue
                    marketplace = str(result.get("marketplace") or "").strip().upper()
                    snapshot = result.get("report_view_snapshot")
                    if not marketplace or not isinstance(snapshot, dict):
                        continue
                    rows = snapshot.get(snapshot_key, [])
                    if rows is None:
                        rows = []
                    if not isinstance(rows, list):
                        continue
                    all_expected_rows.extend(rows)
                    market_expected_rows.setdefault(marketplace, []).extend(rows)
                failures.extend(
                    _excel_rows_match_failures(
                        workbook=workbook,
                        sheet_name=sheet_name,
                        expected_rows=all_expected_rows,
                        identity_fields=identity_fields,
                        content_fields=content_fields,
                        label=label,
                    )
                )
                for marketplace, expected_rows in sorted(market_expected_rows.items()):
                    failures.extend(
                        _excel_rows_match_failures(
                            workbook=workbook,
                            sheet_name=f"{marketplace}_{market_sheet_suffix}",
                            expected_rows=expected_rows,
                            identity_fields=identity_fields,
                            content_fields=content_fields,
                            label=f"{marketplace} {label}",
                        )
                    )
    finally:
        workbook.close()
    return failures


def _autoopt_excel_report_date_failures(path: Path, report_date: str) -> list[str]:
    if not path.exists():
        return []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return [f"autoopt Excel report cannot be opened for report_date validation: {exc}"]
    try:
        if "summary" not in workbook.sheetnames:
            return [f"autoopt Excel report missing summary sheet for report_date validation: {path}"]
        sheet = workbook["summary"]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return [f"autoopt Excel report summary sheet is empty: {path}"]
    headers = [str(value or "").strip() for value in rows[0]]
    if "report_date" not in headers:
        return [f"autoopt Excel report summary missing report_date column: {path}"]
    report_date_index = headers.index("report_date")
    for row in rows[1:]:
        values = [str(value or "").strip() for value in row]
        embedded_report_date = values[report_date_index] if report_date_index < len(values) else ""
        if not embedded_report_date:
            continue
        if embedded_report_date == report_date:
            return []
        return [
            "autoopt Excel report_date mismatch: "
            f"expected {report_date}, got {embedded_report_date}"
        ]
    return [f"autoopt Excel report summary missing report_date value: {path}"]


def _sheet_data_row_count(sheet) -> int:
    rows = list(sheet.iter_rows(values_only=True))
    meaningful_rows = [row for row in rows if any(str(value or "").strip() for value in row)]
    if not meaningful_rows:
        return 0
    return max(0, len(meaningful_rows) - 1)


def _autoopt_excel_json_consistency_failures(report_date: str, output_dir: Path) -> list[str]:
    date_token = report_date.replace("-", "")
    json_path = output_dir / f"autoopt_log_{date_token}.json"
    xlsx_path = output_dir / f"autoopt_{date_token}.xlsx"
    if not json_path.exists() or not xlsx_path.exists():
        return []
    try:
        payload = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception:
        return []
    if not isinstance(payload, dict):
        return []
    try:
        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        return []
    failures: list[str] = []
    try:
        for sheet_name, payload_key in sorted(AUTOOPT_EXCEL_ROW_COUNT_SHEETS.items()):
            expected_rows = payload.get(payload_key, [])
            if expected_rows is None:
                expected_rows = []
            if not isinstance(expected_rows, list):
                failures.append(f"autoopt log {payload_key} must be a list for Excel consistency")
                continue
            if sheet_name not in workbook.sheetnames:
                continue
            actual_count = _sheet_data_row_count(workbook[sheet_name])
            expected_count = len(expected_rows)
            if actual_count != expected_count:
                failures.append(
                    f"autoopt Excel {sheet_name} row count mismatch: "
                    f"expected {expected_count} from {payload_key}, got {actual_count}"
                )
                continue
            spec = AUTOOPT_EXCEL_CONSISTENCY_SPECS.get(sheet_name)
            if not spec or not expected_rows:
                continue
            identity_fields = list(spec["identity_fields"])
            label = str(spec["label"])
            content_fields = list(spec["content_fields"])
            expected_object_rows = [row for row in expected_rows if isinstance(row, dict)]
            actual_rows = _workbook_sheet_rows(workbook, sheet_name)
            actual_object_rows = [row for row in actual_rows if isinstance(row, dict)]
            if len(expected_object_rows) != len(expected_rows):
                failures.append(f"autoopt log {payload_key} contains non-object rows")
                continue
            expected_identity = _identity_counter(expected_object_rows, identity_fields)
            actual_identity = _identity_counter(actual_object_rows, identity_fields)
            if actual_identity != expected_identity:
                failures.append(
                    f"autoopt Excel {sheet_name} identity mismatch for {label}: "
                    f"expected {sorted(expected_identity.elements())}, got {sorted(actual_identity.elements())}"
                )
                continue
            expected_by_identity, expected_duplicate_failures = _rows_by_unique_identity(
                expected_object_rows,
                identity_fields,
                label=label,
                source="autoopt log",
            )
            actual_by_identity, actual_duplicate_failures = _rows_by_unique_identity(
                actual_object_rows,
                identity_fields,
                label=label,
                source=f"autoopt Excel {sheet_name}",
            )
            failures.extend(expected_duplicate_failures)
            failures.extend(actual_duplicate_failures)
            if expected_duplicate_failures or actual_duplicate_failures:
                continue
            for identity, expected_row in sorted(expected_by_identity.items()):
                actual_row = actual_by_identity.get(identity)
                if actual_row is None:
                    continue
                for field in content_fields:
                    if field not in expected_row:
                        continue
                    expected_value = _canonical_report_value(expected_row.get(field))
                    actual_value = _canonical_report_value(actual_row.get(field))
                    if actual_value != expected_value:
                        failures.append(
                            f"autoopt Excel {sheet_name} field mismatch for {label} "
                            f"{identity} field {field}: expected {expected_value!r}, got {actual_value!r}"
                        )
    finally:
        workbook.close()
    return failures


def _enhanced_requests_excel_consistency_failures(output_dir: Path) -> list[str]:
    analysis_path = output_dir / "latest_analysis.json"
    xlsx_path = output_dir / "enhanced_data_requests.xlsx"
    if not analysis_path.exists() or not xlsx_path.exists():
        return []
    try:
        payload = json.loads(analysis_path.read_text(encoding="utf-8"))
    except Exception:
        return []
    if not isinstance(payload, dict):
        return []
    expected_rows = payload.get("enhanced_data_requests", [])
    if expected_rows is None:
        expected_rows = []
    if not isinstance(expected_rows, list):
        return ["latest analysis enhanced_data_requests must be a list for Excel consistency"]
    try:
        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as exc:
        return [f"enhanced data requests Excel cannot be opened for consistency validation: {exc}"]
    try:
        sheet_name = workbook.sheetnames[0] if workbook.sheetnames else ""
        actual_rows = _workbook_sheet_rows(workbook, sheet_name) if sheet_name else []
        return _json_rows_match_failures(
            expected_rows=expected_rows,
            actual_rows=actual_rows,
            identity_fields=ENHANCED_REQUEST_IDENTITY_FIELDS,
            content_fields=ENHANCED_REQUEST_CONTENT_FIELDS,
            label="enhanced data requests",
            expected_source="latest analysis",
            actual_source="enhanced data requests Excel",
        )
    finally:
        workbook.close()


def _deduped_enhanced_request_rows(rows: list[object]) -> list[dict[str, object]]:
    deduped: list[dict[str, object]] = []
    seen: set[tuple[object, ...]] = set()
    for row in rows:
        if not isinstance(row, dict):
            deduped.append(row)
            continue
        key = (
            row.get("marketplace"),
            row.get("report_type"),
            row.get("period"),
            row.get("expected_filename"),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def _enhanced_request_markdown_display_rows(rows: list[object]) -> list[object]:
    display_rows: list[object] = []
    for row in _deduped_enhanced_request_rows(rows):
        if not isinstance(row, dict):
            display_rows.append(row)
            continue
        required_value = str(row.get("required") or "").strip().lower()
        required_text = "是" if required_value in {"1", "true", "yes", "是"} else "否"
        display_rows.append(
            {
                "站点": row.get("marketplace") or "N/A",
                "状态": row.get("status") or "N/A",
                "报表类型": row.get("report_type") or "N/A",
                "周期": row.get("period") or "N/A",
                "日期范围": f"{row.get('start_date') or 'N/A'} ~ {row.get('end_date') or 'N/A'}",
                "导出后文件名": row.get("expected_filename") or "N/A",
                "目标文件夹": row.get("target_folder") or row.get("target_path") or "N/A",
                "必需": required_text,
            }
        )
    return display_rows


def _parse_markdown_table_rows(text: str, headers: list[str]) -> list[dict[str, object]]:
    parsed: list[dict[str, object]] = []
    in_target_table = False
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line.startswith("|") or not line.endswith("|"):
            if in_target_table:
                break
            continue
        cells = [cell.strip() for cell in line.strip("|").split("|")]
        if cells == headers:
            in_target_table = True
            continue
        if not in_target_table:
            continue
        if all(set(cell) <= {"-", ":"} for cell in cells):
            continue
        if len(cells) < len(headers):
            continue
        parsed.append({header: cells[idx] for idx, header in enumerate(headers)})
    return parsed


def _enhanced_requests_markdown_consistency_failures(output_dir: Path) -> list[str]:
    analysis_path = output_dir / "latest_analysis.json"
    markdown_path = output_dir / "enhanced_data_requests.md"
    if not analysis_path.exists() or not markdown_path.exists():
        return []
    try:
        payload = json.loads(analysis_path.read_text(encoding="utf-8"))
    except Exception:
        return []
    if not isinstance(payload, dict):
        return []
    expected_rows = payload.get("enhanced_data_requests", [])
    if expected_rows is None:
        expected_rows = []
    if not isinstance(expected_rows, list):
        return ["latest analysis enhanced_data_requests must be a list for Markdown consistency"]
    try:
        text = markdown_path.read_text(encoding="utf-8", errors="replace")
    except Exception as exc:
        return [f"enhanced data requests Markdown cannot be read for consistency validation: {exc}"]
    expected_display_rows = _enhanced_request_markdown_display_rows(expected_rows)
    actual_rows = _parse_markdown_table_rows(text, ENHANCED_REQUEST_MARKDOWN_HEADERS)
    if not expected_display_rows and not actual_rows:
        if "当前没有需要额外导出的增强数据请求" not in text:
            return ["enhanced data requests Markdown missing empty-state message"]
        return []
    return _json_rows_match_failures(
        expected_rows=expected_display_rows,
        actual_rows=actual_rows,
        identity_fields=ENHANCED_REQUEST_MARKDOWN_IDENTITY_FIELDS,
        content_fields=ENHANCED_REQUEST_MARKDOWN_CONTENT_FIELDS,
        label="enhanced data requests Markdown display",
        expected_source="latest analysis display",
        actual_source="enhanced data requests Markdown",
    )


def _date_scoped_workbook_failures(report_date: str, output_dir: Path) -> list[str]:
    date_token = report_date.replace("-", "")
    failures: list[str] = []
    daily_report_path = output_dir / f"amazon_ops_report_{report_date}.xlsx"
    daily_report_failures = _workbook_content_failures(
        daily_report_path,
        ["总览", "Metrics_Validation", "UK_今日总览", "US_今日总览", "DE_今日总览"],
        "daily Excel report",
    )
    failures.extend(daily_report_failures)
    if not daily_report_failures:
        failures.extend(_daily_excel_report_date_failures(daily_report_path, report_date))
        failures.extend(_daily_excel_frontend_coverage_failures(daily_report_path, output_dir))
        failures.extend(_daily_excel_latest_analysis_consistency_failures(report_date, output_dir))
    autoopt_excel_path = output_dir / f"autoopt_{date_token}.xlsx"
    autoopt_excel_failures = _workbook_content_failures(
        autoopt_excel_path,
        ["autoopt_log", "summary", "action_review", "keyword_action_review", "final_decisions"],
        "autoopt Excel report",
    )
    failures.extend(autoopt_excel_failures)
    if not autoopt_excel_failures:
        failures.extend(_autoopt_excel_report_date_failures(autoopt_excel_path, report_date))
        failures.extend(_autoopt_excel_json_consistency_failures(report_date, output_dir))
    return failures


def _dates_from_manifest_row(row: dict) -> list[date]:
    text = " ".join(
        str(row.get(field) or "")
        for field in ["detected_date_range", "original_filename", "target_path"]
    )
    dates: list[date] = []
    for token in DATE_RE.findall(text):
        try:
            dates.append(date.fromisoformat(token))
        except ValueError:
            continue
    return dates


def _suspicious_core_import_type(row: dict) -> str:
    text = " ".join(
        str(row.get(field) or "")
        for field in ["original_filename", "target_path"]
    )
    for pattern, expected_type in SUSPICIOUS_CORE_IMPORT_PATTERNS:
        if pattern.search(text):
            return expected_type
    return ""


def _manifest_core_import_date_failures(
    manifest_path: Path,
    *,
    report_date: str,
    require_successful_core_import: bool = False,
) -> list[str]:
    if not manifest_path.exists():
        return [f"import manifest missing for report date validation: {manifest_path}"]
    try:
        rows = json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception as exc:
        return [f"import manifest cannot be read for report date validation: {exc}"]
    if not isinstance(rows, list):
        return ["import manifest root must be a list for report date validation"]
    try:
        report_day = date.fromisoformat(report_date)
    except ValueError:
        return [f"latest analysis has invalid report_date: {report_date}"]

    imported_core_dates: list[date] = []
    failures: list[str] = []
    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            continue
        status = str(row.get("status") or "").strip()
        detected_type = str(row.get("detected_type") or "").strip()
        if status not in IMPORT_SUCCESS_STATUSES:
            continue
        label = str(row.get("original_filename") or row.get("target_path") or f"row {idx}").strip()
        if not detected_type:
            failures.append(f"import manifest successful row {idx} {label} missing detected_type")
            continue
        suspicious_core_type = _suspicious_core_import_type(row)
        if suspicious_core_type and detected_type != suspicious_core_type:
            failures.append(
                f"import manifest successful row {idx} {label} looks like "
                f"{suspicious_core_type} but detected_type is {detected_type}"
            )
            continue
        if detected_type not in CORE_IMPORT_TYPES:
            continue
        row_dates = _dates_from_manifest_row(row)
        if not row_dates:
            failures.append(f"import manifest successful core row {idx} {label} missing parseable date")
            continue
        imported_core_dates.extend(row_dates)
    if failures:
        return failures
    if not imported_core_dates:
        if require_successful_core_import:
            return ["import manifest has no successful core ads or ERP import rows for this daily update"]
        return []
    latest_core_date = max(imported_core_dates)
    if report_day < latest_core_date:
        return [
            "latest analysis report_date "
            f"{report_date} is older than imported core data ending {latest_core_date.isoformat()}"
        ]
    return []


def _output_refresh_snapshot(paths: list[Path] | None = None) -> dict[Path, int | None]:
    paths = paths or REQUIRED_REFRESHED_OUTPUTS
    tracked = set(paths)
    analysis_path = _analysis_path_from_required_outputs(paths)
    output_dir = analysis_path.parent
    for pattern in DATE_SCOPED_OUTPUT_PATTERNS:
        tracked.update(output_dir.glob(pattern))
    return {path: _mtime_ns(path) for path in tracked}


def output_content_snapshot(paths: list[Path] | None = None) -> dict[Path, bytes | None]:
    return {
        path: path.read_bytes() if path.exists() and path.is_file() else None
        for path in _output_refresh_snapshot(paths)
    }


def file_content_snapshot(paths: list[Path] | None = None) -> dict[Path, bytes | None]:
    return {
        path: path.read_bytes() if path.exists() and path.is_file() else None
        for path in (paths or DB_STATE_FILES)
    }


def archive_file_snapshot(paths: list[Path] | None = None) -> dict[Path, bytes]:
    files: dict[Path, bytes] = {}
    for root in paths or ARCHIVE_STATE_DIRS:
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if path.is_file():
                files[path] = path.read_bytes()
    return files


def report_state_snapshot(paths: list[Path] | None = None) -> dict[str, object]:
    return {
        "outputs": output_content_snapshot(paths),
        "state_files": file_content_snapshot(),
        "archive_files": archive_file_snapshot(),
    }


def restore_output_content_snapshot(snapshot: dict[Path, bytes | None]) -> list[str]:
    failures: list[str] = []
    for path in _output_refresh_snapshot(list(snapshot.keys())):
        if path in snapshot:
            continue
        try:
            if path.exists() and path.is_file():
                path.unlink()
        except Exception as exc:
            failures.append(f"{path}: {exc}")
    for path, content in snapshot.items():
        try:
            if content is None:
                if path.exists():
                    path.unlink()
                continue
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(content)
        except Exception as exc:
            failures.append(f"{path}: {exc}")
    return failures


def restore_file_content_snapshot(snapshot: dict[Path, bytes | None]) -> list[str]:
    failures: list[str] = []
    for path, content in snapshot.items():
        try:
            if content is None:
                if path.exists():
                    path.unlink()
                continue
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(content)
        except Exception as exc:
            failures.append(f"{path}: {exc}")
    return failures


def restore_archive_file_snapshot(snapshot: dict[Path, bytes]) -> list[str]:
    failures: list[str] = []
    current_files = set(archive_file_snapshot())
    snapshot_files = set(snapshot)
    for path in sorted(current_files.difference(snapshot_files)):
        try:
            path.unlink()
        except Exception as exc:
            failures.append(f"{path}: {exc}")
    for path, content in snapshot.items():
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(content)
        except Exception as exc:
            failures.append(f"{path}: {exc}")
    return failures


def restore_report_state_snapshot(snapshot: dict[str, object]) -> list[str]:
    failures: list[str] = []
    outputs = snapshot.get("outputs")
    state_files = snapshot.get("state_files")
    archive_files = snapshot.get("archive_files")
    if isinstance(outputs, dict):
        failures.extend(restore_output_content_snapshot(outputs))
    if isinstance(state_files, dict):
        failures.extend(restore_file_content_snapshot(state_files))
    if isinstance(archive_files, dict):
        failures.extend(restore_archive_file_snapshot(archive_files))
    return failures


def report_refresh_failures(
    paths: list[Path] | None = None,
    *,
    previous_mtimes_ns: dict[Path, int | None] | None = None,
    import_manifest_path: Path | None = None,
    require_successful_core_import: bool = False,
) -> list[str]:
    paths = list(paths or REQUIRED_REFRESHED_OUTPUTS)
    analysis_path = _analysis_path_from_required_outputs(paths)
    report_date, date_failures = _report_date_from_analysis(analysis_path)
    if report_date:
        paths.extend(_date_scoped_required_outputs(report_date, analysis_path.parent))
    failures = date_failures + output_refresh_failures(paths, previous_mtimes_ns=previous_mtimes_ns)
    failures.extend(_asset_content_failures(paths))
    if report_date:
        failures.extend(_latest_analysis_content_failures(analysis_path, report_date))
        failures.extend(_text_report_date_failures(paths, report_date))
        failures.extend(_dashboard_marketplace_summary_failures(report_date, analysis_path.parent))
        failures.extend(_marketplace_summary_markdown_failures(report_date, analysis_path.parent))
        failures.extend(_summary_snapshot_failures(report_date, analysis_path.parent))
        failures.extend(_latest_recommendations_snapshot_failures(report_date, analysis_path.parent))
        failures.extend(_marketplace_report_snapshot_failures(report_date, analysis_path.parent))
        failures.extend(_autoopt_log_content_failures(report_date, analysis_path.parent))
        failures.extend(_feedback_review_trace_failures(report_date, analysis_path.parent))
        failures.extend(_autoopt_latest_analysis_consistency_failures(report_date, analysis_path.parent))
        failures.extend(_standalone_review_json_failures(report_date, analysis_path.parent))
        failures.extend(_standalone_learning_json_failures(report_date, analysis_path.parent))
        failures.extend(_enhanced_requests_excel_consistency_failures(analysis_path.parent))
        failures.extend(_enhanced_requests_markdown_consistency_failures(analysis_path.parent))
        failures.extend(_date_scoped_workbook_failures(report_date, analysis_path.parent))
        if import_manifest_path is not None:
            failures.extend(
                _manifest_core_import_date_failures(
                    import_manifest_path,
                    report_date=report_date,
                    require_successful_core_import=require_successful_core_import,
                )
            )
    return failures


def _receipt_path_label(path: Path) -> str:
    try:
        return path.relative_to(OUTPUT_DIR).as_posix()
    except ValueError:
        try:
            return path.relative_to(ROOT).as_posix()
        except ValueError:
            return str(path)


def _file_receipt_fingerprint(path: Path) -> dict[str, object]:
    if not path.exists() or not path.is_file():
        return {
            "path": str(path),
            "exists": False,
            "size": 0,
            "mtime_ns": None,
            "sha256": "",
        }
    data = path.read_bytes()
    stat = path.stat()
    return {
        "path": str(path),
        "exists": True,
        "size": stat.st_size,
        "mtime_ns": stat.st_mtime_ns,
        "sha256": sha256(data).hexdigest(),
    }


def _import_manifest_receipt_summary(path: Path = IMPORT_MANIFEST_JSON) -> dict[str, object]:
    try:
        rows = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        rows = []
    status_counts: Counter[str] = Counter()
    success_rows = 0
    if isinstance(rows, list):
        for row in rows:
            if not isinstance(row, dict):
                continue
            status = str(row.get("status") or "<missing>").strip() or "<missing>"
            status_counts[status] += 1
            if status in IMPORT_SUCCESS_STATUSES:
                success_rows += 1
    return {
        "row_count": len(rows) if isinstance(rows, list) else 0,
        "success_row_count": success_rows,
        "status_counts": dict(sorted(status_counts.items())),
    }


def _daily_update_receipt_outputs(paths: list[Path] | None = None) -> dict[str, dict[str, object]]:
    return {
        _receipt_path_label(path): _file_receipt_fingerprint(path)
        for path in sorted(_output_refresh_snapshot(paths), key=lambda item: str(item))
    }


def clear_daily_update_validation_receipt(path: Path | None = None) -> None:
    receipt_path = path or DAILY_UPDATE_VALIDATION_RECEIPT
    if receipt_path.exists():
        receipt_path.unlink()


def write_daily_update_validation_receipt(
    *,
    report_date: str,
    output_paths: list[Path] | None = None,
    receipt_path: Path | None = None,
) -> None:
    receipt_path = receipt_path or DAILY_UPDATE_VALIDATION_RECEIPT
    payload = {
        "schema_version": DAILY_UPDATE_RECEIPT_SCHEMA_VERSION,
        "result": "passed",
        "completed_at": datetime.now().isoformat(timespec="seconds"),
        "report_date": report_date,
        "import_manifest": {
            "json": _file_receipt_fingerprint(IMPORT_MANIFEST_JSON),
            "xlsx": _file_receipt_fingerprint(IMPORT_MANIFEST_XLSX),
            **_import_manifest_receipt_summary(IMPORT_MANIFEST_JSON),
        },
        "frontend_results": {
            "generated_at": _frontend_results_generated_at(FRONTEND_RESULTS_JSON),
            "file": _file_receipt_fingerprint(FRONTEND_RESULTS_JSON),
        },
        "outputs": _daily_update_receipt_outputs(output_paths),
    }
    receipt_path.parent.mkdir(parents=True, exist_ok=True)
    receipt_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def import_manifest_failures(path: Path | None = None, *, previous_mtime_ns: int | None = None) -> list[str]:
    path = path or IMPORT_MANIFEST_JSON
    if not path.exists():
        return [f"import manifest missing: {path}"]
    if previous_mtime_ns is not None and _manifest_mtime_ns(path) == previous_mtime_ns:
        return [f"import manifest was not refreshed by current import step: {path}"]
    try:
        rows = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        return [f"import manifest cannot be read: {exc}"]
    if not isinstance(rows, list):
        return ["import manifest root must be a list"]

    failures: list[str] = []
    for row in rows:
        if not isinstance(row, dict):
            failures.append("import manifest contains non-object row")
            continue
        status = str(row.get("status") or "").strip()
        if status == "unknown" or status.startswith("error"):
            filename = str(row.get("original_filename") or row.get("original_path") or "unknown_file")
            reason = str(row.get("reason") or "")
            failures.append(f"{filename}: {status}{' | ' + reason if reason else ''}")
        elif status not in ALLOWED_IMPORT_STATUSES:
            filename = str(row.get("original_filename") or row.get("original_path") or "unknown_file")
            failures.append(f"{filename}: unexpected import status {status or '<missing>'}")
    return failures


def _resolve_manifest_file_path(value: object) -> Path | None:
    text = str(value or "").strip()
    if not text:
        return None
    path = Path(text)
    return path if path.is_absolute() else ROOT / path


def _parse_manifest_created_at(value: object) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None
    if parsed.tzinfo is not None:
        parsed = parsed.replace(tzinfo=None)
    return parsed


def _import_manifest_success_row_file_failures(
    rows: list[object],
    *,
    import_started_at: datetime | None = None,
) -> list[str]:
    if import_started_at is None:
        return []
    failures: list[str] = []
    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            continue
        status = str(row.get("status") or "").strip()
        if status not in IMPORT_SUCCESS_STATUSES:
            continue
        label = str(row.get("original_filename") or row.get("target_path") or f"row {idx}").strip()
        created_at = _parse_manifest_created_at(row.get("created_at"))
        if created_at is None:
            failures.append(f"import manifest successful row {idx} {label} missing valid created_at")
        elif created_at < import_started_at:
            failures.append(
                f"import manifest successful row {idx} {label} created_at {created_at.isoformat(sep=' ')} "
                f"is older than current import step started at {import_started_at.isoformat(sep=' ')}"
            )
        for field in ["target_path", "archive_path"]:
            resolved = _resolve_manifest_file_path(row.get(field))
            if resolved is None:
                failures.append(f"import manifest successful row {idx} {label} missing {field}")
            elif not resolved.exists():
                failures.append(f"import manifest successful row {idx} {label} {field} does not exist: {resolved}")
    return failures


def import_manifest_audit_failures(
    json_path: Path | None = None,
    xlsx_path: Path | None = None,
    *,
    previous_json_mtime_ns: int | None = None,
    previous_xlsx_mtime_ns: int | None = None,
    import_started_at: datetime | None = None,
) -> list[str]:
    json_path = json_path or IMPORT_MANIFEST_JSON
    xlsx_path = xlsx_path or IMPORT_MANIFEST_XLSX
    failures = import_manifest_failures(json_path, previous_mtime_ns=previous_json_mtime_ns)
    if failures:
        return failures
    try:
        rows = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception as exc:
        return [f"import manifest cannot be read: {exc}"]
    if not isinstance(rows, list):
        return ["import manifest root must be a list"]

    if not xlsx_path.exists():
        return [f"import manifest xlsx missing: {xlsx_path}"]
    if previous_xlsx_mtime_ns is not None and _mtime_ns(xlsx_path) == previous_xlsx_mtime_ns:
        return [f"import manifest xlsx was not refreshed by current import step: {xlsx_path}"]
    if xlsx_path.stat().st_size <= 0:
        return [f"import manifest xlsx is empty: {xlsx_path}"]
    try:
        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as exc:
        return [f"import manifest xlsx cannot be opened: {exc}"]
    try:
        if "import_manifest" not in workbook.sheetnames:
            return [f"import manifest xlsx missing sheet import_manifest: {xlsx_path}"]
        sheet = workbook["import_manifest"]
        sheet_rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not sheet_rows:
        return [f"import manifest xlsx has no header row: {xlsx_path}"]
    headers = [str(value or "").strip() for value in sheet_rows[0]]
    required_headers = {"original_filename", "status", "created_at", "target_path", "archive_path"}
    missing_headers = sorted(header for header in required_headers if header not in headers)
    if missing_headers:
        return [f"import manifest xlsx missing headers: {', '.join(missing_headers)}"]
    data_row_count = max(0, len(sheet_rows) - 1)
    if data_row_count != len(rows):
        return [f"import manifest xlsx row count mismatch: expected {len(rows)}, got {data_row_count}"]
    actual_rows = []
    for values in sheet_rows[1:]:
        row = {
            header: values[idx]
            for idx, header in enumerate(headers)
            if header and idx < len(values)
        }
        if any(str(value or "").strip() for value in row.values()):
            actual_rows.append(row)
    consistency_failures = _json_rows_match_failures(
        expected_rows=rows,
        actual_rows=actual_rows,
        identity_fields=["original_filename", "target_path", "archive_path"],
        content_fields=["original_filename", "status", "created_at", "target_path", "archive_path"],
        label="import manifest rows",
        expected_source="import manifest json",
        actual_source="import manifest xlsx",
    )
    if consistency_failures:
        return consistency_failures
    return _import_manifest_success_row_file_failures(rows, import_started_at=import_started_at)


def main() -> int:
    clear_daily_update_validation_receipt()
    python = sys.executable
    pre_import_steps = [
        [python, "scripts/check_daily_update_preflight.py", "--allow-inbox-business-files"],
        [python, "scripts/import_inbox_files.py"],
    ]
    post_import_steps = [
        [python, "main.py", "--marketplace", "ALL"],
    ]
    import_started_at: datetime | None = None
    for step in pre_import_steps:
        previous_manifest_mtime = _manifest_mtime_ns(IMPORT_MANIFEST_JSON) if step[1] == "scripts/import_inbox_files.py" else None
        previous_manifest_xlsx_mtime = _mtime_ns(IMPORT_MANIFEST_XLSX) if step[1] == "scripts/import_inbox_files.py" else None
        if step[1] == "scripts/import_inbox_files.py":
            import_started_at = datetime.now().replace(microsecond=0)
        code = run_step(step)
        if code != 0:
            return code
    manifest_failures = import_manifest_audit_failures(
        previous_json_mtime_ns=previous_manifest_mtime,
        previous_xlsx_mtime_ns=previous_manifest_xlsx_mtime,
        import_started_at=import_started_at,
    )
    if manifest_failures:
        for failure in manifest_failures:
            print(f"[fail] import manifest blocker: {failure}", flush=True)
        return 1
    pre_report_state_snapshot = report_state_snapshot()
    previous_output_mtimes = _output_refresh_snapshot()
    for step in post_import_steps:
        code = run_step(step)
        if code != 0:
            restore_failures = restore_report_state_snapshot(pre_report_state_snapshot)
            if restore_failures:
                for failure in restore_failures:
                    print(f"[fail] state restore blocker: {failure}", flush=True)
            else:
                print(
                    "[restore] report outputs restored to pre-report snapshot after failure; database/archive state restored when tracked",
                    flush=True,
            )
            return code
    refresh_failures = report_refresh_failures(
        previous_mtimes_ns=previous_output_mtimes,
        import_manifest_path=IMPORT_MANIFEST_JSON,
    )
    if refresh_failures:
        for failure in refresh_failures:
            print(f"[fail] report refresh blocker: {failure}", flush=True)
        restore_failures = restore_report_state_snapshot(pre_report_state_snapshot)
        if restore_failures:
            for failure in restore_failures:
                print(f"[fail] state restore blocker: {failure}", flush=True)
        else:
            print(
                "[restore] report outputs restored to pre-report snapshot after failure; database/archive state restored when tracked",
                flush=True,
            )
        return 1
    report_date, date_failures = _report_date_from_analysis(_analysis_path_from_required_outputs(REQUIRED_REFRESHED_OUTPUTS))
    if date_failures:
        for failure in date_failures:
            print(f"[fail] receipt blocker: {failure}", flush=True)
        return 1
    write_daily_update_validation_receipt(
        report_date=report_date,
        output_paths=REQUIRED_REFRESHED_OUTPUTS,
    )
    output = OUTPUT_DIR
    print("[done] daily update completed with inbox import and cached market evidence", flush=True)
    print(f"[open] {output / 'latest_recommendations.html'}", flush=True)
    print(f"[open] {output / 'dashboard.html'}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

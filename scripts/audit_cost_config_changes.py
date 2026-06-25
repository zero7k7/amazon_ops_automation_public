from __future__ import annotations

import json
import subprocess
import sys
import argparse
from collections import Counter, defaultdict
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "config" / "product_cost_config.xlsx"
OUTPUT_PATH = ROOT / "data" / "output" / "cost_config_diff_summary.json"
MARKDOWN_OUTPUT_PATH = ROOT / "data" / "output" / "cost_config_diff_summary.md"
KEY_HEADERS = {"marketplace", "sku", "asin", "product_name"}
REVIEW_FIELDS = [
    "selling_price",
    "purchase_cost_rmb",
    "purchase_cost_local",
    "first_leg_cost_rmb",
    "first_leg_cost_local",
    "landed_cost_excl_amazon",
    "amazon_fees_excl_ads",
    "total_cost_before_ads",
    "profit_before_ads",
    "profit_after_10pct_ads",
    "break_even_acos",
    "suggested_target_acos",
    "current_inventory",
    "sea_inventory",
    "inventory_note",
]
SENSITIVE_HEADERS = {
    "selling_price",
    "target_acos",
    "purchase_cost_rmb",
    "purchase_cost_local",
    "first_leg_cost_rmb",
    "first_leg_cost_local",
    "shipping_cost",
    "handling_fee",
    "profit_before_ads_per_unit",
    "profit_before_ads",
    "current_inventory",
    "sea_inventory",
    "inventory_note",
}


def _head_file_bytes(path: str) -> bytes:
    return subprocess.check_output(["git", "show", f"HEAD:{path}"], cwd=ROOT)


def _normalize(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float):
        return round(value, 10)
    return value


def _display(value: Any, limit: int = 120) -> str:
    text = str(value if value is not None else "")
    if len(text) > limit:
        return text[: limit - 3].rstrip() + "..."
    return text


def _used_bounds(ws) -> tuple[int, int]:
    max_row = ws.max_row or 0
    max_column = ws.max_column or 0
    last_row = 0
    last_column = 0
    for row in range(1, max_row + 1):
        row_has_value = False
        for column in range(1, max_column + 1):
            value = ws.cell(row, column).value
            if _normalize(value) == "":
                continue
            row_has_value = True
            last_column = max(last_column, column)
        if row_has_value:
            last_row = row
    return last_row, last_column


def _headers(ws) -> dict[int, str]:
    headers: dict[int, str] = {}
    _, max_column = _used_bounds(ws)
    for column in range(1, max_column + 1):
        value = ws.cell(1, column).value
        header = str(value or f"column_{column}").strip()
        headers[column] = header
    return headers


def _row_identity(ws, row_index: int, headers: dict[int, str]) -> dict[str, str]:
    normalized_headers = {column: header.strip().lower() for column, header in headers.items()}
    identity: dict[str, str] = {}
    for column, header in normalized_headers.items():
        if header in KEY_HEADERS:
            identity[header] = _display(ws.cell(row_index, column).value, limit=80)
    return identity


def _compare_sheet(head_ws, current_ws) -> dict[str, Any]:
    head_headers = _headers(head_ws)
    current_headers = _headers(current_ws)
    head_max_row, head_max_column = _used_bounds(head_ws)
    current_max_row, current_max_column = _used_bounds(current_ws)
    max_row = max(head_max_row, current_max_row)
    max_column = max(head_max_column, current_max_column)
    changed_cells: list[dict[str, Any]] = []
    by_column: Counter[str] = Counter()
    by_marketplace: Counter[str] = Counter()
    sensitive_changes: list[dict[str, Any]] = []

    for row_index in range(1, max_row + 1):
        identity = _row_identity(current_ws, row_index, current_headers) or _row_identity(head_ws, row_index, head_headers)
        marketplace = identity.get("marketplace") or ""
        for column in range(1, max_column + 1):
            before = head_ws.cell(row_index, column).value if row_index <= head_max_row and column <= head_max_column else None
            after = current_ws.cell(row_index, column).value if row_index <= current_max_row and column <= current_max_column else None
            if _normalize(before) == _normalize(after):
                continue
            header = current_headers.get(column) or head_headers.get(column) or f"column_{column}"
            by_column[header] += 1
            if marketplace:
                by_marketplace[marketplace] += 1
            record = {
                "row": row_index,
                "column": column,
                "header": header,
                "before": _display(before),
                "after": _display(after),
                "identity": identity,
            }
            if len(changed_cells) < 50:
                changed_cells.append(record)
            if str(header).strip().lower() in SENSITIVE_HEADERS and len(sensitive_changes) < 80:
                sensitive_changes.append(record)

    result = {
        "changed_cell_count": sum(by_column.values()),
        "changed_columns": dict(by_column.most_common()),
        "changed_marketplaces": dict(by_marketplace.most_common()),
        "sample_changes": changed_cells,
        "sensitive_changes": sensitive_changes,
    }
    key_result = _compare_sheet_by_business_key(head_ws, current_ws)
    if key_result:
        result["keyed_diff"] = key_result
    return result


def _sheet_rows_by_key(ws) -> tuple[dict[str, dict[str, Any]], list[str]]:
    headers = _headers(ws)
    header_lookup = {str(header).strip().lower(): column for column, header in headers.items()}
    required = ["marketplace", "sku", "asin"]
    if not all(header in header_lookup for header in required):
        return {}, []
    max_row, _ = _used_bounds(ws)
    rows: dict[str, dict[str, Any]] = {}
    duplicate_keys: list[str] = []
    for row_index in range(2, max_row + 1):
        record = {header: ws.cell(row_index, column).value for column, header in headers.items()}
        key_parts = [
            str(record.get("marketplace") or "").strip().upper(),
            str(record.get("sku") or "").strip(),
            str(record.get("asin") or "").strip().upper(),
        ]
        if not any(key_parts):
            continue
        key = "||".join(key_parts)
        if key in rows:
            duplicate_keys.append(key)
            continue
        rows[key] = record
    return rows, duplicate_keys


def _compare_sheet_by_business_key(head_ws, current_ws) -> dict[str, Any] | None:
    head_rows, head_duplicates = _sheet_rows_by_key(head_ws)
    current_rows, current_duplicates = _sheet_rows_by_key(current_ws)
    if not head_rows and not current_rows:
        return None
    keys = sorted(set(head_rows) | set(current_rows))
    added = [key for key in keys if key not in head_rows]
    removed = [key for key in keys if key not in current_rows]
    changed_fields: Counter[str] = Counter()
    changed_marketplaces: Counter[str] = Counter()
    changed_records: list[dict[str, Any]] = []
    sensitive_records: list[dict[str, Any]] = []
    for key in keys:
        if key in added or key in removed:
            continue
        before = head_rows[key]
        after = current_rows[key]
        headers = sorted(set(before) | set(after))
        field_changes: list[dict[str, str]] = []
        for header in headers:
            if _normalize(before.get(header)) == _normalize(after.get(header)):
                continue
            changed_fields[header] += 1
            field_change = {
                "field": header,
                "before": _display(before.get(header)),
                "after": _display(after.get(header)),
            }
            field_changes.append(field_change)
        if not field_changes:
            continue
        marketplace = key.split("||", 1)[0]
        if marketplace:
            changed_marketplaces[marketplace] += 1
        record = {
            "key": key,
            "marketplace": marketplace,
            "sku": str(after.get("sku") or before.get("sku") or ""),
            "asin": str(after.get("asin") or before.get("asin") or ""),
            "product_name": str(after.get("product_name") or before.get("product_name") or ""),
            "changed_fields": field_changes,
        }
        if len(changed_records) < 500:
            changed_records.append(record)
        if any(str(item.get("field") or "").strip().lower() in SENSITIVE_HEADERS for item in field_changes) and len(sensitive_records) < 80:
            sensitive_records.append(record)
    return {
        "added_keys": added[:80],
        "removed_keys": removed[:80],
        "added_count": len(added),
        "removed_count": len(removed),
        "changed_record_count": sum(changed_marketplaces.values()),
        "changed_fields": dict(changed_fields.most_common()),
        "changed_marketplaces": dict(changed_marketplaces.most_common()),
        "duplicate_keys_in_head": head_duplicates[:40],
        "duplicate_keys_in_current": current_duplicates[:40],
        "sample_changed_records": changed_records,
        "sensitive_changed_records": sensitive_records,
    }


def build_audit(config_path: Path = CONFIG_PATH) -> dict[str, Any]:
    head_bytes = _head_file_bytes("config/product_cost_config.xlsx")
    head_wb = load_workbook(BytesIO(head_bytes), data_only=False, read_only=False)
    current_wb = load_workbook(config_path, data_only=False, read_only=False)
    try:
        sheets = sorted(set(head_wb.sheetnames) | set(current_wb.sheetnames))
        sheet_results: dict[str, Any] = {}
        missing_or_added: dict[str, str] = {}
        for sheet in sheets:
            if sheet not in head_wb.sheetnames:
                missing_or_added[sheet] = "added"
                continue
            if sheet not in current_wb.sheetnames:
                missing_or_added[sheet] = "removed"
                continue
            sheet_results[sheet] = _compare_sheet(head_wb[sheet], current_wb[sheet])
        total_changed = sum(int(result.get("changed_cell_count") or 0) for result in sheet_results.values())
        sensitive_total = sum(len(result.get("sensitive_changes") or []) for result in sheet_results.values())
        return {
            "config_path": str(config_path),
            "baseline": "HEAD:config/product_cost_config.xlsx",
            "total_changed_cells": total_changed,
            "sensitive_sample_count": sensitive_total,
            "sheet_status": missing_or_added,
            "sheets": sheet_results,
        }
    finally:
        head_wb.close()
        current_wb.close()


def _markdown_escape(value: Any) -> str:
    return str(value if value is not None else "").replace("|", "\\|").replace("\n", " ")


def _change_lookup(row: dict[str, Any]) -> dict[str, dict[str, str]]:
    return {str(item.get("field") or ""): item for item in row.get("changed_fields", []) or []}


def _to_float(value: Any) -> float | None:
    text = str(value if value is not None else "").strip()
    if not text or text.startswith("="):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _format_pct(value: float) -> str:
    return f"{value * 100:.1f}%"


def _risk_change_value(
    changes: dict[str, dict[str, str]],
    inventory_changes: dict[str, dict[str, str]],
    field: str,
) -> dict[str, str] | None:
    return changes.get(field) or inventory_changes.get(field)


def _build_high_risk_rows(
    records: list[dict[str, Any]],
    inventory_by_key: dict[str, dict[str, dict[str, str]]],
    limit: int = 12,
) -> list[dict[str, Any]]:
    risk_rows: list[dict[str, Any]] = []
    for row in records:
        changes = _change_lookup(row)
        inventory_changes = inventory_by_key.get(str(row.get("key") or ""), {})
        score = 0
        reasons: list[str] = []
        evidence: list[str] = []

        price_change = _risk_change_value(changes, inventory_changes, "selling_price")
        price_before = _to_float((price_change or {}).get("before"))
        price_after = _to_float((price_change or {}).get("after"))
        if price_before is not None and price_after is not None and price_before:
            price_delta = (price_after - price_before) / abs(price_before)
            if abs(price_delta) >= 0.30:
                score += 35
                reasons.append("售价变动超过 30%")
            elif abs(price_delta) >= 0.15:
                score += 20
                reasons.append("售价变动超过 15%")
            if abs(price_delta) >= 0.15:
                evidence.append(f"selling_price {price_before:g} -> {price_after:g} ({_format_pct(price_delta)})")

        target_change = _risk_change_value(changes, inventory_changes, "suggested_target_acos")
        target_after = _to_float((target_change or {}).get("after"))
        if target_after is not None:
            if target_after <= 0:
                score += 35
                reasons.append("建议 target ACOS 变为 0")
            elif target_after < 0.05:
                score += 25
                reasons.append("建议 target ACOS 低于 5%")
            elif target_after < 0.10:
                score += 12
                reasons.append("建议 target ACOS 低于 10%")
            if target_after < 0.10:
                evidence.append(f"suggested_target_acos {target_after:.4f}")

        profit_after_change = _risk_change_value(changes, inventory_changes, "profit_after_10pct_ads")
        profit_after = _to_float((profit_after_change or {}).get("after"))
        if profit_after is not None and profit_after < 0:
            score += 35
            reasons.append("10% 广告费后利润为负")
            evidence.append(f"profit_after_10pct_ads {profit_after:g}")

        profit_before_change = _risk_change_value(changes, inventory_changes, "profit_before_ads")
        profit_before = _to_float((profit_before_change or {}).get("after"))
        if profit_before is not None and profit_before < 0:
            score += 25
            reasons.append("广告前利润为负")
            evidence.append(f"profit_before_ads {profit_before:g}")

        inventory_change = _risk_change_value(changes, inventory_changes, "current_inventory")
        inventory_before = _to_float((inventory_change or {}).get("before"))
        inventory_after = _to_float((inventory_change or {}).get("after"))
        if inventory_before is not None and inventory_after is not None:
            if inventory_before <= 0 < inventory_after:
                score += 25
                reasons.append("库存从 0 变为有货")
                evidence.append(f"current_inventory {inventory_before:g} -> {inventory_after:g}")
            elif inventory_before > 0 and inventory_after <= 0:
                score += 25
                reasons.append("库存从有货变为 0")
                evidence.append(f"current_inventory {inventory_before:g} -> {inventory_after:g}")

        total_cost_change = _risk_change_value(changes, inventory_changes, "total_cost_before_ads")
        total_cost_after = _to_float((total_cost_change or {}).get("after"))
        if price_after is not None and total_cost_after is not None and price_after < total_cost_after:
            score += 30
            reasons.append("售价低于广告前总成本")
            evidence.append(f"price {price_after:g} < total_cost_before_ads {total_cost_after:g}")

        first_leg_change = _risk_change_value(changes, inventory_changes, "first_leg_cost_rmb")
        first_leg_before = _to_float((first_leg_change or {}).get("before"))
        first_leg_after = _to_float((first_leg_change or {}).get("after"))
        if first_leg_before is not None and first_leg_after is not None and first_leg_before:
            first_leg_delta = (first_leg_after - first_leg_before) / abs(first_leg_before)
            if abs(first_leg_delta) >= 0.25:
                score += 12
                reasons.append("头程 RMB 成本变动超过 25%")
                evidence.append(f"first_leg_cost_rmb {first_leg_before:g} -> {first_leg_after:g}")

        if "inventory_note" in changes or "inventory_note" in inventory_changes:
            score += 5
            reasons.append("库存备注变化")

        if not score:
            continue
        risk_rows.append(
            {
                "score": score,
                "marketplace": row.get("marketplace"),
                "sku": row.get("sku"),
                "asin": row.get("asin"),
                "product_name": row.get("product_name"),
                "reasons": "；".join(dict.fromkeys(reasons)),
                "evidence": "；".join(dict.fromkeys(evidence)),
            }
        )
    return sorted(risk_rows, key=lambda item: (-int(item["score"]), str(item.get("marketplace") or ""), str(item.get("sku") or "")))[:limit]


def build_markdown_summary(audit: dict[str, Any]) -> str:
    product_diff = ((audit.get("sheets") or {}).get("product_cost_config") or {}).get("keyed_diff") or {}
    inventory_diff = ((audit.get("sheets") or {}).get("SKU匹配检查") or {}).get("keyed_diff") or {}
    inventory_by_key = {
        str(row.get("key") or ""): _change_lookup(row)
        for row in inventory_diff.get("sample_changed_records", []) or []
    }
    records = product_diff.get("sample_changed_records", []) or []
    lines = [
        "# product_cost_config 差异人工确认表",
        "",
        f"- 基准：`{audit.get('baseline')}`",
        f"- 当前文件：`{audit.get('config_path')}`",
        f"- 总单元格差异：{audit.get('total_changed_cells')}",
        f"- 产品记录变化：{product_diff.get('changed_record_count', 0)}",
        f"- 站点分布：{json.dumps(product_diff.get('changed_marketplaces', {}), ensure_ascii=False)}",
        "",
        "## 核对规则",
        "",
        "每个产品至少确认售价、采购成本、头程成本、利润、break-even ACOS、建议 target ACOS、当前库存和库存备注。确认后再决定是否保留 `config/product_cost_config.xlsx` 的变更。",
        "",
    ]
    if not records:
        lines.extend(["## 产品差异", "", "未发现按 marketplace + sku + asin 对齐后的产品字段变化。"])
        return "\n".join(lines) + "\n"

    high_risk_rows = _build_high_risk_rows(records, inventory_by_key)
    if high_risk_rows:
        lines.extend(
            [
                "## 最高风险差异",
                "",
                "| 风险分 | 站点 | SKU | ASIN | 产品 | 原因 | 证据 |",
                "|---:|---|---|---|---|---|---|",
            ]
        )
        for row in high_risk_rows:
            lines.append(
                "| "
                + " | ".join(
                    [
                        _markdown_escape(row.get("score")),
                        _markdown_escape(row.get("marketplace")),
                        _markdown_escape(row.get("sku")),
                        _markdown_escape(row.get("asin")),
                        _markdown_escape(row.get("product_name")),
                        _markdown_escape(row.get("reasons")),
                        _markdown_escape(row.get("evidence")),
                    ]
                )
                + " |"
            )
        lines.append("")

    lines.extend(
        [
            "## 产品差异",
            "",
            "| 站点 | SKU | ASIN | 产品 | 字段 | 旧值 | 新值 |",
            "|---|---|---|---|---|---|---|",
        ]
    )
    for row in records:
        changes = _change_lookup(row)
        inventory_changes = inventory_by_key.get(str(row.get("key") or ""), {})
        for field in REVIEW_FIELDS:
            change = changes.get(field) or inventory_changes.get(field)
            if not change:
                continue
            lines.append(
                "| "
                + " | ".join(
                    [
                        _markdown_escape(row.get("marketplace")),
                        _markdown_escape(row.get("sku")),
                        _markdown_escape(row.get("asin")),
                        _markdown_escape(row.get("product_name")),
                        _markdown_escape(field),
                        _markdown_escape(change.get("before")),
                        _markdown_escape(change.get("after")),
                    ]
                )
                + " |"
            )
    lines.extend(
        [
            "",
            "## 全字段变化 Top",
            "",
            "| 字段 | 变化产品数 |",
            "|---|---:|",
        ]
    )
    for field, count in list((product_diff.get("changed_fields") or {}).items())[:30]:
        lines.append(f"| {_markdown_escape(field)} | {count} |")
    return "\n".join(lines) + "\n"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a product_cost_config.xlsx difference audit.")
    parser.add_argument("--config-path", type=Path, default=CONFIG_PATH)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    audit = build_audit(args.config_path)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    MARKDOWN_OUTPUT_PATH.write_text(build_markdown_summary(audit), encoding="utf-8")
    print(f"[check] total changed cells: {audit['total_changed_cells']}", flush=True)
    print(f"[check] wrote: {OUTPUT_PATH}", flush=True)
    print(f"[check] wrote: {MARKDOWN_OUTPUT_PATH}", flush=True)
    for sheet, result in audit["sheets"].items():
        count = int(result.get("changed_cell_count") or 0)
        if count:
            print(f"[sheet] {sheet}: {count}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

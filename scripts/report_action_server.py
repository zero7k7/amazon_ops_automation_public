from __future__ import annotations

import cgi
import html
import json
import mimetypes
import os
import re
import secrets
import select
import shutil
import subprocess
import sys
import threading
import time
from datetime import datetime, timedelta
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook

try:
    from scripts.chrome_cdp_helper import DEFAULT_ENDPOINT as CHROME_CDP_ENDPOINT
    from scripts.chrome_cdp_helper import MAC_CHROME_APP
    from scripts.chrome_cdp_helper import endpoint_available as _chrome_cdp_endpoint_available
except ModuleNotFoundError:  # pragma: no cover - used when executed as scripts/report_action_server.py
    from chrome_cdp_helper import DEFAULT_ENDPOINT as CHROME_CDP_ENDPOINT
    from chrome_cdp_helper import MAC_CHROME_APP
    from chrome_cdp_helper import endpoint_available as _chrome_cdp_endpoint_available


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.autoopt_feedback import add_action_identity, is_executable_action
from src.sellersprite_competitor_discovery import (
    SELLERSPRITE_COMPETITOR_DISCOVERY_CACHE_PATH,
    discovery_record_needs_refresh,
    load_competitor_discovery_records,
)
from src.sellersprite_fusion import load_sellersprite_records
from src.sellersprite_history import sellersprite_cache_max_age_days_for_row, upsert_sellersprite_history
from scripts.frontend_check_results import boolish_flag
from scripts.sellersprite_reverse_asin_fetch import _cached_recent as _sellersprite_cached_recent
from scripts.sellersprite_reverse_asin_fetch import _competitor_rows as _sellersprite_competitor_rows

HOST = "127.0.0.1"
PORT = 8765
OUTPUT_DIR = ROOT / "data" / "output"
ACTION_TOKEN_FILENAME = ".report_action_token"
ACTION_TOKEN_HEADER = "X-Report-Action-Token"
CONFIG_DIR = ROOT / "config"
CONFIG_REVIEW_DIR = ROOT / "data" / "config_review"
CONFIG_ARCHIVE_DIR = ROOT / "data" / "archive" / "config_updates"
INBOX_DIR = ROOT / "data" / "inbox"
UPLOAD_STAGING_DIR = INBOX_DIR / "_upload_staging"
SUBMISSION_STATUS_PATH = OUTPUT_DIR / "local_submission_status.json"
REPORT_BASE_URL = f"http://{HOST}:{PORT}/report"
ALLOWED_UPLOAD_SUFFIXES = {".csv", ".xlsx"}
CONFIG_UPLOAD_TARGETS = {
    "cost": {
        "label": "成本配置",
        "target_path": CONFIG_DIR / "product_cost_config.xlsx",
        "review_path": CONFIG_REVIEW_DIR / "product_cost_config.pending.xlsx",
        "allowed_filenames": {"product_cost_config.xlsx"},
        "required_sheets": {
            "product_cost_config": {"marketplace", "sku", "asin", "product_name", "currency"},
        },
        "review_command": [
            sys.executable,
            "scripts/audit_cost_config_changes.py",
            "--config-path",
            str(CONFIG_REVIEW_DIR / "product_cost_config.pending.xlsx"),
        ],
        "refresh_reports_after_apply": False,
        "requires_confirm": True,
    },
    "alias": {
        "label": "SKU 别名映射",
        "target_path": CONFIG_DIR / "sku_alias_map.xlsx",
        "review_path": CONFIG_REVIEW_DIR / "sku_alias_map.pending.xlsx",
        "allowed_filenames": {"sku_alias_map.xlsx"},
        "required_sheets": {
            "Sheet1": {"marketplace", "source_sku", "canonical_sku", "asin"},
        },
        "review_command": None,
        "refresh_reports_after_apply": True,
        "requires_confirm": False,
    },
}
MAX_UPLOAD_BYTES = 50 * 1024 * 1024
REPORT_LINKS = {
    "dashboard": f"{REPORT_BASE_URL}/dashboard.html",
    "summary": f"{REPORT_BASE_URL}/summary.html",
    "latest_recommendations": f"{REPORT_BASE_URL}/latest_recommendations.html",
    "uk_report": f"{REPORT_BASE_URL}/uk_report.html",
    "us_report": f"{REPORT_BASE_URL}/us_report.html",
    "de_report": f"{REPORT_BASE_URL}/de_report.html",
}
FRONTEND_QUEUE_ACCEPTANCE_THRESHOLD = 0.80
FRONTEND_RETRY_BATCH_LIMIT = 3
FRONTEND_RETRY_MAX_BATCH_LIMIT = 7
SELLERSPRITE_REVERSE_ASIN_TARGET_COUNT = 20
SELLERSPRITE_STATUS_SUMMARY_KEYS = (
    "sellersprite_queue_total",
    "sellersprite_cached_count",
    "sellersprite_missing_count",
    "sellersprite_cached_labels",
    "sellersprite_missing_labels",
    "sellersprite_discovery_queue_total",
    "sellersprite_discovery_cached_count",
    "sellersprite_discovery_missing_count",
    "sellersprite_discovery_cached_labels",
    "sellersprite_discovery_missing_labels",
)
ENABLE_BROWSER_FRONTEND_ENV = "AMAZON_OPS_ENABLE_BROWSER_FRONTEND"
FEEDBACK_INPUT_FILENAME = "autoopt_feedback_input.json"
FEEDBACK_AUDIT_PREFIX = "feedback_audit_log"
SIDE_EFFECT_GET_PATHS = {
    "/run/report-refresh",
    "/run/frontend-retry",
    "/run/frontend-check-one",
    "/run/battle-diagnosis-one",
}
SIDE_EFFECT_POST_PATHS = {
    "/upload/today-data",
    "/upload/config",
    "/apply/config",
    "/copy/text",
    "/run/report-refresh",
    "/run/daily-update",
    "/run/frontend-retry",
    "/run/frontend-check-one",
    "/run/battle-diagnosis-one",
    "/feedback/ad-action-complete",
    "/feedback/ad-action-cancel",
}
LOCAL_CORS_ORIGINS = {
    f"http://{HOST}:{PORT}",
    f"http://localhost:{PORT}",
}

_lock = threading.Lock()
_frontend_async_lock = threading.Lock()
_sellersprite_async_lock = threading.Lock()
_last_result: dict[str, object] = {"running": False, "message": "ready"}
_frontend_async_status: dict[str, object] = {"running": False, "message": "P0 前台后台检查未运行。"}
_sellersprite_async_status: dict[str, object] = {"running": False, "message": "卖家精灵后台反查未运行。"}
_RUNTIME_STATUS_KEYS = {"detail", "elapsed_seconds", "started_at_epoch", "step", "total_steps"}
_LEGACY_FRONTEND_RETRY_FAILURES = (
    "前台数据重试失败：urllib 读取 Amazon 前台",
    "前台数据重试失败：",
)
_REPORT_RESTORE_MARKER = "[restore] report outputs restored to pre-report snapshot after failure"
_REPORT_REFRESH_BLOCKER_MARKER = "[fail] report refresh blocker:"


def _browser_frontend_enabled() -> bool:
    value = os.environ.get(ENABLE_BROWSER_FRONTEND_ENV, "").strip().lower()
    if value in {"0", "false", "no", "off", "urllib"}:
        return False
    return True


def _frontend_refresh_method() -> str:
    return "chrome-cdp" if _browser_frontend_enabled() else "urllib"


def _frontend_method_label() -> str:
    return "Chrome CDP 后台" if _browser_frontend_enabled() else "urllib 静默"


def _frontend_command(
    *,
    priority: str = "",
    require_competitor_samples: bool = False,
    search_policy: str = "always",
    limit: int | None = None,
    only_stale: bool = True,
) -> list[str]:
    method = _frontend_refresh_method()
    command = [
        sys.executable,
        "scripts/run_frontend_checks.py",
        "--method",
        method,
        "--timeout",
        "18",
        "--sleep",
        "0.5",
        "--search-policy",
        search_policy,
    ]
    if only_stale:
        command.append("--only-stale")
    if method == "chrome-cdp":
        command.extend(["--cdp-endpoint", CHROME_CDP_ENDPOINT, "--cdp-attempts", "1"])
    else:
        command.extend(["--retries", "3"])
    if require_competitor_samples:
        command.append("--require-competitor-samples")
    if priority:
        command.extend(["--priority", priority])
    command.extend(["--limit", str(limit or FRONTEND_RETRY_BATCH_LIMIT)])
    return command


def _frontend_refresh_batch_limit(needed: int) -> int:
    if needed <= 0:
        return FRONTEND_RETRY_BATCH_LIMIT
    return min(max(needed, FRONTEND_RETRY_BATCH_LIMIT), FRONTEND_RETRY_MAX_BATCH_LIMIT)


def _sellersprite_reverse_command(*, priority: str = "", params: dict[str, str] | None = None) -> list[str]:
    command = [
        sys.executable,
        "scripts/sellersprite_reverse_asin_fetch.py",
        "--target-count",
        str(SELLERSPRITE_REVERSE_ASIN_TARGET_COUNT),
        "--include-competitors",
        "--competitor-limit-per-product",
        "3",
        "--competitor-cache-days",
        "7",
    ]
    params = params or {}
    marketplace = str(params.get("marketplace") or "").strip().upper()
    sku = str(params.get("sku") or "").strip()
    asin = str(params.get("asin") or "").strip().upper()
    if priority:
        command.extend(["--priority", priority])
    if marketplace:
        command.extend(["--marketplace", marketplace])
    if sku:
        command.extend(["--sku", sku])
    if asin:
        command.extend(["--asin", asin])
    return command


def _action_token_path() -> Path:
    return OUTPUT_DIR / ACTION_TOKEN_FILENAME


def _load_or_create_action_token() -> str:
    path = _action_token_path()
    try:
        token = path.read_text(encoding="utf-8").strip()
        if len(token) >= 32:
            return token
    except OSError:
        pass
    token = secrets.token_urlsafe(32)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(token + "\n", encoding="utf-8")
    try:
        path.chmod(0o600)
    except OSError:
        pass
    return token


def _is_valid_action_token(token: str) -> bool:
    if not token:
        return False
    return secrets.compare_digest(token, _load_or_create_action_token())


def _allowed_cors_origin(origin: str | None) -> str:
    origin = str(origin or "").strip()
    return origin if origin in LOCAL_CORS_ORIGINS else ""


def _base_status(message: str = "ready") -> dict[str, object]:
    return {"running": False, "message": message, "uploaded_files": []}


def _load_submission_status() -> dict[str, object]:
    try:
        payload = json.loads(SUBMISSION_STATUS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return _base_status()
    return payload if isinstance(payload, dict) else _base_status()


def _clear_stale_running_status_on_startup() -> None:
    payload = _load_submission_status()
    cleaned = _strip_runtime_status(payload) if payload.get("running") else dict(payload)
    if payload.get("running"):
        cleaned.update(
            {
                "running": False,
                "message": "上次本地任务已随服务重启中断；请重新点击按钮启动。",
                "returncode": -15,
                "status_scope": str(payload.get("status_scope") or "local_workflow_interrupted"),
                "failure_mode": "local_service_restarted",
                "updated_at_epoch": time.time(),
            }
        )
    elif cleaned.get("status_scope") == "frontend_retry" and str(cleaned.get("failure_mode") or "").startswith("chrome_cdp_frontend_check_timeout"):
        cleaned.update(
            {
                "running": False,
                "message": "前台刷新未运行；上次本轮刷新超时已结束，下次点击会按 3 个一批继续刷新。",
                "returncode": 0,
                "original_returncode": cleaned.get("returncode") or 124,
                "soft_failure": True,
                "failure_mode": "chrome_cdp_frontend_check_timeout_partial",
                "updated_at_epoch": time.time(),
            }
        )
    elif cleaned.get("status_scope") == "frontend_retry" and cleaned.get("failure_mode") == "frontend_refresh_not_needed":
        cleaned = _fresh_status(
            "前台刷新未运行；点击按钮会重新读取当前队列。",
            status_scope="frontend_retry_idle",
            failure_mode="frontend_retry_stale_result_cleared",
            updated_at_epoch=time.time(),
        )
    cleaned["frontend_async_status"] = {"running": False, "message": "P0 前台后台检查未运行。"}
    cleaned["sellersprite_async_status"] = {"running": False, "message": "卖家精灵后台反查未运行。"}
    _write_submission_status(cleaned)


def _fresh_status(message: str, **extra: object) -> dict[str, object]:
    return {
        "running": False,
        "message": message,
        "uploaded_files": [],
        "config_upload": {},
        "report_links": REPORT_LINKS,
        **extra,
    }


def _strip_runtime_status(payload: dict[str, object]) -> dict[str, object]:
    cleaned = dict(payload)
    for key in _RUNTIME_STATUS_KEYS:
        cleaned.pop(key, None)
    return cleaned


def _latest_recommendations_available() -> bool:
    path = OUTPUT_DIR / "latest_recommendations.html"
    try:
        return path.exists() and path.stat().st_size > 0
    except OSError:
        return False


def _daily_update_restored_report_failure(payload: dict[str, object], message: str, detail: str) -> bool:
    if payload.get("returncode") in (None, 0):
        return False
    combined = "\n".join(
        str(payload.get(key) or "")
        for key in ("stdout_tail", "stderr_tail")
    )
    return (
        "daily update 失败" in message
        and _REPORT_RESTORE_MARKER in combined
        and (_REPORT_REFRESH_BLOCKER_MARKER in combined or "report refresh blocker" in combined or "missing parseable date" in combined or "missing parseable date" in detail)
        and _latest_recommendations_available()
    )


def _normalize_status_payload(payload: dict[str, object]) -> dict[str, object]:
    cleaned = dict(payload)
    message = str(cleaned.get("message") or "")
    detail = str(cleaned.get("detail") or "")
    if _daily_update_restored_report_failure(cleaned, message, detail):
        cleaned["original_returncode"] = cleaned.get("returncode")
        cleaned["returncode"] = 0
        cleaned["message"] = (
            "报告可用：上次 daily update 被导入校验拦截，已恢复到失败前报告；"
            "请修正缺少日期的导入文件后再刷新。"
        )
        cleaned["status_scope"] = "daily_update_restored_report"
        cleaned["failure_mode"] = "import_manifest_date_blocker_restored_report"
        cleaned["soft_failure"] = True
        return cleaned
    is_legacy_frontend_retry = any(marker in message for marker in _LEGACY_FRONTEND_RETRY_FAILURES)
    is_legacy_chrome_retry = "真实 Chrome 20次前台检查" in message or "真实 Chrome CDP 20次前台检查" in message
    is_legacy_twenty_gate_message = "当前前台队列 20次验收" in message or "当前前台队列 Chrome CDP 20次验收" in message
    is_urllib_frontend_retry = (
        "urllib 读取 Amazon 前台" in message
        or "urllib 读取 Amazon 前台" in detail
        or message.startswith("urllib 快速重试")
        or message.startswith("urllib 无浏览器快速重试")
    )
    if is_legacy_frontend_retry or is_urllib_frontend_retry or is_legacy_chrome_retry or is_legacy_twenty_gate_message:
        cleaned["status_scope"] = "frontend_retry"
    if cleaned.get("returncode") not in (None, 0) and (is_legacy_frontend_retry or is_urllib_frontend_retry):
        cleaned["message"] = (
            "urllib 快速重试失败：Amazon 可能拦截无浏览器读取；"
            "已保留原有前台缓存和 Chrome 实测证据。"
        )
        cleaned["soft_failure"] = True
        cleaned["failure_mode"] = "urllib_frontend_blocked"
    elif is_legacy_chrome_retry:
        cleaned["message"] = message.replace("真实 Chrome CDP 20次前台检查", "当前前台队列 Chrome CDP 20次验收").replace(
            "真实 Chrome 20次前台检查",
            "当前前台队列 20次验收",
        )
    if cleaned.get("status_scope") == "frontend_retry" and (is_legacy_chrome_retry or is_legacy_twenty_gate_message):
        summary = _frontend_queue_status_summary()
        status_message = _frontend_retry_status_message(summary)
        if status_message:
            cleaned.update(summary)
            cleaned["message"] = status_message
            cleaned["failure_mode"] = (
                "chrome_cdp_frontend_check_passed_with_pending"
                if summary.get("frontend_queue_passed") and int(summary.get("frontend_pending_count") or 0)
                else ("chrome_cdp_frontend_check_passed" if summary.get("frontend_queue_passed") else "chrome_cdp_frontend_check_partial")
            )
            if cleaned.get("returncode") not in (None, 0):
                cleaned["soft_failure"] = True
    if cleaned.get("status_scope") == "frontend_retry" and "当前前台队列刷新" in message:
        summary = _frontend_queue_status_summary()
        summary.update(_frontend_refresh_result_summary())
        for key in (
            "frontend_refresh_total",
            "frontend_refresh_live_checked",
            "frontend_refresh_skipped",
            "frontend_refresh_cache_used",
            "frontend_refresh_failed",
        ):
            if cleaned.get(key) not in (None, ""):
                summary[key] = cleaned.get(key)
        status_message = _frontend_retry_status_message(summary)
        if status_message:
            cleaned.update(summary)
            cleaned["message"] = status_message
            cleaned["failure_mode"] = (
                "chrome_cdp_frontend_check_passed_with_pending"
                if summary.get("frontend_queue_passed") and int(summary.get("frontend_pending_count") or 0)
                else ("chrome_cdp_frontend_check_passed" if summary.get("frontend_queue_passed") else "chrome_cdp_frontend_check_partial")
            )
    if (
        cleaned.get("returncode") not in (None, 0)
        and cleaned.get("status_scope") == "frontend_retry"
        and cleaned.get("failure_mode") in {"chrome_cdp_frontend_check_failed", "chrome_cdp_frontend_check_partial", ""}
    ):
        summary = _frontend_queue_status_summary()
        status_message = _frontend_retry_status_message(summary)
        if status_message:
            cleaned.update(summary)
            cleaned["message"] = status_message
            cleaned["failure_mode"] = (
                "chrome_cdp_frontend_check_passed_with_pending"
                if summary.get("frontend_queue_passed")
                else "chrome_cdp_frontend_check_partial"
            )
            cleaned["soft_failure"] = True
    return cleaned


def _write_submission_status(payload: dict[str, object]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    payload = {**payload, "updated_at_epoch": time.time(), "report_links": REPORT_LINKS}
    SUBMISSION_STATUS_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _sync_last_result(payload: dict[str, object]) -> None:
    global _last_result
    _last_result = payload
    if "frontend_async_status" not in payload and _frontend_async_status:
        payload = {**payload, "frontend_async_status": _frontend_async_status}
    if "sellersprite_async_status" not in payload and _sellersprite_async_status:
        payload = {**payload, "sellersprite_async_status": _sellersprite_async_status}
    _write_submission_status(payload)


def _set_progress(message: str, step: int, total: int, detail: str = "") -> None:
    now = time.time()
    started_at = float(_last_result.get("started_at_epoch") or now)
    _sync_last_result({
        **_last_result,
        "running": True,
        "message": message,
        "detail": detail,
        "step": step,
        "total_steps": total,
        "started_at_epoch": started_at,
        "elapsed_seconds": max(0, int(now - started_at)),
    })


def _sync_frontend_async_status(payload: dict[str, object]) -> None:
    global _frontend_async_status, _last_result
    _frontend_async_status = {**payload, "updated_at_epoch": time.time()}
    _last_result = {**_last_result, "frontend_async_status": _frontend_async_status}
    persisted = _load_submission_status()
    persisted["frontend_async_status"] = _frontend_async_status
    _write_submission_status(persisted)


def _sellersprite_summary_from_payload(payload: dict[str, object]) -> dict[str, object]:
    return {key: payload[key] for key in SELLERSPRITE_STATUS_SUMMARY_KEYS if key in payload}


def _sync_sellersprite_async_status(payload: dict[str, object]) -> None:
    global _sellersprite_async_status, _last_result
    _sellersprite_async_status = {**payload, "updated_at_epoch": time.time()}
    summary = _sellersprite_summary_from_payload(payload)
    _last_result = {**_last_result, **summary, "sellersprite_async_status": _sellersprite_async_status}
    persisted = _load_submission_status()
    persisted.update(summary)
    persisted["sellersprite_async_status"] = _sellersprite_async_status
    _write_submission_status(persisted)


def _set_frontend_async_progress(message: str, step: int, total: int, detail: str = "") -> None:
    now = time.time()
    started_at = float(_frontend_async_status.get("started_at_epoch") or now)
    _sync_frontend_async_status(
        {
            **_frontend_async_status,
            "running": True,
            "message": message,
            "detail": detail,
            "step": step,
            "total_steps": total,
            "started_at_epoch": started_at,
            "elapsed_seconds": max(0, int(now - started_at)),
            "status_scope": "frontend_async",
        }
    )


def _run_command(command: list[str], timeout: int) -> subprocess.CompletedProcess[str]:
    try:
        return subprocess.run(command, cwd=ROOT, env=_subprocess_env(), capture_output=True, text=True, timeout=timeout)
    except OSError as exc:
        return subprocess.CompletedProcess(command, 127, "", f"cannot start command: {exc}")


def _subprocess_env() -> dict[str, str]:
    env = os.environ.copy()
    existing = env.get("PYTHONPATH", "")
    parts = [str(ROOT)]
    if existing:
        parts.append(existing)
    env["PYTHONPATH"] = os.pathsep.join(parts)
    return env


def _run_command_with_status(
    command: list[str],
    timeout: int,
    *,
    step: int,
    total_steps: int,
    message: str,
) -> subprocess.CompletedProcess[str]:
    stdout_parts: list[str] = []
    stderr_parts: list[str] = []
    try:
        process = subprocess.Popen(
            command,
            cwd=ROOT,
            env=_subprocess_env(),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            bufsize=1,
        )
    except OSError as exc:
        return subprocess.CompletedProcess(command, 127, "", f"cannot start command: {exc}")
    deadline = time.monotonic() + timeout
    last_status_at = 0.0
    streams = [stream for stream in (process.stdout, process.stderr) if stream is not None]

    def publish(force: bool = False) -> None:
        nonlocal last_status_at
        now = time.monotonic()
        if not force and now - last_status_at < 1.0:
            return
        last_status_at = now
        _sync_last_result(
            {
                **_last_result,
                "running": True,
                "message": message,
                "detail": " ".join(command),
                "step": step,
                "total_steps": total_steps,
                "elapsed_seconds": max(0, int(time.time() - float(_last_result.get("started_at_epoch") or time.time()))),
                "stdout_tail": "".join(stdout_parts)[-4000:],
                "stderr_tail": "".join(stderr_parts)[-4000:],
            }
        )

    try:
        while streams and process.poll() is None:
            if time.monotonic() > deadline:
                process.kill()
                raise subprocess.TimeoutExpired(
                    command,
                    timeout,
                    output="".join(stdout_parts),
                    stderr="".join(stderr_parts),
                )
            readable, _, _ = select.select(streams, [], [], 0.25)
            if not readable:
                publish()
                continue
            for stream in readable:
                line = stream.readline()
                if not line:
                    streams.remove(stream)
                    continue
                if stream is process.stdout:
                    stdout_parts.append(line)
                else:
                    stderr_parts.append(line)
                publish(force=True)
        remaining_stdout, remaining_stderr = process.communicate(timeout=max(0.1, deadline - time.monotonic()))
    except subprocess.TimeoutExpired:
        process.kill()
        remaining_stdout, remaining_stderr = process.communicate()
        stdout_parts.append(remaining_stdout or "")
        stderr_parts.append(remaining_stderr or "")
        raise subprocess.TimeoutExpired(command, timeout, output="".join(stdout_parts), stderr="".join(stderr_parts))
    stdout_parts.append(remaining_stdout or "")
    stderr_parts.append(remaining_stderr or "")
    publish(force=True)
    return subprocess.CompletedProcess(command, process.returncode or 0, "".join(stdout_parts), "".join(stderr_parts))


def _run_frontend_command_with_progress(
    command: list[str],
    timeout: int,
    *,
    step: int,
    total_steps: int,
    progress_writer=None,
) -> subprocess.CompletedProcess[str]:
    stdout_parts: list[str] = []
    stderr_parts: list[str] = []
    try:
        process = subprocess.Popen(
            command,
            cwd=ROOT,
            env=_subprocess_env(),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            bufsize=1,
        )
    except OSError as exc:
        return subprocess.CompletedProcess(command, 127, "", f"cannot start command: {exc}")
    deadline = time.monotonic() + timeout
    streams = [stream for stream in (process.stdout, process.stderr) if stream is not None]
    try:
        while streams and process.poll() is None:
            if time.monotonic() > deadline:
                process.kill()
                raise subprocess.TimeoutExpired(command, timeout, output="".join(stdout_parts), stderr="".join(stderr_parts))
            readable, _, _ = select.select(streams, [], [], 0.25)
            for stream in readable:
                line = stream.readline()
                if not line:
                    streams.remove(stream)
                    continue
                if stream is process.stdout:
                    stdout_parts.append(line)
                    match = re.search(r"\[frontend-progress\]\s+(\d+)/(\d+)\s+(\w+):\s*(.*)", line)
                    if match:
                        current, total, action, label = match.groups()
                        if action == "skip":
                            message = f"跳过已有今日前台证据 {current}/{total}：{label.strip()}"
                        else:
                            message = f"正在刷新 {current}/{total}：{label.strip()}"
                        writer = progress_writer or _set_progress
                        writer(message, step, total_steps, " ".join(command))
                else:
                    stderr_parts.append(line)
        remaining_stdout, remaining_stderr = process.communicate(timeout=max(0.1, deadline - time.monotonic()))
    except subprocess.TimeoutExpired:
        process.kill()
        remaining_stdout, remaining_stderr = process.communicate()
        stdout_parts.append(remaining_stdout or "")
        stderr_parts.append(remaining_stderr or "")
        raise subprocess.TimeoutExpired(command, timeout, output="".join(stdout_parts), stderr="".join(stderr_parts))
    stdout_parts.append(remaining_stdout or "")
    stderr_parts.append(remaining_stderr or "")
    return subprocess.CompletedProcess(command, process.returncode or 0, "".join(stdout_parts), "".join(stderr_parts))


def _chrome_cdp_available(endpoint: str = CHROME_CDP_ENDPOINT) -> bool:
    return _chrome_cdp_endpoint_available(endpoint)


def _start_chrome_cdp_if_needed(endpoint: str = CHROME_CDP_ENDPOINT, wait_seconds: float = 12.0) -> bool:
    if _chrome_cdp_available(endpoint):
        return True
    if not Path(MAC_CHROME_APP).exists():
        return False
    profile_dir = OUTPUT_DIR / "chrome_cdp_profile"
    profile_dir.mkdir(parents=True, exist_ok=True)
    chrome_args = [
        "--remote-debugging-port=9222",
        f"--user-data-dir={profile_dir}",
        "--no-first-run",
        "--no-default-browser-check",
        "--window-position=-32000,-32000",
        "--window-size=1280,900",
        "about:blank",
    ]
    command = [
        "/usr/bin/open",
        "-g",
        "-na",
        "Google Chrome",
        "--args",
        *chrome_args,
    ]
    log_path = OUTPUT_DIR / "chrome_cdp_launch.log"
    log_file = log_path.open("a", encoding="utf-8")
    try:
        subprocess.Popen(command, cwd=ROOT, env=_subprocess_env(), stdout=log_file, stderr=subprocess.STDOUT, start_new_session=True)
    except OSError as exc:
        log_file.write(f"cannot start Chrome CDP: {exc}\n")
        log_file.close()
        return False
    deadline = time.monotonic() + max(wait_seconds, 1.0)
    while time.monotonic() < deadline:
        if _chrome_cdp_available(endpoint):
            return True
        time.sleep(0.5)
    return False


def _status_payload() -> dict[str, object]:
    persisted = _load_submission_status()
    if _last_result.get("message") == "ready" and persisted.get("updated_at_epoch"):
        payload = dict(persisted)
    else:
        payload = {**persisted, **_last_result}
    if payload.get("running") and payload.get("started_at_epoch"):
        payload["elapsed_seconds"] = max(0, int(time.time() - float(payload["started_at_epoch"])))
    elif not payload.get("running"):
        payload = _strip_runtime_status(payload)
    payload = _normalize_status_payload(payload)
    if "frontend_async_status" not in payload and _frontend_async_status:
        payload["frontend_async_status"] = _frontend_async_status
    if "sellersprite_async_status" not in payload and _sellersprite_async_status:
        payload["sellersprite_async_status"] = _sellersprite_async_status
    seller_status = payload.get("sellersprite_async_status")
    if isinstance(seller_status, dict) and not seller_status.get("running"):
        seller_message = str(seller_status.get("message") or "")
        try:
            seller_summary = _sellersprite_reverse_needed_summary()
        except Exception:
            seller_summary = {}
        if seller_summary:
            payload.update(seller_summary)
            missing_count = int(seller_summary.get("sellersprite_missing_count") or 0)
            refreshed_status = {**seller_status, **seller_summary, "running": False, "status_scope": "sellersprite_async"}
            if missing_count > 0:
                refreshed_status.update(
                    {
                        "returncode": 0,
                            "message": f"本次需抓 {missing_count} 个 ASIN。",
                    }
                )
            elif int(seller_summary.get("sellersprite_queue_total") or 0) > 0:
                refreshed_status.update({"returncode": 0, "message": "卖家精灵当前队列已有有效反查，无需重复运行。"})
            payload["sellersprite_async_status"] = refreshed_status
    payload["report_links"] = REPORT_LINKS
    return payload


def _report_file_link(path: Path) -> str:
    try:
        return f"{REPORT_BASE_URL}/{path.relative_to(OUTPUT_DIR).as_posix()}"
    except ValueError:
        return ""


def _config_audit_links(kind: str) -> dict[str, str]:
    if kind != "cost":
        return {}
    return {
        "cost_diff_markdown": _report_file_link(OUTPUT_DIR / "cost_config_diff_summary.md"),
        "cost_diff_json": _report_file_link(OUTPUT_DIR / "cost_config_diff_summary.json"),
    }


def _config_status_payload() -> dict[str, object]:
    items: dict[str, object] = {}
    for kind, config in CONFIG_UPLOAD_TARGETS.items():
        review_path = Path(config["review_path"])
        target_path = Path(config["target_path"])
        items[kind] = {
            "label": config.get("label"),
            "target_path": str(target_path.relative_to(ROOT)),
            "pending_path": str(review_path.relative_to(ROOT)),
            "has_pending": review_path.exists(),
            "pending_size": review_path.stat().st_size if review_path.exists() else 0,
            "requires_confirm": bool(config.get("requires_confirm")),
            "audit_links": _config_audit_links(kind),
        }
    return {"configs": items}


def _completed_payload(
    completed: subprocess.CompletedProcess[str],
    message_ok: str,
    message_fail: str,
    **extra: object,
) -> dict[str, object]:
    return _fresh_status(
        message_ok if completed.returncode == 0 else message_fail,
        returncode=completed.returncode,
        stdout_tail=completed.stdout[-4000:],
        stderr_tail=completed.stderr[-4000:],
        report_url=f"{REPORT_BASE_URL}/latest_recommendations.html",
        **extra,
    )


def _today_iso() -> str:
    return time.strftime("%Y-%m-%d")


def _feedback_input_path() -> Path:
    return OUTPUT_DIR / FEEDBACK_INPUT_FILENAME


def _feedback_audit_path(event_at: str | None = None) -> Path:
    timestamp = str(event_at or "").strip() or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        day = datetime.fromisoformat(timestamp[:10]).date()
    except Exception:
        day = datetime.now().date()
    return OUTPUT_DIR / f"{FEEDBACK_AUDIT_PREFIX}_{day.strftime('%Y%m%d')}.json"


def _load_audit_rows(path: Path) -> list[dict[str, object]]:
    if not path.exists():
        return []
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []
    if isinstance(payload, list):
        return [row for row in payload if isinstance(row, dict)]
    if isinstance(payload, dict) and isinstance(payload.get("rows"), list):
        return [row for row in payload["rows"] if isinstance(row, dict)]
    return []


def append_feedback_audit_event(
    event: str,
    feedback: dict[str, object],
    *,
    previous_status: str,
    new_status: str,
    source: str = "latest_recommendations.html",
) -> dict[str, object]:
    event_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    audit_row = {
        "event": event,
        "marketplace": str(feedback.get("marketplace") or "").strip().upper(),
        "sku": str(feedback.get("sku") or "").strip(),
        "asin": str(feedback.get("asin") or "").strip().upper(),
        "product_name": str(feedback.get("product_name") or "").strip(),
        "action_id": str(feedback.get("action_id") or "").strip(),
        "action_scope": str(feedback.get("action_scope") or "").strip(),
        "normalized_action": str(feedback.get("normalized_action") or "").strip(),
        "search_term_or_target": str(feedback.get("search_term_or_target") or "").strip(),
        "previous_status": previous_status,
        "new_status": new_status,
        "event_at": event_at,
        "source": source,
        "confirmed_at": str(feedback.get("confirmed_at") or "").strip(),
        "report_date": str(feedback.get("report_date") or "").strip(),
    }
    path = _feedback_audit_path(event_at)
    path.parent.mkdir(parents=True, exist_ok=True)
    rows = _load_audit_rows(path)
    rows.append(audit_row)
    path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    return audit_row


def _latest_report_date() -> str:
    analysis_path = OUTPUT_DIR / "latest_analysis.json"
    try:
        payload = json.loads(analysis_path.read_text(encoding="utf-8"))
    except Exception:
        return _today_iso()
    if isinstance(payload, dict):
        report_date = str(payload.get("report_date") or "").strip()
        if re.match(r"^\d{4}-\d{2}-\d{2}$", report_date):
            return report_date
        for result in payload.get("marketplace_results") or []:
            if not isinstance(result, dict):
                continue
            summary = result.get("summary") or {}
            if isinstance(summary, dict):
                report_date = str(summary.get("report_date") or "").strip()
                if re.match(r"^\d{4}-\d{2}-\d{2}$", report_date):
                    return report_date
    return _today_iso()


def _review_date_from(confirmed_at: str, days: int) -> str:
    try:
        day = datetime.fromisoformat(confirmed_at[:10]).date()
    except Exception:
        day = datetime.now().date()
    return (day + timedelta(days=days)).isoformat()


def _load_feedback_payload(path: Path | None = None) -> tuple[object, list[dict[str, object]]]:
    feedback_path = path or _feedback_input_path()
    if not feedback_path.exists():
        return {"rows": []}, []
    try:
        payload = json.loads(feedback_path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise ValueError(f"反馈文件无法读取：{exc}") from exc
    if isinstance(payload, list):
        rows = [row for row in payload if isinstance(row, dict)]
        return payload, rows
    if isinstance(payload, dict):
        raw_rows = payload.get("rows")
        if raw_rows is None:
            raw_rows = []
        if not isinstance(raw_rows, list):
            raise ValueError("反馈文件 rows 必须是数组。")
        rows = [row for row in raw_rows if isinstance(row, dict)]
        return payload, rows
    raise ValueError("反馈文件根节点必须是数组或包含 rows 的对象。")


def _write_feedback_payload(original_payload: object, rows: list[dict[str, object]], path: Path | None = None) -> None:
    feedback_path = path or _feedback_input_path()
    feedback_path.parent.mkdir(parents=True, exist_ok=True)
    if isinstance(original_payload, list):
        payload: object = rows
    elif isinstance(original_payload, dict):
        payload = {**original_payload, "rows": rows}
    else:
        payload = {"rows": rows}
    feedback_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _feedback_duplicate_key(row: dict[str, object]) -> tuple[str, str, str, str, str, str]:
    return (
        str(row.get("marketplace") or "").upper(),
        str(row.get("sku") or "").strip(),
        str(row.get("asin") or "").strip().upper(),
        str(row.get("action_id") or "").strip(),
        str(row.get("search_term_or_target") or "").strip().lower(),
        str(row.get("confirmed_at") or "")[:10],
    )


def build_ad_completion_feedback(row: dict[str, object]) -> dict[str, object]:
    if not isinstance(row, dict):
        raise ValueError("请求体必须是 JSON 对象。")
    base = {
        "marketplace": str(row.get("marketplace") or "").strip().upper(),
        "sku": str(row.get("sku") or "").strip(),
        "asin": str(row.get("asin") or "").strip().upper(),
        "product_name": str(row.get("product_name") or "").strip(),
        "search_term_or_target": str(row.get("search_term_or_target") or "").strip(),
        "suggested_action": str(row.get("suggested_action") or row.get("manual_action_taken") or "").strip(),
        "manual_action_taken": str(row.get("manual_action_taken") or row.get("suggested_action") or "").strip(),
        "confirmed_note": str(row.get("confirmed_note") or "").strip(),
        "report_date": str(row.get("report_date") or "").strip() or _latest_report_date(),
    }
    missing = [field for field in ["marketplace", "sku", "asin", "search_term_or_target", "suggested_action"] if not base[field]]
    if missing:
        raise ValueError("缺少必要字段：" + "、".join(missing))
    identified = add_action_identity(
        {**row, **base},
        base["manual_action_taken"] or base["suggested_action"],
        str(row.get("action_scope") or "").strip() or None,
    )
    if not is_executable_action(identified):
        raise ValueError("该动作不属于可执行广告动作，不能标记为已完成。")
    confirmed_at = str(row.get("confirmed_at") or "").strip() or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    note = base["confirmed_note"] or "网页勾选已完成；未满3天不判断，3天初查，7天正式复盘。"
    experiment_fields = {
        key: row.get(key)
        for key in [
            "experiment_type",
            "term_source",
            "evidence_level",
            "suggested_daily_budget",
            "suggested_bid_min",
            "suggested_bid_max",
            "test_days",
            "stop_loss_rule",
            "success_rule",
            "reason",
        ]
        if row.get(key) not in (None, "")
    }
    return {
        **base,
        **experiment_fields,
        "action_id": str(identified.get("action_id") or ""),
        "action_scope": str(identified.get("action_scope") or ""),
        "normalized_action": str(identified.get("normalized_action") or ""),
        "confirmed_status": "已执行",
        "confirmed_at": confirmed_at,
        "confirmed_note": note,
        "next_review": str(row.get("next_review") or "").strip() or _review_date_from(confirmed_at, 3),
        "cooldown_days": int(float(row.get("cooldown_days") or 7)),
    }


def append_ad_completion_feedback(row: dict[str, object], path: Path | None = None) -> tuple[dict[str, object], bool, int]:
    feedback_path = path or _feedback_input_path()
    original_payload, rows = _load_feedback_payload(feedback_path)
    feedback = build_ad_completion_feedback(row)
    new_key = _feedback_duplicate_key(feedback)
    for existing in rows:
        if _feedback_duplicate_key(existing) == new_key and str(existing.get("confirmed_status") or "") == "已执行":
            return existing, False, len(rows)
    rows.append(feedback)
    _write_feedback_payload(original_payload, rows, feedback_path)
    append_feedback_audit_event(
        "complete_action",
        feedback,
        previous_status="待确认",
        new_status="已执行",
    )
    return feedback, True, len(rows)


def append_ad_completion_feedback_batch(
    payload: dict[str, object],
    path: Path | None = None,
) -> tuple[list[dict[str, object]], list[bool], int]:
    actions = payload.get("actions")
    if actions is None:
        feedback, appended, row_count = append_ad_completion_feedback(payload, path=path)
        return [feedback], [appended], row_count
    if not isinstance(actions, list) or not actions:
        raise ValueError("批量完成记录必须包含 actions 数组。")

    feedbacks: list[dict[str, object]] = []
    appended_flags: list[bool] = []
    row_count = 0
    for index, action in enumerate(actions, start=1):
        if not isinstance(action, dict):
            raise ValueError(f"第 {index} 条完成记录必须是 JSON 对象。")
        feedback, appended, row_count = append_ad_completion_feedback(action, path=path)
        feedbacks.append(feedback)
        appended_flags.append(appended)
    return feedbacks, appended_flags, row_count


def cancel_ad_completion_feedback(row: dict[str, object], path: Path | None = None) -> tuple[list[dict[str, object]], int, int]:
    feedback_path = path or _feedback_input_path()
    original_payload, rows = _load_feedback_payload(feedback_path)
    expected = build_ad_completion_feedback(row)
    expected_key = _feedback_duplicate_key(expected)
    kept: list[dict[str, object]] = []
    removed: list[dict[str, object]] = []
    for existing in rows:
        if str(existing.get("confirmed_status") or "") == "已执行" and _feedback_duplicate_key(existing) == expected_key:
            removed.append(existing)
            continue
        kept.append(existing)
    if removed:
        _write_feedback_payload(original_payload, kept, feedback_path)
        for feedback in removed:
            append_feedback_audit_event(
                "cancel_completed_action",
                feedback,
                previous_status="已执行",
                new_status="待确认",
            )
    return removed, len(removed), len(kept)


def cancel_ad_completion_feedback_batch(
    payload: dict[str, object],
    path: Path | None = None,
) -> tuple[list[dict[str, object]], int, int]:
    actions = payload.get("actions")
    if actions is None:
        return cancel_ad_completion_feedback(payload, path=path)
    if not isinstance(actions, list) or not actions:
        raise ValueError("批量取消记录必须包含 actions 数组。")

    removed_rows: list[dict[str, object]] = []
    removed_count = 0
    row_count = 0
    for index, action in enumerate(actions, start=1):
        if not isinstance(action, dict):
            raise ValueError(f"第 {index} 条取消记录必须是 JSON 对象。")
        removed, count, row_count = cancel_ad_completion_feedback(action, path=path)
        removed_rows.extend(removed)
        removed_count += count
    return removed_rows, removed_count, row_count


def _record_data_date(row: dict[str, object]) -> str:
    for key in ("frontend_data_date", "checked_at", "generated_at"):
        value = str(row.get(key) or "").strip()
        if re.match(r"^\d{4}-\d{2}-\d{2}", value):
            return value[:10]
    text = str(row.get("frontend_data_freshness") or row.get("frontend_check_status") or "")
    match = re.search(r"\d{4}-\d{2}-\d{2}", text)
    return match.group(0) if match else ""


def _frontend_label(row: dict[str, object]) -> str:
    return " ".join(
        part
        for part in [
            str(row.get("marketplace") or "").strip().upper(),
            str(row.get("asin") or "").strip().upper(),
        ]
        if part
    )


def _has_competitor_samples(row: dict[str, object], minimum: int = 2) -> bool:
    try:
        count = int(float(row.get("frontend_competitor_count") or 0))
    except (TypeError, ValueError):
        count = 0
    competitors = row.get("frontend_competitors") or []
    if isinstance(competitors, list):
        count = max(count, len([item for item in competitors if isinstance(item, dict)]))
    return count >= minimum


def _is_today_frontend_evidence(
    row: dict[str, object],
    today: str | None = None,
    *,
    require_competitor_samples: bool = False,
) -> bool:
    today = today or _today_iso()
    method = str(row.get("frontend_check_method") or "").strip().lower()
    accepted_methods = {"urllib", "chrome-cdp", "chrome", "chrome-persistent", "playwright"}
    is_today = (
        str(row.get("frontend_check_status") or "") == "已自动检查"
        and method in accepted_methods
        and _record_data_date(row) == today
        and not boolish_flag(row.get("frontend_cache_used"))
    )
    if not is_today:
        return False
    if require_competitor_samples:
        return _has_competitor_samples(row)
    return True


def _is_today_live_chrome_cdp(
    row: dict[str, object],
    today: str | None = None,
    *,
    require_competitor_samples: bool = False,
) -> bool:
    return _is_today_frontend_evidence(row, today, require_competitor_samples=require_competitor_samples)


def _load_frontend_queue_rows(priority: str = "") -> list[dict[str, object]]:
    analysis_path = OUTPUT_DIR / "latest_analysis.json"
    try:
        payload = json.loads(analysis_path.read_text(encoding="utf-8"))
    except Exception:
        return []
    rows: list[dict[str, object]] = []
    seen: set[tuple[str, str, str]] = set()
    for result in payload.get("marketplace_results", []) or []:
        if not isinstance(result, dict):
            continue
        view = result.get("report_view_snapshot") or {}
        if not isinstance(view, dict):
            continue
        for row in view.get("frontend_check_queue_rows", []) or []:
            if not isinstance(row, dict):
                continue
            key = (
                str(row.get("marketplace") or "").strip().upper(),
                str(row.get("sku") or "").strip(),
                str(row.get("asin") or "").strip().upper(),
            )
            if key[0] and key[2] and key not in seen:
                seen.add(key)
                rows.append(row)
    priority_filter = str(priority or "").strip().upper()
    if priority_filter:
        rows = [row for row in rows if str(row.get("priority") or "").strip().upper() == priority_filter]
    return rows


def _frontend_refresh_needed_summary(priority: str = "", *, require_competitor_samples: bool = False) -> dict[str, object]:
    rows = _load_frontend_queue_rows(priority=priority)
    today = _today_iso()
    fresh: list[str] = []
    stale: list[str] = []
    for row in rows:
        label = _frontend_label(row)
        if _is_today_frontend_evidence(row, today, require_competitor_samples=require_competitor_samples):
            fresh.append(label)
        else:
            stale.append(label)
    return {
        "frontend_queue_total": len(rows),
        "frontend_today_live_count": len(fresh),
        "frontend_refresh_needed_count": len(stale),
        "frontend_today_live_labels": fresh,
        "frontend_refresh_needed_labels": stale,
    }


def _sellersprite_reverse_needed_summary(priority: str = "") -> dict[str, object]:
    base_rows = _load_frontend_queue_rows(priority=priority)
    discovery_existing = load_competitor_discovery_records(SELLERSPRITE_COMPETITOR_DISCOVERY_CACHE_PATH)
    discovery_cached: list[str] = []
    discovery_missing: list[str] = []
    discovery_seen: set[tuple[str, str, str]] = set()
    for row in base_rows:
        key = (
            str(row.get("marketplace") or "").strip().upper(),
            str(row.get("sku") or "").strip(),
            str(row.get("asin") or "").strip().upper(),
        )
        if not key[0] or not key[2] or key in discovery_seen:
            continue
        discovery_seen.add(key)
        label = _frontend_label(row)
        recommended_steps = str(row.get("market_survey_recommended_fetch_steps") or "")
        discovery_gap = "补抓卖家精灵竞品池" in recommended_steps or not recommended_steps
        if discovery_gap and discovery_record_needs_refresh(discovery_existing.get(key), max_age_days=7):
            discovery_missing.append(label)
        else:
            discovery_cached.append(label)
    rows = [
        *base_rows,
        *_sellersprite_competitor_rows(
            base_rows,
            competitor_limit_per_product=3,
            competitor_discovery_records=discovery_existing,
        ),
    ]
    existing = load_sellersprite_records(OUTPUT_DIR / "sellersprite_reverse_asin_results.json")
    cached: list[str] = []
    missing: list[str] = []
    seen: set[tuple[str, str]] = set()
    for row in rows:
        key = (
            str(row.get("marketplace") or "").strip().upper(),
            str(row.get("asin") or "").strip().upper(),
        )
        if not key[0] or not key[1] or key in seen:
            continue
        seen.add(key)
        label = _frontend_label(row)
        recommended_steps = str(row.get("market_survey_recommended_fetch_steps") or "")
        source_role = str(row.get("source_role") or "own").strip()
        if source_role == "competitor":
            reverse_gap = "补抓竞品 ASIN 反查" in recommended_steps or not recommended_steps
        else:
            reverse_gap = "补抓卖家精灵自己 ASIN" in recommended_steps or not recommended_steps
        if not reverse_gap:
            cached.append(label)
        elif _sellersprite_cached_recent(
            existing.get(key),
            max_age_days=sellersprite_cache_max_age_days_for_row(row, competitor_cache_days=7),
        ):
            cached.append(label)
        else:
            missing.append(label)
    return {
        "sellersprite_queue_total": len(seen),
        "sellersprite_cached_count": len(cached),
        "sellersprite_missing_count": len(missing) + len(discovery_missing),
        "sellersprite_cached_labels": cached,
        "sellersprite_missing_labels": missing,
        "sellersprite_discovery_queue_total": len(discovery_seen),
        "sellersprite_discovery_cached_count": len(discovery_cached),
        "sellersprite_discovery_missing_count": len(discovery_missing),
        "sellersprite_discovery_cached_labels": discovery_cached,
        "sellersprite_discovery_missing_labels": discovery_missing,
    }


def _sellersprite_progress_from_line(line: str) -> dict[str, object]:
    text = line.strip()
    discovery_start_match = re.search(r"\[sellersprite-discovery\]\s+(\d+)/(\d+)\s+([A-Z]{2})\s+([A-Z0-9]{8,})\s+开始竞品发现", text)
    if discovery_start_match:
        index = int(discovery_start_match.group(1))
        total = int(discovery_start_match.group(2))
        marketplace = discovery_start_match.group(3)
        asin = discovery_start_match.group(4)
        return {
            "event": "discovery_start",
            "sellersprite_discovery_step": index,
            "sellersprite_discovery_total": total,
            "current_label": f"{marketplace} {asin}",
            "message": f"卖家精灵竞品发现进行中：{index}/{total} {marketplace} {asin}",
        }
    discovery_finish_match = re.search(
        r"\[sellersprite-discovery\]\s+([A-Z]{2})\s+([A-Z0-9]{8,})\s+(.+?)\s+competitors=(\d+)\s+error=(.*)",
        text,
    )
    if discovery_finish_match:
        marketplace = discovery_finish_match.group(1)
        asin = discovery_finish_match.group(2)
        status = discovery_finish_match.group(3).strip()
        competitor_count = int(discovery_finish_match.group(4))
        error = discovery_finish_match.group(5).strip()
        return {
            "event": "discovery_finish",
            "current_label": f"{marketplace} {asin}",
            "competitor_discovery_status": status,
            "competitor_discovery_count": competitor_count,
            "last_error": error,
            "message": f"卖家精灵竞品发现已返回：{marketplace} {asin} {status}，竞品 {competitor_count}",
        }
    discovery_skipped_match = re.search(r"\[sellersprite-discovery\]\s+skipped cached rows:\s+(\d+)", text)
    if discovery_skipped_match:
        skipped = int(discovery_skipped_match.group(1))
        return {
            "event": "discovery_skip_cached",
            "sellersprite_discovery_skipped_cached": skipped,
            "message": f"卖家精灵竞品发现跳过已有缓存：{skipped} 个",
        }
    discovery_wrote_match = re.search(r"\[sellersprite-discovery\]\s+wrote\s+(.+?);\s+success=(\d+)/(\d+)\s+cached_cover=(\d+)/(\d+)", text)
    if discovery_wrote_match:
        success = int(discovery_wrote_match.group(2))
        total = int(discovery_wrote_match.group(3))
        covered = int(discovery_wrote_match.group(4))
        cover_total = int(discovery_wrote_match.group(5))
        return {
            "event": "discovery_wrote",
            "sellersprite_discovery_success_count": success,
            "sellersprite_discovery_processed_count": total,
            "sellersprite_discovery_cached_cover": covered,
            "sellersprite_discovery_cached_cover_total": cover_total,
            "message": f"卖家精灵竞品发现写入缓存：成功 {success}/{total}，覆盖 {covered}/{cover_total}",
        }
    start_match = re.search(r"\[sellersprite\]\s+(\d+)/(\d+)\s+([A-Z]{2})\s+([A-Z0-9]{8,})\s+开始反查", text)
    if start_match:
        index = int(start_match.group(1))
        total = int(start_match.group(2))
        marketplace = start_match.group(3)
        asin = start_match.group(4)
        return {
            "event": "start",
            "step": index,
            "total_steps": total,
            "current_label": f"{marketplace} {asin}",
            "message": f"卖家精灵后台反查进行中：{index}/{total} {marketplace} {asin}",
        }
    finish_match = re.search(
        r"\[sellersprite\]\s+([A-Z]{2})\s+([A-Z0-9]{8,})\s+(.+?)\s+captured=(\d+)\s+total=([^\s]*)\s+error=(.*)",
        text,
    )
    if finish_match:
        marketplace = finish_match.group(1)
        asin = finish_match.group(2)
        status = finish_match.group(3).strip()
        captured = int(finish_match.group(4))
        reported_total = finish_match.group(5)
        error = finish_match.group(6).strip()
        return {
            "event": "finish",
            "current_label": f"{marketplace} {asin}",
            "record_status": status,
            "captured_count": captured,
            "reported_total": reported_total,
            "last_error": error,
            "message": f"卖家精灵后台反查已返回：{marketplace} {asin} {status}，抓词 {captured}",
        }
    skipped_match = re.search(r"\[sellersprite\]\s+skipped cached rows:\s+(\d+)", text)
    if skipped_match:
        skipped = int(skipped_match.group(1))
        return {
            "event": "skip_cached",
            "sellersprite_skipped_cached": skipped,
            "message": f"卖家精灵后台反查跳过已有缓存：{skipped} 个",
        }
    wrote_match = re.search(r"\[sellersprite\]\s+wrote\s+(.+?);\s+success=(\d+)/(\d+)", text)
    if wrote_match:
        success = int(wrote_match.group(2))
        total = int(wrote_match.group(3))
        return {
            "event": "wrote",
            "sellersprite_success_count": success,
            "sellersprite_processed_count": total,
            "message": f"卖家精灵后台反查写入缓存：成功 {success}/{total}",
        }
    return {"event": "log", "message": text} if text else {}


def _run_sellersprite_command_with_progress(command: list[str], timeout: int, label: str) -> subprocess.CompletedProcess[str]:
    stdout_parts: list[str] = []
    stderr_parts: list[str] = []
    started_at = time.time()
    processed = 0
    success = 0
    failed = 0
    last_progress: dict[str, object] = {}
    try:
        process = subprocess.Popen(
            command,
            cwd=ROOT,
            env=_subprocess_env(),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            bufsize=1,
        )
    except OSError as exc:
        return subprocess.CompletedProcess(command, 127, "", f"cannot start command: {exc}")

    deadline = time.monotonic() + timeout
    last_status_at = 0.0
    streams = [stream for stream in (process.stdout, process.stderr) if stream is not None]

    def publish(message: str | None = None, *, force: bool = False, extra: dict[str, object] | None = None) -> None:
        nonlocal last_status_at
        now = time.monotonic()
        if not force and now - last_status_at < 1.0:
            return
        last_status_at = now
        payload = {
            **_sellersprite_async_status,
            "running": True,
            "message": message or str(last_progress.get("message") or f"卖家精灵后台反查运行中：{label}"),
            "detail": " ".join(command),
            "started_at_epoch": started_at,
            "elapsed_seconds": max(0, int(time.time() - started_at)),
            "stdout_tail": "".join(stdout_parts)[-4000:],
            "stderr_tail": "".join(stderr_parts)[-4000:],
            "sellersprite_processed_count": processed,
            "sellersprite_success_count": success,
            "sellersprite_failed_count": failed,
            "status_scope": "sellersprite_async",
        }
        if extra:
            payload.update(extra)
        _sync_sellersprite_async_status(payload)

    try:
        while streams and process.poll() is None:
            if time.monotonic() > deadline:
                process.kill()
                raise subprocess.TimeoutExpired(command, timeout, output="".join(stdout_parts), stderr="".join(stderr_parts))
            readable, _, _ = select.select(streams, [], [], 0.25)
            if not readable:
                publish()
                continue
            for stream in readable:
                line = stream.readline()
                if not line:
                    streams.remove(stream)
                    continue
                if stream is process.stdout:
                    stdout_parts.append(line)
                    progress = _sellersprite_progress_from_line(line)
                    if progress:
                        last_progress.update(progress)
                        if progress.get("event") == "finish":
                            processed += 1
                            if str(progress.get("record_status") or "") == "已抓取":
                                success += 1
                            else:
                                failed += 1
                        publish(str(progress.get("message") or "").strip() or None, force=True, extra=progress)
                    else:
                        publish(force=True)
                else:
                    stderr_parts.append(line)
                    publish(force=True, extra={"message": f"卖家精灵后台反查 stderr：{line.strip()[:120]}"})
        remaining_stdout, remaining_stderr = process.communicate(timeout=max(0.1, deadline - time.monotonic()))
    except subprocess.TimeoutExpired:
        process.kill()
        remaining_stdout, remaining_stderr = process.communicate()
        stdout_parts.append(remaining_stdout or "")
        stderr_parts.append(remaining_stderr or "")
        raise subprocess.TimeoutExpired(command, timeout, output="".join(stdout_parts), stderr="".join(stderr_parts))
    stdout_parts.append(remaining_stdout or "")
    stderr_parts.append(remaining_stderr or "")
    publish(force=True)
    return subprocess.CompletedProcess(command, process.returncode or 0, "".join(stdout_parts), "".join(stderr_parts))


def _run_sellersprite_reverse_async(command: list[str], label: str) -> None:
    log_path = OUTPUT_DIR / "sellersprite_reverse_async.log"
    started_at = time.time()
    try:
        _sync_sellersprite_async_status(
            {
                "running": True,
                "message": f"卖家精灵后台反查已启动：{label}",
                "detail": " ".join(command),
                "started_at_epoch": started_at,
                "elapsed_seconds": 0,
                "status_scope": "sellersprite_async",
            }
        )
        completed = _run_sellersprite_command_with_progress(command, timeout=1200, label=label)
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        log_path.write_text(
            "\n".join(
                [
                    f"[command] {' '.join(command)}",
                    f"[returncode] {completed.returncode}",
                    "[stdout]",
                    completed.stdout,
                    "[stderr]",
                    completed.stderr,
                ]
            ),
            encoding="utf-8",
        )
        refresh_returncode = None
        if completed.returncode == 0:
            if _lock.acquire(blocking=False):
                try:
                    _sync_sellersprite_async_status(
                        {
                            **_sellersprite_async_status,
                            "running": True,
                            "message": "卖家精灵后台反查已完成，正在刷新 HTML / Excel 报告。",
                            "elapsed_seconds": max(0, int(time.time() - started_at)),
                            "status_scope": "sellersprite_async",
                        }
                    )
                    refresh = _run_command([sys.executable, "main.py", "--marketplace", "ALL"], timeout=180)
                    refresh_returncode = refresh.returncode
                finally:
                    _lock.release()
        elapsed = max(0, int(time.time() - started_at))
        try:
            seller_summary = _sellersprite_reverse_needed_summary()
        except Exception:
            seller_summary = {}
        _sync_sellersprite_async_status(
            {
                "running": False,
                "message": (
                    "卖家精灵后台反查完成，报告已刷新。"
                    if completed.returncode == 0 and refresh_returncode == 0
                    else "卖家精灵后台反查完成；报告将在下次刷新后显示最新融合结果。"
                    if completed.returncode == 0
                    else "卖家精灵后台反查失败；已保留旧缓存，不影响前台检查。"
                ),
                "returncode": completed.returncode,
                "refresh_returncode": refresh_returncode,
                "elapsed_seconds": elapsed,
                "stdout_tail": completed.stdout[-2000:],
                "stderr_tail": completed.stderr[-2000:],
                "log_path": str(log_path.relative_to(ROOT)),
                "status_scope": "sellersprite_async",
                **seller_summary,
            }
        )
    except subprocess.TimeoutExpired as exc:
        _sync_sellersprite_async_status(
            _timeout_payload(
                "卖家精灵后台反查超时；已保留旧缓存，不影响前台检查。",
                exc,
                status_scope="sellersprite_async",
                failure_mode="sellersprite_reverse_timeout",
            )
        )
    finally:
        _sellersprite_async_lock.release()


def _start_sellersprite_reverse_async(*, priority: str = "", params: dict[str, str] | None = None) -> dict[str, object]:
    if not _sellersprite_async_lock.acquire(blocking=False):
        return {
            **_sellersprite_async_status,
            "running": True,
            "message": "卖家精灵后台反查已在运行。",
            "status_scope": "sellersprite_async",
        }
    summary = _sellersprite_reverse_needed_summary(priority=priority) if not params else {}
    if not params and int(summary.get("sellersprite_missing_count") or 0) == 0:
        snapshot_count = _snapshot_cached_sellersprite_queue(priority=priority)
        payload = {
            "running": False,
            "message": "卖家精灵后台反查无需运行：当前前台队列已有缓存。",
            "returncode": 0,
            "status_scope": "sellersprite_async",
            "sellersprite_history_snapshot_count": snapshot_count,
            **summary,
        }
        _sync_sellersprite_async_status(payload)
        _sellersprite_async_lock.release()
        return payload
    command = _sellersprite_reverse_command(priority=priority, params=params)
    label = "单产品" if params else "当前前台队列"
    payload = {
        "running": True,
        "message": f"卖家精灵后台反查已排队：{label}",
        "detail": " ".join(command),
        "started_at_epoch": time.time(),
        "elapsed_seconds": 0,
        "status_scope": "sellersprite_async",
        **summary,
    }
    _sync_sellersprite_async_status(payload)
    thread = threading.Thread(target=_run_sellersprite_reverse_async, args=(command, label), daemon=True)
    thread.start()
    return payload


def _frontend_refresh_result_summary() -> dict[str, object]:
    path = OUTPUT_DIR / "frontend_check_results.json"
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    if not isinstance(payload, dict):
        return {}
    summary = payload.get("refresh_summary") or {}
    return dict(summary) if isinstance(summary, dict) else {}


def _snapshot_cached_sellersprite_queue(priority: str = "") -> int:
    base_rows = _load_frontend_queue_rows(priority=priority)
    discovery_existing = load_competitor_discovery_records(SELLERSPRITE_COMPETITOR_DISCOVERY_CACHE_PATH)
    rows = [
        *base_rows,
        *_sellersprite_competitor_rows(
            base_rows,
            competitor_limit_per_product=3,
            competitor_discovery_records=discovery_existing,
        ),
    ]
    existing = load_sellersprite_records(OUTPUT_DIR / "sellersprite_reverse_asin_results.json")
    snapshots: list[dict] = []
    seen: set[tuple[str, str, str]] = set()
    today = _today_iso()
    for row in rows:
        market = str(row.get("marketplace") or "").strip().upper()
        asin = str(row.get("asin") or "").strip().upper()
        parent_asin = str(row.get("parent_asin") or "").strip().upper()
        key = (market, asin)
        seen_key = (market, asin, parent_asin)
        if not market or not asin or seen_key in seen:
            continue
        seen.add(seen_key)
        record = existing.get(key)
        if record:
            snapshot = dict(record)
            for field in [
                "marketplace",
                "sku",
                "asin",
                "source_role",
                "parent_marketplace",
                "parent_sku",
                "parent_asin",
                "parent_product_name",
                "competitor_discovery_source",
                "competitor_pool_confidence",
                "competitor_source_keyword",
            ]:
                value = row.get(field)
                if value not in (None, ""):
                    snapshot[field] = value
            data_date = str(snapshot.get("data_date") or snapshot.get("checked_at") or "")[:10]
            status = "已抓取" if data_date == today else "沿用缓存"
            snapshots.append({**snapshot, "seller_sprite_check_status": status})
        elif str(row.get("source_role") or "") == "competitor":
            snapshots.append(
                {
                    **row,
                    "seller_sprite_check_status": "竞品池快照",
                    "data_date": today,
                    "checked_at": datetime.now().isoformat(timespec="seconds"),
                    "source": "sellersprite_competitor_pool_snapshot",
                    "keywords": [],
                }
            )
    upsert_sellersprite_history(snapshots)
    return len(snapshots)


def _frontend_queue_status_summary(priority: str = "") -> dict[str, object]:
    rows = _load_frontend_queue_rows(priority=priority)
    total = len(rows)
    if not total:
        return {}

    live_checked: list[str] = []
    stability_passed: list[str] = []
    pending: list[str] = []
    today = _today_iso()
    for row in rows:
        label = _frontend_label(row)
        is_live_checked = _is_today_frontend_evidence(row, today)
        is_stability_passed = (
            is_live_checked
            and boolish_flag(row.get("frontend_stability_passed"))
            and int(row.get("frontend_stability_total_attempts") or 0) >= 20
            and float(row.get("frontend_stability_success_rate") or 0) >= 0.8
        )
        if is_live_checked:
            live_checked.append(label)
            if is_stability_passed:
                stability_passed.append(label)
        else:
            pending.append(label)
    return {
        "frontend_queue_total": total,
        "frontend_live_passed_count": len(live_checked),
        "frontend_pending_count": len(pending),
        "frontend_queue_success_rate": len(live_checked) / total if total else 0,
        "frontend_queue_acceptance_threshold": FRONTEND_QUEUE_ACCEPTANCE_THRESHOLD,
        "frontend_queue_passed": (len(live_checked) / total if total else 0) >= FRONTEND_QUEUE_ACCEPTANCE_THRESHOLD,
        "frontend_stability_passed_count": len(stability_passed),
        "frontend_live_passed_labels": live_checked,
        "frontend_stability_passed_labels": stability_passed,
        "frontend_pending_labels": pending,
    }


def _frontend_retry_status_message(summary: dict[str, object]) -> str:
    total = int(summary.get("frontend_queue_total") or 0)
    passed = int(summary.get("frontend_live_passed_count") or 0)
    pending = int(summary.get("frontend_pending_count") or 0)
    refresh_total = int(summary.get("frontend_refresh_total") or 0)
    refreshed = int(summary.get("frontend_refresh_live_checked") or 0)
    skipped = int(summary.get("frontend_refresh_skipped") or 0)
    cache_used = int(summary.get("frontend_refresh_cache_used") or 0)
    failed = int(summary.get("frontend_refresh_failed") or 0)
    if not total:
        return ""
    success_rate = float(summary.get("frontend_queue_success_rate") or 0)
    threshold = float(summary.get("frontend_queue_acceptance_threshold") or FRONTEND_QUEUE_ACCEPTANCE_THRESHOLD)
    pending_labels = [str(item) for item in summary.get("frontend_pending_labels") or [] if str(item)]
    pending_text = "、".join(pending_labels[:3])
    if len(pending_labels) > 3:
        pending_text += f" 等 {len(pending_labels)} 个"
    if refresh_total and skipped == refresh_total and not refreshed and not cache_used and not failed:
        return f"无需刷新：本轮队列 {skipped}/{refresh_total} 已有今日证据。"
    if refresh_total:
        prefix = "调查完成" if success_rate >= threshold else "调查待补"
        message = f"{prefix}：本轮队列 {passed}/{total}，新读 {refreshed}，失败 {failed}"
        if pending:
            message += f"，待补 {pending}"
        elif cache_used:
            message += f"，缓存 {cache_used}"
        message += "。"
        if pending and pending_text:
            message += f" 待补：{pending_text}"
        return message
    if not passed:
        return ""
    if pending:
        prefix = "调查完成" if success_rate >= threshold else "调查待补"
        return (
            f"{prefix}：本轮队列 {passed}/{total}，待补 {pending}。"
            + (f" 待补：{pending_text}" if pending_text else "")
        )
    return f"调查完成：本轮队列 {passed}/{total}。"


def _frontend_async_status_message(summary: dict[str, object]) -> str:
    message = _frontend_retry_status_message(summary)
    if not message:
        return ""
    return (
        message.replace("当前前台队列刷新", "P0 前台后台检查")
        .replace("当前前台队列", "P0 前台后台检查")
        .replace("无需刷新", "P0 前台后台检查无需刷新")
    )


def _timeout_payload(message: str, exc: subprocess.TimeoutExpired, **extra: object) -> dict[str, object]:
    stdout = exc.stdout or ""
    stderr = exc.stderr or ""
    if isinstance(stdout, bytes):
        stdout = stdout.decode("utf-8", errors="replace")
    if isinstance(stderr, bytes):
        stderr = stderr.decode("utf-8", errors="replace")
    return _fresh_status(
        message,
        returncode=124,
        stdout_tail=str(stdout)[-4000:],
        stderr_tail=str(stderr)[-4000:],
        report_url=f"{REPORT_BASE_URL}/latest_recommendations.html",
        **extra,
    )


def _safe_upload_name(filename: str) -> str:
    name = Path(str(filename or "")).name.strip()
    name = re.sub("[^A-Za-z0-9._()\\-\u4e00-\u9fff ]+", "_", name)
    name = re.sub(r"\\s+", " ", name).strip(" .")
    return name or "uploaded_file"


def _validate_upload_name(filename: str, size: int) -> tuple[str, str]:
    safe_name = _safe_upload_name(filename)
    suffix = Path(safe_name).suffix.lower()
    if suffix not in ALLOWED_UPLOAD_SUFFIXES:
        return safe_name, f"只允许上传 .csv 或 .xlsx：{safe_name}"
    if size > MAX_UPLOAD_BYTES:
        return safe_name, f"文件超过 50MB：{safe_name}"
    if size <= 0:
        return safe_name, f"文件为空：{safe_name}"
    return safe_name, ""


def _unique_inbox_path(filename: str) -> Path:
    INBOX_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = _safe_upload_name(filename)
    candidate = INBOX_DIR / safe_name
    if not candidate.exists():
        return candidate
    stem = candidate.stem
    suffix = candidate.suffix
    stamp = time.strftime("%Y%m%d_%H%M%S")
    index = 1
    while True:
        renamed = INBOX_DIR / f"{stem}_{stamp}_{index}{suffix}"
        if not renamed.exists():
            return renamed
        index += 1


def _target_config(kind: str) -> dict[str, object]:
    config = CONFIG_UPLOAD_TARGETS.get(str(kind or "").strip().lower())
    if not config:
        raise ValueError("未知配置类型。")
    return config


def _validate_config_upload_name(kind: str, filename: str, size: int) -> tuple[str, str]:
    safe_name = _safe_upload_name(filename)
    suffix = Path(safe_name).suffix.lower()
    if suffix != ".xlsx":
        return safe_name, f"配置文件只允许上传 .xlsx：{safe_name}"
    if size > MAX_UPLOAD_BYTES:
        return safe_name, f"文件超过 50MB：{safe_name}"
    if size <= 0:
        return safe_name, f"文件为空：{safe_name}"
    allowed = set(_target_config(kind).get("allowed_filenames") or set())
    if allowed and safe_name not in allowed:
        return safe_name, f"文件名必须是 {', '.join(sorted(allowed))}：{safe_name}"
    return safe_name, ""


def _workbook_headers(path: Path, sheet_name: str) -> set[str]:
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            return set()
        ws = wb[sheet_name]
        return {str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1), [])}
    finally:
        wb.close()


def _validate_config_workbook(kind: str, path: Path) -> list[str]:
    config = _target_config(kind)
    errors: list[str] = []
    required_sheets = config.get("required_sheets") or {}
    for sheet_name, required_headers in dict(required_sheets).items():
        headers = _workbook_headers(path, str(sheet_name))
        missing = sorted(set(required_headers) - headers)
        if not headers:
            errors.append(f"缺少 sheet：{sheet_name}")
        elif missing:
            errors.append(f"{sheet_name} 缺少列：{', '.join(missing)}")
    return errors


def _archive_config_target(target_path: Path, kind: str) -> Path | None:
    if not target_path.exists():
        return None
    CONFIG_ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    stamp = time.strftime("%Y%m%d_%H%M%S")
    archive_path = CONFIG_ARCHIVE_DIR / f"{target_path.stem}_{kind}_{stamp}{target_path.suffix}"
    shutil.copy2(target_path, archive_path)
    return archive_path


def _apply_pending_config(kind: str) -> tuple[dict[str, object], subprocess.CompletedProcess[str] | None]:
    config = _target_config(kind)
    review_path = Path(config["review_path"])
    target_path = Path(config["target_path"])
    if not review_path.exists():
        raise FileNotFoundError(f"没有待应用配置：{review_path.relative_to(ROOT)}")
    errors = _validate_config_workbook(kind, review_path)
    if errors:
        raise ValueError("；".join(errors))

    archive_path = _archive_config_target(target_path, kind)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(review_path, target_path)
    review_path.unlink()

    completed: subprocess.CompletedProcess[str] | None = None
    if config.get("refresh_reports_after_apply"):
        completed = _run_command([sys.executable, "scripts/validate_showcase_mvp.py"], timeout=900)
    return {
        "kind": kind,
        "target_path": str(target_path.relative_to(ROOT)),
        "archive_path": str(archive_path.relative_to(ROOT)) if archive_path else "",
        "refresh_returncode": completed.returncode if completed else None,
        "stdout_tail": completed.stdout[-4000:] if completed else "",
        "stderr_tail": completed.stderr[-4000:] if completed else "",
    }, completed


def _run_submission_preflight() -> subprocess.CompletedProcess[str]:
    return _run_command(
        [
            sys.executable,
            "scripts/check_daily_update_preflight.py",
            "--allow-inbox-business-files",
        ],
        timeout=60,
    )


def _save_uploaded_files(form: cgi.FieldStorage) -> tuple[list[dict[str, object]], list[str]]:
    fields = form["files"] if "files" in form else []
    if not isinstance(fields, list):
        fields = [fields]
    staged: list[tuple[str, str, bytes]] = []
    errors: list[str] = []
    for field in fields:
        filename = str(getattr(field, "filename", "") or "")
        if not filename:
            continue
        raw = field.file.read(MAX_UPLOAD_BYTES + 1)
        safe_name, validation_error = _validate_upload_name(filename, len(raw))
        if validation_error:
            errors.append(validation_error)
            continue
        staged.append((filename, safe_name, raw))
    if errors:
        return [], errors

    saved: list[dict[str, object]] = []
    UPLOAD_STAGING_DIR.mkdir(parents=True, exist_ok=True)
    for filename, safe_name, raw in staged:
        staging_path = UPLOAD_STAGING_DIR / f"{int(time.time() * 1000)}_{safe_name}"
        staging_path.write_bytes(raw)
        target_path = _unique_inbox_path(safe_name)
        shutil.move(str(staging_path), str(target_path))
        saved.append(
            {
                "original_filename": filename,
                "saved_filename": target_path.name,
                "path": str(target_path.relative_to(ROOT)),
                "size": len(raw),
            }
        )
    return saved, errors


def _save_config_upload(form: cgi.FieldStorage, kind: str) -> tuple[dict[str, object], list[str]]:
    fields = form["file"] if "file" in form else []
    if isinstance(fields, list):
        fields = [field for field in fields if getattr(field, "filename", "")]
        if len(fields) > 1:
            return {}, ["配置上传一次只允许一个文件。"]
        field = fields[0] if fields else None
    else:
        field = fields
    filename = str(getattr(field, "filename", "") or "") if field is not None else ""
    if not filename:
        return {}, ["没有收到配置文件。"]
    raw = field.file.read(MAX_UPLOAD_BYTES + 1)
    safe_name, validation_error = _validate_config_upload_name(kind, filename, len(raw))
    if validation_error:
        return {}, [validation_error]

    config = _target_config(kind)
    CONFIG_REVIEW_DIR.mkdir(parents=True, exist_ok=True)
    review_path = Path(config["review_path"])
    staging_path = CONFIG_REVIEW_DIR / f"{int(time.time() * 1000)}_{safe_name}"
    staging_path.write_bytes(raw)
    errors = _validate_config_workbook(kind, staging_path)
    if errors:
        staging_path.unlink(missing_ok=True)
        return {}, errors
    shutil.move(str(staging_path), str(review_path))
    return {
        "kind": kind,
        "label": str(config.get("label") or kind),
        "original_filename": filename,
        "saved_filename": review_path.name,
        "path": str(review_path.relative_to(ROOT)),
        "size": len(raw),
        "requires_confirm": bool(config.get("requires_confirm")),
    }, []


def _run_daily_update() -> None:
    global _last_result
    python = sys.executable
    try:
        _set_progress("运行 daily update", 1, 1, f"{python} scripts/run_daily_update.py")
        completed = _run_command_with_status(
            [python, "scripts/run_daily_update.py"],
            timeout=900,
            step=1,
            total_steps=1,
            message="运行 daily update",
        )
        payload = _completed_payload(
            completed,
            "daily update 完成，报告已刷新。",
            "daily update 失败，请查看输出摘要。",
        )
        _sync_last_result(payload)
        if completed.returncode == 0:
            try:
                _start_p0_frontend_async_if_needed()
            except Exception as exc:
                _sync_frontend_async_status(
                    {
                        "running": False,
                        "message": f"P0 前台后台检查启动失败：{exc}。日报已完成，前台证据保留缓存。",
                        "returncode": 1,
                        "status_scope": "frontend_async",
                        "failure_mode": "frontend_async_start_failed",
                    }
                )
    except subprocess.TimeoutExpired as exc:
        _sync_last_result(_timeout_payload("daily update 超时，已保留当前报告。", exc))
    finally:
        _lock.release()


def _run_report_refresh() -> None:
    global _last_result
    python = sys.executable
    try:
        _set_progress("刷新报告", 1, 1, f"{python} main.py --marketplace ALL")
        completed = _run_command_with_status(
            [python, "main.py", "--marketplace", "ALL"],
            timeout=180,
            step=1,
            total_steps=1,
            message="刷新报告",
        )
        payload = _completed_payload(
            completed,
            "报告刷新完成，广告完成记录已进入冷却和复盘。",
            "报告刷新失败，请查看输出摘要。",
            status_scope="report_refresh",
            frontend_async_status={
                "running": False,
                "message": "轻量报告刷新未访问 Amazon 前台；如需实时前台证据，请用前台证据区刷新按钮。",
                "returncode": 0,
                "status_scope": "frontend_async",
            },
        )
        _sync_last_result(payload)
    except subprocess.TimeoutExpired as exc:
        _sync_last_result(_timeout_payload("报告刷新超时，已保留当前报告。", exc, status_scope="report_refresh"))
    finally:
        _lock.release()


def _run_p0_frontend_async() -> None:
    python = sys.executable
    browser_enabled = _browser_frontend_enabled()
    refresh_command = [python, "main.py", "--marketplace", "ALL"]
    stdout_parts: list[str] = []
    stderr_parts: list[str] = []
    try:
        needed_summary = {
            **_frontend_refresh_needed_summary(priority="P0", require_competitor_samples=browser_enabled),
            **_sellersprite_reverse_needed_summary(priority="P0"),
        }
        total = int(needed_summary.get("frontend_queue_total") or 0)
        needed = int(needed_summary.get("frontend_refresh_needed_count") or 0)
        sellersprite_needed = int(needed_summary.get("sellersprite_missing_count") or 0)
        frontend_command = _frontend_command(
            priority="P0",
            require_competitor_samples=browser_enabled,
            search_policy="always",
            limit=_frontend_refresh_batch_limit(needed),
        )
        if total == 0:
            _sync_frontend_async_status(
                {
                    "running": False,
                    "message": "P0 前台后台检查无需运行：当前没有 P0 前台队列。",
                    "returncode": 0,
                    "status_scope": "frontend_async",
                    **needed_summary,
                }
            )
            return
        if total:
            try:
                _start_sellersprite_reverse_async(priority="P0")
            except Exception as exc:
                _sync_sellersprite_async_status(
                    {
                        "running": False,
                        "message": f"卖家精灵后台反查启动失败：{exc}",
                        "returncode": 1,
                        "status_scope": "sellersprite_async",
                    }
                )
        if needed == 0:
            summary = {**needed_summary, **_frontend_queue_status_summary(priority="P0")}
            message = (
                f"P0 前台后台检查无需刷新：{total}/{total} 个已有今日前台证据；"
                + ("卖家精灵反查交给后台运行。" if sellersprite_needed else "未访问 Amazon。")
            )
            _sync_frontend_async_status(
                {
                    "running": False,
                    "message": message,
                    "returncode": 0,
                    "status_scope": "frontend_async",
                    "failure_mode": "frontend_refresh_not_needed",
                    "frontend_refresh_total": total,
                    "frontend_refresh_live_checked": 0,
                    "frontend_refresh_skipped": total,
                    "frontend_refresh_cache_used": 0,
                    "frontend_refresh_failed": 0,
                    **summary,
                }
            )
            return
        if browser_enabled:
            _set_frontend_async_progress("检查本机 Chrome CDP 会话", 1, 2, CHROME_CDP_ENDPOINT)
        if browser_enabled and not _start_chrome_cdp_if_needed(CHROME_CDP_ENDPOINT):
            _sync_frontend_async_status(
                {
                    "running": False,
                    "message": "P0 前台后台检查未启动：本机 Chrome CDP 端口不可用。日报已完成，前台证据保留缓存。",
                    "returncode": 1,
                    "status_scope": "frontend_async",
                    "failure_mode": "chrome_cdp_unavailable",
                    **needed_summary,
                }
            )
            return
        _set_frontend_async_progress(f"P0 前台后台检查（{_frontend_method_label()}）", 1, 2, " ".join(frontend_command))
        frontend = _run_frontend_command_with_progress(
            frontend_command,
            timeout=600,
            step=1,
            total_steps=2,
            progress_writer=_set_frontend_async_progress,
        )
        stdout_parts.append(frontend.stdout)
        stderr_parts.append(frontend.stderr)
        _set_frontend_async_progress("刷新 HTML / Excel 报告", 2, 2, " ".join(refresh_command))
        refresh = _run_command(refresh_command, timeout=180)
        stdout_parts.append(refresh.stdout)
        stderr_parts.append(refresh.stderr)
        summary = _frontend_queue_status_summary(priority="P0")
        summary.update(_frontend_refresh_result_summary())
        status_message = _frontend_async_status_message(summary)
        returncode = refresh.returncode if refresh.returncode != 0 else frontend.returncode
        failure_mode = (
            "chrome_cdp_frontend_check_passed_with_pending"
            if summary.get("frontend_queue_passed") and int(summary.get("frontend_pending_count") or 0)
            else ("chrome_cdp_frontend_check_passed" if summary.get("frontend_queue_passed") else "chrome_cdp_frontend_check_partial")
        )
        _sync_frontend_async_status(
            {
                "running": False,
                "message": status_message or "P0 前台后台检查完成，报告已刷新。",
                "returncode": returncode,
                "stdout_tail": "\n".join(stdout_parts)[-4000:],
                "stderr_tail": "\n".join(stderr_parts)[-4000:],
                "status_scope": "frontend_async",
                "failure_mode": failure_mode,
                "soft_failure": frontend.returncode != 0 and refresh.returncode == 0,
                **summary,
            }
        )
    except subprocess.TimeoutExpired as exc:
        payload = _timeout_payload(
            "P0 前台后台检查超时，日报已完成，前台证据保留缓存。",
            exc,
            status_scope="frontend_async",
            failure_mode="chrome_cdp_frontend_check_timeout",
        )
        payload["running"] = False
        _sync_frontend_async_status(payload)
    finally:
        _frontend_async_lock.release()


def _start_p0_frontend_async_if_needed() -> dict[str, object]:
    if not _frontend_async_lock.acquire(blocking=False):
        return {
            **_frontend_async_status,
            "running": True,
            "message": "P0 前台后台检查已在运行，日报已完成。",
            "status_scope": "frontend_async",
        }
    payload = {
        "running": True,
        "message": "P0 前台后台检查已启动，日报已完成；该检查不会阻塞报告使用。",
        "step": 0,
        "total_steps": 2,
        "started_at_epoch": time.time(),
        "elapsed_seconds": 0,
        "status_scope": "frontend_async",
    }
    _sync_frontend_async_status(payload)
    thread = threading.Thread(target=_run_p0_frontend_async, daemon=True)
    thread.start()
    return payload


def _run_frontend_retry() -> None:
    global _last_result
    python = sys.executable
    stdout_parts: list[str] = []
    stderr_parts: list[str] = []
    browser_enabled = _browser_frontend_enabled()
    try:
        needed_summary = {
            **_frontend_refresh_needed_summary(),
            **_sellersprite_reverse_needed_summary(),
        }
        total = int(needed_summary.get("frontend_queue_total") or 0)
        needed = int(needed_summary.get("frontend_refresh_needed_count") or 0)
        sellersprite_needed = int(needed_summary.get("sellersprite_missing_count") or 0)
        frontend_step = (
            f"当前前台队列{_frontend_method_label()}手动刷新",
            _frontend_command(search_policy="always", limit=_frontend_refresh_batch_limit(total), only_stale=False),
            600,
        )
        sellersprite_step = (
            "卖家精灵本次反查",
            _sellersprite_reverse_command(),
            1200,
        ) if sellersprite_needed else None
        refresh_step = ("刷新 HTML / Excel 报告", [python, "main.py", "--marketplace", "ALL"], 180)
        steps = [frontend_step]
        if sellersprite_step is not None:
            steps.append(sellersprite_step)
        steps.append(refresh_step)
        frontend_failure: subprocess.CompletedProcess[str] | None = None
        if browser_enabled:
            _set_progress("检查本机 Chrome CDP 会话", 1, len(steps), CHROME_CDP_ENDPOINT)
        if browser_enabled and not _start_chrome_cdp_if_needed(CHROME_CDP_ENDPOINT):
            frontend_failure = subprocess.CompletedProcess(
                [python, "scripts/run_frontend_checks.py", "--method", "chrome-cdp"],
                1,
                stdout="\n".join(stdout_parts),
                stderr="\n".join(stderr_parts + [f"Chrome CDP endpoint is not available: {CHROME_CDP_ENDPOINT}"]),
            )
            stdout_parts.append(frontend_failure.stdout)
            stderr_parts.append(frontend_failure.stderr)
            steps = [step for step in steps if step[0] != frontend_step[0]]
        sellersprite_failure: subprocess.CompletedProcess[str] | None = None
        for index, (label, command, timeout) in enumerate(steps, start=1):
            _set_progress(label, index, len(steps), " ".join(command))
            if label == frontend_step[0]:
                completed = _run_frontend_command_with_progress(command, timeout, step=index, total_steps=len(steps))
            elif sellersprite_step is not None and label == sellersprite_step[0]:
                seller_started_at = time.time()
                seller_status_owned = False
                if not _sellersprite_async_lock.acquire(blocking=False):
                    _sync_sellersprite_async_status(
                        {
                            **_sellersprite_async_status,
                            "running": True,
                            "message": "卖家精灵已有任务在运行，本次调查等待现有任务结果。",
                            "status_scope": "sellersprite_async",
                        }
                    )
                    completed = subprocess.CompletedProcess(command, 0, stdout="", stderr="")
                else:
                    seller_status_owned = True
                    try:
                        _sync_sellersprite_async_status(
                            {
                                "running": True,
                                "message": f"卖家精灵本次反查已启动：需抓 {sellersprite_needed} 个 ASIN",
                                "detail": " ".join(command),
                                "started_at_epoch": seller_started_at,
                                "elapsed_seconds": 0,
                                "status_scope": "sellersprite_async",
                                **needed_summary,
                            }
                        )
                        completed = _run_sellersprite_command_with_progress(command, timeout=timeout, label="本次调查")
                    finally:
                        _sellersprite_async_lock.release()
                if seller_status_owned:
                    try:
                        seller_summary = _sellersprite_reverse_needed_summary()
                    except Exception:
                        seller_summary = {}
                    _sync_sellersprite_async_status(
                        {
                            "running": False,
                            "message": (
                                "卖家精灵本次反查完成。"
                                if completed.returncode == 0
                                else "卖家精灵本次反查失败，已保留缓存。"
                            ),
                            "returncode": completed.returncode,
                            "elapsed_seconds": max(0, int(time.time() - seller_started_at)),
                            "stdout_tail": completed.stdout[-2000:],
                            "stderr_tail": completed.stderr[-2000:],
                            "status_scope": "sellersprite_async",
                            **seller_summary,
                        }
                    )
            else:
                completed = _run_command(command, timeout=timeout)
            stdout_parts.append(completed.stdout)
            stderr_parts.append(completed.stderr)
            if completed.returncode != 0:
                if label == frontend_step[0]:
                    frontend_failure = completed
                    continue
                if sellersprite_step is not None and label == sellersprite_step[0]:
                    sellersprite_failure = completed
                    continue
                combined = subprocess.CompletedProcess(
                    completed.args,
                    completed.returncode,
                    stdout="\n".join(stdout_parts),
                    stderr="\n".join(stderr_parts),
                )
                _sync_last_result(
                    _completed_payload(
                        combined,
                        "当前前台队列刷新完成，报告已刷新。",
                        f"当前前台队列刷新失败，已保留现有前台缓存。失败步骤：{label}",
                        status_scope="frontend_retry",
                        failure_mode="chrome_cdp_frontend_check_failed",
                    )
                )
                return
        if frontend_failure is not None:
            combined = subprocess.CompletedProcess(
                frontend_failure.args,
                frontend_failure.returncode,
                stdout="\n".join(stdout_parts),
                stderr="\n".join(stderr_parts),
            )
            frontend_summary = _frontend_queue_status_summary()
            frontend_summary.update(_frontend_refresh_result_summary())
            status_message = _frontend_retry_status_message(frontend_summary)
            frontend_unavailable = "Chrome CDP endpoint is not available" in str(frontend_failure.stderr or "")
            _sync_last_result(
                _completed_payload(
                    combined,
                    "当前前台队列刷新完成，报告已刷新。",
                    status_message
                    or (
                        "当前前台队列刷新未启动：本机 Chrome CDP 端口不可用。已刷新报告，卖家精灵按本次调查继续处理。"
                        if frontend_unavailable
                        else "当前前台队列刷新未全部读取成功，已刷新报告并保留失败产品为待前台检查。"
                    ),
                    status_scope="frontend_retry",
                    failure_mode=(
                        "chrome_cdp_unavailable"
                        if frontend_unavailable
                        else (
                            "chrome_cdp_frontend_check_passed_with_pending"
                            if frontend_summary.get("frontend_queue_passed") and int(frontend_summary.get("frontend_pending_count") or 0)
                            else ("chrome_cdp_frontend_check_partial" if status_message else "chrome_cdp_frontend_check_failed")
                        )
                    ),
                    soft_failure=bool(status_message) or frontend_unavailable,
                    **frontend_summary,
                )
            )
            return
        combined = subprocess.CompletedProcess(
            steps[-1][1],
            0,
            stdout="\n".join(stdout_parts),
            stderr="\n".join(stderr_parts),
        )
        frontend_summary = _frontend_queue_status_summary()
        frontend_summary.update(_frontend_refresh_result_summary())
        status_message = _frontend_retry_status_message(frontend_summary)
        if sellersprite_failure is not None:
            status_message = (status_message + " " if status_message else "") + "卖家精灵反查失败，已保留缓存。"
        _sync_last_result(
            _completed_payload(
                combined,
                status_message or "当前前台队列刷新完成，报告已刷新。",
                "当前前台队列刷新失败，已保留现有前台缓存。",
                status_scope="frontend_retry",
                failure_mode=(
                    "chrome_cdp_frontend_check_passed_with_pending"
                    if frontend_summary.get("frontend_queue_passed") and int(frontend_summary.get("frontend_pending_count") or 0)
                    else ("chrome_cdp_frontend_check_passed" if frontend_summary.get("frontend_queue_passed") else "chrome_cdp_frontend_check_partial")
                ),
                soft_failure=bool(sellersprite_failure),
                **frontend_summary,
            )
        )
    except subprocess.TimeoutExpired as exc:
        stdout = exc.stdout or ""
        stderr = exc.stderr or ""
        if isinstance(stdout, bytes):
            stdout = stdout.decode("utf-8", errors="replace")
        if isinstance(stderr, bytes):
            stderr = stderr.decode("utf-8", errors="replace")
        refresh = _run_command([python, "main.py", "--marketplace", "ALL"], timeout=180)
        frontend_summary = _frontend_queue_status_summary()
        frontend_summary.update(_frontend_refresh_result_summary())
        status_message = _frontend_retry_status_message(frontend_summary)
        combined = subprocess.CompletedProcess(
            getattr(exc, "cmd", [python, "scripts/run_frontend_checks.py"]),
            0 if refresh.returncode == 0 else refresh.returncode,
            stdout=str(stdout) + "\n" + refresh.stdout,
            stderr=str(stderr) + "\n" + refresh.stderr,
        )
        _sync_last_result(
            _completed_payload(
                combined,
                status_message or "本轮前台刷新超时，已停止本轮；报告已刷新，下一次继续刷剩余队列。",
                "本轮前台刷新超时，报告刷新失败；旧前台缓存仍保留。",
                status_scope="frontend_retry",
                failure_mode="chrome_cdp_frontend_check_timeout_partial",
                soft_failure=True,
                original_returncode=124,
                **frontend_summary,
            )
        )
    finally:
        _lock.release()


def _run_frontend_check_one(params: dict[str, str]) -> None:
    global _last_result
    python = sys.executable
    command = [
        python,
        "scripts/run_frontend_checks.py",
        "--method",
        "urllib",
        "--search-policy",
        "never",
        "--timeout",
        "30",
        "--retries",
        "3",
        "--marketplace",
        params["marketplace"],
        "--sku",
        params["sku"],
        "--asin",
        params["asin"],
    ]
    try:
        try:
            _start_sellersprite_reverse_async(params=params)
        except Exception as exc:
            _sync_sellersprite_async_status(
                {
                    "running": False,
                    "message": f"卖家精灵后台反查启动失败：{exc}",
                    "returncode": 1,
                    "status_scope": "sellersprite_async",
                }
            )
        _set_progress("单产品前台检查", 1, 2, " ".join(command))
        check = _run_command(command, timeout=150)
        if check.returncode == 0:
            refresh_command = [python, "main.py", "--marketplace", "ALL"]
            _set_progress("刷新 HTML / Excel 报告", 2, 2, " ".join(refresh_command))
            refresh = _run_command(
                refresh_command,
                timeout=180,
            )
            stdout = check.stdout + "\n" + refresh.stdout
            stderr = check.stderr + "\n" + refresh.stderr
            combined = subprocess.CompletedProcess(refresh.args, refresh.returncode, stdout=stdout, stderr=stderr)
            _sync_last_result(_completed_payload(combined, "快速前台检查完成，报告已刷新。", "报告刷新失败，前台结果已写入。"))
        else:
            _sync_last_result(_completed_payload(check, "单产品前台检查完成", "单产品前台检查失败，已保留旧数据。"))
    except subprocess.TimeoutExpired as exc:
        _sync_last_result(_timeout_payload("单产品前台检查超时，已保留旧数据。", exc))
    finally:
        _lock.release()


def _run_battle_diagnosis_one(params: dict[str, str]) -> None:
    global _last_result
    python = sys.executable
    command = [
        python,
        "scripts/run_frontend_checks.py",
        "--method",
        "urllib",
        "--search-policy",
        "always",
        "--timeout",
        "30",
        "--retries",
        "3",
        "--marketplace",
        params["marketplace"],
        "--sku",
        params["sku"],
        "--asin",
        params["asin"],
    ]
    try:
        try:
            _start_sellersprite_reverse_async(params=params)
        except Exception as exc:
            _sync_sellersprite_async_status(
                {
                    "running": False,
                    "message": f"卖家精灵后台反查启动失败：{exc}",
                    "returncode": 1,
                    "status_scope": "sellersprite_async",
                }
            )
        _set_progress("单产品作战前台检查", 1, 2, " ".join(command))
        check = _run_command(command, timeout=120)
        refresh_command = [python, "main.py", "--marketplace", "ALL"]
        _set_progress("刷新 HTML / Excel 报告", 2, 2, " ".join(refresh_command))
        refresh = _run_command(
            refresh_command,
            timeout=220,
        )
        stdout = check.stdout + "\n" + refresh.stdout
        stderr = check.stderr + "\n" + refresh.stderr
        returncode = refresh.returncode if check.returncode == 0 else check.returncode
        combined = subprocess.CompletedProcess(refresh.args, returncode, stdout=stdout, stderr=stderr)
        _sync_last_result(_completed_payload(combined, "产品作战诊断完成，报告已刷新。", "产品作战诊断失败，已保留旧数据。"))
    except subprocess.TimeoutExpired as exc:
        _sync_last_result(_timeout_payload("产品作战诊断超时，已保留旧数据。", exc))
    finally:
        _lock.release()


class Handler(BaseHTTPRequestHandler):
    server_version = "AmazonOpsReportActionServer/1.0"

    def log_message(self, format: str, *args: object) -> None:
        return

    def _send_json(self, status: int, payload: dict[str, object]) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.send_header("Pragma", "no-cache")
        allowed_origin = _allowed_cors_origin(self.headers.get("Origin"))
        if allowed_origin:
            self.send_header("Access-Control-Allow-Origin", allowed_origin)
            self.send_header("Vary", "Origin")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", f"Content-Type, {ACTION_TOKEN_HEADER}")
        self.end_headers()
        self.wfile.write(body)

    def _send_empty_report_page(self) -> None:
        text = """<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Amazon Ops Public Console</title>
  <style>
    body {
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      background: #f6f8fb;
      color: #142033;
    }
    main {
      max-width: 760px;
      margin: 72px auto;
      padding: 0 20px;
    }
    .panel {
      background: #fff;
      border: 1px solid #d9e2ef;
      border-radius: 8px;
      padding: 28px;
      box-shadow: 0 10px 30px rgba(20, 32, 51, 0.08);
    }
    h1 {
      margin: 0 0 12px;
      font-size: 26px;
      line-height: 1.25;
    }
    p {
      margin: 10px 0;
      line-height: 1.7;
      color: #41516a;
    }
    code {
      background: #eef3f9;
      border-radius: 5px;
      padding: 2px 6px;
      color: #142033;
    }
  </style>
</head>
<body>
  <main>
    <section class="panel">
      <h1>公共版当前没有报告数据</h1>
      <p>测试商品和 demo 输出已经清除，服务可访问，但尚未生成正式报告。</p>
      <p>正式使用时，把广告、ERP、成本和 SKU 映射文件放入 <code>data/inbox</code>，然后运行 <code>.venv_mac/bin/python scripts/run_daily_update.py</code>。</p>
      <p>报告生成后，本页面会被新的 <code>latest_recommendations.html</code> 覆盖。</p>
    </section>
  </main>
</body>
</html>"""
        body = text.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.send_header("Pragma", "no-cache")
        self.end_headers()
        self.wfile.write(body)

    def _send_file(self, path: Path) -> None:
        try:
            resolved = path.resolve()
            output_root = OUTPUT_DIR.resolve()
            if output_root not in resolved.parents and resolved != output_root:
                self._send_json(403, {"ok": False, "message": "forbidden"})
                return
            if not resolved.exists() or not resolved.is_file():
                if resolved == (output_root / "latest_recommendations.html").resolve():
                    self._send_empty_report_page()
                    return
                self._send_json(404, {"ok": False, "message": "report file not found"})
                return
            content_type = mimetypes.guess_type(str(resolved))[0] or "application/octet-stream"
            if resolved.suffix == ".html":
                content_type = "text/html; charset=utf-8"
                text = resolved.read_text(encoding="utf-8")
                if 'name="report-action-token"' not in text:
                    token_meta = (
                        '<meta name="report-action-token" content="'
                        + html.escape(_load_or_create_action_token(), quote=True)
                        + '">\n'
                    )
                    if "</head>" in text:
                        text = text.replace("</head>", token_meta + "</head>", 1)
                    else:
                        text = token_meta + text
                body = text.encode("utf-8")
            elif resolved.suffix in {".css", ".js"}:
                content_type = f"text/{resolved.suffix[1:]}; charset=utf-8"
                body = resolved.read_bytes()
            else:
                body = resolved.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(len(body)))
            self.send_header("Cache-Control", "no-store")
            self.send_header("Pragma", "no-cache")
            self.end_headers()
            self.wfile.write(body)
        except OSError as exc:
            self._send_json(500, {"ok": False, "message": str(exc)})

    def do_OPTIONS(self) -> None:
        self._send_json(200, {"ok": True})

    def _request_action_token(self, parsed) -> str:
        header_token = str(self.headers.get(ACTION_TOKEN_HEADER) or "").strip()
        if header_token:
            return header_token
        values = parse_qs(parsed.query).get("token") or []
        return str(values[0] if values else "").strip()

    def _require_action_token(self, parsed) -> bool:
        if _is_valid_action_token(self._request_action_token(parsed)):
            return True
        self._send_json(
            403,
            {
                "ok": False,
                "message": "本地确认服务 token 缺失或无效，请通过本地报告服务页面操作。",
            },
        )
        return False

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        if path == "/health":
            self._send_json(200, {"ok": True, "service": "report_action_server", **_status_payload()})
            return
        if path == "/submission/status":
            self._send_json(200, {"ok": True, "service": "report_action_server", **_status_payload()})
            return
        if path == "/config/status":
            self._send_json(200, {"ok": True, "service": "report_action_server", **_config_status_payload()})
            return
        if path == "/" or path == "/report":
            self._send_file(OUTPUT_DIR / "latest_recommendations.html")
            return
        if path.startswith("/report/"):
            relative = path.removeprefix("/report/").lstrip("/")
            self._send_file(OUTPUT_DIR / relative)
            return
        if path in SIDE_EFFECT_GET_PATHS:
            self._send_json(405, {"ok": False, "message": "该接口只接受 POST，GET 不会触发本地操作。"})
            return
        self._send_json(404, {"ok": False, "message": "unknown endpoint"})

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        if path in SIDE_EFFECT_POST_PATHS and not self._require_action_token(parsed):
            return
        if path == "/upload/today-data":
            self._handle_today_upload()
            return
        if path == "/upload/config":
            params = {key: values[0] if values else "" for key, values in parse_qs(parsed.query).items()}
            self._handle_config_upload(params)
            return
        if path == "/apply/config":
            params = {key: values[0] if values else "" for key, values in parse_qs(parsed.query).items()}
            self._handle_config_apply(params)
            return
        if path == "/copy/text":
            self._handle_copy_text()
            return
        if path == "/run/report-refresh":
            self._start_report_refresh()
            return
        if path == "/run/daily-update":
            self._start_daily_update()
            return
        if path == "/run/frontend-retry":
            self._start_frontend_retry()
            return
        if path == "/run/frontend-check-one":
            params = {key: values[0] if values else "" for key, values in parse_qs(parsed.query).items()}
            self._start_frontend_check_one(params)
            return
        if path == "/run/battle-diagnosis-one":
            params = {key: values[0] if values else "" for key, values in parse_qs(parsed.query).items()}
            self._start_battle_diagnosis_one(params)
            return
        if path == "/feedback/ad-action-complete":
            self._handle_ad_action_complete()
            return
        if path == "/feedback/ad-action-cancel":
            self._handle_ad_action_cancel()
            return
        self._send_json(404, {"ok": False, "message": "unknown endpoint"})

    def _read_json_body(self) -> dict[str, object]:
        length = int(self.headers.get("Content-Length") or "0")
        if length <= 0:
            raise ValueError("请求体不能为空。")
        body = self.rfile.read(length)
        try:
            payload = json.loads(body.decode("utf-8"))
        except Exception as exc:
            raise ValueError(f"请求体不是有效 JSON：{exc}") from exc
        if not isinstance(payload, dict):
            raise ValueError("请求体必须是 JSON 对象。")
        return payload

    def _handle_ad_action_complete(self) -> None:
        try:
            payload = self._read_json_body()
            feedbacks, appended_flags, row_count = append_ad_completion_feedback_batch(payload)
            appended_count = sum(1 for appended in appended_flags if appended)
            self._send_json(
                200,
                {
                    "ok": True,
                    "message": "已写入完成记录。重新生成报告后会进入冷却、自我优化和复盘。"
                    if appended_count
                    else "该动作今天已经记录为已完成。",
                    "appended": appended_count > 0,
                    "appended_count": appended_count,
                    "row_count": row_count,
                    "feedback": feedbacks[0] if len(feedbacks) == 1 else None,
                    "feedbacks": feedbacks,
                    "report_links": REPORT_LINKS,
                },
            )
        except ValueError as exc:
            self._send_json(400, {"ok": False, "message": str(exc)})
        except Exception as exc:
            self._send_json(500, {"ok": False, "message": f"写入完成记录失败：{exc}"})

    def _handle_ad_action_cancel(self) -> None:
        try:
            payload = self._read_json_body()
            removed_rows, removed_count, row_count = cancel_ad_completion_feedback_batch(payload)
            self._send_json(
                200,
                {
                    "ok": True,
                    "message": "已取消完成记录。重新生成报告后会恢复待确认。"
                    if removed_count
                    else "没有找到可取消的完成记录。",
                    "removed_count": removed_count,
                    "row_count": row_count,
                    "removed_feedbacks": removed_rows,
                    "report_links": REPORT_LINKS,
                },
            )
        except ValueError as exc:
            self._send_json(400, {"ok": False, "message": str(exc)})
        except Exception as exc:
            self._send_json(500, {"ok": False, "message": f"取消完成记录失败：{exc}"})

    def _handle_copy_text(self) -> None:
        try:
            payload = self._read_json_body()
            text = str(payload.get("text") or "")
            if not text.strip():
                self._send_json(400, {"ok": False, "message": "复制内容为空。"})
                return
            if len(text) > 200_000:
                self._send_json(400, {"ok": False, "message": "复制内容过长。"})
                return
            try:
                completed = subprocess.run(
                    ["pbcopy"],
                    cwd=ROOT,
                    input=text,
                    text=True,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    timeout=3,
                    check=False,
                )
            except OSError as exc:
                self._send_json(500, {"ok": False, "message": f"本机复制命令不可用：{exc}"})
                return
            if completed.returncode != 0:
                self._send_json(
                    500,
                    {
                        "ok": False,
                        "message": "系统剪贴板写入失败。",
                        "stderr_tail": completed.stderr[-500:],
                    },
                )
                return
            self._send_json(200, {"ok": True, "message": "已复制到系统剪贴板。", "chars": len(text)})
        except ValueError as exc:
            self._send_json(400, {"ok": False, "message": str(exc)})
        except subprocess.TimeoutExpired:
            self._send_json(500, {"ok": False, "message": "系统剪贴板写入超时。"})
        except Exception as exc:
            self._send_json(500, {"ok": False, "message": f"复制失败：{exc}"})

    def _handle_today_upload(self) -> None:
        if not _lock.acquire(blocking=False):
            self._send_json(409, {"ok": False, "running": True, "message": "已有上传或 daily update 正在运行，请稍等。"})
            return
        global _last_result
        try:
            content_type = self.headers.get("Content-Type", "")
            if "multipart/form-data" not in content_type:
                self._send_json(400, {"ok": False, "message": "上传请求必须使用 multipart/form-data。"})
                return
            _sync_last_result(
                {
                    **_load_submission_status(),
                    "running": True,
                    "message": "正在接收今日数据文件",
                    "step": 1,
                    "total_steps": 2,
                    "started_at_epoch": time.time(),
                    "elapsed_seconds": 0,
                }
            )
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": content_type,
                    "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
                },
            )
            saved, errors = _save_uploaded_files(form)
            if errors:
                for row in saved:
                    path = ROOT / str(row.get("path") or "")
                    if path.exists() and path.is_file() and INBOX_DIR.resolve() in path.resolve().parents:
                        path.unlink()
                payload = _fresh_status(
                    "上传失败：" + "；".join(errors),
                    returncode=1,
                    uploaded_files=saved,
                    upload_errors=errors,
                )
                _sync_last_result(payload)
                self._send_json(400, {"ok": False, **payload})
                return
            if not saved:
                payload = _fresh_status("没有收到可上传的 .csv 或 .xlsx 文件。", returncode=1)
                _sync_last_result(payload)
                self._send_json(400, {"ok": False, **payload})
                return

            _set_progress("运行 daily preflight", 2, 2, "scripts/check_daily_update_preflight.py --allow-inbox-business-files")
            preflight = _run_submission_preflight()
            payload = _fresh_status(
                "上传完成，preflight 通过。"
                if preflight.returncode == 0
                else "上传完成，但 preflight 未通过，不能运行 daily update。",
                returncode=preflight.returncode,
                uploaded_files=saved,
                stdout_tail=preflight.stdout[-4000:],
                stderr_tail=preflight.stderr[-4000:],
            )
            _sync_last_result(payload)
            self._send_json(200 if preflight.returncode == 0 else 409, {"ok": preflight.returncode == 0, **payload})
        except Exception as exc:
            payload = _fresh_status(f"上传处理失败：{exc}", returncode=1)
            _sync_last_result(payload)
            self._send_json(500, {"ok": False, **payload})
        finally:
            _lock.release()

    def _handle_config_upload(self, params: dict[str, str]) -> None:
        kind = str(params.get("kind") or "").strip().lower()
        try:
            config = _target_config(kind)
        except ValueError as exc:
            self._send_json(400, {"ok": False, "message": str(exc)})
            return
        if not _lock.acquire(blocking=False):
            self._send_json(409, {"ok": False, "running": True, "message": "已有上传或检查正在运行，请稍等。"})
            return
        try:
            content_type = self.headers.get("Content-Type", "")
            if "multipart/form-data" not in content_type:
                self._send_json(400, {"ok": False, "message": "上传请求必须使用 multipart/form-data。"})
                return
            _sync_last_result(
                _fresh_status(
                    f"正在接收{config.get('label')}文件",
                    running=True,
                    step=1,
                    total_steps=2,
                    started_at_epoch=time.time(),
                    elapsed_seconds=0,
                )
            )
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    "REQUEST_METHOD": "POST",
                    "CONTENT_TYPE": content_type,
                    "CONTENT_LENGTH": self.headers.get("Content-Length", "0"),
                },
            )
            saved, errors = _save_config_upload(form, kind)
            if errors:
                payload = _fresh_status(
                    "配置上传失败：" + "；".join(errors),
                    returncode=1,
                    config_upload=saved,
                    upload_errors=errors,
                    config_status=_config_status_payload(),
                    audit_links=_config_audit_links(kind),
                )
                _sync_last_result(payload)
                self._send_json(400, {"ok": False, **payload})
                return

            review_command = config.get("review_command")
            completed: subprocess.CompletedProcess[str] | None = None
            if review_command:
                _set_progress("生成成本配置差异审计", 2, 2, " ".join(str(item) for item in review_command))
                completed = _run_command([str(item) for item in review_command], timeout=90)
            payload = _fresh_status(
                (
                    "成本配置已进入待审核区，差异审计已生成；确认后再应用正式配置。"
                    if kind == "cost" and (not completed or completed.returncode == 0)
                    else "配置已进入待审核区，可以应用。"
                    if not completed
                    else "成本配置上传成功，但差异审计失败，请先查看输出。"
                ),
                returncode=completed.returncode if completed else 0,
                config_upload=saved,
                stdout_tail=completed.stdout[-4000:] if completed else "",
                stderr_tail=completed.stderr[-4000:] if completed else "",
                audit_links=_config_audit_links(kind),
                config_status=_config_status_payload(),
            )
            _sync_last_result(payload)
            self._send_json(200 if payload["returncode"] == 0 else 409, {"ok": payload["returncode"] == 0, **payload})
        except Exception as exc:
            payload = _fresh_status(
                f"配置上传处理失败：{exc}",
                returncode=1,
                config_status=_config_status_payload(),
                audit_links=_config_audit_links(kind),
            )
            _sync_last_result(payload)
            self._send_json(500, {"ok": False, **payload})
        finally:
            _lock.release()

    def _handle_config_apply(self, params: dict[str, str]) -> None:
        kind = str(params.get("kind") or "").strip().lower()
        confirm = str(params.get("confirm") or "").strip()
        try:
            config = _target_config(kind)
        except ValueError as exc:
            self._send_json(400, {"ok": False, "message": str(exc)})
            return
        if config.get("requires_confirm") and confirm != "apply":
            self._send_json(400, {"ok": False, "message": "成本配置应用必须显式确认。"})
            return
        if not _lock.acquire(blocking=False):
            self._send_json(409, {"ok": False, "running": True, "message": "已有上传或检查正在运行，请稍等。"})
            return
        try:
            _sync_last_result(
                _fresh_status(
                    f"正在应用{config.get('label')}",
                    running=True,
                    step=1,
                    total_steps=1,
                    started_at_epoch=time.time(),
                    elapsed_seconds=0,
                )
            )
            applied, completed = _apply_pending_config(kind)
            ok = completed.returncode == 0 if completed else True
            payload = _fresh_status(
                (
                    "成本配置已应用到正式 config；请运行 safe-run 或 daily update 后再看正式报告。"
                    if kind == "cost"
                    else "SKU 别名映射已应用，并已完成 safe-run 验证。"
                    if ok
                    else "SKU 别名映射已应用，但 safe-run 验证失败。"
                ),
                returncode=completed.returncode if completed else 0,
                applied_config=applied,
                config_status=_config_status_payload(),
            )
            _sync_last_result(payload)
            self._send_json(200 if ok else 409, {"ok": ok, **payload})
        except Exception as exc:
            payload = _fresh_status(
                f"配置应用失败：{exc}",
                returncode=1,
                config_status=_config_status_payload(),
                audit_links=_config_audit_links(kind),
            )
            _sync_last_result(payload)
            self._send_json(500, {"ok": False, **payload})
        finally:
            _lock.release()

    def _start_daily_update(self) -> None:
        if not _lock.acquire(blocking=False):
            self._send_json(409, {"ok": False, "running": True, "message": "已有上传或 daily update 正在运行，请稍等。"})
            return
        global _last_result
        _last_result = {
            **_load_submission_status(),
            "running": True,
            "message": "daily update 已启动",
            "step": 0,
            "total_steps": 1,
            "started_at_epoch": time.time(),
            "elapsed_seconds": 0,
            "returncode": None,
        }
        _write_submission_status(_last_result)
        thread = threading.Thread(target=_run_daily_update, daemon=True)
        thread.start()
        self._send_json(202, {"ok": True, "running": True, "message": "已开始导入 inbox 并刷新报告。"})

    def _start_report_refresh(self) -> None:
        if not _lock.acquire(blocking=False):
            self._send_json(409, {"ok": False, "running": True, "message": "已有上传或刷新正在运行，请稍等。"})
            return
        global _last_result
        _last_result = {
            **_load_submission_status(),
            "running": True,
            "message": "轻量报告刷新已启动",
            "step": 0,
            "total_steps": 1,
            "started_at_epoch": time.time(),
            "elapsed_seconds": 0,
            "returncode": None,
            "status_scope": "report_refresh",
        }
        _write_submission_status(_last_result)
        thread = threading.Thread(target=_run_report_refresh, daemon=True)
        thread.start()
        self._send_json(202, {"ok": True, "running": True, "message": "已开始刷新报告；不会导入 inbox 或访问 Amazon 前台。"})

    def _start_frontend_retry(self) -> None:
        if not _lock.acquire(blocking=False):
            self._send_json(409, {"ok": False, "running": True, "message": "前台数据重试正在运行，请稍等。"})
            return
        global _last_result
        _last_result = {
            "running": True,
            "message": "市场调查已启动",
            "step": 0,
            "total_steps": 3,
            "started_at_epoch": time.time(),
            "elapsed_seconds": 0,
            "status_scope": "frontend_retry",
        }
        thread = threading.Thread(target=_run_frontend_retry, daemon=True)
        thread.start()
        self._send_json(202, {"ok": True, "running": True, "message": "已开始市场调查；会一次执行商品页、卖家精灵和报告刷新。"})

    def _start_frontend_check_one(self, params: dict[str, str]) -> None:
        required = {
            "marketplace": str(params.get("marketplace") or "").strip().upper(),
            "sku": str(params.get("sku") or "").strip(),
            "asin": str(params.get("asin") or "").strip().upper(),
        }
        if not required["marketplace"] or not required["asin"]:
            self._send_json(400, {"ok": False, "message": "缺少 marketplace 或 asin。"})
            return
        if not _lock.acquire(blocking=False):
            self._send_json(409, {"ok": False, "running": True, "message": "前台检查正在运行，请稍等。"})
            return
        global _last_result
        _last_result = {
            "running": True,
            "message": "单产品前台检查已启动",
            "step": 0,
            "total_steps": 2,
            "started_at_epoch": time.time(),
            "elapsed_seconds": 0,
            **required,
        }
        thread = threading.Thread(target=_run_frontend_check_one, args=(required,), daemon=True)
        thread.start()
        self._send_json(202, {"ok": True, "running": True, "message": "正在检查这个产品的前台，完成后会刷新报告。"})

    def _start_battle_diagnosis_one(self, params: dict[str, str]) -> None:
        required = {
            "marketplace": str(params.get("marketplace") or "").strip().upper(),
            "sku": str(params.get("sku") or "").strip(),
            "asin": str(params.get("asin") or "").strip().upper(),
        }
        if not required["marketplace"] or not required["asin"]:
            self._send_json(400, {"ok": False, "message": "缺少 marketplace 或 asin。"})
            return
        if not _lock.acquire(blocking=False):
            self._send_json(409, {"ok": False, "running": True, "message": "已有前台/作战诊断正在运行，请稍等。"})
            return
        global _last_result
        _last_result = {
            "running": True,
            "message": "产品作战诊断已启动",
            "step": 0,
            "total_steps": 2,
            "started_at_epoch": time.time(),
            "elapsed_seconds": 0,
            **required,
        }
        thread = threading.Thread(target=_run_battle_diagnosis_one, args=(required,), daemon=True)
        thread.start()
        self._send_json(202, {"ok": True, "running": True, "message": "正在做这个产品的作战诊断，完成后会刷新报告。"})


def main() -> int:
    _clear_stale_running_status_on_startup()
    server = ThreadingHTTPServer((HOST, PORT), Handler)
    print(f"[report-action-server] listening on http://{HOST}:{PORT}", flush=True)
    print("[report-action-server] keep this window open while using page buttons", flush=True)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n[report-action-server] stopped", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

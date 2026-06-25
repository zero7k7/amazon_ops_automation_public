from __future__ import annotations

import argparse
import json
import sys
from dataclasses import asdict
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.import_inbox_files import (
    INBOX,
    OUTPUT,
    UNKNOWN_DIR,
    load_asin_marketplace_map,
    process_file,
)


JSON_OUTPUT = OUTPUT / "unknown_inbox_audit.json"
MARKDOWN_OUTPUT = OUTPUT / "unknown_inbox_audit.md"
BUSINESS_SUFFIXES = {".csv", ".xlsx"}
PROMOTION_METRIC_HEADERS = {
    "促销类型",
    "促销编号",
    "促销名称",
    "促销开始时间",
    "促销结束时间",
    "促销批准状态",
}


def _relative(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def _unknown_business_files() -> list[Path]:
    if not UNKNOWN_DIR.exists():
        return []
    return sorted(
        path
        for path in UNKNOWN_DIR.rglob("*")
        if path.is_file() and path.suffix.lower() in BUSINESS_SUFFIXES
    )


def _clean_cell(value: object) -> str:
    return str(value or "").strip()


def _inspect_xlsx(path: Path) -> dict[str, object]:
    info: dict[str, object] = {
        "sheet_names": [],
        "first_sheet": "",
        "header_sample": [],
        "row_sample": [],
        "likely_report_type": "unknown",
        "likely_marketplace": "UNKNOWN",
        "business_interpretation": "",
    }
    if path.suffix.lower() != ".xlsx":
        return info
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        info["sheet_names"] = workbook.sheetnames
        if not workbook.sheetnames:
            return info
        worksheet = workbook[workbook.sheetnames[0]]
        info["first_sheet"] = worksheet.title
        iterator = worksheet.iter_rows(values_only=True)
        header = [_clean_cell(value) for value in next(iterator, [])]
        row_sample = [_clean_cell(value) for value in next(iterator, [])]
        info["header_sample"] = [value for value in header if value][:20]
        info["row_sample"] = [value for value in row_sample if value][:20]
        header_set = {value for value in header if value}
        if PROMOTION_METRIC_HEADERS.issubset(header_set):
            info["likely_report_type"] = "seller_central_promotion_metrics"
            sales_headers = [value for value in header if "促销销售额" in value]
            if any("(€)" in value for value in sales_headers):
                info["likely_marketplace"] = "DE_or_EUR"
            info["business_interpretation"] = (
                "Seller Central 促销表现 metric-data，包含促销浏览、订购促销商品数量和促销销售额；"
                "当前导入流程未使用该格式。"
            )
    except Exception as exc:  # noqa: BLE001
        info["preview_error"] = f"{type(exc).__name__}: {exc}"
    finally:
        if workbook:
            workbook.close()
    return info


def _recommendation(row: dict[str, object]) -> str:
    likely_report_type = str(row.get("likely_report_type") or "")
    if likely_report_type == "seller_central_promotion_metrics":
        return "这是促销表现报告，当前不参与 daily import。若只做参考，移出 data/inbox/_unknown/；若要纳入诊断，先补新格式解析、字段口径和测试。"
    detected_type = str(row.get("detected_type") or "")
    detected_marketplace = str(row.get("detected_marketplace") or "")
    status = str(row.get("status") or "")
    reason = str(row.get("reason") or "")
    if status.startswith("error"):
        return "先确认文件是否损坏、加密或被占用，再重新下载。"
    if detected_type == "unknown":
        return "人工判断文件来源。若是无关文件，移出 data/inbox/_unknown/；若是新报表格式，先补识别规则和测试。"
    if detected_marketplace in {"UNKNOWN", "MULTI"}:
        return "人工确认站点。若是增强数据，多站点文件需拆分后再导入。"
    if reason:
        return "按原因处理后再运行 daily preflight。"
    return "确认目标路径和数据日期后，再决定是否重新放回 data/inbox/ 导入。"


def build_audit_rows(files: list[Path]) -> list[dict[str, object]]:
    asin_map = load_asin_marketplace_map()
    rows: list[dict[str, object]] = []
    for path in files:
        row, _ = process_file(
            path,
            dry_run=True,
            asin_map=asin_map,
            selected_ads=None,
            selected_erp=None,
        )
        payload = asdict(row)
        payload.update(_inspect_xlsx(path))
        payload["blocks_daily_update"] = True
        payload["recommendation"] = _recommendation(payload)
        rows.append(payload)
    return rows


def build_markdown(rows: list[dict[str, object]]) -> str:
    lines = [
        "# unknown inbox 审计",
        "",
        "该文件用于解释 `data/inbox/_unknown/` 中阻断 daily update 的业务文件。审计只读，不移动 inbox 文件。",
        "",
        f"- 阻断 daily update 的业务文件数：{len(rows)}",
        "",
    ]
    if not rows:
        lines.extend(["## 结果", "", "未发现阻断 daily update 的 unknown 业务文件。"])
        return "\n".join(lines) + "\n"

    lines.extend(
        [
            "## 文件明细",
            "",
            "| 文件 | 类型 | 站点 | 状态 | 原因 | 建议 |",
            "|---|---|---|---|---|---|",
        ]
    )
    for row in rows:
        lines.append(
            "| "
            + " | ".join(
                [
                    _relative(Path(str(row.get("original_path") or ""))),
                    str(row.get("detected_type") or ""),
                    str(row.get("likely_marketplace") or row.get("detected_marketplace") or ""),
                    str(row.get("status") or ""),
                    str(row.get("business_interpretation") or row.get("reason") or row.get("preview_error") or ""),
                    str(row.get("recommendation") or ""),
                ]
            )
            + " |"
        )
    lines.extend(
        [
            "",
            "## 处理规则",
            "",
            "1. 无关文件移出 `data/inbox/_unknown/` 后再跑 preflight",
            "2. 新报表格式先补识别规则和测试，再允许导入",
            "3. 多站点增强数据先拆分站点，再放回 `data/inbox/`",
            "4. 不要直接运行 `scripts/run_daily_update.py` 绕过 preflight",
        ]
    )
    return "\n".join(lines) + "\n"


def write_outputs(rows: list[dict[str, object]]) -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    JSON_OUTPUT.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    MARKDOWN_OUTPUT.write_text(build_markdown(rows), encoding="utf-8")


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Audit business files currently parked in data/inbox/_unknown.")
    parser.add_argument("--fail-on-blocker", action="store_true", help="Return exit code 1 when unknown business files exist.")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    rows = build_audit_rows(_unknown_business_files())
    write_outputs(rows)
    print(f"[check] unknown business files: {len(rows)}", flush=True)
    print(f"[check] wrote: {JSON_OUTPUT}", flush=True)
    print(f"[check] wrote: {MARKDOWN_OUTPUT}", flush=True)
    if rows and args.fail_on_blocker:
        print("[fail] unknown inbox business files block daily update", flush=True)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import time
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
INBOX = DATA / "inbox"
UNKNOWN_DIR = INBOX / "_unknown"
RAW_ADS = DATA / "raw_ads"
RAW_ERP = DATA / "raw_erp"
RAW_CUSTOM = DATA / "raw_amazon_custom"
RAW_CUSTOM_UK = RAW_CUSTOM / "UK"
RAW_CUSTOM_US = RAW_CUSTOM / "US"
RAW_CUSTOM_DE = RAW_CUSTOM / "DE"
ARCHIVE_IMPORTED = DATA / "archive" / "inbox_imported"
ARCHIVE_OVERWRITTEN = DATA / "archive" / "overwritten"
OUTPUT = DATA / "output"

IMPORT_MANIFEST_JSON = OUTPUT / "import_manifest.json"
IMPORT_MANIFEST_XLSX = OUTPUT / "import_manifest.xlsx"

SCAN_EXTENSIONS = {".csv", ".xlsx"}
PREVIEW_ROWS = 20
PER_FILE_TIMEOUT_SEC = 30
TOTAL_TIMEOUT_SEC = 120

ADS_FIELDS = {
    "广告活动编号",
    "广告活动名称",
    "推广的商品 SKU",
    "推广的商品编号",
    "推广的商品站点",
    "搜索词",
    "展示量",
    "点击量",
    "总成本",
    "销售额",
}
ERP_FIELDS = {
    "时间",
    "ASIN",
    "PARENTASIN",
    "MSKU",
    "SKU",
    "品名",
    "标题",
    "国家",
    "店铺",
    "FBA可售",
    "可用库存",
    "FBM可售",
    "销量",
    "订单量",
    "销售额",
}
TRAFFIC_FIELDS = {
    "ASIN",
    "商品名称",
    "转化率",
    "推荐报价浏览量",
    "推荐报价率",
    "已订购商品数量",
    "已发货商品数量",
}
QUERY_FIELDS = {
    "搜索查询",
    "ASIN",
    "商品名称",
    "展示次数",
    "点击数量",
    "添加购物车",
    "购买数量",
}

HYPERLINK_RE = re.compile(
    r'^\s*=HYPERLINK\(\s*"(?P<url>[^"]+)"\s*,\s*"(?P<label>[^"]*)"\s*\)\s*$',
    re.IGNORECASE,
)
URL_ASIN_RE = re.compile(r"/dp/([A-Z0-9]{10})", re.IGNORECASE)
ASIN_RE = re.compile(r"\b([A-Z0-9]{10})\b", re.IGNORECASE)

import sys

if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.amazon_custom_enhancements import (  # noqa: E402
    _classify_report_file,
    _detect_header_metadata,
    _read_first_sheet,
)


@dataclass
class ManifestRow:
    original_filename: str
    original_path: str
    detected_type: str
    detected_marketplace: str
    period_type: str
    detected_date_range: str
    target_path: str
    archive_path: str
    status: str
    reason: str
    modified_time: str
    rows: int
    columns: int
    asin_count: int
    created_at: str
    target_conflict: str = ""
    overwritten: int = 0
    enhanced_detected_from: str = ""
    enhanced_freshness: str = ""
    diagnosis_usage: str = ""


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def ts_compact() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H%M%S")


def ensure_dirs() -> None:
    for p in [
        INBOX,
        UNKNOWN_DIR,
        RAW_ADS,
        RAW_ERP,
        RAW_CUSTOM,
        RAW_CUSTOM_UK,
        RAW_CUSTOM_US,
        RAW_CUSTOM_DE,
        ARCHIVE_IMPORTED,
        ARCHIVE_OVERWRITTEN,
        OUTPUT,
    ]:
        p.mkdir(parents=True, exist_ok=True)


def normalize_headers(headers: Iterable[object]) -> list[str]:
    out: list[str] = []
    for h in headers:
        if h is None:
            continue
        s = str(h).strip().replace("\n", " ")
        if s:
            out.append(s)
    return out


def score_headers(headers: list[str], expected: set[str]) -> int:
    score = 0
    for exp in expected:
        if any(exp in h for h in headers):
            score += 1
    return score


def detect_marketplace_from_name(name: str) -> str:
    n = name.lower()
    if re.search(r"(^|[_\-\s])uk([_\-\s]|$)|英国", n):
        return "UK"
    if re.search(r"(^|[_\-\s])us([_\-\s]|$)|美国", n):
        return "US"
    if re.search(r"(^|[_\-\s])de([_\-\s]|$)|德国", n):
        return "DE"
    return "UNKNOWN"


def detect_period_type(headers: list[str]) -> str:
    text = " | ".join(headers).lower()
    if "wow" in text or "周环比" in text:
        return "wow"
    dates = re.findall(r"20\d{2}[-/年.]\d{1,2}[-/月.]\d{1,2}", text)
    if len(dates) >= 4:
        return "compare"
    if len(dates) >= 2:
        return "single_period"
    return "unknown"


def detect_date_range_from_name(name: str) -> str:
    dates = re.findall(r"(20\d{2}-\d{2}-\d{2})", name)
    if len(dates) >= 2:
        return f"{dates[0]}~{dates[1]}"
    return "unknown"


def read_csv_preview(path: Path) -> tuple[list[str], int, int]:
    for enc in ["utf-8-sig", "utf-8", "gbk", "gb18030"]:
        try:
            with path.open("r", encoding=enc, newline="") as f:
                reader = csv.reader(f)
                rows = []
                for i, row in enumerate(reader):
                    rows.append(row)
                    if i >= PREVIEW_ROWS:
                        break
            if not rows:
                return [], 0, 0
            headers = normalize_headers(rows[0])
            return headers, max(0, len(rows) - 1), len(headers)
        except Exception:
            continue
    raise RuntimeError("CSV preview failed")


def read_xlsx_preview(path: Path) -> tuple[list[str], int, int]:
    wb = None
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
        ws = wb[wb.sheetnames[0]]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            rows.append(row)
            if i >= PREVIEW_ROWS:
                break
        if not rows:
            return [], 0, 0
        headers = normalize_headers(rows[0])
        return headers, max(0, len(rows) - 1), len(headers)
    finally:
        if wb:
            wb.close()


def infer_type(path: Path, headers: list[str]) -> str:
    name = path.name.lower()
    if "traffic_sales" in name:
        return "traffic_sales"
    if "search_query_performance" in name:
        return "search_query_performance"

    ads_score = score_headers(headers, ADS_FIELDS)
    erp_score = score_headers(headers, ERP_FIELDS)
    traffic_score = score_headers(headers, TRAFFIC_FIELDS)
    query_score = score_headers(headers, QUERY_FIELDS)

    if traffic_score >= 4:
        return "traffic_sales"
    if query_score >= 4:
        return "search_query_performance"
    if ads_score >= 5:
        return "ads_report_all"
    if erp_score >= 6 and path.suffix.lower() == ".xlsx":
        return "erp_sales_all"
    return "unknown"


def domain_to_marketplace(url: str) -> str:
    u = url.lower()
    if "amazon.co.uk" in u:
        return "UK"
    if "amazon.com" in u:
        return "US"
    if "amazon.de" in u:
        return "DE"
    return "UNKNOWN"


def is_weak_amazon_com_marketplace(url: str) -> bool:
    return "amazon.com" in str(url or "").lower()


def marketplace_content_scores(text: str) -> Counter[str]:
    value = f" {text.lower()} "
    scores: Counter[str] = Counter()
    if re.search(r"[äöüß]", value):
        scores["DE"] += 4
    patterns = {
        "DE": [
            r"\bfür\b",
            r"\bund\b",
            r"\bmit\b",
            r"\bgröße\b",
            r"\bfarbe\b",
            r"\beur\b",
            r"€",
        ],
        "UK": [
            r"\bcolour\b",
            r"\borganiser\b",
            r"\bcentre\b",
            r"\bpostcode\b",
            r"\buk\b",
            r"£",
        ],
        "US": [
            r"\bcolor\b",
            r"\borganizer\b",
            r"\bcenter\b",
            r"\binch\b",
            r"\bus\b",
            r"\$",
        ],
    }
    for marketplace, marketplace_patterns in patterns.items():
        for pattern in marketplace_patterns:
            scores[marketplace] += len(re.findall(pattern, value))
    return scores


def confident_marketplace_from_scores(scores: Counter[str]) -> str:
    if not scores:
        return "UNKNOWN"
    ranked = scores.most_common(2)
    winner, top_score = ranked[0]
    second_score = ranked[1][1] if len(ranked) > 1 else 0
    if top_score >= 3 and top_score >= second_score + 2:
        return winner
    return "UNKNOWN"


def extract_from_hyperlink_formula(formula: str) -> tuple[str, str]:
    """
    Returns (asin, marketplace).
    """
    m = HYPERLINK_RE.match(formula.strip())
    if not m:
        return "", "UNKNOWN"
    url = m.group("url")
    label = m.group("label")
    mkt = domain_to_marketplace(url)
    asin = ""
    uasin = URL_ASIN_RE.search(url)
    if uasin:
        asin = uasin.group(1).upper()
    if not asin:
        lasin = ASIN_RE.search(label or "")
        if lasin:
            asin = lasin.group(1).upper()
    return asin, mkt


def extract_url_from_hyperlink_formula(formula: str) -> str:
    m = HYPERLINK_RE.match(formula.strip())
    return m.group("url") if m else ""


def load_asin_marketplace_map() -> dict[str, str]:
    out: dict[str, str] = {}
    for path in [ROOT / "config" / "product_cost_config.xlsx", ROOT / "config" / "sku_asin_map.xlsx"]:
        if not path.exists():
            continue
        wb = None
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = ws.iter_rows(values_only=True)
            headers = [str(c).strip().lower() if c is not None else "" for c in next(rows, [])]
            if "asin" not in headers or "marketplace" not in headers:
                continue
            i_asin = headers.index("asin")
            i_mkt = headers.index("marketplace")
            for i, row in enumerate(rows, start=1):
                if i > 50000:
                    break
                asin = row[i_asin] if i_asin < len(row) else None
                mkt = row[i_mkt] if i_mkt < len(row) else None
                if asin is None or mkt is None:
                    continue
                a = str(asin).strip().upper()
                m = str(mkt).strip().upper()
                if a and m and a not in out:
                    out[a] = m
        except Exception:
            pass
        finally:
            if wb:
                wb.close()
    return out


def infer_enhanced_marketplaces_and_asin_count(path: Path, asin_map: dict[str, str]) -> tuple[set[str], int]:
    """
    Read first sheet only; parse ASIN column and hyperlink formulas.
    """
    wb = None
    strong_domain_mkts: set[str] = set()
    mapped_mkt_counts: Counter[str] = Counter()
    weak_us_domain_count = 0
    text_parts: list[str] = []
    asin_seen: set[str] = set()
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
        ws = wb[wb.sheetnames[0]]
        row_iter = ws.iter_rows(values_only=False)
        first = next(row_iter, None)
        if first is None:
            return set(), 0
        headers = [str(c.value).strip() if c.value is not None else "" for c in first]
        asin_idx = next((i for i, h in enumerate(headers) if "ASIN" in h.upper()), -1)
        if asin_idx < 0:
            return set(), 0

        for i, row in enumerate(row_iter, start=1):
            if i > 3000:
                break
            if i <= 80:
                text_parts.extend(str(c.value) for c in row if c.value is not None)
            if asin_idx >= len(row):
                continue
            cell = row[asin_idx]
            v = cell.value
            if v is None:
                continue
            txt = str(v).strip()
            if not txt:
                continue
            if txt.lower() == "total":
                continue

            asin = ""
            mkt = "UNKNOWN"
            if txt.startswith("=") and "HYPERLINK" in txt.upper():
                asin, mkt = extract_from_hyperlink_formula(txt)
                url = extract_url_from_hyperlink_formula(txt)
                if is_weak_amazon_com_marketplace(url):
                    weak_us_domain_count += 1
                    mkt = "UNKNOWN"
            else:
                m = ASIN_RE.search(txt)
                if m:
                    asin = m.group(1).upper()

            if not asin:
                # fallback: try from text by dp
                uasin = URL_ASIN_RE.search(txt)
                if uasin:
                    asin = uasin.group(1).upper()

            if asin:
                asin_seen.add(asin)
                if mkt in {"UK", "DE"}:
                    strong_domain_mkts.add(mkt)
                mapped = asin_map.get(asin, "UNKNOWN")
                if mapped in {"UK", "US", "DE"}:
                    mapped_mkt_counts[mapped] += 1
    finally:
        if wb:
            wb.close()
    if strong_domain_mkts:
        return strong_domain_mkts, len(asin_seen)
    content_winner = confident_marketplace_from_scores(marketplace_content_scores(" ".join(text_parts)))
    if content_winner != "UNKNOWN":
        return {content_winner}, len(asin_seen)
    if mapped_mkt_counts:
        winner, count = mapped_mkt_counts.most_common(1)[0]
        total = sum(mapped_mkt_counts.values())
        if count >= 2 and count / max(total, 1) >= 0.6:
            return {winner}, len(asin_seen)
    if weak_us_domain_count and not mapped_mkt_counts:
        return {"US"}, len(asin_seen)
    return set(), len(asin_seen)


def build_enhanced_target(file_type: str, marketplace: str, date_range: str, period_type: str) -> Path:
    folder = {"UK": RAW_CUSTOM_UK, "US": RAW_CUSTOM_US, "DE": RAW_CUSTOM_DE}.get(marketplace, RAW_CUSTOM_UK)
    prefix = "traffic_sales" if file_type == "traffic_sales" else "search_query_performance"
    if date_range == "unknown":
        fname = f"{prefix}_{marketplace.lower()}_unknown.xlsx"
    else:
        start, end = date_range.split("~")
        if period_type == "wow":
            fname = f"{prefix}_{marketplace.lower()}_wow_{start}_{end}.xlsx"
        elif period_type == "compare":
            fname = f"{prefix}_{marketplace.lower()}_compare_{start}_{end}.xlsx"
        elif period_type == "single_period":
            fname = f"{prefix}_{marketplace.lower()}_single_{start}_{end}.xlsx"
        else:
            fname = f"{prefix}_{marketplace.lower()}_{start}_{end}.xlsx"
    return folder / fname


def detect_enhanced_header_info(path: Path, marketplace: str) -> dict[str, object] | None:
    try:
        _, rows = _read_first_sheet(path)
    except Exception:
        return None
    if not rows:
        return None
    filename_marketplace, filename_data_type, filename_period = _classify_report_file(path)
    columns = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    info = _detect_header_metadata(
        path=path,
        columns=columns,
        preview_rows=rows[:8],
        marketplace=marketplace,
        filename_marketplace=filename_marketplace,
        filename_data_type=filename_data_type,
        filename_period_hint=filename_period,
        common_end=None,
    )
    if info.get("data_type") not in {"traffic_sales", "search_query_performance"}:
        return None
    return info


def build_enhanced_target_from_header(info: dict[str, object]) -> Path | None:
    data_type = str(info.get("data_type") or "")
    marketplace = str(info.get("marketplace") or "").upper()
    recent_start = info.get("recent_start")
    recent_end = info.get("recent_end")
    prior_start = info.get("prior_start")
    prior_end = info.get("prior_end")
    if data_type not in {"traffic_sales", "search_query_performance"} or marketplace not in {"UK", "US", "DE"}:
        return None
    if not (recent_start and recent_end and prior_start and prior_end):
        return None
    folder = {"UK": RAW_CUSTOM_UK, "US": RAW_CUSTOM_US, "DE": RAW_CUSTOM_DE}[marketplace]
    return folder / f"{data_type}_{marketplace.lower()}_wow_{recent_start}_{recent_end}_vs_{prior_start}_{prior_end}.xlsx"


def build_target_path(file_type: str, marketplace: str, date_range: str, period_type: str, original: Path) -> Path:
    if file_type == "ads_report_all":
        return RAW_ADS / "ads_report_all.csv"
    if file_type == "erp_sales_all":
        return RAW_ERP / "sales_report_all.xlsx"
    if file_type in {"traffic_sales", "search_query_performance"}:
        return build_enhanced_target(file_type, marketplace, date_range, period_type)
    return UNKNOWN_DIR / original.name


def timestamped_path(base: Path, name: str) -> Path:
    stem = Path(name).stem
    ext = Path(name).suffix
    return base / f"{stem}_{ts_compact()}{ext}"


def safe_filename_part(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    cleaned = cleaned.strip("._-")
    return cleaned[:80] or "upload"


def disambiguate_same_batch_target(target: Path, original: Path, reserved_targets: set[Path]) -> tuple[Path, str]:
    resolved = target.resolve()
    if resolved not in reserved_targets:
        return target, ""
    suffix = safe_filename_part(original.stem)
    candidate = target.with_name(f"{target.stem}__{suffix}{target.suffix}")
    index = 2
    while candidate.resolve() in reserved_targets:
        candidate = target.with_name(f"{target.stem}__{suffix}_{index}{target.suffix}")
        index += 1
    return candidate, f"same_batch_target_conflict:{target.name}"


def safe_copy(src: Path, dst: Path) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)


def safe_move(src: Path, dst: Path) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(src), str(dst))


def backup_if_exists(target: Path, dry_run: bool) -> tuple[str, int]:
    if not target.exists():
        return "", 0
    backup = timestamped_path(ARCHIVE_OVERWRITTEN, target.name)
    if not dry_run:
        safe_copy(target, backup)
    return str(backup), 1


def convert_ads_xlsx_to_csv(src_xlsx: Path, target_csv: Path, dry_run: bool) -> None:
    if dry_run:
        return
    wb = None
    try:
        wb = load_workbook(src_xlsx, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        target_csv.parent.mkdir(parents=True, exist_ok=True)
        with target_csv.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if c is None else c for c in row])
    finally:
        if wb:
            wb.close()


def save_manifest(rows: list[ManifestRow]) -> None:
    payload = [asdict(x) for x in rows]
    with IMPORT_MANIFEST_JSON.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    wb = Workbook()
    ws = wb.active
    ws.title = "import_manifest"
    headers = list(ManifestRow.__dataclass_fields__.keys())
    ws.append(headers)
    for r in rows:
        d = asdict(r)
        ws.append([d.get(h, "") for h in headers])
    wb.save(IMPORT_MANIFEST_XLSX)


def process_file(
    path: Path,
    *,
    dry_run: bool,
    asin_map: dict[str, str],
    selected_ads: Path | None,
    selected_erp: Path | None,
    reserved_targets: set[Path] | None = None,
) -> tuple[ManifestRow, int]:
    t0 = time.perf_counter()
    modified = datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
    created_at = now_str()
    backup_count = 0

    def guard() -> None:
        if time.perf_counter() - t0 > PER_FILE_TIMEOUT_SEC:
            raise TimeoutError("error_timeout")

    detected_type = "unknown"
    detected_marketplace = detect_marketplace_from_name(path.name)
    period_type = "unknown"
    detected_date_range = detect_date_range_from_name(path.name)
    target_path = ""
    archive_path = ""
    status = "unknown"
    reason = ""
    rows = 0
    cols = 0
    asin_count = 0
    overwritten = 0
    target_conflict = ""
    enhanced_detected_from = ""
    enhanced_freshness = ""
    diagnosis_usage = ""

    print(f"正在处理：{path.name}")
    try:
        guard()
        headers: list[str]
        if path.suffix.lower() == ".csv":
            headers, rows, cols = read_csv_preview(path)
        else:
            headers, rows, cols = read_xlsx_preview(path)
        guard()

        detected_type = infer_type(path, headers)
        enhanced_header_info: dict[str, object] | None = None
        if detected_type in {"traffic_sales", "search_query_performance"} and path.suffix.lower() == ".xlsx":
            period_type = detect_period_type(headers)
            mkts, asin_count = infer_enhanced_marketplaces_and_asin_count(path, asin_map)
            if detected_marketplace == "UNKNOWN":
                if len(mkts) == 1:
                    detected_marketplace = next(iter(mkts))
                elif len(mkts) > 1:
                    detected_marketplace = "MULTI"
                else:
                    detected_marketplace = "UNKNOWN"
            if detected_marketplace in {"UK", "US", "DE"}:
                enhanced_header_info = detect_enhanced_header_info(path, detected_marketplace)
                if enhanced_header_info and enhanced_header_info.get("detected_from") == "header":
                    period_type = str(enhanced_header_info.get("format_type") or period_type)
                    detected_date_range = str(enhanced_header_info.get("detected_date_range") or detected_date_range)
                    enhanced_detected_from = str(enhanced_header_info.get("detected_from") or "")
                    enhanced_freshness = str(enhanced_header_info.get("freshness") or "")

        target = (
            build_enhanced_target_from_header(enhanced_header_info)
            if enhanced_header_info
            else None
        )
        if target is None:
            target = build_target_path(detected_type, detected_marketplace if detected_marketplace != "MULTI" else "UK", detected_date_range, period_type, path)
        if detected_type in {"traffic_sales", "search_query_performance"} and reserved_targets is not None:
            target, target_conflict = disambiguate_same_batch_target(target, path, reserved_targets)
        target_path = str(target)

        if detected_type == "ads_report_all" and selected_ads and path != selected_ads:
            status = "skipped_old_duplicate"
            reason = "older_ads_file"
        elif detected_type == "erp_sales_all" and selected_erp and path != selected_erp:
            status = "skipped_old_duplicate"
            reason = "older_erp_file"
        elif detected_type == "unknown":
            status = "unknown"
            reason = "无法识别文件类型"
        elif detected_marketplace == "MULTI":
            status = "error"
            reason = "增强文件包含多个站点，当前版本未自动拆分"
        elif detected_type in {"traffic_sales", "search_query_performance"} and detected_marketplace == "UNKNOWN":
            status = "unknown"
            reason = "无法识别站点"
        else:
            status = "imported"

        if status == "imported" and detected_type in {"traffic_sales", "search_query_performance"}:
            diagnosis_usage = "pending_report_refresh"

        if reserved_targets is not None and status == "imported":
            reserved_targets.add(target.resolve())

        if not dry_run:
            if status == "imported":
                if detected_type == "ads_report_all":
                    bpath, b = backup_if_exists(target, dry_run=False)
                    backup_count += b
                    overwritten = b
                    if bpath:
                        archive_path = bpath
                    if path.suffix.lower() == ".xlsx":
                        convert_ads_xlsx_to_csv(path, target, dry_run=False)
                    else:
                        safe_copy(path, target)
                else:
                    bpath, b = backup_if_exists(target, dry_run=False)
                    backup_count += b
                    overwritten = b
                    if bpath:
                        archive_path = bpath
                    safe_copy(path, target)

                moved = timestamped_path(ARCHIVE_IMPORTED, path.name)
                safe_move(path, moved)
                if not archive_path:
                    archive_path = str(moved)
            elif status in {"unknown", "skipped_old_duplicate"}:
                folder = UNKNOWN_DIR if status == "unknown" else ARCHIVE_IMPORTED
                moved = timestamped_path(folder, path.name)
                safe_move(path, moved)
                archive_path = str(moved)
    except TimeoutError:
        status = "error_timeout"
        reason = "file_detection_timeout_gt_30s"
    except PermissionError:
        status = "error"
        reason = "file_locked_or_permission_denied"
    except Exception as exc:  # noqa: BLE001
        status = "error"
        reason = f"{type(exc).__name__}: {exc}"

    print(f"检测类型：{detected_type}")
    print(f"目标路径：{target_path or '-'}")
    print(f"状态：{status}")

    row = ManifestRow(
        original_filename=path.name,
        original_path=str(path),
        detected_type=detected_type,
        detected_marketplace=detected_marketplace,
        period_type=period_type,
        detected_date_range=detected_date_range,
        target_path=target_path,
        archive_path=archive_path,
        status=status,
        reason=reason,
        modified_time=modified,
        rows=rows,
        columns=cols,
        asin_count=asin_count,
        created_at=created_at,
        target_conflict=target_conflict,
        overwritten=overwritten,
        enhanced_detected_from=enhanced_detected_from,
        enhanced_freshness=enhanced_freshness,
        diagnosis_usage=diagnosis_usage,
    )
    return row, backup_count


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="仅识别，不移动文件")
    args = parser.parse_args()

    ensure_dirs()
    started = time.perf_counter()

    snapshot = [p for p in INBOX.iterdir() if p.is_file() and p.suffix.lower() in SCAN_EXTENSIONS]
    snapshot.sort(key=lambda x: x.stat().st_mtime, reverse=True)

    selected_ads = max(
        (p for p in snapshot if re.search(r"ads_report_all|amazon_ads_all", p.name, re.IGNORECASE)),
        key=lambda x: x.stat().st_mtime,
        default=None,
    )
    selected_erp = max(
        (p for p in snapshot if p.suffix.lower() == ".xlsx" and re.search(r"销量统计|sales_report|asin", p.name, re.IGNORECASE)),
        key=lambda x: x.stat().st_mtime,
        default=None,
    )

    asin_map = load_asin_marketplace_map()
    manifest_rows: list[ManifestRow] = []
    backup_count_total = 0
    reserved_targets: set[Path] = set()

    for p in snapshot:
        if time.perf_counter() - started > TOTAL_TIMEOUT_SEC:
            manifest_rows.append(
                ManifestRow(
                    original_filename=p.name,
                    original_path=str(p),
                    detected_type="unknown",
                    detected_marketplace="UNKNOWN",
                    period_type="unknown",
                    detected_date_range="unknown",
                    target_path="",
                    archive_path="",
                    status="error_timeout",
                    reason="total_timeout_gt_120s",
                    modified_time=datetime.fromtimestamp(p.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                    rows=0,
                    columns=0,
                    asin_count=0,
                    created_at=now_str(),
                )
            )
            continue

        row, bc = process_file(
            p,
            dry_run=args.dry_run,
            asin_map=asin_map,
            selected_ads=selected_ads,
            selected_erp=selected_erp,
            reserved_targets=reserved_targets,
        )
        manifest_rows.append(row)
        backup_count_total += bc

    save_manifest(manifest_rows)

    imported = sum(1 for r in manifest_rows if r.status in {"imported", "split_imported"})
    unknown = sum(1 for r in manifest_rows if r.status == "unknown")
    error = sum(1 for r in manifest_rows if r.status.startswith("error"))

    print("\n========== Import Summary ==========")
    print(f"dry_run: {'YES' if args.dry_run else 'NO'}")
    print(f"扫描到的文件数: {len(snapshot)}")
    print(f"成功导入/识别: {imported}")
    print(f"unknown: {unknown}")
    print(f"error: {error}")
    print(f"覆盖备份文件数: {backup_count_total}")
    print(f"manifest: {IMPORT_MANIFEST_XLSX}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

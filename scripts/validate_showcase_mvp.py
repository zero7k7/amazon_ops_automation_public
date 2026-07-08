from __future__ import annotations

import ast
import html
import json
import math
import os
import re
import subprocess
import sys
import time
from collections import Counter
from datetime import datetime, timezone
from hashlib import sha256
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.autoopt_feedback import (
    ACTION_REVIEW_REQUIRED_FIELDS,
    KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS,
    NEGATIVE_DISPLAY_OUTCOMES,
    POSITIVE_DISPLAY_OUTCOMES,
)
from src.html_pages.components_review import (
    _action_effect_review_sort_key as _html_action_effect_review_sort_key,
    _keyword_review_sort_key as _html_keyword_review_sort_key,
    _review_display_judgement as _html_review_display_judgement,
)
from src.product_decision_layer import FRONTEND_GATED_GROWTH_ACTIONS as GROWTH_ACTIONS
from src.product_decision_layer import PRODUCT_FINAL_DECISION_REQUIRED_FIELDS
from scripts.run_daily_update import (
    DAILY_EXCEL_CONSISTENCY_SPECS,
    ACTION_REVIEW_IDENTITY_FIELDS as DAILY_ACTION_REVIEW_IDENTITY_FIELDS,
    DIAGNOSIS_ROW_CONTENT_FIELDS as DAILY_DIAGNOSIS_ROW_CONTENT_FIELDS,
    DIAGNOSIS_ROW_IDENTITY_FIELDS as DAILY_DIAGNOSIS_ROW_IDENTITY_FIELDS,
    KEYWORD_REVIEW_IDENTITY_FIELDS as DAILY_KEYWORD_REVIEW_IDENTITY_FIELDS,
    PRODUCT_DECISION_BLOCKED_ACTION_LABELS,
    REQUIRED_REFRESHED_OUTPUTS as DAILY_REQUIRED_REFRESHED_OUTPUTS,
    REQUIRED_REPORT_VIEW_SNAPSHOT_DICT_KEYS as DAILY_REQUIRED_SNAPSHOT_DICT_KEYS,
    REQUIRED_REPORT_VIEW_SNAPSHOT_LIST_KEYS as DAILY_REQUIRED_SNAPSHOT_LIST_KEYS,
    TODAY_TASK_QUEUE_CONTENT_FIELDS as DAILY_TODAY_TASK_QUEUE_CONTENT_FIELDS,
    TODAY_TASK_QUEUE_IDENTITY_FIELDS as DAILY_TODAY_TASK_QUEUE_IDENTITY_FIELDS,
    TOMORROW_REVIEW_CONTENT_FIELDS as DAILY_TOMORROW_REVIEW_CONTENT_FIELDS,
    TOMORROW_REVIEW_IDENTITY_FIELDS as DAILY_TOMORROW_REVIEW_IDENTITY_FIELDS,
    _asset_content_failures as daily_asset_content_failures,
    _frontend_coverage_aggregate_from_snapshots as daily_frontend_coverage_aggregate_from_snapshots,
    _frontend_coverage_display_counts as daily_frontend_coverage_display_counts,
    _frontend_coverage_display_tokens as daily_frontend_coverage_display_tokens,
    _frontend_coverage_excel_expected as daily_frontend_coverage_excel_expected,
    _frontend_coverage_summary_failures as daily_frontend_coverage_summary_failures,
    _html_contains_any_token_in_contexts as daily_html_contains_any_token_in_contexts,
    _html_contains_token_near as daily_html_contains_token_near,
    _html_contains_token_in_contexts as daily_html_contains_token_in_contexts,
    _listing_frontend_evidence_failures as daily_listing_frontend_evidence_failures,
    _marketplace_sort_key as daily_marketplace_sort_key,
    _positive_review_policy_claim as daily_positive_review_policy_claim,
    _product_decision_contract_attr_failures as daily_product_decision_contract_attr_failures,
    _product_decision_contexts as daily_product_decision_contexts,
    _product_frontend_evidence_tokens as daily_product_frontend_evidence_tokens,
    _product_operation_ad_rows_from_snapshot as daily_product_operation_ad_rows_from_snapshot,
    _product_operation_card_failures as daily_product_operation_card_failures,
    _review_html_forbidden_metric_tokens as daily_review_html_forbidden_metric_tokens,
    _summary_review_display_judgement as daily_summary_review_display_judgement,
    _summary_review_display_next_step as daily_summary_review_display_next_step,
    _summary_review_forbidden_early_metric_tokens as daily_summary_review_forbidden_early_metric_tokens,
    _summary_review_needs_display_guard as daily_summary_review_needs_display_guard,
    report_refresh_failures,
)

SAFE_RUN_ROOT = ROOT / "data" / "output" / "safe_run"
PRODUCT_GATED_AD_ACTIONS = GROWTH_ACTIONS | {"create_exact_low_budget"}
VALIDATION_RECEIPT = ROOT / "data" / "output" / "showcase_validation_receipt.json"
VALIDATION_RECEIPT_SCHEMA_VERSION = 1


def _daily_required_output_name(path: Path) -> str:
    try:
        return path.relative_to(ROOT / "data" / "output").as_posix()
    except ValueError:
        return path.name


DAILY_REFRESH_REQUIRED_FILES = [
    _daily_required_output_name(path)
    for path in DAILY_REQUIRED_REFRESHED_OUTPUTS
]
REQUIRED_FILES = [name for name in DAILY_REFRESH_REQUIRED_FILES if not name.startswith("assets/")]
REQUIRED_ASSET_FILES = [name for name in DAILY_REFRESH_REQUIRED_FILES if name.startswith("assets/")]
REQUIRED_SNAPSHOT_KEYS = DAILY_REQUIRED_SNAPSHOT_LIST_KEYS | DAILY_REQUIRED_SNAPSHOT_DICT_KEYS
HTML_ERROR_MARKERS = ["Traceback", "KeyError", "NoneType"]
REQUIRED_HTML_MARKERS = {
    "dashboard.html": ["运营状态入口", "打开三分钟摘要", "打开 ALL 运营控制台"],
    "latest_recommendations.html": ["今天广告动作", "市场调查", "提交今日数据", "系统结论", "融合诊断"],
    "summary.html": ["三分钟摘要"],
    "uk_report.html": ["亚马逊运营日报｜UK", "广告状态", "数据质量与增强数据"],
    "us_report.html": ["亚马逊运营日报｜US", "广告状态", "数据质量与增强数据"],
    "de_report.html": ["亚马逊运营日报｜DE", "广告状态", "数据质量与增强数据"],
}
FORBIDDEN_HTML_MARKERS = {
    "latest_recommendations.html": [
        "需要确认的问题",
        "疑似前台竞争力需要确认",
        "<h2>前台数据更新</h2>",
        "<h2>刷新前台",
        "<h2>重新检查前台",
    ],
    "uk_report.html": ["<h2>前台数据更新</h2>"],
    "us_report.html": ["<h2>前台数据更新</h2>"],
    "de_report.html": ["<h2>前台数据更新</h2>"],
}
FRONTEND_CHECK_OK_STATUSES = {"已自动检查"}
OBSERVATION_ACTIONS = {"观察", "保留", "保留观察", "无需操作", "仅背景参考"}
AD_COPY_ANCHOR = 'id="today-ad-actions-all"'
DATE_RE = re.compile(r"20\d{2}-\d{2}-\d{2}")
PRODUCT_DECISION_IDENTITY_FIELDS = list(
    DAILY_EXCEL_CONSISTENCY_SPECS["product_final_decision_rows"]["identity_fields"]
)
INVENTORY_REPLENISHMENT_IDENTITY_FIELDS = list(
    DAILY_EXCEL_CONSISTENCY_SPECS["inventory_replenishment_rows"]["identity_fields"]
)
INVENTORY_REPLENISHMENT_REQUIRED_FIELDS = [
    "marketplace",
    "sku",
    "asin",
    "product_name",
    "available_stock",
    "days_of_cover",
    "recommended_reorder_qty",
    "stock_risk_level",
    "stock_status_label",
    "replenishment_advice",
]
PRODUCT_DECISION_EXCEL_CONTENT_FIELDS = list(
    DAILY_EXCEL_CONSISTENCY_SPECS["product_final_decision_rows"]["content_fields"]
)
INVENTORY_REPLENISHMENT_EXCEL_CONTENT_FIELDS = list(
    DAILY_EXCEL_CONSISTENCY_SPECS["inventory_replenishment_rows"]["content_fields"]
)
TODAY_TASK_QUEUE_IDENTITY_FIELDS = list(DAILY_TODAY_TASK_QUEUE_IDENTITY_FIELDS)
TODAY_TASK_QUEUE_CONTENT_FIELDS = list(DAILY_TODAY_TASK_QUEUE_CONTENT_FIELDS)
TOMORROW_REVIEW_IDENTITY_FIELDS = list(DAILY_TOMORROW_REVIEW_IDENTITY_FIELDS)
TOMORROW_REVIEW_REQUIRED_FIELDS = [
    "marketplace",
    "sku",
    "asin",
    "review_reason",
    "tomorrow_check",
    "trigger_action",
]
TOMORROW_REVIEW_CONTENT_FIELDS = list(DAILY_TOMORROW_REVIEW_CONTENT_FIELDS)
DIAGNOSIS_ROW_IDENTITY_FIELDS = list(DAILY_DIAGNOSIS_ROW_IDENTITY_FIELDS)
DIAGNOSIS_ROW_REQUIRED_FIELDS = [
    "marketplace",
    "SKU",
    "ASIN",
    "诊断类型",
    "主因",
    "关键证据",
    "建议动作",
]
DIAGNOSIS_ROW_CONTENT_FIELDS = list(DAILY_DIAGNOSIS_ROW_CONTENT_FIELDS)
ACTION_REVIEW_IDENTITY_FIELDS = list(DAILY_ACTION_REVIEW_IDENTITY_FIELDS)
KEYWORD_REVIEW_IDENTITY_FIELDS = list(DAILY_KEYWORD_REVIEW_IDENTITY_FIELDS)


def _is_frontend_fallback_status(status: str, freshness: str, findings: str) -> bool:
    text = " ".join([status, freshness, findings])
    return "待前台检查" in text or "沿用" in text


def _validate_frontend_rows(marketplace: str, rows: list[dict]) -> int:
    for idx, row in enumerate(rows, start=1):
        status = str(row.get("frontend_check_status") or "").strip()
        freshness = str(row.get("frontend_data_freshness") or "").strip()
        findings = str(row.get("frontend_findings") or "").strip()
        asin = row.get("asin") or "unknown_asin"
        tier = str(row.get("frontend_evidence_tier") or "").strip()
        display_tier = str(row.get("frontend_evidence_display_tier") or "").strip()
        decision_tier = str(row.get("frontend_decision_evidence_tier") or display_tier or tier).strip()
        auto_conclusion = str(row.get("frontend_auto_conclusion") or "").strip()
        strong_marked = (
            decision_tier == "强诊断可用"
            or _truthy(row.get("frontend_evidence_is_strong"))
            or auto_conclusion == "FRONTEND_OK"
        )
        cached_or_pending = status != "已自动检查" or status.startswith("沿用") or _truthy(row.get("frontend_cache_used"))
        if cached_or_pending and strong_marked:
            return fail(f"{marketplace} frontend row {idx} {asin} cached or pending evidence marked strong")
        if strong_marked:
            warning = str(row.get("frontend_price_currency_warning") or "").strip()
            location_warning = str(row.get("frontend_location_warning") or "").strip()
            failure = str(row.get("frontend_failure_category") or "").strip()
            location_scope = str(row.get("frontend_location_scope") or "").strip().lower()
            location_exact_explicit = "frontend_location_exact" in row
            location_verified = _truthy(row.get("frontend_location_verified"))
            location_uncertain = _truthy(row.get("frontend_location_uncertain"))
            partial_search = _truthy(row.get("frontend_search_partial_evidence")) or str(row.get("frontend_search_status") or "").strip() == "已读取部分结果"
            search_status = str(row.get("frontend_search_status") or "").strip()
            competitor_count = _number(
                row.get("frontend_competitor_count")
                if row.get("frontend_competitor_count") not in (None, "")
                else row.get("comparable_competitor_count")
            )
            comparable_competitor_count = _number(row.get("comparable_competitor_count"))
            comparability = str(row.get("competitor_comparability") or "").strip().lower()
            quality_score = _number(row.get("frontend_evidence_quality_score"))
            if decision_tier != "强诊断可用":
                return fail(f"{marketplace} frontend row {idx} {asin} strong evidence missing strong decision tier")
            if not _truthy(row.get("frontend_evidence_is_strong")):
                return fail(f"{marketplace} frontend row {idx} {asin} strong evidence missing explicit strong flag")
            if warning:
                return fail(f"{marketplace} frontend row {idx} {asin} currency warning marked strong")
            if location_warning:
                return fail(f"{marketplace} frontend row {idx} {asin} location warning marked strong")
            if failure and failure != "none":
                return fail(f"{marketplace} frontend row {idx} {asin} failure category marked strong")
            if location_uncertain or location_scope in {"wrong", "missing", "unknown", "marketplace"} or (
                location_exact_explicit and not _truthy(row.get("frontend_location_exact"))
            ):
                return fail(f"{marketplace} frontend row {idx} {asin} uncertain location marked strong")
            if location_scope != "exact" and not _truthy(row.get("frontend_location_exact")):
                return fail(f"{marketplace} frontend row {idx} {asin} strong evidence missing exact location")
            if not location_verified:
                return fail(f"{marketplace} frontend row {idx} {asin} strong evidence missing verified location")
            if partial_search:
                return fail(f"{marketplace} frontend row {idx} {asin} partial search evidence marked strong")
            if search_status != "已自动检查":
                return fail(f"{marketplace} frontend row {idx} {asin} strong evidence missing successful search page")
            if competitor_count is None:
                return fail(f"{marketplace} frontend row {idx} {asin} strong evidence missing competitor count")
            if not comparability:
                return fail(f"{marketplace} frontend row {idx} {asin} strong evidence missing competitor comparability")
            if (
                comparability != "high"
                or competitor_count < 2
                or (
                    comparable_competitor_count is not None
                    and comparable_competitor_count < 2
                )
            ):
                return fail(f"{marketplace} frontend row {idx} {asin} weak competitor evidence marked strong")
            if quality_score is None or quality_score < 75:
                return fail(f"{marketplace} frontend row {idx} {asin} strong evidence quality score below threshold")
        if status in FRONTEND_CHECK_OK_STATUSES:
            continue
        if not _is_frontend_fallback_status(status, freshness, findings):
            return fail(
                f"{marketplace} frontend row {idx} {asin} missing cached-date or pending-check status"
            )
        if status.startswith("沿用") and not DATE_RE.search(" ".join([status, freshness, findings])):
            return fail(f"{marketplace} frontend row {idx} {asin} cache status missing date")
    return 0


def _number(value: object) -> float | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if text.endswith("%"):
        try:
            number = float(text[:-1].strip()) / 100
        except ValueError:
            return None
        return None if math.isnan(number) else number
    try:
        number = float(text)
    except ValueError:
        return None
    return None if math.isnan(number) else number


def _ratio_number(value: object) -> float | None:
    number = _number(value)
    if number is None:
        return None
    return number / 100 if number > 1 else number


def _action_set(value: object) -> set[str]:
    if isinstance(value, list):
        return {str(item).strip() for item in value if str(item).strip()}
    if isinstance(value, tuple) or isinstance(value, set):
        return {str(item).strip() for item in value if str(item).strip()}
    text = str(value or "").strip()
    if not text:
        return set()
    if text.startswith("[") and text.endswith("]"):
        for parser in (json.loads, ast.literal_eval):
            try:
                parsed = parser(text)
            except (ValueError, SyntaxError, TypeError, json.JSONDecodeError):
                continue
            if isinstance(parsed, (list, tuple, set)):
                return {str(item).strip() for item in parsed if str(item).strip()}
    for separator in ["；", ";", ",", "，", "|", "/"]:
        text = text.replace(separator, "\n")
    return {item.strip() for item in text.splitlines() if item.strip()}


def _growth_action_bucket(row: dict[str, object]) -> str:
    normalized = str(row.get("normalized_action") or "").strip()
    if normalized in GROWTH_ACTIONS:
        return normalized
    if normalized in {"growth_test", "create_exact", "create_exact_low_budget"}:
        return "create_exact_low_budget"
    text = " ".join(
        str(row.get(field) or "")
        for field in [
            "suggested_action",
            "scale_action",
            "copy_action_line",
            "copy_block",
            "today_action",
            "action_detail",
            "manual_action_taken",
        ]
    ).strip().lower()
    if not text:
        return ""
    negative_markers = [
        "不加价",
        "不提高竞价",
        "停止加价",
        "禁止加价",
        "不能加价",
        "不追加预算",
        "不加预算",
        "不提高预算",
        "停止追加",
        "禁止放量",
        "不放量",
        "不能放量",
        "不推大词放量",
    ]
    if any(marker in text for marker in negative_markers):
        return ""
    if any(
        marker in text
        for marker in [
            "小预算",
            "精准测试",
            "拉精准",
            "创建精准",
            "新建精准",
            "开精准",
            "create exact",
            "exact test",
            "low budget",
        ]
    ):
        return "create_exact_low_budget"
    if (
        "budget up" in text
        or "increase budget" in text
        or "raise budget" in text
        or "加预算" in text
        or "追加预算" in text
        or "提高预算" in text
    ):
        return "budget_up"
    if "broad scale" in text or "放量" in text:
        return "broad_scale"
    bid_raise_markers = [
        "bid up",
        "increase bid",
        "raise bid",
        "加价",
        "提高竞价",
        "上调竞价",
        "调高竞价",
        "提高出价",
        "上调出价",
        "调高出价",
    ]
    price_or_coupon_context = any(marker in text for marker in ["优惠券", "coupon", "价格", "售价", "price"])
    if any(marker in text for marker in bid_raise_markers) or (
        not price_or_coupon_context and re.search(r"加\s*\d+(?:\.\d+)?\s*%", text)
    ):
        return "bid_up"
    return ""


def _validate_product_final_decision_rows(marketplace: str, rows: list[dict]) -> int:
    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            return fail(f"{marketplace} product final decision row {idx} must be an object")
        asin = row.get("asin") or "unknown_asin"
        missing = [field for field in PRODUCT_FINAL_DECISION_REQUIRED_FIELDS if field not in row]
        if missing:
            return fail(
                f"{marketplace} product final decision row {idx} {asin} missing fields: {', '.join(missing)}"
            )
        row_marketplace = str(row.get("marketplace") or "").strip().upper()
        if marketplace in {"UK", "US", "DE"}:
            if not row_marketplace:
                return fail(f"{marketplace} product final decision row {idx} {asin} missing marketplace value")
            if row_marketplace != marketplace:
                return fail(
                    f"{marketplace} product final decision row {idx} {asin} contains {row_marketplace} marketplace data"
                )
        allowed = _action_set(row.get("today_allowed_actions"))
        blocked = _action_set(row.get("today_blocked_actions"))
        overlap = sorted(allowed & blocked)
        if overlap:
            return fail(
                f"{marketplace} product final decision row {idx} {asin} has actions both allowed and blocked: {', '.join(overlap)}"
            )
        state = str(row.get("frontend_evidence_state") or "").strip()
        tier = str(row.get("frontend_evidence_tier") or "").strip()
        display_tier = str(row.get("frontend_evidence_display_tier") or "").strip()
        decision_tier = str(row.get("frontend_decision_evidence_tier") or display_tier or tier).strip()
        failure = str(row.get("frontend_failure_category") or "").strip()
        check_status = str(row.get("frontend_check_status") or "").strip()
        search_status = str(row.get("frontend_search_status") or "").strip()
        partial_search = _truthy(row.get("frontend_search_partial_evidence")) or search_status == "已读取部分结果"
        currency_warning = str(row.get("frontend_price_currency_warning") or "").strip()
        location_warning = str(row.get("frontend_location_warning") or "").strip()
        auto_conclusion = str(row.get("frontend_auto_conclusion") or "").strip()
        location_scope = str(row.get("frontend_location_scope") or "").strip().lower()
        location_exact = (
            _truthy(row.get("frontend_location_exact"))
            if "frontend_location_exact" in row
            else location_scope == "exact"
        )
        location_verified = _truthy(row.get("frontend_location_verified"))
        competitor_count = _number(row.get("comparable_competitor_count"))
        comparability = str(row.get("competitor_comparability") or "").strip().lower()
        quality_score = _number(row.get("frontend_evidence_quality_score"))
        growth_allowed = sorted(allowed & GROWTH_ACTIONS)
        strong_marked = (
            state == "ok_high"
            or decision_tier == "强诊断可用"
            or _truthy(row.get("frontend_evidence_is_strong"))
        )
        if not growth_allowed and strong_marked:
            if display_tier != "强诊断可用":
                return fail(f"{marketplace} product final decision row {idx} {asin} marks strong frontend evidence without strong frontend display tier")
            if decision_tier != "强诊断可用":
                return fail(f"{marketplace} product final decision row {idx} {asin} marks strong frontend evidence without strong frontend decision tier")
            if not _truthy(row.get("frontend_evidence_is_strong")):
                return fail(f"{marketplace} product final decision row {idx} {asin} marks strong frontend evidence without explicit strong frontend flag")
            if check_status != "已自动检查" or _truthy(row.get("frontend_cache_used")):
                return fail(f"{marketplace} product final decision row {idx} {asin} marks strong frontend evidence with cached frontend evidence")
            if partial_search:
                return fail(f"{marketplace} product final decision row {idx} {asin} marks strong frontend evidence with partial search evidence")
            if search_status != "已自动检查":
                return fail(f"{marketplace} product final decision row {idx} {asin} marks strong frontend evidence without successful search page")
            if currency_warning:
                return fail(f"{marketplace} product final decision row {idx} {asin} marks strong frontend evidence with currency warning")
            if location_warning:
                return fail(f"{marketplace} product final decision row {idx} {asin} marks strong frontend evidence with location warning")
            if failure and failure != "none":
                return fail(f"{marketplace} product final decision row {idx} {asin} marks strong frontend evidence with frontend failure")
            if auto_conclusion != "FRONTEND_OK":
                return fail(f"{marketplace} product final decision row {idx} {asin} marks strong frontend evidence without FRONTEND_OK conclusion")
            if not location_exact:
                return fail(f"{marketplace} product final decision row {idx} {asin} marks strong frontend evidence without exact frontend location")
            if not location_verified:
                return fail(f"{marketplace} product final decision row {idx} {asin} marks strong frontend evidence without verified frontend location")
            if not comparability:
                return fail(f"{marketplace} product final decision row {idx} {asin} marks strong frontend evidence without competitor comparability")
            if competitor_count is None or competitor_count < 2 or comparability != "high":
                return fail(f"{marketplace} product final decision row {idx} {asin} marks strong frontend evidence with weak competitor evidence")
            if quality_score is None or quality_score < 75:
                return fail(f"{marketplace} product final decision row {idx} {asin} marks strong frontend evidence without high frontend quality score")
        if (
            not growth_allowed
            and (state != "ok_high" or decision_tier != "强诊断可用" or auto_conclusion != "FRONTEND_OK")
            and not GROWTH_ACTIONS.issubset(blocked)
        ):
            missing = sorted(GROWTH_ACTIONS.difference(blocked))
            return fail(
                f"{marketplace} product final decision row {idx} {asin} does not explicitly block growth actions under weak frontend evidence: {missing}"
            )
        if not growth_allowed:
            continue
        if state != "ok_high":
            return fail(
                f"{marketplace} product final decision row {idx} {asin} allows growth without ok_high frontend evidence"
            )
        if display_tier != "强诊断可用":
            return fail(f"{marketplace} product final decision row {idx} {asin} allows growth without strong frontend display tier")
        if decision_tier != "强诊断可用":
            return fail(f"{marketplace} product final decision row {idx} {asin} allows growth without strong frontend decision tier")
        if not _truthy(row.get("frontend_evidence_is_strong")):
            return fail(f"{marketplace} product final decision row {idx} {asin} allows growth without explicit strong frontend flag")
        if check_status != "已自动检查":
            return fail(f"{marketplace} product final decision row {idx} {asin} allows growth without current frontend check")
        if partial_search:
            return fail(f"{marketplace} product final decision row {idx} {asin} allows growth with partial search evidence")
        if search_status != "已自动检查":
            return fail(f"{marketplace} product final decision row {idx} {asin} allows growth without successful search page")
        if _truthy(row.get("frontend_cache_used")):
            return fail(f"{marketplace} product final decision row {idx} {asin} allows growth with cached frontend evidence")
        if currency_warning:
            return fail(f"{marketplace} product final decision row {idx} {asin} allows growth with currency warning")
        if location_warning:
            return fail(f"{marketplace} product final decision row {idx} {asin} allows growth with location warning")
        if failure and failure != "none":
            return fail(f"{marketplace} product final decision row {idx} {asin} allows growth with frontend failure")
        if auto_conclusion != "FRONTEND_OK":
            return fail(f"{marketplace} product final decision row {idx} {asin} allows growth without FRONTEND_OK conclusion")
        if not location_exact:
            return fail(f"{marketplace} product final decision row {idx} {asin} allows growth without exact frontend location")
        if not location_verified:
            return fail(f"{marketplace} product final decision row {idx} {asin} allows growth without verified frontend location")
        if not comparability:
            return fail(f"{marketplace} product final decision row {idx} {asin} allows growth without competitor comparability")
        if competitor_count is None or competitor_count < 2 or comparability != "high":
            return fail(f"{marketplace} product final decision row {idx} {asin} allows growth with weak competitor evidence")
        if quality_score is None or quality_score < 75:
            return fail(f"{marketplace} product final decision row {idx} {asin} allows growth without high frontend quality score")
    return 0


def _validate_task_queue_growth_gate(
    marketplace: str,
    task_rows: list[dict],
    product_decision_rows: list[dict],
    *,
    row_source: str = "today_task_queue_rows",
) -> int:
    decisions: dict[tuple[str, str, str], dict[str, object]] = {}
    for row in product_decision_rows:
        if not isinstance(row, dict):
            continue
        key = (
            str(row.get("marketplace") or marketplace).strip().upper(),
            str(row.get("sku") or "").strip(),
            str(row.get("asin") or "").strip().upper(),
        )
        if key[0] and key[2]:
            decisions[key] = row

    for idx, row in enumerate(task_rows, start=1):
        if not isinstance(row, dict):
            continue
        confirmed_status = str(row.get("confirmed_status") or "").strip()
        if confirmed_status in {"已执行", "已核查", "已忽略", "仅背景参考"}:
            continue
        bucket = _growth_action_bucket(row)
        if bucket not in PRODUCT_GATED_AD_ACTIONS:
            continue
        key = (
            str(row.get("marketplace") or marketplace).strip().upper(),
            str(row.get("sku") or "").strip(),
            str(row.get("asin") or "").strip().upper(),
        )
        asin = key[2] or "unknown_asin"
        decision = decisions.get(key) or decisions.get((key[0], "", key[2]))
        if not decision:
            return fail(
                f"{marketplace} {row_source} row {idx} {asin} contains growth action {bucket} without matching product final decision gate"
            )
        allowed = _action_set(decision.get("today_allowed_actions"))
        blocked = _action_set(decision.get("today_blocked_actions"))
        if bucket in blocked:
            return fail(
                f"{marketplace} {row_source} row {idx} {asin} contains growth action {bucket} blocked by product final decision"
            )
        if bucket not in allowed:
            return fail(
                f"{marketplace} {row_source} row {idx} {asin} contains growth action {bucket} not allowed by product final decision"
            )
    return 0


def _validate_marketplace_scoped_rows(marketplace: str, rows: list[dict], label: str) -> int:
    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            return fail(f"{marketplace} {label} row {idx} must be an object")
        target = str(row.get("product_name") or row.get("asin") or row.get("sku") or "unknown_row")
        row_marketplace = str(row.get("marketplace") or "").strip().upper()
        if not row_marketplace:
            return fail(f"{marketplace} {label} row {idx} {target} missing marketplace value")
        if row_marketplace != marketplace:
            return fail(f"{marketplace} {label} row {idx} {target} contains {row_marketplace} marketplace data")
    return 0


def _iter_ad_action_rows(snapshot: dict) -> list[dict]:
    groups = snapshot.get("today_action_groups") or {}
    rows = groups.get("广告动作") if isinstance(groups, dict) else []
    return rows if isinstance(rows, list) else []


def _iter_aux_ad_workbench_rows(snapshot: dict) -> list[dict]:
    rows: list[dict] = []
    for key in ["html_search_term_processing_queue_rows", "scale_keyword_rows", "growth_test_rows"]:
        value = snapshot.get(key) or []
        if isinstance(value, list):
            rows.extend(row for row in value if isinstance(row, dict))
    return rows


def _is_observation_ad_row(row: dict) -> bool:
    action = str(row.get("suggested_action") or row.get("normalized_action") or row.get("today_action") or "").strip()
    return action in OBSERVATION_ACTIONS or any(token in action for token in ["观察", "保留", "无需操作"])


def _validate_ad_observation_rows(marketplace: str, rows: list[dict]) -> int:
    for idx, row in enumerate(rows, start=1):
        if not _is_observation_ad_row(row):
            continue
        if str(row.get("confirmed_status") or "").strip() == "已执行":
            target = row.get("search_term_or_target") or row.get("targeting") or row.get("asin") or "unknown_target"
            return fail(f"{marketplace} ad observation row {idx} {target} is marked executed")
    return 0


def _truthy(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "是", "已验证"}


def _positive_score(value: object) -> bool:
    number = _number(value)
    return bool(number is not None and number > 0)


def _review_target_ratio(row: dict) -> float:
    target = _ratio_number(row.get("current_7d_target_acos") or row.get("target_acos") or row.get("suggested_target_acos"))
    if target is None or target <= 0:
        return 0.10
    return target


def _review_timing_failure(marketplace: str, row: dict, idx: int, target: str, days_since: float | None) -> str | None:
    if days_since is None:
        return None
    window = str(row.get("review_window") or "").strip()
    phase = str(row.get("review_phase") or "").strip()
    status = str(row.get("review_status") or "").strip()
    combined = " ".join(token for token in [window, phase, status] if token)
    if days_since < 3 and any(token in combined for token in ["3天后复盘", "3d_check", "day_3_check", "可做3天复查"]):
        return f"{marketplace} review row {idx} {target} claims 3-day review window before 3 days"
    if days_since < 7 and any(token in combined for token in ["7天后复盘", "day_7_review", "7d_review", "7d_check", "可做7天复查", "进入7天复盘"]):
        return f"{marketplace} review row {idx} {target} claims 7-day review window before 7 days"
    if days_since >= 7 and any(token in combined for token in ["未满3天", "under_3_days", "3天后复盘", "3d_check", "day_3_check"]):
        return f"{marketplace} review row {idx} {target} keeps early review window after 7 days"
    return None


def _review_metric_tracking_failure(marketplace: str, row: dict, idx: int, target: str, days_since: float | None) -> str | None:
    if days_since is None or days_since < 3:
        return None
    required_fields = [
        "current_7d_promoted_ad_orders",
        "current_7d_acos",
        "current_7d_tacos",
        "current_7d_total_orders",
        "current_7d_available_stock",
    ]
    if days_since >= 7:
        required_fields.extend(
            [
                "current_14d_promoted_ad_orders",
                "current_14d_acos",
                "current_14d_tacos",
                "current_14d_total_orders",
                "current_14d_available_stock",
            ]
        )
    for field in required_fields:
        if _number(row.get(field)) is None and not _review_allows_missing_nonfinal_metric(row, field):
            return f"{marketplace} review row {idx} {target} reached review window but missing numeric {field}"
    return None


def _review_allows_missing_nonfinal_metric(row: dict, field: str) -> bool:
    if field not in {
        "current_7d_acos",
        "current_7d_tacos",
        "current_7d_available_stock",
        "current_14d_acos",
        "current_14d_tacos",
        "current_14d_available_stock",
    }:
        return False
    review_outcome = str(row.get("review_outcome") or "").strip()
    if review_outcome in {"effective", "ineffective"}:
        return False
    if _positive_score(row.get("effectiveness_score")):
        return False
    evidence_text = " ".join(
        str(row.get(key) or "")
        for key in ["block_reason", "effect_evidence", "review_status", "review_outcome", "outcome", "judgement"]
    )
    if not any(token in evidence_text for token in ["缺少", "N/A", "不可计算", "人工复查", "样本不足", "数据不足", "insufficient_sample"]):
        return False
    if field.endswith("_tacos"):
        total_field = field.replace("_tacos", "_total_orders")
        return _number(row.get(total_field)) == 0 or any(token in evidence_text for token in ["TACOS", "tacos", "分母"])
    if field.endswith("_acos"):
        orders_field = field.replace("_acos", "_ad_orders")
        sales_field = field.replace("_acos", "_ad_sales")
        return _number(row.get(orders_field)) == 0 or _number(row.get(sales_field)) == 0 or "ACOS" in evidence_text
    if field.endswith("_available_stock"):
        return any(token in evidence_text for token in ["库存", "人工复查", "样本不足", "数据不足", "insufficient_sample"])
    return False


def _review_metric_consistency_failure(marketplace: str, row: dict, idx: int, target: str, days_since: float | None) -> str | None:
    if days_since is None or days_since < 3:
        return None
    allow_total_gap = _review_allows_nonfinal_total_order_gap(row)
    promoted_orders = _number(row.get("current_7d_promoted_ad_orders"))
    ad_orders = _number(row.get("current_7d_ad_orders"))
    if promoted_orders is not None and ad_orders is not None and promoted_orders > ad_orders:
        return f"{marketplace} review row {idx} {target} review metrics have promoted SKU orders above ad orders"
    total_orders = _number(row.get("current_7d_total_orders"))
    if promoted_orders is not None and total_orders is not None and promoted_orders > total_orders and not allow_total_gap:
        return f"{marketplace} review row {idx} {target} review metrics have promoted SKU orders above total orders"
    if days_since < 7:
        return None
    promoted_orders_14d = _number(row.get("current_14d_promoted_ad_orders"))
    if promoted_orders is not None and promoted_orders_14d is not None and promoted_orders_14d < promoted_orders:
        return f"{marketplace} review row {idx} {target} review metrics have 14-day promoted SKU orders below 7-day orders"
    ad_orders_14d = _number(row.get("current_14d_ad_orders"))
    if promoted_orders_14d is not None and ad_orders_14d is not None and promoted_orders_14d > ad_orders_14d:
        return f"{marketplace} review row {idx} {target} review metrics have 14-day promoted SKU orders above ad orders"
    total_orders_14d = _number(row.get("current_14d_total_orders"))
    if (
        promoted_orders_14d is not None
        and total_orders_14d is not None
        and promoted_orders_14d > total_orders_14d
        and not allow_total_gap
    ):
        return f"{marketplace} review row {idx} {target} review metrics have 14-day promoted SKU orders above total orders"
    return None


def _review_allows_nonfinal_total_order_gap(row: dict) -> bool:
    review_outcome = str(row.get("review_outcome") or "").strip()
    if review_outcome in {"effective", "ineffective"}:
        return False
    if _positive_score(row.get("effectiveness_score")):
        return False
    evidence_text = " ".join(
        str(row.get(key) or "")
        for key in ["block_reason", "effect_evidence", "review_status", "outcome", "judgement"]
    )
    return any(token in evidence_text for token in ["总单 0", "缺少TACOS", "人工复查", "ERP", "基线"])


def _review_anchor_failure(marketplace: str, row: dict, idx: int, target: str, *, effective_claim: bool) -> str | None:
    if not effective_claim:
        return None
    if str(row.get("review_data_source") or "").strip() != "execution_anchored_daily":
        return f"{marketplace} review row {idx} {target} effective outcome missing execution anchored daily source"
    required_fields = [
        "pre_7d_start",
        "pre_7d_end",
        "post_3d_start",
        "post_3d_end",
        "post_7d_start",
        "post_7d_end",
        "pre_7d_promoted_ad_orders",
        "pre_7d_total_orders",
        "pre_7d_tacos",
        "post_3d_days",
        "post_3d_promoted_ad_orders",
        "post_3d_total_orders",
        "post_3d_acos",
        "post_3d_tacos",
        "post_3d_available_stock",
        "post_7d_days",
        "post_7d_promoted_ad_orders",
        "post_7d_total_orders",
        "post_7d_acos",
        "post_7d_tacos",
        "post_7d_available_stock",
    ]
    for field in required_fields:
        if str(row.get(field) or "").strip() == "":
            return f"{marketplace} review row {idx} {target} effective outcome missing anchored field {field}"
    post_promoted = _number(row.get("post_7d_promoted_ad_orders"))
    current_promoted = _number(row.get("current_7d_promoted_ad_orders"))
    if post_promoted is None or post_promoted <= 0:
        return f"{marketplace} review row {idx} {target} effective outcome missing positive post 7-day promoted SKU orders"
    post_total = _number(row.get("post_7d_total_orders"))
    if post_total is None or post_total <= 0:
        return f"{marketplace} review row {idx} {target} effective outcome missing positive post 7-day total orders"
    if post_total < post_promoted:
        return f"{marketplace} review row {idx} {target} effective outcome has post 7-day promoted SKU orders above total orders"
    if _ratio_number(row.get("post_7d_acos")) is None:
        return f"{marketplace} review row {idx} {target} effective outcome missing post 7-day ACOS"
    if _ratio_number(row.get("post_7d_tacos")) is None:
        return f"{marketplace} review row {idx} {target} effective outcome missing post 7-day TACOS"
    post_stock = _number(row.get("post_7d_available_stock"))
    if post_stock is None or post_stock <= 0:
        return f"{marketplace} review row {idx} {target} effective outcome missing positive post 7-day available stock"
    if current_promoted is not None and post_promoted != current_promoted:
        return f"{marketplace} review row {idx} {target} effective outcome current 7-day promoted SKU orders differ from post 7-day orders"
    return None


def _validate_review_attribution(marketplace: str, row: dict, idx: int, target: str) -> int:
    review_outcome = str(row.get("review_outcome") or "").strip()
    display_outcome = str(row.get("outcome") or row.get("judgement") or "").strip()
    days_since = _number(row.get("days_since_execution"))
    timing_failure = _review_timing_failure(marketplace, row, idx, target, days_since)
    if timing_failure:
        return fail(timing_failure)
    effectiveness_score = _number(row.get("effectiveness_score"))
    negative_final_claim = review_outcome == "ineffective" or (
        effectiveness_score is not None
        and effectiveness_score < 0
        and review_outcome not in {"needs_manual_review", "not_ready"}
    )
    if negative_final_claim and (days_since is None or days_since < 7):
        return fail(
            f"{marketplace} review row {idx} {target} marks ineffective before 7-day review window"
        )
    if display_outcome in POSITIVE_DISPLAY_OUTCOMES and (days_since is None or days_since < 7):
        return fail(
            f"{marketplace} review row {idx} {target} shows positive display outcome before 7-day review window"
        )
    if display_outcome in NEGATIVE_DISPLAY_OUTCOMES and (days_since is None or days_since < 3):
        return fail(
            f"{marketplace} review row {idx} {target} shows negative display outcome before 3-day review window"
        )
    clicks7 = _number(row.get("current_7d_clicks"))
    spend7 = _number(row.get("current_7d_spend"))
    if (
        display_outcome in {"样本不足", "数据不足"}
        and days_since is not None
        and days_since >= 7
        and ((clicks7 is not None and clicks7 >= 8) or (spend7 is not None and spend7 >= 5))
    ):
        return fail(
            f"{marketplace} review row {idx} {target} shows insufficient sample after 7-day sufficient traffic"
        )
    effective_claim = (
        review_outcome == "effective"
        or display_outcome in POSITIVE_DISPLAY_OUTCOMES
        or _positive_score(row.get("effectiveness_score"))
        or (
            bool(daily_positive_review_policy_claim(row))
            and not (review_outcome == "not_ready" and (days_since is None or days_since < 7))
        )
    )
    if not effective_claim:
        metric_tracking_failure = _review_metric_tracking_failure(marketplace, row, idx, target, days_since)
        if metric_tracking_failure:
            return fail(metric_tracking_failure)
        metric_consistency_failure = _review_metric_consistency_failure(marketplace, row, idx, target, days_since)
        if metric_consistency_failure:
            return fail(metric_consistency_failure)
        review_window = str(row.get("review_window") or "").strip()
        early_review_window = (days_since is not None and 3 <= days_since < 7) or review_window in {
            "3天后复盘",
            "3d_check",
            "day_3_check",
        }
        if early_review_window:
            evidence_text = str(row.get("effect_evidence") or row.get("current_evidence") or "")
            if "3天复查口径" not in evidence_text or "7天结论待补" not in evidence_text:
                return fail(
                    f"{marketplace} review row {idx} {target} 3-day review evidence missing early-window qualifier"
                )
        return 0
    if days_since is None:
        return fail(f"{marketplace} review row {idx} {target} effective outcome missing days_since_execution")
    if days_since < 7:
        return fail(f"{marketplace} review row {idx} {target} marks effective before 7-day review window")
    if _truthy(row.get("halo_only_conversion")):
        return fail(f"{marketplace} review row {idx} {target} marks halo-only conversion as effective")
    if _truthy(row.get("target_sku_not_converted")):
        return fail(f"{marketplace} review row {idx} {target} marks target SKU not converted as effective")
    if not _truthy(row.get("promoted_conversion_improved")):
        return fail(f"{marketplace} review row {idx} {target} effective outcome missing promoted SKU conversion")
    promoted_orders = _number(row.get("current_7d_promoted_ad_orders"))
    if promoted_orders is None or promoted_orders <= 0:
        return fail(f"{marketplace} review row {idx} {target} effective outcome missing positive promoted SKU orders")
    anchor_failure = _review_anchor_failure(marketplace, row, idx, target, effective_claim=effective_claim)
    if anchor_failure:
        return fail(anchor_failure)
    required_numeric_fields = [
        "current_7d_acos",
        "current_7d_tacos",
        "current_7d_total_orders",
        "current_7d_available_stock",
        "current_14d_promoted_ad_orders",
        "current_14d_acos",
        "current_14d_total_orders",
        "current_14d_tacos",
        "current_14d_available_stock",
    ]
    for field in required_numeric_fields:
        if _number(row.get(field)) is None:
            return fail(f"{marketplace} review row {idx} {target} effective outcome missing numeric {field}")
    ad_orders = _number(row.get("current_7d_ad_orders"))
    if ad_orders is not None and promoted_orders > ad_orders:
        return fail(f"{marketplace} review row {idx} {target} effective outcome has promoted SKU orders above ad orders")
    promoted_orders_14d = _number(row.get("current_14d_promoted_ad_orders"))
    if promoted_orders_14d is None or promoted_orders_14d <= 0:
        return fail(f"{marketplace} review row {idx} {target} effective outcome missing positive 14-day promoted SKU orders")
    if promoted_orders_14d < promoted_orders:
        return fail(f"{marketplace} review row {idx} {target} effective outcome has 14-day promoted SKU orders below 7-day orders")
    ad_orders_14d = _number(row.get("current_14d_ad_orders"))
    if ad_orders_14d is not None and promoted_orders_14d > ad_orders_14d:
        return fail(f"{marketplace} review row {idx} {target} effective outcome has 14-day promoted SKU orders above ad orders")
    total_orders = _number(row.get("current_7d_total_orders"))
    if total_orders is None or total_orders <= 0:
        return fail(f"{marketplace} review row {idx} {target} effective outcome missing positive total orders")
    if promoted_orders > total_orders:
        return fail(f"{marketplace} review row {idx} {target} effective outcome has promoted SKU orders above total orders")
    total_orders_14d = _number(row.get("current_14d_total_orders"))
    if promoted_orders_14d > total_orders_14d:
        return fail(f"{marketplace} review row {idx} {target} effective outcome has 14-day promoted SKU orders above total orders")
    available_stock = _number(row.get("current_7d_available_stock"))
    if available_stock is None or available_stock <= 0:
        return fail(f"{marketplace} review row {idx} {target} effective outcome missing positive available stock")
    target_ratio = _review_target_ratio(row)
    acos = _ratio_number(row.get("current_7d_acos"))
    if acos is not None and acos > target_ratio:
        return fail(f"{marketplace} review row {idx} {target} effective outcome has ACOS above target")
    tacos = _ratio_number(row.get("current_7d_tacos"))
    if tacos is not None and tacos > target_ratio:
        return fail(f"{marketplace} review row {idx} {target} effective outcome has TACOS above target")
    return 0


def _validate_keyword_action_review_rows(marketplace: str, rows: list[dict]) -> int:
    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            return fail(f"{marketplace} keyword action review row {idx} must be an object")
        missing = [field for field in KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS if field not in row]
        if missing:
            target = row.get("search_term_or_target") or row.get("asin") or "unknown_target"
            return fail(
                f"{marketplace} keyword action review row {idx} {target} missing fields: {', '.join(missing)}"
            )
        target = str(row.get("search_term_or_target") or row.get("asin") or "unknown_target")
        row_marketplace = str(row.get("marketplace") or "").strip().upper()
        if not row_marketplace:
            return fail(f"{marketplace} keyword action review row {idx} {target} missing marketplace value")
        if row_marketplace != marketplace:
            return fail(
                f"{marketplace} keyword action review row {idx} {target} contains {row_marketplace} marketplace data"
            )
        if not str(row.get("action_id") or "").strip():
            return fail(f"{marketplace} keyword action review row {idx} {target} missing action_id value")
        code = _validate_review_attribution(marketplace, row, idx, target)
        if code != 0:
            return code
    return 0


def _validate_action_review_rows(marketplace: str, rows: list[dict]) -> int:
    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            return fail(f"{marketplace} action review row {idx} must be an object")
        missing = [field for field in ACTION_REVIEW_REQUIRED_FIELDS if field not in row]
        if missing:
            target = row.get("product_name") or row.get("asin") or "unknown_product"
            return fail(f"{marketplace} action review row {idx} {target} missing fields: {', '.join(missing)}")
        target = str(row.get("product_name") or row.get("asin") or "unknown_product")
        row_marketplace = str(row.get("marketplace") or "").strip().upper()
        if not row_marketplace:
            return fail(f"{marketplace} action review row {idx} {target} missing marketplace value")
        if row_marketplace != marketplace:
            return fail(f"{marketplace} action review row {idx} {target} contains {row_marketplace} marketplace data")
        if not str(row.get("action_id") or "").strip():
            return fail(f"{marketplace} action review row {idx} {target} missing action_id value")
        code = _validate_review_attribution(marketplace, row, idx, target)
        if code != 0:
            return code
    return 0


def _validate_workbook_required_sheet(
    workbook,
    sheet_name: str,
    required_fields: list[str],
    *,
    expected_rows: int,
) -> int:
    if sheet_name not in workbook.sheetnames:
        return fail(f"Excel workbook missing {sheet_name} sheet")
    sheet = workbook[sheet_name]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    missing = [field for field in required_fields if field not in headers]
    if missing:
        return fail(f"Excel sheet {sheet_name} missing fields: {', '.join(missing)}")
    rows = [
        values
        for values in sheet.iter_rows(min_row=2, values_only=True)
        if any(value not in (None, "") for value in values)
    ]
    if len(rows) != expected_rows:
        return fail(f"Excel sheet {sheet_name} has {len(rows)} data rows, expected exactly {expected_rows}")
    return 0


def _workbook_sheet_rows(workbook, sheet_name: str) -> list[dict[str, object]]:
    sheet = workbook[sheet_name]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def _workbook_sheet_headers(workbook, sheet_name: str) -> list[str]:
    sheet = workbook[sheet_name]
    return [str(cell.value or "").strip() for cell in sheet[1] if str(cell.value or "").strip()]


def _expected_workbook_fields(
    expected_rows: list[dict[str, object]],
    identity_fields: list[str],
    content_fields: list[str],
) -> list[str]:
    fields: list[str] = []
    seen: set[str] = set()
    for field in [*identity_fields, *content_fields]:
        if field in seen:
            continue
        if any(field in row for row in expected_rows):
            fields.append(field)
            seen.add(field)
    return fields


def _identity_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    text = str(value).strip()
    if text.endswith(".0"):
        number = _number(text)
        if number is not None and number.is_integer():
            return str(int(number))
    return text


def _identity_tuple(row: dict[str, object], fields: list[str]) -> tuple[str, ...]:
    return tuple(_identity_value(row.get(field)) for field in fields)


def _canonical_report_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        if math.isnan(float(value)):
            return ""
        return round(float(value), 6)
    if isinstance(value, dict):
        return {str(key): _canonical_report_value(item) for key, item in sorted(value.items())}
    if isinstance(value, (list, tuple, set)):
        return [_canonical_report_value(item) for item in value]
    text = str(value).strip()
    if not text:
        return ""
    if text.lower() in {"true", "false"}:
        return text.lower() == "true"
    try:
        numeric = float(text)
        if math.isnan(numeric):
            return ""
        return round(numeric, 6)
    except ValueError:
        pass
    if text.startswith(("[", "{", "(")) and text.endswith(("]", "}", ")")):
        for parser in (json.loads, ast.literal_eval):
            try:
                parsed = parser(text)
            except (ValueError, SyntaxError, TypeError, json.JSONDecodeError):
                continue
            return _canonical_report_value(parsed)
    return text


def _format_identity(identity: tuple[str, ...], fields: list[str]) -> str:
    parts = [
        f"{field}={value or '<blank>'}"
        for field, value in zip(fields, identity)
    ]
    return "{" + ", ".join(parts) + "}"


def _validate_workbook_identity(
    workbook,
    sheet_name: str,
    expected_rows: list[dict[str, object]],
    identity_fields: list[str],
    label: str,
) -> int:
    expected = Counter(_identity_tuple(row, identity_fields) for row in expected_rows)
    actual = Counter(_identity_tuple(row, identity_fields) for row in _workbook_sheet_rows(workbook, sheet_name))
    if expected == actual:
        return 0
    missing = list((expected - actual).elements())[:3]
    unexpected = list((actual - expected).elements())[:3]
    details = []
    if missing:
        details.append("missing " + "; ".join(_format_identity(item, identity_fields) for item in missing))
    if unexpected:
        details.append("unexpected " + "; ".join(_format_identity(item, identity_fields) for item in unexpected))
    return fail(f"Excel sheet {sheet_name} row identity mismatch for {label}: " + " | ".join(details))


def _rows_by_unique_identity(
    rows: list[dict[str, object]],
    identity_fields: list[str],
    label: str,
    source: str,
) -> tuple[dict[tuple[str, ...], dict[str, object]], str]:
    by_identity: dict[tuple[str, ...], dict[str, object]] = {}
    counts = Counter(_identity_tuple(row, identity_fields) for row in rows)
    duplicate = next((identity for identity, count in counts.items() if count > 1), None)
    if duplicate is not None:
        return {}, (
            f"{source} duplicate identity for {label}: "
            f"{_format_identity(duplicate, identity_fields)}"
        )
    for row in rows:
        by_identity[_identity_tuple(row, identity_fields)] = row
    return by_identity, ""


def _validate_workbook_content(
    workbook,
    sheet_name: str,
    expected_rows: list[dict[str, object]],
    identity_fields: list[str],
    content_fields: list[str],
    label: str,
) -> int:
    headers = set(_workbook_sheet_headers(workbook, sheet_name))
    missing = [
        field
        for field in _expected_workbook_fields(expected_rows, identity_fields, content_fields)
        if field not in headers
    ]
    if missing:
        return fail(f"Excel sheet {sheet_name} missing fields for {label}: {', '.join(missing)}")
    actual_rows = _workbook_sheet_rows(workbook, sheet_name)
    expected_by_identity, expected_duplicate = _rows_by_unique_identity(
        expected_rows,
        identity_fields,
        label,
        "latest_analysis.json",
    )
    if expected_duplicate:
        return fail(expected_duplicate)
    actual_by_identity, actual_duplicate = _rows_by_unique_identity(
        actual_rows,
        identity_fields,
        label,
        f"Excel sheet {sheet_name}",
    )
    if actual_duplicate:
        return fail(actual_duplicate)
    for identity, expected_row in sorted(expected_by_identity.items()):
        actual_row = actual_by_identity.get(identity)
        if actual_row is None:
            continue
        for field in content_fields:
            if field not in expected_row:
                continue
            expected_value = _canonical_report_value(expected_row.get(field))
            actual_value = _canonical_report_value(actual_row.get(field))
            if actual_value != expected_value:
                return fail(
                    f"Excel sheet {sheet_name} field mismatch for {label} "
                    f"{_format_identity(identity, identity_fields)} field {field}: "
                    f"expected {expected_value!r}, got {actual_value!r}"
                )
    return 0


def _validate_workbook_marketplace_scope(workbook, sheet_name: str, expected_marketplace: str) -> int:
    for idx, row in enumerate(_workbook_sheet_rows(workbook, sheet_name), start=1):
        target = str(row.get("product_name") or row.get("asin") or row.get("sku") or "unknown_row")
        row_marketplace = str(row.get("marketplace") or "").strip().upper()
        if not row_marketplace:
            return fail(f"Excel sheet {sheet_name} row {idx} {target} missing marketplace value")
        if row_marketplace != expected_marketplace:
            return fail(f"Excel sheet {sheet_name} row {idx} {target} contains {row_marketplace} marketplace data")
    return 0


def _validate_top_level_snapshot_union(
    data: dict,
    top_level_key: str,
    identity_fields: list[str],
    label: str,
    content_fields: list[str] | None = None,
) -> int:
    top_level_rows = data.get(top_level_key)
    if not isinstance(top_level_rows, list):
        return fail(f"latest_analysis.json top-level {top_level_key} must be a list")

    snapshot_rows: list[dict[str, object]] = []
    marketplace_results = [
        result for result in data.get("marketplace_results") or [] if isinstance(result, dict)
    ]
    marketplace_results = sorted(marketplace_results, key=lambda result: daily_marketplace_sort_key(result.get("marketplace")))
    for result in marketplace_results:
        snapshot = result.get("report_view_snapshot") or {}
        if not isinstance(snapshot, dict):
            continue
        snapshot_rows.extend(row for row in snapshot.get(top_level_key) or [] if isinstance(row, dict))

    actual_rows = [row for row in top_level_rows if isinstance(row, dict)]
    if len(actual_rows) != len(top_level_rows):
        return fail(f"latest_analysis.json top-level {top_level_key} contains non-object rows")
    if len(actual_rows) != len(snapshot_rows):
        return fail(
            f"latest_analysis.json top-level {top_level_key} count mismatch for {label}: "
            f"expected {len(snapshot_rows)} from marketplace snapshots, got {len(actual_rows)}"
        )

    expected = Counter(_identity_tuple(row, identity_fields) for row in snapshot_rows)
    actual = Counter(_identity_tuple(row, identity_fields) for row in actual_rows)
    if expected == actual:
        duplicate_identity = next((identity for identity, count in expected.items() if count > 1), None)
        if duplicate_identity is not None:
            return fail(
                f"latest_analysis.json top-level {top_level_key} duplicate identity for {label}: "
                f"{_format_identity(duplicate_identity, identity_fields)}"
            )
        for identity in sorted(expected):
            expected_row = next(row for row in snapshot_rows if _identity_tuple(row, identity_fields) == identity)
            actual_row = next(row for row in actual_rows if _identity_tuple(row, identity_fields) == identity)
            for field in content_fields or []:
                if field not in expected_row:
                    continue
                expected_value = _canonical_report_value(expected_row.get(field))
                actual_value = _canonical_report_value(actual_row.get(field))
                if actual_value != expected_value:
                    return fail(
                        f"latest_analysis.json top-level {top_level_key} field mismatch for {label} "
                        f"{_format_identity(identity, identity_fields)} field {field}: "
                        f"expected {expected_value!r} from marketplace snapshots, got {actual_value!r}"
                    )
        return 0
    missing = list((expected - actual).elements())[:3]
    unexpected = list((actual - expected).elements())[:3]
    details = []
    if missing:
        details.append("missing " + "; ".join(_format_identity(item, identity_fields) for item in missing))
    if unexpected:
        details.append("unexpected " + "; ".join(_format_identity(item, identity_fields) for item in unexpected))
    return fail(
        f"latest_analysis.json top-level {top_level_key} identity mismatch for {label}: "
        + " | ".join(details)
    )


def _validate_top_level_marketplace_snapshot_mapping(
    data: dict,
    top_level_key: str,
    label: str,
) -> int:
    top_level_value = data.get(top_level_key)
    if not isinstance(top_level_value, dict):
        return fail(f"latest_analysis.json top-level {top_level_key} must be a dict")

    expected_marketplaces = {
        str(result.get("marketplace") or "").upper()
        for result in data.get("marketplace_results") or []
        if isinstance(result, dict) and result.get("marketplace")
    }
    actual_marketplaces = {str(key).upper() for key in top_level_value.keys()}
    if actual_marketplaces != expected_marketplaces:
        return fail(
            f"latest_analysis.json top-level {top_level_key} marketplace mismatch for {label}: "
            f"expected {sorted(expected_marketplaces)}, got {sorted(actual_marketplaces)}"
        )

    for result in data.get("marketplace_results") or []:
        if not isinstance(result, dict):
            continue
        marketplace = str(result.get("marketplace") or "").upper()
        snapshot = result.get("report_view_snapshot") or {}
        if not isinstance(snapshot, dict):
            continue
        snapshot_value = snapshot.get(top_level_key)
        top_value = top_level_value.get(marketplace)
        if top_value != snapshot_value:
            return fail(
                f"latest_analysis.json top-level {top_level_key} mismatch for {marketplace} {label}: "
                f"expected snapshot value {snapshot_value}, got {top_value}"
            )
    return 0


def _validate_decision_count_summary(marketplace: str, snapshot: dict, summary_key: str) -> int:
    rows = snapshot.get("product_final_decision_rows") or []
    if not isinstance(rows, list):
        return fail(f"{marketplace} product_final_decision_rows must be a list")
    expected = dict(Counter(str(row.get("final_decision") or "").strip() for row in rows if isinstance(row, dict)))
    actual = snapshot.get(summary_key)
    if not isinstance(actual, dict):
        return fail(f"{marketplace} {summary_key} must be a dict")
    if actual != expected:
        return fail(
            f"{marketplace} {summary_key} does not match product_final_decision_rows final_decision counts: "
            f"expected {expected}, got {actual}"
        )
    return 0


def _validate_workbook_product_decision_sheet(
    workbook,
    sheet_name: str,
    *,
    expected_rows: int,
) -> int:
    code = _validate_workbook_required_sheet(
        workbook,
        sheet_name,
        PRODUCT_FINAL_DECISION_REQUIRED_FIELDS,
        expected_rows=expected_rows,
    )
    if code != 0:
        return code
    return _validate_product_final_decision_rows("Excel", _workbook_sheet_rows(workbook, sheet_name))


def _validate_workbook_review_sheet(
    workbook,
    sheet_name: str,
    required_fields: list[str],
    *,
    expected_rows: int,
    expected_marketplace: str | None = None,
) -> int:
    code = _validate_workbook_required_sheet(
        workbook,
        sheet_name,
        required_fields,
        expected_rows=expected_rows,
    )
    if code != 0:
        return code
    for idx, row in enumerate(_workbook_sheet_rows(workbook, sheet_name), start=1):
        target = str(
            row.get("search_term_or_target")
            or row.get("product_name")
            or row.get("asin")
            or "unknown_review_row"
        )
        row_marketplace = str(row.get("marketplace") or "").strip().upper()
        if expected_marketplace:
            if not row_marketplace:
                return fail(f"Excel sheet {sheet_name} row {idx} {target} missing marketplace value")
            if row_marketplace != expected_marketplace:
                return fail(f"Excel sheet {sheet_name} row {idx} {target} contains {row_marketplace} marketplace data")
        code = _validate_review_attribution("Excel", row, idx, target)
        if code != 0:
            return code
    return 0


def _ad_action_label_for_validation(row: dict) -> str:
    fallback = ""
    for field in ["suggested_action", "scale_action", "copy_action_line", "copy_block", "today_action"]:
        action = str(row.get(field) or "").strip()
        if not action:
            continue
        if "否" in action and "不直接否" not in action:
            return "否定精准"
        if "暂停" in action and "ASIN" in action.upper():
            return "暂停 ASIN 定向"
        if "降竞价" in action or "降价竞价" in action:
            return "降竞价"
        if "降价" in action:
            return "降价"
        if (
            "growth_test" in action
            or "小预算" in action
            or "新建精准" in action
            or "拉精准" in action
            or "开精准" in action
        ):
            return "小预算试投"
        blocks_bid_up = any(marker in action for marker in ["不加价", "不提高竞价", "不能加价", "禁止加价"])
        if ("加价" in action or "提高竞价" in action) and not blocks_bid_up:
            return "加价"
        if "保留" in action:
            fallback = fallback or "保留观察"
        elif "观察" in action:
            fallback = fallback or "观察"
    return fallback or "观察"


def _is_pending_ad_workbench_row(row: dict) -> bool:
    if str(row.get("confirmed_status") or "") in {"已执行", "已核查", "已忽略", "仅背景参考"}:
        return False
    return _ad_action_label_for_validation(row) not in {"观察", "保留观察"}


def _ad_workbench_rows_from_snapshot(snapshot: dict) -> list[dict]:
    rows: list[dict] = []
    for key in ["html_search_term_processing_queue_rows", "scale_keyword_rows", "growth_test_rows"]:
        value = snapshot.get(key) or []
        if isinstance(value, list):
            rows.extend(row for row in value if isinstance(row, dict))
    groups = snapshot.get("today_action_groups") or {}
    if isinstance(groups, dict):
        group_rows = groups.get("广告动作") or []
        if isinstance(group_rows, list):
            rows.extend(row for row in group_rows if isinstance(row, dict))
    return rows


def _safe_run_has_pending_ad_workbench_rows(data: dict) -> bool:
    results = data.get("marketplace_results") or []
    if not isinstance(results, list):
        return False
    for result in results:
        if not isinstance(result, dict):
            continue
        snapshot = result.get("report_view_snapshot") or {}
        if not isinstance(snapshot, dict):
            continue
        if any(_is_pending_ad_workbench_row(row) for row in _ad_workbench_rows_from_snapshot(snapshot)):
            return True
    return False


def _pending_ad_identity_token(row: dict) -> tuple[str, str] | None:
    target = str(row.get("search_term_or_target") or "").strip()
    if target:
        return ("target", target)
    asin = str(row.get("asin") or "").strip().upper()
    if asin:
        return ("ASIN", asin)
    product_name = str(row.get("product_name") or "").strip()
    if product_name:
        return ("product name", product_name)
    return None


def _tomorrow_review_identity_token(row: dict) -> tuple[str, str] | None:
    target = str(row.get("search_term_or_target") or "").strip()
    if target:
        return ("target", target)
    asin = str(row.get("asin") or "").strip().upper()
    if asin:
        return ("ASIN", asin)
    product_name = str(row.get("product_name") or "").strip()
    if product_name:
        return ("product name", product_name)
    return None


def _validate_latest_recommendations_ad_area(text: str, *, has_pending_ad_rows: bool) -> int:
    if AD_COPY_ANCHOR not in text:
        return fail("latest_recommendations.html missing today-ad-actions-all anchor")
    if has_pending_ad_rows:
        if "复制到广告后台" not in text:
            return fail("latest_recommendations.html missing copy area for pending ad rows")
    else:
        has_empty_copy_state = "无待确认动作" in text or "当前没有需要复制执行的广告动作" in text
        has_zero_pending_status = "待确认 0" in text or "待处理</span><strong>0</strong>" in text
        if not (has_empty_copy_state or has_zero_pending_status):
            return fail("latest_recommendations.html missing zero-pending ad copy status")
    return 0


def _html_contains_token(text: str, token: object) -> bool:
    value = str(token or "").strip()
    return not value or value in text or html.escape(value) in text


def _html_contains_any_token(text: str, tokens: list[str]) -> bool:
    return any(_html_contains_token(text, token) for token in tokens)


def _product_decision_blocking_reason_tokens(row: dict[str, object]) -> list[str]:
    reasons = row.get("frontend_blocking_reasons")
    if isinstance(reasons, list):
        return [str(reason).strip() for reason in reasons if str(reason).strip()]
    reason_text = str(reasons or "").strip()
    if not reason_text:
        return []
    return [token.strip() for token in re.split(r"[；;]\s*", reason_text) if token.strip()]


def _review_display_judgement_for_sort(row: dict[str, object]) -> str:
    return _html_review_display_judgement(row)


def _action_review_sort_key(row: dict[str, object]) -> tuple[int, str, str]:
    return _html_action_effect_review_sort_key(row)


def _keyword_review_sort_key(row: dict[str, object]) -> tuple[int, int, str, str]:
    return _html_keyword_review_sort_key(row)


def _keyword_review_html_title_anchor(row: dict[str, object], target: str, asin: str) -> str:
    marketplace = str(row.get("marketplace") or "").strip().upper()
    if marketplace and target:
        return f"<strong>{html.escape(f'{marketplace}｜{target}')}</strong>"
    return target or asin


def _keyword_review_forbidden_metric_visible(
    text: str,
    row: dict[str, object],
    target: str,
    asin: str,
    token: str,
) -> bool:
    title_anchor = _keyword_review_html_title_anchor(row, target, asin)
    if title_anchor.startswith("<strong>") and _html_contains_token(text, title_anchor):
        return daily_html_contains_token_near(text, title_anchor, token)
    fallback_anchor = target or asin
    if fallback_anchor:
        return daily_html_contains_token_near(text, fallback_anchor, token)
    return _html_contains_token(text, token)


def _review_html_metric_tokens(row: dict[str, object]) -> list[str]:
    days_since = _number(row.get("days_since_execution"))
    if days_since is not None and days_since < 3:
        return []
    tokens: list[str] = []
    if days_since is not None and 3 <= days_since < 7:
        tokens.append("3 天窗口只做初步判断")
    field_labels = [
        ("current_7d_promoted_ad_orders", "本 SKU 单"),
        ("current_7d_acos", "ACOS"),
        ("current_7d_target_acos", "目标 ACOS"),
        ("current_7d_tacos", "TACOS"),
        ("current_7d_total_orders", "总单"),
        ("current_7d_available_stock", "库存"),
    ]
    if days_since is not None and days_since >= 7:
        field_labels.extend(
            [
                ("current_14d_promoted_ad_orders", "14天 本 SKU 单"),
                ("current_14d_acos", "14天 ACOS"),
                ("current_14d_tacos", "14天 TACOS"),
                ("current_14d_total_orders", "14天 总单"),
                ("current_14d_available_stock", "14天 库存"),
            ]
        )
    for field, label in field_labels:
        value = str(row.get(field) or "").strip()
        number = _number(value)
        if value and (number is not None or value.lower() not in {"nan", "none", "null"}):
            tokens.append(f"{label} {value}")
    if str(row.get("review_data_source") or "").strip() == "execution_anchored_daily":
        anchor_fields = [
            ("pre_7d_promoted_ad_orders", "执行前本 SKU 单"),
            ("pre_7d_total_orders", "执行前总单"),
            ("pre_7d_tacos", "执行前 TACOS"),
            ("post_3d_days", "执行后3天覆盖天数"),
            ("post_3d_promoted_ad_orders", "执行后3天本 SKU 单"),
            ("post_3d_total_orders", "执行后3天总单"),
            ("post_3d_acos", "执行后3天 ACOS"),
            ("post_3d_tacos", "执行后3天 TACOS"),
            ("post_3d_available_stock", "执行后3天库存"),
        ]
        if days_since is None or days_since >= 7:
            anchor_fields.extend(
                [
                    ("post_7d_days", "执行后覆盖天数"),
                    ("post_7d_promoted_ad_orders", "执行后本 SKU 单"),
                    ("post_7d_total_orders", "执行后总单"),
                    ("post_7d_acos", "执行后 ACOS"),
                    ("post_7d_tacos", "执行后 TACOS"),
                    ("post_7d_available_stock", "执行后库存"),
                ]
            )
        pre_start = str(row.get("pre_7d_start") or "").strip()
        pre_end = str(row.get("pre_7d_end") or "").strip()
        post_3d_start = str(row.get("post_3d_start") or "").strip()
        post_3d_end = str(row.get("post_3d_end") or "").strip()
        post_start = str(row.get("post_7d_start") or "").strip()
        post_end = str(row.get("post_7d_end") or "").strip()
        if pre_start and pre_end:
            tokens.append(f"执行前7天 {pre_start} 至 {pre_end}")
        if post_3d_start and post_3d_end:
            tokens.append(f"执行后3天 {post_3d_start} 至 {post_3d_end}")
        if (days_since is None or days_since >= 7) and post_start and post_end:
            tokens.append(f"执行后7天 {post_start} 至 {post_end}")
        for field, label in anchor_fields:
            value = str(row.get(field) or "").strip()
            number = _number(value)
            if value and (number is not None or value.lower() not in {"nan", "none", "null"}):
                tokens.append(f"{label} {value}")
    return tokens


def _format_summary_metric_number(value: object) -> str:
    text = str(value or "").strip()
    if text.lower() in {"", "nan", "none", "null"}:
        return ""
    number = _number(text)
    if number is None:
        return text
    if number.is_integer():
        return str(int(number))
    return f"{number:.2f}".rstrip("0").rstrip(".")


def _format_summary_metric_percent(value: object) -> str:
    text = str(value or "").strip()
    if text.lower() in {"", "nan", "none", "null"}:
        return ""
    number = _number(text)
    if number is None:
        return text
    percent = number * 100 if abs(number) <= 1 else number
    if percent.is_integer():
        return f"{int(percent)}%"
    return f"{percent:.1f}".rstrip("0").rstrip(".") + "%"


def _summary_review_metric_tokens(row: dict[str, object]) -> list[str]:
    days_since = _number(row.get("days_since_execution"))
    if days_since is not None and days_since < 3:
        return []
    tokens: list[str] = []
    field_labels = [
        ("current_7d_promoted_ad_orders", "本 SKU 单", _format_summary_metric_number),
        ("current_7d_acos", "ACOS", _format_summary_metric_percent),
        ("current_7d_target_acos", "目标 ACOS", _format_summary_metric_percent),
        ("current_7d_tacos", "TACOS", _format_summary_metric_percent),
        ("current_7d_total_orders", "总单", _format_summary_metric_number),
        ("current_7d_available_stock", "库存", _format_summary_metric_number),
    ]
    if days_since is None or days_since >= 7:
        field_labels.extend(
            [
                ("current_14d_promoted_ad_orders", "14天本 SKU 单", _format_summary_metric_number),
                ("current_14d_acos", "14天 ACOS", _format_summary_metric_percent),
                ("current_14d_tacos", "14天 TACOS", _format_summary_metric_percent),
                ("current_14d_total_orders", "14天总单", _format_summary_metric_number),
                ("current_14d_available_stock", "14天库存", _format_summary_metric_number),
            ]
        )
    for field, label, formatter in field_labels:
        value = formatter(row.get(field))
        if value:
            tokens.append(f"{label} {value}")
    return tokens


def _summary_review_watch_sort_key(row: dict[str, object]) -> tuple[int, int, int, str]:
    judgement = daily_summary_review_display_judgement(row)
    next_step = daily_summary_review_display_next_step(row)
    window = str(row.get("review_window") or "")
    product = str(row.get("product_name") or row.get("search_term_or_target") or "")
    score = 0
    if _truthy(row.get("halo_only_conversion")) or _truthy(row.get("target_sku_not_converted")) or judgement in {
        "本 SKU 未验证",
        "待人工复查",
    }:
        score += 8
    if judgement in {"暂未改善", "初步有效", "有改善迹象"}:
        score += 5
    if judgement in {"待7天确认", "待人工判定有效/无效"}:
        score += 4
    if (
        "不要继续加价" in next_step
        or "不追加预算" in next_step
        or "不追加竞价" in next_step
        or "回到原竞价" in next_step
        or "保留当前竞价" in next_step
    ):
        score += 4
    if "优先要求补竞品/页面证据" in next_step or "确认是否否词匹配类型" in next_step:
        score += 3
    if window == "3天后复盘":
        score += 2
    days_number = _number(row.get("days_since_execution"))
    days = int(days_number) if days_number is not None else 0
    has_target_rank = 0 if str(row.get("search_term_or_target") or "") else 1
    return (-score, -days, has_target_rank, product)


def _summary_inventory_sort_key(row: dict[str, object]) -> tuple[int, float, float, str]:
    status = str(row.get("stock_status_label") or "")
    level = str(row.get("stock_risk_level") or "")
    coverage_source = row.get("days_of_cover")
    if coverage_source in ("", None):
        coverage_source = row.get("coverage_days")
    qty_source = row.get("recommended_reorder_qty")
    if qty_source in ("", None):
        qty_source = row.get("recommended_replenishment_qty")
    coverage = _number(coverage_source)
    qty = _number(qty_source) or 0
    if level == "OUT_OF_STOCK":
        priority = 0
    elif level == "LOW_STOCK" or "低库存" in status:
        priority = 1
    elif level == "REPLENISH_SOON" or "进入补货窗口" in status:
        priority = 2
    else:
        priority = 9
    return (priority, coverage if coverage is not None else 99999, -qty, str(row.get("product_name") or ""))


def _is_summary_inventory_row(row: dict[str, object]) -> bool:
    level = str(row.get("stock_risk_level") or "")
    status = str(row.get("stock_status_label") or "")
    if level in {"OUT_OF_STOCK", "LOW_STOCK", "REPLENISH_SOON"}:
        return True
    return not level and any(token in status for token in ["断货", "低库存", "进入补货窗口"])


def _validate_summary_snapshot_rows(text: str, data: dict) -> int:
    action_review_rows: list[dict[str, object]] = []
    keyword_review_rows: list[dict[str, object]] = []
    inventory_rows: list[dict[str, object]] = []
    marketplace_results = [
        result for result in data.get("marketplace_results") or [] if isinstance(result, dict)
    ]
    marketplace_results = sorted(marketplace_results, key=lambda result: daily_marketplace_sort_key(result.get("marketplace")))
    for result in marketplace_results:
        snapshot = result.get("report_view_snapshot") or {}
        if not isinstance(snapshot, dict):
            continue
        action_review_rows.extend(row for row in snapshot.get("action_effect_review_rows") or [] if isinstance(row, dict))
        keyword_review_rows.extend(row for row in snapshot.get("keyword_action_effect_review_rows") or [] if isinstance(row, dict))
        inventory_rows.extend(row for row in snapshot.get("inventory_replenishment_rows") or [] if isinstance(row, dict))

    for token in daily_frontend_coverage_display_tokens(data):
        if not _html_contains_token(text, token):
            return fail(f"summary.html missing frontend coverage token {token}")

    review_candidates = sorted([*keyword_review_rows, *action_review_rows], key=_summary_review_watch_sort_key)
    for idx, row in enumerate(review_candidates[:3], start=1):
        target = str(row.get("search_term_or_target") or row.get("product_name") or "").strip()
        if target and not _html_contains_token(text, target):
            return fail(f"summary.html missing review watch row {idx} target {target}")
        if daily_summary_review_needs_display_guard(row):
            display_judgement = daily_summary_review_display_judgement(row)
            display_next_step = daily_summary_review_display_next_step(row)
            if target:
                if not daily_html_contains_token_near(text, target, display_judgement):
                    return fail(f"summary.html missing review watch row {idx} judgement {display_judgement}")
                if display_next_step and not daily_html_contains_token_near(text, target, display_next_step):
                    return fail(f"summary.html missing review watch row {idx} next step {display_next_step}")
            else:
                if display_judgement and not _html_contains_token(text, display_judgement):
                    return fail(f"summary.html missing review watch row {idx} judgement {display_judgement}")
                if display_next_step and not _html_contains_token(text, display_next_step):
                    return fail(f"summary.html missing review watch row {idx} next step {display_next_step}")
        for token in _summary_review_metric_tokens(row):
            if not _html_contains_token(text, token):
                return fail(f"summary.html missing review watch row {idx} metric token {token}")
        for token in daily_summary_review_forbidden_early_metric_tokens(row):
            if target:
                if daily_html_contains_token_near(text, target, token):
                    return fail(f"summary.html early review row {idx} should not show metric token {token}")
            elif _html_contains_token(text, token):
                return fail(f"summary.html early review row {idx} should not show metric token {token}")

    replenishment_rows = [
        row
        for row in sorted(inventory_rows, key=_summary_inventory_sort_key)
        if "暂不需要" not in str(row.get("replenishment_advice") or "")
        and _is_summary_inventory_row(row)
    ]
    for idx, row in enumerate(replenishment_rows[:3], start=1):
        product_name = str(row.get("product_name") or "").strip()
        status = str(row.get("stock_status_label") or "").strip()
        if product_name and not _html_contains_token(text, product_name):
            return fail(f"summary.html missing replenishment row {idx} product name {product_name}")
        if status and not _html_contains_token(text, status):
            return fail(f"summary.html missing replenishment row {idx} status {status}")
    return 0


def _validate_latest_recommendations_snapshot_rows(text: str, data: dict) -> int:
    product_rows: list[dict[str, object]] = []
    action_review_rows: list[dict[str, object]] = []
    keyword_review_rows: list[dict[str, object]] = []
    pending_ad_rows: list[dict[str, object]] = []
    tomorrow_review_rows: list[dict[str, object]] = []
    marketplace_results = [
        result for result in data.get("marketplace_results") or [] if isinstance(result, dict)
    ]
    marketplace_results = sorted(marketplace_results, key=lambda result: daily_marketplace_sort_key(result.get("marketplace")))
    for result in marketplace_results:
        snapshot = result.get("report_view_snapshot") or {}
        if not isinstance(snapshot, dict):
            continue
        product_rows.extend(row for row in snapshot.get("product_final_decision_rows") or [] if isinstance(row, dict))
        action_review_rows.extend(row for row in snapshot.get("action_effect_review_rows") or [] if isinstance(row, dict))
        keyword_review_rows.extend(row for row in snapshot.get("keyword_action_effect_review_rows") or [] if isinstance(row, dict))
        tomorrow_review_rows.extend(row for row in snapshot.get("tomorrow_review_rows") or [] if isinstance(row, dict))
        pending_ad_rows.extend(
            row
            for row in _ad_workbench_rows_from_snapshot(snapshot)
            if _is_pending_ad_workbench_row(row)
        )

    for token in daily_frontend_coverage_display_tokens(data):
        if not _html_contains_token(text, token):
            return fail(f"latest_recommendations.html missing frontend coverage token {token}")

    for idx, row in enumerate(product_rows[:6], start=1):
        marketplace = str(row.get("marketplace") or "").strip().upper()
        sku = str(row.get("sku") or "").strip()
        asin = str(row.get("asin") or "").strip().upper()
        product_name = str(row.get("product_name") or "").strip()
        contexts = daily_product_decision_contexts(text, asin, marketplace, sku, container="article")
        if asin and not _html_contains_token(text, asin):
            return fail(f"latest_recommendations.html missing product decision row {idx} ASIN {asin}")
        if asin and not contexts:
            return fail(
                f"latest_recommendations.html missing structured product gate row {idx} {marketplace} {sku} {asin}"
            )
        if product_name and not _html_contains_token(text, product_name):
            return fail(f"latest_recommendations.html missing product decision row {idx} product name {product_name}")
        if product_name and contexts and not daily_html_contains_token_in_contexts(contexts, product_name):
            return fail(f"latest_recommendations.html product gate row {idx} product name not bound to ASIN {asin}")
        for reason in _product_decision_blocking_reason_tokens(row)[:3]:
            if not _html_contains_token(text, reason):
                return fail(
                    f"latest_recommendations.html missing product decision row {idx} blocking reason {reason}"
                )
            if contexts and not daily_html_contains_token_in_contexts(contexts, reason):
                return fail(f"latest_recommendations.html product gate row {idx} blocking reason not bound to ASIN {asin}")
        blocked_actions = _action_set(row.get("today_blocked_actions"))
        for action, labels in PRODUCT_DECISION_BLOCKED_ACTION_LABELS.items():
            if action in blocked_actions and not _html_contains_any_token(text, labels):
                return fail(
                    f"latest_recommendations.html missing product decision row {idx} blocked action label {action}"
                )
            if action in blocked_actions and contexts and not daily_html_contains_any_token_in_contexts(contexts, labels):
                return fail(
                    f"latest_recommendations.html product gate row {idx} blocked action label {action} not bound to ASIN {asin}"
                )
        for label, token in daily_product_frontend_evidence_tokens(row):
            if not _html_contains_token(text, token):
                return fail(f"latest_recommendations.html missing product gate row {idx} {label} {token}")
            if contexts and not daily_html_contains_token_in_contexts(contexts, token):
                return fail(
                    f"latest_recommendations.html product gate row {idx} {label} {token} not bound to ASIN {asin}"
                )
        attr_failures = daily_product_decision_contract_attr_failures(
            "latest_recommendations.html",
            row,
            contexts,
            row_label=f"product gate row {idx}",
        )
        if attr_failures:
            return fail(attr_failures[0])

    sorted_action_rows = sorted(action_review_rows, key=_action_review_sort_key)
    for idx, row in enumerate(sorted_action_rows[:5], start=1):
        asin = str(row.get("asin") or "").strip().upper()
        product_name = str(row.get("product_name") or "").strip()
        anchor = asin or product_name
        if asin and not _html_contains_token(text, asin):
            return fail(f"latest_recommendations.html missing action review row {idx} ASIN {asin}")
        if product_name and not _html_contains_token(text, product_name):
            return fail(f"latest_recommendations.html missing action review row {idx} product name {product_name}")
        for token in _review_html_metric_tokens(row):
            if not _html_contains_token(text, token):
                return fail(f"latest_recommendations.html missing action review row {idx} metric token {token}")
        for token in daily_review_html_forbidden_metric_tokens(row):
            if anchor and daily_html_contains_token_near(text, anchor, token):
                return fail(f"latest_recommendations.html action review row {idx} should not show metric token {token}")
            if not anchor and _html_contains_token(text, token):
                return fail(f"latest_recommendations.html action review row {idx} should not show metric token {token}")

    for idx, row in enumerate(sorted(keyword_review_rows, key=_keyword_review_sort_key)[:50], start=1):
        target = str(row.get("search_term_or_target") or "").strip()
        asin = str(row.get("asin") or "").strip().upper()
        if target and not _html_contains_token(text, target):
            return fail(f"latest_recommendations.html missing keyword review row {idx} target {target}")
        if asin and not _html_contains_token(text, asin):
            return fail(f"latest_recommendations.html missing keyword review row {idx} ASIN {asin}")
        for token in _review_html_metric_tokens(row):
            if not _html_contains_token(text, token):
                return fail(f"latest_recommendations.html missing keyword review row {idx} metric token {token}")
        for token in daily_review_html_forbidden_metric_tokens(row):
            if _keyword_review_forbidden_metric_visible(text, row, target, asin, token):
                return fail(f"latest_recommendations.html keyword review row {idx} should not show metric token {token}")

    for idx, row in enumerate(tomorrow_review_rows[:50], start=1):
        identity = _tomorrow_review_identity_token(row)
        if identity is None:
            continue
        label, token = identity
        if not _html_contains_token(text, token):
            return fail(f"latest_recommendations.html missing tomorrow review row {idx} {label} {token}")

    for idx, row in enumerate(pending_ad_rows[:50], start=1):
        identity = _pending_ad_identity_token(row)
        if identity is None:
            continue
        label, token = identity
        if not _html_contains_token(text, token):
            return fail(f"latest_recommendations.html missing pending ad row {idx} {label} {token}")
    return 0


def _validate_marketplace_report_snapshot_rows(page_name: str, text: str, result: dict) -> int:
    marketplace = str(result.get("marketplace") or "").upper()
    snapshot = result.get("report_view_snapshot") or {}
    if not isinstance(snapshot, dict):
        return 0
    action_rows = [
        row
        for row in snapshot.get("action_effect_review_rows") or []
        if isinstance(row, dict) and str(row.get("marketplace") or "").upper() == marketplace
    ]
    product_rows = [
        row
        for row in snapshot.get("product_final_decision_rows") or []
        if isinstance(row, dict) and str(row.get("marketplace") or "").upper() == marketplace
    ]
    keyword_rows = [
        row
        for row in snapshot.get("keyword_action_effect_review_rows") or []
        if isinstance(row, dict) and str(row.get("marketplace") or "").upper() == marketplace
    ]

    for idx, row in enumerate(product_rows[:6], start=1):
        marketplace = str(row.get("marketplace") or "").strip().upper()
        sku = str(row.get("sku") or "").strip()
        asin = str(row.get("asin") or "").strip().upper()
        product_name = str(row.get("product_name") or "").strip()
        contexts = daily_product_decision_contexts(text, asin, marketplace, sku, container="article")
        if asin and not _html_contains_token(text, asin):
            return fail(f"{page_name} missing product decision row {idx} ASIN {asin}")
        if asin and not contexts:
            return fail(f"{page_name} missing structured product decision row {idx} {marketplace} {sku} {asin}")
        if product_name and not _html_contains_token(text, product_name):
            return fail(f"{page_name} missing product decision row {idx} product name {product_name}")
        if product_name and contexts and not daily_html_contains_token_in_contexts(contexts, product_name):
            return fail(f"{page_name} product decision row {idx} product name not bound to ASIN {asin}")
        for reason in _product_decision_blocking_reason_tokens(row)[:3]:
            if not _html_contains_token(text, reason):
                return fail(f"{page_name} missing product decision row {idx} blocking reason {reason}")
            if contexts and not daily_html_contains_token_in_contexts(contexts, reason):
                return fail(f"{page_name} product decision row {idx} blocking reason not bound to ASIN {asin}")
        blocked_actions = _action_set(row.get("today_blocked_actions"))
        for action, labels in PRODUCT_DECISION_BLOCKED_ACTION_LABELS.items():
            if action in blocked_actions and not _html_contains_any_token(text, labels):
                return fail(f"{page_name} missing product decision row {idx} blocked action label {action}")
            if action in blocked_actions and contexts and not daily_html_contains_any_token_in_contexts(contexts, labels):
                return fail(f"{page_name} product decision row {idx} blocked action label {action} not bound to ASIN {asin}")
        for label, token in daily_product_frontend_evidence_tokens(row):
            if not _html_contains_token(text, token):
                return fail(f"{page_name} missing product decision row {idx} {label} {token}")
            if contexts and not daily_html_contains_token_in_contexts(contexts, token):
                return fail(f"{page_name} product decision row {idx} {label} {token} not bound to ASIN {asin}")
        attr_failures = daily_product_decision_contract_attr_failures(
            page_name,
            row,
            contexts,
            row_label=f"product decision row {idx}",
        )
        if attr_failures:
            return fail(attr_failures[0])

    for idx, row in enumerate(sorted(action_rows, key=_action_review_sort_key)[:3], start=1):
        asin = str(row.get("asin") or "").strip().upper()
        product_name = str(row.get("product_name") or "").strip()
        anchor = asin or product_name
        if asin and not _html_contains_token(text, asin):
            return fail(f"{page_name} missing action review row {idx} ASIN {asin}")
        if product_name and not _html_contains_token(text, product_name):
            return fail(f"{page_name} missing action review row {idx} product name {product_name}")
        for token in _review_html_metric_tokens(row):
            if not _html_contains_token(text, token):
                return fail(f"{page_name} missing action review row {idx} metric token {token}")
        for token in daily_review_html_forbidden_metric_tokens(row):
            if anchor and daily_html_contains_token_near(text, anchor, token):
                return fail(f"{page_name} action review row {idx} should not show metric token {token}")
            if not anchor and _html_contains_token(text, token):
                return fail(f"{page_name} action review row {idx} should not show metric token {token}")

    for idx, row in enumerate(sorted(keyword_rows, key=_keyword_review_sort_key)[:10], start=1):
        target = str(row.get("search_term_or_target") or "").strip()
        asin = str(row.get("asin") or "").strip().upper()
        if target and not _html_contains_token(text, target):
            return fail(f"{page_name} missing keyword review row {idx} target {target}")
        if asin and not _html_contains_token(text, asin):
            return fail(f"{page_name} missing keyword review row {idx} ASIN {asin}")
        for token in _review_html_metric_tokens(row):
            if not _html_contains_token(text, token):
                return fail(f"{page_name} missing keyword review row {idx} metric token {token}")
        for token in daily_review_html_forbidden_metric_tokens(row):
            if _keyword_review_forbidden_metric_visible(text, row, target, asin, token):
                return fail(f"{page_name} keyword review row {idx} should not show metric token {token}")
    return 0


def _validate_top_level_frontend_coverage_summary(data: dict) -> int:
    expected = daily_frontend_coverage_aggregate_from_snapshots(data.get("marketplace_results") or [])
    actual = data.get("frontend_coverage_summary")
    if not isinstance(actual, dict):
        return fail("latest_analysis.json top-level frontend_coverage_summary must be a dict")
    for field, expected_value in expected.items():
        actual_value = actual.get(field)
        if isinstance(expected_value, int):
            if int(_number(actual_value) or 0) != expected_value:
                return fail(
                    f"latest_analysis.json top-level frontend_coverage_summary field {field}: "
                    f"expected {expected_value}, got {actual_value}"
                )
        elif isinstance(expected_value, float):
            actual_number = _number(actual_value)
            if actual_number is None or abs(float(actual_number) - expected_value) > 1e-9:
                return fail(
                    f"latest_analysis.json top-level frontend_coverage_summary field {field}: "
                    f"expected {expected_value}, got {actual_value}"
                )
        elif actual_value != expected_value:
            return fail(
                f"latest_analysis.json top-level frontend_coverage_summary field {field}: "
                f"expected {expected_value}, got {actual_value}"
            )
    return 0


def _validate_dashboard_marketplace_summary(text: str, data: dict) -> int:
    for result in data.get("marketplace_results") or []:
        if not isinstance(result, dict):
            continue
        summary = result.get("summary") or {}
        marketplace = str(summary.get("marketplace") or result.get("marketplace") or "").upper()
        if not marketplace:
            return fail("dashboard.html marketplace summary row missing marketplace")
        report_link = f'{marketplace.lower()}_report.html'
        if report_link not in text:
            return fail(f"dashboard.html missing report link {report_link}")
        fields = [
            marketplace,
            summary.get("ads_row_count"),
            summary.get("erp_row_count"),
            summary.get("sku_count"),
            summary.get("asin_count"),
        ]
        cells = "".join(f"<td>{html.escape(str(value))}</td>" for value in fields)
        pattern = rf"<tr[^>]*>{re.escape(cells)}"
        if not re.search(pattern, text):
            return fail(
                f"dashboard.html missing marketplace summary row for {marketplace}: "
                f"ads={summary.get('ads_row_count')} erp={summary.get('erp_row_count')} "
                f"sku={summary.get('sku_count')} asin={summary.get('asin_count')}"
            )
    for token in daily_frontend_coverage_display_tokens(data, include_title=True):
        if not _html_contains_token(text, token):
            return fail(f"dashboard.html missing frontend coverage token {token}")
    return 0


def _validate_marketplace_summary_markdown(text: str, data: dict) -> int:
    for result in data.get("marketplace_results") or []:
        if not isinstance(result, dict):
            continue
        summary = result.get("summary") or {}
        marketplace = str(summary.get("marketplace") or result.get("marketplace") or "").upper()
        if not marketplace:
            return fail("marketplace_summary.md summary row missing marketplace")
        fields = [
            marketplace,
            summary.get("ads_row_count"),
            summary.get("erp_row_count"),
            summary.get("sku_count"),
            summary.get("asin_count"),
        ]
        row_pattern = r"\|\s*" + r"\s*\|\s*".join(re.escape(str(value)) for value in fields) + r"\s*\|"
        if not re.search(row_pattern, text):
            return fail(
                f"marketplace_summary.md missing marketplace summary row for {marketplace}: "
                f"ads={summary.get('ads_row_count')} erp={summary.get('erp_row_count')} "
                f"sku={summary.get('sku_count')} asin={summary.get('asin_count')}"
            )
    for token in daily_frontend_coverage_display_tokens(data, include_title=True):
        if token not in text:
            return fail(f"marketplace_summary.md missing frontend coverage token {token}")
    return 0


def _validate_workbook_frontend_coverage_summary(workbook, data: dict) -> int:
    expected = daily_frontend_coverage_excel_expected(data)
    if not expected:
        return 0
    if "总览" not in workbook.sheetnames:
        return fail("Excel workbook missing 总览 sheet for frontend coverage summary")
    actual: dict[str, str] = {}
    for row in workbook["总览"].iter_rows(values_only=True):
        values = [str(value or "").strip() for value in row]
        if len(values) < 3 or values[0] != "前台证据覆盖":
            continue
        actual[values[1]] = values[2]
    for metric, expected_value in expected.items():
        actual_value = actual.get(metric)
        if actual_value != expected_value:
            return fail(
                f"Excel sheet 总览 missing frontend coverage metric {metric}: "
                f"expected {expected_value}, got {actual_value or 'missing'}"
            )
    return 0


def run_step(args: list[str], extra_env: dict[str, str] | None = None) -> int:
    print("[run]", " ".join(args), flush=True)
    env = os.environ.copy()
    if extra_env:
        env.update(extra_env)
    completed = subprocess.run(args, cwd=ROOT, env=env)
    print("[exit]", completed.returncode, flush=True)
    return completed.returncode


def fail(message: str) -> int:
    print(f"[fail] {message}", flush=True)
    return 1


def git_text(args: list[str]) -> str:
    completed = subprocess.run(
        ["git", *args],
        cwd=ROOT,
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    return completed.stdout


def git_bytes(args: list[str]) -> bytes:
    completed = subprocess.run(
        ["git", *args],
        cwd=ROOT,
        check=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    return completed.stdout


def git_lines(args: list[str]) -> list[str]:
    return [line for line in git_text(args).splitlines() if line.strip()]


def workspace_fingerprint() -> dict[str, object]:
    head = git_text(["rev-parse", "HEAD"]).strip()
    tracked_changed = git_lines(["diff", "--name-only", "HEAD", "--"])
    untracked = git_lines(["ls-files", "--others", "--exclude-standard"])
    changed_paths = sorted(set(tracked_changed + untracked))
    status_lines = git_lines(["status", "--porcelain=v1", "--untracked-files=all"])

    digest = sha256()
    digest.update(b"git-head\0")
    digest.update(head.encode("utf-8"))
    digest.update(b"\0tracked-diff\0")
    digest.update(git_bytes(["diff", "--binary", "HEAD", "--"]))
    digest.update(b"\0untracked-files\0")
    for relative_path in sorted(untracked):
        path = ROOT / relative_path
        digest.update(relative_path.encode("utf-8", errors="surrogateescape"))
        digest.update(b"\0")
        if path.is_file():
            digest.update(sha256(path.read_bytes()).hexdigest().encode("ascii"))
        digest.update(b"\0")

    return {
        "git_head": head,
        "workspace_hash": digest.hexdigest(),
        "changed_paths": changed_paths,
        "status_porcelain": status_lines,
    }


def _analysis_summary(safe_dir: Path) -> dict[str, object]:
    analysis_path = safe_dir / "latest_analysis.json"
    try:
        data = json.loads(analysis_path.read_text(encoding="utf-8"))
    except Exception:
        return {"report_date": "", "marketplaces": []}
    results = data.get("marketplace_results") or []
    marketplaces = []
    if isinstance(results, list):
        marketplaces = sorted(
            {
                str(result.get("marketplace") or "").upper()
                for result in results
                if isinstance(result, dict) and result.get("marketplace")
            }
        )
    return {
        "report_date": str(data.get("report_date") or ""),
        "marketplaces": marketplaces,
    }


def write_validation_receipt(
    safe_dir: Path,
    *,
    pytest_exit_code: int,
    safe_run_exit_code: int,
    output_validation_exit_code: int,
) -> None:
    state = workspace_fingerprint()
    payload = {
        "schema_version": VALIDATION_RECEIPT_SCHEMA_VERSION,
        "result": "passed",
        "validated_at": datetime.now(timezone.utc).isoformat(),
        "pytest_exit_code": pytest_exit_code,
        "safe_run_exit_code": safe_run_exit_code,
        "output_validation_exit_code": output_validation_exit_code,
        "safe_run_dir": str(safe_dir),
        **_analysis_summary(safe_dir),
        **state,
    }
    VALIDATION_RECEIPT.parent.mkdir(parents=True, exist_ok=True)
    VALIDATION_RECEIPT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[check] wrote validation receipt: {VALIDATION_RECEIPT}", flush=True)


def clear_validation_receipt() -> None:
    if VALIDATION_RECEIPT.exists():
        VALIDATION_RECEIPT.unlink()


def _safe_run_daily_refresh_paths(safe_dir: Path) -> list[Path]:
    return [safe_dir / name for name in DAILY_REFRESH_REQUIRED_FILES]


def validate_safe_run_outputs(safe_dir: Path) -> int:
    missing = [name for name in [*REQUIRED_FILES, *REQUIRED_ASSET_FILES] if not (safe_dir / name).exists()]
    if missing:
        return fail(f"safe-run missing files: {', '.join(missing)}")
    for name in REQUIRED_ASSET_FILES:
        if (safe_dir / name).stat().st_size <= 0:
            return fail(f"safe-run asset is empty: {name}")
    asset_failures = daily_asset_content_failures([safe_dir / name for name in REQUIRED_ASSET_FILES])
    if asset_failures:
        return fail(f"safe-run {asset_failures[0]}")

    excel_files = sorted(safe_dir.glob("amazon_ops_report_*.xlsx"))
    if not excel_files:
        return fail("safe-run missing amazon_ops_report_*.xlsx")

    analysis_path = safe_dir / "latest_analysis.json"
    try:
        data = json.loads(analysis_path.read_text(encoding="utf-8"))
    except Exception as exc:
        return fail(f"cannot read latest_analysis.json: {exc}")

    report_date = str(data.get("report_date") or "").strip()
    if not report_date:
        return fail("latest_analysis.json missing report_date")

    results = data.get("marketplace_results")
    if not isinstance(results, list) or len(results) != 3:
        return fail("latest_analysis.json must contain three marketplace_results")
    marketplace_sequence = [
        str(result.get("marketplace") or "").upper()
        for result in results
        if isinstance(result, dict)
    ]
    if len(marketplace_sequence) != 3:
        return fail("latest_analysis.json marketplace_results must contain three object rows")
    marketplace_counts = Counter(marketplace_sequence)
    duplicate_marketplaces = sorted(
        marketplace
        for marketplace, count in marketplace_counts.items()
        if marketplace and count > 1
    )
    if duplicate_marketplaces or len(marketplace_sequence) != len({"UK", "US", "DE"}):
        return fail(
            "marketplace_results must contain exactly one UK, one US, and one DE; "
            f"got {marketplace_sequence}"
        )
    marketplaces = set(marketplace_sequence)
    if marketplaces != {"UK", "US", "DE"}:
        return fail(f"marketplace_results must contain UK, US, DE; got {sorted(marketplaces)}")
    result_by_marketplace = {
        str(result.get("marketplace") or "").upper(): result
        for result in results
        if isinstance(result, dict)
    }
    has_pending_ad_rows = _safe_run_has_pending_ad_workbench_rows(data)
    all_product_decision_rows: list[dict[str, object]] = []
    all_inventory_replenishment_rows: list[dict[str, object]] = []
    all_today_task_rows: list[dict[str, object]] = []
    all_tomorrow_review_rows: list[dict[str, object]] = []
    all_listing_diagnosis_rows: list[dict[str, object]] = []
    all_cost_diagnosis_rows: list[dict[str, object]] = []
    all_action_review_rows: list[dict[str, object]] = []
    all_keyword_review_rows: list[dict[str, object]] = []
    product_decision_row_count = 0
    inventory_replenishment_row_count = 0
    today_task_row_count = 0
    tomorrow_review_row_count = 0
    listing_diagnosis_row_count = 0
    cost_diagnosis_row_count = 0
    action_review_row_count = 0
    keyword_review_row_count = 0
    market_product_decision_counts: dict[str, int] = {}
    market_inventory_replenishment_counts: dict[str, int] = {}
    market_today_task_counts: dict[str, int] = {}
    market_tomorrow_review_counts: dict[str, int] = {}
    market_listing_diagnosis_counts: dict[str, int] = {}
    market_cost_diagnosis_counts: dict[str, int] = {}
    market_action_review_counts: dict[str, int] = {}
    market_keyword_review_counts: dict[str, int] = {}
    market_product_decision_rows: dict[str, list[dict[str, object]]] = {}
    market_inventory_replenishment_rows: dict[str, list[dict[str, object]]] = {}
    market_today_task_rows: dict[str, list[dict[str, object]]] = {}
    market_tomorrow_review_rows: dict[str, list[dict[str, object]]] = {}
    market_listing_diagnosis_rows: dict[str, list[dict[str, object]]] = {}
    market_cost_diagnosis_rows: dict[str, list[dict[str, object]]] = {}
    market_action_review_rows: dict[str, list[dict[str, object]]] = {}
    market_keyword_review_rows: dict[str, list[dict[str, object]]] = {}

    import_summary = data.get("import_summary") or {}
    for key in ["ads_imported_rows", "erp_imported_rows"]:
        if key not in import_summary:
            return fail(f"import_summary missing {key}")

    for result in results:
        marketplace = str(result.get("marketplace") or "").upper()
        summary = result.get("summary") or {}
        if not summary.get("report_date"):
            return fail(f"{marketplace} summary missing report_date")
        snapshot = result.get("report_view_snapshot") or {}
        missing_snapshot = sorted(REQUIRED_SNAPSHOT_KEYS.difference(snapshot.keys()))
        if missing_snapshot:
            return fail(f"{marketplace} report_view_snapshot missing {', '.join(missing_snapshot)}")
        frontend_rows = snapshot.get("frontend_check_queue_rows") or []
        if not isinstance(frontend_rows, list):
            return fail(f"{marketplace} frontend_check_queue_rows must be a list")
        code = _validate_frontend_rows(marketplace, frontend_rows)
        if code != 0:
            return code
        coverage_failures = daily_frontend_coverage_summary_failures(
            marketplace,
            frontend_rows,
            snapshot.get("frontend_coverage_summary"),
        )
        if coverage_failures:
            return fail(coverage_failures[0])
        today_task_rows = snapshot.get("today_task_queue_rows") or []
        if not isinstance(today_task_rows, list):
            return fail(f"{marketplace} today_task_queue_rows must be a list")
        all_today_task_rows.extend(today_task_rows)
        today_task_row_count += len(today_task_rows)
        market_today_task_counts[marketplace] = len(today_task_rows)
        market_today_task_rows[marketplace] = today_task_rows
        tomorrow_review_rows = snapshot.get("tomorrow_review_rows") or []
        if not isinstance(tomorrow_review_rows, list):
            return fail(f"{marketplace} tomorrow_review_rows must be a list")
        all_tomorrow_review_rows.extend(tomorrow_review_rows)
        tomorrow_review_row_count += len(tomorrow_review_rows)
        market_tomorrow_review_counts[marketplace] = len(tomorrow_review_rows)
        market_tomorrow_review_rows[marketplace] = tomorrow_review_rows
        listing_diagnosis_rows = snapshot.get("listing_price_diagnosis_rows") or []
        if not isinstance(listing_diagnosis_rows, list):
            return fail(f"{marketplace} listing_price_diagnosis_rows must be a list")
        listing_frontend_failures = daily_listing_frontend_evidence_failures(
            marketplace,
            listing_diagnosis_rows,
            frontend_rows,
        )
        if listing_frontend_failures:
            return fail(listing_frontend_failures[0])
        all_listing_diagnosis_rows.extend(listing_diagnosis_rows)
        listing_diagnosis_row_count += len(listing_diagnosis_rows)
        market_listing_diagnosis_counts[marketplace] = len(listing_diagnosis_rows)
        market_listing_diagnosis_rows[marketplace] = listing_diagnosis_rows
        cost_diagnosis_rows = snapshot.get("cost_profit_diagnosis_rows") or []
        if not isinstance(cost_diagnosis_rows, list):
            return fail(f"{marketplace} cost_profit_diagnosis_rows must be a list")
        all_cost_diagnosis_rows.extend(cost_diagnosis_rows)
        cost_diagnosis_row_count += len(cost_diagnosis_rows)
        market_cost_diagnosis_counts[marketplace] = len(cost_diagnosis_rows)
        market_cost_diagnosis_rows[marketplace] = cost_diagnosis_rows
        code = _validate_ad_observation_rows(marketplace, _iter_ad_action_rows(snapshot))
        if code != 0:
            return code
        product_decision_rows = snapshot.get("product_final_decision_rows") or []
        if not isinstance(product_decision_rows, list):
            return fail(f"{marketplace} product_final_decision_rows must be a list")
        all_product_decision_rows.extend(product_decision_rows)
        product_decision_row_count += len(product_decision_rows)
        market_product_decision_counts[marketplace] = len(product_decision_rows)
        market_product_decision_rows[marketplace] = product_decision_rows
        code = _validate_product_final_decision_rows(marketplace, product_decision_rows)
        if code != 0:
            return code
        product_operation_rows = snapshot.get("product_operation_cards") or []
        if not isinstance(product_operation_rows, list):
            return fail(f"{marketplace} product_operation_cards must be a list")
        operation_failures = daily_product_operation_card_failures(
            marketplace,
            product_operation_rows,
            product_decision_rows,
            daily_product_operation_ad_rows_from_snapshot(snapshot),
        )
        if operation_failures:
            return fail(operation_failures[0])
        code = _validate_task_queue_growth_gate(marketplace, today_task_rows, product_decision_rows)
        if code != 0:
            return code
        aux_ad_rows = _iter_aux_ad_workbench_rows(snapshot)
        code = _validate_task_queue_growth_gate(
            marketplace,
            aux_ad_rows,
            product_decision_rows,
            row_source="ad_workbench_rows",
        )
        if code != 0:
            return code
        code = _validate_decision_count_summary(marketplace, snapshot, "final_decision_summary")
        if code != 0:
            return code
        code = _validate_decision_count_summary(marketplace, snapshot, "decision_gate_counts")
        if code != 0:
            return code
        inventory_replenishment_rows = snapshot.get("inventory_replenishment_rows") or []
        if not isinstance(inventory_replenishment_rows, list):
            return fail(f"{marketplace} inventory_replenishment_rows must be a list")
        code = _validate_marketplace_scoped_rows(
            marketplace,
            inventory_replenishment_rows,
            "inventory replenishment",
        )
        if code != 0:
            return code
        all_inventory_replenishment_rows.extend(inventory_replenishment_rows)
        inventory_replenishment_row_count += len(inventory_replenishment_rows)
        market_inventory_replenishment_counts[marketplace] = len(inventory_replenishment_rows)
        market_inventory_replenishment_rows[marketplace] = inventory_replenishment_rows
        action_review_rows = snapshot.get("action_effect_review_rows") or []
        if not isinstance(action_review_rows, list):
            return fail(f"{marketplace} action_effect_review_rows must be a list")
        all_action_review_rows.extend(action_review_rows)
        action_review_row_count += len(action_review_rows)
        market_action_review_counts[marketplace] = len(action_review_rows)
        market_action_review_rows[marketplace] = action_review_rows
        code = _validate_action_review_rows(marketplace, action_review_rows)
        if code != 0:
            return code
        keyword_review_rows = snapshot.get("keyword_action_effect_review_rows") or []
        if not isinstance(keyword_review_rows, list):
            return fail(f"{marketplace} keyword_action_effect_review_rows must be a list")
        all_keyword_review_rows.extend(keyword_review_rows)
        keyword_review_row_count += len(keyword_review_rows)
        market_keyword_review_counts[marketplace] = len(keyword_review_rows)
        market_keyword_review_rows[marketplace] = keyword_review_rows
        code = _validate_keyword_action_review_rows(marketplace, keyword_review_rows)
        if code != 0:
            return code

    code = _validate_top_level_snapshot_union(
        data,
        "product_final_decision_rows",
        PRODUCT_DECISION_IDENTITY_FIELDS,
        "product final decisions",
        PRODUCT_DECISION_EXCEL_CONTENT_FIELDS,
    )
    if code != 0:
        return code
    code = _validate_top_level_snapshot_union(
        data,
        "inventory_replenishment_rows",
        INVENTORY_REPLENISHMENT_IDENTITY_FIELDS,
        "inventory replenishment rows",
        INVENTORY_REPLENISHMENT_EXCEL_CONTENT_FIELDS,
    )
    if code != 0:
        return code
    code = _validate_top_level_marketplace_snapshot_mapping(
        data,
        "final_decision_summary",
        "final decision summary",
    )
    if code != 0:
        return code
    code = _validate_top_level_marketplace_snapshot_mapping(
        data,
        "decision_gate_counts",
        "decision gate counts",
    )
    if code != 0:
        return code
    code = _validate_top_level_frontend_coverage_summary(data)
    if code != 0:
        return code

    try:
        workbook = load_workbook(excel_files[-1], read_only=True)
    except Exception as exc:
        return fail(f"cannot open Excel workbook: {exc}")
    try:
        if "Metrics_Validation" not in workbook.sheetnames:
            return fail("Excel workbook missing Metrics_Validation sheet")
        code = _validate_workbook_frontend_coverage_summary(workbook, data)
        if code != 0:
            return code
        if product_decision_row_count:
            code = _validate_workbook_product_decision_sheet(
                workbook,
                "产品最终决策",
                expected_rows=product_decision_row_count,
            )
            if code != 0:
                return code
            code = _validate_workbook_identity(
                workbook,
                "产品最终决策",
                all_product_decision_rows,
                PRODUCT_DECISION_IDENTITY_FIELDS,
                "product final decisions",
            )
            if code != 0:
                return code
            code = _validate_workbook_content(
                workbook,
                "产品最终决策",
                all_product_decision_rows,
                PRODUCT_DECISION_IDENTITY_FIELDS,
                PRODUCT_DECISION_EXCEL_CONTENT_FIELDS,
                "product final decisions",
            )
            if code != 0:
                return code
        if inventory_replenishment_row_count:
            code = _validate_workbook_required_sheet(
                workbook,
                "库存补货提醒",
                INVENTORY_REPLENISHMENT_REQUIRED_FIELDS,
                expected_rows=inventory_replenishment_row_count,
            )
            if code != 0:
                return code
            code = _validate_workbook_identity(
                workbook,
                "库存补货提醒",
                all_inventory_replenishment_rows,
                INVENTORY_REPLENISHMENT_IDENTITY_FIELDS,
                "inventory replenishment rows",
            )
            if code != 0:
                return code
            code = _validate_workbook_content(
                workbook,
                "库存补货提醒",
                all_inventory_replenishment_rows,
                INVENTORY_REPLENISHMENT_IDENTITY_FIELDS,
                INVENTORY_REPLENISHMENT_EXCEL_CONTENT_FIELDS,
                "inventory replenishment rows",
            )
            if code != 0:
                return code
        if today_task_row_count:
            code = _validate_workbook_required_sheet(
                workbook,
                "今日动作清单",
                TODAY_TASK_QUEUE_IDENTITY_FIELDS,
                expected_rows=today_task_row_count,
            )
            if code != 0:
                return code
            code = _validate_workbook_identity(
                workbook,
                "今日动作清单",
                all_today_task_rows,
                TODAY_TASK_QUEUE_IDENTITY_FIELDS,
                "today task queue rows",
            )
            if code != 0:
                return code
            code = _validate_workbook_content(
                workbook,
                "今日动作清单",
                all_today_task_rows,
                TODAY_TASK_QUEUE_IDENTITY_FIELDS,
                TODAY_TASK_QUEUE_CONTENT_FIELDS,
                "today task queue rows",
            )
            if code != 0:
                return code
        if tomorrow_review_row_count:
            code = _validate_workbook_required_sheet(
                workbook,
                "明日复查清单",
                TOMORROW_REVIEW_REQUIRED_FIELDS,
                expected_rows=tomorrow_review_row_count,
            )
            if code != 0:
                return code
            code = _validate_workbook_identity(
                workbook,
                "明日复查清单",
                all_tomorrow_review_rows,
                TOMORROW_REVIEW_IDENTITY_FIELDS,
                "tomorrow review rows",
            )
            if code != 0:
                return code
            code = _validate_workbook_content(
                workbook,
                "明日复查清单",
                all_tomorrow_review_rows,
                TOMORROW_REVIEW_IDENTITY_FIELDS,
                TOMORROW_REVIEW_CONTENT_FIELDS,
                "tomorrow review rows",
            )
            if code != 0:
                return code
        if listing_diagnosis_row_count:
            code = _validate_workbook_required_sheet(
                workbook,
                "Listing待确认",
                DIAGNOSIS_ROW_REQUIRED_FIELDS,
                expected_rows=listing_diagnosis_row_count,
            )
            if code != 0:
                return code
            code = _validate_workbook_identity(
                workbook,
                "Listing待确认",
                all_listing_diagnosis_rows,
                DIAGNOSIS_ROW_IDENTITY_FIELDS,
                "listing price diagnosis rows",
            )
            if code != 0:
                return code
            code = _validate_workbook_content(
                workbook,
                "Listing待确认",
                all_listing_diagnosis_rows,
                DIAGNOSIS_ROW_IDENTITY_FIELDS,
                DIAGNOSIS_ROW_CONTENT_FIELDS,
                "listing price diagnosis rows",
            )
            if code != 0:
                return code
        if cost_diagnosis_row_count:
            code = _validate_workbook_required_sheet(
                workbook,
                "成本利润诊断",
                DIAGNOSIS_ROW_REQUIRED_FIELDS,
                expected_rows=cost_diagnosis_row_count,
            )
            if code != 0:
                return code
            code = _validate_workbook_identity(
                workbook,
                "成本利润诊断",
                all_cost_diagnosis_rows,
                DIAGNOSIS_ROW_IDENTITY_FIELDS,
                "cost profit diagnosis rows",
            )
            if code != 0:
                return code
            code = _validate_workbook_content(
                workbook,
                "成本利润诊断",
                all_cost_diagnosis_rows,
                DIAGNOSIS_ROW_IDENTITY_FIELDS,
                DIAGNOSIS_ROW_CONTENT_FIELDS,
                "cost profit diagnosis rows",
            )
            if code != 0:
                return code
        if action_review_row_count:
            code = _validate_workbook_review_sheet(
                workbook,
                "执行后效果复盘",
                ACTION_REVIEW_REQUIRED_FIELDS,
                expected_rows=action_review_row_count,
            )
            if code != 0:
                return code
            code = _validate_workbook_identity(
                workbook,
                "执行后效果复盘",
                all_action_review_rows,
                ACTION_REVIEW_IDENTITY_FIELDS,
                "action reviews",
            )
            if code != 0:
                return code
            code = _validate_workbook_content(
                workbook,
                "执行后效果复盘",
                all_action_review_rows,
                ACTION_REVIEW_IDENTITY_FIELDS,
                ACTION_REVIEW_REQUIRED_FIELDS,
                "action reviews",
            )
            if code != 0:
                return code
        if keyword_review_row_count:
            code = _validate_workbook_review_sheet(
                workbook,
                "词级执行复盘",
                KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS,
                expected_rows=keyword_review_row_count,
            )
            if code != 0:
                return code
            code = _validate_workbook_identity(
                workbook,
                "词级执行复盘",
                all_keyword_review_rows,
                KEYWORD_REVIEW_IDENTITY_FIELDS,
                "keyword action reviews",
            )
            if code != 0:
                return code
            code = _validate_workbook_content(
                workbook,
                "词级执行复盘",
                all_keyword_review_rows,
                KEYWORD_REVIEW_IDENTITY_FIELDS,
                KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS,
                "keyword action reviews",
            )
            if code != 0:
                return code
        for marketplace, row_count in sorted(market_product_decision_counts.items()):
            if not row_count:
                continue
            code = _validate_workbook_product_decision_sheet(
                workbook,
                f"{marketplace}_产品最终决策",
                expected_rows=row_count,
            )
            if code != 0:
                return code
            code = _validate_workbook_marketplace_scope(
                workbook,
                f"{marketplace}_产品最终决策",
                marketplace,
            )
            if code != 0:
                return code
            code = _validate_workbook_identity(
                workbook,
                f"{marketplace}_产品最终决策",
                market_product_decision_rows.get(marketplace, []),
                PRODUCT_DECISION_IDENTITY_FIELDS,
                f"{marketplace} product final decisions",
            )
            if code != 0:
                return code
            code = _validate_workbook_content(
                workbook,
                f"{marketplace}_产品最终决策",
                market_product_decision_rows.get(marketplace, []),
                PRODUCT_DECISION_IDENTITY_FIELDS,
                PRODUCT_DECISION_EXCEL_CONTENT_FIELDS,
                f"{marketplace} product final decisions",
            )
            if code != 0:
                return code
        for marketplace, row_count in sorted(market_inventory_replenishment_counts.items()):
            if not row_count:
                continue
            code = _validate_workbook_required_sheet(
                workbook,
                f"{marketplace}_库存补货提醒",
                INVENTORY_REPLENISHMENT_REQUIRED_FIELDS,
                expected_rows=row_count,
            )
            if code != 0:
                return code
            code = _validate_workbook_marketplace_scope(
                workbook,
                f"{marketplace}_库存补货提醒",
                marketplace,
            )
            if code != 0:
                return code
            code = _validate_workbook_identity(
                workbook,
                f"{marketplace}_库存补货提醒",
                market_inventory_replenishment_rows.get(marketplace, []),
                INVENTORY_REPLENISHMENT_IDENTITY_FIELDS,
                f"{marketplace} inventory replenishment rows",
            )
            if code != 0:
                return code
            code = _validate_workbook_content(
                workbook,
                f"{marketplace}_库存补货提醒",
                market_inventory_replenishment_rows.get(marketplace, []),
                INVENTORY_REPLENISHMENT_IDENTITY_FIELDS,
                INVENTORY_REPLENISHMENT_EXCEL_CONTENT_FIELDS,
                f"{marketplace} inventory replenishment rows",
            )
            if code != 0:
                return code
        for marketplace, row_count in sorted(market_today_task_counts.items()):
            if not row_count:
                continue
            sheet_name = f"{marketplace}_今日动作清单"
            code = _validate_workbook_required_sheet(
                workbook,
                sheet_name,
                TODAY_TASK_QUEUE_IDENTITY_FIELDS,
                expected_rows=row_count,
            )
            if code != 0:
                return code
            code = _validate_workbook_marketplace_scope(
                workbook,
                sheet_name,
                marketplace,
            )
            if code != 0:
                return code
            code = _validate_workbook_identity(
                workbook,
                sheet_name,
                market_today_task_rows.get(marketplace, []),
                TODAY_TASK_QUEUE_IDENTITY_FIELDS,
                f"{marketplace} today task queue rows",
            )
            if code != 0:
                return code
            code = _validate_workbook_content(
                workbook,
                sheet_name,
                market_today_task_rows.get(marketplace, []),
                TODAY_TASK_QUEUE_IDENTITY_FIELDS,
                TODAY_TASK_QUEUE_CONTENT_FIELDS,
                f"{marketplace} today task queue rows",
            )
            if code != 0:
                return code
        for marketplace, row_count in sorted(market_tomorrow_review_counts.items()):
            if not row_count:
                continue
            sheet_name = f"{marketplace}_明日复查清单"
            code = _validate_workbook_required_sheet(
                workbook,
                sheet_name,
                TOMORROW_REVIEW_REQUIRED_FIELDS,
                expected_rows=row_count,
            )
            if code != 0:
                return code
            code = _validate_workbook_marketplace_scope(
                workbook,
                sheet_name,
                marketplace,
            )
            if code != 0:
                return code
            code = _validate_workbook_identity(
                workbook,
                sheet_name,
                market_tomorrow_review_rows.get(marketplace, []),
                TOMORROW_REVIEW_IDENTITY_FIELDS,
                f"{marketplace} tomorrow review rows",
            )
            if code != 0:
                return code
            code = _validate_workbook_content(
                workbook,
                sheet_name,
                market_tomorrow_review_rows.get(marketplace, []),
                TOMORROW_REVIEW_IDENTITY_FIELDS,
                TOMORROW_REVIEW_CONTENT_FIELDS,
                f"{marketplace} tomorrow review rows",
            )
            if code != 0:
                return code
        for marketplace, row_count in sorted(market_listing_diagnosis_counts.items()):
            if not row_count:
                continue
            sheet_name = f"{marketplace}_Listing待确认"
            code = _validate_workbook_required_sheet(
                workbook,
                sheet_name,
                DIAGNOSIS_ROW_REQUIRED_FIELDS,
                expected_rows=row_count,
            )
            if code != 0:
                return code
            code = _validate_workbook_marketplace_scope(workbook, sheet_name, marketplace)
            if code != 0:
                return code
            code = _validate_workbook_identity(
                workbook,
                sheet_name,
                market_listing_diagnosis_rows.get(marketplace, []),
                DIAGNOSIS_ROW_IDENTITY_FIELDS,
                f"{marketplace} listing price diagnosis rows",
            )
            if code != 0:
                return code
            code = _validate_workbook_content(
                workbook,
                sheet_name,
                market_listing_diagnosis_rows.get(marketplace, []),
                DIAGNOSIS_ROW_IDENTITY_FIELDS,
                DIAGNOSIS_ROW_CONTENT_FIELDS,
                f"{marketplace} listing price diagnosis rows",
            )
            if code != 0:
                return code
        for marketplace, row_count in sorted(market_cost_diagnosis_counts.items()):
            if not row_count:
                continue
            sheet_name = f"{marketplace}_成本利润诊断"
            code = _validate_workbook_required_sheet(
                workbook,
                sheet_name,
                DIAGNOSIS_ROW_REQUIRED_FIELDS,
                expected_rows=row_count,
            )
            if code != 0:
                return code
            code = _validate_workbook_marketplace_scope(workbook, sheet_name, marketplace)
            if code != 0:
                return code
            code = _validate_workbook_identity(
                workbook,
                sheet_name,
                market_cost_diagnosis_rows.get(marketplace, []),
                DIAGNOSIS_ROW_IDENTITY_FIELDS,
                f"{marketplace} cost profit diagnosis rows",
            )
            if code != 0:
                return code
            code = _validate_workbook_content(
                workbook,
                sheet_name,
                market_cost_diagnosis_rows.get(marketplace, []),
                DIAGNOSIS_ROW_IDENTITY_FIELDS,
                DIAGNOSIS_ROW_CONTENT_FIELDS,
                f"{marketplace} cost profit diagnosis rows",
            )
            if code != 0:
                return code
        for marketplace, row_count in sorted(market_action_review_counts.items()):
            if not row_count:
                continue
            code = _validate_workbook_review_sheet(
                workbook,
                f"{marketplace}_执行后复盘",
                ACTION_REVIEW_REQUIRED_FIELDS,
                expected_rows=row_count,
                expected_marketplace=marketplace,
            )
            if code != 0:
                return code
            code = _validate_workbook_identity(
                workbook,
                f"{marketplace}_执行后复盘",
                market_action_review_rows.get(marketplace, []),
                ACTION_REVIEW_IDENTITY_FIELDS,
                f"{marketplace} action reviews",
            )
            if code != 0:
                return code
            code = _validate_workbook_content(
                workbook,
                f"{marketplace}_执行后复盘",
                market_action_review_rows.get(marketplace, []),
                ACTION_REVIEW_IDENTITY_FIELDS,
                ACTION_REVIEW_REQUIRED_FIELDS,
                f"{marketplace} action reviews",
            )
            if code != 0:
                return code
        for marketplace, row_count in sorted(market_keyword_review_counts.items()):
            if not row_count:
                continue
            code = _validate_workbook_review_sheet(
                workbook,
                f"{marketplace}_词级执行复盘",
                KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS,
                expected_rows=row_count,
                expected_marketplace=marketplace,
            )
            if code != 0:
                return code
            code = _validate_workbook_identity(
                workbook,
                f"{marketplace}_词级执行复盘",
                market_keyword_review_rows.get(marketplace, []),
                KEYWORD_REVIEW_IDENTITY_FIELDS,
                f"{marketplace} keyword action reviews",
            )
            if code != 0:
                return code
            code = _validate_workbook_content(
                workbook,
                f"{marketplace}_词级执行复盘",
                market_keyword_review_rows.get(marketplace, []),
                KEYWORD_REVIEW_IDENTITY_FIELDS,
                KEYWORD_ACTION_REVIEW_REQUIRED_FIELDS,
                f"{marketplace} keyword action reviews",
            )
            if code != 0:
                return code
    finally:
        workbook.close()

    for name in REQUIRED_FILES:
        if not name.endswith(".html"):
            continue
        text = (safe_dir / name).read_text(encoding="utf-8", errors="replace")
        if report_date not in text:
            return fail(f"{name} does not show report_date {report_date}")
        marker = next((item for item in HTML_ERROR_MARKERS if item in text), "")
        if marker:
            return fail(f"{name} contains error marker {marker}")
        for required_marker in REQUIRED_HTML_MARKERS.get(name, []):
            if required_marker not in text:
                return fail(f"{name} missing required marker {required_marker}")
        if name == "dashboard.html":
            code = _validate_dashboard_marketplace_summary(text, data)
            if code != 0:
                return code
        if name == "summary.html":
            code = _validate_summary_snapshot_rows(text, data)
            if code != 0:
                return code
        if name == "latest_recommendations.html":
            code = _validate_latest_recommendations_ad_area(text, has_pending_ad_rows=has_pending_ad_rows)
            if code != 0:
                return code
            code = _validate_latest_recommendations_snapshot_rows(text, data)
            if code != 0:
                return code
        marketplace_name = name.split("_", 1)[0].upper() if name.endswith("_report.html") else ""
        if marketplace_name in result_by_marketplace:
            code = _validate_marketplace_report_snapshot_rows(name, text, result_by_marketplace[marketplace_name])
            if code != 0:
                return code
        for forbidden_marker in FORBIDDEN_HTML_MARKERS.get(name, []):
            if forbidden_marker in text:
                return fail(f"{name} contains forbidden marker {forbidden_marker}")

    summary_markdown_path = safe_dir / "marketplace_summary.md"
    if summary_markdown_path.exists():
        text = summary_markdown_path.read_text(encoding="utf-8", errors="replace")
        if report_date not in text:
            return fail(f"marketplace_summary.md does not show report_date {report_date}")
        code = _validate_marketplace_summary_markdown(text, data)
        if code != 0:
            return code

    formal_failures = report_refresh_failures(
        _safe_run_daily_refresh_paths(safe_dir),
        previous_mtimes_ns={},
    )
    if formal_failures:
        shown = "; ".join(formal_failures[:5])
        suffix = f"; ... {len(formal_failures) - 5} more" if len(formal_failures) > 5 else ""
        return fail(f"safe-run formal daily validation failed: {shown}{suffix}")

    print(f"[check] safe-run outputs validated: {safe_dir}", flush=True)
    return 0


def main() -> int:
    clear_validation_receipt()
    python = sys.executable
    code = run_step([python, "-m", "pytest"])
    if code != 0:
        return fail("pytest failed")
    pytest_exit_code = code

    safe_run_id = f"showcase_{int(time.time() * 1000)}_{os.getpid()}"
    safe_dir = SAFE_RUN_ROOT / safe_run_id
    if safe_dir.exists():
        return fail(f"safe-run output directory already exists: {safe_dir}")

    code = run_step(
        [python, "main.py", "--marketplace", "ALL", "--safe-run"],
        extra_env={"AMAZON_OPS_SAFE_RUN_ID": safe_run_id},
    )
    if code != 0:
        return fail("ALL safe-run failed")
    safe_run_exit_code = code

    if not safe_dir.exists():
        return fail(f"cannot find expected safe-run output directory: {safe_dir}")

    code = validate_safe_run_outputs(safe_dir)
    if code != 0:
        return code
    output_validation_exit_code = code

    write_validation_receipt(
        safe_dir,
        pytest_exit_code=pytest_exit_code,
        safe_run_exit_code=safe_run_exit_code,
        output_validation_exit_code=output_validation_exit_code,
    )
    print("[done] showcase MVP validation passed: pytest, ALL safe-run, and generated outputs checked", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
